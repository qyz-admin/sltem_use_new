import pandas as pd
import os
import datetime
import xlwings
import win32api, win32con
import win32com.client as win32
import requests
import json
import sys
from sso_updata import Query_sso_updata
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel

from sqlalchemy import create_engine
from settings import Settings
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色

from mysqlControl import MysqlControl
from settings_sso import Settings_sso
# -*- coding:utf-8 -*-
class QueryUpdate(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
        self.m = MysqlControl()
        # self.sso = QueryTwo()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    # 获取签收表内容---港澳台更新签收总表
    def readFormHost(self, upload):
        match = {'查询-运单号': r'D:\Users\Administrator\Desktop\需要用到的文件\A查询导表',
                 '查询运费': r'F:\神龙签收率\A运费-核实',
                 '其他': r'D:\Users\Administrator\Desktop\需要用到的文件\B客服工作表'}
        path = match[upload]
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                self.wbsheetHost(filePath, upload)
                if upload == '查询运费':
                    excel = win32.gencache.EnsureDispatch('Excel.Application')
                    wb = excel.Workbooks.Open(filePath)
                    file_path = os.path.join(path, "~$ " + dir)
                    wb.SaveAs(file_path, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
                    wb.Close()  # FileFormat = 56 is for .xls extension
                    excel.Application.Quit()
                    os.remove(filePath)
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, upload):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    # print(db.columns)
                    if upload == '查询-订单号':
                        db = db[['订单编号']]
                    elif upload == '查询-运单号':
                        db = db[['运单编号']]
                    elif upload == '查询运费':
                        db = db[['订单编号', '选品人', '备注', '产品id', '规格中文']]
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在导入查询：' + sht.name + '表； 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    self.writeCacheHost(db, upload)
                    print('++++正在获取：' + sht.name + '--->>>到查询缓存表')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()

    # 写入更新缓存表
    def writeCacheHost(self, dataFrame, upload):
        month_begin = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y-%m-%d')
        df = None
        if upload == '查询-订单号':
            dataFrame.to_sql('sheet1_iphone', con=self.engine1, index=False, if_exists='replace')
            sql = '''SELECT gat_zqsb.订单编号,gat_zqsb.系统订单状态,gat_zqsb.`系统物流状态`,gat_zqsb.`物流状态`,gat_zqsb.`最终状态`
                            FROM sheet1_iphone
            	            LEFT JOIN gat_zqsb ON sheet1_iphone.`订单编号` = gat_zqsb.`订单编号`;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
        elif upload == '查询-运单号':
            dataFrame.to_sql('sheet1_iphone_cy', con=self.engine1, index=False, if_exists='replace')
            print('正在获取查询数据内容…………')
            sql = '''SELECT b.订单编号, b.运单编号, IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态
                    FROM sheet1_iphone_cy a
                    LEFT JOIN (SELECT * FROM gat WHERE id IN (SELECT MAX(id) FROM gat WHERE gat.添加时间 > '{0} 00:00:00' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                    LEFT JOIN gat_logisitis_match c ON b.物流状态 = c.签收表物流状态;'''.format(month_begin)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
        elif upload == '查询运费':
            dataFrame.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            columns = list(dataFrame.columns)
            columns = ','.join(columns)
            sql = '''REPLACE INTO 运费核实({}, 记录时间) SELECT *, NOW() 记录时间 FROM customer;'''.format(columns)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('++++成功导入缓存表')
            return

        print('正在写入excel…………')
        rq = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
        df.to_excel('G:\\输出文件\\订单检索-查询{}.xlsx'.format(rq), sheet_name='查询', index=False)
        print('----已写入excel')

    # 停用
    def trans_way_cost(self, team):
        match = {'gat': '港台', 'slsc': '品牌'}
        output = datetime.datetime.now().strftime('%m.%d')
        month_yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        month_yesterday = '2021-09-14'
        print(month_yesterday)
        month_now = (datetime.datetime.now()).strftime('%Y%m')
        month_now = '202109'
        print(month_now)

        sql = '''SELECT 年月,日期,团队,币种,订单编号,数量,电话号码,运单编号,是否改派,物流方式,商品id,ds.产品id,产品名称,价格,下单时间,审核时间,仓储扫描时间,完结状态,完结状态时间,物流花费,包裹重量,包裹体积,规格中文,产品量
                        FROM gat_order_list ds
                        LEFT JOIN (SELECT 产品id, COUNT(订单编号) 产品量
                                    FROM gat_order_list ds
        					        WHERE ds.`日期` = '{0}' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND ds.是否改派 = '直发'
        					        GROUP BY ds.`产品id`
        				) dds on ds.`产品id` = dds.`产品id`
                        WHERE ds.`年月` = '{1}' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND ds.`包裹重量` <> 0 AND ds.`是否改派` = '直发'
        			        AND ds.`规格中文` IN (SELECT `规格中文`
                                                 FROM (SELECT 产品id,`规格中文`,COUNT(订单编号) 单量, MIN(包裹重量), MAX(包裹重量),  MAX(包裹重量)-MIN(包裹重量) as 重量差
                                                      FROM gat_order_list d 
        											  WHERE d.`年月` = '{1}' and d.`是否改派` = '直发' AND d.`产品id` <> 0 AND d.`包裹重量` <> 0
        											    AND d.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND d.是否改派 = '直发'
        											  GROUP BY d.`产品id`,d.`规格中文`
        											  ORDER BY d. 产品id
        										    ) s1
        										WHERE s1.`重量差` > 100 AND s1.`单量` >= 2
        									   )	
        			        AND ds.`产品id` IN (SELECT s.`产品id`
        										FROM (SELECT 年月,日期,团队,币种,订单编号,数量,电话号码,运单编号,是否改派,物流方式,商品id,产品id,产品名称,价格,下单时间,审核时间,仓储扫描时间,完结状态,完结状态时间,物流花费,包裹重量,包裹体积,规格中文
        											 FROM gat_order_list ds
        											 WHERE ds.`年月` = '{1}' AND ds.`产品id` <> 0 AND ds.`包裹重量` <> 0 and ds.`是否改派` = '直发'
        											   AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND ds.是否改派 = '直发'
        											   AND ds.`规格中文` IN (SELECT `规格中文`
        																	 FROM (SELECT 产品id,`规格中文`,COUNT(订单编号) 单量, MIN(包裹重量), MAX(包裹重量),  MAX(包裹重量)-MIN(包裹重量) as 重量差
        																		  FROM gat_order_list d 
        																		  WHERE d.`年月` = '{1}' and d.`是否改派` = '直发' AND d.`产品id` <> 0 AND d.`包裹重量` <> 0
        																			AND d.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND d.是否改派 = '直发'
        																		  GROUP BY d.`产品id`,d.`规格中文`
        																		  ORDER BY d. 产品id
        																		 ) s1
        																	WHERE s1.`重量差` > 100 AND s1.`单量` >= 2
        																   ) ORDER BY 日期
        											 ) s
        											 GROUP BY  s.`产品id`
        											 HAVING count(s.`产品id`) >1
        										)
                        ORDER BY 日期;'''.format(month_yesterday, month_now)
        print('正在获取 ' + match[team] + ' 运费总直发情况…………')
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('gat_trans_way', con=self.engine1, index=False, if_exists='replace')
        print('正在获取运费总直发内容…………')
        sql = '''SELECT * 	
                        FROM( SELECT * 
        			            FROM gat_trans_way ds
        			            LEFT JOIN (SELECT *
        								    FROM (SELECT 产品id '产品id2',`规格中文` '规格中文2',包裹重量 '包裹重量2', COUNT(订单编号) 单量, MIN(包裹重量), MAX(包裹重量),  MAX(包裹重量)-MIN(包裹重量) as 重量差
        											FROM gat_trans_way d 
        											GROUP BY d.`产品id`,d.`规格中文`
        											ORDER BY d. 产品id
        											) s1
        								    WHERE s1.`重量差` > 100 AND s1.`单量` >= 2
        								    ) dss ON ds.`产品ID` = dss.`产品id2` and ds.`规格中文`= dss.`规格中文2`
                        ) s WHERE s.重量差 IS not null;'''
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df = df[['年月', '团队',  '日期', '币种', '订单编号', '数量', '电话号码', '运单编号', '是否改派', '物流方式', '商品id', '产品id', '产品名称',
                 '价格', '仓储扫描时间', '完结状态', '物流花费', '包裹重量', '包裹体积', '规格中文',
                 '产品量', 'MIN(包裹重量)', 'MAX(包裹重量)', '重量差']]
        print('正在写入excel…………')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        df.to_excel('G:\\输出文件\\{} 运费总直发-查询{}.xlsx'.format(output, rq),
                    sheet_name='查询', index=False)
        print('----已写入excel')

        sql = '''SELECT 年月,日期,团队,币种,订单编号,数量,电话号码,运单编号,是否改派,物流方式,商品id,ds.产品id,产品名称,价格,下单时间,审核时间,仓储扫描时间,完结状态,完结状态时间,物流花费,包裹重量,包裹体积,规格中文,产品量
                FROM gat_order_list ds
                LEFT JOIN (SELECT 产品id, COUNT(订单编号) 产品量
                            FROM gat_order_list ds
					        WHERE ds.`日期` = '{0}' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND ds.`币种` = '台湾' AND ds.是否改派 = '直发'
					        GROUP BY ds.`产品id`
				) dds on ds.`产品id` = dds.`产品id`
                WHERE ds.`年月` = '{1}' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND ds.`包裹重量` <> 0 AND ds.`是否改派` = '直发'
			        AND ds.`规格中文` IN (SELECT `规格中文`
                                         FROM (SELECT 产品id,`规格中文`,COUNT(订单编号) 单量, MIN(包裹重量), MAX(包裹重量),  MAX(包裹重量)-MIN(包裹重量) as 重量差
                                              FROM gat_order_list d 
											  WHERE d.`年月` = '{1}' and d.`是否改派` = '直发' AND d.`产品id` <> 0 AND d.`包裹重量` <> 0
											    AND d.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND d.`币种` = '台湾' AND d.是否改派 = '直发'
											  GROUP BY d.`产品id`,d.`规格中文`
											  ORDER BY d. 产品id
										    ) s1
										WHERE s1.`重量差` > 100 AND s1.`单量` >= 2
									   )	
			        AND ds.`产品id` IN (SELECT s.`产品id`
										FROM (SELECT 年月,日期,团队,币种,订单编号,数量,电话号码,运单编号,是否改派,物流方式,商品id,产品id,产品名称,价格,下单时间,审核时间,仓储扫描时间,完结状态,完结状态时间,物流花费,包裹重量,包裹体积,规格中文
											 FROM gat_order_list ds
											 WHERE ds.`年月` = '{1}' AND ds.`产品id` <> 0 AND ds.`包裹重量` <> 0 and ds.`是否改派` = '直发'
											   AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')  AND ds.`币种` = '台湾' AND ds.是否改派 = '直发'
											   AND ds.`规格中文` IN (SELECT `规格中文`
																	 FROM (SELECT 产品id,`规格中文`,COUNT(订单编号) 单量, MIN(包裹重量), MAX(包裹重量),  MAX(包裹重量)-MIN(包裹重量) as 重量差
																		  FROM gat_order_list d 
																		  WHERE d.`年月` = '{1}' and d.`是否改派` = '直发' AND d.`产品id` <> 0 AND d.`包裹重量` <> 0
																			AND d.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND d.`币种` = '台湾' AND d.是否改派 = '直发'
																		  GROUP BY d.`产品id`,d.`规格中文`
																		  ORDER BY d. 产品id
																		 ) s1
																	WHERE s1.`重量差` > 100 AND s1.`单量` >= 2
																   ) ORDER BY 日期
											 ) s
											 GROUP BY  s.`产品id`
											 HAVING count(s.`产品id`) >1
										)
                ORDER BY 日期;'''.format(month_yesterday, month_now)
        print('正在获取 ' + match[team] + ' 运费台湾直发-情况…………')
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('gat_trans_way', con=self.engine1, index=False, if_exists='replace')
        print('正在获取-运费台湾直发-内容…………')
        sql = '''SELECT * 	
                FROM( SELECT * 
			            FROM gat_trans_way ds
			            LEFT JOIN (SELECT *
								    FROM (SELECT 产品id '产品id2',`规格中文` '规格中文2',包裹重量 '包裹重量2', COUNT(订单编号) 单量, MIN(包裹重量), MAX(包裹重量),  MAX(包裹重量)-MIN(包裹重量) as 重量差
											FROM gat_trans_way d 
											GROUP BY d.`产品id`,d.`规格中文`
											ORDER BY d. 产品id
											) s1
								    WHERE s1.`重量差` > 100 AND s1.`单量` >= 2
								    ) dss ON ds.`产品ID` = dss.`产品id2` and ds.`规格中文`= dss.`规格中文2`
                ) s WHERE s.重量差 IS not null;'''
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df = df[['年月', '团队', '日期', '币种', '订单编号', '数量', '电话号码', '运单编号', '是否改派', '物流方式', '商品id', '产品id', '产品名称',
                 '价格', '仓储扫描时间', '完结状态', '物流花费', '包裹重量', '包裹体积', '规格中文',
                 '产品量', 'MIN(包裹重量)', 'MAX(包裹重量)', '重量差']]
        print('正在写入excel…………')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        df.to_excel('G:\\输出文件\\{} 运费台湾直发-查询{}.xlsx'.format(output, rq),
                    sheet_name='查询', index=False)
        print('----已写入excel')

        sql = '''SELECT 年月,日期,团队,币种,订单编号,数量,电话号码,运单编号,是否改派,物流方式,商品id,ds.产品id,产品名称,价格,下单时间,审核时间,仓储扫描时间,完结状态,完结状态时间,物流花费,包裹重量,包裹体积,规格中文,产品量
                FROM gat_order_list ds
                LEFT JOIN (SELECT 产品id, COUNT(订单编号) 产品量
                            FROM gat_order_list ds
					        WHERE ds.`日期` = '{0}' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
					        GROUP BY ds.`产品id`
				) dds on ds.`产品id` = dds.`产品id`
                WHERE ds.`年月` = '{1}' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND ds.`包裹重量` <> 0 AND ds.`是否改派` = '直发'
			        AND ds.`规格中文` IN (SELECT `规格中文`
                                         FROM (SELECT 产品id,`规格中文`,COUNT(订单编号) 单量, MIN(包裹重量), MAX(包裹重量),  MAX(包裹重量)-MIN(包裹重量) as 重量差
                                              FROM gat_order_list d 
											  WHERE d.`年月` = '{1}' and d.`是否改派` = '直发' AND d.`产品id` <> 0 AND d.`包裹重量` <> 0
											    AND d.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
											  GROUP BY d.`产品id`,d.`规格中文`
											  ORDER BY d. 产品id
										    ) s1
										WHERE s1.`重量差` > 100 AND s1.`单量` >= 2
									   )	
			        AND ds.`产品id` IN (SELECT s.`产品id`
										FROM (SELECT 年月,日期,团队,币种,订单编号,数量,电话号码,运单编号,是否改派,物流方式,商品id,产品id,产品名称,价格,下单时间,审核时间,仓储扫描时间,完结状态,完结状态时间,物流花费,包裹重量,包裹体积,规格中文
											 FROM gat_order_list ds
											 WHERE ds.`年月` = '{1}' AND ds.`产品id` <> 0 AND ds.`包裹重量` <> 0 and ds.`是否改派` = '直发'
											   AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')  AND ds.`币种` = '台湾' AND ds.是否改派 = '直发'
											   AND ds.`规格中文` IN (SELECT `规格中文`
																	 FROM (SELECT 产品id,`规格中文`,COUNT(订单编号) 单量, MIN(包裹重量), MAX(包裹重量),  MAX(包裹重量)-MIN(包裹重量) as 重量差
																		  FROM gat_order_list d 
																		  WHERE d.`年月` = '{1}' and d.`是否改派` = '直发' AND d.`产品id` <> 0 AND d.`包裹重量` <> 0
																			AND d.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
																		  GROUP BY d.`产品id`,d.`规格中文`
																		  ORDER BY d. 产品id
																		 ) s1
																	WHERE s1.`重量差` > 100 AND s1.`单量` >= 2
																   ) ORDER BY 日期
											 ) s
											 GROUP BY  s.`产品id`
											 HAVING count(s.`产品id`) >1
										)
                ORDER BY 日期;'''.format(month_yesterday, month_now)
        print('正在获取 ' + match[team] + ' 运费情况…………')
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('gat_trans_way', con=self.engine1, index=False, if_exists='replace')
        print('正在获取运费内容…………')
        sql = '''SELECT * 	
                FROM( SELECT * 
			            FROM gat_trans_way ds
			            LEFT JOIN (SELECT *
								    FROM (SELECT 产品id '产品id2',`规格中文` '规格中文2',包裹重量 '包裹重量2', COUNT(订单编号) 单量, MIN(包裹重量), MAX(包裹重量),  MAX(包裹重量)-MIN(包裹重量) as 重量差
											FROM gat_trans_way d 
											GROUP BY d.`产品id`,d.`规格中文`
											ORDER BY d. 产品id
											) s1
								    WHERE s1.`重量差` > 100 AND s1.`单量` >= 2
								    ) dss ON ds.`产品ID` = dss.`产品id2` and ds.`规格中文`= dss.`规格中文2`
                ) s WHERE s.重量差 IS not null;'''
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df = df[['年月', '团队',  '日期', '币种', '订单编号', '数量', '电话号码', '运单编号', '是否改派', '物流方式', '商品id', '产品id', '产品名称',
                 '价格', '仓储扫描时间', '完结状态', '物流花费', '包裹重量', '包裹体积', '规格中文',
                 '产品量', 'MIN(包裹重量)', 'MAX(包裹重量)', '重量差']]
        print('正在写入excel…………')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        df.to_excel('G:\\输出文件\\{} 运费-查询{}.xlsx'.format(output, rq),
                    sheet_name='查询', index=False)
        print('----已写入excel')
    def trans_way_cost_new(self, team):
        match = {'gat': '港台', 'slsc': '品牌'}
        output = datetime.datetime.now().strftime('%m.%d')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')

        month_yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        month_now = (datetime.datetime.now()).strftime('%Y%m')
        # month_yesterday = '2021-12-01'
        # month_now = '202211'
        print(month_yesterday)
        print(month_now)
        listT = []  # 查询sql的结果 存放池
        print('正在获取 新增的' + match[team] + ' 运费核实…………')
        sql = '''SELECT *
                FROM( SELECT * 
                        FROM ( SELECT *,yu.订单包裹重量 - yu.`同规格最小包裹重量` as 差量
        				        FROM ( SELECT 年月,日期,团队,币种,订单编号,数量,电话号码,运单编号,是否改派,物流方式,商品id,ds.产品id,产品名称,价格,下单时间,审核时间,仓储扫描时间,完结状态,完结状态时间,物流花费,包裹重量 as 订单包裹重量,包裹体积,ds.规格中文,产品量, 
        				                    单量,重量小 as '同规格最小包裹重量', 重量大 as '同规格最大包裹重量',重量差, 选品人, null 备注
        								FROM (SELECT *
											FROM gat_order_list g
											WHERE g.年月>= '{1}'
								        ) ds
        								LEFT JOIN (SELECT 产品id, COUNT(订单编号) 产品量
        											FROM gat_order_list ds
        											WHERE ds.`日期` = '{0}' AND ds.是否改派 = '直发' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') 
        											GROUP BY ds.`产品id`
        								) dds on ds.`产品id` = dds.`产品id`
        								LEFT JOIN (SELECT 产品id,`规格中文`,COUNT(订单编号) 单量, MIN(包裹重量) as 重量小, MAX(包裹重量) as 重量大,  MAX(包裹重量)-MIN(包裹重量) as 重量差
        											FROM gat_order_list d 
        											WHERE d.`年月` = '{1}' and d.`是否改派` = '直发' AND d.`产品id` <> 0 AND d.`包裹重量` <> 0 AND d.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        											GROUP BY d.`产品id`,d.`规格中文`
        											ORDER BY d. 产品id
        							 ) dds2 on ds.`产品id` = dds2.`产品id` AND ds.`规格中文` = dds2.`规格中文`
        						WHERE ds.`年月` = '{1}' AND ds.`包裹重量` <> 0 AND ds.`是否改派` = '直发' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        						GROUP BY ds.`订单编号`
        				        ) yu
                        ) y
                        WHERE y.单量 >=2 and y.差量 > 100
                    ) s
	            WHERE s.产品id NOT IN ( SELECT  DISTINCT  产品id FROM 运费核实 y)
	            ORDER BY s.选品人, s.产品id, s.规格中文;'''.format(month_yesterday, month_now)
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        df1 = df1[['年月', '团队',  '日期', '币种', '订单编号', '数量', '电话号码', '运单编号', '是否改派', '物流方式', '商品id', '产品id', '产品名称',
                 '价格', '仓储扫描时间', '完结状态', '物流花费', '包裹体积', '规格中文', '产品量', '订单包裹重量',  '同规格最小包裹重量', '差量', '选品人', '备注']]
        df1 = df1.loc[df1["币种"] == "台湾"]
        listT.append(df1)

        # print('正在获取 ' + match[team] + ' 运费总直发情况…………')
        # sql = '''SELECT *
        #                 FROM ( SELECT *,yu.包裹重量 - yu.`MIN(包裹重量)` as 差量
        # 				        FROM ( SELECT 年月,日期,团队,币种,订单编号,数量,电话号码,运单编号,是否改派,物流方式,商品id,ds.产品id,产品名称,价格,下单时间,审核时间,仓储扫描时间,完结状态,完结状态时间,物流花费,包裹重量,包裹体积,ds.规格中文,产品量,
        #                                       单量,重量小 as 'MIN(包裹重量)', 重量大 as 'MAX(包裹重量)',重量差, 选品人
        # 								FROM (SELECT *
		# 									FROM gat_order_list g
		# 									WHERE g.年月 = '{1}'
		# 						        ) ds
        # 								LEFT JOIN (SELECT 产品id, COUNT(订单编号) 产品量
        # 											FROM gat_order_list ds
        # 											WHERE ds.`日期` = '{0}' AND ds.是否改派 = '直发' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        # 											GROUP BY ds.`产品id`
        # 								) dds on ds.`产品id` = dds.`产品id`
        # 								LEFT JOIN (SELECT 产品id,`规格中文`,COUNT(订单编号) 单量, MIN(包裹重量) as 重量小, MAX(包裹重量) as 重量大,  MAX(包裹重量)-MIN(包裹重量) as 重量差
        # 											FROM gat_order_list d
        # 											WHERE d.`年月` = '{1}' and d.`是否改派` = '直发' AND d.`产品id` <> 0 AND d.`包裹重量` <> 0 AND d.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        # 											GROUP BY d.`产品id`,d.`规格中文`
        # 											ORDER BY d. 产品id
        # 							 ) dds2 on ds.`产品id` = dds2.`产品id` AND ds.`规格中文` = dds2.`规格中文`
        # 						WHERE ds.`年月` = '{1}' AND ds.`包裹重量` <> 0 AND ds.`是否改派` = '直发' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        # 						GROUP BY ds.`订单编号`
        # 				        ) yu
        #                 ) y
        #                 WHERE y.单量 >=2 and y.差量 > 100;'''.format(month_yesterday, month_now)
        # df2 = pd.read_sql_query(sql=sql, con=self.engine1)
        # df2 = df2[['年月', '团队',  '日期', '币种', '订单编号', '数量', '电话号码', '运单编号', '是否改派', '物流方式', '商品id', '产品id', '产品名称',
        #          '价格', '仓储扫描时间', '完结状态', '物流花费', '包裹体积', '规格中文', '产品量', '包裹重量',  'MIN(包裹重量)', '差量', '选品人']]
        # listT.append(df2)

        print('正在写入excel…………')
        file_path = 'F:\\神龙签收率\\A运费-核实\\{} 运费差异-查询{}.xlsx'.format(output, rq)
        # sheet_name = ['运费查询', '运费总']
        sheet_name = ['运费查询']
        df0 = pd.DataFrame([])                                      # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)                        # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')       # 初始化写入对象
        book = load_workbook(file_path)                             # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book                                          # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:                             # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('----已写入excel ')

        # sql = '''SELECT *
        #                 FROM ( SELECT *,yu.包裹重量 - yu.`MIN(包裹重量)` as 差量
        # 				        FROM ( SELECT 年月,日期,团队,币种,订单编号,数量,电话号码,运单编号,是否改派,物流方式,商品id,ds.产品id,产品名称,价格,下单时间,审核时间,仓储扫描时间,完结状态,完结状态时间,物流花费,包裹重量,包裹体积,ds.规格中文,产品量,
        #                                       单量,重量小 as 'MIN(包裹重量)', 重量大 as 'MAX(包裹重量)',重量差
        # 								FROM gat_order_list ds
        # 								LEFT JOIN (SELECT 产品id, COUNT(订单编号) 产品量
        # 											FROM gat_order_list ds
        # 											WHERE ds.`日期` = '{0}' AND ds.是否改派 = '直发' AND ds.币种 = '台湾' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        # 											GROUP BY ds.`产品id`
        # 								) dds on ds.`产品id` = dds.`产品id`
        # 								LEFT JOIN (SELECT 产品id,`规格中文`,COUNT(订单编号) 单量, MIN(包裹重量) as 重量小, MAX(包裹重量) as 重量大,  MAX(包裹重量)-MIN(包裹重量) as 重量差
        # 											FROM gat_order_list d
        # 											WHERE d.`年月` = '{1}' and d.`是否改派` = '直发' AND d.币种 = '台湾' AND d.`产品id` <> 0 AND d.`包裹重量` <> 0 AND d.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        # 											GROUP BY d.`产品id`,d.`规格中文`
        # 											ORDER BY d. 产品id
        # 							 ) dds2 on ds.`产品id` = dds2.`产品id` AND ds.`规格中文` = dds2.`规格中文`
        # 						WHERE ds.`年月` = '{1}' AND ds.`包裹重量` <> 0 AND ds.`是否改派` = '直发' AND ds.币种 = '台湾' AND ds.系统订单状态 IN ('已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        # 						GROUP BY ds.`订单编号`
        # 				        ) yu
        #                 ) y
        #                 WHERE y.单量 >=2 and y.差量 > 100;'''.format(month_yesterday, month_now)
        # print('正在获取 ' + match[team] + ' 运费台湾直发-情况…………')
        # df = pd.read_sql_query(sql=sql, con=self.engine1)
        # df = df[['年月', '团队',  '日期', '币种', '订单编号', '数量', '电话号码', '运单编号', '是否改派', '物流方式', '商品id', '产品id', '产品名称',
        #          '价格', '仓储扫描时间', '完结状态', '物流花费', '包裹体积', '规格中文', '产品量', '包裹重量',  'MIN(包裹重量)', '差量']]
        # print('正在写入excel…………')
        # rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        # df.to_excel('G:\\输出文件\\{} 运费台湾直发-查询{}.xlsx'.format(output, rq),
        #             sheet_name='查询', index=False)
        # print('----已写入excel')

    def onrount_online(self, team, login_TmpCode, handle):
        print('正在获取查询 在途-未上线 数据…………')
        rq = datetime.datetime.now().strftime('%Y.%m.%d')
        month_begin = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y-%m-%d')
        sql = '''SELECT 日期,团队,币种,订单编号,
                        IF(物流方式 LIKE "%天马%" AND LENGTH(运单编号) = 20, CONCAT(861, RIGHT(运单编号, 8)), IF((物流方式 LIKE "%速派%" or 物流方式 LIKE "%易速配%" or 物流方式 LIKE "%龟山%") AND (运单编号 LIKE "A%" OR 运单编号 LIKE "B%"), RIGHT(运单编号, LENGTH(运单编号) - 1), UPPER(运单编号))) 运单编号,
                        是否改派,物流方式, 系统订单状态,系统物流状态,物流状态,签收表物流状态, 最终状态,下单时间,审核时间,仓储扫描时间,出货时间,上线时间,状态时间,完结状态时间,在途未上线
                FROM ( SELECT *, IF(最终状态 = '在途',
						    IF(币种 = '香港',IF((出货时间 IS NULL AND 仓储扫描时间 <= DATE_SUB(CURDATE(), INTERVAL 2 DAY)) OR
                                              (出货时间 IS NOT NULL AND 仓储扫描时间 <= DATE_SUB(CURDATE(), INTERVAL 3 DAY)),1,0)
                                          ,IF((出货时间 IS NULL AND 仓储扫描时间 <= DATE_SUB(CURDATE(), INTERVAL 4 DAY)) OR
                                              (出货时间 IS NOT NULL AND 仓储扫描时间 <= DATE_SUB(CURDATE(), INTERVAL 5 DAY)),1,0)),
                            IF(币种 = '香港',IF((出货时间 IS NULL AND 仓储扫描时间 <= DATE_SUB(CURDATE(), INTERVAL 1 DAY)) OR
                                              (出货时间 IS NOT NULL AND 仓储扫描时间 <= DATE_SUB(CURDATE(), INTERVAL 2 DAY)),1,0)
                                          ,IF((出货时间 IS NULL AND 仓储扫描时间 <= DATE_SUB(CURDATE(), INTERVAL 2 DAY)) OR
                                              (出货时间 IS NOT NULL AND 仓储扫描时间 <= DATE_SUB(CURDATE(), INTERVAL 3 DAY)),1,0))
                            ) AS 在途未上线
						FROM (SELECT * FROM gat_zqsb gz WHERE gz.年月>= '{0}' AND gz.最终状态 IN ('在途','未上线')) ss
                ) ss
                WHERE 在途未上线= 1;'''.format(month_begin)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
        df.to_excel('F:\\神龙签收率\\(物   流) 在途未上线-订单催促\\{} 在途未上线(催促).xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
        print('正在获取更新 在途-未上线 数据…………')
        # 获取更新订单的语句
        sql = '''SELECT 订单编号 FROM customer;'''
        data_df = ['orderNumber',  'orderStatus', 'logisticsStatus']
        data_df2 = ['订单编号',  '订单状态', '物流状态']
        # 获取更新表的语句
        sql2 = '''update customer a, cache b
                set a.`系统订单状态`= IF(b.`订单状态` = '', NULL, b.`订单状态`),
                    a.`系统物流状态`= IF(b.`物流状态` = '' or b.`物流状态` = '发货中', NULL, b.`物流状态`)
                where a.`订单编号`=b.`订单编号`;'''.format('gat_waybill_list')
        # 调用更新库 函数
        up = Settings_sso()
        if handle == '手动':
            print('请输入口令Token:  回车确认')
            login_TmpCode = str(input())
        up.updata(sql, sql2, team, data_df, data_df2, login_TmpCode, handle)
        print('更新完成…………')

        print('正在获取写入excel内容…………')
        sql = '''SELECT 订单编号,运单编号, 是否改派,发货时间,物流方式,当前状态,最终状态, NULL 查询状态结果,NULL 配送问题, NULL 状态时间
                FROM (  SELECT c.订单编号,c.运单编号,c.系统订单状态, c.系统物流状态, c.是否改派, g.标准物流状态,g.签收表物流状态, c.仓储扫描时间 AS 发货时间, 物流方式, b.出货时间 AS 新出货时间,
							IF(ISNULL(系统物流状态), IF(ISNULL(g.标准物流状态) OR g.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , 
													IF(物流方式 like '%天马%' and g.签收表物流状态 = '在途','未上线', g.标准物流状态)
                            ), 系统物流状态) AS 当前状态,最终状态
                        FROM customer c
                        LEFT JOIN gat_wl_data b ON c.`运单编号` = b.`运单编号`
                        LEFT JOIN gat_logisitis_match g ON b.物流状态 = g.签收表物流状态
                ) s
                WHERE 最终状态 NOT IN ('拒收','已签收')
                ORDER BY 发货时间;'''.format()
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        print('正在写入excel…………')
        file_pathT = 'G:\\输出文件\\{0} 在途未上线-总表.xlsx'.format(rq)
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_pathT, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_pathT, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_pathT)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        df[(df['最终状态'].str.contains('在途'))].to_excel(excel_writer=writer, sheet_name='在途', index=False)
        df[(df['最终状态'].str.contains('未上线'))].to_excel(excel_writer=writer, sheet_name='未上线', index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()

        # print(df)
        waybill = ['天马&天马', '速派&速派', '龟山|易速配&易速配', '铱熙无敌&协来运', '香港-立邦&立邦', '香港-圆通&圆通']
        # waybill = ['立邦']
        for wy in waybill:
            wy1 = wy.split('&')[0]
            wy2 = wy.split('&')[1]
            db1 = df[(df['物流方式'].str.contains(wy1))]
            file_path = 'G:\\输出文件\\{0}{1} 在途未上线.xlsx'.format(rq, wy2)
            writer2 = pd.ExcelWriter(file_path, engine='openpyxl')
            db1[['订单编号', '运单编号', '是否改派', '发货时间', '当前状态', '查询状态结果', '配送问题', '状态时间']].to_excel(writer2, sheet_name='查询', index=False, startrow=0)
            writer2.save()
            writer2.close()
        print('----已写入excel')


if __name__ == '__main__':
    m = QueryUpdate()
    start: datetime = datetime.datetime.now()
    week: datetime = datetime.datetime.now()
    match1 = {'slgat': '神龙-港台',
              'slgat_hfh': '火凤凰-港台',
              'slgat_hs': '红杉-港台',
              'slgat_js': '金狮-港台',
              'gat': '港台'}
    team = 'gat'
    '''
    # -----------------------------------------------手动查询状态运行（一）----------------------------------------
    # upload = '查询-订单号'
    # m.trans_way_cost(team)  # 同产品下的规格运费查询
    '''
    select = 2
    if int(select) == 1:
            upload = '查询-运单号'
            m.readFormHost(upload)

    elif int(select) == 2:
        m.readFormHost('查询运费')
        m.trans_way_cost_new(team)  # 同产品下的规格运费查询

    elif int(select) == 3:
        if week.isoweekday() == 1 or week.isoweekday() == 3 or week.isoweekday() == 5:
            upload = '查询-运单号'
            m.readFormHost(upload)
        m.readFormHost('查询运费')
        m.trans_way_cost_new(team)  # 同产品下的规格运费查询


    elif int(select) == 4:
        m.readFormHost('查询运费')
        m.trans_way_cost_new(team)  # 同产品下的规格运费查询
        
        # if week.isoweekday() == 2 or week.isoweekday() == 5:
        upload = '查询-运单号'    # 获取在途未上线 催促的
        team = 'gat'
        login_TmpCode = '5071d283e9f03534a7548c8ab2701682'
        handle = '手动0'
        m.onrount_online(team, login_TmpCode, handle)

        

    print('输出耗时：', datetime.datetime.now() - start)
    # win32api.MessageBox(0, "注意:>>>    程序运行结束， 请查看表  ！！！", "提 醒", win32con.MB_OK)
