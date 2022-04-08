import pandas as pd
import os
import datetime
import xlwings
import win32com.client as win32
import requests
import json
import sys
from sso_updata import QueryTwo
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel

from sqlalchemy import create_engine
from settings import Settings
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色

from mysqlControl import MysqlControl
# -*- coding:utf-8 -*-
class QueryUpdate(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
        self.m = MysqlControl()
        # self.sso = QueryTwo()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    # 获取签收表内容---港澳台更新签收总表
    def readFormHost(self, startday, team):
        match = {'拒收核实': r'F:\神龙签收率\(订   单) 拒收原因-核实\(上传)订单客户反馈-核实原因 & 再次克隆下单汇总',
                 '拒收缓存': r'F:\神龙签收率\(订   单) 拒收原因-核实\每日使用',
                 '其他': r'D:\Users\Administrator\Desktop\需要用到的文件\B客服工作表'}
        path = match[team]
        dirs = os.listdir(path=path)

        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if 'xlsx' not in filePath:
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                wb = excel.Workbooks.Open(filePath)
                wb.SaveAs(filePath + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
                wb.Close()  # FileFormat = 56 is for .xls extension
                excel.Application.Quit()
                filePath = filePath + "x"
                print(filePath)
                print('****** 已成功将 xls 转换成 xlsx 格式 ******')
            if dir[:2] != '~$':
                wb_data = None
                if '换货' in dir or '换补' in dir:
                    print(filePath)
                    wb_data = '换货表'
                elif '退货' in dir:
                    print(filePath)
                    wb_data = '退货表'
                elif '工单' in dir:
                    print(filePath)
                    wb_data = '工单收集表'
                elif '台湾系统' in dir or '香港系统' in dir or '问题件+客诉件' in dir or '問題件+客訴件' in dir or 'export' in dir or '理赔订单' in dir:
                    print(filePath)
                    wb_data = '客服电话处理'
                elif '压单反馈' in dir and startday in dir:
                    print(filePath)
                    wb_data = '压单反馈'
                elif '需核实拒收-每日上传' in dir:
                    print(filePath)
                    wb_data = '拒收核实'
                elif '利英' in dir or '慧霞' in dir or '贵敏' in dir or '嘉仪' in dir:
                    print(filePath)
                    wb_data = '需核实拒收_缓存每日'
                if wb_data is None:
                    print('***不符合上传格式，跳过此表！！！')
                    pass
                else:
                    self.wbsheetHost(filePath, wb_data)

                excel = win32.gencache.EnsureDispatch('Excel.Application')
                wb = excel.Workbooks.Open(filePath)
                file_path = os.path.join(path, "~$ " + dir)
                wb.SaveAs(file_path, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
                wb.Close()  # FileFormat = 56 is for .xls extension
                excel.Application.Quit()
                os.remove(filePath)

        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, wb_data):
        fileType = os.path.splitext(filePath)[1]
        fileName = os.path.split(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                if sht.api.Visible == -1:
                    try:
                        team = None
                        db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                        # print(db.columns)
                        if wb_data == '换货表':
                            team = '换货表'
                            db = db[['订单编号', '产品Id', '产品名称', '数量', '反馈方式', '金额', '克隆后金额', '反馈问题类型',
                                    '新订单编号', '登记人', '导入时间', '处理人', '处理时间', '下单时间', '币种', '团队']]
                            db['币种'] = db['币种'].astype(str)
                            db = db[(db['币种'].str.contains('台币|港币'))]
                        elif wb_data == '退货表':
                            team = '退货表'
                            db = db[['订单编号', '产品Id', '产品名称', '数量', '反馈方式', '金额', '反馈问题类型', '退款金额',
                                     '登记人', '导入时间', '处理人', '处理时间', '下单时间', '币种', '团队', '退款类型']]
                            db['币种'] = db['币种'].astype(str)
                            db = db[(db['币种'].str.contains('台币|港币'))]
                        elif wb_data == '工单收集表':
                            team = '工单收集表'
                            db = db[['订单编号', '产品id', '产品名称', '问题类型', '环节问题', '订单金额', '订单状态', '物流状态', '签收时间', '所属团队',
                                    '提交形式', '提交时间', '登记人', '币种', '数量']]
                            db['币种'] = db['币种'].astype(str)
                            db = db[(db['币种'].str.contains('台币|港币'))]

                        elif wb_data == '拒收核实':
                            team = '拒收核实'
                            db = db[['处理日期', '订单编号', '核实原因', '具体原因', '再次克隆下单', '处理人']]
                        elif wb_data == '需核实拒收_缓存每日':
                            team = '需核实拒收_缓存每日'
                            db = db[['订单编号']]

                        elif wb_data == '压单反馈':
                            team = '压单反馈'
                            db = db[['订单编号', '产品ID', '产品名称', '币种', '团队', '状态', '反馈时间', '压单原因', '其他原因', '采购员', '入库时间', '下单时间', '其他原因最后更新时间']]
                            db['币种'] = db['币种'].astype(str)
                            db = db[(db['币种'].str.contains('台币|港币'))]
                        elif wb_data == '客服电话处理':
                            db, team = self.infoSheet(db, sht.name, fileName)
                            if team == '采购异常':
                                db = db[(db['币种'].str.contains('台币|港币'))]
                                db.drop(labels=['币种'], axis=1, inplace=True)
                            print('    导入的数据库表：' + str(team))             # 类型错误:只能连接str(不是“列表”)到str
                        if db is not None and len(db) > 0:
                            if wb_data in ('换货表', '退货表', '工单收集表'):
                                db.to_sql(wb_data, con=self.engine1, index=False, if_exists='replace')
                                print('++++成功导入: ' + wb_data + '表')
                            elif wb_data == '拒收核实':
                                self.jushou_write(db, wb_data)
                            else:
                                print('++++正在导入：' + sht.name + ' 表； 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                                db.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
                                print('++++成功导入缓存表')
                                columns = list(db.columns)
                                columns = ','.join(columns)
                                sql = '''REPLACE INTO {}({}, 记录时间) SELECT *, NOW() 记录时间 FROM customer;'''.format(team, columns)
                                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                                print('++++：' + sht.name + '表--->>>更新成功')
                        else:
                            print('----------数据为空导入失败：' + sht.name + ' 表；')
                    except Exception as e:
                        print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                else:
                    print('----不用导入：' + sht.name)
            wb.close()
        app.quit()

    def infoSheet(self, df, shtName, fileName):
        math = {'系统问题件': {'处理时间': [True, ['处理时间'], []],
                            '订单编号': [True, ['订单编号'], []],
                            '问题原因': [True, ['问题原因', '问題原因', '問題原因', '問题原因'], []],
                            '备注': [True, ['备注', '备註', '備注', '備註'], []]},
                '物流客诉件': {'处理时间': [True, ['处理时间'], []],
                            '物流反馈时间': [True, ['物流反馈时间'], []],
                            '处理人': [True, ['处理人'], []],
                            '订单编号': [True, ['订单编号'], []],
                            '处理方案': [False, ['处理方案'], []],
                            '处理结果': [True, ['处理结果'], []],
                            '客诉原因': [False, ['客诉汇总', '客诉原因'], []]},
                '物流问题件': {'处理时间': [True, ['处理时间'], []],
                            '物流反馈时间': [True, ['物流反馈时间'], []],
                            '处理人': [True, ['处理人'], []],
                            '订单编号': [True, ['订单编号', '客户单号'], []],
                            '处理结果': [True, ['处理结果', '跟踪号'], []],
                            '拒收原因': [False, ['拒收原因'], []]},
                '采购异常': {'订单编号': [True, ['订单编号'], []],
                          '处理结果': [True, ['处理结果'], []],
                          '反馈时间': [True, ['反馈时间'], []],
                          '处理时间': [True, ['处理时间'], []],
                          '取消原因': [True, ['取消原因'], []],
                          '币种': [False, ['币种'], []]},
                '丢件_破损_扣货': {'订单编号': [True, ['订单编号'], []],
                                    '处理结果': [True, ['处理结果'], []],
                                    '具体原因': [True, ['具体原因'], []],
                                    '登记时间': [True, ['登记时间'], []],
                                    '新订单编号': [False, ['新订单编号'], []]}
                }
        if '规格(中文)' in list(df.columns):
            df.drop(labels=['规格(中文)'], axis=1, inplace=True)  # 去掉多余的旬列表
        elif '规格' in list(df.columns):
            df.drop(labels=['规格'], axis=1, inplace=True)
        team = None        # 初始化需导入的数据库表
        if '台湾系统' in fileName:
            team = '系统问题件'
        elif '香港系统' in fileName:
            if '系统' in shtName or '已核实' in shtName or '已删除' in shtName or '已刪除' in shtName:
                team = '系统问题件'
            elif '问题件' in shtName or '問題件' in shtName or '问題件' in shtName:
                team = '物流问题件'
            elif '客诉件' in shtName or '客訴件' in shtName:
                team = '物流客诉件'
        elif '问题件+客诉件' in fileName or '問題件+客訴件' in fileName or '问題件+客訴件' in fileName:
            if '问题件' in shtName or '問題件' in shtName or '问題件' in shtName or '拒收' in shtName:
                team = '物流问题件'
            elif '客诉件' in shtName or '客訴件' in shtName:
                team = '物流客诉件'
        elif 'export' in fileName:
            team = '采购异常'
        elif '理赔订单' in fileName:
            team = '丢件_破损_扣货'
        # 添加需要的列
        if '处理人' not in df:
            df.insert(0, '处理人', "")
        # print(df)
        necessary = 0               # 初始化字段是否是必须的字段计数
        unnecessary = 0             # 初始化字段是否是非必须的字段计数
        needDrop = []
        columns = list(df.columns)
        # print(columns)
        # 保留一个列名，后面要用
        if team is not None:
            for index, column in enumerate(columns):
                # print(column)
                if not column:
                    # 如果列名为空，肯定不是需要的列，起一个名字，标记，后面要删除
                    columns[index] = 'needDrop' + str(index)
                    column = 'needDrop' + str(index)
                for k, v in math[team].items():
                    # 遍历字段匹配字典
                    # print(v)
                    if column in v[1]:
                        # 如果列名完全匹配需要的字段，则，字段重命名为标准字段名
                        columns[index] = k
                        if k in columns[:index]:
                            # 如果这个需要的字段，之前出现过，则添加到需要删除的列表里面
                            tem = k + str(columns.index(k, 0, index))
                            columns[columns.index(k, 0, index)] = tem
                            needDrop.append(tem)
                            if v[0]:
                                necessary -= 1
                        break
                if k != columns[index]:
                    needDrop.append(columns[index])
                else:
                    if v[0]:
                        necessary += 1
                    else:
                        unnecessary += 1
        # print(df)
        # print(columns)
        df['订单编号'] = df['订单编号'].fillna(value='null')
        df = df[~df['订单编号'].isin(['null'])]
        # print(df['订单编号'])
        if necessary >= 4:
            df.columns = columns
            df.drop(labels=needDrop, axis=1, inplace=True)
            # df.dropna(axis=0, subset=['运单编号'], inplace=True)
            # print(df.columns)
            # df['订单编号'] = df['订单编号'].str.replace('-', '')
            # print(df)
            return df, team
        else:
            return None, None  # 注意返回值和需要接收的返回值的对等
    def jushou_write(self, db, wb_data):    # 更新-总表(地区签收率使用)
        columns = list(db)
        columns = ', '.join(columns)
        try:
            print('正在导入缓存表......')
            db.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            print('正在更新总表中......')
            sql = 'REPLACE INTO {}({}, 记录时间) SELECT *, NOW() 记录时间 FROM customer; '.format(wb_data, columns)
            # sql = '''update {0} a, customer b
            #                 set a.`处理日期`= IF(b.`处理日期` = '', NULL, b.`处理日期`),
            #                     a.`核实原因`= IF(b.`核实原因` = '', NULL, b.`核实原因`),
            #                     a.`具体原因`= IF(b.`具体原因` = '', NULL, b.`具体原因`),
            #                     a.`再次克隆下单`= IF(b.`再次克隆下单` = '', NULL, b.`再次克隆下单`),
            #                     a.`处理人`= IF(b.`处理人` = '', NULL, b.`处理人`)
            #         where a.`订单编号`= b.`订单编号`;'''.format(wb_data)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')

    # 写入更新缓存表
    def writeSql(self):
        data_now = (datetime.datetime.now() - relativedelta(months=1)).strftime('%Y%m')
        # data_now = datetime.datetime.now().strftime('%Y%m')
        print(data_now)
        listT = []  # 查询sql的结果 存放池
        print('正在获取 第一部分 信息…………')
        print('正在获取 产品前十（分币种&分家族） 信息…………')
        sql = '''SELECT *
                FROM (
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, '合计' 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾'
                    GROUP BY d.`年月`,d.`币种`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10 ) 
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '神龙家族%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10 ) 
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '火凤凰-港澳台'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10 )
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '金鹏%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10)
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '金狮%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '红杉%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 			
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '火凤凰-港台(繁体)%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '台湾' and d.`团队` LIKE '神龙-运营1组%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, '合计' 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港'
                    GROUP BY d.`年月`,d.`币种`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港' and d.`团队` LIKE '神龙家族%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港' and d.`团队` LIKE '火凤凰-港澳台'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10) 
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港' and d.`团队` LIKE '金狮%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10)
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港' and d.`团队` LIKE '金鹏%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10)
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港' and d.`团队` LIKE '红杉%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10)
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港' and d.`团队` LIKE '火凤凰-港台(繁体)%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10)
                UNION all 
                (SELECT 年月,IF(币种 = '台湾','台币','港币') 币种, 团队,产品id,产品名称,COUNT(订单编号) as 单量,SUM(数量) as 商品数量
                    FROM {1} d
                    WHERE d.`年月` ='{0}' and d.`币种` = '香港' and d.`团队` LIKE '神龙-运营1组%'
                    GROUP BY d.`年月`,d.`币种`,d.`团队`, d.`产品id`
                    ORDER BY 单量 DESC
                    LIMIT 10)				
                ) ss;'''.format(data_now, 'gat_zqsb')
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('product_info', con=self.engine1, index=False, if_exists='replace')

        print('正在获取 上月产品前十（总） 信息…………')
        sql1 = '''SELECT *
                FROM (
                    (SELECT *,concat(ROUND(IF(已处理数量 = 0,NULL,已处理数量)  / 工单数量 * 100,2),'%') as 处理占比
                    FROM (SELECT ss.年月, ss.币种, ss.团队, CONCAT(ss.产品id,'#',ss.产品名称) as  产品信息,
                                IF(ss.商品数量 = 0,NULL,ss.商品数量) as 商品数量,
                                IF(换货数量 = 0,NULL,换货数量) as 换货数量,
                                IF(退货数量 = 0,NULL,退货数量) as 退货数量,
                                IF((IF(换货数量 IS NULL, 0,换货数量) + IF(退货数量 IS NULL, 0,退货数量)) = 0,NULL,(IF(换货数量 IS NULL, 0,换货数量) + IF(退货数量 IS NULL, 0,退货数量))) as 已处理数量,
                                IF(工单数量 = 0,NULL,工单数量) as 工单数量				
                        FROM product_info ss
                    LEFT JOIN
                        (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'合计' 团队,产品id,COUNT(订单编号) 换货单量,SUM(数量) as 换货数量
                            FROM 换货表 th
                            GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 产品id
                            ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 换货单量 DESC
                        ) cx ON ss.年月 = cx.年月 AND ss.币种 = cx.币种 AND ss.团队 = cx.团队 AND ss.产品id = cx.产品id
                    LEFT JOIN
                        (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'合计' 团队,产品id,COUNT(订单编号) 退货单量,SUM(数量) as 退货数量
                            FROM 退货表 th
                            GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 产品id
                            ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 退货单量 DESC
                        ) cx2 ON ss.年月 = cx2.年月 AND ss.币种 = cx2.币种 AND ss.团队 = cx2.团队 AND ss.产品id = cx2.产品id
                    LEFT JOIN
                        (SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,'合计' 所属团队,产品id,COUNT(订单编号) 工单单量,SUM(数量) as 工单数量
                            FROM 工单收集表 th
                            GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 产品id
                            ORDER BY DATE_FORMAT(提交时间,'%Y%m') DESC,币种, 工单单量 DESC
                        ) cx3 ON ss.年月 = cx3.年月 AND ss.币种 = cx3.币种 AND ss.团队 = cx3.所属团队 AND ss.产品id = cx3.产品id
                        WHERE ss.团队 = '合计'
                        GROUP BY ss.年月,ss.币种,ss.团队,ss.产品id
                        ORDER BY ss.年月,ss.币种,工单数量 DESC
                    ) s
                )
                UNION all
                (SELECT *,concat(ROUND(IF(已处理数量 = 0,NULL,已处理数量)  / 工单数量 * 100,2),'%') as 处理占比
                    FROM (SELECT ss.年月, ss.币种, ss.团队,  CONCAT(ss.产品id,'#',ss.产品名称) as  产品信息,
                                IF(ss.商品数量 = 0,NULL,ss.商品数量) as 商品数量,
                                IF(换货数量 = 0,NULL,换货数量) as 换货数量,
                                IF(退货数量 = 0,NULL,退货数量) as 退货数量,
                                IF((IF(换货数量 IS NULL, 0,换货数量) + IF(退货数量 IS NULL, 0,退货数量)) = 0,NULL,(IF(换货数量 IS NULL, 0,换货数量) + IF(退货数量 IS NULL, 0,退货数量))) as 已处理数量,
                                IF(工单数量 = 0,NULL,工单数量) as 工单数量
                        FROM product_info ss
                    LEFT JOIN
                        (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,产品id,COUNT(订单编号) 换货单量,SUM(数量) as 换货数量
                            FROM 换货表 th
                            GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队, 产品id
                            ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 团队 , 换货单量 DESC
                        ) cx ON ss.年月 = cx.年月 AND ss.币种 = cx.币种 AND ss.团队 = cx.团队 AND ss.产品id = cx.产品id
                    LEFT JOIN
                        (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,产品id,COUNT(订单编号) 退货单量,SUM(数量) as 退货数量
                            FROM 退货表 th
                            GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队, 产品id
                            ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 团队 , 退货单量 DESC
                        ) cx2 ON ss.年月 = cx2.年月 AND ss.币种 = cx2.币种 AND ss.团队 = cx2.团队 AND ss.产品id = cx2.产品id
                    LEFT JOIN
                        (SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,所属团队,产品id,COUNT(订单编号) 工单单量,SUM(数量) as 工单数量
                            FROM 工单收集表 th
                            GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 所属团队, 产品id
                            ORDER BY DATE_FORMAT(提交时间,'%Y%m') DESC,币种, 所属团队 , 工单单量 DESC
                        ) cx3 ON ss.年月 = cx3.年月 AND ss.币种 = cx3.币种 AND ss.团队 = cx3.所属团队 AND ss.产品id = cx3.产品id
                        WHERE ss.团队 != '合计'
                        GROUP BY ss.年月,ss.币种,ss.团队,ss.产品id 
                        ORDER BY ss.年月,ss.币种,ss.团队,工单数量 DESC
                    ) s
                )
                ) sx;'''
        df1 = pd.read_sql_query(sql=sql1, con=self.engine1)
        listT.append(df1)
        print('正在获取 上月产品前十（明细） 信息…………')
        sql2 = '''SELECT s.年月, s.币种, s.团队,s.产品id, s.商品数量, s.换货数量, s.退货数量,s.已处理数量,s.工单数量,
                        concat(ROUND(IF(已处理数量 = 0,NULL,已处理数量)  / 工单数量 * 100,2),'%') as 处理占比,
                        IF(下错订单 = 0,NULL,下错订单) 下错订单,concat(ROUND(IF(下错订单 = 0,NULL,下错订单) / 工单数量 * 100,2),'%') as 占比,
                        IF(重复订单 = 0,NULL,重复订单) 重复订单,concat(ROUND(IF(重复订单 = 0,NULL,重复订单) / 工单数量 * 100,2),'%') as 占比,
                        IF(尺寸不合 = 0,NULL,尺寸不合) 尺寸不合,concat(ROUND(IF(尺寸不合 = 0,NULL,尺寸不合) / 工单数量 * 100,2),'%') as 占比,
                        IF(尺码偏大 = 0,NULL,尺码偏大) 尺码偏大,concat(ROUND(IF(尺码偏大 = 0,NULL,尺码偏大) / 工单数量 * 100,2),'%') as 占比,
                        IF(尺码偏小 = 0,NULL,尺码偏小) 尺码偏小,concat(ROUND(IF(尺码偏小 = 0,NULL,尺码偏小) / 工单数量 * 100,2),'%') as 占比,
                        IF(价格较高 = 0,NULL,价格较高) 价格较高,concat(ROUND(IF(价格较高 = 0,NULL,价格较高) / 工单数量 * 100,2),'%') as 占比,
                        IF(产品质量不合格 = 0,NULL,产品质量不合格) 产品质量不合格,concat(ROUND(IF(产品质量不合格 = 0,NULL,产品质量不合格) / 工单数量 * 100,2),'%') as 占比,
                        IF(产品瑕疵 = 0,NULL,产品瑕疵) 产品瑕疵,concat(ROUND(IF(产品瑕疵 = 0,NULL,产品瑕疵) / 工单数量 * 100,2),'%') as 占比,
                        IF(漏发错发 = 0,NULL,漏发错发) 漏发错发,concat(ROUND(IF(漏发错发 = 0,NULL,漏发错发) / 工单数量 * 100,2),'%') as 占比,
                        IF(产品不符合客户预期 = 0,NULL,产品不符合客户预期) 产品不符合客户预期,concat(ROUND(IF(产品不符合客户预期 = 0,NULL,产品不符合客户预期) / 工单数量 * 100,2),'%') as 占比,
                        IF(与网站不符 = 0,NULL,与网站不符) 与网站不符,concat(ROUND(IF(与网站不符 = 0,NULL,与网站不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(中国制造 = 0,NULL,中国制造) 中国制造,concat(ROUND(IF(中国制造 = 0,NULL,中国制造) / 工单数量 * 100,2),'%') as 占比,
                        IF(非正品拒收 = 0,NULL,非正品拒收) 非正品拒收,concat(ROUND(IF(非正品拒收 = 0,NULL,非正品拒收) / 工单数量 * 100,2),'%') as 占比,
                        IF(产品到货无法使用 = 0,NULL,产品到货无法使用) 产品到货无法使用,concat(ROUND(IF(产品到货无法使用 = 0,NULL,产品到货无法使用) / 工单数量 * 100,2),'%') as 占比,
                        IF(客户不会使用该产品 = 0,NULL,客户不会使用该产品) 客户不会使用该产品,concat(ROUND(IF(客户不会使用该产品 = 0,NULL,客户不会使用该产品) / 工单数量 * 100,2),'%') as 占比,
                        IF(客户自身原因 = 0,NULL,客户自身原因) 客户自身原因,concat(ROUND(IF(客户自身原因 = 0,NULL,客户自身原因) / 工单数量 * 100,2),'%') as 占比,
                        IF(没有产品说明书 = 0,NULL,没有产品说明书) 没有产品说明书,concat(ROUND(IF(没有产品说明书 = 0,NULL,没有产品说明书) / 工单数量 * 100,2),'%') as 占比,
                        IF(不喜欢 = 0,NULL,不喜欢) 不喜欢,concat(ROUND(IF(不喜欢 = 0,NULL,不喜欢) / 工单数量 * 100,2),'%') as 占比,
                        IF(无订购 = 0,NULL,无订购) 无订购,concat(ROUND(IF(无订购 = 0,NULL,无订购) / 工单数量 * 100,2),'%') as 占比,
                        IF(无理由拒收退货 = 0,NULL,无理由拒收退货) 无理由拒收退货,concat(ROUND(IF(无理由拒收退货 = 0,NULL,无理由拒收退货) / 工单数量 * 100,2),'%') as 占比,
                        IF(已在其他地方购买 = 0,NULL,已在其他地方购买) 已在其他地方购买,concat(ROUND(IF(已在其他地方购买 = 0,NULL,已在其他地方购买) / 工单数量 * 100,2),'%') as 占比,

                        IF(做工瑕疵 = 0,NULL,做工瑕疵) 做工瑕疵,concat(ROUND(IF(做工瑕疵 = 0,NULL,做工瑕疵) / 工单数量 * 100,2),'%') as 占比,
                        IF(修改规格 = 0,NULL,修改规格) 修改规格,concat(ROUND(IF(修改规格 = 0,NULL,修改规格) / 工单数量 * 100,2),'%') as 占比,
                        IF(取消订单 = 0,NULL,取消订单) 取消订单,concat(ROUND(IF(取消订单 = 0,NULL,取消订单) / 工单数量 * 100,2),'%') as 占比,
                        IF(功效与广告不符 = 0,NULL,功效与广告不符) 功效与广告不符,concat(ROUND(IF(功效与广告不符 = 0,NULL,功效与广告不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(包装商品损坏有污渍 = 0,NULL,包装商品损坏有污渍) 包装商品损坏有污渍,concat(ROUND(IF(包装商品损坏有污渍 = 0,NULL,包装商品损坏有污渍) / 工单数量 * 100,2),'%') as 占比,
                        IF(修改收货信息 = 0,NULL,修改收货信息) 修改收货信息,concat(ROUND(IF(修改收货信息 = 0,NULL,修改收货信息) / 工单数量 * 100,2),'%') as 占比,
                        IF(材质与广告不符 = 0,NULL,材质与广告不符) 材质与广告不符,concat(ROUND(IF(材质与广告不符 = 0,NULL,材质与广告不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(没有效果 = 0,NULL,没有效果) 没有效果,concat(ROUND(IF(没有效果 = 0,NULL,没有效果) / 工单数量 * 100,2),'%') as 占比,
                        IF(已留言取消订单 = 0,NULL,已留言取消订单) 已留言取消订单,concat(ROUND(IF(已留言取消订单 = 0,NULL,已留言取消订单) / 工单数量 * 100,2),'%') as 占比,
                        IF(物流反馈客诉件 = 0,NULL,物流反馈客诉件) 物流反馈客诉件,concat(ROUND(IF(物流反馈客诉件 = 0,NULL,物流反馈客诉件) / 工单数量 * 100,2),'%') as 占比,
                        IF(款式与广告不符 = 0,NULL,款式与广告不符) 款式与广告不符,concat(ROUND(IF(款式与广告不符 = 0,NULL,款式与广告不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(功能与广告不符 = 0,NULL,功能与广告不符) 功能与广告不符,concat(ROUND(IF(功能与广告不符 = 0,NULL,功能与广告不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(无证书 = 0,NULL,无证书) 无证书,concat(ROUND(IF(无证书 = 0,NULL,无证书) / 工单数量 * 100,2),'%') as 占比,
                        IF(商品使用后过敏 = 0,NULL,商品使用后过敏) 商品使用后过敏,concat(ROUND(IF(商品使用后过敏 = 0,NULL,商品使用后过敏) / 工单数量 * 100,2),'%') as 占比,
                        IF(不需要 = 0,NULL,不需要) 不需要,concat(ROUND(IF(不需要 = 0,NULL,不需要) / 工单数量 * 100,2),'%') as 占比,
                        IF(非正品 = 0,NULL,非正品) 非正品,concat(ROUND(IF(非正品 = 0,NULL,非正品) / 工单数量 * 100,2),'%') as 占比,
                        IF(物流时间长 = 0,NULL,物流时间长) 物流时间长,concat(ROUND(IF(物流时间长 = 0,NULL,物流时间长) / 工单数量 * 100,2),'%') as 占比,
                        IF(过期 = 0,NULL,过期) 过期,concat(ROUND(IF(过期 = 0,NULL,过期) / 工单数量 * 100,2),'%') as 占比,
                        IF(产品丢失 = 0,NULL,产品丢失) 产品丢失,concat(ROUND(IF(产品丢失 = 0,NULL,产品丢失) / 工单数量 * 100,2),'%') as 占比,
                        IF(尺寸容量与页面描述不符 = 0,NULL,尺寸容量与页面描述不符) 尺寸容量与页面描述不符,concat(ROUND(IF(尺寸容量与页面描述不符 = 0,NULL,尺寸容量与页面描述不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(其他 = 0,NULL,其他) 其他,concat(ROUND(IF(其他 = 0,NULL,其他) / 工单数量 * 100,2),'%') as 占比
                    FROM (SELECT ss.年月, ss.币种, ss.团队, CONCAT(ss.产品id,'#',ss.产品名称) 产品id, 
                                IF(ss.商品数量 = 0,NULL,ss.商品数量) as 商品数量,
                                IF(换货数量 = 0,NULL,换货数量) as 换货数量,
                                IF(退货数量 = 0,NULL,退货数量) as 退货数量,
                                IF((IF(换货数量 IS NULL, 0,换货数量) + IF(退货数量 IS NULL, 0,退货数量)) = 0,NULL,(IF(换货数量 IS NULL, 0,换货数量) + IF(退货数量 IS NULL, 0,退货数量))) as 已处理数量,
                                IF(工单数量 = 0,NULL,工单数量) as 工单数量,
                                下错订单,重复订单,尺寸不合,尺码偏大,尺码偏小,价格较高,产品质量不合格,产品瑕疵,漏发错发,产品不符合客户预期,与网站不符,中国制造,
                                非正品拒收,产品到货无法使用,客户不会使用该产品,客户自身原因,没有产品说明书,不喜欢,无订购,无理由拒收退货,已在其他地方购买,
                                做工瑕疵,修改规格,取消订单,功效与广告不符,包装商品损坏有污渍,修改收货信息,材质与广告不符,没有效果,已留言取消订单,物流反馈客诉件,
                                款式与广告不符,功能与广告不符,无证书,商品使用后过敏,不需要,非正品,物流时间长,过期,产品丢失,尺寸容量与页面描述不符,其他
                        FROM product_info ss
                    LEFT JOIN
                        (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'合计' 团队,产品id,COUNT(订单编号) 换货单量,SUM(数量) as 换货数量
                            FROM 换货表 th
                            GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 产品id
                            ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 换货单量 DESC
                        ) cx ON ss.年月 = cx.年月 AND ss.币种 = cx.币种 AND ss.团队 = cx.团队 AND ss.产品id = cx.产品id
                    LEFT JOIN
                        (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'合计' 团队,产品id,COUNT(订单编号) 退货单量,SUM(数量) as 退货数量
                            FROM 退货表 th
                            GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 产品id
                            ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 退货单量 DESC
                        ) cx2 ON ss.年月 = cx2.年月 AND ss.币种 = cx2.币种 AND ss.团队 = cx2.团队 AND ss.产品id = cx2.产品id
                    LEFT JOIN
                        (SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,'合计' 所属团队,产品id,COUNT(订单编号) 工单单量,SUM(数量) as 工单数量,
                                        SUM(IF(`问题类型` = '下错订单',数量,0)) AS 下错订单,
                                        SUM(IF(`问题类型` = '重复订单',数量,0)) AS 重复订单,
                                        SUM(IF(`问题类型` = '尺寸不合',数量,0)) AS 尺寸不合,
                                        SUM(IF(`问题类型` = '尺码偏大',数量,0)) AS 尺码偏大,
                                        SUM(IF(`问题类型` = '尺码偏小',数量,0)) AS 尺码偏小,
                                        SUM(IF(`问题类型` = '价格较高',数量,0)) AS 价格较高,
                                        SUM(IF(`问题类型` = '产品质量不合格',数量,0)) AS 产品质量不合格,
                                        SUM(IF(`问题类型` = '产品瑕疵',数量,0)) AS 产品瑕疵,
                                        SUM(IF(`问题类型` = '漏发错发',数量,0)) AS 漏发错发,
                                        SUM(IF(`问题类型` = '产品不符合客户预期',数量,0)) AS 产品不符合客户预期,
                                        SUM(IF(`问题类型` = '与网站不符',数量,0)) AS 与网站不符,
                                        SUM(IF(`问题类型` = '中国制造',数量,0)) AS 中国制造,
                                        SUM(IF(`问题类型` = '非正品拒收',数量,0)) AS 非正品拒收,
                                        SUM(IF(`问题类型` = '产品到货无法使用',数量,0)) AS 产品到货无法使用,
                                        SUM(IF(`问题类型` = '客户不会使用该产品',数量,0)) AS 客户不会使用该产品,
                                        SUM(IF(`问题类型` = '客户自身原因',数量,0)) AS 客户自身原因,
                                        SUM(IF(`问题类型` = '没有产品说明书',数量,0)) AS 没有产品说明书,
                                        SUM(IF(`问题类型` = '不喜欢',数量,0)) AS 不喜欢,
                                        SUM(IF(`问题类型` = '无订购',数量,0)) AS 无订购,
                                        SUM(IF(`问题类型` = '无理由拒收退货',数量,0)) AS 无理由拒收退货,
                                        SUM(IF(`问题类型` = '已在其他地方购买',数量,0)) AS 已在其他地方购买,
                                        SUM(IF(`问题类型` = '做工瑕疵',数量,0)) AS 做工瑕疵,
                                        SUM(IF(`问题类型` = '修改规格',数量,0)) AS 修改规格,
                                        SUM(IF(`问题类型` = '取消订单',数量,0)) AS 取消订单,
                                        SUM(IF(`问题类型` = '功效与广告不符',数量,0)) AS 功效与广告不符,
                                        SUM(IF(`问题类型` = '包装/商品损坏/有污渍',数量,0)) AS 包装商品损坏有污渍,
                                        SUM(IF(`问题类型` = '修改收货信息',数量,0)) AS 修改收货信息,
                                        SUM(IF(`问题类型` = '材质与广告不符',数量,0)) AS 材质与广告不符,
                                        SUM(IF(`问题类型` = '没有效果',数量,0)) AS 没有效果,
                                        SUM(IF(`问题类型` = '已留言取消订单',数量,0)) AS 已留言取消订单,
                                        SUM(IF(`问题类型` = '物流反馈客诉件',数量,0)) AS 物流反馈客诉件,
                                        SUM(IF(`问题类型` = '款式与广告不符',数量,0)) AS 款式与广告不符,
                                        SUM(IF(`问题类型` = '功能与广告不符',数量,0)) AS 功能与广告不符,
                                        SUM(IF(`问题类型` = '无证书',数量,0)) AS 无证书,
                                        SUM(IF(`问题类型` = '商品使用后过敏',数量,0)) AS 商品使用后过敏,
                                        SUM(IF(`问题类型` = '不需要',数量,0)) AS 不需要,
                                        SUM(IF(`问题类型` = '非正品',数量,0)) AS 非正品,
                                        SUM(IF(`问题类型` = '物流时间长',数量,0)) AS 物流时间长,
                                        SUM(IF(`问题类型` = '过期',数量,0)) AS 过期,
                                        SUM(IF(`问题类型` = '产品丢失',数量,0)) AS 产品丢失,
                                        SUM(IF(`问题类型` = '尺寸/容量与页面描述不符',数量,0)) AS 尺寸容量与页面描述不符,
                                        SUM(IF(`问题类型` = '其他',数量,0)) AS 其他
                            FROM 工单收集表 th
                            GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 产品id
                            ORDER BY DATE_FORMAT(提交时间,'%Y%m') DESC,币种, 工单单量 DESC
                        ) cx3 ON ss.年月 = cx3.年月 AND ss.币种 = cx3.币种 AND ss.团队 = cx3.所属团队 AND ss.产品id = cx3.产品id
                        WHERE ss.团队 = '合计'
                        GROUP BY ss.年月,ss.币种,ss.团队,ss.产品id
                        ORDER BY ss.年月,ss.币种,工单数量 DESC
                    ) s
                        GROUP BY s.年月,s.币种,s.团队,s.产品id;'''
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
        listT.append(df2)
        sql3 = '''SELECT s.年月, s.币种, s.团队,s.产品id, s.商品数量, s.换货数量, s.退货数量,s.已处理数量,s.工单数量,
                        concat(ROUND(IF(已处理数量 = 0,NULL,已处理数量) / 工单数量 * 100,2),'%') as 处理占比,
                        IF(下错订单 = 0,NULL,下错订单) 下错订单,concat(ROUND(IF(下错订单 = 0,NULL,下错订单) / 工单数量 * 100,2),'%') as 占比,
                        IF(重复订单 = 0,NULL,重复订单) 重复订单,concat(ROUND(IF(重复订单 = 0,NULL,重复订单) / 工单数量 * 100,2),'%') as 占比,
                        IF(尺寸不合 = 0,NULL,尺寸不合) 尺寸不合,concat(ROUND(IF(尺寸不合 = 0,NULL,尺寸不合) / 工单数量 * 100,2),'%') as 占比,
                        IF(尺码偏大 = 0,NULL,尺码偏大) 尺码偏大,concat(ROUND(IF(尺码偏大 = 0,NULL,尺码偏大) / 工单数量 * 100,2),'%') as 占比,
                        IF(尺码偏小 = 0,NULL,尺码偏小) 尺码偏小,concat(ROUND(IF(尺码偏小 = 0,NULL,尺码偏小) / 工单数量 * 100,2),'%') as 占比,
                        IF(价格较高 = 0,NULL,价格较高) 价格较高,concat(ROUND(IF(价格较高 = 0,NULL,价格较高) / 工单数量 * 100,2),'%') as 占比,
                        IF(产品质量不合格 = 0,NULL,产品质量不合格) 产品质量不合格,concat(ROUND(IF(产品质量不合格 = 0,NULL,产品质量不合格) / 工单数量 * 100,2),'%') as 占比,
                        IF(产品瑕疵 = 0,NULL,产品瑕疵) 产品瑕疵,concat(ROUND(IF(产品瑕疵 = 0,NULL,产品瑕疵) / 工单数量 * 100,2),'%') as 占比,
                        IF(漏发错发 = 0,NULL,漏发错发) 漏发错发,concat(ROUND(IF(漏发错发 = 0,NULL,漏发错发) / 工单数量 * 100,2),'%') as 占比,
                        IF(产品不符合客户预期 = 0,NULL,产品不符合客户预期) 产品不符合客户预期,concat(ROUND(IF(产品不符合客户预期 = 0,NULL,产品不符合客户预期) / 工单数量 * 100,2),'%') as 占比,
                        IF(与网站不符 = 0,NULL,与网站不符) 与网站不符,concat(ROUND(IF(与网站不符 = 0,NULL,与网站不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(中国制造 = 0,NULL,中国制造) 中国制造,concat(ROUND(IF(中国制造 = 0,NULL,中国制造) / 工单数量 * 100,2),'%') as 占比,
                        IF(非正品拒收 = 0,NULL,非正品拒收) 非正品拒收,concat(ROUND(IF(非正品拒收 = 0,NULL,非正品拒收) / 工单数量 * 100,2),'%') as 占比,
                        IF(产品到货无法使用 = 0,NULL,产品到货无法使用) 产品到货无法使用,concat(ROUND(IF(产品到货无法使用 = 0,NULL,产品到货无法使用) / 工单数量 * 100,2),'%') as 占比,
                        IF(客户不会使用该产品 = 0,NULL,客户不会使用该产品) 客户不会使用该产品,concat(ROUND(IF(客户不会使用该产品 = 0,NULL,客户不会使用该产品) / 工单数量 * 100,2),'%') as 占比,
                        IF(客户自身原因 = 0,NULL,客户自身原因) 客户自身原因,concat(ROUND(IF(客户自身原因 = 0,NULL,客户自身原因) / 工单数量 * 100,2),'%') as 占比,
                        IF(没有产品说明书 = 0,NULL,没有产品说明书) 没有产品说明书,concat(ROUND(IF(没有产品说明书 = 0,NULL,没有产品说明书) / 工单数量 * 100,2),'%') as 占比,
                        IF(不喜欢 = 0,NULL,不喜欢) 不喜欢,concat(ROUND(IF(不喜欢 = 0,NULL,不喜欢) / 工单数量 * 100,2),'%') as 占比,
                        IF(无订购 = 0,NULL,无订购) 无订购,concat(ROUND(IF(无订购 = 0,NULL,无订购) / 工单数量 * 100,2),'%') as 占比,
                        IF(无理由拒收退货 = 0,NULL,无理由拒收退货) 无理由拒收退货,concat(ROUND(IF(无理由拒收退货 = 0,NULL,无理由拒收退货) / 工单数量 * 100,2),'%') as 占比,
                        IF(已在其他地方购买 = 0,NULL,已在其他地方购买) 已在其他地方购买,concat(ROUND(IF(已在其他地方购买 = 0,NULL,已在其他地方购买) / 工单数量 * 100,2),'%') as 占比,

                        IF(做工瑕疵 = 0,NULL,做工瑕疵) 做工瑕疵,concat(ROUND(IF(做工瑕疵 = 0,NULL,做工瑕疵) / 工单数量 * 100,2),'%') as 占比,
                        IF(修改规格 = 0,NULL,修改规格) 修改规格,concat(ROUND(IF(修改规格 = 0,NULL,修改规格) / 工单数量 * 100,2),'%') as 占比,
                        IF(取消订单 = 0,NULL,取消订单) 取消订单,concat(ROUND(IF(取消订单 = 0,NULL,取消订单) / 工单数量 * 100,2),'%') as 占比,
                        IF(功效与广告不符 = 0,NULL,功效与广告不符) 功效与广告不符,concat(ROUND(IF(功效与广告不符 = 0,NULL,功效与广告不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(包装商品损坏有污渍 = 0,NULL,包装商品损坏有污渍) 包装商品损坏有污渍,concat(ROUND(IF(包装商品损坏有污渍 = 0,NULL,包装商品损坏有污渍) / 工单数量 * 100,2),'%') as 占比,
                        IF(修改收货信息 = 0,NULL,修改收货信息) 修改收货信息,concat(ROUND(IF(修改收货信息 = 0,NULL,修改收货信息) / 工单数量 * 100,2),'%') as 占比,
                        IF(材质与广告不符 = 0,NULL,材质与广告不符) 材质与广告不符,concat(ROUND(IF(材质与广告不符 = 0,NULL,材质与广告不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(没有效果 = 0,NULL,没有效果) 没有效果,concat(ROUND(IF(没有效果 = 0,NULL,没有效果) / 工单数量 * 100,2),'%') as 占比,
                        IF(已留言取消订单 = 0,NULL,已留言取消订单) 已留言取消订单,concat(ROUND(IF(已留言取消订单 = 0,NULL,已留言取消订单) / 工单数量 * 100,2),'%') as 占比,
                        IF(物流反馈客诉件 = 0,NULL,物流反馈客诉件) 物流反馈客诉件,concat(ROUND(IF(物流反馈客诉件 = 0,NULL,物流反馈客诉件) / 工单数量 * 100,2),'%') as 占比,
                        IF(款式与广告不符 = 0,NULL,款式与广告不符) 款式与广告不符,concat(ROUND(IF(款式与广告不符 = 0,NULL,款式与广告不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(功能与广告不符 = 0,NULL,功能与广告不符) 功能与广告不符,concat(ROUND(IF(功能与广告不符 = 0,NULL,功能与广告不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(无证书 = 0,NULL,无证书) 无证书,concat(ROUND(IF(无证书 = 0,NULL,无证书) / 工单数量 * 100,2),'%') as 占比,
                        IF(商品使用后过敏 = 0,NULL,商品使用后过敏) 商品使用后过敏,concat(ROUND(IF(商品使用后过敏 = 0,NULL,商品使用后过敏) / 工单数量 * 100,2),'%') as 占比,
                        IF(不需要 = 0,NULL,不需要) 不需要,concat(ROUND(IF(不需要 = 0,NULL,不需要) / 工单数量 * 100,2),'%') as 占比,
                        IF(非正品 = 0,NULL,非正品) 非正品,concat(ROUND(IF(非正品 = 0,NULL,非正品) / 工单数量 * 100,2),'%') as 占比,
                        IF(物流时间长 = 0,NULL,物流时间长) 物流时间长,concat(ROUND(IF(物流时间长 = 0,NULL,物流时间长) / 工单数量 * 100,2),'%') as 占比,
                        IF(过期 = 0,NULL,过期) 过期,concat(ROUND(IF(过期 = 0,NULL,过期) / 工单数量 * 100,2),'%') as 占比,
                        IF(产品丢失 = 0,NULL,产品丢失) 产品丢失,concat(ROUND(IF(产品丢失 = 0,NULL,产品丢失) / 工单数量 * 100,2),'%') as 占比,
                        IF(尺寸容量与页面描述不符 = 0,NULL,尺寸容量与页面描述不符) 尺寸容量与页面描述不符,concat(ROUND(IF(尺寸容量与页面描述不符 = 0,NULL,尺寸容量与页面描述不符) / 工单数量 * 100,2),'%') as 占比,
                        IF(其他 = 0,NULL,其他) 其他,concat(ROUND(IF(其他 = 0,NULL,其他) / 工单数量 * 100,2),'%') as 占比
                    FROM (SELECT ss.年月, ss.币种, ss.团队, CONCAT(ss.产品id,'#',ss.产品名称) 产品id, 
                                IF(ss.商品数量 = 0,NULL,ss.商品数量) as 商品数量,
                                IF(换货数量 = 0,NULL,换货数量) as 换货数量,
                                IF(退货数量 = 0,NULL,退货数量) as 退货数量,
                                IF((IF(换货数量 IS NULL, 0,换货数量) + IF(退货数量 IS NULL, 0,退货数量)) = 0,NULL,(IF(换货数量 IS NULL, 0,换货数量) + IF(退货数量 IS NULL, 0,退货数量))) as 已处理数量,
                                IF(工单数量 = 0,NULL,工单数量) as 工单数量,
                                下错订单,重复订单,尺寸不合,尺码偏大,尺码偏小,价格较高,产品质量不合格,产品瑕疵,漏发错发,产品不符合客户预期,与网站不符,中国制造,
                                非正品拒收,产品到货无法使用,客户不会使用该产品,客户自身原因,没有产品说明书,不喜欢,无订购,无理由拒收退货,已在其他地方购买,
                                做工瑕疵,修改规格,取消订单,功效与广告不符,包装商品损坏有污渍,修改收货信息,材质与广告不符,没有效果,已留言取消订单,物流反馈客诉件,
                                款式与广告不符,功能与广告不符,无证书,商品使用后过敏,不需要,非正品,物流时间长,过期,产品丢失,尺寸容量与页面描述不符,其他
                        FROM product_info ss
                    LEFT JOIN
                        (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,产品id,COUNT(订单编号) 换货单量,SUM(数量) as 换货数量
                            FROM 换货表 th
                            GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队, 产品id
                            ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 团队 , 换货单量 DESC
                        ) cx ON ss.年月 = cx.年月 AND ss.币种 = cx.币种 AND ss.团队 = cx.团队 AND ss.产品id = cx.产品id
                    LEFT JOIN
                        (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,产品id,COUNT(订单编号) 退货单量,SUM(数量) as 退货数量
                            FROM 退货表 th
                            GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队, 产品id
                            ORDER BY DATE_FORMAT(导入时间,'%Y%m') DESC,币种, 团队 , 退货单量 DESC
                        ) cx2 ON ss.年月 = cx2.年月 AND ss.币种 = cx2.币种 AND ss.团队 = cx2.团队 AND ss.产品id = cx2.产品id
                    LEFT JOIN
                        (SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,所属团队,产品id,COUNT(订单编号) 工单单量,SUM(数量) as 工单数量,
                                        SUM(IF(`问题类型` = '下错订单',数量,0)) AS 下错订单,
                                        SUM(IF(`问题类型` = '重复订单',数量,0)) AS 重复订单,
                                        SUM(IF(`问题类型` = '尺寸不合',数量,0)) AS 尺寸不合,
                                        SUM(IF(`问题类型` = '尺码偏大',数量,0)) AS 尺码偏大,
                                        SUM(IF(`问题类型` = '尺码偏小',数量,0)) AS 尺码偏小,
                                        SUM(IF(`问题类型` = '价格较高',数量,0)) AS 价格较高,
                                        SUM(IF(`问题类型` = '产品质量不合格',数量,0)) AS 产品质量不合格,
                                        SUM(IF(`问题类型` = '产品瑕疵',数量,0)) AS 产品瑕疵,
                                        SUM(IF(`问题类型` = '漏发错发',数量,0)) AS 漏发错发,
                                        SUM(IF(`问题类型` = '产品不符合客户预期',数量,0)) AS 产品不符合客户预期,
                                        SUM(IF(`问题类型` = '与网站不符',数量,0)) AS 与网站不符,
                                        SUM(IF(`问题类型` = '中国制造',数量,0)) AS 中国制造,
                                        SUM(IF(`问题类型` = '非正品拒收',数量,0)) AS 非正品拒收,
                                        SUM(IF(`问题类型` = '产品到货无法使用',数量,0)) AS 产品到货无法使用,
                                        SUM(IF(`问题类型` = '客户不会使用该产品',数量,0)) AS 客户不会使用该产品,
                                        SUM(IF(`问题类型` = '客户自身原因',数量,0)) AS 客户自身原因,
                                        SUM(IF(`问题类型` = '没有产品说明书',数量,0)) AS 没有产品说明书,
                                        SUM(IF(`问题类型` = '不喜欢',数量,0)) AS 不喜欢,
                                        SUM(IF(`问题类型` = '无订购',数量,0)) AS 无订购,
                                        SUM(IF(`问题类型` = '无理由拒收退货',数量,0)) AS 无理由拒收退货,
                                        SUM(IF(`问题类型` = '已在其他地方购买',数量,0)) AS 已在其他地方购买,
                                        SUM(IF(`问题类型` = '做工瑕疵',数量,0)) AS 做工瑕疵,
                                        SUM(IF(`问题类型` = '修改规格',数量,0)) AS 修改规格,
                                        SUM(IF(`问题类型` = '取消订单',数量,0)) AS 取消订单,
                                        SUM(IF(`问题类型` = '功效与广告不符',数量,0)) AS 功效与广告不符,
                                        SUM(IF(`问题类型` = '包装/商品损坏/有污渍',数量,0)) AS 包装商品损坏有污渍,
                                        SUM(IF(`问题类型` = '修改收货信息',数量,0)) AS 修改收货信息,
                                        SUM(IF(`问题类型` = '材质与广告不符',数量,0)) AS 材质与广告不符,
                                        SUM(IF(`问题类型` = '没有效果',数量,0)) AS 没有效果,
                                        SUM(IF(`问题类型` = '已留言取消订单',数量,0)) AS 已留言取消订单,
                                        SUM(IF(`问题类型` = '物流反馈客诉件',数量,0)) AS 物流反馈客诉件,
                                        SUM(IF(`问题类型` = '款式与广告不符',数量,0)) AS 款式与广告不符,
                                        SUM(IF(`问题类型` = '功能与广告不符',数量,0)) AS 功能与广告不符,
                                        SUM(IF(`问题类型` = '无证书',数量,0)) AS 无证书,
                                        SUM(IF(`问题类型` = '商品使用后过敏',数量,0)) AS 商品使用后过敏,
                                        SUM(IF(`问题类型` = '不需要',数量,0)) AS 不需要,
                                        SUM(IF(`问题类型` = '非正品',数量,0)) AS 非正品,
                                        SUM(IF(`问题类型` = '物流时间长',数量,0)) AS 物流时间长,
                                        SUM(IF(`问题类型` = '过期',数量,0)) AS 过期,
                                        SUM(IF(`问题类型` = '产品丢失',数量,0)) AS 产品丢失,
                                        SUM(IF(`问题类型` = '尺寸/容量与页面描述不符',数量,0)) AS 尺寸容量与页面描述不符,
                                        SUM(IF(`问题类型` = '其他',数量,0)) AS 其他
                            FROM 工单收集表 th
                            GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 所属团队, 产品id
                            ORDER BY DATE_FORMAT(提交时间,'%Y%m') DESC,币种, 所属团队 , 工单单量 DESC
                        ) cx3 ON ss.年月 = cx3.年月 AND ss.币种 = cx3.币种 AND ss.团队 = cx3.所属团队 AND ss.产品id = cx3.产品id
                        WHERE ss.团队 != '合计'
                        GROUP BY ss.年月,ss.币种,ss.团队,ss.产品id 
                        ORDER BY ss.年月,ss.币种,ss.团队,工单数量 DESC
                    ) s
                    GROUP BY s.年月,s.币种,s.团队,s.产品id;'''
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)

        print('正在获取 第二部分 信息…………')
        print('正在获取 总体单量 信息…………')
        sql5 = '''SELECT s1.`年月`,s1.`币种`,s1.`团队`,
                        IF(s1.`0%单量` = 0,NULL,s1.`0%单量`) '0%单量',
                        IF(s1.`<10%单量` = 0,NULL,s1.`<10%单量`) '<10%单量',
                        IF(s1.`<20%单量` = 0,NULL,s1.`<20%单量`) '<20%单量',
                        IF(s1.`<30%单量` = 0,NULL,s1.`<30%单量`) '<30%单量',
                        IF(s1.`<40%单量` = 0,NULL,s1.`<40%单量`) '<40%单量',
                        IF(s1.`<50%单量` = 0,NULL,s1.`<50%单量`) '<50%单量',
                        IF(s1.`>=50%单量` = 0,NULL,s1.`>=50%单量`) '>=50%单量',
                        IF(s1.`非换补单量` = 0,NULL,s1.`非换补单量`) '非换补单量',
                        IF(s1.`换货单量` = 0,NULL,s1.`换货单量`) '换货单量',
                        IF(s2.`0%单量` = 0,NULL,s2.`0%单量`) '0%单量',
                        IF(s2.`<10%单量` = 0,NULL,s2.`<10%单量`) '<10%单量',
                        IF(s2.`<20%单量` = 0,NULL,s2.`<20%单量`) '<20%单量',
                        IF(s2.`<30%单量` = 0,NULL,s2.`<30%单量`) '<30%单量',
                        IF(s2.`<40%单量` = 0,NULL,s2.`<40%单量`) '<40%单量',
                        IF(s2.`<50%单量` = 0,NULL,s2.`<50%单量`) '<50%单量',
                        IF(s2.`>=50%单量` = 0,NULL,s2.`>=50%单量`) '>=50%单量',
                        IF(s2.`不全款单量` = 0,NULL,s2.`不全款单量`) '不全款单量',
                        IF(s2.`退货单量` = 0,NULL,s2.`退货单量`) '换货单量',
                        s1.`换货单量` + s2.退货单量 as 退换补单量,工单单量
                FROM (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'总体' 团队,
                            SUM(IF(`占比` = '0%',1,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',1,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',1,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',1,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',1,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',1,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',1,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',1,0)) AS '非换补单量',
                            COUNT(订单编号) 换货单量
                        FROM ( SELECT *,IF(克隆后金额/金额 = 0 OR 克隆后金额/金额 IS null,'0%',
                                        IF(克隆后金额/金额 > 0 AND 克隆后金额/金额 <= 0.1,'<10%',
                                        IF(克隆后金额/金额 > 0.1 AND 克隆后金额/金额 <= 0.2,'<20%',
                                        IF(克隆后金额/金额 > 0.2 AND 克隆后金额/金额 <= 0.3,'<30%',
                                        IF(克隆后金额/金额 > 0.3 AND 克隆后金额/金额 <= 0.4,'<40%',
                                        IF(克隆后金额/金额 > 0.4 AND 克隆后金额/金额 <= 0.5,'<50%',
                                        IF(克隆后金额/金额 > 0.5,'>=50%',克隆后金额/金额))))))) as 占比
                                FROM 换货表 th WHERE th.`币种` IN ('台币','港币')
                            ) th
                        GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种
                    ) s1
                JOIN
                    (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'总体' 团队,
                        SUM(IF(`占比` = '0%',1,0)) AS '0%单量',
                        SUM(IF(`占比` = '<10%',1,0)) AS '<10%单量',
                        SUM(IF(`占比` = '<20%',1,0)) AS '<20%单量',
                        SUM(IF(`占比` = '<30%',1,0)) AS '<30%单量',
                        SUM(IF(`占比` = '<40%',1,0)) AS '<40%单量',
                        SUM(IF(`占比` = '<50%',1,0)) AS '<50%单量',
                        SUM(IF(`占比` = '>=50%',1,0)) AS '>=50%单量',
                        SUM(IF(`占比` != '0%',1,0)) AS '不全款单量',
                        COUNT(订单编号) 退货单量
                    FROM ( SELECT *,IF(退款金额/金额 = 0 OR 退款金额/金额 IS null,'0%',
                                    IF(退款金额/金额 > 0 AND 退款金额/金额 <= 0.1,'<10%',
                                    IF(退款金额/金额 > 0.1 AND 退款金额/金额 <= 0.2,'<20%',
                                    IF(退款金额/金额 > 0.2 AND 退款金额/金额 <= 0.3,'<30%',
                                    IF(退款金额/金额 > 0.3 AND 退款金额/金额 <= 0.4,'<40%',
                                    IF(退款金额/金额 > 0.4 AND 退款金额/金额 <= 0.5,'<50%',
                                    IF(退款金额/金额 > 0.5,'>=50%',退款金额/金额))))))) as 占比
                            FROM 退货表 th WHERE th.`币种` IN ('台币','港币')
                        ) th
                    GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种
                    ) s2 ON s1.年月 = s2.年月 AND s1.币种 = s2.币种 AND s1.团队 = s2.团队
                JOIN
                    (SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,'总体' 所属团队,COUNT(订单编号) 工单单量
                        FROM 工单收集表
                        GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种
                    ) s3 ON s1.年月 = s3.年月 AND s1.币种 = s3.币种 AND s1.团队 = s3.所属团队;'''
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)
        print('正在获取 总体克隆金额 信息…………')
        sql6 = '''SELECT s1.`年月`,s1.`币种`,s1.`团队`,
                        IF(s1.`0%单量` = 0,NULL,s1.`0%单量`) '0%单量',
                        IF(s1.`<10%单量` = 0,NULL,s1.`<10%单量`) '<10%单量',
                        IF(s1.`<20%单量` = 0,NULL,s1.`<20%单量`) '<20%单量',
                        IF(s1.`<30%单量` = 0,NULL,s1.`<30%单量`) '<30%单量',
                        IF(s1.`<40%单量` = 0,NULL,s1.`<40%单量`) '<40%单量',
                        IF(s1.`<50%单量` = 0,NULL,s1.`<50%单量`) '<50%单量',
                        IF(s1.`>=50%单量` = 0,NULL,s1.`>=50%单量`) '>=50%单量',
                        IF(s1.`非换补单量` = 0,NULL,s1.`非换补单量`) '非换补单量',
                        IF(s1.`换货单量` = 0,NULL,s1.`换货单量`) '换货单量',
                        IF(s2.`0%单量` = 0,NULL,s2.`0%单量`) '0%单量',
                        IF(s2.`<10%单量` = 0,NULL,s2.`<10%单量`) '<10%单量',
                        IF(s2.`<20%单量` = 0,NULL,s2.`<20%单量`) '<20%单量',
                        IF(s2.`<30%单量` = 0,NULL,s2.`<30%单量`) '<30%单量',
                        IF(s2.`<40%单量` = 0,NULL,s2.`<40%单量`) '<40%单量',
                        IF(s2.`<50%单量` = 0,NULL,s2.`<50%单量`) '<50%单量',
                        IF(s2.`>=50%单量` = 0,NULL,s2.`>=50%单量`) '>=50%单量',
                        IF(s2.`不全款单量` = 0,NULL,s2.`不全款单量`) '不全款单量',
                        IF(s2.`退货单量` = 0,NULL,s2.`退货单量`) '换货单量',
                        s1.`换货单量` + s2.退货单量 as 退换补单量, 
                        s1.`换货单量` + s2.退货单量 + 挽回金额 as 工单单量
                FROM (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'总体' 团队,
                            SUM(IF(`占比` = '0%',克隆后金额,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',克隆后金额,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',克隆后金额,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',克隆后金额,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',克隆后金额,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',克隆后金额,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',克隆后金额,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',克隆后金额,0)) AS '非换补单量',
                            SUM(克隆后金额) 换货单量
                        FROM ( SELECT *,IF(克隆后金额/金额 = 0 OR 克隆后金额/金额 IS null,'0%',
                                        IF(克隆后金额/金额 > 0 AND 克隆后金额/金额 <= 0.1,'<10%',
                                        IF(克隆后金额/金额 > 0.1 AND 克隆后金额/金额 <= 0.2,'<20%',
                                        IF(克隆后金额/金额 > 0.2 AND 克隆后金额/金额 <= 0.3,'<30%',
                                        IF(克隆后金额/金额 > 0.3 AND 克隆后金额/金额 <= 0.4,'<40%',
                                        IF(克隆后金额/金额 > 0.4 AND 克隆后金额/金额 <= 0.5,'<50%',
                                        IF(克隆后金额/金额 > 0.5,'>=50%',克隆后金额/金额))))))) as 占比
                                FROM 换货表 th WHERE th.`币种` IN ('台币','港币')
                            ) th
                        GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种
                ) s1
                JOIN
                (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'总体' 团队,
                        SUM(IF(`占比` = '0%',退款金额,0)) AS '0%单量',
                        SUM(IF(`占比` = '<10%',退款金额,0)) AS '<10%单量',
                        SUM(IF(`占比` = '<20%',退款金额,0)) AS '<20%单量',
                        SUM(IF(`占比` = '<30%',退款金额,0)) AS '<30%单量',
                        SUM(IF(`占比` = '<40%',退款金额,0)) AS '<40%单量',
                        SUM(IF(`占比` = '<50%',退款金额,0)) AS '<50%单量',
                        SUM(IF(`占比` = '>=50%',退款金额,0)) AS '>=50%单量',
                        SUM(IF(`占比` != '0%',退款金额,0)) AS '不全款单量',
                        SUM(退款金额) 退货单量
                FROM ( SELECT *,IF(退款金额/金额 = 0 OR 退款金额/金额 IS null,'0%',
                                IF(退款金额/金额 > 0 AND 退款金额/金额 <= 0.1,'<10%',
                                IF(退款金额/金额 > 0.1 AND 退款金额/金额 <= 0.2,'<20%',
                                IF(退款金额/金额 > 0.2 AND 退款金额/金额 <= 0.3,'<30%',
                                IF(退款金额/金额 > 0.3 AND 退款金额/金额 <= 0.4,'<40%',
                                IF(退款金额/金额 > 0.4 AND 退款金额/金额 <= 0.5,'<50%',
                                IF(退款金额/金额 > 0.5,'>=50%',退款金额/金额))))))) as 占比
                        FROM 退货表 th WHERE th.`币种` IN ('台币','港币')
                    ) th
                GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种
                ) s2 ON s1.年月 = s2.年月 AND s1.币种 = s2.币种 AND s1.团队 = s2.团队
                JOIN
                ( SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,'总体' 所属团队, SUM(订单金额) 挽回金额
                    FROM 工单收集表
                    GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种
                ) s3 ON s1.年月 = s3.年月 AND s1.币种 = s3.币种 AND s1.团队 = s3.所属团队;'''
        df6 = pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)
        print('正在获取 总体金额 信息…………')
        sql7 = '''SELECT s1.`年月`,s1.`币种`,s1.`团队`,
                        IF(s1.`0%单量` = 0,NULL,s1.`0%单量`) '0%单量',
                        IF(s1.`<10%单量` = 0,NULL,s1.`<10%单量`) '<10%单量',
                        IF(s1.`<20%单量` = 0,NULL,s1.`<20%单量`) '<20%单量',
                        IF(s1.`<30%单量` = 0,NULL,s1.`<30%单量`) '<30%单量',
                        IF(s1.`<40%单量` = 0,NULL,s1.`<40%单量`) '<40%单量',
                        IF(s1.`<50%单量` = 0,NULL,s1.`<50%单量`) '<50%单量',
                        IF(s1.`>=50%单量` = 0,NULL,s1.`>=50%单量`) '>=50%单量',
                        IF(s1.`非换补单量` = 0,NULL,s1.`非换补单量`) '非换补单量',
                        IF(s1.`换货单量` = 0,NULL,s1.`换货单量`) '换货单量',
                        IF(s2.`0%单量` = 0,NULL,s2.`0%单量`) '0%单量',
                        IF(s2.`<10%单量` = 0,NULL,s2.`<10%单量`) '<10%单量',
                        IF(s2.`<20%单量` = 0,NULL,s2.`<20%单量`) '<20%单量',
                        IF(s2.`<30%单量` = 0,NULL,s2.`<30%单量`) '<30%单量',
                        IF(s2.`<40%单量` = 0,NULL,s2.`<40%单量`) '<40%单量',
                        IF(s2.`<50%单量` = 0,NULL,s2.`<50%单量`) '<50%单量',
                        IF(s2.`>=50%单量` = 0,NULL,s2.`>=50%单量`) '>=50%单量',
                        IF(s2.`不全款单量` = 0,NULL,s2.`不全款单量`) '不全款单量',
                        IF(s2.`退货单量` = 0,NULL,s2.`退货单量`) '换货单量',
                        s1.`换货单量` + s2.退货单量 as 退换补单量,
                        工单单量
                FROM (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'总体' 团队,
                            SUM(IF(`占比` = '0%',金额,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',金额,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',金额,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',金额,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',金额,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',金额,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',金额,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',金额,0)) AS '非换补单量',
                            SUM(金额) 换货单量
                        FROM ( SELECT *,IF(克隆后金额/金额 = 0 OR 克隆后金额/金额 IS null,'0%',
                                        IF(克隆后金额/金额 > 0 AND 克隆后金额/金额 <= 0.1,'<10%',
                                        IF(克隆后金额/金额 > 0.1 AND 克隆后金额/金额 <= 0.2,'<20%',
                                        IF(克隆后金额/金额 > 0.2 AND 克隆后金额/金额 <= 0.3,'<30%',
                                        IF(克隆后金额/金额 > 0.3 AND 克隆后金额/金额 <= 0.4,'<40%',
                                        IF(克隆后金额/金额 > 0.4 AND 克隆后金额/金额 <= 0.5,'<50%',
                                        IF(克隆后金额/金额 > 0.5,'>=50%',克隆后金额/金额))))))) as 占比
                                FROM 换货表 th WHERE th.`币种` IN ('台币','港币')
                            ) th
                        GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种
                ) s1
                JOIN
                (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,'总体' 团队,
                        SUM(IF(`占比` = '0%',金额,0)) AS '0%单量',
                        SUM(IF(`占比` = '<10%',金额,0)) AS '<10%单量',
                        SUM(IF(`占比` = '<20%',金额,0)) AS '<20%单量',
                        SUM(IF(`占比` = '<30%',金额,0)) AS '<30%单量',
                        SUM(IF(`占比` = '<40%',金额,0)) AS '<40%单量',
                        SUM(IF(`占比` = '<50%',金额,0)) AS '<50%单量',
                        SUM(IF(`占比` = '>=50%',金额,0)) AS '>=50%单量',
                        SUM(IF(`占比` != '0%',金额,0)) AS '不全款单量',
                        SUM(金额) 退货单量
                FROM ( SELECT *,IF(退款金额/金额 = 0 OR 退款金额/金额 IS null,'0%',
                                IF(退款金额/金额 > 0 AND 退款金额/金额 <= 0.1,'<10%',
                                IF(退款金额/金额 > 0.1 AND 退款金额/金额 <= 0.2,'<20%',
                                IF(退款金额/金额 > 0.2 AND 退款金额/金额 <= 0.3,'<30%',
                                IF(退款金额/金额 > 0.3 AND 退款金额/金额 <= 0.4,'<40%',
                                IF(退款金额/金额 > 0.4 AND 退款金额/金额 <= 0.5,'<50%',
                                IF(退款金额/金额 > 0.5,'>=50%',退款金额/金额))))))) as 占比
                        FROM 退货表 th WHERE th.`币种` IN ('台币','港币')
                    ) th
                GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种
                ) s2 ON s1.年月 = s2.年月 AND s1.币种 = s2.币种 AND s1.团队 = s2.团队
                JOIN
                ( SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,'总体' 所属团队,SUM(订单金额) 工单单量
                    FROM 工单收集表
                    GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种
                ) s3 ON s1.年月 = s3.年月 AND s1.币种 = s3.币种 AND s1.团队 = s3.所属团队;'''
        df7 = pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)

        print('正在获取 第三部分 信息…………')
        print('正在获取 分团队单量 信息…………')
        sql5 = '''SELECT s1.`年月`,s1.`币种`,s1.`所属团队`,
                        IF(s2.`0%单量` = 0,NULL,s2.`0%单量`) '0%单量',
                        IF(s2.`<10%单量` = 0,NULL,s2.`<10%单量`) '<10%单量',
                        IF(s2.`<20%单量` = 0,NULL,s2.`<20%单量`) '<20%单量',
                        IF(s2.`<30%单量` = 0,NULL,s2.`<30%单量`) '<30%单量',
                        IF(s2.`<40%单量` = 0,NULL,s2.`<40%单量`) '<40%单量',
                        IF(s2.`<50%单量` = 0,NULL,s2.`<50%单量`) '<50%单量',
                        IF(s2.`>=50%单量` = 0,NULL,s2.`>=50%单量`) '>=50%单量',
                        IF(s2.`非换补单量` = 0,NULL,s2.`非换补单量`) '非换补单量',
                        IF(s2.`换货单量` = 0,NULL,s2.`换货单量`) '换货单量',	
                        IF(s3.`0%单量` = 0,NULL,s3.`0%单量`) '0%单量',
                        IF(s3.`<10%单量` = 0,NULL,s3.`<10%单量`) '<10%单量',
                        IF(s3.`<20%单量` = 0,NULL,s3.`<20%单量`) '<20%单量',
                        IF(s3.`<30%单量` = 0,NULL,s3.`<30%单量`) '<30%单量',
                        IF(s3.`<40%单量` = 0,NULL,s3.`<40%单量`) '<40%单量',
                        IF(s3.`<50%单量` = 0,NULL,s3.`<50%单量`) '<50%单量',
                        IF(s3.`>=50%单量` = 0,NULL,s3.`>=50%单量`) '>=50%单量',
                        IF(s3.`不全款单量` = 0,NULL,s3.`不全款单量`) '不全款单量',
                        IF(s3.`退货单量` = 0,NULL,s3.`退货单量`) '换货单量',
                        s2.`换货单量` + s3.退货单量 as 退换补单量,工单单量
                FROM (SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,所属团队,COUNT(订单编号) 工单单量
                        FROM 工单收集表
                        GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 所属团队
                ) s1
               LEFT JOIN
                    (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                            SUM(IF(`占比` = '0%',1,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',1,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',1,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',1,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',1,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',1,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',1,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',1,0)) AS '非换补单量',
                            COUNT(订单编号) 换货单量
                        FROM ( SELECT *,IF(克隆后金额/金额 = 0 OR 克隆后金额/金额 IS null,'0%',
                                        IF(克隆后金额/金额 > 0 AND 克隆后金额/金额 <= 0.1,'<10%',
                                        IF(克隆后金额/金额 > 0.1 AND 克隆后金额/金额 <= 0.2,'<20%',
                                        IF(克隆后金额/金额 > 0.2 AND 克隆后金额/金额 <= 0.3,'<30%',
                                        IF(克隆后金额/金额 > 0.3 AND 克隆后金额/金额 <= 0.4,'<40%',
                                        IF(克隆后金额/金额 > 0.4 AND 克隆后金额/金额 <= 0.5,'<50%',
                                        IF(克隆后金额/金额 > 0.5,'>=50%',克隆后金额/金额))))))) as 占比
                                FROM 换货表 th WHERE th.`币种` IN ('台币','港币')
                            ) th
                        GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                ) s2 ON s1.年月 = s2.年月 AND s1.币种 = s2.币种 AND s1.所属团队 = s2.团队
              LEFT  JOIN
                    (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                            SUM(IF(`占比` = '0%',1,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',1,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',1,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',1,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',1,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',1,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',1,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',1,0)) AS '不全款单量',
                            COUNT(订单编号) 退货单量
                        FROM ( SELECT *,IF(退款金额/金额 = 0 OR 退款金额/金额 IS null,'0%',
                                    IF(退款金额/金额 > 0 AND 退款金额/金额 <= 0.1,'<10%',
                                    IF(退款金额/金额 > 0.1 AND 退款金额/金额 <= 0.2,'<20%',
                                    IF(退款金额/金额 > 0.2 AND 退款金额/金额 <= 0.3,'<30%',
                                    IF(退款金额/金额 > 0.3 AND 退款金额/金额 <= 0.4,'<40%',
                                    IF(退款金额/金额 > 0.4 AND 退款金额/金额 <= 0.5,'<50%',
                                    IF(退款金额/金额 > 0.5,'>=50%',退款金额/金额))))))) as 占比
                            FROM 退货表 th WHERE th.`币种` IN ('台币','港币')
                        ) th
                    GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
              )  s3 ON s1.年月 = s3.年月 AND s1.币种 = s3.币种 AND s1.所属团队 = s3.团队;'''
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)
        print('正在获取 分团队克隆金额 信息…………')
        sql6 = '''SELECT s3.`年月`,s3.`币种`,s3.`所属团队`,
                        IF(s1.`0%单量` = 0,NULL,s1.`0%单量`) '0%单量',
                        IF(s1.`<10%单量` = 0,NULL,s1.`<10%单量`) '<10%单量',
                        IF(s1.`<20%单量` = 0,NULL,s1.`<20%单量`) '<20%单量',
                        IF(s1.`<30%单量` = 0,NULL,s1.`<30%单量`) '<30%单量',
                        IF(s1.`<40%单量` = 0,NULL,s1.`<40%单量`) '<40%单量',
                        IF(s1.`<50%单量` = 0,NULL,s1.`<50%单量`) '<50%单量',
                        IF(s1.`>=50%单量` = 0,NULL,s1.`>=50%单量`) '>=50%单量',
                        IF(s1.`非换补单量` = 0,NULL,s1.`非换补单量`) '非换补单量',
                        IF(s1.`换货单量` = 0,NULL,s1.`换货单量`) '换货单量',
                        IF(s2.`0%单量` = 0,NULL,s2.`0%单量`) '0%单量',
                        IF(s2.`<10%单量` = 0,NULL,s2.`<10%单量`) '<10%单量',
                        IF(s2.`<20%单量` = 0,NULL,s2.`<20%单量`) '<20%单量',
                        IF(s2.`<30%单量` = 0,NULL,s2.`<30%单量`) '<30%单量',
                        IF(s2.`<40%单量` = 0,NULL,s2.`<40%单量`) '<40%单量',
                        IF(s2.`<50%单量` = 0,NULL,s2.`<50%单量`) '<50%单量',
                        IF(s2.`>=50%单量` = 0,NULL,s2.`>=50%单量`) '>=50%单量',
                        IF(s2.`不全款单量` = 0,NULL,s2.`不全款单量`) '不全款单量',
                        IF(s2.`退货单量` = 0,NULL,s2.`退货单量`) '换货单量',
                        s1.`换货单量` + s2.退货单量 as 退换补单量,
                        s1.`换货单量` + s2.退货单量 + 挽回金额 as 工单单量
                FROM ( SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,所属团队,SUM(订单金额) 挽回金额
                    FROM 工单收集表
                    GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 所属团队
                ) s3
                LEFT JOIN
                    (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                            SUM(IF(`占比` = '0%',克隆后金额,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',克隆后金额,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',克隆后金额,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',克隆后金额,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',克隆后金额,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',克隆后金额,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',克隆后金额,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',克隆后金额,0)) AS '非换补单量',
                            SUM(克隆后金额) 换货单量
                        FROM ( SELECT *,IF(克隆后金额/金额 = 0 OR 克隆后金额/金额 IS null,'0%',
                                        IF(克隆后金额/金额 > 0 AND 克隆后金额/金额 <= 0.1,'<10%',
                                        IF(克隆后金额/金额 > 0.1 AND 克隆后金额/金额 <= 0.2,'<20%',
                                        IF(克隆后金额/金额 > 0.2 AND 克隆后金额/金额 <= 0.3,'<30%',
                                        IF(克隆后金额/金额 > 0.3 AND 克隆后金额/金额 <= 0.4,'<40%',
                                        IF(克隆后金额/金额 > 0.4 AND 克隆后金额/金额 <= 0.5,'<50%',
                                        IF(克隆后金额/金额 > 0.5,'>=50%',克隆后金额/金额))))))) as 占比
                                FROM 换货表 th WHERE th.`币种` IN ('台币','港币')
                        ) th
                        GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                    ) s1 ON s3.年月 = s1.年月 AND s3.币种 = s1.币种 AND s3.所属团队 = s1.团队
                LEFT JOIN
                    (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                            SUM(IF(`占比` = '0%',退款金额,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',退款金额,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',退款金额,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',退款金额,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',退款金额,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',退款金额,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',退款金额,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',退款金额,0)) AS '不全款单量',
                            SUM(退款金额) 退货单量
                        FROM ( SELECT *,IF(退款金额/金额 = 0 OR 退款金额/金额 IS null,'0%',
                                IF(退款金额/金额 > 0 AND 退款金额/金额 <= 0.1,'<10%',
                                IF(退款金额/金额 > 0.1 AND 退款金额/金额 <= 0.2,'<20%',
                                IF(退款金额/金额 > 0.2 AND 退款金额/金额 <= 0.3,'<30%',
                                IF(退款金额/金额 > 0.3 AND 退款金额/金额 <= 0.4,'<40%',
                                IF(退款金额/金额 > 0.4 AND 退款金额/金额 <= 0.5,'<50%',
                                IF(退款金额/金额 > 0.5,'>=50%',退款金额/金额))))))) as 占比
                        FROM 退货表 th WHERE th.`币种` IN ('台币','港币')
                    ) th
                    GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                ) s2 ON s3.年月 = s2.年月 AND s3.币种 = s2.币种 AND s3.所属团队 = s2.团队;'''
        df6 = pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)
        print('正在获取 分团队金额 信息…………')
        sql7 = '''SELECT s3.`年月`,s3.`币种`,s3.`所属团队`,
                        IF(s1.`0%单量` = 0,NULL,s1.`0%单量`) '0%单量',
                        IF(s1.`<10%单量` = 0,NULL,s1.`<10%单量`) '<10%单量',
                        IF(s1.`<20%单量` = 0,NULL,s1.`<20%单量`) '<20%单量',
                        IF(s1.`<30%单量` = 0,NULL,s1.`<30%单量`) '<30%单量',
                        IF(s1.`<40%单量` = 0,NULL,s1.`<40%单量`) '<40%单量',
                        IF(s1.`<50%单量` = 0,NULL,s1.`<50%单量`) '<50%单量',
                        IF(s1.`>=50%单量` = 0,NULL,s1.`>=50%单量`) '>=50%单量',
                        IF(s1.`非换补单量` = 0,NULL,s1.`非换补单量`) '非换补单量',
                        IF(s1.`换货单量` = 0,NULL,s1.`换货单量`) '换货单量',
                        IF(s2.`0%单量` = 0,NULL,s2.`0%单量`) '0%单量',
                        IF(s2.`<10%单量` = 0,NULL,s2.`<10%单量`) '<10%单量',
                        IF(s2.`<20%单量` = 0,NULL,s2.`<20%单量`) '<20%单量',
                        IF(s2.`<30%单量` = 0,NULL,s2.`<30%单量`) '<30%单量',
                        IF(s2.`<40%单量` = 0,NULL,s2.`<40%单量`) '<40%单量',
                        IF(s2.`<50%单量` = 0,NULL,s2.`<50%单量`) '<50%单量',
                        IF(s2.`>=50%单量` = 0,NULL,s2.`>=50%单量`) '>=50%单量',
                        IF(s2.`不全款单量` = 0,NULL,s2.`不全款单量`) '不全款单量',
                        IF(s2.`退货单量` = 0,NULL,s2.`退货单量`) '换货单量',
                        s1.`换货单量` + s2.退货单量 as 退换补单量,工单单量
                FROM ( SELECT DATE_FORMAT(提交时间,'%Y%m') as 年月,币种,所属团队,SUM(订单金额) 工单单量
                    FROM 工单收集表
                    GROUP BY DATE_FORMAT(提交时间,'%Y%m'),币种, 所属团队
                ) s3
                LEFT JOIN
                    (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                            SUM(IF(`占比` = '0%',金额,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',金额,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',金额,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',金额,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',金额,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',金额,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',金额,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',金额,0)) AS '非换补单量',
                            SUM(金额) 换货单量
                        FROM ( SELECT *,IF(克隆后金额/金额 = 0 OR 克隆后金额/金额 IS null,'0%',
                                        IF(克隆后金额/金额 > 0 AND 克隆后金额/金额 <= 0.1,'<10%',
                                        IF(克隆后金额/金额 > 0.1 AND 克隆后金额/金额 <= 0.2,'<20%',
                                        IF(克隆后金额/金额 > 0.2 AND 克隆后金额/金额 <= 0.3,'<30%',
                                        IF(克隆后金额/金额 > 0.3 AND 克隆后金额/金额 <= 0.4,'<40%',
                                        IF(克隆后金额/金额 > 0.4 AND 克隆后金额/金额 <= 0.5,'<50%',
                                        IF(克隆后金额/金额 > 0.5,'>=50%',克隆后金额/金额))))))) as 占比
                                FROM 换货表 th WHERE th.`币种` IN ('台币','港币')
                            ) th
                        GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                    ) s1 ON s3.年月 = s1.年月 AND s3.币种 = s1.币种 AND s3.所属团队 = s1.团队
               LEFT JOIN
                    (SELECT DATE_FORMAT(导入时间,'%Y%m') as 年月,币种,团队,
                            SUM(IF(`占比` = '0%',金额,0)) AS '0%单量',
                            SUM(IF(`占比` = '<10%',金额,0)) AS '<10%单量',
                            SUM(IF(`占比` = '<20%',金额,0)) AS '<20%单量',
                            SUM(IF(`占比` = '<30%',金额,0)) AS '<30%单量',
                            SUM(IF(`占比` = '<40%',金额,0)) AS '<40%单量',
                            SUM(IF(`占比` = '<50%',金额,0)) AS '<50%单量',
                            SUM(IF(`占比` = '>=50%',金额,0)) AS '>=50%单量',
                            SUM(IF(`占比` != '0%',金额,0)) AS '不全款单量',
                            SUM(金额) 退货单量
                        FROM ( SELECT *,IF(退款金额/金额 = 0 OR 退款金额/金额 IS null,'0%',
                                IF(退款金额/金额 > 0 AND 退款金额/金额 <= 0.1,'<10%',
                                IF(退款金额/金额 > 0.1 AND 退款金额/金额 <= 0.2,'<20%',
                                IF(退款金额/金额 > 0.2 AND 退款金额/金额 <= 0.3,'<30%',
                                IF(退款金额/金额 > 0.3 AND 退款金额/金额 <= 0.4,'<40%',
                                IF(退款金额/金额 > 0.4 AND 退款金额/金额 <= 0.5,'<50%',
                                IF(退款金额/金额 > 0.5,'>=50%',退款金额/金额))))))) as 占比
                        FROM 退货表 th WHERE th.`币种` IN ('台币','港币')
                    ) th
                    GROUP BY DATE_FORMAT(导入时间,'%Y%m'),币种, 团队
                ) s2 ON s3.年月 = s2.年月 AND s3.币种 = s2.币种 AND s3.所属团队 = s2.团队;'''
        df7 = pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)

        print('正在写入excel…………')
        today = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        file_path = 'G:\\输出文件\\\\客服处理汇总 {}.xlsx'.format(today)
        sheet_name = ['总团队', '分团队', '分团队2', '总体单量', '总体克隆金额', '总体金额', '分团队单量', '分团队克隆金额', '分团队金额']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        # try:
        #     print('正在运行表宏…………')
        #     app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
        #     app.display_alerts = False
        #     wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
        #     wbsht1 = app.books.open(file_path)
        #     wbsht.macro('zl_report_day')()
        #     wbsht1.save()
        #     wbsht1.close()
        #     wbsht.close()
        #     app.quit()
        # except Exception as e:
        #     print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')


    # 客服介入订单处理结果报告
    def month_reporrt_data(self):
        data_now = (datetime.datetime.now() - relativedelta(months=1)).strftime('%Y%m')
        listT = []  # 查询sql的结果 存放池
        print('正在获取 采购异常-压单-丢件-扣件 信息…………')
        sql = '''SELECT *
            FROM (
                    (SELECT '采购异常' AS 工作,NULL 日期, 单量,客服介入量,单量-发货量 AS 删除单量,concat(ROUND((单量-发货量) / 客服介入量 * 100,2),'%') AS 删除率, 发货量,签收量,在途量,concat(ROUND(签收量 / 单量 * 100,2),'%') AS 签收率
                    FROM (SELECT '采购异常' AS 采购异常,  COUNT(s.`订单编号`) AS 单量, COUNT(s.`订单编号`) AS 客服介入量,
					            SUM(IF(s.`系统订单状态` IN ('已完成','已退货(销售)','已发货','已退货(物流)'),1,0)) AS 发货量,
					            SUM(IF(s.`系统物流状态` = '已签收',1,0)) AS 签收量,  SUM(IF(s.`系统订单状态` = '已发货',1,0)) AS 在途量
		                FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`
					            FROM (SELECT * 
								    FROM 采购异常 
								    WHERE id IN (SELECT MAX(id) FROM 采购异常 w WHERE w.`处理时间` BETWEEN DATE_ADD(curdate()-day(curdate())+1,interval -1 month) AND last_day(date_add(curdate()-day(curdate())+1,interval -1 month))  
								    GROUP BY 订单编号) 
								    ORDER BY id
						        ) cg
					    LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
		                ) s
                    ) ss
                )
                UNION ALL
                ( SELECT '压单' AS 工作,NULL 日期, 单量,客服介入量,单量-发货量 AS 删除单量,concat(ROUND((单量-发货量)  / 客服介入量 * 100,2),'%') AS 删除率, 发货量,签收量,在途量,concat(ROUND(签收量 / 单量 * 100,2),'%') AS 签收率
                    FROM ( SELECT COUNT(s.`订单编号`) AS 单量, COUNT(s.`订单编号`) AS 客服介入量,
			                    SUM(IF(s.`系统订单状态` IN ('已完成','已退货(销售)','已发货','已退货(物流)'),1,0)) AS 发货量,
			                    SUM(IF(s.`系统物流状态` = '已签收',1,0)) AS 签收量,  SUM(IF(s.`系统订单状态` = '已发货',1,0)) AS 在途量
                        FROM ( SELECT gt.*
			                    FROM (SELECT cg.订单编号
						        FROM 压单反馈 cg
						        WHERE cg.`反馈时间` BETWEEN DATE_ADD(curdate()-day(curdate())+1,interval -1 month) AND last_day(date_add(curdate()-day(curdate())+1,interval -1 month)) 
			                    )  lp
			            LEFT JOIN gat_order_list gt ON lp.`订单编号` = gt.`订单编号`
	                    ) s
                    ) ss
                )
                UNION ALL
                (SELECT ss.处理结果,NULL 日期, 单量,客服介入量,单量-发货量 AS 删除单量,concat(ROUND((单量-发货量)  / 客服介入量 * 100,2),'%') AS 删除率, 发货量,签收量,在途量,concat(ROUND(签收量 / 单量 * 100,2),'%') AS 签收率
                    FROM (SELECT 处理结果, COUNT(lp.`订单编号`) AS 单量, COUNT(lp.`订单编号`) AS 客服介入量
			                FROM 丢件_破损_扣货 lp
			                WHERE lp.`登记时间` BETWEEN DATE_ADD(curdate()-day(curdate())+1,interval -1 month) AND last_day(date_add(curdate()-day(curdate())+1,interval -1 month)) 
			                GROUP BY lp.`处理结果`
                    ) ss
                    LEFT JOIN
                    (SELECT 处理结果, SUM(IF(s.`系统订单状态` IN ('已完成','已退货(销售)','已发货','已退货(物流)'),1,0)) AS 发货量,
			                SUM(IF(s.`系统物流状态` = '已签收',1,0)) AS 签收量, SUM(IF(s.`系统订单状态` = '已发货',1,0)) AS 在途量
                    FROM (SELECT lp.`处理结果`, gt.*
			                FROM 丢件_破损_扣货 lp
			                LEFT JOIN gat_order_list gt ON lp.`新订单编号` = gt.`订单编号` 
			                WHERE lp.`登记时间` BETWEEN DATE_ADD(curdate()-day(curdate())+1,interval -1 month) AND last_day(date_add(curdate()-day(curdate())+1,interval -1 month)) AND lp.`新订单编号` IS NOT NULL
                    ) s
                    GROUP BY s.处理结果
                    ) ss1 ON ss.`处理结果` = ss1.`处理结果`
                )
            ) sx;'''.format(data_now, 'gat_zqsb')
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df1)

        print('正在获取 市场--客诉 信息…………')
        sql2 = '''SELECT '港台' AS '市场--客诉件', NULL 日期, 单量, 挽单量, concat(ROUND(挽单量 / 单量 * 100,2),'%') AS 挽单率, 	挽单发货量, 签收量,	concat(ROUND(签收量 / 挽单量 * 100,2),'%') AS 签收率	
                FROM (SELECT '客诉件' AS 客诉件, COUNT(s.`订单编号`) AS 单量, 
						SUM(IF(s.`处理方案` = '赠品' or s.`处理方案` LIKE '%补发%' or s.`处理方案` LIKE '%换货%',1,0)) AS 挽单量,
						SUM(IF(s.`处理方案` = '赠品' or s.`处理方案` LIKE '%补发%' or s.`处理方案` LIKE '%换货%',价格,0)) AS 挽单金额,
						SUM(IF((s.`处理方案` = '赠品' or s.`处理方案` LIKE '%补发%' or s.`处理方案` LIKE '%换货%') AND s.`系统订单状态` IN ('已完成','已退货(销售)','已发货','已退货(物流)'),1,0)) AS 挽单发货量,
						SUM(IF((s.`处理方案` = '赠品' or s.`处理方案` LIKE '%补发%' or s.`处理方案` LIKE '%换货%') AND s.`系统物流状态` = '已签收',1,0)) AS 签收量
		            FROM (SELECT lp.*,gt.系统订单状态,gt.系统物流状态,gt.`价格`
					        FROM (SELECT *
								    FROM 物流客诉件 
								    WHERE id IN (SELECT MAX(id) FROM 物流客诉件 w WHERE w.`处理时间` BETWEEN DATE_ADD(curdate()-day(curdate())+1,interval -1 month) AND last_day(date_add(curdate()-day(curdate())+1,interval -1 month)) 
								    GROUP BY 订单编号) 
								    ORDER BY id
					        ) lp
					    LEFT JOIN gat_order_list gt ON lp.`订单编号` = gt.`订单编号`
		            ) s
                ) ss;'''.format(data_now, 'gat_zqsb')
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
        listT.append(df2)

        print('正在获取 市场--拒收订单 信息…………')
        sql3 = '''SELECT '港台' AS '市场--拒收订单', NULL 日期,已完结拒收单量 AS 拒收单量,客服联系单量,	concat(ROUND(客服联系单量 / 已完结拒收单量 * 100,2),'%') AS 联系率,	客服挽单量,
                        concat(ROUND(客服挽单量 / 客服有效联系量 * 100,2),'%') AS 挽单率, 挽单发货量,签收量,concat(ROUND(签收量 / 客服挽单量 * 100,2),'%') AS 签收率
                FROM (SELECT '拒收订单' 拒收订单, SUM(IF(最终状态 = "拒收" AND EXTRACT(YEAR_MONTH FROM `完结状态时间`) = DATE_FORMAT( DATE_ADD(curdate(),interval -day(curdate()) day), '%Y%m'),1,0)) as 已完结拒收单量
			            FROM ((SELECT 年月, 订单编号,最终状态,完结状态时间 FROM gat_zqsb g WHERE g.`年月` = DATE_FORMAT( DATE_ADD(curdate(),interval -day(curdate()) day), '%Y%m'))
						        UNION
						        (SELECT 年月, 订单编号,最终状态,完结状态时间 FROM gat_zqsb g WHERE g.`年月` = DATE_FORMAT( CURDATE( ),'%Y%m'))
			            ) lp
		        ) ss
                LEFT JOIN
		        (SELECT '拒收订单' 拒收订单, COUNT(lp.`订单编号`) AS 客服联系单量, SUM(IF(lp.`核实原因` NOT IN ('无人接听','无效号码'),1,0)) AS 客服有效联系量
			        FROM 拒收核实 lp
			        WHERE lp.`处理日期` BETWEEN DATE_ADD(curdate()-day(curdate())+1,interval -1 month)  AND  last_day(date_add(curdate()-day(curdate())+1,interval -1 month))
		        ) ss1  ON ss.`拒收订单` = ss1.`拒收订单`
                LEFT JOIN
		        (SELECT '拒收订单' 拒收订单, COUNT(s.`订单编号`) AS 客服挽单量, SUM(IF(gt.`系统订单状态` IN ('已完成','已退货(销售)','已发货','已退货(物流)'),1,0)) AS 挽单发货量, SUM(IF(gt.`系统物流状态` = '已签收',1,0)) AS 签收量
			        FROM (SELECT *
						    FROM 拒收核实 js
						    WHERE (js.`处理日期` BETWEEN DATE_ADD(curdate()-day(curdate())+1,interval -1 month)  AND  last_day(date_add(curdate()-day(curdate())+1,interval -1 month))) 
							AND (js.`再次克隆下单` IS NOT NULL AND js.`再次克隆下单` <> '')
			        ) s
			        LEFT JOIN gat_order_list gt ON s.`再次克隆下单` = gt.`订单编号` 
		        ) ss2 ON ss.`拒收订单` = ss2.`拒收订单`;'''.format(data_now, 'gat_zqsb')
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)

        print('正在获取 市场--派送问题件 信息…………')
        sql4 = '''SELECT '港台' AS '市场--派送问题件', NULL 日期, 物流反馈问题件总量, 签收, 	concat(ROUND(签收 / 物流反馈问题件总量 * 100,2),'%') AS 签收率, 客服联系单量,签收量,
			            concat(ROUND(签收量 / 客服联系单量 * 100,2),'%') AS 客服联系签收率,	核实后再派单量, 核实后再派单签收量,	
			            concat(ROUND(核实后再派单签收量 / 核实后再派单量 * 100,2),'%') AS 核实后再派单签收率
		        FROM (SELECT '派送问题件' AS 派送问题件,  COUNT(s.`订单编号`) AS 物流反馈问题件总量, 
							SUM(IF(s.`系统物流状态` = '已签收',1,0)) AS 签收,
							SUM(IF(s.`处理结果` NOT like '%无人接听%' AND s.`处理结果` NOT like '%已签收%' AND s.`处理结果` NOT like '%无效号码%',1,0)) AS 客服联系单量,
							SUM(IF(s.`处理结果` NOT like '%无人接听%' AND s.`处理结果` NOT like '%已签收%' AND s.`处理结果` NOT like '%无效号码%' AND s.`系统物流状态` = '已签收',1,0)) AS 签收量,
							SUM(IF(s.`处理结果` like '%送货%' OR s.`处理结果` like '%配送%' OR s.`处理结果` like '%自取%',1,0)) AS 核实后再派单量,
							SUM(IF((s.`处理结果` like '%送货%' OR s.`处理结果` like '%配送%' OR s.`处理结果` like '%自取%') AND s.`系统物流状态` = '已签收',1,0)) AS 核实后再派单签收量
				    FROM (SELECT lp.*,gt.系统订单状态,gt.系统物流状态
							FROM (SELECT * 
										FROM 物流问题件 
										WHERE id IN (SELECT MAX(id) FROM 物流问题件 w WHERE w.`处理时间` BETWEEN DATE_ADD(curdate()-day(curdate())+1,interval -1 month) AND last_day(date_add(curdate()-day(curdate())+1,interval -1 month))
										GROUP BY 订单编号) 
										ORDER BY id
							) lp
					        LEFT JOIN gat_order_list gt ON lp.`订单编号` = gt.`订单编号`
				    ) s
	            ) ss;'''.format(data_now, 'gat_zqsb')
        df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        listT.append(df4)

        print('正在获取 市场--邮件 信息…………')
        df5 = pd.DataFrame([['', '', '', '', '', '']], columns=['市场--邮件', '日期', '月回复量（客服回复的邮件量）', '工单量（当月工单录入总量，不区分售中和售后）', '工单已完成量（已处理完成的工单里）','工单已完成率'])
        listT.append(df5)

        print('正在获取 市场--退换货 信息…………')
        sql6 = '''SELECT '港台' AS '市场--退换货', NULL 日期, 
			            SUM(IF(售后类型 = '换货' AND 占比 =0,1,0)) AS '0元换货单量',
			            SUM(IF(售后类型 = '换货' AND 占比 <>0,1,0)) AS '非0元换货单量',
			            SUM(IF(售后类型 = '补发' AND 占比 =0,1,0)) AS '0元补发单量',
			            SUM(IF(售后类型 = '补发' AND 占比 <>0,1,0)) AS '非0元补发单量',
			            SUM(IF(售后类型 = '退货' AND 占比 >0.5,1,0)) AS '全额退款单量（退款大于50%）',
			            SUM(IF(售后类型 = '退货' AND 占比 <=0.5,1,0)) AS '部分退款量（退款小于50%）'			
                FROM (SELECT *, IF(金额 = 0,1,IF(售后类型 = '退货',退款金额/金额,克隆后金额/金额)) AS 占比
			            FROM 退换货表 g
			            WHERE g.`导入时间` BETWEEN TIMESTAMP(DATE_ADD(curdate()-day(curdate())+1,interval -1 month)) AND TIMESTAMP(DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY))
                ) s;'''.format(data_now, 'gat_zqsb')
        df6 = pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)

        print('正在获取 市场-问题订单 信息…………')
        sql7 = '''SELECT '港台' AS '市场--问题订单', NULL 日期, 问题订单, concat(ROUND(转化有效订单 / 问题订单 * 100,2),'%') AS 转化率, 转化有效订单,
			            concat(ROUND(转化有效订单签收 / 转化有效订单 * 100,2),'%') AS 转化有效订单签收率, 签收订单
                FROM (SELECT 日期,COUNT(订单编号) AS 问题订单,
						    SUM(IF(g.`系统订单状态` NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货'),1,0)) AS 转化有效订单,
						    SUM(IF((g.`系统订单状态` NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货')) AND g.`系统物流状态` = '已签收',1,0)) AS 转化有效订单签收,
						    SUM(IF(g.`系统物流状态` = '已签收',1,0)) AS 签收订单,
						    SUM(IF(g.`系统订单状态` = '已删除',1,0)) AS 已删除
			        FROM gat_order_list g
			        WHERE (g.`日期` BETWEEN DATE_ADD(curdate()-day(curdate())+1,interval -1 month)  AND  last_day(date_add(curdate()-day(curdate())+1,interval -1 month)) ) 
			            AND (g.`问题时间` BETWEEN TIMESTAMP(DATE_ADD(curdate()-day(curdate())+1,interval -1 month))  AND TIMESTAMP(DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY))) 
			            AND g.`问题原因` IS NOT NULL
                ) s;'''.format(data_now, 'gat_zqsb')
        df7 = pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)

        print('正在获取 市场--line 信息…………')
        df8 = pd.DataFrame([['', '', '', '', '', '', '', '', '']], columns=['市场--邮件', '日期', '加入客户量','有效客户量','封锁客户量','对话量','接收的讯息','传送的讯息','发送占接收比'])
        listT.append(df8)

        print('正在获取 市场--zendesk 信息…………')
        df9 = pd.DataFrame([['', '', '', '', '', '', '']], columns=['市场--邮件', '日期', '接待量','总咨询量','处理占比','评价量','满意度'])
        listT.append(df9)

        print('正在获取 总订单-审核率-删单率-问题转化率 信息…………')
        sql10 = '''SELECT NULL 工作数量, 年月,币种 AS 市场, COUNT(订单编号)  as 总订单量,
						SUM(IF(gs.`系统订单状态` <> '已删除',1,0)) AS 有效订单量,
						concat(ROUND(SUM(IF(gs.`系统订单状态` <> '已删除',1,0)) / COUNT(订单编号) * 100,2),'%') as 有效订单转化率,
						SUM(IF(gs.`审单类型` = '是',1,0)) as 自动审单量,
						SUM(IF(gs.`审单类型` = '是' AND gs.`系统订单状态` = '已删除',1,0)) as 自动审核订单删除量,
						SUM(IF(gs.`审单类型` = '否' AND gs.`系统订单状态` = '已删除',1,0)) as 人工审核订单删除量,
						SUM(IF(gs.`系统订单状态` = '已删除',1,0)) as 总删除量,
						
						SUM(IF(gs.`系统物流状态` = '已签收',1,0)) AS 总订单签收量,
						SUM(IF(gs.`系统订单状态` = '已发货',1,0)) AS 总订单在途量,
						SUM(IF(gs.`系统订单状态` IN ('待审核','问题订单','已转采购','未支付','支付失败','截单'),1,0)) AS 总订单未发货量,
						
						SUM(IF(gs.`审单类型` = '是' AND gs.`系统物流状态` = '已签收',1,0)) AS 自动审单签收量,
						SUM(IF(gs.`审单类型` = '是' AND gs.`系统订单状态` = '已发货',1,0)) AS 自动审单在途量,
						SUM(IF(gs.`审单类型` = '是' AND gs.`系统订单状态` IN ('待审核','问题订单','已转采购','未支付','支付失败','截单'),1,0)) AS 自动审单未发货量,
						
						SUM(IF(gs.`审单类型` = '否' AND gs.`系统物流状态` = '已签收',1,0)) AS 人工签收量,
						SUM(IF(gs.`审单类型` = '否' AND gs.`系统订单状态` = '已发货',1,0)) AS 人工在途量,
						SUM(IF(gs.`审单类型` = '否' AND gs.`系统订单状态` IN ('待审核','问题订单','已转采购','未支付','支付失败','截单'),1,0)) AS 人工未发货量
			FROM  gat_order_list gs
			WHERE gs.`年月` >= DATE_FORMAT( DATE_ADD(curdate(),interval -day(curdate()) day), '%Y%m' )
			GROUP BY gs.`年月`,gs.币种
			ORDER BY gs.`年月`,币种;'''.format(data_now, 'gat_zqsb')
        df10 = pd.read_sql_query(sql=sql10, con=self.engine1)
        listT.append(df10)

        print('正在写入excel…………')
        today = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        file_path = 'G:\\输出文件\\\\工作数量 {}.xlsx'.format(today)

        writer2 = pd.ExcelWriter(file_path, engine='openpyxl')
        df1.to_excel(writer2, index=False)                  # 采购
        df2.to_excel(writer2, index=False, startrow=10)     # 客诉
        df3.to_excel(writer2, index=False, startrow=14)     # 拒收
        df4.to_excel(writer2, index=False, startrow=18)     # 派送问题件
        df5.to_excel(writer2, index=False, startrow=22)     # 邮件

        df6.to_excel(writer2, index=False, startrow=26)     # 退换货
        df7.to_excel(writer2, index=False, startrow=30)     # 问题订单
        df8.to_excel(writer2, index=False, startrow=34)     # line
        df9.to_excel(writer2, index=False, startrow=38)     # zendesk
        df10.to_excel(writer2, index=False, startcol=16)    # 审核率
        writer2.save()
        writer2.close()

        wb = load_workbook(file_path)
        sheet = wb.get_sheet_by_name("Sheet1")
        sheet.column_dimensions['A'].width = 15.82
        sheet.column_dimensions['B'].width = 8.38
        sheet.column_dimensions['C'].width = 8.38
        sheet.column_dimensions['D'].width = 10.5
        sheet.column_dimensions['E'].width = 10.5
        sheet.column_dimensions['F'].width = 9.88
        sheet.column_dimensions['J'].width = 9.88
        sheet.column_dimensions['K'].width = 12
        sheet.column_dimensions['P'].width = 8.38
        sheet.column_dimensions['Q'].width = 8.38
        sheet.column_dimensions['R'].width = 8.38
        sheet.column_dimensions['S'].width = 8.38
        sheet.column_dimensions['T'].width = 8.38
        sheet.column_dimensions['U'].width = 8.38
        sheet.column_dimensions['V'].width = 8.38
        sheet.column_dimensions['W'].width = 8.38
        sheet.column_dimensions['X'].width = 8.38
        sheet.column_dimensions['Y'].width = 8.38
        sheet.column_dimensions['Z'].width = 8.38
        sheet.column_dimensions['AA'].width = 8.38
        sheet.column_dimensions['AB'].width = 8.38
        sheet.column_dimensions['AC'].width = 8.38
        sheet.column_dimensions['AD'].width = 8.38
        sheet.column_dimensions['AE'].width = 8.38
        sheet.column_dimensions['AF'].width = 8.38
        sheet.column_dimensions['AG'].width = 8.38
        wb.save(file_path)

        # sheet_name = ['工作数量']
        # df0 = pd.DataFrame([])                                      # 创建空的dataframe数据框
        # df0.to_excel(file_path, index=False)                        # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')       # 初始化写入对象
        # book = load_workbook(file_path)                             # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book                                          # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # for i in range(len(listT)):
        #     listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        # if 'Sheet1' in book.sheetnames:                             # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()

        # try:
        #     print('正在运行表宏…………')
        #     app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
        #     app.display_alerts = False
        #     wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
        #     wbsht1 = app.books.open(file_path)
        #     wbsht.macro('zl_report_day')()
        #     wbsht1.save()
        #     wbsht1.close()
        #     wbsht.close()
        #     app.quit()
        # except Exception as e:
        #     print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')

    # 拒收核实-查询需要的产品id
    def jushou(self):
        print('正在查询需核实订单…………')
        listT = []  # 查询sql的结果 存放池
        sql = '''SELECT *
                FROM (SELECT g.*,c.`家族`,c.`月份`,c.`拒收`,c.`总订单`,c.`退货率`,c.`拒收率`
			            FROM  需核实拒收_每日新增订单 g
			            LEFT JOIN (SELECT *
								 FROM(SELECT IFNULL(s1.家族, '合计') 家族, IFNULL(s1.地区, '合计') 地区, IFNULL(s1.月份, '合计') 月份,
											IFNULL(s1.产品id, '合计') 产品id,
											IFNULL(s1.产品名称, '合计') 产品名称,
									--		IFNULL(s1.父级分类, '合计') 父级分类,
									--		IFNULL(s1.二级分类, '合计') 二级分类,
											SUM(s1.已签收) as 已签收,
											SUM(s1.拒收) as 拒收,
											SUM(s1.已退货) as 已退货,
											SUM(s1.已完成) as 已完成,
						                    SUM(s1.总订单) as 总订单,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						                    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						                    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						                    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率
                                    FROM(SELECT IFNULL(cx.`家族`, '合计') 家族, IFNULL(cx.币种, '合计') 地区, IFNULL(cx.`年月`, '合计') 月份,
						                        IFNULL(cx.产品id, '合计') 产品id,
						                        IFNULL(cx.产品名称, '合计') 产品名称,
						                        IFNULL(cx.父级分类, '合计') 父级分类,
						                        IFNULL(cx.二级分类, '合计') 二级分类,
						                        COUNT(cx.`订单编号`) as 总订单,
						                        SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
						                        SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
						                        SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
						                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成		
		                                FROM (SELECT *,IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-低价%","神龙-低价",IF(cc.团队 LIKE "金鹏%","小虎队",cc.团队)))))) as 家族
                                            FROM gat_zqsb cc 
					                        WHERE cc.年月 >=  DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m') AND cc.`币种` = '台湾' AND cc.`运单编号` is not null
		                                ) cx
                                        GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                    ) s1
                                    GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                    WITH ROLLUP 
                                ) s 
                                HAVING s.月份 != '合计' AND s.产品id != '合计' AND s.`拒收` >= '1'
                                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙-运营1组','红杉','金狮','合计'),
                                FIELD(s.`地区`,'台湾','香港','合计'),
                                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'合计'),
                                FIELD(s.`产品id`,'合计'),
                                s.拒收 DESC
			            ) c ON g.`团队` = c.`家族` AND EXTRACT(YEAR_MONTH FROM g.`下单时间`) = c.`月份` AND g.`产品Id` =c.`产品Id`
                ) s WHERE s.`家族` is not null;'''
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df)
        print('正在查询两月拒收订单…………')
        sql2 = '''SELECT * FROM 需核实拒收_获取最近两个月订单;'''
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
        listT.append(df2)
        print('正在查询两月拒收产品id…………')
        sql3 = '''SELECT *
								 FROM(SELECT IFNULL(s1.家族, '合计') 家族, IFNULL(s1.地区, '合计') 地区, IFNULL(s1.月份, '合计') 月份,
											IFNULL(s1.产品id, '合计') 产品id,
											IFNULL(s1.产品名称, '合计') 产品名称,
										--	IFNULL(s1.父级分类, '合计') 父级分类,
										--	IFNULL(s1.二级分类, '合计') 二级分类,
											SUM(s1.已签收) as 已签收,
											SUM(s1.拒收) as 拒收,
											SUM(s1.已退货) as 已退货,
											SUM(s1.已完成) as 已完成,
						                    SUM(s1.总订单) as 总订单,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						                    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						                    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						                    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率
                                    FROM(SELECT IFNULL(cx.`家族`, '合计') 家族, IFNULL(cx.币种, '合计') 地区, IFNULL(cx.`年月`, '合计') 月份,
						                        IFNULL(cx.产品id, '合计') 产品id,
						                        IFNULL(cx.产品名称, '合计') 产品名称,
						                        IFNULL(cx.父级分类, '合计') 父级分类,
						                        IFNULL(cx.二级分类, '合计') 二级分类,
						                        COUNT(cx.`订单编号`) as 总订单,
						                        SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
						                        SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
						                        SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
						                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成		
		                                FROM (SELECT *,IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-低价%","神龙-低价",IF(cc.团队 LIKE "金鹏%","小虎队",cc.团队)))))) as 家族
                                            FROM gat_zqsb cc 
					                        WHERE cc.年月 >=  DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m') AND cc.`币种` = '台湾' AND cc.`运单编号` is not null
		                                ) cx
                                        GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                    ) s1
                                    GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                    WITH ROLLUP 
                                ) s 
                                HAVING s.月份 != '合计' AND s.产品id != '合计' AND s.`拒收` >= '1'
                                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙-运营1组','红杉','金狮','合计'),
                                FIELD(s.`地区`,'台湾','香港','合计'),
                                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'合计'),
                                FIELD(s.`产品id`,'合计'),
                                s.拒收 DESC;'''
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)
        # print('正在查询需核实拒收_每日新增订单…………')
        # sql4 = '''SELECT * FROM 需核实拒收_每日新增订单;'''
        # df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        # listT.append(df4)
        print('正在写入excel…………')
        today = datetime.date.today().strftime('%m.%d')
        file_path = 'G:\\输出文件\\{} 需核实拒收-每日数据源.xlsx'.format(today)
        if os.path.exists(file_path):  # 判断是否有需要的表格
            print("正在清除重复文件......")
            os.remove(file_path)
        # sheet_name = ['查询', '两月拒收', '两月拒收产品id', '每日新增订单']
        sheet_name = ['查询', '两月拒收', '两月拒收产品id']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('----已写入excel ')

if __name__ == '__main__':
    m = QueryUpdate()
    start: datetime = datetime.datetime.now()
    # -----------------------------------------------手动查询状态运行（一）-----------------------------------------
    # m.readFormHost('202110')                   # 读取需要的工作表内容（工单、退货、换补发； 系统问题件、物流问题件、物流客诉件； 系统采购异常； 压单反馈表）
    # m.writeSql()                               # 获取工单和退换货的客服处理记录
    '''
        1、 上传文件；  读取需要的工作表内容（工单、退货、换补发； 系统问题件、物流问题件、物流客诉件； 系统采购异常； 压单反馈表）
        2、 上传文件-按日期；            
        3、 获取工单和退换货的客服处理记录；
        4、 拒收核实-查询需要的产品id；  获取前 记得上传发过的核实表和返回的核实表；以及客诉件和问题件表
    '''
    select = 5
    if int(select) == 1:
        m.readFormHost('202110', '拒收核实')        # 上传每日核实过的
        m.readFormHost('202110', '拒收缓存')        # 上传核实的表

    elif int(select) == 2:
        m.readFormHost('202110', '其他')

    elif int(select) == 3:
        begin = datetime.date(2021, 12, 1)      # 压单反馈上传使用
        print(begin)
        end = datetime.date(2021, 12, 31)
        print(end)
        for i in range((end - begin).days):     # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            upload = str(day)
            startday = str(day).replace('-', '')
            print(startday)
            m.readFormHost(startday, '其他')

    elif int(select) == 4:
        m.readFormHost('202110', '拒收核实')
        m.readFormHost('202110', '拒收缓存')
        m.readFormHost('202110', '其他')
        m.jushou()

    elif int(select) == 5:
        m.readFormHost('202110', '其他')
        m.writeSql()                    # 月客服介入订单处理结果报告

    elif int(select) == 6:
        m.month_reporrt_data()                    # 月工作数量

    elif int(select) == 7:
        m.month_reporrt_data()                    # 月工作数量

    print('输出耗时：', datetime.datetime.now() - start)



    # print(999999999999999)
    # begin = datetime.date(2021, 11, 1)       # 1、手动设置时间；若无法查询，切换代理和直连的网络
    # print(begin)
    # end = datetime.date(2021, 12, 14)
    # print(end)
    # print(datetime.datetime.now())
    # for i in range((end - begin).days):  # 按天循环获取订单状态
    # # for i in range((end - begin).days / 5):  # 按天循环获取订单状态
    #     print(i)
    #     last_month = begin + datetime.timedelta(days=5 * i)
    #     now_month = begin + datetime.timedelta(days=(i+1) * 5)
    #     if end >= now_month:
    #         print('正在更新 ' + str(last_month) + ' 号 --- ' + str(now_month) + ' 号信息…………')
    #     else:
    #         now_month = last_month + datetime.timedelta(days=(end - last_month).days)
    #         print('正在更新 ' + str(last_month) + ' 号 --- ' + str(now_month) + ' 号信息…………')
    #         break

    # https: // oapi.dingtalk.com / robot / send?access_token = c2c6e41d9e35775b24e2a3867ab977e3b4c14aec14f6f0651866e163390f3a0a