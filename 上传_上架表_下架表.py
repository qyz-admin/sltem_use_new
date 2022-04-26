import pandas as pd
import os
import datetime
import time
import xlwings
import xlsxwriter
import math
import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel
import win32api,win32con
import math
from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Color,Alignment ,PatternFill # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色

# -*- coding:utf-8 -*-
class Updata_return_bill(Settings, Settings_sso):
    def __init__(self, userMobile, password):
        Settings.__init__(self)
        Settings_sso.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue(maxsize=10)  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        # self.sso_online_Two()
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
    def readFormHost(self):
        start = datetime.datetime.now()
        path = r'D:\Users\Administrator\Desktop\需要用到的文件\数据库'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                tem_data = ''
                team = ''
                if '吉客印退仓' in dir:
                    team = 'gat_return_bill'
                    tem_data = '速派'
                elif '吉客印过期退仓' in dir:
                    team = 'gat_return_bill_over'
                    tem_data = '速派'

                elif '吉客印上架总表' in dir:
                    tem_data = '天马'

                elif '吉客印龟山库存总表' in dir:
                    team = 'gat_return_bill'
                    tem_data = '易速配'

                elif 'HSA045' in dir:
                    team = 'gat_return_bill'
                    tem_data = '协来运'


                self.wbsheetHost(filePath, team, tem_data)
                # os.remove(filePath)
                print('已清除上传文件…………')
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, team, tem_data):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    if tem_data == '速派':
                        if team == 'gat_return_bill':   # 上架表
                            db.rename(columns={'订单号': '订单编号', '承运单号': '运单编号'}, inplace=True)
                            db.insert(0, '物流渠道', tem_data)
                            db = db[['物流渠道', '订单编号', '运单编号', '退货单号', '退货上架货架', '上架时间', '仓库名称']]

                        elif team == 'gat_return_bill_over':    # 下架表
                            db.rename(columns={'订单号': '订单编号', '承运单号': '运单编号'}, inplace=True)
                            db.insert(0, '物流渠道', tem_data)
                            db = db[['物流渠道', '订单编号', '运单编号', '退货单号', '退货上架货架', '上架时间', '仓库名称', '在仓天数', '末条状态']]

                    elif tem_data == '易速配':
                        if team == 'gat_return_bill':
                            db.rename(columns={'内部单号': '订单编号', '原单号': '运单编号', '龟山入库单号': '退货单号', '库位': '退货上架货架'}, inplace=True)
                            db.insert(0, '物流渠道', tem_data)
                            db.insert(0, '仓库名称', '龟山')
                            db = db[['物流渠道', '订单编号', '运单编号', '退货单号', '退货上架货架', '上架时间', '仓库名称']]

                    elif tem_data == '协来运':
                        if team == 'gat_return_bill':
                            drop = {'订单编号': [True, ['訂單號'], []],
                                    '运单编号': [True, ['配編'], []],
                                    '退货单号': [True, ['條碼號'], []],
                                    '退货上架货架': [True, ['倉位'], []],
                                    '上架时间': [True, ['入倉日期'], []]
                                    }
                            necessary = 0       # 初始化字段是否是必须的字段计数
                            unnecessary = 0     # 初始化字段是否是非必须的字段计数
                            needDrop = []
                            columns = list(db.columns)              # 保留一个列名，后面要用
                            for index, column in enumerate(columns):
                                if not column:                      # 如果列名为空，肯定不是需要的列，起一个名字，标记，后面要删除
                                    columns[index] = 'needDrop' + str(index)
                                    column = 'needDrop' + str(index)
                                for k, v in drop.items():  # 遍历字段匹配字典
                                    if column in v[1]:                 # 如果列名完全匹配需要的字段，则，字段重命名为标准字段名
                                        columns[index] = k
                                        if k in columns[:index]:    # 如果这个需要的字段，之前出现过，则添加到需要删除的列表里面
                                            tem = k + str(columns.index(k, 0, index))
                                            columns[columns.index(k, 0, index)] = tem
                                            needDrop.append(tem)
                                            if v[0]:
                                                necessary -= 1
                                        break
                                if k != columns[index]:
                                    needDrop.append(columns[index])
                                else:
                                    if v[0]:
                                        necessary += 1
                                    else:
                                        unnecessary += 1
                            print(db.columns)
                            db = db[db.columns].T.drop_duplicates().T   # DataFrame删除重复列
                            print(db.columns)
                            db.rename(columns={'訂單號': '订单编号', '配編': '运单编号', '條碼號': '退货单号', '倉位': '退货上架货架', '入倉日期': '上架时间'}, inplace=True)
                            db.insert(0, '物流渠道', tem_data)
                            db.insert(0, '仓库名称', '协来运')
                            db = db[['物流渠道', '订单编号', '运单编号', '退货单号', '退货上架货架', '上架时间', '仓库名称']]
                    elif team == '':
                        if '上架' in sht:
                            team = 'gat_return_bill'
                            db.rename(columns={'内部单号': '订单编号', '转单号码': '运单编号', '通知日期': '上架时间', '所属仓库': '仓库名称'}, inplace=True)
                            db.insert(0, '退货单号', '')
                            db.insert(0, '退货上架货架', '')
                            db['退货单号'] = db['运单编号'].copy()
                            db = db[['订单编号', '运单编号', '退货单号', '退货上架货架', '上架时间', '仓库名称']]
                        elif '下架' in sht:
                            team = 'gat_return_bill_over'
                            db.rename(columns={'内部单号': '订单编号', '转单号码': '运单编号', '下架日期': '上架时间'}, inplace=True)
                            db.insert(0, '退货单号', '')
                            db.insert(0, '退货上架货架', '')
                            db.insert(0, '仓库名称', '')
                            db['退货单号'] = db['运单编号'].copy()
                            db = db[['订单编号', '运单编号', '退货单号', '退货上架货架', '上架时间', '仓库名称']]
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
                    print('++++正在导入：' + sht.name + ' 表；共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    db.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
                    print('++++成功导入缓存表')
                    sql = '''SELECT * FROM customer c WHERE c.订单编号 IS  NULL OR c.运单编号 IS  NULL OR c.退货单号 IS  NULL;'''.format(team)
                    df = pd.read_sql_query(sql=sql, con=self.engine1)
                    if len(df.index) > 0:
                        df.to_excel('G:\\输出文件\\{0} 上下架数据不全表 {1} .xlsx'.format(tem_data, rq), sheet_name='查询', index=False,engine='xlsxwriter')
                        sql = '''DELETE FROM customer c WHERE c.订单编号 IS  NULL OR c.运单编号 IS  NULL OR c.退货单号 IS  NULL;'''.format(team)
                        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                        print('已导出 并清除 上下架数据不全的订单......')

                    if team == 'ga t_return_bill':
                        sql = '''SELECT c.运单编号, 退货单号 
                                FROM customer c
                                LEFT JOIN  gat_order_list g ON c.订单编号 =g.订单编号
                                WHERE g.订单编号 IS NOT NULL;'''.format(team)
                        df = pd.read_sql_query(sql=sql, con=self.engine1)
                        df.to_excel('G:\\输出文件\\{0} 导入收货表 {1} .xlsx'.format(tem_data, rq), sheet_name='查询', index=False,engine='xlsxwriter')
                        print('收货表导出')

                        sql = '''SELECT 退货单号, 退货上架货架
                                FROM customer c
                                LEFT JOIN  gat_order_list g ON c.订单编号 =g.订单编号
                                WHERE g.订单编号 IS NOT NULL;'''.format(team)
                        df = pd.read_sql_query(sql=sql, con=self.engine1)
                        df.to_excel('G:\\输出文件\\{0} 导入上架表 {1} .xlsx'.format(tem_data, rq), sheet_name='查询', index=False,engine='xlsxwriter')
                        print('上架表导出')
                    elif team == 'ga gat_return_bill_over':
                        sql = '''SELECT c.*
                                FROM customer c
                                LEFT JOIN  gat_order_list g ON c.订单编号 =g.订单编号
                                WHERE g.订单编号 IS NOT NULL;'''.format(team)
                        df = pd.read_sql_query(sql=sql, con=self.engine1)
                        df.to_excel('G:\\输出文件\\{0} 下架表 {1} .xlsx'.format(tem_data, rq), sheet_name='查询', index=False,engine='xlsxwriter')
                        print('下架表导出')

                    columns = list(db.columns)
                    columns = ','.join(columns)
                    sql = '''REPLACE INTO {0}({1}, 记录时间) SELECT *, NOW() 记录时间 FROM customer;'''.format(team, columns)
                    pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                    print('++++：' + sht.name + '表--->>>上传成功')
                else:
                    print('----------数据为空,不需导入：' + sht.name)
            wb.close()
        app.quit()



if __name__ == '__main__':
    m = Updata_return_bill('+86-18538110674', 'qyz04163510')
    start: datetime = datetime.datetime.now()
    '''
    # -----------------------------------------------查询状态运行（一）-----------------------------------------
    # 1、 点到表上传 team = 'gat_logisitis_googs'；2、上架表上传；；3、订单跟进上传 team = 'gat_waybill_list'--->>数据更新切换
    '''

    select = 1
    if int(select) == 1:
        m.readFormHost()


    print('查询耗时：', datetime.datetime.now() - start)