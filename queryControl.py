# coding:utf-8
import pandas as pd
import os
import datetime
import xlwings as xl

import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel

from sqlalchemy import create_engine
from settings import Settings
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, \
    Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色


# -*- coding:utf-8 -*-
class QueryControl(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()

    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    def writeSqlReplace(self, dataFrame):
        dataFrame.to_sql('tem', con=self.engine1, index=False, if_exists='replace')

    def replaceInto(self, team, dfColumns):
        columns = list(dfColumns)
        columns = ', '.join(columns)
        if team == 'slrb':
            print(team + '---9')
            sql = 'REPLACE INTO {}({}, 添加时间) SELECT *, NOW() 添加时间 FROM tem; '.format(team, columns)
        else:
            print(team)
            sql = 'INSERT IGNORE INTO {}({}, 添加时间) SELECT *, NOW() 添加时间 FROM tem; '.format(team, columns)
        try:
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)
        except Exception as e:
            print('插入失败：', str(Exception) + str(e))

    def readSql(self, sql):
        db = pd.read_sql_query(sql=sql, con=self.engine1)
        # db = pd.read_sql(sql=sql, con=self.engine1) or team == 'slgat'
        return db
    # 团队花费明细查询（公用）（现用）
    def sl_tem_cost(self, team, tem):
        match = {'slgat_zqsb': '"台湾", "香港"',
                 'sltg_zqsb': '泰国',
                 'slxmt_zqsb': '"新加坡", "马来西亚", "菲律宾"',
                 'slrb_zqsb_rb': '日本'}
        match3 = {'新加坡': 'SG',
                  '马来西亚': 'MY',
                  '菲律宾': 'PH',
                  '日本': 'JP',
                  '香港': 'HK',
                  '台湾': 'TW',
                  '泰国': 'TH'}
        emailAdd = {'slgat_zqsb': 'giikinliujun@163.com',
                    'sltg_zqsb': 'zhangjing@giikin.com',
                    'slxmt_zqsb': 'zhangjing@giikin.com',
                    'slzb': '直播团队',
                    'slyn': '越南',
                    'slrb_zqsb_rb': 'sunyaru@giikin.com'}
        start = datetime.datetime.now()
        endDate = (datetime.datetime.now()).strftime('%Y%m')
        print(endDate)
        startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y%m')
        print(startDate)
        if datetime.datetime.now().day >= 9:
            end_Date = [startDate, endDate]
            start_Date = [startDate, endDate]
        else:
            end_Date = [startDate]
            start_Date = [startDate]
        listT = []  # 查询sql 存放池
        show_name = []  # 打印进度需要
        for i in range(len(end_Date)):
            # 总花费明细表---查询
            # sql20 = '''SELECT *
            #         FROM (
            #             SELECT sl_zong.币种,
            #                     IFNULL(sl_zong.年月,'合计') 年月,
            #                     IFNULL(sl_zong.父级分类,'合计') 父级分类,
            #                     IFNULL(sl_zong.二级分类,'合计') 二级分类,
            #                     IFNULL(sl_zong.三级分类,'合计') 三级分类,
            #                     IFNULL(sl_zong.产品id,'合计') 产品id,
            #                     IFNULL(sl_zong.产品名称,'合计') 产品名称,
            #                     IFNULL(sl_zong.物流方式,'合计') 物流方式,
            #                     IFNULL(sl_zong.旬,'合计') 旬,
            #                     SUM(sl_zong.订单量) 订单量,
            #                     IFNULL(SUM(sl_zong_zf.`直发订单量`),0) 直发订单量,
            #                     (SUM(sl_zong.订单量) - IFNULL(SUM(sl_zong_zf.`直发订单量`),0)) AS 改派订单量,
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) 签收订单量,
            #                     IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0) 拒收订单量,
            #                     SUM(sl_zong.总成本) / SUM(sl_zong.销售额)  AS '采购/销售额',
            #                     IFNULL(SUM(sl_zong_zf.`直发成本`),0) / SUM(sl_zong.销售额)  AS '直发采购/销售额',
            #                     SUM(sl_zong.物流运费) / SUM(sl_zong.销售额)  AS '运费占比',
            #                     SUM(sl_zong.手续费) / SUM(sl_zong.销售额)  AS '手续费占比',
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) AS '金额签收/完成',
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / SUM(sl_zong.销售额) AS '金额签收/总计',
            #                     (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) / SUM(sl_zong.销售额) AS '金额完成占比',
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) AS '数量签收/完成',
            #                     (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) / SUM(sl_zong.订单量) AS '数量完成占比'
            #             FROM (
            #                     SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 订单量,
            #                             SUM(`价格RMB`) 销售额,
            #                             SUM(`成本价`) 总成本,
            #                             SUM(`物流花费`) 物流运费,
            #                             SUM(`打包花费`) 打包花费,
            #                             SUM(`其它花费`) 手续费
            #                     FROM  {0} sl_cx
            #                     WHERE sl_cx.`币种` = '{1}'
            #                         AND sl_cx.`年月` >= '{start_Date}'
            #                         AND sl_cx.`年月` <= '{end_Date}'
            #                         AND sl_cx.`系统订单状态`!="已删除"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #                 ) sl_zong
            #         LEFT JOIN
            #                 (   SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 直发订单量,
            #                             SUM(`价格RMB`) 销售额,
            #                             SUM(`成本价`) 直发成本,
            #                             SUM(`物流花费`) 物流运费,
            #                             SUM(`打包花费`) 打包花费,
            #                             SUM(`其它花费`) 手续费
            #                     FROM  {0} sl_cx_zf
            #                     WHERE sl_cx_zf.`币种` = '{1}'
            #                         AND sl_cx_zf.`年月` >= '{start_Date}'
            #                         AND sl_cx_zf.`年月` <= '{end_Date}'
            #                         AND sl_cx_zf.`系统订单状态`!="已删除"
            #                         AND sl_cx_zf.`是否改派` = "直发"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #             ) sl_zong_zf
            #                 ON sl_zong_zf.`币种` = sl_zong.`币种`
            #                     AND sl_zong_zf.`年月` = sl_zong.`年月`
            #                     AND sl_zong_zf.`父级分类` = sl_zong.`父级分类`
            #                     AND sl_zong_zf.`二级分类` = sl_zong.`二级分类`
            #                     AND sl_zong_zf.`三级分类` = sl_zong.`三级分类`
            #                     AND sl_zong_zf.`产品id` = sl_zong.`产品id`
            #                     AND sl_zong_zf.`产品名称` = sl_zong.`产品名称`
            #                     AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
            #                     AND sl_zong_zf.`旬` = sl_zong.`旬`
            #         LEFT JOIN
            #                 (   SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 已签收订单量,
            #                             SUM(`价格RMB`) 已签收销售额,
            #                             SUM(`成本价`) 已签收成本,
            #                             SUM(`物流花费`) 已签收物流运费,
            #                             SUM(`打包花费`) 已签收打包花费,
            #                             SUM(`其它花费`) 已签收手续费
            #                     FROM  {0} sl_cx_zhifa
            #                     WHERE sl_cx_zhifa.`币种` = '{1}'
            #                         AND sl_cx_zhifa.`年月` >= '{start_Date}'
            #                         AND sl_cx_zhifa.`年月` <= '{end_Date}'
            #                         AND sl_cx_zhifa.`系统订单状态`!="已删除"
            #                         AND sl_cx_zhifa.`最终状态` = "已签收"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #             ) sl_zong_zhifa
            #             ON sl_zong_zhifa.`币种` = sl_zong.`币种`
            #                     AND sl_zong_zhifa.`年月` = sl_zong.`年月`
            #                     AND sl_zong_zhifa.`父级分类` = sl_zong.`父级分类`
            #                     AND sl_zong_zhifa.`二级分类` = sl_zong.`二级分类`
            #                     AND sl_zong_zhifa.`三级分类` = sl_zong.`三级分类`
            #                     AND sl_zong_zhifa.`产品id` = sl_zong.`产品id`
            #                     AND sl_zong_zhifa.`产品名称` = sl_zong.`产品名称`
            #                     AND sl_zong_zhifa.`物流方式` = sl_zong.`物流方式`
            #                     AND sl_zong_zhifa.`旬` = sl_zong.`旬`
            #         LEFT JOIN
            #                 (   SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 拒收订单量,
            #                             SUM(`价格RMB`) 拒收销售额,
            #                             SUM(`成本价`) 拒收成本,
            #                             SUM(`物流花费`) 拒收物流运费,
            #                             SUM(`打包花费`) 拒收打包花费,
            #                             SUM(`其它花费`) 拒收手续费
            #                     FROM  {0} sl_cx_jushou
            #                     WHERE sl_cx_jushou.`币种` = '{1}'
            #                         AND sl_cx_jushou.`年月` >= '{start_Date}'
            #                         AND sl_cx_jushou.`年月` <= '{end_Date}'
            #                         AND sl_cx_jushou.`系统订单状态`!="已删除"
            #                         AND sl_cx_jushou.`最终状态` = "拒收"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #             ) sl_zong_jushou
            #             ON sl_zong_jushou.`币种` = sl_zong.`币种`
            #                     AND sl_zong_jushou.`年月` = sl_zong.`年月`
            #                     AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类`
            #                     AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类`
            #                     AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类`
            #                     AND sl_zong_jushou.`产品id` = sl_zong.`产品id`
            #                     AND sl_zong_jushou.`产品名称` = sl_zong.`产品名称`
            #                     AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式`
            #                     AND sl_zong_jushou.`旬` = sl_zong.`旬`
            #         GROUP BY sl_zong.父级分类,sl_zong.三级分类,sl_zong.产品名称,sl_zong.物流方式,sl_zong.旬
            #         with rollup
            #         ) sl_zong_wl
            #         WHERE sl_zong_wl.`旬` = '合计';'''.format(team, tem, start_Date=start_Date[i], end_Date=end_Date[i])
            # listT.append(sql20)
            # show_name.append(start_Date[i] + '月（总）详细花费数据…………')

            # 直发花费明细表---查询
            # sql30 = '''SELECT sl.`币种`,
            #                     sl.`年月`,
            #                     sl.`父级分类`,
            #                     sl.`三级分类`,
            #                     sl.`产品名称`,
            #                     sl.`物流方式`,
            #                     sl.`总订单量`,
            #                     sl.`直发订单量`,
            #                     sl.`直发退货量`,
            #                     sl.`已签收订单量`,
            #                     sl.`拒收订单量`,
            #                     sl.`总销售额` / sl.`总订单量` AS '总客单价',
            #                     sl.`直发成本` / sl.`总销售额` AS '直发采购/销售额',
            #                     sl.`直发物流运费` / sl.`总销售额` AS '运费占比',
            #                     sl.`直发手续费` / sl.`总销售额` AS '手续费占比',
            #                     sl.`已签收销售额` / (sl.`已签收销售额` + sl.`拒收销售额`) AS '金额签收/完成',
            #                     sl.`已签收销售额` / sl.`总销售额` AS '金额签收/总计',
            #                     (sl.`已签收销售额` + sl.`拒收销售额`) /  sl.`总销售额`  AS '金额完成占比',
            #                     sl.`已签收订单量` /  (sl.`已签收订单量` + sl.`拒收订单量`) AS '数量签收/完成',
            #                     (sl.`已签收订单量` + sl.`拒收订单量`) / sl.`总订单量` AS '数量完成占比'
            #         FROM {0} sl
            #          WHERE sl.`币种` = '{1}'
            #             AND sl.`旬` = CONVERT('合计' USING utf8) COLLATE utf8_general_ci
            #             AND sl.`年月` = CONVERT('{end_Date}' USING utf8) COLLATE utf8_general_ci;'''.format(match1[team], tem, end_Date=end_Date[i])

            sql30 = '''SELECT *
                    FROM (
                        SELECT sl_zong.币种,
                            IFNULL(sl_zong.年月,'合计') 年月,
                            IFNULL(sl_zong.父级分类,'合计') 父级分类,
                            IFNULL(sl_zong.二级分类,'合计') 二级分类,
                            IFNULL(sl_zong.三级分类,'合计') 三级分类,
                            IFNULL(sl_zong.产品id,'合计') 产品id,
                            IFNULL(sl_zong.产品名称,'合计') 产品名称,
                            IFNULL(sl_zong.物流方式,'合计') 物流方式,
                            IFNULL(sl_zong.旬,'合计') 旬,
                            SUM(sl_zong.直发订单量) 直发订单量,
                            IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) 签收订单量,
                            IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0) 拒收订单量,
                            SUM(sl_zong.直发成本) / SUM(sl_zong.销售额) AS '直发采购/销售额',
                            SUM(sl_zong.物流运费) / SUM(sl_zong.销售额) AS '运费占比',
                            SUM(sl_zong.手续费) / SUM(sl_zong.销售额) AS '手续费占比',
                            IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) AS '金额签收/完成',
                            IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / SUM(sl_zong.销售额) AS '金额签收/总计',
                            (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) / SUM(sl_zong.销售额) AS '金额完成占比',
                            IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) AS '数量签收/完成',
                            (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) / SUM(sl_zong.直发订单量) AS '数量完成占比'
                        FROM (
                            SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 直发订单量,
                                    SUM(`价格RMB`) 销售额,
                                    SUM(`成本价`) 直发成本,
                                    SUM(`物流花费`) 物流运费,
                                    SUM(`打包花费`) 打包花费,
                                    SUM(`其它花费`) 手续费
                            FROM  {0} sl_cx
                            WHERE sl_cx.`币种` = '{1}'
                                AND sl_cx.`年月` >= '{start_Date}'
                                AND sl_cx.`年月` <= '{end_Date}'
                                AND sl_cx.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx.`是否改派` = "直发"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong
                        LEFT JOIN
                            (SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 已签收订单量,
                                    SUM(`价格RMB`) 已签收销售额,
                                    SUM(`成本价`) 已签收成本,
                                    SUM(`物流花费`) 已签收物流运费,
                                    SUM(`打包花费`) 已签收打包花费,
                                    SUM(`其它花费`) 已签收手续费
                            FROM  {0}	sl_cx_zhifa
                            WHERE sl_cx_zhifa.`币种` = '{1}'
                                AND sl_cx_zhifa.`年月` >= '{start_Date}'
                                AND sl_cx_zhifa.`年月` <= '{end_Date}'
                                AND sl_cx_zhifa.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx_zhifa.`是否改派` = "直发"
                                AND sl_cx_zhifa.`最终状态` = "已签收"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong_zhifa
                            ON sl_zong_zhifa.`币种` = sl_zong.`币种`
                                AND sl_zong_zhifa.`年月` = sl_zong.`年月`
                                AND sl_zong_zhifa.`父级分类` = sl_zong.`父级分类`
                                AND sl_zong_zhifa.`二级分类` = sl_zong.`二级分类`
                                AND sl_zong_zhifa.`三级分类` = sl_zong.`三级分类`
                                AND sl_zong_zhifa.`产品id` = sl_zong.`产品id`
                                AND sl_zong_zhifa.`产品名称` = sl_zong.`产品名称`
                                AND sl_zong_zhifa.`物流方式` = sl_zong.`物流方式`
                                AND sl_zong_zhifa.`旬` = sl_zong.`旬`
                        LEFT JOIN
                            (SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 拒收订单量,
                                    SUM(`价格RMB`) 拒收销售额,
                                    SUM(`成本价`) 拒收成本,
                                    SUM(`物流花费`) 拒收物流运费,
                                    SUM(`打包花费`) 拒收打包花费,
                                    SUM(`其它花费`) 拒收手续费
                            FROM  {0} sl_cx_jushou
                            WHERE sl_cx_jushou.`币种` = '{1}'
                                AND sl_cx_jushou.`年月` >= '{start_Date}'
                                AND sl_cx_jushou.`年月` <= '{end_Date}'
                                AND sl_cx_jushou.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx_jushou.`是否改派` = "直发"
                                AND sl_cx_jushou.`最终状态` = "拒收"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong_jushou
                        ON sl_zong_jushou.`币种` = sl_zong.`币种`
                            AND sl_zong_jushou.`年月` = sl_zong.`年月`
                            AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类`
                            AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类`
                            AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类`
                            AND sl_zong_jushou.`产品id` = sl_zong.`产品id`
                            AND sl_zong_jushou.`产品名称` = sl_zong.`产品名称`
                            AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式`
                            AND sl_zong_jushou.`旬` = sl_zong.`旬`
                    GROUP BY sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.产品名称,sl_zong.物流方式,sl_zong.旬
                    with rollup
                    ) sl_zong_wl
                    WHERE sl_zong_wl.`旬` = '合计';'''.format(team, tem, start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql30)
            show_name.append(start_Date[i] + '月（直发）详细花费数据…………')
            # 总成本父级
            sql40 = '''SELECT *
                        FROM(
                            (SELECT s1.团队,
                                    s1.年月,
                                    s1.品类,
                                    s1.销售额,
                                    s1.订单量,
                                    (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                    (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                    s1.改派订单量,
                                    s1.改派订单量 / s1.订单量 AS '改派占比',
                                    s1.销售额 / s1.订单量 AS '客单价',
                                    s1.销售额 / s1.广告成本 AS 'ROI',
                                    s1.活跃产品数,
                                    s1.订单量 / s1.活跃产品数 AS 产能,
                                    s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                    s1.广告成本 / s1.销售额 AS '广告占比',
                                    s1.物流成本 / s1.销售额 AS '运费占比',
                                    s1.手续费 / s1.销售额 AS '手续费占比',
                                    (s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                    (s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                    s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                    s1.签收额 / s1.销售额 AS '金额签收/总计',
                                    (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                    s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                    s1.签收量 / s1.订单量 AS '数量签收/总计',
                                    (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                    s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                    (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                                FROM (SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.ppname AS 品类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费 
                                    FROM gk_order_day a
                                        LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                        LEFT JOIN dim_area c on c.id = a.area_id
                                        LEFT JOIN dim_cate d on d.id = a.third_cate_id
                            --          LEFT JOIN gk_product e on e.id = a.product_id
                                        LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                    WHERE b.pcode = '{0}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                        AND c.uname = '王冰'
                                        AND a.beform <> 'mf'
                                        AND c.uid <> 10099  -- 过滤翼虎
                                     GROUP BY b.pname, c.uname, a.cate_id
                                    ORDER BY a.product_id
                                ) s1 WHERE s1.订单量 > 0 ORDER BY s1.订单量)
                            UNION ALL
                                (SELECT s3.团队,
						                s3.年月,
						                s3.品类,
						                s3.销售额,
						                s3.订单量,
						                (s3.订单量 - s3.改派订单量) AS 直发订单量,
						                (s3.订单量 - s3.改派订单量) / s3.订单量 AS 直发占比,
						                s3.改派订单量,
						                s3.改派订单量 / s3.订单量 AS 改派占比,
						                s3.销售额 / s3.订单量 客单价,
						                s3.销售额 / s3.广告成本 ROI,
						                S3.活跃产品数,
						                s3.订单量 / S3.活跃产品数 AS 产能,
						                s3.直发采购额 / s3.销售额 AS '直发采购/总销售额',
						                s3.广告成本 / s3.销售额 AS '广告占比',
						                s3.物流成本 / s3.销售额 AS '运费占比',
						                s3.手续费 / s3.销售额 AS '手续费占比',
						                (s3.广告成本 + s3.物流成本 + s3.手续费 + s3.直发采购额 ) AS '总成本',
						                (s3.广告成本 + s3.物流成本 + s3.手续费 + s3.直发采购额 ) / s3.销售额 AS '总成本占比',
						                s3.签收额 / (s3.拒收额 + s3.签收额) '金额签收/完成',
						                s3.签收额 / s3.销售额 '金额签收/总计',
						                (s3.签收额 + s3.拒收额) / s3.销售额 '金额完成占比',
						                s3.签收量 / (s3.拒收量 + s3.签收量) '数量签收/完成',
						                s3.签收量 / s3.订单量 '数量签收/总计',
						                (s3.拒收量 + s3.签收量) / s3.订单量 '数量完成占比',
						                s3.签收额 / (s3.签收额 + s3.拒收额) -( s3.直发采购额 + s3.广告成本 + s3.物流成本 + s3.手续费 ) AS 利润率,
						                (s3.签收额 / (s3.签收额 + s3.拒收额) -( s3.直发采购额 + s3.广告成本 + s3.物流成本 + s3.手续费 ) / s3.销售额) * (s3.销售额 / s3.订单量) AS 利润值
                                FROM (SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
						                b.pname AS 团队,
						                '合计' AS 品类,
						                SUM(a.salesRMB) 销售额,
						                SUM(a.orders) AS 订单量,
						                COUNT(DISTINCT a.product_id) AS 活跃产品数,
						                '' 产能,
						                SUM(a.yqs) AS 签收量,
						                SUM(a.yjs) AS 拒收量,
						                SUM(a.salesRMB_yqs) AS 签收额,
						                SUM(a.salesRMB_yjs) AS 拒收额,
						                SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
						                SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费 
			                        FROM gk_order_day a
				                        LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
				                        LEFT JOIN dim_area c on c.id = a.area_id
				                        LEFT JOIN dim_cate d on d.id = a.third_cate_id
                            --          LEFT JOIN gk_product e on e.id = a.product_id
                                        LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                    WHERE b.pcode = '{0}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
						                AND c.uname = '王冰'
						                AND a.beform <> 'mf'
						                AND c.uid <> 10099  -- 过滤翼虎
			                        GROUP BY b.pname, c.uname
                                ) s3)
                            ) s ORDER BY s.订单量'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql40)
            show_name.append(start_Date[i] + '月（父级）成本数据…………')
            # 总成本二级
            sql41 = '''SELECT s1.团队,
                                s1.年月,
                                s1.二级分类,
                                s1.销售额,
                                s1.订单量,
                                (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                s1.改派订单量,
                                s1.改派订单量 / s1.订单量 AS '改派占比',
                                s1.销售额 / s1.订单量 AS '客单价',
                                s1.销售额 / s1.广告成本 AS 'ROI',
                                s1.活跃产品数,
                                s1.订单量 / s1.活跃产品数 AS 产能,
                                s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                s1.广告成本 / s1.销售额 AS '广告占比',
                                s1.物流成本 / s1.销售额 AS '运费占比',
                                s1.手续费 / s1.销售额 AS '手续费占比',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                s1.签收额 / s1.销售额 AS '金额签收/总计',
                                (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                s1.签收量 / s1.订单量 AS '数量签收/总计',
                                (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                        FROM (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.pname AS 二级分类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费  
                                FROM gk_order_day a
                                    LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                    LEFT JOIN dim_area c on c.id = a.area_id
                                    LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        --          LEFT JOIN gk_product e on e.id = a.product_id
                                    LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                WHERE b.pcode = '{0}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                    AND c.uname = '王冰'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099  -- 过滤翼虎
                                GROUP BY b.pname, c.uname, a.second_cate_id
                                ORDER BY a.product_id
                        ) s1
                        WHERE s1.订单量 > 0
                        ORDER BY s1.订单量;'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql41)
            show_name.append(start_Date[i] + '月（二级）成本数据…………')
            # 总成本三级
            sql42 = '''SELECT s1.团队,
                                s1.年月,
                                s1.三级分类,
                                s1.销售额,
                                s1.订单量,
                                (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                s1.改派订单量,
                                s1.改派订单量 / s1.订单量 AS '改派占比',
                                s1.销售额 / s1.订单量 AS '客单价',
                                s1.销售额 / s1.广告成本 AS 'ROI',
                                s1.活跃产品数,
                                s1.订单量 / s1.活跃产品数 AS 产能,
                                s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                s1.广告成本 / s1.销售额 AS '广告占比',
                                s1.物流成本 / s1.销售额 AS '运费占比',
                                s1.手续费 / s1.销售额 AS '手续费占比',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                s1.签收额 / s1.销售额 AS '金额签收/总计',
                                (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                s1.签收量 / s1.订单量 AS '数量签收/总计',
                                (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                        FROM (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.`name` AS 三级分类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费 
                                FROM gk_order_day a
                                    LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                    LEFT JOIN dim_area c on c.id = a.area_id
                                    LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        --          LEFT JOIN gk_product e on e.id = a.product_id
                                    LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                WHERE b.pcode = '{0}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                    AND c.uname = '王冰'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099  -- 过滤翼虎
                                GROUP BY b.pname, c.uname, a.third_cate_id
                                ORDER BY a.product_id
                        ) s1
                        WHERE s1.订单量 > 0
                        ORDER BY s1.订单量;'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql42)
            show_name.append(start_Date[i] + '月（三级）成本数据…………')
        listTValue = []  # 查询sql的结果 存放池
        for i, sql in enumerate(listT):
            print(i)
            print('正在获取 ' + tem + show_name[i])
            if i == 0 or i == 4:
                df = pd.read_sql_query(sql=sql, con=self.engine1)
            else:
                df = pd.read_sql_query(sql=sql, con=self.engine2)
            # print(df)
            columns = list(df.columns)  # 获取数据的标题名，转为列表
            columns_value = ['直发占比', '改派占比', '直发采购/总销售额', '广告占比', '总成本占比', '数量签收/总计',  '利润率', '采购/销售额', '直发采购/销售额', '运费占比', '手续费占比', '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比']
            if '旬' in columns:
                df.drop(labels=['旬'], axis=1, inplace=True)  # 去掉多余的旬列表
            for column_val in columns_value:
                if column_val in columns:
                    df[column_val] = df[column_val].fillna(value=0)
                    df[column_val] = df[column_val].apply(lambda x: format(x, '.2%'))
            listTValue.append(df)
        print('查询耗时：', datetime.datetime.now() - start)
        today = datetime.date.today().strftime('%Y.%m.%d')
        sheet_name = ['直发成本', '父级成本', '二级成本', '三级成本']  # 生成的工作表的表名
        if len(listTValue) == 4:
            file_Path = []  # 发送邮箱文件使用
            filePath = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}上月产品花费表.xlsx'.format(today, tem)
            # if os.path.exists(filePath):  # 判断是否有需要的表格
            #     print("正在使用(上月-单月)文件......")
            #     filePath = filePath
            # else:  # 判断是否无需要的表格，进行初始化创建
            #     print("正在创建文件......")
            #     df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            #     df0.to_excel(filePath, sheet_name='缓存使用', index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            #     filePath = filePath
            # print('正在写入excel…………')
            # writer = pd.ExcelWriter(filePath, engine='openpyxl')  # 初始化写入对象
            # book = load_workbook(filePath)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            # for i in range(len(listTValue)):
            #     listTValue[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            # if '缓存使用' in book.sheetnames:  # 删除新建文档时的第一个工作表
            #     del book['缓存使用']
            # writer.save()
            # writer.close()
            with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
                for i in range(len(listTValue)):
                    listTValue[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            print('输出文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePath)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePath, sheet_name[1])
            file_Path.append(filePath)
            self.e.send(tem + '产品花费表', file_Path,
                        emailAdd[team])
        else:
            file_Path = []  # 发送邮箱文件使用
            filePath = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}上月产品花费表.xlsx'.format(today, tem)
            # if os.path.exists(filePath):  # 判断是否有需要的表格
            #     print("正在使用(上月)文件......")
            #     filePath = filePath
            # else:  # 判断是否无需要的表格，进行初始化创建
            #     print("正在创建文件......")
            #     df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            #     df0.to_excel(filePath, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            #     filePath = filePath
            # print('正在写入excel…………')
            # writer = pd.ExcelWriter(filePath, engine='openpyxl')  # 初始化写入对象
            # book = load_workbook(filePath)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            # listTValue[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
            # listTValue[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
            # listTValue[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
            # listTValue[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
            # # listTValue[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
            # writer.save()
            # writer.close()
            with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
                listTValue[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
                listTValue[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
                listTValue[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
                listTValue[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
                # listTValue[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)

            print('输出(上月)文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePath)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePath, sheet_name[1])
            file_Path.append(filePath)
            print('------分割线------')
            filePathT = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}本月产品花费表.xlsx'.format(today, tem)
            if os.path.exists(filePathT):  # 判断是否有需要的表格
                print("正在使用(本月)文件......")
                filePathT = filePathT
            else:  # 判断是否无需要的表格，进行初始化创建
                print("正在创建文件......")
                df0T = pd.DataFrame([])  # 创建空的dataframe数据框-2
                df0T.to_excel(filePathT, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）-2
                filePathT = filePathT
            print('正在写入excel…………')
            writerT = pd.ExcelWriter(filePathT, engine='openpyxl')  # 初始化写入对象-2
            bookT = load_workbook(filePathT)  # 可以向不同的sheet写入数据（对现有工作表的追加）-2
            writerT.book = bookT  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet-2
            listTValue[4].to_excel(excel_writer=writerT, sheet_name=sheet_name[0], index=False)
            listTValue[5].to_excel(excel_writer=writerT, sheet_name=sheet_name[1], index=False)
            listTValue[6].to_excel(excel_writer=writerT, sheet_name=sheet_name[2], index=False)
            listTValue[7].to_excel(excel_writer=writerT, sheet_name=sheet_name[3], index=False)
            writerT.save()
            writerT.close()
            print('输出(本月)文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePathT)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePathT, sheet_name[1])
            file_Path.append(filePathT)
            self.e.send(tem + '产品花费表', file_Path,
                        emailAdd[team])
            print('处理耗时：', datetime.datetime.now() - start)
    def sl_tem_costHFH(self, team, tem):
        match = {'slgat_hfh_zqsb': '"台湾", "香港"',
                 'sltg_zqsb': '泰国',
                 'slxmt_hfh_zqsb': '"新加坡", "马来西亚", "菲律宾"',
                 'slrb_zqsb_rb': '日本'}
        match3 = {'新加坡': 'SG',
                  '马来西亚': 'MY',
                  '菲律宾': 'PH',
                  '日本': 'JP',
                  '香港': 'HK',
                  '台湾': 'TW',
                  '泰国': 'TH'}
        emailAdd = {'slgat_hfh_zqsb': 'giikinliujun@163.com',
                    'sltg_zqsb': 'zhangjing@giikin.com',
                    'slxmt_hfh_zqsb': 'zhangjing@giikin.com',
                    'slzb': '直播团队',
                    'slyn': '越南',
                    'slrb_zqsb_rb': 'sunyaru@giikin.com'}
        start = datetime.datetime.now()
        endDate = (datetime.datetime.now()).strftime('%Y%m')
        print(endDate)
        startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y%m')
        print(startDate)
        if datetime.datetime.now().day >= 9:
            end_Date = [startDate, endDate]
            start_Date = [startDate, endDate]
        else:
            end_Date = [startDate]
            start_Date = [startDate]
        listT = []  # 查询sql 存放池
        show_name = []  # 打印进度需要
        for i in range(len(end_Date)):
            # 总花费明细表---查询
            # sql20 = '''SELECT *
            #         FROM (
            #             SELECT sl_zong.币种,
            #                     IFNULL(sl_zong.年月,'合计') 年月,
            #                     IFNULL(sl_zong.父级分类,'合计') 父级分类,
            #                     IFNULL(sl_zong.二级分类,'合计') 二级分类,
            #                     IFNULL(sl_zong.三级分类,'合计') 三级分类,
            #                     IFNULL(sl_zong.产品id,'合计') 产品id,
            #                     IFNULL(sl_zong.产品名称,'合计') 产品名称,
            #                     IFNULL(sl_zong.物流方式,'合计') 物流方式,
            #                     IFNULL(sl_zong.旬,'合计') 旬,
            #                     SUM(sl_zong.订单量) 订单量,
            #                     IFNULL(SUM(sl_zong_zf.`直发订单量`),0) 直发订单量,
            #                     (SUM(sl_zong.订单量) - IFNULL(SUM(sl_zong_zf.`直发订单量`),0)) AS 改派订单量,
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) 签收订单量,
            #                     IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0) 拒收订单量,
            #                     SUM(sl_zong.总成本) / SUM(sl_zong.销售额)  AS '采购/销售额',
            #                     IFNULL(SUM(sl_zong_zf.`直发成本`),0) / SUM(sl_zong.销售额)  AS '直发采购/销售额',
            #                     SUM(sl_zong.物流运费) / SUM(sl_zong.销售额)  AS '运费占比',
            #                     SUM(sl_zong.手续费) / SUM(sl_zong.销售额)  AS '手续费占比',
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) AS '金额签收/完成',
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / SUM(sl_zong.销售额) AS '金额签收/总计',
            #                     (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) / SUM(sl_zong.销售额) AS '金额完成占比',
            #                     IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) AS '数量签收/完成',
            #                     (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) / SUM(sl_zong.订单量) AS '数量完成占比'
            #             FROM (
            #                     SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 订单量,
            #                             SUM(`价格RMB`) 销售额,
            #                             SUM(`成本价`) 总成本,
            #                             SUM(`物流花费`) 物流运费,
            #                             SUM(`打包花费`) 打包花费,
            #                             SUM(`其它花费`) 手续费
            #                     FROM  {0} sl_cx
            #                     WHERE sl_cx.`币种` = '{1}'
            #                         AND sl_cx.`年月` >= '{start_Date}'
            #                         AND sl_cx.`年月` <= '{end_Date}'
            #                         AND sl_cx.`系统订单状态`!="已删除"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #                 ) sl_zong
            #         LEFT JOIN
            #                 (   SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 直发订单量,
            #                             SUM(`价格RMB`) 销售额,
            #                             SUM(`成本价`) 直发成本,
            #                             SUM(`物流花费`) 物流运费,
            #                             SUM(`打包花费`) 打包花费,
            #                             SUM(`其它花费`) 手续费
            #                     FROM  {0} sl_cx_zf
            #                     WHERE sl_cx_zf.`币种` = '{1}'
            #                         AND sl_cx_zf.`年月` >= '{start_Date}'
            #                         AND sl_cx_zf.`年月` <= '{end_Date}'
            #                         AND sl_cx_zf.`系统订单状态`!="已删除"
            #                         AND sl_cx_zf.`是否改派` = "直发"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #             ) sl_zong_zf
            #                 ON sl_zong_zf.`币种` = sl_zong.`币种`
            #                     AND sl_zong_zf.`年月` = sl_zong.`年月`
            #                     AND sl_zong_zf.`父级分类` = sl_zong.`父级分类`
            #                     AND sl_zong_zf.`二级分类` = sl_zong.`二级分类`
            #                     AND sl_zong_zf.`三级分类` = sl_zong.`三级分类`
            #                     AND sl_zong_zf.`产品id` = sl_zong.`产品id`
            #                     AND sl_zong_zf.`产品名称` = sl_zong.`产品名称`
            #                     AND sl_zong_zf.`物流方式` = sl_zong.`物流方式`
            #                     AND sl_zong_zf.`旬` = sl_zong.`旬`
            #         LEFT JOIN
            #                 (   SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 已签收订单量,
            #                             SUM(`价格RMB`) 已签收销售额,
            #                             SUM(`成本价`) 已签收成本,
            #                             SUM(`物流花费`) 已签收物流运费,
            #                             SUM(`打包花费`) 已签收打包花费,
            #                             SUM(`其它花费`) 已签收手续费
            #                     FROM  {0} sl_cx_zhifa
            #                     WHERE sl_cx_zhifa.`币种` = '{1}'
            #                         AND sl_cx_zhifa.`年月` >= '{start_Date}'
            #                         AND sl_cx_zhifa.`年月` <= '{end_Date}'
            #                         AND sl_cx_zhifa.`系统订单状态`!="已删除"
            #                         AND sl_cx_zhifa.`最终状态` = "已签收"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #             ) sl_zong_zhifa
            #             ON sl_zong_zhifa.`币种` = sl_zong.`币种`
            #                     AND sl_zong_zhifa.`年月` = sl_zong.`年月`
            #                     AND sl_zong_zhifa.`父级分类` = sl_zong.`父级分类`
            #                     AND sl_zong_zhifa.`二级分类` = sl_zong.`二级分类`
            #                     AND sl_zong_zhifa.`三级分类` = sl_zong.`三级分类`
            #                     AND sl_zong_zhifa.`产品id` = sl_zong.`产品id`
            #                     AND sl_zong_zhifa.`产品名称` = sl_zong.`产品名称`
            #                     AND sl_zong_zhifa.`物流方式` = sl_zong.`物流方式`
            #                     AND sl_zong_zhifa.`旬` = sl_zong.`旬`
            #         LEFT JOIN
            #                 (   SELECT 币种,
            #                             年月,
            #                             父级分类,
            #                             二级分类,
            #                             三级分类,
            #                             产品id,
            #                             CONCAT(产品id, '#' ,产品名称) 产品名称,
            #                             物流方式,
            #                             旬,
            #                             COUNT(`订单编号`) 拒收订单量,
            #                             SUM(`价格RMB`) 拒收销售额,
            #                             SUM(`成本价`) 拒收成本,
            #                             SUM(`物流花费`) 拒收物流运费,
            #                             SUM(`打包花费`) 拒收打包花费,
            #                             SUM(`其它花费`) 拒收手续费
            #                     FROM  {0} sl_cx_jushou
            #                     WHERE sl_cx_jushou.`币种` = '{1}'
            #                         AND sl_cx_jushou.`年月` >= '{start_Date}'
            #                         AND sl_cx_jushou.`年月` <= '{end_Date}'
            #                         AND sl_cx_jushou.`系统订单状态`!="已删除"
            #                         AND sl_cx_jushou.`最终状态` = "拒收"
            #                     GROUP BY 币种,年月,父级分类,三级分类,产品名称,物流方式,旬
            #                     ORDER BY 币种,年月
            #             ) sl_zong_jushou
            #             ON sl_zong_jushou.`币种` = sl_zong.`币种`
            #                     AND sl_zong_jushou.`年月` = sl_zong.`年月`
            #                     AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类`
            #                     AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类`
            #                     AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类`
            #                     AND sl_zong_jushou.`产品id` = sl_zong.`产品id`
            #                     AND sl_zong_jushou.`产品名称` = sl_zong.`产品名称`
            #                     AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式`
            #                     AND sl_zong_jushou.`旬` = sl_zong.`旬`
            #         GROUP BY sl_zong.父级分类,sl_zong.三级分类,sl_zong.产品名称,sl_zong.物流方式,sl_zong.旬
            #         with rollup
            #         ) sl_zong_wl
            #         WHERE sl_zong_wl.`旬` = '合计';'''.format(team, tem, start_Date=start_Date[i], end_Date=end_Date[i])
            # listT.append(sql20)
            # show_name.append(start_Date[i] + '月（总）详细花费数据…………')

            # 直发花费明细表---查询
            # sql30 = '''SELECT sl.`币种`,
            #                     sl.`年月`,
            #                     sl.`父级分类`,
            #                     sl.`三级分类`,
            #                     sl.`产品名称`,
            #                     sl.`物流方式`,
            #                     sl.`总订单量`,
            #                     sl.`直发订单量`,
            #                     sl.`直发退货量`,
            #                     sl.`已签收订单量`,
            #                     sl.`拒收订单量`,
            #                     sl.`总销售额` / sl.`总订单量` AS '总客单价',
            #                     sl.`直发成本` / sl.`总销售额` AS '直发采购/销售额',
            #                     sl.`直发物流运费` / sl.`总销售额` AS '运费占比',
            #                     sl.`直发手续费` / sl.`总销售额` AS '手续费占比',
            #                     sl.`已签收销售额` / (sl.`已签收销售额` + sl.`拒收销售额`) AS '金额签收/完成',
            #                     sl.`已签收销售额` / sl.`总销售额` AS '金额签收/总计',
            #                     (sl.`已签收销售额` + sl.`拒收销售额`) /  sl.`总销售额`  AS '金额完成占比',
            #                     sl.`已签收订单量` /  (sl.`已签收订单量` + sl.`拒收订单量`) AS '数量签收/完成',
            #                     (sl.`已签收订单量` + sl.`拒收订单量`) / sl.`总订单量` AS '数量完成占比'
            #         FROM {0} sl
            #          WHERE sl.`币种` = '{1}'
            #             AND sl.`旬` = CONVERT('合计' USING utf8) COLLATE utf8_general_ci
            #             AND sl.`年月` = CONVERT('{end_Date}' USING utf8) COLLATE utf8_general_ci;'''.format(match1[team], tem, end_Date=end_Date[i])

            sql30 = '''SELECT *
                    FROM (
                        SELECT sl_zong.币种,
                            IFNULL(sl_zong.年月,'合计') 年月,
                            IFNULL(sl_zong.父级分类,'合计') 父级分类,
                            IFNULL(sl_zong.二级分类,'合计') 二级分类,
                            IFNULL(sl_zong.三级分类,'合计') 三级分类,
                            IFNULL(sl_zong.产品id,'合计') 产品id,
                            IFNULL(sl_zong.产品名称,'合计') 产品名称,
                            IFNULL(sl_zong.物流方式,'合计') 物流方式,
                            IFNULL(sl_zong.旬,'合计') 旬,
                            SUM(sl_zong.直发订单量) 直发订单量,
                            IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) 签收订单量,
                            IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0) 拒收订单量,
                            SUM(sl_zong.直发成本) / SUM(sl_zong.销售额) AS '直发采购/销售额',
                            SUM(sl_zong.物流运费) / SUM(sl_zong.销售额) AS '运费占比',
                            SUM(sl_zong.手续费) / SUM(sl_zong.销售额) AS '手续费占比',
                            IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) AS '金额签收/完成',
                            IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / SUM(sl_zong.销售额) AS '金额签收/总计',
                            (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) / SUM(sl_zong.销售额) AS '金额完成占比',
                            IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) AS '数量签收/完成',
                            (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) / SUM(sl_zong.直发订单量) AS '数量完成占比'
                        FROM (
                            SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 直发订单量,
                                    SUM(`价格RMB`) 销售额,
                                    SUM(`成本价`) 直发成本,
                                    SUM(`物流花费`) 物流运费,
                                    SUM(`打包花费`) 打包花费,
                                    SUM(`其它花费`) 手续费
                            FROM  {0} sl_cx
                            WHERE sl_cx.`币种` = '{1}'
                                AND sl_cx.`年月` >= '{start_Date}'
                                AND sl_cx.`年月` <= '{end_Date}'
                                AND sl_cx.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx.`是否改派` = "直发"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong
                        LEFT JOIN
                            (SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 已签收订单量,
                                    SUM(`价格RMB`) 已签收销售额,
                                    SUM(`成本价`) 已签收成本,
                                    SUM(`物流花费`) 已签收物流运费,
                                    SUM(`打包花费`) 已签收打包花费,
                                    SUM(`其它花费`) 已签收手续费
                            FROM  {0}	sl_cx_zhifa
                            WHERE sl_cx_zhifa.`币种` = '{1}'
                                AND sl_cx_zhifa.`年月` >= '{start_Date}'
                                AND sl_cx_zhifa.`年月` <= '{end_Date}'
                                AND sl_cx_zhifa.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx_zhifa.`是否改派` = "直发"
                                AND sl_cx_zhifa.`最终状态` = "已签收"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong_zhifa
                            ON sl_zong_zhifa.`币种` = sl_zong.`币种`
                                AND sl_zong_zhifa.`年月` = sl_zong.`年月`
                                AND sl_zong_zhifa.`父级分类` = sl_zong.`父级分类`
                                AND sl_zong_zhifa.`二级分类` = sl_zong.`二级分类`
                                AND sl_zong_zhifa.`三级分类` = sl_zong.`三级分类`
                                AND sl_zong_zhifa.`产品id` = sl_zong.`产品id`
                                AND sl_zong_zhifa.`产品名称` = sl_zong.`产品名称`
                                AND sl_zong_zhifa.`物流方式` = sl_zong.`物流方式`
                                AND sl_zong_zhifa.`旬` = sl_zong.`旬`
                        LEFT JOIN
                            (SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 拒收订单量,
                                    SUM(`价格RMB`) 拒收销售额,
                                    SUM(`成本价`) 拒收成本,
                                    SUM(`物流花费`) 拒收物流运费,
                                    SUM(`打包花费`) 拒收打包花费,
                                    SUM(`其它花费`) 拒收手续费
                            FROM  {0} sl_cx_jushou
                            WHERE sl_cx_jushou.`币种` = '{1}'
                                AND sl_cx_jushou.`年月` >= '{start_Date}'
                                AND sl_cx_jushou.`年月` <= '{end_Date}'
                                AND sl_cx_jushou.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx_jushou.`是否改派` = "直发"
                                AND sl_cx_jushou.`最终状态` = "拒收"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong_jushou
                        ON sl_zong_jushou.`币种` = sl_zong.`币种`
                            AND sl_zong_jushou.`年月` = sl_zong.`年月`
                            AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类`
                            AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类`
                            AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类`
                            AND sl_zong_jushou.`产品id` = sl_zong.`产品id`
                            AND sl_zong_jushou.`产品名称` = sl_zong.`产品名称`
                            AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式`
                            AND sl_zong_jushou.`旬` = sl_zong.`旬`
                    GROUP BY sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.产品名称,sl_zong.物流方式,sl_zong.旬
                    with rollup
                    ) sl_zong_wl
                    WHERE sl_zong_wl.`旬` = '合计';'''.format(team, tem, start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql30)
            show_name.append(start_Date[i] + '月（直发）详细花费数据…………')
            # 总成本父级
            sql40 = '''SELECT *
                        FROM(
                            (SELECT s1.团队,
                                    s1.年月,
                                    s1.品类,
                                    s1.销售额,
                                    s1.订单量,
                                    (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                    (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                    s1.改派订单量,
                                    s1.改派订单量 / s1.订单量 AS '改派占比',
                                    s1.销售额 / s1.订单量 AS '客单价',
                                    s1.销售额 / s1.广告成本 AS 'ROI',
                                    s1.活跃产品数,
                                    s1.订单量 / s1.活跃产品数 AS 产能,
                                    s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                    s1.广告成本 / s1.销售额 AS '广告占比',
                                    s1.物流成本 / s1.销售额 AS '运费占比',
                                    s1.手续费 / s1.销售额 AS '手续费占比',
                                    (s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                    (s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                    s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                    s1.签收额 / s1.销售额 AS '金额签收/总计',
                                    (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                    s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                    s1.签收量 / s1.订单量 AS '数量签收/总计',
                                    (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                    s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                    (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                                FROM (SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.ppname AS 品类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费 
                                    FROM gk_order_day a
                                        LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                        LEFT JOIN dim_area c on c.id = a.area_id
                                        LEFT JOIN dim_cate d on d.id = a.third_cate_id
                            --          LEFT JOIN gk_product e on e.id = a.product_id
                                        LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                    WHERE b.pcode = '{0}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                        AND c.uname = '罗超源'
                                        AND a.beform <> 'mf'
                                        AND c.uid <> 10099  -- 过滤翼虎
                                     GROUP BY b.pname, c.uname, a.cate_id
                                    ORDER BY a.product_id
                                ) s1 WHERE s1.订单量 > 0 ORDER BY s1.订单量)
                            UNION ALL
                                (SELECT s3.团队,
						                s3.年月,
						                s3.品类,
						                s3.销售额,
						                s3.订单量,
						                (s3.订单量 - s3.改派订单量) AS 直发订单量,
						                (s3.订单量 - s3.改派订单量) / s3.订单量 AS 直发占比,
						                s3.改派订单量,
						                s3.改派订单量 / s3.订单量 AS 改派占比,
						                s3.销售额 / s3.订单量 客单价,
						                s3.销售额 / s3.广告成本 ROI,
						                S3.活跃产品数,
						                s3.订单量 / S3.活跃产品数 AS 产能,
						                s3.直发采购额 / s3.销售额 AS '直发采购/总销售额',
						                s3.广告成本 / s3.销售额 AS '广告占比',
						                s3.物流成本 / s3.销售额 AS '运费占比',
						                s3.手续费 / s3.销售额 AS '手续费占比',
						                (s3.广告成本 + s3.物流成本 + s3.手续费 + s3.直发采购额 ) AS '总成本',
						                (s3.广告成本 + s3.物流成本 + s3.手续费 + s3.直发采购额 ) / s3.销售额 AS '总成本占比',
						                s3.签收额 / (s3.拒收额 + s3.签收额) '金额签收/完成',
						                s3.签收额 / s3.销售额 '金额签收/总计',
						                (s3.签收额 + s3.拒收额) / s3.销售额 '金额完成占比',
						                s3.签收量 / (s3.拒收量 + s3.签收量) '数量签收/完成',
						                s3.签收量 / s3.订单量 '数量签收/总计',
						                (s3.拒收量 + s3.签收量) / s3.订单量 '数量完成占比',
						                s3.签收额 / (s3.签收额 + s3.拒收额) -( s3.直发采购额 + s3.广告成本 + s3.物流成本 + s3.手续费 ) AS 利润率,
						                (s3.签收额 / (s3.签收额 + s3.拒收额) -( s3.直发采购额 + s3.广告成本 + s3.物流成本 + s3.手续费 ) / s3.销售额) * (s3.销售额 / s3.订单量) AS 利润值
                                FROM (SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
						                b.pname AS 团队,
						                '合计' AS 品类,
						                SUM(a.salesRMB) 销售额,
						                SUM(a.orders) AS 订单量,
						                COUNT(DISTINCT a.product_id) AS 活跃产品数,
						                '' 产能,
						                SUM(a.yqs) AS 签收量,
						                SUM(a.yjs) AS 拒收量,
						                SUM(a.salesRMB_yqs) AS 签收额,
						                SUM(a.salesRMB_yjs) AS 拒收额,
						                SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
						                SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费 
			                        FROM gk_order_day a
				                        LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
				                        LEFT JOIN dim_area c on c.id = a.area_id
				                        LEFT JOIN dim_cate d on d.id = a.third_cate_id
                            --          LEFT JOIN gk_product e on e.id = a.product_id
                                        LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                    WHERE b.pcode = '{0}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
						                AND c.uname = '罗超源'
						                AND a.beform <> 'mf'
						                AND c.uid <> 10099  -- 过滤翼虎
			                        GROUP BY b.pname, c.uname
                                ) s3)
                            ) s ORDER BY s.订单量'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql40)
            show_name.append(start_Date[i] + '月（父级）成本数据…………')
            # 总成本二级
            sql41 = '''SELECT s1.团队,
                                s1.年月,
                                s1.二级分类,
                                s1.销售额,
                                s1.订单量,
                                (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                s1.改派订单量,
                                s1.改派订单量 / s1.订单量 AS '改派占比',
                                s1.销售额 / s1.订单量 AS '客单价',
                                s1.销售额 / s1.广告成本 AS 'ROI',
                                s1.活跃产品数,
                                s1.订单量 / s1.活跃产品数 AS 产能,
                                s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                s1.广告成本 / s1.销售额 AS '广告占比',
                                s1.物流成本 / s1.销售额 AS '运费占比',
                                s1.手续费 / s1.销售额 AS '手续费占比',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                s1.签收额 / s1.销售额 AS '金额签收/总计',
                                (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                s1.签收量 / s1.订单量 AS '数量签收/总计',
                                (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                        FROM (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.pname AS 二级分类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费  
                                FROM gk_order_day a
                                    LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                    LEFT JOIN dim_area c on c.id = a.area_id
                                    LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        --          LEFT JOIN gk_product e on e.id = a.product_id
                                    LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                WHERE b.pcode = '{0}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                    AND c.uname = '罗超源'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099  -- 过滤翼虎
                                GROUP BY b.pname, c.uname, a.second_cate_id
                                ORDER BY a.product_id
                        ) s1
                        WHERE s1.订单量 > 0
                        ORDER BY s1.订单量;'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql41)
            show_name.append(start_Date[i] + '月（二级）成本数据…………')
            # 总成本三级
            sql42 = '''SELECT s1.团队,
                                s1.年月,
                                s1.三级分类,
                                s1.销售额,
                                s1.订单量,
                                (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                s1.改派订单量,
                                s1.改派订单量 / s1.订单量 AS '改派占比',
                                s1.销售额 / s1.订单量 AS '客单价',
                                s1.销售额 / s1.广告成本 AS 'ROI',
                                s1.活跃产品数,
                                s1.订单量 / s1.活跃产品数 AS 产能,
                                s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                s1.广告成本 / s1.销售额 AS '广告占比',
                                s1.物流成本 / s1.销售额 AS '运费占比',
                                s1.手续费 / s1.销售额 AS '手续费占比',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                s1.签收额 / s1.销售额 AS '金额签收/总计',
                                (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                s1.签收量 / s1.订单量 AS '数量签收/总计',
                                (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                        FROM (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.`name` AS 三级分类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费 
                                FROM gk_order_day a
                                    LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                    LEFT JOIN dim_area c on c.id = a.area_id
                                    LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        --          LEFT JOIN gk_product e on e.id = a.product_id
                                    LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                WHERE b.pcode = '{0}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                    AND c.uname = '罗超源'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099  -- 过滤翼虎
                                GROUP BY b.pname, c.uname, a.third_cate_id
                                ORDER BY a.product_id
                        ) s1
                        WHERE s1.订单量 > 0
                        ORDER BY s1.订单量;'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql42)
            show_name.append(start_Date[i] + '月（三级）成本数据…………')
        listTValue = []  # 查询sql的结果 存放池
        for i, sql in enumerate(listT):
            print(i)
            print('正在获取 ' + tem + show_name[i])
            if i == 0 or i == 4:
                df = pd.read_sql_query(sql=sql, con=self.engine1)
            else:
                df = pd.read_sql_query(sql=sql, con=self.engine2)
            # print(df)
            columns = list(df.columns)  # 获取数据的标题名，转为列表
            columns_value = ['直发占比', '改派占比', '直发采购/总销售额', '广告占比', '总成本占比', '数量签收/总计',  '利润率', '采购/销售额', '直发采购/销售额', '运费占比', '手续费占比', '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比']
            if '旬' in columns:
                df.drop(labels=['旬'], axis=1, inplace=True)  # 去掉多余的旬列表
            for column_val in columns_value:
                if column_val in columns:
                    df[column_val] = df[column_val].fillna(value=0)
                    df[column_val] = df[column_val].apply(lambda x: format(x, '.2%'))
            listTValue.append(df)
        print('查询耗时：', datetime.datetime.now() - start)
        today = datetime.date.today().strftime('%Y.%m.%d')
        sheet_name = ['直发成本', '父级成本', '二级成本', '三级成本']  # 生成的工作表的表名
        if len(listTValue) == 4:
            file_Path = []  # 发送邮箱文件使用
            filePath = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} 火凤凰-{}上月产品花费表.xlsx'.format(today, tem)
            if os.path.exists(filePath):  # 判断是否有需要的表格
                print("正在使用(上月-单月)文件......")
                filePath = filePath
            else:  # 判断是否无需要的表格，进行初始化创建
                print("正在创建文件......")
                df0 = pd.DataFrame([])  # 创建空的dataframe数据框
                df0.to_excel(filePath, sheet_name='缓存使用', index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
                filePath = filePath
            print('正在写入excel…………')
            writer = pd.ExcelWriter(filePath, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(filePath)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listTValue)):
                listTValue[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if '缓存使用' in book.sheetnames:  # 删除新建文档时的第一个工作表
                del book['缓存使用']
            writer.save()
            writer.close()
            print('输出文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePath)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePath, sheet_name[1])
            file_Path.append(filePath)
            self.e.send('火凤凰-' + tem + '产品花费表', file_Path,
                        emailAdd[team])
        else:
            file_Path = []  # 发送邮箱文件使用
            filePath = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} 火凤凰-{}上月产品花费表.xlsx'.format(today, tem)
            if os.path.exists(filePath):  # 判断是否有需要的表格
                print("正在使用(上月)文件......")
                filePath = filePath
            else:  # 判断是否无需要的表格，进行初始化创建
                print("正在创建文件......")
                df0 = pd.DataFrame([])  # 创建空的dataframe数据框
                df0.to_excel(filePath, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
                filePath = filePath
            print('正在写入excel…………')
            writer = pd.ExcelWriter(filePath, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(filePath)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            listTValue[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
            listTValue[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
            listTValue[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
            listTValue[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
            # listTValue[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
            writer.save()
            writer.close()
            print('输出(上月)文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePath)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePath, sheet_name[1])
            file_Path.append(filePath)
            print('------分割线------')
            filePathT = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} 火凤凰-{}本月产品花费表.xlsx'.format(today, tem)
            if os.path.exists(filePathT):  # 判断是否有需要的表格
                print("正在使用(本月)文件......")
                filePathT = filePathT
            else:  # 判断是否无需要的表格，进行初始化创建
                print("正在创建文件......")
                df0T = pd.DataFrame([])  # 创建空的dataframe数据框-2
                df0T.to_excel(filePathT, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）-2
                filePathT = filePathT
            print('正在写入excel…………')
            writerT = pd.ExcelWriter(filePathT, engine='openpyxl')  # 初始化写入对象-2
            bookT = load_workbook(filePathT)  # 可以向不同的sheet写入数据（对现有工作表的追加）-2
            writerT.book = bookT  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet-2
            listTValue[4].to_excel(excel_writer=writerT, sheet_name=sheet_name[0], index=False)
            listTValue[5].to_excel(excel_writer=writerT, sheet_name=sheet_name[1], index=False)
            listTValue[6].to_excel(excel_writer=writerT, sheet_name=sheet_name[2], index=False)
            listTValue[7].to_excel(excel_writer=writerT, sheet_name=sheet_name[3], index=False)
            writerT.save()
            writerT.close()
            print('输出(本月)文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePathT)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePathT, sheet_name[1])
            file_Path.append(filePathT)
            self.e.send('火凤凰-' + tem + '产品花费表', file_Path,
                        emailAdd[team])
            print('处理耗时：', datetime.datetime.now() - start)
    def sl_tem_costT(self, team, tem):
        match = {'slxmt_hfh_zqsb': '"新加坡", "马来西亚", "菲律宾"',
                 'slxmt_t': '"新加坡", "马来西亚", "菲律宾"'}
        match3 = {'新加坡': 'SG',
                  '马来西亚': 'MY',
                  '菲律宾': 'PH'}
        emailAdd = {'slxmt_hfh_zqsb': 'zhangjing@giikin.com',
                    'slxmt_t_zqsb': 'zhangjing@giikin.com'}
        start = datetime.datetime.now()
        endDate = (datetime.datetime.now()).strftime('%Y%m')
        print(endDate)
        startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y%m')
        print(startDate)
        if datetime.datetime.now().day >= 9:
            end_Date = [startDate, endDate]
            start_Date = [startDate, endDate]
        else:
            end_Date = [startDate]
            start_Date = [startDate]
        listT = []  # 查询sql 存放池
        show_name = []  # 打印进度需要
        for i in range(len(end_Date)):
            sql30 = '''SELECT *
                    FROM (
                        SELECT sl_zong.币种,
                            IFNULL(sl_zong.年月,'合计') 年月,
                            IFNULL(sl_zong.父级分类,'合计') 父级分类,
                            IFNULL(sl_zong.二级分类,'合计') 二级分类,
                            IFNULL(sl_zong.三级分类,'合计') 三级分类,
                            IFNULL(sl_zong.产品id,'合计') 产品id,
                            IFNULL(sl_zong.产品名称,'合计') 产品名称,
                            IFNULL(sl_zong.物流方式,'合计') 物流方式,
                            IFNULL(sl_zong.旬,'合计') 旬,
                            SUM(sl_zong.直发订单量) 直发订单量,
                            IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) 签收订单量,
                            IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0) 拒收订单量,
                            SUM(sl_zong.直发成本) / SUM(sl_zong.销售额) AS '直发采购/销售额',
                            SUM(sl_zong.物流运费) / SUM(sl_zong.销售额) AS '运费占比',
                            SUM(sl_zong.手续费) / SUM(sl_zong.销售额) AS '手续费占比',
                            IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) AS '金额签收/完成',
                            IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) / SUM(sl_zong.销售额) AS '金额签收/总计',
                            (IFNULL(SUM(sl_zong_zhifa.`已签收销售额`),0) + IFNULL(SUM(sl_zong_jushou.`拒收销售额`),0)) / SUM(sl_zong.销售额) AS '金额完成占比',
                            IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) / (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) AS '数量签收/完成',
                            (IFNULL(SUM(sl_zong_zhifa.`已签收订单量`),0) + IFNULL(SUM(sl_zong_jushou.`拒收订单量`),0)) / SUM(sl_zong.直发订单量) AS '数量完成占比'
                        FROM (
                            SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 直发订单量,
                                    SUM(`价格RMB`) 销售额,
                                    SUM(`成本价`) 直发成本,
                                    SUM(`物流花费`) 物流运费,
                                    SUM(`打包花费`) 打包花费,
                                    SUM(`其它花费`) 手续费
                            FROM  {0} sl_cx
                            WHERE sl_cx.`币种` = '{1}'
                                AND sl_cx.`年月` >= '{start_Date}'
                                AND sl_cx.`年月` <= '{end_Date}'
                                AND sl_cx.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx.`是否改派` = "直发"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong
                        LEFT JOIN
                            (SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 已签收订单量,
                                    SUM(`价格RMB`) 已签收销售额,
                                    SUM(`成本价`) 已签收成本,
                                    SUM(`物流花费`) 已签收物流运费,
                                    SUM(`打包花费`) 已签收打包花费,
                                    SUM(`其它花费`) 已签收手续费
                            FROM  {0}	sl_cx_zhifa
                            WHERE sl_cx_zhifa.`币种` = '{1}'
                                AND sl_cx_zhifa.`年月` >= '{start_Date}'
                                AND sl_cx_zhifa.`年月` <= '{end_Date}'
                                AND sl_cx_zhifa.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx_zhifa.`是否改派` = "直发"
                                AND sl_cx_zhifa.`最终状态` = "已签收"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong_zhifa
                            ON sl_zong_zhifa.`币种` = sl_zong.`币种`
                                AND sl_zong_zhifa.`年月` = sl_zong.`年月`
                                AND sl_zong_zhifa.`父级分类` = sl_zong.`父级分类`
                                AND sl_zong_zhifa.`二级分类` = sl_zong.`二级分类`
                                AND sl_zong_zhifa.`三级分类` = sl_zong.`三级分类`
                                AND sl_zong_zhifa.`产品id` = sl_zong.`产品id`
                                AND sl_zong_zhifa.`产品名称` = sl_zong.`产品名称`
                                AND sl_zong_zhifa.`物流方式` = sl_zong.`物流方式`
                                AND sl_zong_zhifa.`旬` = sl_zong.`旬`
                        LEFT JOIN
                            (SELECT 币种,
                                    年月,
                                    父级分类,
                                    二级分类,
                                    三级分类,
                                    产品id,
                                    CONCAT(产品id, '#' ,产品名称) 产品名称,
                                    物流方式,
                                    旬,
                                    COUNT(`订单编号`) 拒收订单量,
                                    SUM(`价格RMB`) 拒收销售额,
                                    SUM(`成本价`) 拒收成本,
                                    SUM(`物流花费`) 拒收物流运费,
                                    SUM(`打包花费`) 拒收打包花费,
                                    SUM(`其它花费`) 拒收手续费
                            FROM  {0} sl_cx_jushou
                            WHERE sl_cx_jushou.`币种` = '{1}'
                                AND sl_cx_jushou.`年月` >= '{start_Date}'
                                AND sl_cx_jushou.`年月` <= '{end_Date}'
                                AND sl_cx_jushou.`系统订单状态` IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核') 
                                AND sl_cx_jushou.`是否改派` = "直发"
                                AND sl_cx_jushou.`最终状态` = "拒收"
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,产品名称,物流方式,旬
                            ORDER BY 币种,年月
                            ) sl_zong_jushou
                        ON sl_zong_jushou.`币种` = sl_zong.`币种`
                            AND sl_zong_jushou.`年月` = sl_zong.`年月`
                            AND sl_zong_jushou.`父级分类` = sl_zong.`父级分类`
                            AND sl_zong_jushou.`二级分类` = sl_zong.`二级分类`
                            AND sl_zong_jushou.`三级分类` = sl_zong.`三级分类`
                            AND sl_zong_jushou.`产品id` = sl_zong.`产品id`
                            AND sl_zong_jushou.`产品名称` = sl_zong.`产品名称`
                            AND sl_zong_jushou.`物流方式` = sl_zong.`物流方式`
                            AND sl_zong_jushou.`旬` = sl_zong.`旬`
                    GROUP BY sl_zong.父级分类,sl_zong.二级分类,sl_zong.三级分类,sl_zong.产品名称,sl_zong.物流方式,sl_zong.旬
                    with rollup
                    ) sl_zong_wl
                    WHERE sl_zong_wl.`旬` = '合计';'''.format(team, tem, start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql30)
            show_name.append(start_Date[i] + '月（直发）详细花费数据…………')
            # 总成本父级
            sql40 = '''SELECT *
                        FROM(
                            (SELECT s1.团队,
                                    s1.年月,
                                    s1.品类,
                                    s1.销售额,
                                    s1.订单量,
                                    (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                    (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                    s1.改派订单量,
                                    s1.改派订单量 / s1.订单量 AS '改派占比',
                                    s1.销售额 / s1.订单量 AS '客单价',
                                    s1.销售额 / s1.广告成本 AS 'ROI',
                                    s1.活跃产品数,
                                    s1.订单量 / s1.活跃产品数 AS 产能,
                                    s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                    s1.广告成本 / s1.销售额 AS '广告占比',
                                    s1.物流成本 / s1.销售额 AS '运费占比',
                                    s1.手续费 / s1.销售额 AS '手续费占比',
                                    (s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                    (s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                    s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                    s1.签收额 / s1.销售额 AS '金额签收/总计',
                                    (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                    s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                    s1.签收量 / s1.订单量 AS '数量签收/总计',
                                    (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                    s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                    (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                                FROM (SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.ppname AS 品类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费 
                                    FROM gk_order_day a
                                        LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                        LEFT JOIN dim_area c on c.id = a.area_id
                                        LEFT JOIN dim_cate d on d.id = a.third_cate_id
                            --          LEFT JOIN gk_product e on e.id = a.product_id
                                        LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                    WHERE b.pcode = '{0}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                        AND c.uname = '王冰'
                                        AND a.beform <> 'mf'
                                        AND c.uid <> 10099  -- 过滤翼虎
                                     GROUP BY b.pname, c.uname, a.cate_id
                                    ORDER BY a.product_id
                                ) s1 WHERE s1.订单量 > 0 ORDER BY s1.订单量)
                            UNION ALL
                                (SELECT s3.团队,
						                s3.年月,
						                s3.品类,
						                s3.销售额,
						                s3.订单量,
						                (s3.订单量 - s3.改派订单量) AS 直发订单量,
						                (s3.订单量 - s3.改派订单量) / s3.订单量 AS 直发占比,
						                s3.改派订单量,
						                s3.改派订单量 / s3.订单量 AS 改派占比,
						                s3.销售额 / s3.订单量 客单价,
						                s3.销售额 / s3.广告成本 ROI,
						                S3.活跃产品数,
						                s3.订单量 / S3.活跃产品数 AS 产能,
						                s3.直发采购额 / s3.销售额 AS '直发采购/总销售额',
						                s3.广告成本 / s3.销售额 AS '广告占比',
						                s3.物流成本 / s3.销售额 AS '运费占比',
						                s3.手续费 / s3.销售额 AS '手续费占比',
						                (s3.广告成本 + s3.物流成本 + s3.手续费 + s3.直发采购额 ) AS '总成本',
						                (s3.广告成本 + s3.物流成本 + s3.手续费 + s3.直发采购额 ) / s3.销售额 AS '总成本占比',
						                s3.签收额 / (s3.拒收额 + s3.签收额) '金额签收/完成',
						                s3.签收额 / s3.销售额 '金额签收/总计',
						                (s3.签收额 + s3.拒收额) / s3.销售额 '金额完成占比',
						                s3.签收量 / (s3.拒收量 + s3.签收量) '数量签收/完成',
						                s3.签收量 / s3.订单量 '数量签收/总计',
						                (s3.拒收量 + s3.签收量) / s3.订单量 '数量完成占比',
						                s3.签收额 / (s3.签收额 + s3.拒收额) -( s3.直发采购额 + s3.广告成本 + s3.物流成本 + s3.手续费 ) AS 利润率,
						                (s3.签收额 / (s3.签收额 + s3.拒收额) -( s3.直发采购额 + s3.广告成本 + s3.物流成本 + s3.手续费 ) / s3.销售额) * (s3.销售额 / s3.订单量) AS 利润值
                                FROM (SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
						                b.pname AS 团队,
						                '合计' AS 品类,
						                SUM(a.salesRMB) 销售额,
						                SUM(a.orders) AS 订单量,
						                COUNT(DISTINCT a.product_id) AS 活跃产品数,
						                '' 产能,
						                SUM(a.yqs) AS 签收量,
						                SUM(a.yjs) AS 拒收量,
						                SUM(a.salesRMB_yqs) AS 签收额,
						                SUM(a.salesRMB_yjs) AS 拒收额,
						                SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
						                SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费 
			                        FROM gk_order_day a
				                        LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
				                        LEFT JOIN dim_area c on c.id = a.area_id
				                        LEFT JOIN dim_cate d on d.id = a.third_cate_id
                            --          LEFT JOIN gk_product e on e.id = a.product_id
                                        LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                    WHERE b.pcode = '{0}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                        AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
						                AND c.uname = '王冰'
						                AND a.beform <> 'mf'
						                AND c.uid <> 10099  -- 过滤翼虎
			                        GROUP BY b.pname, c.uname
                                ) s3)
                            ) s ORDER BY s.订单量'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql40)
            show_name.append(start_Date[i] + '月（父级）成本数据…………')
            # 总成本二级
            sql41 = '''SELECT s1.团队,
                                s1.年月,
                                s1.二级分类,
                                s1.销售额,
                                s1.订单量,
                                (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                s1.改派订单量,
                                s1.改派订单量 / s1.订单量 AS '改派占比',
                                s1.销售额 / s1.订单量 AS '客单价',
                                s1.销售额 / s1.广告成本 AS 'ROI',
                                s1.活跃产品数,
                                s1.订单量 / s1.活跃产品数 AS 产能,
                                s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                s1.广告成本 / s1.销售额 AS '广告占比',
                                s1.物流成本 / s1.销售额 AS '运费占比',
                                s1.手续费 / s1.销售额 AS '手续费占比',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                s1.签收额 / s1.销售额 AS '金额签收/总计',
                                (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                s1.签收量 / s1.订单量 AS '数量签收/总计',
                                (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                        FROM (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.pname AS 二级分类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费  
                                FROM gk_order_day a
                                    LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                    LEFT JOIN dim_area c on c.id = a.area_id
                                    LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        --          LEFT JOIN gk_product e on e.id = a.product_id
                                    LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                WHERE b.pcode = '{0}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                    AND c.uname = '王冰'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099  -- 过滤翼虎
                                GROUP BY b.pname, c.uname, a.second_cate_id
                                ORDER BY a.product_id
                        ) s1
                        WHERE s1.订单量 > 0
                        ORDER BY s1.订单量;'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql41)
            show_name.append(start_Date[i] + '月（二级）成本数据…………')
            # 总成本三级
            sql42 = '''SELECT s1.团队,
                                s1.年月,
                                s1.三级分类,
                                s1.销售额,
                                s1.订单量,
                                (s1.订单量 - s1.改派订单量) AS '直发订单量',
                                (s1.订单量 - s1.改派订单量) / s1.订单量 AS '直发占比',
                                s1.改派订单量,
                                s1.改派订单量 / s1.订单量 AS '改派占比',
                                s1.销售额 / s1.订单量 AS '客单价',
                                s1.销售额 / s1.广告成本 AS 'ROI',
                                s1.活跃产品数,
                                s1.订单量 / s1.活跃产品数 AS 产能,
                                s1.直发采购额 / s1.销售额 AS '直发采购/总销售额',
                                s1.广告成本 / s1.销售额 AS '广告占比',
                                s1.物流成本 / s1.销售额 AS '运费占比',
                                s1.手续费 / s1.销售额 AS '手续费占比',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS '总成本',
                                ( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) / s1.销售额 AS '总成本占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) AS '金额签收/完成',
                                s1.签收额 / s1.销售额 AS '金额签收/总计',
                                (s1.签收额 + s1.拒收额) / s1.销售额 AS '金额完成占比',
                                s1.签收量 / (s1.签收量 + s1.拒收量) AS '数量签收/完成',
                                s1.签收量 / s1.订单量 AS '数量签收/总计',
                                (s1.签收量 + s1.拒收量) / s1.订单量 AS '数量完成占比',
                                s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) AS 利润率,
                                (s1.签收额 / (s1.签收额 + s1.拒收额) -( s1.直发采购额 + s1.广告成本 + s1.物流成本 + s1.手续费 ) /s1.销售额) * (s1.销售额 / s1.订单量) AS 利润值
                        FROM (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS 年月,
                                        b.pname AS 团队,
                                        c.uname AS leader,
                                        d.`name` AS 三级分类,
                                        SUM(a.orders) AS 订单量,
                                        COUNT(DISTINCT a.product_id) AS 活跃产品数,
                                        SUM(a.yqs) AS 签收量,
                                        SUM(a.yjs) AS 拒收量,
                                        SUM(a.salesRMB) AS 销售额,
                                        SUM(a.salesRMB_yqs) AS 签收额,
                                        SUM(a.salesRMB_yjs) AS 拒收额,
                                        SUM(a.gps) AS 改派订单量,
                            --          SUM(a.cgcost) AS 总采购额,
                            --          SUM(a.cgcost_zf) AS 直发采购额,
                            --          SUM(a.adcost) AS 广告成本,
                                        null 总采购额,
                                        null 直发采购额,
                                        null 广告成本,
                                        SUM(a.wlcost) AS 物流成本,
                            --          SUM(a.qtcost) AS 手续费 
                                        null 手续费 
                                FROM gk_order_day a
                                    LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                                    LEFT JOIN dim_area c on c.id = a.area_id
                                    LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        --          LEFT JOIN gk_product e on e.id = a.product_id
                                    LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.id = a.product_id
                                WHERE b.pcode = '{0}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) >= '{start_Date}'
                                    AND EXTRACT(YEAR_MONTH FROM a.rq) <= '{end_Date}'
                                    AND c.uname = '王冰'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099  -- 过滤翼虎
                                GROUP BY b.pname, c.uname, a.third_cate_id
                                ORDER BY a.product_id
                        ) s1
                        WHERE s1.订单量 > 0
                        ORDER BY s1.订单量;'''.format(match3[tem], start_Date=start_Date[i], end_Date=end_Date[i])
            listT.append(sql42)
            show_name.append(start_Date[i] + '月（三级）成本数据…………')
        listTValue = []  # 查询sql的结果 存放池
        for i, sql in enumerate(listT):
            print(i)
            print('正在获取 ' + tem + show_name[i])
            if i == 0 or i == 4:
                df = pd.read_sql_query(sql=sql, con=self.engine1)
            else:
                df = pd.read_sql_query(sql=sql, con=self.engine2)
            # print(df)
            columns = list(df.columns)  # 获取数据的标题名，转为列表
            columns_value = ['直发占比', '改派占比', '直发采购/总销售额', '广告占比', '总成本占比', '数量签收/总计',  '利润率', '采购/销售额', '直发采购/销售额', '运费占比', '手续费占比', '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比']
            if '旬' in columns:
                df.drop(labels=['旬'], axis=1, inplace=True)  # 去掉多余的旬列表
            for column_val in columns_value:
                if column_val in columns:
                    df[column_val] = df[column_val].fillna(value=0)
                    df[column_val] = df[column_val].apply(lambda x: format(x, '.2%'))
            listTValue.append(df)
        print('查询耗时：', datetime.datetime.now() - start)
        today = datetime.date.today().strftime('%Y.%m.%d')
        sheet_name = ['直发成本', '父级成本', '二级成本', '三级成本']  # 生成的工作表的表名
        if len(listTValue) == 4:
            file_Path = []  # 发送邮箱文件使用
            filePath = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙T-{}上月产品花费表.xlsx'.format(today, tem)
            if os.path.exists(filePath):  # 判断是否有需要的表格
                print("正在使用(上月-单月)文件......")
                filePath = filePath
            else:  # 判断是否无需要的表格，进行初始化创建
                print("正在创建文件......")
                df0 = pd.DataFrame([])  # 创建空的dataframe数据框
                df0.to_excel(filePath, sheet_name='缓存使用', index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
                filePath = filePath
            print('正在写入excel…………')
            writer = pd.ExcelWriter(filePath, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(filePath)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listTValue)):
                listTValue[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if '缓存使用' in book.sheetnames:  # 删除新建文档时的第一个工作表
                del book['缓存使用']
            writer.save()
            writer.close()
            print('输出文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePath)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePath, sheet_name[1])
            file_Path.append(filePath)
            self.e.send('神龙T-' + tem + '产品花费表', file_Path,
                        emailAdd[team])
        else:
            file_Path = []  # 发送邮箱文件使用
            filePath = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙T-{}上月产品花费表.xlsx'.format(today, tem)
            if os.path.exists(filePath):  # 判断是否有需要的表格
                print("正在使用(上月)文件......")
                filePath = filePath
            else:  # 判断是否无需要的表格，进行初始化创建
                print("正在创建文件......")
                df0 = pd.DataFrame([])  # 创建空的dataframe数据框
                df0.to_excel(filePath, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
                filePath = filePath
            print('正在写入excel…………')
            writer = pd.ExcelWriter(filePath, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(filePath)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            listTValue[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
            listTValue[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
            listTValue[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
            listTValue[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
            # listTValue[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
            writer.save()
            writer.close()
            print('输出(上月)文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePath)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePath, sheet_name[1])
            file_Path.append(filePath)
            print('------分割线------')
            filePathT = 'D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙T-{}本月产品花费表.xlsx'.format(today, tem)
            if os.path.exists(filePathT):  # 判断是否有需要的表格
                print("正在使用(本月)文件......")
                filePathT = filePathT
            else:  # 判断是否无需要的表格，进行初始化创建
                print("正在创建文件......")
                df0T = pd.DataFrame([])  # 创建空的dataframe数据框-2
                df0T.to_excel(filePathT, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）-2
                filePathT = filePathT
            print('正在写入excel…………')
            writerT = pd.ExcelWriter(filePathT, engine='openpyxl')  # 初始化写入对象-2
            bookT = load_workbook(filePathT)  # 可以向不同的sheet写入数据（对现有工作表的追加）-2
            writerT.book = bookT  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet-2
            listTValue[4].to_excel(excel_writer=writerT, sheet_name=sheet_name[0], index=False)
            listTValue[5].to_excel(excel_writer=writerT, sheet_name=sheet_name[1], index=False)
            listTValue[6].to_excel(excel_writer=writerT, sheet_name=sheet_name[2], index=False)
            listTValue[7].to_excel(excel_writer=writerT, sheet_name=sheet_name[3], index=False)
            writerT.save()
            writerT.close()
            print('输出(本月)文件成功…………')
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)           # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(filePathT)
            wbsht.macro('花费运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            # self.xiugaiyangshi(filePathT, sheet_name[1])
            file_Path.append(filePathT)
            self.e.send('神龙T-' + tem + '产品花费表', file_Path,
                        emailAdd[team])
            print('处理耗时：', datetime.datetime.now() - start)

    # 更新团队产品明细（新后台的第一部分）
    def productIdInfo(self, tokenid, searchType, team):  # 进入查询界面，
        print('正在获取需要更新的产品id信息')
        start = datetime.datetime.now()
        month_begin = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y-%m-%d')
        sql = '''SELECT id,`订单编号`  FROM {0}_order_list sl 
    			WHERE sl.`日期`> '{1}' 
    				AND (sl.`产品名称` IS NULL or sl.`父级分类` IS NULL or  sl.`物流方式` IS NULL)
    				AND ( NOT sl.`系统订单状态` IN ('已删除','问题订单','支付失败','未支付'));'''.format(team, month_begin)
        ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
        if ordersDict.empty:
            print('无需要更新的产品id信息！！！')
            # sys.exit()
            return
        orderId = list(ordersDict['订单编号'])
        print('获取耗时：', datetime.datetime.now() - start)
        max_count = len(orderId)    # 使用len()获取列表的长度，上节学的
        n = 0
        while n < max_count:        # 这里用到了一个while循环，穿越过来的
            ord = ', '.join(orderId[n:n + 10])
            print(ord)
            n = n + 10
            self.productIdquery(tokenid, ord, searchType, team)

    def productIdquery(self, tokenid, orderId, searchType, team):  # 进入查询界面，
        start = datetime.datetime.now()
        url = r'http://gimp.giikin.com/service?service=gorder.customer&action=getQueryOrder'
        data = {'phone': None,
                'email': None,
                'ip': None,
                'page': 1,
                'pageSize': 100,
                '_token': tokenid}
        if searchType == '订单号':
            data.update({'orderPrefix': orderId,
                         'shippingNumber': None})
        elif searchType == '运单号':
            data.update({'order_number': None,
                         'shippingNumber': orderId})
        proxy = '39.105.167.0:40005'    # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36',
            'Referer': 'http://gimp.giikin.com/front/orderToolsServiceQuery'}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('已成功发送请求++++++')
        print('正在处理json数据…………')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        print('正在转化数据为dataframe…………')
        # print(req)
        ordersDict = []
        for result in req['data']['list']:
            # print(result)
            # 添加新的字典键-值对，为下面的重新赋值用
            result['productId'] = 0
            result['saleName'] = 0
            result['saleProduct'] = 0
            result['spec'] = 0
            result['link'] = 0
            # print(result['specs'])
            # spe = ''
            # spe2 = ''
            # spe3 = ''
            # spe4 = ''
            # # 产品详细的获取
            # for ind, re in enumerate(result['specs']):
            #     print(ind)
            #     print(re)
            #     print(result['specs'][ind])
            #     spe = spe + ';' + result['specs'][ind]['saleName']
            #     spe2 = spe2 + ';' + result['specs'][ind]['saleProduct']
            #     spe3 = spe3 + ';' + result['specs'][ind]['spec']
            #     spe4 = spe4 + ';' + result['specs'][ind]['link']
            #     spe = spe + ';' + result['specs'][ind]['saleProduct'] + result['specs'][ind]['spec'] + result['specs'][ind]['link'] + result['specs'][ind]['saleName']
            # result['specs'] = spe
            # # del result['specs']             # 删除多余的键值对
            # result['saleName'] = spe
            # result['saleProduct'] = spe2
            # result['spec'] = spe3
            # result['link'] = spe4
            result['saleName'] = result['specs'][0]['saleName']
            result['saleProduct'] = result['specs'][0]['saleProduct']
            result['spec'] = result['specs'][0]['spec']
            result['link'] = result['specs'][0]['link']
            result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
            quest = ''
            for re in result['questionReason']:
                quest = quest + ';' + re
            result['questionReason'] = quest
            delr = ''
            for re in result['delReason']:
                delr = delr + ';' + re
            result['delReason'] = delr
            auto = ''
            for re in result['autoVerify']:
                auto = auto + ';' + re
            result['autoVerify'] = auto
            self.q.put(result)
        # print(len(req['data']['list']))
        for i in range(len(req['data']['list'])):
            ordersDict.append(self.q.get())
        data = pd.json_normalize(ordersDict)
        df = data[['orderNumber', 'wayBillNumber', 'logisticsName', 'logisticsStatus', 'orderStatus', 'isSecondSend',
                   'currency', 'area', 'currency', 'shipInfo.shipPhone', 'quantity', 'productId']]
        print(df)
        print('正在写入缓存中......')
        try:
            df.to_sql('d1_cp', con=self.engine1, index=False, if_exists='replace')
            sql = '''SELECT orderNumber,
                                logisticsName 物流方式,
                                dim_trans_way.simple_name 物流名称,
    					        dim_trans_way.remark 运输方式,
    					        productId,
    					        dp.`name`,
    					        dc.ppname cate,
    					        dc.pname second_cate,
    					        dc.`name` third_cate
    				    FROM d1_cp
    				        LEFT JOIN dim_product dp ON  d1_cp.productId = dp.id
    				        LEFT JOIN dim_cate dc ON  dc.id = dp.third_cate_id
    				        LEFT JOIN dim_trans_way ON  dim_trans_way.all_name = d1_cp.logisticsName;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print(df)
            df.to_sql('tem_product_id', con=self.engine1, index=False, if_exists='replace')
            print('正在更新产品详情…………')
            sql = '''update {0}_order_list a, tem_product_id b
    		                        set a.`物流方式`= IF(b.`物流方式` = '',NULL, b.`物流方式`),
    		                            a.`物流名称`= IF(b.`物流名称` = '',NULL, b.`物流名称`),
    		                            a.`运输方式`= b.`运输方式`,
    		                            a.`产品id`= b.`productId`,
    		                            a.`产品名称`= IF(b.`name` = '',NULL, b.`name`),
    				                    a.`父级分类`= IF(b.`cate` = '',NULL, b.`cate`),
    				                    a.`二级分类`= IF(b.`second_cate` = '',NULL, b.`second_cate`),
    				                    a.`三级分类`= IF(b.`third_cate` = '',NULL, b.`third_cate`)
    				                where a.`订单编号`= b.`orderNumber`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')
        print('更新耗时：', datetime.datetime.now() - start)

    # 更新团队品类明细（新后台的第二部分）
    def cateIdInfo(self, tokenid, team):  # 进入产品检索界面，
        print('正在获取需要更新的品类id信息')
        start = datetime.datetime.now()
        month_begin = (datetime.datetime.now() - relativedelta(months=4)).strftime('%Y-%m-%d')
        sql = '''SELECT id,`订单编号`, `产品id` , `产品名称` ,null 父级分类, null 二级分类, null 三级分类 FROM {0}_order_list sl 
    			WHERE sl.`日期`> '{1}' AND (sl.`父级分类` IS NULL or sl.`父级分类` = '')
    				AND ( NOT sl.`系统订单状态` IN ('已删除','问题订单','支付失败','未支付'));'''.format(team, month_begin)
        ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
        ordersDict.to_sql('d1_cp_cate', con=self.engine1, index=False, if_exists='replace')  # 写入临时品类缓存表中
        if ordersDict.empty:
            print('无需要更新的品类id信息！！！')
            return
        orderId = list(ordersDict['产品id'].drop_duplicates())
        # print(orderId)
        orderId = [str(i) for i in orderId]  # join函数就是字符串的函数,参数和插入的都要是字符串
        print('获取耗时：', datetime.datetime.now() - start)
        max_count = len(orderId)    # 使用len()获取列表的长度，上节学的
        n = 0
        while n < max_count:        # 这里用到了一个while循环，穿越过来的
            cateid = ', '.join(orderId[n:n + 1])
            print(cateid)
            n = n + 1
            self.cateIdquery(tokenid, cateid, team)

    def cateIdquery(self, tokenid, cateid, team):  # 进入产品检索界面，
        start = datetime.datetime.now()
        # productid = '508746'
        # token = '7dd7c0085722cf49493c5ab2ecbc6234'
        proxy = '39.105.167.0:40005'    # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy
                   }
        url = r'http://gimp.giikin.com/service?service=gorder.customer&action=getProductList&page=1&pageSize=90'\
              r'&productName=&status=&source=&isSensitive=&isGift=&isDistribution=&chooserId=&buyerId='\
              r'&productId=' + str(cateid) + '&_token=' + str(tokenid)
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36',
            'Referer': 'http://gimp.giikin.com/front/orderToolsServiceQuery'}
        # rq = requests.get(url=url, headers=r_header, proxies=proxies)
        rq = requests.get(url=url, headers=r_header)
        print('已成功发送请求++++++')
        req = rq.json()  # json类型数据
        print('正在转化数据为dataframe…………')
        # print(req)
        ordersDict = []
        for result in req['data']['list']:
            # print(result)
            result['cate_id'] = 0
            result['second_cate_id'] = 0
            result['third_cate_id'] = 0
            result['cate_id'] = (result['categorys']).split('>')[2]
            result['second_cate_id'] = (result['categorys']).split('>')[1]
            result['third_cate_id'] = (result['categorys']).split('>')[0]
            self.q.put(result)
            # 添加新的字典键-值对，为下面的重新赋值用
        for i in range(len(req['data']['list'])):
            ordersDict.append(self.q.get())
        data = pd.json_normalize(ordersDict)        # 多层结构字典Mixing dicts转化df
        print('正在写入缓存中......')
        data['name'] = data['name'].str.strip()
        data['cate_id'] = data['cate_id'].str.strip()
        data['second_cate_id'] = data['second_cate_id'].str.strip()
        data['third_cate_id'] = data['third_cate_id'].str.strip()
        df = data[['id', 'name', 'categorys', 'cate_id', 'second_cate_id', 'third_cate_id', 'status', 'price', 'createTime']]
        print(df)
        try:
            df.to_sql('d1_cp', con=self.engine1, index=False, if_exists='replace')
            print('正在更新品类缓存中......')
            sql = '''update d1_cp_cate a, d1_cp b
                            set a.`产品名称`= b.`name`,
                                a.`父级分类`= b.`cate_id`,
                                a.`二级分类`= b.`second_cate_id`,
                                a.`三级分类`= b.`third_cate_id`
                    where a.`产品id`=b.`id`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
            print('正在更新总表中......')
            sql = '''update {0}_order_list a, d1_cp_cate b
                            set a.`产品名称`= b.`产品名称`,
                                a.`父级分类`= b.`父级分类`,
                                a.`二级分类`= b.`二级分类`,
                                a.`三级分类`= b.`三级分类`
                    where a.`订单编号`=b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('+++更新成功…………')

if __name__ == '__main__':
    m = QueryControl()
    match1 = {'slgat': '港台',
              'sltg': '泰国',
              'slxmt': '新马',
              'slzb': '直播团队',
              'slyn': '越南',
              'slrb': '日本'}
    # messagebox.showinfo("提示！！！", "当前查询已完成--->>> 请前往（ 输出文件 ）查看")
    #  各团队全部订单表-函数
    # m.tgOrderQuan('sltg')

    # team = 'slgat'
    # for tem in ['台湾', '香港']:
    #     m.OrderQuan(team, tem)

    #  订单花费明细查询
    # match9 = {'slgat_zqsb': '港台',
    #           'sltg_zqsb': '泰国',
    #           'slxmt_zqsb': '新马',
    #           'slrb_zqsb_rb': '日本'}
    # team = 'sltg_zqsb'
    # m.sl_tem_cost(team, match9[team])

    team = 'slrb_js'  # 第一部分查询
    token = '93da2bbc59940c03804d04d30b7e6ce4'
    # m.productIdquery(token, 'NJ210330085757094517', '订单号', team)
    m.cateIdInfo(token, team)
    # m.productIdInfo(token, '订单号', team)

    # team = 'slgat_hfh'  # 第二部分查询
    # token = 'b28f877ee2b82c6bc9253e1b4a676001'
    # m.productIdqueryTT()
    # m.productIdInfoTT(token, '订单号', team)