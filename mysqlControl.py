import pandas as pd
from sqlalchemy import create_engine
from settings import Settings
from queryControl import QueryControl
from emailControl import EmailControl
from bpsControl import BpsControl
from sltemMonitoring import SltemMonitoring
import win32api,win32con
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from tkinter import messagebox
import os
import zipfile
import xlwings as xl
import datetime
from dateutil.relativedelta import relativedelta
import time


class MysqlControl(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                     self.mysql20['password'],
                                                                                     self.mysql20['host'],
                                                                                     self.mysql20['port'],
                                                                                     self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.engine4 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql4['user'],
                                                                                    self.mysql4['password'],
                                                                                    self.mysql4['host'],
                                                                                    self.mysql4['port'],
                                                                                    self.mysql4['datebase']))
        self.e = EmailControl()
        self.d = QueryControl()

    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    def writeSqlReplace(self, dataFrame):
        dataFrame.to_sql('tem', con=self.engine1, index=False, if_exists='replace')

    def replaceInto(self, team, dfColumns):
        columns = list(dfColumns)
        columns = ', '.join(columns)
        if team in ('slrb', 'slsc'):
            print(team + '---9')
            sql = 'REPLACE INTO {}({}, 添加时间) SELECT *, NOW() 添加时间 FROM tem; '.format(team, columns)
        elif team in ('slgat', 'gat'):
            print(team + '---909')  # 当天和前天的添加时间比较，判断是否一样数据
            sql = 'INSERT IGNORE INTO {}({}, 添加时间, 更新时间) SELECT *, CURDATE() 添加时间, NOW() 更新时间 FROM tem; '.format(team, columns)
        else:
            print(team)
            sql = 'INSERT IGNORE INTO {}({}, 添加时间) SELECT *, NOW() 添加时间 FROM tem; '.format(team, columns)
            # sql = 'INSERT IGNORE INTO {}_copy({}, 添加时间) SELECT *, NOW() 添加时间 FROM tem; '.format(team, columns)
        try:
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('插入失败：', str(Exception) + str(e))

    def readSql(self, sql):
        db = pd.read_sql_query(sql=sql, con=self.engine1)
        # db = pd.read_sql(sql=sql, con=self.engine1) or team == 'slgat'
        return db

    def update_gk_product(self):  # 更新产品id的列表
        sql = '''DELETE FROM gat_zqsb
                WHERE gat_zqsb.`订单编号` IN (SELECT 订单编号
											FROM gat_order_list 
											WHERE gat_order_list.`系统订单状态` NOT IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
											);'''
        print('正在清除港澳台-总表的可能删除了的订单…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''DELETE FROM slsc_zqsb
                        WHERE slsc_zqsb.`订单编号` IN (SELECT 订单编号
        											FROM slsc_order_list 
        											WHERE slsc_order_list.`系统订单状态` NOT IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        											);'''
        print('正在清除品牌-总表的可能删除了的订单…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1))
        yy = int((datetime.datetime.now() - datetime.timedelta(days=30)).strftime('%Y'))
        mm = int((datetime.datetime.now() - datetime.timedelta(days=30)).strftime('%m'))
        dd = int((datetime.datetime.now() - datetime.timedelta(days=30)).strftime('%d'))
        begin = datetime.date(yy, mm, dd)
        # begin = datetime.date(2017, 9, 25)
        # begin = datetime.date(2021, 6, 15)
        print(begin)
        yy2 = int(datetime.datetime.now().strftime('%Y'))
        mm2 = int(datetime.datetime.now().strftime('%m'))
        dd2 = int(datetime.datetime.now().strftime('%d'))
        end = datetime.date(yy2, mm2, dd2)
        # end = datetime.date(2021, 6, 17)
        print(end)
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            month_last = str(day)
            sql = '''SELECT id,
				            rq,
                            product_id,
				            product_name,
				            cate_id,
				            second_cate_id,
				            third_cate_id,
				            null seller_id,
				            null selector,
				            null buyer_id,
				            price,
				            gs.`status`,
				            id sale_id
		            FROM gk_sale gs
		            WHERE gs.rq = '{0}';'''.format(month_last)
            print('正在获取 ' + month_last + ' 号以后的产品详情…………')
            df = pd.read_sql_query(sql=sql, con=self.engine2)
            print('正在写入产品缓存中…………')
            df.to_sql('tem_product', con=self.engine1, index=False, if_exists='replace')
            try:
                print('正在更新中…………')
                sql = 'REPLACE INTO dim_product SELECT *, NOW() 更新时间 FROM tem_product; '
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            except Exception as e:
                print('插入失败：', str(Exception) + str(e))
            print('港澳台产品信息更新完成…………')

            # sql = '''SELECT id,
			# 	            rq,
            #                 product_id,
			# 	            product_name,
			# 	            cate_id,
			# 	            second_cate_id,
			# 	            third_cate_id,
			# 	            null seller_id,
			# 	            null selector,
			# 	            null buyer_id,
			# 	            price,
			# 	            gs.`status`,
			# 	            id sale_id
            # 		    FROM gk_sale gs
            # 		    WHERE gs.rq = '{0}';'''.format(month_last)
            # print('正在获取 ' + month_last + ' 号以后的产品详情…………')
            # df = pd.read_sql_query(sql=sql, con=self.engine20)
            # print('正在写入产品缓存中…………')
            # df.to_sql('tem_product', con=self.engine1, index=False, if_exists='replace')
            # try:
            #     print('正在更新中…………')
            #     sql = 'REPLACE INTO dim_product SELECT *, NOW() 更新时间 FROM tem_product; '
            #     pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            # except Exception as e:
            #     print('插入失败：', str(Exception) + str(e))
            # print('日本产品信息更新完成…………')
        # try:
        #     print('正在更新中…………')
        #     sql = 'REPLACE INTO dim_product_slsc SELECT *  FROM dim_product WHERE id IN (SELECT MAX(id) FROM dim_product  GROUP BY product_id);'
        #     pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        # except Exception as e:
        #     print('插入失败：', str(Exception) + str(e))
        # print('商城信息更新完成…………')

        print('正在获取物流信息中…………')
        try:
            sql = '''SELECT * FROM dim_trans_way gs;'''
            df = pd.read_sql_query(sql=sql, con=self.engine2)
            df.to_sql('tem_product', con=self.engine1, index=False, if_exists='replace')
            print('正在更新中…………')
            sql = 'REPLACE INTO dim_trans_way SELECT * FROM tem_product; '
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('插入失败：', str(Exception) + str(e))
        print('物流信息更新完成…………')

    def update_gk_sign_rate(self):  # 更新产品签收率列表
        print('正在获取产品签收率的历史信息…………')
        # sql = '''SELECT * FROM  gk_stat_sign_rate;'''
        sql = '''SELECT a.id,
                        goods_id,
                        dim_area.`name` AS area_id,
                        dim_currency_lang.pname AS currency_id,
                        cod_order_count,
                        cod_order_sign_count,
                        cod_sign_rate,
                        online_order_count,
                        online_order_sign_count,
                        online_sign_rate,
                        avg_sign_rate,
                        sign_rate_last_month,
                        finish_rate,
                        data_rq,
                        update_time
                FROM  gk_stat_sign_rate a
                    left join dim_area ON dim_area.id = a.area_id
                    left join dim_currency_lang ON dim_currency_lang.id = a.currency_id
                WHERE dim_area.name IN ("神龙家族-港澳台", "火凤凰-港澳台", "红杉家族-港澳台", "红杉家族-港澳台2", "金狮-港澳台", "金鹏家族-小虎队", "火凤凰-港台(繁体)", "神龙-低价")'''
        df = pd.read_sql_query(sql=sql, con=self.engine2)
        # df['data_rq'] = df['data_rq'].fillna(value=datetime.datetime(1990, 1, 1, 0, 0))
        df['data_rq'] = df['data_rq'].replace(to_replace='', value=datetime.datetime(1990, 1, 1, 0, 0))
        df.to_sql('gk_stat_cache', con=self.engine1, index=False, if_exists='replace')
        try:
            print('正在更新中…………')
            sql = '''REPLACE INTO gk_stat_sign_rate SELECT *, NOW() 获取时间 FROM gk_stat_cache;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('插入失败：', str(Exception) + str(e))
        print('历史产品签收率更新完成…………')

        print('正在获取产品签收率的预测信息…………')
        # sql = '''SELECT * FROM  gk_bi_estimate_goods;'''
        sql = '''SELECT a.id,
			            dim_area.`name` AS area_id,
			            dim_currency_lang.pname AS currency_id,
			            goods_id,
			            sign_rate,
			            create_time,
			            update_time
                FROM  gk_bi_estimate_goods a
		            left join dim_area ON dim_area.id = a.area_id
		            left join dim_currency_lang ON dim_currency_lang.id = a.currency_id
                WHERE dim_area.name IN ("神龙家族-港澳台", "火凤凰-港澳台", "红杉家族-港澳台", "红杉家族-港澳台2", "金狮-港澳台", "金鹏家族-小虎队", "火凤凰-港台(繁体)", "神龙-低价");'''
        df = pd.read_sql_query(sql=sql, con=self.engine2)
        df.to_sql('gk_stat_cache', con=self.engine1, index=False, if_exists='replace')
        try:
            print('正在更新中…………')
            sql = '''REPLACE INTO gk_bi_estimate_goods SELECT *, NOW() 获取时间 FROM gk_stat_cache;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('插入失败：', str(Exception) + str(e))
        print('预测产品签收率更新完成…………')



    def creatMyOrderSl(self, team):  # 最近五天的全部订单信息
        match = {'gat': '"神龙家族-港澳台", "火凤凰-港澳台", "红杉家族-港澳台", "红杉家族-港澳台2", "金狮-港澳台", "金鹏家族-小虎队", "火凤凰-港台(繁体)", "神龙-低价", "神龙-主页运营1组", "神龙-运营1组", "神龙-主页运营"',
                 'slsc': '"金鹏家族-品牌", "金鹏家族-品牌1组", "金鹏家族-品牌2组", "金鹏家族-品牌3组"',
                 'sl_rb': '"神龙家族-日本团队", "金狮-日本", "红杉家族-日本", "红杉家族-日本666", "精灵家族-日本", "精灵家族-韩国", "精灵家族-品牌", "火凤凰-日本", "金牛家族-日本", "金鹏家族-小虎队", "奎蛇-日本", "奎蛇-韩国", "神龙-韩国"'
                 }
        # 12-1月的
        if team in ('slsc', 'gat', 'sl_rb'):
            # 获取日期时间
            sql = 'SELECT MAX(`日期`) 日期 FROM {0}_order_list;'.format(team)
            rq = pd.read_sql_query(sql=sql, con=self.engine1)
            rq = pd.to_datetime(rq['日期'][0])
            yy = int((rq - datetime.timedelta(days=4)).strftime('%Y'))
            mm = int((rq - datetime.timedelta(days=4)).strftime('%m'))
            dd = int((rq - datetime.timedelta(days=4)).strftime('%d'))
            # print(dd)
            begin = datetime.date(yy, mm, dd)
            print(begin)
            yy2 = int(datetime.datetime.now().strftime('%Y'))
            mm2 = int(datetime.datetime.now().strftime('%m'))
            dd2 = int(datetime.datetime.now().strftime('%d'))
            end = datetime.date(yy2, mm2, dd2)
            print(end)
        else:
            # 11-12月的
            begin = datetime.date(2022, 1, 1)
            print(begin)
            end = datetime.date(2022, 2, 22)
            print(end)
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            # print(str(day))
            yesterday = str(day) + ' 23:59:59'
            last_month = str(day)
            if team == 'slsc':
                sql = '''SELECT a.id,
                                        a.month 年月,
                                        a.month_mid 旬,
                                        a.rq 日期,
                                        dim_area.name 团队,
                                        a.region_code 区域,
                                        dim_currency_lang.pname 币种,
                                        a.beform 订单来源,
                                        a.order_number 订单编号,
                                        a.qty 数量,
                                        a.ship_phone 电话号码,
                                        UPPER(a.waybill_number) 运单编号,
                                        a.order_status 系统订单状态id,
                                        IF(a.logistics_status = 1, 0, a.logistics_status) 系统物流状态id,
                                        IF(a.second=0,'直发','改派') 是否改派,
                                        dim_trans_way.all_name 物流方式,
                                        dim_trans_way.simple_name 物流名称,
                                        dim_trans_way.remark 运输方式,
                                        a.logistics_type 货物类型,
                                        IF(a.low_price=0,'否','是') 是否低价,
                                        a.sale_id 商品id,
                                        a.product_id 产品id,
                         		        gk_sale.product_name 产品名称,
                                        dim_cate.ppname 父级分类,
                                        dim_cate.pname 二级分类,
                                        dim_cate.name 三级分类,
                                        dim_payment.pay_name 付款方式,
                                        a.amount 价格,
                                        a.addtime 下单时间,
                                        a.verity_time 审核时间,
                                        a.delivery_time 仓储扫描时间,
                                        a.online_time 上线时间,
                                        a.finish_status 完结状态,
                                        a.endtime 完结状态时间,
                                        a.salesRMB 价格RMB,
                                        null 价格区间,
                                        null 成本价,
                                        a.logistics_cost 物流花费,
                                        null 打包花费,
                                        a.other_fee 其它花费,
                                        a.weight 包裹重量,
                                        a.volume 包裹体积,
                                        a.ship_zip 邮编,
                                        a.turn_purchase_time 添加物流单号时间,
                                        null 省洲,
                                        a.del_reason 订单删除原因,
                                        IF(dim_area.name = '精灵家族-品牌',IF(a.coll_id=1000000269,'饰品','内衣'),a.coll_id) 站点ID
                                FROM gk_order a
                                        left join dim_area ON dim_area.id = a.area_id
                                        left join dim_payment ON dim_payment.id = a.payment_id
            	                        LEFT JOIN gk_sale ON gk_sale.id = a.sale_id
                                        left join dim_trans_way ON dim_trans_way.id = a.logistics_id
                                        left join dim_cate ON dim_cate.id = a.third_cate_id
                                        left join dim_currency_lang ON dim_currency_lang.id = a.currency_lang_id
                                WHERE  a.rq = '{0}' AND a.rq <= '{1}'
                                    AND dim_area.name IN ({2});'''.format(last_month, yesterday, match[team])
                print('正在获取 ' + match[team] + last_month[5:7] + '-' + yesterday[8:10] + ' 号订单…………')
                df = pd.read_sql_query(sql=sql, con=self.engine4)
            elif team == 'sl_rb':
                sql = '''SELECT a.id,
                                            a.month 年月,
                                            a.month_mid 旬,
                                            a.rq 日期,
                                            dim_area.name 团队,
                                            a.region_code 区域,
                                            dim_currency_lang.pname 币种,
                                            a.beform 订单来源,
                                            a.order_number 订单编号,
                                            a.qty 数量,
                                            a.ship_phone 电话号码,
                                            UPPER(a.waybill_number) 运单编号,
                                            a.order_status 系统订单状态id,
                                            IF(a.logistics_status = 1, 0, a.logistics_status) 系统物流状态id,
                                            IF(a.second=0,'直发','改派') 是否改派,
                                            dim_trans_way.all_name 物流方式,
                                            dim_trans_way.simple_name 物流名称,
                                            dim_trans_way.remark 运输方式,
                                            a.logistics_type 货物类型,
                                            IF(a.low_price=0,'否','是') 是否低价,
                                            a.sale_id 商品id,
                                            a.product_id 产品id,
                             		        gk_sale.product_name 产品名称,
                                            dim_cate.ppname 父级分类,
                                            dim_cate.pname 二级分类,
                                            dim_cate.name 三级分类,
                                            dim_payment.pay_name 付款方式,
                                            a.amount 价格,
                                            a.addtime 下单时间,
                                            a.verity_time 审核时间,
                                            a.delivery_time 仓储扫描时间,
                                            a.online_time 上线时间,
                                            a.finish_status 完结状态,
                                            a.endtime 完结状态时间,
                                            a.salesRMB 价格RMB,
                                            null 价格区间,
                                            null 成本价,
                                            a.logistics_cost 物流花费,
                                            null 打包花费,
                                            a.other_fee 其它花费,
                                            a.weight 包裹重量,
                                            a.volume 包裹体积,
                                            a.ship_zip 邮编,
                                            a.turn_purchase_time 添加物流单号时间,
                                            null 省洲,
                                            a.del_reason 订单删除原因,
                                            IF(dim_area.name = '精灵家族-品牌',IF(a.coll_id=1000000269,'饰品','内衣'),a.coll_id) 站点ID
                                    FROM gk_order a
                                            left join dim_area ON dim_area.id = a.area_id
                                            left join dim_payment ON dim_payment.id = a.payment_id
                	                        LEFT JOIN gk_sale ON gk_sale.id = a.sale_id
                                            left join dim_trans_way ON dim_trans_way.id = a.logistics_id
                                            left join dim_cate ON dim_cate.id = a.third_cate_id
                                            left join dim_currency_lang ON dim_currency_lang.id = a.currency_lang_id
                                    WHERE  a.rq = '{0}' AND a.rq <= '{1}'
                                        AND dim_area.name IN ({2});'''.format(last_month, yesterday, match[team])
                print('正在获取 ' + match[team] + last_month[5:7] + '-' + yesterday[8:10] + ' 号订单…………')
                df = pd.read_sql_query(sql=sql, con=self.engine20)
            elif team == 'gat':
                sql = '''SELECT a.id,
                            a.month 年月,
                            a.month_mid 旬,
                            a.rq 日期,
                            IF(dim_area.name LIKE "红杉家族-港澳台2%","红杉家族-港澳台",dim_area.name) 团队,
                            a.region_code 区域,
                            dim_currency_lang.pname 币种,
                            a.beform 订单来源,
                            a.order_number 订单编号,
                            a.qty 数量,
                            a.ship_phone 电话号码,
                            UPPER(a.waybill_number) 运单编号,
                            a.order_status 系统订单状态id,
                            IF(a.logistics_status = 1, 0, a.logistics_status) 系统物流状态id,
                            IF(a.second=0,'直发','改派') 是否改派,
                            dim_trans_way.all_name 物流方式,
                            dim_trans_way.simple_name 物流名称,
                            dim_trans_way.remark 运输方式,
                            a.logistics_type 货物类型,
                            IF(a.low_price=0,'否','是') 是否低价,
                            a.sale_id 商品id,
                            a.product_id 产品id,
             		        gk_sale.product_name 产品名称,
                            dim_cate.ppname 父级分类,
                            dim_cate.pname 二级分类,
                            dim_cate.name 三级分类,
                            dim_payment.pay_name 付款方式,
                            a.amount 价格,
                            a.addtime 下单时间,
                            a.verity_time 审核时间,
                            a.delivery_time 仓储扫描时间,
                            IF(a.finish_status=0,'未收款',IF(a.finish_status=2,'收款',IF(a.finish_status=4,'退款',a.finish_status))) 完结状态,
                            a.endtime 完结状态时间,   
                            a.salesRMB 价格RMB,
                            intervals.intervals 价格区间,
                            null 成本价,
                            a.logistics_cost 物流花费,
                            null 打包花费,
                            a.other_fee 其它花费,
                            a.weight 包裹重量,
                            null 包裹体积,
                            a.ship_zip 邮编,
                            a.turn_purchase_time 添加物流单号时间,
                            null 规格中文,
                            a.ship_state 省洲,
                            null 审单类型,
                            a.del_reason 删除原因,
                            null 删除时间,
                            a.question_reason 问题原因,
                            null 问题时间,
                            null 下单人,
                            null 克隆人,
                            a.stock_type 下架类型,
                            a.lower_time 下架时间,
                            a.tihuo_time 物流提货时间,
                            a.fahuo_time 物流发货时间,
                            a.online_time 上线时间,
                            a.guonei_time 国内清关时间,
                            a.mudidi_time 目的清关时间,
                            a.receipt_time 回款时间,
                            a.ip IP,
                            null 选品人
                    FROM gk_order a
                            left join dim_area ON dim_area.id = a.area_id
                            left join dim_payment ON dim_payment.id = a.payment_id
	                        LEFT JOIN gk_sale ON gk_sale.id = a.sale_id
                            left join dim_trans_way ON dim_trans_way.id = a.logistics_id
                            left join dim_cate ON dim_cate.id = a.third_cate_id
                            left join intervals ON intervals.id = a.intervals
                            left join dim_currency_lang ON dim_currency_lang.id = a.currency_lang_id
                    WHERE  a.rq = '{0}' AND a.rq <= '{1}' AND dim_area.name IN ({2});'''.format(last_month, yesterday, match[team])
                print('正在获取 ' + match[team] + last_month[5:7] + '-' + yesterday[8:10] + ' 号订单…………')
                df = pd.read_sql_query(sql=sql, con=self.engine2)
            sql = 'SELECT * FROM dim_order_status;'
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            print('+++合并订单状态中…………')
            df = pd.merge(left=df, right=df1, left_on='系统订单状态id', right_on='id', how='left')
            sql = 'SELECT * FROM dim_logistics_status;'
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            print('+++合并物流状态中…………')
            df = pd.merge(left=df, right=df1, left_on='系统物流状态id', right_on='id', how='left')
            df = df.drop(labels=['id', 'id_y', '系统订单状态id', '系统物流状态id'], axis=1)
            df.rename(columns={'id_x': 'id', 'name_x': '系统订单状态', 'name_y': '系统物流状态'}, inplace=True)
            print('++++++正在将 ' + yesterday[8:10] + ' 号订单写入数据库++++++')
            # 这一句会报错,需要修改my.ini文件中的[mysqld]段中的"max_allowed_packet = 1024M"
            try:
                df.to_sql('sl_order', con=self.engine1, index=False, if_exists='replace')
                sql = 'REPLACE INTO {}_order_list SELECT *, NOW() 记录时间 FROM sl_order; '.format(team)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            except Exception as e:
                print('插入失败：', str(Exception) + str(e))
            print('写入完成…………')
        return '写入完成'

    def creatMyOrderSlTWO(self, team, begin, end):  # 最近两个月的更新订单信息
        match = {'gat': '"神龙家族-港澳台", "火凤凰-港澳台", "红杉家族-港澳台", "红杉家族-港澳台2", "金狮-港澳台", "金鹏家族-小虎队", "火凤凰-港台(繁体)", "神龙-低价", "神龙-主页运营1组", "神龙-运营1组", "神龙-主页运营"',
                 'slsc': '"金鹏家族-品牌", "金鹏家族-品牌1组", "金鹏家族-品牌2组", "金鹏家族-品牌3组"',
                 'sl_rb': '"神龙家族-日本团队", "金狮-日本", "红杉家族-日本", "红杉家族-日本666", "精灵家族-日本", "精灵家族-韩国", "精灵家族-品牌", "火凤凰-日本", "金牛家族-日本", "金鹏家族-小虎队", "奎蛇-日本", "奎蛇-韩国", "神龙-韩国"',
                 }
        today = datetime.date.today().strftime('%Y.%m.%d')
        # if team in ('sltg', 'slsc', 'slrb', 'slrb_jl', 'slrb_js', 'slrb_hs', 'gat', 'slgat', 'slgat_hfh', 'slgat_hs', 'slxmt', 'slxmt_t', 'slxmt_hfh'):
        #     yy = int((datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y'))
        #     mm = int((datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%m'))
        #     begin = datetime.date(yy, mm, 1)
        #     print(begin)
        #     yy2 = int(datetime.datetime.now().strftime('%Y'))
        #     mm2 = int(datetime.datetime.now().strftime('%m'))
        #     dd2 = int(datetime.datetime.now().strftime('%d'))
        #     end = datetime.date(yy2, mm2, dd2)
        #     print(end)
        # else:
        #     begin = datetime.date(2021, 5, 1)
        #     print(begin)
        #     end = datetime.date(2021, 7, 10)
        #     print(end)
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            # print(str(day))
            yesterday = str(day) + ' 23:59:59'
            last_month = str(day)
            print('正在更新 ' + match[team] + last_month[5:7] + '-' + yesterday[8:10] + ' 号订单信息…………')
            if team == 'sl_rb':
                sql = '''SELECT DISTINCT a.id,
                                a.rq 日期,
                                dim_currency_lang.pname 币种,
                                a.order_number 订单编号,
                                a.qty 数量,
                                a.ship_phone 电话号码,
                                UPPER(a.waybill_number) 运单编号,
                                a.order_status 系统订单状态id,
                                IF(a.logistics_status = 1, 0, a.logistics_status) 系统物流状态id,
                                IF(a.second=0,'直发','改派') 是否改派,
                                dim_trans_way.all_name 物流方式,
                                dim_trans_way.simple_name 物流名称,
                                a.sale_id 商品id,
                                a.product_id 产品id,
             		            gk_sale.product_name 产品名称,
                                dim_cate.ppname 父级分类,
                                dim_cate.pname 二级分类,
                                dim_cate.name 三级分类,
                                dim_payment.pay_name 付款方式,
                                a.amount 价格,
                                a.logistics_type 货物类型,
                                a.verity_time 审核时间,
                                a.delivery_time 仓储扫描时间,
                                a.online_time 上线时间,
                                a.finish_status 完结状态,
                                a.endtime 完结状态时间,
                                a.salesRMB 价格RMB,
                                a.logistics_cost 物流花费,
                                a.weight 包裹重量,
                                null 省洲
                        FROM gk_order a
                                left join dim_area ON dim_area.id = a.area_id
                                left join dim_payment on dim_payment.id = a.payment_id
                                LEFT JOIN gk_sale ON gk_sale.id = a.sale_id
                 		--	    left join (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) gs ON gs.product_id = a.product_id
                                left join dim_trans_way on dim_trans_way.id = a.logistics_id
                                left join dim_cate on dim_cate.id = a.third_cate_id
                        --      left join intervals on intervals.id = a.intervals
                                left join dim_currency_lang on dim_currency_lang.id = a.currency_lang_id
                        WHERE a.rq = '{0}' AND a.rq <= '{1}'
                            AND dim_area.name IN ({2});'''.format(last_month, yesterday, match[team])
                df = pd.read_sql_query(sql=sql, con=self.engine20)
            elif team == 'slsc':
                sql = '''SELECT DISTINCT a.id,
                                a.rq 日期,
                                dim_currency_lang.pname 币种,
                                a.order_number 订单编号,
                                a.qty 数量,
                                a.ship_phone 电话号码,
                                UPPER(a.waybill_number) 运单编号,
                                a.order_status 系统订单状态id,
                                IF(a.logistics_status = 1, 0, a.logistics_status) 系统物流状态id,
                                IF(a.second=0,'直发','改派') 是否改派,
                                dim_trans_way.all_name 物流方式,
                                dim_trans_way.simple_name 物流名称,
                                a.sale_id 商品id,
                                a.product_id 产品id,
             		            gk_sale.product_name 产品名称,
                                dim_cate.ppname 父级分类,
                                dim_cate.pname 二级分类,
                                dim_cate.name 三级分类,
                                dim_payment.pay_name 付款方式,
                                a.amount 价格,
                                a.logistics_type 货物类型,
                                a.verity_time 审核时间,
                                a.delivery_time 仓储扫描时间,
                                a.online_time 上线时间,
                                a.finish_status 完结状态,
                                a.endtime 完结状态时间,
                                a.salesRMB 价格RMB,
                                a.logistics_cost 物流花费,
                                a.weight 包裹重量,
                                null 省洲
                        FROM gk_order a
                                left join dim_area ON dim_area.id = a.area_id
                                left join dim_payment on dim_payment.id = a.methods
                                LEFT JOIN gk_sale ON gk_sale.id = a.sale_id
                 		--	    left join (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) gs ON gs.product_id = a.product_id
                                left join dim_trans_way on dim_trans_way.id = a.logistics_id
                                left join dim_cate on dim_cate.id = a.third_cate_id
                        --      left join intervals on intervals.id = a.intervals
                                left join dim_currency_lang on dim_currency_lang.id = a.currency_lang_id
                        WHERE a.rq = '{0}' AND a.rq <= '{1}'
                            AND dim_area.name IN ({2});'''.format(last_month, yesterday, match[team])
                df = pd.read_sql_query(sql=sql, con=self.engine4)
            elif team == 'gat':
                sql = '''SELECT a.id,
                                            a.rq 日期,
                                            dim_currency_lang.pname 币种,
                                            a.order_number 订单编号,
                                            a.qty 数量,
                                            a.ship_phone 电话号码,
                                            UPPER(a.waybill_number) 运单编号,
                                            a.order_status 系统订单状态id,
                                            IF(a.logistics_status = 1, 0, a.logistics_status) 系统物流状态id,
                                            IF(a.second=0,'直发','改派') 是否改派,
                                            dim_trans_way.all_name 物流方式,
                                            dim_trans_way.simple_name 物流名称,
                                            a.logistics_type 货物类型,
                                            a.sale_id 商品id,
                                            a.product_id 产品id,
                             		        gk_sale.product_name 产品名称,
                                            a.amount 价格,
                                            a.verity_time 审核时间,
                                            a.delivery_time 仓储扫描时间,
                                            IF(a.finish_status=0,'未收款',IF(a.finish_status=2,'收款',IF(a.finish_status=4,'退款',a.finish_status))) 完结状态,
                                            a.endtime 完结状态时间,   
                                            a.salesRMB 价格RMB,
                                            a.logistics_cost 物流花费,
                                            a.other_fee 其它花费,
                                            a.weight 包裹重量,
                                            a.turn_purchase_time 添加物流单号时间,
                                            a.del_reason 删除原因,
                                            a.question_reason 问题原因,
                                            a.stock_type 下架类型,
                                            a.lower_time 下架时间,
                                            a.tihuo_time 物流提货时间,
                                            a.fahuo_time 物流发货时间,
                                            a.online_time 上线时间,
                                            a.guonei_time 国内清关时间,
                                            a.mudidi_time 目的清关时间,
                                            a.receipt_time 回款时间,
                                            a.ip IP
                                    FROM gk_order a
                                            left join dim_area ON dim_area.id = a.area_id
                                            left join dim_payment ON dim_payment.id = a.payment_id
                	                        LEFT JOIN gk_sale ON gk_sale.id = a.sale_id
                                            left join dim_trans_way ON dim_trans_way.id = a.logistics_id
                                            left join dim_cate ON dim_cate.id = a.third_cate_id
                                            left join intervals ON intervals.id = a.intervals
                                            left join dim_currency_lang ON dim_currency_lang.id = a.currency_lang_id
                                    WHERE  a.rq = '{0}' AND a.rq <= '{1}' AND dim_area.name IN ({2});'''.format(last_month, yesterday, match[team])
                df = pd.read_sql_query(sql=sql, con=self.engine2)
            sql = 'SELECT * FROM dim_order_status;'
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            print('++++更新订单状态中…………')
            df = pd.merge(left=df, right=df1, left_on='系统订单状态id', right_on='id', how='left')
            sql = 'SELECT * FROM dim_logistics_status;'
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            print('++++更新物流状态中…………')
            df = pd.merge(left=df, right=df1, left_on='系统物流状态id', right_on='id', how='left')
            df = df.drop(labels=['id', 'id_y', '系统订单状态id', '系统物流状态id'], axis=1)
            df.rename(columns={'id_x': 'id', 'name_x': '系统订单状态', 'name_y': '系统物流状态'}, inplace=True)
            print('+++++++正在将 ' + yesterday[8:10] + ' 号订单更新到数据库++++++')
            # 这一句会报错,需要修改my.ini文件中的[mysqld]段中的"max_allowed_packet = 1024M"
            df.to_sql('sl_order2', con=self.engine1, index=False, if_exists='replace')
            try:
                if team == 'gat':
                    sql = '''update {0}_order_list a, sl_order2 b
                        set a.`币种`=b.`币种`,
                            a.`数量`=b.`数量`,
		                    a.`电话号码`=b.`电话号码` ,
		                    a.`运单编号`=b.`运单编号`,
		                    a.`系统订单状态`=b.`系统订单状态`,
		                    a.`系统物流状态`=b.`系统物流状态`,
		                    a.`是否改派`=b.`是否改派`,
		                    a.`物流方式`=b.`物流方式`,
		                    a.`物流名称`=b.`物流名称`,
		                    a.`商品id`=b.`商品id`,
		                    a.`产品id`=b.`产品id`,
		                    a.`产品名称`=b.`产品名称`,
		                    a.`价格`=b.`价格`,
		                    a.`审核时间`=b.`审核时间`,
		                    a.`仓储扫描时间`=b.`仓储扫描时间`,
		                    a.`完结状态`=b.`完结状态`,
		                    a.`完结状态时间`=b.`完结状态时间`,
		                    a.`价格RMB`=b.`价格RMB`,
		                    a.`物流花费`=b.`物流花费`,
		                    a.`包裹重量`=b.`包裹重量`,
		                    a.`添加物流单号时间`=b.`添加物流单号时间`,
		                    a.`下架类型`=b.`下架类型`,
		                    a.`下架时间`=b.`下架时间`,
		                    a.`物流提货时间`=b.`物流提货时间`,
		                    a.`物流发货时间`=b.`物流发货时间`,
		                    a.`上线时间`=b.`上线时间`,
		                    a.`国内清关时间`=b.`国内清关时间`,
		                    a.`目的清关时间`=b.`目的清关时间`,
		                    a.`回款时间`=b.`回款时间`
		                where a.`订单编号`=b.`订单编号`;'''.format(team)
                elif team in ('slsc', 'sl_rb'):
                    sql = '''update {0}_order_list a, sl_order2 b
                        set a.`币种`=b.`币种`,
                            a.`数量`=b.`数量`,
		                    a.`电话号码`=b.`电话号码` ,
		                    a.`运单编号`=b.`运单编号`,
		                    a.`系统订单状态`=b.`系统订单状态`,
		                    a.`系统物流状态`=b.`系统物流状态`,
		                    a.`是否改派`=b.`是否改派`,
		                    a.`物流方式`=b.`物流方式`,
		                    a.`物流名称`=b.`物流名称`,
		                    a.`商品id`=b.`商品id`,
		                    a.`产品id`=b.`产品id`,
		                    a.`产品名称`=b.`产品名称`,
		                    a.`价格`=b.`价格`,
		                    a.`审核时间`=b.`审核时间`,
		                    a.`仓储扫描时间`=b.`仓储扫描时间`,
		                    a.`完结状态`=b.`完结状态`,
		                    a.`完结状态时间`=b.`完结状态时间`,
		                    a.`价格RMB`=b.`价格RMB`,
		                    a.`物流花费`=b.`物流花费`,
		                    a.`包裹重量`=b.`包裹重量`,
		                    a.`上线时间`=b.`上线时间`
		                where a.`订单编号`=b.`订单编号`;'''.format(team)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            except Exception as e:
                print('插入失败：', str(Exception) + str(e))
            print('----更新完成----')
        return '更新完成'

    def connectOrder(self, team, month_last, month_yesterday, month_begin):
        match = {'slgat': '神龙-港台',
                 'slgat_hfh': '火凤凰-港台',
                 'slgat_hs': '红杉-港台',
                 'slgat_js': '金狮-港台',
                 'slgat_jp': '小虎队-港台',
                 'slgat_run': '神龙-主页运营1组',
                 'gat': '港台',
                 'slsc': '品牌',
                 'sl_rb': '日本',
                 'slrb': '神龙-日本',
                 'slrb_jl': '精灵-日本',
                 'slrb_js': '金狮-日本',
                 'slrb_hs': '红杉-日本',
                 'slrb_hfh': '火凤凰-日本',
                 'slrb_jn': '金牛家族-日本',
                 'slrb_xhd': '金鹏家族-小虎队',
                 'slrb_ks': '奎蛇-日本',
                 'slrb_ks_hg': '奎蛇-韩国',
                 'slrb_sl': '神龙-韩国'}
        emailAdd = {'slgat': 'giikinliujun@163.com',
                    'slgat_hfh': 'giikinliujun@163.com',
                    'slgat_hs': 'giikinliujun@163.com',
                    'slsc': 'sunyaru@giikin.com'}
        # if team in ('slsc', 'slrb', 'slrb_jl', 'slrb_js', 'slrb_hs', 'gat0', 'slgat', 'slgat_hfh', 'slgat_hs'):
        #     month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        #     month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
        #     month_begin = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y-%m-%d')
        #     print(month_begin)
        # else:
        #     month_last = '2021-06-01'
        #     month_yesterday = '2021-07-31'
        #     month_begin = '2021-02-01'
        print('正在检查父级分类为空的信息---')
        # sql = '''SELECT 订单编号,商品id,
		# 		        dp.product_id, dp.`name` product_name, dp.third_cate_id,
        #                 dc.ppname cate, dc.pname second_cate, dc.`name` third_cate
        #         FROM (SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
        #             FROM {0}_order_list sl
        #             WHERE sl.`日期`> '{1}' AND (sl.`父级分类` IS NULL or sl.`父级分类`= '') AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'))
		# 	        ) s
        #         LEFT JOIN dim_product_gat dp ON  dp.product_id = s.`产品id`
        #         LEFT JOIN dim_cate dc ON  dc.id = dp.third_cate_id;'''.format(team, month_begin)
        # df = pd.read_sql_query(sql=sql, con=self.engine1)
        # df.to_sql('tem_product_id', con=self.engine1, index=False, if_exists='replace')
        # print('正在更新父级分类的详情…………')
        # sql = '''update {0}_order_list a, tem_product_id b
        #     		    set a.`父级分类`= IF(b.`cate` = '', a.`父级分类`, b.`cate`),
        #     				a.`二级分类`= IF(b.`second_cate` = '', a.`二级分类`, b.`second_cate`),
        #     				a.`三级分类`= IF(b.`third_cate` = '', a.`三级分类`, b.`third_cate`)
        #     			where a.`订单编号`= b.`订单编号`;'''.format(team)
        # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=1000)
        # print('更新完成+++')
        #
        # print('正在检查产品id为空的信息---')
        # sql = '''SELECT 订单编号,商品id,
		# 		        dp.product_id, dp.`name` product_name, dp.third_cate_id
        #         FROM (SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
        #             FROM {0}_order_list sl
        #             WHERE sl.`日期`> '{1}' AND (sl.`产品名称` IS NULL or sl.`产品名称`= '') AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'))
		# 	        ) s
        #         LEFT JOIN dim_product_gat dp ON dp.product_id = s.`产品id`;'''.format(
        #     team, month_begin)
        # df = pd.read_sql_query(sql=sql, con=self.engine1)
        # df.to_sql('tem_product_id', con=self.engine1, index=False, if_exists='replace')
        # print('正在更新产品详情…………')
        # sql = '''update {0}_order_list a, tem_product_id b
        #     		    set a.`产品id`= IF(b.`product_id` = '',a.`产品id`, b.`product_id`),
        #     		        a.`产品名称`= IF(b.`product_name` = '',a.`产品名称`, b.`product_name`)
        #     			where a.`订单编号`= b.`订单编号`;'''.format(team)
        # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        # print('更新完成+++')

        token = 'fc246aa95068f486c7d11368d12e0dbb'  # 补充查询产品信息需要
        if team == 'slxmt':  # 新马物流查询函数导出
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                        IF(ISNULL(b.出货时间) or b.出货时间='1899-12-29 00:00:00' or b.出货时间='0000-00-00 00:00:00' or b.状态时间='1990-01-01 00:00:00', g.出货时间, b.出货时间) 出货时间,
                        IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                        IF(b.状态时间='1990-01-01 00:00:00' or b.状态时间='1899-12-30 00:00:00' or b.状态时间='0000-00-00 00:00:00', '', b.状态时间) 状态时间,
                        IF(ISNULL(b.上线时间), a.上线时间, b.上线时间) 上线时间, 系统订单状态,
                        IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态, IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        是否改派,物流方式,物流名称,运输方式,货物类型,是否低价,付款方式,产品id,产品名称,父级分类,
                        二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,价格区间,
                        包裹重量,包裹体积,邮编,IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
                        b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态, b.添加时间, a.成本价, a.物流花费, a.打包花费, a.其它花费, a.添加物流单号时间,数量, a.省洲
                    FROM {0}_order_list a
                        LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} WHERE {0}.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
                        LEFT JOIN (SELECT * FROM {0}wl WHERE id IN (SELECT MAX(id) FROM {0}wl  WHERE {0}wl.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) g ON a.运单编号 = g.运单编号
                    WHERE a.日期 >= '{2}' AND a.日期 <= '{3}'
                        AND a.系统订单状态 IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核')
                    ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)
        elif team == 'slxmt_hfh' or team == 'slxmt_t':
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                        IF(ISNULL(b.出货时间) or b.出货时间='1899-12-29 00:00:00' or b.出货时间='0000-00-00 00:00:00' or b.状态时间='1990-01-01 00:00:00', g.出货时间, b.出货时间) 出货时间,
                        IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                        IF(b.状态时间='1990-01-01 00:00:00' or b.状态时间='1899-12-30 00:00:00' or b.状态时间='0000-00-00 00:00:00', '', b.状态时间) 状态时间,
                        IF(ISNULL(b.上线时间), a.上线时间, b.上线时间) 上线时间, 系统订单状态, IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态, IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        是否改派,物流方式,物流名称,运输方式,货物类型,是否低价,付款方式,产品id,产品名称,父级分类,
                        二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,价格区间,
                        包裹重量,包裹体积,邮编,IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
                        b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态, b.添加时间, a.成本价, a.物流花费, a.打包花费, a.其它花费, a.添加物流单号时间,数量, a.省洲
                    FROM {0}_order_list a
                        LEFT JOIN (SELECT * FROM {1} WHERE id IN (SELECT MAX(id) FROM {1} WHERE {1}.添加时间 > '{2}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {1}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {1}_return d ON a.订单编号 = d.订单编号
                        LEFT JOIN (SELECT * FROM {1}wl WHERE id IN (SELECT MAX(id) FROM {1}wl  WHERE {1}wl.添加时间 > '{2}' GROUP BY 运单编号) ORDER BY id) g ON a.运单编号 = g.运单编号
                    WHERE a.日期 >= '{3}' AND a.日期 <= '{4}'
                        AND a.系统订单状态 IN ('已审核', '待发货', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)', '待发货转审核')
                    ORDER BY a.`下单时间`;'''.format(team, 'slxmt', month_begin, month_last, month_yesterday)
        elif team == 'sltg':
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                            IF(出货时间='1990-01-01 00:00:00' or 出货时间='1899-12-29 00:00:00' or 出货时间='1899-12-30 00:00:00' or 出货时间='0000-00-00 00:00:00', null, 出货时间) 出货时间,
                            IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                            IF(状态时间='1990-01-01 00:00:00' or 状态时间='1899-12-30 00:00:00' or 状态时间='0000-00-00 00:00:00', '', 状态时间) 状态时间,
                            IF(b.上线时间='1990-01-01 00:00:00' or b.上线时间='1899-12-29 00:00:00' or b.上线时间='1899-12-30 00:00:00' or b.上线时间='0000-00-00 00:00:00', '', IF(ISNULL(b.上线时间), a.上线时间, b.上线时间)) 上线时间, 系统订单状态,
                            IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                            IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                            IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                            是否改派,物流方式,物流名称,运输方式,货物类型,是否低价,付款方式,产品id,产品名称,父级分类,
                            二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,价格区间,
                            包裹重量,包裹体积,邮编,IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
                            b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态, b.添加时间, a.成本价, a.物流花费, a.打包花费, a.其它花费, a.添加物流单号时间, 数量, b.问题明细
                    FROM {0}_order_list a
                        LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} WHERE {0}.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
                    WHERE a.日期 >= '{2}' AND a.日期 <= '{3}'
                        AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                    ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)
        elif team == 'sl_r9b':
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                        IF(出货时间='1990-01-01 00:00:00' or 出货时间='1899-12-29 00:00:00' or 出货时间='1899-12-30 00:00:00' or 出货时间='0000-00-00 00:00:00', null, 出货时间) 出货时间,
                        IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                        IF(状态时间='1990-01-01 00:00:00' or 状态时间='1899-12-30 00:00:00' or 状态时间='0000-00-00 00:00:00', '', 状态时间) 状态时间,
                        IF(ISNULL(a.上线时间), IF(b.上线时间='1990-01-01 00:00:00' or b.上线时间='1899-12-29 00:00:00' or b.上线时间='1899-12-30 00:00:00' or b.上线时间='0000-00-00 00:00:00', '',b.上线时间), a.上线时间) 上线时间, 系统订单状态,
                        IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                        IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        是否改派,物流方式,物流名称,运输方式,'货到付款' AS 货物类型,是否低价,付款方式,产品id,产品名称,父级分类,
                        二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,null 价格区间,
                        null 包裹重量,null 包裹体积,邮编,IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
                        b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态,b.添加时间, null 成本价, null 物流花费, null 打包花费, null 其它花费, a.添加物流单号时间,省洲, 数量, a.站点ID
                    FROM {0}_order_list a
                        LEFT JOIN (SELECT * FROM {1} WHERE id IN (SELECT MAX(id) FROM {1} WHERE {1}.添加时间 > '{2}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {1}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {1}_return d ON a.订单编号 = d.订单编号
                    WHERE a.日期 >= '{3}' AND a.日期 <= '{4}'
                        AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                    ORDER BY a.`下单时间`;'''.format(team, team, month_begin, month_last, month_yesterday)
        elif team in ('slsc'):
            # print(month_yesterday)
            # print('正在获取台湾的物流信息......')
            # sql = '''REPLACE INTO slsc SELECT null,订单编号,原运单号,运单编号,出货时间,物流状态,状态时间,航班时间,清关时间,上线时间,更新时间 添加时间
            #         FROM gat WHERE gat.`添加时间` = '{0} 00:00:00';'''.format(month_yesterday)
            # df = pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                        IF(出货时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), a.仓储扫描时间, 出货时间) 出货时间,
                        IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                        IF(状态时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), '', 状态时间) 状态时间,
                        IF(ISNULL(a.上线时间), IF(b.上线时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), '',b.上线时间), a.上线时间) 上线时间, 系统订单状态, 
                        IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                        IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        IF(是否改派='二次改派', '改派', 是否改派) 是否改派,物流方式,
                        IF(物流名称 like '天马物流%','天马顺丰',IF(物流名称 like '天马运通%','天马新竹',IF(物流方式 like '台湾-大黄蜂普货头程-森鸿尾程%','大黄蜂',IF(物流方式 like '台湾-立邦普货头程-易速配尾程%','立邦国际',IF(物流方式 like '日本-圆通国际-黑猫普货%','圆通国际',物流名称))))) as 物流名称,运输方式,货物类型,是否低价,
                        IF(a.物流名称 like '%义达%' or a.物流名称 like '%圆通国际%', '在线付款' ,IF(a.付款方式 not like '货到付款', '在线付款' ,'货到付款')) 付款方式,
                        产品id,产品名称,父级分类,二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,价格区间,包裹重量,包裹体积,邮编,
                        IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
                        b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态,b.添加时间, a.成本价, a.物流花费, a.打包花费, a.其它花费, a.添加物流单号时间, 省洲, 数量, 站点ID
                    FROM {0}_order_list a
                        LEFT JOIN (SELECT * FROM {0} WHERE id IN (SELECT MAX(id) FROM {0} WHERE {0}.添加时间 > '{1}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
                    WHERE a.日期 >= '{2}' AND a.日期 <= '{3}'
                    AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                    ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)
        elif team in ('gat'):
            # self.d.productIdInfo(token, '订单号', team)   # 产品id详情更新   （参数一需要手动更换）
            # self.d.cateIdInfo(token, team)  # 进入产品检索界面（参数一需要手动更换）
            sql = '''DELETE FROM gat_zqsb
                            WHERE gat_zqsb.`订单编号` IN (SELECT 订单编号
            											FROM gat_order_list 
            											WHERE gat_order_list.`系统订单状态` NOT IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
            											);'''
            print('正在清除港澳台-总表的可能删除了的订单…………')
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            sql = '''SELECT 年月, 旬, 日期, 团队, 币种, null 区域, 订单来源, a.订单编号, 电话号码, a.运单编号,
                            IF(出货时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), a.仓储扫描时间, 出货时间) 出货时间,
                            IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                            IF(状态时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), '', 状态时间) 状态时间,
                            IF(ISNULL(a.上线时间), IF(b.上线时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), null,b.上线时间), a.上线时间) 上线时间, 系统订单状态,
                            IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                            IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                            IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                            IF(是否改派='二次改派', '改派', 是否改派) 是否改派,
                            物流方式,物流名称,null 运输方式,null 货物类型,是否低价,付款方式,产品id,产品名称,父级分类, 二级分类,三级分类, 下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB, null 价格区间, null 包裹重量, null 包裹体积,null 邮编, 
                            IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在, null 签收表订单编号, null 签收表运单编号, null 原运单号, b.物流状态 签收表物流状态, null 添加时间, null 成本价, null 物流花费, null 打包花费, null 其它花费, 添加物流单号时间,
                            省洲,数量, a.下架时间, a.物流提货时间, a.完结状态, a.回款时间
                        FROM (SELECT * 
        					    FROM {0}_order_list g
        						WHERE g.日期 >= '{2}' AND g.日期 <= '{3}' AND g.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        					) a
                        LEFT JOIN gat_wl_data b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
                        ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)

        if team != 'sl_rb':
            print('正在获取---' + match[team] + ' ---全部导出数据内容…………')
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print('正在写入---' + match[team] + ' ---临时缓存…………')  # 备用临时缓存表 'slgat_run': '神龙-主页运营1组',
            df.to_sql('d1_{0}'.format(team), con=self.engine1, index=False, if_exists='replace')
        today = datetime.date.today().strftime('%Y.%m.%d')
        print('正在写入excel…………')
        filePath = []
        if team in ('gat'):
            for tem in ('"神龙家族-港澳台"|slgat', '"红杉家族-港澳台", "红杉家族-港澳台2"|slgat_hs', '"火凤凰-港台(繁体)", "火凤凰-港澳台"|slgat_hfh', '"金狮-港澳台"|slgat_js', '"金鹏家族-小虎队"|slgat_jp',  '"神龙-主页运营1组"|slgat_run'):
                tem1 = tem.split('|')[0]
                tem2 = tem.split('|')[1]
                sql = '''SELECT * FROM d1_{0} sl WHERE sl.`团队`in ({1});'''.format(team, tem1)
                df = pd.read_sql_query(sql=sql, con=self.engine1)
                df.to_sql('d1_{0}'.format(tem2), con=self.engine1, index=False, if_exists='replace')
                df.to_excel('G:\\输出文件\\{} {}签收表.xlsx'.format(today, match[tem2]),
                            sheet_name=match[tem2], index=False)
                print(tem2 + '----已写入excel')
                # print('正在打印' + match[tem2] + ' 物流时效…………')
                # self.data_wl(tem2)
        elif team in ('sl_rb'):
            for tem in ('"神龙家族-日本团队"|slrb', '"金狮-日本"|slrb_js', '"红杉家族-日本","红杉家族-日本666"|slrb_hs', '"精灵家族-日本", "精灵家族-韩国", "精灵家族-品牌"|slrb_jl', '"火凤凰-日本"|slrb_hfh', '"金牛家族-日本"|slrb_jn', '"金鹏家族-小虎队"|slrb_xhd', '"奎蛇-日本"|slrb_ks', '"奎蛇-韩国"|slrb_ks_hg', '"神龙-韩国"|slrb_sl'):
                tem1 = tem.split('|')[0]
                tem2 = tem.split('|')[1]
                sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                                IF(出货时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), null, 出货时间) 出货时间,
                                IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                                IF(状态时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), '', 状态时间) 状态时间,
                                IF(ISNULL(a.上线时间), IF(b.上线时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), '',b.上线时间), a.上线时间) 上线时间, 系统订单状态,
                                IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                                IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                                IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                                是否改派,物流方式,物流名称,运输方式,'货到付款' AS 货物类型,是否低价,付款方式,产品id,产品名称,父级分类,
                                二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,null 价格区间,
                                null 包裹重量,null 包裹体积,邮编,IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,
                                b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态,b.添加时间, null 成本价, null 物流花费, null 打包花费, null 其它花费, a.添加物流单号时间,省洲, 数量, a.站点ID
                            FROM (SELECT * 
							        FROM {0}_order_list g
							        WHERE g.日期 >= '{3}' AND g.日期 <= '{4}' 
							            AND g.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
							            AND g.团队 in ({5})
						    ) a
                            LEFT JOIN slrb_wl_data b ON a.`运单编号` = b.`运单编号`
                            LEFT JOIN {1}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                            LEFT JOIN {1}_return d ON a.订单编号 = d.订单编号
                            ORDER BY a.`下单时间`;'''.format('sl_rb', 'sl_rb', month_begin, month_last, month_yesterday, tem1)
                df = pd.read_sql_query(sql=sql, con=self.engine1)
                df.to_excel('G:\\输出文件\\{} {}签收表.xlsx'.format(today, match[tem2]), sheet_name=match[tem2], index=False)
                print(tem2 + '----已写入excel')
                df.to_sql('d1_{0}'.format(tem2), con=self.engine1, index=False, if_exists='replace')
                sql = 'REPLACE INTO {0}_zqsb_rb SELECT *, NOW() 更新时间 FROM d1_{1};'.format('sl_rb', tem2)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                print('----已写入' + match[team] + '全部签收表中')
                # print('正在打印' + match[tem2] + ' 物流时效…………')
                # self.data_wl(tem2)
        else:
            df.to_excel('G:\\输出文件\\{} {}签收表.xlsx'.format(today, match[team]),
                        sheet_name=match[team], index=False)
            print('----已写入excel')
            filePath = ['G:\\输出文件\\{} {}签收表.xlsx'.format(today, match[team])]
        if team == 'slsc':
            self.e.send('{} {}签收表.xlsx'.format(today, match[team]), filePath, emailAdd[team])

        # 导入签收率表中和输出物流时效（不包含全部的订单状态）
        print('输出文件成功…………')
        if team in ('gat', 'slsc'):
            print('正在写入' + match[team] + ' 全部签收表中…………')
            sql = 'REPLACE INTO {0}_zqsb SELECT *, NOW() 更新时间 FROM d1_{0};'.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('----已写入' + match[team] + '全部签收表中')

        # 商城订单的获取---暂时使用的
        if team in ('slgat0', 'gat0'):  # IG和UP订单
            emailAdd2 = {'slgat': 'service@igeehome.com', 'gat': 'service@igeehome.com'}
            today = datetime.date.today().strftime('%Y.%m.%d')
            month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
            month_begin = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y-%m-%d')
            print(month_begin)
            # month_last = '2021-04-01'
            # month_yesterday = '2021-06-02'
            # month_begin = '2021-2-01'
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 区域, 订单来源, a.订单编号 订单编号, 电话号码, a.运单编号 运单编号,
                        IF(出货时间='1990-01-01 00:00:00' or 出货时间='1899-12-30 00:00:00' or 出货时间='0000-00-00 00:00:00', a.仓储扫描时间, 出货时间) 出货时间,
                        IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                        IF(状态时间='1990-01-01 00:00:00' or 状态时间='1899-12-30 00:00:00' or 状态时间='0000-00-00 00:00:00', '', 状态时间) 状态时间,
                        IF(b.上线时间='1990-01-01 00:00:00' or b.上线时间='1899-12-30 00:00:00' or b.上线时间='0000-00-00 00:00:00', '', IF(ISNULL(b.上线时间), a.上线时间, b.上线时间)) 上线时间, 系统订单状态, 
                        IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                        IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                        IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , c.标准物流状态), 系统物流状态), '已退货') 最终状态,
                        IF(是否改派='二次改派', '改派', 是否改派) 是否改派,物流方式,物流名称,运输方式,货物类型,是否低价,付款方式,产品id,产品名称,父级分类,
                        二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB,价格区间,
                        IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在,b.订单编号 签收表订单编号, b.运单编号 签收表运单编号, 原运单号, b.物流状态 签收表物流状态
                    FROM {0}_order_list a
                        LEFT JOIN (SELECT * FROM {1} WHERE id IN (SELECT MAX(id) FROM {1} WHERE {1}.添加时间 > '{2}' GROUP BY 运单编号) ORDER BY id) b ON a.`运单编号` = b.`运单编号`
                        LEFT JOIN {1}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {1}_return d ON a.订单编号 = d.订单编号
                    WHERE a.日期 >= '{3}' AND a.日期 <= '{4}' AND (a.订单编号 like 'UP%' or a.订单编号 like 'IG%') AND a.团队 = '神龙家族-港澳台'
                        AND a.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)','已退货(物流)', '已退货(不拆包物流)')
                    ORDER BY a.`下单时间`;'''.format(team, 'gat', month_begin, month_last, month_yesterday)
            print('正在获取---' + match[team] + ' ---商城IG和UP订单数据内容…………')
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            df.to_excel('G:\\输出文件\\{} 商城-{}签收表.xlsx'.format(today, match[team]),
                        sheet_name=match[team], index=False)
            print('----已写入excel')
            filePath = ['G:\\输出文件\\{} 商城-{}签收表.xlsx'.format(today, match[team])]
            self.e.send('{} 商城-{}签收表.xlsx'.format(today, match[team]), filePath,
                        emailAdd2[team])

    # 物流时效
    def data_wl(self, team):  # 获取各团队近两个月的物流数据
        match = {'slgat': ['台湾', '香港'],
                 'slgat_hfh': ['台湾', '香港'],
                 'slgat_hs': ['台湾', '香港'],
                 'slgat_js': ['台湾', '香港'],
                 'slgat_jp': ['台湾', '香港'],
                 'slgat_low': ['台湾', '香港'],
                 'sltg': ['泰国'],
                 'slxmt': ['新加坡', '马来西亚', '菲律宾'],
                 'slxmt_t': ['新加坡', '马来西亚', '菲律宾'],
                 'slxmt_hfh': ['新加坡', '马来西亚', '菲律宾'],
                 'slrb': ['日本'],
                 'slrb_js': ['日本'],
                 'slrb_hs': ['日本'],
                 'slrb_jl': ['日本', '韩国']}
        match1 = {'slgat': ['台湾|神龙家族-港澳台', '香港|神龙家族-港澳台'],
                  'slgat_hfh': ['台湾|火凤凰-港澳台', '香港|火凤凰-港澳台'],
                  'slgat_hs': ['台湾|红杉家族-港澳台', '香港|红杉家族-港澳台'],
                  'slgat_js': ['台湾|金狮-港澳台', '香港|金狮-港澳台'],
                  'slgat_jp': ['台湾|小虎队-港澳台', '香港|小虎队-港澳台'],
                  'slgat_low': ['台湾|神龙-低价', '香港|神龙-低价'],
                  'sltg': ['泰国|神龙家族-泰国'],
                  'slxmt': ['新加坡|神龙家族-新加坡', '马来西亚|神龙家族-马来西亚', '菲律宾|神龙家族-菲律宾'],
                  'slxmt_t': ['新加坡|神龙-T新马菲', '马来西亚|神龙-T新马菲', '菲律宾|神龙-T新马菲'],
                  'slxmt_hfh': ['新加坡|火凤凰-新加坡', '马来西亚|火凤凰-马来西亚', '菲律宾|火凤凰-菲律宾'],
                  'slrb': ['日本|神龙家族-日本团队'],
                  'slrb_js': ['日本|金狮-日本'],
                  'slrb_hs': ['日本|红杉-日本'],
                  'slrb_jl': ['日本|精灵家族-日本', '韩国|精灵家族-韩国', '品牌|精灵家族-品牌']}
        emailAdd = {'台湾': 'giikinliujun@163.com',
                    '香港': 'giikinliujun@163.com',
                    '新加坡': 'zhangjing@giikin.com',
                    '马来西亚': 'zhangjing@giikin.com',
                    '菲律宾': 'zhangjing@giikin.com',
                    '泰国': 'zhangjing@giikin.com',
                    '日本': 'sunyaru@giikin.com',
                    '韩国': 'sunyaru@giikin.com',
                    '品牌': 'sunyaru@giikin.com'}
        if team in ('sltg', 'slrb', 'slrb_jl', 'slgat', 'slgat_hfh', 'slgat_hs', 'slgat_js', 'slgat_jp', 'slxmt', 'slxmt_t', 'slxmt_hfh'):
            month_last = (datetime.datetime.now().replace(day=1)).strftime('%Y-%m-%d')
        else:
            pass
        for tem in match1[team]:
            tem1 = tem.split('|')[0]
            tem2 = tem.split('|')[1]
            filePath = []
            listT = []  # 查询sql的结果 存放池
            print('正在获取---' + tem2 + '---物流时效…………')
            # 总月
            sql = '''SELECT 年月,
			                IFNULL(币种,'总计') as 币种,
			                IFNULL(物流方式,'总计') as 物流方式,
			                IF(下单出库时 = 90,null,IFNULL(下单出库时,'总计')) as 天数,
			                SUM(订单量) as 总计,
			                SUM(签收量) as 签收量,
			                SUM(完成量) as 完成量,				  
			                SUM(签收量) / SUM(完成量) AS '签收率完成',
			                SUM(签收量) / SUM(订单量)  AS '签收率总计',
			                null 累计完成占比
                    FROM( SELECT  年月,
					            币种,
					            物流方式,
					            IF(ISNULL(DATEDIFF(仓储扫描时间, 下单时间)), 90, DATEDIFF(仓储扫描时间, 下单时间))  AS 下单出库时,
					            COUNT(订单编号) AS 订单量, 
					            SUM(最终状态 = '已签收') as 签收量, 
					            SUM(最终状态 in ('已签收','拒收','理赔','已退货', '自发头程丢件')) as 完成量
			            FROM  d1_{0} cx 
			            WHERE cx.`币种` = '{1}'	AND cx.`团队` = '{2}'
				            AND cx.`是否改派` = '直发' 
				            AND cx.系统订单状态 IN ( '已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)' ) 
				        GROUP BY 年月,币种,物流方式,下单出库时 
				        ORDER BY 年月,币种,物流方式,下单出库时 
		                ) sl
		            GROUP BY 年月,币种,物流方式,下单出库时 
			        with rollup
			        HAVING (`币种` IS NOT null  AND `年月` IS NOT null);'''.format(team, tem1, tem2)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            listT.append(df)
            sql2 = '''SELECT 年月,
			                IFNULL(币种,'总计') as 币种,
			                IFNULL(物流方式,'总计') as 物流方式,
			                IF(出库完成时 = 90,null,IFNULL(出库完成时,'总计')) as 天数,
			                SUM(订单量) as 总计,
			                SUM(签收量) as 签收量,
			                SUM(完成量) as 完成量,				  
			                SUM(签收量) / SUM(完成量) AS '签收率完成',
			                SUM(签收量) / SUM(订单量)  AS '签收率总计',
			                null 累计完成占比
                    FROM( SELECT  年月,
					            币种,
					            物流方式,
					            IF(ISNULL(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)), 90, DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`))  AS 出库完成时,
					            COUNT(订单编号) AS 订单量, 
					            SUM(最终状态 = '已签收') as 签收量, 
					            SUM(最终状态 in ('已签收','拒收','理赔','已退货', '自发头程丢件')) as 完成量
			                FROM  d1_{0} cx 
			                WHERE cx.`币种` = '{1}' AND cx.`团队` = '{2}' 
				                AND cx.`是否改派` = '直发' 
				                AND cx.系统订单状态 IN ( '已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)' ) 
				            GROUP BY 年月,币种,物流方式,出库完成时 
				            ORDER BY 年月,币种,物流方式,出库完成时 
		                ) sl
		            GROUP BY 年月,币种,物流方式,出库完成时 
			        with rollup
			        HAVING (`币种` IS NOT null  AND `年月` IS NOT null);'''.format(team, tem1, tem2)
            df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
            listT.append(df2)
            sql3 = '''SELECT 年月,
			                IFNULL(币种,'总计') as 币种,
			                IFNULL(物流方式,'总计') as 物流方式,
			                IF(下单完成时 = 90,null,IFNULL(下单完成时,'总计')) as 天数,
			                SUM(订单量) as 总计,
			                SUM(签收量) as 签收量,
			                SUM(完成量) as 完成量,				  
			                SUM(签收量) / SUM(完成量) AS '签收率完成',
			                SUM(签收量) / SUM(订单量)  AS '签收率总计',
			                null 累计完成占比
                    FROM( SELECT  年月,
					            币种,
					            物流方式,
					            IF(ISNULL(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)), 90, DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`))  AS 下单完成时,
					            COUNT(订单编号) AS 订单量, 
					            SUM(最终状态 = '已签收') as 签收量, 
					            SUM(最终状态 in ('已签收','拒收','理赔','已退货', '自发头程丢件')) as 完成量
			            FROM  d1_{0} cx 
			            WHERE cx.`币种` = '{1}' AND cx.`团队` = '{2}' 
				            AND cx.`是否改派` = '直发' 
				            AND cx.系统订单状态 IN ( '已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)' ) 
				        GROUP BY 年月,币种,物流方式,下单完成时 
				        ORDER BY 年月,币种,物流方式,下单完成时 
		            ) sl
		            GROUP BY 年月,币种,物流方式,下单完成时 
			        with rollup
			        HAVING (`币种` IS NOT null  AND `年月` IS NOT null);'''.format(team, tem1, tem2)
            df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
            listT.append(df3)
            sql4 = '''SELECT 年月,
			                IFNULL(币种,'总计') as 币种,
			                IFNULL(物流方式,'总计') as 物流方式,
			                IF(下单完成时 = 90,null,IFNULL(下单完成时,'总计')) as 天数,
			                SUM(订单量) as 总计,
			                SUM(签收量) as 签收量,
			                SUM(完成量) as 完成量,				  
			                SUM(签收量) / SUM(完成量) AS '签收率完成',
			                SUM(签收量) / SUM(订单量)  AS '签收率总计',
			                null 累计完成占比
                    FROM( SELECT  年月,
					            币种,
					            物流方式,
					            IF(ISNULL(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)), 90, DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`))  AS 下单完成时,
					            COUNT(订单编号) AS 订单量, 
					            SUM(最终状态 = '已签收') as 签收量, 
					            SUM(最终状态 in ('已签收','拒收','理赔','已退货', '自发头程丢件')) as 完成量
			            FROM  d1_{0} cx
			            WHERE cx.`币种` = '{1}' AND cx.`团队` = '{2}' 
				            AND cx.`是否改派` = '改派' 
				            AND cx.系统订单状态 IN ( '已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)' ) 
				        GROUP BY 年月,币种,物流方式,下单完成时 
				        ORDER BY 年月,币种,物流方式,下单完成时 
		            ) sl
		            GROUP BY 年月,币种,物流方式,下单完成时 
			        with rollup
			        HAVING (`币种` IS NOT null  AND `年月` IS NOT null);'''.format(team, tem1, tem2)
            df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
            listT.append(df4)
            # 分旬
            print('正在获取---' + tem + '---物流分旬时效…………')
            sql10 = '''SELECT 年月,
			                IFNULL(币种,'总计') as 币种,
			                IFNULL(物流方式,'总计') as 物流方式,
			                IFNULL(旬,'总计') as 旬,
			                IF(下单出库时 = 90,null,IFNULL(下单出库时,'总计')) as 天数,
			                SUM(订单量) as 总计,
			                SUM(签收量) as 签收量,
			                SUM(完成量) as 完成量,				  
			                SUM(签收量) / SUM(完成量) AS '签收率完成',
			                SUM(签收量) / SUM(订单量)  AS '签收率总计',
			                null 累计完成占比
                    FROM( SELECT  年月,
					            币种,
					            物流方式,
					            旬,
					            IF(ISNULL(DATEDIFF(仓储扫描时间, 下单时间)), 90, DATEDIFF(仓储扫描时间, 下单时间))  AS 下单出库时,
					            COUNT(订单编号) AS 订单量, 
					            SUM(最终状态 = '已签收') as 签收量, 
					            SUM(最终状态 in ('已签收','拒收','理赔','已退货', '自发头程丢件')) as 完成量
			            FROM  d1_{0} cx 
			            WHERE cx.`币种` = '{1}' AND cx.`团队` = '{2}' 
				            AND cx.`是否改派` = '直发' 
				            AND cx.系统订单状态 IN ( '已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)' ) 
				        GROUP BY 年月,币种,物流方式,旬,下单出库时 
				        ORDER BY 年月,币种,物流方式,旬,下单出库时 
		            ) sl
		            GROUP BY 年月,币种,物流方式,旬,下单出库时 
			        with rollup
			        HAVING (`币种` IS NOT null  AND `年月` IS NOT null);'''.format(team, tem1, tem2)
            df10 = pd.read_sql_query(sql=sql10, con=self.engine1)
            listT.append(df10)
            sql20 = '''SELECT 年月,
			                IFNULL(币种,'总计') as 币种,
			                IFNULL(物流方式,'总计') as 物流方式,
			                IFNULL(旬,'总计') as 旬,
			                IF(出库完成时 = 90,null,IFNULL(出库完成时,'总计')) as 天数,
			                SUM(订单量) as 总计,
			                SUM(签收量) as 签收量,
			                SUM(完成量) as 完成量,				  
			                SUM(签收量) / SUM(完成量) AS '签收率完成',
			                SUM(签收量) / SUM(订单量)  AS '签收率总计',
			                null 累计完成占比
                    FROM( SELECT  年月,
					            币种,
					            物流方式,
					            旬,
					            IF(ISNULL(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)), 90, DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`))  AS 出库完成时,
					            COUNT(订单编号) AS 订单量, 
					            SUM(最终状态 = '已签收') as 签收量, 
					            SUM(最终状态 in ('已签收','拒收','理赔','已退货', '自发头程丢件')) as 完成量
			            FROM  d1_{0} cx 
			            WHERE cx.`币种` = '{1}' AND cx.`团队` = '{2}' 
				            AND cx.`是否改派` = '直发' 
				            AND cx.系统订单状态 IN ( '已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)' ) 
				            GROUP BY 年月,币种,物流方式,旬,出库完成时 
				        ORDER BY 年月,币种,物流方式,旬,出库完成时 
		            ) sl
		            GROUP BY 年月,币种,物流方式,旬,出库完成时 
			        with rollup
			        HAVING (`币种` IS NOT null  AND `年月` IS NOT null);'''.format(team, tem1, tem2)
            df20 = pd.read_sql_query(sql=sql20, con=self.engine1)
            listT.append(df20)
            sql30 = '''SELECT 年月,
			                IFNULL(币种,'总计') as 币种,
			                IFNULL(物流方式,'总计') as 物流方式,
			                IFNULL(旬,'总计') as 旬,
			                IF(下单完成时 = 90,null,IFNULL(下单完成时,'总计')) as 天数,
			                SUM(订单量) as 总计,
			                SUM(签收量) as 签收量,
			                SUM(完成量) as 完成量,				  
			                SUM(签收量) / SUM(完成量) AS '签收率完成',
			                SUM(签收量) / SUM(订单量)  AS '签收率总计',
			                null 累计完成占比
                    FROM( SELECT  年月,
					            币种,
					            物流方式,
					            旬,
					            IF(ISNULL(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)), 90, DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`))  AS 下单完成时,
					            COUNT(订单编号) AS 订单量, 
					            SUM(最终状态 = '已签收') as 签收量, 
					            SUM(最终状态 in ('已签收','拒收','理赔','已退货', '自发头程丢件')) as 完成量
			            FROM  d1_{0} cx 
			            WHERE cx.`币种` = '{1}' AND cx.`团队` = '{2}' 
				            AND cx.`是否改派` = '直发' 
				            AND cx.系统订单状态 IN ( '已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)' ) 
				        GROUP BY 年月,币种,物流方式,旬,下单完成时 
				        ORDER BY 年月,币种,物流方式,旬,下单完成时 
		            ) sl
		            GROUP BY 年月,币种,物流方式,旬,下单完成时 
			        with rollup
			        HAVING (`币种` IS NOT null  AND `年月` IS NOT null);'''.format(team, tem1, tem2)
            df30 = pd.read_sql_query(sql=sql30, con=self.engine1)
            listT.append(df30)
            sql40 = '''SELECT 年月,
			                IFNULL(币种,'总计') as 币种,
			                IFNULL(物流方式,'总计') as 物流方式,
			                IFNULL(旬,'总计') as 旬,
			                IF(下单完成时 = 90,null,IFNULL(下单完成时,'总计')) as 天数,
			                SUM(订单量) as 总计,
			                SUM(签收量) as 签收量,
			                SUM(完成量) as 完成量,				  
			                SUM(签收量) / SUM(完成量) AS '签收率完成',
			                SUM(签收量) / SUM(订单量)  AS '签收率总计',
			                null 累计完成占比
                    FROM( SELECT  年月,
					            币种,
					            物流方式,
					            旬,
					            IF(ISNULL(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)), 90, DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`))  AS 下单完成时,
					            COUNT(订单编号) AS 订单量, 
					            SUM(最终状态 = '已签收') as 签收量, 
					            SUM(最终状态 in ('已签收','拒收','理赔','已退货', '自发头程丢件')) as 完成量
			            FROM  d1_{0} cx 
			            WHERE cx.`币种` = '{1}' AND cx.`团队` = '{2}' 
				            AND cx.`是否改派` = '改派' 
				            AND cx.系统订单状态 IN ( '已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)' ) 
				        GROUP BY 年月,币种,物流方式,旬,下单完成时 
				        ORDER BY 年月,币种,物流方式,旬,下单完成时 
		            ) sl
		            GROUP BY 年月,币种,物流方式,旬,下单完成时 
			        with rollup
			        HAVING (`币种` IS NOT null  AND `年月` IS NOT null);'''.format(team, tem1, tem2)
            df40 = pd.read_sql_query(sql=sql40, con=self.engine1)
            listT.append(df40)
            print('正在写入excel…………')
            today = datetime.date.today().strftime('%Y.%m.%d')
            if team in ('slgat_hfh', 'slxmt_hfh'):
                file_path = 'G:\\输出文件\\{} 火凤凰-{}物流时效.xlsx'.format(today, tem1)
            elif team in ('slgat_hs', 'slrb_hs'):
                file_path = 'G:\\输出文件\\{} 红杉-{}物流时效.xlsx'.format(today, tem1)
            elif team == 'slxmt_t':
                file_path = 'G:\\输出文件\\{} 神龙T-{}物流时效.xlsx'.format(today, tem1)
            elif team in ('slgat_js', 'slrb_js'):
                file_path = 'G:\\输出文件\\{} 金狮-{}物流时效.xlsx'.format(today, tem1)
            elif team in ('slgat_jp'):
                file_path = 'G:\\输出文件\\{} 金鹏-{}物流时效.xlsx'.format(today, tem1)
            elif team == 'slrb_jl':
                file_path = 'G:\\输出文件\\{} 精灵-{}物流时效.xlsx'.format(today, tem1)
            else:
                file_path = 'G:\\输出文件\\{} 神龙-{}物流时效.xlsx'.format(today, tem1)
            sheet_name = ['下单出库时', '出库完成时', '下单完成时', '改派下单完成时', '下单出库(分旬)', '出库完成(分旬)', '下单完成(分旬)', '改派下单完成(分旬)']
            df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listT)):
                listT[i]['签收率完成'] = listT[i]['签收率完成'].fillna(value=0)
                listT[i]['签收率总计'] = listT[i]['签收率总计'].fillna(value=0)
                listT[i]['签收率完成'] = listT[i]['签收率完成'].apply(lambda x: format(x, '.2%'))
                listT[i]['签收率总计'] = listT[i]['签收率总计'].apply(lambda x: format(x, '.2%'))
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
                del book['Sheet1']
            writer.save()
            writer.close()
            print('正在运行宏…………')
            app = xl.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('sltem物流时效')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            print('----已写入excel ')
            # filePath.append(file_path)
            # self.e.send('{} 神龙-{}物流时效.xlsx'.format(today, tem1), filePath,
            #             emailAdd[tem1])

    # report报表
    def qsb_wl(self, team):  # 获取各团队近两个月的物流数据
        match = {'gat': '港台'}
        emailAdd = {'台湾': 'giikinliujun@163.com',
                    '香港': 'giikinliujun@163.com',
                    '品牌': 'sunyaru@giikin.com'}
        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 每日
        sql0 = '''SELECT 月份,地区, 家族,
                        SUM(s.昨日订单量) as 昨日订单量,
                        SUM(s.直发签收) as 直发签收,
                        SUM(s.直发完成) as 直发完成,
                        SUM(s.直发总订单) as 直发总订单,
                        IFNULL(SUM(s.直发签收) / SUM(s.直发完成), 0) as 直发完成签收,
                        IFNULL(SUM(s.直发签收) / SUM(s.直发总订单), 0) as 直发总计签收,
                        IFNULL(SUM(s.直发完成) / SUM(s.直发总订单), 0) as 直发完成占比,
                        SUM(s.改派签收) as 改派签收,
                        SUM(s.改派完成) as 改派完成,
                        SUM(s.改派总订单) as 改派总订单,
                        IFNULL(SUM(s.改派签收) / SUM(s.改派完成), 0) as 改派完成签收,
                        IFNULL(SUM(s.改派签收) / SUM(s.改派总订单), 0) as 改派总计签收,
                        IFNULL(SUM(s.改派完成) / SUM(s.改派总订单), 0) as 改派完成占比
                FROM( SELECT IFNULL(cx.`年月`, '总计') 月份,
                            IFNULL(cx.币种, '总计') 地区,
                            IFNULL(cx.家族, '总计') 家族,
                            SUM(IF(cx.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY),1,0)) as 昨日订单量,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) as 直发签收,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 直发完成,
                            SUM(IF(`是否改派` = '直发',1,0)) as 直发总订单,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) as 改派签收,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 改派完成,
                            SUM(IF(`是否改派` = '改派',1,0)) as 改派总订单
                        FROM  (SELECT *,
                               IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",cc.团队)))) as 家族 
                                FROM d1_gat cc
                              ) cx
                        GROUP BY cx.年月,cx.币种,cx.家族
                        WITH ROLLUP 
                    ) s
                    GROUP BY 月份,地区,家族
                    ORDER BY 月份 DESC,
                            FIELD( 地区, '台湾', '香港', '总计'),
                            FIELD( 家族, '神龙', '火凤凰', '金狮', '红杉', '总计' );'''.format(team)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)

        # 总表
        sql = '''SELECT cx.币种 线路,
			                cx.团队 家族,
			                cx.年月 月份,
			                count(订单编号) as 总订单,
			                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
			                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) * 100,2),'%') as 总计签收,
			                concat(ROUND(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) * 100,2),'%') as 完成占比,
			                null 序号
                    FROM d1_gat cx
                    GROUP BY cx.币种,cx.团队,cx.年月
                    ORDER BY cx.币种,cx.团队,cx.年月;'''.format(team)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df)
        # 总表-上月
        sql2 = '''SELECT 线路,家族,月份,总订单,完成签收,总计签收,完成占比,@rownum:=@rownum+1 AS 序号
	            FROM (SELECT cx.币种 线路,
        			        cx.团队 家族,
        			        cx.年月 月份,
        			        count(订单编号) as 总订单,
        			        concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
        			        concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) * 100,2),'%') as 总计签收,
        			        concat(ROUND(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) * 100,2),'%') as 完成占比,
        			        @rownum:=0 
                        FROM d1_gat_cp cx
                        GROUP BY cx.币种,cx.团队,cx.年月
                    ) s
                ORDER BY s.线路,s.家族,s.月份;'''.format(team)
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
        listT.append(df2)

        # 物流
        sql3 = '''SELECT s2.币种,
							s2.团队,
							s2.年月,
							s2.是否改派,
							s2.物流方式,
							s2.总订单,
							concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
							concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
							concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
							concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
							concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
							concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			                null 序号
				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.物流方式,'总计') as 物流方式,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.物流方式 as 物流方式,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                 FROM d1_gat cx
                                LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                            FROM d1_gat dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                            ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
						    ) s1
						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
					   	    with rollup
					    ) s2
                ORDER BY    FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`物流方式`,'总计'),
							s2.总订单 DESC;'''.format(team)
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)
        # 物流-上月
        sql4 = '''SELECT 币种,团队,年月,是否改派,物流方式,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as '总计签收(金额)',累计占比, @rownum:=@rownum+1 AS 序号
		        FROM ( SELECT s2.币种,
        							s2.团队,
        							s2.年月,
        							s2.是否改派,
        							s2.物流方式,
        							s2.总订单,
        							concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        							concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        							concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        							concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        							concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
        							concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.物流方式,'总计') as 物流方式,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.物流方式 as 物流方式,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                         FROM d1_gat_cp cx
                                            LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                                    FROM d1_gat_cp dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                    ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                            GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                            ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        						    ) s1
        						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
        					   	    with rollup
        					    ) s2
                        ) s
                        ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        					    FIELD(s.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        					    FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        					    FIELD(s.`是否改派`,'直发','改派','总计'),
        					    FIELD(s.`物流方式`,'总计'),
        					    s.总订单 DESC;'''.format(team)
        df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        listT.append(df4)

        # 品类
        sql5 = '''SELECT s2.币种,
								s2.团队,
								s2.年月,
								s2.是否改派,
								s2.父级分类,
								s2.总订单,
								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
                                concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			                    null 序号
				 FROM (
                        SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.父级分类,'总计') as 父级分类,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	 SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.父级分类 as 父级分类,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM d1_gat cx
                                LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                            FROM d1_gat dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                            ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
							) s1
						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
					   	with rollup
				 ) s2
				 ORDER BY	FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`父级分类`,'总计'),
							s2.总订单 DESC;'''.format(team)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)
        # 品类-上月
        sql5 = '''SELECT 币种,团队,年月,是否改派,父级分类,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as `总计签收(金额)`,累计占比, @rownum:=@rownum+1 AS 序号
		        FROM (SELECT s2.币种,
        								s2.团队,
        								s2.年月,
        								s2.是否改派,
        								s2.父级分类,
        								s2.总订单,
        								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
                                        concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM (
                                SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.父级分类,'总计') as 父级分类,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	 SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.父级分类 as 父级分类,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM d1_gat_cp cx
                                        LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                                    FROM d1_gat_cp dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                    ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        							) s1
        						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
        					   	with rollup
        				) s2 
        		) s
                ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        				FIELD(s.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        				FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        				FIELD(s.`是否改派`,'直发','改派','总计'),
        				FIELD(s.`父级分类`,'总计'),
        				s.总订单 DESC;'''.format(team)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)

        # 产品
        sql6 = '''SELECT * 
				    FROM ( SELECT   IFNULL( cx.`币种`,'总计') as 币种,
                                    IFNULL( cx.`团队`,'总计') as 团队,
                                    IFNULL( cx.`年月`,'总计') as 年月,
                                    IFNULL( cx.`产品id`,'总计') as 产品id,
                                    cx.`产品名称`,
							        cx.`父级分类`,
                                    count(订单编号) as 总订单,
                                    SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成签收,
                                    SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) as 总计签收,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) as 完成占比,
                                    count(订单编号) /总订单2 单量占比,
                                    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)),0) as 直发完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发完成占比,
                                    IFNULL(SUM(IF(是否改派 = '直发',1,0))  / 直发总订单2,0) as 直发单量占比,
                                    SUM(IF(是否改派 = '改派',1,0)) as 改派总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)),0) as 改派完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派完成占比,
                                    IFNULL(SUM(IF(是否改派 = '改派',1,0)) / 改派总订单2,0) 改派单量占比
                            FROM d1_gat cx
                            LEFT JOIN  (SELECT 币种,团队,年月,count(订单编号) as 总订单2 , 
											    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单2 , 
												SUM(IF(是否改派 = '改派',1,0)) as 改派总订单2 
										FROM d1_gat da GROUP BY da.币种,da.团队,da.年月
									) cx2  ON cx.币种 = cx2.币种 AND cx.团队 = cx2.团队 AND cx.年月 = cx2.年月
                            GROUP BY cx.币种,cx.团队,cx.年月,`产品id`
	                        with rollup
					) s1
	                ORDER BY	FIELD(s1.`币种`,'台湾','香港','总计'),
								FIELD(s1.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
								FIELD(s1.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
								总订单 DESC;'''.format(team)
        df6 = pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)

        # 产品明细-台湾
        sql7 = '''SELECT 币种,团队,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
			            concat(ROUND(速派签收量 / 速派完成量 * 100,2),'%')  AS 速派完成签收,
			            concat(ROUND(速派签收量 / 速派单量 * 100,2),'%')  AS 速派总计签收,
			            concat(ROUND(速派完成量 / 速派单量 * 100,2),'%')  AS 速派完成占比,
			            concat(ROUND(速派单量 / 订单量 * 100,2),'%')  AS 速派单量占比,
			            concat(ROUND(711签收量 / 711完成量 * 100,2),'%')  AS 711完成签收,
			            concat(ROUND(711签收量 / 711单量 * 100,2),'%')  AS 711总计签收,
			            concat(ROUND(711完成量 / 711单量 * 100,2),'%')  AS 711完成占比,
			            concat(ROUND(711单量 / 订单量 * 100,2),'%')  AS 711单量占比,
			            concat(ROUND(天马签收量 / 天马完成量 * 100,2),'%')  AS 天马完成签收,
			            concat(ROUND(天马签收量 / 天马单量 * 100,2),'%')  AS 天马总计签收,
			            concat(ROUND(天马完成量 / 天马单量 * 100,2),'%')  AS 天马完成占比,
			            concat(ROUND(天马单量 / 订单量 * 100,2),'%')  AS 天马单量占比,
			            concat(ROUND(易速配签收量 / 易速配完成量 * 100,2),'%')  AS 易速配完成签收,
			            concat(ROUND(易速配签收量 / 易速配单量 * 100,2),'%')  AS 易速配总计签收,
			            concat(ROUND(易速配完成量 / 易速配单量 * 100,2),'%')  AS 易速配完成占比,
			            concat(ROUND(易速配单量 / 订单量 * 100,2),'%')  AS 易速配单量占比,
			            concat(ROUND(森鸿签收量 / 森鸿完成量 * 100,2),'%')  AS 森鸿完成签收,
			            concat(ROUND(森鸿签收量 / 森鸿单量 * 100,2),'%')  AS 森鸿总计签收,
			            concat(ROUND(森鸿完成量 / 森鸿单量 * 100,2),'%')  AS 森鸿完成占比,
			            concat(ROUND(森鸿单量 / 订单量 * 100,2),'%')  AS 森鸿单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
				            SUM(速派单量) 速派单量,  SUM(速派签收量) 速派签收量,  SUM(速派完成量) 速派完成量,
				            SUM(711单量) 711单量,  SUM(711签收量) 711签收量,  SUM(711完成量) 711完成量,
				            SUM(天马单量) 天马单量,  SUM(天马签收量) 天马签收量,  SUM(天马完成量) 天马完成量,
				            SUM(易速配单量) 易速配单量,  SUM(易速配签收量) 易速配签收量,  SUM(易速配完成量) 易速配完成量,
				            SUM(森鸿单量) 森鸿单量,  SUM(森鸿签收量) 森鸿签收量,  SUM(森鸿完成量) 森鸿完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派签收量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  速派完成量,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS '711单量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as  '711签收量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  '711完成量',
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马单量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as  天马签收量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  天马完成量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配单量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as  易速配签收量,
							    SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配完成量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿单量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as  森鸿签收量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿完成量
	                        FROM  d1_gat cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  d1_gat cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '台湾'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-港澳台', '火凤凰-港澳台', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team)
        df7 = pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)
        # 产品明细-香港
        sql8 = '''SELECT 币种,团队,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
						concat(ROUND(立邦签收量 / 立邦完成量 * 100,2),'%')  AS 立邦完成签收,
						concat(ROUND(立邦签收量 / 立邦单量 * 100,2),'%')  AS 立邦总计签收,
						concat(ROUND(立邦完成量 / 立邦单量 * 100,2),'%')  AS 立邦完成占比,
						concat(ROUND(立邦单量 / 订单量 * 100,2),'%')  AS 立邦单量占比,
						concat(ROUND(森鸿SF签收量 / 森鸿SF完成量 * 100,2),'%')  AS 森鸿SF完成签收,
						concat(ROUND(森鸿SF签收量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF总计签收,
						concat(ROUND(森鸿SF完成量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF完成占比,
						concat(ROUND(森鸿SF单量 / 订单量 * 100,2),'%')  AS 森鸿SF单量占比,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH完成量 * 100,2),'%')  AS 森鸿SH完成签收,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH总计签收,
					    concat(ROUND(森鸿SH完成量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH完成占比,
					    concat(ROUND(森鸿SH单量 / 订单量 * 100,2),'%')  AS 森鸿SH单量占比,
					    concat(ROUND(易速配SF签收量 / 易速配SF完成量 * 100,2),'%')  AS 易速配SF完成签收,
					    concat(ROUND(易速配SF签收量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF总计签收,
					    concat(ROUND(易速配SF完成量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF完成占比,
					    concat(ROUND(易速配SF单量 / 订单量 * 100,2),'%')  AS 易速配SF单量占比,
					    concat(ROUND(易速配YC签收量 / 易速配YC完成量 * 100,2),'%')  AS 易速配YC完成签收,
					    concat(ROUND(易速配YC签收量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC总计签收,
					    concat(ROUND(易速配YC完成量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC完成占比,
					    concat(ROUND(易速配YC单量 / 订单量 * 100,2),'%')  AS 易速配YC单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
							SUM(立邦单量) 立邦单量,  SUM(立邦签收量) 立邦签收量,  SUM(立邦完成量) 立邦完成量,
				            SUM(森鸿SF单量) 森鸿SF单量,  SUM(森鸿SF签收量) 森鸿SF签收量,  SUM(森鸿SF完成量) 森鸿SF完成量,
				            SUM(森鸿SH单量) 森鸿SH单量,  SUM(森鸿SH签收量) 森鸿SH签收量,  SUM(森鸿SH完成量) 森鸿SH完成量,					
				            SUM(易速配SF单量) 易速配SF单量,  SUM(易速配SF签收量) 易速配SF签收量,  SUM(易速配SF完成量) 易速配SF完成量,				
				            SUM(易速配YC单量) 易速配YC单量,  SUM(易速配YC签收量) 易速配YC签收量,  SUM(易速配YC完成量) 易速配YC完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦签收量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  立邦完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿SF单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SF签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SF完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SH签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SH完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配SF单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as  易速配SF签收量,
							    SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配SF完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" ,1,0)) AS 易速配YC单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 = "已签收",1,0)) as  易速配YC签收量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配YC完成量
	                        FROM  d1_gat cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  d1_gat cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '香港'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-港澳台', '火凤凰-港澳台', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team)
        df8 = pd.read_sql_query(sql=sql8, con=self.engine1)
        listT.append(df8)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        for wbbook in ['神龙', '火凤凰', '红杉', '金狮']:
            file_path = 'G:\\输出文件\\{} {}-签收率.xlsx'.format(today, wbbook)
            sheet_name = ['每日', '总表', '总表上月', '物流', '物流上月', '品类', '品类上月', '产品', '产品明细台湾', '产品明细香港']
            df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listT)):
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
                del book['Sheet1']
            writer.save()
            writer.close()
            print('正在运行' + wbbook + '表宏…………')
            app = xl.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('py_sl_总运行')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        print('----已写入excel ')
        # filePath.append(file_path)
        # self.e.send('{} 神龙-{}物流时效.xlsx'.format(today, tem1), filePath,
        #                 emailAdd[tem1])

    def qsb_report(self, team):  # 报表各团队近两个月的物流数据
        match = {'gat': '港台-每日'}
        emailAdd = {'台湾': 'giikinliujun@163.com',
                    '香港': 'giikinliujun@163.com',
                    '品牌': 'sunyaru@giikin.com'}
        sql = '''DELETE FROM gat_zqsb
                        WHERE gat_zqsb.`订单编号` IN (SELECT 订单编号
        											FROM gat_order_list 
        											WHERE gat_order_list.`系统订单状态` NOT IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
        											);'''
        print('正在清除总表的可能删除了的订单…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)

        sql = '''UPDATE gat_zqsb d
                SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-易速配-顺丰%','香港-易速配-顺丰', IF(d.`物流方式` LIKE '台湾-天马-711%','台湾-天马-新竹', d.`物流方式`) )
                WHERE d.`是否改派` ='直发';'''
        print('正在修改-直发的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)
        sql = '''UPDATE gat_zqsb d
                SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-森鸿%','香港-森鸿-改派',
                                IF(d.`物流方式` LIKE '香港-立邦%','香港-立邦-改派',
								IF(d.`物流方式` LIKE '香港-易速配%','香港-易速配-改派',
								IF(d.`物流方式` LIKE '台湾-立邦普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-大黄蜂普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-森鸿-新竹%','森鸿',
								IF(d.`物流方式` LIKE '台湾-天马-顺丰%','天马顺丰',
								IF(d.`物流方式` LIKE '台湾-天马-新竹%' OR d.`物流方式` LIKE '台湾-天马-711%','天马新竹',
								IF(d.`物流方式` LIKE '台湾-天马-黑猫%','天马黑猫',
								IF(d.`物流方式` LIKE '台湾-易速配-龟山%' OR d.`物流方式` LIKE '台湾-易速配-新竹%' OR d.`物流方式` = '易速配','龟山',
								IF(d.`物流方式` LIKE '台湾-速派-新竹%' OR d.`物流方式` LIKE '台湾-速派-711超商%','速派', 
								IF(d.`物流方式` LIKE '台湾-大黄蜂普货头程-易速配尾程%' OR d.`物流方式` LIKE '台湾-立邦普货头程-易速配尾程%','龟山', d.`物流方式`)))  )  )  )  )  )  )  )
                WHERE d.`是否改派` ='改派';'''
        print('正在修改-改派的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=100)

        month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        month_now = (datetime.datetime.now()).strftime('%Y-%m-%d')
        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 每日各线路
        print('正在获取---每日各线路…………')
        sql0 = '''SELECT 月份,地区, 家族,
                        SUM(s.昨日订单量) as 昨日订单量,
                        SUM(s.直发签收) as 直发签收,
                        SUM(s.直发拒收) as 直发拒收,
                        SUM(s.直发完成) as 直发完成,
                        SUM(s.直发总订单) as 直发总订单,
                        concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发完成), 0) * 100,2),'%') as 直发完成签收,
                        concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发总订单), 0) * 100,2),'%') as 直发总计签收,
                        concat(ROUND(IFNULL(SUM(s.直发完成) / SUM(s.直发总订单), 0) * 100,2),'%')as 直发完成占比,
                        SUM(s.改派签收) as 改派签收,
                        SUM(s.改派拒收) as 改派拒收,
                        SUM(s.改派完成) as 改派完成,
                        SUM(s.改派总订单) as 改派总订单,
                        concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派完成), 0) * 100,2),'%') as 改派完成签收,
                        concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派总计签收,
                        concat(ROUND(IFNULL(SUM(s.改派完成) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派完成占比
                FROM( SELECT IFNULL(cx.`年月`, '总计') 月份,
                            IFNULL(cx.币种, '总计') 地区,
                            IFNULL(cx.家族, '总计') 家族,  
                            SUM(IF(cx.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY),1,0)) as 昨日订单量,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) as 直发签收,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 = "拒收",1,0)) as 直发拒收,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 直发完成,
                            SUM(IF(`是否改派` = '直发',1,0)) as 直发总订单,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) as 改派签收,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 = "拒收",1,0)) as 改派拒收,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 改派完成,
                            SUM(IF(`是否改派` = '改派',1,0)) as 改派总订单
                        FROM (SELECT *,
                               IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                                FROM gat_zqsb cc
                                where cc.日期 >= '{0}'
                              ) cx
                        GROUP BY cx.年月,cx.币种,cx.家族
                        WITH ROLLUP 
                    ) s
                    GROUP BY 月份,地区,家族
                    ORDER BY 月份 DESC,
                            FIELD( 地区, '台湾', '香港', '总计' ),
                            FIELD( 家族, '神龙', '火凤凰', '金狮', '金鹏', '红杉', '总计' );'''.format(month_last, team)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)

        # 各月各线路
        print('正在获取---各月各线路…………')
        sql10 = '''SELECT *
                        FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                    IFNULL(cx.`币种`, '总计') 地区,
                                    IFNULL(cx.家族, '总计') 家族,
                                    COUNT(cx.`订单编号`) as 总单量,
        			                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
        			                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
        			                concat(ROUND(SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
        			                concat(ROUND(SUM(IF( 最终状态 = "已退货",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 退款率,
        			                concat(ROUND(SUM(IF( 最终状态 = "已签收",价格RMB,0)) / SUM(价格RMB) * 100,2),'%') as '总计签收(金额)',
        			                ROUND(SUM(价格RMB) / COUNT(cx.`订单编号`),2) as 平均客单价,

                                    SUM(IF(`是否改派` = '直发',1,0))  as 直发单量,
        			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 直发完成签收,
        			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发总计签收,
        			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发完成占比,
                                    concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%')as 改派占比,
        			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 改派完成签收,
        			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派总计签收,
        			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派完成占比
                            FROM (SELECT *,
                                     IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                                 FROM gat_zqsb cc
                                  ) cx									
                            GROUP BY cx.年月,cx.币种,cx.家族
                            WITH ROLLUP 
        	            ) s
                        ORDER BY 月份 DESC,
                                FIELD( 地区, '台湾', '香港', '总计' ),
                                FIELD( s.家族, '神龙', '火凤凰','金狮', '金鹏', '红杉', '总计' ),
                                s.总单量 DESC;'''.format(team)
        df10 = pd.read_sql_query(sql=sql10, con=self.engine1)
        listT.append(df10)

        # 各月各线路---分旬
        print('正在获取---各月各线路---分旬…………')
        sql11 = '''SELECT *
                        FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                    IFNULL(cx.`旬`, '总计') 旬,
                                    IFNULL(cx.`币种`, '总计') 地区,
                                    IFNULL(cx.家族, '总计') 家族,
                                    COUNT(cx.`订单编号`) as 总单量,
        			                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
        			                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
        			                concat(ROUND(SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
        			                concat(ROUND(SUM(IF( 最终状态 = "已退货",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 退款率,
        			                concat(ROUND(SUM(IF( 最终状态 = "已签收",价格RMB,0)) / SUM(价格RMB) * 100,2),'%') as '总计签收(金额)',
        			                ROUND(SUM(价格RMB) / COUNT(cx.`订单编号`),2) as 平均客单价,

                                    SUM(IF(`是否改派` = '直发',1,0))  as 直发单量,
        			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 直发完成签收,
        			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发总计签收,
        			                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发完成占比,
                                    concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%')as 改派占比,
        			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 改派完成签收,
        			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派总计签收,
        			                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派完成占比
                            FROM (SELECT *,
                                      IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                                 FROM gat_zqsb cc
                                  )  cx									
                            GROUP BY cx.年月,cx.旬,cx.币种, cx.家族
                            WITH ROLLUP 
        	            ) s
                        ORDER BY 月份 DESC,旬,
                                FIELD( 地区, '台湾', '香港', '总计' ),
                                FIELD( s.家族, '神龙', '火凤凰','金狮', '金鹏', '红杉', '总计' ),
                                s.总单量 DESC;'''.format(team)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)
        listT.append(df11)

        # 各品类各线路
        print('正在获取---各品类各线路…………')
        sql12 = '''SELECT *
                        FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                    IFNULL(cx.`币种`, '总计') 地区,
                                    IFNULL(cx.`父级分类`, '总计') 父级分类,
                                    IFNULL(cx.家族, '总计') 家族,
                                    COUNT(cx.`订单编号`) as 总单量,
                                    concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
                                    concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
                                    concat(ROUND(SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
                                    concat(ROUND(SUM(IF( 最终状态 = "已退货",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 退款率,
                                    concat(ROUND(SUM(IF( 最终状态 = "已签收",价格RMB,0)) / SUM(价格RMB) * 100,2),'%') as '总计签收(金额)',
                                    ROUND(SUM(价格RMB) / COUNT(cx.`订单编号`),2) as 平均客单价,

                                    SUM(IF(`是否改派` = '直发',1,0))  as 直发单量,
                                    concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 直发完成签收,
                                    concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发总计签收,
                                    concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发完成占比,
                                    concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%')as 改派占比,
                                    concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 改派完成签收,
                                    concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派总计签收,
                                    concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派完成占比
                            FROM (SELECT *,
                                      IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                                 FROM gat_zqsb cc
                                  where cc.日期 >= '{0}'
                                ) cx                                  
                            GROUP BY cx.年月,cx.币种,cx.父级分类,cx.家族
                            WITH ROLLUP 
                        ) s
                        ORDER BY 月份 DESC,
                                FIELD( 地区, '台湾', '香港', '总计' ),
                                FIELD( 父级分类, '居家百货', '电子电器', '服饰', '医药保健',  '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','总计' ),
                                FIELD( s.家族, '神龙', '火凤凰','金狮', '金鹏', '红杉', '总计' ),
                                s.总单量 DESC;'''.format(month_last, team)
        df12 = pd.read_sql_query(sql=sql12, con=self.engine1)
        listT.append(df12)

        # 各物流各线路
        print('正在获取---各物流各线路…………')
        sql13 = '''SELECT *
                    FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                IFNULL(cx.`币种`, '总计') 地区,
                                IFNULL(cx.`是否改派`, '总计') 是否改派,
                                IFNULL(cx.`物流方式`, '总计') 物流方式,
                                IFNULL(cx.家族, '总计') 家族,
                                COUNT(cx.`订单编号`) as 总单量,
                                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
                                concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
                                concat(ROUND(SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
                                concat(ROUND(SUM(IF( 最终状态 = "已退货",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 退款率,
                                concat(ROUND(SUM(IF( 最终状态 = "已签收",价格RMB,0)) / SUM(价格RMB) * 100,2),'%') as '总计签收(金额)',
                                ROUND(SUM(价格RMB) / COUNT(cx.`订单编号`),2) as 平均客单价,
                        
                                SUM(IF(`是否改派` = '直发',1,0))  as 直发单量,
                                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发总计签收,
                                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 直发完成签收,
                                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发完成占比,
                                concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%')as 改派占比,
                                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派总计签收,
                                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 改派完成签收,
                                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派完成占比
                        FROM (SELECT *, IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族
                                FROM gat_zqsb cc
                                where cc.日期 >= '{0}'
                            ) cx                                  
                        GROUP BY cx.年月,cx.币种,cx.是否改派,cx.物流方式,cx.家族
                        WITH ROLLUP
                    ) s
                    ORDER BY FIELD(月份, '202108', '202107', '202106', '202105', '202104', '总计' ),
                            FIELD(地区, '台湾', '香港', '总计' ),
                            FIELD(是否改派, '直发', '改派', '总计' ),
                            FIELD(物流方式, '台湾-大黄蜂普货头程-森鸿尾程','台湾-大黄蜂普货头程-易速配尾程', '台湾-立邦普货头程-森鸿尾程','台湾-立邦普货头程-易速配尾程', '台湾-森鸿-新竹-自发头程', '台湾-速派-711超商', '台湾-速派-新竹','台湾-天马-新竹','台湾-天马-顺丰','台湾-天马-黑猫','台湾-易速配-新竹',
                                '香港-立邦-顺丰','香港-森鸿-SH渠道','香港-森鸿-顺丰渠道','香港-易速配-顺丰', '龟山','森鸿','速派','天马顺丰','天马新竹','香港-立邦-改派','香港-森鸿-改派','香港-易速配-改派','总计' ),
                            FIELD( s.家族, '神龙', '火凤凰','金狮','金鹏', '红杉', '总计' ),
                            s.总单量 DESC;'''.format(month_last, team)
        df13 = pd.read_sql_query(sql=sql13, con=self.engine1)
        listT.append(df13)

        # 同产品各团队的对比
        print('正在获取---同产品各团队的对比…………')
        sql14 = '''SELECT *,
			            IF(神龙完成签收 = '0.00%' OR 神龙完成签收 IS NULL, 神龙完成签收, concat(ROUND(神龙完成签收-完成签收,2),'%')) as 神龙对比,
			            IF(火凤凰完成签收 = '0.00%' OR 火凤凰完成签收 IS NULL, 火凤凰完成签收, concat(ROUND(火凤凰完成签收-完成签收,2),'%')) as 火凤凰对比,
			            IF(金狮完成签收 = '0.00%' OR 金狮完成签收 IS NULL, 金狮完成签收, concat(ROUND(金狮完成签收-完成签收,2),'%')) as 金狮对比,
			            IF(红杉完成签收 = '0.00%' OR 红杉完成签收 IS NULL,红杉完成签收, concat(ROUND(红杉完成签收-完成签收,2),'%')) as 红杉对比
                FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                            IFNULL(cx.币种, '总计') 地区,
                            IFNULL(cx.产品id, '总计') 产品id,
                            IFNULL(cx.产品名称, '总计') 产品名称,
                            IFNULL(cx.父级分类, '总计') 父级分类,
                            COUNT(cx.`订单编号`) as 总单量,
                            SUM(IF( 最终状态 = "已签收",1,0)) as 签收,
                            SUM(IF( 最终状态 = "拒收",1,0)) as 拒收,
                            concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
                        concat(ROUND(SUM(IF( 最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
                        SUM(IF(cx.家族 LIKE '神龙%',1,0)) as 神龙单量,
                            SUM(IF( cx.家族 LIKE '神龙%' AND 最终状态 = "已签收",1,0)) as 神龙签收,
                            SUM(IF( cx.家族 LIKE '神龙%' AND 最终状态 = "拒收",1,0)) as 神龙拒收,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '神龙%' AND `是否改派` = '改派',1,0)) / SUM(IF(cx.家族 LIKE '神龙%',1,0)) * 100,2),'%') as 神龙改派占比,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '神龙%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '神龙%',1,0)) * 100,2),'%') as 神龙签收率,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '神龙%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '神龙%' AND 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 神龙完成签收,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '神龙%' AND  最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / SUM(IF(cx.家族 LIKE '神龙%',1,0)) * 100,2),'%') as 神龙完成占比,
                        SUM(IF(cx.家族 LIKE '火凤凰%',1,0)) as 火凤凰单量,
                            SUM(IF( cx.家族 LIKE '火凤凰%' AND 最终状态 = "已签收",1,0)) as 火凤凰签收,
                            SUM(IF( cx.家族 LIKE '火凤凰%' AND 最终状态 = "拒收",1,0)) as 火凤凰拒收,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '火凤凰%' AND `是否改派` = '改派',1,0)) / SUM(IF(cx.家族 LIKE '火凤凰%',1,0)) * 100,2),'%') as 火凤凰改派占比,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '火凤凰%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '火凤凰%',1,0)) * 100,2),'%') as 火凤凰签收率,
                             concat(ROUND(SUM(IF(cx.家族 LIKE '火凤凰%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '火凤凰%' AND 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 火凤凰完成签收,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '火凤凰%' AND  最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / SUM(IF(cx.家族 LIKE '火凤凰%',1,0)) * 100,2),'%') as 火凤凰完成占比,
                        SUM(IF(cx.家族 LIKE '金狮%',1,0)) as 金狮单量,
                            SUM(IF( cx.家族 LIKE '金狮%' AND 最终状态 = "已签收",1,0)) as 金狮签收,
                            SUM(IF( cx.家族 LIKE '金狮%' AND 最终状态 = "拒收",1,0)) as 金狮拒收,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '金狮%' AND `是否改派` = '改派',1,0)) / SUM(IF(cx.家族 LIKE '金狮%',1,0)) * 100,2),'%') as 金狮改派占比,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '金狮%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '金狮%',1,0)) * 100,2),'%') as 金狮签收率,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '金狮%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '金狮%' AND 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 金狮完成签收,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '金狮%' AND  最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / SUM(IF(cx.家族 LIKE '金狮%',1,0)) * 100,2),'%') as 金狮完成占比,
                        SUM(IF(cx.家族 LIKE '金鹏%',1,0)) as 金鹏单量,
                            SUM(IF( cx.家族 LIKE '金鹏%' AND 最终状态 = "已签收",1,0)) as 金鹏签收,
                            SUM(IF( cx.家族 LIKE '金鹏%' AND 最终状态 = "拒收",1,0)) as 金鹏拒收,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '金鹏%' AND `是否改派` = '改派',1,0)) / SUM(IF(cx.家族 LIKE '金鹏%',1,0)) * 100,2),'%') as 金鹏改派占比,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '金鹏%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '金鹏%',1,0)) * 100,2),'%') as 金鹏签收率,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '金鹏%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '金鹏%' AND 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 金鹏完成签收,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '金鹏%' AND  最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / SUM(IF(cx.家族 LIKE '金鹏%',1,0)) * 100,2),'%') as 金鹏完成占比,
                        SUM(IF(cx.家族 LIKE '红杉%',1,0)) as 红杉单量,
                            SUM(IF( cx.家族 LIKE '红杉%' AND 最终状态 = "已签收",1,0)) as 红杉签收,
                            SUM(IF( cx.家族 LIKE '红杉%' AND 最终状态 = "拒收",1,0)) as 红杉拒收,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '红杉%' AND `是否改派` = '改派',1,0)) / SUM(IF(cx.家族 LIKE '红杉%',1,0)) * 100,2),'%') as 红杉改派占比,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '红杉%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '红杉%',1,0)) * 100,2),'%') as 红杉签收率,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '红杉%' AND  最终状态 = "已签收",1,0)) / SUM(IF(cx.家族 LIKE '红杉%' AND 最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) * 100,2),'%') as 红杉完成签收,
                            concat(ROUND(SUM(IF(cx.家族 LIKE '红杉%' AND  最终状态 IN ("已签收","拒收","已退货","理赔"),1,0)) / SUM(IF(cx.家族 LIKE '红杉%',1,0)) * 100,2),'%') as 红杉完成占比
                    FROM (SELECT *,IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                        FROM gat_zqsb cc
                        where cc.日期 >= '2021-06-01'
                        ) cx
                    GROUP BY cx.年月,cx.币种,cx.产品id
                WITH ROLLUP ) s
                ORDER BY FIELD(月份,DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'),'总计'),
                        FIELD(地区,'台湾','香港','总计'),
                        总单量 DESC;'''.format(month_last, team)
        df14 = pd.read_sql_query(sql=sql14, con=self.engine1)
        listT.append(df14)

        # 同产品各月的对比
        print('正在获取---同产品各月的对比…………')
        sql15 = '''SELECT *
                FROM(SELECT IFNULL(cx.`家族`, '总计') 家族,
                            IFNULL(cx.币种, '总计') 地区,
                            IFNULL(cx.产品id, '总计') 产品id,
                            IFNULL(cx.产品名称, '总计') 产品名称,
                            IFNULL(cx.父级分类, '总计') 父级分类,
                            COUNT(cx.`订单编号`) as 总单量,
                        SUM(IF(cx.年月 = '202104',1,0)) as 04总单量,
                            concat(ROUND(SUM(IF(cx.年月 = '202104' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202104',1,0)) * 100,2),'%') as 04总计签收,
                            concat(ROUND(SUM(IF(cx.年月 = '202104' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202104' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 04完成签收,
                            concat(ROUND(SUM(IF(cx.年月 = '202104' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(cx.年月 = '202104',1,0)) * 100,2),'%') as 04完成占比,
                        SUM(IF(cx.年月 = '202105',1,0)) as 05总单量,
                            concat(ROUND(SUM(IF(cx.年月 = '202105' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202105',1,0)) * 100,2),'%') as 05总计签收,
                            concat(ROUND(SUM(IF(cx.年月 = '202105' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202105' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 05完成签收,
                            concat(ROUND(SUM(IF(cx.年月 = '202105' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(cx.年月 = '202105',1,0)) * 100,2),'%') as 05完成占比,
                        SUM(IF(cx.年月 = '202106',1,0)) as 06总单量,
                            concat(ROUND(SUM(IF(cx.年月 = '202106' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202106',1,0)) * 100,2),'%') as 06总计签收,
                            concat(ROUND(SUM(IF(cx.年月 = '202106' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202106' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 06完成签收,
                            concat(ROUND(SUM(IF(cx.年月 = '202106' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(cx.年月 = '202106',1,0)) * 100,2),'%') as 06完成占比,        
                        SUM(IF(cx.年月 = '202107',1,0)) as 07总单量,
                            concat(ROUND(SUM(IF(cx.年月 = '202107' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202107',1,0)) * 100,2),'%') as 07总计签收,
                            concat(ROUND(SUM(IF(cx.年月 = '202107' AND 最终状态 = "已签收",1,0)) / SUM(IF(cx.年月 = '202107' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 07完成签收,
                            concat(ROUND(SUM(IF(cx.年月 = '202107' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(cx.年月 = '202107',1,0)) * 100,2),'%') as 07完成占比
                    FROM (SELECT *,
                                IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","金鹏",cc.团队))))) as 家族 
                         FROM gat_zqsb cc
                         )  cx
                    GROUP BY cx.家族,cx.币种,cx.产品id
                    WITH ROLLUP 
                ) s
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','金狮','金鹏','红杉','总计'),
                        FIELD( 地区, '台湾', '香港', '总计' ),
                        s.总单量 DESC;'''
        df15 = pd.read_sql_query(sql=sql15, con=self.engine1)
        listT.append(df15)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        file_path = 'G:\\输出文件\\{} {}-签收率.xlsx'.format(today, match[team])
        sheet_name = ['每日各线路', '各月各线路', '各月各线路分旬', '各品类各线路', '各物流各线路', '同产品各团队', '同产品各月']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('正在运行' + match[team] + '表宏…………')
        app = xl.App(visible=False, add_book=False)  # 运行宏调整
        app.display_alerts = False
        wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
        wbsht1 = app.books.open(file_path)
        wbsht.macro('zl_report_day')()
        wbsht1.save()
        wbsht1.close()
        wbsht.close()
        app.quit()
        print('----已写入excel ')
        # filePath.append(file_path)
        # self.e.send('{} 神龙-{}物流时效.xlsx'.format(today, tem1), filePath,
        #                 emailAdd[tem1])


    # 无运单号查询
    def noWaybillNumber(self, team):
        match1 = {'slgat': '神龙-港台',
                  'slgat_hfh': '火凤凰-港台',
                  'sltg': '神龙-泰国',
                  'slxmt': '神龙-新马',
                  'slxmt_t': '神龙T-新马',
                  'slxmt_hfh': '火凤凰-新马',
                  'slrb': '神龙-日本',
                  'slrb_jl': '精灵-日本'}
        match = {'slgat': '"神龙家族-港澳台"',
                 'slgat_hfh': '"火凤凰-港澳台"',
                 'sltg': '"神龙家族-泰国"',
                 'slxmt': '"神龙家族-新加坡", "神龙家族-马来西亚", "神龙家族-菲律宾"',
                 'slxmt_t': '"神龙-T新马菲"',
                 'slxmt_hfh': '"火凤凰-新加坡", "火凤凰-马来西亚", "火凤凰-菲律宾"',
                 'slrb': '"神龙家族-日本团队"',
                 'slrb_jl': '"精灵家族-日本"'}
        emailAdd = {'slgat': 'giikinliujun@163.com',
                    'slgat_hfh': 'giikinliujun@163.com',
                    'sltg': 'zhangjing@giikin.com',
                    'slxmt': 'zhangjing@giikin.com',
                    'slxmt_t': 'zhangjing@giikin.com',
                    'slxmt_hfh': 'zhangjing@giikin.com',
                    'slrb': 'sunyaru@giikin.com',
                    'slrb_jl': 'sunyaru@giikin.com'}
        emailAdd2 = {'sltg': 'libin@giikin.com'}
        print('正在查询{}无运单订单列表…………'.format(match[team]))
        yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        last_month = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        sql = '''SELECT a.rq 日期,
                        dim_area.name 团队,
                        a.order_number 订单编号,
                        a.waybill_number 运单编号,
                        a.order_status 系统订单状态id,
                        IF(a.second=0,'直发','改派') 是否改派,
                        dim_trans_way.all_name 物流方式,
                        a.addtime 下单时间,
                        a.verity_time 审核时间
                FROM gk_order a
                    left join dim_area ON dim_area.id = a.area_id
                    left join dim_trans_way on dim_trans_way.id = a.logistics_id
                WHERE
                    a.rq >= '{}' AND a.rq <= '{}'
                    AND dim_area.name IN ({})
                    AND ISNULL(waybill_number)
                    AND order_status NOT IN (1, 8, 11, 14, 16)
                ORDER BY a.rq;'''.format(last_month, yesterday, match[team])
        df = pd.read_sql_query(sql=sql, con=self.engine2)
        sql = 'SELECT * FROM dim_order_status;'
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        print('正在合并订单状态…………')
        df = pd.merge(left=df, right=df1, left_on='系统订单状态id', right_on='id', how='left')
        df = df.drop(labels=['id', '系统订单状态id', ], axis=1)
        df.rename(columns={'name': '系统订单状态'}, inplace=True)
        today = datetime.date.today().strftime('%Y.%m.%d')
        df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}无运单号列表.xlsx'.format(today, match1[team]),
                    sheet_name=match1[team], index=False)
        filePath = ['D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}无运单号列表.xlsx'.format(today, match1[team])]
        print('输出文件成功…………')
        self.e.send('{} {}无运单号列表.xlsx'.format(today, match1[team]), filePath,
                    emailAdd[team])
        if team == 'sltg':
            self.e.send('{} {}无运单号列表.xlsx'.format(today, match1[team]), filePath,
                        emailAdd2[team])
    # 产品花费表
    def orderCost(self, team):
        if datetime.datetime.now().day >= 9:
            endDate = datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')
            startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            endDate = [endDate, datetime.datetime.now().strftime('%Y-%m-%d')]
            startDate = [startDate, datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')]
        else:
            endDate = datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')
            startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            endDate = [endDate]
            startDate = [startDate]
        match = {'SG': '新加坡',
                 'MY': '马来西亚',
                 'PH': '菲律宾',
                 'JP': '日本',
                 'HK': '香港',
                 'TW': '台湾',
                 'TH': '泰国'}
        match2 = {'SG': 'slxmt_zqsb',
                  'MY': 'slxmt_zqsb',
                  'PH': 'slxmt_zqsb',
                  'JP': 'slrb_zqsb_rb',
                  'HK': 'slgat_zqsb',
                  'TW': 'slgat_zqsb',
                  'TH': 'sltg_zqsb'}
        emailAdd = {'SG': 'zhangjing@giikin.com',
                    'MY': 'zhangjing@giikin.com',
                    'PH': 'zhangjing@giikin.com',
                    'JP': 'sunyaru@giikin.com',
                    'HK': 'giikinliujun@163.com',
                    'TW': 'giikinliujun@163.com',
                    'TH': 'zhangjing@giikin.com'}
        # filePath = []
        for i in range(len(endDate)):
            print('正在查询' + match[team] + '产品花费表…………')
            sql = '''SELECT s1.`month` AS '月份',
                         s1.area AS '地区',
                         s1.leader AS '负责人',
                         s1.pid AS '产品ID',
                         s1.pname AS '产品名称',
                         s1.cate1 AS '一级品类',
                         s1.cate2 AS '二级品类',/   
                         s1.cate3 AS '三级品类',
                         s1.orders AS '订单量',
                         s1.orders - s1.gps AS '直发订单量',
                         s1.gps AS '改派订单量',
                         s1.salesRMB / s1.orders AS '客单价',
                         s1.salesRMB / s1.adcost AS 'ROI',
                         s1.orders / s2.orders AS '订单品类占比',
                         s1.cgcost_zf / s1.salesRMB AS '直发采购/销售额',
                         s1.adcost / s1.salesRMB AS '花费占比',
                         s1.wlcost / s1.salesRMB AS '运费占比',
                         s1.qtcost / s1.salesRMB AS '手续费占比',
                         ( s1.cgcost_zf + s1.adcost + s1.wlcost + s1.qtcost ) / s1.salesRMB AS '总成本占比',
                         s1.salesRMB_yqs / (s1.salesRMB_yqs + s1.salesRMB_yjs) AS '金额签收/完成',
                         s1.salesRMB_yqs / s1.salesRMB AS '金额签收/总计',
                         (s1.salesRMB_yqs + s1.salesRMB_yjs) / s1.salesRMB AS '金额完成占比',
                         s1.yqs / (s1.yqs + s1.yjs) AS '数量签收/完成',
                         (s1.yqs + s1.yjs) / s1.orders AS '数量完成占比',
                         s3.orders AS '昨日订单量'
            FROM (  SELECT DISTINCT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     a.product_id AS pid,
                                     e.`product_name` AS pname,
            --                       e.`name` AS pname,
                                     d.ppname AS cate1,
                                     d.pname AS cate2,
                                     d.`name` AS cate3,
            -- 						 GROUP_CONCAT(DISTINCT a.low_price) AS low_price,
                                     SUM(a.orders) AS orders,
                                     SUM(a.yqs) AS yqs,
                                     SUM(a.yjs) AS yjs,
                                     SUM(a.salesRMB) AS salesRMB,
                                     SUM(a.salesRMB_yqs) AS salesRMB_yqs,
                                     SUM(a.salesRMB_yjs) AS salesRMB_yjs,
                                     SUM(a.gps) AS gps,
            --                       SUM(a.cgcost_zf) AS cgcost_zf,
            --                       SUM(a.adcost) AS adcost,
                                     null cgcost_zf,
                                     null adcost,
                                     SUM(a.wlcost) AS wlcost,
            --                       SUM(a.qtcost) AS qtcost
                                     null qtcost
                        FROM gk_order_day a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                            LEFT JOIN dim_cate d on d.id = a.third_cate_id
                          	LEFT JOIN gk_sale e ON e.id = a.sale_id
                        WHERE a.rq >= '{startDate}'
                            AND a.rq < '{endDate}'
                            AND b.pcode = '{team}'
                            AND c.uname = '王冰'
                            AND a.beform <> 'mf'
                            AND c.uid <> 10099  -- 过滤翼虎
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, a.product_id
                     ) s1
                      LEFT JOIN
                     (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     d.ppname AS cate1,
                                     SUM(a.orders) AS orders
                        FROM gk_order_day a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                            LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        WHERE a.rq >= '{startDate}'
                            AND a.rq < '{endDate}'
                            AND b.pcode = '{team}'
                            AND c.uname = '王冰'
                            AND a.beform <> 'mf'
                            AND c.uid <> 10099
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, d.ppname
                     ) s2  ON s1.`month`=s2.`month` AND s1.area=s2.area AND s1.leader=s2.leader AND s1.cate1=s2.cate1
                      LEFT JOIN
                     (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     a.product_id AS pid,
                                     SUM(a.orders) AS orders
                        FROM gk_order_day a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                        WHERE a.rq = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                    AND b.pcode = '{team}'
                                    AND c.uname = '王冰'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, a.product_id
                     ) s3 ON s1.area=s3.area AND s1.leader=s3.leader AND s1.pid=s3.pid
            WHERE s1.orders > 0
            ORDER BY s1.orders DESC'''.format(team=team, startDate=startDate[i], endDate=endDate[i])
            df = pd.read_sql_query(sql=sql, con=self.engine2)
            print('正在输出' + match[team] + '产品花费表…………')
            columns = ['订单品类占比', '直发采购/销售额', '花费占比', '运费占比', '手续费占比', '总成本占比',
                       '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比']
            for column in columns:
                df[column] = df[column].fillna(value=0)
                df[column] = df[column].apply(lambda x: format(x, '.2%'))
            today = datetime.date.today().strftime('%Y.%m.%d')
            if i == 0:
                df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}上月产品花费表.xlsx'.format(today, match[team]),
                            sheet_name=match[team], index=False)
                # filePath.append('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}上月产品花费表.xlsx'.format(today, match[team]))
            else:
                df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}本月产品花费表.xlsx'.format(today, match[team]),
                            sheet_name=match[team], index=False)
                # filePath.append('D:\\Users\\Administrator\\Desktop\\输出文件\\{} {}本月产品花费表.xlsx'.format(today, match[team]))
        # self.e.send(match[team] + '产品花费表', filePath,
        #             emailAdd[team])
        # self.d.sl_tem_cost(match2[team], match[team])
    def orderCostHFH(self, team):
        if datetime.datetime.now().day >= 9:
            endDate = datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')
            startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            endDate = [endDate, datetime.datetime.now().strftime('%Y-%m-%d')]
            startDate = [startDate, datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')]
        else:
            endDate = datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')
            startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            endDate = [endDate]
            startDate = [startDate]
        match = {'SG': '新加坡',
                 'MY': '马来西亚',
                 'PH': '菲律宾',
                 'JP': '日本',
                 'HK': '香港',
                 'TW': '台湾',
                 'TH': '泰国'}
        match2 = {'SG': 'slxmt_hfh_zqsb',
                  'MY': 'slxmt_hfh_zqsb',
                  'PH': 'slxmt_hfh_zqsb',
                  'JP': 'slrb_zqsb_rb',
                  'HK': 'slgat_hfh_zqsb',
                  'TW': 'slgat_hfh_zqsb',
                  'TH': 'sltg_zqsb'}
        emailAdd = {'SG': 'zhangjing@giikin.com',
                    'MY': 'zhangjing@giikin.com',
                    'PH': 'zhangjing@giikin.com',
                    'JP': 'sunyaru@giikin.com',
                    'HK': 'giikinliujun@163.com',
                    'TW': 'giikinliujun@163.com',
                    'TH': 'zhangjing@giikin.com'}
        # filePath = []
        for i in range(len(endDate)):
            print('正在查询' + match[team] + '产品花费表…………')
            sql = '''SELECT s1.`month` AS '月份',
                         s1.area AS '地区',
                         s1.leader AS '负责人',
                         s1.pid AS '产品ID',
                         s1.pname AS '产品名称',
                         s1.cate1 AS '一级品类',
                         s1.cate2 AS '二级品类',
                         s1.cate3 AS '三级品类',
                         s1.orders AS '订单量',
                         s1.orders - s1.gps AS '直发订单量',
                         s1.gps AS '改派订单量',
                         s1.salesRMB / s1.orders AS '客单价',
                         s1.salesRMB / s1.adcost AS 'ROI',
                         s1.orders / s2.orders AS '订单品类占比',
                         s1.cgcost_zf / s1.salesRMB AS '直发采购/销售额',
                         s1.adcost / s1.salesRMB AS '花费占比',
                         s1.wlcost / s1.salesRMB AS '运费占比',
                         s1.qtcost / s1.salesRMB AS '手续费占比',
                         ( s1.cgcost_zf + s1.adcost + s1.wlcost + s1.qtcost ) / s1.salesRMB AS '总成本占比',
                         s1.salesRMB_yqs / (s1.salesRMB_yqs + s1.salesRMB_yjs) AS '金额签收/完成',
                         s1.salesRMB_yqs / s1.salesRMB AS '金额签收/总计',
                         (s1.salesRMB_yqs + s1.salesRMB_yjs) / s1.salesRMB AS '金额完成占比',
                         s1.yqs / (s1.yqs + s1.yjs) AS '数量签收/完成',
                         (s1.yqs + s1.yjs) / s1.orders AS '数量完成占比',
                         s3.orders AS '昨日订单量'
            FROM (  SELECT DISTINCT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     a.product_id AS pid,
                                     e.`product_name` AS pname,
            --                       e.`name` AS pname,
                                     d.ppname AS cate1,
                                     d.pname AS cate2,
                                     d.`name` AS cate3,
            -- 						 GROUP_CONCAT(DISTINCT a.low_price) AS low_price,
                                     SUM(a.orders) AS orders,
                                     SUM(a.yqs) AS yqs,
                                     SUM(a.yjs) AS yjs,
                                     SUM(a.salesRMB) AS salesRMB,
                                     SUM(a.salesRMB_yqs) AS salesRMB_yqs,
                                     SUM(a.salesRMB_yjs) AS salesRMB_yjs,
                                     SUM(a.gps) AS gps,
            --                       SUM(a.cgcost_zf) AS cgcost_zf,
            --                       SUM(a.adcost) AS adcost,
                                     null cgcost_zf,
                                     null adcost,
                                     SUM(a.wlcost) AS wlcost,
            --                       SUM(a.qtcost) AS qtcost
                                     null qtcost
                        FROM gk_order_day a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                            LEFT JOIN dim_cate d on d.id = a.third_cate_id
                            LEFT JOIN gk_sale e ON e.id = a.sale_id
                        WHERE a.rq >= '{startDate}'
                            AND a.rq < '{endDate}'
                            AND b.pcode = '{team}'
                            AND c.uname = '罗超源'
                            AND a.beform <> 'mf'
                            AND c.uid <> 10099  -- 过滤翼虎
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, a.product_id
                     ) s1
                      LEFT JOIN
                     (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     d.ppname AS cate1,
                                     SUM(a.orders) AS orders
                        FROM gk_order_day a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                            LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        WHERE a.rq >= '{startDate}'
                            AND a.rq < '{endDate}'
                            AND b.pcode = '{team}'
                            AND c.uname = '罗超源'
                            AND a.beform <> 'mf'
                            AND c.uid <> 10099
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, d.ppname
                     ) s2  ON s1.`month`=s2.`month` AND s1.area=s2.area AND s1.leader=s2.leader AND s1.cate1=s2.cate1
                      LEFT JOIN
                     (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     a.product_id AS pid,
                                     SUM(a.orders) AS orders
                        FROM gk_order_day a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                        WHERE a.rq = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                    AND b.pcode = '{team}'
                                    AND c.uname = '罗超源'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, a.product_id
                     ) s3 ON s1.area=s3.area AND s1.leader=s3.leader AND s1.pid=s3.pid
            WHERE s1.orders > 0
            ORDER BY s1.orders DESC'''.format(team=team, startDate=startDate[i], endDate=endDate[i])
            df = pd.read_sql_query(sql=sql, con=self.engine2)
            print('正在输出' + match[team] + '产品花费表…………')
            columns = ['订单品类占比', '直发采购/销售额', '花费占比', '运费占比', '手续费占比', '总成本占比',
                       '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比']
            for column in columns:
                df[column] = df[column].fillna(value=0)
                df[column] = df[column].apply(lambda x: format(x, '.2%'))
            today = datetime.date.today().strftime('%Y.%m.%d')
            if i == 0:
                df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 火凤凰-{}上月产品花费表.xlsx'.format(today, match[team]),
                            sheet_name=match[team], index=False)
                # filePath.append('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 火凤凰{}上月产品花费表.xlsx'.format(today, match[team]))
            else:
                df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 火凤凰-{}本月产品花费表.xlsx'.format(today, match[team]),
                            sheet_name=match[team], index=False)
                # filePath.append('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 火凤凰{}本月产品花费表.xlsx'.format(today, match[team]))
        # self.e.send(match[team] + '产品花费表', filePath,
        #             emailAdd[team])
        # self.d.sl_tem_costHFH(match2[team], match[team])
    def orderCostT(self, team):
        if datetime.datetime.now().day >= 9:
            endDate = datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')
            startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            endDate = [endDate, datetime.datetime.now().strftime('%Y-%m-%d')]
            startDate = [startDate, datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')]
        else:
            endDate = datetime.datetime.now().replace(day=1).strftime('%Y-%m-%d')
            startDate = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            endDate = [endDate]
            startDate = [startDate]
        match = {'SG': '新加坡',
                 'MY': '马来西亚',
                 'PH': '菲律宾'}
        match2 = {'SG': 'slxmt_t_zqsb',
                  'MY': 'slxmt_t_zqsb',
                  'PH': 'slxmt_t_zqsb'}
        emailAdd = {'SG': 'zhangjing@giikin.com',
                    'MY': 'zhangjing@giikin.com',
                    'PH': 'zhangjing@giikin.com'}
        # filePath = []
        for i in range(len(endDate)):
            print('正在查询' + match[team] + '产品花费表…………')
            sql = '''SELECT s1.`month` AS '月份',
                         s1.area AS '地区',
                         s1.leader AS '负责人',
                         s1.pid AS '产品ID',
                         s1.pname AS '产品名称',
                         s1.cate1 AS '一级品类',
                         s1.cate2 AS '二级品类',
                         s1.cate3 AS '三级品类',
                         s1.orders AS '订单量',
                         s1.orders - s1.gps AS '直发订单量',
                         s1.gps AS '改派订单量',
                         s1.salesRMB / s1.orders AS '客单价',
                         s1.salesRMB / s1.adcost AS 'ROI',
                         s1.orders / s2.orders AS '订单品类占比',
                         s1.cgcost_zf / s1.salesRMB AS '直发采购/销售额',
                         s1.adcost / s1.salesRMB AS '花费占比',
                         s1.wlcost / s1.salesRMB AS '运费占比',
                         s1.qtcost / s1.salesRMB AS '手续费占比',
                         ( s1.cgcost_zf + s1.adcost + s1.wlcost + s1.qtcost ) / s1.salesRMB AS '总成本占比',
                         s1.salesRMB_yqs / (s1.salesRMB_yqs + s1.salesRMB_yjs) AS '金额签收/完成',
                         s1.salesRMB_yqs / s1.salesRMB AS '金额签收/总计',
                         (s1.salesRMB_yqs + s1.salesRMB_yjs) / s1.salesRMB AS '金额完成占比',
                         s1.yqs / (s1.yqs + s1.yjs) AS '数量签收/完成',
                         (s1.yqs + s1.yjs) / s1.orders AS '数量完成占比',
                         s3.orders AS '昨日订单量'
            FROM (  SELECT DISTINCT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     a.product_id AS pid,
                                     e.`product_name` AS pname,
            --                       e.`name` AS pname,
                                     d.ppname AS cate1,
                                     d.pname AS cate2,
                                     d.`name` AS cate3,
            -- 						 GROUP_CONCAT(DISTINCT a.low_price) AS low_price,
                                     SUM(a.orders) AS orders,
                                     SUM(a.yqs) AS yqs,
                                     SUM(a.yjs) AS yjs,
                                     SUM(a.salesRMB) AS salesRMB,
                                     SUM(a.salesRMB_yqs) AS salesRMB_yqs,
                                     SUM(a.salesRMB_yjs) AS salesRMB_yjs,
                                     SUM(a.gps) AS gps,
            --                       SUM(a.cgcost_zf) AS cgcost_zf,
            --                       SUM(a.adcost) AS adcost,
                                     null cgcost_zf,
                                     null adcost,
                                     SUM(a.wlcost) AS wlcost,
            --                       SUM(a.qtcost) AS qtcost
                                     null qtcost
                        FROM gk_order_day a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                            LEFT JOIN dim_cate d on d.id = a.third_cate_id
            --              LEFT JOIN gk_product e on e.id = a.product_id
                          LEFT JOIN (SELECT * FROM gk_sale WHERE id IN (SELECT MAX(id) FROM gk_sale GROUP BY product_id ) ORDER BY id) e on e.product_id = a.product_id
                        WHERE a.rq >= '{startDate}'
                            AND a.rq < '{endDate}'
                            AND b.pcode = '{team}'
                            AND c.uname = '王冰'
                            AND a.beform <> 'mf'
                            AND c.uid <> 10099  -- 过滤翼虎
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, a.product_id
                     ) s1
                      LEFT JOIN
                     (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     d.ppname AS cate1,
                                     SUM(a.orders) AS orders
                        FROM gk_order_day a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                            LEFT JOIN dim_cate d on d.id = a.third_cate_id
                        WHERE a.rq >= '{startDate}'
                            AND a.rq < '{endDate}'
                            AND b.pcode = '{team}'
                            AND c.uname = '王冰'
                            AND a.beform <> 'mf'
                            AND c.uid <> 10099
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, d.ppname
                     ) s2  ON s1.`month`=s2.`month` AND s1.area=s2.area AND s1.leader=s2.leader AND s1.cate1=s2.cate1
                      LEFT JOIN
                     (  SELECT EXTRACT(YEAR_MONTH FROM a.rq) AS `month`,
                                     b.pname AS area,
                                     c.uname AS leader,
                                     a.product_id AS pid,
                                     SUM(a.orders) AS orders
                        FROM gk_order_day a
                            LEFT JOIN dim_currency_lang b ON a.currency_lang_id = b.id
                            LEFT JOIN dim_area c on c.id = a.area_id
                        WHERE a.rq = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                    AND b.pcode = '{team}'
                                    AND c.uname = '王冰'
                                    AND a.beform <> 'mf'
                                    AND c.uid <> 10099
                        GROUP BY EXTRACT(YEAR_MONTH FROM a.rq), b.pname, c.uname, a.product_id
                     ) s3 ON s1.area=s3.area AND s1.leader=s3.leader AND s1.pid=s3.pid
            WHERE s1.orders > 0
            ORDER BY s1.orders DESC'''.format(team=team, startDate=startDate[i], endDate=endDate[i])
            df = pd.read_sql_query(sql=sql, con=self.engine2)
            print('正在输出' + match[team] + '产品花费表…………')
            columns = ['订单品类占比', '直发采购/销售额', '花费占比', '运费占比', '手续费占比', '总成本占比',
                       '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比']
            for column in columns:
                df[column] = df[column].fillna(value=0)
                df[column] = df[column].apply(lambda x: format(x, '.2%'))
            today = datetime.date.today().strftime('%Y.%m.%d')
            if i == 0:
                df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙T-{}上月产品花费表.xlsx'.format(today, match[team]),
                            sheet_name=match[team], index=False)
                # filePath.append('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 火凤凰{}上月产品花费表.xlsx'.format(today, match[team]))
            else:
                df.to_excel('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 神龙T-{}本月产品花费表.xlsx'.format(today, match[team]),
                            sheet_name=match[team], index=False)
                # filePath.append('D:\\Users\\Administrator\\Desktop\\输出文件\\{} 火凤凰{}本月产品花费表.xlsx'.format(today, match[team]))
        # self.e.send(match[team] + '产品花费表', filePath,
        #             emailAdd[team])
        # self.d.sl_tem_costT(match2[team], match[team])


if __name__ == '__main__':
    #  messagebox.showinfo("提示！！！", "当前查询已完成--->>> 请前往（ 输出文件 ）查看")200
    m = MysqlControl()
    start = datetime.datetime.now()

    # 更新产品id的列表
    m.update_gk_product()
    m.update_gk_sign_rate()
    # m.qsb_report('gat')

    # 测试物流时效
    # team = 'sltg'
    # m.data_wl(team)
    # for team in ['slgat', 'slgat_hfh', 'slgat_hs','slrb', 'slrb_jl', 'sltg', 'slxmt', 'slxmt_hfh']:
    # for team in ['slgat_hs']:
    #     m.data_wl(team)

    # -----------------------------------------暂停使用-------------------------------
    # m.qsb_wl('gat')
    # m.qsb_wl2('gat')


    # for team in ['slrb', 'slxmt', 'slxmt_t', 'slxmt_hfh']:  # 无运单号查询200
    #     m.noWaybillNumber(team)
    #
    # match = {'SG': '新加坡',
    #          'MY': '马来西亚',
    #          'PH': '菲律宾',
    #          'JP': '日本'}
    # match = {'HK': '香港',
    #          'TW': '台湾'}
    # for team in match.keys():  # 产品花费表200
    #     if team == 'JP':
    #         m.orderCost(team)
    #     elif team in ('HK', 'TW'):
    #         m.orderCost(team)
    #         m.orderCostHFH(team)
    #     else:
    #         m.orderCost(team)
    #         m.orderCostHFH(team)
    #         m.orderCostT(team)

    # sm = SltemMonitoring()  # 成本查询
    # for team in ['菲律宾', '新加坡', '马来西亚', '日本', '香港', '台湾']:
    #     sm.costWaybill(team)
    print('耗时：', datetime.datetime.now() - start)
    win32api.MessageBox(0, "注意:>>>    程序运行结束， 请查看表  ！！！", "提 醒",win32con.MB_OK)

