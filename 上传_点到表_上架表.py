import pandas as pd
import os
import datetime
import time
import xlwings
import xlsxwriter
import math
import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel
import win32api,win32con
import math
from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Color,Alignment ,PatternFill # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色


# -*- coding:utf-8 -*-
class QueryTwo(Settings, Settings_sso):
    def __init__(self, userMobile, password):
        Settings.__init__(self)
        Settings_sso.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue(maxsize=10)  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        # self.sso_online_Two()
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
    def readFormHost(self, team):
        start = datetime.datetime.now()
        path = r'D:\Users\Administrator\Desktop\需要用到的文件\数据库'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                self.wbsheetHost(filePath, team)
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, team):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    if team == 'gat_logisitis_googs':
                        db.rename(columns={'订单号': '订单编号'}, inplace=True)
                        db.rename(columns={'承运单号': '运单编号'}, inplace=True)
                        print(db.columns)
                        if '物流状态' not in db.columns and '末条时间' not in db.columns and '末条信息' not in db.columns:
                            db.insert(0, '物流状态', '')
                            db.insert(0, '末条时间', '')
                            db.insert(0, '末条信息', '')
                        db = db[['下单时间', '订单编号', '运单编号', '核重时间', '物流状态', '末条时间', '末条信息']]
                        db.dropna(axis=0, how='any', inplace=True)  # 空值（缺失值），将空值所在的行/列删除后
                    elif team == 'gat_waybill_list':
                        db = db[['订单编号', '物流', '物流状态', '订单状态', '下单时间', '出库时间', '提货时间','上线时间','完成时间']]
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在导入：' + sht.name + ' 表；共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    db.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
                    print('++++成功导入缓存表')
                    columns = list(db.columns)
                    columns = ','.join(columns)
                    if team == 'gat_logisitis_googs':
                        sql = '''REPLACE INTO {0}({1}, 记录时间) SELECT *, NOW() 记录时间 FROM customer;'''.format(team, columns)
                        pd.read_sql_query(sql=sql, con=self.engine1)
                        print('++++：' + sht.name + '表--->>>更新成功')
                    elif team == 'gat_waybill_list':
                        sql = '''REPLACE INTO {0}({1},添加时间,记录时间) SELECT *, CURDATE() 添加时间,NOW() 记录时间 FROM customer;'''.format(team,columns)
                        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                        # self.waybill_updata()
                        print('++++成功导出订单跟进明细')
                else:
                    print('----------数据为空,不需导入：' + sht.name)
            wb.close()
        app.quit()

    def waybill_updata(self):
        today = datetime.date.today().strftime('%Y.%m.%d')
        listT = []  # 查询sql的结果 存放池
        print('正在获取 订单跟进汇总…………')
        sql = '''SELECT IFNULL(物流, '总计') 物流,出库,提货,上线,完成,合计
                FROM( SELECT 物流,
                            sum(IF(节点类型 = '出库',1,0)) AS 出库,
                            sum(IF(节点类型 = '提货',1,0)) AS 提货,
                            sum(IF(节点类型 = '上线',1,0)) AS 上线,
                            sum(IF(节点类型 = '完成',1,0)) AS 完成,
                            COUNT(订单编号) AS 合计
                    FROM( SELECT *,IF(出库时间 IS NULL,'出库',IF(提货时间 IS NULL,'提货',
                                    IF(上线时间 IS NULL,'上线',IF(完成时间 IS NULL,'完成',完成时间)))) AS 节点类型
                        FROM gat_waybill_list s
                    ) ss
                    GROUP BY 物流
                    WITH ROLLUP
                ) sss
                GROUP BY 物流
                ORDER BY FIELD(物流,'台湾-立邦普货头程-易速配尾程','台湾-优美宇通-新竹代收普货','台湾-优美宇通-新竹代收特货',
                            '台湾-速派-新竹','台湾-速派-711超商','台湾-天马-新竹','台湾-天马-711','合计');'''.format()
        df0 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df0)

        print('正在获取 订单跟进明细…………')
        sql = '''SELECT *,null 原因汇总
                FROM( SELECT s1.*,单量
                    FROM gat_waybill s1
                    LEFT JOIN ( SELECT 物流,节点类型, COUNT(订单编号) AS 单量
                                FROM( SELECT *,
                                            IF(出库时间 IS NULL,'出库',IF(提货时间 IS NULL,'提货',
                                            IF(上线时间 IS NULL,'上线',IF(完成时间 IS NULL,'完成',完成时间)))) AS 节点类型
                                    FROM gat_waybill_list s
                                    ) ss
                                GROUP BY 物流,节点类型
                    ) s2 ON s1.物流=s2.物流 AND s1.节点类型=s2.节点类型
                ) g
                ORDER BY FIELD(物流,'台湾-立邦普货头程-易速配尾程','台湾-优美宇通-新竹代收普货','台湾-优美宇通-新竹代收特货',
                                    '台湾-速派-新竹','台湾-速派-711超商','台湾-天马-新竹','台湾-天马-711','合计'),
                        FIELD(节点类型,'出库','提货','上线','完成','合计');'''.format()
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df1)

        print('正在获取 订单跟进明细表…………')
        sql = '''SELECT *,IF(出库时间 IS NULL,'出库',IF(提货时间 IS NULL,'提货',
                            IF(上线时间 IS NULL,'上线',IF(完成时间 IS NULL,'完成',完成时间)))) AS 节点类型
                FROM gat_waybill_list s'''.format()
        df2 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df2)

        print('正在写入excel…………')
        today = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        file_path = 'G:\\输出文件\\{} 订单跟进.xlsx'.format(today)

        writer2 = pd.ExcelWriter(file_path, engine='openpyxl')
        df0.to_excel(writer2, sheet_name='汇总',index=False, startrow=1)
        df1.to_excel(writer2, sheet_name='汇总',index=False, startrow=1, startcol=7)
        df2.to_excel(writer2, sheet_name='明细表',index=False)
        writer2.save()
        writer2.close()

        # 初始化赋值   https://openpyxl.readthedocs.io/en/stable/index.html
        # from openpyxl.utils import get_column_letter, column_index_from_string
        # # 根据列的数字返回字母
        # print(get_column_letter(2))  # B
        # # 根据字母返回列的数字
        # print(column_index_from_string('D'))  # 4

        month_yesterday = (datetime.datetime.now() - datetime.timedelta(days=5)).strftime('%m.%d')
        month_begin = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%m') + '.01-'
        # 指定了等线24号，加粗斜体，字体颜色红色。直接使用cell的font属性
        bold_itatic_24_font = Font(name='宋体', size=14, italic=False, color=colors.COLOR_INDEX[2], bold=True)
        # cell的属性aligment，这里指定垂直居中和水平居中。除了center，还可以使用right、left等
        cell_alignment = Alignment(horizontal='center', vertical='center')
        # 颜色背景
        # Fillsytle = PatternFill('solid', fgColor='#AABBCC', bgColor='#DDEEFF')
        # Fillsytle = PatternFill('solid', fgColor='#93b6f7', bgColor='#93b6f7')

        border = Border(left=Side(border_style=None,color='FF000000'),
                        right=Side(border_style=None,color='FF000000'),
                        top=Side(border_style=None,color='FF000000'),
                        bottom=Side(border_style=None,color='FF000000'),
                        diagonal=Side(border_style=None,color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style=None,color='FF000000'),
                        vertical=Side(border_style=None,color='FF000000'),
                        horizontal=Side(border_style=None, color='FF000000'),
                        diagonalDown=False,
                        start=None,
                        end=None)

        wb = load_workbook(file_path)
        # sheet = wb.get_sheet_by_name("汇总")
        sheet = wb["汇总"]
        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 50
        sheet.row_dimensions[2].alignment = cell_alignment
        # sheet.row_dimensions[2].fill = Fillsytle  # 设定行的颜色

        sheet["A1"] = month_begin + month_yesterday +'台湾直发订单跟进 汇总'
        sheet['A1'].font = bold_itatic_24_font
        sheet['A1'].alignment = cell_alignment
        # sheet['A2:K30'].alignment = cell_alignment
        sheet.merge_cells('A1:F1')
        sheet.column_dimensions['A'].width = 28
        for cl in ['B','C','D','E','F','C','C']:
            sheet.column_dimensions[cl].width = 11.13
        for row in sheet.iter_rows(min_row=3, max_row=10, min_col=2, max_col=6):
            for cell in row:
                cell.alignment = cell_alignment

        sheet["H1"] = month_begin + month_yesterday +'台湾直发订单跟进 明细'
        sheet['H1'].font = bold_itatic_24_font
        sheet['H1'].alignment = cell_alignment
        sheet.merge_cells('H1:K1')
        sheet.column_dimensions['H'].width = 28
        for cl in ['I','J','K']:
            sheet.column_dimensions[cl].width = 11.13
        for row in sheet.iter_rows(min_row=3, max_row=30, min_col=9, max_col=11):
            for cell in row:
                cell.alignment = cell_alignment
        # sheet.column_dimensions['A:f'].width = 25
        # sheet.columns[2].width = 11.13
        # sheet.rows(2,6).width = 11.13
        # sheet['2:6'].width = 11.13

        wb.save(file_path)
        print('----已写入excel ')


if __name__ == '__main__':
    m = QueryTwo('+86-18538110674', 'qyz04163510')
    start: datetime = datetime.datetime.now()
    '''
    # -----------------------------------------------查询状态运行（一）-----------------------------------------
    # 1、 点到表上传；2、上架表上传；；3、订单跟进上传--->>数据更新切换
    '''

    select = 3
    if int(select) == 1:
        team = 'gat_logisitis_googs'
        m.readFormHost(team)
    elif int(select) == 2:
        print("2-->>> 正在按时间查询+++")
        timeStart = '2022-03-28'
        timeEnd = '2022-03-29'
        # m.order_TimeQuery(timeStart, timeEnd)
    elif int(select) == 3:
        team = 'gat_waybill_list'
        m.readFormHost(team)
        m.waybill_updata()
    print('查询耗时：', datetime.datetime.now() - start)