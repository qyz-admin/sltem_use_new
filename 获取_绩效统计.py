# coding=utf-8
import pandas as pd
import os
import datetime
import time
import xlwings
import xlsxwriter
import math
import requests
from requests.adapters import HTTPAdapter
import json
import sys
import zhconv  # transform2_zh_hant：转为繁体;transform2_zh_hans：转为简体
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread  # 使用 threading 模块创建线程
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
import pandas.io.formats.excel
import win32api, win32con
import math
from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, \
    Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色
from 查询_订单检索 import QueryOrder


# -*- coding:utf-8 -*-
class QueryOrder_Code(Settings, Settings_sso):
    def __init__(self, userMobile, password, login_TmpCode, handle, proxy_handle, proxy_id, select):
        Settings.__init__(self)
        Settings_sso.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue(maxsize=10)  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        # self.sso_online_Two()
        # self._online_Two()
        # self.sso__online_auto()

        if select >= 10:
            if proxy_handle == '代理服务器':
                if handle == '手动':
                    self.sso__online_handle_proxy(login_TmpCode, proxy_id)
                else:
                    self.sso__online_auto_proxy(proxy_id)
            else:
                if handle == '手动':
                    self.sso__online_handle(login_TmpCode)
                else:
                    self.sso__online_auto()

        # self.sso__online_handle(login_TmpCode)
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                     self.mysql20['password'],
                                                                                     self.mysql20['host'],
                                                                                     self.mysql20['port'],
                                                                                     self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()

    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    def readFormHost(self, team, searchType, pople_Query, timeStart, timeEnd):
        start = datetime.datetime.now()
        path = r'D:\Users\Administrator\Desktop\需要用到的文件\A查询导表'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                self.wbsheetHost(filePath, team, searchType)
                # self.cs_wbsheetHost(filePath, team, searchType)
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, team, searchType):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                if sht.api.Visible == -1:
                    try:
                        tem = None
                        db = None
                        db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                        print(db.columns)
                        # db = db[['订单编号']]
                        columns_value = list(db.columns)  # 获取数据的标题名，转为列表
                        if searchType == '订单号':
                            tem = '订单编号'
                            if '订单号' in columns_value:
                                db.rename(columns={'订单号': '订单编号'}, inplace=True)
                            for column_val in columns_value:
                                if '订单编号' != column_val:
                                    db.drop(labels=[column_val], axis=1, inplace=True)  # 去掉多余的旬列表
                            db.dropna(axis=0, how='any', inplace=True)  # 空值（缺失值），将空值所在的行/列删除后
                        elif searchType == '运单号':
                            tem = '运单编号'
                            if '运单号' in columns_value:
                                db.rename(columns={'运单号': '运单编号'}, inplace=True)
                            elif '查件单号' in columns_value:
                                db.rename(columns={'查件单号': '运单编号'}, inplace=True)
                    except Exception as e:
                        print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                    if db is not None and len(db) > 0:
                        # print(db)
                        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
                        print('++++正在获取：' + sht.name + ' 表；共：' + str(len(db)) + '行',
                              'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        orderId = list(db[tem])
                        max_count = len(orderId)  # 使用len()获取列表的长度，上节学的
                        # print(orderId)
                        # print(max_count)
                        if max_count > 500:
                            ord = ','.join(orderId[0:500])
                            # print(ord)
                            df = self.orderInfoQuery(ord, searchType)
                            # print(df)
                            dlist = []
                            n = 0
                            while n < max_count - 500:  # 这里用到了一个while循环，穿越过来的
                                n = n + 500
                                ord = ','.join(orderId[n:n + 500])
                                data = self.orderInfoQuery(ord, searchType)
                                dlist.append(data)
                            print('正在写入......')
                            # print(dlist)
                            dp = df.append(dlist, ignore_index=True)
                        else:
                            ord = ','.join(orderId[0:max_count])
                            dp = self.orderInfoQuery(ord, searchType)
                        dp.columns = ['订单编号', '币种', '运营团队', '产品id', '产品名称', '出货单名称', '规格(中文)', '收货人', '联系电话', '拉黑率',
                                      '电话长度',
                                      '配送地址', '应付金额', '数量', '订单状态', '运单号', '支付方式', '下单时间', '审核人', '审核时间', '物流渠道',
                                      '货物类型',
                                      '是否低价', '站点ID', '商品ID', '订单类型', '物流状态', '重量', '删除原因', '问题原因', '下单人', '转采购时间',
                                      '发货时间', '上线时间',
                                      '完成时间', '销售退货时间', '备注', 'IP', '体积', '省洲', '市/区', '选品人', '优化师', '审单类型', '克隆人',
                                      '克隆ID', '发货仓库', '是否发送短信',
                                      '物流渠道预设方式', '拒收原因', '物流更新时间', '状态时间', '来源域名', '订单来源类型', '更新时间', '异常提示', '异常拉黑率',
                                      '拉黑率总量', '拉黑率签收', '拉黑率拒收', '留言']
                        dp.to_excel('F:\\输出文件\\订单检索-查询{}.xlsx'.format(rq), sheet_name='查询', index=False,
                                    engine='xlsxwriter')  # Xlsx是python用来构造xlsx文件的模块，可以向excel2007+中写text，numbers，formulas 公式以及hyperlinks超链接。
                        print('查询已导出+++')
                    else:
                        print('----------数据为空,查询失败：' + sht.name)
                else:
                    print('----不需查询：' + sht.name)
            wb.close()
        app.quit()

    # 绩效-查询 促单（一.1）
    def service_id_order(self, timeStart, timeEnd, proxy_handle, proxy_id):  # 进入订单检索界面     促单查询
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        pople_Query = '促单查询'
        print('正在查询 促单订单 起止时间：' + str(timeStart) + " *** " + str(timeEnd))
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = { 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500, 'orderPrefix': None, 'orderNumberFuzzy': None, 'shipUsername': None,'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None,
                'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None,  'lowerstatus': None, 'warehouse': None,
                'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None,  'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None,
                'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': -1, 'autoVerifyStatus': None,'shipZip': None, 'remark': None,
                'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'isChangeMark': None, 'timeStart': timeStart + ' 00:00:00', 'timeEnd': timeEnd + ' 23:59:59'}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        max_count = req['data']['count']
        print('共...' + str(max_count) + '...单量')
        if max_count != 0 and max_count != []:
            df = pd.DataFrame([])
            dlist = []
            in_count = math.ceil(max_count / 500)
            n = 1
            while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                data = self._service_id_order(timeStart, timeEnd, n, proxy_handle, proxy_id, pople_Query, "")
                dlist.append(data)
                print('剩余查询次数' + str(in_count - n))
                n = n + 1
            dp = df.append(dlist, ignore_index=True)
            dp = dp[['orderNumber', 'currency', 'addTime', 'service', 'cloneUser', 'orderStatus', 'logisticsStatus', 'cloneTypeName']]
            dp.columns = ['订单编号', '币种', '下单时间', '代下单客服', '克隆人', '订单状态', '物流状态', '克隆类型']
            dp.to_excel('F:\\输出文件\\绩效促单-下单时间{}.xlsx'.format(rq), sheet_name='促单', index=False, engine='xlsxwriter')
            dp.to_sql('cache_check', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 促单_下单时间(订单编号,币种, 下单时间, 代下单客服, 克隆人, 克隆类型, 订单状态, 物流状态, 统计月份,记录时间) 
                    SELECT 订单编号,币种, 下单时间, 代下单客服, 克隆人, 克隆类型, 订单状态, 物流状态, DATE_FORMAT(下单时间,'%Y%m') 统计月份,NOW() 记录时间 
                    FROM cache_check;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
        else:
            print('没有需要获取的信息！！！')
            return
        print('-' * 50)
        print('-' * 50)
    def _service_id_order(self, timeStart, timeEnd, n, proxy_handle, proxy_id, pople_Query, ord):
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = { 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        if pople_Query == '促单查询':
            data = {'page': n, 'pageSize': 500, 'orderPrefix': None, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                    'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None,
                    'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': None, 'warehouse': None,
                    'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None, 'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None,
                    'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': -1, 'autoVerifyStatus': None, 'shipZip': None, 'remark': None,
                    'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                    'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'isChangeMark': None, 'timeStart': timeStart + ' 00:00:00', 'timeEnd': timeEnd + ' 23:59:59'}
        elif pople_Query == '物流客诉查询':
            data = {'page': 1, 'pageSize': 500, 'orderPrefix': ord, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                    'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None,
                    'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': None, 'warehouse': None,
                    'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None, 'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None,
                    'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': None, 'autoVerifyStatus': None, 'shipZip': None, 'remark': None,
                    'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                    'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'isChangeMark': None, 'timeStart': None, 'timeEnd': None}
        else:
            data = None
        if proxy_handle == '代理服务器':     # print('+++已成功发送请求......')
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersdict = []
        try:    # print('正在处理json数据转化为dataframe…………')
            for result in req['data']['list']:
                if result['specs'] != []:
                    result['saleId'] = 0  # 添加新的字典键-值对，为下面的重新赋值用
                    result['saleName'] = 0
                    result['productId'] = 0
                    result['saleProduct'] = 0
                    result['spec'] = 0
                    result['chooser'] = 0
                    result['saleId'] = result['specs'][0]['saleId']
                    result['saleName'] = result['specs'][0]['saleName']
                    result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
                    result['saleProduct'] = (result['specs'][0]['saleProduct']).split('#')[2]
                    result['spec'] = result['specs'][0]['spec']
                    result['chooser'] = result['specs'][0]['chooser']
                else:
                    result['saleId'] = ''  # 添加新的字典键-值对，为下面的重新赋值用
                    result['saleName'] = ''
                    result['productId'] = ''
                    result['saleProduct'] = ''
                    result['spec'] = ''
                    result['chooser'] = ''
                quest = ''
                for re in result['questionReason']:
                    quest = quest + ';' + re
                result['questionReason'] = quest
                delr = ''
                for re in result['delReason']:
                    delr = delr + ';' + re
                result['delReason'] = delr
                auto = ''
                for re in result['autoVerify']:
                    auto = auto + ';' + re
                result['autoVerify'] = auto

                result['auto_VerifyTip'] = ''
                result['auto_VerifyTip_zl'] = ''
                result['auto_VerifyTip_qs'] = ''
                result['auto_VerifyTip_js'] = ''
                if result['autoVerifyTip'] == "":
                    result['auto_VerifyTip'] = '0.00%'
                else:
                    if '未读到拉黑表记录' in result['autoVerifyTip']:
                        result['auto_VerifyTip'] = '0.00%'
                    else:
                        t3 = result['autoVerifyTip']
                        result['auto_VerifyTip_zl'] = (t3.split('订单配送总量：')[1]).split(',')[0]
                        result['auto_VerifyTip_qs'] = (t3.split('送达订单量：')[1]).split(',')[0]
                        result['auto_VerifyTip_js'] = (t3.split('拒收订单量：')[1]).split(',')[0]
                        if '拉黑率问题' not in result['autoVerifyTip']:
                            t2 = result['autoVerifyTip'].split(',拉黑率')[1]
                            result['auto_VerifyTip'] = t2.split('%;')[0] + '%'
                        else:
                            t2 = result['autoVerifyTip'].split('拒收订单量：')[1]
                            t2 = t2.split('%;')[0]
                            result['auto_VerifyTip'] = t2.split('拉黑率')[1] + '%'
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e", "订单检索-绩效数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        df = pd.json_normalize(ordersdict)
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return df

    # 绩效-查询 挽单列表（一.2）
    def service_id_getRedeemOrderList(self, timeStart, timeEnd, proxy_handle, proxy_id):  # 进入订单检索界面     挽单列表查询
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在查询 挽单列表 起止时间：' + str(timeStart) + " *** " + str(timeEnd))
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getRedeemOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/saveOrder'}
        data = {'order_number': None, 'type': None, 'order_status': None, 'logistics_status': None, 'old_order_status': None, 'old_logistics_status': None, 'operator': None,
                'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59', 'is_del': None, 'page': 1, 'pageSize': 10, 'area_id': None}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        max_count = req['data']['count']
        print('共...' + str(max_count) + '...单量')
        if max_count != 0:
            df = pd.DataFrame([])
            n = 1
            in_count = math.ceil(max_count / 90)
            # print(in_count)
            dlist = []
            while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                data = self._service_id_getRedeemOrderList(timeStart, timeEnd, n, proxy_handle, proxy_id)
                dlist.append(data)
                print('剩余查询次数' + str(in_count - n))
                n = n + 1
            dp = df.append(dlist, ignore_index=True)
            dp = dp[['id', 'order_number', 'redeemType', 'oldOrderStatus', 'oldLogisticsStatus', 'oldAmount', 'orderStatus', 'logisticsStatus', 'amount', 'logisticsName', 'operatorName',
                     'create_time', 'save_money', 'currencyName', 'delOperatorName', 'del_reason']]
            dp.columns = ['id', '订单编号', '挽单类型', '原订单状态', '原物流状态', '原订单金额', '当前订单状态', '当前物流状态', '当前订单金额', '当前物流渠道', '创建人', '创建时间', '挽单金额', '币种', '删除人', '删除原因']
            dp.to_excel('F:\\输出文件\\挽单列表-查询{}.xlsx'.format(rq), sheet_name='挽单', index=False, engine='xlsxwriter')
            dp.to_sql('cache_check', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 挽单列表_创建时间(id, 订单编号,币种, 创建时间, 创建人, 挽单类型, 挽单金额, 当前订单状态, 当前物流状态, 回款状态, 删除人, 删除原因, 统计月份,记录时间) 
                    SELECT id, 订单编号,币种, 创建时间, 创建人, 挽单类型, 挽单金额, 当前订单状态, 当前物流状态,NULL as 回款状态, 删除人, 删除原因, DATE_FORMAT(创建时间,'%Y%m') 统计月份,NOW() 记录时间 
                    FROM cache_check;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
        print('-' * 50)
        print('-' * 50)
    def _service_id_getRedeemOrderList(self, timeStart, timeEnd, n, proxy_handle, proxy_id):
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getRedeemOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'order_number': None, 'type': None, 'order_status': None, 'logistics_status': None,'old_order_status': None, 'old_logistics_status': None, 'operator': None,
                'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59', 'is_del': None, 'page': n, 'pageSize': 90, 'area_id': None}
        if proxy_handle == '代理服务器':        # print('+++已成功发送请求......')
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersdict = []
        try:        # print('正在处理json数据转化为dataframe…………')
            for result in req['data']['list']:
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e","挽单列表-绩效数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        df = pd.json_normalize(ordersdict)
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return df

    # 绩效-查询 采购异常             （二.1.1）
    def service_id_ssale(self, timeStart, timeEnd, proxy_handle, proxy_id, order_time):  # 进入采购问题件界面   # 筛选币种
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在查询 采购订单(' + order_time + ') 起止时间：' + str(timeStart) + " *** " + str(timeEnd))
        url = r'https://gimp.giikin.com/service?service=gorder.afterSale&action=getPurchaseAbnormalList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerComplaint'}
        data = {'page': 1, 'pageSize': 90, 'areaId': None, 'userId': None, 'dealUser': None, 'currencyId': "6,13", 'orderNumber': None,
                'productId': None, 'timeStart': None, 'timeEnd': None, 'add_time_start': None, 'add_time_end': None,
                'orderType': None, 'lastProcess': None, 'logisticsStatus': None, 'update_time_start': None, 'update_time_end': None}
        data_woks = None
        data_woks2 = None
        if order_time == '跟进时间':        # print('+++已成功发送请求......')
            data.update({'update_time_start': timeStart + ' 00:00:00', 'update_time_end': timeEnd + ' 23:59:59'})
            data_woks = '采购问题件_跟进时间'
            data_woks2 = '处理时间'
        elif order_time == '创建时间':
            data.update({'timeStart': timeStart + ' 00:00:00', 'timeEnd': timeEnd + ' 23:59:59'})
            data_woks = '采购异常_创建时间'
            data_woks2 = '创建时间'
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        if req['data'] != []:
            max_count = req['data']['total']
            print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
            print('*' * 50)
            if max_count != 0 and max_count != []:
                df = pd.DataFrame([])
                dlist = []
                in_count = math.ceil(max_count / 90)
                n = 1
                while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                    data = self._service_id_ssale(timeStart, timeEnd, n, proxy_handle, proxy_id, order_time)  # 分页获取详情
                    dlist.append(data)
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                dp = df.append(dlist, ignore_index=True)
                dp = dp[['orderNumber', 'currencyName', 'addtime', 'orderStatus', 'logisticsStatus', 'dealTime', 'dealName','dealProcess', 'description', 'create_time', 'fbName']]
                dp.columns = ['订单编号', '币种', '下单时间', '订单状态', '物流状态', '处理时间', '处理人', '处理结果', '反馈描述', '创建时间', '采购反馈人']
                dp.to_excel('F:\\输出文件\\绩效采购问题件-{0}{1}.xlsx'.format(order_time, rq), sheet_name='采购', index=False, engine='xlsxwriter')
                dp.to_sql('cache_check', con=self.engine1, index=False, if_exists='replace')
                sql = '''REPLACE INTO {0}(订单编号,币种,下单时间,订单状态,物流状态,处理时间,处理人, 处理结果, 反馈描述, 创建时间, 采购反馈人,客服处理时间,客服处理人, 客服处理结果,客服反馈描述,统计月份,记录时间) 
                         SELECT 订单编号,币种,下单时间,订单状态,物流状态,处理时间,处理人, 处理结果, 反馈描述, 创建时间, 采购反馈人,NULL 客服处理时间,NULL 客服处理人, NULL 客服处理结果,NULL 客服反馈描述,DATE_FORMAT({1},'%Y%m') 统计月份, NOW() 记录时间 
                        FROM cache_check;'''.format(data_woks, data_woks2)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                print('写入成功......')
        else:
            print('没有需要获取的信息！！！')
            return
        print('-' * 50)
        print('-' * 50)
    def _service_id_ssale(self, timeStart, timeEnd, n, proxy_handle, proxy_id, order_time):  # 进入物流问题件界面
        print('+++正在查询第 ' + str(n) + ' 页信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.afterSale&action=getPurchaseAbnormalList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerComplaint'}
        data = {'page': n, 'pageSize': 90, 'areaId': None, 'userId': None, 'dealUser': None, 'currencyId': "6,13", 'orderNumber': None,
                'productId': None, 'timeStart': None, 'timeEnd': None, 'add_time_start': None, 'add_time_end': None,
                'orderType': None, 'lastProcess': None, 'logisticsStatus': None, 'update_time_start': None,'update_time_end': None}
        if order_time == '跟进时间':        # print('+++已成功发送请求......')
            data.update({'update_time_start': timeStart + ' 00:00:00', 'update_time_end': timeEnd + ' 23:59:59'})
        elif order_time == '创建时间':
            data.update({'timeStart': timeStart + ' 00:00:00', 'timeEnd': timeEnd + ' 23:59:59'})
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersDict = []
        try:
            for result in req['data']['data']:  # 添加新的字典键-值对，为下面的重新赋值用
                ordersDict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e","采购异常-绩效数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        data = pd.json_normalize(ordersDict)
        print('++++++单次查询成功+++++++')
        print('*' * 50)
        return data
    # 绩效-查询 采购异常 补充查询             （二.1.2）
    def service_id_ssale_info(self, proxy_handle, proxy_id, data_name, user_caigou):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('采购异常-绩效 处理详情 获取中......')
        sql2 = '''SELECT 订单编号 FROM cache_check s1;'''
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
        orderId = list(df2['订单编号'])
        max_count = len(orderId)  # 使用len()获取列表的长度，上节学的
        if max_count > 0:
            print('++++++本批次更新;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
            df = pd.DataFrame([])  # 创建空的dataframe数据框
            dlist = []
            for ord in orderId:
                data = self._service_id_ssale_info(ord, proxy_handle, proxy_id, user_caigou)
                if data is not None and len(data) > 0:
                    dlist.append(data)
            dp = df.append(dlist, ignore_index=True)
            print(dp)
            dp.to_excel('F:\\输出文件\\绩效采购-查询详情{}.xlsx'.format(rq), sheet_name='采购', index=False, engine='xlsxwriter')
            dp = dp[['orderNumber', 'addTime', 'name', 'dealProcess', 'content']]
            dp.columns = ['订单编号', '客服处理时间', '客服处理人', '客服处理结果', '客服反馈描述']
            dp.to_sql('cache_check_cp', con=self.engine1, index=False, if_exists='replace')
            print('正在更新表处理详情中......')
            sql = '''update {0} a, {1} b
                                   set a.`客服处理时间`= IF(b.`客服处理时间` = '', NULL,  b.`客服处理时间`),
                                       a.`客服处理人`= IF(b.`客服处理人` = '', NULL,  b.`客服处理人`),
                                       a.`客服处理结果`= IF(b.`客服处理结果` = '', NULL, b.`客服处理结果`),
                                       a.`客服反馈描述`= IF(b.`客服反馈描述` = '', NULL, b.`客服反馈描述`)
                           where a.`订单编号`=b.`订单编号`;'''.format(data_name, 'cache_check_cp')
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('更新成功......')
    def _service_id_ssale_info(self, ord, proxy_handle, proxy_id, user_caigou):  # 进入采购问题件界面
        print('正在查询 ' + str(ord) + ' 处理详情中')
        url = r'https://gimp.giikin.com/service?service=gorder.afterSale&action=abnormalDisposeLog'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https://gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/purchaseFeedback'}
        data = {'orderNumber': ord}
        if proxy_handle == '代理服务器':        # print('+++已成功发送请求......')
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersDict = []
        try:
            for result in req['data']:  # 添加新的字典键-值对，为下面的重新赋值用
                # print(result)
                # if result['name'] == '蔡利英' or result['name'] == '张陈平' or result['name'] == '杨嘉仪' or result['name'] == '李晓青':
                # print(user_caigou)
                if result['name'] in user_caigou:
                    # print(77)
                    # print(result['name'])
                    ordersDict.append(result)
                    break
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e","采购异常补充明细-绩效数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        data = pd.json_normalize(ordersDict)
        print('++++++单次查询成功+++++++')
        print('*' * 50)
        return data

    # 绩效-查询 物流问题件 压单核实  （二.2）
    def service_id_waybill(self, timeStart, timeEnd, proxy_handle, proxy_id, order_time):  # 进入物流问题件界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在查询 物流问题件&压单核实 起止时间：' + str(timeStart) + " *** " + str(timeEnd))
        url = r'https://gimp.giikin.com/service?service=gorder.customerQuestion&action=getCustomerComplaintList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerQuestion'}
        data = {'order_number': None, 'waybill_no': None, 'transfer_no': None, 'gift_reissue_order_number': None, 'is_gift_reissue': None, 'order_trace_id': None, 'question_type': None, 'critical': None,
                'read_status': None, 'operator_type': None, 'operator': None, 'create_time': None, 'trace_time': None, 'is_collection': None, 'logistics_status': None, 'user_id': None, 'page': 1, 'pageSize': 90}
        data_woks = None
        data_woks2 = None
        data_woks3 = None
        if order_time == '跟进时间':
            data.update({'trace_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
            data_woks = '物流问题件_跟进时间'
            data_woks2 = '最新处理时间'
            data_woks3 = '压单核实_跟进时间'
        elif order_time == '创建时间':
            data.update({'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
            data_woks = '物流问题件_创建时间'
            data_woks2 = '导入时间'
            data_woks3 = '压单核实_创建时间'
        if proxy_handle == '代理服务器':        # print('+++已成功发送请求......')
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        max_count = req['data']['count']
        print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        print('*' * 50)
        if max_count != 0:
            df = pd.DataFrame([])
            in_count = math.ceil(max_count / 500)
            dlist = []
            n = 1
            while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                print('剩余查询次数' + str(in_count - n))
                data = self._service_id_waybill(timeStart, timeEnd, n, proxy_handle, proxy_id, order_time)
                dlist.append(data)
                n = n + 1
            dp = df.append(dlist, ignore_index=True)
            dp = dp[['order_number', 'currency', 'addtime', 'orderStatus', 'logisticsStatus', 'create_time', 'traceUserName', 'trace_UserName', 'contact', 'questionType', 'dealStatus',
                     'dealContent', 'deal_Content', 'dealTime', 'deal_time', 'result_info', 'result_reson', 'gift_reissue_order_number', 'giftStatus', 'questionTypeName', 'traceRecord']]
            dp.columns = ['订单编号', '币种', '下单时间', '订单状态', '物流状态', '导入时间', '最新处理人', '最新客服处理人', '联系方式', '跟进问题类型', '最新处理状态',
                          '最新处理结果', '最新客服处理', '最新处理时间', '最新客服处理日期', '拒收原因', '具体原因', '赠品补发订单编号', '赠品补发订单状态', '问题类型', '历史处理记录']
            print('正在写入 物流问题件......')
            dp1 = dp[~(dp['问题类型'].str.contains('订单压单（giikin内部专用）', na=False))]  # 筛选 问题类型
            dp1.to_excel('F:\\输出文件\\物流问题件-{0}{1}.xlsx'.format(data_woks2, rq), sheet_name='查询', index=False, engine='xlsxwriter')
            dp1.to_sql('cache_check', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO {0}(订单编号, 币种,下单时间, 订单状态,  物流状态, 导入时间,  最新处理人, 最新客服处理人,联系方式, 跟进问题类型, 最新处理状态, 最新处理结果, 最新客服处理,
                                        最新处理时间, 最新客服处理日期,拒收原因,具体原因,赠品补发订单编号, 赠品补发订单状态, 问题类型,历史处理记录,统计月份,记录时间)
                    SELECT 订单编号, 币种,下单时间, 订单状态,  物流状态, 导入时间,  最新处理人, 最新客服处理人,联系方式, 跟进问题类型,最新处理状态, 最新处理结果, 最新客服处理,
                            IF(最新处理时间 = '',NULL,最新处理时间) AS 最新处理时间, IF(最新客服处理日期 = '',NULL,最新客服处理日期) AS 最新客服处理日期,拒收原因,具体原因,
                            赠品补发订单编号, 赠品补发订单状态,问题类型,历史处理记录,DATE_FORMAT({1},'%Y%m') 统计月份,NOW() 记录时间
                    FROM cache_check;'''.format(data_woks, data_woks2)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

            print('正在写入 压单核实......')
            dp2 = dp[(dp['问题类型'].str.contains('订单压单（giikin内部专用）', na=False))]  # 筛选 问题类型
            dp2.to_excel('F:\\输出文件\\压单核实-{0}{1}.xlsx'.format(data_woks2, rq), sheet_name='查询', index=False, engine='xlsxwriter')
            dp2.to_sql('cache_check', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO {0}(订单编号, 币种,下单时间, 订单状态,  物流状态, 导入时间,  最新处理人, 最新客服处理人,联系方式, 跟进问题类型, 最新处理状态, 最新处理结果, 最新客服处理,
                                    最新处理时间, 最新客服处理日期,拒收原因,具体原因,赠品补发订单编号, 赠品补发订单状态, 问题类型,历史处理记录,统计月份,记录时间)
                    SELECT 订单编号, 币种,下单时间, 订单状态,  物流状态, 导入时间,  最新处理人, 最新客服处理人,联系方式, 跟进问题类型,最新处理状态, 最新处理结果, 最新客服处理,
                            IF(最新处理时间 = '',NULL,最新处理时间) AS 最新处理时间, IF(最新客服处理日期 = '',NULL,最新客服处理日期) AS 最新客服处理日期,拒收原因,具体原因,
                            赠品补发订单编号, 赠品补发订单状态,问题类型,历史处理记录,DATE_FORMAT({1},'%Y%m') 统计月份,NOW() 记录时间
                    FROM cache_check;'''.format(data_woks3, data_woks2)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
        else:
            print('没有需要获取的信息！！！')
            return
        print('-' * 50)
        print('-' * 50)
    def _service_id_waybill(self, timeStart, timeEnd, n, proxy_handle, proxy_id, order_time):  # 进入物流问题件界面
        print('+++正在查询第 ' + str(n) + ' 页信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customerQuestion&action=getCustomerComplaintList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerQuestion'}
        data = {'order_number': None, 'waybill_no': None, 'transfer_no': None, 'gift_reissue_order_number': None, 'is_gift_reissue': None, 'order_trace_id': None, 'question_type': None, 'critical': None,
                'read_status': None, 'operator_type': None, 'operator': None, 'create_time': None, 'trace_time': None,'is_collection': None, 'logistics_status': None, 'user_id': None, 'page': n, 'pageSize': 500}
        if order_time == '跟进时间':
            data.update({'trace_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
        elif order_time == '创建时间':
            data.update({'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
        if proxy_handle == '代理服务器':        # print('+++已成功发送请求......')
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersDict = []
        try:
            for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值
                if 'traceRecord' in result:
                    result['dealContent'] = zhconv.convert(result['dealContent'], 'zh-hans')
                    result['traceRecord'] = zhconv.convert(result['traceRecord'], 'zh-hans')
                    if '地址;' in result['traceRecord']:
                        result['traceRecord'] = result['traceRecord'].replace('地址;', '地址:')
                    if ';20' in result['traceRecord']:
                        trace_record = result['traceRecord'].split(";20")
                        result['deal_time'] = ''
                        result['result_reson'] = ''
                        result['result_info'] = ''
                        result['deal_Content'] = ''
                        result['trace_UserName'] = ''
                        for i in range(len(trace_record)):
                            if i == 0:
                                record = trace_record[i]
                            else:
                                record = '20' + trace_record[i]
                    # if ';' in result['traceRecord'] and '地址' not in result['traceRecord']:
                    #     # print(55)
                    #     # print(result['order_number'])
                    #     trace_record = result['traceRecord'].split(";")
                    #     result['deal_time'] = ''
                    #     result['result_reson'] = ''
                    #     result['result_info'] = ''
                    #     result['deal_Content'] = ''
                    #     result['trace_UserName'] = ''
                    #     for record in trace_record:
                            if '物流' not in record and '香港立邦' not in record:
                                if record.split("#处理结果：")[1] != '' and len(record.split("#处理结果：")[1]) > 1:
                                    result['deal_time'] = record.split()[0]
                                    rec = record.split("#处理结果：")[1]
                                    if len(rec.split()) > 2:
                                        result['result_info'] = rec.split()[2]  # 客诉原因
                                    if len(rec.split()) > 1:
                                        result['result_reson'] = rec.split()[1]  # 处理内容
                                    result['deal_Content'] = rec.split()[0]  # 最新处理结果
                                    rec_name = record.split("#处理结果：")[0]
                                    if '客服' in rec_name:
                                        recname = (rec_name.split())[2]
                                        result['trace_UserName'] = recname.replace('(客服)', '')
                        ordersDict.append(result.copy())
                    else:
                        # print(44)
                        # print(result['order_number'])
                        result['deal_time'] = ''
                        result['result_reson'] = ''
                        result['result_info'] = ''
                        result['deal_Content'] = ''
                        result['trace_UserName'] = ''
                        if '拒收' in result['dealContent']:
                            if len(result['dealContent'].split()) > 2:
                                result['result_info'] = result['dealContent'].split()[2]
                            if len(result['dealContent'].split()) > 1:
                                result['result_reson'] = result['dealContent'].split()[1]
                            result['deal_Content'] = result['dealContent'].split()[0]
                        else:
                            result['deal_Content'] = result['dealContent'].strip()
                        if result['traceRecord'] != '' and result['traceRecord'] != []:
                            result['deal_time'] = result['traceRecord'].split()[0]
                        if result['traceUserName'] != '' and result['traceUserName'] != []:
                            result['trace_UserName'] = result['traceUserName'].replace('客服：', '')
                        ordersDict.append(result.copy())
                else:
                    result['deal_time'] = result['update_time']
                    result['result_reson'] = ''
                    result['result_info'] = ''
                    result['deal_Content'] = ''
                    result['trace_UserName'] = ''
                    ordersDict.append(result.copy())
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e","压单核实-绩效数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        data = pd.json_normalize(ordersDict)
        print('++++++第 ' + str(n) + ' 批次查询成功+++++++')
        print('*' * 50)
        return data

    # 绩效-查询 物流客诉件           （二.3）
    def service_id_waybill_Query(self, timeStart, timeEnd, proxy_handle, proxy_id, order_time):  # 进入物流客诉件界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在查询 物流客诉件(' + order_time + ') 起止时间：' + str(timeStart) + " *** " + str(timeEnd))
        url = r'https://gimp.giikin.com/service?service=gorder.orderCustomerComplaint&action=getCustomerComplaintList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerComplaint'}
        data = {'order_number': None, 'waybill_no': None, 'transfer_no': None, 'order_trace_id': None,'question_type': None, 'critical': None, 'read_status': None, 'operator_type': None, 'operator': None,
                'create_time': None, 'trace_time': None,'is_gift_reissue': None,'is_collection': None, 'logistics_status': None, 'user_id': None, 'page': 1, 'pageSize': 90}
        data_woks = None
        data_woks2 = None
        if order_time == '跟进时间':
            data.update({'trace_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
            data_woks = '物流客诉件_跟进时间'
            data_woks2 = '最新处理时间'
        elif order_time == '创建时间':
            data.update({'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
            data_woks = '物流客诉件_创建时间'
            data_woks2 = '导入时间'
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        max_count = req['data']['count']
        print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        print('*' * 50)
        dp = None
        if max_count > 0:
            in_count = math.ceil(max_count / 500)
            df = pd.DataFrame([])
            dlist = []
            n = 1
            while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                print('剩余查询次数' + str(in_count - n))
                data = self._service_id_waybill_Query(timeStart, timeEnd, n, proxy_handle, proxy_id, order_time)
                n = n + 1
                dlist.append(data)
            dp = df.append(dlist, ignore_index=True)
        if dp.empty:
            print("今日无更新数据")
        else:
            dp = dp[['id', 'order_number', 'currency', 'addtime', 'areaName', 'payType', 'reassignmentTypeName', 'orderStatus', 'logisticsStatus', 'logisticsName', 'questionTypeName', 'create_time',
                     'dealStatus', 'dealTime', 'deal_time', 'traceUserName', 'trace_UserName', 'dealContent',
                     'deal_Content', 'result_content', 'result_info', 'result_reson','gift_reissue_order_number', 'giftStatus', 'contact', 'traceRecord']]
            dp.columns = ['id', '订单编号', '币种', '下单时间', '归属团队', '支付类型', '订单类型', '订单状态', '物流状态', '物流渠道', '问题类型', '导入时间', '最新处理状态', '最新处理时间', '最新客服处理日期', '最新处理人',
                          '最新客服处理人', '最新处理结果', '最新客服处理', '最新客服处理结果', '客诉原因', '具体原因', '赠品补发订单编号', '赠品补发订单状态', '联系方式', '历史处理记录']
            print('正在写入......')
            dp.to_sql('cache_check', con=self.engine1, index=False, if_exists='replace')
            dp.to_excel('F:\\输出文件\\物流客诉件-{0}{1}.xlsx'.format(order_time, rq), sheet_name='查询', index=False,  engine='xlsxwriter')
            sql = '''REPLACE INTO {0}(id,订单编号,币种,下单时间,归属团队,支付类型, 订单类型, 订单状态, 物流状态, 物流渠道,问题类型, 导入时间,
                                        最新处理状态,最新处理时间,最新客服处理日期,最新处理人,最新客服处理人,最新处理结果,最新客服处理,最新客服处理结果,客诉原因,具体原因,
                                        赠品补发订单编号,赠品补发订单状态,赠品补发物流状态,联系方式,历史处理记录,统计月份,记录时间) 
                    SELECT id,订单编号,币种,下单时间,归属团队,支付类型, 订单类型, 订单状态, 物流状态, 物流渠道,问题类型, 导入时间,
                            最新处理状态,IF(最新处理时间 = '' OR 最新处理时间 IS NULL,导入时间,最新处理时间) as 最新处理时间, IF(最新客服处理日期 = '' OR 最新客服处理日期 IS NULL,DATE_FORMAT(导入时间,'%Y-%m-%d'),最新客服处理日期) as 最新客服处理日期,
                            最新处理人,最新客服处理人,最新处理结果,最新客服处理,最新客服处理结果,客诉原因,具体原因, 赠品补发订单编号,赠品补发订单状态,null 赠品补发物流状态,联系方式,历史处理记录,DATE_FORMAT({1},'%Y%m') 统计月份,NOW() 记录时间 
                    FROM cache_check;'''.format(data_woks, data_woks2)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')

            print('++++++订单状态明细查询中+++++++')
            order_list = list(dp['赠品补发订单编号'])
            max_count = len(order_list)
            in_count = math.ceil((max_count + 500) / 500)
            pople_Query = '物流客诉查询'
            df2 = pd.DataFrame([])
            dtlist = []
            n = 0
            while n < max_count + 500:
                ord = ','.join(order_list[n:n + 500])
                data = self._service_id_order("", "", 0, proxy_handle, proxy_id, pople_Query, ord)  # 查询全部订单信息
                dtlist.append(data)
                n = n + 500
                print('剩余查询次数' + str(in_count - ((n + 500)/500)))
            dp2 = df2.append(dtlist, ignore_index=True)
            print(99)
            print(dp2)
            dp3 = dp2[['orderNumber', 'addTime', 'orderStatus', 'logisticsStatus']]
            dp3.columns = ['订单编号', '下单时间', '订单状态', '物流状态']
            dp3.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            sql = '''update {0} a, customer b
                       set a.`赠品补发订单状态`= IF(b.`订单状态` = '', NULL, b.`订单状态`),
                           a.`赠品补发物流状态`= IF(b.`物流状态` = '', NULL, b.`物流状态`)
                    where a.`赠品补发订单编号`=b.`订单编号`;'''.format(data_woks)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('订单更新成功......')
        print('*' * 50)
    def _service_id_waybill_Query(self, timeStart, timeEnd, n, proxy_handle, proxy_id, order_time):  # 进入物流客诉件界面
        print('+++正在查询第 ' + str(n) + ' 页信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.orderCustomerComplaint&action=getCustomerComplaintList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerComplaint'}
        data = {'order_number': None, 'waybill_no': None, 'transfer_no': None, 'order_trace_id': None,'question_type': None, 'critical': None, 'read_status': None,'operator_type': None, 'operator': None,
                'create_time': None, 'trace_time': None,'is_gift_reissue': None,'is_collection': None, 'logistics_status': None, 'user_id': None, 'page': n, 'pageSize': 500}
        if order_time == '跟进时间':
            data.update({'trace_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
        elif order_time == '创建时间':
            data.update({'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersDict = []
        try:
            for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值用
                # print(result['order_number'])
                # print(result)
                if 'traceRecord' in result:
                    result['dealContent'] = zhconv.convert(result['dealContent'], 'zh-hans')
                    result['traceRecord'] = zhconv.convert(result['traceRecord'], 'zh-hans')
                    if ';' in result['traceRecord']:
                        trace_record = result['traceRecord'].split(";")
                        result['deal_time'] = ''
                        result['result_reson'] = ''
                        result['result_info'] = ''
                        result['result_content'] = ''
                        result['deal_Content'] = ''
                        result['trace_UserName'] = ''
                        for record in trace_record:
                            if '物流' not in record:
                                rec = record.split("#处理结果：")[1]
                                # print(record)
                                # print(rec)
                                if rec != "" and rec != " ":
                                    result['deal_time'] = record.split()[0]
                                    if len(rec.split()) > 3:
                                        result['result_reson'] = rec.split()[3]  # 最新客服 具体原因
                                    if len(rec.split()) > 2:
                                        result['result_info'] = rec.split()[2]  # 最新客服 客诉原因
                                    if len(rec.split()) > 1:
                                        result['result_content'] = rec.split()[1]  # 最新客服 处理结果
                                    result['deal_Content'] = rec.split()[0]  # 最新客服 处理
                                    rec_name = record.split("#处理结果：")[0]
                                    if '客服' in rec_name:
                                        recname = (rec_name.split())[2]
                                        result['trace_UserName'] = recname.replace('(客服)', '')
                        ordersDict.append(
                            result.copy())  # append()方法只是将字典的地址存到list中，而键赋值的方式就是修改地址，所以才导致覆盖的问题;  使用copy() 或者 deepcopy()  当字典中存在list的时候需要使用deepcopy()
                    else:
                        result['deal_time'] = ''
                        result['result_reson'] = ''
                        result['result_info'] = ''
                        result['result_content'] = ''
                        result['deal_Content'] = ''
                        result['trace_UserName'] = ''
                        if len(result['dealContent'].split()) > 3:
                            result['result_reson'] = result['dealContent'].split()[3]  # 最新客服 具体原因
                        if len(result['dealContent'].split()) > 2:
                            result['result_info'] = result['dealContent'].split()[2]  # 最新客服 客诉原因
                        if len(result['dealContent'].split()) > 1:
                            result['result_content'] = result['dealContent'].split()[1]  # 最新客服 处理内容
                        result['deal_Content'] = result['dealContent'].split()[0]  # 最新客服 处理

                        if result['traceRecord'] != '' or result['traceRecord'] != []:
                            result['deal_time'] = result['traceRecord'].split()[0]
                        if result['traceUserName'] != '' or result['traceUserName'] != []:
                            # if '赠品' in result['traceRecord'] or '退款' in result['traceRecord'] or '补发' in result['traceRecord'] or '换货' in result['traceRecord']:
                            if '客服' in result['traceRecord']:
                                result['trace_UserName'] = result['traceUserName'].replace('客服：', '')
                        ordersDict.append(result.copy())
                else:
                    result['deal_time'] = ''
                    result['result_reson'] = ''
                    result['result_info'] = ''
                    result['result_content'] = ''
                    result['deal_Content'] = ''
                    result['trace_UserName'] = ''
                    ordersDict.append(result.copy())
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e","物流客诉件-绩效数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        data = pd.json_normalize(ordersDict)
        print('++++++第 ' + str(n) + ' 批次查询成功+++++++')
        print('*' * 50)
        return data

    # 绩效-查询 派送问题件           （二.4）
    def service_id_getDeliveryList(self, timeStart, timeEnd, order_time, proxy_handle, proxy_id):  # 进入订单检索界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在查询 派送问题件(' + order_time + ') 起止时间：' + str(timeStart) + " *** " + str(timeEnd))
        url = r'https://gimp.giikin.com/service?service=gorder.deliveryQuestion&action=getDeliveryList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/deliveryProblemPackage'}
        data = {'order_number': None, 'waybill_number': None, 'question_level': None, 'question_type': None, 'order_trace_id': None, 'ship_phone': None, 'page': 1, 'pageSize': 90, 'addtime': None,  'question_time': None, 'trace_time': None,
                'create_time': None,'finishtime': None, 'sale_id': None, 'product_id': None,'logistics_id': None, 'area_id': None, 'currency_id': None, 'order_status': None,'logistics_status': None}
        data_woks = None
        data_woks2 = None
        if order_time == '处理时间':
            data.update({'trace_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
            data_woks = '派送问题件_处理时间'
            data_woks2 = '处理时间'
        elif order_time == '创建时间':
            data.update({'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
            data_woks = '派送问题件_创建时间'
            data_woks2 = '创建时间'
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        if req['data'] != []:
            max_count = req['data']['count']  # 获取 请求订单量
            print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
            print('*' * 50)
            if max_count != 0 and max_count != []:
                df = pd.DataFrame([])  # 创建空的dataframe
                dlist = []  # 创建空的列表 放每次查询的结果
                in_count = math.ceil(max_count / 90)
                n = 1
                while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                    data = self._service_id_getDeliveryList(timeStart, timeEnd, n, order_time, proxy_handle, proxy_id)
                    dlist.append(data)
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                    time.sleep(1)
                dp = df.append(dlist, ignore_index=True)
                dp = dp[['id', 'order_number', 'currency', 'addtime', 'orderStatus', 'logisticsStatus', 'logisticsName', 'create_time', 'lastQuestionName', 'contactName', 'userName', 'traceName', 'content',
                         'traceTime', 'failNum', 'questionAddtime', 'questionTypeName']]
                dp.columns = ['id', '订单编号', '币种', '下单时间', '订单状态', '物流状态', '物流渠道', '创建时间', '派送问题类型', '联系方式', '最新处理人',
                              '最新处理状态', '最新处理结果', '处理时间', '派送次数', '最新抓取时间', '最新问题类型']
                print('正在写入......')
                dp.to_sql('cache_check', con=self.engine1, index=False, if_exists='replace')
                dp.to_excel('F:\\输出文件\\派送问题件-{0}{1}.xlsx'.format(order_time, rq), sheet_name='查询', index=False, engine='xlsxwriter')
                sql = '''REPLACE INTO {0}(id,订单编号,币种, 下单时间,订单状态,物流状态,物流渠道,创建时间,派送问题类型, 联系方式,最新处理人, 最新处理状态, 最新处理结果,处理时间,派送次数,最新抓取时间,最新问题类型,统计月份, 记录时间) 
                        SELECT id,订单编号,币种, 下单时间,订单状态,物流状态,物流渠道,创建时间,派送问题类型, 联系方式,最新处理人, 最新处理状态, 最新处理结果,IF(处理时间 = '',NULL,处理时间) 处理时间,派送次数,IF(最新抓取时间 = '',NULL,最新抓取时间) 最新抓取时间,最新问题类型,DATE_FORMAT({1},'%Y%m') 统计月份, NOW() 记录时间 
                        FROM cache_check;'''.format(data_woks, data_woks2)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                print('写入成功......')
        else:
            print('没有需要获取的信息！！！')
            return
        print('-' * 50)
        print('-' * 50)
    def _service_id_getDeliveryList(self, timeStart, timeEnd, n, order_time, proxy_handle, proxy_id):  # 进入派送问题件界面
        print('+++正在查询第 ' + str(n) + ' 页信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.deliveryQuestion&action=getDeliveryList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/deliveryProblemPackage'}
        data = {'order_number': None, 'waybill_number': None, 'question_level': None, 'question_type': None, 'order_trace_id': None, 'ship_phone': None, 'page': n,'pageSize': 90, 'addtime': None, 'question_time': None,
                'trace_time': None, 'create_time': None, 'finishtime': None, 'sale_id': None, 'product_id': None, 'logistics_id': None, 'area_id': None, 'currency_id': None, 'order_status': None, 'logistics_status': None}
        if order_time == '处理时间':
            data.update({'trace_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
        elif order_time == '创建时间':
            data.update({'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersDict = []
        try:
            if req['data'] != []:
                for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值
                    # print(result['order_number'])
                    # result['traceRecord'] = zhconv.convert(result['traceRecord'], 'zh-hans')
                    ordersDict.append(result.copy())
            else:
                return None
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e","派送问题件-绩效数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        data = pd.json_normalize(ordersDict)
        print('++++++第 ' + str(n) + ' 批次查询成功+++++++')
        print('*' * 50)
        return data

    # 绩效-查询 拒收问题件           （二.50）
    def service_id_order_js_Query(self, timeStart, timeEnd, proxy_handle, proxy_id, order_time):  # 进入拒收问题件界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在查询 拒收问题件(' + order_time + ') 起止时间：' + str(timeStart) + " *** " + str(timeEnd))
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getRejectList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36 Edg/111.0.1661.62',
                    'origin': 'https://gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerRejection'}
        data = {'page': 1, 'pageSize': 100, 'orderPrefix': None, 'shipUsername': None, 'shippingNumber': None,'email': None, 'saleIds': None, 'ip': None,'productIds': None, 'phone': None,
                'optimizer': None, 'payment': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None, 'orderStatus': None,
                'timeStart': None,'timeEnd': None, 'payType': None, 'questionId': None,'autoVerifys': None, 'reassignmentType': None, 'logisticsStatus': None, 'logisticsId': None,
                'traceItemIds': None, 'finishTimeStart': None,'finishTimeEnd': None, 'traceTimeStart': None, 'traceTimeEnd': None, 'newCloneNumber': None}
        self.session.mount('http://', HTTPAdapter(max_retries=5))
        self.session.mount('https://', HTTPAdapter(max_retries=5))
        data_woks = None
        data_woks2 = None
        if order_time == '跟进时间':
            data.update(
                {'traceItemIds': -1, 'traceTimeStart': timeStart + ' 00:00:00,', 'traceTimeEnd': timeEnd + ' 23:59:59'})
            data_woks = '拒收问题件_跟进时间'
            data_woks2 = '处理时间'
        elif order_time == '下单跟进时间':
            data.update({'traceItemIds': -1, 'timeStart': timeStart + ' 00:00:00,', 'timeEnd': timeEnd + ' 23:59:59'})
            data_woks = '拒收问题件_下单跟进时间'
            data_woks2 = '下单时间'
        elif order_time == '下单时间':
            # for i in range((timeEnd - timeStart).days):  # 按天循环获取订单状态
            #     day = timeStart + datetime.timedelta(days=i)
            #     day_time = str(day)
            #     data.update({'timeStart': day_time + ' 00:00:00,', 'timeEnd': day_time + ' 23:59:59'})
            data.update({'timeStart': timeStart + ' 00:00:00,', 'timeEnd': timeEnd + ' 23:59:59'})
            data_woks = '拒收问题件_下单时间'
            data_woks2 = '下单时间'
        try:
            if proxy_handle == '代理服务器':
                proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
                req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies, timeout=None)
            else:
                req = self.session.post(url=url, headers=r_header, data=data, timeout=None)
            print(req)
            req = json.loads(req.text)  # json类型数据转换为dict字典
            max_count = req['data']['count']
        except requests.exceptions.RequestException as e:
            print(e)
        # print('+++已成功发送请求......')
        print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        # print(req)
        if max_count != 0:
            df = pd.DataFrame([])
            in_count = math.ceil(max_count / 100)
            dlist = []
            n = 1
            while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                print('剩余查询次数' + str(in_count - n))
                data = self._service_id_order_js_Query(timeStart, timeEnd, n, proxy_handle, proxy_id, order_time)
                n = n + 1
                dlist.append(data)
                time.sleep(3)
            dp = df.append(dlist, ignore_index=True)
            dp = dp[['id', '订单编号', 'currency', 'percentInfo.orderCount', 'percentInfo.rejectCount','percentInfo.arriveCount', 'addTime', 'finishTime', 'tel_phone', 'shipInfo.shipPhone', 'ip',
                     'cloneUser', 'newCloneUser', 'newCloneStatus', 'newCloneLogisticsStatus', '再次克隆下单', '跟进人', '时间', '联系方式', '问题类型', '问题原因', '内容', '处理结果', '是否需要商品']]
            dp.columns = ['id', '订单编号', '币种', '订单总量', '拒收量', '签收量', '下单时间', '完成时间', '电话', '联系电话', 'ip', '本单克隆人',
                          '新单克隆人', '新单订单状态', '新单物流状态', '再次克隆下单', '处理人', '处理时间', '联系方式', '核实原因', '具体原因', '备注', '处理结果','是否需要商品']
            print('正在写入......')
            dp.to_excel('F:\\输出文件\\拒收问题件-{0}{1}.xlsx'.format(order_time, rq), sheet_name='查询', index=False, engine='xlsxwriter')
            dp.to_sql('cache_check', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO {0}(id, 订单编号,币种, 订单总量, 拒收量, 签收量, 下单时间, 完成时间, 电话, 联系电话, ip, 本单克隆人, 新单克隆人, 新单订单状态, 新单物流状态, 再次克隆下单,处理人,处理时间,联系方式, 核实原因, 具体原因, 备注, 处理结果, 是否需要商品,统计月份,记录时间)
                     SELECT id, 订单编号,币种, 订单总量, 拒收量, 签收量, 下单时间, 完成时间, IF(电话 LIKE "852%",RIGHT(电话,LENGTH(电话)-3),IF(电话 LIKE "886%",RIGHT(电话,LENGTH(电话)-3),电话)) 电话, 联系电话, ip, 本单克隆人, 新单克隆人, 新单订单状态, 新单物流状态,  IF(再次克隆下单 = '',NULL,再次克隆下单) 再次克隆下单,处理人,IF(处理时间 = '',NULL,处理时间) AS 处理时间,联系方式, 核实原因, 具体原因, 备注, 处理结果,是否需要商品, DATE_FORMAT({1},'%Y%m') 统计月份,NOW() 记录时间
                    FROM cache_check;'''.format(data_woks, data_woks2)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
        else:
            print('****** 没有信息！！！')
        print('*' * 50)
    def _service_id_order_js_Query(self, timeStart, timeEnd, n, proxy_handle, proxy_id, order_time):  # 进入拒收问题件界面
        print('+++正在查询第 ' + str(n) + ' 页信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getRejectList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36 Edg/111.0.1661.62',
                    'origin': 'https://gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerRejection'}
        data = {'page': n, 'pageSize': 100, 'orderPrefix': None, 'shipUsername': None, 'shippingNumber': None, 'email': None, 'saleIds': None, 'ip': None,'productIds': None, 'phone': None,
                'optimizer': None, 'payment': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None, 'orderStatus': None,
                'timeStart': None,'timeEnd': None, 'payType': None, 'questionId': None, 'autoVerifys': None, 'reassignmentType': None, 'logisticsStatus': None, 'logisticsId': None,
                'traceItemIds': None, 'finishTimeStart': None, 'finishTimeEnd': None, 'traceTimeStart': None, 'traceTimeEnd': None, 'newCloneNumber': None}
        self.session.mount('http://', HTTPAdapter(max_retries=5))
        self.session.mount('https://', HTTPAdapter(max_retries=5))
        if order_time == '跟进时间':
            data.update(
                {'traceItemIds': -1, 'traceTimeStart': timeStart + ' 00:00:00,', 'traceTimeEnd': timeEnd + ' 23:59:59'})
        elif order_time == '下单跟进时间':
            data.update({'traceItemIds': -1, 'timeStart': timeStart + ' 00:00:00,', 'timeEnd': timeEnd + ' 23:59:59'})
        elif order_time == '下单时间':
            data.update({'timeStart': timeStart + ' 00:00:00,', 'timeEnd': timeEnd + ' 23:59:59'})
        # print(req)
        # print(req.text)
        try:
            if proxy_handle == '代理服务器':
                proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
                req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies, timeout=None)
            else:
                req = self.session.post(url=url, headers=r_header, data=data, timeout=None)
            print('+++已成功发送请求......')
            req = json.loads(req.text)  # json类型数据转换为dict字典
        except Exception as e:
            print('转化失败： 请求失败', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e", "拒收问题件-绩效数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        # print(req)
        ordersDict = []
        try:
            for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值用
                # print(result['orderNumber'])
                result['订单编号'] = result['orderNumber']
                result['再次克隆下单'] = result['newCloneNumber']
                result['跟进人'] = ''
                result['时间'] = ''
                result['内容'] = ''
                result['联系方式'] = ''
                result['问题类型'] = ''
                result['问题原因'] = ''
                result['处理结果'] = ''
                result['是否需要商品'] = ''
                if result['traceItems'] != []:
                    # print(result['traceItems'])
                    for res in result['traceItems']:
                        resval = res.split(':')[0]
                        if '跟进人' in resval:
                            result['跟进人'] = (res.split('跟进人:')[1]).strip()  # 跟进人
                        if '时间' in resval and '跟进' not in resval:
                            result['时间'] = (res.split('时间:')[1]).strip()  # 跟进人
                        if '内容' in resval:
                            result['内容'] = (res.split('内容:')[1]).strip()  # 跟进人
                        if '联系方式' in resval:
                            result['联系方式'] = (res.split('联系方式:')[1]).strip()  # 跟进人
                        if '问题类型' in resval:
                            result['问题类型'] = (res.split('问题类型:')[1]).strip()  # 跟进人
                        if '问题原因' in resval:
                            result['问题原因'] = (res.split('问题原因:')[1]).strip()  # 跟进人
                        if '处理结果' in resval:
                            result['处理结果'] = (res.split('处理结果:')[1]).strip()  # 跟进人
                        if '是否需要商品' in resval:
                            result['是否需要商品'] = (res.split('是否需要商品:')[1]).strip()  # 跟进人

                ordersDict.append(result.copy())
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e", "拒收问题件-绩效数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        data = pd.json_normalize(ordersDict)
        print('++++++第 ' + str(n) + ' 批次查询成功+++++++')
        print('*' * 50)
        return data

    # 绩效-查询系统问题件         （三）
    def service_id_orderInfo(self, timeStart, timeEnd, proxy_handle, proxy_id):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在查询 系统问题件 起止时间：' + str(timeStart) + " *** " + str(timeEnd))
        sql = '''SELECT 订单编号 FROM gat_order_list g WHERE (g.`日期` BETWEEN '{0}' AND '{1}')  AND g.`问题原因` IS NOT NULL;'''.format(timeStart, timeEnd)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        orderId = list(df['订单编号'])
        max_count = len(orderId)
        print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        n = 0
        if max_count > 0:
            df = pd.DataFrame([])
            dlist = []
            while n <= max_count:  # 这里用到了一个while循环，穿越过来的
                ord = ','.join(orderId[n:n + 10])
                data = self._service_id_orderInfo(ord, proxy_handle, proxy_id)
                dlist.append(data)
                print('剩余查询次数' + str(math.ceil((max_count - n) / 10)))
                n = n + 10
            dp = df.append(dlist, ignore_index=True)
            print('正在写入......')
            dp.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
            dp.to_excel('F:\\输出文件\\系统问题件-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
            sql = '''REPLACE INTO gat_order_list_log(id, orderId, orderNumber, orderStatus, updateTime, uid, remark, name, 统计月份,记录时间) 
                                              SELECT id, orderId, orderNumber, orderStatus, updateTime, uid, remark, name, DATE_FORMAT('{0}','%Y%m') 统计月份, NOW() 记录时间 
                    FROM cache_ch;'''.format(timeEnd)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
            print('*' * 100)
            self._service_id_orderInfoTWO(timeEnd)
            print('*' * 100)
            self._service_id_orderInfoThree(timeEnd)
        else:
            print('无需查询......')
    def _service_id_orderInfo(self, ord, proxy_handle, proxy_id):  # 进入订单检索界面
        print('+++正在查询订单信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getOrderLog'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'orderKey': ord}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        try:
            for result in req['data']:
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e","系统问题件-绩效数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        data = pd.json_normalize(ordersdict)
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return data
    def _service_id_orderInfoTWO(self, timeEnd):  # 进入订单检索界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('+++正在查询订单 转化人 信息中')
        sql = '''SELECT *
                FROM gat_order_list_log c
                WHERE c.统计月份 = DATE_FORMAT('{0}','%Y%m') AND c.orderStatus IS NOT NULL AND c.orderStatus <> ""
                ORDER BY orderNumber, id;'''.format(timeEnd)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        print(82)
        print(df)
        dict = {}
        for index, x in df.iterrows():
            print(index, x['id'], x['orderNumber'], x['orderStatus'], x['updateTime'], x['remark'], x['name'])
            order_Number = x['orderNumber']
            dict_info = {}
            if order_Number not in dict:
                dict_info['订单编号'] = x['orderNumber']
                dict_info['id'] = x['id']
                dict_info['订单状态'] = x['orderStatus']
                dict_info['转化时间'] = x['updateTime']
                dict_info['备注'] = x['remark']
                dict_info['转化人'] = '0'
                dict[order_Number] = dict_info
                step = 0
            else:
                if step >= 1:
                    continue
                order_Number_last = dict[order_Number]['订单编号']
                id_Status = x['id']
                id_Status_last = dict[order_Number]['id']
                order_Status = x['orderStatus']
                order_Status_last = dict[order_Number]['订单状态']
                update_Time = x['updateTime']
                update_Time_last = dict[order_Number]['转化时间']
                remark_Status = x['remark']
                remark_Status_last = dict[order_Number]['备注']
                name_Status = x['name']
                name_Status_last = dict[order_Number]['转化人']
                # print(name_Status_last)
                if order_Status == '问题订单':
                    # print('已删除 待发货不在')
                    # print('转化人:' + str(name_Status_last))
                    dict_info['订单编号'] = x['orderNumber']
                    dict_info['id'] = x['id']
                    dict_info['订单状态'] = x['orderStatus']
                    dict_info['转化时间'] = x['updateTime']
                    dict_info['备注'] = x['remark']
                    if name_Status_last != '0':
                        if name_Status_last == '蔡利英' or name_Status_last == '杨嘉仪' or name_Status_last == '张陈平' or name_Status_last == '李晓青':
                            dict_info['id'] = id_Status_last
                            dict_info['订单编号'] = order_Number_last
                            dict_info['订单状态'] = order_Status_last
                            dict_info['转化时间'] = update_Time_last
                            dict_info['备注'] = remark_Status_last
                            dict_info['转化人'] = name_Status_last
                        else:
                            dict_info['转化人'] = x['name']
                    else:
                        dict_info['转化人'] = x['name']
                    dict[order_Number] = dict_info
                elif '已删除' in order_Status or '待发货' in order_Status:
                    if order_Status_last == "问题订单":
                        step = step + 1
                        if '修改order_status' in remark_Status:
                            if '蔡利英' in name_Status or '杨嘉仪' in name_Status or '张陈平' in name_Status or '李晓青' in name_Status:
                                dict_info['id'] = x['id']
                                dict_info['订单编号'] = x['orderNumber']
                                dict_info['订单状态'] = x['orderStatus']
                                dict_info['转化时间'] = x['updateTime']
                                dict_info['备注'] = x['remark']
                                dict_info['转化人'] = x['name']
                            else:
                                if '修改remark,->张' in remark_Status_last or '修改remark,->楊' in remark_Status_last or '修改remark,->英' in remark_Status_last or '修改remark,->李' in remark_Status_last:
                                    dict_info['订单编号'] = order_Number_last
                                    dict_info['id'] = id_Status_last
                                    dict_info['订单状态'] = order_Status_last
                                    dict_info['转化时间'] = update_Time_last
                                    dict_info['备注'] = remark_Status_last
                                    dict_info['转化人'] = name_Status_last
                                else:
                                    dict_info['id'] = x['id']
                                    dict_info['订单编号'] = x['orderNumber']
                                    dict_info['订单状态'] = x['orderStatus']
                                    dict_info['转化时间'] = x['updateTime']
                                    dict_info['备注'] = x['remark']
                                    dict_info['转化人'] = x['name']
                            pass
                        else:
                            pass
                    else:
                        pass
                    # print(dict_info)
                    if dict_info != {}:
                        dict[order_Number] = dict_info
        print('*' * 52)
        data = list(dict.values())
        data = pd.json_normalize(data)
        print(data)
        data.to_excel('F:\\输出文件\\系统问题件-下单时间{0}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
        data.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO 系统问题件_下单时间( id, 订单编号, 订单状态, 转化时间, 备注, 转化人,  系统订单状态, 系统物流状态, 统计月份,记录时间) 
                 SELECT id, 订单编号, 订单状态, 转化时间, 备注, 转化人, NULL AS 系统订单状态, NULL AS 系统物流状态, DATE_FORMAT('{0}','%Y%m') 统计月份, NOW() 记录时间 
                 FROM cache_ch;'''.format(timeEnd)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('写入成功......')
    def _service_id_orderInfoThree(self, timeEnd):  # 进入订单检索界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在更新 订单状态 信息…………')
        sql = '''SELECT 订单编号 FROM 系统问题件_下单时间 c WHERE c.统计月份 = DATE_FORMAT('{1}','%Y%m');'''.format('gat', timeEnd)
        ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
        print(ordersDict)

        if ordersDict.empty:
            print(' ****** 没有要补充的订单信息; ****** ')
        else:
            print('！！！ 请再次更新订单编号数据！！！')
            proxy_handle = '代理服务器0'
            proxy_id = '192.168.13.89:37467'  # 输入代理服务器节点和端口
            handle = '手0动'
            login_TmpCode = '517e55c6fb6c34ca99a69874aaf5ec25'  # 输入登录口令Tkoen
            js = QueryOrder('+86-18538110674', 'qyz04163510.', login_TmpCode, handle, proxy_handle, proxy_id)

            orders_Dict = list(ordersDict['订单编号'])
            max_count = len(orders_Dict)
            if max_count > 0:
                in_count = math.ceil(max_count / 500)
                df = pd.DataFrame([])
                dlist = []
                n = 0
                while n < in_count:  # 这里用到了一个while循环，穿越过来的
                    print('查询第 ' + str(n + 1) + ' 页中，剩余次数' + str(in_count - n - 1))
                    n1 = n * 500
                    n2 = (n + 1) * 500
                    ord = ','.join(orders_Dict[n1:n2])
                    data = js.orderInfoQuery(ord, '订单号', proxy_id, proxy_handle)
                    # print(data)
                    dlist.append(data)
                    n = n + 1
                    print(n)
                print('正在写入......')
                dp = df.append(dlist, ignore_index=True)
                dp = dp[['orderNumber', 'logisticsStatus', 'orderStatus']]
                dp.to_sql('cache_check', con=self.engine1, index=False, if_exists='replace')
                sql = '''update `系统问题件_下单时间` a, cache_check b
                            SET a.`系统订单状态` = IF(b.`orderStatus` = '', NULL, b.`orderStatus`),
                                a.`系统物流状态` = IF(b.`logisticsStatus` = '', NULL, b.`logisticsStatus`)
                        where a.`订单编号`=b.`orderNumber`;'''
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            else:
                print('无需补充数据')
        print('更新成功......')
    '''
         # 绩效-汇总输出 - 单独获取使用    # 单独 本月绩效数据使用 不包含上月的留底数据
    '''
    # 单独 本月绩效数据使用 不包含上月的留底数据
    def service_check(self, username_Cudan, username_Jushou, username_caigou_yadan_wentijian, month_time, day_time):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        listT = []
        print('挽单列表-绩效 数据整理 获取中 中（零）......')
        sql11 = '''SELECT *, IF(当前物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 当前物流状态, IF(当前物流状态 IN ('已签收','理赔'), IF(当前订单状态 = '已退货(销售)','拒收',当前物流状态), 
                            IF(当前物流状态 = '发货中','在途', IF(当前物流状态 = '' or 当前物流状态 IS NULL or 当前物流状态 = '暂无物流状态', IF(当前订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),当前物流状态)))) as 最终状态
                FROM 挽单列表_创建时间 s1
                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND 删除人 = '';'''.format(month_time, day_time)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)

        print('促单-绩效 源数据 获取中（一.1）......')
        sql21 = '''SELECT *, IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                            IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态
                FROM 促单_下单时间 s1
                WHERE s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}';'''.format(month_time, day_time, username_Cudan)
        df21 = pd.read_sql_query(sql=sql21, con=self.engine1)
        listT.append(df21)
        print('促单-绩效 统计分析 获取中（一.2）......')
        sql22 = '''SELECT 代下单客服, IF(已签收 = 0,NULL, 已签收) AS 已签收,  IF(拒收 = 0,NULL, 拒收) AS 拒收, IF(已退货 = 0,NULL, 已退货) AS 已退货,  IF(理赔 = 0,NULL, 理赔) AS 理赔, 
                        IF(未发货 = 0,NULL, 未发货) AS 未发货, IF(在途 = 0,NULL, 在途) AS 在途, IF(已完成 = 0,NULL, 已完成) AS 已完成,  (总计-无效订单) AS 有效单量, 总计, 
                        concat(ROUND(IFNULL(已签收 / 已完成,0) * 100,2),'%') AS 签收率,
                        concat(ROUND(IFNULL(已完成 / (总计-无效订单),0) * 100,2),'%') AS 完成占比,IF(无效订单 = 0,NULL, 无效订单) AS 无效订单, 
                        concat(ROUND(IFNULL((总计-无效订单) / 总计,0) * 100,2),'%') AS 转换率
                FROM ( SELECT 代下单客服, count(订单编号) as 总计,
                                SUM(IF(最终状态 = "已签收",1,0)) as 已签收, SUM(IF(最终状态 = "拒收",1,0)) as 拒收, SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔, SUM(IF(最终状态 = "未发货",1,0)) as 未发货, SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成, SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单
                        FROM (
                                (	SELECT '促单' as 类型, 代下单客服, 订单编号, 订单状态,物流状态, 
                                            IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                                            IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态, 统计月份
                                    FROM 促单_下单时间 s1
                                    WHERE s1.`统计月份` = '{0}' AND DATE_FORMAT( s1.`记录时间`, '%Y-%m-%d') = '{1}' AND s1.克隆人 = '' AND s1.代下单客服 in ({2})
                                )
                                UNION
                                (	SELECT '挽单列表' as 类型, 创建人, 订单编号,  当前订单状态, 当前物流状态, 
                                            IF(当前物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 当前物流状态, IF(当前物流状态 IN ('已签收','理赔'), IF(当前订单状态 = '已退货(销售)','拒收',当前物流状态),  IF(当前物流状态 = '发货中','在途',
                                            IF(当前物流状态 = '' or 当前物流状态 IS NULL or 当前物流状态 = '暂无物流状态', IF(当前订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),当前物流状态)))) as 最终状态, 统计月份
                                    FROM 挽单列表_创建时间 w
                                    WHERE  w.`统计月份` = '{0}' and DATE_FORMAT(w.`记录时间`, '%Y-%m-%d') = '{1}' AND w.创建人 in ({2}) AND w.删除人 = '' AND w.挽单类型 IN ("取消挽单","未支付/支付失败挽单")
                                )
                        ) s1
                        GROUP BY  代下单客服
                ) s ORDER BY FIELD(代下单客服,{2},'合计');'''.format(month_time, day_time, username_Cudan)
        df22 = pd.read_sql_query(sql=sql22, con=self.engine1)
        listT.append(df22)

        print('物流客诉-绩效 源数据 获取中（二.1）......')
        sql31 = '''SELECT *, IF(赠品补发订单编号 <> "",IF(最新客服处理结果 LIKE '%补发海外仓%','统计','不统计'),'不统计') AS 是否统计,										
                            IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                                IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态,
                            IF(赠品补发物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 赠品补发物流状态, IF(赠品补发物流状态 IN ('已签收','理赔'), IF(赠品补发订单状态 = '已退货(销售)','拒收',赠品补发物流状态), 
							    IF(赠品补发物流状态 = '发货中','在途',  IF(赠品补发物流状态 = '' or 赠品补发物流状态 IS NULL or 赠品补发物流状态 = '暂无物流状态', IF(赠品补发订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),赠品补发物流状态)))) as 赠品补发最终状态
                FROM 物流客诉件_创建时间 s1
                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}';'''.format(month_time, day_time)
        df31 = pd.read_sql_query(sql=sql31, con=self.engine1)
        listT.append(df31)
        print('物流客诉-绩效 统计分析 获取中（二.2）......')
        sql32 = '''SELECT 最新客服处理人,
						SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
						SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
						SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
						SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
						SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
						SUM(IF(最终状态 = "在途",1,0)) as 在途,
						SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
						SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单,
						count(订单编号) as 总计
			    FROM (
						(	SELECT '客诉退货' as 类型, 最新客服处理人, 订单编号, 赠品补发订单编号, 订单状态, 物流状态, 赠品补发订单状态, 赠品补发物流状态,
                                    IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                                        IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态,
                                    IF(赠品补发物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 赠品补发物流状态, IF(赠品补发物流状态 IN ('已签收','理赔'), IF(赠品补发订单状态 = '已退货(销售)','拒收',赠品补发物流状态), 
                                        IF(赠品补发物流状态 = '发货中','在途', IF(赠品补发物流状态 = '' or 赠品补发物流状态 IS NULL or 赠品补发物流状态 = '暂无物流状态', IF(赠品补发订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),赠品补发物流状态)))) as 赠品补发最终状态, 统计月份
							FROM (	SELECT *, IF(赠品补发订单编号 <> "",IF(最新客服处理结果 LIKE '%补发海外仓%','统计','不统计'),'不统计') AS 是否统计	
									FROM 物流客诉件_创建时间 k1
									WHERE  k1.`统计月份` = '{0}' and DATE_FORMAT(k1.`记录时间`,'%Y-%m-%d') = '{1}' AND k1.最新客服处理人 in ({3}) AND k1.最新处理状态 <> ""
							) k	
							WHERE  是否统计 = "统计"
						)
						UNION
						(	SELECT '挽单列表' as 类型, 创建人, 订单编号, NULL AS 赠品补发订单编号, 当前订单状态, 当前物流状态, NULL AS 赠品补发订单状态,	NULL AS 赠品补发物流状态,
											IF(当前物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 当前物流状态, IF(当前物流状态 IN ('已签收','理赔'), IF(当前订单状态 = '已退货(销售)','拒收',当前物流状态),  IF(当前物流状态 = '发货中','在途',
												IF(当前物流状态 = '' or 当前物流状态 IS NULL or 当前物流状态 = '暂无物流状态', IF(当前订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),当前物流状态)))) as 最终状态, NULL AS 赠品补发最终状态, 统计月份
							FROM 挽单列表_创建时间 w
							WHERE  w.`统计月份` = '{0}' and DATE_FORMAT(w.`记录时间`, '%Y-%m-%d') = '{1}' AND w.创建人 in ({2}) AND w.删除人 = '' AND w.挽单类型 IN ("退换补挽单")
						)
				) k	
				GROUP BY  最新客服处理人
				ORDER BY FIELD(最新客服处理人,{2},'合计');'''.format(month_time, day_time, username_Jushou, username_caigou_yadan_wentijian)
        df32 = pd.read_sql_query(sql=sql32, con=self.engine1)
        listT.append(df32)

        print('拒收问题件-绩效 源数据 获取中（三.1）......')
        sql41 = '''SELECT *, IF(新单物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 新单物流状态,IF(新单物流状态 IN ('已签收','理赔'), IF(新单订单状态 = '已退货(销售)','拒收',新单物流状态), IF(新单物流状态 = '发货中','在途',
                            IF(新单物流状态 = '' or 新单物流状态 IS NULL or 新单物流状态 = '暂未物流状态', IF(新单订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),新单物流状态)))) as 最终状态
                FROM 拒收问题件_跟进时间 s1
                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}';'''.format(month_time, day_time)
        df41 = pd.read_sql_query(sql=sql41, con=self.engine1)
        listT.append(df41)
        print('拒收问题件-绩效 统计分析 获取中（三.2）......')
        sql42 = '''SELECT 新单克隆人, IF(已签收 = 0,NULL, 已签收) AS 已签收,  IF(拒收 = 0,NULL, 拒收) AS 拒收, IF(已退货 = 0,NULL, 已退货) AS 已退货,  IF(理赔 = 0,NULL, 理赔) AS 理赔, 
                            IF(未发货 = 0,NULL, 未发货) AS 未发货, IF(在途 = 0,NULL, 在途) AS 在途, IF(已完成 = 0,NULL, 已完成) AS 已完成,  (总计-无效订单) AS 有效单量, 总计, 
                            concat(ROUND(IFNULL(已签收 / 已完成,0) * 100,2),'%') AS 签收率,
                            concat(ROUND(IFNULL(已完成 / (总计-无效订单),0) * 100,2),'%') AS 完成占比,IF(无效订单 = 0,NULL, 无效订单) AS 无效订单,
                            concat(ROUND(IFNULL((总计-无效订单) / 总计,0) * 100,2),'%') AS 转换率
                FROM ( SELECT 新单克隆人, count(订单编号) as 总计,
                                SUM(IF(最终状态 = "已签收",1,0)) as 已签收, SUM(IF(最终状态 = "拒收",1,0)) as 拒收, SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔, SUM(IF(最终状态 = "未发货",1,0)) as 未发货, SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成, SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单
                        FROM (					
                                (
                                    SELECT '拒收件' AS 类型, 新单克隆人, 订单编号, 再次克隆下单 AS 克隆后新订单号, 新单订单状态, 新单物流状态, 
                                            IF(新单物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 新单物流状态,IF(新单物流状态 IN ('已签收','理赔'), IF(新单订单状态 = '已退货(销售)','拒收',新单物流状态), IF(新单物流状态 = '发货中','在途',
                                            IF(新单物流状态 = '' or 新单物流状态 IS NULL or 新单物流状态 = '暂未物流状态', IF(新单订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),新单物流状态)))) as 最终状态, 统计月份, 记录时间
                                    FROM (
                                            SELECT *
                                            FROM 拒收问题件_跟进时间 s1
                                            WHERE  s1.`统计月份` = '{0}' AND DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.新单克隆人 in ({2}) AND s1.再次克隆下单 <> "" 
                                    ) s2
                                )
                                UNION
                                (	SELECT '挽单列表' as 类型, 创建人, 订单编号, NULL 克隆后新订单号, 当前订单状态, 当前物流状态, 
                                            IF(当前物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 当前物流状态, IF(当前物流状态 IN ('已签收','理赔'), IF(当前订单状态 = '已退货(销售)','拒收',当前物流状态),  IF(当前物流状态 = '发货中','在途',
                                            IF(当前物流状态 = '' or 当前物流状态 IS NULL or 当前物流状态 = '暂无物流状态', IF(当前订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),当前物流状态)))) as 最终状态, 统计月份, 记录时间
                                    FROM 挽单列表_创建时间 w
                                    WHERE  w.`统计月份` = '{0}' and DATE_FORMAT(w.`记录时间`, '%Y-%m-%d') = '{1}' AND w.创建人 in ({2}) AND w.删除人 = '' AND w.挽单类型 = "拒收挽单"
                                )
                        ) s1
                        GROUP BY  新单克隆人
                ) s 
                ORDER BY FIELD(新单克隆人,{2},'合计');'''.format(month_time, day_time, username_Jushou)
        df42 = pd.read_sql_query(sql=sql42, con=self.engine1)
        listT.append(df42)

        print('采购异常-绩效 源数据 获取中（四.1）......')
        sql51 = '''SELECT *,IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                            IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态
                FROM 采购异常_创建时间 s1
                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}';'''.format(month_time, day_time)
        df51 = pd.read_sql_query(sql=sql51, con=self.engine1)
        listT.append(df51)
        print('采购异常-绩效 统计分析 获取中（四.2）......')
        sql52 = '''SELECT 类型, 客服处理人, IF(已签收 = 0,NULL, 已签收) AS 已签收,  IF(拒收 = 0,NULL, 拒收) AS 拒收, IF(已退货 = 0,NULL, 已退货) AS 已退货,  IF(理赔 = 0,NULL, 理赔) AS 理赔, 
                            IF(未发货 = 0,NULL, 未发货) AS 未发货, IF(在途 = 0,NULL, 在途) AS 在途, IF(已完成 = 0,NULL, 已完成) AS 已完成,  (总计-无效订单) AS 有效单量, 总计, 
                            concat(ROUND(IFNULL(已签收 / 已完成,0) * 100,2),'%') AS 签收率,
                            concat(ROUND(IFNULL(已完成 / (总计-无效订单),0) * 100,2),'%') AS 完成占比,IF(无效订单 = 0,NULL, 无效订单) AS 无效订单, 
                            concat(ROUND(IFNULL((总计-无效订单) / 总计,0) * 100,2),'%') AS 转换率
                FROM (		
                        SELECT '采购异常' AS 类型, 客服处理人,  count(订单编号) as 总计,
                                SUM(IF(最终状态 = "已签收",1,0)) as 已签收, SUM(IF(最终状态 = "拒收",1,0)) as 拒收, SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔, SUM(IF(最终状态 = "未发货",1,0)) as 未发货, SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成, SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单
                        FROM (
                                SELECT *,IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                                         IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态
                                FROM 采购异常_创建时间 s1
                                WHERE  s1.`统计月份` = '{0}' AND DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.客服处理人 in ({2}) AND s1.客服处理结果 NOT IN ("已发货","改派","无须处理") 
                        ) s2
                        GROUP BY  客服处理人	
                ) s
                ORDER BY FIELD(客服处理人,{2},'合计');;'''.format(month_time, day_time, username_caigou_yadan_wentijian)
        df52 = pd.read_sql_query(sql=sql52, con=self.engine1)
        listT.append(df52)

        print('压单核实-绩效 源数据 获取中（五.1）......')
        sql61 = '''SELECT *,IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                            IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态
                FROM 压单核实_创建时间 s1
                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}';'''.format(month_time, day_time)
        df61 = pd.read_sql_query(sql=sql61, con=self.engine1)
        listT.append(df61)
        print('压单核实-绩效 统计分析 获取中（五.2）......')
        sql62 = '''SELECT 类型, 最新客服处理人, IF(已签收 = 0,NULL, 已签收) AS 已签收,  IF(拒收 = 0,NULL, 拒收) AS 拒收, IF(已退货 = 0,NULL, 已退货) AS 已退货,  IF(理赔 = 0,NULL, 理赔) AS 理赔, 
                            IF(未发货 = 0,NULL, 未发货) AS 未发货, IF(在途 = 0,NULL, 在途) AS 在途, IF(已完成 = 0,NULL, 已完成) AS 已完成,  (总计-无效订单) AS 有效单量, 总计, 
                            concat(ROUND(IFNULL(已签收 / 已完成,0) * 100,2),'%') AS 签收率,
                            concat(ROUND(IFNULL(已完成 / (总计-无效订单),0) * 100,2),'%') AS 完成占比,IF(无效订单 = 0,NULL, 无效订单) AS 无效订单, 
                            concat(ROUND(IFNULL((总计-无效订单) / 总计,0) * 100,2),'%') AS 转换率
                FROM (
                        SELECT '压单核实' AS 类型, 最新客服处理人,  count(订单编号) as 总计,
                                SUM(IF(最终状态 = "已签收",1,0)) as 已签收, SUM(IF(最终状态 = "拒收",1,0)) as 拒收, SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔, SUM(IF(最终状态 = "未发货",1,0)) as 未发货, SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成, SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单
                        FROM (
                                SELECT *,IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                                            IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态,
                                        IF(最新处理结果 NOT LIKE '%取消%' AND (最新处理结果 LIKE '%无人接听%' OR 最新处理结果 LIKE '%无效号码%' OR 最新处理结果 LIKE '%电话暂停使用%' OR 最新处理结果 LIKE '%电话停止使用%'),'不统计','统计') AS 是否统计		
                                FROM 压单核实_创建时间 s1
                                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' and s1.最新客服处理人 in ({2}) AND s1.最新处理结果 <> ""
                        ) s2
                        WHERE  是否统计 = '统计'
                        GROUP BY  最新客服处理人	
                ) s
                ORDER BY FIELD(最新客服处理人,{2},'合计');'''.format(month_time, day_time, username_caigou_yadan_wentijian)
        df62 = pd.read_sql_query(sql=sql62, con=self.engine1)
        listT.append(df62)

        print('系统问题件-绩效 源数据 获取中（六.1）......')
        sql71 = '''SELECT *,IF(系统物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 系统物流状态, IF(系统物流状态 IN ('已签收','理赔'), IF(系统订单状态 = '已退货(销售)','拒收',系统物流状态), IF(系统物流状态 = '发货中','在途',
                            IF(系统物流状态 = '' or 系统物流状态 IS NULL or 系统物流状态 = '暂无物流状态', IF(系统订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),系统物流状态)))) as 最终状态
                FROM 系统问题件_下单时间 s1
                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}';'''.format(month_time, day_time)
        df71 = pd.read_sql_query(sql=sql71, con=self.engine1)
        listT.append(df71)
        print('系统问题件-绩效 统计分析 获取中（六.2）......')
        sql72 = '''SELECT 类型, 转化人, IF(已签收 = 0,NULL, 已签收) AS 已签收,  IF(拒收 = 0,NULL, 拒收) AS 拒收, IF(已退货 = 0,NULL, 已退货) AS 已退货,  IF(理赔 = 0,NULL, 理赔) AS 理赔, 
                            IF(未发货 = 0,NULL, 未发货) AS 未发货, IF(在途 = 0,NULL, 在途) AS 在途, IF(已完成 = 0,NULL, 已完成) AS 已完成,  (总计-无效订单) AS 有效单量, 总计, 
                            concat(ROUND(IFNULL(已签收 / 已完成,0) * 100,2),'%') AS 签收率,
                            concat(ROUND(IFNULL(已完成 / (总计-无效订单),0) * 100,2),'%') AS 完成占比,IF(无效订单 = 0,NULL, 无效订单) AS 无效订单, 
                            concat(ROUND(IFNULL((总计-无效订单) / 总计,0) * 100,2),'%') AS 转换率
                FROM (		
                        SELECT '系统问题件' AS 类型, 转化人,  count(订单编号) as 总计,
                                SUM(IF(最终状态 = "已签收",1,0)) as 已签收, SUM(IF(最终状态 = "拒收",1,0)) as 拒收, SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔, SUM(IF(最终状态 = "未发货",1,0)) as 未发货, SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成, SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单
                        FROM (
                                SELECT *,IF(系统物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 系统物流状态, IF(系统物流状态 IN ('已签收','理赔'), IF(系统订单状态 = '已退货(销售)','拒收',系统物流状态), IF(系统物流状态 = '发货中','在途',
                                        IF(系统物流状态 = '' or 系统物流状态 IS NULL or 系统物流状态 = '暂无物流状态', IF(系统订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),系统物流状态)))) as 最终状态
                                FROM 系统问题件_下单时间 s1
                                WHERE  s1.`统计月份` = '{0}' AND DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.转化人 in ({2})
                        ) s2
                        GROUP BY  转化人	
                ) s
                ORDER BY FIELD(转化人,{2},'合计');'''.format(month_time, day_time, username_caigou_yadan_wentijian)
        df72 = pd.read_sql_query(sql=sql72, con=self.engine1)
        listT.append(df72)

        print('物流问题-绩效 源数据 获取中（三.1）......')
        sql81 = '''SELECT *,IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                            IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态
                FROM 物流问题件_创建时间 s1
                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}';'''.format(month_time, day_time)
        df81 = pd.read_sql_query(sql=sql81, con=self.engine1)
        listT.append(df81)

        print('派送问题-绩效 源数据 获取中（三.3）......')
        sql82 = '''SELECT *,IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                            IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态
                FROM 派送问题件_处理时间 s1
                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}';'''.format(month_time, day_time)
        df82 = pd.read_sql_query(sql=sql82, con=self.engine1)
        listT.append(df82)

        print('物流问题 & 派送问题-绩效 统计分析 获取中（三.3）......')
        sql83 = '''SELECT 类型, 最新客服处理人, IF(已签收 = 0,NULL, 已签收) AS 已签收,  IF(拒收 = 0,NULL, 拒收) AS 拒收, IF(已退货 = 0,NULL, 已退货) AS 已退货,  IF(理赔 = 0,NULL, 理赔) AS 理赔, 
                        IF(未发货 = 0,NULL, 未发货) AS 未发货, IF(在途 = 0,NULL, 在途) AS 在途, IF(已完成 = 0,NULL, 已完成) AS 已完成,  (总计-无效订单) AS 有效单量, 总计, 
                        concat(ROUND(IFNULL(已签收 / 已完成,0) * 100,2),'%') AS 签收率,
                        concat(ROUND(IFNULL(已完成 / (总计-无效订单),0) * 100,2),'%') AS 完成占比,IF(无效订单 = 0,NULL, 无效订单) AS 无效订单, 
                        concat(ROUND(IFNULL((总计-无效订单) / 总计,0) * 100,2),'%') AS 转换率
                FROM (		
                        SELECT '物流问题件' AS 类型, 最新客服处理人,  count(订单编号) as 总计,
                                SUM(IF(最终状态 = "已签收",1,0)) as 已签收, SUM(IF(最终状态 = "拒收",1,0)) as 拒收, SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔, SUM(IF(最终状态 = "未发货",1,0)) as 未发货, SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成, SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单
                        FROM (    
                                (
                                    SELECT '物流派送问题件' AS 类型, 订单编号, 币种, 订单状态, 物流状态, 最新客服处理人, 最新处理结果, 最终状态,统计月份
                                    FROM (
                                            SELECT *,IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                                                    IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态,

                                                    IF(最新客服处理 LIKE '已处理%' OR 最新客服处理 LIKE '货件拒收%' OR 最新客服处理 LIKE '货态拒收%' OR 最新客服处理 LIKE '货态签收%' OR 最新客服处理 LIKE '货态已签收%' 
                                                    OR 最新客服处理 LIKE '已通知客户%' OR 最新客服处理 LIKE '已告知客户%' OR 最新客服处理 LIKE '通知客户取货%' OR 最新客服处理 LIKE '通知客户自取%' OR 最新客服处理 LIKE '请通知客户取货%'
                                                    OR 最新客服处理 LIKE '%暂停使用%' OR 最新客服处理 LIKE '%停止使用%' OR 最新客服处理 LIKE '%没有登记%' OR 最新客服处理 LIKE '%无登记%' OR 最新客服处理 LIKE '%电话停机%'
                                                    OR 最新客服处理 LIKE '已签收' OR 最新客服处理 LIKE '以帮客户下单'   OR 最新客服处理 LIKE '已发图片' 
                                                    OR 最新客服处理 LIKE '无人额急停%' OR 最新客服处理 LIKE '无人接听%' OR 最新客服处理 LIKE '无效号码%','不统计', 
                                                    IF(最新客服处理 NOT LIKE '%拒收%',
                                                    IF(最新客服处理 LIKE '无人额急停%' OR 最新客服处理 LIKE '无人接听%' OR 最新客服处理 LIKE '无效号码%' OR 最新客服处理 LIKE '%停机%' OR 最新客服处理 LIKE '%暂停使用%' OR 最新客服处理 LIKE '电话无登记%' ,'不统计',
                                                    IF(最新处理结果 = '已签收','不统计','统计')),'统计')) AS 是否统计
                                            FROM 物流问题件_创建时间 s1
                                            WHERE  s1.`统计月份` = '{0}' AND DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.最新客服处理人 in ({2})
                                    ) s2
                                    WHERE 是否统计 = '统计'
                                )
                                UNION
                                (
                                    SELECT '物流派送问题件' AS 类型, 订单编号, 币种, 订单状态, 物流状态, 最新处理人, 最新处理结果, 最终状态,统计月份
                                    FROM (
                                            SELECT *,IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                                                    IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态,

                                                    IF(最新处理结果 LIKE '已处理%' OR 最新处理结果 LIKE '货件拒收%' OR 最新处理结果 LIKE '货态拒收%' OR 最新处理结果 LIKE '货态签收%' OR 最新处理结果 LIKE '货态已签收%' 
                                                    OR 最新处理结果 LIKE '已通知客户%' OR 最新处理结果 LIKE '已告知客户%' OR 最新处理结果 LIKE '通知客户取货%' OR 最新处理结果 LIKE '通知客户自取%' OR 最新处理结果 LIKE '请通知客户取货%'
                                                    OR 最新处理结果 LIKE '已签收' OR 最新处理结果 LIKE '以帮客户下单'   OR 最新处理结果 LIKE '已发图片' 
                                                    OR 最新处理结果 LIKE '无人额急停%' OR 最新处理结果 LIKE '无人接听%' OR 最新处理结果 LIKE '无效号码%','不统计', 
                                                    IF(最新处理结果 NOT LIKE '%拒收%',
                                                    IF(最新处理结果 LIKE '无人额急停%' OR 最新处理结果 LIKE '无人接听%' OR 最新处理结果 LIKE '无效号码%' OR 最新处理结果 LIKE '%停机%' OR 最新处理结果 LIKE '%暂停使用%' 
                                                    OR 最新处理结果 LIKE '电话无登记%' OR 最新处理结果 LIKE '%停止使用%' OR 最新处理结果 LIKE '%没有登记%' OR 最新处理结果 LIKE '%无登记%' OR 最新处理结果 LIKE '%电话停机%','不统计',
                                                    IF(最新处理结果 = '已签收','不统计','统计')),'统计')) AS 是否统计
                                            FROM 派送问题件_处理时间 s1
                                            WHERE  s1.`统计月份` = '{0}' AND DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.最新处理人 in ({2})
                                    ) s2
                                    WHERE 是否统计 = '统计'
                                )
                        ) ss1
                        GROUP BY  最新客服处理人	
                ) s
                ORDER BY FIELD(最新客服处理人,{2},'合计');'''.format(month_time, day_time, username_caigou_yadan_wentijian)
        df83 = pd.read_sql_query(sql=sql83, con=self.engine1)
        listT.append(df83)

        file_path = r'''F:\\输出文件\\{0}绩效数据明细 {1}.xlsx'''.format(rq_month, rq)
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df11.to_excel(excel_writer=writer, sheet_name='挽单', index=False)
            df21.to_excel(excel_writer=writer, sheet_name='促单', index=False)
            df22.to_excel(excel_writer=writer, sheet_name='促单分析', index=False)
            df31.to_excel(excel_writer=writer, sheet_name='物流客诉', index=False)
            df32.to_excel(excel_writer=writer, sheet_name='物流客诉分析', index=False)
            df41.to_excel(excel_writer=writer, sheet_name='拒收问题件', index=False)
            df42.to_excel(excel_writer=writer, sheet_name='拒收问题件分析', index=False)
            df51.to_excel(excel_writer=writer, sheet_name='采购异常', index=False)
            df52.to_excel(excel_writer=writer, sheet_name='采购异常分析', index=False)
            df61.to_excel(excel_writer=writer, sheet_name='压单核实', index=False)
            df62.to_excel(excel_writer=writer, sheet_name='压单核实分析', index=False)
            df71.to_excel(excel_writer=writer, sheet_name='系统问题件', index=False)
            df72.to_excel(excel_writer=writer, sheet_name='系统问题件分析', index=False)
            df81.to_excel(excel_writer=writer, sheet_name='物流问题', index=False)
            df82.to_excel(excel_writer=writer, sheet_name='派送问题', index=False)
            df83.to_excel(excel_writer=writer, sheet_name='物流问题&派送问题分析', index=False)
    '''
        # 先更新 获取上月的订单，再去更新之前未完结的订单状态，然后再去更新 需要统计的时间
    '''
    # 先更新 获取上月的订单，再去更新之前未完结的订单状态，然后再去更新 需要统计的时间
    def read_write_workbook(self):    # 读取工作表的数据 写入数据库(一)
        start = datetime.datetime.now()
        path = r'F:\神龙签收率\(A) 各种报表汇总\更新客诉退货回款'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                fileType = os.path.splitext(filePath)[1]
                app = xlwings.App(visible=False, add_book=False)
                app.display_alerts = False
                if 'xls' in fileType:
                    wb = app.books.open(filePath, update_links=False, read_only=True)
                    for sht in wb.sheets:
                        if sht.api.Visible == -1:
                            db = None
                            try:
                                db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                                print(db.columns)
                                print(db)
                            except Exception as e:
                                print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                            if db is not None and len(db) > 0:
                                print('++++正在获取：' + sht.name + ' 表；共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                                db = db[['订单编号', '回款状态', '更新月份']]
                                db.to_sql('cache_ch_cp', con=self.engine1, index=False, if_exists='replace')
                                print('数据更新中')
                                sql = '''update {0} a, cache_ch_cp b
                                            set a.`回款状态`= IF(b.`回款状态` = '', NULL, b.`回款状态`),
                                                a.`更新月份`= b.更新月份
                                         where a.`订单编号`=b.`订单编号`;'''.format('物流客诉_挽单列表_退货_计算统计')
                                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                                print('更新成功')
                            else:
                                print('----------数据为空,获取失败：' + sht.name)
                        else:
                            print('----不需获取：' + sht.name)
                    wb.close()
                app.quit()

        print('处理耗时：', datetime.datetime.now() - start)
    def userid_performance_old_upadata(self, update_old, update_new):
        print('正在获取上月未完结订单的数据......')
        sql11 = '''SELECT *
                    FROM (
                            (
                                SELECT 订单编号
                                FROM 促单_挽单列表_下单时间_计算统计 s1
                                WHERE  s1.`是否计算` = '否'  AND s1.更新月份 = '{0}'
                            )
                            UNION
                            (
                                SELECT 订单编号
                                FROM 采购异常_问题订单_压单_计算统计 s1
                                WHERE  s1.`是否计算` = '否'  AND s1.更新月份 = '{0}'
                            )
                            UNION
                            (
                                SELECT 克隆后新订单号
                                FROM 拒收挽单_挽单列表_计算统计 s1
                                WHERE  s1.`是否计算` = '否'  AND s1.更新月份 = '{0}'
                            )
                            UNION
                            (
                                SELECT 订单编号
                                FROM 物流_派送_问题件_计算统计 s1
                                WHERE  s1.`是否计算` = '否'  AND s1.更新月份 = '{0}'
                            )
                            UNION
                            (
                                SELECT 订单编号
                                FROM 物流客诉_挽单列表_退货_计算统计 s1
                                WHERE  s1.`是否计算` = '否'  AND s1.更新月份 = '{0}'
                            )
                            UNION
                            (
                                SELECT 赠品补发订单编号
                                FROM 物流客诉_挽单列表_退货_计算统计 s1
                                WHERE  s1.`是否计算` = '否'  AND s1.更新月份 = '{0}'
                            )
                    ) s;'''.format(update_old, 'day_time')
        dp = pd.read_sql_query(sql=sql11, con=self.engine1)

        print('++++++订单状态明细查询中+++++++')
        order_list = list(dp['订单编号'])
        # print(order_list)
        max_count = len(order_list)
        pople_Query = '物流客诉查询'
        df2 = pd.DataFrame([])
        dtlist = []
        n = 0
        while n < max_count + 500:
            ord = ','.join(order_list[n:n + 500])
            # print(ord)
            data = self._service_id_order("", "", 0, proxy_handle, proxy_id, pople_Query, ord)  # 查询全部订单信息
            dtlist.append(data)
            n = n + 500
        dp2 = df2.append(dtlist, ignore_index=True)
        print(99)
        print(dp2)
        dp3 = dp2[['orderNumber', 'addTime', 'orderStatus', 'logisticsStatus']]
        dp3.columns = ['订单编号', '下单时间', '订单状态', '物流状态']
        dp3.to_sql('cache_ch_cp', con=self.engine1, index=False, if_exists='replace')
        sql = '''SELECT *, '{0}' as 更新月份, 
                        IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), 
                        IF(物流状态 = '发货中','在途',IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态
                FROM cache_ch_cp;'''.format(update_new, 'day_time')
        dp3 = pd.read_sql_query(sql=sql, con=self.engine1)
        dp3.to_sql('customer', con=self.engine1, index=False, if_exists='replace')

        print('正在更新 各表的数据......')
        sql = '''update {0} a, customer b
                    set a.`订单状态`= IF(b.`订单状态` = '', NULL, b.`订单状态`),
                        a.`物流状态`= IF(b.`物流状态` = '', NULL, b.`物流状态`),
                        a.`最终状态`= IF(b.`最终状态` = '', NULL, b.`最终状态`),
                        a.`是否计算`= IF(b.最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否'),
                        a.`更新月份`= b.更新月份
                 where a.`订单编号`=b.`订单编号` AND a.是否计算 = '否'; '''.format('促单_挽单列表_下单时间_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('促单_挽单列表_下单时间_计算统计    更新成功......')

        sql = '''update {0} a, customer b
                    set a.`订单状态`= IF(b.`订单状态` = '', NULL, b.`订单状态`),
                        a.`物流状态`= IF(b.`物流状态` = '', NULL, b.`物流状态`),
                        a.`最终状态`= IF(b.`最终状态` = '', NULL, b.`最终状态`),
                        a.`是否计算`= IF(b.最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否'),
                        a.`更新月份`= b.更新月份
                 where a.`订单编号`=b.`订单编号` AND a.是否计算 = '否';'''.format('采购异常_问题订单_压单_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('采购异常_问题订单_压单_计算统计    更新成功......')

        sql = '''update {0} a, customer b
                    set a.`新单订单状态`= IF(b.`订单状态` = '', NULL, b.`订单状态`),
                        a.`新单物流状态`= IF(b.`物流状态` = '', NULL, b.`物流状态`),
                        a.`最终状态`= IF(b.`最终状态` = '', NULL, b.`最终状态`),
                        a.`是否计算`= IF(b.最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否'),
                        a.`更新月份`= b.更新月份
                 where a.`克隆后新订单号`=b.`订单编号` AND a.是否计算 = '否';'''.format('拒收挽单_挽单列表_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('拒收挽单_挽单列表_计算统计    更新成功......')

        sql = '''update {0} a, customer b
                    set a.`订单状态`= IF(b.`订单状态` = '', NULL, b.`订单状态`),
                        a.`物流状态`= IF(b.`物流状态` = '', NULL, b.`物流状态`),
                        a.`最终状态`= IF(b.`最终状态` = '', NULL, b.`最终状态`),
                        a.`是否计算`= IF(b.最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否'),
                        a.`更新月份`= b.更新月份
                 where a.`订单编号`=b.`订单编号` AND a.是否计算 = '否';'''.format('物流_派送_问题件_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('物流_派送_问题件_计算统计    更新成功......')

        sql = '''update {0} a, customer b
                    set a.`订单状态`= IF(b.`订单状态` = '', NULL, b.`订单状态`),
                        a.`物流状态`= IF(b.`物流状态` = '', NULL, b.`物流状态`),
                        a.`最终状态`= IF(b.`最终状态` = '', NULL, b.`最终状态`),
                        a.`是否计算`= IF(b.最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否')
                 where a.`订单编号`=b.`订单编号` AND a.是否计算 = '否';'''.format('物流客诉_挽单列表_退货_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('物流客诉_挽单列表_退货_计算统计    更新成功......')

        sql = '''update {0} a, customer b
                    set a.`赠品补发订单状态`= IF(b.`订单状态` = '', NULL, b.`订单状态`),
                        a.`赠品补发物流状态`= IF(b.`物流状态` = '', NULL, b.`物流状态`),
                        a.`赠品补发最终状态`= IF(b.`最终状态` = '', NULL, b.`最终状态`),
                        a.`是否计算`= IF(b.最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否')
                 where a.`赠品补发订单编号`=b.`订单编号` AND a.是否计算 = '否';'''.format('物流客诉_挽单列表_退货_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('物流客诉_挽单列表_退货_计算统计 赠品订单信息    更新成功......')
    def userid_performance_New(self, username_Cudan, username_Jushou, username_caigou_yadan_wentijian, month_time, day_time):
        listT = []
        print('挽单  列表-绩效 数据整理 写入各 计算统计表 中（零）......')
        sql11 = '''SELECT *, IF(当前物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 当前物流状态, IF(当前物流状态 IN ('已签收','理赔'), IF(当前订单状态 = '已退货(销售)','拒收',当前物流状态), 
                            IF(当前物流状态 = '发货中','在途',IF(当前物流状态 = '' or 当前物流状态 IS NULL or 当前物流状态 = '暂无物流状态', IF(当前订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),当前物流状态)))) as 最终状态
                FROM 挽单列表_创建时间 s1
                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' and s1.创建人 in ({2});'''.format(month_time, day_time, username_Cudan)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)

        db12 = df11[(df11['挽单类型'].str.contains('取消挽单|未支付/支付失败挽单'))]  # 归为促单
        db12.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO {0}(类型, 代下单客服, 订单编号, 订单状态,物流状态, 最终状态, 统计月份, 是否计算, 计算月份, 更新月份, 记录时间, 更新时间) 
                SELECT 挽单类型 as 类型,  创建人 as 代下单客服, 订单编号,  当前订单状态 as 订单状态, 当前物流状态 as 物流状态, 最终状态, 统计月份,
                        IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 
                        IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份, 
                        DATE_FORMAT(curdate(),'%Y%m') as 更新月份, 记录时间, NOW() 更新时间
                FROM cache_ch 
                WHERE 删除人 = '' and 创建人 in ({1});'''.format('促单_挽单列表_下单时间_计算统计', username_Cudan)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        db13 = df11[(df11['挽单类型'].str.contains('退换补挽单'))]  # 归为退货挽单
        db13.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO {0}(类型, 客服处理人, 订单编号, 赠品补发订单编号,订单状态,物流状态, 赠品补发订单状态, 赠品补发物流状态,最终状态, 赠品补发最终状态,统计月份, 是否计算, 计算月份, 更新月份, 回款状态,记录时间, 更新时间) 
                 SELECT 挽单类型 as 类型, 创建人 as 客服处理人, 订单编号, NULL as 赠品补发订单编号, 当前订单状态 as 订单状态, 当前物流状态 as 物流状态, NULL as 赠品补发订单状态, NULL as 赠品补发物流状态,最终状态, NULL as 赠品补发最终状态, 统计月份,
                        IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 
                        IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份,  DATE_FORMAT(curdate(),'%Y%m') as 更新月份, NULL as 回款状态, 记录时间, NOW() 更新时间
                FROM cache_ch 
                WHERE 删除人 = '' ;'''.format('物流客诉_挽单列表_退货_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        db14 = df11[(df11['挽单类型'].str.contains('拒收挽单'))]  # 归为拒收挽单
        db14.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO {0}(类型, 新单克隆人, 订单编号, 克隆后新订单号,新单订单状态,新单物流状态, 最终状态, 统计月份, 是否计算, 计算月份, 更新月份, 记录时间, 更新时间) 
                 SELECT 挽单类型 as 类型, 创建人 as 新单克隆人, 订单编号, null as 克隆后新订单号, 当前订单状态 as 新单订单状态, 当前物流状态 as 新单物流状态,  最终状态, 统计月份,
                        IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 
                        IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份,  
                        DATE_FORMAT(curdate(),'%Y%m') as 更新月份, 记录时间, NOW() 更新时间
                FROM cache_ch 
                WHERE 删除人 = '' ;'''.format('拒收挽单_挽单列表_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('促      单-绩效 数据整理 写入计算统计表 中 （一）......')  # 不同类型计算两次
        sql = '''SELECT '促单' as 类型, 代下单客服, 订单编号, 订单状态,物流状态, 
                        IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                        IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态, 统计月份, 记录时间
                FROM 促单_下单时间 s1
                WHERE  s1.代下单客服 in ({0}) AND (s1.克隆人 = '' or s1.克隆类型 = "扣货克隆") AND s1.`统计月份` = '{1}' AND DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{2}';'''.format(username_Cudan, month_time, day_time)
        df2 = pd.read_sql_query(sql=sql, con=self.engine1)
        df2.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO {0}(类型, 代下单客服, 订单编号, 订单状态,物流状态, 最终状态, 统计月份, 是否计算, 计算月份, 更新月份, 记录时间, 更新时间) 
                           SELECT 类型, 代下单客服, 订单编号, 订单状态,物流状态, 最终状态, 统计月份,
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份, 
                                DATE_FORMAT(curdate(),'%Y%m') as 更新月份, 记录时间, NOW() 更新时间
                FROM cache_ch;'''.format('促单_挽单列表_下单时间_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('物流客诉件-绩效 数据整理 写入计算统计表 中 （二.一）......')  # 不同类型计算一次  --  和挽单，以最后克隆人为准，若有两个克隆人则导出看原因
        sql = '''SELECT '客诉件' AS 类型, 最新客服处理人 AS 客服处理人, 订单编号, 赠品补发订单编号,订单状态, 物流状态, 赠品补发订单状态, 赠品补发物流状态,
                        IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                        IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂未物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态, 
                        IF(赠品补发物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 赠品补发物流状态, IF(赠品补发物流状态 IN ('已签收','理赔'), IF(赠品补发订单状态 = '已退货(销售)','拒收',赠品补发物流状态), IF(赠品补发物流状态 = '发货中','在途', 
                        IF(赠品补发物流状态 = '' or 赠品补发物流状态 IS NULL or 赠品补发物流状态 = '暂无物流状态', IF(赠品补发订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),赠品补发物流状态)))) as 赠品补发最终状态,统计月份, 记录时间
                FROM (
                        SELECT *, IF(赠品补发订单编号 <> "",IF(最新处理结果 LIKE '%补发海外仓%','统计','不统计'),'不统计') AS 是否统计	
                        FROM 物流客诉件_创建时间 s1
                        WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.最新处理状态 <> ""
                ) s
                WHERE s.`是否统计` = '统计' and s.最新客服处理人 in ({2});'''.format(month_time, day_time, username_caigou_yadan_wentijian)
        df3 = pd.read_sql_query(sql=sql, con=self.engine1)
        df3.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO {0}(类型, 客服处理人, 订单编号, 赠品补发订单编号, 订单状态, 物流状态, 赠品补发订单状态, 赠品补发物流状态,最终状态, 赠品补发最终状态, 是否计算, 统计月份, 计算月份, 更新月份, 回款状态,记录时间, 更新时间) 
                           SELECT 类型, 客服处理人, 订单编号, 赠品补发订单编号, 订单状态, 物流状态, 赠品补发订单状态, 赠品补发物流状态,最终状态, 赠品补发最终状态,
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 统计月份,
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份, DATE_FORMAT(curdate(),'%Y%m') as 更新月份, NULL as 回款状态, 记录时间, NOW() 更新时间
                           FROM cache_ch;'''.format('物流客诉_挽单列表_退货_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('拒收问题件-绩效 数据整理 写入计算统计表 中 （三.一）......')  # 不同类型计算一次  --  和挽单，以最后克隆人为准，若有两个克隆人则导出看原因
        sql = '''SELECT '拒收件' AS 类型, 新单克隆人, 订单编号, 再次克隆下单 AS 克隆后新订单号, 新单订单状态, 新单物流状态, 
                        IF(新单物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 新单物流状态, IF(新单物流状态 IN ('已签收','理赔'), IF(新单订单状态 = '已退货(销售)','拒收',新单物流状态), IF(新单物流状态 = '发货中','在途',
                        IF(新单物流状态 = '' or 新单物流状态 IS NULL or 新单物流状态 = '暂未物流状态', IF(新单订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),新单物流状态)))) as 最终状态, 统计月份, 记录时间
                FROM (
                        SELECT *
                        FROM 拒收问题件_跟进时间 s1
                        WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.再次克隆下单 <> "" and s1.新单克隆人 in ({2})
                ) s;'''.format(month_time, day_time, username_Jushou)
        df4 = pd.read_sql_query(sql=sql, con=self.engine1)
        df4.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO {0}(类型, 新单克隆人, 订单编号, 克隆后新订单号, 新单订单状态,新单物流状态, 最终状态, 是否计算, 统计月份, 计算月份, 更新月份, 记录时间, 更新时间) 
                           SELECT 类型, 新单克隆人, 订单编号, 克隆后新订单号, 新单订单状态,新单物流状态, 最终状态, 
                                 IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 统计月份,
                                 IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份,  
                                 DATE_FORMAT(curdate(),'%Y%m') as 更新月份, 记录时间, NOW() 更新时间
                           FROM cache_ch;'''.format('拒收挽单_挽单列表_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)


        print('采购  异常-绩效 数据整理 写入计算统计表 中 （二.一）......')
        sql = '''SELECT *
                FROM ( SELECT '采购异常' AS 类型, 客服处理人, 订单编号, 订单状态, 物流状态, 
                            IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
                            IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态, 
                            IF(客服处理结果 LIKE '%已发货%' OR 客服处理结果 LIKE '%改派%' OR 客服处理结果 LIKE '%无须处理%','不统计','统计') AS 是否统计, 统计月份, 记录时间
                    FROM 采购异常_创建时间 s1
                    WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' and s1.客服处理人 in ({2})
                ) s	
                WHERE 是否统计 = '统计';'''.format(month_time, day_time, username_caigou_yadan_wentijian)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO {0}(类型, 客服处理人, 订单编号, 订单状态,物流状态, 最终状态, 是否计算, 统计月份, 计算月份, 更新月份, 记录时间, 更新时间) 
                           SELECT 类型, 客服处理人, 订单编号, 订单状态,物流状态, 最终状态, 
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 统计月份,
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份,  
                                DATE_FORMAT(curdate(),'%Y%m') as 更新月份, 记录时间, NOW() 更新时间
                            FROM cache_ch;'''.format('采购异常_问题订单_压单_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('压单  核实-绩效 数据整理 写入计算统计表 中 （二.二）......')
        sql = '''SELECT *
                FROM ( SELECT '压单' AS 类型, 最新客服处理人 AS 客服处理人, 订单编号, 订单状态, 物流状态, 
                            IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态,
                            IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
			                IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态, 
			                IF(最新处理结果 NOT LIKE '%取消%' AND (最新处理结果 LIKE '%无人接听%' OR 最新处理结果 LIKE '%无效号码%' OR 最新处理结果 LIKE '%电话暂停使用%' OR 最新处理结果 LIKE '%电话停止使用%'),'不统计','统计') AS 是否统计, 统计月份, 记录时间
                    FROM 压单核实_创建时间 s1
                    WHERE s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.最新处理结果 <> "" and s1.最新客服处理人 in ({2}) 
                ) s	
                WHERE 是否统计 = '统计';'''.format(month_time, day_time, username_caigou_yadan_wentijian)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO {0}(类型, 客服处理人, 订单编号, 订单状态,物流状态, 最终状态, 是否计算, 统计月份, 计算月份, 更新月份, 记录时间, 更新时间) 
                           SELECT 类型, 客服处理人, 订单编号, 订单状态,物流状态, 最终状态,
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 统计月份,
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份,  
                                DATE_FORMAT(curdate(),'%Y%m') as 更新月份, 记录时间, NOW() 更新时间
                            FROM cache_ch;;'''.format('采购异常_问题订单_压单_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('系统问题件-绩效 数据整理 写入计算统计表 中 （二.三）......')
        sql = '''SELECT '问题订单' AS 类型, 转化人 AS 客服处理人, 订单编号, 系统订单状态 AS 订单状态, 系统物流状态 AS 物流状态, 
                        IF(系统物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 系统物流状态, IF(系统物流状态 IN ('已签收','理赔'), IF(系统订单状态 = '已退货(销售)','拒收',系统物流状态), IF(系统物流状态 = '发货中','在途',
			            IF(系统物流状态 = '' or 系统物流状态 IS NULL or 系统物流状态 = '暂无物流状态', IF(系统订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),系统物流状态)))) as 最终状态, 统计月份, 记录时间
                FROM 系统问题件_下单时间 s1
                WHERE  s1.`统计月份` = '{0}' and DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.转化人 IN ({2});'''.format(month_time, day_time, username_caigou_yadan_wentijian)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO {0}(类型, 客服处理人, 订单编号, 订单状态,物流状态, 最终状态, 是否计算, 统计月份, 计算月份, 更新月份, 记录时间, 更新时间) 
                           SELECT 类型, 客服处理人, 订单编号, 订单状态,物流状态, 最终状态, 
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 统计月份,
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份, DATE_FORMAT(curdate(),'%Y%m') as 更新月份, 记录时间, NOW() 更新时间
                FROM cache_ch;'''.format('采购异常_问题订单_压单_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('物流问题件-绩效 数据整理 写入计算统计表 中 （三.一）......')
        sql = '''SELECT '物流问题件' AS 类型, 最新客服处理人 AS 客服处理人, 订单编号, 订单状态, 物流状态, 
                        IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
			            IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态, 统计月份, 记录时间
                FROM ( SELECT *, IF(最新客服处理 LIKE '已处理%' OR 最新客服处理 LIKE '货件拒收%' OR 最新客服处理 LIKE '货态拒收%' OR 最新客服处理 LIKE '货态签收%' OR 最新客服处理 LIKE '货态已签收%' 
                                OR 最新客服处理 LIKE '已通知客户%'  OR 最新客服处理 LIKE '已告知客户%' OR 最新客服处理 LIKE '通知客户取货%' OR 最新客服处理 LIKE '通知客户自取%' OR 最新客服处理 LIKE '请通知客户取货%' 
                                OR 最新客服处理 LIKE '%暂停使用%' OR 最新客服处理 LIKE '%停止使用%'  OR 最新客服处理 LIKE '%没有登记%' OR 最新客服处理 LIKE '%无登记%' OR 最新客服处理 LIKE '%电话停机%' OR 最新客服处理 LIKE '已签收' 
                                OR 最新客服处理 LIKE '以帮客户下单' OR 最新客服处理 LIKE '已发图片' OR 最新客服处理 LIKE '无人额急停%' OR 最新客服处理 LIKE '无人接听%' OR 最新客服处理 LIKE '无效号码%','不统计', 
                                IF(最新客服处理 NOT LIKE '%拒收%',
                                IF(最新客服处理 LIKE '无人额急停%' OR 最新客服处理 LIKE '无人接听%' OR 最新客服处理 LIKE '无效号码%' OR 最新客服处理 LIKE '%停机%' OR 最新客服处理 LIKE '%暂停使用%' OR 最新客服处理 LIKE '电话无登记%' ,'不统计',
                                IF(最新处理结果 = '已签收','不统计','统计')),'统计')) AS 是否统计
                        FROM 物流问题件_创建时间 s1
                        WHERE  s1.`统计月份` = '{0}' AND DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.最新处理状态 <> "" AND s1.最新客服处理人 in ({2})
                ) s
                WHERE s.`是否统计` = '统计';'''.format(month_time, day_time, username_caigou_yadan_wentijian)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('cache_ch', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO {0}(类型, 客服处理人, 订单编号, 订单状态,物流状态, 最终状态, 是否计算, 统计月份, 计算月份, 更新月份, 记录时间, 更新时间) 
                           SELECT 类型, 客服处理人, 订单编号, 订单状态,物流状态, 最终状态, 
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 统计月份,
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份, 
                                DATE_FORMAT(curdate(),'%Y%m') as 更新月份, 记录时间, NOW() 更新时间
                           FROM cache_ch;'''.format('物流_派送_问题件_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('派送问题件-绩效 数据整理 写入计算统计表 中 （三.二）......')
        sql = '''SELECT '派送问题件' AS 类型, 最新处理人 AS 客服处理人, 订单编号, 订单状态, 物流状态, 
                        IF(物流状态 IN ('已退货','拒收', '自发头程丢件', '客户取消'), 物流状态, IF(物流状态 IN ('已签收','理赔'), IF(订单状态 = '已退货(销售)','拒收',物流状态), IF(物流状态 = '发货中','在途',
			            IF(物流状态 = '' or 物流状态 IS NULL or 物流状态 = '暂无物流状态', IF(订单状态 IN ('已删除','未支付','支付失败'),'无效订单','未发货'),物流状态)))) as 最终状态, 统计月份, 记录时间
                FROM (
                        SELECT *, IF(最新处理结果 LIKE '已处理%' OR 最新处理结果 LIKE '货件拒收%' OR 最新处理结果 LIKE '货态拒收%' OR 最新处理结果 LIKE '货态签收%' OR 最新处理结果 LIKE '货态已签收%' 
                                    OR 最新处理结果 LIKE '已通知客户%' OR 最新处理结果 LIKE '已告知客户%' OR 最新处理结果 LIKE '通知客户取货%' OR 最新处理结果 LIKE '通知客户自取%' OR 最新处理结果 LIKE '请通知客户取货%'
                                    OR 最新处理结果 LIKE '已签收' OR 最新处理结果 LIKE '以帮客户下单' OR 最新处理结果 LIKE '已发图片' OR 最新处理结果 LIKE '无人额急停%' OR 最新处理结果 LIKE '无人接听%' OR 最新处理结果 LIKE '无效号码%','不统计', 
                                  IF(最新处理结果 NOT LIKE '%拒收%',
                                     IF(最新处理结果 LIKE '无人额急停%' OR 最新处理结果 LIKE '无人接听%' OR 最新处理结果 LIKE '无效号码%' OR 最新处理结果 LIKE '%停机%' OR 最新处理结果 LIKE '%暂停使用%' 
										OR 最新处理结果 LIKE '电话无登记%' OR 最新处理结果 LIKE '%停止使用%' OR 最新处理结果 LIKE '%没有登记%' OR 最新处理结果 LIKE '%无登记%' OR 最新处理结果 LIKE '%电话停机%','不统计',
                                     IF(最新处理结果 = '已签收','不统计','统计')),'统计')) AS 是否统计
                        FROM 派送问题件_处理时间 s1
                        WHERE  s1.`统计月份` = '{0}' AND DATE_FORMAT(s1.`记录时间`,'%Y-%m-%d') = '{1}' AND s1.最新处理状态 <> "" AND s1.最新处理人 in ({2})
                ) s
                WHERE s.`是否统计` = '统计';'''.format(month_time, day_time, username_caigou_yadan_wentijian)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('cache_ch_cp', con=self.engine1, index=False, if_exists='replace')

        print('派送问题件-绩效 数据整理 写入计算统计表 中 （三.三去掉物流与派送重复的，以物流为准）......')
        sql = '''SELECT *  FROM cache_ch_cp p WHERE p.订单编号 NOT IN (SELECT 订单编号 FROM cache_ch);'''
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df.to_sql('cache_check_cp', con=self.engine1, index=False, if_exists='replace')

        sql = '''REPLACE INTO {0}(类型, 客服处理人, 订单编号, 订单状态,物流状态, 最终状态, 是否计算, 统计月份, 计算月份, 更新月份, 记录时间, 更新时间) 
                           SELECT 类型, 客服处理人, 订单编号, 订单状态,物流状态, 最终状态, 
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),'是','否') as 是否计算, 统计月份,
                                IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"), 统计月份, '-') as 计算月份, 
                                DATE_FORMAT(curdate(),'%Y%m') as 更新月份, 记录时间, NOW() 更新时间
                           FROM cache_check_cp;'''.format('物流_派送_问题件_计算统计')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('各绩效表 更新完成......')
    def userid_performance_New_export(self, username_Cudan, username_Jushou, username_caigou_yadan_wentijian, month_time):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        listT = []
        print('促单-绩效 源数据 获取中（一.1）......')
        sql11 = '''SELECT * FROM 促单_挽单列表_下单时间_计算统计 s1  WHERE  s1.`更新月份` = '{0}';'''.format(month_time)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)
        listT.append(df11)
        print('促单-绩效 统计分析 获取中（一.2）......')
        sql12 = '''SELECT 代下单客服, IF(已签收 = 0,NULL, 已签收) AS 已签收,  IF(拒收 = 0,NULL, 拒收) AS 拒收, IF(已退货 = 0,NULL, 已退货) AS 已退货,  IF(理赔 = 0,NULL, 理赔) AS 理赔, 
                        IF(未发货 = 0,NULL, 未发货) AS 未发货, IF(在途 = 0,NULL, 在途) AS 在途, IF(已完成 = 0,NULL, 已完成) AS 已完成,  (总计-无效订单) AS 有效单量, 总计, 
                        concat(ROUND(IFNULL(已签收 / 已完成,0) * 100,2),'%') AS 签收率,
                        concat(ROUND(IFNULL(已完成 / (总计-无效订单),0) * 100,2),'%') AS 完成占比,IF(无效订单 = 0,NULL, 无效订单) AS 无效订单, 
                        concat(ROUND(IFNULL((总计-无效订单) / 总计,0) * 100,2),'%') AS 转换率
                FROM ( SELECT 代下单客服, count(订单编号) as 总计,
                                SUM(IF(最终状态 = "已签收",1,0)) as 已签收, SUM(IF(最终状态 = "拒收",1,0)) as 拒收, SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔, SUM(IF(最终状态 = "未发货",1,0)) as 未发货, SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成, SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单
                        FROM (
                                SELECT *
                                FROM 促单_挽单列表_下单时间_计算统计 s1
                                WHERE  s1.`更新月份` = '{0}'
                        ) s1
                        GROUP BY  代下单客服
                ) s ORDER BY FIELD(代下单客服,{1},'合计');'''.format(month_time, username_Cudan)
        df12 = pd.read_sql_query(sql=sql12, con=self.engine1)
        listT.append(df12)

        print('物流客诉-绩效 源数据 获取中（二.1）......')
        sql31 = '''SELECT * FROM 物流客诉_挽单列表_退货_计算统计 s1  WHERE  s1.`更新月份` = '{0}';'''.format(month_time)
        df31 = pd.read_sql_query(sql=sql31, con=self.engine1)
        listT.append(df31)
        print('物流客诉-绩效 统计分析 获取中（二.2）......')
        sql32 = '''SELECT 客服处理人,
                        SUM(IF(回款状态 = "已回款",1,0)) as 已回款,
                        SUM(IF(回款状态 = "已退款",1,0)) as 已退款,
                        SUM(IF(回款状态 = "售后订单",1,0)) as 售后订单,
                        SUM(IF(回款状态 = "未回款",1,0)) as 未回款,
                        count(订单编号) as 总计
                FROM (
                        SELECT * 
                        FROM 物流客诉_挽单列表_退货_计算统计 s1  
                        WHERE  s1.`更新月份` = '{0}'
                ) k	
                GROUP BY  客服处理人
                ORDER BY FIELD(客服处理人,{1},'合计');'''.format(month_time, username_Jushou)
        df32 = pd.read_sql_query(sql=sql32, con=self.engine1)
        listT.append(df32)

        print('拒收问题件-绩效 源数据 获取中（三.1）......')
        sql41 = '''SELECT * FROM 拒收挽单_挽单列表_计算统计 s1  WHERE  s1.`更新月份` = '{0}';'''.format(month_time)
        df41 = pd.read_sql_query(sql=sql41, con=self.engine1)
        listT.append(df41)
        print('拒收问题件-绩效 统计分析 获取中（三.2）......')
        sql42 = '''SELECT 新单克隆人, IF(已签收 = 0,NULL, 已签收) AS 已签收,  IF(拒收 = 0,NULL, 拒收) AS 拒收, IF(已退货 = 0,NULL, 已退货) AS 已退货,  IF(理赔 = 0,NULL, 理赔) AS 理赔, 
                            IF(未发货 = 0,NULL, 未发货) AS 未发货, IF(在途 = 0,NULL, 在途) AS 在途, IF(已完成 = 0,NULL, 已完成) AS 已完成,  (总计-无效订单) AS 有效单量, 总计, 
                            concat(ROUND(IFNULL(已签收 / 已完成,0) * 100,2),'%') AS 签收率,
                            concat(ROUND(IFNULL(已完成 / (总计-无效订单),0) * 100,2),'%') AS 完成占比,IF(无效订单 = 0,NULL, 无效订单) AS 无效订单,
                            concat(ROUND(IFNULL((总计-无效订单) / 总计,0) * 100,2),'%') AS 转换率
                FROM ( SELECT 新单克隆人, count(订单编号) as 总计,
                                SUM(IF(最终状态 = "已签收",1,0)) as 已签收, SUM(IF(最终状态 = "拒收",1,0)) as 拒收, SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔, SUM(IF(最终状态 = "未发货",1,0)) as 未发货, SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成, SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单
                        FROM (					
                                SELECT *
                                FROM 拒收挽单_挽单列表_计算统计 s1
                                WHERE  s1.`更新月份` = '{0}'
                        ) s1
                        GROUP BY  新单克隆人
                ) s 
                ORDER BY FIELD(新单克隆人,{1},'合计');'''.format(month_time, username_Jushou)
        df42 = pd.read_sql_query(sql=sql42, con=self.engine1)
        listT.append(df42)

        print('采购异常-绩效 源数据 获取中（四.1）......')
        sql51 = '''SELECT * FROM 采购异常_问题订单_压单_计算统计 s1  WHERE  s1.`更新月份` = '{0}';'''.format(month_time)
        df51 = pd.read_sql_query(sql=sql51, con=self.engine1)
        listT.append(df51)
        print('采购异常-绩效 统计分析 获取中（四.2）......')
        sql52 = '''SELECT 类型, 客服处理人, IF(已签收 = 0,NULL, 已签收) AS 已签收,  IF(拒收 = 0,NULL, 拒收) AS 拒收, IF(已退货 = 0,NULL, 已退货) AS 已退货,  IF(理赔 = 0,NULL, 理赔) AS 理赔, 
                            IF(未发货 = 0,NULL, 未发货) AS 未发货, IF(在途 = 0,NULL, 在途) AS 在途, IF(已完成 = 0,NULL, 已完成) AS 已完成,  (总计-无效订单) AS 有效单量, 总计, 
                            concat(ROUND(IFNULL(已签收 / 已完成,0) * 100,2),'%') AS 签收率,
                            concat(ROUND(IFNULL(已完成 / (总计-无效订单),0) * 100,2),'%') AS 完成占比,IF(无效订单 = 0,NULL, 无效订单) AS 无效订单, 
                            concat(ROUND(IFNULL((总计-无效订单) / 总计,0) * 100,2),'%') AS 转换率
                FROM (		
                        SELECT  类型, 客服处理人,  count(订单编号) as 总计,
                                SUM(IF(最终状态 = "已签收",1,0)) as 已签收, SUM(IF(最终状态 = "拒收",1,0)) as 拒收, SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔, SUM(IF(最终状态 = "未发货",1,0)) as 未发货, SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成, SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单
                        FROM (
                                SELECT * 
                                FROM 采购异常_问题订单_压单_计算统计 s1  
                                WHERE  s1.`更新月份` = '{0}'
                        ) s2
                        GROUP BY  类型,客服处理人	
                ) s
                ORDER BY 类型,FIELD(客服处理人,{1},'合计');;'''.format(month_time, username_caigou_yadan_wentijian)
        df52 = pd.read_sql_query(sql=sql52, con=self.engine1)
        listT.append(df52)

        print('物流问题-绩效 源数据 获取中（三.1）......')
        sql81 = '''SELECT * FROM 物流_派送_问题件_计算统计 s1 WHERE  s1.`更新月份` = '{0}';'''.format(month_time)
        df81 = pd.read_sql_query(sql=sql81, con=self.engine1)
        listT.append(df81)
        print('物流问题 & 派送问题-绩效 统计分析 获取中（三.3）......')
        sql82 = '''SELECT 类型, 客服处理人, IF(已签收 = 0,NULL, 已签收) AS 已签收,  IF(拒收 = 0,NULL, 拒收) AS 拒收, IF(已退货 = 0,NULL, 已退货) AS 已退货,  IF(理赔 = 0,NULL, 理赔) AS 理赔, 
                        IF(未发货 = 0,NULL, 未发货) AS 未发货, IF(在途 = 0,NULL, 在途) AS 在途, IF(已完成 = 0,NULL, 已完成) AS 已完成,  (总计-无效订单) AS 有效单量, 总计, 
                        concat(ROUND(IFNULL(已签收 / 已完成,0) * 100,2),'%') AS 签收率,
                        concat(ROUND(IFNULL(已完成 / (总计-无效订单),0) * 100,2),'%') AS 完成占比,IF(无效订单 = 0,NULL, 无效订单) AS 无效订单, 
                        concat(ROUND(IFNULL((总计-无效订单) / 总计,0) * 100,2),'%') AS 转换率
                FROM (
                        SELECT 类型, 客服处理人,  count(订单编号) as 总计,
                                SUM(IF(最终状态 = "已签收",1,0)) as 已签收, SUM(IF(最终状态 = "拒收",1,0)) as 拒收, SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔, SUM(IF(最终状态 = "未发货",1,0)) as 未发货, SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成, SUM(IF(最终状态 = "无效订单",1,0)) as 无效订单
                        FROM (    
                                SELECT * 
                                FROM 物流_派送_问题件_计算统计 s1 
                                WHERE  s1.`更新月份` = '{0}'
                        ) ss1
                        GROUP BY  客服处理人	
                ) s
                ORDER BY FIELD(客服处理人,{1},'合计');;'''.format(month_time, username_caigou_yadan_wentijian)
        df82 = pd.read_sql_query(sql=sql82, con=self.engine1)
        listT.append(df82)

        file_path = r'''F:\\输出文件\\{0} 绩效数据（明细） {1}.xlsx'''.format(month_time, rq)
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df11.to_excel(excel_writer=writer, sheet_name='促单', index=False)
            df12.to_excel(excel_writer=writer, sheet_name='促单', index=False, startcol=15)
            df31.to_excel(excel_writer=writer, sheet_name='物流客诉', index=False)
            df32.to_excel(excel_writer=writer, sheet_name='物流客诉', index=False, startcol=20)
            df41.to_excel(excel_writer=writer, sheet_name='拒收问题件', index=False)
            df42.to_excel(excel_writer=writer, sheet_name='拒收问题件', index=False, startcol=16)
            df51.to_excel(excel_writer=writer, sheet_name='采购异常', index=False)
            df52.to_excel(excel_writer=writer, sheet_name='采购异常', index=False, startcol=15)
            df81.to_excel(excel_writer=writer, sheet_name='物流派送', index=False)
            df82.to_excel(excel_writer=writer, sheet_name='物流派送', index=False, startcol=15)

if __name__ == '__main__':
    # select = input("请输入需要查询的选项：1=> 按订单查询； 2=> 按时间查询；\n")
    handle = '手动0'
    login_TmpCode = '4b84b336ab9739218a563cde0be598ee'  # 输入登录口令Tkoen
    proxy_handle = '代理服务器0'
    proxy_id = '192.168.13.89:37469'  # 输入代理服务器节点和端口
    select = 8
    m = QueryOrder_Code('+86-18538110674', 'qyz04163510.', login_TmpCode, handle, proxy_handle, proxy_id, select)
    # m = QueryOrder('+86-15565053520', 'sunan1022wang.@&')
    start: datetime = datetime.datetime.now()
    match1 = {'gat': '港台', 'gat_order_list': '港台', 'slsc': '品牌'}
    '''
        # -----------------------------------------------查询状态运行（一）-----------------------------------------
    '''
    # 1、 正在按订单查询；2、正在按时间查询；--->>数据更新切换
    if int(select) == 1:
        print("1-->>> 正在按订单查询+++")
        team = 'gat'
        searchType = '订单号'
        pople_Query = '订单检索'  # 客服查询；订单检索
        m.readFormHost(team, searchType, pople_Query, 'timeStart', 'timeEnd')  # 导入；，更新--->>数据更新切换

    elif int(select) == 99:             # 绩效使用   更新数据使用
        '''
            开始获取 绩效数据
        '''
        timeStart = '2023-05-01'
        timeEnd = '2023-05-31'
        print('开始获取 绩效数据:>>>' + timeStart + "---" + timeEnd)
        m.service_id_order(timeStart, timeEnd, proxy_handle, proxy_id)  # 促单查询；下单时间 @~@ok

        order_time = '创建时间'
        user_caigou = '"蔡利英", "张陈平", "杨嘉仪", "李晓青"'
        m.service_id_ssale(timeStart, timeEnd, proxy_handle, proxy_id, order_time)  # 采购查询；创建时间 （一、获取订单内容）@~@ok
        m.service_id_ssale_info(proxy_handle, proxy_id, '采购异常_创建时间', user_caigou)  # 采购查询；创建时间 （二、获取处理详情）@~@ok

        order_time = '创建时间'
        m.service_id_getRedeemOrderList(timeStart, timeEnd, proxy_handle, proxy_id)  # 挽单列表  查询@~@ok

        order_time = '处理时间'  # 派送问题  处理时间:登记结果处理时间； 创建时间： 订单放入时间@~@
        m.service_id_getDeliveryList(timeStart, timeEnd, order_time, proxy_handle, proxy_id)  # (需处理两次)
        m.service_id_getDeliveryList(timeStart, timeEnd, order_time, proxy_handle, proxy_id)  # (需处理两次)

        order_time = '创建时间'
        m.service_id_waybill_Query(timeStart, timeEnd, proxy_handle, proxy_id, order_time)  # 物流客诉件  查询；订单检索@~@ok

        order_time = '创建时间'
        m.service_id_waybill(timeStart, timeEnd, proxy_handle, proxy_id, order_time)  # 物流问题  压单核实 查询；订单检索ok

        order_time = '创建时间'
        m.service_id_orderInfo(timeStart, timeEnd, proxy_handle, proxy_id)  # 系统问题件  查询；订单检索（处理人员： "蔡利英", "张陈平", "杨嘉仪", "李晓青"）

        order_time = '跟进时间'  # 拒收问题  查询；订单检索@~@ok
        m.service_id_order_js_Query(timeStart, timeEnd, proxy_handle, proxy_id, order_time)  # (需处理两次)

    elif int(select) == 999:  # 更新数据使用
        '''
            开始获取 其他数据
        '''
        timeStart = '2023-05-01'
        timeEnd = '2023-05-10'
        order_time = '跟进时间'
        user_caigou = '"蔡利英", "张陈平", "杨嘉仪", "李晓青"'
        m.service_id_ssale(timeStart, timeEnd, proxy_handle, proxy_id, order_time)  # 采购查询；处理时间 （一、获取订单内容）@~@ok
        m.service_id_ssale_info(proxy_handle, proxy_id, '采购问题件_跟进时间', user_caigou)                             # 采购查询；处理时间 （二、获取处理详情）@~@ok

        order_time = '跟进时间'
        m.service_id_waybill(timeStart, timeEnd, proxy_handle, proxy_id, order_time)              # 物流问题  压单核实 查询；订单检索ok

        order_time = '跟进时间'
        m.service_id_waybill_Query(timeStart, timeEnd, proxy_handle, proxy_id, order_time)       # 物流客诉件  查询；订单检索@~@ok

        order_time = '创建时间'                                                                 # 派送问题   创建时间： 订单放入时间（每次导出时需要更新数据）@~@
        m.service_id_getDeliveryList(timeStart, timeEnd, order_time, proxy_handle, proxy_id)

        order_time = '下单跟进时间'
        m.service_id_order_js_Query(timeStart, timeEnd, proxy_handle, proxy_id, order_time)      # 拒收问题  查询；订单检索@~@ok
        order_time = '下单时间'
        begin = datetime.date(2023, 4, 22)  # 单点更新
        end = datetime.date(2023, 4, 23)
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            day_time = str(day)
            m.service_id_order_js_Query(day_time, day_time, proxy_handle, proxy_id, order_time)      # 拒收问题  查询；订单检索@~@ok

    elif int(select) == 8:      # 单独 本月绩效数据使用 不包含上月的留底数据
        username_Cudan = '"刘文君","马育慧","曲开拓","闫凯歌","杨昊","周浩迪","曹可可"'                                         # 促单人
        username_Jushou = '"刘文君","马育慧","曲开拓","闫凯歌","杨昊","周浩迪","曹可可","蔡利英","杨嘉仪","张陈平","李晓青"'        # 拒收挽单
        username_caigou_yadan_wentijian = '"蔡利英","杨嘉仪","张陈平","李晓青"'                                             # 采购问题压单
        rq_month = '202305'  # 统计月份
        rq_day = '2023-06-05'  # 统计日期
        m.service_check(username_Cudan, username_Jushou, username_caigou_yadan_wentijian,  rq_month, rq_day)  # 绩效数据导出


    elif int(select) == 66:    # 本月 更新上月 留底的未完结数据 绩效数据使用（二）
        update_old = '202305'  # 上月的 更新的月份  --  上月未完结的订单
        update_new = '202306'  # 本月的 更新月份  --   上月未完结的订单
        m.userid_performance_old_upadata(update_old, update_new)

    elif int(select) == 5:      #  本月 统计上月 绩效数据使用（一）
        username_Cudan = '"刘文君","马育慧","曲开拓","闫凯歌","杨昊","周浩迪","曹可可","曲开拓"'                                         # 促单人
        username_Jushou = '"刘文君","马育慧","曲开拓","闫凯歌","杨昊","周浩迪","曹可可","蔡利英","杨嘉仪","张陈平","李晓青","曲开拓"'        # 拒收挽单
        username_caigou_yadan_wentijian = '"蔡利英","杨嘉仪","张陈平","李晓青"'                                             # 采购问题压单
        rq_month = '202305'  # 统计月份
        rq_day = '2023-06-05'  # 统计日期
        m.userid_performance_New(username_Cudan, username_Jushou, username_caigou_yadan_wentijian, rq_month, rq_day)


    elif int(select) == 7:    # 本月  绩效数据 导出（三）
        m.read_write_workbook()         # 先更新客诉的 回款状态
        username_Cudan = '"刘文君","马育慧","曲开拓","闫凯歌","杨昊","周浩迪","曹可可","曲开拓"'                                         # 促单人
        username_Jushou = '"刘文君","马育慧","曲开拓","闫凯歌","杨昊","周浩迪","曹可可","蔡利英","杨嘉仪","张陈平","李晓青","曲开拓"'        # 拒收挽单
        username_caigou_yadan_wentijian = '"蔡利英","杨嘉仪","张陈平","李晓青"'                                             # 采购问题压单
        month_time = '202306'  # 更新月份
        m.userid_performance_New_export(username_Cudan, username_Jushou, username_caigou_yadan_wentijian, month_time)




    elif int(select) == 2:
        # m._service_id_orderInfoTWO('2023-01-01')              # 系统问题件  查询；订单检索  单独测试使用
        # m._service_id_orderInfoThree('2023-01-01')            # 系统问题件  查询；订单检索  单独测试使用
        pass

    print('查询耗时：', datetime.datetime.now() - start)