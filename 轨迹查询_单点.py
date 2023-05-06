import pandas as pd
import os
import datetime
import time
import xlwings
import xlsxwriter
import math
import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel
from bs4 import BeautifulSoup # 抓标签里面元素的方法
from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色


# -*- coding:utf-8 -*-
class QueryTwo(Settings, Settings_sso):
    def __init__(self, userMobile, password, login_TmpCode, handle, proxy_handle, proxy_id):
        Settings.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        # self._online()
        # self.sso_online_Two()
        # self.sso__online_handle(login_TmpCode)
        # self.sso__online_auto()
        if proxy_handle == '代理服务器':
            if handle == '手动':
                self.sso__online_handle_proxy(login_TmpCode, proxy_id)
            else:
                self.sso__online_auto_proxy(proxy_id)
        else:
            if handle == '手动':
                self.sso__online_handle(login_TmpCode)
            else:
                self.sso__online_auto()

        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
    # 获取签收表内容
    def readFormHost(self, isReal, proxy_handle, proxy_id, cat):
        start = datetime.datetime.now()
        path = r'F:\需要用到的文件\A查询导表'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                self.wbsheetHost(filePath, isReal, proxy_handle, proxy_id, cat)
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, isReal, proxy_handle, proxy_id, cat):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                if sht.api.Visible == -1:
                    try:
                        tem = ''
                        db = None
                        db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                        # columns_value = list(db.columns)                             # 获取数据的标题名，转为列表
                        # for column_val in columns_value:
                        #     if '运单编号' == column_val:
                        #         tem = column_val
                        #     elif '运单号' == column_val:
                        #         tem = column_val
                        if '运单编号' in db.columns:
                            tem = '运单编号'
                        if '回执单号' in db.columns:
                            tem = '回执单号'
                        elif '查件单号' in db.columns:
                            tem = '查件单号'
                        elif '运单号' in db.columns:
                            tem = '运单号'
                        elif '物流单号' in db.columns:
                            tem = '物流单号'
                        db = db[[tem]]
                        db.dropna(axis=0, how='any', inplace=True)                  # 空值（缺失值），将空值所在的行/列删除后
                    except Exception as e:
                        print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                    if db is not None and len(db) > 0:
                        print(db)
                        print('++++正在获取：' + sht.name + ' 表；共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        # 将获取到的运单号 查询轨迹
                        self.Search_online(db, isReal, tem, proxy_handle, proxy_id, cat)
                    else:
                        print('----------数据为空,查询失败：' + sht.name)
                else:
                    print('----不需查询：' + sht.name)
            wb.close()
        app.quit()

    #  查询运单轨迹-按订单查询（一）
    def Search_online(self, db, isReal, tem, proxy_handle, proxy_id, cat):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        orderId = list(db[tem])
        print(orderId)
        max_count = len(orderId)                 # 使用len()获取列表的长度，上节学的
        if max_count > 0:
            df = pd.DataFrame([])                # 创建空的dataframe数据框
            dlist = []
            for ord in orderId:
                print(ord)
                ord = str(ord)
                print(type(ord))
                # if type(ord) == "<class 'str'>":
                #     print('字符串')
                # else:
                #     print(ord[:3])
                if cat == 0:
                    if ord[:3] == '620' or ord[:3] == '901':
                        print('单独 - 黑猫查询中')
                        data = self._SearchGoods_heimao(ord, proxy_handle, proxy_id)
                    else:
                        print('单独 - 单点查询中')
                        data = self._order_online(ord, isReal, proxy_handle, proxy_id)
                        print(data)
                elif cat == 1:
                    print('全体 - 单点查询中')
                    data = self._order_online(ord, isReal, proxy_handle, proxy_id)
                    print(data)
                if data is not None and len(data) > 0:
                    dlist.append(data)
            dp = df.append(dlist, ignore_index=True)
            # dp.dropna(axis=0, how='any', inplace=True)
        else:
            dp = None
        print(dp)
        dp = dp[['orderNumber', 'wayBillNumber', 'track_date', '出货时间', '上线时间', '保管时间', '完成时间', 'track_info', 'track_status', '负责营业所', '轨迹备注', '序号', '便利店']]
        dp.columns = ['订单编号', '运单号', '物流轨迹时间', '出货时间', '上线时间', '保管时间', '完成时间', '物流轨迹', '轨迹代码', '负责营业所', '轨迹备注', '序号', '便利店']
        dp.to_excel('F:\\输出文件\\运单轨迹-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')   # Xlsx是python用来构造xlsx文件的模块，可以向excel2007+中写text，numbers，formulas 公式以及hyperlinks超链接。
        print('查询已导出+++')
        print('*' * 50)

    def Search_online_write(self, isReal, proxy_handle, proxy_id):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        sql = '''SELECT 运单编号
                FROM {0} s
                WHERE s.`运单编号` NOT IN (SELECT DISTINCT 运单号 FROM 轨迹查询);'''.format('运单号')
        ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
        if ordersDict.empty:
            print('无需要更新订单信息！！！')
            return
        orderId = list(ordersDict['运单编号'])
        max_count = len(orderId)                 # 使用len()获取列表的长度，上节学的
        print('++++++本次需查询;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        n = 0
        while n <= max_count + 10:  # 这里用到了一个while循环，穿越过来的
            order = orderId[n:n + 10]
            df = pd.DataFrame([])                # 创建空的dataframe数据框
            dlist = []
            for ord in order:
                print(ord)
                ord = str(ord)
                print(type(ord))
                if ord[:3] == '620' or ord[:3] == '901':
                    print('黑猫查询中')
                    data = self._SearchGoods_heimao(ord, proxy_handle, proxy_id)
                else:
                    print('单点查询中')
                    data = self._order_online(ord, isReal, proxy_handle, proxy_id)
                    print(data)
                if data is not None and len(data) > 0:
                    dlist.append(data)
            dp = df.append(dlist, ignore_index=True)
            print(dp)
            dp = dp[['orderNumber', 'wayBillNumber', 'track_date', '出货时间', '上线时间', '保管时间', '完成时间', 'track_info', 'track_status', '负责营业所', '轨迹备注', '序号']]
            dp.columns = ['订单编号', '运单号', '物流轨迹时间', '出货时间', '上线时间', '保管时间', '完成时间', '物流轨迹', '轨迹代码', '负责营业所', '轨迹备注', '序号']
            dp.to_sql('cache_waybill', con=self.engine1, index=False, if_exists='replace')
            columns = list(dp.columns)
            columns = ','.join(columns)
            sql = 'REPLACE INTO {0}({1}, 添加时间) SELECT *, NOW() 添加时间 FROM cache_waybill; '.format('轨迹查询', columns)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            # dp.to_excel('F:\\输出文件\\运单轨迹-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')   # Xlsx是python用来构造xlsx文件的模块，可以向excel2007+中写text，numbers，formulas 公式以及hyperlinks超链接。
            print('查询已导出+++')
            print('*' * 50)
            n = n + 10
            print('剩余查询信息：' + str(max_count - n) + ' 条信息+++++++')

    #  查询运单轨迹-按时间查询（二）
    def order_online(self, timeStart, timeEnd, isReal, proxy_handle, proxy_id):  # 进入运单轨迹界面
        # print('正在获取需要订单信息......')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        start = datetime.datetime.now()
        begin = datetime.datetime.strptime(timeStart, '%Y-%m-%d')
        end = datetime.datetime.strptime(timeEnd, '%Y-%m-%d')
        for i in range((end - begin).days):
            day = begin + datetime.timedelta(days=i + 1)
            day = day.strftime('%Y-%m-%d')
            print('正在获取 ' + day + ' 号订单信息…………')
            sql = '''SELECT id,`运单编号`  FROM gat_order_list sl WHERE sl.`日期` = '{0}' AND sl.运单编号 IS NOT NULL;'''.format(day)
            # sql = '''SELECT id,`运单编号`  FROM gat_order_list sl WHERE sl.`下单时间` BETWEEN  '2022-02-03 09:00:00' AND '2022-02-03 09:30:00' AND sl.运单编号 IS NOT NULL;'''.format(day)
            ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
            if ordersDict.empty:
                print('无需要更新订单信息！！！')
                return
            # print(ordersDict['运单编号'][0])
            orderId = list(ordersDict['运单编号'])
            max_count = len(orderId)            # 使用len()获取列表的长度，上节学的
            if max_count > 0:
                df = pd.DataFrame([])           # 创建空的dataframe数据框
                dlist = []
                for ord in orderId:
                    print(ord)
                    data = self._order_online(ord, isReal, proxy_handle, proxy_id)
                    if data is not None and len(data) > 0:
                        dlist.append(data)
                dp = df.append(dlist, ignore_index=True)
                dp.to_excel('F:\\输出文件\\运单轨迹 {0} .xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
        print('++++++查询成功+++++++')
        print('查询耗时：', datetime.datetime.now() - start)
        print('*' * 50)

    #  查询运单轨迹-按订单查询（一 、1.单点）
    def _order_online(self, ord, isReal, proxy_handle, proxy_id):  # 进入订单检索界面
        print('+++实时_搜索轨迹信息中')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getLogisticsTrace'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/logisticsTrajectory'}
        data = {'numbers': ord,
                'searchType': 1,
                'isReal': isReal
                }
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        # print(5)
        print(req['data'])
        if req['data'] == []:
            data = self._order_online_data(ord, 0, proxy_handle, proxy_id)
            return data
        elif req['data']['list'] == []:
            data = self._order_online_data(ord, 0, proxy_handle, proxy_id)
            return data
        else:
            ordersDict = []
            try:
                if req['data']['list'][0]['list'] == []:
                    # print(req['data']['list'])
                    return None
                else:
                    step = 0
                    for result in req['data']['list']:
                        # for res in result['list']:
                        for index, res in enumerate(result['list']):
                            res['序号'] = index + 1
                            res['负责营业所'] = ''
                            res['轨迹备注'] = ''
                            res['出货时间'] = ''
                            res['上线时间'] = ''
                            res['完成时间'] = ''
                            res['保管时间'] = ''
                            res['orderNumber'] = result['order_number']
                            res['wayBillNumber'] = result['track_no']
                            res['便利店'] = ''
                            if '.' in res['track_date']:
                                res['track_date'] = res['track_date'].split('.')[0]
                            else:
                                res['track_date'] = res['track_date']

                            if '二次出貨' in res['track_info'] or '二次出貨貼標' in res['track_info'] or '出貨理貨中' in res['track_info'] or '首發出貨貼標' in res['track_info'] or '已发货' in res['track_info'] or '已集貨' in res['track_info'] or '顺丰速运 已收取快件' in res['track_info'] or '已核单打包出库' in res['track_info']:
                                res['出货时间'] = res['track_date']
                            elif '已核重,集运仓' in res['track_info'] or '已核重-集运仓' in res['track_info'] or '取件中' in res['track_info'] or '【已核重】-集运仓发货' in res['track_info']:
                                res['出货时间'] = res['track_date']

                            if '貨件已抵達 土城營業所，貨件整理中' in res['track_info'] or '貨件已抵達土城營業所，貨件整理中' in res['track_info'] or '轉交配送中'in res['track_info'] or '進驗成功 包裹已送達物流中心，進行理貨中' in res['track_info'] or '貨件已抵達桃園營業所，貨件整理中。貨物件數' in res['track_info']  or '貨件已抵達桃園營業所 ，貨件整理中。貨物件數' in res['track_info'] \
                                or '快件在青衣航运路速運營業點（不對公眾開放）完成分拣' in res['track_info'] or '進驗成功 包裹已送達物流中心，進行理貨中' in res['track_info'] or 'SRP 出货回档 ' in res['track_info']:
                                res['上线时间'] = res['track_date']
                            elif '貨件已抵達 大園營業所 ，貨件整理中' in res['track_info'] or '貨件已抵達大園營業所，貨件整理中。貨物件數' in res['track_info'] or '轉運中' in res['track_info'] or '快件到达 【香港青衣中轉場】' in res['track_info']:
                                res['上线时间'] = res['track_date']
                            elif '營業所發出，前往配送站途中' in res['track_info']:
                                res['上线时间'] = res['track_date']
                            elif '營業所，分貨中' in res['track_info'] or 'PPS 到达门店' in res['track_info']:
                                res['上线时间'] = res['track_date']

                            if ('送達客戶不在' in res['track_info'] or '貨件拒收' in res['track_info'] or '與客戶另約時間' in res['track_info'] or '收貨人' in res['track_info']) and '保管中' in res['track_info'] or '營業所送至' in res['track_info'] or '集配站送至' in res['track_info'] or '已与收方客户约定' in res['track_info'] or '快件到达指定自取点' in res['track_info']:
                                res['保管时间'] = res['track_date']
                            elif ('貨件由' in res['track_info'] and '保管中' in res['track_info']) or '收件地址為公司行號，本日休假' in res['track_info'] or '暫置營業所' in res['track_info'] or (('收貨人' in res['track_info'] or '該地址' in res['track_info']) and '營業所處理中' in res['track_info']) or '快件派送不成功' in res['track_info']:
                                res['保管时间'] = res['track_date']

                            if '送達。貨物件數共' in res['track_info'] or '貨件已退回，退貨號碼' in res['track_info'] or '順利送達' in res['track_info'] or '在官网"运单资料&签收图"，可查看签收人信息' in res['track_info'] or '已完成包裹成功取件' in res['track_info'] or 'ESP 成功签收已取件' in res['track_info']:
                                res['完成时间'] = res['track_date']
                            elif '退件' in res['track_info'] or '货物退回-已上架' in res['track_info']:
                                res['完成时间'] = res['track_date']

                            # print(res)track_info
                            if step >= 1:
                                continue
                            else:
                                if '送至' in res['track_info']:
                                    vt = res['track_info'].split('送至')[1]
                                    vt2 = vt.split('(')[0]
                                    if '7-11' in vt2:
                                        vt3 = vt2.replace('7-11', 'Seven Eleven')
                                    else:
                                        vt3 = vt2
                                    res['便利店'] = vt3
                                    step = step + 1
                                elif '與客戶另約時間配送' in res['track_info']:
                                    res['便利店'] = '与客户另约时间配送'
                                elif '與客戶另約時間到站自領' in res['track_info']:
                                    res['便利店'] = '与客户另约时间到站自领'
                            ordersDict.append(res)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersDict)
            data.sort_values(by=["orderNumber", "track_date"], inplace=True, ascending=[True, True])  # inplace: 原地修改; ascending：升序 （是否升序排序，默认为true，降序则为false。如果是列表，则需和by指定的列表数量相同，指明每一列的排序方式）
            print('++++++本次获取成功+++++++')
            # print('*' * 50)
            return data
    #  物流轨迹数据库         （一 、2.单点）
    def _order_online_data(self, ord, isReal, proxy_handle, proxy_id):  # 进入订单检索界面
        print('+++数据库_搜索轨迹信息中')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getLogisticsTrace'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/logisticsTrajectory'}
        data = {'numbers': ord,
                'searchType': 1,
                'isReal': isReal
                }
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        if req['data']['list'] == []:
            return None
        else:
            ordersDict = []
            try:
                if req['data']['list'][0]['list'] == []:
                    # print(req['data']['list'])
                    return None
                else:
                    step = 0
                    for result in req['data']['list']:
                        # for res in result['list']:
                        for index, res in enumerate(result['list']):
                            res['序号'] = index + 1
                            # res['出货上线时间'] = ''
                            # res['orderNumber'] = result['order_number']
                            # res['wayBillNumber'] = result['track_no']
                            # if '.' in res['track_date']:
                            #     res['track_date'] = res['track_date'].split('.')[0]
                            # else:
                            #     res['track_date'] = res['track_date']
                            # if '已核重-集运仓发货' in res['track_info'] or '顺丰速运 已收取快件' in res['track_info'] or '貨件整理中' in res['track_info'] or '二次出貨貼標' in res['track_info']:
                            #     res['出货上线时间'] = res['track_date']
                            res['负责营业所'] = ''
                            res['轨迹备注'] = ''
                            res['出货时间'] = ''
                            res['上线时间'] = ''
                            res['完成时间'] = ''
                            res['保管时间'] = ''
                            res['orderNumber'] = result['order_number']
                            res['wayBillNumber'] = result['track_no']
                            res['便利店'] = ''
                            if '.' in res['track_date']:
                                res['track_date'] = res['track_date'].split('.')[0]
                            else:
                                res['track_date'] = res['track_date']

                            if '二次出貨' in res['track_info'] or '二次出貨貼標' in res['track_info'] or '出貨理貨中' in res['track_info'] or '首發出貨貼標' in res['track_info'] or '已发货' in res['track_info'] or '已集貨' in res['track_info'] or '顺丰速运 已收取快件' in res['track_info'] or '已核单打包出库' in res['track_info']:
                                res['出货时间'] = res['track_date']
                            elif '已核重,集运仓' in res['track_info'] or '已核重-集运仓' in res['track_info'] or '取件中' in res['track_info'] or '【已核重】-集运仓发货'  in res['track_info']:
                                res['出货时间'] = res['track_date']

                            if '貨件已抵達 土城營業所，貨件整理中' in res['track_info'] or '貨件已抵達土城營業所，貨件整理中' in res['track_info'] or '轉交配送中'in res['track_info'] or '進驗成功 包裹已送達物流中心，進行理貨中' in res['track_info'] or '貨件已抵達桃園營業所，貨件整理中。貨物件數' in res['track_info']  or '貨件已抵達桃園營業所 ，貨件整理中。貨物件數' in res['track_info'] \
                                or '快件在青衣航运路速運營業點（不對公眾開放）完成分拣' in res['track_info'] or '進驗成功 包裹已送達物流中心，進行理貨中' in res['track_info']:
                                res['上线时间'] = res['track_date']
                            elif '貨件已抵達 大園營業所 ，貨件整理中' in res['track_info'] or '貨件已抵達大園營業所，貨件整理中。貨物件數' in res['track_info'] or '轉運中' in res['track_info'] or '快件到达 【香港青衣中轉場】' in res['track_info']:
                                res['上线时间'] = res['track_date']
                            elif '營業所發出，前往配送站途中' in res['track_info']:
                                res['上线时间'] = res['track_date']
                            elif '營業所，分貨中' in res['track_info']:
                                res['上线时间'] = res['track_date']

                            if ('送達客戶不在' in res['track_info'] or '貨件拒收' in res['track_info'] or '與客戶另約時間' in res['track_info'] or '收貨人' in res['track_info']) and '保管中' in res['track_info'] or '營業所送至' in res['track_info'] or '集配站送至' in res['track_info'] or '已与收方客户约定' in res['track_info'] or '快件到达指定自取点' in res['track_info']:
                                res['保管时间'] = res['track_date']
                            elif ('貨件由' in res['track_info'] and '保管中' in res['track_info']) or '收件地址為公司行號，本日休假' in res['track_info'] or '暫置營業所' in res['track_info'] or (('收貨人' in res['track_info'] or '該地址' in res['track_info']) and '營業所處理中' in res['track_info']) or '快件派送不成功' in res['track_info']:
                                res['保管时间'] = res['track_date']

                            if '送達。貨物件數共' in res['track_info'] or '貨件已退回，退貨號碼' in res['track_info'] or '順利送達' in res['track_info'] or '在官网"运单资料&签收图"，可查看签收人信息' in res['track_info'] or '已完成包裹成功取件' in res['track_info']:
                                res['完成时间'] = res['track_date']
                            elif '退件' in res['track_info'] or '货物退回-已上架' in res['track_info']:
                                res['完成时间'] = res['track_date']

                            if step >= 1:
                                continue
                            else:
                                if '送至' in res['track_info']:
                                    vt = res['track_info'].split('送至')[1]
                                    vt2 = vt.split('(')[0]
                                    if '7-11' in vt2:
                                        vt3 = vt2.replace('7-11', 'Seven Eleven')
                                    else:
                                        vt3 = vt2
                                    res['便利店'] = vt3
                                    step = step + 1
                                elif '與客戶另約時間配送' in res['track_info']:
                                    res['便利店'] = '与客户另约时间配送'
                                elif '與客戶另約時間到站自領' in res['track_info']:
                                    res['便利店'] = '与客户另约时间到站自领'

                            ordersDict.append(res)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersDict)
            data.sort_values(by=["orderNumber", "track_date"], inplace=True, ascending=[True, True])  # inplace: 原地修改; ascending：升序 （是否升序排序，默认为true，降序则为false。如果是列表，则需和by指定的列表数量相同，指明每一列的排序方式）
            print('++++++本次获取成功+++++++')
            # print('*' * 50)
            return data

    #  查询运单轨迹-按运单查询（一 、3.黑猫官网）
    def _SearchGoods_heimao(self,wayBillNumber, proxy_handle, proxy_id):
        #1、构建url 、请求数据
        url = "https://www.t-cat.com.tw/Inquire/TraceDetail.aspx?BillID=" + wayBillNumber   #url为机器人的webhook
        #2、构建一下请求头部
        r_header = {"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3",
                    "Accept-Encoding": "gzip, deflate, br",
                    "Accept-Language": "zh-CN,zh;q=0.9",
                    "Cache-Control": "no-cache",
                    "Connection": "keep-alive",
                    # "Connection": "close",
                    # "Content-Length": '92',
                    # "Content-Type": "application/x-www-form-urlencoded",
                    "Host": "www.t-cat.com.tw",
                    'Origin': 'https://www.t-cat.com.tw',
                    # "Pragma": "no-cache",
                    # "Pragma": "1",
                    "Referer": "https://www.t-cat.com.tw/Inquire/trace.aspx",
                    "Sec-Fetch-Mode": "navigate",
                    "Sec-Fetch-Site": "same-origin",
                    "Sec-Fetch-User": "?1",
                    "Upgrade-Insecure-Requests": '1',
                    "Charset": "UTF-8",
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36"
                    # "User-Agent": user_agent()
                    # "User-Agent": random.choice(user_agent_list)
                    }
        #3、构建请求数据
        data = {'BillID': 620430597712}
        if proxy_handle == '代理服务器':             # 使用代理ip发送请求
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # req = self.session.get(url=url, headers=r_header, data=data, allow_redirects=False)
        # print(req)
        # print('----------数据获取返回成功-----------')
        rq = BeautifulSoup(req.text, 'lxml')
        rq = rq.find_all('tr')
        ordersDict = []
        L_waybill = ''
        L_time = ''
        L_info = ''
        L_info2 = ''
        L_info3 = ''
        for index, val in enumerate(rq):
            result = {}
            # print(index)
            # print(val)
            # print('-' * 10)
            if "top" in str(val) and "time" not in str(val):            # 没有查询到货态
                result['序号'] = ""
                result['orderNumber'] = ""
                result['wayBillNumber'] = ""
                result['track_date'] = ""
                result['出货时间'] = ""
                result['上线时间'] = ""
                result['保管时间'] = ""
                result['完成时间'] = ""
                result['track_info'] = ""
                result['负责营业所'] = ""
                result['轨迹备注'] = ""
                result['便利店'] = ''
            if "height" in str(val) and "rowspan" in str(val):       # 查询到货态（一）
                L_waybill = str(val).split('</span>')[0].split('bl12">')[1]
                L_time_val = str(val).split('<br/>')
                L_time10 = L_time_val[0].split('bl12">')
                L_time11 = L_time_val[0].split('bl12">')[len(L_time10)-1]
                L_time22 = L_time_val[1].split('</span>')[0]
                L_time = L_time11 + '' + L_time22
                if 'strong' in str(val):
                    L_info = str(val).split('<strong>')[1].split('</strong>')[0]
                else:
                    L_info = str(val).split('bl12">')[2].split('</span>')[0]
                L_info2 = str(val).split('foothold.aspx?n=')[1].split('</a>')[0]
                L_info2 = L_info2.split('>')[1]
                L_info3 = str(val).split('title=')[1].split('>')[0]
                result['序号'] = index
                result['orderNumber'] = ""
                result['wayBillNumber'] = L_waybill
                result['track_date'] = L_time
                result['出货时间'] = ""
                result['上线时间'] = ""
                result['保管时间'] = ""
                result['完成时间'] = ""
                result['track_info'] = L_info
                result['轨迹备注'] = L_info3
                result['负责营业所'] = L_info2
                result['轨迹备注'] = L_info3
                result['便利店'] = ''
                if '包裹已經送達收件人' in L_info3:
                    result['完成时间'] = L_time
                elif '順利送達' in L_info:
                    result['完成时间'] = L_time
            if "height" not in str(val) and "rowspan" not in str(val) and "<br/>" in str(val):
                L_time_val = str(val).split('<br/>')
                L_time10 = L_time_val[0].split('bl12">')
                L_time11 = L_time_val[0].split('bl12">')[len(L_time10)-1]
                L_time22 = L_time_val[1].split('</span>')[0]
                L_time = L_time11 + '' + L_time22
                if 'strong' in str(val):
                    L_info = str(val).split('<strong>')[1].split('</strong>')[0]
                else:
                    L_info = str(val).split('bl12">')[1].split('</span>')[0]
                L_info2 = str(val).split('foothold.aspx?n=')[1].split('</a>')[0]
                L_info2 = L_info2.split('>')[1]
                L_info3 = str(val).split('title=')[1].split('>')[0]
                result['序号'] = index
                result['orderNumber'] = ""
                result['wayBillNumber'] = L_waybill
                result['track_date'] = L_time
                result['出货时间'] = ""
                result['上线时间'] = ""
                result['保管时间'] = ""
                result['完成时间'] = ""
                result['track_info'] = L_info
                result['负责营业所'] = L_info2
                result['轨迹备注'] = L_info3
                result['便利店'] = ''
                if 'sd已經至寄件人指定地點收到包裹' in L_info3:
                    result['出货时间'] = L_time
                elif '已集貨' in L_info:
                    result['出货时间'] = L_time

                if '包裹正從營業所送到轉運中心，或從轉運中心送到營業所' in L_info3:
                    result['上线时间'] = L_time
                elif '轉運中' in L_info or 'sd正在將包裹配送到收件人途中' in L_info3:
                    result['上线时间'] = L_time

                if '暫置營業所保管中' in L_info:
                    result['保管时间'] = L_time
                elif '調查處理中' in L_info or '另約時間' in L_info:
                    result['保管时间'] = L_time

                if '包裹已經送達收件人' in L_info3:
                    result['完成时间'] = L_time
                elif '順利送達' in L_info:
                    result['完成时间'] = L_time
            ordersDict.append(result)
        data = pd.json_normalize(ordersDict)
        print(data)
        nan_value = float("NaN")  # 用null替换所有空位，然后用dropna函数删除所有空值列。
        data.replace("", nan_value, inplace=True)
        data.dropna(axis=0, how='any', inplace=True, subset=['wayBillNumber'])
        data.sort_values(by=["wayBillNumber", "track_date"], inplace=True, ascending=[True, True])
        # print(99)
        # print(data)
        # rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        # data.to_excel('F:\\输出文件\\黑猫宅急便 {0} .xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
        return data




    #  查询运单轨迹-上线及派送（三）
    def order_bind_status(self, timeStart, timeEnd):  # 进入运单轨迹界面
        # print('正在获取需要订单信息......')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        start = datetime.datetime.now()
        begin = datetime.datetime.strptime(timeStart, '%Y-%m-%d')
        end = datetime.datetime.strptime(timeEnd, '%Y-%m-%d')
        for i in range((end - begin).days):
            day = begin + datetime.timedelta(days=i + 1)
            day = day.strftime('%Y-%m-%d')
            print('正在获取 ' + day + ' 号订单信息…………')
            # sql = '''SELECT id,`运单编号`  FROM gat_order_list sl WHERE sl.`日期` = '{0}' AND sl.运单编号 IS NOT NULL;'''.format(day)
            sql = '''SELECT id,`运单编号`  FROM gat_order_list sl WHERE sl.`下单时间` BETWEEN  '2022-02-03 00:00:00' AND '2022-02-03 23:59:59' AND sl.运单编号 IS NOT NULL;'''.format(day)
            ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
            if ordersDict.empty:
                print('无需要更新订单信息！！！')
                return
            # print(ordersDict['运单编号'][0])
            orderId = list(ordersDict['运单编号'])
            max_count = len(orderId)            # 使用len()获取列表的长度，上节学的
            if max_count > 0:
                df = pd.DataFrame([])           # 创建空的dataframe数据框
                dlist = []
                for ord in orderId:
                    print(ord)
                    data = self._order_bind_status(ord)
                    if data is not None and len(data) > 0:
                        dlist.append(data)
                dp = df.append(dlist, ignore_index=True)
                dp.to_excel('F:\\输出文件\\运单轨迹 {0} .xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
        print('++++++查询成功+++++++')
        print('查询耗时：', datetime.datetime.now() - start)
        print('*' * 50)
    def _order_bind_status(self, ord):  # 进入订单检索界面
        print('+++正在查询订单信息中')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getLogisticsTrace'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/logisticsTrajectory'}
        data = {'numbers': ord,
                'searchType': 1,
                'isReal': 1
                }
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersDict = []
        val = {}
        try:
            if req['data']['list'] == []:
                print(req['data']['list'])
                return None
            else:
                for result in req['data']['list']:
                    # print(result)
                    # print(999)
                    k = 0
                    t = 0
                    g = 0
                    j = 0
                    h = 0
                    val['订单编号'] = result['order_number']
                    val['运单编号'] = result['track_no']
                    for res in result['list']:
                        # print(0)
                        # print(res)
                        if '快件到达 【香港青衣中轉場】' in res['track_info']:         # 立邦顺丰
                            while t == 0:
                                val['上线时间'] = res['track_date']
                                t = t + 1
                            for res in result['list']:
                                # print(11)
                                # print(res)
                                if '正在派送途中' in res['track_info']:
                                    g = g + 1
                                    val[str(g) +'派'] = res['track_date']
                                    # print('g:' + str(g) +' 派')

                        elif '項目營業部' in res['track_info']:                       # 天马顺丰
                            val['上线时间'] = res['track_date']
                            for res in result['list']:
                                # print(22)
                                # print(res)
                                if '正在派件...' in res['track_info'] or '快件派送不成功' in res['track_info'] or '快件已签收' in res['track_info'] or '快件代签收' in res['track_info']:
                                    j = j + 1
                                    val[str(j) +'派'] = res['track_date']
                                    # print('j:' + str(j))

                        elif '貨件已抵達土城營業所，貨件整理中' in res['track_info'] or '貨件已抵達桃園營業所，貨件整理中' in res['track_info'] or '集荷' in res['track_info']:      # 新竹 貨件已抵達桃園營業所，貨件整理中
                            val['上线时间'] = res['track_date']
                            for res in result['list']:
                                # print(33)
                                # print(res)
                                if '配送中' in res['track_info']:
                                    k = k + 1
                                    val[str(k) +'派'] = res['track_date']
                                    # print('k:' + str(k))

                        elif '結轉物流中心' in res['track_info']:                         # 7-11 
                            val['上线时间'] = res['track_date']
                            for res in result['list']:
                                # print(44)
                                # print(res)
                                if '门市配达' in res['track_info']:
                                    h = h + 1
                                    val[str(h) +'派'] = res['track_date']
                                    # print('h:' + str(h))
                        elif '已安排航班' in res['track_info'] or '起飞' in res['track_info']:                         # 7-11 
                            val['上线时间'] = res['track_date']
                            for res in result['list']:
                                # print(55)
                                # print(res)
                                if '配送中' in res['track_info']:
                                    h = h + 1
                                    val[str(h) +'派'] = res['track_date']
                                    # print('h:' + str(h))
                        # if '貨件已抵達土城營業所，貨件整理中' in res['track_info'] or '貨件已抵達桃園營業所，貨件整理中' in res['track_info']:      # 新竹
                        #     val['上线时间'] = res['track_date']
                        # if '配送中' in res['track_info']:
                        #     k = k + 1
                        #     val[str(k) +'派'] = res['track_date']

                        # if '結轉物流中心' in res['track_info']:         # 7-11 
                        #     val['上线时间'] = res['track_date']
                        # if '门市配达' in res['track_info']:
                        #     val[str(k) +'派'] = res['track_date']

                        # if '項目營業部' in res['track_info']:         # 顺丰
                        #     val['上线时间'] = res['track_date']
                        # if '正在派件' in res['track_info'] or '快件已' in res['track_info'] or '快件代' in res['track_info'] or '快件派送' in res['track_info']:
                        #     val[str(k) +'派'] = res['track_date']

                        # if '快件到达 【香港青衣中轉場】' in res['track_info']:         # 立邦
                        #     while t == 0:
                        #         t= t + 1
                        #         val['上线时间'] = res['track_date']
                        # if '正在派送途中' in res['track_info']:
                        #     g = g + 1
                        #     val[str(g) +'派'] = res['track_date']
                    ordersDict.append(val)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersDict)
        print(data)
        # data.sort_values(by="track_date", inplace=True, ascending=True)  # inplace: 原地修改; ascending：升序
        # data['name'] = data['name'].str.strip()
        # data['cate_id'] = data['cate_id'].str.strip()
        # data['second_cate_id'] = data['second_cate_id'].str.strip()
        # data['third_cate_id'] = data['third_cate_id'].str.strip()
        # data = data[['id', 'name', 'cate_id', 'second_cate_id', 'third_cate_id', 'status', 'price', 'selectionName',
        #              'sellerCount', 'buyerName', 'saleCount', 'logisticsCost', 'lender', 'isGift', 'createTime',
        #              'categorys', 'image']]
        # data.columns = ['产品id', '产品名称', '一级分类', '二级分类', '三级分类', '产品状态', '价格(￥)', '选品人',
        #                 '供应商数', '采购人', '商品数', 'logisticsCost', '出借人', '特殊信息', '添加时间',
        #                 '产品分类', '产品图片']
        # data.to_excel('F:\\输出文件\\运单轨迹 {0} .xlsx'.format(rq), sheet_name='查询', index=False,engine='xlsxwriter')
        print('++++++本批次查询成功+++++++')
        # print('*' * 50)
        return data



    # 语音外呼 派送问题件           （一.1）
    def getDeliveryList(self, timeStart, timeEnd, order_time, proxy_handle, proxy_id, tm_data, message):  # 进入订单检索界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在查询 派送问题件(' + order_time + ') 起止时间：' + str(timeStart) + " *** " + str(timeEnd))
        url = r'https://gimp.giikin.com/service?service=gorder.deliveryQuestion&action=getDeliveryList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/deliveryProblemPackage'}
        data = {'order_number': None, 'waybill_number': None, 'question_level': None, 'question_type': None,'order_trace_id': None, 'ship_phone': None, 'page': 1,
                'pageSize': 90,'addtime': None, 'question_time': None, 'trace_time': None,'create_time': None, 'finishtime': None, 'sale_id': None, 'product_id': None,
                'logistics_id': None, 'area_id': None, 'currency_id': None,'order_status': None, 'logistics_status': None}
        data_woks = None
        data_woks2 = None
        if order_time == '处理时间':
            data.update({'trace_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
            data_woks = '派送问题件_处理时间'
            data_woks2 = '处理时间'
        elif order_time == '创建时间':
            data.update({'create_time': timeStart + ',' + timeEnd})
            data_woks = '派送问题件_语音外呼'
            data_woks2 = '创建时间'
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)          # json类型数据转换为dict字典
        if req['data'] != []:
            max_count = req['data']['count']    # 获取 请求订单量
            print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
            print('*' * 50)
            if max_count != 0 and max_count != []:
                df = pd.DataFrame([])               # 创建空的dataframe
                dlist = []                          # 创建空的列表 放每次查询的结果
                in_count = math.ceil(max_count / 90)
                n = 1
                while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                    data = self._getDeliveryList(timeStart, timeEnd, n, order_time, proxy_handle, proxy_id)
                    dlist.append(data)
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                    time.sleep(1)
                dp = df.append(dlist, ignore_index=True)
                dp = dp[['id','order_number',  'waybill_number', 'currency', 'addtime', 'amount','orderStatus', 'logisticsStatus', 'logisticsName','create_time', 'lastQuestionName', 'contactName','userName', 'traceName',  'content', 'traceTime', 'failNum', 'questionAddtime', 'questionTypeName']]
                dp.columns = ['id','订单编号', '运单编号', '币种', '下单时间', '金额','订单状态', '物流状态', '物流渠道','创建时间', '派送问题类型', '联系方式', '最新处理人', '最新处理状态', '最新处理结果', '处理时间', '派送次数', '最新抓取时间', '最新问题类型']
                print('正在写入......')
                dp.to_sql('cache_check', con=self.engine1, index=False, if_exists='replace')
                dp.to_excel('F:\\输出文件\\派送问题件_语音外呼-{0}{1}.xlsx'.format(order_time,rq), sheet_name='查询', index=False, engine='xlsxwriter')
                sql = '''REPLACE INTO {0}(id,订单编号, 运单编号,币种, 下单时间,订单状态,物流状态,物流渠道,创建时间,派送问题类型, 联系方式,最新处理人, 最新处理状态, 最新处理结果,处理时间,派送次数,最新抓取时间,最新问题类型,统计日期, 物流轨迹时间, 物流轨迹,便利店, 商品名, 来源渠道,姓名,电话,地址,金额,付款类型,团队,时间,外呼类型,是否短信,商品简称,记录时间) 
                        SELECT id,订单编号,运单编号,币种, 下单时间,订单状态,物流状态,物流渠道,创建时间,派送问题类型, 联系方式,最新处理人, 最新处理状态, 最新处理结果,IF(处理时间 = '',NULL,处理时间) 处理时间,派送次数,IF(最新抓取时间 = '',NULL,最新抓取时间) 最新抓取时间,最新问题类型,DATE_FORMAT({1},'%Y%m%d') 统计日期, 
                                null 物流轨迹时间, null 物流轨迹, null 便利店, null 商品名, null 来源渠道,null 姓名,null 电话,null 地址,null 金额,null 付款类型,null 团队, '{2}' AS 时间, null 外呼类型,null 是否短信,null 商品简称,NOW() 记录时间 
                        FROM cache_check;'''.format(data_woks, data_woks2, tm_data)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                print('写入成功......')
        else:
            print('没有需要获取的信息！！！')
            return
        print('-' * 50)
        print('-' * 50)
    # 语音外呼 派送问题件           （一.2）
    def _getDeliveryList(self, timeStart, timeEnd, n, order_time, proxy_handle, proxy_id):  # 进入派送问题件界面
        print('+++正在查询第 ' + str(n) + ' 页信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.deliveryQuestion&action=getDeliveryList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/deliveryProblemPackage'}
        data = {'order_number': None, 'waybill_number': None, 'question_level': None, 'question_type': None,'order_trace_id': None, 'ship_phone': None, 'page': n,
                'pageSize': 90,'addtime': None, 'question_time': None, 'trace_time': None,'create_time': None, 'finishtime': None, 'sale_id': None, 'product_id': None,
                'logistics_id': None, 'area_id': None, 'currency_id': None,'order_status': None, 'logistics_status': None}
        if order_time == '处理时间':
            data.update({'trace_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59'})
        elif order_time == '创建时间':
            data.update({'create_time': timeStart + ',' + timeEnd})
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(88)
        # print(req)
        ordersDict = []
        try:
            if req['data'] !=[]:
                for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值
                    # print(result['order_number'])
                    # result['traceRecord'] = zhconv.convert(result['traceRecord'], 'zh-hans')
                    ordersDict.append(result.copy())
            else:
                return None
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersDict)
        print('++++++第 ' + str(n) + ' 批次查询成功+++++++')
        print('*' * 50)
        return data

    # 语音外呼 派送的物流轨迹信息（二.1）
    def _getDeliveryList_info(self, proxy_handle, proxy_id, tm_data):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        sql = '''SELECT id FROM {0} s WHERE s.`时间` = '{1}';'''.format('派送问题件_语音外呼', tm_data)
        # sql = '''SELECT id FROM {0} s;'''.format('派送问题件_语音外呼', tm_data)
        ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
        if ordersDict.empty:
            print('无需要更新订单信息！！！')
            return
        orderId = list(ordersDict['id'])
        max_count = len(orderId)                 # 使用len()获取列表的长度，上节学的
        print('++++++本次需查询;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        print(orderId)

        df = pd.DataFrame([])
        dict = []
        for ord in orderId:
            ord = str(ord)
            res = {}
            res['id'] = ord
            res['物流轨迹'] = ''
            res['便利店'] = ''
            res['外呼类型'] = ''
            ship_info, ship_time, ship_name, ship_type, ship_message = self._getDeliveryList_info2(ord, proxy_handle, proxy_id)
            res['物流轨迹'] = ship_info
            res['物流轨迹时间'] = ship_time
            res['便利店'] = ship_name
            res['外呼类型'] = ship_type
            res['是否短信'] = ship_message
            dict.append(res)
        data = pd.json_normalize(dict)
        dp = df.append(data, ignore_index=True)

        dp.to_sql('ps', con=self.engine1, index=False, if_exists='replace')
        sql = '''update 派送问题件_语音外呼 a, ps b
                        set a.`物流轨迹`= IF(b.`物流轨迹` = '', NULL, b.`物流轨迹`),
                            a.`物流轨迹时间`= IF(b.`物流轨迹时间` = '', NULL, b.`物流轨迹时间`),
                            a.`便利店`= IF(b.`便利店` = '', NULL, b.`便利店`),
                            a.`外呼类型`= IF(b.`外呼类型` = '', NULL, b.`外呼类型`),
                            a.`是否短信`= IF(b.`是否短信` = '', NULL, b.`是否短信`)
                where a.`id`= b.`id`;'''.format('team')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('查询已更新+++')
        print('*' * 50)
    # 语音外呼 派送的物流轨迹信息（二.2）
    def _getDeliveryList_info2(self, ord, proxy_handle, proxy_id):  # 进入派送问题件界面
        print('+++正在查询 派送轨迹 信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.deliveryQuestion&action=getDeliveryQuestionRecord'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https://gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/deliveryProblemPackage'}
        data = {'id': ord}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(88)
        # print(req)
        ship_info = ''
        ship_time = ''
        ship_name = ''
        ship_type = ''
        ship_message = ''
        step = 0
        try:
            if req['data'] !=[]:
                for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值
                    print(result)
                    print(ord)
                    if step >= 1:
                        continue
                    if '送至' in result['track_info']:
                        vt = result['track_info'].split('送至')[1]
                        vt2 = vt.split('(')[0]
                        if '7-11' in vt2:
                            ship_name = vt2.replace('7-11', 'Seven Eleven')
                        else:
                            ship_name = vt2
                        ship_info = result['track_info']
                        ship_time = result['addtime']
                        ship_type = '超商'
                        ship_message = '否'
                        step = step + 1
                    elif '與客戶另約時間到站自領' in result['track_info']:
                        vt = result['track_info'].split('到站自領，')[1]
                        ship_name = vt.split('保管中')[0]
                        ship_info = result['track_info']
                        ship_time = result['addtime']
                        ship_type = '到站自取'
                        ship_message = '否'
                    elif '與客戶另約時間配送' in result['track_info']:
                        ship_name = '与客户另约时间配送'
                        ship_info = result['track_info']
                        ship_time = result['addtime']
                        ship_type = '预约时间配送'
                        ship_message = '否'
            else:
                return ship_info, ship_time, ship_name, ship_type, ship_message
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        print('*' * 50)
        return ship_info, ship_time, ship_name, ship_type, ship_message

    # 语音外呼 派送的物流轨迹 订单信息（三.1）
    def _getDeliveryList_order(self, proxy_handle, proxy_id, tm_data):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        sql = '''SELECT 订单编号 FROM {0} s WHERE s.`时间` = '{1}';'''.format('派送问题件_语音外呼', tm_data)
        # sql = '''SELECT 订单编号 FROM {0} s;'''.format('派送问题件_语音外呼', tm_data)
        ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
        if ordersDict.empty:
            print('无需要更新订单信息！！！')
            return
        orderId = list(ordersDict['订单编号'])
        max_count = len(orderId)                 # 使用len()获取列表的长度，上节学的
        print('++++++本次需查询;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量

        in_count = math.ceil(max_count / 500)
        print(in_count)
        df = pd.DataFrame([])
        dlist = []
        n = 1
        while n <= in_count:  # 这里用到了一个while循环，穿越过来的
            n1 = 500 * (n-1)
            n2 = 500 * n
            ord = ','.join(orderId[n1:n2])
            data = self._getDeliveryList_order2(ord, '订单号', proxy_handle, proxy_id)
            dlist.append(data)
            print('剩余查询数' + str(in_count - n))
            n = n + 1
        dp = df.append(dlist, ignore_index=True)
        dp = dp[['orderNumber', 'wayBillNumber','currency', 'area', 'befrom', 'amount','tel_phone', 'shipInfo.shipName', 'shipInfo.shipAddress', 'saleName', 'orderStatus', 'logisticsStatus', 'payType','abbreviation']]
        print(dp)
        dp.to_sql('ps', con=self.engine1, index=False, if_exists='replace')
        sql = '''update 派送问题件_语音外呼 a, ps b
                        set a.`商品名`= IF(b.`saleName` = '', NULL, b.`saleName`),
                            a.`来源渠道`= IF(b.`befrom` = '', NULL, b.`befrom`),
                            a.`姓名`= IF(b.`shipInfo.shipName` = '', NULL, b.`shipInfo.shipName`),
                            a.`电话`= IF(b.`tel_phone` = '', NULL, b.`tel_phone`),
                            a.`地址`= IF(b.`shipInfo.shipAddress` = '', NULL, b.`shipInfo.shipAddress`),
                            a.`金额`= IF(b.`amount` = '', NULL, b.`amount`),
                            a.`付款类型`= IF(b.`payType` = '', NULL, b.`payType`),
                            a.`订单状态`= IF(b.`orderStatus` = '', NULL, b.`orderStatus`),
                            a.`物流状态`= IF(b.`logisticsStatus` = '', NULL, b.`logisticsStatus`),
                            a.`团队`= IF(b.`area` = '', NULL, b.`area`),
                            a.`运单编号`= IF(b.`wayBillNumber` = '', NULL, b.`wayBillNumber`),
                            a.`商品简称`= IF(b.`abbreviation` = '', NULL, b.`abbreviation`)
                where a.`订单编号`= b.`orderNumber`;'''.format('team')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('查询已更新+++')
        print('*' * 50)
        # 语音外呼 派送的物流轨迹信息（三.2）
    # 语音外呼 派送的物流轨迹 订单信息（三.2）
    def _getDeliveryList_order2(self, ord, searchType, proxy_id, proxy_handle):  # 进入订单检索界面
        print('+++正在查询 订单信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500,
                'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None,
                'type': None, 'collId': None, 'isClone': None,
                'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '',
                'warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None,
                'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None,
                'service_id': None, 'autoVerifyStatus': None, 'shipZip': None, 'remark': None, 'shipState': None, 'weightStart': None,
                'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None}
        if searchType == '订单号':
            data.update({'orderPrefix': ord, 'shippingNumber': None})
        elif searchType == '运单号':
            data.update({'order_number': None, 'shippingNumber': ord})
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        # print('正在处理json数据转化为dataframe…………')
        try:
            for result in req['data']['list']:
                result['saleId'] = 0        # 添加新的字典键-值对，为下面的重新赋值用
                result['saleName'] = 0
                result['productId'] = 0
                result['saleProduct'] = 0
                result['spec'] = 0
                result['chooser'] = 0
                result['saleId'] = result['specs'][0]['saleId']
                result['saleName'] = result['specs'][0]['saleName']
                result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
                result['saleProduct'] = (result['specs'][0]['saleProduct']).split('#')[2]
                result['spec'] = result['specs'][0]['spec']
                result['chooser'] = result['specs'][0]['chooser']
                quest = ''
                for re in result['questionReason']:
                    quest = quest + ';' + re
                result['questionReason'] = quest
                delr = ''
                for re in result['delReason']:
                    delr = delr + ';' + re
                result['delReason'] = delr
                auto = ''
                for re in result['autoVerify']:
                    auto = auto + ';' + re
                result['autoVerify'] = auto

                result['auto_VerifyTip'] = ''
                result['auto_VerifyTip_zl'] = ''
                result['auto_VerifyTip_qs'] = ''
                result['auto_VerifyTip_js'] = ''
                if result['autoVerifyTip'] == "":
                    result['auto_VerifyTip'] = '0.00%'
                else:
                    if '未读到拉黑表记录' in result['autoVerifyTip']:
                        result['auto_VerifyTip'] = '0.00%'
                    else:
                        t3 = result['autoVerifyTip']
                        result['auto_VerifyTip_zl'] = (t3.split('订单配送总量：')[1]).split(',')[0]
                        result['auto_VerifyTip_qs'] = (t3.split('送达订单量：')[1]).split(',')[0]
                        result['auto_VerifyTip_js'] = (t3.split('拒收订单量：')[1]).split(',')[0]
                        if '拉黑率问题' not in result['autoVerifyTip']:
                            t2 = result['autoVerifyTip'].split(',拉黑率')[1]
                            result['auto_VerifyTip'] = t2.split('%;')[0] + '%'
                        else:
                            t2 = result['autoVerifyTip'].split('拒收订单量：')[1]
                            t2 = t2.split('%;')[0]
                            result['auto_VerifyTip'] = t2.split('拉黑率')[1] + '%'
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        df = pd.json_normalize(ordersdict)
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return df
    # 查询导出 （四）
    def _getDeliveryList_out(self, tm_data):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在获取 超商 语音外呼…………')
        sql = '''SELECT 姓名 AS 客户名,	NULL 性别,	电话 AS 'contact number', NULL 客户分组, NULL 微信, 来源渠道 AS 购买渠道, substring_index(商品名,'#',-1) AS 商品名, 便利店 AS 超商名称, IF(付款类型 NOT LIKE '%货到付款%', 0, 金额) as 金额, 
                       NULL 副号码1, NULL 副号码2, NULL 副号码3, NULL 副号码4, NULL 副号码5, NULL 副号码6, NULL 副号码7, NULL 副号码8, NULL 副号码9, 订单编号, 团队,商品简称
                FROM 派送问题件_语音外呼 s 
                WHERE s.`时间` = '{1}' AND s.币种 = '台币' and s.订单状态 not in ('已完成','已退货(销售)','已退货(物流)','已退货(不拆包物流)')
                  AND s.外呼类型 = '超商';'''.format('派送问题件_语音外呼', tm_data)
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)

        print('正在获取 预约时间配送 语音外呼…………')
        sql = '''SELECT 姓名 AS 客户名,	NULL 性别,	电话 AS 'contact number', NULL 客户分组, NULL 微信, 来源渠道 AS 购买渠道, substring_index(商品名,'#',-1) AS 商品名, 地址, IF(付款类型 NOT LIKE '%货到付款%', 0, 金额) as 金额, 
                       NULL 副号码1, NULL 副号码2, NULL 副号码3, NULL 副号码4, NULL 副号码5, NULL 副号码6, NULL 副号码7, NULL 副号码8, NULL 副号码9, 订单编号, 团队,商品简称
                FROM 派送问题件_语音外呼 s 
                WHERE s.`时间` = '{1}' AND s.币种 = '台币' and s.订单状态 not in ('已完成','已退货(销售)','已退货(物流)','已退货(不拆包物流)')
                  AND s.外呼类型 = '预约时间配送';'''.format('派送问题件_语音外呼', tm_data)
        df2 = pd.read_sql_query(sql=sql, con=self.engine1)

        print('正在获取 到站自取 语音外呼…………')
        sql = '''SELECT 姓名 AS 客户名,	NULL 性别,	电话 AS 'contact number', NULL 客户分组, NULL 微信, 便利店 AS 营业所, 来源渠道 AS 购买渠道, 运单编号 AS 快递单号, substring_index(商品名,'#',-1) AS 商品名, IF(付款类型 NOT LIKE '%货到付款%', 0, 金额) as 金额, 
                       NULL 副号码1, NULL 副号码2, NULL 副号码3, NULL 副号码4, NULL 副号码5, NULL 副号码6, NULL 副号码7, NULL 副号码8, NULL 副号码9, 订单编号, 团队,商品简称
                FROM 派送问题件_语音外呼 s 
                WHERE s.`时间` = '{1}' AND s.币种 = '台币' and s.订单状态 not in ('已完成','已退货(销售)','已退货(物流)','已退货(不拆包物流)')
                  AND s.外呼类型 = '到站自取';'''.format('派送问题件_语音外呼', tm_data)
        df3 = pd.read_sql_query(sql=sql, con=self.engine1)

        df1.to_excel('F:\\输出文件\\送至便利店-语音外呼{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')  # Xlsx是python用来构造xlsx文件的模块，可以向excel2007+中写text，numbers，formulas 公式以及hyperlinks超链接。
        df2.to_excel('F:\\输出文件\\与客户另约时间配送-语音外呼{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')  # Xlsx是python用来构造xlsx文件的模块，可以向excel2007+中写text，numbers，formulas 公式以及hyperlinks超链接。
        df3.to_excel('F:\\输出文件\\到站自取-语音外呼{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')

        filePath = r'F:\\输出文件\\语音外呼-派送问题件{}.xlsx'.format(rq)
        # df0 = pd.DataFrame([])                                  # 创建空的dataframe数据框
        # df0.to_excel(filePath, index=False)                     # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        print('正在写入excel…………')
        # writer = pd.ExcelWriter(filePath, engine='openpyxl')    # 初始化写入对象
        # book = load_workbook(filePath)                          # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book                                      # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # df1.to_excel(excel_writer=writer, sheet_name='超商', index=False)
        # df2.to_excel(excel_writer=writer, sheet_name='预约时间配送', index=False)
        # df3.to_excel(excel_writer=writer, sheet_name='到站自取', index=False)
        # if 'Sheet1' in book.sheetnames:                         # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()
        with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
            df1.to_excel(excel_writer=writer, sheet_name='超商', index=False)
            df2.to_excel(excel_writer=writer, sheet_name='预约时间配送', index=False)
            df3.to_excel(excel_writer=writer, sheet_name='到站自取', index=False)
        print('查询已导出+++')
        print('*' * 50)
        # 语音外呼 派送的物流轨迹信息（三.2）


if __name__ == '__main__':
    # TODO------------------------------------单点更新配置------------------------------------
    handle = '手动0'
    login_TmpCode = '0b04de569eb6395e88a34a2e9cde8e92'  # 输入登录口令Tkoen
    proxy_handle = '代理服务器0'
    proxy_id = '192.168.13.89:37466'  # 输入代理服务器节点和端口

    # TODO------------------------------------单点更新读取------------------------------------
    m = QueryTwo('+86-18538110674', 'qyz04163510.', login_TmpCode, handle, proxy_handle, proxy_id)
    start: datetime = datetime.datetime.now()
    match1 = {'gat': '港台', 'gat_order_list': '港台', 'slsc': '品牌'}
    '''
    # -----------------------------------------------手动导入状态运行（一）-----------------------------------------
    # 1、 正在按订单查询；2、正在按时间查询；--->>数据更新切换
    # isReal: 0 查询后台保存的运单轨迹； 1 查询物流的实时运单轨迹 ；  cat = 1 、黑猫切换是否使用后台数据  0 、还是官网数据 
    '''
    select = 1
    isReal = 1
    cat = 0
    if int(select) == 1:
        print("1-->>> 正在按运单号查询+++")
        m.readFormHost(isReal, proxy_handle, proxy_id, cat)       # 导入；，更新--->>数据更新切换

    elif int(select) == 3:
        print("1-->>> 正在按运单号查询+++")
        m.Search_online_write(isReal, proxy_handle, proxy_id)       # 写入数据库中；，可中断重启查询

    elif int(select) == 2:
        print("2-->>> 正在按时间查询+++")
        m.order_online('2022-01-01', '2022-01-05', isReal, proxy_handle, proxy_id)

    elif int(select) in (51,52,53,54):
        print("1-->>> 正在按运单号查询+++")
        tm = ''
        tm2 = ''
        tm_data = ''
        if select == 51:
            tm = (datetime.datetime.now() - relativedelta(days=1)).strftime('%Y-%m-%d') + ' 16:00:00'
            tm2 = (datetime.datetime.now().strftime('%Y-%m-%d')) + ' 11:00:00'
            tm_data = (datetime.datetime.now().strftime('%Y-%m-%d')) + 'AM'
        elif select == 52:
            tm = (datetime.datetime.now().strftime('%Y-%m-%d')) + ' 11:00:00'
            tm2 = (datetime.datetime.now().strftime('%Y-%m-%d')) + ' 16:00:00'
            tm_data = (datetime.datetime.now().strftime('%Y-%m-%d')) + 'PM'
        elif select == 53:
            tm = '2023-03-27 10:00:00'
            tm2 = '2023-03-27 11:00:00'
            tm_data = (datetime.datetime.now().strftime('%Y-%m-%d')) + 'AM'
        elif select == 54:
            tm = '2023-04-17 00:00:00'
            tm2 = '2023-04-17 16:00:00'
            tm_data = (datetime.datetime.now().strftime('%Y-%m-%d')) + 'PM'
        print('****** 查询      起止时间：' + tm + ' - ' + tm2 + ' ******')
        message = '带短信'
        m.getDeliveryList(tm, tm2, '创建时间', proxy_handle, proxy_id, tm_data, message)
        m._getDeliveryList_info(proxy_handle, proxy_id, tm_data)
        m._getDeliveryList_order(proxy_handle, proxy_id, tm_data)
        m._getDeliveryList_out(tm_data)


    # m.order_bind_status('2022-01-01', '2022-01-02')

    # m._order_bind_status('7449201841')

    print('查询耗时：', datetime.datetime.now() - start)