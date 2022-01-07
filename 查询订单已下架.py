import pandas as pd
import os
import datetime
import time
import xlwings
import xlsxwriter
import math
import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel

from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色
from bs4 import BeautifulSoup # 抓标签里面元素的方法

# -*- coding:utf-8 -*-
class QueryTwoLower(Settings, Settings_sso):
    def __init__(self, userMobile, password):
        Settings.__init__(self)
        Settings_sso.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        self.sso_online_cang()
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
    #  登录后台中
    def _online(self):  # 登录系统保持会话状态
        print('正在登录后台系统中......')
        # print('第一阶段获取-钉钉用户信息......')
        url = r'https://login.dingtalk.com/login/login_with_pwd'
        data = {'mobile': self.userMobile,
                'pwd': self.password,
                'goto': 'https://oapi.dingtalk.com/connect/oauth2/sns_authorize?appid=dingoajqpi5bp2kfhekcqm&response_type=code&scope=snsapi_login&state=STATE&redirect_uri=http://gsso.giikin.com/admin/dingtalk_service/getunionidbytempcode',
                'pdmToken': '',
                'araAppkey': '1917',
                'araToken': '0#19171628645731266586976965831628645747396525G1E2B0816DEBF96BC4199761B6A1F3C0FCD91FB',
                'araScene': 'login',
                'captchaImgCode': '',
                'captchaSessionId': '',
                'type': 'h5'}
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Origin': 'https://login.dingtalk.com',
                    'Referer': 'https://login.dingtalk.com/'}
        req = self.session.post(url=url, headers=r_header, data=data, allow_redirects=False)
        req = req.json()
        # print(req)
        req_url = req['data']
        loginTmpCode = req_url.split('loginTmpCode=')[1]        # 获取loginTmpCode值
        # print(loginTmpCode)
        # print('+++已获取loginTmpCode值+++')

        time.sleep(1)
        # print('第二阶段请求-登录页面......')
        url = r'http://gsso.giikin.com/admin/dingtalk_service/gettempcodebylogin.html'
        data = {'tmpCode': loginTmpCode,
                'system': 1,
                'url': '',
                'ticker': '',
                'companyId': 1}
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Origin': 'https://login.dingtalk.com',
                    'Referer': 'http://gsso.giikin.com/admin/login/logout.html'}
        req = self.session.post(url=url, headers=r_header, data=data, allow_redirects=False)
        # print(req.text)
        # print('+++请求登录页面url成功+++')

        time.sleep(1)
        # print('第三阶段请求-dingtalk服务器......')
        # print('（一）加载dingtalk_service跳转页面......')
        url = req.text
        data = {'tmpCode': loginTmpCode,
                'system': 1,
                'url': '',
                'ticker': '',
                'companyId': 1}
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': 'http://gsso.giikin.com/'}
        req = self.session.get(url=url, headers=r_header, data=data, allow_redirects=False)
        # print(req.headers)
        gimp = req.headers['Location']
        # print('+++已获取跳转页面+++')
        time.sleep(1)
        # print('（二）请求dingtalk_service的cookie值......')
        url = gimp
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': 'http://gsso.giikin.com/'}
        req = self.session.get(url=url, headers=r_header, allow_redirects=False)
        # print(req)
        # print('+++已获取cookie值+++')

        time.sleep(1)
        # print('第四阶段页面-重定向跳转中......')
        # print('（一）加载chooselogin.html页面......')
        url = r'http://gsso.giikin.com/admin/login_by_dingtalk/chooselogin.html'
        data = {'user_id': 1343}
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': gimp,
                    'Origin': 'http://gsso.giikin.com'}
        req = self.session.post(url=url, headers=r_header, data=data, allow_redirects=False)
        # print(req.headers)
        index = req.headers['Location']
        # print('+++已获取gimp.giikin.com页面')
        time.sleep(1)
        # print('（二）加载gimp.giikin.com页面......')
        url = index
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': index}
        req = self.session.get(url=url, headers=r_header, allow_redirects=False)
        # print(req.headers)
        index2 = req.headers['Location']
        # print('+++已获取index.html页面')

        time.sleep(1)
        # print('（三）加载index.html页面......')
        url = 'http://gimp.giikin.com/' + index2
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': 'http://gsso.giikin.com/'}
        req = self.session.get(url=url, headers=r_header, allow_redirects=False)
        # print(req.headers)
        index_system = req.headers['Location']
        # print('+++已获取index.html?_system=18正式页面')

        time.sleep(1)
        # print('第五阶段正式页面-重定向跳转中......')
        # print('（一）加载index.html?_system页面......')
        url = index_system
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': 'http://gsso.giikin.com/'}
        req = self.session.get(url=url, headers=r_header, allow_redirects=False)
        # print(req.headers)
        index_system2 = req.headers['Location']
        # print('+++已获取index.html?_ticker=页面......')
        time.sleep(1)
        # print('（二）加载index.html?_ticker=页面......')
        url = index_system2
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': 'http://gsso.giikin.com/'}
        req = self.session.get(url=url, headers=r_header, allow_redirects=False)
        # print(req)
        # print(req.headers)
        print('++++++已成功登录++++++')

    # 查询压单（仓储的获取）
    def orderInfoQuery(self, searchType):  # 进入订单检索界面
        print('+++正在查询订单信息中')
        url = r'http://gwms.giikin.cn/order/pressure/index'
        url = r'http://gwms-v3.giikin.cn/order/pressure/index'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'ttp://gwms-v3.giikin.cn/order/pressure/index'}
        data = {'selectStr': '1=1',
                'page': '1',
                'limit': 500,
                'startDate': '',
                'endDate': ''}
        if searchType == '台币':
            data.update({'selectStr': '1=1 and oc.currency_id= "1"'})
        elif searchType == '港币':
            data.update({'selectStr': '1=1 and oc.currency_id= "2"'})
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        # print('正在处理json数据转化为dataframe…………')
        try:
            for result in req['data']:
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            # time.sleep(10)
            # self.orderInfoQuery(ord, searchType)
        #     self.q.put(result)
        # for i in range(len(req['data']['list'])):
        #     ordersdict.append(self.q.get())
        data = pd.json_normalize(ordersdict)
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return data


    def order_lower(self, timeStart, timeEnd, auto_time):  # 进入已下架界面
        start: datetime = datetime.datetime.now()
        team_whid = ['龟山易速配', '速派八股仓', '天马新竹仓', '立邦香港顺丰', '香港易速配', '龟山-神龙备货', '龟山-火凤凰备货', '天马顺丰仓']
        # team_whid = ['天马顺丰仓']
        team_stock_type = [1, 2]
        # team_stock_type = [2]
        match = {1: 'SKU库存',
                 2: '组合库存',
                 3: '混合库存'}
        match2 = {'龟山易速配': 70,
                 '速派八股仓': 95,
                 '天马新竹仓': 102,
                 '立邦香港顺丰': 117,
                 '香港易速配': 134,
                 '龟山-神龙备货': 166,
                 '龟山-火凤凰备货': 198,
                 '天马顺丰仓': 204}
        if auto_time == '自动':
            sql = '''SELECT DISTINCT 统计时间 FROM 已下架表 d GROUP BY 统计时间 ORDER BY 统计时间 DESC'''
            rq = pd.read_sql_query(sql=sql, con=self.engine1)
            rq = pd.to_datetime(rq['统计时间'][0])

            begin = (rq + datetime.timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
            begin = datetime.datetime.strptime(begin, '%Y-%m-%d %H:%M:%S')
            end = (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M:%S')
            end = datetime.datetime.strptime(end, '%Y-%m-%d %H:%M:%S')
            print('****** 总起止时间：' + begin.strftime('%Y-%m-%d') + ' - ' + end.strftime('%Y-%m-%d') + ' ******')

            for i in range((end - begin).days + 1):  # 按天循环获取订单状态
                day = begin + datetime.timedelta(days=i)
                timeStart = (day - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
                timeEnd = day.strftime('%Y-%m-%d')
                print('正在查询日期---起止时间：' + timeStart + ' - ' + timeEnd)
                for tem in team_whid:
                    if tem in ('龟山易速配', '龟山-神龙备货', '龟山-火凤凰备货'):
                        for tem_type in team_stock_type:
                            print('+++正在查询仓库： ' + tem + '；库存类型:' + match[tem_type] + ' 信息')
                            self._order_lower_info(match2[tem], tem_type, timeStart, timeEnd, tem, match[tem_type])
                    else:
                        print('+++正在查询仓库： ' + tem + '；库存类型:组合库存 信息')
                        self._order_lower_info(match2[tem], 2, timeStart, timeEnd, tem, '组合库存')
        else:
            print('正在查询日期---起止时间：' + timeStart + ' - ' + timeEnd)
            for tem in team_whid:
                if tem in ('龟山易速配', '龟山-神龙备货', '龟山-火凤凰备货'):
                    for tem_type in team_stock_type:
                        print('+++正在查询仓库： ' + tem + '；库存类型:' + match[tem_type] + ' 信息')
                        self._order_lower_info(match2[tem], tem_type, timeStart, timeEnd, tem, match[tem_type])
                else:
                    print('+++正在查询仓库： ' + tem + '；库存类型:组合库存 信息')
                    self._order_lower_info(match2[tem], 2, timeStart, timeEnd, tem, '组合库存')
        print('查询耗时：', datetime.datetime.now() - start)

    def _order_lower_info(self, tem, tem_type, timeStart, timeEnd, tem_name, type_name):  # 进入已下架界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        # print('+++正在查询信息中')
        url = r'http://gwms-v3.giikin.cn/order/order/shelves'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': 1, 'limit': 500, 'startDate': timeStart + ' 08:30:00', 'endDate':  timeEnd + ' 08:30:00', 'selectStr': '1=1 and ob.whid = ' + str(tem) + ' and ob.stock_type = ' + str(tem_type)}
        proxy = '47.75.114.218:10020'  # 使用代理服务器
        # proxies = {'http': 'socks5://' + proxy, 'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print(req.text)
        # print(json.loads(f'"{req.text}"'))
        # req = req.text.encode('utf-8').decode("unicode_escape")
        # print('+++已成功发送请求......')              # 转码使用
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['data']
        if max_count != []:
            ordersDict = []
            try:
                for result in req['data']:              # 添加新的字典键-值对，为下面的重新赋值
                    result['count_time'] = timeEnd
                    if type_name == 'SKU库存':
                        result['waill_name'] = '龟山备货'
                    else:
                        if '龟山易速配' in result['whid']:
                            result['waill_name'] = '龟山'
                        elif '速派八股仓' in result['whid']:
                            result['waill_name'] = '速派'
                        elif '天马新竹仓' in result['whid']:
                            result['waill_name'] = '天马新竹'
                        elif '立邦香港顺丰' in result['whid']:
                            result['waill_name'] = '立邦'
                        elif '香港易速配' in result['whid']:
                            result['waill_name'] = '易速配'
                        elif '神龙备货' in result['whid']:
                            result['waill_name'] = '龟山备货'
                        elif '火凤凰备货' in result['whid']:
                            result['waill_name'] = '龟山备货'
                        elif '天马顺丰仓' in result['whid']:
                            result['waill_name'] = '天马顺丰'
                    # print(result)
                    ordersDict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersDict)
            data = data[['order_number', 'addtime', 'billno', 'old_billno', 'goods_id', 'product_name', 'intime', 'whid', 'waill_name', 'currency_id', 'count_time']]
            data.columns = ['订单编号', '下单时间', '新运单号', '原运单号', '产品id', '商品名称', '下架时间', '仓库', '物流渠道', '币种', '统计时间']
            print(data)
            print('>>>' + tem_name + '-' + type_name + ' <<< 查询完结！！！')
            data.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 已下架表(订单编号,下单时间,新运单号,原运单号, 产品id, 商品名称, 下架时间, 仓库, 物流渠道,币种,统计时间,记录时间)
                    SELECT 订单编号,下单时间,新运单号,原运单号, 产品id, 商品名称, 下架时间, 仓库, 物流渠道,币种, 统计时间,NOW() 记录时间
                    FROM customer'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
            data.to_excel('G:\\输出文件\\已下架 {0} {1}-{2}.xlsx'.format(tem_name, type_name, rq), sheet_name='查询', index=False, engine='xlsxwriter')
        else:
            print('****** 没有新增的改派订单！！！')
            return None
        print('*' * 50)

    # 改派-查询未发货的订单
    def gp_order(self):
        print('正在查询改派未发货订单…………')
        listT = []  # 查询sql的结果 存放池
        sql = '''SELECT xj.*, '未发货' AS 状态
                FROM 已下架表  xj
                LEFT JOIN gat_zqsb gz ON xj.订单编号= gz.订单编号
			    LEFT JOIN gat_order_list gs ON xj.订单编号= gs.订单编号
                WHERE xj.下单时间 >= TIMESTAMP(DATE_ADD(curdate()-day(curdate())+1,interval -2 month)) 
					AND xj.币种 = '台币' AND (最终状态 = '未发货' or 最终状态 IS NULL)  
					AND  gs.系统订单状态 NOT IN ('已删除', '问题订单', '待发货', '截单') or gs.系统订单状态 IS NULL;'''
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df)
        print('正在写入excel…………')
        today = datetime.date.today().strftime('%m.%d')
        file_path = 'G:\\输出文件\\{} 改派未发货.xlsx'.format(today)
        if os.path.exists(file_path):  # 判断是否有需要的表格
            print("正在清除重复文件......")
            os.remove(file_path)
        sheet_name = ['台湾']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('----已写入excel ')

if __name__ == '__main__':
    m = QueryTwoLower('+86-18538110674', 'qyz04163510')
    start: datetime = datetime.datetime.now()
    match1 = {'gat': '港台', 'gat_order_list': '港台', 'slsc': '品牌'}
    # -----------------------------------------------手动设置时间；若无法查询，切换代理和直连的网络-----------------------------------------
    m.order_lower('2021-12-31', '2022-01-01', '自动')


    print('查询耗时：', datetime.datetime.now() - start)