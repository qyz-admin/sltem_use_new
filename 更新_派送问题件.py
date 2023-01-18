import pandas as pd
import os
import datetime
import time
from tqdm import tqdm

import xlwings
import xlwings as xl
import xlsxwriter
import math
import requests
import json
import re
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel
import zhconv          # transform2_zh_hant：转为繁体;transform2_zh_hans：转为简体
import win32com.client as win32
import win32com.client

from mysqlControl import MysqlControl
from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色
from 更新_已下架_压单_驳回发货_头程导入提单号 import QueryTwoLower
from 查询_订单检索 import QueryOrder

# -*- coding:utf-8 -*-
class QueryTwo(Settings, Settings_sso):
    def __init__(self, userMobile, password, login_TmpCode,handle, select, proxy_handle, proxy_id):
        Settings.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        # self._online()
        # self.sso_online_Two()
        # self.sso__online_handle(login_TmpCode)
        # # self.sso__online_auto()
        if select == 99:
            if proxy_handle == '代理服务器':
                if handle == '手动':
                    self.sso__online_handle_proxy(login_TmpCode, proxy_id)
                else:
                    self.sso__online_auto_proxy(proxy_id)
            else:
                if handle == '手动':
                    self.sso__online_handle(login_TmpCode)
                else:
                    self.sso__online_auto()

        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
        # self.my = MysqlControl()

    # 获取查询时间
    def readInfo(self, team):
        print('>>>>>>正式查询中<<<<<<')
        print('正在获取需要订单信息......')
        start = datetime.datetime.now()
        if team == '派送问题件_导出':
            last_time = datetime.datetime.now().strftime('%Y-%m') + '-01'
            now_time = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')

        elif team == '派送问题件_更新':
            sql = 'SELECT DISTINCT 日期 FROM 派送问题件_跟进表2_cp d  ORDER BY 日期 DESC'
            rq = pd.read_sql_query(sql=sql, con=self.engine1)
            rq = pd.to_datetime(rq['日期'][0])
            last_time = (rq - datetime.timedelta(days=15)).strftime('%Y-%m-%d')
            now_time = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')

        elif team == '派送问题件_订单完成单量&短信发送单量':
            sql = 'SELECT DISTINCT 日期 FROM 派送问题件_跟进表2_cp d  ORDER BY 日期 DESC'
            rq = pd.read_sql_query(sql=sql, con=self.engine1)
            rq = pd.to_datetime(rq['日期'][0])
            last_time = rq.strftime('%Y-%m-%d')
            now_time = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')

        elif team == '物流问题件_更新':
            sql = 'SELECT DISTINCT 日期 FROM 派送问题件_跟进表2_cp d  ORDER BY 日期 DESC'
            rq = pd.read_sql_query(sql=sql, con=self.engine1)
            rq = pd.to_datetime(rq['日期'][0])
            last_time = (rq - datetime.timedelta(days=15)).strftime('%Y-%m-%d')
            now_time = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')

        elif team == '拒收问题件_更新':
            sql = 'SELECT DISTINCT 日期 FROM 派送问题件_跟进表2_cp d  ORDER BY 日期 DESC'
            rq = pd.read_sql_query(sql=sql, con=self.engine1)
            rq = pd.to_datetime(rq['日期'][0])
            last_time = (rq - datetime.timedelta(days=15)).strftime('%Y-%m-%d')
            now_time = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')

        else:
            sql = '''SELECT DISTINCT 处理时间 FROM {0} d GROUP BY 处理时间 ORDER BY 处理时间 DESC'''.format(team)
            rq = pd.read_sql_query(sql=sql, con=self.engine1)
            rq = pd.to_datetime(rq['处理时间'][0])
            last_time = (rq + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            now_time = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        print('******************起止时间：' + team + last_time + ' - ' + now_time + ' ******************')
        return last_time, now_time

    def outport_getDeliveryList(self, timeStart, timeEnd, logisticsN_begin, logisticsN_end):
        rq = datetime.datetime.now().strftime('%m.%d')
        # self.getOrderList(timeStart, timeEnd)
        # self.getDeliveryList(timeStart, timeEnd)
        month = (datetime.datetime.now()).strftime('%Y%m')
        # print(month)
        # print(type(month))
        time_Start = (datetime.datetime.now() - relativedelta(months=12)).strftime('%Y-%m') + '-01'        # 派送问题件签收率
        if (datetime.datetime.now()).strftime('%d') == 1:
            timeStart = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y-%m') + '-01'
            timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            timeStart = (datetime.datetime.now() - relativedelta(months=2)).strftime('%Y-%m') + '-01'
            timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')

        print('派送问题件 各类型签收率，导出时间》》》 ' + time_Start + "---" + timeEnd)
        sql8 = '''SELECT s2.派送类型, s2.月份, s2.总订单,
                        concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        -- concat(ROUND(IFNULL(s2.签收退货 / s2.已完成,0) * 100,2),'%') as 完成签收退货,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.总订单 / ss2.单量,0) * 100,2),'%') as 订单占比,NULL 处理方式
                FROM( SELECT s1.月份, s1.派送类型, COUNT(ss3.订单编号) AS 总订单,
                                    SUM(IF(ss3.系统物流状态 = "已签收" OR ss3.系统物流状态 = "已退货",1,0)) as 签收退货,
                                    SUM(IF(ss3.系统物流状态 = "已签收",1,0)) as 签收,
                                    SUM(IF(ss3.系统物流状态 = "拒收",1,0)) as 拒收,
                                    SUM(IF(ss3.系统物流状态 = "已退货",1,0)) as 已退货,
                                    SUM(IF(ss3.系统物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成
                    FROM (  SELECT *,EXTRACT(YEAR_MONTH FROM 创建日期) AS 月份, IF(派送问题 = '送至便利店' OR 派送问题 = '客户自取','送至便利店',IF(派送问题 = '客户长期不在' OR 派送问题 = '送达客户不在','送达客户不在',派送问题)) AS 派送类型
                            FROM 派送问题件_跟进表 g
                            WHERE g.`创建日期` >= '{0}'  AND g.`创建日期` <= '{1}' AND g.币种 ='台币'
                    ) s1
                    LEFT JOIN (SELECT * FROM gat_order_list) ss3 ON s1.订单编号 = ss3.订单编号
                    GROUP BY s1.月份, s1.派送类型
                ) s2
                LEFT JOIN (  SELECT 年月 AS 月份,  COUNT(订单编号) AS 单量
                            FROM gat_zqsb g
                            WHERE g.`日期` >= '{0}'  AND g.`日期` <= '{1}'  AND g.币种 ='台湾'
                            GROUP BY 月份
                ) ss2  ON s2.月份 = ss2.月份
                ORDER BY 
                FIELD(派送类型,'送至便利店','客户要求更改派送时间或者地址','客户不接电话','地址问题','送达客户不在','拒收','合计'),
				月份, 总订单;'''.format(time_Start, timeEnd)
        df8 = pd.read_sql_query(sql=sql8, con=self.engine1)

        sql81 = '''SELECT s2.派送类型, s2.月份, s2.总订单,
                                concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                              --  concat(ROUND(IFNULL(s2.签收退货 / s2.已完成,0) * 100,2),'%') as 完成签收退货,
                                concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                                concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                                concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                                concat(ROUND(IFNULL(s2.总订单 / ss2.单量,0) * 100,2),'%') as 订单占比,NULL 处理方式
                        FROM( SELECT s1.月份, s1.派送类型, COUNT(ss3.订单编号) AS 总订单,
                                            SUM(IF(ss3.系统物流状态 = "已签收" OR ss3.系统物流状态 = "已退货",1,0)) as 签收退货,
                                            SUM(IF(ss3.系统物流状态 = "已签收",1,0)) as 签收,
                                            SUM(IF(ss3.系统物流状态 = "拒收",1,0)) as 拒收,
                                            SUM(IF(ss3.系统物流状态 = "已退货",1,0)) as 已退货,
                                            SUM(IF(ss3.系统物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成
                            FROM (  SELECT *,EXTRACT(YEAR_MONTH FROM 创建日期) AS 月份,  IF(派送问题 = '客户长期不在' OR 派送问题 = '送达客户不在','送达客户不在',派送问题) AS 派送类型
                                    FROM 派送问题件_跟进表 g
                                    WHERE g.`创建日期` >= '{0}'  AND g.`创建日期` <= '{1}' AND g.币种 ='港币'
                            ) s1
                            LEFT JOIN (SELECT * FROM gat_order_list) ss3 ON s1.订单编号 = ss3.订单编号
                            GROUP BY s1.月份, s1.派送类型
                        ) s2
                        LEFT JOIN (  SELECT 年月 AS 月份,  COUNT(订单编号) AS 单量
                                    FROM gat_zqsb g
                                    WHERE g.`日期` >= '{0}'  AND g.`日期` <= '{1}'  AND g.币种 ='香港'
                                    GROUP BY 月份
                        ) ss2  ON s2.月份 = ss2.月份
                        ORDER BY 
                        FIELD(派送类型,'预约送达时间','客户要求更改派送时间或者地址','送达客户不在','客户自取','客户不接电话','地址问题','合计'),
        				月份, 总订单;'''.format(time_Start, timeEnd)
        df81 = pd.read_sql_query(sql=sql81, con=self.engine1)

        print('正在获取excel内容…………')
        sql = '''SELECT *, IF(派送问题 LIKE "地址问题" OR 派送问题 LIKE "客户要求更改派送时间或者地址","地址问题/客户要求更改派送时间或者地址",派送问题) AS 问题件类型, 
                        IF(备注 <> "", IF(备注 LIKE "已签收%" OR 备注 LIKE "已完结%","已完结",IF(备注 LIKE "无人接听%" OR 备注 LIKE "无效号码%","无人接听", IF(备注 LIKE "已通知%" OR 备注 LIKE "已告知%" OR 备注 LIKE "请告知%" OR 备注 LIKE "请通知%","已发短信", 
                        IF(备注 LIKE "%*%","未回复",IF((备注 NOT LIKE "%*%" AND 备注 NOT LIKE "%拒收%") AND (备注 LIKE "%客%取%" OR 备注 LIKE "%客%拿%" OR 备注 LIKE "%送货%" OR 备注 LIKE "%送貨%" OR 备注 LIKE "%取件%" OR 备注 LIKE "%取货%" OR 备注 LIKE "%取貨%"),"回复",""))))),备注) AS 回复类型
                 FROM 派送问题件_跟进表 p
                 WHERE p.创建日期 >= '{0}'
                 ORDER BY 币种, 创建日期 , 
                 FIELD(问题件类型,'送至便利店','地址问题/客户要求更改派送时间或者地址','客户长期不在','送达客户不在','客户不接电话','拒收','合计');'''.format(timeStart)
        sql = '''SELECT *
                FROM ( SELECT s1.*, IF(s1.派送问题 LIKE "地址问题" OR s1.派送问题 LIKE "客户要求更改派送时间或者地址","地址问题/客户要求更改派送时间或者地址",s1.派送问题) AS 问题件类型, 
					        IF(s1.备注 <> "", IF(s1.备注 LIKE "已签收%" OR s1.备注 LIKE "已完结%" OR s1.备注 LIKE "已拒收%"  OR s1.备注 LIKE "显示拒收%","已完结",IF(s1.备注 LIKE "无人接听%" OR s1.备注 LIKE "无效号码%","无人接听", IF(s1.备注 LIKE "已通知%" OR s1.备注 LIKE "已告知%" OR s1.备注 LIKE "请告知%" OR s1.备注 LIKE "请通知%","已发短信", IF(s1.备注 LIKE "%*%","未回复",
					        IF((s1.备注 NOT LIKE "%*%" AND s1.备注 NOT LIKE "%拒收%") AND (s1.备注 LIKE "%客%取%" OR s1.备注 LIKE "%客%拿%" OR s1.备注 LIKE "%送货%" OR s1.备注 LIKE "%送貨%" OR s1.备注 LIKE "%取件%" OR s1.备注 LIKE "%取货%" OR s1.备注 LIKE "%取貨%"),"回复",""))))),s1.备注) AS 回复类型, s2.核实原因
		            FROM ( SELECT ss1.订单编号,	ss1.币种,	ss1.下单时间,	ss1.订单状态,	ss1.物流状态,	ss1.物流渠道,	ss1.创建日期,	ss1.创建时间,	ss1.派送问题, ss1.处理人,	ss1.处理记录,	ss1.处理时间, IF(ss1.备注 = "",ss2.处理结果,ss1.备注) AS 备注
				            FROM ( SELECT *
                                    FROM 派送问题件_跟进表 p
						            WHERE p.创建日期 >= '{0}'  
				            ) ss1
				            LEFT JOIN 物流问题件 ss2 ON ss1.订单编号 = ss2.订单编号
				            WHERE ss1.处理人 <> ""
		            ) s1
		            LEFT JOIN 
		            (SELECT *
						FROM 拒收问题件
						WHERE id IN (SELECT MAX(id) FROM 拒收问题件 y WHERE y.处理时间 >= DATE_SUB(CURDATE(), INTERVAL 2 month) GROUP BY 订单编号) 
					) s2 ON s1.订单编号 = s2.订单编号
                ) s
                WHERE s.回复类型= "回复" AND (s.`核实原因` IS NULL OR s.`核实原因`= "未联系上客户") AND s.物流状态 = '拒收';'''.format(timeStart)
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)

        print('正在获取拒收内容…………')
        sql = '''SELECT 创建日期, IFNULL(具体原因,'合计') AS 拒收原因, 单量,concat(ROUND(IFNULL(单量 / 总单量,0) * 100,2),'%') as 占比
                FROM (	
                    SELECT 创建日期, 具体原因,COUNT(s1.订单编号) AS 单量
                    FROM(SELECT *, IF(派送问题 LIKE "地址问题" OR 派送问题 LIKE "客户要求更改派送时间或者地址","地址问题/客户要求更改派送时间或者地址",IF(派送问题 LIKE "送达客户不在" OR 派送问题 LIKE "客户长期不在","送达客户不在/客户长期不在",派送问题)) AS 问题件类型,
                                IF(备注 <> "", IF(备注 LIKE "已签收%" OR 备注 LIKE "已完结%","已完结",IF(备注 LIKE "无人接听%" OR 备注 LIKE "无效号码%","无人接听", IF(备注 LIKE "已通知%" OR 备注 LIKE "已告知%" OR 备注 LIKE "请告知%" OR 备注 LIKE "请通知%","已发短信", 
                                IF(备注 LIKE "%*%","未回复",IF((备注 NOT LIKE "%*%" AND 备注 NOT LIKE "%拒收%") AND (备注 LIKE "%客%取%" OR 备注 LIKE "%客%拿%" OR 备注 LIKE "%送货%" OR 备注 LIKE "%送貨%" OR 备注 LIKE "%取件%" OR 备注 LIKE "%取货%" OR 备注 LIKE "%取貨%"),"回复",""))))),备注) AS 回复类型
                             FROM 派送问题件_跟进表 p
                             WHERE p.创建日期 >= '{0}'    AND p.物流状态 = "拒收"
                    ) s1
                    LEFT JOIN 拒收问题件 js ON s1.订单编号 =js.订单编号
                    WHERE s1.回复类型 = "回复" AND js.具体原因 <> '未联系上客户' AND js.具体原因 IS not NULL
                    GROUP BY 创建日期, 具体原因
                    WITH ROLLUP
                ) s
                LEFT JOIN 
                ( SELECT 创建日期 日期, 具体原因 具体,COUNT(s1.订单编号) AS 总单量
                   FROM(SELECT *, IF(备注 <> "", IF(备注 LIKE "已签收%" OR 备注 LIKE "已完结%","已完结",IF(备注 LIKE "无人接听%" OR 备注 LIKE "无效号码%","无人接听", IF(备注 LIKE "已通知%" OR 备注 LIKE "已告知%" OR 备注 LIKE "请告知%" OR 备注 LIKE "请通知%","已发短信", IF(备注 LIKE "%*%","未回复",IF((备注 NOT LIKE "%*%" AND 备注 NOT LIKE "%拒收%") AND (备注 LIKE "%客%取%" OR 备注 LIKE "%客%拿%" OR 备注 LIKE "%送货%" OR 备注 LIKE "%送貨%" OR 备注 LIKE "%取件%" OR 备注 LIKE "%取货%" OR 备注 LIKE "%取貨%"),"回复",""))))),备注) AS 回复类型
                        FROM 派送问题件_跟进表 p
                        WHERE p.创建日期 >= '{0}'    AND p.物流状态 = "拒收"
                   ) s1
                   LEFT JOIN 拒收问题件 js ON s1.订单编号 =js.订单编号
                   WHERE s1.回复类型 = "回复" AND js.具体原因 <> '未联系上客户' AND js.具体原因 IS not NULL
                   GROUP BY 创建日期
                ) ss ON s.创建日期 =ss.日期
                HAVING 创建日期 IS NOT NULL
                ORDER BY 创建日期, FIELD(拒收原因,'合计') DESC, 单量 DESC;'''.format(timeStart)
        sql = '''SELECT 创建日期, IFNULL(具体原因,'合计') AS 拒收原因, 单量,concat(ROUND(IFNULL(单量 / 总单量,0) * 100,2),'%') as 占比
                FROM ( SELECT 创建日期, 具体原因,COUNT(s1.订单编号) AS 单量
                     FROM( SELECT 订单编号,币种, IF(完结状态时间 IS NULL,状态时间,完结状态时间) AS 完结时间,  DATE_FORMAT(IF(完结状态时间 IS NULL,状态时间,完结状态时间), '%Y-%m-%d') AS 创建日期
                            FROM d1_gat d
                            WHERE d.最终状态 = '拒收'
                     ) s1
                    LEFT JOIN 
                     (SELECT *
						FROM 拒收问题件
						WHERE id IN (SELECT MAX(id) FROM 拒收问题件 y WHERE y.处理时间 >= DATE_SUB(CURDATE(), INTERVAL 2 month) GROUP BY 订单编号) 
					) js ON s1.订单编号 =js.订单编号
                    WHERE js.核实原因 <> '未联系上客户' AND js.具体原因 IS not NULL
                    GROUP BY 创建日期, 具体原因
                    WITH ROLLUP
                ) s
                LEFT JOIN 
                ( SELECT 创建日期 日期, 具体原因 具体,COUNT(s1.订单编号) AS 总单量
                    FROM(  SELECT 订单编号,币种, IF(完结状态时间 IS NULL,状态时间,完结状态时间) AS 完结时间,  DATE_FORMAT(IF(完结状态时间 IS NULL,状态时间,完结状态时间), '%Y-%m-%d') AS 创建日期
                            FROM d1_gat d
                            WHERE d.最终状态 = '拒收'
                     ) s1
                     LEFT JOIN 
                    (SELECT *
						FROM 拒收问题件
						WHERE id IN (SELECT MAX(id) FROM 拒收问题件 y WHERE y.处理时间 >= DATE_SUB(CURDATE(), INTERVAL 2 month) GROUP BY 订单编号) 
					)  js ON s1.订单编号 =js.订单编号
                     WHERE js.核实原因 <> '未联系上客户' AND js.具体原因 IS not NULL
                     GROUP BY 创建日期
                ) ss ON s.创建日期 =ss.日期
                WHERE 创建日期 IS NOT NULL
                ORDER BY 创建日期, FIELD(拒收原因,'合计') DESC, 单量 DESC;'''.format(timeStart)
        df11 = pd.read_sql_query(sql=sql, con=self.engine1)

        print('正在获取物流内容…………')
        sql = '''SELECT 币种, 日期,周, 
                        全部签收 AS 签收单量, 全部拒收 拒收单量, 
                        concat(ROUND(IFNULL(全部签收 / 全部单量,0) * 100,2),'%') as 签收率,
                        concat(ROUND(IFNULL(全部退货 / 全部单量,0) * 100,2),'%') as 退款率,
                    速派签收单量, 速派拒收单量, 
                        concat(ROUND(IFNULL(速派签收单量 / 速派单量,0) * 100,2),'%') as 速派签收率,
                        concat(ROUND(IFNULL(速派退货单量 / 速派单量,0) * 100,2),'%') as 速派退款率,
                    天马签收单量, 天马拒收单量, 
                        concat(ROUND(IFNULL(天马签收单量 / 天马单量,0) * 100,2),'%') as 天马签收率,
                        concat(ROUND(IFNULL(天马退货单量 / 天马单量,0) * 100,2),'%') as 天马退款率,
                    协来运签收单量, 协来运拒收单量, 
                        concat(ROUND(IFNULL(协来运签收单量 / 协来运单量,0) * 100,2),'%') as 协来运签收率,
                        concat(ROUND(IFNULL(协来运退货单量 / 协来运单量,0) * 100,2),'%') as 协来运退款率,
                    易速配签收单量, 易速配拒收单量, 
                        concat(ROUND(IFNULL(易速配签收单量 / 易速配单量,0) * 100,2),'%') as 易速配签收率,
                        concat(ROUND(IFNULL(易速配退货单量 / 易速配单量,0) * 100,2),'%') as 易速配退款率,
                    立邦签收单量, 立邦拒收单量, 
                        concat(ROUND(IFNULL(立邦签收单量 / 立邦单量,0) * 100,2),'%') as 立邦签收率,
                        concat(ROUND(IFNULL(立邦退货单量 / 立邦单量,0) * 100,2),'%') as 立邦退款率,
                    圆通签收单量, 圆通拒收单量, 
                        concat(ROUND(IFNULL(圆通签收单量 / 圆通单量,0) * 100,2),'%') as 圆通签收率,
                        concat(ROUND(IFNULL(圆通退货单量 / 圆通单量,0) * 100,2),'%') as 圆通退款率
				FROM (  SELECT 币种, 日期, CASE DATE_FORMAT(日期,'%w')	WHEN 1 THEN '星期一' WHEN 2 THEN '星期二' WHEN 3 THEN '星期三' WHEN 4 THEN '星期四' WHEN 5 THEN '星期五' WHEN 6 THEN '星期六' WHEN 0 THEN '星期日' END as 周,
                            IF(`物流名称` = '全部',总单量,0) AS 全部单量,
                            IF(`物流名称` = '全部',签收单量,0) AS 全部签收,
                            IF(`物流名称` = '全部',拒收单量,0) AS 全部拒收,
                            IF(`物流名称` = '全部',退货单量,0) AS 全部退货,
                        SUM(IF(`物流名称` = '速派',总单量,0)) AS 速派单量,
                            SUM(IF(`物流名称` = '速派',签收单量,0)) AS 速派签收单量,
                            SUM(IF(`物流名称` = '速派',拒收单量,0)) AS 速派拒收单量,
                            SUM(IF(`物流名称` = '速派',退货单量,0)) AS 速派退货单量,
                        SUM(IF(`物流名称` = '天马',总单量,0)) AS 天马单量,
                            SUM(IF(`物流名称` = '天马',签收单量,0)) AS 天马签收单量,
                            SUM(IF(`物流名称` = '天马',拒收单量,0)) AS 天马拒收单量,
                            SUM(IF(`物流名称` = '天马',退货单量,0)) AS 天马退货单量,
                        SUM(IF(`物流名称` = '协来运',总单量,0)) AS 协来运单量,
                            SUM(IF(`物流名称` = '协来运',签收单量,0)) AS 协来运签收单量,
                            SUM(IF(`物流名称` = '协来运',拒收单量,0)) AS 协来运拒收单量,
                            SUM(IF(`物流名称` = '协来运',退货单量,0)) AS 协来运退货单量,
                        SUM(IF(`物流名称` = '易速配',总单量,0)) AS 易速配单量,
                            SUM(IF(`物流名称` = '易速配',签收单量,0)) AS 易速配签收单量,
                            SUM(IF(`物流名称` = '易速配',拒收单量,0)) AS 易速配拒收单量,
                            SUM(IF(`物流名称` = '易速配',退货单量,0)) AS 易速配退货单量,
                        SUM(IF(`物流名称` = '立邦',总单量,0)) AS 立邦单量,
                            SUM(IF(`物流名称` = '立邦',签收单量,0)) AS 立邦签收单量,
                            SUM(IF(`物流名称` = '立邦',拒收单量,0)) AS 立邦拒收单量,
                            SUM(IF(`物流名称` = '立邦',退货单量,0)) AS 立邦退货单量,
                        SUM(IF(`物流名称` = '圆通',总单量,0)) AS 圆通单量,
                            SUM(IF(`物流名称` = '圆通',签收单量,0)) AS 圆通签收单量,
                            SUM(IF(`物流名称` = '圆通',拒收单量,0)) AS 圆通拒收单量,
                            SUM(IF(`物流名称` = '圆通',退货单量,0)) AS 圆通退货单量
                        FROM 派送问题件_跟进表2_cp p 
						WHERE p.`日期` >= '{0}'
                        GROUP BY 币种, 日期
				) s1
                ORDER BY 币种, 日期;'''.format(timeStart)
        sql = '''SELECT 币种, 年月 月份, 日期, 周,
                        全部签收 AS 签收单量, 全部拒收 拒收单量, 
                            concat(ROUND(IFNULL(全部签收 / 全部单量,0) * 100,2),'%') as 签收率,
                            concat(ROUND(IFNULL(全部退货 / 全部单量,0) * 100,2),'%') as 退款率,
                        速派签收单量, 速派拒收单量, 
                            concat(ROUND(IFNULL(速派签收单量 / 速派单量,0) * 100,2),'%') as 速派签收率,
                            concat(ROUND(IFNULL(速派退货单量 / 速派单量,0) * 100,2),'%') as 速派退款率,
                        天马签收单量, 天马拒收单量, 
                            concat(ROUND(IFNULL(天马签收单量 / 天马单量,0) * 100,2),'%') as 天马签收率,
                            concat(ROUND(IFNULL(天马退货单量 / 天马单量,0) * 100,2),'%') as 天马退款率,
                        协来运签收单量, 协来运拒收单量, 
                            concat(ROUND(IFNULL(协来运签收单量 / 协来运单量,0) * 100,2),'%') as 协来运签收率,
                            concat(ROUND(IFNULL(协来运退货单量 / 协来运单量,0) * 100,2),'%') as 协来运退款率,
                        易速配签收单量, 易速配拒收单量, 
                            concat(ROUND(IFNULL(易速配签收单量 / 易速配单量,0) * 100,2),'%') as 易速配签收率,
                            concat(ROUND(IFNULL(易速配退货单量 / 易速配单量,0) * 100,2),'%') as 易速配退款率,
                                                    
                        上月全部签收 AS 上月签收单量, 上月全部拒收 上月拒收单量, 
                            concat(ROUND(IFNULL(上月全部签收 / 上月全部单量,0) * 100,2),'%') as 上月签收率,
                            concat(ROUND(IFNULL(上月全部退货 / 上月全部单量,0) * 100,2),'%') as 上月退款率,
                        上月速派签收单量, 上月速派拒收单量, 
                            concat(ROUND(IFNULL(上月速派签收单量 / 上月速派单量,0) * 100,2),'%') as 上月速派签收率,
                            concat(ROUND(IFNULL(上月速派退货单量 / 上月速派单量,0) * 100,2),'%') as 上月速派退款率,
                        上月天马签收单量, 上月天马拒收单量, 
                            concat(ROUND(IFNULL(上月天马签收单量 / 上月天马单量,0) * 100,2),'%') as 上月天马签收率,
                            concat(ROUND(IFNULL(上月天马退货单量 / 上月天马单量,0) * 100,2),'%') as 上月天马退款率,
                        上月协来运签收单量, 上月协来运拒收单量, 
                            concat(ROUND(IFNULL(上月协来运签收单量 / 上月协来运单量,0) * 100,2),'%') as 上月协来运签收率,
                            concat(ROUND(IFNULL(上月协来运退货单量 / 上月协来运单量,0) * 100,2),'%') as 上月协来运退款率,
                         上月易速配签收单量, 上月易速配拒收单量, 
                            concat(ROUND(IFNULL(上月易速配签收单量 / 上月易速配单量,0) * 100,2),'%') as 上月易速配签收率,
                            concat(ROUND(IFNULL(上月易速配退货单量 / 上月易速配单量,0) * 100,2),'%') as 上月易速配退款率,
                                                                        
                        立邦签收单量, 立邦拒收单量, 
                            concat(ROUND(IFNULL(立邦签收单量 / 立邦单量,0) * 100,2),'%') as 立邦签收率,
                            concat(ROUND(IFNULL(立邦退货单量 / 立邦单量,0) * 100,2),'%') as 立邦退款率,
                        圆通签收单量, 圆通拒收单量, 
                            concat(ROUND(IFNULL(圆通签收单量 / 圆通单量,0) * 100,2),'%') as 圆通签收率,
                            concat(ROUND(IFNULL(圆通退货单量 / 圆通单量,0) * 100,2),'%') as 圆通退款率,
                                                    
                        上月立邦签收单量, 上月立邦拒收单量, 
                            concat(ROUND(IFNULL(上月立邦签收单量 / 上月立邦单量,0) * 100,2),'%') as 上月立邦签收率,
                            concat(ROUND(IFNULL(上月立邦退货单量 / 上月立邦单量,0) * 100,2),'%') as 上月立邦退款率,
                        上月圆通签收单量, 上月圆通拒收单量, 
                            concat(ROUND(IFNULL(上月圆通签收单量 / 上月圆通单量,0) * 100,2),'%') as 上月圆通签收率,
                            concat(ROUND(IFNULL(上月圆通退货单量 / 上月圆通单量,0) * 100,2),'%') as 上月圆通退款率					
				FROM (  SELECT IFNULL(币种,'合计') 币种,IFNULL(年月,'合计') 年月,IFNULL(日期,'合计') 日期, CASE DATE_FORMAT(日期,'%w')	WHEN 1 THEN '星期一' WHEN 2 THEN '星期二' WHEN 3 THEN '星期三' WHEN 4 THEN '星期四' WHEN 5 THEN '星期五' WHEN 6 THEN '星期六' WHEN 0 THEN '星期日' END as 周,
                            SUM(IF(`物流名称` = '全部',总单量,0)) AS 全部单量,
                                SUM(IF(`物流名称` = '全部',签收单量,0)) AS 全部签收,
                                SUM(IF(`物流名称` = '全部',拒收单量,0)) AS 全部拒收,
                                SUM(IF(`物流名称` = '全部',退货单量,0)) AS 全部退货,
                            SUM(IF(`物流名称` = '速派',总单量,0)) AS 速派单量,
                                SUM(IF(`物流名称` = '速派',签收单量,0)) AS 速派签收单量,
                                SUM(IF(`物流名称` = '速派',拒收单量,0)) AS 速派拒收单量,
                                SUM(IF(`物流名称` = '速派',退货单量,0)) AS 速派退货单量,
                            SUM(IF(`物流名称` = '天马',总单量,0)) AS 天马单量,
                                SUM(IF(`物流名称` = '天马',签收单量,0)) AS 天马签收单量,
                                SUM(IF(`物流名称` = '天马',拒收单量,0)) AS 天马拒收单量,
                                SUM(IF(`物流名称` = '天马',退货单量,0)) AS 天马退货单量,
                            SUM(IF(`物流名称` = '协来运',总单量,0)) AS 协来运单量,
                                SUM(IF(`物流名称` = '协来运',签收单量,0)) AS 协来运签收单量,
                                SUM(IF(`物流名称` = '协来运',拒收单量,0)) AS 协来运拒收单量,
                                SUM(IF(`物流名称` = '协来运',退货单量,0)) AS 协来运退货单量,
                            SUM(IF(`物流名称` = '易速配',总单量,0)) AS 易速配单量,
                                SUM(IF(`物流名称` = '易速配',签收单量,0)) AS 易速配签收单量,
                                SUM(IF(`物流名称` = '易速配',拒收单量,0)) AS 易速配拒收单量,
                                SUM(IF(`物流名称` = '易速配',退货单量,0)) AS 易速配退货单量,
														
                            SUM(IF(`物流名称` = '全部',上月总单量,0)) AS 上月全部单量,
                                SUM(IF(`物流名称` = '全部',上月签收单量,0)) AS 上月全部签收,
                                SUM(IF(`物流名称` = '全部',上月拒收单量,0)) AS 上月全部拒收,
                                SUM(IF(`物流名称` = '全部',上月退货单量,0)) AS 上月全部退货,
                            SUM(IF(`物流名称` = '速派',上月总单量,0)) AS 上月速派单量,
                                SUM(IF(`物流名称` = '速派',上月签收单量,0)) AS 上月速派签收单量,
                                SUM(IF(`物流名称` = '速派',上月拒收单量,0)) AS 上月速派拒收单量,
                                SUM(IF(`物流名称` = '速派',上月退货单量,0)) AS 上月速派退货单量,
                            SUM(IF(`物流名称` = '天马',上月总单量,0)) AS 上月天马单量,
                                SUM(IF(`物流名称` = '天马',上月签收单量,0)) AS 上月天马签收单量,
                                SUM(IF(`物流名称` = '天马',上月拒收单量,0)) AS 上月天马拒收单量,
                                SUM(IF(`物流名称` = '天马',上月退货单量,0)) AS 上月天马退货单量,
                            SUM(IF(`物流名称` = '协来运',上月总单量,0)) AS 上月协来运单量,
                                SUM(IF(`物流名称` = '协来运',上月签收单量,0)) AS 上月协来运签收单量,
                                SUM(IF(`物流名称` = '协来运',上月拒收单量,0)) AS 上月协来运拒收单量,
                                SUM(IF(`物流名称` = '协来运',上月退货单量,0)) AS 上月协来运退货单量,
                            SUM(IF(`物流名称` = '易速配',上月总单量,0)) AS 上月易速配单量,
                                SUM(IF(`物流名称` = '易速配',上月签收单量,0)) AS 上月易速配签收单量,
                                SUM(IF(`物流名称` = '易速配',上月拒收单量,0)) AS 上月易速配拒收单量,
                                SUM(IF(`物流名称` = '易速配',上月退货单量,0)) AS 上月易速配退货单量,
														
                            SUM(IF(`物流名称` = '立邦',总单量,0)) AS 立邦单量,
                                SUM(IF(`物流名称` = '立邦',签收单量,0)) AS 立邦签收单量,
                                SUM(IF(`物流名称` = '立邦',拒收单量,0)) AS 立邦拒收单量,
                                SUM(IF(`物流名称` = '立邦',退货单量,0)) AS 立邦退货单量,
                            SUM(IF(`物流名称` = '圆通',总单量,0)) AS 圆通单量,
                                SUM(IF(`物流名称` = '圆通',签收单量,0)) AS 圆通签收单量,
                                SUM(IF(`物流名称` = '圆通',拒收单量,0)) AS 圆通拒收单量,
                                SUM(IF(`物流名称` = '圆通',退货单量,0)) AS 圆通退货单量,
                                                            
                            SUM(IF(`物流名称` = '立邦',上月总单量,0)) AS 上月立邦单量,
                                SUM(IF(`物流名称` = '立邦',上月签收单量,0)) AS 上月立邦签收单量,
                                SUM(IF(`物流名称` = '立邦',上月拒收单量,0)) AS 上月立邦拒收单量,
                                SUM(IF(`物流名称` = '立邦',上月退货单量,0)) AS 上月立邦退货单量,
                            SUM(IF(`物流名称` = '圆通',上月总单量,0)) AS 上月圆通单量,
                                SUM(IF(`物流名称` = '圆通',上月签收单量,0)) AS 上月圆通签收单量,
                                SUM(IF(`物流名称` = '圆通',上月拒收单量,0)) AS 上月圆通拒收单量,
                                SUM(IF(`物流名称` = '圆通',上月退货单量,0)) AS 上月圆通退货单量
                        FROM ( SELECT s1.*,s2.`总单量` AS 上月总单量 ,s2.`签收单量` AS 上月签收单量 ,s2.`拒收单量` AS 上月拒收单量 ,s2.`退货单量` AS 上月退货单量 
								FROM( SELECT *,EXTRACT(YEAR_MONTH FROM p.日期) AS 年月
									 FROM 派送问题件_跟进表2_cp p 
									 WHERE p.`日期` >= '{0}'
							    ) s1
							    LEFT JOIN  派送问题件_跟进表2_cp s2 on s1.币种 = s2.币种 AND s1.物流名称 = s2.物流名称 AND s1.日期 = DATE_SUB(s2.日期,INTERVAL -1 MONTH)
						) p 
                        GROUP BY 币种, 年月, 日期
						WITH ROLLUP
				) s1
				WHERE s1.年月 <> '合计'
                ORDER BY 币种, 年月, 日期;'''.format(timeStart)
        df12 = pd.read_sql_query(sql=sql, con=self.engine1)
        df121 = df12[(df12['币种'].str.contains('台币'))]
        # df1211 = df121[["月份","日期","周","签收单量","拒收单量","签收率","退款率", "上月签收单量","上月拒收单量","上月签收率","上月退款率",
        #                 "速派签收单量","速派拒收单量","速派签收率","速派退款率", "上月速派签收单量","上月速派拒收单量","上月速派签收率","上月速派退款率",
        #                 "天马签收单量","天马拒收单量","天马签收率","天马退款率", "上月天马签收单量","上月天马拒收单量","上月天马签收率","上月天马退款率",
        #                 "协来运签收单量","协来运拒收单量","协来运签收率","协来运退款率", "上月协来运签收单量","上月协来运拒收单量","上月协来运签收率","上月协来运退款率",
        #                 "易速配签收单量","易速配拒收单量","易速配签收率","易速配退款率", "上月易速配签收单量","上月易速配拒收单量","上月易速配签收率","上月易速配退款率"]].copy()
        df1211 = df121[["月份","日期","周","签收单量","拒收单量","签收率","退款率", "上月签收单量","上月拒收单量","上月签收率","上月退款率",
                        "速派签收单量","速派拒收单量","速派签收率","速派退款率", "上月速派签收单量","上月速派拒收单量","上月速派签收率","上月速派退款率",
                        "天马签收单量","天马拒收单量","天马签收率","天马退款率", "上月天马签收单量","上月天马拒收单量","上月天马签收率","上月天马退款率",
                        "协来运签收单量","协来运拒收单量","协来运签收率","协来运退款率", "上月协来运签收单量","上月协来运拒收单量","上月协来运签收率","上月协来运退款率"]].copy()
        df122 = df12[(df12['币种'].str.contains('港币'))]
        df1222 = df122[["月份","日期","周","签收单量","拒收单量","签收率","退款率","上月签收单量","上月拒收单量","上月签收率","上月退款率",
                        "立邦签收单量","立邦拒收单量","立邦签收率","立邦退款率","上月立邦签收单量","上月立邦拒收单量","上月立邦签收率","上月立邦退款率",
                        "圆通签收单量","圆通拒收单量","圆通签收率","圆通退款率","上月圆通签收单量","上月圆通拒收单量","上月圆通签收率","上月圆通退款率"]].copy()

        print('正在获取跟进内容…………')
        sql = '''SELECT 币种, EXTRACT(YEAR_MONTH FROM 创建日期) AS 月份, 创建日期, CASE DATE_FORMAT(创建日期,'%w')	WHEN 1 THEN '星期一' WHEN 2 THEN '星期二' WHEN 3 THEN '星期三' WHEN 4 THEN '星期四' WHEN 5 THEN '星期五' WHEN 6 THEN '星期六' WHEN 0 THEN '星期日' END as 上月周,
                        总单量, 签收单量, 拒收单量, concat(ROUND(IFNULL(签收单量 / 总单量,0) * 100,2),'%') as 签收率, 派送问题件单量, 问题件类型,单量,短信,邮件,在线, 
                        IF(电话 = 0,NULL,电话) AS 电话,IF(客户回复再派量 = 0,NULL,客户回复再派量) AS 客户回复再派量,
                        concat(ROUND(IFNULL(物流再派签收 / 物流再派,0) * 100,2),'%') as 物流再派签收率,
                        concat(ROUND(IFNULL(物流3派签收 / 物流3派,0) * 100,2),'%') as 物流3派签收率,
                 --       IF(问题件类型 = '客户不接电话' and 未派 <> 0,CONCAT(未派,'单处理时已完结'),IF(单量 >= 短信,"获取物流轨迹信息后，后台会排队处理；若30-40分钟内订单状态变为已完结，则不发送短信。",IF(单量 < 短信,"物流轨迹更新后， 根据派送问题类型的更改，会再次发送短信。", NULL))) 未派, 											
                        IF(问题件类型 = '送达客户不在/客户长期不在' AND 创建日期 >='{1}' AND 创建日期 <='{2}','暂未处理。物流已2派或3派',IF(问题件类型 = '客户不接电话' and 未派 <> 0,CONCAT(未派,'单处理时已完结'),IF(单量 >= 短信,"获取物流轨迹信息后，后台会排队处理；若30-40分钟内订单状态变为已完结，则不发送短信。",IF(单量 < 短信,"物流轨迹更新后， 根据派送问题类型的更改，会再次发送短信。", NULL))))  未派, 
                        异常, 上月总单量, 上月签收单量, 上月拒收单量, 
                        concat(ROUND(IFNULL(上月签收单量 / 上月总单量,0) * 100,2),'%') as 上月签收率, 上月派送问题件单量,上月周
                FROM ( SELECT s1.币种, s1.创建日期, s3.签收单量, s3.拒收单量, s3.总单量, 派送问题件单量, 问题件类型,
                            COUNT(订单编号) AS 单量, 发送量 短信, NULL AS 邮件, NULL AS 在线, 
                            SUM(IF(s1.备注 <> "" AND s1.回复类型 <> "已完结" AND s1.回复类型 <> "已发短信",1,0)) AS 电话, 
                            SUM(IF(回复类型 = "回复",1,0)) AS 客户回复再派量, 物流再派, 物流再派签收, 物流3派, 物流3派签收, 
                            SUM(IF(s1.回复类型 = "已完结" OR s1.回复类型 = "已发短信",1,0)) AS 未派, 异常,
                            s4.签收单量 AS 上月签收单量, s4.拒收单量 AS 上月拒收单量, s4.总单量 AS 上月总单量, s5.上月派送问题件单量, s5.上月周
                    FROM( SELECT *, IF(派送问题 LIKE "地址问题" OR 派送问题 LIKE "客户要求更改派送时间或者地址","地址问题/客户要求更改派送时间或者地址",IF(派送问题 LIKE "送达客户不在" OR 派送问题 LIKE "客户长期不在","送达客户不在/客户长期不在",派送问题)) AS 问题件类型,
                                IF(备注 <> "", IF(备注 LIKE "已签收%" OR 备注 LIKE "已完结%" OR 备注 LIKE "已拒收%"  OR 备注 LIKE "显示拒收%" ,"已完结", IF(备注 LIKE "无人接听%" OR 备注 LIKE "无效号码%","无人接听", IF(备注 LIKE "已通知%" OR 备注 LIKE "已告知%" OR 备注 LIKE "请告知%" OR 备注 LIKE "请通知%","已发短信", 
	                            IF(备注 LIKE "%*%","未回复",IF((备注 NOT LIKE "%*%" AND 备注 NOT LIKE "%拒收%") AND (备注 LIKE "%客%取%" OR 备注 LIKE "%客%拿%" OR 备注 LIKE "%送货%" OR 备注 LIKE "%送貨%" OR 备注 LIKE "%取件%" OR 备注 LIKE "%取货%" OR 备注 LIKE "%取貨%"),"回复",""))))),备注) AS 回复类型
                        FROM ( SELECT ss1.订单编号,	ss1.币种,	ss1.下单时间,	ss1.订单状态,	ss1.物流状态,	ss1.物流渠道,	ss1.创建日期,	ss1.创建时间,	ss1.派送问题, ss1.派送次数, ss1.处理人,	ss1.处理记录,	ss1.处理时间, IF(ss1.备注 = "",ss2.处理结果,ss1.备注) AS 备注
								FROM ( SELECT * FROM 派送问题件_跟进表 p WHERE p.创建日期 >= '{0}' ) ss1
								LEFT JOIN 物流问题件 ss2 ON ss1.订单编号 = ss2.订单编号
							) p
                        ) s1
                        LEFT JOIN 
                        ( SELECT 币种, 创建日期, COUNT(订单编号) AS 派送问题件单量,
                                SUM(IF(派送次数 = 2,1,0)) AS 物流再派,
                                SUM(IF(物流状态 = "已签收" AND 派送次数 = 2,1,0)) AS 物流再派签收,
                                SUM(IF(派送次数 > 2,1,0)) AS 物流3派,
                                SUM(IF(物流状态 = "已签收" AND 派送次数 > 2,1,0)) AS 物流3派签收,
                                SUM(IF(回复类型 = "回复" AND 物流状态 = "拒收",1,0)) AS 异常
                          FROM ( SELECT *, IF(备注 <> "", IF(备注 LIKE "已签收%" OR 备注 LIKE "已完结%" OR 备注 LIKE "已拒收%"  OR 备注 LIKE "显示拒收%","已完结",IF(备注 LIKE "无人接听%" OR 备注 LIKE "无效号码%","无人接听", IF(备注 LIKE "已通知%" OR 备注 LIKE "已告知%" OR 备注 LIKE "请告知%" OR 备注 LIKE "请通知%","已发短信", 
	                                    IF(备注 LIKE "%*%","未回复",IF((备注 NOT LIKE "%*%" AND 备注 NOT LIKE "%拒收%") AND (备注 LIKE "%客%取%" OR 备注 LIKE "%客%拿%" OR 备注 LIKE "%送货%" OR 备注 LIKE "%送貨%" OR 备注 LIKE "%取件%" OR 备注 LIKE "%取货%" OR 备注 LIKE "%取貨%"),"回复",""))))),备注) AS 回复类型
                                 FROM ( SELECT ss1.订单编号,	ss1.币种,	ss1.下单时间,	ss1.订单状态,	ss1.物流状态,	ss1.物流渠道,	ss1.创建日期,	ss1.创建时间,	ss1.派送问题, ss1.派送次数, ss1.处理人,	ss1.处理记录,	ss1.处理时间, IF(ss1.备注 = "",ss2.处理结果,ss1.备注) AS 备注
										FROM ( SELECT * FROM 派送问题件_跟进表 p WHERE p.创建日期 >= '{0}' ) ss1
										LEFT JOIN 物流问题件 ss2 ON ss1.订单编号 = ss2.订单编号
								) p
                            ) PP
                          GROUP BY 币种, 创建日期
                        ) s2 on s1.币种 =s2.币种 AND s1.创建日期 =s2.创建日期
                        LEFT JOIN (SELECT * FROM 派送问题件_跟进表2_cp p WHERE p.`物流名称` = '全部') s3 on s1.币种 = s3.币种 AND s1.创建日期 = s3.日期
                        LEFT JOIN (SELECT * FROM 派送问题件_跟进表2_cp p WHERE p.`物流名称` = '全部') s4 on s1.币种 = s4.币种 AND s1.创建日期 = DATE_SUB(s4.日期,INTERVAL -1 MONTH)
                        LEFT JOIN (SELECT 币种, 创建日期, CASE DATE_FORMAT(创建日期,'%w')	WHEN 1 THEN '星期一' WHEN 2 THEN '星期二' WHEN 3 THEN '星期三' WHEN 4 THEN '星期四' WHEN 5 THEN '星期五' WHEN 6 THEN '星期六' WHEN 0 THEN '星期日' END as 上月周, COUNT(订单编号) AS 上月派送问题件单量
                                    FROM 派送问题件_跟进表 p
                                    WHERE p.创建日期 >= DATE_SUB('{0}',INTERVAL 1 MONTH)  AND p.创建日期 < '{0}'
                                    GROUP BY 币种, 创建日期
                        ) s5 on s1.币种 = s5.币种 AND s1.创建日期 = DATE_SUB(s5.创建日期,INTERVAL -1 MONTH)
                        LEFT JOIN 派送问题件_跟进表_message s6 on s1.币种 = s6.币种 AND s1.创建日期 = s6.日期 AND s1.问题件类型 =s6.短信模板
                        GROUP BY s1.币种, s1.创建日期, s1.问题件类型
                ) s
                ORDER BY s.币种, s.创建日期 , 
                FIELD(s.问题件类型,'送至便利店','地址问题/客户要求更改派送时间或者地址','客户自取','客户不接电话','送达客户不在/客户长期不在','拒收','合计');'''.format(timeStart, logisticsN_begin, logisticsN_end)
        sql = '''SELECT 币种, EXTRACT(YEAR_MONTH FROM 日期) AS 月份, 日期 AS 创建日期, 
                        CASE DATE_FORMAT(日期,'%w')	WHEN 1 THEN '星期一' WHEN 2 THEN '星期二' WHEN 3 THEN '星期三' WHEN 4 THEN '星期四' WHEN 5 THEN '星期五' WHEN 6 THEN '星期六' WHEN 0 THEN '星期日' END as 上月周,
                        总单量, 签收单量, 拒收单量, concat(ROUND(IFNULL(签收单量 / 总单量,0) * 100,2),'%') as 签收率, 派送问题件单量, 问题件类型,单量,短信,邮件,在线, IF(电话 = 0,NULL,电话) AS 电话,IF(客户回复再派量 = 0,NULL,客户回复再派量) AS 客户回复再派量,
                        concat(ROUND(IFNULL(物流再派签收 / 物流再派,0) * 100,2),'%') as 物流再派签收率,
                        concat(ROUND(IFNULL(物流3派签收 / 物流3派,0) * 100,2),'%') as 物流3派签收率,									
                        IF(问题件类型 = '送达客户不在/客户长期不在' AND 日期 >='{1}' AND 日期 <='{2}','暂未处理。物流已2派或3派',
                            IF(问题件类型 = '客户不接电话' and 未派 <> 0,CONCAT(未派,'单处理时已完结'),IF(单量 >= 短信,"获取物流轨迹信息后，后台会排队处理；若30-40分钟内订单状态变为已完结，则不发送短信。",
                            IF(单量 < 短信,"物流轨迹更新后， 根据派送问题类型的更改，会再次发送短信。", NULL))))  未派, 
                        异常, 上月总单量, 上月签收单量, 上月拒收单量, 
                        concat(ROUND(IFNULL(上月签收单量 / 上月总单量,0) * 100,2),'%') as 上月签收率, 上月派送问题件单量,上月周
            FROM ( SELECT s1.币种, s1.日期, s1.签收单量, s1.拒收单量, s1.总单量,  派送问题件单量, 问题件类型,
                            COUNT(订单编号) AS 单量, 发送量 短信, NULL AS 邮件, NULL AS 在线, 
                            SUM(IF(s3.备注 <> "" AND s3.回复类型 <> "已完结" AND s3.回复类型 <> "已发短信",1,0)) AS 电话, 
                            SUM(IF(s3.回复类型 = "回复",1,0)) AS 客户回复再派量, 物流再派, 物流再派签收, 物流3派, 物流3派签收, 
                            SUM(IF(s3.回复类型 = "已完结" OR s3.回复类型 = "已发短信",1,0)) AS 未派, 异常,
                            s2.签收单量 AS 上月签收单量, s2.拒收单量 AS 上月拒收单量, s2.总单量 AS 上月总单量, s5.上月派送问题件单量, s5.上月周
                    FROM (SELECT * 
                            FROM 派送问题件_跟进表2_cp p 
                            WHERE p.`物流名称` = '全部' AND p.日期 >= '{0}' 
                    ) s1
                    LEFT JOIN (SELECT * 
                                FROM 派送问题件_跟进表2_cp p 
                                WHERE p.`物流名称` = '全部'
                    ) s2 on s1.币种 = s2.币种 AND s1.日期 = DATE_SUB(s2.日期,INTERVAL -1 MONTH)
                    LEFT JOIN 
                    ( SELECT *, IF(派送问题 LIKE "地址问题" OR 派送问题 LIKE "客户要求更改派送时间或者地址","地址问题/客户要求更改派送时间或者地址",IF(派送问题 LIKE "送达客户不在" OR 派送问题 LIKE "客户长期不在","送达客户不在/客户长期不在",派送问题)) AS 问题件类型,
                                IF(备注 <> "", IF(备注 LIKE "已签收%" OR 备注 LIKE "已完结%" OR 备注 LIKE "已拒收%"  OR 备注 LIKE "显示拒收%" ,"已完结", 
                                    IF(备注 LIKE "无人接听%" OR 备注 LIKE "无效号码%","无人接听", IF(备注 LIKE "已通知%" OR 备注 LIKE "已告知%" OR 备注 LIKE "请告知%" OR 备注 LIKE "请通知%","已发短信", 
                                    IF(备注 LIKE "%*%","未回复",IF((备注 NOT LIKE "%*%" AND 备注 NOT LIKE "%拒收%") AND 
                                      (备注 LIKE "%客%取%" OR 备注 LIKE "%客%拿%" OR 备注 LIKE "%送货%" OR 备注 LIKE "%送貨%" OR 备注 LIKE "%取件%" OR 备注 LIKE "%取货%" OR 备注 LIKE "%取貨%"),"回复",""))))),备注) AS 回复类型
                        FROM ( SELECT ss1.订单编号,	ss1.币种,	ss1.下单时间,	ss1.订单状态,	ss1.物流状态,	ss1.物流渠道,	ss1.创建日期,	ss1.创建时间,	ss1.派送问题, ss1.派送次数, ss1.处理人,	ss1.处理记录,	ss1.处理时间, IF(ss1.备注 = "",ss2.处理结果,ss1.备注) AS 备注
                                FROM ( SELECT * 
                                        FROM 派送问题件_跟进表 p 
                                        WHERE p.创建日期 >= '{0}' 
                                ) ss1
                                LEFT JOIN 物流问题件 ss2 ON ss1.订单编号 = ss2.订单编号
                            ) p
                    ) s3 on s1.币种 = s3.币种 AND s1.日期 = s3.创建日期
					LEFT JOIN 
					( SELECT 币种, 创建日期, COUNT(订单编号) AS 派送问题件单量,
                            SUM(IF(派送次数 = 2,1,0)) AS 物流再派,
                            SUM(IF(物流状态 = "已签收" AND 派送次数 = 2,1,0)) AS 物流再派签收,
                            SUM(IF(派送次数 > 2,1,0)) AS 物流3派,
                            SUM(IF(物流状态 = "已签收" AND 派送次数 > 2,1,0)) AS 物流3派签收,
                            SUM(IF(回复类型 = "回复" AND 物流状态 = "拒收",1,0)) AS 异常
                        FROM ( SELECT *, IF(备注 <> "", IF(备注 LIKE "已签收%" OR 备注 LIKE "已完结%" OR 备注 LIKE "已拒收%"  OR 备注 LIKE "显示拒收%","已完结",
                                            IF(备注 LIKE "无人接听%" OR 备注 LIKE "无效号码%","无人接听", IF(备注 LIKE "已通知%" OR 备注 LIKE "已告知%" OR 备注 LIKE "请告知%" OR 备注 LIKE "请通知%","已发短信", 
                                            IF(备注 LIKE "%*%","未回复",IF((备注 NOT LIKE "%*%" AND 备注 NOT LIKE "%拒收%") AND 
                                              (备注 LIKE "%客%取%" OR 备注 LIKE "%客%拿%" OR 备注 LIKE "%送货%" OR 备注 LIKE "%送貨%" OR 备注 LIKE "%取件%" OR 备注 LIKE "%取货%" OR 备注 LIKE "%取貨%"),"回复",""))))),备注) AS 回复类型
                                 FROM ( SELECT ss1.订单编号,	ss1.币种,	ss1.下单时间,	ss1.订单状态,	ss1.物流状态,	ss1.物流渠道,	ss1.创建日期,	ss1.创建时间,	ss1.派送问题, ss1.派送次数, ss1.处理人,	ss1.处理记录,	ss1.处理时间, IF(ss1.备注 = "",ss2.处理结果,ss1.备注) AS 备注
                                        FROM ( SELECT * 
                                                FROM 派送问题件_跟进表 p 
                                                WHERE p.创建日期 >= '{0}' 
                                        ) ss1
                                        LEFT JOIN 物流问题件 ss2 ON ss1.订单编号 = ss2.订单编号
                                ) p
                            ) PP
                        GROUP BY 币种, 创建日期
					) s4 on s3.币种 =s4.币种 AND s3.创建日期 =s4.创建日期
                     LEFT JOIN ( SELECT 币种, 创建日期, CASE DATE_FORMAT(创建日期,'%w')	WHEN 1 THEN '星期一' WHEN 2 THEN '星期二' WHEN 3 THEN '星期三' WHEN 4 THEN '星期四' WHEN 5 THEN '星期五' WHEN 6 THEN '星期六' WHEN 0 THEN '星期日' END as 上月周, COUNT(订单编号) AS 上月派送问题件单量
                                        FROM 派送问题件_跟进表 p
                                        WHERE p.创建日期 >= DATE_SUB('{0} ',INTERVAL 1 MONTH)  AND p.创建日期 < '{0}'
                                        GROUP BY 币种, 创建日期
                     ) s5 on s3.币种 = s5.币种 AND s3.创建日期 = DATE_SUB(s5.创建日期,INTERVAL -1 MONTH)
                    LEFT JOIN 派送问题件_跟进表_message s6 on s3.币种 = s6.币种 AND s3.创建日期 = s6.日期 AND s3.问题件类型 =s6.短信模板
                    GROUP BY s1.币种, s1.日期, s3.问题件类型
            ) s
            ORDER BY s.币种, s.日期 , 
	            FIELD(s.问题件类型,'送至便利店','地址问题/客户要求更改派送时间或者地址','客户自取','客户不接电话','送达客户不在/客户长期不在','拒收','合计');'''.format(timeStart, logisticsN_begin, logisticsN_end)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        db2 = df[(df['币种'].str.contains('台币'))]
        # db22 = db2[(db2['月份'].str.contains("'" + month + "'"))]
        db3 = df[(df['币种'].str.contains('港币'))]
        # db33 = db3[(db3['月份'].str.contains("'" + month + "'"))]

        print('正在获取趋势图数据…………')
        db2.to_sql('cp', con=self.engine1, index=False, if_exists='replace')
        sql = '''SELECT DISTINCT 创建日期, 总单量, 上月总单量, 派送问题件单量, 上月派送问题件单量, 签收率, 上月签收率 FROM cp;'''.format(timeStart)
        df5 = pd.read_sql_query(sql=sql, con=self.engine1)
        df51 = df5[['创建日期', '总单量', '上月总单量']].copy()
        df52 = df5[['创建日期', '派送问题件单量', '上月派送问题件单量']].copy()
        df53 = df5[['创建日期', '签收率', '上月签收率']].copy()
        df54 = df5[['创建日期', '总单量', '派送问题件单量']].copy()
        df55 = df5[['创建日期', '上月总单量', '上月派送问题件单量']].copy()

        df51.rename(columns={'总单量': '当日', '上月总单量': '上月'}, inplace=True)
        df52.rename(columns={'派送问题件单量': '当日', '上月派送问题件单量': '上月'}, inplace=True)
        df53.rename(columns={'签收率': '当日', '上月签收率': '上月'}, inplace=True)
        df54.rename(columns={'总单量': '完成单量', '派送问题件单量': '派送问题件单量'}, inplace=True)
        df55.rename(columns={'上月总单量': '完成单量', '上月派送问题件单量': '派送问题件单量'}, inplace=True)

        print('正在写入excel…………')
        file_pathT = 'F:\\神龙签收率\\A订单改派跟进\\{0} 派送问题件跟进情况.xlsx'.format(rq)
        df0 = pd.DataFrame([])
        df0.to_excel(file_pathT, index=False)
        writer = pd.ExcelWriter(file_pathT, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_pathT)
        writer.book = book
        db2.drop(['币种', '总单量', '上月总单量'], axis=1).to_excel(excel_writer=writer, sheet_name='台湾', index=False)
        # db2.drop(['币种', '月份', '总单量', '上月总单量'], axis=1).to_excel(excel_writer=writer, sheet_name='台湾各月', index=False)
        db3.drop(['币种', '总单量', '上月总单量'], axis=1).to_excel(excel_writer=writer, sheet_name='香港', index=False)
        # db3.drop(['币种', '月份', '总单量', '上月总单量'], axis=1).to_excel(excel_writer=writer, sheet_name='香港各月', index=False)
        df1.to_excel(excel_writer=writer, sheet_name='明细', index=False)
        df11.to_excel(excel_writer=writer, sheet_name='拒收', index=False)
        df1211.to_excel(excel_writer=writer, sheet_name='台湾-物流', index=False)
        df1222.to_excel(excel_writer=writer, sheet_name='香港-物流', index=False)
        # df5[['创建日期', '总单量', '上月总单量', '派送问题件单量', '上月派送问题件单量', '签收率', '上月签收率']].to_excel(excel_writer=writer, sheet_name='趋势图', index=False)
        # df5[['创建日期', '上月总单量', '上月派送问题件单量']].to_excel(excel_writer=writer, sheet_name='趋势图', index=False, startcol=10)
        # df5[['创建日期', '总单量', '派送问题件单量']].to_excel(excel_writer=writer, sheet_name='趋势图', index=False, startcol=15)
        df51.to_excel(excel_writer=writer, sheet_name='同期完成订单', index=False)
        df52.to_excel(excel_writer=writer, sheet_name='同期派送问题件', index=False)
        df53.to_excel(excel_writer=writer, sheet_name='同期签收率', index=False)
        df54.to_excel(excel_writer=writer, sheet_name='当日', index=False)
        df55.to_excel(excel_writer=writer, sheet_name='上月', index=False)
        df8.to_excel(excel_writer=writer, sheet_name='台湾 问题类型 签收率', index=False)
        df81.to_excel(excel_writer=writer, sheet_name='香港 问题类型 签收率', index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表 cp
            del book['Sheet1']
        writer.save()
        writer.close()
        try:
            print('正在运行 派送问题件表 宏…………')
            # app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            # app.display_alerts = False
            # wbsht = app.books.open('D:/Users/Administrator/Desktop/slgat_签收计算(ver5.24).xlsm')
            # wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            # wbsht1 = app.books.open(file_pathT)
            # wbsht.macro('派送问题件_修饰')()
            # wbsht1.save()
            # wbsht1.close()
            # wbsht.close()
            # app.quit()
            # app.quit()

            # print('正在运行 派送问题件表 宏…………')
            # # 通过Win32的方式并不限制xls和xlsx（因为操作是wps在做）  https://wenku.baidu.com/view/3d298b06de36a32d7375a417866fb84ae45cc3ef.html
            # # excel =win32com.client.Dispatch('Excel.Application')  # word、excel、powerpoint对应的是微软的文字、表格和演示
            # excel = win32com.client.Dispatch('Ket.Application')  # wps、et、wpp对应的是金山文件、表格和演示
            # excel.Visible = False  # 可视化选项
            # Path = r"D:/Users/Administrator/Desktop/slgat_签收计算(ver5.24).xlsm"
            # workbook = excel.Workbooks.Open(Path)
            # workbook1 = excel.Workbooks.Open(file_pathT)
            # workbook.Application.Run("'D:/Users/Administrator/Desktop/slgat_签收计算(ver5.24).xlsm'!派送问题件_修饰")
            # workbook1.Save()
            # excel.Quit()

        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        print('----已写入excel')

    # 查询更新（新后台的获取-派送问题件更新）
    def getDeliveryList(self, timeStart, timeEnd, proxy_handle, proxy_id):  # 进入订单检索界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('+++正在查询信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.deliveryQuestion&action=getDeliveryList'
        url = r'https://gimp.giikin.com/service?service=gorder.deliveryQuestion&action=getDeliveryList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/deliveryProblemPackage'}
        data = {'order_number': None, 'waybill_number': None, 'question_level': None, 'question_type': None, 'order_trace_id': None, 'ship_phone': None, 'page': 1, 'pageSize': 90,
                'addtime': None, 'question_time': None, 'trace_time': None, 'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59', 'finishtime': None,
                'sale_id': None, 'product_id': None, 'logistics_id': None, 'area_id': None, 'currency_id': None, 'order_status': None, 'logistics_status': None}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)          # json类型数据转换为dict字典
        max_count = req['data']['count']    # 获取 请求订单量

        ordersDict = []
        if max_count != 0 and max_count != []:
            try:
                for result in req['data']['list']:                  # 添加新的字典键-值对，为下面的重新赋值
                    ordersDict.append(result.copy())
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            df = pd.json_normalize(ordersDict)
            print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
            print('*' * 50)
            if max_count > 90:
                in_count = math.ceil(max_count/90)
                dlist = []
                n = 1
                while n < in_count:  # 这里用到了一个while循环，穿越过来的
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                    data = self._getDeliveryList(timeStart, timeEnd, n,proxy_handle,proxy_id)
                    dlist.append(data)
                dp = df.append(dlist, ignore_index=True)
            else:
                dp = df
            dp = dp[['order_number',  'currency', 'ship_phone', 'addtime', 'create_time', 'finishtime', 'lastQuestionName', 'orderStatus', 'logisticsStatus',
                     'reassignmentTypeName', 'logisticsName',  'questionAddtime', 'userName', 'traceName', 'traceTime', 'content','failNum']]
            dp.columns = ['订单编号', '币种', '联系电话', '下单时间', '创建时间', '完成时间', '派送问题', '订单状态', '物流状态',
                          '订单类型', '物流渠道',  '派送问题首次时间', '处理人', '处理记录', '处理时间', '备注', '派送次数']
            print('正在写入......')
            dp.to_sql('cache_info', con=self.engine1, index=False, if_exists='replace')
            # dp.to_excel('G:\\输出文件\\派送问题件-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
            sql = '''REPLACE INTO 派送问题件_跟进表(订单编号,币种, 联系电话, 下单时间,完成时间,订单状态,物流状态,订单类型,物流渠道, 创建日期, 创建时间, 派送问题, 派送问题首次时间, 派送次数,处理人, 处理记录, 处理时间,备注, 记录时间) 
                    SELECT 订单编号,币种, 联系电话, 下单时间,完成时间,订单状态,物流状态,订单类型,物流渠道, DATE_FORMAT(创建时间,'%Y-%m-%d') 创建日期, 创建时间, 派送问题, 派送问题首次时间, 派送次数, 处理人, 处理记录, IF(处理时间 = '',NULL,处理时间) 处理时间,备注,NOW() 记录时间 
                    FROM cache_info;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
        else:
            print('没有需要获取的信息！！！')
            return
        print('*' * 50)
    def _getDeliveryList(self, timeStart, timeEnd, n,proxy_handle,proxy_id):  # 进入派送问题件界面
        print('+++正在查询第 ' + str(n) + ' 页信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.deliveryQuestion&action=getDeliveryList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/deliveryProblemPackage'}
        data = {'order_number': None, 'waybill_number': None, 'question_level': None, 'question_type': None, 'order_trace_id': None, 'ship_phone': None, 'page': n, 'pageSize': 90,
                'addtime': None, 'question_time': None, 'trace_time': None, 'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59', 'finishtime': None,
                'sale_id': None, 'product_id': None, 'logistics_id': None, 'area_id': None, 'currency_id': None, 'order_status': None, 'logistics_status': None}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersDict = []
        try:
            for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值
                ordersDict.append(result.copy())
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersDict)
        print('++++++第 ' + str(n) + ' 批次查询成功+++++++')
        print('*' * 50)
        return data

    # 查询更新（新后台的获取-订单完成 一）
    def getOrderList(self, timeStart, timeEnd,proxy_handle,proxy_id):  # 进入订单检索界面
        begin = datetime.datetime.strptime(timeStart, '%Y-%m-%d')
        begin = begin.date()
        end = datetime.datetime.strptime(timeEnd, '%Y-%m-%d')
        end = (end + datetime.timedelta(days=1)).date()
        print('正在查询日期---起止时间：' + timeStart + ' - ' + timeEnd)
        currencyId = [13, 6]            # 6 是港币；13 是台币
        logisticsStatus = [2, 3]
        match = {6: '港币', 13: '台币'}
        match2 = {2: '已签收', 3: '拒收'}
        dlist = []
        df =pd.DataFrame([])
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            day_time = str(day)
            for id in currencyId:
                print('+++正在查询： ' + day_time + match[id] + '完成 信息')
                dict = []
                res = {}
                count = self._getOrderList(id, None, day_time, day_time,proxy_handle,proxy_id)
                res['币种'] = match[id]
                res['日期'] = day_time
                res['总单量'] = count
                res['签收单量'] = ''
                res['拒收单量'] = ''
                dict.append(res)
                for log in logisticsStatus:
                        print('+++正在查询： ' + match[id] + match2[log] + ' 信息')
                        count2 = self._getOrderList(id, log,  day_time, day_time,proxy_handle,proxy_id)
                        if log == 2:
                            res['签收单量'] = count2
                        elif log == 3:
                            res['拒收单量'] = count2
                        # dict.append(res)
                data = pd.json_normalize(dict)
                print(data)
                dlist.append(data)
        dp = df.append(dlist, ignore_index=True)
        dp.to_sql('cache_info', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO 派送问题件_跟进表2(币种,日期,总单量,签收单量, 拒收单量) 
                SELECT 币种,日期,总单量,签收单量, 拒收单量
                FROM cache_info;'''
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
    def _getOrderList(self, id, log, timeStart, timeEnd,proxy_handle,proxy_id):  # 进入订单检索界面
        print('+++正在查询信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500, 'order_number': None, 'shippingNumber': None, 'orderNumberFuzzy': None, 'shipUsername': None,
                'phone': None, 'email': None, 'ip': None, 'productIds': None, 'saleIds': None, 'payType': None, 'logisticsId': None,
                'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': id, 'emailStatus': None,
                'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '', 'warehouse': None, 'isEmptyWayBillNumber': None,
                'logisticsStatus': log, 'orderStatus': None, 'tuan': None, 'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None,
                'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': None, 'autoVerifyStatus': None, 'shipZip': None,
                'remark': None, 'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None,
                'order': None, 'sortField': None, 'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None,
                'timeStart': None, 'timeEnd': None, 'finishTimeStart': timeStart + '00:00:00', 'finishTimeEnd': timeEnd + '23:59:59'}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        max_count = req['data']['count']  # 获取 请求订单量
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return max_count

    # 查询更新（新后台的获取-订单完成 二）
    def getOrderList_T(self, timeStart, timeEnd, proxy_handle, proxy_id):  # 进入订单检索界面
        begin = datetime.datetime.strptime(timeStart, '%Y-%m-%d')
        begin = begin.date()
        end = datetime.datetime.strptime(timeEnd, '%Y-%m-%d')
        end = (end + datetime.timedelta(days=1)).date()
        print('正在查询日期---起止时间：' + timeStart + ' - ' + timeEnd)
        currencyId = [13, 6]            # 13 是台币； 6 是港币
        logisticsStatus = [9999, 2, 3, 4]
        logisticsId_tw = [9999, '85,191,348,703,711,722', '198,199,229,356,376,380', '555,556,557,724,768,769,770', '367,383,255']
        logisticsId_hk = [9999, '230,277', '665,693']

        match = {6: '港币', 13: '台币'}
        match2 = {9999: '全部', 2: '已签收', 3: '拒收', 4: '已退货'}
        match3 = {9999: '全部', '85,191,348,703,711,722': '速派', '198,199,229,356,376,380': '天马', '555,556,557,724,768,769,770': '协来运', '367,383,255': '易速配', '230,277': '立邦', '665,693': '圆通'}

        dlist = []
        df =pd.DataFrame([])
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            day_time = str(day)
            for id in currencyId:
                logisticsId = ''
                if id == 13:
                    logisticsId = logisticsId_tw
                elif id == 6:
                    logisticsId = logisticsId_hk

                dict = []
                for log_Id in logisticsId:
                    print('+++正在查询： ' + day_time + match[id] + match3[log_Id] + '完成 信息')
                    res = {}
                    res['币种'] = match[id]
                    res['日期'] = day_time
                    res['物流名称'] = match3[log_Id]
                    res['总单量'] = ''
                    res['签收单量'] = ''
                    res['拒收单量'] = ''
                    res['退货单量'] = ''
                    for log in logisticsStatus:
                        print('+++正在查询： ' + match[id] + match3[log_Id] + match2[log] + ' 信息')
                        count2 = self._getOrderList_T(id, log_Id, log, day_time, day_time, proxy_handle, proxy_id)
                        if log == 9999:
                            res['总单量'] = count2
                        elif log == 2:
                            res['签收单量'] = count2
                        elif log == 3:
                            res['拒收单量'] = count2
                        elif log == 4:
                            res['退货单量'] = count2
                        dict.append(res)

                data = pd.json_normalize(dict)
                print(data)
                dlist.append(data)
        dp = df.append(dlist, ignore_index=True)
        dp.to_sql('cache_info', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO 派送问题件_跟进表2_cp(币种,日期, 物流名称, 总单量,签收单量, 拒收单量, 退货单量) 
                SELECT 币种,日期,物流名称, 总单量,签收单量, 拒收单量, 退货单量
                FROM cache_info;'''
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
    def _getOrderList_T(self, id, log_Id, log, timeStart, timeEnd, proxy_handle, proxy_id):  # 进入订单检索界面
        print('+++正在查询信息中')
        if log_Id == 9999:
            log_Id = None
        if log == 9999:
            log = None
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500, 'order_number': None, 'shippingNumber': None, 'orderNumberFuzzy': None, 'shipUsername': None,
                'phone': None, 'email': None, 'ip': None, 'productIds': None, 'saleIds': None, 'payType': None, 'logisticsId': log_Id,
                'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': id, 'emailStatus': None,
                'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '', 'warehouse': None, 'isEmptyWayBillNumber': None,
                'logisticsStatus': log, 'orderStatus': None, 'tuan': None, 'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None,
                'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': None, 'autoVerifyStatus': None, 'shipZip': None,
                'remark': None, 'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None,
                'order': None, 'sortField': None, 'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None,
                'timeStart': None, 'timeEnd': None, 'finishTimeStart': timeStart + '00:00:00', 'finishTimeEnd': timeEnd + '23:59:59'}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        max_count = req['data']['count']  # 获取 请求订单量
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return max_count

    # 查询更新（新后台的获取-短信发送）
    def getMessageLog(self, timeStart, timeEnd, proxy_handle, proxy_id):  # 进入订单检索界面
        begin = datetime.datetime.strptime(timeStart, '%Y-%m-%d')
        begin = begin.date()
        end = datetime.datetime.strptime(timeEnd, '%Y-%m-%d')
        end = (end + datetime.timedelta(days=1)).date()
        print('台币：正在查询日期---起止时间：' + timeStart + ' - ' + timeEnd)
        # currencyId = [13, 6]            # 6 是港币；13 是台币
        template_id = ['90,89,88', '49,73,77', '50,72,78']
        match = {'90,89,88': '客户自取', '49,73,77': '地址问题/客户要求更改派送时间或者地址', '50,72,78': '送至便利店'}
        # match2 = {2: '已签收', 3: '拒收'}
        dlist = []
        df = pd.DataFrame([])
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            day_time = str(day)
            dict = []
            for id in template_id:
                print('+++正在查询： ' + day_time + match[id] + ' 短信发送量')
                res = {}
                count = self._getMessageLog(id, day_time, day_time, proxy_handle, proxy_id)
                res['币种'] = '台币'
                res['日期'] = day_time
                res['短信模板'] = match[id]
                res['发送量'] = count
                dict.append(res)
            data = pd.json_normalize(dict)
            print(data)
            dlist.append(data)
        dp = df.append(dlist, ignore_index=True)
        dp.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO 派送问题件_跟进表_message(币种,日期,短信模板,发送量)  SELECT * FROM cache_cp;'''
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('港币：正在查询日期---起止时间：' + timeStart + ' - ' + timeEnd)
        template_id = ['84,85', '82,87', '151']
        match = {'84,85': '预约送达时间', '82,87': '送达客户不在', '151': '客户自取'}
        # match2 = {2: '已签收', 3: '拒收'}
        dlist = []
        df = pd.DataFrame([])
        for i in range((end - begin).days):  # 按天循环获取订单状态
            day = begin + datetime.timedelta(days=i)
            day_time = str(day)
            dict = []
            for id in template_id:
                print('+++正在查询： ' + day_time + match[id] + ' 短信发送量')
                res = {}
                count = self._getMessageLog(id, day_time, day_time, proxy_handle, proxy_id)
                res['币种'] = '港币'
                res['日期'] = day_time
                res['短信模板'] = match[id]
                res['发送量'] = count
                dict.append(res)
            data = pd.json_normalize(dict)
            print(data)
            dlist.append(data)
        dp = df.append(dlist, ignore_index=True)
        dp.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        sql = '''REPLACE INTO 派送问题件_跟进表_message(币种,日期,短信模板,发送量)  SELECT * FROM cache_cp;'''
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
    def _getMessageLog(self, id, timeStart, timeEnd, proxy_handle, proxy_id):  # 进入订单检索界面
        print('+++正在查询信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.sms&action=getMessageLog'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'order_number': None, 'waybill_number': None, 'to_phone': None, 'add_date': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59',
                'send_status': None, 'msgid': None, 'template_id': id, 'page': 1, 'pageSize': 10}
        # print(data)
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        max_count = req['data']['count']  # 获取 请求订单量
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return max_count

    def outport_List(self, timeStart, timeEnd):
        rq = datetime.datetime.now().strftime('%m.%d')

        print('正在获取拒收内容…………')
        sql = '''SELECT 创建日期, IFNULL(具体原因,'合计') AS 拒收原因, 单量,concat(ROUND(IFNULL(单量 / 总单量,0) * 100,2),'%') as 占比
                FROM ( SELECT 创建日期, 具体原因,COUNT(s1.订单编号) AS 单量
                     FROM( SELECT 订单编号,币种, IF(完结状态时间 IS NULL,状态时间,完结状态时间) AS 完结时间,  DATE_FORMAT(IF(完结状态时间 IS NULL,状态时间,完结状态时间), '%Y-%m-%d') AS 创建日期
                            FROM d1_gat d
                            WHERE d.最终状态 = '拒收'
                     ) s1
                    LEFT JOIN 拒收问题件 js ON s1.订单编号 =js.订单编号
                    WHERE js.拒收问题件 <> '未联系上客户' AND js.具体原因 IS not NULL
                    GROUP BY 创建日期, 具体原因
                    WITH ROLLUP
                ) s
                LEFT JOIN 
                ( SELECT 创建日期 日期, 具体原因 具体,COUNT(s1.订单编号) AS 总单量
                    FROM(  SELECT 订单编号,币种, IF(完结状态时间 IS NULL,状态时间,完结状态时间) AS 完结时间,  DATE_FORMAT(IF(完结状态时间 IS NULL,状态时间,完结状态时间), '%Y-%m-%d') AS 创建日期
                            FROM d1_gat d
                            WHERE d.最终状态 = '拒收'
                     ) s1
                     LEFT JOIN 拒收问题件 js ON s1.订单编号 =js.订单编号
                     WHERE js.拒收问题件 <> '未联系上客户' AND js.具体原因 IS not NULL
                     GROUP BY 创建日期
                ) ss ON s.创建日期 =ss.日期
                WHERE 创建日期 IS NOT NULL
                ORDER BY 创建日期, FIELD(拒收原因,'合计') DESC, 单量 DESC;'''.format(timeStart)
        df11 = pd.read_sql_query(sql=sql, con=self.engine1)
        print('正在写入excel…………')
        file_pathT = 'F:\\神龙签收率\\A订单改派跟进\\{0} 派送问题件跟进情况.xlsx'.format(rq)
        df0 = pd.DataFrame([])
        df0.to_excel(file_pathT, index=False)
        writer = pd.ExcelWriter(file_pathT, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_pathT)
        writer.book = book

        df11.to_excel(excel_writer=writer, sheet_name='拒收', index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表 cp
            del book['Sheet1']
        writer.save()
        writer.close()
        try:
            print('正在运行 派送问题件表 宏…………')
            # # 通过Win32的方式并不限制xls和xlsx（因为操作是wps在做）  https://wenku.baidu.com/view/3d298b06de36a32d7375a417866fb84ae45cc3ef.html
            # # excel =win32com.client.Dispatch('Excel.Application')  # word、excel、powerpoint对应的是微软的文字、表格和演示
            # excel = win32com.client.Dispatch('Ket.Application')  # wps、et、wpp对应的是金山文件、表格和演示
            # excel.Visible = False  # 可视化选项
            # Path = r"D:/Users/Administrator/Desktop/slgat_签收计算(ver5.24).xlsm"
            # workbook = excel.Workbooks.Open(Path)
            # workbook1 = excel.Workbooks.Open(file_pathT)
            # workbook.Application.Run("'D:/Users/Administrator/Desktop/slgat_签收计算(ver5.24).xlsm'!派送问题件_修饰")
            # workbook1.Save()
            # excel.Quit()

        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        print('----已写入excel')


    # 查询更新（新后台的获取-拒收问题件更新）
    def order_js_QueryUpdata(self, timeStart, timeEnd, proxy_handle, proxy_id):  # 进入拒收问题件界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('+++正在查询信息中---拒收问题件更新')
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getRejectList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerRejection'}
        data = {'page': 1, 'pageSize': 500, 'orderPrefix': None, 'shipUsername': None, 'shippingNumber': None, 'email': None, 'saleIds': None, 'ip': None,
                'productIds': None, 'phone': None, 'optimizer': None, 'payment': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None,
                'emailStatus': None, 'befrom': None, 'areaId': None, 'orderStatus': None, 'timeStart': None, 'timeEnd': None, 'payType': None, 'questionId': None,
                'autoVerifys': None, 'reassignmentType': None, 'logisticsStatus': None, 'logisticsId': None, 'traceItemIds': -1,
                'finishTimeStart': timeStart + ' 00:00:00', 'finishTimeEnd': timeEnd + ' 23:59:59', 'traceTimeStart': None, 'traceTimeEnd': None,'newCloneNumber': None}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        max_count = req['data']['count']
        print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        # print(req)
        ordersDict = []
        if max_count != 0:
            try:
                for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值用
                    # print(result['orderNumber'])
                    result['订单编号'] = result['orderNumber']
                    result['再次克隆下单'] = result['newCloneNumber']
                    result['跟进人'] = ''
                    result['时间'] = ''
                    result['内容'] = ''
                    result['联系方式'] = ''
                    result['问题类型'] = ''
                    result['问题原因'] = ''
                    result['处理结果'] = ''
                    result['是否需要商品'] = ''
                    if result['traceItems'] != []:
                        for res in result['traceItems']:
                            resval = res.split(':')[0]
                            if '跟进人' in resval:
                                result['跟进人'] = (res.split('跟进人:')[1]).strip()  # 跟进人
                            if '时间' in resval and '跟进' not in resval:
                                result['时间'] = (res.split('时间:')[1]).strip()  # 跟进人
                            if '内容' in resval:
                                result['内容'] = (res.split('内容:')[1]).strip()  # 跟进人
                            if '联系方式' in resval:
                                result['联系方式'] = (res.split('联系方式:')[1]).strip()  # 跟进人
                            if '问题类型' in resval:
                                result['问题类型'] = (res.split('问题类型:')[1]).strip()  # 跟进人
                            if '问题原因' in resval:
                                result['问题原因'] = (res.split('问题原因:')[1]).strip()  # 跟进人
                            if '处理结果' in res:
                                result['处理结果'] = (res.split('处理结果:')[1]).strip()  # 跟进人
                            if '是否需要商品' in res:
                                result['是否需要商品'] = (res.split('是否需要商品:')[1]).strip()  # 跟进人
                    ordersDict.append(result.copy())
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            df = pd.json_normalize(ordersDict)
            print('*' * 50)
            if max_count > 500:
                in_count = math.ceil(max_count/500)
                dlist = []
                n = 1
                while n < in_count:  # 这里用到了一个while循环，穿越过来的
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                    data = self._order_js_QueryUpdata(timeStart, timeEnd, n,proxy_handle,proxy_id)
                    dlist.append(data)
                dp = df.append(dlist, ignore_index=True)
            else:
                dp = df
            dp.to_excel('G:\\输出文件\\拒收问题件-更新2{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
            dp = dp[['订单编号', 'currency', 'percentInfo.orderCount', 'percentInfo.rejectCount', 'percentInfo.arriveCount', 'addTime', 'finishTime', 'tel_phone', 'shipInfo.shipPhone', 'ip','newCloneUser', 'newCloneStatus', 'newCloneLogisticsStatus', '再次克隆下单', '跟进人', '时间', '联系方式', '问题类型', '问题原因', '内容', '处理结果', '是否需要商品']]
            dp.columns = ['订单编号', '币种', '订单总量', '拒收量', '签收量','下单时间', '完成时间', '电话', '联系电话', 'ip','新单克隆人', '新单订单状态', '新单物流状态', '再次克隆下单', '处理人', '处理时间', '联系方式', '核实原因', '具体原因', '备注', '处理结果', '是否需要商品']
            print('正在写入......')
            dp.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 拒收问题件_check_iphone(订单编号,币种,订单总量, 拒收量, 签收量, 下单时间, 完成时间, 电话, 联系电话, ip, 新单克隆人, 新单订单状态, 新单物流状态, 再次克隆下单,处理人,处理时间,联系方式, 核实原因, 具体原因, 备注, 处理结果, 是否需要商品,记录时间)
                    SELECT 订单编号,币种, 订单总量, 拒收量, 签收量, 下单时间, 完成时间, IF(电话 LIKE "852%",RIGHT(电话,LENGTH(电话)-3),IF(电话 LIKE "886%",RIGHT(电话,LENGTH(电话)-3),电话)) 电话, 联系电话, ip,新单克隆人, 新单订单状态, 新单物流状态,  IF(再次克隆下单 = '',NULL,再次克隆下单) 再次克隆下单,处理人,处理时间,联系方式, 核实原因, 具体原因, 备注, 处理结果,是否需要商品, NOW() 记录时间
                    FROM customer;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
        else:
            print('****** 没有信息！！！')
        print('*' * 50)
    def _order_js_QueryUpdata(self, timeStart, timeEnd, n, proxy_handle, proxy_id):  # 进入拒收问题件界面
        print('+++正在查询第 ' + str(n) + ' 页信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getRejectList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerRejection'}
        data = {'page': n, 'pageSize': 500, 'orderPrefix': None, 'shipUsername': None, 'shippingNumber': None, 'email': None, 'saleIds': None, 'ip': None,
                'productIds': None, 'phone': None, 'optimizer': None, 'payment': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None,
                'emailStatus': None, 'befrom': None, 'areaId': None, 'orderStatus': None, 'timeStart': None, 'timeEnd': None, 'payType': None, 'questionId': None,
                'autoVerifys': None, 'reassignmentType': None, 'logisticsStatus': None, 'logisticsId': None, 'traceItemIds': -1,
                'finishTimeStart': timeStart + ' 00:00:00', 'finishTimeEnd': timeEnd + ' 23:59:59', 'traceTimeStart': None, 'traceTimeEnd': None,'newCloneNumber': None}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersDict = []
        try:
            for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值用
                # print(result['orderNumber'])
                result['订单编号'] = result['orderNumber']
                result['再次克隆下单'] = result['newCloneNumber']
                result['跟进人'] = ''
                result['时间'] = ''
                result['内容'] = ''
                result['联系方式'] = ''
                result['问题类型'] = ''
                result['问题原因'] = ''
                result['处理结果'] = ''
                result['是否需要商品'] = ''
                if result['traceItems'] != []:
                    for res in result['traceItems']:
                        resval = res.split(':')[0]
                        if '跟进人' in resval:
                            result['跟进人'] = (res.split('跟进人:')[1]).strip()  # 跟进人
                        if '时间' in resval and '跟进' not in resval:
                            result['时间'] = (res.split('时间:')[1]).strip()  # 跟进人
                        if '内容' in resval:
                            result['内容'] = (res.split('内容:')[1]).strip()  # 跟进人
                        if '联系方式' in resval:
                            result['联系方式'] = (res.split('联系方式:')[1]).strip()  # 跟进人
                        if '问题类型' in resval:
                            result['问题类型'] = (res.split('问题类型:')[1]).strip()  # 跟进人
                        if '问题原因' in resval:
                            result['问题原因'] = (res.split('问题原因:')[1]).strip()  # 跟进人
                        if '处理结果' in res:
                            result['处理结果'] = (res.split('处理结果:')[1]).strip()  # 跟进人
                        if '是否需要商品' in res:
                            result['是否需要商品'] = (res.split('是否需要商品:')[1]).strip()  # 跟进人
                ordersDict.append(result.copy())
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersDict)
        print('++++++第 ' + str(n) + ' 批次查询成功+++++++')
        print('*' * 50)
        return data

    # 查询更新（新后台的获取-物流问题件更新）
    def waybill_InfoQueryUpdata(self, timeStart, timeEnd,proxy_handle,proxy_id):  # 进入物流问题件界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('+++正在查询信息中---物流问题件更新')
        url = r'https://gimp.giikin.com/service?service=gorder.customerQuestion&action=getCustomerComplaintList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerQuestion'}
        data = {'order_number': None, 'waybill_no': None, 'transfer_no': None, 'gift_reissue_order_number': None, 'is_gift_reissue': None, 'order_trace_id': None,
                'question_type': None, 'critical': None, 'read_status': None, 'operator_type': None, 'operator': None, 'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59',
                'trace_time': None, 'is_collection': None, 'logistics_status': None, 'user_id': None,
                'page': 1, 'pageSize': 90}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        max_count = req['data']['count']
        ordersDict = []
        try:
            for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值
                # print(result)
                # print(11)
                # print(result['order_number'])
                result['dealContent'] = zhconv.convert(result['dealContent'], 'zh-hans')
                if 'traceRecord' in result:
                    result['traceRecord'] = zhconv.convert(result['traceRecord'], 'zh-hans')
                    if ';' in result['traceRecord']:
                        trace_record = result['traceRecord'].split(";")
                        for record in trace_record:
                            if record.split("#处理结果：")[1] != '':
                                result['deal_time'] = record.split()[0]
                                result['result_reson'] = ''
                                result['result_info'] = ''

                                rec = record.split("#处理结果：")[1]
                                if len(rec.split()) > 2:
                                    result['result_info'] = rec.split()[2]        # 客诉原因
                                if len(rec.split()) > 1:
                                    result['result_reson'] = rec.split()[1]     # 处理内容
                                result['dealContent'] = rec.split()[0]            # 最新处理结果
                                rec_name = record.split("#处理结果：")[0]
                                if len(rec_name.split()) > 1:
                                    if (rec_name.split())[2] != '' and (rec_name.split())[2] != []:
                                        if '客服' in (rec_name.split())[2]:
                                            result['traceUserName'] = ((rec_name.split())[2]).split("(客服)")[0]
                                        else:
                                            result['traceUserName'] = (rec_name.split())[2]
                                else:
                                    result['traceUserName'] = ''
                                ordersDict.append(result.copy())
                    else:
                        result['deal_time'] = ''
                        result['result_reson'] = ''
                        result['result_info'] = ''
                        if '拒收' in result['dealContent']:
                            if len(result['dealContent'].split()) > 2:
                                result['result_info'] = result['dealContent'].split()[2]
                            if len(result['dealContent'].split()) > 1:
                                result['result_reson'] = result['dealContent'].split()[1]
                            result['dealContent'] = result['dealContent'].split()[0]
                        if result['traceRecord'] != '' or result['traceRecord'] != []:
                            result['deal_time'] = result['traceRecord'].split()[0]
                        if result['traceUserName'] != '' or result['traceUserName'] != []:
                            result['traceUserName'] = result['traceUserName'].replace('客服：', '')
                        result['dealContent'] = result['dealContent'].strip()
                        ordersDict.append(result.copy())
                else:
                    result['deal_time'] = result['update_time']
                    result['result_reson'] = ''
                    result['result_info'] = ''
                    result['dealContent'] = ''
                    result['traceUserName'] = ''
                    ordersDict.append(result.copy())
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        df = pd.json_normalize(ordersDict)
        print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        print('*' * 50)
        if max_count != 0:
            if max_count > 90:
                in_count = math.ceil(max_count/90)
                dlist = []
                n = 1
                while n < in_count:  # 这里用到了一个while循环，穿越过来的
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                    data = self._waybillInfoQueryUpdata(timeStart, timeEnd, n,proxy_handle,proxy_id)
                    dlist.append(data)
                dp = df.append(dlist, ignore_index=True)
            else:
                dp = df
            dp.to_excel('G:\\输出文件\\物流问题件-更新2{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
            dp = dp[['order_number',  'currency', 'ship_phone', 'amount', 'customer_name', 'customer_mobile', 'arrived_address', 'arrived_time', 'create_time', 'dealStatus', 'dealContent',
                     'deal_time', 'dealTime', 'result_reson', 'result_info', 'questionType', 'questionTypeName', 'question_desc', 'traceRecord', 'traceUserName', 'giftStatus', 'operatorName','contact',
                     'gift_reissue_order_number',  'addtime','update_time']]
            dp.columns = ['订单编号', '币种', '联系电话', '订单金额', '客户姓名', '客户电话', '客户地址', '送达时间', '导入时间', '最新处理状态', '最新处理结果',
                          '处理时间', '处理日期时间', '拒收原因', '具体原因', '问题类型状态', '问题类型', '问题描述', '历史处理记录', '处理人', '赠品补发订单状态', '导入人', '联系方式',
                          '赠品补发订单编号', '下单时间', '更新时间']
            # dp = dp[(dp['处理人'].str.contains('蔡利英|杨嘉仪|蔡贵敏|刘慧霞|张陈平', na=False))]
            print('正在写入......')
            dp.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            # dp.to_excel('G:\\输出文件\\物流问题件-更新2{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
            sql = '''REPLACE INTO 物流问题件_check_iphone(订单编号, 下单时间, 联系电话, 币种, 问题类型, 问题描述, 物流反馈时间, 导入人,处理时间, 处理日期时间, 处理人, 联系方式,  处理结果,拒收原因, 克隆订单编号, 记录时间) 
                    SELECT 订单编号, 下单时间, 联系电话, 币种, 问题类型, NULL AS 问题描述, 导入时间 AS 物流反馈时间, 导入人,IF(处理时间 = '',NULL,处理时间) AS 处理时间, IF(处理日期时间 = '',NULL,处理日期时间) AS 处理日期时间, 处理人, 联系方式, IF(最新处理结果 = '',问题类型状态,最新处理结果) AS 处理结果,拒收原因, 赠品补发订单编号 AS 克隆订单编号, NOW() 记录时间 
                    FROM customer;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
        else:
            print('没有需要获取的信息！！！')
            return
        print('*' * 50)
    def _waybillInfoQueryUpdata(self, timeStart, timeEnd, n,proxy_handle,proxy_id):  # 进入物流问题件界面
        print('+++正在查询第 ' + str(n) + ' 页信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customerQuestion&action=getCustomerComplaintList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/customerQuestion'}
        data = {'order_number': None, 'waybill_no': None, 'transfer_no': None, 'gift_reissue_order_number': None, 'is_gift_reissue': None, 'order_trace_id': None,
                'question_type': None, 'critical': None, 'read_status': None, 'operator_type': None, 'operator': None, 'create_time': timeStart + ' 00:00:00,' + timeEnd + ' 23:59:59',
                'trace_time': None, 'is_collection': None, 'logistics_status': None, 'user_id': None,
                'page': n, 'pageSize': 90}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersDict = []
        try:
            for result in req['data']['list']:  # 添加新的字典键-值对，为下面的重新赋值
                # print(55)
                # print(result['order_number'])
                result['dealContent'] = zhconv.convert(result['dealContent'], 'zh-hans')
                if 'traceRecord' in result:
                    result['traceRecord'] = zhconv.convert(result['traceRecord'], 'zh-hans')
                    if ';' in result['traceRecord']:
                        trace_record = result['traceRecord'].split(";")
                        for record in trace_record:
                            if record.split("#处理结果：")[1] != '':
                                result['deal_time'] = record.split()[0]
                                result['result_reson'] = ''
                                result['result_info'] = ''

                                rec = record.split("#处理结果：")[1]
                                if len(rec.split()) > 2:
                                    result['result_info'] = rec.split()[2]        # 客诉原因
                                if len(rec.split()) > 1:
                                    result['result_reson'] = rec.split()[1]     # 处理内容
                                result['dealContent'] = rec.split()[0]            # 最新处理结果
                                rec_name = record.split("#处理结果：")[0]
                                if len(rec_name.split()) > 1:
                                    if (rec_name.split())[2] != '' and (rec_name.split())[2] != []:
                                        if '客服' in (rec_name.split())[2]:
                                            result['traceUserName'] = ((rec_name.split())[2]).split("(客服)")[0]
                                        else:
                                            result['traceUserName'] = (rec_name.split())[2]
                                else:
                                    result['traceUserName'] = ''
                                ordersDict.append(result.copy())
                    else:
                        result['deal_time'] = ''
                        result['result_reson'] = ''
                        result['result_info'] = ''
                        if '拒收' in result['dealContent']:
                            if len(result['dealContent'].split()) > 2:
                                result['result_info'] = result['dealContent'].split()[2]
                            if len(result['dealContent'].split()) > 1:
                                result['result_reson'] = result['dealContent'].split()[1]
                            result['dealContent'] = result['dealContent'].split()[0]
                        if result['traceRecord'] != '' or result['traceRecord'] != []:
                            result['deal_time'] = result['traceRecord'].split()[0]
                        if result['traceUserName'] != '' or result['traceUserName'] != []:
                            result['traceUserName'] = result['traceUserName'].replace('客服：', '')
                        result['dealContent'] = result['dealContent'].strip()
                        ordersDict.append(result.copy())
                else:
                    result['deal_time'] = result['update_time']
                    result['result_reson'] = ''
                    result['result_info'] = ''
                    result['dealContent'] = ''
                    result['traceUserName'] = ''
                    ordersDict.append(result.copy())
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersDict)
        print('++++++第 ' + str(n) + ' 批次查询成功+++++++')
        print('*' * 50)
        return data

    def Check_Iphone_Updata(self, timeStart, timeEnd):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在输出 拒收&派送|物流问题件-电话检测 --- 数据中,起止：' + timeStart + ':' + timeEnd)
        sql2 = '''SELECT sss2.币种, sss2.标准电话, sss2.拒收原因, sss2.近两月拒收量, 
                         sss3.近两月订单量, sss3.近两月签收量, sss3.近两月退货量,  
                         concat(ROUND(IFNULL(sss3.近两月签收量 / sss3.近两月订单量,0) * 100,2),'%') as 近两月签收率,
                         concat(ROUND(IFNULL(sss3.近两月拒收量 / sss3.近两月订单量,0) * 100,2),'%') as 近两月拒收率,
                         concat(ROUND(IFNULL(sss3.近两月退货量 / sss3.近两月订单量,0) * 100,2),'%') as 近两月退货率,
                         sss1.总单量, sss1.总签收量, sss1.总拒收量,
                         concat(ROUND(IFNULL(sss1.总签收量 / sss1.总单量,0) * 100,2),'%') as 总签收率, sss1.下单拒收率
                FROM (		
                        SELECT s1.币种, s1.标准电话, 具体原因, 总单量, 总签收量, 总拒收量, s1.下单拒收率, COUNT(s1.订单编号) AS 次数
                        FROM (SELECT 订单编号, 币种, 标准电话,下单时间, 订单配送总量 AS 总单量, 下单拒收率, 签收量 AS 总签收量, 拒收量 AS 总拒收量
                                FROM (SELECT *
                                            FROM gat_order_list g
                                            WHERE g.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 6 MONTH),'%Y%m') AND g.系统物流状态 = '拒收'
                                ) ss1
                                WHERE ss1.`完结状态时间` >= '{0} 00:00:00' AND ss1.`完结状态时间` <= '{1} 23:59:59'
                        ) s1
                        LEFT JOIN 
                        (  SELECT * 
                            FROM 拒收问题件_check_iphone js
                        ) s2 ON s1.订单编号 = s2.订单编号
                        GROUP BY s1.币种, s1.标准电话
                        HAVING 次数 >=5
                        ORDER BY 币种, 次数 DESC
                ) sss1
                LEFT JOIN 
                ( 	SELECT IFNULL(币种,'合计') 币种, IFNULL(标准电话,'合计') 标准电话, IFNULL(拒收原因,'合计') 拒收原因, COUNT(订单编号) AS 近两月拒收量
                        FROM (
                                SELECT s1.订单编号, s1.币种, s1.标准电话, s1.下单时间, IF(具体原因 IS NULL,'-',具体原因) AS 拒收原因
                                FROM (SELECT 订单编号, 币种, 标准电话,下单时间
                                            FROM (SELECT *
                                                        FROM gat_order_list g
                                                        WHERE g.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 6 MONTH),'%Y%m') AND g.系统物流状态 = '拒收'
                                            ) ss1
                                            WHERE ss1.`完结状态时间` >= '{0} 00:00:00' AND ss1.`完结状态时间` <= '{1} 23:59:59'
                                ) s1
                                LEFT JOIN 
                                ( SELECT * 
                                    FROM 拒收问题件_check_iphone js
                                ) s2 ON s1.订单编号 = s2.订单编号
                    ) ss1
                    GROUP BY 币种, 标准电话, 拒收原因
                    WITH ROLLUP
                ) sss2 ON sss1.标准电话 = sss2.标准电话
                LEFT JOIN (SELECT 标准电话,
                                    COUNT(订单编号)  as 近两月订单量,
                                    SUM(IF(`系统物流状态` = '已签收',1,0)) AS 近两月签收量,
                                    SUM(IF(`系统物流状态` = '已退货',1,0)) AS 近两月退货量,
                                    SUM(IF(`系统物流状态` = '拒收',1,0)) AS 近两月拒收量
                            FROM ( SELECT *
                                    FROM gat_order_list g
                                    WHERE g.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 6 MONTH),'%Y%m') AND g.系统物流状态 = '拒收'
                            ) ss1
                            WHERE ss1.`完结状态时间` >= '{0} 00:00:00' AND ss1.`完结状态时间` <= '{1} 23:59:59'
                            GROUP BY 标准电话
                ) sss3 ON sss1.标准电话 = sss3.标准电话
                ORDER BY sss1.币种, sss1.次数 DESC, sss1.标准电话, sss2.近两月拒收量 DESC;'''.format(timeStart, timeEnd)
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)

        sql3 = '''SELECT s2.*,
                         s1.近两月订单量, s1.近两月签收量, s1.近两月拒收量, s1.近两月退货量, 
                         concat(ROUND(IFNULL(s1.近两月签收量 / s1.近两月订单量,0) * 100,2),'%') as 近两月签收率,
                         concat(ROUND(IFNULL(s1.近两月拒收量 / s1.近两月订单量,0) * 100,2),'%') as 近两月拒收率,
                         concat(ROUND(IFNULL(s1.近两月退货量 / s1.近两月订单量,0) * 100,2),'%') as 近两月退货率,
                         s1.总单量, s1.总签收量, s1.总拒收量, 
                         concat(ROUND(IFNULL(s1.总签收量 / s1.总单量,0) * 100,2),'%') as 总签收率,s1.下单拒收率
                FROM (   SELECT ss2.币种, ss2.标准电话,  
                                 COUNT(ss2.订单编号) AS 近两月订单量,
                                 SUM(IF(ss2.`系统物流状态` = '已签收',1,0)) AS 近两月签收量,
                                 SUM(IF(ss2.`系统物流状态` = '已退货',1,0)) AS 近两月退货量,
                                 SUM(IF(ss2.`系统物流状态` = '拒收',1,0)) AS 近两月拒收量, 
                                 ss2.订单配送总量 AS 总单量, ss2.签收量 AS 总签收量,  ss2.拒收量 AS 总拒收量, ss2.下单拒收率
                        FROM (
                                ( SELECT 订单编号, 派送问题
                                    FROM 派送问题件_跟进表 ps
                                    WHERE ps.创建日期 >= '{0}' AND ps.创建日期 <= '{1}'
                                ) 
                                union 
                                ( SELECT 订单编号, IF(问题描述 IS NULL ,'-',问题描述) 问题描述
                                    FROM 物流问题件_check_iphone wl
                                    WHERE wl.物流反馈时间 >= '{0}' AND wl.物流反馈时间 <= '{1}' AND wl.问题类型 <= '派送问题件' 
                                      AND wl.订单编号 NOT IN (
                                                                SELECT 订单编号
                                                                FROM 派送问题件_跟进表 ps
                                                                WHERE ps.创建日期 >= '{0}' AND ps.创建日期 <= '{1}'
                                                            )
                                ) 
                        ) ss1
                        LEFT JOIN
                        (	SELECT *
                            FROM gat_order_list g
                            WHERE g.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 6 MONTH),'%Y%m')
                        ) ss2 ON ss1.订单编号 = ss2.订单编号
                        GROUP BY ss2.币种, ss2.标准电话
                        HAVING 近两月订单量 >= 5
                ) s1
                LEFT JOIN 
                (
                        SELECT IFNULL(ss2.币种,'合计') AS 币种, IFNULL(ss2.标准电话,'合计') AS 标准电话, IFNULL(派送问题,'合计') AS 派送问题, COUNT(ss1.订单编号) AS 出现次数
                        FROM (
                                ( SELECT 订单编号, 派送问题
                                    FROM 派送问题件_跟进表 ps
                                    WHERE ps.创建日期 >= '{0}' AND ps.创建日期 <= '{1}'
                                ) 
                                union 
                                ( SELECT 订单编号, IF(问题描述 IS NULL ,'-',问题描述) 问题描述
                                    FROM 物流问题件_check_iphone wl
                                    WHERE wl.物流反馈时间 >= '{0}' AND wl.物流反馈时间 <= '{1}' AND wl.问题类型 <= '派送问题件' 
                                      AND wl.订单编号 NOT IN (
                                                                SELECT 订单编号
                                                                FROM 派送问题件_跟进表 ps
                                                                WHERE ps.创建日期 >= '{0}' AND ps.创建日期 <= '{1}'
                                                            )
                                )
                        ) ss1
                        LEFT JOIN
                        (	SELECT *
                            FROM gat_order_list g
                            WHERE g.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 6 MONTH),'%Y%m')
                        ) ss2 ON ss1.订单编号 = ss2.订单编号
                        GROUP BY ss2.币种, ss2.标准电话, 派送问题
                        WITH ROLLUP
                        ORDER BY ss2.币种, ss2.标准电话, 出现次数 DESC
                ) s2 ON s1.标准电话 = s2.标准电话
                ORDER BY s1.币种, s1.近两月订单量 DESC, s1.标准电话, s2.出现次数 DESC;'''.format(timeStart, timeEnd)
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)

        file_path = 'G:\\输出文件\\拒收&派送|物流问题件-电话检测 {}.xlsx'.format(rq)
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        df2.to_excel(excel_writer=writer, sheet_name='拒收-电话', index=False)
        df3.to_excel(excel_writer=writer, sheet_name='派送&物流-电话', index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('输出成功')

if __name__ == '__main__':
    start: datetime = datetime.datetime.now()
    '''
    # -----------------------------------------------自动获取 问题件 状态运行（一）-----------------------------------------
    # 1、 物流问题件；2、物流客诉件；3、物流问题件；4、全部；--->>数据更新切换
    '''
    select = 99
    if int(select) == 99:
        handle = '手动'
        login_TmpCode = '2bb2a3ecea7932ddb588d29b39c8d8ba'          # 输入登录口令Tkoen
        proxy_handle = '代理服务器'
        proxy_id = '192.168.13.89:37466'                            # 输入代理服务器节点和端口

        m = QueryTwo('+86-18538110674', 'qyz04163510.', login_TmpCode, handle, select, proxy_handle, proxy_id)
        start: datetime = datetime.datetime.now()

        if int(select) == 11:
            timeStart, timeEnd = m.readInfo('物流问题件')

        elif int(select) == 99:         # 查询更新-派送问题件
            timeStart, timeEnd = m.readInfo('派送问题件_订单完成单量&短信发送单量')
            # m.getOrderList_T('2022-07-25', '2022-07-26')
            m.getOrderList_T(timeStart, timeEnd, proxy_handle, proxy_id)                      # 订单完成单量 更新

            # m.getMessageLog('2022-07-25', '2022-07-26')
            m.getMessageLog(timeStart, timeEnd, proxy_handle, proxy_id)                       # 短信发送单量 更新

            timeStart, timeEnd = m.readInfo('派送问题件_更新')
            # m.getDeliveryList('2022-06-12', '2022-06-30')
            # m.getDeliveryList('2022-07-10', '2022-07-26')
            m.getDeliveryList(timeStart, timeEnd, proxy_handle, proxy_id)                     # 派送问题件 更新

            # timeStart, timeEnd = m.readInfo('派送问题件_导出')
            logisticsN_begin = '2022-07-11'                         # 送达客户不在/客户长期不在  物流轨迹查询时间
            logisticsN_end = '2022-07-31'

            # timeStart = '2022-09-01'
            # timeEnd = '2022-10-25'
            if (datetime.datetime.now()).strftime('%d') == 1:
                timeStart = (datetime.datetime.now() - relativedelta(months=2)).strftime('%Y-%m') + '-01'
                timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            else:
                timeStart = (datetime.datetime.now() - relativedelta(months=1)).strftime('%Y-%m') + '-01'
                timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')

            print('派送问题件，导出时间》》》 ' + timeStart + "---" + timeEnd)
            m.outport_getDeliveryList('2022-10-01', timeEnd, logisticsN_begin, logisticsN_end)
            # m.outport_getDeliveryList(timeStart, timeEnd)             # 派送问题件跟进表 导出

        elif int(select) == 909:         # 查询更新-派送问题件
            # timeStart, timeEnd = m.readInfo('派送问题件_更新')
            m.getDeliveryList('2022-04-01', '2022-04-30', proxy_handle, proxy_id)
            # m.getDeliveryList('2022-07-10', '2022-07-26')
            # m.getDeliveryList(timeStart, timeEnd)                     # 派送问题件 更新

    elif int(select) == 1:
        proxy_handle = '代理服务器0'
        proxy_id = '192.168.13.89:37467'                            # 输入代理服务器节点和端口
        m = QueryTwo('+86-18538110674', 'qyz04163510.', "", "", select, proxy_handle, proxy_id)
        # timeStart, timeEnd = m.readInfo('派送问题件_跟进表')
        # m.getOrderList_T('2022-06-01', '2022-06-30')
        logisticsN_begin = '2022-07-11'                             # 送达客户不在/客户长期不在  物流轨迹查询时间
        logisticsN_end = '2022-07-31'
        m.outport_getDeliveryList('2022-10-01', '2022-12-11', logisticsN_begin, logisticsN_end)
        # m.getMessageLog('2022-07-01', '2022-07-15')



    elif int(select) == 2:
        proxy_handle = '代理服务器0'
        proxy_id = '192.168.13.89:37467'                            # 输入代理服务器节点和端口
        m = QueryTwo('+86-18538110674', 'qyz04163510.', "", "", select, proxy_handle, proxy_id)
        # timeStart, timeEnd = m.readInfo('派送问题件_跟进表')
        # m.getOrderList_T('2022-06-01', '2022-06-30')
        m.outport_List('2022-07-20', '2022-08-17')
        # m.getMessageLog('2022-07-01', '2022-07-15')


    elif int(select) == 3:              # 拒收问题件、物流问题件、物流问题件 检测同一个客户（电话） 的订单
        proxy_handle = '代理服务器0'
        proxy_id = '192.168.13.89:37467'                            # 输入代理服务器节点和端口
        m = QueryTwo('+86-18538110674', 'qyz04163510.', "", "", 99, proxy_handle, proxy_id)

        # timeStart, timeEnd = m.readInfo('拒收问题件_更新')
        m.order_js_QueryUpdata('2022-10-01', '2022-11-30', proxy_handle, proxy_id)
        # m.order_js_QueryUpdata(timeStart, timeEnd)  # 拒收问题件 更新

        # timeStart, timeEnd = m.readInfo('物流问题件_更新')
        m.waybill_InfoQueryUpdata('2022-10-01', '2022-10-31', proxy_handle, proxy_id)
        m.waybill_InfoQueryUpdata('2022-11-01', '2022-11-30', proxy_handle, proxy_id)
        # m.waybill_InfoQueryUpdata(timeStart, timeEnd)  # 物流问题件 更新

        # timeStart, timeEnd = m.readInfo('派送问题件_更新')
        m.getDeliveryList('2022-10-01', '2022-11-30', proxy_handle, proxy_id)
        # m.getDeliveryList(timeStart, timeEnd)  # 派送问题件 更新

        timeStart = (datetime.datetime.now() - relativedelta(months=1)).strftime('%Y-%m') + '-01'
        timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        m.Check_Iphone_Updata(timeStart, timeEnd)              # 拒收&派送|物流问题件-电话检测


    print('查询耗时：', datetime.datetime.now() - start)