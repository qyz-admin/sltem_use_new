import pandas as pd
import os
import xlwings as xl
import pandas.io.formats.excel
from sqlalchemy import create_engine
from settings import Settings
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from dateutil.relativedelta import relativedelta
import datetime
import xlwings
import win32api,win32con

# -*- coding:utf-8 -*-
class SltemMonitoring(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()

    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    def check_time(self, team):
        match = {'日本': 'slrb',
                 '港台': 'gat'}
        Time_day = []
        for i in range(1, datetime.datetime.now().month + 1):  # 获取当年当前的月份时间
            try:
                daytime = (datetime.datetime.now().replace(month=i)).strftime('%Y-%m') + (
                    (datetime.datetime.now()).strftime('-%d'))
                Time_day.append(daytime)
            except Exception as e:
                print('xxxx时间配置出错,已手动调整：' + str(i) + '月份', str(Exception) + str(e))
                Time_day.append(
                    str(int(datetime.datetime.now().year)) + '-' + str(i) + (datetime.datetime.now().strftime('-%d')))
        for i in range(datetime.datetime.now().month + 1, 13):  # 获取往年当前的月份时间
            try:
                daytime = str(int(datetime.datetime.now().year) -1) + (datetime.datetime.now().replace(month=i)).strftime('-%m') + (
                    (datetime.datetime.now()).strftime('-%d'))
                Time_day.append(daytime)
            except Exception as e:
                print('xxxx时间配置出错失败00：' + str(i) + '月份', str(Exception) + str(e))
                Time_day.append(str(int(datetime.datetime.now().year) - 1) + '-' + str(i) + (
                    datetime.datetime.now().strftime('-%d')))
        #  对时间数组进行排序  list.sort(cmp=None, key=None, reverse=False)；reverse -- 排序规则，reverse = True 降序， reverse = False 升序（默认）
        Time_day.sort()
        print('正在获取本次同期比较需要的---具体时间......')
        print(Time_day[11])
        print(Time_day[10])
        # 获取监控表是否有同期上传时间的数据
        rq_day = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        print('正在检查监控表是否有需要的---具体日期......')
        sql = '''SELECT distinct qsb.`记录时间` FROM qsb_{0} qsb WHERE qsb.`记录时间`>='{1}';'''.format(match[team], rq_day)
        rq = pd.read_sql_query(sql=sql, con=self.engine1)
        df = rq['记录时间'].values              # datafram转为数组
        info = ''
        for r in df:
            # print(type(r.strftime('%Y')))
            if Time_day[10] == r.strftime('%Y-%m-%d'):
                print(r)
                info = '---已确认，可以进行同期数据对比'
                break
            else:
                info = '---需要手动上传需要时间的数据'

        if info == '---已确认，可以进行同期数据对比':
            print('++++++完成时间确认++++++')
            print('===>>>已确认---' + team + '---团队开始运行<<<===')
            self.order_Monitoring(team)  # 各月缓存
            self.sl_Monitoring(team)  # 输出数据
            print('===>>>---' + team + '---团队运行结束<<<===')
        else:
            print(info)


    def match_time(self, team):
        match = {'港台': 'qsb_gat',
                  '品牌': 'qsb_slsc'}
        sql = '''SELECT DISTINCT 记录时间
                FROM {0} d where d.记录时间 >= '{1}'
                GROUP BY 记录时间
                ORDER BY 记录时间 DESC;'''.format(match[team], (datetime.datetime.now() - relativedelta(months=2)).strftime('%Y-%m-%d'))
        rq = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(rq)
        now_time = datetime.datetime.now().strftime('%Y-%m-%d')
        now_day = int(datetime.datetime.now().strftime('%d'))
        print('本期时间： ' + now_time)
        last_time = (datetime.datetime.now() - relativedelta(months=1)).strftime('%Y-%m-%d')
        last_time2 = datetime.datetime.now() - relativedelta(months=1)
        last_day = int((datetime.datetime.now() - relativedelta(months=1)).strftime('%d'))
        print('上期时间： ' + last_time)
        last_month = ''
        if now_day > last_day:
            print('请确认对比时间的 上期时间')
        else:
            count = 0

            print('即将获取对比数据-上期时间')
            for i in range(0, 31):  # 按天循环获取订单状态
                day = last_time2 + datetime.timedelta(days=i)
                day2 = datetime.date(int(day.strftime('%Y')), int(day.strftime('%m')), int(day.strftime('%d')))
                # print(count)
                # print(day2)
                if count == 1:
                    # print('结束')
                    break
                for k in rq['记录时间']:
                    if day2 == k:
                        # print('有')
                        last_month = k
                        count = count + 1
                        break
            print(last_month)

        # 各月缓存（整体一）
        self.order_Monitoring(team)
        for team in ['神龙-台湾', '神龙-香港', '火凤凰-台湾', '火凤凰-香港', '金鹏-台湾']:
            self.sl_Monitoring(team, now_time, last_month)      # 输出数据--每月正常使用的时间（二）

    # 获取各团队各月的签收表缓存数据（一）
    def order_Monitoring(self, team):
        match = {'品牌': 'slsc',
                 '港台': 'gat'}
        start: datetime = datetime.datetime.now()
        print('正在获取 ' + team + ' 每月（全部）签收数据…………')
        if match[team] == 'gat':
            sql = '''SELECT LEFT(年月,4) AS 年, 年月, 旬, 日期, 团队, 币种, 订单来源, 订单编号, 出货时间, 状态时间, 上线时间, 最终状态, 是否改派,物流方式,产品id,
                            父级分类,二级分类,三级分类,下单时间,审核时间,仓储扫描时间,完结状态时间,价格RMB
                    FROM {0}_zqsb a 
                    WHERE a.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 6 MONTH),'%Y%m') AND a.年月 <= DATE_FORMAT(curdate(),'%Y%m')
                    ORDER BY a.`下单时间`;'''.format(match[team])
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        print('----写入中......')
        df.to_sql('qsb_缓存_month', con=self.engine1, index=False, if_exists='replace', chunksize=20000)
        columns = list(df.columns)
        columns = ','.join(columns)
        sql = 'REPLACE INTO {0}({1}) SELECT * FROM qsb_缓存_month; '.format('qsb_缓存_month_cp', columns)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=20000)

        print('写入缓存耗时：', datetime.datetime.now() - start)

    # 获取每月正常使用的时间（二）
    def sl_Monitoring(self, team, now_month, last_month, ready):
        match = {'品牌-日本': '"金鹏家族-品牌", "金鹏家族-品牌1组", "金鹏家族-品牌2组", "金鹏家族-品牌3组"',
                 '品牌-香港': '"金鹏家族-品牌", "金鹏家族-品牌1组", "金鹏家族-品牌2组", "金鹏家族-品牌3组"',
                 '品牌-台湾': '"金鹏家族-品牌", "金鹏家族-品牌1组", "金鹏家族-品牌2组", "金鹏家族-品牌3组"',
                 '品牌-马来西亚': '"金鹏家族-品牌", "金鹏家族-品牌1组", "金鹏家族-品牌2组", "金鹏家族-品牌3组"',
                 '品牌-新加坡': '"金鹏家族-品牌", "金鹏家族-品牌1组", "金鹏家族-品牌2组", "金鹏家族-品牌3组"',
                 '品牌-菲律宾': '"金鹏家族-品牌", "金鹏家族-品牌1组", "金鹏家族-品牌2组", "金鹏家族-品牌3组"',
                 '港台-台湾':'"神龙家族-港澳台", "火凤凰-港澳台", "红杉家族-港澳台", "红杉家族-港澳台2", "金狮-港澳台", "金鹏家族-小虎队", "火凤凰-港台(繁体)", "神龙-低价", "神龙-主页运营1组", "神龙-运营1组", "神龙-主页运营"',
                 '神龙-香港': '"神龙家族-港澳台"',
                 '神龙-台湾': '"神龙家族-港澳台"',
                 '小虎队-香港': '"金鹏家族-小虎队"',
                 '小虎队-台湾': '"金鹏家族-小虎队"',
                 '神龙运营1组-台湾': '"神龙-运营1组"',
                 '神龙火凤凰-台湾': '"神龙家族-港澳台","火凤凰-港澳台", "火凤凰-港台(繁体)"',
                 '火凤凰-台湾': '"火凤凰-港澳台", "火凤凰-港台(繁体)"',
                 '火凤凰-香港': '"火凤凰-港澳台", "火凤凰-港台(繁体)"'}
        emailAdd = {'神龙香港': 'giikinliujun@163.com', '神龙台湾': 'giikinliujun@163.com',
                    '火凤凰香港': 'giikinliujun@163.com', '火凤凰台湾': 'giikinliujun@163.com',
                    '品牌-日本': 'sunyaru@giikin.com', '品牌-台湾': 'sunyaru@giikin.com', '品牌-香港': 'sunyaru@giikin.com',
                    '品牌-马来西亚': 'sunyaru@giikin.com', '品牌-新加坡': 'sunyaru@giikin.com', '品牌-菲律宾': 'sunyaru@giikin.com'}
        # 初始化配置
        start: datetime = datetime.datetime.now()
        family = ""
        if team in ('港台-台湾', '神龙火凤凰-台湾', '神龙-香港', '神龙-台湾', '火凤凰-香港', '火凤凰-台湾', '小虎队-香港', '小虎队-台湾', '神龙运营1组-台湾'):
            family = 'qsb_gat'
        elif team in ('品牌-日本', '品牌-马来西亚', '品牌-新加坡', '品牌-菲律宾', '品牌-台湾', '品牌-香港'):
            family = 'qsb_slsc'
        currency = team.split('-')[1]
        print('*********开始运行监控对比表*********')
        # 获取对比时间-本期
        sql = '''SELECT DISTINCT 年月,日期
                    FROM {0} d
                    WHERE d.`记录时间` ='{1}'
                    GROUP BY 年月
                    ORDER BY 年月 DESC'''.format(family, now_month)
        rq = pd.read_sql_query(sql=sql, con=self.engine1)
        now_month_new = ''
        now_month_old = ''
        if ready == '本期宏':
            now_month_new = rq['年月'][0]
            now_month_old = rq['年月'][1]
        elif ready == '本期上月宏':
            now_month_new = rq['年月'][0]
            now_month_old = rq['年月'][1]
        elif ready == '上期宏':
            now_month_new = rq['年月'][1]
            now_month_old = rq['年月'][2]
        print('本期时间：' + now_month)
        print('当月: ', end="")
        print(now_month_new)
        print('上月: ', end="")
        print(now_month_old)
        # 获取对比时间-上期
        sql = '''SELECT DISTINCT 年月,日期
                    FROM {0} d
                    WHERE d.`记录时间` ='{1}'
                    GROUP BY 年月
                    ORDER BY 年月 DESC'''.format(family, last_month)
        rq = pd.read_sql_query(sql=sql, con=self.engine1)
        last_month_new = ''
        last_month_old = ''
        if ready == '本期宏':
            last_month_new = rq['年月'][0]
            last_month_old = rq['年月'][1]
        elif ready == '本期上月宏':
            last_month_new = rq['年月'][0]
            last_month_old = rq['年月'][1]
        elif ready == '上期宏':
            last_month_new = rq['年月'][1]
            last_month_old = rq['年月'][2]
        print('上期时间：' + last_month)
        print('当月: ', end="")
        print(last_month_new)
        print('上月: ', end="")
        print(last_month_old)

        listT = []  # 查询sql 存放池
        show_name = []  # 打印进度需要
        # 月签收率（天）---查询
        sqlqsb2 = '''SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,
                                总订单量, 
                                签收量 / 完成量 AS '总签收/完成',
                                签收量 / 总订单量 AS '总签收/总计',
                                退货量 / 总订单量 AS '退款率',
                                完成量 / 总订单量 AS '总完成占比',
                                直发量 直发总计,
                                直发签收量 / 直发完成量 AS '直发签收/完成',
                                直发签收量 / 直发量 AS '直发签收/总计',
                                直发完成量 / 直发量 AS '直发完成占比',
                                改派量 改派总计,
                                改派签收量 / 改派完成量 AS '改派签收/完成',
                                改派签收量 / 改派量 AS '改派签收/总计',
                                改派完成量 / 改派量 AS '改派完成占比'
                    FROM( SELECT IFNULL(币种,'合计') 币种, IFNULL(年月,'合计') 年月,IFNULL(父级分类,'合计') 父级分类,IFNULL(二级分类,'合计') 二级分类, IFNULL(三级分类,'合计') 三级分类,IFNULL(物流方式,'合计') 物流方式,IFNULL(旬,'合计') 旬,
                                COUNT(`订单编号`) 总订单量, 
                                SUM(IF(`是否改派` = '直发',1,0)) as 直发量,
                                SUM(IF(`是否改派` = '直发' AND `最终状态` = '已签收',1,0)) as 直发签收量,
                                SUM(IF(`是否改派` = '直发' AND `最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') ,1,0)) as 直发完成量,
                                SUM(IF(`是否改派` = '改派',1,0)) as 改派量,
                                SUM(IF(`是否改派` = '改派' AND `最终状态` = '已签收',1,0)) as 改派签收量,
                                SUM(IF(`是否改派` = '改派' AND `最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') ,1,0)) as 改派完成量,
                                SUM(IF(`最终状态` = '已签收',1,0)) as 签收量,
                                SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) as 完成量,
                                SUM(IF(`最终状态` = '已退货',1,0)) as 退货量
                        FROM {0} sl_cx
                        WHERE (sl_cx.`记录时间`= '{1}' AND sl_cx.`年月` = '{2}' OR sl_cx.`记录时间`= '{3}' AND sl_cx.`年月` = '{4}')
                            AND sl_cx.`币种` = '{5}'  AND sl_cx.`团队` IN ({6})
                            AND sl_cx.`父级分类` IS NOT NULL 
                        GROUP BY 年月,父级分类,二级分类,三级分类,物流方式,旬
                        with rollup ) sl;'''.format(family, now_month, now_month_new, last_month, last_month_new, currency, match[team])
        listT.append(sqlqsb2)
        show_name.append(' 月（天）签收率_…………')
        # 月签收率（整月）---查询
        sqlqsb3 = '''SELECT 币种,年月,父级分类,二级分类,三级分类,物流方式,旬,
                                    总订单量, 
                                    签收量 / 完成量 AS '总签收/完成',
                                    签收量 / 总订单量 AS '总签收/总计',
                                    退货量 / 总订单量 AS '退款率',
                                    完成量 / 总订单量 AS '总完成占比',
                                    直发量 直发总计,
                                    直发签收量 / 直发完成量 AS '直发签收/完成',
                                    直发签收量 / 直发量 AS '直发签收/总计',
                                    直发完成量 / 直发量 AS '直发完成占比',
                                    改派量 改派总计,
                                    改派签收量 / 改派完成量 AS '改派签收/完成',
                                    改派签收量 / 改派量 AS '改派签收/总计',
                                    改派完成量 / 改派量 AS '改派完成占比'
                        FROM( SELECT IFNULL(币种,'合计') 币种,IFNULL(年月,'合计') 年月,IFNULL(父级分类,'合计') 父级分类,IFNULL(二级分类,'合计') 二级分类, IFNULL(三级分类,'合计') 三级分类,IFNULL(物流方式,'合计') 物流方式,IFNULL(旬,'合计') 旬,
                                    COUNT(`订单编号`) 总订单量, 
                                    SUM(IF(`是否改派` = '直发',1,0)) as 直发量,
                                    SUM(IF(`是否改派` = '直发' AND `最终状态` = '已签收',1,0)) as 直发签收量,
                                    SUM(IF(`是否改派` = '直发' AND `最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') ,1,0)) as 直发完成量,
                                    SUM(IF(`是否改派` = '改派',1,0)) as 改派量,
                                    SUM(IF(`是否改派` = '改派' AND `最终状态` = '已签收',1,0)) as 改派签收量,
                                    SUM(IF(`是否改派` = '改派' AND `最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') ,1,0)) as 改派完成量,
                                    SUM(IF(`最终状态` = '已签收',1,0)) as 签收量,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) as 完成量,
                                    SUM(IF(`最终状态` = '已退货',1,0)) as 退货量
                        FROM {0} sl_cx
                        WHERE (sl_cx.`记录时间`= '{1}' AND sl_cx.`年月` = '{2}' OR sl_cx.`记录时间`= '{3}' AND sl_cx.`年月` = '{4}')
                            AND sl_cx.`币种` = '{5}'  AND sl_cx.`团队` IN ({6})
                            AND sl_cx.`父级分类` IS NOT NULL 
                        GROUP BY 年月,父级分类,二级分类,三级分类,物流方式,旬
                        with rollup ) sl;'''.format(family, now_month, now_month_old, last_month, last_month_old, currency, match[team])
        listT.append(sqlqsb3)
        show_name.append(' 月（整月）签收率_…………')
        # 月签收率（旬）---查询
        sqlqsb4 = '''SELECT 年月,旬,币种,父级分类,二级分类,三级分类,物流方式,
                                总订单量, 
                                签收量 / 完成量 AS '总签收/完成',
                                签收量 / 总订单量 AS '总签收/总计',
                                退货量 / 总订单量 AS '退款率',
                                完成量 / 总订单量 AS '总完成占比',
                                直发量 直发总计,
                                直发签收量 / 直发完成量 AS '直发签收/完成',
                                直发签收量 / 直发量 AS '直发签收/总计',
                                直发完成量 / 直发量 AS '直发完成占比',
                                改派量 改派总计,
                                改派签收量 / 改派完成量 AS '改派签收/完成',
                                改派签收量 / 改派量 AS '改派签收/总计',
                                改派完成量 / 改派量 AS '改派完成占比'
                    FROM( SELECT IFNULL(年月,'合计') 年月,IFNULL(旬,'合计') 旬,IFNULL(币种,'合计') 币种,IFNULL(父级分类,'合计') 父级分类, IFNULL(二级分类,'合计') 二级分类,IFNULL(三级分类,'合计') 三级分类,IFNULL(物流方式,'合计') 物流方式,
                                COUNT(`订单编号`) 总订单量, 
                                SUM(IF(`是否改派` = '直发',1,0)) as 直发量,
                                SUM(IF(`是否改派` = '直发' AND `最终状态` = '已签收',1,0)) as 直发签收量,
                                SUM(IF(`是否改派` = '直发' AND `最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') ,1,0)) as 直发完成量,
                                SUM(IF(`是否改派` = '改派',1,0)) as 改派量,
                                SUM(IF(`是否改派` = '改派' AND `最终状态` = '已签收',1,0)) as 改派签收量,
                                SUM(IF(`是否改派` = '改派' AND `最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') ,1,0)) as 改派完成量,
                                SUM(IF(`最终状态` = '已签收',1,0)) as 签收量,
                                SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) as 完成量,
                                SUM(IF(`最终状态` = '已退货',1,0)) as 退货量
                        FROM {0} sl_cx
                        WHERE sl_cx.`记录时间` = '{1}'
                            AND sl_cx.`币种` = '{2}'  AND sl_cx.`团队` IN ({3})
                            AND sl_cx.`父级分类` IS NOT NULL 
                        GROUP BY 年月,旬,父级分类,二级分类,三级分类,物流方式
                        with rollup ) sl;'''.format(family, now_month, currency, match[team])
        listT.append(sqlqsb4)
        show_name.append(' 月（旬）签收率_…………')
        # 月签收率（各月）---查询
        sqlqsb5 = '''SELECT sl_gat.`币种`,sl_gat.`年月`,sl_gat.父级分类,sl_gat.二级分类,sl_gat.三级分类,'' 产品名称,sl_gat.物流方式,sl_gat.旬,
                            sl_gat.`总订单量`,
                            sl_gat.`已签收订单量` / sl_gat.`完成订单量` AS '总签收/完成',
                            sl_gat.`已签收订单量` / sl_gat.`总订单量` AS '总签收/总计',
                            sl_gat.`退货订单量` / sl_gat.`总订单量` AS '退款率',
                            sl_gat.`完成订单量` / sl_gat.`总订单量` AS '总完成占比',
                            sl_gat.`直发订单量` 直发总计,
                            sl_gat.`直发已签收订单量` / sl_gat.`直发完成订单量` AS '直发签收/完成',
                            sl_gat.`直发已签收订单量` / sl_gat.`直发订单量` AS '直发签收/总计', 
                            sl_gat.`直发完成订单量` / sl_gat.`直发订单量` AS '直发完成占比',
                            sl_gat.`改派订单量` 改派总计,
                            sl_gat.`改派已签收订单量` / sl_gat.`改派完成订单量` AS '改派签收/完成',
                            sl_gat.`改派已签收订单量` / sl_gat.`改派订单量` AS '改派签收/总计',
                            sl_gat.`改派完成订单量` / sl_gat.`改派订单量` AS '改派完成占比',
                            sl_gat.`总销售额`,
                            sl_gat.`已签收销售额` / sl_gat.`完成销售额` AS '总签收/完成(金额)',
                            sl_gat.`已签收销售额` / sl_gat.`总销售额` AS '总签收/总计(金额)',
                            sl_gat.`退货销售额` / sl_gat.`总销售额` AS '退款率(金额)',
                            sl_gat.`完成销售额` / sl_gat.`总销售额` AS '总完成占比(金额)',
                            sl_gat.`直发销售额`,
                            sl_gat.`直发已签收销售额` / sl_gat.`直发完成销售额` AS '直发签收/完成(金额)',
                            sl_gat.`直发已签收销售额` / sl_gat.`直发销售额` AS '直发签收/总计(金额)',
                            sl_gat.`直发完成销售额` / sl_gat.`直发销售额` AS '直发完成占比(金额)',
                            sl_gat.`改派销售额`,
                            sl_gat.`改派已签收销售额` / sl_gat.`改派完成销售额` AS '改派签收/完成(金额)',
                            sl_gat.`改派已签收销售额` / sl_gat.`改派销售额` AS '改派签收/总计(金额)',
                            sl_gat.`改派完成销售额` / sl_gat.`改派销售额` AS '改派完成占比(金额)'
                    FROM (SELECT  币种,IFNULL(年月,'合计') 年月,IFNULL(父级分类,'合计') 父级分类,IFNULL(二级分类,'合计') 二级分类, IFNULL(三级分类,'合计') 三级分类,IFNULL(物流方式,'合计') 物流方式,IFNULL(旬,'合计') 旬,
                                SUM(总订单量) 总订单量,
                                SUM(总销售额) 总销售额,
                                IFNULL(SUM(直发订单量),0) 直发订单量, IFNULL(SUM(直发销售额),0) 直发销售额,
                                IFNULL(SUM(直发已签收订单量),0) 直发已签收订单量, IFNULL(SUM(直发已签收销售额),0) 直发已签收销售额,
                                IFNULL(SUM(直发完成订单量),0) 直发完成订单量, IFNULL(SUM(直发完成销售额),0) 直发完成销售额,
                                (SUM(总订单量) - IFNULL(SUM(直发订单量),0)) AS 改派订单量, (SUM(总销售额) - IFNULL(SUM(直发销售额),0)) AS 改派销售额,
                                IFNULL(SUM(改派已签收订单量),0) 改派已签收订单量, IFNULL(SUM(改派已签收销售额),0) 改派已签收销售额,
                                IFNULL(SUM(改派完成订单量),0) 改派完成订单量, IFNULL(SUM(改派完成销售额),0) 改派完成销售额,
                                IFNULL(SUM(已签收订单量),0) 已签收订单量, IFNULL(SUM(已签收销售额),0) 已签收销售额,
                                IFNULL(SUM(完成订单量),0) 完成订单量, IFNULL(SUM(完成销售额),0) 完成销售额,
                                IFNULL(SUM(退货订单量),0) 退货订单量, IFNULL(SUM(退货销售额),0) 退货销售额
                        FROM (SELECT  币种,年月,父级分类,二级分类,三级分类,物流方式,旬,
                                    COUNT(`订单编号`) 总订单量,
                                    SUM(`价格RMB`) 总销售额,
                                    SUM(IF(`最终状态` = "已签收",1,0)) 已签收订单量,
                                    SUM(IF(`最终状态` = "已签收",1,0)) 已签收销售额,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) 完成订单量,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) 完成销售额,
                                    SUM(IF(`最终状态` = "已退货",1,0)) 退货订单量,
                                    SUM(IF(`最终状态` = "已退货",1,0)) 退货销售额,

                                    SUM(IF(`是否改派` = "直发",1,0)) 直发订单量,
                                    SUM(IF(`是否改派` = "直发",`价格RMB`,0)) 直发销售额,
                                    SUM(IF(`是否改派` = "直发" AND `最终状态` = "已签收",1,0)) 直发已签收订单量,
                                    SUM(IF(`是否改派` = "直发" AND `最终状态` = "已签收",`价格RMB`,0)) 直发已签收销售额,
                                    SUM(IF(`是否改派` = "直发" AND `最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) 直发完成订单量,
                                    SUM(IF(`是否改派` = "直发" AND `最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),`价格RMB`,0)) 直发完成销售额,
                                    
                                    SUM(IF(`是否改派` = "改派",1,0)) 改派订单量,
                                    SUM(IF(`是否改派` = "改派",`价格RMB`,0)) 改派销售额,
                                    SUM(IF(`是否改派` = "改派" AND `最终状态` = "已签收",1,0)) 改派已签收订单量,
                                    SUM(IF(`是否改派` = "改派" AND `最终状态` = "已签收",`价格RMB`,0)) 改派已签收销售额,
                                    SUM(IF(`是否改派` = "改派" AND `最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) 改派完成订单量,
                                    SUM(IF(`是否改派` = "改派" AND `最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),`价格RMB`,0)) 改派完成销售额	
                            FROM {0} sl_cx
                            WHERE sl_cx.`币种` = '{1}' AND sl_cx.`团队` IN ({2}) AND sl_cx.`父级分类` IS NOT NULL
                            GROUP BY 币种,年月,父级分类,二级分类,三级分类,物流方式,旬
                            ORDER BY 币种,年月
						) s
						GROUP BY s.年月,s.父级分类,s.二级分类,s.三级分类,s.物流方式,s.旬
                        with rollup
					) sl_gat 
                    ORDER BY sl_gat.`年月` DESC;'''.format('qsb_缓存_month_cp', currency, match[team])
        listT.append(sqlqsb5)
        show_name.append(' 月（各月）签收率_…………')

        # 月物流（天）---查询
        sqlWl2 = '''SELECT 币种,年月,物流方式,父级分类,旬,
                                总订单量 总订单, 
                                null '总签收/完成',null '总签收/总计',null '退款率',null  '总完成占比',
                                直发量 总计,
                                直发签收量 / 直发完成量 AS '直发签收/完成',
                                直发签收量 / 直发量 AS '直发签收/总计',
                                直发完成量 / 直发量 AS '直发完成占比',
                                null  改派总计,null '改派签收/完成',null '改派签收/总计',null '改派完成占比'
                    FROM( SELECT IFNULL(币种,'合计') 币种,IFNULL(年月,'合计') 年月,IFNULL(物流方式,'合计') 物流方式,IFNULL(父级分类,'合计') 父级分类,IFNULL(旬,'合计') 旬,
                                null 总订单量, 
                                COUNT(`订单编号`) as 直发量,
                                SUM(IF(`最终状态` = '已签收',1,0)) as 直发签收量,
                                SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') ,1,0)) as 直发完成量
                    FROM {0} sl_cx
                    WHERE (sl_cx.`记录时间`= '{1}' AND sl_cx.`年月` = '{2}' OR sl_cx.`记录时间`= '{3}' AND sl_cx.`年月` = '{4}')
                        AND sl_cx.`币种` = '{5}'  AND sl_cx.`团队` IN ({6})
                        AND sl_cx.`是否改派` = "直发"
                        AND sl_cx.`父级分类` IS NOT NULL 
                    GROUP BY 币种,年月,物流方式,父级分类,旬
                    with rollup ) sl;'''.format(family, now_month, now_month_new, last_month, last_month_new, currency, match[team])
        listT.append(sqlWl2)
        show_name.append(' 月（天）物流…………')
        # 月物流（整月）---查询
        sqlWl3 = '''SELECT 币种,年月,物流方式,父级分类,旬,
                                总订单量 总订单, 
                                null '总签收/完成',null '总签收/总计',null '退款率',null '总完成占比',
                                直发量 总计,
                                直发签收量 / 直发完成量 AS '直发签收/完成',
                                直发签收量 / 直发量 AS '直发签收/总计',
                                直发完成量 / 直发量 AS '直发完成占比',
                                null  改派总计,null '改派签收/完成',null '改派签收/总计',null '改派完成占比'
                    FROM( SELECT IFNULL(币种,'合计') 币种,IFNULL(年月,'合计') 年月,IFNULL(物流方式,'合计') 物流方式,IFNULL(父级分类,'合计') 父级分类,IFNULL(旬,'合计') 旬,
                                null 总订单量, 
                                COUNT(`订单编号`) as 直发量,
                                SUM(IF(`最终状态` = '已签收',1,0)) as 直发签收量,
                                SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') ,1,0)) as 直发完成量
                        FROM {0} sl_cx
                        WHERE (sl_cx.`记录时间`= '{1}' AND sl_cx.`年月` = '{2}' OR sl_cx.`记录时间`= '{3}' AND sl_cx.`年月` = '{4}')
                            AND sl_cx.`币种` = '{5}'  AND sl_cx.`团队` IN ({6})
                            AND sl_cx.`是否改派` = "直发"
                            AND sl_cx.`父级分类` IS NOT NULL 
                    GROUP BY 年月,物流方式,父级分类,旬
                    with rollup ) sl;'''.format(family, now_month, now_month_old, last_month, last_month_old, currency, match[team])
        listT.append(sqlWl3)
        show_name.append(' 月（整月）物流…………')

        # 月时效（天）---查询
        sqltime2 = '''SELECT 币种,年月,物流方式,父级分类,旬,
                                总订单量 总单量, 
                                直发订单量 AS 直发下单出库量,
                                IFNULL(`直发下单-出库时`,0) / IFNULL(`直发订单量`,0) AS 下单出库时效,
                                直发出库完成量,
                                IFNULL(`直发出库-完成时`,0) / IFNULL(`直发出库完成量`,0) 出库完成时效,
                                直发下单完成量,
                                IFNULL(`直发下单-完成时`,0) / IFNULL(`直发下单完成量`,0) 下单完成时效,
                                直发出货上线量,
                                IFNULL(`直发出货-上线时`,0) / IFNULL(`直发出货上线量`,0) 出货上线时效,
                                直发上线完成量,
                                IFNULL(`直发上线-完成时`,0) / IFNULL(`直发上线完成量`,0) 上线完成时效,					
                                直发已签收订单量 / 直发下单完成量 AS '签收/完成',
                                直发已签收订单量 / 直发出库完成量 AS '签收/总计'
                        FROM( SELECT IFNULL(币种,'合计') 币种,IFNULL(年月,'合计') 年月,IFNULL(物流方式,'合计') 物流方式,IFNULL(父级分类,'合计') 父级分类,IFNULL(旬,'合计') 旬,
                                    NULL 总订单量, 
                                    SUM(IF(`最终状态` = '已签收',1,0)) AS 直发已签收订单量,
                                    COUNT(`订单编号`) 直发订单量,
                                    SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时',
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发出库完成量,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`),0)) AS '直发出库-完成时',
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发下单完成量,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`),0)) AS '直发下单-完成时',
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发出货上线量,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(`上线时间`, IFNULL(`仓储扫描时间`,`出货时间`)),0)) AS '直发出货-上线时',
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发上线完成量,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`上线时间`),0)) AS '直发上线-完成时'
                            FROM {0} sl_cx
                            WHERE (sl_cx.`记录时间`= '{1}' AND sl_cx.`年月` = '{2}' OR sl_cx.`记录时间`= '{3}' AND sl_cx.`年月` = '{4}')
                                AND sl_cx.`币种` = '{5}' AND sl_cx.`团队` IN ({6})
                                AND sl_cx.`是否改派` = "直发"
                                AND sl_cx.`父级分类` IS NOT NULL 
                                AND sl_cx.`仓储扫描时间` IS NOT NULL 
                            GROUP BY 年月,物流方式,旬
                        with rollup ) sl;'''.format(family, now_month, now_month_new, last_month, last_month_new, currency, match[team])
        listT.append(sqltime2)
        show_name.append(' 月（天）时效…………')
        # 月时效（旬）---查询
        sqltime3 = '''SELECT 币种,年月,旬,物流方式,父级分类,
                                总订单量 总单量, 
                                直发订单量 AS 直发下单出库量,
                                IFNULL(`直发下单-出库时`,0) / IFNULL(`直发订单量`,0) as 下单出库时效,
                                直发出库完成量,
                                IFNULL(`直发出库-完成时`,0) / IFNULL(`直发出库完成量`,0) 出库完成时效,
                                直发下单完成量,
                                IFNULL(`直发下单-完成时`,0) / IFNULL(`直发下单完成量`,0) 下单完成时效,
                                直发出货上线量,
                                IFNULL(`直发出货-上线时`,0) / IFNULL(`直发出货上线量`,0) 出货上线时效,
                                直发上线完成量,
                                IFNULL(`直发上线-完成时`,0) / IFNULL(`直发上线完成量`,0) 上线完成时效
                        FROM( SELECT IFNULL(币种,'合计') 币种,IFNULL(年月,'合计') 年月,IFNULL(旬,'合计') 旬,IFNULL(物流方式,'合计') 物流方式,IFNULL(父级分类,'合计') 父级分类,
                                    NULL 总订单量, 
                                    SUM(IF(`最终状态` = '已签收',1,0))  as 直发已签收订单量,
                                    COUNT(`订单编号`) 直发订单量,
                                    SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时',
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发出库完成量,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`),0)) AS '直发出库-完成时',
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发下单完成量,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`),0)) AS '直发下单-完成时',
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发出货上线量,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(`上线时间`, IFNULL(`仓储扫描时间`,`出货时间`)),0)) AS '直发出货-上线时',
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发上线完成量,
                                    SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`上线时间`),0)) AS '直发上线-完成时'
                            FROM {0} sl_cx
                            WHERE (sl_cx.`记录时间`= '{1}' AND sl_cx.`年月` = '{2}' OR sl_cx.`记录时间`= '{3}' AND sl_cx.`年月` = '{4}')
                                AND sl_cx.`币种` = '{5}' AND sl_cx.`团队` IN ({6})
                                AND sl_cx.`是否改派` = "直发"
                                AND sl_cx.`父级分类` IS NOT NULL 
                                AND sl_cx.`仓储扫描时间` IS NOT NULL 
                            GROUP BY 年月,旬,物流方式,父级分类
                            with rollup ) sl;'''.format(family, now_month, now_month_new, last_month, last_month_new, currency, match[team])
        listT.append(sqltime3)
        show_name.append(' 月（旬）时效…………')
        # 月时效(各月)---查询
        sqltime4 = '''SELECT sl_rb.`币种`,sl_rb.`年月`,sl_rb.`物流方式`,sl_rb.`父级分类`,sl_rb.`旬`,
                                sl_rb.`总单量`,
                                sl_rb.`直发下单出库单量`,sl_rb.`直发下单出库时效`,
                                sl_rb.`直发出货上线量`,sl_rb.`直发出货上线时效`,
                                sl_rb.`直发上线完成量`,sl_rb.`直发上线完成时效`,
                                sl_rb.`直发出库完成单量`,sl_rb.`直发出库完成时效`,
                                sl_rb.`直发下单完成时效`,sl_rb.`直发下单完成单量`,
                                sl_rb.`直发已签收订单量` / sl_rb.`直发下单完成单量` AS '签收/完成',
                                sl_rb.`直发已签收订单量`/ sl_rb.`直发下单出库单量` AS '签收/总计'
                    FROM (SELECT sl_zong.币种 币种,IFNULL(sl_zong.年月,'合计') 年月,IFNULL(sl_zong.物流方式,'合计') 物流方式,IFNULL(sl_zong.父级分类,'合计') 父级分类,IFNULL(sl_zong.旬,'合计') 旬,
                                SUM(sl_zong.`总订单量`) 总单量,
                                SUM(IFNULL(sl_cx_zf_qs.`直发已签收订单量`,0)) 直发已签收订单量,
                                SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库单量,
                                SUM(IFNULL(sl_zong_zf.`直发下单-出库时`,0)) / SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库时效,
                                SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成单量,
                                SUM(IFNULL(sl_cx_zf_wc.`直发出库-完成时`,0)) / SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成时效,
                                SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成单量,
                                SUM(IFNULL(sl_cx_zf_wc.`直发下单-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成时效,
                                SUM(IFNULL(sl_cx_zf_wc.`直发出货上线量`,0)) 直发出货上线量,
                                SUM(IFNULL(sl_cx_zf_wc.`直发出货-上线时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发出货上线量`,0)) 直发出货上线时效,
                                SUM(IFNULL(sl_cx_zf_wc.`直发上线完成量`,0)) 直发上线完成量,
                                SUM(IFNULL(sl_cx_zf_wc.`直发上线-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发上线完成量`,0)) 直发上线完成时效
                        FROM (SELECT  币种,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 总订单量
                                FROM  {0} sl_cx
                                WHERE sl_cx.`币种` = '{1}' AND sl_cx.`团队` IN ({2}) AND sl_cx.`父级分类` IS NOT NULL AND sl_cx.`是否改派` = "直发"
                                GROUP BY 币种,年月,物流方式,父级分类,旬
                                ORDER BY 币种,年月
                                ) sl_zong
                        LEFT JOIN
                                (SELECT 币种,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 直发订单量, SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时'
                                FROM  {0} sl_cx_zf
                                WHERE sl_cx_zf.`币种` = '{1}' AND sl_cx_zf.`团队` IN ({2}) AND sl_cx_zf.`父级分类` IS NOT NULL  AND sl_cx_zf.`是否改派` = "直发" AND sl_cx_zf.`仓储扫描时间` is not null
                                GROUP BY 币种,年月,物流方式,父级分类,旬
                                ORDER BY 币种,年月
                            ) sl_zong_zf
                             ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年月` = sl_zong.`年月` AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`父级分类` = sl_zong.`父级分类`  AND sl_zong_zf.`旬` = sl_zong.`旬` 	
                            LEFT JOIN
                                (SELECT 币种,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 直发已签收订单量
                                FROM  {0}	sl_cx_zf_qianshou
                                WHERE sl_cx_zf_qianshou.`币种` = '{1}' AND sl_cx_zf_qianshou.`团队` IN ({2}) AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL AND sl_cx_zf_qianshou.`是否改派` = "直发" AND sl_cx_zf_qianshou.`仓储扫描时间` is not null AND sl_cx_zf_qianshou.`最终状态` = "已签收"
                                GROUP BY 币种,年月,物流方式,父级分类,旬
                                ORDER BY 币种,年月
                            ) sl_cx_zf_qs
                             ON sl_cx_zf_qs.`币种` = sl_zong.`币种`  AND sl_cx_zf_qs.`年月` = sl_zong.`年月`  AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类`  AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 	
                        LEFT JOIN
                                (SELECT 币种,年月,物流方式,父级分类,旬,
                                        COUNT(`订单编号`) 直发出库完成量,
                                        SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)) AS '直发出库-完成时',
                                        COUNT(`订单编号`) 直发下单完成量,
                                        SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)) AS '直发下单-完成时',
                                        COUNT(`订单编号`) 直发出货上线量,
                                        SUM(DATEDIFF(`上线时间`, IFNULL(`仓储扫描时间`,`出货时间`))) AS '直发出货-上线时',
                                        COUNT(`订单编号`) 直发上线完成量,
                                        SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`上线时间`)) AS '直发上线-完成时'
                                FROM  {0}	sl_cx_zf_wancheng
                                WHERE sl_cx_zf_wancheng.`币种` = '{1}' AND sl_cx_zf_wancheng.`团队` IN ({2}) AND sl_cx_zf_wancheng.`父级分类` IS NOT NULL AND sl_cx_zf_wancheng.`是否改派` = "直发" AND sl_cx_zf_wancheng.`最终状态`IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') 
                                GROUP BY 币种,年月,物流方式,父级分类,旬
                                ORDER BY 币种,年月
                            ) sl_cx_zf_wc
                             ON sl_cx_zf_wc.`币种` = sl_zong.`币种`  AND sl_cx_zf_wc.`年月` = sl_zong.`年月` AND sl_cx_zf_wc.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_wc.`父级分类` = sl_zong.`父级分类`  AND sl_cx_zf_wc.`旬` = sl_zong.`旬`
                        GROUP BY sl_zong.年月,sl_zong.物流方式,sl_zong.父级分类,sl_zong.旬
                        with rollup) sl_rb;'''.format('qsb_缓存_month_cp', currency, match[team])
        sqltime4 = '''SELECT sl_rb.`币种`,sl_rb.`年`,sl_rb.`年月`,sl_rb.`物流方式`,sl_rb.`父级分类`,sl_rb.`旬`,
                                sl_rb.`总单量`,
                                sl_rb.`直发下单出库单量`,sl_rb.`直发下单出库时效`,
                                sl_rb.`直发出货上线量`,sl_rb.`直发出货上线时效`,
                                sl_rb.`直发上线完成量`,sl_rb.`直发上线完成时效`,
                                sl_rb.`直发出库完成单量`,sl_rb.`直发出库完成时效`,
                                sl_rb.`直发下单完成时效`,sl_rb.`直发下单完成单量`,
                                sl_rb.`直发已签收订单量` / sl_rb.`直发下单完成单量` AS '签收/完成',
                                sl_rb.`直发已签收订单量`/ sl_rb.`直发下单出库单量` AS '签收/总计'
                    FROM (SELECT sl_zong.币种 币种,IFNULL(sl_zong.年,'合计') 年,IFNULL(sl_zong.年月,'合计') 年月,IFNULL(sl_zong.物流方式,'合计') 物流方式,IFNULL(sl_zong.父级分类,'合计') 父级分类,IFNULL(sl_zong.旬,'合计') 旬,
                                SUM(sl_zong.`总订单量`) 总单量,
                                SUM(IFNULL(sl_cx_zf_qs.`直发已签收订单量`,0)) 直发已签收订单量,
                                SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库单量,
                                SUM(IFNULL(sl_zong_zf.`直发下单-出库时`,0)) / SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库时效,
                                SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成单量,
                                SUM(IFNULL(sl_cx_zf_wc.`直发出库-完成时`,0)) / SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成时效,
                                SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成单量,
                                SUM(IFNULL(sl_cx_zf_wc.`直发下单-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成时效,
                                SUM(IFNULL(sl_cx_zf_wc.`直发出货上线量`,0)) 直发出货上线量,
                                SUM(IFNULL(sl_cx_zf_wc.`直发出货-上线时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发出货上线量`,0)) 直发出货上线时效,
                                SUM(IFNULL(sl_cx_zf_wc.`直发上线完成量`,0)) 直发上线完成量,
                                SUM(IFNULL(sl_cx_zf_wc.`直发上线-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发上线完成量`,0)) 直发上线完成时效
                        FROM (SELECT  币种,年,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 总订单量
                                FROM  {0} sl_cx
                                WHERE sl_cx.`币种` = '{1}' AND sl_cx.`团队` IN ({2}) AND sl_cx.`父级分类` IS NOT NULL AND sl_cx.`是否改派` = "直发"
                                GROUP BY 币种,年,年月,物流方式,父级分类,旬
                                ORDER BY 币种,年,年月
                                ) sl_zong
                        LEFT JOIN
                                (SELECT 币种,年,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 直发订单量, SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时'
                                FROM  {0} sl_cx_zf
                                WHERE sl_cx_zf.`币种` = '{1}' AND sl_cx_zf.`团队` IN ({2}) AND sl_cx_zf.`父级分类` IS NOT NULL  AND sl_cx_zf.`是否改派` = "直发" AND sl_cx_zf.`仓储扫描时间` is not null
                                GROUP BY 币种,年,年月,物流方式,父级分类,旬
                                ORDER BY 币种,年,年月
                            ) sl_zong_zf  ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年` = sl_zong.`年` AND sl_zong_zf.`年月` = sl_zong.`年月` AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`父级分类` = sl_zong.`父级分类`  AND sl_zong_zf.`旬` = sl_zong.`旬` 	
                       LEFT JOIN
                                (SELECT 币种,年,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 直发已签收订单量
                                FROM  {0} sl_cx_zf_qianshou
                                WHERE sl_cx_zf_qianshou.`币种` = '{1}' AND sl_cx_zf_qianshou.`团队` IN ({2}) AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL AND sl_cx_zf_qianshou.`是否改派` = "直发" AND sl_cx_zf_qianshou.`仓储扫描时间` is not null AND sl_cx_zf_qianshou.`最终状态` = "已签收"
                                GROUP BY 币种,年,年月,物流方式,父级分类,旬
                                ORDER BY 币种,年,年月
                            ) sl_cx_zf_qs  ON sl_cx_zf_qs.`币种` = sl_zong.`币种`  AND sl_cx_zf_qs.`年` = sl_zong.`年`  AND sl_cx_zf_qs.`年月` = sl_zong.`年月`  AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类`  AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 	
                       LEFT JOIN
                                (SELECT 币种,年,年月,物流方式,父级分类,旬,
                                        COUNT(`订单编号`) 直发出库完成量,
                                        SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)) AS '直发出库-完成时',
                                        COUNT(`订单编号`) 直发下单完成量,
                                        SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)) AS '直发下单-完成时',
                                        COUNT(`订单编号`) 直发出货上线量,
                                        SUM(DATEDIFF(`上线时间`, IFNULL(`仓储扫描时间`,`出货时间`))) AS '直发出货-上线时',
                                        COUNT(`订单编号`) 直发上线完成量,
                                        SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`上线时间`)) AS '直发上线-完成时'
                                FROM  {0} sl_cx_zf_wancheng
                                WHERE sl_cx_zf_wancheng.`币种` = '{1}' AND sl_cx_zf_wancheng.`团队` IN ({2}) AND sl_cx_zf_wancheng.`父级分类` IS NOT NULL AND sl_cx_zf_wancheng.`是否改派` = "直发" AND sl_cx_zf_wancheng.`最终状态`IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') 
                                GROUP BY 币种,年,年月,物流方式,父级分类,旬
                                ORDER BY 币种,年,年月
                            ) sl_cx_zf_wc ON sl_cx_zf_wc.`币种` = sl_zong.`币种` AND sl_cx_zf_wc.`年` = sl_zong.`年` AND sl_cx_zf_wc.`年月` = sl_zong.`年月` AND sl_cx_zf_wc.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_wc.`父级分类` = sl_zong.`父级分类`  AND sl_cx_zf_wc.`旬` = sl_zong.`旬`
                        GROUP BY  sl_zong.年, sl_zong.年月,sl_zong.物流方式,sl_zong.父级分类,sl_zong.旬
                        with rollup
												) sl_rb;'''.format('qsb_缓存_month_cp', currency, match[team])
        listT.append(sqltime4)
        show_name.append(' 月(各月)时效…………')

        # 月时效（天）---改派 查询
        sqltime7 = '''SELECT 币种,年月,物流方式,父级分类,旬,
                                        总订单量 总单量, 
                                        直发订单量 AS 直发下单出库量,
                                        IFNULL(`直发下单-出库时`,0) / IFNULL(`直发订单量`,0) AS 下单出库时效,
                                        直发出库完成量,
                                        IFNULL(`直发出库-完成时`,0) / IFNULL(`直发出库完成量`,0) 出库完成时效,
                                        直发下单完成量,
                                        IFNULL(`直发下单-完成时`,0) / IFNULL(`直发下单完成量`,0) 下单完成时效,
                                        直发出货上线量,
                                        IFNULL(`直发出货-上线时`,0) / IFNULL(`直发出货上线量`,0) 出货上线时效,
                                        直发上线完成量,
                                        IFNULL(`直发上线-完成时`,0) / IFNULL(`直发上线完成量`,0) 上线完成时效,					
                                        直发已签收订单量 / 直发下单完成量 AS '签收/完成',
                                        直发已签收订单量 / 直发出库完成量 AS '签收/总计'
                                FROM( SELECT IFNULL(币种,'合计') 币种,IFNULL(年月,'合计') 年月,IFNULL(物流方式,'合计') 物流方式,IFNULL(父级分类,'合计') 父级分类,IFNULL(旬,'合计') 旬,
                                            NULL 总订单量, 
                                            SUM(IF(`最终状态` = '已签收',1,0)) AS 直发已签收订单量,
                                            COUNT(`订单编号`) 直发订单量,
                                            SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时',
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发出库完成量,
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`),0)) AS '直发出库-完成时',
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发下单完成量,
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`),0)) AS '直发下单-完成时',
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发出货上线量,
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(`上线时间`, IFNULL(`仓储扫描时间`,`出货时间`)),0)) AS '直发出货-上线时',
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发上线完成量,
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`上线时间`),0)) AS '直发上线-完成时'
                                    FROM {0} sl_cx
                                    WHERE (sl_cx.`记录时间`= '{1}' AND sl_cx.`年月` = '{2}' OR sl_cx.`记录时间`= '{3}' AND sl_cx.`年月` = '{4}')
                                        AND sl_cx.`币种` = '{5}' AND sl_cx.`团队` IN ({6})
                                        AND sl_cx.`是否改派` = "改派"
                                        AND sl_cx.`父级分类` IS NOT NULL 
                                        AND sl_cx.`仓储扫描时间` IS NOT NULL 
                                    GROUP BY 年月,物流方式,旬
                                with rollup ) sl;'''.format(family, now_month, now_month_new, last_month, last_month_new, currency, match[team])
        listT.append(sqltime7)
        show_name.append(' 月（改派天）时效…………')
        # 月时效（旬）---改派 查询
        sqltime71 = '''SELECT 币种,年月,旬,物流方式,父级分类,
                                        总订单量 总单量, 
                                        直发订单量 AS 直发下单出库量,
                                        IFNULL(`直发下单-出库时`,0) / IFNULL(`直发订单量`,0) as 下单出库时效,
                                        直发出库完成量,
                                        IFNULL(`直发出库-完成时`,0) / IFNULL(`直发出库完成量`,0) 出库完成时效,
                                        直发下单完成量,
                                        IFNULL(`直发下单-完成时`,0) / IFNULL(`直发下单完成量`,0) 下单完成时效,
                                        直发出货上线量,
                                        IFNULL(`直发出货-上线时`,0) / IFNULL(`直发出货上线量`,0) 出货上线时效,
                                        直发上线完成量,
                                        IFNULL(`直发上线-完成时`,0) / IFNULL(`直发上线完成量`,0) 上线完成时效
                                FROM( SELECT IFNULL(币种,'合计') 币种,IFNULL(年月,'合计') 年月,IFNULL(旬,'合计') 旬,IFNULL(物流方式,'合计') 物流方式,IFNULL(父级分类,'合计') 父级分类,
                                            NULL 总订单量, 
                                            SUM(IF(`最终状态` = '已签收',1,0))  as 直发已签收订单量,
                                            COUNT(`订单编号`) 直发订单量,
                                            SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时',
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发出库完成量,
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`),0)) AS '直发出库-完成时',
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发下单完成量,
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`),0)) AS '直发下单-完成时',
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发出货上线量,
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(`上线时间`, IFNULL(`仓储扫描时间`,`出货时间`)),0)) AS '直发出货-上线时',
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0))  as 直发上线完成量,
                                            SUM(IF(`最终状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`上线时间`),0)) AS '直发上线-完成时'
                                    FROM {0} sl_cx
                                    WHERE (sl_cx.`记录时间`= '{1}' AND sl_cx.`年月` = '{2}' OR sl_cx.`记录时间`= '{3}' AND sl_cx.`年月` = '{4}')
                                        AND sl_cx.`币种` = '{5}' AND sl_cx.`团队` IN ({6})
                                        AND sl_cx.`是否改派` = "改派"
                                        AND sl_cx.`父级分类` IS NOT NULL 
                                        AND sl_cx.`仓储扫描时间` IS NOT NULL 
                                    GROUP BY 年月,旬,物流方式,父级分类
                                    with rollup ) sl;'''.format(family, now_month, now_month_new, last_month, last_month_new, currency, match[team])
        listT.append(sqltime71)
        show_name.append(' 月（改派旬）时效…………')
        # 月时效(各月)---改派 查询
        sqltime72 = '''SELECT sl_rb.`币种`,sl_rb.`年月`,sl_rb.`物流方式`,sl_rb.`父级分类`,sl_rb.`旬`,
                                        sl_rb.`总单量`,
                                        sl_rb.`直发下单出库单量`,sl_rb.`直发下单出库时效`,
                                        sl_rb.`直发出货上线量`,sl_rb.`直发出货上线时效`,
                                        sl_rb.`直发上线完成量`,sl_rb.`直发上线完成时效`,
                                        sl_rb.`直发出库完成单量`,sl_rb.`直发出库完成时效`,
                                        sl_rb.`直发下单完成时效`,sl_rb.`直发下单完成单量`,
                                        sl_rb.`直发已签收订单量` / sl_rb.`直发下单完成单量` AS '签收/完成',
                                        sl_rb.`直发已签收订单量`/ sl_rb.`直发下单出库单量` AS '签收/总计'
                            FROM (SELECT sl_zong.币种 币种,IFNULL(sl_zong.年月,'合计') 年月,IFNULL(sl_zong.物流方式,'合计') 物流方式,IFNULL(sl_zong.父级分类,'合计') 父级分类,IFNULL(sl_zong.旬,'合计') 旬,
                                        SUM(sl_zong.`总订单量`) 总单量,
                                        SUM(IFNULL(sl_cx_zf_qs.`直发已签收订单量`,0)) 直发已签收订单量,
                                        SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库单量,
                                        SUM(IFNULL(sl_zong_zf.`直发下单-出库时`,0)) / SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库时效,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成单量,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发出库-完成时`,0)) / SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成时效,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成单量,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发下单-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成时效,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发出货上线量`,0)) 直发出货上线量,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发出货-上线时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发出货上线量`,0)) 直发出货上线时效,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发上线完成量`,0)) 直发上线完成量,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发上线-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发上线完成量`,0)) 直发上线完成时效
                                FROM (SELECT  币种,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 总订单量
                                        FROM  {0} sl_cx
                                        WHERE sl_cx.`币种` = '{1}' AND sl_cx.`团队` IN ({2}) AND sl_cx.`父级分类` IS NOT NULL AND sl_cx.`是否改派` = "改派"
                                        GROUP BY 币种,年月,物流方式,父级分类,旬
                                        ORDER BY 币种,年月
                                        ) sl_zong
                                LEFT JOIN
                                        (SELECT 币种,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 直发订单量, SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时'
                                        FROM  {0} sl_cx_zf
                                        WHERE sl_cx_zf.`币种` = '{1}' AND sl_cx_zf.`团队` IN ({2}) AND sl_cx_zf.`父级分类` IS NOT NULL  AND sl_cx_zf.`是否改派` = "改派" AND sl_cx_zf.`仓储扫描时间` is not null
                                        GROUP BY 币种,年月,物流方式,父级分类,旬
                                        ORDER BY 币种,年月
                                    ) sl_zong_zf
                                     ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年月` = sl_zong.`年月` AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`父级分类` = sl_zong.`父级分类`  AND sl_zong_zf.`旬` = sl_zong.`旬` 	
                                    LEFT JOIN
                                        (SELECT 币种,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 直发已签收订单量
                                        FROM  {0}	sl_cx_zf_qianshou
                                        WHERE sl_cx_zf_qianshou.`币种` = '{1}' AND sl_cx_zf_qianshou.`团队` IN ({2}) AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL AND sl_cx_zf_qianshou.`是否改派` = "改派" AND sl_cx_zf_qianshou.`仓储扫描时间` is not null AND sl_cx_zf_qianshou.`最终状态` = "已签收"
                                        GROUP BY 币种,年月,物流方式,父级分类,旬
                                        ORDER BY 币种,年月
                                    ) sl_cx_zf_qs
                                     ON sl_cx_zf_qs.`币种` = sl_zong.`币种`  AND sl_cx_zf_qs.`年月` = sl_zong.`年月`  AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类`  AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 	
                                LEFT JOIN
                                        (SELECT 币种,年月,物流方式,父级分类,旬,
                                                COUNT(`订单编号`) 直发出库完成量,
                                                SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)) AS '直发出库-完成时',
                                                COUNT(`订单编号`) 直发下单完成量,
                                                SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)) AS '直发下单-完成时',
                                                COUNT(`订单编号`) 直发出货上线量,
                                                SUM(DATEDIFF(`上线时间`, IFNULL(`仓储扫描时间`,`出货时间`))) AS '直发出货-上线时',
                                                COUNT(`订单编号`) 直发上线完成量,
                                                SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`上线时间`)) AS '直发上线-完成时'
                                        FROM  {0}	sl_cx_zf_wancheng
                                        WHERE sl_cx_zf_wancheng.`币种` = '{1}' AND sl_cx_zf_wancheng.`团队` IN ({2}) AND sl_cx_zf_wancheng.`父级分类` IS NOT NULL AND sl_cx_zf_wancheng.`是否改派` = "改派" AND sl_cx_zf_wancheng.`最终状态`IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') 
                                        GROUP BY 币种,年月,物流方式,父级分类,旬
                                        ORDER BY 币种,年月
                                    ) sl_cx_zf_wc
                                     ON sl_cx_zf_wc.`币种` = sl_zong.`币种`  AND sl_cx_zf_wc.`年月` = sl_zong.`年月` AND sl_cx_zf_wc.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_wc.`父级分类` = sl_zong.`父级分类`  AND sl_cx_zf_wc.`旬` = sl_zong.`旬`
                                GROUP BY sl_zong.年月,sl_zong.物流方式,sl_zong.父级分类,sl_zong.旬
                                with rollup) sl_rb;'''.format('qsb_缓存_month_cp', currency, match[team])
        sqltime72 = '''SELECT sl_rb.`币种`,sl_rb.`年`,sl_rb.`年月`,sl_rb.`物流方式`,sl_rb.`父级分类`,sl_rb.`旬`,
                                        sl_rb.`总单量`,
                                        sl_rb.`直发下单出库单量`,sl_rb.`直发下单出库时效`,
                                        sl_rb.`直发出货上线量`,sl_rb.`直发出货上线时效`,
                                        sl_rb.`直发上线完成量`,sl_rb.`直发上线完成时效`,
                                        sl_rb.`直发出库完成单量`,sl_rb.`直发出库完成时效`,
                                        sl_rb.`直发下单完成时效`,sl_rb.`直发下单完成单量`,
                                        sl_rb.`直发已签收订单量` / sl_rb.`直发下单完成单量` AS '签收/完成',
                                        sl_rb.`直发已签收订单量`/ sl_rb.`直发下单出库单量` AS '签收/总计'
                            FROM (SELECT sl_zong.币种 币种,IFNULL(sl_zong.年,'合计') 年,IFNULL(sl_zong.年月,'合计') 年月,IFNULL(sl_zong.物流方式,'合计') 物流方式,IFNULL(sl_zong.父级分类,'合计') 父级分类,IFNULL(sl_zong.旬,'合计') 旬,
                                        SUM(sl_zong.`总订单量`) 总单量,
                                        SUM(IFNULL(sl_cx_zf_qs.`直发已签收订单量`,0)) 直发已签收订单量,
                                        SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库单量,
                                        SUM(IFNULL(sl_zong_zf.`直发下单-出库时`,0)) / SUM(IFNULL(sl_zong_zf.`直发订单量`,0)) 直发下单出库时效,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成单量,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发出库-完成时`,0)) / SUM(IFNULL(sl_cx_zf_wc.`直发出库完成量`,0)) 直发出库完成时效,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成单量,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发下单-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发下单完成量`,0)) 直发下单完成时效,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发出货上线量`,0)) 直发出货上线量,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发出货-上线时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发出货上线量`,0)) 直发出货上线时效,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发上线完成量`,0)) 直发上线完成量,
                                        SUM(IFNULL(sl_cx_zf_wc.`直发上线-完成时`,0)) /SUM(IFNULL(sl_cx_zf_wc.`直发上线完成量`,0)) 直发上线完成时效
                                FROM (SELECT  币种,年,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 总订单量
                                        FROM  {0} sl_cx
                                        WHERE sl_cx.`币种` = '{1}' AND sl_cx.`团队` IN ({2}) AND sl_cx.`父级分类` IS NOT NULL AND sl_cx.`是否改派` = "改派"
                                        GROUP BY 币种,年,年月,物流方式,父级分类,旬
                                        ORDER BY 币种,年,年月
                                        ) sl_zong
                                LEFT JOIN
                                        (SELECT 币种,年,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 直发订单量, SUM(DATEDIFF(`仓储扫描时间`,`下单时间`)) AS '直发下单-出库时'
                                        FROM  {0} sl_cx_zf
                                        WHERE sl_cx_zf.`币种` = '{1}' AND sl_cx_zf.`团队` IN ({2}) AND sl_cx_zf.`父级分类` IS NOT NULL  AND sl_cx_zf.`是否改派` = "改派" AND sl_cx_zf.`仓储扫描时间` is not null
                                        GROUP BY 币种,年,年月,物流方式,父级分类,旬
                                        ORDER BY 币种,年,年月
                                    ) sl_zong_zf  ON sl_zong_zf.`币种` = sl_zong.`币种` AND sl_zong_zf.`年` = sl_zong.`年` AND sl_zong_zf.`年月` = sl_zong.`年月` AND sl_zong_zf.`物流方式` = sl_zong.`物流方式` AND sl_zong_zf.`父级分类` = sl_zong.`父级分类`  AND sl_zong_zf.`旬` = sl_zong.`旬` 	
                               LEFT JOIN
                                        (SELECT 币种,年,年月,物流方式,父级分类,旬,COUNT(`订单编号`) 直发已签收订单量
                                        FROM  {0} sl_cx_zf_qianshou
                                        WHERE sl_cx_zf_qianshou.`币种` = '{1}' AND sl_cx_zf_qianshou.`团队` IN ({2}) AND sl_cx_zf_qianshou.`父级分类` IS NOT NULL AND sl_cx_zf_qianshou.`是否改派` = "改派" AND sl_cx_zf_qianshou.`仓储扫描时间` is not null AND sl_cx_zf_qianshou.`最终状态` = "已签收"
                                        GROUP BY 币种,年,年月,物流方式,父级分类,旬
                                        ORDER BY 币种,年,年月
                                    ) sl_cx_zf_qs  ON sl_cx_zf_qs.`币种` = sl_zong.`币种`  AND sl_cx_zf_qs.`年` = sl_zong.`年`  AND sl_cx_zf_qs.`年月` = sl_zong.`年月`  AND sl_cx_zf_qs.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_qs.`父级分类` = sl_zong.`父级分类`  AND sl_cx_zf_qs.`旬` = sl_zong.`旬` 	
                               LEFT JOIN
                                        (SELECT 币种,年,年月,物流方式,父级分类,旬,
                                                COUNT(`订单编号`) 直发出库完成量,
                                                SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`仓储扫描时间`)) AS '直发出库-完成时',
                                                COUNT(`订单编号`) 直发下单完成量,
                                                SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`下单时间`)) AS '直发下单-完成时',
                                                COUNT(`订单编号`) 直发出货上线量,
                                                SUM(DATEDIFF(`上线时间`, IFNULL(`仓储扫描时间`,`出货时间`))) AS '直发出货-上线时',
                                                COUNT(`订单编号`) 直发上线完成量,
                                                SUM(DATEDIFF(IFNULL(`完结状态时间`,`状态时间`),`上线时间`)) AS '直发上线-完成时'
                                        FROM  {0} sl_cx_zf_wancheng
                                        WHERE sl_cx_zf_wancheng.`币种` = '{1}' AND sl_cx_zf_wancheng.`团队` IN ({2}) AND sl_cx_zf_wancheng.`父级分类` IS NOT NULL AND sl_cx_zf_wancheng.`是否改派` = "改派" AND sl_cx_zf_wancheng.`最终状态`IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件') 
                                        GROUP BY 币种,年,年月,物流方式,父级分类,旬
                                        ORDER BY 币种,年,年月
                                    ) sl_cx_zf_wc ON sl_cx_zf_wc.`币种` = sl_zong.`币种` AND sl_cx_zf_wc.`年` = sl_zong.`年` AND sl_cx_zf_wc.`年月` = sl_zong.`年月` AND sl_cx_zf_wc.`物流方式` = sl_zong.`物流方式` AND sl_cx_zf_wc.`父级分类` = sl_zong.`父级分类`  AND sl_cx_zf_wc.`旬` = sl_zong.`旬`
                                GROUP BY  sl_zong.年, sl_zong.年月,sl_zong.物流方式,sl_zong.父级分类,sl_zong.旬
                                with rollup
        												) sl_rb;;'''.format('qsb_缓存_month_cp', currency, match[team])
        listT.append(sqltime72)
        show_name.append(' 月(改派各月)时效…………')

        listTValue = []  # 查询sql的结果 存放池
        for i, sql in enumerate(listT):
            print('正在获取 ' + team + show_name[i])
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print(df)
            columns = list(df.columns)  # 获取数据的标题名，转为列表
            columns_value = ['采购/销售额', '直发采购/销售额', '运费占比', '手续费占比', '金额签收/完成', '金额签收/总计', '金额完成占比', '数量签收/完成', '数量完成占比',
                             '签收/完成', '签收/总计', '完成占比', '总签收/完成', '总签收/总计', '退款率', '总完成占比', '直发签收/完成', '直发签收/总计', '直发完成占比',
                             '改派签收/完成', '改派签收/总计', '改派完成占比', '总签收/完成(金额)', '总签收/总计(金额)', '退款率(金额)', '总完成占比(金额)', '直发签收/完成(金额)',
                             '直发签收/总计(金额)', '直发完成占比(金额)', '改派签收/完成(金额)', '改派签收/总计(金额)', '改派完成占比(金额)', '订单品类占比', '直发采购额/销售额',
                             '花费占比', '总成本', '利润率', '改派占比', '采购占比', '广告占比', '总成本占比', '签收/完成', '签收/总计', '完成占比']
            for column_val in columns_value:
                if column_val in columns:
                    try:
                        df[column_val] = df[column_val].fillna(value=0)
                        df[column_val] = df[column_val].apply(lambda x: format(x, '.2%'))
                    except Exception as e:
                        print('修改失败：', str(Exception) + str(e) + df[column_val])
            listTValue.append(df)
        print('查询耗时：', datetime.datetime.now() - start)
        today = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        sheet_name = ['签率(天)_', '签率(月)_', '签率(旬)_', '签率(总)_', '物流(天)_', '物流(月)_', '时效(天)_', '时效(旬)_', '时效(总)_', '时效(改派天)_', '时效(改派旬)_', '时效(改派总)_']  # 生成的工作表的表名
        file_Path = []  # 发送邮箱文件使用
        filePath = ''
        if "品牌" in team:
            filePath = 'F:\\查询\\品牌监控\\{}{} {} 监控表.xlsx'.format(today, team, ready)
        elif "神龙" in team or "火凤凰" in team or "小虎队" in team or "港台" in team:
            filePath = 'F:\\查询\\港台监控\\{}{} {} 监控表.xlsx'.format(today, team, ready)
        if os.path.exists(filePath):                            # 判断是否有需要的表格，进行初始化创建
            print("正在清除重复文件......")
            os.remove(filePath)
        print("正在创建文件......")
        df0 = pd.DataFrame([])                                  # 创建空的dataframe数据框
        df0.to_excel(filePath, index=False)                     # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        print('正在写入excel…………')
        writer = pd.ExcelWriter(filePath, engine='openpyxl')    # 初始化写入对象
        book = load_workbook(filePath)                          # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book                                      # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listTValue)):
            listTValue[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i] + team, index=False)
        if 'Sheet1' in book.sheetnames:                         # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('正在运行宏…………')
        app = xl.App(visible=False, add_book=False)             # 运行宏调整
        app.display_alerts = False
        wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
        wbsht1 = app.books.open(filePath)
        if ready == '本期宏':
            wbsht.macro('sl_总监控运行')()
        elif ready == '本期上月宏':
            wbsht.macro('sl_总监控运行3')()
        else:
            wbsht.macro('sl_总监控运行3')()
        wbsht1.save()
        wbsht1.close()
        wbsht.close()
        app.quit()
        print('输出(监控)文件成功…………')
        file_Path.append(filePath)
        if team in ['品牌-日本', '品牌-台湾', '品牌-香港', '品牌-马来西亚', '品牌-新加坡', '品牌-菲律宾']:
            self.e.send('{} {}监控表.xlsx'.format(today, team), file_Path,
                        emailAdd[team])
        print('处理耗时：', datetime.datetime.now() - start)


    # 获取签收表内容
    def readForm(self, team, startday):
        match3 = {'品牌': 'slsc',
                  '火凤凰-港台': 'slgat_hfh',
                  '神龙-港台': 'slgat',
                  '小虎队-港台': 'slgat_jp',
                  '港台': 'slgat'}
        start = datetime.datetime.now()
        family = ""
        if team in ('神龙-港台', '火凤凰-港台', '小虎队-港台', '红杉-港台', '金狮-港台', '神龙-主页运营1组'):
            family = 'gat'
        elif team in ('品牌签收表'):
            family = 'slsc'
        path = r'F:\\查询\\订单数据'
        dirs = os.listdir(path=path)
        for dir in dirs:            # ---读取execl文件---
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                if team in dir and startday in dir:
                    print(filePath)
                    self.wbsheet(filePath, family, startday)
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheet(self, filePath, team, startday):
        print('---正在获取签收表的详情++++++')
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    # db = sht.used_range.value
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    db = db[['年月', '旬', '日期', '团队', '币种', '订单来源', '订单编号', '出货时间', '状态时间', '上线时间', '最终状态', '是否改派',
                             '物流方式', '产品id', '父级分类', '二级分类', '三级分类', '下单时间', '审核时间', '仓储扫描时间', '完结状态时间', '价格RMB']]
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在导入：' + sht.name + ' 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    self.writeCache(db)                                         # 将返回的dateFrame导入数据库的临时表
                    print('++++正在更新：' + sht.name + '--->>>到总订单')
                    self.replaceSql(team, list(db.columns), startday)     # 将数据库的临时表替换进指定的总表
                    print('++++----->>>' + sht.name + '：订单更新完成++++')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()
    # 写入临时缓存表
    def writeCache(self, dataFrame):
        dataFrame.to_sql('qsb_缓存', con=self.engine1, index=False, if_exists='replace')
    # 写入总表
    def replaceSql(self, team, dfColumns, startday):
        columns = list(dfColumns)
        columns = ', '.join(columns)
        try:
            sql = '''INSERT IGNORE INTO qsb_{0}({1}, 记录时间) SELECT *, '{2}' 记录时间 FROM qsb_缓存; '''.format(team, columns, startday)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=2000)
        except Exception as e:
            print('插入失败：', str(Exception) + str(e))


if __name__ == '__main__':
    m = SltemMonitoring()
    start: datetime = datetime.datetime.now()
    match1 = {'gat': '港台',
              'slsc': '品牌'}
    # -----------------------------------------------监控运行的主要程序和步骤-----------------------------------------
    # 获取签收表内容（一）qsb_slgat
    last_month = '2022.06.20'
    now_month = '2022.07.19'
    # for team in ['神龙-港台', '火凤凰-港台', '小虎队-港台', '红杉-港台', '金狮-港台', '神龙-主页运营1组']:
        # m.readForm(team, last_month)      # 上月上传
        # m.readForm(team, now_month)       # 本月上传

    # 测试监控运行（二）-- 第一种手动方式
    m.order_Monitoring('港台')        # 各月缓存（整体一）
    # for team in ['神龙-台湾', '神龙-香港', '神龙运营1组-台湾', '火凤凰-台湾', '火凤凰-香港', '小虎队-台湾']:
    for team in ['神龙-台湾', '神龙-香港', '火凤凰-台湾', '火凤凰-香港', '神龙运营1组-台湾']:
    # for team in ['神龙-台湾', '神龙-香港', '神龙运营1组-台湾']:
    # for team in ['港台-台湾']:
        now_month = now_month.replace('.', '-')           # 修改配置时间
        last_month = last_month.replace('.', '-')
        m.sl_Monitoring(team, now_month, last_month, '本期宏')      # 输出数据--每月正常使用的时间（二）
        # m.sl_Monitoring(team, now_month, last_month, '本期上月宏')      # 输出数据--每月正常使用的时间（二）
        # m.sl_Monitoring(team, now_month, last_month, '上期宏')      # 输出数据--每月正常使用的时间（二）

    # for team in ['火凤凰-台湾', '火凤凰-香港', '神龙-台湾', '神龙-香港', '神龙运营1组-台湾', '港台-台湾']:
    #     now_month = now_month.replace('.', '-')           # 修改配置时间
    #     last_month = last_month.replace('.', '-')
    #     m.sl_Monitoring(team, now_month, last_month, '本期上月宏')      # 输出数据--每月正常使用的时间（二）


    # 测试监控运行（三）-- 第二种自动方式
    # m.match_time('港台')      # 检测时间
    # win32api.MessageBox(0, "注意:>>>    程序运行结束， 请查看表  ！！！", "提 醒", win32con.MB_OK)
