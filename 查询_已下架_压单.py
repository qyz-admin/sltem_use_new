import pandas as pd
import os
import datetime
import time
import xlwings
import xlsxwriter
import math
import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel
import win32api,win32con
import win32com.client as win32

from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色
from bs4 import BeautifulSoup # 抓标签里面元素的方法

# -*- coding:utf-8 -*-
class QueryTwoLower(Settings, Settings_sso):
    def __init__(self, userMobile, password, login_TmpCode,handle):
        Settings.__init__(self)
        Settings_sso.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        # self.sso_online_cang()
        self.bulid_file()
        if handle == '手动':
            self.sso_online_cang_handle(login_TmpCode)
        else:
            self.sso_online_cang_auto()

        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
    #  登录后台中
    def _online(self):  # 登录系统保持会话状态
        print('正在登录后台系统中......')
        # print('第一阶段获取-钉钉用户信息......')
        url = r'https://login.dingtalk.com/login/login_with_pwd'
        data = {'mobile': self.userMobile,
                'pwd': self.password,
                'goto': 'https://oapi.dingtalk.com/connect/oauth2/sns_authorize?appid=dingoajqpi5bp2kfhekcqm&response_type=code&scope=snsapi_login&state=STATE&redirect_uri=http://gsso.giikin.com/admin/dingtalk_service/getunionidbytempcode',
                'pdmToken': '',
                'araAppkey': '1917',
                'araToken': '0#19171628645731266586976965831628645747396525G1E2B0816DEBF96BC4199761B6A1F3C0FCD91FB',
                'araScene': 'login',
                'captchaImgCode': '',
                'captchaSessionId': '',
                'type': 'h5'}
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Origin': 'https://login.dingtalk.com',
                    'Referer': 'https://login.dingtalk.com/'}
        req = self.session.post(url=url, headers=r_header, data=data, allow_redirects=False)
        req = req.json()
        # print(req)
        req_url = req['data']
        loginTmpCode = req_url.split('loginTmpCode=')[1]        # 获取loginTmpCode值
        # print(loginTmpCode)
        # print('+++已获取loginTmpCode值+++')

        time.sleep(1)
        # print('第二阶段请求-登录页面......')
        url = r'http://gsso.giikin.com/admin/dingtalk_service/gettempcodebylogin.html'
        data = {'tmpCode': loginTmpCode,
                'system': 1,
                'url': '',
                'ticker': '',
                'companyId': 1}
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Origin': 'https://login.dingtalk.com',
                    'Referer': 'http://gsso.giikin.com/admin/login/logout.html'}
        req = self.session.post(url=url, headers=r_header, data=data, allow_redirects=False)
        # print(req.text)
        # print('+++请求登录页面url成功+++')

        time.sleep(1)
        # print('第三阶段请求-dingtalk服务器......')
        # print('（一）加载dingtalk_service跳转页面......')
        url = req.text
        data = {'tmpCode': loginTmpCode,
                'system': 1,
                'url': '',
                'ticker': '',
                'companyId': 1}
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': 'http://gsso.giikin.com/'}
        req = self.session.get(url=url, headers=r_header, data=data, allow_redirects=False)
        # print(req.headers)
        gimp = req.headers['Location']
        # print('+++已获取跳转页面+++')
        time.sleep(1)
        # print('（二）请求dingtalk_service的cookie值......')
        url = gimp
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': 'http://gsso.giikin.com/'}
        req = self.session.get(url=url, headers=r_header, allow_redirects=False)
        # print(req)
        # print('+++已获取cookie值+++')

        time.sleep(1)
        # print('第四阶段页面-重定向跳转中......')
        # print('（一）加载chooselogin.html页面......')
        url = r'http://gsso.giikin.com/admin/login_by_dingtalk/chooselogin.html'
        data = {'user_id': 1343}
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': gimp,
                    'Origin': 'http://gsso.giikin.com'}
        req = self.session.post(url=url, headers=r_header, data=data, allow_redirects=False)
        # print(req.headers)
        index = req.headers['Location']
        # print('+++已获取gimp.giikin.com页面')
        time.sleep(1)
        # print('（二）加载gimp.giikin.com页面......')
        url = index
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': index}
        req = self.session.get(url=url, headers=r_header, allow_redirects=False)
        # print(req.headers)
        index2 = req.headers['Location']
        # print('+++已获取index.html页面')

        time.sleep(1)
        # print('（三）加载index.html页面......')
        url = 'http://gimp.giikin.com/' + index2
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': 'http://gsso.giikin.com/'}
        req = self.session.get(url=url, headers=r_header, allow_redirects=False)
        # print(req.headers)
        index_system = req.headers['Location']
        # print('+++已获取index.html?_system=18正式页面')

        time.sleep(1)
        # print('第五阶段正式页面-重定向跳转中......')
        # print('（一）加载index.html?_system页面......')
        url = index_system
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': 'http://gsso.giikin.com/'}
        req = self.session.get(url=url, headers=r_header, allow_redirects=False)
        # print(req.headers)
        index_system2 = req.headers['Location']
        # print('+++已获取index.html?_ticker=页面......')
        time.sleep(1)
        # print('（二）加载index.html?_ticker=页面......')
        url = index_system2
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'Referer': 'http://gsso.giikin.com/'}
        req = self.session.get(url=url, headers=r_header, allow_redirects=False)
        # print(req)
        # print(req.headers)
        print('++++++已成功登录++++++')

    def readFile(self,select):
        path = ''
        if select == 1:
            path = r'F:\神龙签收率\(未发货) 直发-仓库-压单\每日压单核实汇总'
        elif select == 2:
            path = r'D:\Users\Administrator\Desktop\需要用到的文件\数据库'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                rq = ''
                if select == 1:
                    rq = (dir.split('压单')[0]).strip()
                    rq = datetime.datetime.strptime(rq, '%Y.%m.%d')
                    rq = rq.strftime('%Y.%m.%d')
                    self._readFile(filePath, rq)
                elif select == 2:
                    self._readFile_select(filePath, rq)
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                wb = excel.Workbooks.Open(filePath)
                file_path = os.path.join(path, "~$ " + dir)
                wb.SaveAs(file_path, FileFormat=51)              # FileFormat = 51 is for .xlsx extension
                wb.Close()                                      # FileFormat = 56 is for .xls extension
                excel.Application.Quit()
                os.remove(filePath)
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def _readFile(self, filePath, rq):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    # print(db.columns)
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在导入查询：' + sht.name + '表； 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    if '处理时间' not in db.columns:
                        db.insert(0, '处理时间', rq)
                    db = db[['订单编号', '处理时间', '备注（压单核实是否需要）','处理人']]
                    db.rename(columns={'备注（压单核实是否需要）': '处理结果'}, inplace=True)
                    db.dropna(axis=0, subset=['处理结果'], how='any', inplace=True)
                    db.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
                    # sql = '''update 压单表 a, customer b set a.`处理时间` = b.`处理时间`, a.`处理结果` = b.`处理结果` where a.`订单编号`= b.`订单编号`;'''
                    sql = '''REPLACE INTO 压单表_已核实(订单编号,处理时间,处理结果,处理人, 记录时间) 
                            SELECT 订单编号,处理时间,处理结果,IF(处理人 = '' OR 处理人 IS NULL,'-',处理人) 处理人, NOW() 记录时间 
                            FROM customer;'''
                    pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                    print('++++成功导入：' + sht.name + '--->>>到压单表')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()

    # 工作表的订单信息
    def _readFile_select(self, filePath, rq):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    # print(db.columns)
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在导入更新：' + sht.name + '表； 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    if '提货物流' not in db.columns:
                        db.insert(0, '提货物流', '立邦国际')
                    db = db[['提货物流', '出貨日期', '件數', '主號','航班號','航班情况','清關情況','全清時間', '出貨日期']]
                    # db.rename(columns={'备注（压单核实是否需要）': '处理结果'}, inplace=True)
                    # db.dropna(axis=0, subset=['处理结果'], how='any', inplace=True)
                    db.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
                    sql = '''update gat_take_delivery a, customer b 
                            set a.`主號` = IF(b.`主號` = '' or  b.`主號` is NULL, a.`主號`, b.`主號`),
                                a.`航班號` = IF(b.`航班號` = '' or  b.`航班號` is NULL, a.`航班號`, b.`航班號`)
                            where a.`提货日期`= b.`出貨日期` and a.`提货物流`= b.`提货物流`;'''
                    # sql = '''REPLACE INTO 压单表_已核实(订单编号,处理时间,处理结果,处理人, 记录时间)
                    #         SELECT 订单编号,处理时间,处理结果,IF(处理人 = '' OR 处理人 IS NULL,'-',处理人) 处理人, NOW() 记录时间
                    #         FROM customer;'''     出貨日期	件數	重量	主號	航班號	'航班情况',	'清關情況',	'全清時間',
                    pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                    print('++++成功更新：' + sht.name + '--->>>到头程物流表')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()

    # 查询压单（仓储的获取）
    def order_spec(self):  # 进入压单检索界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        timeStart = ((datetime.datetime.now() + datetime.timedelta(days=1)) - relativedelta(months=2)).strftime('%Y-%m-%d')
        timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
        print('正在查询 港台 压单订单信息中......')
        url = r'http://gwms-v3.giikin.cn/order/pressure/index'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': '1',
                'limit': 500,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        print('++++++本次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        if max_count != [] or max_count != 0:
            # 首次查询
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            df = pd.json_normalize(ordersdict)
            # 剩余查询
            if max_count > 500:
                in_count = math.ceil(max_count/500)
                dlist = []
                n = 1
                while n < in_count:  # 这里用到了一个while循环，穿越过来的
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                    data = self._order_spec(timeStart, timeEnd, n)                     # 分页获取详情
                    dlist.append(data)
                dp = df.append(dlist, ignore_index=True)
            else:
                dp = df
            print('正在写入......')
            dp = dp[['order_number', 'goods_id', 'goods_name', 'currency_id', 'area_id', 'ydtime', 'purid', 'other_reason', 'buyer', 'intime', 'addtime', 'is_lower', 'below_time', 'cate']]
            dp.columns = ['订单编号', '产品ID', '产品名称', '币种', '团队', '反馈时间', '压单原因', '其他原因', '采购员', '入库时间', '下单时间', '是否下架', '下架时间', '品类']
            dp = dp[(dp['币种'].str.contains('港币|台币', na=False))]
            # print(dp)
            dp.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 压单表(订单编号,产品ID,产品名称,币种,团队, 反馈时间, 压单原因, 其他原因, 采购员, 入库时间, 下单时间, 是否下架, 下架时间,记录时间) 
                    SELECT 订单编号,产品ID,产品名称,币种,团队, 反馈时间, 压单原因, 其他原因, 采购员, 入库时间, 下单时间, 是否下架, IF(下架时间 = '',NULL,下架时间) 下架时间, NOW() 记录时间
                    FROM customer'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('共有 ' + str(len(dp)) + '条 成功写入数据库+++++++')

            print('正在获取 压单反馈 信息中......')
            time_path: datetime = datetime.datetime.now()
            mkpath = r"F:\神龙签收率\(未发货) 直发-仓库-压单\\" + time_path.strftime('%m.%d')
            sql = '''SELECT s.*,s1.处理结果,s1.处理时间,NULL 备注,DATEDIFF(curdate(),入库时间) 压单天数,DATE_FORMAT(入库时间,'%Y-%m-%d') 入库
                    FROM ( SELECT * FROM 压单表 g WHERE g.`记录时间` >= CURDATE() and g.是否下架 <> '已下架'
                    ) s
                    LEFT JOIN (SELECT *
							    FROM 压单表_已核实
								WHERE id IN (SELECT MAX(id) FROM 压单表_已核实 y WHERE y.处理时间 >= DATE_SUB(CURDATE(), INTERVAL 2 month) GROUP BY 订单编号) 
					) s1 ON s.订单编号 = s1.订单编号
					ORDER BY 压单天数;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            isExists = os.path.exists(mkpath)
            if not isExists:
                os.makedirs(mkpath)
            else:
                print(mkpath + ' 目录已存在')
            file_path = mkpath + '\\压单反馈 {0}.xlsx'.format(rq)
            df.to_excel(file_path, sheet_name='查询', index=False, engine='xlsxwriter')
            print('输出成功......')
            print('*' * 50)
        else:
            print('****** 没有新增的改派订单！！！')
            return None
        print('*' * 50)
    def _order_spec(self, timeStart, timeEnd, n):  # 进入压单检索界面
        print('+++正在查询订单信息中')
        url = r'http://gwms-v3.giikin.cn/order/pressure/index'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': n,
                'limit': 500,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        if max_count != [] or max_count != 0:
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersdict)
            # print(data)
        else:
            data = None
            print('****** 没有信息！！！')
        return data


    # 进入已下架界面
    def order_lower(self, timeStart, timeEnd, auto_time):  # 进入已下架界面
        start: datetime = datetime.datetime.now()
        team_whid = ['龟山易速配', '速派八股仓', '天马新竹仓', '立邦香港顺丰', '香港易速配', '龟山-神龙备货', '龟山-火凤凰备货', '天马顺丰仓', '协来运', '协来运（废弃）','易速配-桃园仓']
        # team_whid = ['协来运']
        team_stock_type = [1, 2]
        # team_stock_type = [2]
        match = {1: 'SKU库存',
                 2: '组合库存',
                 3: '混合库存'}
        match2 = {'龟山易速配': 70,
                  '速派八股仓': 95,
                  '天马新竹仓': 102,
                  '立邦香港顺丰': 117,
                  '香港易速配': 134,
                  '龟山-神龙备货': 166,
                  '龟山-火凤凰备货': 198,
                  '天马顺丰仓': 204,
                  '协来运': 241,
                  '协来运（废弃）': 49,
                  '易速配-桃园仓': 253
                  }
        if auto_time == '自动':
            # sql = '''SELECT DISTINCT 统计时间 FROM 已下架表 d GROUP BY 统计时间 ORDER BY 统计时间 DESC'''
            # rq = pd.read_sql_query(sql=sql, con=self.engine1)
            # rq = pd.to_datetime(rq['统计时间'][0])
            #
            # begin = (rq + datetime.timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
            # begin = datetime.datetime.strptime(begin, '%Y-%m-%d %H:%M:%S')
            # end = (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M:%S')
            # end = datetime.datetime.strptime(end, '%Y-%m-%d %H:%M:%S')
            # print('****** 总起止时间：' + begin.strftime('%Y-%m-%d') + ' - ' + end.strftime('%Y-%m-%d') + ' ******')
            #
            # for i in range((end - begin).days + 1):  # 按天循环获取订单状态
            #     day = begin + datetime.timedelta(days=i)
            #     timeStart = (day - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            #     timeEnd = day.strftime('%Y-%m-%d')
            timeStart = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
            print('正在查询日期---起止时间：' + timeStart + ' - ' + timeEnd)
            for tem in team_whid:
                if tem in ('龟山易速配', '龟山-神龙备货', '龟山-火凤凰备货','易速配-桃园仓'):
                    for tem_type in team_stock_type:
                        print('+++正在查询仓库： ' + tem + '；库存类型:' + match[tem_type] + ' 信息')
                        self._order_lower_info(match2[tem], tem_type, timeStart, timeEnd, tem, match[tem_type])
                else:
                    print('+++正在查询仓库： ' + tem + '；库存类型:组合库存 信息')
                    self._order_lower_info(match2[tem], 2, timeStart, timeEnd, tem, '组合库存')
        else:
            print('正在查询日期---起止时间：' + timeStart + ' - ' + timeEnd)
            for tem in team_whid:
                if tem in ('龟山易速配', '龟山-神龙备货', '龟山-火凤凰备货','易速配-桃园仓'):
                    for tem_type in team_stock_type:
                        print('+++正在查询仓库： ' + tem + '；库存类型:' + match[tem_type] + ' 信息')
                        self._order_lower_info(match2[tem], tem_type, timeStart, timeEnd, tem, match[tem_type])
                else:
                    print('+++正在查询仓库： ' + tem + '；库存类型:组合库存 信息')
                    self._order_lower_info(match2[tem], 2, timeStart, timeEnd, tem, '组合库存')
        print('查询耗时：', datetime.datetime.now() - start)
    def _order_lower_info(self, tem, tem_type, timeStart, timeEnd, tem_name, type_name):  # 进入已下架界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        # print('+++正在查询信息中')
        url = r'http://gwms-v3.giikin.cn/order/order/shelves'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': 1, 'limit': 500, 'startDate': timeStart + ' 08:30:00', 'endDate':  timeEnd + ' 23:59:59', 'selectStr': '1=1 and ob.whid = ' + str(tem) + ' and ob.stock_type = ' + str(tem_type)}
        proxy = '47.75.114.218:10020'  # 使用代理服务器
        # proxies = {'http': 'socks5://' + proxy, 'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print(req.text)
        # print(json.loads(f'"{req.text}"'))
        # req = req.text.encode('utf-8').decode("unicode_escape")
        # print('+++已成功发送请求......')              # 转码使用
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        max_count = req['data']
        if max_count != []:
            ordersDict = []
            try:
                for result in req['data']:              # 添加新的字典键-值对，为下面的重新赋值
                    if result['intime'] > (result['intime']).split()[0] + ' 08:30:00':      # 判断修改统计时间
                        result['count_time'] = (datetime.datetime.strptime(result['intime'], '%Y-%m-%d %H:%M:%S') + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
                    else:
                        result['count_time'] = (result['intime']).split()[0]
                    # result['count_time'] = timeEnd
                    if type_name == 'SKU库存':
                        result['waill_name'] = '龟山备货'
                    else:
                        if '龟山易速配' in result['whid']:
                            result['waill_name'] = '龟山'
                        if '易速配-桃园仓' in result['whid']:
                            result['waill_name'] = '易速配桃园'
                        elif '速派八股仓' in result['whid']:
                            result['waill_name'] = '速派'
                        elif '天马新竹仓' in result['whid']:
                            result['waill_name'] = '天马新竹'
                        elif '立邦香港顺丰' in result['whid']:
                            result['waill_name'] = '立邦'
                        elif '香港易速配' in result['whid']:
                            result['waill_name'] = '易速配'
                        elif '神龙备货' in result['whid']:
                            result['waill_name'] = '龟山备货'
                        elif '火凤凰备货' in result['whid']:
                            result['waill_name'] = '龟山备货'
                        elif '天马顺丰仓' in result['whid']:
                            result['waill_name'] = '天马顺丰'
                        elif '协来运' in result['whid']:
                            result['waill_name'] = '协来运'
                    # print(result)
                    ordersDict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersDict)
            data = data[['order_number', 'addtime', 'billno', 'old_billno', 'goods_id', 'product_name', 'intime', 'whid', 'waill_name', 'currency_id', 'area_id', 'product_spec', 'quantity', 'ship_name', 'ship_address', 'ship_phone', 'amount', 'count_time']]
            data.columns = ['订单编号', '下单时间', '新运单号', '原运单号', '产品id', '商品名称', '下架时间', '仓库', '物流渠道', '币种', '团队', '商品规格', '购买数量', '收货人', '收货地址', '联系电话', '订单金额', '统计时间']
            print(data)
            print('>>>' + tem_name + '-' + type_name + ' <<< 查询完结！！！')
            data.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 已下架表(订单编号,下单时间,新运单号,原运单号, 产品id, 商品名称, 下架时间, 仓库, 物流渠道,币种, 团队,商品规格, 购买数量, 收货人, 收货地址, 联系电话,订单金额,统计时间,记录时间)
                    SELECT 订单编号,下单时间,新运单号,原运单号, 产品id, 商品名称, 下架时间, 仓库, 物流渠道,币种, 团队, 商品规格, 购买数量, 收货人, 收货地址, 联系电话, 订单金额,统计时间,NOW() 记录时间
                    FROM customer'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
            data.to_excel('G:\\输出文件\\已下架 {0} {1}-{2}.xlsx'.format(tem_name, type_name, rq), sheet_name='查询', index=False, engine='xlsxwriter')

            print('获取每日新增龟山备货表......')
            rq = datetime.datetime.now().strftime('%m.%d')
            sql = '''SELECT CURDATE() '序號(無用途)',NULL '訂單號長度限制: 20碼請勿使用中文）', 收货人 AS '收件人姓名(必填)長度限制: 20碼', 收货地址 AS '收件人地址(必填)中文限制: 50字', 
                            联系电话 AS '收件人電話長度限制: 15碼',商品名称 AS '託運備註中文限制: 50字', NULL '(商品別編號)勿填', 购买数量 AS '商品數量(必填)(限數字)', NULL '才積重量限數字', 
                            订单金额 AS '代收貨款限數字',NULL '指定配送日期YYYYMMDD範例: 20140220    ->2月20號', NULL '指定配送時間範例:   1   (上午 -> 09~13) 2   (下午 -> 13~17)3   (晚上 -> 17~20)',
			                订单编号 , 商品规格, 产品id AS '产品ID', NULL '原运单号', 团队,下架时间,统计时间
                    FROM 已下架表 yx
                    WHERE yx.记录时间 >= TIMESTAMP(CURDATE()) AND yx.物流渠道 = '龟山备货' AND yx.`新运单号` IS NULL;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            if df is not None and len(df) > 0:
                df.to_excel('G:\\输出文件\\{} 龟山备货.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
                print('获取成功......')
            else:
                print('****** 今日无新增龟山备货数据！！！')
        else:
            print('****** 没有新增的改派订单！！！')
            return None
        print('*' * 50)

    # 改派-查询未发货的订单
    def gp_order(self):
        print('正在查询改派未发货订单…………')
        today = datetime.date.today().strftime('%m.%d')
        listT = []  # 查询sql的结果 存放池
        sql = '''SELECT *
                                FROM ( SELECT xj.订单编号, xj.下单时间, gs.运单编号, xj.产品id, xj.商品名称, xj.下架时间, xj.仓库, xj.物流渠道, xj.币种, xj.统计时间, xj.记录时间, gz.最终状态 ,gs.系统订单状态 , gs.是否改派
                                        FROM (SELECT *
                			                FROM 已下架表  x
                			                WHERE x.下单时间 >= TIMESTAMP(DATE_ADD(curdate()-day(curdate())+1,interval -2 month)) AND x.币种 = '台币'
                                        )  xj
                                        LEFT JOIN gat_zqsb gz ON xj.订单编号= gz.订单编号
                                        LEFT JOIN gat_order_list gs ON xj.订单编号= gs.订单编号
                                        WHERE 最终状态 = '未发货' or 最终状态 IS NULL
                                ) ss
                                WHERE 是否改派 = '改派' AND (系统订单状态 NOT IN ('已删除', '问题订单', '待发货', '截单')) OR 是否改派 IS NULL
                                ORDER BY FIELD(物流渠道,'龟山','龟山备货','天马顺丰','天马新竹','速派','立邦');'''
        sql = '''SELECT xj.订单编号, xj.下单时间, gs.运单编号, xj.产品id, xj.商品名称, xj.下架时间, xj.仓库, xj.物流渠道, xj.币种, xj.统计时间, xj.记录时间, gz.最终状态 ,gs.系统订单状态 , gs.是否改派
        				FROM (SELECT * FROM 已下架表  x WHERE x.记录时间 >=  TIMESTAMP(CURDATE()) AND x.币种 = '台币')  xj
                        LEFT JOIN gat_zqsb gz ON xj.订单编号= gz.订单编号
                        LEFT JOIN gat_order_list gs ON xj.订单编号= gs.订单编号
                        WHERE 最终状态 NOT IN ("已签收","拒收","已退货","理赔","自发头程丢件","在途") or 最终状态 IS NULL;'''
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df = df.loc[df["币种"] == "台币"]
        listT.append(df)

        print('正在写入excel…………')
        file_path = 'F:\\神龙签收率\\(未发货) 改派-物流\\{} 改派未发货.xlsx'.format(today)
        if os.path.exists(file_path):  # 判断是否有需要的表格
            print("正在清除重复文件......")
            os.remove(file_path)
        sheet_name = ['台湾']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('----已写入excel ')

    # 查询改派无运单好（仓储的获取）
    def get_billno_res(self):  # 进入仓储界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        timeStart = ((datetime.datetime.now() + datetime.timedelta(days=1)) - relativedelta(months=2)).strftime('%Y-%m-%d')
        timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
        print('正在查询 港台 改派无运单号......')
        url = r'http://gwms-v3.giikin.cn/order/order/secondsendbillnonone'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': '1',
                'limit': 500,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        print('++++++本次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        if max_count != [] or max_count != 0:
            # 首次查询
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            df = pd.json_normalize(ordersdict)
            # 剩余查询
            if max_count > 500:
                in_count = math.ceil(max_count/500)
                dlist = []
                n = 1
                while n < in_count:  # 这里用到了一个while循环，穿越过来的
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                    data = self._order_spec(timeStart, timeEnd, n)                     # 分页获取详情
                    dlist.append(data)
                dp = df.append(dlist, ignore_index=True)
            else:
                dp = df
            print('正在写入......')
            dp.to_excel('G:\\输出文件\\改派无运单号 {}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
            # dp = dp[['order_number', 'goods_id', 'goods_name', 'currency_id', 'area_id', 'ydtime', 'purid', 'other_reason', 'buyer', 'intime', 'addtime', 'is_lower', 'below_time', 'cate']]
            # dp.columns = ['订单编号', '产品ID', '产品名称', '币种', '团队', '反馈时间', '压单原因', '其他原因', '采购员', '入库时间', '下单时间', '是否下架', '下架时间', '品类']
            # dp = dp[(dp['币种'].str.contains('港币|台币', na=False))]
            # print(dp)
            # dp.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 压单表(订单编号,产品ID,产品名称,币种,团队, 反馈时间, 压单原因, 其他原因, 采购员, 入库时间, 下单时间, 是否下架, 下架时间,记录时间) 
                    SELECT 订单编号,产品ID,产品名称,币种,团队, 反馈时间, 压单原因, 其他原因, 采购员, 入库时间, 下单时间, 是否下架, IF(下架时间 = '',NULL,下架时间) 下架时间, NOW() 记录时间
                    FROM customer'''
            # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('共有 ' + str(len(dp)) + '条 成功写入数据库+++++++')

            # print('正在获取 压单反馈 信息中......')
            # time_path: datetime = datetime.datetime.now()
            # mkpath = r"F:\神龙签收率\(未发货) 直发-仓库-压单\\" + time_path.strftime('%m.%d')
            # sql = '''SELECT s.*,s1.处理结果,s1.处理时间
            #         FROM ( SELECT * FROM 压单表 g WHERE g.`记录时间` >= CURDATE() and g.是否下架 <> '已下架'
            #         ) s
            #         LEFT JOIN 压单表_已核实 s1 ON s.订单编号 = s1.订单编号;'''
            # df = pd.read_sql_query(sql=sql, con=self.engine1)
            # isExists = os.path.exists(mkpath)
            # if not isExists:
            #     os.makedirs(mkpath)
            # else:
            #     print(mkpath + ' 目录已存在')
            # file_path = mkpath + '\\压单反馈 {0}.xlsx'.format(rq)
            # df.to_excel(file_path, sheet_name='查询', index=False, engine='xlsxwriter')
            # print('输出成功......')
            # print('*' * 50)
        else:
            print('****** 没有新增的改派订单！！！')
            return None
        print('*' * 50)
    def _get_billno_res(self, timeStart, timeEnd, n):  # 进入压单检索界面
        print('+++正在查询订单信息中')
        url = r'http://gwms-v3.giikin.cn/order/order/secondsendbillnonone'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': '1',
                'limit': 500,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        if max_count != [] or max_count != 0:
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersdict)
            # print(data)
        else:
            data = None
            print('****** 没有信息！！！')
        return data

    def bulid_file(self):
        print('正在生成每日新文件夹......')
        time_path: datetime = datetime.datetime.now()
        mkpath = "F:\\神龙签收率\\" + time_path.strftime('%m.%d')
        isExists = os.path.exists(mkpath)
        if not isExists:
            os.makedirs(mkpath)
            os.makedirs(mkpath + "\\产品签收率")
            os.makedirs(mkpath + "\\产品签收率\\直发&改派")
            os.makedirs(mkpath + "\\导运单号&提货时间")
            os.makedirs(mkpath + "\\导状态")
            os.makedirs(mkpath + "\\签收率")
            os.makedirs(mkpath + "\\物流表")
            print('创建成功')
            file_path = mkpath + '\\导运单号&提货时间\\{} 龟山 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path2 = mkpath + '\\导运单号&提货时间\\{} 立邦 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path3 = mkpath + '\\导运单号&提货时间\\{} 天马 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path4 = mkpath + '\\导运单号&提货时间\\{} 速派 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path5 = mkpath + '\\导运单号&提货时间\\{} 协来运普货 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path50 = mkpath + '\\导运单号&提货时间\\{} 协来运特货 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            df = pd.DataFrame([['', '']], columns=['订单编号', '物流单号'])
            df.to_excel(file_path, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path2, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path3, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path4, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path5, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path50, sheet_name='查询', index=False, engine='xlsxwriter')

            file_path31 = mkpath + '\\导运单号&提货时间\\{} 天马 换新运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path32 = mkpath + '\\导运单号&提货时间\\{} 协来运 换新运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path33 = mkpath + '\\导运单号&提货时间\\{} 立邦 换新运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path34 = mkpath + '\\导运单号&提货时间\\{} 速派 换新运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path35 = mkpath + '\\导运单号&提货时间\\{} 龟山 换新运单号.xlsx'.format(time_path.strftime('%m.%d'))
            df2 = pd.DataFrame([['', '', '']], columns=['订单编号', '旧运单号', '新运单号'])
            df2.to_excel(file_path31, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path32, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path33, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path34, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path35, sheet_name='查询', index=False, engine='xlsxwriter')

            file_path91 = mkpath + '\\导运单号&提货时间\\{} 导入提货时间 龟山.xlsx'.format(time_path.strftime('%m.%d'))
            file_path92 = mkpath + '\\导运单号&提货时间\\{} 导入提货时间 天马.xlsx'.format(time_path.strftime('%m.%d'))
            file_path93 = mkpath + '\\导运单号&提货时间\\{} 导入提货时间 速派.xlsx'.format(time_path.strftime('%m.%d'))
            file_path94 = mkpath + '\\导运单号&提货时间\\{} 导入提货时间 协来运.xlsx'.format(time_path.strftime('%m.%d'))
            file_path95 = mkpath + '\\导运单号&提货时间\\{} 导入提货时间 立邦.xlsx'.format(time_path.strftime('%m.%d'))
            df2 = pd.DataFrame([['', '', '']], columns=['订单号', '物流单号', '提货时间'])
            df2.to_excel(file_path91, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path92, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path93, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path94, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path95, sheet_name='查询', index=False, engine='xlsxwriter')
            print('创建文件')
        else:
            print(mkpath + ' 目录已存在')
        print('*' * 50)

    def get_take_delivery_no(self):  # 进入 头程物流跟踪 界面
        print('+++正在查询头程物流信息中')
        timeStart = (datetime.datetime.now() - datetime.timedelta(days=10)).strftime('%Y-%m-%d')
        timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
        url = r'http://gwms-v3.giikin.cn/order/delivery/firstLegTrace'
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
            'origin': 'http://gwms-v3.giikin.cn',
            'Referer': 'http://gwms-v3.giikin.cn/order/delivery/takeDeliveryRegister?id=8755'}
        data = {'page': '1',
                'limit': 100,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1 and a.country= "TW"'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print(req)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        max_count = req['count']
        if max_count != [] or max_count != 0:
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersdict)
            # print(data)
            df = data[['id', 'take_delivery_no', 'take_delivery_date', 'take_delivery_company', 'take_delivery_company_id', 'transport_mode', 'product_type', 'transport_type',
                       'batch', 'barcode', 'country', 'boxCount', 'analy', 'deliverytime', 'send_first_logistics_comment', 'uptime']]
            df.columns = ['id', '提货单号', '提货时间', '提货物流', '提货物流id', '运输方式', '货物类型', '运输公司',
                          '运输班次', '箱号', '线路', '箱数', '统计', '交货时间', '报关资料发送结果', '更新时间']

            print('共有 ' + str(len(df)) + '条 正在写入......')
            df.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            # df.to_excel('G:\\输出文件\\{0}-查询{1}.xlsx'.format(match[team], rq), sheet_name='查询', index=False,engine='xlsxwriter')
            sql = '''REPLACE INTO gat_take_delivery(id,提货单号,提货时间,提货日期,提货物流,提货物流id,运输方式,货物类型,运输公司,运输班次,箱号,线路,箱数,统计,交货时间,报关资料发送结果,更新时间, 主號,航班號,记录时间)
                     SELECT id,提货单号,提货时间,DATE_FORMAT(提货时间,'%Y-%m-%d') 提货日期,提货物流,提货物流id,运输方式,货物类型,IF(运输公司 = '',NULL,运输公司) 运输公司,IF(运输班次 = '',NULL,运输班次) 运输班次,箱号,线路,箱数,统计,
                            IF(交货时间 = '',NULL,交货时间) 交货时间,报关资料发送结果,更新时间,NULL 主號,NULL 航班號,NOW() 记录时间
                    FROM customer;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
        else:
            df = None
            print('****** 没有信息！！！')
        return df


    def _get_take_delivery_no(self):  # 进入头程检索界面
        timeStart = (datetime.datetime.now() - datetime.timedelta(days=30)).strftime('%Y-%m-%d')
        start = datetime.datetime.now()
        print('正在更新 头程提货单号 信息…………')
        sql = '''SELECT id, 提货单号,主號, 航班號, 提货日期  
                FROM {0} g 
                WHERE g.运输公司 IS NULL AND g.`航班號` IS NOT NULL AND g.`提货日期` >= '{1}';'''.format('gat_take_delivery', timeStart)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        if df.empty:
            print('无需要更新订单信息！！！')
            return
        for row in df.itertuples():
            ord_id = getattr(row, '提货单号')
            id = getattr(row, 'id')
            take_delivery_no = getattr(row, '主號')
            batch = getattr(row, '航班號')
            self._upload_take_delivery_no(ord_id, id, take_delivery_no, batch)
        print('单次更新耗时：', datetime.datetime.now() - start)

    def _upload_take_delivery_no(self, ord_id, id, take_delivery_no, batch):  # 进入头程检索界面
        print('+++正在更新中')
        url = r'http://gwms-v3.giikin.cn/order/delivery/takedeliveryregister'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/delivery/takeDeliveryRegister?id=8755'}
        transport_type = batch[:2]
        print('提货单号：' + ord_id, 'id：' + str(id), ';主號：' + take_delivery_no, '；航班號：'+transport_type, '；航班信息：'+batch)
        data = {'id': id,
                'take_delivery_no': take_delivery_no,
                'transport_type': transport_type,
                'batch': batch,
                'departed_time': '',
                'departed_place': '',
                'arrived_time': '',
                'arrived_place': '',
                'product_type': ''
                }
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print(req)
        # print(req.text)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        # print(req['code'])
        # print(req['comment'])
        val = req['comment']
        if val == 'success':
            print('头程物流 更新成功！！！')
        else:
            print('头程物流 更新失败！！！')


if __name__ == '__main__':
    m = QueryTwoLower('+86-18538110674', 'qyz35100416','84c3a0212a7b3de386b2a20d4a46b0ea','手0动')
    start: datetime = datetime.datetime.now()
    match1 = {'gat': '港台', 'gat_order_list': '港台', 'slsc': '品牌'}

    # -----------------------------------------------手动设置时间；若无法查询，切换代理和直连的网络-----------------------------------------

    # m.order_lower('2022-02-17', '2022-02-18', '自动')   # 已下架
    select = 1
    if select == 1:
        m.readFile(select)            # 上传每日压单核实结果
        m.order_spec()       # 压单反馈  （备注（压单核实是否需要））

    elif select == 2:
        m.readFile(select)
        m._get_take_delivery_no()


    elif select == 3:
        m.get_take_delivery_no()
        m.readFile(select)
        m._get_take_delivery_no()

        # m. _upload_take_delivery_no(8637, '297-82680091', 'CI', 'CI6844')


    # m.get_billno_res()      # 改派无运单号

    #h  ttp://gwms-v3.giikin.cn/order/delivery/batchImportLegDeliveryBoxData

    print('查询耗时：', datetime.datetime.now() - start)