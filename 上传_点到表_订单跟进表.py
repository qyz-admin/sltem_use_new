import pandas as pd
import os
import datetime
import time
import xlwings
import xlsxwriter
import math
import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel
import win32api,win32con
import math
from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Color,Alignment ,PatternFill # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色

# -*- coding:utf-8 -*-
class QueryTwo(Settings, Settings_sso):
    def __init__(self, userMobile, password):
        Settings.__init__(self)
        Settings_sso.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue(maxsize=10)  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        # self.sso_online_Two()
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
    def readFormHost(self, team):
        start = datetime.datetime.now()
        path = r'D:\Users\Administrator\Desktop\需要用到的文件\数据库'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                if '时长区间订单明细' in dir:
                    team = 'gat_waybill_list'
                else:
                    team = 'gat_logisitis_googs'
                self.wbsheetHost(filePath, team)
                os.remove(filePath)
                print('已清除上传文件…………')
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, team):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    if team == 'gat_logisitis_googs':
                        db.rename(columns={'订单号': '订单编号'}, inplace=True)
                        db.rename(columns={'承运单号': '运单编号'}, inplace=True)
                        # print(db.columns)
                        if '物流状态' not in db.columns and '末条时间' not in db.columns and '末条信息' not in db.columns:
                            db.insert(0, '物流状态', '')
                            db.insert(0, '末条时间', '')
                            db.insert(0, '末条信息', '')
                        if '下单时间' not in db.columns:
                            db.insert(0, '下单时间', '')
                            db['下单时间'] = db['核重时间'].copy()
                        if '核重时间' not in db.columns:
                            db.insert(0, '核重时间', '')
                            db['核重时间'] = db['下单时间'].copy()
                        db = db[['下单时间', '订单编号', '运单编号', '核重时间', '物流状态', '末条时间', '末条信息']]
                        db.dropna(axis=0, how='any', inplace=True)  # 空值（缺失值），将空值所在的行/列删除后
                    elif team == 'gat_waybill_list':
                        db.insert(0, '运单编号', '')
                        db = db[['订单编号', '运单编号', '物流', '物流状态', '订单状态', '下单时间', '出库时间', '提货时间','上线时间','完成时间']]
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在导入：' + sht.name + ' 表；共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    db.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
                    print('++++成功导入缓存表')
                    columns = list(db.columns)
                    columns = ','.join(columns)
                    if team == 'gat_logisitis_googs':
                        sql = '''REPLACE INTO {0}({1}, 记录时间) SELECT *, NOW() 记录时间 FROM customer;'''.format(team, columns)
                        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                        print('++++：' + sht.name + '表--->>>更新成功')
                    elif team == 'gat_waybill_list':
                        sql = '''REPLACE INTO {0}({1},添加时间,记录时间) SELECT *, CURDATE() 添加时间,NOW() 记录时间 FROM customer;'''.format(team,columns)
                        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                        # self.waybill_info()
                        # self.waybill_updata()
                        print('++++成功导出订单跟进明细')
                else:
                    print('----------数据为空,不需导入：' + sht.name)
            wb.close()
        app.quit()

    # 更新订单状态
    def waybill_info(self, login_TmpCode, handle):
        # lw = Settings_sso()
        # lw.sso_online_Two()
        # self.sso_online_Two()
        if handle == '手动':
            self.sso__online_handle(login_TmpCode)
        else:
            self.sso__online_auto()
        print('正在更新 订单跟进 信息……………………………………………………………………………………………………………………………………………………………………………………')
        start = datetime.datetime.now()
        sql = '''SELECT 订单编号 FROM {0} s WHERE s.`添加时间` = CURDATE();'''.format('gat_waybill_list')
        db = pd.read_sql_query(sql=sql, con=self.engine1)
        if db.empty:
            print('无需要更新订单信息！！！')
            return
        print(db['订单编号'][0])
        orderId = list(db['订单编号'])
        max_count = len(orderId)  # 使用len()获取列表的长度，上节学的
        if max_count > 500:
            ord = ', '.join(orderId[0:500])
            df = self.order_Info_Query(ord)
            dlist = []
            n = 0
            while n < max_count - 500:  # 这里用到了一个while循环，穿越过来的
                n = n + 500
                ord = ','.join(orderId[n:n + 500])
                data = self.order_Info_Query(ord)
                dlist.append(data)
            print('正在写入......')
            dp = df.append(dlist, ignore_index=True)
        else:
            ord = ','.join(orderId[0:max_count])
            dp = self.order_Info_Query(ord)
        dp.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
        print('正在更新订单跟进表中......')
        sql = '''update {0} a, customer b
                        set a.`运单编号`= IF(b.`运单号` = '', NULL, b.`运单号`),
                            a.`订单状态`= IF(b.`订单状态` = '', NULL, b.`订单状态`),
                            a.`物流状态`= IF(b.`物流状态` = '', NULL, b.`物流状态`)
                where a.`订单编号`=b.`订单编号`;'''.format('gat_waybill_list')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('查询耗时：', datetime.datetime.now() - start)
    def order_Info_Query(self, ord):  # 更新订单跟进 的状态信息
        print('+++正在查询订单信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500,
                'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None,
                'type': None, 'collId': None, 'isClone': None,
                'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '',
                'warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None,
                'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None,
                'service_id': None, 'autoVerifyStatus': None, 'shipZip': None, 'remark': None, 'shipState': None, 'weightStart': None,
                'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None}
        data.update({'orderPrefix': ord,
                    'shippingNumber': None})
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersdict = []
        # print('正在处理json数据转化为dataframe…………')
        try:
            for result in req['data']['list']:
                result['saleId'] = 0        # 添加新的字典键-值对，为下面的重新赋值用
                result['saleName'] = 0
                result['productId'] = 0
                result['saleProduct'] = 0
                result['spec'] = 0
                result['chooser'] = 0
                result['saleId'] = result['specs'][0]['saleId']
                result['saleName'] = result['specs'][0]['saleName']
                result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
                result['saleProduct'] = (result['specs'][0]['saleProduct']).split('#')[2]
                result['spec'] = result['specs'][0]['spec']
                result['chooser'] = result['specs'][0]['chooser']
                quest = ''
                for re in result['questionReason']:
                    quest = quest + ';' + re
                result['questionReason'] = quest
                delr = ''
                for re in result['delReason']:
                    delr = delr + ';' + re
                result['delReason'] = delr
                auto = ''
                for re in result['autoVerify']:
                    auto = auto + ';' + re
                result['autoVerify'] = auto
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            # time.sleep(10)
            # print(team)
            # print(searchType)
            # self.readFormHost(team, searchType)
            # self.orderInfoQuery(ord, searchType)
        #     self.q.put(result)
        # for i in range(len(req['data']['list'])):
        #     ordersdict.append(self.q.get())
        data = pd.json_normalize(ordersdict)
        df = None
        try:
            df = data[['orderNumber', 'currency', 'area', 'productId', 'saleName', 'percent',
                       'amount', 'quantity', 'orderStatus', 'wayBillNumber', 'payType', 'addTime', 'username', 'verifyTime',
                       'logisticsName', 'dpeStyle', 'reassignmentTypeName', 'logisticsStatus', 'deliveryTime', 'onlineTime', 'finishTime',
                       'logisticsUpdateTime', 'stateTime', 'update_time']]
            df.columns = ['订单编号', '币种', '运营团队', '产品id', '产品名称', '拉黑率',
                          '应付金额', '数量', '订单状态', '运单号', '支付方式', '下单时间', '审核人', '审核时间',
                          '物流渠道', '货物类型', '订单类型', '物流状态', '发货时间', '上线时间', '完成时间',
                          '物流更新时间', '状态时间', '更新时间']
        except Exception as e:
            print('------查询为空')
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return df

    # 更新压单状态
    def chuhuo_info(self,login_TmpCode,login_TmpCode2, handle):
        print('正在更新 压单订单 信息…………………………………………………………………………………………………………………………………………………………………………………………………………')
        start = datetime.datetime.now()
        # 获取更新订单的语句
        sql = '''SELECT 订单编号 FROM {0} s WHERE s.`添加时间` = CURDATE() and s.`出库时间` IS NULL;'''.format('gat_waybill_list')
        data_df = ['order_number', 'goods_id', 'goods_name', 'currency_id', 'area_id', 'ydtime', 'purid', 'other_reason',
                 'buyer', 'intime', 'addtime', 'is_lower', 'below_time', 'cate']
        data_df2 = ['订单编号', '产品ID', '产品名称', '币种', '团队', '反馈时间', '压单原因', '其他原因',
                    '采购员', '入库时间', '下单时间', '是否下架', '下架时间', '品类']
        # 获取更新表的语句
        sql2 = '''update {0} a, cache b
                set a.`订单状态`= '压单'
                where a.`订单编号`=b.`订单编号`;'''.format('gat_waybill_list')
        # 调用更新库 函数
        up = Settings_sso()
        up.updata_yadan(sql, sql2, 'gat', data_df, data_df2,login_TmpCode, handle)
        print('更新完成…………')

        print('正在更新 出库订单 信息…………………………………………………………………………………………………………………………………………………………………………………………………………')
        # 获取更新订单的语句
        sql = '''SELECT 订单编号 FROM {0} s WHERE s.`添加时间` = CURDATE() and s.`出库时间` IS NULL and s.`订单状态`<> '压单';'''.format('gat_waybill_list')
        data_df = ['order_number', 'addtime', 'billno', 'status_desc']
        data_df2 = ['订单编号', '运单扫描时间', '运单编号', '扫描状态']
        # 获取更新表的语句
        sql2 = '''update {0} a, cache b
                set a.`订单状态`= '今日出库'
                where a.`订单编号`=b.`订单编号`;'''.format('gat_waybill_list')
        # 调用更新库 函数
        up = Settings_sso()
        up.updata_chuku(sql, sql2, 'gat', data_df, data_df2,login_TmpCode2, handle)
        print('更新完成…………')

        # print('正在更新 提货订单 信息…………………………………………………………………………………………………………………………………………………………………………………………………………')
        # # 获取更新订单的语句
        # sql = '''SELECT 订单编号 FROM {0} s WHERE s.`添加时间` = CURDATE() and s.`提货时间` IS NULL;'''.format('gat_waybill_list')
        # data_df = ['order_number', 'billno', 'country_code', 'intime', 'logistics_id', 'is_exception', 'is_deal']
        # data_df2 = ['订单编号', '运单编号', '币种', '提货时间', '物流渠道', '是否异常', '是否处理']
        # # 获取更新表的语句
        # sql2 = '''update {0} a, cache b
        #         set a.`订单状态`= '今日提货'
        #         where a.`订单编号`=b.`订单编号`;'''.format('gat_waybill_list')
        # # 调用更新库 函数
        # up = Settings_sso()
        # up.updata_tihuo(sql, sql2, 'gat', data_df, data_df2)
        # print('更新完成…………')

        print('查询耗时：', datetime.datetime.now() - start)

    # 订单跟进明细
    def waybill_updata(self):
        today = datetime.date.today().strftime('%Y.%m.%d')
        listT = []  # 查询sql的结果 存放池
        print('正在获取 订单跟进汇总……………………………………………………')
        sql = '''SELECT IFNULL(物流未完成, '总计') 物流未完成,出库,提货,上线,完成,合计
                FROM( SELECT IFNULL(物流未完成, '总计') 物流未完成,
                            sum(IF(节点类型 = '出库',1,0)) AS 出库,
                            sum(IF(节点类型 = '提货',1,0)) AS 提货,
                            sum(IF(节点类型 = '上线',1,0)) AS 上线,
                            sum(IF(节点类型 = '完成',1,0)) AS 完成,
                            COUNT(订单编号) AS 合计
                    FROM( SELECT *,IF(出库时间 IS NULL,'出库',IF(提货时间 IS NULL,'提货',
                                    IF(上线时间 IS NULL,'上线',IF(完成时间 IS NULL,'完成',完成时间)))) AS 节点类型,
									IF(物流 LIKE '%速派%','台湾-速派-新竹&711超商',
									IF(物流 LIKE '%天马%','台湾-天马-新竹&711',
									IF(物流 LIKE '%优美宇通%' or 物流 LIKE '%铱熙无敌%','台湾-铱熙无敌-新竹普货&特货',物流))) AS 物流未完成
                        FROM gat_waybill_list s
                        WHERE s.`添加时间` = CURDATE()
                    ) ss
                    GROUP BY 物流未完成
                    WITH ROLLUP
                ) sss
                GROUP BY 物流未完成
                ORDER BY FIELD(物流未完成,'台湾-立邦普货头程-易速配尾程','台湾-速派-新竹&711超商', '台湾-天马-新竹&711','台湾-铱熙无敌-新竹普货&特货','总计');'''.format()
        df0 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df0)

        print('正在获取 订单跟进明细……………………………………………………')
        sql = '''SELECT *,null 原因汇总
                FROM( SELECT s1.*,单量,
							IF(压单量 = 0,NULL,压单量) AS 是否压单,
							IF(今日出库量 = 0,IF(今日提货量 = 0,NULL,今日提货量),今日出库量) AS '今日出库/提货',
							IF(今日提货量 = 0,NULL,今日提货量) AS 今日提货,
							IF(取消量 = 0,NULL,取消量) AS 取消量,
							IF(物流已上线 = 0,NULL,物流已上线) AS 物流已上线,
							IF(物流已提货待出货 = 0,NULL,物流已提货待出货) AS 物流已提货待出货,
							IF(物流已出货待上线 = 0,NULL,物流已出货待上线) AS 物流已出货待上线
                    FROM gat_waybill s1
                    LEFT JOIN ( SELECT 物流未完成,节点类型, COUNT(订单编号) AS 单量,SUM(IF(订单状态 = '压单',1,0)) AS 压单量,SUM(IF(订单状态 = '今日出库',1,0)) AS 今日出库量,
                                        SUM(IF(节点类型 = '提货' AND 出库时间 >= TIMESTAMP(DATE_SUB(CURDATE(), INTERVAL 1 DAY)),1,0)) AS 今日提货量,
										SUM(IF(订单状态 = '已删除',1,0)) AS 取消量,
										SUM(IF(节点类型 = '上线' AND 物流出货时间 IS NULL,1,0)) AS 物流已提货待出货,
										SUM(IF(节点类型 = '上线' AND 物流出货时间 IS NOT NULL AND 最终状态 <> '在途',1,0)) AS 物流已出货待上线,
										SUM(IF(节点类型 = '上线' AND 最终状态 = '在途',1,0)) AS 物流已上线
                                FROM( SELECT s.*,z.最终状态,z.出货时间 AS 物流出货时间,
                                            IF(s.出库时间 IS NULL,'出库',IF(s.提货时间 IS NULL,'提货',
                                            IF(s.上线时间 IS NULL,'上线',IF(s.完成时间 IS NULL,'完成',s.完成时间)))) AS 节点类型,
											IF(物流 LIKE '%速派%','台湾-速派-新竹&711超商',
											IF(物流 LIKE '%天马%','台湾-天马-新竹&711',
											IF(物流 LIKE '%优美宇通%' or 物流 LIKE '%铱熙无敌%','台湾-铱熙无敌-新竹普货&特货',物流))) AS 物流未完成
                                    FROM gat_waybill_list s
									LEFT JOIN gat_zqsb z ON s.订单编号 = z.订单编号
                                    WHERE s.`添加时间` = CURDATE()
                                ) ss
                                GROUP BY 物流未完成,节点类型
                    ) s2 ON s1.物流=s2.物流未完成 AND s1.节点类型=s2.节点类型
                ) g
                ORDER BY FIELD(物流,'台湾-立邦普货头程-易速配尾程','台湾-速派-新竹&711超商', '台湾-天马-新竹&711','台湾-铱熙无敌-新竹普货&特货','合计'),
                        FIELD(节点类型,'出库','提货','上线','完成','合计');'''.format()
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df1)

        print('正在获取 订单跟进明细表…………………………………………')
        sql = '''SELECT ss.*,
				        g.`下单时间` AS 物流下单时间, g.`核重时间` AS 物流核重时间, g.`物流状态` AS 物流核重状态, g.`末条时间` AS 物流末条时间, g.`末条信息` AS 物流末条信息,
						z.出货时间 as 物流出货时间, z.上线时间 AS 物流上线时间, z.签收表物流状态, z.最终状态
           FROM ( SELECT *,
						IF(出库时间 IS NULL,'出库',IF(提货时间 IS NULL,'提货',IF(上线时间 IS NULL,'上线',IF(完成时间 IS NULL,'完成',完成时间)))) AS 节点类型,
						IF(物流 LIKE '%速派%','台湾-速派-新竹&711超商',IF(物流 LIKE '%天马%','台湾-天马-新竹&711',IF(物流 LIKE '%优美宇通%' or 物流 LIKE '%铱熙无敌%','台湾-铱熙无敌-新竹普货&特货',物流))) AS 物流未完成
	            FROM gat_waybill_list s
	            WHERE s.`添加时间` = CURDATE()
           ) ss
           LEFT JOIN gat_logisitis_googs g ON ss.订单编号 = g.订单编号
		   LEFT JOIN gat_zqsb z ON ss.订单编号 = z.订单编号;'''.format()
        df2 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df2)

        print('正在写入excel…………')
        today = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        file_path = 'G:\\输出文件\\{} 订单跟进.xlsx'.format(today)

        writer2 = pd.ExcelWriter(file_path, engine='openpyxl')
        df0.to_excel(writer2, sheet_name='汇总',index=False, startrow=1)
        df1.to_excel(writer2, sheet_name='汇总',index=False, startrow=1, startcol=7)
        df2.to_excel(writer2, sheet_name='明细表',index=False)
        writer2.save()
        writer2.close()

        # 初始化赋值   https://openpyxl.readthedocs.io/en/stable/index.html
        # from openpyxl.utils import get_column_letter, column_index_from_string
        # # 根据列的数字返回字母
        # print(get_column_letter(2))  # B
        # # 根据字母返回列的数字
        # print(column_index_from_string('D'))  # 4

        month_yesterday = (datetime.datetime.now() - datetime.timedelta(days=5)).strftime('%m.%d')
        month_begin = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%m') + '.01-'
        # 指定了等线24号，加粗斜体，字体颜色红色。直接使用cell的font属性
        bold_itatic_24_font = Font(name='宋体', size=14, italic=False, color=colors.COLOR_INDEX[2], bold=True)
        # cell的属性aligment，这里指定垂直居中和水平居中。除了center，还可以使用right、left等
        cell_alignment = Alignment(horizontal='center', vertical='center')
        # 颜色背景
        # Fillsytle = PatternFill('solid', fgColor='#AABBCC', bgColor='#DDEEFF')
        # Fillsytle = PatternFill('solid', fgColor='#93b6f7', bgColor='#93b6f7')

        border = Border(left=Side(border_style=None,color='FF000000'),
                        right=Side(border_style=None,color='FF000000'),
                        top=Side(border_style=None,color='FF000000'),
                        bottom=Side(border_style=None,color='FF000000'),
                        diagonal=Side(border_style=None,color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style=None,color='FF000000'),
                        vertical=Side(border_style=None,color='FF000000'),
                        horizontal=Side(border_style=None, color='FF000000'),
                        diagonalDown=False,
                        start=None,
                        end=None)

        wb = load_workbook(file_path)
        # sheet = wb.get_sheet_by_name("汇总")
        sheet = wb["汇总"]
        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 50
        sheet.row_dimensions[2].alignment = cell_alignment
        # sheet.row_dimensions[2].fill = Fillsytle  # 设定行的颜色

        sheet["A1"] = month_begin + month_yesterday +'台湾直发订单跟进 汇总'
        sheet['A1'].font = bold_itatic_24_font
        sheet['A1'].alignment = cell_alignment
        # sheet['A2:K30'].alignment = cell_alignment
        sheet.merge_cells('A1:F1')
        sheet.column_dimensions['A'].width = 28
        for cl in ['B','C','D','E','F','C','C']:
            sheet.column_dimensions[cl].width = 11.13
        for row in sheet.iter_rows(min_row=3, max_row=10, min_col=2, max_col=6):
            for cell in row:
                cell.alignment = cell_alignment

        sheet["H1"] = month_begin + month_yesterday +'台湾直发订单跟进 明细'
        sheet['H1'].font = bold_itatic_24_font
        sheet['H1'].alignment = cell_alignment
        sheet.merge_cells('H1:K1')
        sheet.column_dimensions['H'].width = 28
        for cl in ['I','J','K']:
            sheet.column_dimensions[cl].width = 11.13
        for row in sheet.iter_rows(min_row=3, max_row=30, min_col=9, max_col=11):
            for cell in row:
                cell.alignment = cell_alignment
        # sheet.column_dimensions['A:f'].width = 25
        # sheet.columns[2].width = 11.13
        # sheet.rows(2,6).width = 11.13
        # sheet['2:6'].width = 11.13

        wb.save(file_path)
        print('----已写入excel ')





if __name__ == '__main__':
    m = QueryTwo('+86-18538110674', 'qyz35100416')
    start: datetime = datetime.datetime.now()
    '''
    # -----------------------------------------------查询状态运行（一）-----------------------------------------
    # 1、 点到表上传 team = 'gat_logisitis_googs'；2、上架表上传；；3、订单跟进上传 team = 'gat_waybill_list'--->>数据更新切换
    '''

    select = 5
    if int(select) == 1:
        team = 'gat_logisitis_googs'
        m.readFormHost(team)

    elif int(select) == 2:
        print("2-->>> 正在按时间查询+++")
        timeStart = '2022-03-28'
        timeEnd = '2022-03-29'
        # m.order_TimeQuery(timeStart, timeEnd)

    elif int(select) == 4:
        team = 'gat_waybill_list'
        handle = '手动'
        m.readFormHost(team)
        m.waybill_info('login_TmpCode',handle)
        m.chuhuo_info('login_TmpCode','login_TmpCode2',handle)

    elif int(select) == 5:
        team = 'gat_waybill_list'
        login_TmpCode = '8f60a2f666d73bb1ac7afedf7d31cb14'
        handle = '手0动'
        # m.readFormHost(team)
        # m.waybill_info(login_TmpCode, handle)
        m.chuhuo_info('217b05a1c8e8345fb1476c28f3fb91ee', 'fbb2cfe70910362d8dc937190da35507', handle)
        m.waybill_updata()


    print('查询耗时：', datetime.datetime.now() - start)