import pandas as pd
import os, shutil
import datetime
import xlwings
import win32api, win32con
import requests
from os import *
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel

from sqlalchemy import create_engine
from settings import Settings
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, \
    Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色
from 查询_产品明细 import QueryTwoT

from mysqlControl import MysqlControl
# -*- coding:utf-8 -*-
class QueryUpdate(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
        self.m = MysqlControl()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    # 获取签收表内容---港澳台更新签收总表(一)
    def readFormHost(self, team, write, last_time):
        start = datetime.datetime.now()
        path = r'D:\Users\Administrator\Desktop\需要用到的文件\数据库'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                if '需发货的改派订单' in dir or '需发货改派订单' in dir:
                    write = '需发货'
                self.wbsheetHost(filePath, team, write, last_time)
                os.remove(filePath)
                print('已清除上传文件…………')
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, team, write, last_time):
        match2 = {'slgat': '神龙港台',
                  'slgat_hfh': '火凤凰港台',
                  'slgat_hs': '红杉港台',
                  'slsc': '品牌',
                  'gat': '港台'}
        print('---正在获取 ' + match2[team] + ' 签收表的详情++++++')
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    if write == '本期':                 # 将数据库的临时表替换进指定的总表
                        print('++++正在导入更新：' + sht.name + ' 共：' + str(len(db)) + '行','sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        db.to_sql('gat_update', con=self.engine1, index=False, if_exists='replace')  # 将返回的dateFrame导入数据库的临时表
                        self.replacHost(team)
                    elif write == '上期':
                        print('++++正在导入更新：' + sht.name + ' 共：' + str(len(db)) + '行','sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        db.to_sql('gat_update', con=self.engine1, index=False, if_exists='replace')  # 将返回的dateFrame导入数据库的临时表
                        self.replaceHostbefore(team, last_time)
                    elif write == '需发货':
                        db = db[['订单编号']]
                        print('++++正在导入更新：' + sht.name + ' 共：' + str(len(db)) + '行','sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        db.to_sql('gat_update', con=self.engine1, index=False,if_exists='replace')  # 将返回的dateFrame导入数据库的临时表
                        self.online(team)
                    print('++++----->>>' + sht.name + '：订单更新完成++++')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()                                     # 工作表的订单信息

    # 更新-总表(地区签收率使用)
    def repHost(self, team):    # 更新-总表(地区签收率使用)
        try:
            print('正在更新总表中......')
            sql = '''update {0}_zqsb a, gat_update b
                            set a.`系统订单状态`= IF(b.`系统订单状态` = '', NULL, b.`系统订单状态`),
                                a.`系统物流状态`= IF(b.`系统物流状态` = '', NULL, b.`系统物流状态`),
                                a.`最终状态`= IF(b.`系统物流状态` = '', NULL, b.`系统物流状态`),
                                a.`价格`= IF(b.`价格` = '', NULL, b.`价格`),
                                a.`价格RMB`= IF(b.`价格RMB` = '', NULL, b.`价格RMB`)
                    where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')

    # 更新-总表（总体签收率使用）
    def replacHost(self, team):
        try:
            print('正在更新单表中......')
            sql = '''update {0}_order_list a, gat_update b
                                set a.`运单编号`= IF(b.`运单编号` = '', NULL, b.`运单编号`),
                                    a.`系统订单状态`= IF(b.`系统订单状态` = '', NULL, b.`系统订单状态`),
                                    a.`系统物流状态`= IF(b.`系统物流状态` = '', NULL, b.`系统物流状态`),
        		                    a.`是否改派`= IF(b.`是否改派` = '', NULL, b.`是否改派`),
        		                    a.`物流方式`= IF(b.`物流方式` = '', NULL, b.`物流方式`),
        		                    a.`物流名称`= IF(b.`物流名称` = '', NULL, b.`物流名称`),
        		                    a.`付款方式`= IF(b.`付款方式` = '', NULL, b.`付款方式`),
        		                    a.`产品id`= IF(b.`产品id` = '', NULL, b.`产品id`),
        		                    a.`产品名称`= IF(b.`产品名称` = '', NULL, b.`产品名称`),
        		                    a.`父级分类`= IF(b.`父级分类` = '', NULL, b.`父级分类`),
        		                    a.`二级分类`= IF(b.`二级分类` = '', NULL, b.`二级分类`)
        		                where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('正在更新总表中......')
            sql = '''update {0}_zqsb a, gat_update b
                                            set a.`运单编号`= IF(b.`运单编号` = '', NULL, b.`运单编号`),
                                                a.`系统订单状态`= IF(b.`系统订单状态` = '', NULL, b.`系统订单状态`),
                                                a.`系统物流状态`= IF(b.`系统物流状态` = '', NULL, b.`系统物流状态`),
                                                a.`最终状态`= IF(b.`最终状态` = '', NULL, b.`最终状态`),
                    		                    a.`是否改派`= IF(b.`是否改派` = '', NULL, b.`是否改派`),
                    		                    a.`物流方式`= IF(b.`物流方式` = '', NULL, b.`物流方式`),
                    		                    a.`物流名称`= IF(b.`物流名称` = '', NULL, b.`物流名称`),
                    		                    a.`付款方式`= IF(b.`付款方式` = '', NULL, b.`付款方式`),
                    		                    a.`产品id`= IF(b.`产品id` = '', NULL, b.`产品id`),
                    		                    a.`产品名称`= IF(b.`产品名称` = '', NULL, b.`产品名称`),
                    		                    a.`父级分类`= IF(b.`父级分类` = '', NULL, b.`父级分类`),
                    		                    a.`二级分类`= IF(b.`二级分类` = '', NULL, b.`二级分类`)
                    		                where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')
    # 更新-总表（修改已发货 使用）
    def online(self, team):
        try:
            print('正在更新单表中......')
            sql = '''update {0}_order_list a, gat_update b
                                set a.`系统订单状态`= '已发货'
        		                where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('正在更新总表中......')
            sql = '''update {0}_zqsb a, gat_update b
                                set a.`系统订单状态`= '已发货',
                                    a.`最终状态`= '在途'
                    		    where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')


    #  导出需要更新的签收表---港澳台(二)
    def EportOrder(self, team, month_last, month_yesterday, month_begin, check):
        match = {'gat': '港台',
                 'slsc': '品牌'}
        emailAdd = {'gat': 'giikinliujun@163.com',
                    'slsc': 'sunyaru@giikin.com'}
        today = datetime.date.today().strftime('%Y.%m.%d')
        if check == '是':
            print('正在第一次检查父级分类为空的信息---')
            sql = '''SELECT 订单编号,商品id,dp.`product_id`, dp.`name` product_name, dp.third_cate_id, dc.`ppname` cate, dc.`pname` second_cate, dc.`name` third_cate
                    FROM (SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
                            FROM {0}_order_list sl
                            WHERE sl.`日期`> '{1}' AND (sl.`父级分类` IS NULL or sl.`父级分类`= '') AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'))
                         ) s
                    LEFT JOIN dim_product_gat dp ON  dp.product_id = s.`产品id`
                    LEFT JOIN dim_cate dc ON  dc.id = dp.third_cate_id;'''.format(team, month_begin)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            if df.empty:
                print('  第一次检查没有为空的………… ')
            else:
                print('正在更新父级分类的详情…………')
                df.to_sql('tem_product_id', con=self.engine1, index=False, if_exists='replace')
                sql = '''update {0}_order_list a, tem_product_id b
                            set a.`父级分类`= IF(b.`cate` = '', a.`父级分类`, b.`cate`),
                                a.`二级分类`= IF(b.`second_cate` = '', a.`二级分类`, b.`second_cate`),
                                a.`三级分类`= IF(b.`third_cate` = '', a.`三级分类`, b.`third_cate`)
                        where a.`订单编号`= b.`订单编号`;'''.format(team)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                print('更新完成+++')

                print('正在第二次检查父级分类为空的信息---')
                sql = '''SELECT 订单编号,商品id,dp.`product_id`, dp.`name` product_name, dp.third_cate_id, dc.`ppname` cate, dc.`pname` second_cate, dc.`name` third_cate
                        FROM (SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
                                FROM gat_order_list sl
                                WHERE sl.`日期`>= '{1}' AND (sl.`父级分类` IS NULL or sl.`父级分类`= '') AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'))
                             ) s
                        LEFT JOIN dim_product_gat dp ON  dp.product_id = s.`产品id`
                        LEFT JOIN (SELECT * FROM dim_cate GROUP BY pid ) dc ON  dc.pid = dp.second_cate_id;'''.format(team,month_begin)
                df = pd.read_sql_query(sql=sql, con=self.engine1)
                if df.empty:
                    print('  第二次检查没有为空的………… ')
                else:
                    print('正在更新父级分类的详情…………')
                    sql = '''update {0}_order_list a, tem_product_id b
                                set a.`父级分类`= IF(b.`cate` = '', a.`父级分类`, b.`cate`),
                                    a.`二级分类`= IF(b.`second_cate` = '', a.`二级分类`, b.`second_cate`),
                                    a.`三级分类`= IF(b.`third_cate` = '', a.`三级分类`, b.`third_cate`)
                            where a.`订单编号`= b.`订单编号`;'''.format(team)
                    pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                    print('更新完成+++')

            print('正在第一次检查产品id为空的信息---')
            sql = '''SELECT 订单编号,商品id,dp.product_id, dp.`name` product_name, dp.third_cate_id
                    FROM (SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
                            FROM {0}_order_list sl
                            WHERE sl.`日期`>= '{1}' AND (sl.`产品名称` IS NULL or sl.`产品名称`= '') AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'))
                        ) s
                    LEFT JOIN dim_product_gat dp ON dp.product_id = s.`产品id`;'''.format(team, month_begin)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            if df.empty:
                print('  第一次检查没有为空的………… ')
            else:
                print('正在更新产品详情…………')
                df.to_sql('tem_product_id', con=self.engine1, index=False, if_exists='replace')
                sql = '''update {0}_order_list a, tem_product_id b
                            set a.`产品id`= IF(b.`product_id` = '',a.`产品id`, b.`product_id`),
                                a.`产品名称`= IF(b.`product_name` = '',a.`产品名称`, b.`product_name`)
                    where a.`订单编号`= b.`订单编号`;'''.format(team)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                print('更新完成+++')

            print('正在综合检查 父级分类、产品id 为空的信息---')
            sql = '''SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
                    FROM gat_order_list sl
                    WHERE sl.`日期`>= '{1}'
                        AND (sl.`父级分类` IS NULL or sl.`父级分类`= '' OR sl.`产品名称` IS NULL or sl.`产品名称`= '')
                        AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'));'''.format(team, month_begin)
            ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
            if ordersDict.empty:
                print(' ****** 没有要补充的信息; ****** ')
            else:
                print('！！！ 请再次补充缺少的数据中！！！')
                lw = QueryTwoT('+86-18538110674', 'qyz04163510.',"","手动")
                lw.productInfo('gat_order_list', ordersDict)

        if team in ('gat'):
            del_time = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y%m')
            sql = '''DELETE FROM gat_zqsb gt
                    WHERE gt.年月 >= {0}
                      and gt.`订单编号` IN (SELECT 订单编号 
                                            FROM gat_order_list gs
                                            WHERE gs.年月 >= {0}
                                              and gs.`系统订单状态` NOT IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'));'''.format(del_time)
            print('正在清除港澳台-总表的可能删除了的订单…………')
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('正在获取---' + match[team] + '---更新数据内容…………')
            sql = '''SELECT 年月, 旬, 日期, 团队, 币种, null 区域, 订单来源, a.订单编号, 电话号码, a.运单编号,
                            IF(出货时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), a.仓储扫描时间, 出货时间) 出货时间,
                            IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                            IF(状态时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), '', 状态时间) 状态时间,
                            IF(ISNULL(a.上线时间), IF(b.上线时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), null,b.上线时间), a.上线时间) 上线时间, 系统订单状态,
                            IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                            IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                            IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), 
                                                        IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , 
                                                            IF(物流方式 like '%天马%' and c.签收表物流状态 = '在途','未上线', c.标准物流状态)
                                                        ), 系统物流状态), '已退货') 最终状态,
                            IF(是否改派='二次改派', '改派', 是否改派) 是否改派,
                            物流方式,物流名称,null 运输方式,null 货物类型,是否低价,付款方式,产品id,产品名称,父级分类, 二级分类,三级分类, 下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB, null 价格区间, null 包裹重量, null 包裹体积,null 邮编, 
                            IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在, null 签收表订单编号, null 签收表运单编号, null 原运单号, b.物流状态 签收表物流状态, null 添加时间, null 成本价, null 物流花费, null 打包花费, null 其它花费, 添加物流单号时间,
                            省洲,市区,数量, a.下架时间, a.物流提货时间, a.完结状态, a.回款时间
                        FROM (SELECT * 
							FROM {0}_order_list g
							WHERE g.日期 >= '{2}' AND g.日期 <= '{3}' AND g.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
						) a
                        LEFT JOIN gat_wl_data b ON a.`查件单号` = b.`运单编号`
                        LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN {0}_return d ON a.订单编号 = d.订单编号
                        ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print('正在写入---' + match[team] + ' ---临时缓存…………')  # 备用临时缓存表
            df.to_sql('d1_{0}'.format(team), con=self.engine1, index=False, if_exists='replace', chunksize=10000)
            print('正在写入excel…………')
            df = df[['日期', '团队', '币种', '订单编号', '电话号码', '运单编号', '出货时间', '物流状态', '物流状态代码', '状态时间', '上线时间',
                     '系统订单状态', '系统物流状态', '最终状态', '是否改派', '物流方式', '物流名称', '签收表物流状态', '付款方式', '产品id', '产品名称',
                     '父级分类', '二级分类', '下单时间', '审核时间', '仓储扫描时间', '完结状态时间']]
            old_path = 'G:\\输出文件\\{} {} 更新-签收表.xlsx'.format(today, match[team])
            df.to_excel(old_path, sheet_name=match[team], index=False)
            new_path = "F:\\神龙签收率\\" + (datetime.datetime.now()).strftime('%m.%d') + '\\{} {} 更新-签收表.xlsx'.format(today,match[team])
            shutil.copyfile(old_path, new_path)     # copy到指定位置
            print('----已写入excel; 并复制到指定文件夹中')

        sql = '''DELETE FROM d1_gat gt WHERE gt.`订单编号` IN (SELECT 订单编号 FROM gat_易速配退运);'''
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''DELETE FROM gat_zqsb gt WHERE gt.年月 BETWEEN '202206' AND '202207' AND gt.`订单编号` IN (SELECT 订单编号 FROM gat_易速配退运);'''
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('已清除不参与计算的易速配退运订单（总表中）…………')

        print('正在写入' + match[team] + ' 全部签收表中…………')
        sql = 'REPLACE INTO {0}_zqsb SELECT *, NOW() 更新时间 FROM d1_{0};'.format(team)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''DELETE FROM gat_zqsb gz 
                 WHERE gz.`系统订单状态` = '已转采购' and gz.`是否改派` = '改派'
                   and gz.`审核时间` >= '{0} 00:00:00' AND gz.`日期` >= '{1}';'''.format(month_yesterday, month_last)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('已清除不参与计算的今日改派订单…………')

    # 导出总的签收表---各家族-港澳台(三)
    def EportOrderBook(self, team, month_last, month_yesterday):
        today = datetime.date.today().strftime('%Y.%m.%d')
        match = {'slgat': '神龙-港台',
                 'slgat_hfh': '火凤凰-港台',
                 'slgat_hs': '红杉-港台',
                 'slgat_js': '金狮-港台',
                 'slgat_jp': '小虎队-港台',
                 'slgat_run': '神龙-运营1组',
                 'gat': '港台',
                 'slsc': '品牌',
                 'slrb': '神龙-日本',
                 'slrb_jl': '精灵-日本',
                 'slrb_js': '金狮-日本',
                 'slrb_hs': '红杉-日本'}
        # if team in ('gat9'):
        #     month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        #     month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
        # else:
        #     month_last = '2021-08-01'
        #     month_yesterday = '2021-10-01'
        print(month_last)
        print(month_yesterday)
        sql = '''UPDATE gat_zqsb d
                        SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-易速配-顺丰%','香港-易速配-顺丰', 
                                                IF(d.`物流方式` LIKE '台湾-天马-711%','台湾-天马-新竹', 
                                                IF(d.`物流方式` LIKE '%优美宇通%' or d.`物流方式` LIKE '%铱熙无敌%','台湾-铱熙无敌-新竹', d.`物流方式`)) )
                        WHERE d.`是否改派` ='直发';'''
        print('正在修改-直发的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''UPDATE gat_zqsb d
                        SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-森鸿%','香港-森鸿-改派',
                                        IF(d.`物流方式` LIKE '香港-立邦%','香港-立邦-改派',
            							IF(d.`物流方式` LIKE '香港-易速配%','香港-易速配-改派',
            							IF(d.`物流方式` LIKE '台湾-立邦普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-大黄蜂普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-森鸿-新竹%','森鸿',
            							IF(d.`物流方式` LIKE '台湾-天马-顺丰%','天马顺丰',
            							IF(d.`物流方式` LIKE '台湾-天马-新竹%' OR d.`物流方式` LIKE '台湾-天马-711%','天马新竹',
            							IF(d.`物流方式` LIKE '台湾-天马-黑猫%','天马黑猫',
            							IF(d.`物流方式` LIKE '台湾-易速配-龟山%' OR d.`物流方式` LIKE '台湾-易速配-新竹%' OR d.`物流方式` LIKE '新易速配-台湾-改派%' OR d.`物流方式` = '易速配','龟山',
            							IF(d.`物流方式` LIKE '台湾-速派-新竹%' OR d.`物流方式` LIKE '台湾-速派-711超商%','速派',
            							IF(d.`物流方式` LIKE '台湾-大黄蜂普货头程-易速配尾程%' OR d.`物流方式` LIKE '台湾-立邦普货头程-易速配尾程%','龟山', 
            							IF(d.`物流方式` LIKE '台湾-优美宇通-新竹改派%','台湾-铱熙无敌-新竹改派', 
            							IF(d.`物流方式` LIKE '香港-圆通%','香港-圆通-改派', IF(d.`物流方式` LIKE '台湾-速派宅配通%','速派宅配通', d.`物流方式`))))))  )  )  )  )  )  )  )
                        WHERE d.`是否改派` ='改派';'''
        print('正在修改-改派的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('正在获取---' + match[team] + ' ---全部数据内容…………')
        sql = '''SELECT * FROM {0}_zqsb a WHERE a.日期 >= '{1}' AND a.日期 <= '{2}' ORDER BY a.`下单时间`;'''.format(team, month_last, month_yesterday)     # 港台查询函数导出
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        print('正在写入---' + match[team] + ' ---临时缓存…………')             # 备用临时缓存表
        df.to_sql('d1_{0}'.format(team), con=self.engine1, index=False, if_exists='replace', chunksize=10000)

        for tem in ('"神龙家族-港澳台"|slgat', '"红杉家族-港澳台", "红杉家族-港澳台2"|slgat_hs', '"火凤凰-港台(繁体)", "火凤凰-港澳台"|slgat_hfh', '"金狮-港澳台"|slgat_js', '"金鹏家族-小虎队"|slgat_jp', '"神龙-运营1组"|slgat_run'):
            tem1 = tem.split('|')[0]
            tem2 = tem.split('|')[1]
            sql = '''SELECT * FROM d1_{0} sl WHERE sl.`团队`in ({1});'''.format(team, tem1)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            # df.to_sql('d1_{0}'.format(tem2), con=self.engine1, index=False, if_exists='replace', chunksize=10000)
            old_path = 'G:\\输出文件\\{} {}签收表.xlsx'.format(today, match[tem2])
            df.to_excel(old_path, sheet_name=match[tem2], index=False)
            new_path = "F:\\神龙签收率\\" + (datetime.datetime.now()).strftime('%m.%d') + '\\{} {}签收表.xlsx'.format(today, match[tem2])
            shutil.copyfile(old_path, new_path)     # copy到指定位置
            print(tem2 + '----已写入excel; 并复制到指定文件夹中')
            # print('正在打印' + match[tem2] + ' 物流时效…………')
            # self.m.data_wl(tem2)
        try:
            print('正在转存中' + month_yesterday + '最近两个月的订单......')
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 订单来源, 订单编号, 出货时间, IF(`状态时间` = '',NULL,状态时间) as 状态时间, 上线时间, 最终状态,是否改派,物流方式,
                            产品id,父级分类,二级分类,三级分类,下单时间, 审核时间,仓储扫描时间,下架时间, 物流提货时间, 完结状态, 完结状态时间,回款时间, 价格RMB, curdate() 记录时间
                    FROM d1_{0} a WHERE a.`运单编号` is not null ;'''.format(team)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print('正在添加缓存中......')
            df.to_sql('gat_update_cp', con=self.engine1, index=False, if_exists='replace')
            print('正在转存数据中......')
            sql = '''REPLACE INTO qsb_{0} SELECT * FROM gat_update_cp; '''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('转存失败：', str(Exception) + str(e))
        print('转存成功…………')

        print('正在获取预估签收率的数据......')
        week: datetime = datetime.datetime.now()
        if week.isoweekday() == 0 or week.isoweekday() == 0:
            time_path: datetime = datetime.datetime.now()
            mkpath = "F:\\神龙签收率\\A预估签收率\\" + time_path.strftime('%m.%d')
            isExists = os.path.exists(mkpath)
            if not isExists:
                os.makedirs(mkpath)
            else:
                print(mkpath + ' 目录已存在')
            file_path = mkpath + '\\{} 预测_产品签收率_使用版.xlsx'.format(time_path.strftime('%m.%d'))
            sql = '''SELECT 团队 AS 家族, 币种, 产品id, 产品名称, 
			                concat(ROUND(产品金额团队占比* 100,2),'%') AS 产品金额团队占比, 最近3天单量,
			                concat(ROUND(IF(历史平均 = 0 OR 历史平均 IS NULL,预测, 历史平均)* 100,2),'%') AS '预测签收',
			                concat(ROUND(IF(历史平均 = 0 OR 历史平均 IS NULL,产品金额团队占比 * 预测影响, 产品金额团队占比 * 历史平均影响)* 100,2),'%') AS '预测签收影响',
			                IF(历史平均 = 0 OR 历史平均 IS NULL,'预测取值', '历史取值') AS 取值,
			                concat(ROUND(目标签收率* 100,2),'%') AS 目标签收率
                    FROM(SELECT s.团队, s.币种, s.产品id, s.产品名称, 
			                    总金额 / 团队金额 AS 产品金额团队占比,
			                    最近单量 AS 最近3天单量,
			                    avg_sign_rate AS 历史平均,
			                    IF(s.团队 = '神龙家族-港澳台' AND s.币种 = '台湾',(avg_sign_rate-0.825),
			                    IF(s.团队 = '神龙家族-港澳台' AND s.币种 = '香港',(avg_sign_rate-0.89) ,
			                    IF(s.团队 = '火凤凰-港澳台' AND s.币种 = '台湾',(avg_sign_rate-0.87),
			                    IF(s.团队 = '火凤凰-港澳台' AND s.币种 = '香港',(avg_sign_rate-0.89), 
			                    IF(s.团队 = '神龙-运营1组' AND s.币种 = '台湾',(avg_sign_rate-0.86), 
			                    IF(s.团队 = '神龙-运营1组' AND s.币种 = '香港',(avg_sign_rate-0.88), NULL)))))) AS 历史平均影响, 
			                    sign_rate AS 预测, 
			                    IF(s.团队 = '神龙家族-港澳台' AND s.币种 = '台湾',(sign_rate-0.825),
			                    IF(s.团队 = '神龙家族-港澳台' AND s.币种 = '香港',(sign_rate-0.89) ,
			                    IF(s.团队 = '火凤凰-港澳台' AND s.币种 = '台湾',(sign_rate-0.87),
			                    IF(s.团队 = '火凤凰-港澳台' AND s.币种 = '香港',(sign_rate-0.89), 
			                    IF(s.团队 = '神龙-运营1组' AND s.币种 = '台湾',(sign_rate-0.86), 
			                    IF(s.团队 = '神龙-运营1组' AND s.币种 = '香港',(sign_rate-0.88), NULL)))))) AS 预测影响,
			                    IF(s.团队 = '神龙家族-港澳台' AND s.币种 = '台湾',0.825,
			                    IF(s.团队 = '神龙家族-港澳台' AND s.币种 = '香港',0.89,
			                    IF(s.团队 = '火凤凰-港澳台' AND s.币种 = '台湾',0.87,
			                    IF(s.团队 = '火凤凰-港澳台' AND s.币种 = '香港',0.89, 
			                    IF(s.团队 = '神龙-运营1组' AND s.币种 = '台湾',0.86, 
			                    IF(s.团队 = '神龙-运营1组' AND s.币种 = '香港',0.88, NULL)))))) AS 目标签收率
                        FROM (SELECT cc.`团队`, cc.`币种`, cc.`产品id`, cc.`产品名称`, cc.`三级分类`,
						            COUNT(订单编号) AS 本月单量,
						            SUM(`价格RMB`) AS 总金额,
						            SUM(IF(最终状态 = "已签收",价格RMB,0)) as 签收金额,
						            SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
						            SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),价格RMB,0)) as 完成金额						
			                FROM gat_zqsb cc
			                where cc.年月 = DATE_FORMAT(curdate(),'%Y%m')
			                GROUP BY cc.`团队`, cc.`币种`, cc.`产品id`
                        ) s
                        LEFT JOIN 
                        (   SELECT cc.`团队`, cc.`币种`, SUM(`价格RMB`) AS 团队金额
	                        FROM gat_zqsb cc
                            where cc.年月 = DATE_FORMAT(curdate(),'%Y%m')
	                        GROUP BY cc.`团队`, cc.`币种`
                        ) s1  ON s.团队 = s1.团队 AND s.币种 = s1.币种
                        LEFT JOIN
                        (   SELECT cc.`团队`, cc.`币种`, cc.`产品id`, COUNT(订单编号) AS 最近单量
	                        FROM gat_zqsb cc
	                        where cc.日期  BETWEEN DATE_ADD(CURRENT_DATE(), INTERVAL -3 DAY) AND CURDATE()
	                        GROUP BY cc.`团队`, cc.`币种`, cc.`产品id`
                        ) s2 ON s.团队 = s2.团队 AND s.币种 = s2.币种 AND s.产品id = s2.产品id
                        LEFT JOIN
                        (   SELECT * FROM gk_stat_sign_rate) s3 ON s.团队 = s3.area_id AND s.币种 = s3.currency_id AND s.产品id = s3.goods_id
                        LEFT JOIN
                        (   SELECT * FROM gk_bi_estimate_goods) s4 ON s.团队 = s4.area_id AND s.币种 = s4.currency_id AND s.产品id = s4.goods_id
                    ) z;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            df.to_excel(file_path, sheet_name='使用', index=False)
            print('输出成功…………')
            try:
                print('正在运行 预估签收率 表宏…………')
                app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
                app.display_alerts = False
                wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
                wbsht1 = app.books.open(file_path)
                wbsht.macro('预估签收率修饰_使用')()
                wbsht1.save()
                wbsht1.close()
                wbsht.close()
                app.quit()
            except Exception as e:
                print('运行失败：', str(Exception) + str(e))
            print('运行成功…………')
        else:
            print('今日  无需获取预估签收率的数据！！！')

        print('正在获取同产品各团队对比的数据......')
        week: datetime = datetime.datetime.now()
        if week.isoweekday() == 2:
            time_path: datetime = datetime.datetime.now()
            mkpath = "F:\\神龙签收率\\A同产品各团队对比\\" + time_path.strftime('%m.%d')
            isExists = os.path.exists(mkpath)
            if not isExists:
                os.makedirs(mkpath)
            else:
                print(mkpath + ' 目录已存在')
            file_path = mkpath + '\\{} 同产品各团队对比_神龙.xlsx'.format(time_path.strftime('%m.%d'))
            sql = '''SELECT *
					FROM(SELECT	IFNULL(月份, '总计') 月份, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id, IFNULL(产品名称, '总计') 产品名称,
							SUM(神龙单量) 神龙单量, 
                                concat(ROUND(SUM(神龙签收) / SUM(神龙总量) * 100,2),'%') as 神龙总计签收,
                                concat(ROUND(SUM(神龙完成) / SUM(神龙总量) * 100,2),'%') as 神龙完成占比,					
							SUM(火凤凰单量) 火凤凰单量, 
                                concat(ROUND(SUM(火凤凰签收) / SUM(火凤凰总量) * 100,2),'%') as 火凤凰总计签收,
                                concat(ROUND(SUM(火凤凰完成) / SUM(火凤凰总量) * 100,2),'%') as 火凤凰完成占比,					
							SUM(神龙运营单量) 神龙运营单量, 
                                concat(ROUND(SUM(神龙运营签收) / SUM(神龙运营总量) * 100,2),'%') as 神龙运营总计签收,
                                concat(ROUND(SUM(神龙运营完成) / SUM(神龙运营总量) * 100,2),'%') as 神龙运营完成占比					
                        FROM(SELECT 年月 月份,币种 地区, 产品id, 产品名称,
                                    SUM(IF(家族 = '神龙',1,0)) as 神龙单量,
									SUM(IF(家族 = '神龙',价格,0)) as 神龙总量,
                                    SUM(IF(家族 = '神龙' AND 最终状态 = "已签收",价格,0)) as 神龙签收,
                                    SUM(IF(家族 = '神龙' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 神龙完成,
                                    SUM(IF(家族 = '火凤凰',1,0)) as 火凤凰单量,
									SUM(IF(家族 = '火凤凰',价格,0)) as 火凤凰总量,
                                    SUM(IF(家族 = '火凤凰' AND 最终状态 = "已签收",价格,0)) as 火凤凰签收,
                                    SUM(IF(家族 = '火凤凰' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 火凤凰完成,
                                    SUM(IF(家族 = '神龙-运营1组',1,0)) as 神龙运营单量,
									SUM(IF(家族 = '神龙-运营1组',价格,0)) as 神龙运营总量,
                                    SUM(IF(家族 = '神龙-运营1组' AND 最终状态 = "已签收",价格,0)) as 神龙运营签收,
                                    SUM(IF(家族 = '神龙-运营1组' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 神龙运营完成,
                                    SUM(IF(家族 = '小虎队',1,0)) as 小虎队单量,
									SUM(IF(家族 = '小虎队',价格,0)) as 小虎队总量,
                                    SUM(IF(家族 = '小虎队' AND 最终状态 = "已签收",价格,0)) as 小虎队签收,
                                    SUM(IF(家族 = '小虎队' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 小虎队完成
                            FROM gat_zqsb_cache cc
							WHERE cc.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m')
							GROUP BY cc.年月,cc.币种,cc.产品id
						) s
					GROUP BY 月份,地区,产品id		
--                   WITH ROLLUP 
					) ss
                   ORDER BY FIELD(月份,DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'),'总计'),
                            FIELD(地区,'台湾','香港','总计'),
                            神龙单量 DESC;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            df.to_excel(file_path, sheet_name='使用', index=False)
            print('输出成功…………')
            try:
                print('正在运行 同产品各团队对比 表宏…………')
                app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
                app.display_alerts = False
                wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
                wbsht1 = app.books.open(file_path)
                wbsht.macro('同产品各团队对比_使用')()
                wbsht1.save()
                wbsht1.close()
                wbsht.close()
                app.quit()
            except Exception as e:
                print('运行失败：', str(Exception) + str(e))
            print('运行成功…………')

            file_path = mkpath + '\\{} 同产品各团队对比_火凤凰.xlsx'.format(time_path.strftime('%m.%d'))
            sql = '''SELECT *
        					FROM(SELECT	IFNULL(月份, '总计') 月份, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id, IFNULL(产品名称, '总计') 产品名称,
        							SUM(火凤凰单量) 火凤凰单量, 
                                        concat(ROUND(SUM(火凤凰签收) / SUM(火凤凰总量) * 100,2),'%') as 火凤凰总计签收,
                                        concat(ROUND(SUM(火凤凰完成) / SUM(火凤凰总量) * 100,2),'%') as 火凤凰完成占比,	
        							SUM(神龙单量) 神龙单量, 
                                        concat(ROUND(SUM(神龙签收) / SUM(神龙总量) * 100,2),'%') as 神龙总计签收,
                                        concat(ROUND(SUM(神龙完成) / SUM(神龙总量) * 100,2),'%') as 神龙完成占比,									
        							SUM(神龙运营单量) 神龙运营单量, 
                                        concat(ROUND(SUM(神龙运营签收) / SUM(神龙运营总量) * 100,2),'%') as 神龙运营总计签收,
                                        concat(ROUND(SUM(神龙运营完成) / SUM(神龙运营总量) * 100,2),'%') as 神龙运营完成占比					
                                FROM(SELECT 年月 月份,币种 地区, 产品id, 产品名称,
                                            SUM(IF(家族 = '神龙',1,0)) as 神龙单量,
        									SUM(IF(家族 = '神龙',价格,0)) as 神龙总量,
                                            SUM(IF(家族 = '神龙' AND 最终状态 = "已签收",价格,0)) as 神龙签收,
                                            SUM(IF(家族 = '神龙' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 神龙完成,
                                            SUM(IF(家族 = '火凤凰',1,0)) as 火凤凰单量,
        									SUM(IF(家族 = '火凤凰',价格,0)) as 火凤凰总量,
                                            SUM(IF(家族 = '火凤凰' AND 最终状态 = "已签收",价格,0)) as 火凤凰签收,
                                            SUM(IF(家族 = '火凤凰' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 火凤凰完成,
                                            SUM(IF(家族 = '神龙-运营1组',1,0)) as 神龙运营单量,
        									SUM(IF(家族 = '神龙-运营1组',价格,0)) as 神龙运营总量,
                                            SUM(IF(家族 = '神龙-运营1组' AND 最终状态 = "已签收",价格,0)) as 神龙运营签收,
                                            SUM(IF(家族 = '神龙-运营1组' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 神龙运营完成,
                                            SUM(IF(家族 = '小虎队',1,0)) as 小虎队单量,
        									SUM(IF(家族 = '小虎队',价格,0)) as 小虎队总量,
                                            SUM(IF(家族 = '小虎队' AND 最终状态 = "已签收",价格,0)) as 小虎队签收,
                                            SUM(IF(家族 = '小虎队' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 小虎队完成
                                    FROM gat_zqsb_cache cc
        							WHERE cc.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m')
        							GROUP BY cc.年月,cc.币种,cc.产品id
        						) s
        					GROUP BY 月份,地区,产品id		
        --                   WITH ROLLUP 
        					) ss
                           ORDER BY FIELD(月份,DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'),'总计'),
                                    FIELD(地区,'台湾','香港','总计'),
                                    火凤凰单量 DESC;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            df.to_excel(file_path, sheet_name='使用', index=False)
            print('输出成功…………')
            try:
                print('正在运行 同产品各团队对比 表宏…………')
                app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
                app.display_alerts = False
                wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
                wbsht1 = app.books.open(file_path)
                wbsht.macro('同产品各团队对比_使用')()
                wbsht1.save()
                wbsht1.close()
                wbsht.close()
                app.quit()
            except Exception as e:
                print('运行失败：', str(Exception) + str(e))
            print('运行成功…………')

        else:
            print('今日  无需获取 同产品各团队对比 的数据！！！')

    # 新版签收率-报表(自己看的) - 单量计算
    def gat_new(self, team, month_last, month_yesterday):  # 报表各团队近两个月的物流数据
        month_now = datetime.datetime.now().strftime('%Y-%m-%d')
        match = {'gat': '港台'}
        emailAdd = {'台湾': 'giikinliujun@163.com',
                    '香港': 'giikinliujun@163.com',
                    '品牌': 'sunyaru@giikin.com'}
        # if team == 'gat9':
        #     month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        #     month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
        # else:
        #     month_last = '2021-08-01'
        #     month_yesterday = '2021-09-30'
        print(month_last)
        print(month_yesterday)
        sql = '''UPDATE gat_zqsb d
                        SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-易速配-顺丰%','香港-易速配-顺丰', IF(d.`物流方式` LIKE '台湾-天马-711%','台湾-天马-新竹', IF(d.`物流方式` LIKE '%优美宇通%' or d.`物流方式` LIKE '%铱熙无敌%','台湾-铱熙无敌-新竹', d.`物流方式`)) )
                        WHERE d.`是否改派` ='直发';'''
        print('正在修改-直发的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''UPDATE gat_zqsb d
                        SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-森鸿%','香港-森鸿-改派',
                                        IF(d.`物流方式` LIKE '香港-立邦%','香港-立邦-改派',
            							IF(d.`物流方式` LIKE '香港-易速配%','香港-易速配-改派',
            							IF(d.`物流方式` LIKE '台湾-立邦普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-大黄蜂普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-森鸿-新竹%','森鸿',
            							IF(d.`物流方式` LIKE '台湾-天马-顺丰%','天马顺丰',
            							IF(d.`物流方式` LIKE '台湾-天马-新竹%' OR d.`物流方式` LIKE '台湾-天马-711%','天马新竹',
            							IF(d.`物流方式` LIKE '台湾-天马-黑猫%','天马黑猫',
            							IF(d.`物流方式` LIKE '台湾-易速配-龟山%' OR d.`物流方式` LIKE '台湾-易速配-新竹%' OR d.`物流方式` LIKE '新易速配-台湾-改派%' OR d.`物流方式` = '易速配','龟山',
            							IF(d.`物流方式` LIKE '台湾-速派-新竹%' OR d.`物流方式` LIKE '台湾-速派-711超商%','速派',
            							IF(d.`物流方式` LIKE '台湾-大黄蜂普货头程-易速配尾程%' OR d.`物流方式` LIKE '台湾-立邦普货头程-易速配尾程%','龟山', 
            							IF(d.`物流方式` LIKE '台湾-优美宇通-新竹改派%','台湾-铱熙无敌-新竹改派', 
            							IF(d.`物流方式` LIKE '香港-圆通%','香港-圆通-改派', IF(d.`物流方式` LIKE '台湾-速派宅配通%','速派宅配通', d.`物流方式`))))))  )  )  )  )  )  )  )
                        WHERE d.`是否改派` ='改派';'''
        print('正在修改-改派的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 物流分类
        print('正在获取---物流分类…………')
        sql0 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.物流方式,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,					
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                    concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                        concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                        concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                FROM ( SELECT IFNULL(s1.币种,'合计') as 币种,
                            IFNULL(s1.家族,'合计') as 家族,
                            IFNULL(s1.年月,'合计') as 年月,
                            IFNULL(s1.是否改派,'合计') as 是否改派,
                            IFNULL(s1.物流方式,'合计') as 物流方式,
                            SUM(s1.签收) as 签收,
                            SUM(s1.拒收) as 拒收,
                            SUM(s1.在途) as 在途,
                            SUM(s1.未发货) as 未发货,
                            SUM(s1.未上线) as 未上线,
                            SUM(s1.已退货) as 已退货,
                            SUM(s1.理赔) as 理赔,
                            SUM(s1.自发头程丢件) as 自发头程丢件,
                            SUM(s1.已发货) as 已发货,
                            SUM(s1.已完成) as 已完成,
                            SUM(s1.总订单) as 总订单,
                            SUM(s1.签收金额) as 签收金额,
                            SUM(s1.退货金额) as 退货金额,
                            SUM(s1.完成金额) as 完成金额,
                            SUM(s1.发货金额) as 发货金额,
                            SUM(s1.总计金额) as 总计金额
                    FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.物流方式 as 物流方式,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                            ) cx
                            GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                            ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
                    ) s1
                    GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
                    with rollup
                ) s2
                GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`物流方式` 
                HAVING s2.年月 <> '合计'
    ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
            FIELD(s2.`币种`,'台湾','香港','合计'),
            s2.`年月`,
            FIELD(s2.`是否改派`,'改派','直发','合计'),
            FIELD(s2.`物流方式`, '台湾-大黄蜂普货头程-森鸿尾程','台湾-大黄蜂普货头程-易速配尾程', '台湾-立邦普货头程-森鸿尾程','台湾-易速配-TW海快','台湾-铱熙无敌-新竹','台湾-立邦普货头程-易速配尾程', 
                                '台湾-森鸿-新竹-自发头程', '台湾-速派-711超商', '台湾-速派-新竹','台湾-天马-新竹','台湾-天马-顺丰','台湾-天马-黑猫','台湾-易速配-新竹',
                                '香港-圆通', '香港-立邦-顺丰','香港-易速配-顺丰','香港-易速配-顺丰YC', '香港-森鸿-SH渠道','香港-森鸿-顺丰渠道',
                                '龟山','森鸿','速派', '速派宅配通','天马顺丰','天马新竹','香港-圆通-改派','香港-立邦-改派','香港-森鸿-改派','香港-易速配-改派','合计' ),
            s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)
        # 物流分旬
        print('正在获取---物流分旬…………')
        sql11 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.物流方式,s2.旬,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,	
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
                    concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                        concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                        concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'    
                FROM (SELECT IFNULL(s1.币种,'合计') as 币种,
                            IFNULL(s1.家族,'合计') as 家族,
                            IFNULL(s1.年月,'合计') as 年月,
                            IFNULL(s1.是否改派,'合计') as 是否改派,
                            IFNULL(s1.物流方式,'合计') as 物流方式,
                            IFNULL(s1.旬,'合计') as 旬,
                            SUM(s1.签收) as 签收,
                            SUM(s1.拒收) as 拒收,
                            SUM(s1.在途) as 在途,
                            SUM(s1.未发货) as 未发货,
                            SUM(s1.未上线) as 未上线,
                            SUM(s1.已退货) as 已退货,
                            SUM(s1.理赔) as 理赔,
                            SUM(s1.自发头程丢件) as 自发头程丢件,
                            SUM(s1.已发货) as 已发货,
                            SUM(s1.已完成) as 已完成,
                            SUM(s1.总订单) as 总订单,
                            s1.总订单量,
							s1.已发货单量,
							s1.已完成单量,
                            SUM(s1.签收金额) as 签收金额,
                            SUM(s1.退货金额) as 退货金额,
                            SUM(s1.完成金额) as 完成金额,
                            SUM(s1.发货金额) as 发货金额,
                            SUM(s1.总计金额) as 总计金额
                    FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.物流方式 as 物流方式,
                                IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
								总订单量,
								已发货单量,
								已完成单量,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族
                                    FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                            ) cx
                            LEFT JOIN 
							    (SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                                FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                    ) dg  
								    GROUP BY dg.币种,dg.家族,dg.年月
                            ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                            GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派`, cx.`物流方式`, cx.`旬`
                            ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`是否改派`, s1.`物流方式`, s1.`旬`
                        with rollup
                    ) s2 
                    GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`物流方式`, s2.`旬`
                    HAVING s2.是否改派 <> '合计'
        ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                FIELD(s2.`币种`,'台湾','香港','合计'),
                s2.`年月`,
                FIELD(s2.`是否改派`,'改派','直发','合计'),
                FIELD(s2.`物流方式`, '台湾-大黄蜂普货头程-森鸿尾程','台湾-大黄蜂普货头程-易速配尾程', '台湾-立邦普货头程-森鸿尾程','台湾-易速配-TW海快','台湾-铱熙无敌-新竹','台湾-立邦普货头程-易速配尾程', 
                                '台湾-森鸿-新竹-自发头程', '台湾-速派-711超商', '台湾-速派-新竹','台湾-天马-新竹','台湾-天马-顺丰','台湾-天马-黑猫','台湾-易速配-新竹',
                                '香港-圆通', '香港-立邦-顺丰','香港-易速配-顺丰','香港-易速配-顺丰YC', '香港-森鸿-SH渠道','香港-森鸿-顺丰渠道',
                                '龟山','森鸿','速派','速派宅配通','天马顺丰','天马新竹','香港-圆通-改派','香港-立邦-改派','香港-森鸿-改派','香港-易速配-改派','合计' ),
                FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)
        listT.append(df11)

        # 父级分旬
        print('正在获取---父级分旬…………')
        sql12 = '''SELECT s2.家族,s2.币种,s2.年月,s2.父级分类,s2.旬,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,	
                    concat(ROUND(s2.签收 / s2.已完成 * 100,2),'%') as 完成签收,
                        concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') as 总计签收,
                        concat(ROUND(s2.已完成 / s2.总订单 * 100,2),'%') as 完成占比,
                        concat(ROUND(s2.已完成 / s2.已发货 * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(s2.已退货 / s2.总订单 * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
					concat(ROUND(s2.签收金额 / s2.完成金额 * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(s2.完成金额 / s2.总计金额 * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(s2.完成金额 / s2.发货金额 * 100,2),'%') as '已完成/已发货(金额)',
						concat(ROUND(s2.退货金额 / s2.总计金额 * 100,2),'%') as '退货率(金额)'
				 FROM ( SELECT  IFNULL(s1.币种,'合计') as 币种,IFNULL(s1.家族,'合计') as 家族,IFNULL(s1.年月,'合计') as 年月,IFNULL(s1.父级分类,'合计') as 父级分类,IFNULL(s1.旬,'合计') as 旬,
								SUM(s1.签收) as 签收,
								SUM(s1.拒收) as 拒收,
								SUM(s1.在途) as 在途,
								SUM(s1.未发货) as 未发货,
								SUM(s1.未上线) as 未上线,
								SUM(s1.已退货) as 已退货,
								SUM(s1.理赔) as 理赔,
								SUM(s1.自发头程丢件) as 自发头程丢件,
								SUM(s1.已发货) as 已发货,
								SUM(s1.已完成) as 已完成,
								SUM(s1.总订单) as 总订单,
                                s1.总订单量,s1.已发货单量,s1.已完成单量,
								SUM(s1.签收金额) as 签收金额,
								SUM(s1.退货金额) as 退货金额,
								SUM(s1.完成金额) as 完成金额,
								SUM(s1.发货金额) as 发货金额,
								SUM(s1.总计金额) as 总计金额
                        FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.父级分类 as 父级分类,IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
                                    SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                    SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                    SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                    SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                    SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                    SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                    SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                    SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                    count(订单编号) as 总订单,
                                    总订单量,已发货单量,已完成单量,
                                    count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                    SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                    SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                    SUM(`价格RMB`) as 总计金额,
                                    SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                                FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                ) cx
                                LEFT JOIN 
							        (SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                                    FROM (SELECT *,
                                                IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                            FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                        ) dg  
								        GROUP BY dg.币种,dg.家族,dg.年月
                                ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类`, cx.`旬`
                                ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`父级分类`, s1.`旬`
                        with rollup
                ) s2 HAVING s2.年月 <> '合计'
            ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                    FIELD(s2.`币种`,'台湾','香港','合计'),
                    s2.`年月`,
                    FIELD(s2.父级分类, '居家百货', '电子电器', '服饰', '医药保健',  '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','合计' ),
                    FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                    s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df12 = pd.read_sql_query(sql=sql12, con=self.engine1)
        listT.append(df12)
        # 二级分旬
        print('正在获取---二级分旬…………')
        sql13 = '''SELECT s2.家族,s2.币种,s2.年月,s2.父级分类,s2.二级分类,s2.旬,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,	
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
						concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
						concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
				 FROM ( SELECT  IFNULL(s1.币种,'合计') as 币种,IFNULL(s1.家族,'合计') as 家族,IFNULL(s1.年月,'合计') as 年月,IFNULL(s1.父级分类,'合计') as 父级分类,IFNULL(s1.二级分类,'合计') as 二级分类,IFNULL(s1.旬,'合计') as 旬,
								SUM(s1.签收) as 签收,
								SUM(s1.拒收) as 拒收,
								SUM(s1.在途) as 在途,
								SUM(s1.未发货) as 未发货,
								SUM(s1.未上线) as 未上线,
								SUM(s1.已退货) as 已退货,
								SUM(s1.理赔) as 理赔,
								SUM(s1.自发头程丢件) as 自发头程丢件,
								SUM(s1.已发货) as 已发货,
								SUM(s1.已完成) as 已完成,
								SUM(s1.总订单) as 总订单,
                                s1.总订单量,s1.已发货单量,s1.已完成单量,
								SUM(s1.签收金额) as 签收金额,
								SUM(s1.退货金额) as 退货金额,
								SUM(s1.完成金额) as 完成金额,
								SUM(s1.发货金额) as 发货金额,
								SUM(s1.总计金额) as 总计金额
                        FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.父级分类 as 父级分类,cx.二级分类 as 二级分类,IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
								总订单量,已发货单量,已完成单量,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                            ) cx
                            LEFT JOIN 
							    (SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                                FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                    ) dg  
								    GROUP BY dg.币种,dg.家族,dg.年月
                            ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                            GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类`, cx.`二级分类`, cx.`旬`
                            ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类`, cx.`二级分类` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`父级分类`, s1.`二级分类`, s1.`旬`
                        with rollup
                ) s2 HAVING s2.年月 <> '合计'
        ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                FIELD(s2.`币种`,'台湾','香港','合计'),
                s2.`年月`,
                FIELD(s2.父级分类, '居家百货', '电子电器', '服饰', '医药保健', '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','合计' ),
                FIELD(s2.二级分类,'个人洗护','皮鞋','日用百货','影音娱乐','家用电器','药品','上衣','下装'
                            ,'饰品','保健器械','保健食品','彩妆','钱包','休闲运动鞋','内衣','护理护具','凉/拖鞋'
                            ,'裙子','个护电器','配饰','护肤','布艺家纺','母婴用品','厨房用品','汽车用品','双肩包'
                            ,'单肩包','手机外设','电脑外设','成人保健','套装','靴子','手表手环','行李箱包','户外运动'
                            ,'玩具','手表','宠物用品','合计' ),
                FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df13 = pd.read_sql_query(sql=sql13, con=self.engine1)
        listT.append(df13)

        # 产品整月 台湾
        print('正在获取---产品整月 台湾…………')
        sql14 = '''SELECT *
                FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                            IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                            SUM(s1.已签收) as 已签收,
						    SUM(s1.拒收) as 拒收,
						    SUM(s1.已退货) as 已退货,
						    SUM(s1.已完成) as 已完成,
						    SUM(s1.总订单) as 总订单,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						    concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
						SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
						    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
						    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
						    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
						    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
						    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
						    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
						    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
						SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
						    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
						    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
						    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
						    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
						    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
						    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
						    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						    SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						    SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						    SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						    SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						    concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						    concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						    concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
						SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
						    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
						    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
						    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
						    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
						    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
						    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
						    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
						SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
						    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
						    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
						    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
						    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
						    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
						    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
						    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
						SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
						    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
						    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
						    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
						    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
						    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
						    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
						    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
						SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
						    SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
						    SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
						    SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
						    SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
						    concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
						    concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
						    concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
						SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
						    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
						    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
						    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
						    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
						    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
				    		concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
					    	concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
					    	concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
					    	concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
						SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
						    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
					    	SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
					    	SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
					    	SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
				    		concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
				    		concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
					    	concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
						SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
					    	SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
					    	SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
					    	SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
					    	SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
					    	concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
					    	concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
					    	concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
						SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
					    	SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
						    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
					    	SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
					        SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
					    	concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
					    	concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
					    	concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
						SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
					    	SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
					    	SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
					    	SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
					    	SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
					    	concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
					    	concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
					    	concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
						SUM(s1.铱熙无敌已签收) as '铱熙无敌已签收',
					    	SUM(s1.铱熙无敌拒收) as '铱熙无敌拒收',
					    	SUM(s1.铱熙无敌已退货) as '铱熙无敌已退货',
					    	SUM(s1.铱熙无敌已完成) as '铱熙无敌已完成',
					    	SUM(s1.铱熙无敌总订单) as '铱熙无敌总订单',
					    	concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌完成签收',
					    	concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌总计签收',
					    	concat(ROUND(SUM(s1.铱熙无敌已完成) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌完成占比',
					    	concat(ROUND(SUM(s1.铱熙无敌已退货) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌退货率',
					    	concat(ROUND(SUM(s1.铱熙无敌拒收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌拒收率',
						SUM(s1.龟山改派已签收) as '龟山改派已签收',
					    	SUM(s1.龟山改派拒收) as '龟山改派拒收',
					    	SUM(s1.龟山改派已退货) as '龟山改派已退货',
					    	SUM(s1.龟山改派已完成) as '龟山改派已完成',
					    	SUM(s1.龟山改派总订单) as '龟山改派总订单',
					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
					    	concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
					    	concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
					    	concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
				    	SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
					    	SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
					    	SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
					    	SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
					    	SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
					    	concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
					    	concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
					    	concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
						SUM(s1.速派改派已签收) as '速派改派已签收',
					    	SUM(s1.速派改派拒收) as '速派改派拒收',
					    	SUM(s1.速派改派已退货) as '速派改派已退货',
					    	SUM(s1.速派改派已完成) as '速派改派已完成',
					    	SUM(s1.速派改派总订单) as '速派改派总订单',
					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
					    	concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
					    	concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
					    	concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
						SUM(s1.速派宅配通已签收) as '速派宅配通改派已签收',
					    	SUM(s1.速派宅配通拒收) as '速派宅配通改派拒收',
					    	SUM(s1.速派宅配通已退货) as '速派宅配通改派已退货',
					    	SUM(s1.速派宅配通已完成) as '速派宅配通改派已完成',
					    	SUM(s1.速派宅配通总订单) as '速派宅配通改派总订单',
					    	concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派完成签收',
					    	concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派总计签收',
					    	concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派完成占比',
					    	concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派退货率',
					    	concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派拒收率',
						SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
					    	SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
					    	SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
					    	SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
					    	SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
					    	concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
					    	concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
					    	concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
						SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
					    	SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
					    	SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
					    	SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
					    	SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
					    	concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
					    	concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
					    	concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
						SUM(s1.铱熙无敌改派已签收) as '铱熙无敌改派已签收',
					    	SUM(s1.铱熙无敌改派拒收) as '铱熙无敌改派拒收',
					    	SUM(s1.铱熙无敌改派已退货) as '铱熙无敌改派已退货',
					    	SUM(s1.铱熙无敌改派已完成) as '铱熙无敌改派已完成',
					    	SUM(s1.铱熙无敌改派总订单) as '铱熙无敌改派总订单',
					    	concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
					    	concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
					    	concat(ROUND(SUM(s1.铱熙无敌改派已完成) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
					    	concat(ROUND(SUM(s1.铱熙无敌改派已退货) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
					    	concat(ROUND(SUM(s1.铱熙无敌改派拒收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率'
                    FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌总订单,
								SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌已签收,
								SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌拒收,
								SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌已退货,
								SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌已完成,
							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
							SUM(IF(cx.物流方式 = "速派宅配通" ,1,0)) AS 速派宅配通总订单,
								SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
								SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
								SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
								SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
							SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌改派总订单,
								SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌改派已签收,
								SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌改派拒收,
								SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌改派已退货,
								SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌改派已完成
				            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                            ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                        ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                        WITH ROLLUP 
                ) s HAVING s.月份 != '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df14 = pd.read_sql_query(sql=sql14, con=self.engine1)
        listT.append(df14)
        # 产品分旬 台湾
        print('正在获取---产品分旬 台湾…………')
        sql15 = '''SELECT *
                    FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
						SUM(s1.已签收) as 已签收,
						SUM(s1.拒收) as 拒收,
						SUM(s1.已退货) as 已退货,
						SUM(s1.已完成) as 已完成,
						SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
						SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
						SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
						SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
						SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
						concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
						concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
						concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
					SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
						SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
						SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
						SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
						SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
						concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
						concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
						concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
					SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
					SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
						SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
						SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
						SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
						SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
						concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
						concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
						concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
					SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
						SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
						SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
						SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
						SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
						concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
						concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
						concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
					SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
						SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
						SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
						SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
						SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
						concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
						concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
						concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
					SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
						SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
						SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
						SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
						SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
						concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
						concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
						concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
					SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
						SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
						SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
						SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
						SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
						concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
						concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
						concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
					SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
						SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
						SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
						SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
						SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
						concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
						concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
						concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
					SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
						SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
						SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
						SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
						SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
						concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
						concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
						concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
					SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
						SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
						SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
						SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
						SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
						concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
						concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
						concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
					SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
						SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
						SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
						SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
						SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
						concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
						concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
						concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
                    SUM(s1.铱熙无敌已签收) as '铱熙无敌已签收',
                        SUM(s1.铱熙无敌拒收) as '铱熙无敌拒收',
                        SUM(s1.铱熙无敌已退货) as '铱熙无敌已退货',
                        SUM(s1.铱熙无敌已完成) as '铱熙无敌已完成',
                        SUM(s1.铱熙无敌总订单) as '铱熙无敌总订单',
                        concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌完成签收',
                        concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌总计签收',
                        concat(ROUND(SUM(s1.铱熙无敌已完成) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌完成占比',
                        concat(ROUND(SUM(s1.铱熙无敌已退货) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌退货率',
                        concat(ROUND(SUM(s1.铱熙无敌拒收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌拒收率',
					SUM(s1.龟山改派已签收) as '龟山改派已签收',
						SUM(s1.龟山改派拒收) as '龟山改派拒收',
						SUM(s1.龟山改派已退货) as '龟山改派已退货',
						SUM(s1.龟山改派已完成) as '龟山改派已完成',
						SUM(s1.龟山改派总订单) as '龟山改派总订单',
						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
						concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
						concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
						concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
					SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
						SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
						SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
						SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
						SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
						concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
						concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
						concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
					SUM(s1.速派改派已签收) as '速派改派已签收',
						SUM(s1.速派改派拒收) as '速派改派拒收',
						SUM(s1.速派改派已退货) as '速派改派已退货',
						SUM(s1.速派改派已完成) as '速派改派已完成',
						SUM(s1.速派改派总订单) as '速派改派总订单',
						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
						concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
						concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
						concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
                    SUM(s1.速派宅配通已签收) as '速派宅配通改派已签收',
                        SUM(s1.速派宅配通拒收) as '速派宅配通改派拒收',
                        SUM(s1.速派宅配通已退货) as '速派宅配通改派已退货',
                        SUM(s1.速派宅配通已完成) as '速派宅配通改派已完成',
                        SUM(s1.速派宅配通总订单) as '速派宅配通改派总订单',
                        concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派完成签收',
                        concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派总计签收',
                        concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派完成占比',
                        concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派退货率',
                        concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派拒收率',
					SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
						SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
						SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
						SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
						SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
						concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
						concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
						concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
					SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
						SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
						SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
						SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
						SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
						concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
						concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
						concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
                    SUM(s1.铱熙无敌改派已签收) as '铱熙无敌改派已签收',
                        SUM(s1.铱熙无敌改派拒收) as '铱熙无敌改派拒收',
                        SUM(s1.铱熙无敌改派已退货) as '铱熙无敌改派已退货',
                        SUM(s1.铱熙无敌改派已完成) as '铱熙无敌改派已完成',
                        SUM(s1.铱熙无敌改派总订单) as '铱熙无敌改派总订单',
                        concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
                        concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
                        concat(ROUND(SUM(s1.铱熙无敌改派已完成) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
                        concat(ROUND(SUM(s1.铱熙无敌改派已退货) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
                        concat(ROUND(SUM(s1.铱熙无敌改派拒收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率'
                FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
                            SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌总订单,
                                SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌已签收,
                                SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌拒收,
                                SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌已退货,
                                SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌已完成,
							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
							SUM(IF(cx.物流方式 = "速派宅配通" ,1,0)) AS 速派宅配通总订单,
								SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
								SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
								SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
								SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,		
							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
                            SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌改派总订单,
                                SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌改派已签收,
                                SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌改派拒收,
                                SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌改派已退货,
                                SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌改派已完成
				        FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                        ) cx WHERE cx.`币种` = '台湾'
                    GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                    ) s1
                GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                WITH ROLLUP 
            ) s HAVING s.旬 != '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df15 = pd.read_sql_query(sql=sql15, con=self.engine1)
        listT.append(df15)

        # 产品整月 香港
        print('正在获取---产品整月 香港…………')
        sql16 = '''SELECT *
                    FROM(SELECT IFNULL(s1.家族, '合计') 家族,
                                IFNULL(s1.地区, '合计') 地区,
                                IFNULL(s1.月份, '合计') 月份,
                        IFNULL(s1.产品id, '合计') 产品id,
                        IFNULL(s1.产品名称, '合计') 产品名称,
                        IFNULL(s1.父级分类, '合计') 父级分类,
                        IFNULL(s1.二级分类, '合计') 二级分类,
						SUM(s1.已签收) as 已签收,
						SUM(s1.拒收) as 拒收,
						SUM(s1.已退货) as 已退货,
						SUM(s1.已完成) as 已完成,
						SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.香港圆通已签收) as '香港-圆通已签收',
						SUM(s1.香港圆通拒收) as '香港-圆通拒收',
						SUM(s1.香港圆通已退货) as '香港-圆通拒收已退货',
						SUM(s1.香港圆通已完成) as '香港-圆通拒收已完成',
						SUM(s1.香港圆通总订单) as '香港-圆通拒收总订单',
						concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收完成签收',
						concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收总计签收',
						concat(ROUND(SUM(s1.香港圆通已完成) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收完成占比',
						concat(ROUND(SUM(s1.香港圆通已退货) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收退货率',
						concat(ROUND(SUM(s1.香港圆通拒收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收拒收率',
					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
					SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
						SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
						SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
						SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
						SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
						concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
						concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
						concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
						concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
						concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "香港-圆通" ,1,0)) AS 香港圆通总订单,
								SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "已签收",1,0)) as 香港圆通已签收,
								SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "拒收",1,0)) as 香港圆通拒收,
								SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "已退货",1,0)) as 香港圆通已退货,
								SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 香港圆通已完成,
							SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
							SUM(IF(cx.物流方式 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
								SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
								SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
								SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
								SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
				            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                            ) cx WHERE cx.`币种` = '香港'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                        ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.月份 != '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df16 = pd.read_sql_query(sql=sql16, con=self.engine1)
        listT.append(df16)
        # 产品分旬 香港
        print('正在获取---产品分旬 香港…………')
        sql17 = '''SELECT *
                    FROM(SELECT 
						IFNULL(s1.家族, '合计') 家族,
						IFNULL(s1.地区, '合计') 地区,
						IFNULL(s1.月份, '合计') 月份,
						IFNULL(s1.旬, '合计') 旬,
						IFNULL(s1.产品id, '合计') 产品id,
						IFNULL(s1.产品名称, '合计') 产品名称,
						IFNULL(s1.父级分类, '合计') 父级分类,
						IFNULL(s1.二级分类, '合计') 二级分类,
					SUM(s1.已签收) as 已签收,
						SUM(s1.拒收) as 拒收,
						SUM(s1.已退货) as 已退货,
						SUM(s1.已完成) as 已完成,
						SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.香港圆通已签收) as '香港-圆通已签收',
						SUM(s1.香港圆通拒收) as '香港-圆通拒收',
						SUM(s1.香港圆通已退货) as '香港-圆通拒收已退货',
						SUM(s1.香港圆通已完成) as '香港-圆通拒收已完成',
						SUM(s1.香港圆通总订单) as '香港-圆通拒收总订单',
						concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收完成签收',
						concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收总计签收',
						concat(ROUND(SUM(s1.香港圆通已完成) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收完成占比',
						concat(ROUND(SUM(s1.香港圆通已退货) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收退货率',
						concat(ROUND(SUM(s1.香港圆通拒收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收拒收率',
					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
					SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
						SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
						SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
						SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
						SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
						concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
						concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
						concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
						concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
						concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "香港-圆通" ,1,0)) AS 香港圆通总订单,
								SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "已签收",1,0)) as 香港圆通已签收,
								SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "拒收",1,0)) as 香港圆通拒收,
								SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "已退货",1,0)) as 香港圆通已退货,
								SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 香港圆通已完成,
							SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
							SUM(IF(cx.物流方式 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
								SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
								SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
								SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
								SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
				        FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                        ) cx WHERE cx.`币种` = '香港'
                        GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                    ) s1
                    GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                    WITH ROLLUP 
            ) s HAVING s.旬 <> '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df17 = pd.read_sql_query(sql=sql17, con=self.engine1)
        listT.append(df17)

        # 产品整月_直发 台湾
        print('正在获取---产品整月_直发 台湾…………')
        sql18 = '''SELECT *
                        FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                                    IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                                    SUM(s1.已签收) as 已签收,
        						    SUM(s1.拒收) as 拒收,
        						    SUM(s1.已退货) as 已退货,
        						    SUM(s1.已完成) as 已完成,
        						    SUM(s1.总订单) as 总订单,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						            concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        						SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        						SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						        SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						            SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						            SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						            SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						            SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						            concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						            concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						            concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        						SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        						SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        						SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        						SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						    SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						    SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						    SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						    SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						    concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						    concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						    concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        						SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        				    		concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        					    	concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        					    	concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        					    	concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
                                SUM(s1.铱熙无敌已签收) as '铱熙无敌已签收',
                                    SUM(s1.铱熙无敌拒收) as '铱熙无敌拒收',
                                    SUM(s1.铱熙无敌已退货) as '铱熙无敌已退货',
                                    SUM(s1.铱熙无敌已完成) as '铱熙无敌已完成',
                                    SUM(s1.铱熙无敌总订单) as '铱熙无敌总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌已完成) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌已退货) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌拒收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌拒收率',
        						SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        					    	SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        					    	SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        					    	SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        				    		concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        				    		concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        						SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        					    	SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        					    	SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        					    	SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        					    	SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        					    	concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        						SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        					    	SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        					    	SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        					        SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        					    	concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        					    	concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        						SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        					    	SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        					    	SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        					    	SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        					    	SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        					    	concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        					    	concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        						SUM(s1.龟山改派已签收) as '龟山改派已签收',
        					    	SUM(s1.龟山改派拒收) as '龟山改派拒收',
        					    	SUM(s1.龟山改派已退货) as '龟山改派已退货',
        					    	SUM(s1.龟山改派已完成) as '龟山改派已完成',
        					    	SUM(s1.龟山改派总订单) as '龟山改派总订单',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        					    	concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        					    	concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        					    	concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        				    	SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        					    	SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        					    	SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        					    	SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        					    	SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        					    	concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        					    	concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        						SUM(s1.速派改派已签收) as '速派改派已签收',
        					    	SUM(s1.速派改派拒收) as '速派改派拒收',
        					    	SUM(s1.速派改派已退货) as '速派改派已退货',
        					    	SUM(s1.速派改派已完成) as '速派改派已完成',
        					    	SUM(s1.速派改派总订单) as '速派改派总订单',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        					    	concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        					    	concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        					    	concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
                                SUM(s1.速派宅配通已签收) as '速派宅配通改派已签收',
                                    SUM(s1.速派宅配通拒收) as '速派宅配通改派拒收',
                                    SUM(s1.速派宅配通已退货) as '速派宅配通改派已退货',
                                    SUM(s1.速派宅配通已完成) as '速派宅配通改派已完成',
                                    SUM(s1.速派宅配通总订单) as '速派宅配通改派总订单',
                                    concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派完成签收',
                                    concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派总计签收',
                                    concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派完成占比',
                                    concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派退货率',
                                    concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派拒收率',
        						SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        					    	SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        					    	SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        					    	SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        					    	SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        					    	concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        						SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        					    	SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        					    	SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        					    	SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        					    	SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
                                SUM(s1.铱熙无敌改派已签收) as '铱熙无敌改派已签收',
                                    SUM(s1.铱熙无敌改派拒收) as '铱熙无敌改派拒收',
                                    SUM(s1.铱熙无敌改派已退货) as '铱熙无敌改派已退货',
                                    SUM(s1.铱熙无敌改派已完成) as '铱熙无敌改派已完成',
                                    SUM(s1.铱熙无敌改派总订单) as '铱熙无敌改派总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌改派已完成) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌改派已退货) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌改派拒收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率'
                            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
                                    SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌总订单,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌已签收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌拒收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌已退货,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
                                    SUM(IF(cx.物流方式 = "速派宅配通" ,1,0)) AS 速派宅配通总订单,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
                                    SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌改派总订单,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌改派已签收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌改派拒收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌改派已退货,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌改派已完成
        				            FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`是否改派` = '直发' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                    ) cx WHERE cx.`币种` = '台湾'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                        ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df18 = pd.read_sql_query(sql=sql18, con=self.engine1)
        listT.append(df18)
        # 产品分旬_直发 台湾
        print('正在获取---产品分旬_直发 台湾…………')
        sql19 = '''SELECT *
                            FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,
        						SUM(s1.拒收) as 拒收,
        						SUM(s1.已退货) as 已退货,
        						SUM(s1.已完成) as 已完成,
        						SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						        concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        					SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        					SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						    SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						        SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						        SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						        SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						        SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						        concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						        concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						        concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        					SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        					SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        					SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        					SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        					SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        						concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        						concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        						concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
                            SUM(s1.铱熙无敌已签收) as '铱熙无敌已签收',
                                SUM(s1.铱熙无敌拒收) as '铱熙无敌拒收',
                                SUM(s1.铱熙无敌已退货) as '铱熙无敌已退货',
                                SUM(s1.铱熙无敌已完成) as '铱熙无敌已完成',
                                SUM(s1.铱熙无敌总订单) as '铱熙无敌总订单',
                                concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌已完成) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌已退货) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌退货率',
                                concat(ROUND(SUM(s1.铱熙无敌拒收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌拒收率',
        					SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        						SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        						SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        						SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        						concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        						concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        						concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        					SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        						SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        						SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        						SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        						SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        						concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        						concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        						concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        					SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        						SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        						SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        						SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        						concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        						concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        						concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        					SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        						SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        						SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        						SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        						SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        						concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        						concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        						concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        					SUM(s1.龟山改派已签收) as '龟山改派已签收',
        						SUM(s1.龟山改派拒收) as '龟山改派拒收',
        						SUM(s1.龟山改派已退货) as '龟山改派已退货',
        						SUM(s1.龟山改派已完成) as '龟山改派已完成',
        						SUM(s1.龟山改派总订单) as '龟山改派总订单',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        						concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        						concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        						concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        					SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        						SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        						SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        						SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        						SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        						concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        						concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        						concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        					SUM(s1.速派改派已签收) as '速派改派已签收',
        						SUM(s1.速派改派拒收) as '速派改派拒收',
        						SUM(s1.速派改派已退货) as '速派改派已退货',
        						SUM(s1.速派改派已完成) as '速派改派已完成',
        						SUM(s1.速派改派总订单) as '速派改派总订单',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        						concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        						concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        						concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
                            SUM(s1.速派宅配通已签收) as '速派宅配通改派已签收',
                                SUM(s1.速派宅配通拒收) as '速派宅配通改派拒收',
                                SUM(s1.速派宅配通已退货) as '速派宅配通改派已退货',
                                SUM(s1.速派宅配通已完成) as '速派宅配通改派已完成',
                                SUM(s1.速派宅配通总订单) as '速派宅配通改派总订单',
                                concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派完成签收',
                                concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派总计签收',
                                concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派完成占比',
                                concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派退货率',
                                concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派拒收率',
        					SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        						SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        						SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        						SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        						SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        						concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        						concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        						concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        					SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        						SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        						SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        						SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        						SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        						concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        						concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
                            SUM(s1.铱熙无敌改派已签收) as '铱熙无敌改派已签收',
                                SUM(s1.铱熙无敌改派拒收) as '铱熙无敌改派拒收',
                                SUM(s1.铱熙无敌改派已退货) as '铱熙无敌改派已退货',
                                SUM(s1.铱熙无敌改派已完成) as '铱熙无敌改派已完成',
                                SUM(s1.铱熙无敌改派总订单) as '铱熙无敌改派总订单',
                                concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌改派已完成) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌改派已退货) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
                                concat(ROUND(SUM(s1.铱熙无敌改派拒收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率'
                        FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
                                    SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌总订单,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌已签收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌拒收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌已退货,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
                                    SUM(IF(cx.物流方式 = "速派宅配通" ,1,0)) AS 速派宅配通总订单,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
                                    SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌改派总订单,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌改派已签收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌改派拒收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌改派已退货,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌改派已完成
        				        FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                    FROM {0}_zqsb cc where  cc.`是否改派` = '直发' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.旬 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df19 = pd.read_sql_query(sql=sql19, con=self.engine1)
        listT.append(df19)

        # 产品整月_改派 台湾
        print('正在获取---产品整月_直发 台湾…………')
        sql20 = '''SELECT *
                        FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                                    IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                                    SUM(s1.已签收) as 已签收,
        						    SUM(s1.拒收) as 拒收,
        						    SUM(s1.已退货) as 已退货,
        						    SUM(s1.已完成) as 已完成,
        						    SUM(s1.总订单) as 总订单,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						            concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        						SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        						SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						        SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						            SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						            SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						            SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						            SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						            concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						            concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						            concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        						SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        						SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
                                SUM(s1.铱熙无敌已签收) as '铱熙无敌已签收',
                                    SUM(s1.铱熙无敌拒收) as '铱熙无敌拒收',
                                    SUM(s1.铱熙无敌已退货) as '铱熙无敌已退货',
                                    SUM(s1.铱熙无敌已完成) as '铱熙无敌已完成',
                                    SUM(s1.铱熙无敌总订单) as '铱熙无敌总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌已完成) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌已退货) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌拒收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌拒收率',
        						SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        						SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						    SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						    SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						    SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						    SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						    concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						    concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						    concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        						SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        				    		concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        					    	concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        					    	concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        					    	concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
        						SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        					    	SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        					    	SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        					    	SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        				    		concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        				    		concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        						SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        					    	SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        					    	SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        					    	SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        					    	SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        					    	concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        						SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        					    	SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        					    	SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        					        SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        					    	concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        					    	concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        						SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        					    	SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        					    	SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        					    	SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        					    	SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        					    	concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        					    	concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        						SUM(s1.龟山改派已签收) as '龟山改派已签收',
        					    	SUM(s1.龟山改派拒收) as '龟山改派拒收',
        					    	SUM(s1.龟山改派已退货) as '龟山改派已退货',
        					    	SUM(s1.龟山改派已完成) as '龟山改派已完成',
        					    	SUM(s1.龟山改派总订单) as '龟山改派总订单',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        					    	concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        					    	concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        					    	concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        				    	SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        					    	SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        					    	SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        					    	SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        					    	SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        					    	concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        					    	concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        						SUM(s1.速派改派已签收) as '速派改派已签收',
        					    	SUM(s1.速派改派拒收) as '速派改派拒收',
        					    	SUM(s1.速派改派已退货) as '速派改派已退货',
        					    	SUM(s1.速派改派已完成) as '速派改派已完成',
        					    	SUM(s1.速派改派总订单) as '速派改派总订单',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        					    	concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        					    	concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        					    	concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
                                SUM(s1.速派宅配通已签收) as '速派宅配通改派已签收',
                                    SUM(s1.速派宅配通拒收) as '速派宅配通改派拒收',
                                    SUM(s1.速派宅配通已退货) as '速派宅配通改派已退货',
                                    SUM(s1.速派宅配通已完成) as '速派宅配通改派已完成',
                                    SUM(s1.速派宅配通总订单) as '速派宅配通改派总订单',
                                    concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派完成签收',
                                    concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派总计签收',
                                    concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派完成占比',
                                    concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派退货率',
                                    concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派拒收率',
        						SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        					    	SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        					    	SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        					    	SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        					    	SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        					    	concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        						SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        					    	SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        					    	SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        					    	SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        					    	SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
                                SUM(s1.铱熙无敌改派已签收) as '铱熙无敌改派已签收',
                                    SUM(s1.铱熙无敌改派拒收) as '铱熙无敌改派拒收',
                                    SUM(s1.铱熙无敌改派已退货) as '铱熙无敌改派已退货',
                                    SUM(s1.铱熙无敌改派已完成) as '铱熙无敌改派已完成',
                                    SUM(s1.铱熙无敌改派总订单) as '铱熙无敌改派总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌改派已完成) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌改派已退货) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌改派拒收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率'
                            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
                                    SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌总订单,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌已签收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌拒收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌已退货,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
                                    SUM(IF(cx.物流方式 = "速派宅配通" ,1,0)) AS 速派宅配通总订单,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
                                    SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌改派总订单,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌改派已签收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌改派拒收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌改派已退货,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌改派已完成
        				            FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`是否改派` = '改派' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                    ) cx WHERE cx.`币种` = '台湾'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                        ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙运营1组','Line运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df20 = pd.read_sql_query(sql=sql20, con=self.engine1)
        listT.append(df20)
        # 产品分旬_改派 台湾
        print('正在获取---产品分旬_直发 台湾…………')
        sql21 = '''SELECT *
                            FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,
        						SUM(s1.拒收) as 拒收,
        						SUM(s1.已退货) as 已退货,
        						SUM(s1.已完成) as 已完成,
        						SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						        concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        					SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        					SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						    SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						        SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						        SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						        SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						        SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						        concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						        concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						        concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        					SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        					SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        					SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        					SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        					SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        						concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        						concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        						concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
        					SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        						SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        						SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        						SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        						concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        						concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        						concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        					SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        						SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        						SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        						SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        						SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        						concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        						concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        						concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        					SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        						SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        						SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        						SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        						concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        						concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        						concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        					SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        						SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        						SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        						SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        						SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        						concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        						concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        						concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
                            SUM(s1.铱熙无敌已签收) as '铱熙无敌已签收',
                                SUM(s1.铱熙无敌拒收) as '铱熙无敌拒收',
                                SUM(s1.铱熙无敌已退货) as '铱熙无敌已退货',
                                SUM(s1.铱熙无敌已完成) as '铱熙无敌已完成',
                                SUM(s1.铱熙无敌总订单) as '铱熙无敌总订单',
                                concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌已签收) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌已完成) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌已退货) / SUM(s1.铱熙无敌总订单) * 100,2),'%') as '铱熙无敌退货率',
                                concat(ROUND(SUM(s1.铱熙无敌拒收) / SUM(s1.铱熙无敌已完成) * 100,2),'%') as '铱熙无敌拒收率',
        					SUM(s1.龟山改派已签收) as '龟山改派已签收',
        						SUM(s1.龟山改派拒收) as '龟山改派拒收',
        						SUM(s1.龟山改派已退货) as '龟山改派已退货',
        						SUM(s1.龟山改派已完成) as '龟山改派已完成',
        						SUM(s1.龟山改派总订单) as '龟山改派总订单',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        						concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        						concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        						concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        					SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        						SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        						SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        						SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        						SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        						concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        						concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        						concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        					SUM(s1.速派改派已签收) as '速派改派已签收',
        						SUM(s1.速派改派拒收) as '速派改派拒收',
        						SUM(s1.速派改派已退货) as '速派改派已退货',
        						SUM(s1.速派改派已完成) as '速派改派已完成',
        						SUM(s1.速派改派总订单) as '速派改派总订单',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        						concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        						concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        						concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
                            SUM(s1.速派宅配通已签收) as '速派宅配通改派已签收',
                                SUM(s1.速派宅配通拒收) as '速派宅配通改派拒收',
                                SUM(s1.速派宅配通已退货) as '速派宅配通改派已退货',
                                SUM(s1.速派宅配通已完成) as '速派宅配通改派已完成',
                                SUM(s1.速派宅配通总订单) as '速派宅配通改派总订单',
                                concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派完成签收',
                                concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派总计签收',
                                concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派完成占比',
                                concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '速派宅配通改派退货率',
                                concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '速派宅配通改派拒收率',
        					SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        						SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        						SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        						SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        						SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        						concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        						concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        						concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        					SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        						SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        						SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        						SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        						SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        						concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        						concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
                            SUM(s1.铱熙无敌改派已签收) as '铱熙无敌改派已签收',
                                SUM(s1.铱熙无敌改派拒收) as '铱熙无敌改派拒收',
                                SUM(s1.铱熙无敌改派已退货) as '铱熙无敌改派已退货',
                                SUM(s1.铱熙无敌改派已完成) as '铱熙无敌改派已完成',
                                SUM(s1.铱熙无敌改派总订单) as '铱熙无敌改派总订单',
                                concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌改派已签收) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌改派已完成) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌改派已退货) / SUM(s1.铱熙无敌改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
                                concat(ROUND(SUM(s1.铱熙无敌改派拒收) / SUM(s1.铱熙无敌改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率'
                        FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
                                    SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌总订单,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌已签收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌拒收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌已退货,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
                                    SUM(IF(cx.物流方式 = "速派宅配通" ,1,0)) AS 速派宅配通总订单,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
                                        SUM(IF(cx.物流方式 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
                                    SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌改派总订单,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌改派已签收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌改派拒收,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌改派已退货,
                                        SUM(IF(cx.物流方式 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌改派已完成
        				        FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                    FROM {0}_zqsb cc where  cc.`是否改派` = '改派' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.旬 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df21 = pd.read_sql_query(sql=sql21, con=self.engine1)
        listT.append(df21)

        # 产品整月-直发 香港
        print('正在获取---产品整月_直发 香港…………')
        sql31 = '''SELECT *
                            FROM(SELECT IFNULL(s1.家族, '合计') 家族,
                                        IFNULL(s1.地区, '合计') 地区,
                                        IFNULL(s1.月份, '合计') 月份,
                                IFNULL(s1.产品id, '合计') 产品id,
                                IFNULL(s1.产品名称, '合计') 产品名称,
                                IFNULL(s1.父级分类, '合计') 父级分类,
                                IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,
        						SUM(s1.拒收) as 拒收,
        						SUM(s1.已退货) as 已退货,
        						SUM(s1.已完成) as 已完成,
        						SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
        					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
        						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                            SUM(s1.香港圆通已签收) as '香港-圆通已签收',
                                SUM(s1.香港圆通拒收) as '香港-圆通拒收',
                                SUM(s1.香港圆通已退货) as '香港-圆通拒收已退货',
                                SUM(s1.香港圆通已完成) as '香港-圆通拒收已完成',
                                SUM(s1.香港圆通总订单) as '香港-圆通拒收总订单',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收完成签收',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收总计签收',
                                concat(ROUND(SUM(s1.香港圆通已完成) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收完成占比',
                                concat(ROUND(SUM(s1.香港圆通已退货) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收退货率',
                                concat(ROUND(SUM(s1.香港圆通拒收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收拒收率',
        					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
        						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
        						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
        						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
        						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
        						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
        						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
        						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
        					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
        						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
        						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
        						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
        						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
        						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
        						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
        						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
        					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
        						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
        						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
        						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
        						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
        						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
        					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
        						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
        						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
        						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
        						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
        						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
        					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
        						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
        						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
        						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
        						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
        						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
        						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
        						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
                            SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
                                SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
                                SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
                                SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
                                SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
                                concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
                                concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
                                concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
        					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
        						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
        						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
        						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
        						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
        						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
        						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
        						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
        		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
        							SUM(cx.`价格RMB`) as 总订单金额,
        								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
        								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
        								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
                                    SUM(IF(cx.物流方式 = "香港-圆通" ,1,0)) AS 香港圆通总订单,
                                        SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "已签收",1,0)) as 香港圆通已签收,
                                        SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "拒收",1,0)) as 香港圆通拒收,
                                        SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "已退货",1,0)) as 香港圆通已退货,
                                        SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 香港圆通已完成,
        							SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
        							SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
                                    SUM(IF(cx.物流方式 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
        							SUM(IF(cx.物流方式 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
        				            FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.`是否改派` = '直发' AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                    ) cx WHERE cx.`币种` = '香港'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                            ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df31 = pd.read_sql_query(sql=sql31, con=self.engine1)
        listT.append(df31)
        # 产品分旬-直发 香港
        print('正在获取---产品分旬_直发 香港…………')
        sql32 = '''SELECT *
                            FROM(SELECT 
        						IFNULL(s1.家族, '合计') 家族,
        						IFNULL(s1.地区, '合计') 地区,
        						IFNULL(s1.月份, '合计') 月份,
        						IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,
        						IFNULL(s1.产品名称, '合计') 产品名称,
        						IFNULL(s1.父级分类, '合计') 父级分类,
        						IFNULL(s1.二级分类, '合计') 二级分类,
        					SUM(s1.已签收) as 已签收,
        						SUM(s1.拒收) as 拒收,
        						SUM(s1.已退货) as 已退货,
        						SUM(s1.已完成) as 已完成,
        						SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
        					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
        						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                            SUM(s1.香港圆通已签收) as '香港-圆通已签收',
                                SUM(s1.香港圆通拒收) as '香港-圆通拒收',
                                SUM(s1.香港圆通已退货) as '香港-圆通拒收已退货',
                                SUM(s1.香港圆通已完成) as '香港-圆通拒收已完成',
                                SUM(s1.香港圆通总订单) as '香港-圆通拒收总订单',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收完成签收',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收总计签收',
                                concat(ROUND(SUM(s1.香港圆通已完成) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收完成占比',
                                concat(ROUND(SUM(s1.香港圆通已退货) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收退货率',
                                concat(ROUND(SUM(s1.香港圆通拒收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收拒收率',
        					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
        						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
        						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
        						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
        						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
        						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
        						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
        						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
        					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
        						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
        						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
        						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
        						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
        						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
        						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
        						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
        					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
        						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
        						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
        						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
        						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
        						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
        					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
        						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
        						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
        						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
        						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
        						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
        					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
        						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
        						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
        						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
        						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
        						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
        						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
        						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
                            SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
                                SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
                                SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
                                SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
                                SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
                                concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
                                concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
                                concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
        					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
        						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
        						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
        						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
        						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
        						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
        						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
        						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
        		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
        							SUM(cx.`价格RMB`) as 总订单金额,
        								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
        								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
        								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
                                    SUM(IF(cx.物流方式 = "香港-圆通" ,1,0)) AS 香港圆通总订单,
                                        SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "已签收",1,0)) as 香港圆通已签收,
                                        SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "拒收",1,0)) as 香港圆通拒收,
                                        SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 = "已退货",1,0)) as 香港圆通已退货,
                                        SUM(IF(cx.物流方式 = "香港-圆通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 香港圆通已完成,
        							SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
        							SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
                                    SUM(IF(cx.物流方式 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
        							SUM(IF(cx.物流方式 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
        				        FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.`是否改派` = '直发' AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                ) cx WHERE cx.`币种` = '香港'
                                GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                            GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                            WITH ROLLUP 
                    ) s HAVING s.旬 <> '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df32 = pd.read_sql_query(sql=sql32, con=self.engine1)
        listT.append(df32)

        # 产品整月-改派 香港
        print('正在获取---产品整月_改派 香港…………')
        sql41 = '''SELECT *
                            FROM(SELECT IFNULL(s1.家族, '合计') 家族,
                                        IFNULL(s1.地区, '合计') 地区,
                                        IFNULL(s1.月份, '合计') 月份,
                                IFNULL(s1.产品id, '合计') 产品id,
                                IFNULL(s1.产品名称, '合计') 产品名称,
                                IFNULL(s1.父级分类, '合计') 父级分类,
                                IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,
        						SUM(s1.拒收) as 拒收,
        						SUM(s1.已退货) as 已退货,
        						SUM(s1.已完成) as 已完成,
        						SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
        					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
        						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
        						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
        						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
        						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
        						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
        						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
        						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
        						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
        					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
        						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
        						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
        						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
        						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
        						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
        						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
        						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
        					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
        						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
        						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
        						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
        						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
        						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
        					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
        						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
        						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
        						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
        						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
        						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
        					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
        						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
        						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
        						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
        						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
        						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
        						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
        						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
                            SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
                                SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
                                SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
                                SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
                                SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
                                concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
                                concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
                                concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
        					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
        						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
        						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
        						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
        						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
        						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
        						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
        						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
        		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
        							SUM(cx.`价格RMB`) as 总订单金额,
        								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
        								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
        								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
        							SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
                                    SUM(IF(cx.物流方式 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
        							SUM(IF(cx.物流方式 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
        				            FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.`是否改派` = '改派' AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                    ) cx WHERE cx.`币种` = '香港'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                            ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df41 = pd.read_sql_query(sql=sql41, con=self.engine1)
        listT.append(df41)
        # 产品分旬-改派 香港
        print('正在获取---产品分旬_改派 香港…………')
        sql42 = '''SELECT *
                            FROM(SELECT 
        						IFNULL(s1.家族, '合计') 家族,
        						IFNULL(s1.地区, '合计') 地区,
        						IFNULL(s1.月份, '合计') 月份,
        						IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,
        						IFNULL(s1.产品名称, '合计') 产品名称,
        						IFNULL(s1.父级分类, '合计') 父级分类,
        						IFNULL(s1.二级分类, '合计') 二级分类,
        					SUM(s1.已签收) as 已签收,
        						SUM(s1.拒收) as 拒收,
        						SUM(s1.已退货) as 已退货,
        						SUM(s1.已完成) as 已完成,
        						SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
        					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
        						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
        						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
        						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
        						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
        						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
        						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
        						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
        						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
        					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
        						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
        						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
        						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
        						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
        						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
        						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
        						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
        					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
        						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
        						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
        						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
        						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
        						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
        					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
        						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
        						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
        						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
        						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
        						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
        					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
        						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
        						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
        						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
        						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
        						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
        						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
        						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
                            SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
                                SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
                                SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
                                SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
                                SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
                                concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
                                concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
                                concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
        					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
        						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
        						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
        						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
        						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
        						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
        						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
        						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
        		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
        							SUM(cx.`价格RMB`) as 总订单金额,
        								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
        								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
        								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
        								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
        							SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
        								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
        							SUM(IF(cx.物流方式 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
        								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
                                    SUM(IF(cx.物流方式 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
                                        SUM(IF(cx.物流方式 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
        							SUM(IF(cx.物流方式 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
        								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
        				        FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.`是否改派` = '改派' AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                ) cx WHERE cx.`币种` = '香港'
                                GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                            GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                            WITH ROLLUP 
                    ) s HAVING s.旬 <> '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df42 = pd.read_sql_query(sql=sql42, con=self.engine1)
        listT.append(df42)

        today = datetime.date.today().strftime('%Y.%m.%d')
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾', '产品分旬台湾', '产品整月香港', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾']
        print('正在将物流品类写入excel…………')
        file_path = 'G:\\输出文件\\{} {} 物流品类-签收率.xlsx'.format(today, match[team])
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        listT[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
        listT[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
        listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
        listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_总_品类_物流_两月签收率')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        new_path = 'F:\\神龙签收率\\' + (datetime.datetime.now()).strftime('%m.%d') + '\\签收率\\{} {} 物流品类-签收率.xlsx'.format(today, match[team])
        shutil.copyfile(file_path, new_path)        # copy到指定位置
        print('----已写入excel; 并复制到指定文件夹中')

        print('正在将品类分旬写入excel…………')
        file_path = 'G:\\输出文件\\{} {} 品类分旬-签收率.xlsx'.format(today, match[team])
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾', '产品整月香港', '产品分旬台湾', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
        listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_品类直发分旬签收率')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        new_path = 'F:\\神龙签收率\\' + (datetime.datetime.now()).strftime('%m.%d') + '\\签收率\\{} {} 品类分旬-签收率.xlsx'.format(today, match[team])
        shutil.copyfile(file_path, new_path)        # copy到指定位置
        print('----已写入excel; 并复制到指定文件夹中')

        print('正在将产品写入excel…………')
        file_path = 'G:\\输出文件\\{} {} 产品明细-签收率.xlsx'.format(today, match[team])
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾','产品分旬台湾',  '产品整月香港', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾', '产品月_直发香港', '产品旬_直发香港', '产品月_改派香港', '产品旬_改派香港']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        listT[4].to_excel(excel_writer=writer, sheet_name=sheet_name[4], index=False)       # 产品整月台湾
        listT[5].to_excel(excel_writer=writer, sheet_name=sheet_name[5], index=False)       # 产品分旬台湾
        listT[6].to_excel(excel_writer=writer, sheet_name=sheet_name[6], index=False)       # 产品整月香港
        listT[7].to_excel(excel_writer=writer, sheet_name=sheet_name[7], index=False)       # 产品分旬香港
        listT[8].to_excel(excel_writer=writer, sheet_name=sheet_name[8], index=False)       # 产品月_直发台湾
        listT[9].to_excel(excel_writer=writer, sheet_name=sheet_name[9], index=False)       # 产品旬_直发台湾
        listT[10].to_excel(excel_writer=writer, sheet_name=sheet_name[10], index=False)     # 产品月_改派台湾
        listT[11].to_excel(excel_writer=writer, sheet_name=sheet_name[11], index=False)     # 产品旬_改派台湾
        listT[12].to_excel(excel_writer=writer, sheet_name=sheet_name[12], index=False)     # 产品月_直发香港
        listT[13].to_excel(excel_writer=writer, sheet_name=sheet_name[13], index=False)     # 产品旬_直发香港
        listT[14].to_excel(excel_writer=writer, sheet_name=sheet_name[14], index=False)     # 产品月_改派香港
        listT[15].to_excel(excel_writer=writer, sheet_name=sheet_name[15], index=False)     # 产品旬_改派香港
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_产品签收率_总')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        new_path = 'F:\\神龙签收率\\' + (datetime.datetime.now()).strftime('%m.%d') + '\\签收率\\{} {} 产品明细-签收率.xlsx'.format(today, match[team])
        shutil.copyfile(file_path, new_path)        # copy到指定位置
        print('----已写入excel; 并复制到指定文件夹中')

        print("强制关闭Execl后台进程中......")
        system('taskkill /F /IM EXCEL.EXE')

    # 新版签收率-报表(刘姐看的)- 单量计算
    def qsb_new(self, team, month_last):  # 报表各团队近两个月的物流数据
        month_now = datetime.datetime.now().strftime('%Y-%m-%d')
        match = {'gat': '港台-每日'}
        not_team = '"红杉家族-港澳台", "红杉家族-港澳台2", "金狮-港澳台", "金鹏家族-小虎队","Line运营","神龙-主页运营"'
        # if team == 'ga9t':
        #     month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        #     month_now = datetime.datetime.now().strftime('%Y-%m-%d')
        # else:
        #     month_last = '2021-08-01'
        #     month_now = '2021-09-30'
        del_time = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y%m')
        sql = '''DELETE FROM gat_zqsb gt
                WHERE gt.年月 >= {0}
                  and gt.`订单编号` IN (SELECT 订单编号 
                                        FROM gat_order_list gs
                                        WHERE gs.年月 >= {0}
                                          and gs.`系统订单状态` NOT IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'));'''.format(del_time)
        print('正在清除港澳台-总表的可能删除了的订单…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        sql = '''DELETE FROM gat_zqsb gz 
                WHERE gz.`系统订单状态` = '已转采购' and gz.`是否改派` = '改派' and gz.`审核时间` >= '{0} 00:00:00' AND gz.`日期` >= '{1}';'''.format(month_now, month_last)
        print('正在清除不参与计算的今日改派订单…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        sql = '''UPDATE gat_zqsb d
                        SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-易速配-顺丰%','香港-易速配-顺丰', 
                                                IF(d.`物流方式` LIKE '台湾-天马-711%','台湾-天马-新竹', 
                                                IF(d.`物流方式` LIKE '%优美宇通%' or d.`物流方式` LIKE '%铱熙无敌%','台湾-铱熙无敌-新竹', d.`物流方式`)) )
                        WHERE d.`是否改派` ='直发';'''
        print('正在修改-直发的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''UPDATE gat_zqsb d
                        SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-森鸿%','香港-森鸿-改派',
                                        IF(d.`物流方式` LIKE '香港-立邦%','香港-立邦-改派',
            							IF(d.`物流方式` LIKE '香港-易速配%','香港-易速配-改派',
            							IF(d.`物流方式` LIKE '台湾-立邦普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-大黄蜂普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-森鸿-新竹%','森鸿',
            							IF(d.`物流方式` LIKE '台湾-天马-顺丰%','天马顺丰',
            							IF(d.`物流方式` LIKE '台湾-天马-新竹%' OR d.`物流方式` LIKE '台湾-天马-711%','天马新竹',
            							IF(d.`物流方式` LIKE '台湾-天马-黑猫%','天马黑猫',
            							IF(d.`物流方式` LIKE '台湾-易速配-龟山%' OR d.`物流方式` LIKE '台湾-易速配-新竹%' OR d.`物流方式` LIKE '新易速配-台湾-改派%' OR d.`物流方式` = '易速配','龟山',
            							IF(d.`物流方式` LIKE '台湾-速派-新竹%' OR d.`物流方式` LIKE '台湾-速派-711超商%','速派',
            							IF(d.`物流方式` LIKE '台湾-大黄蜂普货头程-易速配尾程%' OR d.`物流方式` LIKE '台湾-立邦普货头程-易速配尾程%','龟山', 
            							IF(d.`物流方式` LIKE '台湾-优美宇通-新竹改派%','台湾-铱熙无敌-新竹改派', 
            							IF(d.`物流方式` LIKE '香港-圆通%','香港-圆通-改派', IF(d.`物流方式` LIKE '台湾-速派宅配通%','速派宅配通', d.`物流方式`))))))  )  )  )  )  )  )  )
                        WHERE d.`是否改派` ='改派';'''
        print('正在修改-改派的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)


        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 0、每日-各团队
        print('正在获取---0、每日各团队…………')
        # sql0 = '''SELECT 月份,地区, 家族,
        #                     SUM(s.昨日订单量) as 有运单号,
        #                     SUM(s.直发签收) as 直发签收,
        #                     SUM(s.直发拒收) as 直发拒收,
        #                     SUM(s.直发完成) as 直发完成,
        #                     SUM(s.直发总订单) as 直发总订单,
        #                     concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发完成), 0) * 100,2),'%') as 直发完成签收,
        #                     concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发总订单), 0) * 100,2),'%') as 直发总计签收,
        #                     concat(ROUND(IFNULL(SUM(s.直发完成) / SUM(s.直发总订单), 0) * 100,2),'%')as 直发完成占比,
        #                     SUM(s.改派签收) as 改派签收,
        #                     SUM(s.改派拒收) as 改派拒收,
        #                     SUM(s.改派完成) as 改派完成,
        #                     SUM(s.改派总订单) as 改派总订单,
        #                     concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派完成), 0) * 100,2),'%') as 改派完成签收,
        #                     concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派总计签收,
        #                     concat(ROUND(IFNULL(SUM(s.改派完成) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派完成占比
        #             FROM( SELECT IFNULL(cx.`年月`, '总计') 月份,
        #                         IFNULL(cx.币种, '总计') 地区,
        #                         IFNULL(cx.家族, '总计') 家族,
        #                         SUM(IF(cx.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY),1,0)) as 昨日订单量,
        #                         SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) as 直发签收,
        #                         SUM(IF(`是否改派` = '直发' AND 最终状态 = "拒收",1,0)) as 直发拒收,
        #                         SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 直发完成,
        #                         SUM(IF(`是否改派` = '直发',1,0)) as 直发总订单,
        #                         SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) as 改派签收,
        #                         SUM(IF(`是否改派` = '改派' AND 最终状态 = "拒收",1,0)) as 改派拒收,
        #                         SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 改派完成,
        #                         SUM(IF(`是否改派` = '改派',1,0)) as 改派总订单
        #                     FROM (SELECT *,
        #                                 IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","金鹏",cc.团队)))))) as 家族
        #                             FROM gat_zqsb cc
        #                             where cc.日期 >= '{0}' and cc.`运单编号` is not null
        #                           ) cx
        #                     GROUP BY cx.年月,cx.币种,cx.家族
        #                     WITH ROLLUP
        #                 ) s
        #                 GROUP BY 月份,地区,家族
        #                 ORDER BY 月份 DESC,
        #                         FIELD( 地区, '台湾', '香港', '总计' ),
        #                         FIELD( 家族, '神龙', '火凤凰', '金狮', '金鹏','神龙运营1组', '红杉', '总计');'''.format(month_last, team)
        sql0 = '''SELECT *
                FROM (SELECT IFNULL(s.`年月`, '总计') 月份,
                            IFNULL(s.币种, '总计') 地区,
                            IFNULL(s.家族, '总计') 家族,  
                            SUM(昨日单量) 昨日单量,
                            SUM(s.直发签收) as 直发签收,
                            SUM(s.直发拒收) as 直发拒收,
                            SUM(s.直发完成) as 直发完成,
                            SUM(s.直发总订单) as 直发总订单,
                            concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发完成), 0) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发总订单), 0) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(IFNULL(SUM(s.直发完成) / SUM(s.直发总订单), 0) * 100,2),'%')as 直发完成占比,
                            SUM(s.改派签收) as 改派签收,
                            SUM(s.改派拒收) as 改派拒收,
                            SUM(s.改派完成) as 改派完成,
                            SUM(s.改派总订单) as 改派总订单,
                            concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派完成), 0) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(IFNULL(SUM(s.改派完成) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派完成占比
                    FROM( SELECT cx.`年月`,
                                cx.`币种`,
                                cx.`家族`,  
                                总订单量 昨日单量,
                                SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) as 直发签收,
                                SUM(IF(`是否改派` = '直发' AND 最终状态 = "拒收",1,0)) as 直发拒收,
                                SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 直发完成,
                                SUM(IF(`是否改派` = '直发',1,0)) as 直发总订单,
                                SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) as 改派签收,
                                SUM(IF(`是否改派` = '改派' AND 最终状态 = "拒收",1,0)) as 改派拒收,
                                SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 改派完成,
                                SUM(IF(`是否改派` = '改派',1,0)) as 改派总订单
                            FROM (SELECT *,IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                    FROM gat_zqsb cc
                                    where cc.日期 >= '{0}' and cc.`运单编号` is not null AND cc.团队 NOT IN ({2})
                            ) cx	
							LEFT JOIN 
							(SELECT 年月,币种,家族,count(订单编号) as 总订单量
								FROM (SELECT *,IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
										FROM gat_order_list cc 
										WHERE  cc.日期 = DATE_SUB(CURDATE(), INTERVAL 1 DAY) AND cc.团队 NOT IN ({2})
								) dg  
								GROUP BY dg.年月,dg.币种,dg.家族
							) cx2 
							ON  cx.年月 = cx2.年月 AND cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族   
                          GROUP BY cx.年月,cx.币种,cx.家族
                        ) s						
                    GROUP BY s.年月,s.币种,s.家族
					WITH ROLLUP 
					HAVING `地区` <> '总计'
				) ss					
                ORDER BY 月份 DESC,
                        FIELD( 地区, '台湾', '香港', '总计' ),
                        FIELD( 家族, '神龙', '火凤凰', '小虎队', '神龙运营1组', '红杉', '金狮', '总计'),
                        直发总订单 DESC;'''.format(month_last, team, not_team)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)

        # 01、各团队-审核率 删单率
        print('正在获取---01、各团队-审核率_删单率…………')
        sql01 = '''SELECT gs.币种,SUBSTRING(删除原因,2) as 删除原因, 
        			            COUNT(订单编号) as 单量, 
        			            SUM(IF(gs.`审单类型` = '是',1,0)) as 自动审单量,
        			            concat(ROUND(SUM(IF(gs.`审单类型` = '是',1,0)) / 总订单量 * 100,2),'%') as 自动审单量率,
        			            SUM(IF(gs.`审单类型` = '否' or gs.`审单类型` IS NULL,1,0)) as 人工审单量,
        			            concat(ROUND(SUM(IF(gs.`审单类型` = '否' or gs.`审单类型` IS NULL,1,0)) / 总订单量 * 100,2),'%') as 人工审单量率,
        			            SUM(IF(gs.`问题原因` IS NOT NULL,1,0)) as 问题订单量,
        			            SUM(IF(gs.`问题原因` IS NOT NULL AND gs.`系统订单状态` IN ("已删除","支付失败","未支付"),1,0)) as 问题订单删单量,
        			            concat(ROUND(SUM(IF(gs.`问题原因` IS NOT NULL AND gs.`系统订单状态` NOT IN ("已删除","支付失败","未支付"),1,0)) / SUM(IF(gs.`问题原因` IS NOT NULL,1,0))  * 100,2),'%') as 问题订单转化率,
        			            SUM(IF(gs.`系统订单状态` = '已删除',1,0)) as 删单量,
        			            concat(ROUND(SUM(IF(gs.`系统订单状态` = '已删除',1,0)) / 总订单量 * 100,2),'%') as 删单率
                            FROM  gat_order_list gs
                            LEFT JOIN (SELECT 币种, COUNT(订单编号)  as 总订单量
        					            FROM  gat_order_list gss
        					            WHERE gss.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY) AND gss.团队 NOT IN ({0})
        					            GROUP BY gss.`币种`
        					) gs2 ON gs.`币种` = gs2.`币种`
                            WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                            GROUP BY gs.`币种`,gs.`删除原因`
                            WITH ROLLUP
                            HAVING gs.`币种` IS NOT null
                            ORDER BY gs.币种,单量 DESC;'''.format(not_team)
        df01 = pd.read_sql_query(sql=sql01, con=self.engine1)
        listT.append(df01)

        # 1、各月-各团队
        print('正在获取---1、各月各团队…………')
        sql10 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(家族, '总计') 家族,		
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,		
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,										
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                        FROM(SELECT 年月 月份,币种 地区,家族,
                                    COUNT(订单编号) as 总单量,
									SUM(IF(最终状态 = "已签收",1,0)) 签收,
									SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
									SUM(IF(最终状态 = "已退货",1,0)) 退货,
									SUM(价格RMB) AS 金额,
									SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                                    SUM(IF(是否改派 = '直发',1,0))  as 直发单量,
									SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
									SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                                    SUM(IF(是否改派 = '改派',1,0))  as 改派单量,
									SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
									SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                            FROM gat_zqsb_cache cx
							WHERE cx.`运单编号` is not null AND cx.团队 NOT IN ({1})                     
                            GROUP BY cx.年月,cx.币种, cx.家族
                        ) s
						GROUP BY 月份,地区,家族
                        WITH ROLLUP		
                ) ss			
                ORDER BY 月份 DESC,
                        FIELD( 地区, '台湾', '香港', '总计' ),
                        FIELD( 家族, '神龙','火凤凰','小虎队','神龙运营1组','红杉','金狮', '总计' ),
                        总单量 DESC;'''.format(team, not_team)
        df10 = pd.read_sql_query(sql=sql10, con=self.engine1)
        listT.append(df10)
        # 2、各月各团队---分旬
        print('正在获取---2、各月各团队---分旬…………')
        sql11 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(旬, '总计') 旬,IFNULL(地区, '总计') 地区,IFNULL(家族, '总计') 家族,		
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,		
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,										
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                        FROM(SELECT 年月 月份, 旬,币种 地区,家族,
                                    COUNT(订单编号) as 总单量,
									SUM(IF(最终状态 = "已签收",1,0)) 签收,
									SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
									SUM(IF(最终状态 = "已退货",1,0)) 退货,
									SUM(价格RMB) AS 金额,
									SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                                    SUM(IF(是否改派 = '直发',1,0))  as 直发单量,
									SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
									SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                                    SUM(IF(是否改派 = '改派',1,0))  as 改派单量,
									SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
									SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                            FROM gat_zqsb_cache cx
							WHERE cx.`运单编号` is not null AND cx.团队 NOT IN ({1})               
                            GROUP BY cx.年月,cx.旬,cx.币种, cx.家族
                        ) s
						GROUP BY 月份,旬,地区,家族
                        WITH ROLLUP	
                ) ss				
                ORDER BY 月份 DESC,旬,
                        FIELD( 地区, '台湾', '香港', '总计' ),
                        FIELD( 家族, '神龙','火凤凰','小虎队','神龙运营1组','红杉','金狮', '总计' ),
                        总单量 DESC;'''.format(team, not_team)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)
        listT.append(df11)

        # 3、各团队-各品类
        print('正在获取---3、各团队-各品类…………')
        sql20 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(s.家族, '总计') 家族,IFNULL(父级分类, '总计') 父级分类,
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							concat(ROUND(SUM(总单量) / 总订单量 * 100,2),'%') as 品类占比,						
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,									
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
							concat(ROUND(SUM(直发单量) / 直发总单量 * 100,2),'%') as 直发品类占比,
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比,
							concat(ROUND(SUM(改派单量) / 改派总单量 * 100,2),'%') as 改派品类占比
                    FROM(SELECT cx.年月 月份, cx.币种 地区, cx.家族, cx.父级分类,
                            总订单量, 
                                COUNT(cx.订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,		
							直发总单量,							
                                SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,
							改派总单量,
                                SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM (SELECT 年月,币种,家族,父级分类,订单编号,最终状态,价格RMB,是否改派
							    FROM gat_zqsb_cache cc 
								WHERE  cc.`运单编号` is not null AND cc.团队 NOT IN ({2})
							) cx 
                        LEFT JOIN 
							( SELECT 币种,家族,年月,count(订单编号) as 总订单量,SUM(IF(`是否改派`= '直发',1,0)) as 直发总单量,SUM(IF(`是否改派` = '改派',1,0)) as 改派总单量
							    FROM (SELECT 年月,币种,家族,订单编号,是否改派
										FROM gat_zqsb_cache cc 
										WHERE  cc.`运单编号` is not null AND cc.团队 NOT IN ({2})
									) dg  
								GROUP BY dg.币种,dg.家族,dg.年月
							) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月   
						GROUP BY cx.年月,cx.币种,cx.家族,cx.父级分类
                    ) s				
					GROUP BY 月份,地区,s.家族,父级分类
                    WITH ROLLUP
            ) ss
			ORDER BY 月份 DESC,
                    FIELD( 地区, '台湾', '香港', '总计' ),
                    FIELD( 家族, '神龙','火凤凰','小虎队','神龙运营1组','红杉','金狮', '总计' ),
                    FIELD( 父级分类, '居家百货', '电子电器', '服饰', '医药保健',  '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','总计' ),
                    总单量 DESC;'''.format(month_last, team, not_team)
        df20 = pd.read_sql_query(sql=sql20, con=self.engine1)
        listT.append(df20)
        # 4、各团队-各物流
        print('正在获取---4、各团队-各物流…………')
        sql21 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(是否改派, '总计') 是否改派,IFNULL(家族, '总计') 家族,IFNULL(物流方式, '总计') 物流方式,
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,											
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                    FROM(SELECT 年月 月份, 币种 地区, 是否改派, 家族, 物流方式,
                                COUNT(订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                            SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                            SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM gat_zqsb_cache cx
						WHERE  cx.`运单编号` is not null  AND cx.团队 NOT IN ({2})       
						GROUP BY cx.年月,cx.币种,cx.是否改派,cx.家族,cx.物流方式
                    ) s
					GROUP BY 月份,地区,是否改派,家族,物流方式
                    WITH ROLLUP
            ) ss
            ORDER BY FIELD(月份, DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'), 
                                DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 4 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 5 MONTH),'%Y%m'), 
			                    DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 6 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 7 MONTH),'%Y%m'), '总计' ),
                            FIELD(地区, '台湾', '香港', '总计' ),
                            FIELD(是否改派, '直发', '改派', '总计' ),
                            FIELD(家族, '神龙','火凤凰','小虎队','神龙运营1组','红杉','金狮', '总计' ),
                            FIELD(物流方式, '台湾-大黄蜂普货头程-森鸿尾程','台湾-大黄蜂普货头程-易速配尾程', '台湾-立邦普货头程-森鸿尾程','台湾-易速配-TW海快','台湾-铱熙无敌-新竹','台湾-立邦普货头程-易速配尾程', 
                                            '台湾-森鸿-新竹-自发头程', '台湾-速派-711超商', '台湾-速派-新竹','台湾-天马-新竹','台湾-天马-顺丰','台湾-天马-黑猫','台湾-易速配-新竹',
                                            '香港-圆通', '香港-立邦-顺丰','香港-易速配-顺丰','香港-易速配-顺丰YC', '香港-森鸿-SH渠道','香港-森鸿-顺丰渠道',
                                            '龟山','森鸿','速派','速派宅配通','天马顺丰','天马新竹','香港-圆通-改派','香港-立邦-改派','香港-森鸿-改派','香港-易速配-改派','合计' ),
                            总单量 DESC;'''.format(month_last, team, not_team)
        df21 = pd.read_sql_query(sql=sql21, con=self.engine1)
        listT.append(df21)

        # 5、各团队-各平台
        print('正在获取---5、各团队-各平台…………')
        sql30 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(家族, '总计') 家族,IFNULL(平台, '总计') 平台,
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,											
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                    FROM(SELECT 年月 月份, 币种 地区, 家族,订单来源 平台,
                                COUNT(订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                            SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                            SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM gat_zqsb_cache cx
						WHERE  cx.`运单编号` is not null AND cx.团队 NOT IN ({2})                 
                        GROUP BY cx.年月,cx.币种,cx.家族,cx.订单来源
                        ) s
					GROUP BY 月份,地区,家族,平台
                    WITH ROLLUP
            ) ss
            ORDER BY 月份 DESC,
                    FIELD(地区, '台湾', '香港', '总计' ),
                    FIELD(家族, '神龙','火凤凰','小虎队','神龙运营1组','红杉','金狮', '总计' ),
                    FIELD(平台, 'google', 'facebook', 'line', 'native',  'Criteo', 'tiktok', 'yahoo','facebookpage','recommend','postsaleclone','recomm','shangwutong','总计' ),
                    总单量 DESC;'''.format(month_last, team, not_team)
        df30 = pd.read_sql_query(sql=sql30, con=self.engine1)
        listT.append(df30)
        # 6、各平台-各团队
        print('正在获取---6、各平台-各团队…………')
        sql31 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(平台, '总计') 平台,IFNULL(家族, '总计') 家族,									
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,											
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                    FROM(SELECT 年月 月份, 币种 地区, 订单来源 平台,家族,
                                COUNT(订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                            SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                            SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM gat_zqsb_cache cx
						WHERE cx.日期 >= '{0}' and cx.`运单编号` is not null AND cx.团队 NOT IN ({2})
                        GROUP BY cx.年月,cx.币种,cx.订单来源,cx.家族
                        ) s
					GROUP BY 月份,地区,平台,家族
                    WITH ROLLUP
            ) ss
            ORDER BY 月份 DESC,
                    FIELD(地区, '台湾', '香港', '总计' ),
                    FIELD(平台, 'google', 'facebook', 'line', 'native',  'Criteo', 'tiktok', 'yahoo','facebookpage','recommend','postsaleclone','recomm','shangwutong','总计' ),
                    FIELD(家族, '神龙','火凤凰','小虎队','神龙运营1组','红杉','金狮', '总计' ),
                    总单量 DESC;'''.format(month_last, team,not_team)
        df31 = pd.read_sql_query(sql=sql31, con=self.engine1)
        listT.append(df31)

        # 7、各品类-各团队
        print('正在获取---7、各品类-各团队…………')
        sql40 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(父级分类, '总计') 父级分类,IFNULL(家族, '总计') 家族,									
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,											
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                    FROM(SELECT 年月 月份, 币种 地区, 父级分类,家族,
                                COUNT(订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                            SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                            SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM gat_zqsb_cache cx
						WHERE cx.日期 >= '{0}' and cx.`运单编号` is not null  AND cx.团队 NOT IN ({2})               
                        GROUP BY cx.年月,cx.币种,cx.父级分类,cx.家族
                        ) s
					GROUP BY 月份,地区,父级分类,家族
                    WITH ROLLUP
            ) ss
            ORDER BY 月份 DESC,
                    FIELD(地区, '台湾', '香港', '总计' ),
                    FIELD(父级分类, '居家百货', '电子电器', '服饰', '医药保健', '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','总计' ),
                    FIELD(家族, '神龙','火凤凰','小虎队','神龙运营1组','红杉','金狮', '总计' ),
                    总单量 DESC;'''.format(month_last, team,not_team)
        df40 = pd.read_sql_query(sql=sql40, con=self.engine1)
        listT.append(df40)
        # 8、各物流-各团队
        print('正在获取---8、各物流-各团队…………')
        sql41 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(是否改派, '总计') 是否改派,IFNULL(物流方式, '总计') 物流方式,IFNULL(家族, '总计') 家族,									
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,											
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                    FROM(SELECT 年月 月份, 币种 地区, 是否改派, 物流方式, 家族,
                                COUNT(订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                            SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                            SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM gat_zqsb_cache cx
						WHERE cx.日期 >= '{0}' and cx.`运单编号` is not null  AND cx.团队 NOT IN ({2})                         
                        GROUP BY cx.年月,cx.币种,cx.是否改派,cx.物流方式,cx.家族
                        ) s
					GROUP BY 月份, 地区, 是否改派, 物流方式, 家族
                    WITH ROLLUP
            ) ss
            ORDER BY FIELD(月份, DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'), 
                                DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 4 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 5 MONTH),'%Y%m'), 
			                    DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 6 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 7 MONTH),'%Y%m'), '总计' ),
                    FIELD(地区, '台湾', '香港', '总计' ),
                    FIELD(是否改派, '直发', '改派', '总计' ),
                    FIELD(物流方式, '台湾-大黄蜂普货头程-森鸿尾程','台湾-大黄蜂普货头程-易速配尾程', '台湾-立邦普货头程-森鸿尾程','台湾-易速配-TW海快','台湾-铱熙无敌-新竹','台湾-立邦普货头程-易速配尾程', 
                                '台湾-森鸿-新竹-自发头程', '台湾-速派-711超商', '台湾-速派-新竹','台湾-天马-新竹','台湾-天马-顺丰','台湾-天马-黑猫','台湾-易速配-新竹',
                                '香港-圆通', '香港-立邦-顺丰','香港-易速配-顺丰','香港-易速配-顺丰YC', '香港-森鸿-SH渠道','香港-森鸿-顺丰渠道',
                                '龟山','森鸿','速派','速派宅配通','天马顺丰','天马新竹','香港-圆通-改派','香港-立邦-改派','香港-森鸿-改派','香港-易速配-改派','合计' ),
                    FIELD(家族, '神龙','火凤凰','小虎队','神龙运营1组','红杉','金狮', '总计' ),
                    总单量 DESC;'''.format(month_last, team, not_team)
        df41 = pd.read_sql_query(sql=sql41, con=self.engine1)
        listT.append(df41)

        # 9、同产品各团队的对比
        print('正在获取---9、同产品各团队的对比…………')
        sql50 = '''SELECT *, IF(神龙完成签收 = '0.00%' OR 神龙完成签收 IS NULL, 神龙完成签收, concat(ROUND(神龙完成签收-完成签收,2),'%')) as 神龙对比,
    			            IF(火凤凰完成签收 = '0.00%' OR 火凤凰完成签收 IS NULL, 火凤凰完成签收, concat(ROUND(火凤凰完成签收-完成签收,2),'%')) as 火凤凰对比,
    			            IF(神龙运营完成签收 = '0.00%' OR 神龙运营完成签收 IS NULL, 神龙运营完成签收, concat(ROUND(神龙运营完成签收-完成签收,2),'%')) as 神龙运营对比,
    			            IF(小虎队完成签收 = '0.00%' OR 小虎队完成签收 IS NULL, 小虎队完成签收, concat(ROUND(小虎队完成签收-完成签收,2),'%')) as 小虎队对比,
    			            IF(红杉完成签收 = '0.00%' OR 红杉完成签收 IS NULL,红杉完成签收, concat(ROUND(红杉完成签收-完成签收,2),'%')) as 红杉对比,
    			            IF(金狮完成签收 = '0.00%' OR 金狮完成签收 IS NULL, 金狮完成签收, concat(ROUND(金狮完成签收-完成签收,2),'%')) as 金狮对比
					FROM(SELECT	IFNULL(月份, '总计') 月份, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id, IFNULL(产品名称, '总计') 产品名称, IFNULL(父级分类, '总计') 父级分类,
							SUM(总单量) 总单量, SUM(签收) 签收, SUM(拒收) 拒收,
                                concat(ROUND(SUM(改派) / SUM(总单量) * 100,2),'%') as 改派占比,
                                concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                                concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                                concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,								
							SUM(神龙单量) 神龙单量, SUM(神龙签收) 神龙签收, SUM(神龙拒收) 神龙拒收,
                                concat(ROUND(SUM(神龙改派) / SUM(神龙单量) * 100,2),'%') as 神龙改派占比,
                                concat(ROUND(SUM(神龙签收) / SUM(神龙单量) * 100,2),'%') as 神龙总计签收,
                                concat(ROUND(SUM(神龙签收) / SUM(神龙完成) * 100,2),'%') as 神龙完成签收,
                                concat(ROUND(SUM(神龙完成) / SUM(神龙单量) * 100,2),'%') as 神龙完成占比,					
							SUM(火凤凰单量) 火凤凰单量, SUM(火凤凰签收) 火凤凰签收, SUM(火凤凰拒收) 火凤凰拒收,
                                concat(ROUND(SUM(火凤凰改派) / SUM(火凤凰单量) * 100,2),'%') as 火凤凰改派占比,
                                concat(ROUND(SUM(火凤凰签收) / SUM(火凤凰单量) * 100,2),'%') as 火凤凰总计签收,
                                concat(ROUND(SUM(火凤凰签收) / SUM(火凤凰完成) * 100,2),'%') as 火凤凰完成签收,
                                concat(ROUND(SUM(火凤凰完成) / SUM(火凤凰单量) * 100,2),'%') as 火凤凰完成占比,					
							SUM(神龙运营单量) 神龙运营单量, SUM(神龙运营签收) 神龙运营签收, SUM(神龙运营拒收) 神龙运营拒收,
                                concat(ROUND(SUM(神龙运营改派) / SUM(神龙运营单量) * 100,2),'%') as 神龙运营改派占比,
                                concat(ROUND(SUM(神龙运营签收) / SUM(神龙运营单量) * 100,2),'%') as 神龙运营总计签收,
                                concat(ROUND(SUM(神龙运营签收) / SUM(神龙运营完成) * 100,2),'%') as 神龙运营完成签收,
                                concat(ROUND(SUM(神龙运营完成) / SUM(神龙运营单量) * 100,2),'%') as 神龙运营完成占比,						
							SUM(小虎队单量) 小虎队单量, SUM(小虎队签收) 小虎队签收, SUM(小虎队拒收) 小虎队拒收,
                                concat(ROUND(SUM(小虎队改派) / SUM(小虎队单量) * 100,2),'%') as 小虎队改派占比,
                                concat(ROUND(SUM(小虎队签收) / SUM(小虎队单量) * 100,2),'%') as 小虎队总计签收,
                                concat(ROUND(SUM(小虎队签收) / SUM(小虎队完成) * 100,2),'%') as 小虎队完成签收,
                                concat(ROUND(SUM(小虎队完成) / SUM(小虎队单量) * 100,2),'%') as 小虎队完成占比,					
							SUM(红杉单量) 红杉单量, SUM(红杉签收) 红杉签收, SUM(红杉拒收) 红杉拒收,
                                concat(ROUND(SUM(红杉改派) / SUM(红杉单量) * 100,2),'%') as 红杉改派占比,
                                concat(ROUND(SUM(红杉签收) / SUM(红杉单量) * 100,2),'%') as 红杉总计签收,
                                concat(ROUND(SUM(红杉签收) / SUM(红杉完成) * 100,2),'%') as 红杉完成签收,
                                concat(ROUND(SUM(红杉完成) / SUM(红杉单量) * 100,2),'%') as 红杉完成占比,						
							SUM(金狮单量) 金狮单量, SUM(金狮签收) 金狮签收, SUM(金狮拒收) 金狮拒收,
                                concat(ROUND(SUM(金狮改派) / SUM(金狮单量) * 100,2),'%') as 金狮改派占比,
                                concat(ROUND(SUM(金狮签收) / SUM(金狮单量) * 100,2),'%') as 金狮总计签收,
                                concat(ROUND(SUM(金狮签收) / SUM(金狮完成) * 100,2),'%') as 金狮完成签收,
                                concat(ROUND(SUM(金狮完成) / SUM(金狮单量) * 100,2),'%') as 金狮完成占比									
                        FROM(SELECT 年月 月份,币种 地区, 产品id, 产品名称, 父级分类,
                                COUNT(订单编号) as 总单量,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 完成,
                                SUM(IF(是否改派 = '改派',1,0)) as 改派,
                            SUM(IF(家族 = '神龙',1,0)) as 神龙单量,
                                SUM(IF(家族 = '神龙' AND 最终状态 = "已签收",1,0)) as 神龙签收,
                                SUM(IF(家族 = '神龙' AND 最终状态 = "拒收",1,0)) as 神龙拒收,
                                SUM(IF(家族 = '神龙' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 神龙完成,
                                SUM(IF(家族 = '神龙' AND 是否改派 = '改派',1,0)) as 神龙改派,
                            SUM(IF(家族 = '火凤凰',1,0)) as 火凤凰单量,
                                SUM(IF(家族 = '火凤凰' AND 最终状态 = "已签收",1,0)) as 火凤凰签收,
                                SUM(IF(家族 = '火凤凰' AND 最终状态 = "拒收",1,0)) as 火凤凰拒收,
                                SUM(IF(家族 = '火凤凰' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 火凤凰完成,
                                SUM(IF(家族 = '火凤凰' AND 是否改派 = '改派',1,0)) as 火凤凰改派,
                            SUM(IF(家族 = '神龙-运营1组',1,0)) as 神龙运营单量,
                                SUM(IF(家族 = '神龙-运营1组' AND 最终状态 = "已签收",1,0)) as 神龙运营签收,
                                SUM(IF(家族 = '神龙-运营1组' AND 最终状态 = "拒收",1,0)) as 神龙运营拒收,
                                SUM(IF(家族 = '神龙-运营1组' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 神龙运营完成,
                                SUM(IF(家族 = '神龙-运营1组' AND 是否改派 = '改派',1,0)) as 神龙运营改派,
                            SUM(IF(家族 = '小虎队',1,0)) as 小虎队单量,
                                SUM(IF(家族 = '小虎队' AND 最终状态 = "已签收",1,0)) as 小虎队签收,
                                SUM(IF(家族 = '小虎队' AND 最终状态 = "拒收",1,0)) as 小虎队拒收,
                                SUM(IF(家族 = '小虎队' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 小虎队完成,
                                SUM(IF(家族 = '小虎队' AND 是否改派 = '改派',1,0)) as 小虎队改派,
                            SUM(IF(家族 = '红杉',1,0)) as 红杉单量,
                                SUM(IF(家族 = '红杉' AND 最终状态 = "已签收",1,0)) as 红杉签收,
                                SUM(IF(家族 = '红杉' AND 最终状态 = "拒收",1,0)) as 红杉拒收,
                                SUM(IF(家族 = '红杉' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 红杉完成,
                                SUM(IF(家族 = '红杉' AND 是否改派 = '改派',1,0)) as 红杉改派,
                            SUM(IF(家族 = '金狮',1,0)) as 金狮单量,
                                SUM(IF(家族 = '金狮' AND 最终状态 = "已签收",1,0)) as 金狮签收,
                                SUM(IF(家族 = '金狮' AND 最终状态 = "拒收",1,0)) as 金狮拒收,
                                SUM(IF(家族 = '金狮' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 金狮完成,
                                SUM(IF(家族 = '金狮' AND 是否改派 = '改派',1,0)) as 金狮改派
                            FROM gat_zqsb_cache cc
						    WHERE cc.日期 >= '{0}' and cc.`运单编号` is not null AND cc.团队 NOT IN ({2})
                            GROUP BY cc.年月,cc.币种,cc.产品id
					    ) s
						GROUP BY 月份,地区,产品id		
                        WITH ROLLUP 
					) ss
                   ORDER BY FIELD(月份,DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'),'总计'),
                            FIELD(地区,'台湾','香港','总计'),
                            总单量 DESC;'''.format(month_last, team, not_team)
        df50 = pd.read_sql_query(sql=sql50, con=self.engine1)
        listT.append(df50)

        # 10、同产品各月的对比
        print('正在获取---10、同产品各月的对比…………')
        sql51 = '''SELECT *
                FROM (SELECT IFNULL(家族, '总计') 家族, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id,  IFNULL(产品名称, '总计') 产品名称, IFNULL(父级分类, '总计') 父级分类,
						    SUM(总单量) 总单量,					
						SUM(03总量) 202108总单量,
							concat(ROUND(SUM(03签收量) / SUM(03总量) * 100,2),'%') as 202108总计签收,
							concat(ROUND(SUM(03签收量) / SUM(03完成量) * 100,2),'%') as 202108完成签收,
							concat(ROUND(SUM(03完成量) / SUM(03总量) * 100,2),'%') as 202108完成占比,						
						SUM(04总量) 202109总单量,
							concat(ROUND(SUM(04签收量) / SUM(04总量) * 100,2),'%') as 202109总计签收,
							concat(ROUND(SUM(04签收量) / SUM(04完成量) * 100,2),'%') as 202109完成签收,
							concat(ROUND(SUM(04完成量) / SUM(04总量) * 100,2),'%') as 202109完成占比,					
						SUM(05总量) 202110总单量,
							concat(ROUND(SUM(05签收量) / SUM(05总量) * 100,2),'%') as 202110总计签收,
							concat(ROUND(SUM(05签收量) / SUM(05完成量) * 100,2),'%') as 202110完成签收,
							concat(ROUND(SUM(05完成量) / SUM(05总量) * 100,2),'%') as 202110完成占比,						
						SUM(06总量) 202111总单量,
							concat(ROUND(SUM(06签收量) / SUM(06总量) * 100,2),'%') as 202111总计签收,
							concat(ROUND(SUM(06签收量) / SUM(06完成量) * 100,2),'%') as 202111完成签收,
							concat(ROUND(SUM(06完成量) / SUM(06总量) * 100,2),'%') as 202111完成占比,				
						SUM(07总量) 202112总单量,
							concat(ROUND(SUM(07签收量) / SUM(07总量) * 100,2),'%') as 202112总计签收,
							concat(ROUND(SUM(07签收量) / SUM(07完成量) * 100,2),'%') as 202112完成签收,
							concat(ROUND(SUM(07完成量) / SUM(07总量) * 100,2),'%') as 202112完成占比,
						SUM(08总量) 202201总单量,
							concat(ROUND(SUM(08签收量) / SUM(08总量) * 100,2),'%') as 202201总计签收,
							concat(ROUND(SUM(08签收量) / SUM(08完成量) * 100,2),'%') as 202201完成签收,
							concat(ROUND(SUM(08完成量) / SUM(08总量) * 100,2),'%') as 202201完成占比,					
						SUM(09总量) 202202总单量,
							concat(ROUND(SUM(09签收量) / SUM(09总量) * 100,2),'%') as 202202总计签收,
							concat(ROUND(SUM(09签收量) / SUM(09完成量) * 100,2),'%') as 202202完成签收,
							concat(ROUND(SUM(09完成量) / SUM(09总量) * 100,2),'%') as 202202完成占比,					
						SUM(10总量) 202203总单量,
							concat(ROUND(SUM(10签收量) / SUM(10总量) * 100,2),'%') as 202203总计签收,
							concat(ROUND(SUM(10签收量) / SUM(10完成量) * 100,2),'%') as 202203完成签收,
							concat(ROUND(SUM(10完成量) / SUM(10总量) * 100,2),'%') as 202203完成占比,					
						SUM(11总量) 202204总单量,
							concat(ROUND(SUM(11签收量) / SUM(11总量) * 100,2),'%') as 202204总计签收,
							concat(ROUND(SUM(11签收量) / SUM(11完成量) * 100,2),'%') as 202204完成签收,
							concat(ROUND(SUM(11完成量) / SUM(11总量) * 100,2),'%') as 202204完成占比,						
						SUM(12总量) 202205总单量,
							concat(ROUND(SUM(12签收量) / SUM(12总量) * 100,2),'%') as 202205总计签收,
							concat(ROUND(SUM(12签收量) / SUM(12完成量) * 100,2),'%') as 202205完成签收,
							concat(ROUND(SUM(12完成量) / SUM(12总量) * 100,2),'%') as 202205完成占比,					
						SUM(13总量) 202206总单量,
							concat(ROUND(SUM(12签收量) / SUM(12总量) * 100,2),'%') as 202206总计签收,
							concat(ROUND(SUM(12签收量) / SUM(12完成量) * 100,2),'%') as 202206完成签收,
							concat(ROUND(SUM(12完成量) / SUM(12总量) * 100,2),'%') as 202206完成占比,
						SUM(14总量) 202207总单量,
							concat(ROUND(SUM(14签收量) / SUM(14总量) * 100,2),'%') as 202207总计签收,
							concat(ROUND(SUM(14签收量) / SUM(14完成量) * 100,2),'%') as 202207完成签收,
							concat(ROUND(SUM(14完成量) / SUM(14总量) * 100,2),'%') as 202207完成占比,
						SUM(15总量) 202208总单量,
							concat(ROUND(SUM(15签收量) / SUM(15总量) * 100,2),'%') as 202208总计签收,
							concat(ROUND(SUM(15签收量) / SUM(15完成量) * 100,2),'%') as 202208完成签收,
							concat(ROUND(SUM(15完成量) / SUM(15总量) * 100,2),'%') as 202208完成占比
                    FROM(SELECT 家族,币种 地区, 产品id, 产品名称, 父级分类,
                                COUNT(cx.`订单编号`) as 总单量,				
                            SUM(IF(年月 = 202108,1,0)) as 03总量,
                                SUM(IF(年月 = 202108 AND 最终状态 = "已签收",1,0)) as 03签收量,
                                SUM(IF(年月 = 202108 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 03完成量,					
                            SUM(IF(年月 = 202109,1,0)) as 04总量,
                                SUM(IF(年月 = 202109 AND 最终状态 = "已签收",1,0)) as 04签收量,
                                SUM(IF(年月 = 202109 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 04完成量,					
                            SUM(IF(年月 = 202110,1,0)) as 05总量,
                                SUM(IF(年月 = 202110 AND 最终状态 = "已签收",1,0)) as 05签收量,
                                SUM(IF(年月 = 202110 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 05完成量,						
                            SUM(IF(年月 = 202111,1,0)) as 06总量,
                                SUM(IF(年月 = 202111 AND 最终状态 = "已签收",1,0)) as 06签收量,
                                SUM(IF(年月 = 202111 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 06完成量,						
                            SUM(IF(年月 = 202112,1,0)) as 07总量,
                                SUM(IF(年月 = 202112 AND 最终状态 = "已签收",1,0)) as 07签收量,
                                SUM(IF(年月 = 202112 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 07完成量,					
                            SUM(IF(年月 = 202201,1,0)) as 08总量,
                                SUM(IF(年月 = 202201 AND 最终状态 = "已签收",1,0)) as 08签收量,
                                SUM(IF(年月 = 202201 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 08完成量,					
                            SUM(IF(年月 = 202202,1,0)) as 09总量,
                                SUM(IF(年月 = 202202 AND 最终状态 = "已签收",1,0)) as 09签收量,
                                SUM(IF(年月 = 202202 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 09完成量,					
                            SUM(IF(年月 = 202203,1,0)) as 10总量,
                                SUM(IF(年月 = 202203 AND 最终状态 = "已签收",1,0)) as 10签收量,
                                SUM(IF(年月 = 202203 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 10完成量,						
                            SUM(IF(年月 = 202204,1,0)) as 11总量,
                                SUM(IF(年月 = 202204 AND 最终状态 = "已签收",1,0)) as 11签收量,
                                SUM(IF(年月 = 202204 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 11完成量,						
                            SUM(IF(年月 = 202205,1,0)) as 12总量,
                                SUM(IF(年月 = 202205 AND 最终状态 = "已签收",1,0)) as 12签收量,
                                SUM(IF(年月 = 202205 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 12完成量,
                            SUM(IF(年月 = 202206,1,0)) as 13总量,
                                SUM(IF(年月 = 202206 AND 最终状态 = "已签收",1,0)) as 13签收量,
                                SUM(IF(年月 = 202206 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 13完成量,
                            SUM(IF(年月 = 202207,1,0)) as 14总量,
                                SUM(IF(年月 = 202207 AND 最终状态 = "已签收",1,0)) as 14签收量,
                                SUM(IF(年月 = 202207 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 14完成量,
                            SUM(IF(年月 = 202208,1,0)) as 15总量,
                                SUM(IF(年月 = 202208 AND 最终状态 = "已签收",1,0)) as 15签收量,
                                SUM(IF(年月 = 202208 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 15完成量			
                        FROM gat_zqsb_cache cx
                        where cx.`运单编号` is not null AND cx.团队 NOT IN ({0})
                        GROUP BY cx.家族,cx.币种,cx.产品id
                    ) s
					GROUP BY s.家族,s.地区,s.产品id
					WITH ROLLUP 
				) ss
                ORDER BY FIELD(ss.`家族`,'神龙','火凤凰','小虎队', '神龙运营1组','红杉','金狮','总计'),
                        FIELD(ss.地区, '台湾', '香港', '总计' ),
                        ss.总单量 DESC;'''.format(not_team)
        df51 = pd.read_sql_query(sql=sql51, con=self.engine1)
        listT.append(df51)

        # 13、各团队-问题率
        # print('正在获取---3、各团队-问题率…………')
        sql02 = '''SELECT *
                FROM (
                    (SELECT 币种,'核实地址' AS 问题原因,核实地址 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实地址%",1,0)) as 核实地址
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实姓名' AS 问题原因,核实姓名 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实姓名%",1,0)) as 核实姓名
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实规格' AS 问题原因,核实规格 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实规格%",1,0)) as 核实规格
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实数量' AS 问题原因,核实数量 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实数量%",1,0)) as 核实数量
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实电话' AS 问题原因,核实电话 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实电话%",1,0)) as 核实电话
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'重复下单' AS 问题原因,重复下单 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%重复下单%",1,0)) as 重复下单
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实拉黑率' AS 问题原因,核实拉黑率 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实拉黑率%",1,0)) as 核实拉黑率
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实IP' AS 问题原因,核实IP AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实IP%",1,0)) as 核实IP
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'回复留言' AS 问题原因,回复留言 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%回复留言%",1,0)) as 回复留言
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实金额' AS 问题原因,核实金额 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实金额%",1,0)) as 核实金额
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'删运单号' AS 问题原因,删运单号 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%删运单号%",1,0)) as 删运单号
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'可疑订单' AS 问题原因,可疑订单 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%可疑订单%",1,0)) as 可疑订单
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实邮箱' AS 问题原因,核实邮箱 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实邮箱%",1,0)) as 核实邮箱
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'无法派送地区' AS 问题原因,无法派送地区 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%无法派送地区%",1,0)) as 无法派送地区
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = '2021-10-02'
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实邮编' AS 问题原因,核实邮编 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实邮编%",1,0)) as 核实邮编
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'拼团未完成' AS 问题原因,拼团未完成 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%拼团未完成%",1,0)) as 拼团未完成
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'支付失败' AS 问题原因,支付失败 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%支付失败%",1,0)) as 支付失败
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'未支付' AS 问题原因,未支付 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%未支付%",1,0)) as 未支付
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )					
                ) scs
                ORDER BY 币种 , 数量 DESC;'''
        # df02 = pd.read_sql_query(sql=sql02, con=self.engine1)
        # listT.append(df02)

        # 11、各团队-各二级品类
        # print('正在获取---3、各团队-各二级品类…………')
        sql20 = '''SELECT *
                    FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                IFNULL(cx.`币种`, '总计') 地区,
                                IFNULL(cx.`家族`, '总计') 家族,
                                IFNULL(cx.`父级分类`, '总计') 父级分类,
                                IFNULL(cx.`二级分类`, '总计') 二级分类,
                                COUNT(cx.`订单编号`) as 总单量,
                                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
                                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
                                concat(ROUND(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
                                concat(ROUND(SUM(IF(最终状态 = "已退货",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 退款率,
                                concat(ROUND(SUM(IF(最终状态 = "已签收",价格RMB,0)) / SUM(价格RMB) * 100,2),'%') as '总计签收(金额)',
                                concat(ROUND(COUNT(cx.`订单编号`) / 总订单量 * 100,2),'%') as 品类占比,
                                ROUND(SUM(价格RMB) / COUNT(cx.`订单编号`),2) as 平均客单价,
                            SUM(IF(`是否改派` = '直发',1,0))  as 直发单量,
                                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 直发完成签收,
                                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发总计签收,
                                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发完成占比,
                                concat(ROUND(SUM(IF(`是否改派` = '直发',1,0)) / 直发总单量 * 100,2),'%') as 直发品类占比,
                            concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%')as 改派占比,
                                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 改派完成签收,
                                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派总计签收,
                                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派完成占比,
                                concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / 改派总单量 * 100,2),'%') as 改派品类占比
                        FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM gat_zqsb cc
                                where cc.`运单编号` is not null AND cc.团队 NOT IN ("红杉家族-港澳台", "红杉家族-港澳台2", "金狮-港澳台", "金鹏家族-小虎队","Line运营")
                            ) cx 
                        LEFT JOIN 
        					(SELECT 币种,家族,年月,count(订单编号) as 总订单量,SUM(IF(`是否改派`= '直发',1,0)) as 直发总单量,SUM(IF(`是否改派` = '改派',1,0)) as 改派总单量
        					FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                    FROM gat_zqsb cc 
        							WHERE cc.`运单编号` is not null AND cc.团队 NOT IN ("红杉家族-港澳台", "红杉家族-港澳台2", "金狮-港澳台", "金鹏家族-小虎队","Line运营")
        						) dg  GROUP BY dg.币种,dg.家族,dg.年月
        					) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月                       
                        GROUP BY cx.年月,cx.币种,cx.家族,cx.父级分类,cx.二级分类
                        WITH ROLLUP 
                    ) s
                    ORDER BY 月份 DESC,
                            FIELD( 地区, '台湾', '香港', '总计' ),
                            FIELD( s.家族, '神龙','火凤凰','小虎队','神龙运营1组','红杉','金狮', '总计' ),
                            FIELD( s.父级分类, '居家百货', '电子电器', '服饰', '医药保健',  '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','总计' ),
                            FIELD( s.二级分类, '厨房用品', '日用百货', '布艺家纺', '宠物用品',  '户外运动', '汽车用品', '手表手环','影音娱乐','电脑外设','手机外设',
                                                '家用电器', '个护电器','上衣', '下装',  '内衣', '套装', '裙子','配饰','母婴服饰','保健食品','护理护具', 
                                                '保健器械', '药品', '成人保健', '凉/拖鞋', '皮鞋', '休闲运动鞋','靴子', '彩妆','护肤','个人洗护','单肩包','双肩包',
                                                '钱包','行李箱包', '手表', '饰品','玩具','母婴用品','总计'),
                            s.总单量 DESC;'''.format(month_last, team)
        # df20 = pd.read_sql_query(sql=sql20, con=self.engine1)
        # listT.append(df20)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        file_path = 'G:\\输出文件\\{} {}-签收率.xlsx'.format(today, match[team])
        sheet_name = ['每日各团队', '审核率_删单率', '各月各团队', '各月各团队分旬', '各团队各品类', '各团队各物流', '各团队各平台', '各平台各团队', '各品类各团队', '各物流各团队', '同产品各团队','同产品各月', '各团队二级品类']
        df0 = pd.DataFrame([])                                          # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)                            # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')           # 初始化写入对象
        book = load_workbook(file_path)                                 # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book                                              # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:                                 # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('zl_report_day')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        new_path = 'F:\\神龙签收率\\' + (datetime.datetime.now()).strftime('%m.%d') + '\\{} {}-签收率.xlsx'.format(today, match[team])
        shutil.copyfile(file_path, new_path)        # copy到指定位置
        print('----已写入excel; 并复制到指定文件夹中')

    # 更新-地区签收率(自己看的)
    def address_repot(self, team, month_last, month_yesterday):    # 更新-地区签收率
        today = datetime.date.today().strftime('%Y.%m.%d')
        match = {'gat': '港台'}
        # if team == 'gat':
        #     month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        #     month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
        # else:
        #     month_last = '2021-06-01'
        #     month_yesterday = '2021-07-31'
        print(month_last)
        print(month_yesterday)
        try:
            print('正在更新单表中......')
            sql = '''update {0}_order_list a, gat_update b
                            set a.`省洲`= IF(b.`省洲` = '', NULL, b.`省洲`),
                            set a.`市区`= IF(b.`市区` = '', NULL, b.`市区`)
        		            where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('正在更新总表中......')
            sql = '''update {0}_zqsb a, gat_update b
                            set a.`省洲`= IF(b.`省洲` = '', NULL, b.`省洲`)
                            set a.`市区`= IF(b.`市区` = '', NULL, b.`市区`)
                    		where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')

        # 地区分类
        sheet_name = ['地区分类', '地区总']
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---地区签收率…………')
        sql0 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.省洲,
    						IF(s2.签收=0,NULL,s2.签收) as 签收,
    						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
    						IF(s2.在途=0,NULL,s2.在途) as 在途,
    						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
    						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
    						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,
    						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
    						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
    						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
    						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
    						IF(s2.总订单=0,NULL,s2.总订单) as 全部,
                        concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                            concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                            concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                            concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                            concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                            concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                            concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                            concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                            concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                    FROM ( SELECT IFNULL(s1.币种,'合计') as 币种,
                                IFNULL(s1.家族,'合计') as 家族,
                                IFNULL(s1.年月,'合计') as 年月,
                                IFNULL(s1.是否改派,'合计') as 是否改派,
                                IFNULL(s1.省洲,'合计') as 省洲,
                                SUM(s1.签收) as 签收,
                                SUM(s1.拒收) as 拒收,
                                SUM(s1.在途) as 在途,
                                SUM(s1.未发货) as 未发货,
                                SUM(s1.未上线) as 未上线,
                                SUM(s1.已退货) as 已退货,
                                SUM(s1.理赔) as 理赔,
                                SUM(s1.自发头程丢件) as 自发头程丢件,
                                SUM(s1.已发货) as 已发货,
                                SUM(s1.已完成) as 已完成,
                                SUM(s1.总订单) as 总订单,
                                SUM(s1.签收金额) as 签收金额,
                                SUM(s1.退货金额) as 退货金额,
                                SUM(s1.完成金额) as 完成金额,
                                SUM(s1.发货金额) as 发货金额,
                                SUM(s1.总计金额) as 总计金额
                        FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.省洲 as 省洲,
                                    SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                    SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                    SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                    SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                    SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                    SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                    SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                    SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                    count(订单编号) as 总订单,
                                    count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                    SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                    SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                    SUM(`价格RMB`) as 总计金额,
                                    SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                                FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族
                                    FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                ) cx
                                GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派`, cx.`省洲`
                                ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`是否改派`, s1.`省洲`
                        with rollup
                    ) s2
                    GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`省洲`
                    HAVING s2.年月 <> '合计'
        ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','小虎队','红杉','神龙运营1组','金狮','合计'),
                FIELD(s2.`币种`,'台湾','香港','合计'),
                s2.`年月`,
                FIELD(s2.`是否改派`,'改派','直发','合计'),
                FIELD(s2.`省洲`,'屏东县','高雄市','新竹市','宜兰县','新北市','花莲县','台东县','基隆市','台北市','新竹县',
                                '桃园市','苗栗县','台中市','彰化县','南投县','嘉义市','嘉义县','云林县','台南市','合计'),
                s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)

        sql1 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.省洲,
            						IF(s2.签收=0,NULL,s2.签收) as 签收,
            						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
            						IF(s2.在途=0,NULL,s2.在途) as 在途,
            						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
            						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
            						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,
            						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
            						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
            						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
            						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
            						IF(s2.总订单=0,NULL,s2.总订单) as 全部,
                                concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                                    concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                                    concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                                    concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                                    concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                                concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                                    concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                                    concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                                    concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                                    concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                            FROM ( SELECT IFNULL(s1.币种,'合计') as 币种,
                                        IFNULL(s1.家族,'合计') as 家族,
                                        IFNULL(s1.年月,'合计') as 年月,
                                        IFNULL(s1.是否改派,'合计') as 是否改派,
                                        IFNULL(s1.省洲,'合计') as 省洲,
                                        SUM(s1.签收) as 签收,
                                        SUM(s1.拒收) as 拒收,
                                        SUM(s1.在途) as 在途,
                                        SUM(s1.未发货) as 未发货,
                                        SUM(s1.未上线) as 未上线,
                                        SUM(s1.已退货) as 已退货,
                                        SUM(s1.理赔) as 理赔,
                                        SUM(s1.自发头程丢件) as 自发头程丢件,
                                        SUM(s1.已发货) as 已发货,
                                        SUM(s1.已完成) as 已完成,
                                        SUM(s1.总订单) as 总订单,
                                        SUM(s1.签收金额) as 签收金额,
                                        SUM(s1.退货金额) as 退货金额,
                                        SUM(s1.完成金额) as 完成金额,
                                        SUM(s1.发货金额) as 发货金额,
                                        SUM(s1.总计金额) as 总计金额
                                FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.省洲 as 省洲,
                                            SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                            SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                            SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                            SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                            SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                            SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                            SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                            SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                            SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                            count(订单编号) as 总订单,
                                            count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                            SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                            SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                            SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                            SUM(`价格RMB`) as 总计金额,
                                            SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                                        FROM (SELECT *,
                                                IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "%火凤凰%","火凤凰",IF(cc.团队 LIKE "%神龙家族%","神龙",IF(cc.团队 LIKE "%金狮%","金狮",IF(cc.团队 LIKE "%金鹏%","小虎队",IF(cc.团队 LIKE "%神龙-运营1组%","神龙运营1组",cc.团队)))))) as 家族
                                            FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                        ) cx
                                        GROUP BY cx.`币种`,cx.`年月`, cx.`是否改派`, cx.`省洲`
                                        ORDER BY cx.`币种`,cx.`年月`, cx.`是否改派` DESC,总订单 DESC
                                ) s1
                            --    GROUP BY s1.`币种`, s1.`年月`, s1.`是否改派`, s1.`省洲`
                                GROUP BY s1.`币种`, s1.`年月`, s1.`省洲`
                                with rollup
                            ) s2
                         --   GROUP BY s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`省洲`
                            GROUP BY s2.`币种`, s2.`年月`, s2.`省洲`
                            HAVING s2.年月 <> '合计'
                ORDER BY FIELD(s2.`币种`,'台湾','香港','合计'),
                        s2.`年月`,
                        FIELD(s2.`是否改派`,'改派','直发','合计'),
                        FIELD(s2.`省洲`,'屏東縣','高雄市','新竹市','宜蘭縣','新北市','花蓮縣','臺東縣','基隆市','臺北市','新竹縣',
                                        '桃園市','苗栗縣','臺中市','彰化縣','南投縣','嘉義市','嘉義縣','雲林縣','臺南市','合计'),
                        s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df1 = pd.read_sql_query(sql=sql1, con=self.engine1)
        listT.append(df1)

        print('正在将 地区签收率 写入excel…………')
        file_path = 'G:\\输出文件\\{} {} 地区-签收率.xlsx'.format(today, match[team])
        df0 = pd.DataFrame([])                                  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)                    # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')   # 初始化写入对象
        book = load_workbook(file_path)                         # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book                                      # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:                         # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_总_地区_两月签收率')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')
    # 更新上期-总表 （备用）
    def replaceHostbefore(self, team, last_time):
        try:
            print('正在获取往昔数据中......')
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 订单来源, 订单编号, 出货时间, 状态时间, 上线时间, 最终状态,是否改派,物流方式, 产品id, 
                            父级分类,二级分类,三级分类,下单时间, 审核时间,仓储扫描时间,完结状态时间,IF(价格RMB = '',null,价格RMB) as 价格RMB, '{0}' as 记录时间
                    FROM gat_update;'''.format(last_time)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print('正在添加缓存中......')
            df.to_sql('gat_update_cp', con=self.engine1, index=False, if_exists='replace')
            print('正在数据添加中......')
            sql = '''REPLACE INTO qsb_{0} SELECT * FROM gat_update_cp; '''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')
    # report报表（备用）
    def qsb_report(self, team, day_yesterday, day_last):  # 获取各团队近两个月的物流数据
        match = {'gat': '港台'}
        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 每日
        sql0 = '''SELECT 月份,地区, 家族,
                        SUM(s.昨日订单量) as 昨日订单量,
                        SUM(s.直发签收) as 直发签收,
                        SUM(s.直发完成) as 直发完成,
                        SUM(s.直发总订单) as 直发总订单,
                        IFNULL(SUM(s.直发签收) / SUM(s.直发完成), 0) as 直发完成签收,
                        IFNULL(SUM(s.直发签收) / SUM(s.直发总订单), 0) as 直发总计签收,
                        IFNULL(SUM(s.直发完成) / SUM(s.直发总订单), 0) as 直发完成占比,
                        SUM(s.改派签收) as 改派签收,
                        SUM(s.改派完成) as 改派完成,
                        SUM(s.改派总订单) as 改派总订单,
                        IFNULL(SUM(s.改派签收) / SUM(s.改派完成), 0) as 改派完成签收,
                        IFNULL(SUM(s.改派签收) / SUM(s.改派总订单), 0) as 改派总计签收,
                        IFNULL(SUM(s.改派完成) / SUM(s.改派总订单), 0) as 改派完成占比
                FROM( SELECT IFNULL(cx.`年月`, '总计') 月份,
                            IFNULL(cx.币种, '总计') 地区,
                            IFNULL(cx.团队, '总计') 家族,
                            SUM(IF(cx.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY),1,0)) as 昨日订单量,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) as 直发签收,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 直发完成,
                            SUM(IF(`是否改派` = '直发',1,0)) as 直发总订单,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) as 改派签收,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 改派完成,
                            SUM(IF(`是否改派` = '改派',1,0)) as 改派总订单
                        FROM  qsb_gat cx
                        WHERE cx.`记录时间` = '{1}'
                        GROUP BY cx.年月,cx.币种,cx.团队
                        WITH ROLLUP 
                    ) s
                    GROUP BY 月份,地区,家族
                    ORDER BY 月份 DESC,
                            FIELD( 地区, '台湾', '香港', '总计' ),
                            FIELD( 家族, '神龙', '火凤凰', '红杉', '金狮', '总计' );'''.format(team, day_yesterday)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)

        # 总表
        sql = '''SELECT cx.币种 线路,
			                cx.团队 家族,
			                cx.年月 月份,
			                count(订单编号) as 总订单,
			                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
			                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) * 100,2),'%') as 总计签收,
			                concat(ROUND(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) * 100,2),'%') as 完成占比,
			                null 序号
                    FROM qsb_gat cx
                    WHERE cx.`记录时间` = '{1}'
                    GROUP BY cx.币种,cx.团队,cx.年月
                    ORDER BY cx.币种,cx.团队,cx.年月;'''.format(team, day_yesterday)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df)
        # 总表-上月
        sql2 = '''SELECT 线路,家族,月份,总订单,完成签收,总计签收,完成占比,@rownum:=@rownum+1 AS 序号
	            FROM (SELECT cx.币种 线路,
        			        cx.团队 家族,
        			        cx.年月 月份,
        			        count(订单编号) as 总订单,
        			        concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
        			        concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) * 100,2),'%') as 总计签收,
        			        concat(ROUND(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) * 100,2),'%') as 完成占比,
        			        @rownum:=0 
                        FROM qsb_gat cx
                        WHERE cx.`记录时间` = '{1}'
                        GROUP BY cx.币种,cx.团队,cx.年月
                    ) s
                ORDER BY s.线路,s.家族,s.月份;'''.format(team, day_last)
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
        listT.append(df2)

        # 物流
        sql3 = '''SELECT s2.币种,s2.团队 家族,s2.年月,s2.是否改派,s2.物流方式,
						s2.总订单,
						concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
						concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
						concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
						concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
						concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
						concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			            null 序号
				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.物流方式,'总计') as 物流方式,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.物流方式 as 物流方式,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                          ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
						    ) s1
						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
					   	    with rollup
					    ) s2
                ORDER BY    FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`物流方式`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)
        # 物流-上月
        sql4 = '''SELECT 币种,团队 家族,年月,是否改派,物流方式,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as '总计签收(金额)',累计占比, @rownum:=@rownum+1 AS 序号
		        FROM ( SELECT s2.币种,
        							s2.团队,
        							s2.年月,
        							s2.是否改派,
        							s2.物流方式,
        							s2.总订单,
        							concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        							concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        							concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        							concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        							concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
        							concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.物流方式,'总计') as 物流方式,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.物流方式 as 物流方式,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        						    ) s1
        						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
        					   	    with rollup
        					    ) s2
                        ) s
                        ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        					    FIELD(s.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        					    FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        					    FIELD(s.`是否改派`,'直发','改派','总计'),
        					    FIELD(s.`物流方式`,'总计'),
        					    s.总订单 DESC;'''.format(team, day_last)
        df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        listT.append(df4)

        # 品类
        sql5 = '''SELECT s2.币种,
								s2.团队 家族,
								s2.年月,
								s2.是否改派,
								s2.父级分类,
								s2.总订单,
								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
                                concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			                    null 序号
				 FROM (
                        SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.父级分类,'总计') as 父级分类,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	 SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.父级分类 as 父级分类,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                            ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
							) s1
						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
					   	with rollup
				 ) s2
				 ORDER BY	FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`父级分类`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)
        # 品类-上月
        sql5 = '''SELECT 币种,团队 家族,年月,是否改派,父级分类,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as `总计签收(金额)`,累计占比, @rownum:=@rownum+1 AS 序号
		        FROM (SELECT s2.币种,
        								s2.团队,
        								s2.年月,
        								s2.是否改派,
        								s2.父级分类,
        								s2.总订单,
        								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
                                        concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM (
                                SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.父级分类,'总计') as 父级分类,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	 SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.父级分类 as 父级分类,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                    ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        							) s1
        						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
        					   	with rollup
        				) s2 
        		) s
                ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        				FIELD(s.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        				FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        				FIELD(s.`是否改派`,'直发','改派','总计'),
        				FIELD(s.`父级分类`,'总计'),
        				s.总订单 DESC;'''.format(team, day_last)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)

        # 产品
        sql6 = '''SELECT * 
				    FROM ( SELECT   IFNULL( cx.`币种`,'总计') as 币种,
                                    IFNULL( cx.`团队`,'总计') as 家族,
                                    IFNULL( cx.`年月`,'总计') as 年月,
                                    IFNULL( cx.`产品id`,'总计') as 产品id,
                                    cx.`产品名称`,
							        cx.`父级分类`,
                                    count(订单编号) as 总订单,
                                    SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成签收,
                                    SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) as 总计签收,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) as 完成占比,
                                    count(订单编号) /总订单2 单量占比,
                                    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)),0) as 直发完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发完成占比,
                                    IFNULL(SUM(IF(是否改派 = '直发',1,0))  / 直发总订单2,0) as 直发单量占比,
                                    SUM(IF(是否改派 = '改派',1,0)) as 改派总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)),0) as 改派完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派完成占比,
                                    IFNULL(SUM(IF(是否改派 = '改派',1,0)) / 改派总订单2,0) 改派单量占比
                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                            LEFT JOIN  (SELECT 币种,团队,年月,count(订单编号) as 总订单2 , 
											    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单2 , 
												SUM(IF(是否改派 = '改派',1,0)) as 改派总订单2 
										FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') da GROUP BY da.币种,da.团队,da.年月
									) cx2  ON cx.币种 = cx2.币种 AND cx.团队 = cx2.团队 AND cx.年月 = cx2.年月
                            GROUP BY cx.币种,cx.团队,cx.年月,`产品id`
	                        with rollup
					) s1
	                ORDER BY	FIELD(s1.`币种`,'台湾','香港','总计'),
								FIELD(s1.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
								FIELD(s1.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
								总订单 DESC;'''.format(team, day_yesterday)
        df6 = pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)

        # 产品明细-台湾
        sql7 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
			            concat(ROUND(速派签收量 / 速派完成量 * 100,2),'%')  AS 速派完成签收,
			            concat(ROUND(速派签收量 / 速派单量 * 100,2),'%')  AS 速派总计签收,
			            concat(ROUND(速派完成量 / 速派单量 * 100,2),'%')  AS 速派完成占比,
			            concat(ROUND(速派单量 / 订单量 * 100,2),'%')  AS 速派单量占比,
			            concat(ROUND(711签收量 / 711完成量 * 100,2),'%')  AS 711完成签收,
			            concat(ROUND(711签收量 / 711单量 * 100,2),'%')  AS 711总计签收,
			            concat(ROUND(711完成量 / 711单量 * 100,2),'%')  AS 711完成占比,
			            concat(ROUND(711单量 / 订单量 * 100,2),'%')  AS 711单量占比,
			            concat(ROUND(天马签收量 / 天马完成量 * 100,2),'%')  AS 天马完成签收,
			            concat(ROUND(天马签收量 / 天马单量 * 100,2),'%')  AS 天马总计签收,
			            concat(ROUND(天马完成量 / 天马单量 * 100,2),'%')  AS 天马完成占比,
			            concat(ROUND(天马单量 / 订单量 * 100,2),'%')  AS 天马单量占比,
			            concat(ROUND(易速配签收量 / 易速配完成量 * 100,2),'%')  AS 易速配完成签收,
			            concat(ROUND(易速配签收量 / 易速配单量 * 100,2),'%')  AS 易速配总计签收,
			            concat(ROUND(易速配完成量 / 易速配单量 * 100,2),'%')  AS 易速配完成占比,
			            concat(ROUND(易速配单量 / 订单量 * 100,2),'%')  AS 易速配单量占比,
			            concat(ROUND(森鸿签收量 / 森鸿完成量 * 100,2),'%')  AS 森鸿完成签收,
			            concat(ROUND(森鸿签收量 / 森鸿单量 * 100,2),'%')  AS 森鸿总计签收,
			            concat(ROUND(森鸿完成量 / 森鸿单量 * 100,2),'%')  AS 森鸿完成占比,
			            concat(ROUND(森鸿单量 / 订单量 * 100,2),'%')  AS 森鸿单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
				            SUM(速派单量) 速派单量,  SUM(速派签收量) 速派签收量,  SUM(速派完成量) 速派完成量,
				            SUM(711单量) 711单量,  SUM(711签收量) 711签收量,  SUM(711完成量) 711完成量,
				            SUM(天马单量) 天马单量,  SUM(天马签收量) 天马签收量,  SUM(天马完成量) 天马完成量,
				            SUM(易速配单量) 易速配单量,  SUM(易速配签收量) 易速配签收量,  SUM(易速配完成量) 易速配完成量,
				            SUM(森鸿单量) 森鸿单量,  SUM(森鸿签收量) 森鸿签收量,  SUM(森鸿完成量) 森鸿完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派签收量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  速派完成量,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS '711单量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as  '711签收量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  '711完成量',
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马单量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as  天马签收量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  天马完成量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配单量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as  易速配签收量,
							    SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配完成量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿单量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as  森鸿签收量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '台湾'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-港澳台', '火凤凰-港澳台', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df7 = pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)
        # 产品明细-香港
        sql8 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
						concat(ROUND(立邦签收量 / 立邦完成量 * 100,2),'%')  AS 立邦完成签收,
						concat(ROUND(立邦签收量 / 立邦单量 * 100,2),'%')  AS 立邦总计签收,
						concat(ROUND(立邦完成量 / 立邦单量 * 100,2),'%')  AS 立邦完成占比,
						concat(ROUND(立邦单量 / 订单量 * 100,2),'%')  AS 立邦单量占比,
						concat(ROUND(森鸿SF签收量 / 森鸿SF完成量 * 100,2),'%')  AS 森鸿SF完成签收,
						concat(ROUND(森鸿SF签收量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF总计签收,
						concat(ROUND(森鸿SF完成量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF完成占比,
						concat(ROUND(森鸿SF单量 / 订单量 * 100,2),'%')  AS 森鸿SF单量占比,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH完成量 * 100,2),'%')  AS 森鸿SH完成签收,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH总计签收,
					    concat(ROUND(森鸿SH完成量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH完成占比,
					    concat(ROUND(森鸿SH单量 / 订单量 * 100,2),'%')  AS 森鸿SH单量占比,
					    concat(ROUND(易速配SF签收量 / 易速配SF完成量 * 100,2),'%')  AS 易速配SF完成签收,
					    concat(ROUND(易速配SF签收量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF总计签收,
					    concat(ROUND(易速配SF完成量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF完成占比,
					    concat(ROUND(易速配SF单量 / 订单量 * 100,2),'%')  AS 易速配SF单量占比,
					    concat(ROUND(易速配YC签收量 / 易速配YC完成量 * 100,2),'%')  AS 易速配YC完成签收,
					    concat(ROUND(易速配YC签收量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC总计签收,
					    concat(ROUND(易速配YC完成量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC完成占比,
					    concat(ROUND(易速配YC单量 / 订单量 * 100,2),'%')  AS 易速配YC单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
							SUM(立邦单量) 立邦单量,  SUM(立邦签收量) 立邦签收量,  SUM(立邦完成量) 立邦完成量,
				            SUM(森鸿SF单量) 森鸿SF单量,  SUM(森鸿SF签收量) 森鸿SF签收量,  SUM(森鸿SF完成量) 森鸿SF完成量,
				            SUM(森鸿SH单量) 森鸿SH单量,  SUM(森鸿SH签收量) 森鸿SH签收量,  SUM(森鸿SH完成量) 森鸿SH完成量,					
				            SUM(易速配SF单量) 易速配SF单量,  SUM(易速配SF签收量) 易速配SF签收量,  SUM(易速配SF完成量) 易速配SF完成量,				
				            SUM(易速配YC单量) 易速配YC单量,  SUM(易速配YC签收量) 易速配YC签收量,  SUM(易速配YC完成量) 易速配YC完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦签收量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  立邦完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿SF单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SF签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SF完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SH签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SH完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配SF单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as  易速配SF签收量,
							    SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配SF完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" ,1,0)) AS 易速配YC单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 = "已签收",1,0)) as  易速配YC签收量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配YC完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '香港'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-港澳台', '火凤凰-港澳台', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df8 = pd.read_sql_query(sql=sql8, con=self.engine1)
        listT.append(df8)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        for wbbook in ['神龙', '火凤凰', '红杉', '金狮']:
            file_path = 'G:\\输出文件\\{} {}-签收率.xlsx'.format(today, wbbook)
            sheet_name = ['每日', '总表', '总表上月', '物流', '物流上月', '品类', '品类上月', '产品', '产品明细台湾', '产品明细香港']
            df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listT)):
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
                del book['Sheet1']
            writer.save()
            writer.close()
            # print('正在运行' + wbbook + '表宏…………')
            # app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            # app.display_alerts = False
            # wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            # wbsht1 = app.books.open(file_path)
            # wbsht.macro('py_sl_总运行')()
            # wbsht1.save()
            # wbsht1.close()
            # wbsht.close()
            # app.quit()
        print('----已写入excel ')
    # 获取各团队近两个月的物流数据（备用）
    def qsb_report_T(self, team, day_yesterday, day_last):
        match = {'gat': '港台'}
        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 物流
        sql3 = '''SELECT s2.币种,s2.团队 家族,s2.年月,s2.是否改派,s2.物流方式,
						s2.总订单,
						concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
						concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
						concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
						concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
						concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
						concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			            null 序号
				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.物流方式,'总计') as 物流方式,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.物流方式 as 物流方式,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                          ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
						    ) s1
						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
					   	    with rollup
					    ) s2
                ORDER BY    FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`物流方式`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)
        # 物流-上月
        sql4 = '''SELECT 币种,团队 家族,年月,是否改派,物流方式,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as '总计签收(金额)',累计占比, @rownum:=@rownum+1 AS 序号
		        FROM ( SELECT s2.币种,
        							s2.团队,
        							s2.年月,
        							s2.是否改派,
        							s2.物流方式,
        							s2.总订单,
        							concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        							concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        							concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        							concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        							concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
        							concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.物流方式,'总计') as 物流方式,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.物流方式 as 物流方式,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        						    ) s1
        						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
        					   	    with rollup
        					    ) s2
                        ) s
                        ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        					    FIELD(s.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        					    FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        					    FIELD(s.`是否改派`,'直发','改派','总计'),
        					    FIELD(s.`物流方式`,'总计'),
        					    s.总订单 DESC;'''.format(team, day_last)
        df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        listT.append(df4)

        # 品类
        sql5 = '''SELECT s2.币种,
								s2.团队 家族,
								s2.年月,
								s2.是否改派,
								s2.父级分类,
								s2.总订单,
								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
                                concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			                    null 序号
				 FROM (
                        SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.父级分类,'总计') as 父级分类,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	 SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.父级分类 as 父级分类,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                            ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
							) s1
						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
					   	with rollup
				 ) s2
				 ORDER BY	FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`父级分类`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)
        # 品类-上月
        sql5 = '''SELECT 币种,团队 家族,年月,是否改派,父级分类,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as `总计签收(金额)`,累计占比, @rownum:=@rownum+1 AS 序号
		        FROM (SELECT s2.币种,
        								s2.团队,
        								s2.年月,
        								s2.是否改派,
        								s2.父级分类,
        								s2.总订单,
        								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
                                        concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM (
                                SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.父级分类,'总计') as 父级分类,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	 SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.父级分类 as 父级分类,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                    ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        							) s1
        						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
        					   	with rollup
        				) s2 
        		) s
                ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        				FIELD(s.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        				FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        				FIELD(s.`是否改派`,'直发','改派','总计'),
        				FIELD(s.`父级分类`,'总计'),
        				s.总订单 DESC;'''.format(team, day_last)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)

        # 产品
        sql6 = '''SELECT * 
				    FROM ( SELECT   IFNULL( cx.`币种`,'总计') as 币种,
                                    IFNULL( cx.`团队`,'总计') as 家族,
                                    IFNULL( cx.`年月`,'总计') as 年月,
                                    IFNULL( cx.`产品id`,'总计') as 产品id,
                                    cx.`产品名称`,
							        cx.`父级分类`,
                                    count(订单编号) as 总订单,
                                    SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成签收,
                                    SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) as 总计签收,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) as 完成占比,
                                    count(订单编号) /总订单2 单量占比,
                                    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)),0) as 直发完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发完成占比,
                                    IFNULL(SUM(IF(是否改派 = '直发',1,0))  / 直发总订单2,0) as 直发单量占比,
                                    SUM(IF(是否改派 = '改派',1,0)) as 改派总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)),0) as 改派完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派完成占比,
                                    IFNULL(SUM(IF(是否改派 = '改派',1,0)) / 改派总订单2,0) 改派单量占比
                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                            LEFT JOIN  (SELECT 币种,团队,年月,count(订单编号) as 总订单2 , 
											    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单2 , 
												SUM(IF(是否改派 = '改派',1,0)) as 改派总订单2 
										FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') da GROUP BY da.币种,da.团队,da.年月
									) cx2  ON cx.币种 = cx2.币种 AND cx.团队 = cx2.团队 AND cx.年月 = cx2.年月
                            GROUP BY cx.币种,cx.团队,cx.年月,`产品id`
	                        with rollup
					) s1
	                ORDER BY	FIELD(s1.`币种`,'台湾','香港','总计'),
								FIELD(s1.`团队`,'神龙家族-港澳台','火凤凰-港澳台','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
								FIELD(s1.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
								总订单 DESC;'''.format(team, day_yesterday)
        df6 = pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)

        # 产品明细-台湾
        sql7 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
			            concat(ROUND(速派签收量 / 速派完成量 * 100,2),'%')  AS 速派完成签收,
			            concat(ROUND(速派签收量 / 速派单量 * 100,2),'%')  AS 速派总计签收,
			            concat(ROUND(速派完成量 / 速派单量 * 100,2),'%')  AS 速派完成占比,
			            concat(ROUND(速派单量 / 订单量 * 100,2),'%')  AS 速派单量占比,
			            concat(ROUND(711签收量 / 711完成量 * 100,2),'%')  AS 711完成签收,
			            concat(ROUND(711签收量 / 711单量 * 100,2),'%')  AS 711总计签收,
			            concat(ROUND(711完成量 / 711单量 * 100,2),'%')  AS 711完成占比,
			            concat(ROUND(711单量 / 订单量 * 100,2),'%')  AS 711单量占比,
			            concat(ROUND(天马签收量 / 天马完成量 * 100,2),'%')  AS 天马完成签收,
			            concat(ROUND(天马签收量 / 天马单量 * 100,2),'%')  AS 天马总计签收,
			            concat(ROUND(天马完成量 / 天马单量 * 100,2),'%')  AS 天马完成占比,
			            concat(ROUND(天马单量 / 订单量 * 100,2),'%')  AS 天马单量占比,
			            concat(ROUND(易速配签收量 / 易速配完成量 * 100,2),'%')  AS 易速配完成签收,
			            concat(ROUND(易速配签收量 / 易速配单量 * 100,2),'%')  AS 易速配总计签收,
			            concat(ROUND(易速配完成量 / 易速配单量 * 100,2),'%')  AS 易速配完成占比,
			            concat(ROUND(易速配单量 / 订单量 * 100,2),'%')  AS 易速配单量占比,
			            concat(ROUND(森鸿签收量 / 森鸿完成量 * 100,2),'%')  AS 森鸿完成签收,
			            concat(ROUND(森鸿签收量 / 森鸿单量 * 100,2),'%')  AS 森鸿总计签收,
			            concat(ROUND(森鸿完成量 / 森鸿单量 * 100,2),'%')  AS 森鸿完成占比,
			            concat(ROUND(森鸿单量 / 订单量 * 100,2),'%')  AS 森鸿单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
				            SUM(速派单量) 速派单量,  SUM(速派签收量) 速派签收量,  SUM(速派完成量) 速派完成量,
				            SUM(711单量) 711单量,  SUM(711签收量) 711签收量,  SUM(711完成量) 711完成量,
				            SUM(天马单量) 天马单量,  SUM(天马签收量) 天马签收量,  SUM(天马完成量) 天马完成量,
				            SUM(易速配单量) 易速配单量,  SUM(易速配签收量) 易速配签收量,  SUM(易速配完成量) 易速配完成量,
				            SUM(森鸿单量) 森鸿单量,  SUM(森鸿签收量) 森鸿签收量,  SUM(森鸿完成量) 森鸿完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派签收量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  速派完成量,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS '711单量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as  '711签收量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  '711完成量',
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马单量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as  天马签收量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  天马完成量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配单量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as  易速配签收量,
							    SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配完成量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿单量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as  森鸿签收量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '台湾'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-港澳台', '火凤凰-港澳台', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df7 = pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)
        # 产品明细-香港
        sql8 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
						concat(ROUND(立邦签收量 / 立邦完成量 * 100,2),'%')  AS 立邦完成签收,
						concat(ROUND(立邦签收量 / 立邦单量 * 100,2),'%')  AS 立邦总计签收,
						concat(ROUND(立邦完成量 / 立邦单量 * 100,2),'%')  AS 立邦完成占比,
						concat(ROUND(立邦单量 / 订单量 * 100,2),'%')  AS 立邦单量占比,
						concat(ROUND(森鸿SF签收量 / 森鸿SF完成量 * 100,2),'%')  AS 森鸿SF完成签收,
						concat(ROUND(森鸿SF签收量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF总计签收,
						concat(ROUND(森鸿SF完成量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF完成占比,
						concat(ROUND(森鸿SF单量 / 订单量 * 100,2),'%')  AS 森鸿SF单量占比,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH完成量 * 100,2),'%')  AS 森鸿SH完成签收,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH总计签收,
					    concat(ROUND(森鸿SH完成量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH完成占比,
					    concat(ROUND(森鸿SH单量 / 订单量 * 100,2),'%')  AS 森鸿SH单量占比,
					    concat(ROUND(易速配SF签收量 / 易速配SF完成量 * 100,2),'%')  AS 易速配SF完成签收,
					    concat(ROUND(易速配SF签收量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF总计签收,
					    concat(ROUND(易速配SF完成量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF完成占比,
					    concat(ROUND(易速配SF单量 / 订单量 * 100,2),'%')  AS 易速配SF单量占比,
					    concat(ROUND(易速配YC签收量 / 易速配YC完成量 * 100,2),'%')  AS 易速配YC完成签收,
					    concat(ROUND(易速配YC签收量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC总计签收,
					    concat(ROUND(易速配YC完成量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC完成占比,
					    concat(ROUND(易速配YC单量 / 订单量 * 100,2),'%')  AS 易速配YC单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
							SUM(立邦单量) 立邦单量,  SUM(立邦签收量) 立邦签收量,  SUM(立邦完成量) 立邦完成量,
				            SUM(森鸿SF单量) 森鸿SF单量,  SUM(森鸿SF签收量) 森鸿SF签收量,  SUM(森鸿SF完成量) 森鸿SF完成量,
				            SUM(森鸿SH单量) 森鸿SH单量,  SUM(森鸿SH签收量) 森鸿SH签收量,  SUM(森鸿SH完成量) 森鸿SH完成量,					
				            SUM(易速配SF单量) 易速配SF单量,  SUM(易速配SF签收量) 易速配SF签收量,  SUM(易速配SF完成量) 易速配SF完成量,				
				            SUM(易速配YC单量) 易速配YC单量,  SUM(易速配YC签收量) 易速配YC签收量,  SUM(易速配YC完成量) 易速配YC完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦签收量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  立邦完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿SF单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SF签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SF完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SH签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SH完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配SF单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as  易速配SF签收量,
							    SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配SF完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" ,1,0)) AS 易速配YC单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 = "已签收",1,0)) as  易速配YC签收量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配YC完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '香港'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-港澳台', '火凤凰-港澳台', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df8 = pd.read_sql_query(sql=sql8, con=self.engine1)
        listT.append(df8)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        for wbbook in ['神龙', '火凤凰', '红杉', '金狮']:
            file_path = 'G:\\输出文件\\{} {}-签收率.xlsx'.format(today, wbbook)
            sheet_name = ['每日', '总表', '总表上月', '物流', '物流上月', '品类', '品类上月', '产品', '产品明细台湾', '产品明细香港']
            df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listT)):
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
                del book['Sheet1']
            writer.save()
            writer.close()
            # print('正在运行' + wbbook + '表宏…………')
            # app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            # app.display_alerts = False
            # wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            # wbsht1 = app.books.open(file_path)
            # wbsht.macro('py_sl_总运行')()
            # wbsht1.save()
            # wbsht1.close()
            # wbsht.close()
            # app.quit()
        print('----已写入excel ')


    # 拒收核实-查询需要的产品id
    def jushou(self):
        print('正在查询需核实订单…………')
        listT = []  # 查询sql的结果 存放池
        sql = '''SELECT *
                FROM (SELECT g.*,c.`家族`,c.`月份`,c.`拒收`,c.`总订单`,c.`退货率`,c.`拒收率`
			            FROM  需核实拒收_每日新增订单 g
			            LEFT JOIN (SELECT *
								 FROM(SELECT IFNULL(s1.家族, '合计') 家族, IFNULL(s1.地区, '合计') 地区, IFNULL(s1.月份, '合计') 月份,
											IFNULL(s1.产品id, '合计') 产品id,
											IFNULL(s1.产品名称, '合计') 产品名称,
											IFNULL(s1.父级分类, '合计') 父级分类,
											IFNULL(s1.二级分类, '合计') 二级分类,
											SUM(s1.已签收) as 已签收,
											SUM(s1.拒收) as 拒收,
											SUM(s1.已退货) as 已退货,
											SUM(s1.已完成) as 已完成,
						                    SUM(s1.总订单) as 总订单,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						                    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						                    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						                    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率
                                    FROM(SELECT IFNULL(cx.`家族`, '合计') 家族, IFNULL(cx.币种, '合计') 地区, IFNULL(cx.`年月`, '合计') 月份,
						                        IFNULL(cx.产品id, '合计') 产品id,
						                        IFNULL(cx.产品名称, '合计') 产品名称,
						                        IFNULL(cx.父级分类, '合计') 父级分类,
						                        IFNULL(cx.二级分类, '合计') 二级分类,
						                        COUNT(cx.`订单编号`) as 总订单,
						                        SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
						                        SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
						                        SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
						                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成		
		                                FROM (SELECT *,IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族
                                            FROM gat_zqsb cc 
					                        WHERE cc.年月 >=  DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m') AND cc.`币种` = '台湾' AND cc.`运单编号` is not null
		                                ) cx
                                        GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                    ) s1
                                    GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                    WITH ROLLUP 
                                ) s 
                                HAVING s.月份 != '合计' AND s.产品id != '合计' AND s.`拒收` >= '1'
                                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                                FIELD(s.`地区`,'台湾','香港','合计'),
                                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'合计'),
                                FIELD(s.`产品id`,'合计'),
                                s.拒收 DESC
			            ) c ON g.`团队` = c.`家族` AND EXTRACT(YEAR_MONTH FROM g.`下单时间`) = c.`月份` AND g.`产品Id` =c.`产品Id`
                ) s WHERE s.`家族` is not null;'''
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df)
        print('正在查询两月拒收订单…………')
        sql2 = '''SELECT * FROM 需核实拒收_获取最近两个月订单;'''
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
        listT.append(df2)
        print('正在查询两月拒收产品id…………')
        sql3 = '''SELECT *
								 FROM(SELECT IFNULL(s1.家族, '合计') 家族, IFNULL(s1.地区, '合计') 地区, IFNULL(s1.月份, '合计') 月份,
											IFNULL(s1.产品id, '合计') 产品id,
											IFNULL(s1.产品名称, '合计') 产品名称,
											IFNULL(s1.父级分类, '合计') 父级分类,
											IFNULL(s1.二级分类, '合计') 二级分类,
											SUM(s1.已签收) as 已签收,
											SUM(s1.拒收) as 拒收,
											SUM(s1.已退货) as 已退货,
											SUM(s1.已完成) as 已完成,
						                    SUM(s1.总订单) as 总订单,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						                    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						                    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						                    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率
                                    FROM(SELECT IFNULL(cx.`家族`, '合计') 家族, IFNULL(cx.币种, '合计') 地区, IFNULL(cx.`年月`, '合计') 月份,
						                        IFNULL(cx.产品id, '合计') 产品id,
						                        IFNULL(cx.产品名称, '合计') 产品名称,
						                        IFNULL(cx.父级分类, '合计') 父级分类,
						                        IFNULL(cx.二级分类, '合计') 二级分类,
						                        COUNT(cx.`订单编号`) as 总订单,
						                        SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
						                        SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
						                        SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
						                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成		
		                                FROM (SELECT *,IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族
                                            FROM gat_zqsb cc 
					                        WHERE cc.年月 >=  DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m') AND cc.`币种` = '台湾' AND cc.`运单编号` is not null
		                                ) cx
                                        GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                    ) s1
                                    GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                    WITH ROLLUP 
                                ) s 
                                HAVING s.月份 != '合计' AND s.产品id != '合计' AND s.`拒收` >= '1'
                                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                                FIELD(s.`地区`,'台湾','香港','合计'),
                                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'合计'),
                                FIELD(s.`产品id`,'合计'),
                                s.拒收 DESC;'''
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)
        print('正在查询需核实拒收_每日新增订单…………')
        sql4 = '''SELECT * FROM 需核实拒收_每日新增订单;'''
        df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        listT.append(df4)
        print('正在写入excel…………')
        today = datetime.date.today().strftime('%m.%d')
        file_path = 'G:\\输出文件\\{} 需核实拒收-每日数据源.xlsx'.format(today)
        if os.path.exists(file_path):  # 判断是否有需要的表格
            print("正在清除重复文件......")
            os.remove(file_path)
        sheet_name = ['查询', '两月拒收', '两月拒收产品id', '每日新增订单']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        for i in range(len(listT)):
            listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        print('----已写入excel ')

    # 获取电话核实日报表 周报表
    def phone_report(self, handle):
        today = datetime.date.today().strftime('%Y.%m.%d')
        match = {'gat': '港台'}
        week: datetime = datetime.datetime.now()
        if week.isoweekday() == 1 or week.isoweekday() == '手动':
            listT = []  # 查询sql的结果 存放池
            print('正在获取 日报表 数据内容…………')
            sql = '''SELECT 日期31天, ss.*, ss1.*, ss2.*, ss3.*, ss4.*
    FROM date
    LEFT JOIN
            (SELECT 日期 AS 系统问题,
                        COUNT(订单编号) AS 问题订单,
                        SUM(IF(g.`系统订单状态` NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货'),1,0)) AS 正常出货,
                        SUM(IF(g.`系统订单状态` = '已删除',1,0)) AS 删除订单,
                        SUM(IF(g.`系统物流状态` = '已签收',1,0)) AS 实际签收
            FROM gat_order_list g
            WHERE (g.日期 BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY)) AND 
                        (g.`问题时间` BETWEEN TIMESTAMP(DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY)) AND TIMESTAMP(CURDATE())) 
                        AND g.`问题原因` IS NOT NULL AND g.币种 = '台湾'
            GROUP BY DATE(日期) 
            ORDER BY DATE(日期)
    ) ss ON  date.`日期31天` = EXTRACT(day FROM ss.`系统问题`)
    
    LEFT JOIN 
    
    (	SELECT ww.* ,物流问题总量, 约派送, 核实拒收, 再派签收, 未接听, 无效号码
        FROM (SELECT 处理时间 AS 物流问题, COUNT(订单编号) AS 物流问题联系量
                    FROM 物流问题件 cg
                    WHERE cg.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) AND cg.币种 = '台币'
                    GROUP BY 处理时间
                    ORDER BY 处理时间
        ) ww
        LEFT JOIN 
        (SELECT 处理时间 AS 物流问题,
                        COUNT(订单编号) AS 物流问题总量,
                        SUM(IF(ks.`处理结果` LIKE '%送货%' or ks.`处理结果` LIKE '%配送%' or ks.`处理结果` LIKE '%自取%',1,0)) AS 约派送,
                        SUM(IF(ks.`处理结果` LIKE '%拒收%' ,1,0)) AS 核实拒收,
                        SUM(IF((ks.`处理结果` LIKE '%送货%' or ks.`处理结果` LIKE '%配送%') AND ks.`系统物流状态` LIKE '已签收%',1,0)) AS 再派签收,
                        SUM(IF(ks.`处理结果` LIKE '%无人接听%',1,0)) AS 未接听,
                        SUM(IF(ks.`处理结果` LIKE '%无效号码%',1,0)) AS 无效号码
            FROM (SELECT wt.*, g.`系统订单状态`, g.`系统物流状态`, g.`完结状态`
                        FROM (SELECT * 
                                    FROM 物流问题件 
                                    WHERE id IN (SELECT MAX(id) FROM 物流问题件 w WHERE w.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) GROUP BY 订单编号) 
                                    ORDER BY id
                        ) wt 
                        LEFT JOIN gat_order_list g ON  wt.`订单编号` = g.`订单编号`
                        WHERE wt.币种 = '台币'
            ) ks
            GROUP BY ks.处理时间
            ORDER BY 处理时间
        ) ww2  ON ww.`物流问题` = ww2.`物流问题`
    ) ss1 ON  date.`日期31天` = EXTRACT(day FROM ss1.`物流问题`)
    
    LEFT JOIN
    
    ( SELECT cc.* ,客诉总量, 挽回单数, 未确认, 退款单数, 实际退款单数, 实际挽回单数
        FROM (SELECT 处理时间 AS 物流客诉, COUNT(订单编号) AS 物流客诉联系量
                    FROM 物流客诉件 cg
                    WHERE cg.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) and cg.币种 = '台币'
                    GROUP BY 处理时间
                    ORDER BY 处理时间
        ) cc 
    LEFT JOIN
        (SELECT 处理时间 AS 物流客诉,
                        COUNT(订单编号) AS 客诉总量,
                        SUM(IF(ks.`处理方案` LIKE '%不退款%' or ks.`处理方案` LIKE '%赠品%' or ks.`处理方案` LIKE '%补发%' or ks.`处理方案` LIKE '%换货%',1,0)) AS 挽回单数,
                        SUM(IF(ks.`处理结果` LIKE '%转语音%' or ks.`处理结果` LIKE '%空号%' or ks.`处理结果` LIKE '%挂断电话%' or ks.`处理结果` LIKE '%无人接听%',1,0)) AS 未确认,
                        SUM(IF(ks.`处理方案` LIKE '%退款%' AND ks.`处理方案` NOT LIKE '%不%',1,0)) AS 退款单数,
                                
                        SUM(IF(ks.`完结状态` = '退款',1,0)) AS 实际退款单数,
                        SUM(IF(ks.`完结状态` = '收款',1,0)) AS 实际挽回单数
            FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`, g.`完结状态`
                        FROM (SELECT * 
                                    FROM 物流客诉件 
                                    WHERE id IN (SELECT MAX(id) FROM 物流客诉件 w WHERE w.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) GROUP BY 订单编号) 
                                    ORDER BY id
                            ) cg
                        LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                        WHERE cg.币种 = '台币'
            ) ks
    
            GROUP BY ks.处理时间
            ORDER BY 处理时间
        ) cc2  ON cc.`物流客诉` = cc2.`物流客诉`
    ) ss2 ON  date.`日期31天` = EXTRACT(day FROM ss2.`物流客诉`)
    
    LEFT JOIN
    
    (SELECT gg.* ,异常单量, 正常发货, 取消订单
        FROM (SELECT cg.处理时间 AS 采购异常, COUNT(cg.订单编号) AS 采购异常联系量
                    FROM 采购异常 cg
                    LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                    WHERE cg.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) AND g.是否改派 = '直发'
                    GROUP BY cg.处理时间
                    ORDER BY cg.处理时间
        ) gg 
        LEFT JOIN
        (SELECT DATE(s.处理时间) AS 采购异常,
                        COUNT(订单编号) AS 异常单量,
                        SUM(IF(s.`系统订单状态` NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货') AND s.处理结果 <> '跟进',1,0)) AS 正常发货,
                        SUM(IF(s.`系统订单状态` = '已删除',1,0)) AS 取消订单,
                        SUM(IF(s.处理结果 = '跟进',1,0)) AS 跟进,
                        
                        SUM(IF(s.`反馈内容` NOT like '%取消%',1,0)) AS 正常发货22,
                        SUM(IF(s.`反馈内容` like '%取消%',1,0)) AS 取消订单22
            FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`, g.`币种`, g.`是否改派`
                        FROM (SELECT * 
                                    FROM 采购异常 
                                    WHERE id IN (SELECT MAX(id) FROM 采购异常 w WHERE w.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) GROUP BY 订单编号) 
                                    ORDER BY id
                            ) cg
                        LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                        WHERE  g.是否改派 = '直发'
            ) s
            GROUP BY DATE(s.处理时间) 
            ORDER BY DATE(s.处理时间) 
        ) gg2 ON gg.`采购异常` = gg2.`采购异常`
    ) ss3 ON  date.`日期31天` = EXTRACT(day FROM ss3.`采购异常`)
    LEFT JOIN
    (SELECT DATE(s.处理时间) AS 拒收问题件,
                        COUNT(订单编号) AS '联系量（有结果）',
                        SUM(IF(s.`再次克隆下单` IS NOT NULL,1,0)) AS 挽单量
            FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`, g.`币种`
                        FROM (SELECT * 
                                    FROM 拒收问题件 
                                  WHERE id IN (SELECT MAX(id) FROM 拒收问题件 w WHERE w.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) GROUP BY 订单编号)	
                                    ORDER BY id
                            ) cg
                        LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                        WHERE g.币种 = '台湾' 
            ) s
            WHERE  s.核实原因 <> '未联系上客户'
            GROUP BY DATE(s.处理时间) 
            ORDER BY DATE(s.处理时间) 
        ) ss4 ON  date.`日期31天` = EXTRACT(day FROM ss4.`拒收问题件`)
    GROUP BY 日期31天
    ORDER BY 日期31天;'''.format()     # 港台查询函数导出
            df0 = pd.read_sql_query(sql=sql, con=self.engine1)
            listT.append(df0)
            print('正在获取 周报表 数据内容…………')
            sql = '''SELECT 日期31天,ss.问题订单,ss.正常出货,ss.删除订单,ss.实际签收,concat(ROUND(IFNULL(ss.删除订单/ss.问题订单,0) * 100,2),'%') as 取消占比,
                            ss1.约派送,ss1.核实拒收,ss1.再派签收,ss2.挽回单数,ss2.未确认,ss2.退款单数,ss2.实际挽回单数,ss3.正常发货,ss3.取消订单,ss4.`联系量（有结果）`,ss4.挽单量,
                            ss4.张联系量 AS '张陈平-联系量（有结果）',ss4.张挽单量 AS '张陈平-挽单量',
                            ss4.蔡联系量 AS '蔡利英-联系量（有结果）',ss4.蔡挽单量 AS '蔡利英-挽单量',
                            ss4.杨联系量 AS '杨嘉仪-联系量（有结果）',ss4.杨挽单量 AS '杨嘉仪-挽单量'
                    FROM date
                    LEFT JOIN
                    (SELECT 日期 AS 系统问题,COUNT(订单编号) AS 问题订单,
                            SUM(IF(g.`系统订单状态` NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货'),1,0)) AS 正常出货,
                            SUM(IF(g.`系统订单状态` = '已删除',1,0)) AS 删除订单,
                            SUM(IF(g.`系统物流状态` = '已签收',1,0)) AS 实际签收
                        FROM gat_order_list g
                        WHERE (g.日期  BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0)) AND 
                            (g.`问题时间` BETWEEN TIMESTAMP(subdate(curdate(),date_format(curdate(),'%w')+6)) AND TIMESTAMP(subdate(curdate(),date_format(curdate(),'%w')-1))) 
                            AND g.`问题原因` IS NOT NULL AND g.币种 = '台湾'
                        GROUP BY DATE(日期) 
                        ORDER BY DATE(日期)
                    ) ss ON  date.`日期31天` = EXTRACT(day FROM ss.`系统问题`)
                    LEFT JOIN 
                    (SELECT ww.* ,物流问题总量, 约派送, 核实拒收, 再派签收, 未接听, 无效号码
                        FROM (SELECT 处理时间 AS 物流问题, COUNT(订单编号) AS 物流问题联系量
                                FROM 物流问题件 cg
                                WHERE cg.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) AND cg.币种 = '台币'
                                GROUP BY 处理时间
                                ORDER BY 处理时间
                        ) ww
                        LEFT JOIN 
                        (SELECT 处理时间 AS 物流问题, COUNT(订单编号) AS 物流问题总量,
                                SUM(IF(ks.`处理结果` LIKE '%送货%' or ks.`处理结果` LIKE '%配送%' or ks.`处理结果` LIKE '%自取%',1,0)) AS 约派送,
                                SUM(IF(ks.`处理结果` LIKE '%拒收%' OR ks.`处理结果` LIKE '%无人接听%' OR ks.`处理结果` LIKE '%无效号码%',1,0)) AS 核实拒收,
                                SUM(IF((ks.`处理结果` LIKE '%送货%' or ks.`处理结果` LIKE '%配送%') AND ks.`系统物流状态` LIKE '已签收%',1,0)) AS 再派签收,
                                SUM(IF(ks.`处理结果` LIKE '%无人接听%',1,0)) AS 未接听,
                                SUM(IF(ks.`处理结果` LIKE '%无效号码%',1,0)) AS 无效号码
                        FROM (SELECT wt.*, g.`系统订单状态`, g.`系统物流状态`, g.`完结状态`
                                FROM (SELECT * 
                                        FROM 物流问题件 
                                        WHERE id IN (SELECT MAX(id) FROM 物流问题件 w WHERE w.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) GROUP BY 订单编号) 
                                        ORDER BY id
                                ) wt 
                            LEFT JOIN gat_order_list g ON  wt.`订单编号` = g.`订单编号`
                            WHERE wt.币种 = '台币'
                        ) ks
                        GROUP BY ks.处理时间
                        ORDER BY 处理时间
                        ) ww2  ON ww.`物流问题` = ww2.`物流问题`
                    ) ss1 ON  date.`日期31天` = EXTRACT(day FROM ss1.`物流问题`)
                    LEFT JOIN
                    ( SELECT cc.* ,客诉总量, 挽回单数, 未确认, 退款单数, 实际退款单数, 实际挽回单数
                        FROM (SELECT 处理时间 AS 物流客诉, COUNT(订单编号) AS 物流客诉联系量
                                FROM 物流客诉件 cg
                                WHERE cg.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) and cg.币种 = '台币'
                                GROUP BY 处理时间
                                ORDER BY 处理时间
                        ) cc 
                        LEFT JOIN
                        (SELECT 处理时间 AS 物流客诉, COUNT(订单编号) AS 客诉总量,
                                SUM(IF(ks.`处理方案` LIKE '%不退款%' or ks.`处理方案` LIKE '%赠品%' or ks.`处理方案` LIKE '%补发%' or ks.`处理方案` LIKE '%换货%',1,0)) AS 挽回单数,
                                SUM(IF(ks.`处理结果` LIKE '%转语音%' or ks.`处理结果` LIKE '%空号%' or ks.`处理结果` LIKE '%挂断电话%' or ks.`处理结果` LIKE '%无人接听%',1,0)) AS 未确认,
                                SUM(IF(ks.`处理方案` LIKE '%退款%' AND ks.`处理方案` NOT LIKE '%不%',1,0)) AS 退款单数,
                                        
                                SUM(IF(ks.`完结状态` = '退款',1,0)) AS 实际退款单数,
                                SUM(IF(ks.`完结状态` = '收款',1,0)) AS 实际挽回单数
                            FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`, g.`完结状态`
                                    FROM (SELECT * 
                                            FROM 物流客诉件 
                                            WHERE id IN (SELECT MAX(id) FROM 物流客诉件 w WHERE w.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) GROUP BY 订单编号) 
                                            ORDER BY id
                                    ) cg
                                    LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                    WHERE cg.币种 = '台币'
                            ) ks
                            GROUP BY ks.处理时间
                            ORDER BY 处理时间
                        ) cc2  ON cc.`物流客诉` = cc2.`物流客诉`
                    ) ss2 ON  date.`日期31天` = EXTRACT(day FROM ss2.`物流客诉`)
                    LEFT JOIN
                    (SELECT gg.* ,异常单量, 正常发货, 取消订单
                        FROM (SELECT 处理时间 AS 采购异常, COUNT(cg.订单编号) AS 采购异常联系量
                                FROM 采购异常 cg
                                LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                WHERE cg.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) AND g.是否改派 = '直发'
                                GROUP BY 处理时间
                                ORDER BY 处理时间
                        ) gg 
                        LEFT JOIN
                        (SELECT DATE(s.处理时间) AS 采购异常, COUNT(订单编号) AS 异常单量,
                                SUM(IF(s.`系统订单状态` NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货') AND s.处理结果 <> '跟进',1,0)) AS 正常发货,
                                SUM(IF(s.`系统订单状态` = '已删除',1,0)) AS 取消订单,
                                SUM(IF(s.处理结果 = '跟进',1,0)) AS 跟进,
                                SUM(IF(s.`反馈内容` NOT like '%取消%',1,0)) AS 正常发货22,
                                SUM(IF(s.`反馈内容` like '%取消%',1,0)) AS 取消订单22
                            FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`, g.`币种`
                                    FROM (SELECT * 
                                            FROM 采购异常 
                                            WHERE id IN (SELECT MAX(id) FROM 采购异常 w WHERE w.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) GROUP BY 订单编号) 
                                            ORDER BY id
                                    ) cg
                                     LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                    WHERE g.是否改派 = '直发'
                            ) s
                            GROUP BY DATE(s.处理时间) 
                            ORDER BY DATE(s.处理时间) 
                        ) gg2 ON gg.`采购异常` = gg2.`采购异常`
                    ) ss3 ON  date.`日期31天` = EXTRACT(day FROM ss3.`采购异常`)
                    LEFT JOIN
                    (SELECT DATE(s.处理时间) AS 拒收问题件,
                            COUNT(订单编号) AS '联系量（有结果）',
                            SUM(IF(s.`再次克隆下单` IS NOT NULL,1,0)) AS 挽单量,
                            SUM(IF(处理人='张陈平',1,0)) AS 张联系量,
                            SUM(IF(处理人='张陈平' AND s.`再次克隆下单` IS NOT NULL,1,0)) AS 张挽单量,
                            SUM(IF(处理人='蔡利英',1,0)) AS 蔡联系量,
                            SUM(IF(处理人='蔡利英' AND s.`再次克隆下单` IS NOT NULL,1,0)) AS 蔡挽单量,
                            SUM(IF(处理人='杨嘉仪',1,0)) AS 杨联系量,
                            SUM(IF(处理人='杨嘉仪' AND s.`再次克隆下单` IS NOT NULL,1,0)) AS 杨挽单量
                        FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`, g.`币种`
                                FROM (SELECT * 
                                        FROM 拒收问题件 
                                        WHERE id IN (SELECT MAX(id) FROM 拒收问题件 w WHERE w.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) GROUP BY 订单编号) 
                                        ORDER BY id
                                    ) cg
                                LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                WHERE g.币种 = '台湾' 
                        ) s
                        WHERE  s.核实原因 <> '未联系上客户'
                        GROUP BY DATE(s.处理时间) 
                        ORDER BY DATE(s.处理时间) 
                    ) ss4 ON  date.`日期31天` = EXTRACT(day FROM ss4.`拒收问题件`)
                    WHERE ss.系统问题 IS NOT NULL
                    GROUP BY 日期31天
                    ORDER BY 系统问题;'''.format()  # 港台查询函数导出
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            listT.append(df1)

            print('正在写入excel…………')
            today = datetime.date.today().strftime('%Y.%m.%d')
            file_path = 'G:\\输出文件\\{} 电话核实 日报表_周报表.xlsx'.format(today)
            sheet_name = ['日报表', '周报表']
            df0 = pd.DataFrame([])                                          # 创建空的dataframe数据框
            df0.to_excel(file_path, index=False)                            # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            writer = pd.ExcelWriter(file_path, engine='openpyxl')           # 初始化写入对象
            book = load_workbook(file_path)                                 # 可以向不同的sheet写入数据（对现有工作表的追加）
            writer.book = book                                              # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            for i in range(len(listT)):
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            if 'Sheet1' in book.sheetnames:                                 # 删除新建文档时的第一个工作表
                del book['Sheet1']
            writer.save()
            writer.close()
            try:
                print('正在运行 日报表、周报表 宏…………')
                app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
                app.display_alerts = False
                wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
                wbsht1 = app.books.open(file_path)
                wbsht.macro('电话核实日报表_周报表')()
                wbsht1.save()
                wbsht1.close()
                wbsht.close()
                app.quit()
            except Exception as e:
                print('运行失败：', str(Exception) + str(e))
            print('----已写入excel ')


    def slrb_new(self, team, month_last, month_yesterday):  # 报表各团队近两个月的物流数据
        month_now = datetime.datetime.now().strftime('%Y-%m-%d')
        match = {'gat': '港台'}
        emailAdd = {'台湾': 'giikinliujun@163.com',
                    '香港': 'giikinliujun@163.com',
                    '品牌': 'sunyaru@giikin.com'}
        # if team == 'gat9':
        #     month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        #     month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
        # else:
        #     month_last = '2021-08-01'
        #     month_yesterday = '2021-09-30'
        print(month_last)
        print(month_yesterday)
        sql = '''UPDATE gat_zqsb d
                        SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-易速配-顺丰%','香港-易速配-顺丰', IF(d.`物流方式` LIKE '台湾-天马-711%','台湾-天马-新竹', d.`物流方式`) )
                        WHERE d.`是否改派` ='直发';'''
        print('正在修改-直发的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''UPDATE gat_zqsb d
                        SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-森鸿%','香港-森鸿-改派',
                                        IF(d.`物流方式` LIKE '香港-立邦%','香港-立邦-改派',
            							IF(d.`物流方式` LIKE '香港-易速配%','香港-易速配-改派',
            							IF(d.`物流方式` LIKE '台湾-立邦普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-大黄蜂普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-森鸿-新竹%','森鸿',
            							IF(d.`物流方式` LIKE '台湾-天马-顺丰%','天马顺丰',
            							IF(d.`物流方式` LIKE '台湾-天马-新竹%' OR d.`物流方式` LIKE '台湾-天马-711%','天马新竹',
            							IF(d.`物流方式` LIKE '台湾-天马-黑猫%','天马黑猫',
            							IF(d.`物流方式` LIKE '台湾-易速配-龟山%' OR d.`物流方式` LIKE '台湾-易速配-新竹%' OR d.`物流方式` = '易速配','龟山',
            							IF(d.`物流方式` LIKE '台湾-速派-新竹%' OR d.`物流方式` LIKE '台湾-速派-711超商%','速派',
            							IF(d.`物流方式` LIKE '台湾-大黄蜂普货头程-易速配尾程%' OR d.`物流方式` LIKE '台湾-立邦普货头程-易速配尾程%','龟山', d.`物流方式`)))  )  )  )  )  )  )  )
                        WHERE d.`是否改派` ='改派';'''
        print('正在修改-改派的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 物流分类
        print('正在获取---物流分类…………')
        sql0 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.物流方式,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,					
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                    concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                        concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                        concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                FROM ( SELECT IFNULL(s1.币种,'合计') as 币种,
                            IFNULL(s1.家族,'合计') as 家族,
                            IFNULL(s1.年月,'合计') as 年月,
                            IFNULL(s1.是否改派,'合计') as 是否改派,
                            IFNULL(s1.物流方式,'合计') as 物流方式,
                            SUM(s1.签收) as 签收,
                            SUM(s1.拒收) as 拒收,
                            SUM(s1.在途) as 在途,
                            SUM(s1.未发货) as 未发货,
                            SUM(s1.未上线) as 未上线,
                            SUM(s1.已退货) as 已退货,
                            SUM(s1.理赔) as 理赔,
                            SUM(s1.自发头程丢件) as 自发头程丢件,
                            SUM(s1.已发货) as 已发货,
                            SUM(s1.已完成) as 已完成,
                            SUM(s1.总订单) as 总订单,
                            SUM(s1.签收金额) as 签收金额,
                            SUM(s1.退货金额) as 退货金额,
                            SUM(s1.完成金额) as 完成金额,
                            SUM(s1.发货金额) as 发货金额,
                            SUM(s1.总计金额) as 总计金额
                    FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.物流方式 as 物流方式,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                            ) cx
                            GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                            ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
                    ) s1
                    GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
                    with rollup
                ) s2
                GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`物流方式` 
                HAVING s2.年月 <> '合计'
    ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
            FIELD(s2.`币种`,'台湾','香港','合计'),
            s2.`年月`,
            FIELD(s2.`是否改派`,'改派','直发','合计'),
            FIELD(s2.`物流方式`, '台湾-大黄蜂普货头程-森鸿尾程','台湾-大黄蜂普货头程-易速配尾程', '台湾-立邦普货头程-森鸿尾程','台湾-易速配-TW海快','台湾-立邦普货头程-易速配尾程', 
                                '台湾-森鸿-新竹-自发头程', '台湾-速派-711超商', '台湾-速派-新竹','台湾-天马-新竹','台湾-天马-顺丰','台湾-天马-黑猫','台湾-易速配-新竹',
                                '香港-立邦-顺丰','香港-易速配-顺丰','香港-易速配-顺丰YC', '香港-森鸿-SH渠道','香港-森鸿-顺丰渠道',
                                '龟山','森鸿','速派','天马顺丰','天马新竹','香港-立邦-改派','香港-森鸿-改派','香港-易速配-改派','合计' ),
            s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)
        # 物流分旬
        print('正在获取---物流分旬…………')
        sql11 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.物流方式,s2.旬,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,	
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
                    concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                        concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                        concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'    
                FROM (SELECT IFNULL(s1.币种,'合计') as 币种,
                            IFNULL(s1.家族,'合计') as 家族,
                            IFNULL(s1.年月,'合计') as 年月,
                            IFNULL(s1.是否改派,'合计') as 是否改派,
                            IFNULL(s1.物流方式,'合计') as 物流方式,
                            IFNULL(s1.旬,'合计') as 旬,
                            SUM(s1.签收) as 签收,
                            SUM(s1.拒收) as 拒收,
                            SUM(s1.在途) as 在途,
                            SUM(s1.未发货) as 未发货,
                            SUM(s1.未上线) as 未上线,
                            SUM(s1.已退货) as 已退货,
                            SUM(s1.理赔) as 理赔,
                            SUM(s1.自发头程丢件) as 自发头程丢件,
                            SUM(s1.已发货) as 已发货,
                            SUM(s1.已完成) as 已完成,
                            SUM(s1.总订单) as 总订单,
                            s1.总订单量,
							s1.已发货单量,
							s1.已完成单量,
                            SUM(s1.签收金额) as 签收金额,
                            SUM(s1.退货金额) as 退货金额,
                            SUM(s1.完成金额) as 完成金额,
                            SUM(s1.发货金额) as 发货金额,
                            SUM(s1.总计金额) as 总计金额
                    FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.物流方式 as 物流方式,
                                IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
								总订单量,
								已发货单量,
								已完成单量,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                    FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                            ) cx
                            LEFT JOIN 
							    (SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                                FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                    ) dg  
								    GROUP BY dg.币种,dg.家族,dg.年月
                            ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                            GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派`, cx.`物流方式`, cx.`旬`
                            ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`是否改派`, s1.`物流方式`, s1.`旬`
                        with rollup
                    ) s2 
                    GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`物流方式`, s2.`旬`
                    HAVING s2.是否改派 <> '合计'
        ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                FIELD(s2.`币种`,'台湾','香港','合计'),
                s2.`年月`,
                FIELD(s2.`是否改派`,'改派','直发','合计'),
                FIELD(s2.`物流方式`,'台湾-大黄蜂普货头程-森鸿尾程','台湾-大黄蜂普货头程-易速配尾程','台湾-立邦普货头程-森鸿尾程','台湾-易速配-TW海快','台湾-立邦普货头程-易速配尾程',
                        '台湾-森鸿-新竹-自发头程','台湾-速派-711超商','台湾-速派-新竹','台湾-天马-新竹','台湾-天马-顺丰','台湾-天马-黑猫','台湾-易速配-新竹',
                        '香港-立邦-顺丰','香港-易速配-顺丰','香港-易速配-顺丰YC','香港-森鸿-SH渠道','香港-森鸿-顺丰渠道','合计'),   
                FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)
        listT.append(df11)

        # 父级分旬
        print('正在获取---父级分旬…………')
        sql12 = '''SELECT s2.家族,s2.币种,s2.年月,s2.父级分类,s2.旬,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,	
                    concat(ROUND(s2.签收 / s2.已完成 * 100,2),'%') as 完成签收,
                        concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') as 总计签收,
                        concat(ROUND(s2.已完成 / s2.总订单 * 100,2),'%') as 完成占比,
                        concat(ROUND(s2.已完成 / s2.已发货 * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(s2.已退货 / s2.总订单 * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
					concat(ROUND(s2.签收金额 / s2.完成金额 * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(s2.完成金额 / s2.总计金额 * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(s2.完成金额 / s2.发货金额 * 100,2),'%') as '已完成/已发货(金额)',
						concat(ROUND(s2.退货金额 / s2.总计金额 * 100,2),'%') as '退货率(金额)'
				 FROM ( SELECT  IFNULL(s1.币种,'合计') as 币种,IFNULL(s1.家族,'合计') as 家族,IFNULL(s1.年月,'合计') as 年月,IFNULL(s1.父级分类,'合计') as 父级分类,IFNULL(s1.旬,'合计') as 旬,
								SUM(s1.签收) as 签收,
								SUM(s1.拒收) as 拒收,
								SUM(s1.在途) as 在途,
								SUM(s1.未发货) as 未发货,
								SUM(s1.未上线) as 未上线,
								SUM(s1.已退货) as 已退货,
								SUM(s1.理赔) as 理赔,
								SUM(s1.自发头程丢件) as 自发头程丢件,
								SUM(s1.已发货) as 已发货,
								SUM(s1.已完成) as 已完成,
								SUM(s1.总订单) as 总订单,
                                s1.总订单量,s1.已发货单量,s1.已完成单量,
								SUM(s1.签收金额) as 签收金额,
								SUM(s1.退货金额) as 退货金额,
								SUM(s1.完成金额) as 完成金额,
								SUM(s1.发货金额) as 发货金额,
								SUM(s1.总计金额) as 总计金额
                        FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.父级分类 as 父级分类,IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
                                    SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                    SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                    SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                    SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                    SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                    SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                    SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                    SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                    count(订单编号) as 总订单,
                                    总订单量,已发货单量,已完成单量,
                                    count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                    SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                    SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                    SUM(`价格RMB`) as 总计金额,
                                    SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                                FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                ) cx
                                LEFT JOIN 
							        (SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                                    FROM (SELECT *,
                                                IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                            FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                        ) dg  
								        GROUP BY dg.币种,dg.家族,dg.年月
                                ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类`, cx.`旬`
                                ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`父级分类`, s1.`旬`
                        with rollup
                ) s2 HAVING s2.年月 <> '合计'
            ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                    FIELD(s2.`币种`,'台湾','香港','合计'),
                    s2.`年月`,
                    FIELD(s2.父级分类, '居家百货', '电子电器', '服饰', '医药保健',  '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','合计' ),
                    FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                    s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df12 = pd.read_sql_query(sql=sql12, con=self.engine1)
        listT.append(df12)
        # 二级分旬
        print('正在获取---二级分旬…………')
        sql13 = '''SELECT s2.家族,s2.币种,s2.年月,s2.父级分类,s2.二级分类,s2.旬,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,	
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
						concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
						concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
				 FROM ( SELECT  IFNULL(s1.币种,'合计') as 币种,IFNULL(s1.家族,'合计') as 家族,IFNULL(s1.年月,'合计') as 年月,IFNULL(s1.父级分类,'合计') as 父级分类,IFNULL(s1.二级分类,'合计') as 二级分类,IFNULL(s1.旬,'合计') as 旬,
								SUM(s1.签收) as 签收,
								SUM(s1.拒收) as 拒收,
								SUM(s1.在途) as 在途,
								SUM(s1.未发货) as 未发货,
								SUM(s1.未上线) as 未上线,
								SUM(s1.已退货) as 已退货,
								SUM(s1.理赔) as 理赔,
								SUM(s1.自发头程丢件) as 自发头程丢件,
								SUM(s1.已发货) as 已发货,
								SUM(s1.已完成) as 已完成,
								SUM(s1.总订单) as 总订单,
                                s1.总订单量,s1.已发货单量,s1.已完成单量,
								SUM(s1.签收金额) as 签收金额,
								SUM(s1.退货金额) as 退货金额,
								SUM(s1.完成金额) as 完成金额,
								SUM(s1.发货金额) as 发货金额,
								SUM(s1.总计金额) as 总计金额
                        FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.父级分类 as 父级分类,cx.二级分类 as 二级分类,IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
								总订单量,已发货单量,已完成单量,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                            ) cx
                            LEFT JOIN 
							    (SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                                FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                    ) dg  
								    GROUP BY dg.币种,dg.家族,dg.年月
                            ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                            GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类`, cx.`二级分类`, cx.`旬`
                            ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类`, cx.`二级分类` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`父级分类`, s1.`二级分类`, s1.`旬`
                        with rollup
                ) s2 HAVING s2.年月 <> '合计'
        ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                FIELD(s2.`币种`,'台湾','香港','合计'),
                s2.`年月`,
                FIELD(s2.父级分类, '居家百货', '电子电器', '服饰', '医药保健', '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','合计' ),
                FIELD(s2.二级分类,'个人洗护','皮鞋','日用百货','影音娱乐','家用电器','药品','上衣','下装'
                            ,'饰品','保健器械','保健食品','彩妆','钱包','休闲运动鞋','内衣','护理护具','凉/拖鞋'
                            ,'裙子','个护电器','配饰','护肤','布艺家纺','母婴用品','厨房用品','汽车用品','双肩包'
                            ,'单肩包','手机外设','电脑外设','成人保健','套装','靴子','手表手环','行李箱包','户外运动'
                            ,'玩具','手表','宠物用品','合计' ),
                FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df13 = pd.read_sql_query(sql=sql13, con=self.engine1)
        listT.append(df13)

        # 产品整月 台湾
        print('正在获取---产品整月 台湾…………')
        sql14 = '''SELECT *
                FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                            IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                            SUM(s1.已签收) as 已签收,
						    SUM(s1.拒收) as 拒收,
						    SUM(s1.已退货) as 已退货,
						    SUM(s1.已完成) as 已完成,
						    SUM(s1.总订单) as 总订单,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						    concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
						SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
						    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
						    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
						    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
						    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
						    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
						    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
						    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
						SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
						    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
						    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
						    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
						    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
						    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
						    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
						    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						    SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						    SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						    SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						    SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						    concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						    concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						    concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
						SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
						    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
						    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
						    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
						    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
						    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
						    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
						    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
						SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
						    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
						    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
						    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
						    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
						    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
						    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
						    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
						SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
						    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
						    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
						    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
						    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
						    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
						    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
						    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
						SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
						    SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
						    SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
						    SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
						    SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
						    concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
						    concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
						    concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
						SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
						    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
						    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
						    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
						    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
						    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
				    		concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
					    	concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
					    	concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
					    	concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
						SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
						    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
					    	SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
					    	SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
					    	SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
				    		concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
				    		concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
					    	concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
						SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
					    	SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
					    	SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
					    	SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
					    	SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
					    	concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
					    	concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
					    	concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
						SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
					    	SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
						    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
					    	SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
					        SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
					    	concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
					    	concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
					    	concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
						SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
					    	SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
					    	SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
					    	SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
					    	SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
					    	concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
					    	concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
					    	concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
						SUM(s1.龟山改派已签收) as '龟山改派已签收',
					    	SUM(s1.龟山改派拒收) as '龟山改派拒收',
					    	SUM(s1.龟山改派已退货) as '龟山改派已退货',
					    	SUM(s1.龟山改派已完成) as '龟山改派已完成',
					    	SUM(s1.龟山改派总订单) as '龟山改派总订单',
					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
					    	concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
					    	concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
					    	concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
				    	SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
					    	SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
					    	SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
					    	SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
					    	SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
					    	concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
					    	concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
					    	concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
						SUM(s1.速派改派已签收) as '速派改派已签收',
					    	SUM(s1.速派改派拒收) as '速派改派拒收',
					    	SUM(s1.速派改派已退货) as '速派改派已退货',
					    	SUM(s1.速派改派已完成) as '速派改派已完成',
					    	SUM(s1.速派改派总订单) as '速派改派总订单',
					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
					    	concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
					    	concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
					    	concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
						SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
					    	SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
					    	SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
					    	SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
					    	SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
					    	concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
					    	concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
					    	concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
						SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
					    	SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
					    	SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
					    	SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
					    	SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
					    	concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
					    	concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
					    	concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                    FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
				            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                            ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                        ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                        WITH ROLLUP 
                ) s HAVING s.月份 != '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df14 = pd.read_sql_query(sql=sql14, con=self.engine1)
        listT.append(df14)
        # 产品分旬 台湾
        print('正在获取---产品分旬 台湾…………')
        sql15 = '''SELECT *
                    FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
						SUM(s1.已签收) as 已签收,
						SUM(s1.拒收) as 拒收,
						SUM(s1.已退货) as 已退货,
						SUM(s1.已完成) as 已完成,
						SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
						SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
						SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
						SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
						SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
						concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
						concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
						concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
					SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
						SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
						SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
						SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
						SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
						concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
						concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
						concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
					SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
					SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
						SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
						SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
						SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
						SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
						concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
						concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
						concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
					SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
						SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
						SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
						SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
						SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
						concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
						concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
						concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
					SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
						SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
						SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
						SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
						SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
						concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
						concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
						concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
					SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
						SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
						SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
						SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
						SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
						concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
						concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
						concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
					SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
						SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
						SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
						SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
						SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
						concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
						concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
						concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
					SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
						SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
						SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
						SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
						SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
						concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
						concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
						concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
					SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
						SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
						SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
						SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
						SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
						concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
						concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
						concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
					SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
						SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
						SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
						SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
						SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
						concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
						concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
						concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
					SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
						SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
						SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
						SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
						SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
						concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
						concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
						concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
					SUM(s1.龟山改派已签收) as '龟山改派已签收',
						SUM(s1.龟山改派拒收) as '龟山改派拒收',
						SUM(s1.龟山改派已退货) as '龟山改派已退货',
						SUM(s1.龟山改派已完成) as '龟山改派已完成',
						SUM(s1.龟山改派总订单) as '龟山改派总订单',
						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
						concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
						concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
						concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
					SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
						SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
						SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
						SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
						SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
						concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
						concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
						concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
					SUM(s1.速派改派已签收) as '速派改派已签收',
						SUM(s1.速派改派拒收) as '速派改派拒收',
						SUM(s1.速派改派已退货) as '速派改派已退货',
						SUM(s1.速派改派已完成) as '速派改派已完成',
						SUM(s1.速派改派总订单) as '速派改派总订单',
						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
						concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
						concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
						concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
					SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
						SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
						SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
						SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
						SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
						concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
						concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
						concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
					SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
						SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
						SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
						SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
						SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
						concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
						concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
						concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
				        FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                        ) cx WHERE cx.`币种` = '台湾'
                    GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                    ) s1
                GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                WITH ROLLUP 
            ) s HAVING s.旬 != '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df15 = pd.read_sql_query(sql=sql15, con=self.engine1)
        listT.append(df15)

        # 产品整月 香港
        print('正在获取---产品整月 香港…………')
        sql16 = '''SELECT *
                    FROM(SELECT IFNULL(s1.家族, '合计') 家族,
                                IFNULL(s1.地区, '合计') 地区,
                                IFNULL(s1.月份, '合计') 月份,
                        IFNULL(s1.产品id, '合计') 产品id,
                        IFNULL(s1.产品名称, '合计') 产品名称,
                        IFNULL(s1.父级分类, '合计') 父级分类,
                        IFNULL(s1.二级分类, '合计') 二级分类,
						SUM(s1.已签收) as 已签收,
						SUM(s1.拒收) as 拒收,
						SUM(s1.已退货) as 已退货,
						SUM(s1.已完成) as 已完成,
						SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
				            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                            ) cx WHERE cx.`币种` = '香港'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                        ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.月份 != '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df16 = pd.read_sql_query(sql=sql16, con=self.engine1)
        listT.append(df16)
        # 产品分旬 香港
        print('正在获取---产品分旬 香港…………')
        sql17 = '''SELECT *
                    FROM(SELECT 
						IFNULL(s1.家族, '合计') 家族,
						IFNULL(s1.地区, '合计') 地区,
						IFNULL(s1.月份, '合计') 月份,
						IFNULL(s1.旬, '合计') 旬,
						IFNULL(s1.产品id, '合计') 产品id,
						IFNULL(s1.产品名称, '合计') 产品名称,
						IFNULL(s1.父级分类, '合计') 父级分类,
						IFNULL(s1.二级分类, '合计') 二级分类,
					SUM(s1.已签收) as 已签收,
						SUM(s1.拒收) as 拒收,
						SUM(s1.已退货) as 已退货,
						SUM(s1.已完成) as 已完成,
						SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
				        FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                        ) cx WHERE cx.`币种` = '香港'
                        GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                    ) s1
                    GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                    WITH ROLLUP 
            ) s HAVING s.旬 <> '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df17 = pd.read_sql_query(sql=sql17, con=self.engine1)
        listT.append(df17)

        # 产品整月_直发 台湾
        print('正在获取---产品整月_直发 台湾…………')
        sql18 = '''SELECT *
                        FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                                    IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                                    SUM(s1.已签收) as 已签收,
        						    SUM(s1.拒收) as 拒收,
        						    SUM(s1.已退货) as 已退货,
        						    SUM(s1.已完成) as 已完成,
        						    SUM(s1.总订单) as 总订单,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						            concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        						SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        						SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						        SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						            SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						            SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						            SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						            SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						            concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						            concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						            concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        						SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        						SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        						SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        						SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						    SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						    SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						    SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						    SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						    concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						    concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						    concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        						SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        				    		concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        					    	concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        					    	concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        					    	concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
        						SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        					    	SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        					    	SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        					    	SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        				    		concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        				    		concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        						SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        					    	SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        					    	SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        					    	SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        					    	SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        					    	concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        						SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        					    	SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        					    	SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        					        SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        					    	concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        					    	concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        						SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        					    	SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        					    	SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        					    	SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        					    	SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        					    	concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        					    	concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        						SUM(s1.龟山改派已签收) as '龟山改派已签收',
        					    	SUM(s1.龟山改派拒收) as '龟山改派拒收',
        					    	SUM(s1.龟山改派已退货) as '龟山改派已退货',
        					    	SUM(s1.龟山改派已完成) as '龟山改派已完成',
        					    	SUM(s1.龟山改派总订单) as '龟山改派总订单',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        					    	concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        					    	concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        					    	concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        				    	SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        					    	SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        					    	SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        					    	SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        					    	SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        					    	concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        					    	concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        						SUM(s1.速派改派已签收) as '速派改派已签收',
        					    	SUM(s1.速派改派拒收) as '速派改派拒收',
        					    	SUM(s1.速派改派已退货) as '速派改派已退货',
        					    	SUM(s1.速派改派已完成) as '速派改派已完成',
        					    	SUM(s1.速派改派总订单) as '速派改派总订单',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        					    	concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        					    	concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        					    	concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
        						SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        					    	SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        					    	SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        					    	SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        					    	SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        					    	concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        						SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        					    	SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        					    	SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        					    	SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        					    	SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
        				            FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`是否改派` = '直发' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                    ) cx WHERE cx.`币种` = '台湾'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                        ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df18 = pd.read_sql_query(sql=sql18, con=self.engine1)
        listT.append(df18)
        # 产品分旬_直发 台湾
        print('正在获取---产品分旬_直发 台湾…………')
        sql19 = '''SELECT *
                            FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,
        						SUM(s1.拒收) as 拒收,
        						SUM(s1.已退货) as 已退货,
        						SUM(s1.已完成) as 已完成,
        						SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						        concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        					SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        					SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						    SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						        SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						        SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						        SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						        SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						        concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						        concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						        concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        					SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        					SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        					SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        					SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        					SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        						concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        						concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        						concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
        					SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        						SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        						SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        						SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        						concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        						concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        						concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        					SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        						SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        						SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        						SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        						SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        						concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        						concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        						concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        					SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        						SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        						SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        						SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        						concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        						concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        						concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        					SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        						SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        						SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        						SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        						SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        						concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        						concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        						concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        					SUM(s1.龟山改派已签收) as '龟山改派已签收',
        						SUM(s1.龟山改派拒收) as '龟山改派拒收',
        						SUM(s1.龟山改派已退货) as '龟山改派已退货',
        						SUM(s1.龟山改派已完成) as '龟山改派已完成',
        						SUM(s1.龟山改派总订单) as '龟山改派总订单',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        						concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        						concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        						concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        					SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        						SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        						SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        						SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        						SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        						concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        						concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        						concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        					SUM(s1.速派改派已签收) as '速派改派已签收',
        						SUM(s1.速派改派拒收) as '速派改派拒收',
        						SUM(s1.速派改派已退货) as '速派改派已退货',
        						SUM(s1.速派改派已完成) as '速派改派已完成',
        						SUM(s1.速派改派总订单) as '速派改派总订单',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        						concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        						concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        						concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
        					SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        						SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        						SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        						SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        						SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        						concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        						concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        						concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        					SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        						SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        						SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        						SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        						SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        						concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        						concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                        FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
        				        FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                    FROM {0}_zqsb cc where  cc.`是否改派` = '直发' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.旬 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df19 = pd.read_sql_query(sql=sql19, con=self.engine1)
        listT.append(df19)

        # 产品整月_改派 台湾
        print('正在获取---产品整月_直发 台湾…………')
        sql20 = '''SELECT *
                        FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                                    IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                                    SUM(s1.已签收) as 已签收,
        						    SUM(s1.拒收) as 拒收,
        						    SUM(s1.已退货) as 已退货,
        						    SUM(s1.已完成) as 已完成,
        						    SUM(s1.总订单) as 总订单,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						            concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        						SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        						SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						        SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						            SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						            SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						            SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						            SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						            concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						            concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						            concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        						SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        						SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        						SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        						SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						    SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						    SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						    SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						    SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						    concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						    concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						    concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        						SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        				    		concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        					    	concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        					    	concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        					    	concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
        						SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        					    	SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        					    	SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        					    	SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        				    		concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        				    		concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        						SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        					    	SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        					    	SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        					    	SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        					    	SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        					    	concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        						SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        					    	SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        					    	SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        					        SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        					    	concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        					    	concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        						SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        					    	SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        					    	SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        					    	SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        					    	SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        					    	concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        					    	concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        						SUM(s1.龟山改派已签收) as '龟山改派已签收',
        					    	SUM(s1.龟山改派拒收) as '龟山改派拒收',
        					    	SUM(s1.龟山改派已退货) as '龟山改派已退货',
        					    	SUM(s1.龟山改派已完成) as '龟山改派已完成',
        					    	SUM(s1.龟山改派总订单) as '龟山改派总订单',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        					    	concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        					    	concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        					    	concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        				    	SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        					    	SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        					    	SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        					    	SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        					    	SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        					    	concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        					    	concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        						SUM(s1.速派改派已签收) as '速派改派已签收',
        					    	SUM(s1.速派改派拒收) as '速派改派拒收',
        					    	SUM(s1.速派改派已退货) as '速派改派已退货',
        					    	SUM(s1.速派改派已完成) as '速派改派已完成',
        					    	SUM(s1.速派改派总订单) as '速派改派总订单',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        					    	concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        					    	concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        					    	concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
        						SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        					    	SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        					    	SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        					    	SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        					    	SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        					    	concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        						SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        					    	SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        					    	SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        					    	SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        					    	SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
        				            FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`是否改派` = '改派' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                    ) cx WHERE cx.`币种` = '台湾'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                        ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df20 = pd.read_sql_query(sql=sql20, con=self.engine1)
        listT.append(df20)
        # 产品分旬_改派 台湾
        print('正在获取---产品分旬_直发 台湾…………')
        sql21 = '''SELECT *
                            FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,
        						SUM(s1.拒收) as 拒收,
        						SUM(s1.已退货) as 已退货,
        						SUM(s1.已完成) as 已完成,
        						SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						        concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        					SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        					SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						    SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						        SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						        SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						        SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						        SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						        concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						        concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						        concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        					SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        					SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        					SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        					SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        					SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        						concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        						concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        						concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
        					SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        						SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        						SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        						SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        						concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        						concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        						concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        					SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        						SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        						SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        						SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        						SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        						concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        						concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        						concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        					SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        						SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        						SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        						SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        						concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        						concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        						concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        					SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        						SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        						SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        						SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        						SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        						concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        						concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        						concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        					SUM(s1.龟山改派已签收) as '龟山改派已签收',
        						SUM(s1.龟山改派拒收) as '龟山改派拒收',
        						SUM(s1.龟山改派已退货) as '龟山改派已退货',
        						SUM(s1.龟山改派已完成) as '龟山改派已完成',
        						SUM(s1.龟山改派总订单) as '龟山改派总订单',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        						concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        						concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        						concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        					SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        						SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        						SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        						SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        						SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        						concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        						concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        						concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        					SUM(s1.速派改派已签收) as '速派改派已签收',
        						SUM(s1.速派改派拒收) as '速派改派拒收',
        						SUM(s1.速派改派已退货) as '速派改派已退货',
        						SUM(s1.速派改派已完成) as '速派改派已完成',
        						SUM(s1.速派改派总订单) as '速派改派总订单',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        						concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        						concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        						concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
        					SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        						SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        						SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        						SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        						SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        						concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        						concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        						concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        					SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        						SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        						SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        						SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        						SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        						concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        						concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                        FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
        				        FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-运营1组%","神龙运营1组",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",cc.团队))))))) as 家族 
                                    FROM {0}_zqsb cc where  cc.`是否改派` = '改派' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.旬 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','小虎队','神龙运营1组','Line运营','神龙主页运营','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df21 = pd.read_sql_query(sql=sql21, con=self.engine1)
        listT.append(df21)

        today = datetime.date.today().strftime('%Y.%m.%d')
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾', '产品整月香港', '产品分旬台湾', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾']
        print('正在将物流品类写入excel…………')
        file_path = 'G:\\输出文件\\{} {} 物流品类-签收率.xlsx'.format(today, match[team])
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        listT[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
        listT[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
        listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
        listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_总_品类_物流_两月签收率')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')

        print('正在将品类分旬写入excel…………')
        file_path = 'G:\\输出文件\\{} {} 品类分旬-签收率.xlsx'.format(today, match[team])
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾', '产品整月香港', '产品分旬台湾', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
        listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_品类直发分旬签收率')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')

        print('正在将产品写入excel…………')
        file_path = 'G:\\输出文件\\{} {} 产品明细-签收率.xlsx'.format(today, match[team])
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾', '产品分旬台湾', '产品整月香港', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾']
        df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        listT[4].to_excel(excel_writer=writer, sheet_name=sheet_name[4], index=False)
        listT[5].to_excel(excel_writer=writer, sheet_name=sheet_name[5], index=False)
        listT[6].to_excel(excel_writer=writer, sheet_name=sheet_name[6], index=False)
        listT[7].to_excel(excel_writer=writer, sheet_name=sheet_name[7], index=False)
        listT[8].to_excel(excel_writer=writer, sheet_name=sheet_name[8], index=False)
        listT[9].to_excel(excel_writer=writer, sheet_name=sheet_name[9], index=False)
        listT[10].to_excel(excel_writer=writer, sheet_name=sheet_name[10], index=False)
        listT[11].to_excel(excel_writer=writer, sheet_name=sheet_name[11], index=False)
        if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            del book['Sheet1']
        writer.save()
        writer.close()
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('D:/Users/Administrator/Desktop/新版-格式转换(工具表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_产品签收率_总')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')

if __name__ == '__main__':
    m = QueryUpdate()
    start: datetime = datetime.datetime.now()
    match1 = {'gat': '港台'}
    team = 'gat'
    '''  
        -----------------------------------------------手动导入状态运行（一）-----------------------------------------
        初始化配置>>> 
        1、dim_product： 切换：总产品- 不包含直发改派；分产品- 包含直发改派 ；
        2、write：       切换：本期- 本期最近两个月的数据 ； 本期并转存-本期最近两个月的数据的转存； 上期 -上期最近两个月的数据的转存
        3、last_time：   切换：更新上传时间；
    '''
    if team == 'gat':
        month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        month_old = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        # month_old = '2021-12-01'  # 获取-每日-报表 开始的时间
        month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
    else:
        month_last = '2022-06-01'
        month_old = '2022-06-01'        # 获取-每日-报表 开始的时间
        month_yesterday = '2022-07-31'

    last_time = '2021-01-01'
    write = '本期'
    m.readFormHost(team, write, last_time)                            # 更新签收表---港澳台（一）

    m.gat_new(team, month_last, month_yesterday)                  # 获取-签收率-报表、
    m.qsb_new(team, month_old)                                    # 获取-每日-报表
    m.EportOrderBook(team, month_last, month_yesterday)           # 导出-总的-签收
    m.phone_report('handle')                                      # 获取电话核实日报表 周报表 handle=手动 自定义时间

    # m.jushou()                                            #  拒收核实-查询需要的产品id
    # m.address_repot(team, month_last, month_yesterday)                       #  获取-地区签收率-报表

     # 停用备用使用
    # m.EportOrder(team)       #  导出需要更新的签收表
    # m.qsb_report(team, '2021-06-26', '2021-05-26')
    print('耗时：', datetime.datetime.now() - start)
    # win32api.MessageBox(0, "注意:>>>    程序运行结束， 请查看表  ！！！", "提 醒",win32con.MB_OK)