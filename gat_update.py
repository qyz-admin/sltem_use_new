import pandas as pd
import os, shutil
import datetime
import time
import xlwings
import win32api, win32con
import win32com.client
import requests
from os import *
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel

from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, \
    Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色
from 查询_产品明细 import QueryTwoT

from mysqlControl import MysqlControl
from sso_updata import Query_sso_updata

# -*- codinF:utf-8 -*-
class QueryUpdate(Settings):
    def __init__(self):
        Settings.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
        self.m = MysqlControl()
        self.dk = Settings_sso()    # 钉钉发送
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))

    # 获取签收表内容---港澳台更新签收总表(一)
    def readFormHost(self, team, write, last_time, up_time):
        start = datetime.datetime.now()
        path = r'F:\需要用到的文件\数据库'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                if '需发货的改派订单' in dir or '需发货改派订单' in dir:
                    write = '需发货'
                elif '订单检索' in dir or '导退货状态-临时' in dir:
                    write = '手动更新数据库'
                elif 'Payment_list' in dir or '港台线付退款' in dir or '拒付统计' in dir:
                    write = '在线支付'
                elif '线上支付重复订单' in dir:
                    write = '线付重复'
                self.wbsheetHost(filePath, team, write, last_time, up_time)
                os.remove(filePath)
                print('已清除上传文件…………')
        print('处理耗时：', datetime.datetime.now() - start)
    # 工作表的订单信息
    def wbsheetHost(self, filePath, team, write, last_time, up_time):
        match2 = {'slgat': '神龙港台',
                  'slgat_hfh': '火凤凰港台',
                  'slgat_hs': '红杉港台',
                  'slsc': '品牌',
                  'gat': '港台'}
        print('---正在获取 ' + match2[team] + ' 签收表的详情++++++')
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                db = None
                try:
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False, keep_default_na=False).value
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    if write == '本期':                 # 将数据库的临时表替换进指定的总表
                        print('++++正在导入更新：' + sht.name + ' 共：' + str(len(db)) + '行','sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        db.to_sql('gat_update', con=self.engine1, index=False, if_exists='replace')  # 将返回的dateFrame导入数据库的临时表
                        self.replacHost(team)
                    elif write == '上期':
                        print('++++正在导入更新：' + sht.name + ' 共：' + str(len(db)) + '行','sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        db.to_sql('gat_update', con=self.engine1, index=False, if_exists='replace')  # 将返回的dateFrame导入数据库的临时表
                        self.replaceHostbefore(team, last_time)
                    elif write == '需发货':
                        db = db[['订单编号']]
                        print('++++正在导入更新：' + sht.name + ' 共：' + str(len(db)) + '行','sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        db.to_sql('gat_update', con=self.engine1, index=False,if_exists='replace')  # 将返回的dateFrame导入数据库的临时表
                        self.online(team)
                    elif write == '手动更新数据库':
                        db = db[['订单编号', '运单号', '订单状态', '物流状态', '发货时间', '收货时间', '上线时间', '完成时间']]
                        print('++++正在 手动更新数据库：' + sht.name + ' 共：' + str(len(db)) + '行','sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        db.to_sql('gat_update', con=self.engine1, index=False,if_exists='replace')  # 将返回的dateFrame导入数据库的临时表
                        self.up_gat_ztl(team, up_time)
                    elif write == '地区签收率':
                        db = db[['订单编号', '省洲', '市区']]
                        print('++++正在 手动更新数据库：' + sht.name + ' 共：' + str(len(db)) + '行','sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        db.to_sql('gat_update', con=self.engine1, index=False,if_exists='replace')  # 将返回的dateFrame导入数据库的临时表
                    elif write == '在线支付':
                        pay = ''
                        if 'Payment_list' in filePath:
                            if '交易清单' in sht.name or 'Sheet' in sht.name:
                                pay = '交易清单'
                                db.rename(columns={'订单号': '订单编号', '退款金额': '交易退款金额'}, inplace=True)
                                db = db[['交易编号','订单编号', '交易币种', '交易金额', '交易状态', '交易创建时间', '订单创建时间', '交易退款金额', '支付方式']]
                                db.insert(0, 'id', None)
                                db.sort_values(by=["订单编号","交易创建时间"], axis=0, ascending=True)
                                print(db)
                                self._online_paly(pay, db)
                        elif '线付退款记录' in filePath:
                            if '港台' in sht.name:
                                pay = '线付退款记录'
                                db = db[['订单编号', '退款时间', '退款原因', '具体原因', '退款金额', '订单金额', '剩余金额', '申请退款人']]
                                print(db)
                                self._online_paly(pay, db)
                        elif '拒付统计' in filePath:
                            if '港台' in sht.name:
                                pay = '拒付统计'
                                print(db.columns)
                                db = db[['订单编号', '拒付时间']]
                                print(db)
                                self._online_paly(pay, db)
                    elif write == '线付重复':
                        if '明细' in sht.name:
                            # print(db)
                            # print(db.columns)
                            list = []
                            for lt in db.columns:
                                dt = lt.replace("\n", "")
                                lt_data = ''
                                if '订单编号' in lt and '此单是重复单' in lt:
                                    lt_data = '订单编号此单是重复单'
                                elif '是否联系处理' in lt:
                                    lt_data = '是否联系处理'
                                elif '备注' in lt and '处理明细' in lt:
                                    lt_data = '备注处理明细'
                                elif '转走日期' in lt:
                                    lt_data = '转走日期'
                                elif '同一客人' in lt and '是/否' in lt:
                                    lt_data = '同一客人是否'
                                elif '订单编号' in lt and '最近的上笔' in lt:
                                    lt_data = '订单编号最近的上笔'
                                else:
                                    lt_data = lt
                                list.append(lt_data)
                            # print(list)
                            db.columns = list
                            db.dropna(axis=0, how='any', inplace=True, subset=['订单编号此单是重复单'])  # 空值（缺失值），将空值所在的行/列删除后
                            print(db.columns)
                            # self.double_online_paly(db)

                    print('++++----->>>' + sht.name + '：订单更新完成++++')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()                                     # 工作表的订单信息

    # 更新-总表(地区签收率使用)
    def repHost(self, team):    # 更新-总表(地区签收率使用)
        try:
            print('正在更新总表中......')
            sql = '''update {0}_zqsb a, gat_update b
                            set a.`系统订单状态`= IF(b.`系统订单状态` = '', NULL, b.`系统订单状态`),
                                a.`系统物流状态`= IF(b.`系统物流状态` = '', NULL, b.`系统物流状态`),
                                a.`最终状态`= IF(b.`系统物流状态` = '', NULL, b.`系统物流状态`),
                                a.`价格`= IF(b.`价格` = '', NULL, b.`价格`),
                                a.`价格RMB`= IF(b.`价格RMB` = '', NULL, b.`价格RMB`)
                    where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')

    # 更新-总表（总体签收率使用）
    def replacHost(self, team):
        try:
            print('正在更新单表中......')
            sql = '''update {0}_order_list a, gat_update b
                                set a.`运单编号`= IF(b.`运单编号` = '', NULL, b.`运单编号`),
                                    a.`系统订单状态`= IF(b.`系统订单状态` = '', NULL, b.`系统订单状态`),
                                    a.`系统物流状态`= IF(b.`系统物流状态` = '', NULL, b.`系统物流状态`),
        		                    a.`是否改派`= IF(b.`是否改派` = '', NULL, b.`是否改派`),
        		                    a.`物流方式`= IF(b.`物流方式` = '', NULL, b.`物流方式`),
        		                    a.`物流渠道`= IF(b.`是否改派` ='直发',
                                                        IF(b.`物流方式` LIKE '香港-易速配-顺丰%','香港-易速配-顺丰', 
                                                            IF(b.`物流方式` LIKE '台湾-天马-711%' or b.`物流方式` LIKE '台湾-天马-新竹%','台湾-天马-新竹', 
                                                            IF(b.`物流方式` LIKE '台湾-铱熙无敌-新竹%' or b.`物流方式` LIKE '%优美宇通-新竹%','台湾-铱熙无敌-新竹', 
                                                            IF(b.`物流方式` LIKE '台湾-铱熙无敌-黑猫%','台湾-铱熙无敌-黑猫', 
                                                            IF(b.`物流方式` LIKE '台湾-铱熙无敌-711%','台湾-铱熙无敌-711超商', 
                                                            IF(b.`物流方式` LIKE '台湾-速派-新竹%','台湾-速派-新竹', 
                                                            IF(b.`物流方式` LIKE '香港-立邦-改派','香港-立邦-顺丰', 
                                                            IF(b.`物流方式` LIKE '香港-圆通-改派','香港-圆通', b.`物流方式`)))))) )),
                                                        IF(b.`物流方式` LIKE '香港-森鸿%','香港-森鸿-改派',
                                                            IF(b.`物流方式` LIKE '香港-立邦-顺丰%','香港-立邦-改派',
                                                            IF(b.`物流方式` LIKE '香港-易速配%','香港-易速配-改派',
                                                            IF(b.`物流方式` LIKE '台湾-立邦普货头程-森鸿尾程%' OR b.`物流方式` LIKE '台湾-大黄蜂普货头程-森鸿尾程%' OR b.`物流方式` LIKE '台湾-森鸿-新竹%','森鸿',
                                                            IF(b.`物流方式` LIKE '台湾-立邦普货头程-易速配尾程%' OR b.`物流方式` LIKE '台湾-大黄蜂普货头程-易速配尾程%','龟山',
                                                            IF(b.`物流方式` LIKE '台湾-易速配-龟山%' OR b.`物流方式` LIKE '台湾-易速配-新竹%' OR b.`物流方式` LIKE '新易速配-台湾-改派%' OR b.`物流方式` = '易速配','龟山',
                                                            IF(b.`物流方式` LIKE '台湾-天马-顺丰%','天马顺丰',
                                                            IF(b.`物流方式` LIKE '台湾-天马-新竹%' OR b.`物流方式` LIKE '台湾-天马-711%','天马新竹',
                                                            IF(b.`物流方式` LIKE '台湾-天马-黑猫%','天马黑猫',
                                                            IF(b.`物流方式` LIKE '台湾-速派-新竹%' OR b.`物流方式` LIKE '台湾-速派-711超商%','速派新竹',
                                                            IF(b.`物流方式` LIKE '台湾-速派宅配通%','速派宅配通',
                                                            IF(b.`物流方式` LIKE '台湾-速派-黑猫%','速派黑猫',
                                                            IF(b.`物流方式` LIKE '香港-圆通%','香港-圆通-改派',
                                                            IF(b.`物流方式` LIKE '台湾-优美宇通-新竹%','台湾-铱熙无敌-新竹改派',
                                                            IF(b.`物流方式` LIKE '台湾-铱熙无敌-黑猫普货' or b.`物流方式` LIKE '台湾-铱熙无敌-黑猫特货','台湾-铱熙无敌-黑猫改派',b.`物流方式`)))))))))))))))),
        		                    a.`物流名称`= IF(b.`物流名称` = '', NULL, b.`物流名称`),
        		                    a.`付款方式`= IF(b.`付款方式` = '', NULL, b.`付款方式`),
        		                    a.`产品id`= IF(b.`产品id` = '', NULL, b.`产品id`),
        		                    a.`产品名称`= IF(b.`产品名称` = '', NULL, b.`产品名称`),
        		                    a.`父级分类`= IF(b.`父级分类` = '', NULL, b.`父级分类`),
        		                    a.`二级分类`= IF(b.`二级分类` = '', NULL, b.`二级分类`)
        		                where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('正在更新总表中......')
            sql = '''update {0}_zqsb a, gat_update b
                                            set a.`运单编号`= IF(b.`运单编号` = '', NULL, b.`运单编号`),
                                                a.`系统订单状态`= IF(b.`系统订单状态` = '', NULL, b.`系统订单状态`),
                                                a.`系统物流状态`= IF(b.`系统物流状态` = '', NULL, b.`系统物流状态`),
                                                a.`最终状态`= IF(b.`最终状态` = '', NULL, b.`最终状态`),
                    		                    a.`是否改派`= IF(b.`是否改派` = '', NULL, b.`是否改派`),
                    		                    a.`物流方式`= IF(b.`物流方式` = '', NULL, b.`物流方式`),
                                                a.`物流渠道`= IF(b.`是否改派` ='直发',
                                                                IF(b.`物流方式` LIKE '香港-易速配-顺丰%','香港-易速配-顺丰', 
                                                                    IF(b.`物流方式` LIKE '台湾-天马-711%' or b.`物流方式` LIKE '台湾-天马-新竹%','台湾-天马-新竹', 
                                                                    IF(b.`物流方式` LIKE '台湾-铱熙无敌-新竹%' or b.`物流方式` LIKE '%优美宇通-新竹%','台湾-铱熙无敌-新竹', 
                                                                    IF(b.`物流方式` LIKE '台湾-铱熙无敌-黑猫%','台湾-铱熙无敌-黑猫', 
                                                                    IF(b.`物流方式` LIKE '台湾-铱熙无敌-711%','台湾-铱熙无敌-711超商', 
                                                                    IF(b.`物流方式` LIKE '台湾-速派-新竹%','台湾-速派-新竹', 
                                                                    IF(b.`物流方式` LIKE '香港-立邦-改派','香港-立邦-顺丰', 
                                                                    IF(b.`物流方式` LIKE '香港-圆通-改派','香港-圆通', b.`物流方式`)))))) )),
                                                                IF(b.`物流方式` LIKE '香港-森鸿%','香港-森鸿-改派',
                                                                    IF(b.`物流方式` LIKE '香港-立邦-顺丰%','香港-立邦-改派',
                                                                    IF(b.`物流方式` LIKE '香港-易速配%','香港-易速配-改派',
                                                                    IF(b.`物流方式` LIKE '台湾-立邦普货头程-森鸿尾程%' OR b.`物流方式` LIKE '台湾-大黄蜂普货头程-森鸿尾程%' OR b.`物流方式` LIKE '台湾-森鸿-新竹%','森鸿',
                                                                    IF(b.`物流方式` LIKE '台湾-立邦普货头程-易速配尾程%' OR b.`物流方式` LIKE '台湾-大黄蜂普货头程-易速配尾程%','龟山',
                                                                    IF(b.`物流方式` LIKE '台湾-易速配-龟山%' OR b.`物流方式` LIKE '台湾-易速配-新竹%' OR b.`物流方式` LIKE '新易速配-台湾-改派%' OR b.`物流方式` = '易速配','龟山',
                                                                    IF(b.`物流方式` LIKE '台湾-天马-顺丰%','天马顺丰',
                                                                    IF(b.`物流方式` LIKE '台湾-天马-新竹%' OR b.`物流方式` LIKE '台湾-天马-711%','天马新竹',
                                                                    IF(b.`物流方式` LIKE '台湾-天马-黑猫%','天马黑猫',
                                                                    IF(b.`物流方式` LIKE '台湾-速派-新竹%' OR b.`物流方式` LIKE '台湾-速派-711超商%','速派新竹',
                                                                    IF(b.`物流方式` LIKE '台湾-速派宅配通%','速派宅配通',
                                                                    IF(b.`物流方式` LIKE '台湾-速派-黑猫%','速派黑猫',
                                                                    IF(b.`物流方式` LIKE '香港-圆通%','香港-圆通-改派',
                                                                    IF(b.`物流方式` LIKE '台湾-优美宇通-新竹%','台湾-铱熙无敌-新竹改派',
                                                                    IF(b.`物流方式` LIKE '台湾-铱熙无敌-黑猫普货' or b.`物流方式` LIKE '台湾-铱熙无敌-黑猫特货','台湾-铱熙无敌-黑猫改派',b.`物流方式`)))))))))))))))),
                    		                    a.`物流名称`= IF(b.`物流名称` = '', NULL, b.`物流名称`),
                    		                    a.`付款方式`= IF(b.`付款方式` = '', NULL, b.`付款方式`),
                    		                    a.`产品id`= IF(b.`产品id` = '', NULL, b.`产品id`),
                    		                    a.`产品名称`= IF(b.`产品名称` = '', NULL, b.`产品名称`),
                    		                    a.`父级分类`= IF(b.`父级分类` = '', NULL, b.`父级分类`),
                    		                    a.`二级分类`= IF(b.`二级分类` = '', NULL, b.`二级分类`)
                    		                where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')

    # 更新-总表（修改已发货 使用）
    def online(self, team):
        try:
            print('正在更新单表中......')
            sql = '''update {0}_order_list a, gat_update b
                                set a.`系统订单状态`= '已发货'
        		                where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('正在更新总表中......')
            sql = '''update {0}_zqsb a, gat_update b
                                set a.`系统订单状态`= '已发货',
                                    a.`最终状态`= '在途'
                    		    where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')
    # 手动更新-总表（修改已完结 使用）
    def up_gat_ztl(self, team, up_time):
        try:
            print('正在更新单表中......')
            sql = '''update gat_order_list a, gat_update b
                        set a.`系统订单状态`= b.`订单状态`,
                            a.`系统物流状态`= b.`物流状态`,
                            a.`仓储扫描时间`= b.`发货时间`,
                            a.`完结状态时间`= b.`完成时间`,
                            a.`上线时间`= b.`上线时间`
                    where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('正在更新总表中......')
            sql = '''update gat_zqsb a, gat_update b
                        set a.`系统订单状态`= b.`订单状态`,
                            a.`系统物流状态`= b.`物流状态`,
                            a.`最终状态`= b.`物流状态`,
                            a.`仓储扫描时间`= b.`发货时间`,
                            a.`完结状态时间`= b.`完成时间`,
                            a.`上线时间`= b.`上线时间`
                    where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('正在更新历史总表中......')
            sql = '''update qsb_gat a, gat_update b
                        set a.`最终状态`= b.`物流状态`,
                            a.`仓储扫描时间`= b.`发货时间`,
                            a.`完结状态时间`= b.`完成时间`,
                            a.`上线时间`= b.`上线时间`
                    where a.`订单编号`= b.`订单编号` AND a.`记录时间` = '{0}';'''.format(up_time)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')


    # 在线支付情况
    def _online_paly(self, pay, db):
        print('正在写入......')
        db.to_sql('pay_cache', con=self.engine1, index=False, if_exists='replace')
        if pay == '交易清单':
            columns = list(db.columns)
            columns = ','.join(columns)
            sql = 'REPLACE INTO 交易清单({0}, 记录时间) SELECT *, NOW() 记录时间 FROM pay_cache x  WHERE x.交易编号 IS NOT NULL; '.format(columns)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        elif pay == '线付退款记录':
            columns = list(db.columns)
            columns = ','.join(columns)
            sql = 'REPLACE INTO 线付退款记录({0}, 记录时间) SELECT *, NOW() 记录时间 FROM pay_cache x  WHERE x.订单编号 IS NOT NULL; '.format(columns)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        elif pay == '拒付统计':
            columns = list(db.columns)
            columns = ','.join(columns)
            sql = 'REPLACE INTO 拒付统计({0}, 记录时间) SELECT 订单编号, left(拒付时间,LENGTH(拒付时间)-8) AS 拒付时间, NOW() 记录时间 FROM pay_cache x  WHERE x.订单编号 IS NOT NULL; '.format(columns)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)


        elif pay == '交易清单0':
            sql = '''SELECT 交易编号,订单编号, 交易币种, 交易金额, 交易状态, 交易创建时间, 订单创建时间, 交易退款金额, 支付方式
            订单编号, 交易币种, 交易金额, 交易状态, left(交易创建时间,LENGTH(交易创建时间)-8) AS 交易创建时间, 交易退款金额, 支付方式, 
                            NULL 退款时间, NULL 退款原因, NULL 详细原因, NULL 具体原因, IF(交易退款金额 IS NOT NULL AND 交易退款金额 <> '','已退款',NULL) 是否退款, NULL 退款金额, NULL 订单金额, NULL 剩余金额, NULL 申请退款人
                    FROM 线付缓存;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            df.to_sql('线付缓存_cache', con=self.engine1, index=False, if_exists='replace')
            columns = list(df.columns)
            columns = ','.join(columns)
            sql = 'REPLACE INTO 交易清单({0}, 记录时间) SELECT *, NOW() 记录时间 FROM 线付缓存_cache ORDER BY 订单编号, 交易创建时间; '.format(columns)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        elif pay == '线付退款记录0':
            sql = '''update 交易清单 a, 线付缓存 b
                    set a.`退款时间`= b.`退款时间`,
                        a.`退款原因`= b.`退款原因`,
                        a.`详细原因`= b.`详细原因` ,
                        a.`具体原因`= b.`具体原因`,
                        a.`退款金额`= b.`退款金额`,
                        a.`订单金额`= b.`订单金额`,
                        a.`剩余金额`= b.`剩余金额`,
                        a.`申请退款人`= b.`申请退款人`
                    where a.`订单编号`=b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('写入成功......')
    # 在线支付情况  写入
    def online_paly(self):
        print('+++单点更新订单状态中......')
        sso = Query_sso_updata('+86-17596568562', 'xhy123456', '4139', '77999c2203a632e8bd2a66d286b83c20', '手0动')
        time_online = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01  00:00:00'
        time_online2 = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y%m')
        sql = '''SELECT 订单编号  FROM {0} sl WHERE sl.`交易创建时间` >= '{1}';'''.format('交易清单', time_online)
        ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
        print(ordersDict['订单编号'][0])
        orderId = list(ordersDict['订单编号'])
        max_count = len(orderId)  # 使用len()获取列表的长度，上节学的
        n = 0
        while n < max_count:  # 这里用到了一个while循环，穿越过来的
            ord = ', '.join(orderId[n:n + 500])
            n = n + 500
            sso.orderInfoQuery(ord)
        print('更新耗时：', datetime.datetime.now() - start)

        print('正在获取 线付月数据......')
        sql = '''SELECT IFNULL(团队,'合计') 团队, 
                        IFNULL(币种,'合计') 币种, 
                        IFNULL(月份,'合计') 月份, 
                        IFNULL(支付成功,'合计') 支付成功, 
                        IF(取消退款=0,NULL,取消退款) 取消退款, 
                        IF(退货退款=0,NULL,退货退款) 退货退款, 
                        IF(拒付=0,NULL,拒付) 拒付, 
                        IF(支付失败=0,NULL,支付失败) 支付失败, 
                        IF(已创建=0,NULL,已创建) 已创建, 
                        总计, 
                        有效订单量, 
                        concat(ROUND(IFNULL(拒付 / (支付成功+取消退款+退货退款+拒付),0) * 100,2),'%') 拒付率, 
                        concat(ROUND(IFNULL((支付成功+取消退款+退货退款+拒付) / 总计,0) * 100,2),'%') 支付成功率, 
                        concat(ROUND(IFNULL(在线有效订单量 / 有效订单量,0) * 100,2),'%') 线付占比
                FROM ( SELECT 团队, 币种, 月份,
                                SUM(IF(交易最终状态 = "成功",1,0)) as 支付成功,
                                SUM(IF(交易最终状态 = "取消退款",1,0)) as 取消退款,
                                SUM(IF(交易最终状态 = "退货退款",1,0)) as 退货退款,
                                SUM(IF(交易最终状态 = "拒付",1,0)) as 拒付,
                                SUM(IF(交易最终状态 = "失败",1,0)) as 支付失败,
                                SUM(IF(交易最终状态 = "已创建",1,0)) as 已创建,
                                COUNT(订单编号) AS 总计,
                                SUM(IF(系统订单状态 NOT IN ('已删除','未支付', '支付失败'),1,0)) AS 有效订单量,
                                SUM(IF(系统订单状态 NOT IN ('已删除','未支付', '支付失败') AND 付款方式 NOT LIKE '%货到付款%',1,0)) AS 在线有效订单量
                    FROM ( SELECT s1.*,s2.`币种`, s2.`系统订单状态`, s2.`系统物流状态`,  s2.`付款方式`, s2.`团队`, 
                                   IF(ISNULL(退款原因), 交易状态,IF(退款原因 = '退款不取件',IF(退款金额 / 交易金额 >=0.2,'退货退款','退款不取件'),IF(系统订单状态 = '已删除', IF(退款原因 = "退货退款",'取消退款',退款原因), IF(退款原因 = "取消退款", '退货退款', 退款原因))))	AS 交易最终状态
                            FROM  ( SELECT *, DATE_FORMAT(交易创建时间,'%Y%m') AS 月份
                                    FROM 交易清单 
                                    WHERE id IN (SELECT MAX(id) FROM 交易清单 GROUP BY 订单编号 ) 
                                    ORDER BY id
                            ) s1 
                            LEFT JOIN (SELECT * FROM gat_order_list g WHERE g.年月 >= '202209') s2 ON s1.订单编号= s2.订单编号
                    ) ss
                GROUP BY 团队, 币种, 月份
                WITH ROLLUP
                ) s;'''
        sql = '''SELECT IFNULL(s.团队,'合计') 团队, 
                        IFNULL(s.币种,'合计') 币种, 
                        IFNULL(s.月份,'合计') 月份, 
                        IF(SUM(支付成功)=0,NULL,SUM(支付成功)) 支付成功, 
                        IF(SUM(取消退款)=0,NULL,SUM(取消退款)) 取消退款, 
                        IF(SUM(退货退款)=0,NULL,SUM(退货退款)) 退货退款, 
                        IF(SUM(拒付)=0,NULL,SUM(拒付)) 拒付, 
                        IF(SUM(支付失败)=0,NULL,SUM(支付失败)) 支付失败, 
                        IF(SUM(已创建)=0,NULL,SUM(已创建)) 已创建, 
                        SUM(总计) 总计,
                        SUM(有效订单量) 有效订单量, 
                        concat(ROUND(IFNULL(SUM(拒付) / (SUM(支付成功)+SUM(取消退款)+SUM(退货退款)+SUM(拒付)),0) * 100,2),'%') 拒付率, 
                        concat(ROUND(IFNULL((SUM(支付成功)+SUM(取消退款)+SUM(退货退款)+SUM(拒付)) / 总计,0) * 100,2),'%') 支付成功率, 
                        concat(ROUND(IFNULL(SUM(在线有效订单量) / SUM(有效订单量) ,0) * 100,2),'%') 线付占比
                FROM ( SELECT ss.团队, ss.币种, ss.月份,
                                SUM(IF(交易最终状态 = "成功",1,0)) as 支付成功,
                                SUM(IF(交易最终状态 = "取消退款",1,0)) as 取消退款,
                                SUM(IF(交易最终状态 = "退货退款",1,0)) as 退货退款,
                                SUM(IF(交易最终状态 = "拒付",1,0)) as 拒付,
                                SUM(IF(交易最终状态 = "失败",1,0)) as 支付失败,
                                SUM(IF(交易最终状态 = "已创建",1,0)) as 已创建,
                                COUNT(订单编号) AS 总计,
                                SUM(IF(系统订单状态 NOT IN ('已删除','未支付', '支付失败') AND 付款方式 NOT LIKE '%货到付款%',1,0)) AS 在线有效订单量
                    FROM ( SELECT s1.*,s2.`币种`, s2.`系统订单状态`, s2.`系统物流状态`,  s2.`付款方式`, s2.`团队`, 
                                   IF(ISNULL(退款原因), 交易状态,IF(退款原因 = '退款不取件',IF(退款金额 / 交易金额 >=0.2,'退货退款','退款不取件'),IF(系统订单状态 = '已删除', IF(退款原因 = "退货退款",'取消退款',退款原因), IF(退款原因 = "取消退款", '退货退款', 退款原因))))	AS 交易最终状态
                            FROM  ( SELECT *, DATE_FORMAT(交易创建时间,'%Y%m') AS 月份
                                    FROM 交易清单 
                                    WHERE id IN (SELECT MAX(id) FROM 交易清单 GROUP BY 订单编号 ) 
                                    ORDER BY id
                            ) s1 
                            LEFT JOIN (SELECT * FROM gat_order_list g WHERE g.年月 >= '{0}') s2 ON s1.订单编号= s2.订单编号
                    ) ss
                    GROUP BY 团队, 币种, 月份
                ) s
                LEFT JOIN (	SELECT 团队, 币种, 年月, SUM(IF(g.`系统订单状态` NOT IN ('已删除','未支付', '支付失败'),1,0)) AS 有效订单量
                            FROM gat_order_list g 
                            WHERE g.年月 >= '{0}' 
                            GROUP BY 团队, 币种, 年月
                ) ss2 ON s.团队 =ss2.团队 AND s.币种 =ss2.币种 AND s.月份 =ss2.年月
                GROUP BY s.团队, s.币种, s.月份
                WITH ROLLUP;'''.format(time_online2)
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)

        print('正在获取 线付每天数据......')
        sql = '''SELECT IFNULL(团队,'合计') 团队, 
                        IFNULL(币种,'合计') 币种, 
                        IFNULL(交易日期,'合计') 交易日期, 
                        IFNULL(支付成功,'合计') 支付成功, 
                        IF(取消退款=0,NULL,取消退款) 取消退款, 
                        IF(退货退款=0,NULL,退货退款) 退货退款, 
                        IF(拒付=0,NULL,拒付) 拒付, 
                        IF(支付失败=0,NULL,支付失败) 支付失败, 
                        IF(已创建=0,NULL,已创建) 已创建, 
                        总计, 
                        有效订单量, 
                        concat(ROUND(IFNULL(拒付 / (支付成功+取消退款+退货退款+拒付),0) * 100,2),'%') 拒付率, 
                        concat(ROUND(IFNULL((支付成功+取消退款+退货退款+拒付) / 总计,0) * 100,2),'%') 支付成功率, 
                        concat(ROUND(IFNULL(在线有效订单量 / 有效订单量,0) * 100,2),'%') 线付占比
                FROM ( SELECT 团队, 币种, 交易日期,
                                SUM(IF(交易最终状态 = "成功",1,0)) as 支付成功,
                                SUM(IF(交易最终状态 = "取消退款",1,0)) as 取消退款,
                                SUM(IF(交易最终状态 = "退货退款",1,0)) as 退货退款,
                                SUM(IF(交易最终状态 = "拒付",1,0)) as 拒付,
                                SUM(IF(交易最终状态 = "失败",1,0)) as 支付失败,
                                SUM(IF(交易最终状态 = "已创建",1,0)) as 已创建,
                                COUNT(订单编号) AS 总计,
                                SUM(IF(系统订单状态 NOT IN ('已删除','未支付', '支付失败'),1,0)) AS 有效订单量,
                                SUM(IF(系统订单状态 NOT IN ('已删除','未支付', '支付失败') AND 付款方式 NOT LIKE '%货到付款%',1,0)) AS 在线有效订单量
                    FROM ( SELECT s1.*,s2.`币种`, s2.`系统订单状态`, s2.`系统物流状态`,  s2.`付款方式`, s2.`团队`, 
                                   IF(ISNULL(退款原因), 交易状态,IF(退款原因 = '退款不取件',IF(退款金额 / 交易金额 >=0.2,'退货退款','退款不取件'),IF(系统订单状态 = '已删除', IF(退款原因 = "退货退款",'取消退款',退款原因), IF(退款原因 = "取消退款", '退货退款', 退款原因))))	AS 交易最终状态
                            FROM  ( SELECT *, DATE_FORMAT(交易创建时间,'%Y-%m-%d') AS 交易日期
                                    FROM 交易清单 
                                    WHERE id IN (SELECT MAX(id) FROM 交易清单 GROUP BY 订单编号 ) 
                                    ORDER BY id
                            ) s1 
                            LEFT JOIN (SELECT * FROM gat_order_list g WHERE g.年月 >= '202209') s2 ON s1.订单编号= s2.订单编号
                    ) ss
                GROUP BY 团队, 币种, 交易日期
                WITH ROLLUP
                ) s;'''
        sql = '''SELECT IFNULL(s.团队,'合计') 团队, 
                        IFNULL(s.币种,'合计') 币种, 
                        IFNULL(s.交易日期,'合计') 交易日期, 
                        IF(SUM(支付成功)=0,NULL,SUM(支付成功)) 支付成功, 
                        IF(SUM(取消退款)=0,NULL,SUM(取消退款)) 取消退款, 
                        IF(SUM(退货退款)=0,NULL,SUM(退货退款)) 退货退款, 
                        IF(SUM(拒付)=0,NULL,SUM(拒付)) 拒付, 
                        IF(SUM(支付失败)=0,NULL,SUM(支付失败)) 支付失败, 
                        IF(SUM(已创建)=0,NULL,SUM(已创建)) 已创建, 
                        SUM(总计) 总计,
                        SUM(有效订单量) 有效订单量, 
                        concat(ROUND(IFNULL(SUM(拒付) / (SUM(支付成功)+SUM(取消退款)+SUM(退货退款)+SUM(拒付)),0) * 100,2),'%') 拒付率, 
                        concat(ROUND(IFNULL((SUM(支付成功)+SUM(取消退款)+SUM(退货退款)+SUM(拒付)) / 总计,0) * 100,2),'%') 支付成功率, 
                        concat(ROUND(IFNULL(SUM(在线有效订单量) / SUM(有效订单量) ,0) * 100,2),'%') 线付占比
                FROM ( SELECT ss.团队, ss.币种, ss.交易日期,
                                SUM(IF(交易最终状态 = "成功",1,0)) as 支付成功,
                                SUM(IF(交易最终状态 = "取消退款",1,0)) as 取消退款,
                                SUM(IF(交易最终状态 = "退货退款",1,0)) as 退货退款,
                                SUM(IF(交易最终状态 = "拒付",1,0)) as 拒付,
                                SUM(IF(交易最终状态 = "失败",1,0)) as 支付失败,
                                SUM(IF(交易最终状态 = "已创建",1,0)) as 已创建,
                                COUNT(订单编号) AS 总计,
                                SUM(IF(系统订单状态 NOT IN ('已删除','未支付', '支付失败') AND 付款方式 NOT LIKE '%货到付款%',1,0)) AS 在线有效订单量
                    FROM ( SELECT s1.*,s2.`币种`, s2.`系统订单状态`, s2.`系统物流状态`,  s2.`付款方式`, s2.`团队`, 
                                   IF(ISNULL(退款原因), 交易状态,IF(退款原因 = '退款不取件',IF(退款金额 / 交易金额 >=0.2,'退货退款','退款不取件'),IF(系统订单状态 = '已删除', IF(退款原因 = "退货退款",'取消退款',退款原因), IF(退款原因 = "取消退款", '退货退款', 退款原因))))	AS 交易最终状态
                            FROM  ( SELECT *, DATE_FORMAT(交易创建时间,'%Y-%m-%d') AS 交易日期
                                    FROM 交易清单 
                                    WHERE id IN (SELECT MAX(id) FROM 交易清单 GROUP BY 订单编号 ) 
                                    ORDER BY id
                            ) s1 
                            LEFT JOIN (SELECT * FROM gat_order_list g WHERE g.年月 >= '{0}') s2 ON s1.订单编号= s2.订单编号
                    ) ss
                    GROUP BY 团队, 币种, 交易日期
                ) s
                LEFT JOIN (	SELECT 团队, 币种, 日期, SUM(IF(g.`系统订单状态` NOT IN ('已删除','未支付', '支付失败'),1,0)) AS 有效订单量
                            FROM gat_order_list g 
                            WHERE g.年月 >= '{0}' 
                            GROUP BY 团队, 币种, 日期
                ) ss2 ON s.团队 =ss2.团队 AND s.币种 =ss2.币种 AND s.交易日期 =ss2.日期
                GROUP BY s.团队, s.币种, s.交易日期
                WITH ROLLUP;'''.format(time_online2)
        df2 = pd.read_sql_query(sql=sql, con=self.engine1)

        today = datetime.date.today().strftime('%Y.%m.%d')
        file_path = 'F:\\输出文件\\\\{} 线付月数据.xlsx'.format(today)
        # writer2 = pd.ExcelWriter(file_path, engine='openpyxl')
        # df1.to_excel(writer2, index=False, startrow=1)     # 月数据
        # df2.to_excel(writer2, index=False, startcol=16)    # 天数据
        # writer2.save()
        # writer2.close()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer2:
            df1.to_excel(writer2, index=False, startrow=1)  # 月数据
            df2.to_excel(writer2, index=False, startcol=16)  # 天数据

        # wb = load_workbook(file_path)
        # sheet = wb.get_sheet_by_name("Sheet1")
        # sheet = wb["Sheet1"]
        # sheet.column_dimensions['A'].width = 15.82
        # sheet.column_dimensions['B'].width = 8.38
        # wb.save(file_path)
        print('写入成功......')

    # 线付重复订单  核实&签收率
    def double_online_paly(self, db):
        print('正在写入......')
        rq = datetime.datetime.now().strftime('%Y.%m.%d')
        # db.to_sql('线付缓存', con=self.engine1, index=False, if_exists='replace')  # 将返回的dateFrame导入数据库的临时表
        # columns = list(db.columns)
        # columns = ','.join(columns)
        # sql = 'REPLACE INTO 线付重复订单({0}, 记录时间) SELECT *, NOW() 记录时间 FROM 线付缓存; '.format(columns)
        # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        db.to_sql('线付重复订单', con=self.engine1, index=False, if_exists='replace')  # 将返回的dateFrame导入数据库的临时表
        listT = []
        sqltime82 = '''SELECT NULL AS '此单是重复订单', IFNULL(年月,'总计') as 年月,单量, 有效订单,
                            concat(ROUND(有效订单 / 单量 * 100,2),'%') AS 有效订单率,签收,拒收,完成,
                            concat(ROUND(签收 / 完成 * 100,2),'%') AS 完成签收,
                            concat(ROUND(完成 / 有效订单 * 100,2),'%') AS 完成占比
                        FROM (
                                SELECT 年月, COUNT(订单编号) AS 单量,
                                            SUM(IF(系统订单状态 NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货'),1,0)) AS 有效订单,
                                            SUM(IF(系统物流状态 = '已签收',1,0)) AS 签收,
                                            SUM(IF(系统物流状态 = '拒收',1,0)) AS 拒收,
                                            SUM(IF(系统物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) AS 完成
                                FROM (
                                        SELECT DISTINCT 订单编号此单是重复单
                                        FROM 线付重复订单 x1
                                ) x
                                LEFT JOIN (
                                            SELECT * 
                                            FROM gat_order_list s1 
                                            WHERE s1.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 6 MONTH),'%Y%m')
                                ) s ON x.订单编号此单是重复单 = s.订单编号
                                WHERE s.年月 IS NOT NULL AND s.日期 <= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                                GROUP BY 年月
                                WITH ROLLUP
                        ) ss ;'''
        df82 = pd.read_sql_query(sql=sqltime82, con=self.engine1)
        listT.append(df82)
        print('线付重复 此单订单的 签收率')

        sqltime83 = '''SELECT NULL AS '上笔订单', IFNULL(年月,'总计') as 年月,单量, 有效订单,
                            concat(ROUND(有效订单 / 单量 * 100,2),'%') AS 有效订单率,签收,拒收,完成,
                            concat(ROUND(签收 / 完成 * 100,2),'%') AS 完成签收,
                            concat(ROUND(完成 / 有效订单 * 100,2),'%') AS 完成占比
                        FROM (
                                SELECT 年月, COUNT(订单编号) AS 单量,
                                            SUM(IF(系统订单状态 NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货'),1,0)) AS 有效订单,
                                            SUM(IF(系统物流状态 = '已签收',1,0)) AS 签收,
                                            SUM(IF(系统物流状态 = '拒收',1,0)) AS 拒收,
                                            SUM(IF(系统物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) AS 完成
                                FROM (
                                            SELECT DISTINCT 订单编号最近的上笔
                                            FROM 线付重复订单 x2
                                            WHERE x2.订单编号最近的上笔 IS NOT NULL
                                ) x
                                LEFT JOIN (
                                                        SELECT * 
                                                        FROM gat_order_list s1 
                                                        WHERE s1.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 6 MONTH),'%Y%m')
                                ) s ON x.订单编号最近的上笔 = s.订单编号
                                WHERE s.年月 IS NOT NULL AND s.日期 <= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                                GROUP BY 年月
                                WITH ROLLUP
                        ) ss;'''
        df83 = pd.read_sql_query(sql=sqltime83, con=self.engine1)
        listT.append(df83)
        print('线付重复 上笔订单的 签收率')

        sqltime84 = '''SELECT NULL AS '合并订单', IFNULL(年月,'总计') as 年月,单量, 有效订单,
                            concat(ROUND(有效订单 / 单量 * 100,2),'%') AS 有效订单率, 签收,拒收,完成,
                            concat(ROUND(签收 / 完成 * 100,2),'%') AS 完成签收,
                            concat(ROUND(完成 / 有效订单 * 100,2),'%') AS 完成占比
                        FROM (
                                SELECT 年月, COUNT(订单编号) AS 单量,
                                            SUM(IF(系统订单状态 NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货'),1,0)) AS 有效订单,
                                            SUM(IF(系统物流状态 = '已签收',1,0)) AS 签收,
                                            SUM(IF(系统物流状态 = '拒收',1,0)) AS 拒收,
                                            SUM(IF(系统物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) AS 完成
                                FROM (
                                        (
                                            SELECT 订单编号此单是重复单
                                            FROM 线付重复订单 x1
                                        )
                                        UNION 
                                        (
                                            SELECT 订单编号最近的上笔
                                            FROM 线付重复订单 x2
                                            WHERE x2.订单编号最近的上笔 IS NOT NULL
                                        )
                                ) x
                                LEFT JOIN (
                                            SELECT * 
                                            FROM gat_order_list s1 
                                            WHERE s1.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 6 MONTH),'%Y%m')
                                ) s ON x.订单编号此单是重复单 = s.订单编号
                                WHERE s.年月 IS NOT NULL AND s.日期 <= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                                GROUP BY 年月
                                WITH ROLLUP
                        ) ss;'''
        df84 = pd.read_sql_query(sql=sqltime84, con=self.engine1)
        listT.append(df84)
        print('线付重复 此单和上笔 去除重复订单的 签收率')

        sqltime85 = '''SELECT * FROM 线付重复订单 x1;'''
        df85 = pd.read_sql_query(sql=sqltime85, con=self.engine1)
        listT.append(df85)
        print('线付重复 此单和上笔 去除重复订单的 签收率')

        sqltime86 = '''SELECT y.交易编号, y.订单编号, y.交易币种, y.交易金额, y.交易状态, y.交易创建时间, y.订单创建时间, y.交易退款金额, y.支付方式,
							j.拒付时间,
							t.退款时间, t.退款原因, t.具体原因, t.退款金额, t.订单金额, t.剩余金额, t.申请退款人, t.是否退款, g.团队, g.币种,g.年月,
							IF(交易状态 = 'CREATED','已创建', 
                            IF(交易状态 = 'FAILED','支付失败', 
                            IF(交易状态 = 'SUCCEEDED','支付成功', 
                            IF(交易状态 = 'FULLY REFUNDED','全额退款', 
                            IF(交易状态 = 'DISPUTED','拒付',
                            IF(交易状态 = 'PARTIALLY REFUNDED','部分退款',交易状态)))))) AS 中文交易状态
                    FROM 交易清单 y
                    LEFT JOIN 拒付统计 j ON y.订单编号 = j.订单编号
                    LEFT JOIN (SELECT *,  IF(退款金额/订单金额 < 0.2,'不退款','已退款') AS 是否退款 FROM 线付退款记录) t ON t.订单编号 = j.订单编号
                    LEFT JOIN (SELECT * FROM gat_order_list WHERE 年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 6 MONTH),'%Y%m')) g ON g.订单编号 = y.订单编号;'''
        df86 = pd.read_sql_query(sql=sqltime86, con=self.engine1)
        print('线付订单明细')
        df86.to_excel('F:\\输出文件\\{0} 线付订单明细.xlsx', sheet_name='查询', index=False, engine='xlsxwriter').format(rq)

        file_path = 'F:\\输出文件\\{} 线付重复订单签收率.xlsx'.format(rq)
        # df0 = pd.DataFrame([])
        # df0.to_excel(file_path, index=False)
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')
        # book = load_workbook(file_path)
        # writer.book = book
        # listT[0].to_excel(excel_writer=writer, sheet_name='签收率', index=False)
        # listT[1].to_excel(excel_writer=writer, sheet_name='签收率', index=False, startcol=10)  # 明细
        # listT[2].to_excel(excel_writer=writer, sheet_name='签收率', index=False, startcol=20)  # 有效单量
        # listT[3].to_excel(excel_writer=writer, sheet_name='线付明细', index=False)  # 有效单量
        # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            listT[0].to_excel(excel_writer=writer, sheet_name='签收率', index=False)
            listT[1].to_excel(excel_writer=writer, sheet_name='签收率', index=False, startcol=10)  # 明细
            listT[2].to_excel(excel_writer=writer, sheet_name='签收率', index=False, startcol=20)  # 有效单量
            listT[3].to_excel(excel_writer=writer, sheet_name='线付明细', index=False)  # 有效单量
        print('输出成功......')


    # 导出总签收表---修改 物流渠道 使用(一)
    def update_logistics_name(self):
        sql = '''UPDATE gat_zqsb d
                SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-易速配-顺丰%','香港-易速配-顺丰', 
                                IF(d.`物流方式` LIKE '台湾-天马-711%' or d.`物流方式` LIKE '台湾-天马-新竹%','台湾-天马-新竹', 
                                IF(d.`物流方式` LIKE '台湾-铱熙无敌-新竹%' or d.`物流方式` LIKE '%优美宇通-新竹%','台湾-铱熙无敌-新竹', 
                                IF(d.`物流方式` LIKE '台湾-铱熙无敌-黑猫%','台湾-铱熙无敌-黑猫', 
                                IF(d.`物流方式` LIKE '台湾-铱熙无敌-711%','台湾-铱熙无敌-711超商', 
                                IF(d.`物流方式` LIKE '台湾-速派-新竹%','台湾-速派-新竹', 
                                IF(d.`物流方式` LIKE '香港-立邦-改派','香港-立邦-顺丰', 
                                IF(d.`物流方式` LIKE '香港-圆通-改派','香港-圆通', d.`物流方式`)))))) ))
                WHERE d.`是否改派` ='直发';'''
        print('正在修改-直发的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''UPDATE gat_zqsb d
                SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-森鸿%','香港-森鸿-改派',
                                IF(d.`物流方式` LIKE '香港-立邦-顺丰%','香港-立邦-改派',
                                IF(d.`物流方式` LIKE '香港-易速配%','香港-易速配-改派',
                                IF(d.`物流方式` LIKE '台湾-立邦普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-大黄蜂普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-森鸿-新竹%','森鸿',
                                IF(d.`物流方式` LIKE '台湾-立邦普货头程-易速配尾程%' OR d.`物流方式` LIKE '台湾-大黄蜂普货头程-易速配尾程%','龟山',
                                IF(d.`物流方式` LIKE '台湾-易速配-龟山%' OR d.`物流方式` LIKE '台湾-易速配-新竹%' OR d.`物流方式` LIKE '新易速配-台湾-改派%' OR d.`物流方式` = '易速配','龟山',
                                IF(d.`物流方式` LIKE '台湾-天马-顺丰%','天马顺丰',
                                IF(d.`物流方式` LIKE '台湾-天马-新竹%' OR d.`物流方式` LIKE '台湾-天马-711%','天马新竹',
                                IF(d.`物流方式` LIKE '台湾-天马-黑猫%','天马黑猫',
                                IF(d.`物流方式` LIKE '台湾-速派-新竹%' OR d.`物流方式` LIKE '台湾-速派-711超商%','速派新竹',
                                IF(d.`物流方式` LIKE '台湾-速派宅配通%','速派宅配通',
                                IF(d.`物流方式` LIKE '台湾-速派-黑猫%','速派黑猫',
                                IF(d.`物流方式` LIKE '香港-圆通%','香港-圆通-改派',
                                IF(d.`物流方式` LIKE '台湾-优美宇通-新竹%','台湾-铱熙无敌-新竹改派',
                                IF(d.`物流方式` LIKE '台湾-铱熙无敌-黑猫普货' or d.`物流方式` LIKE '台湾-铱熙无敌-黑猫特货','台湾-铱熙无敌-黑猫改派',d.`物流方式`)))))))))))))))
                WHERE d.`是否改派` ='改派';'''
        print('正在修改-改派的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

    # 导出需要更新的签收表---港澳台(二)
    def EportOrder(self, team, month_last, month_yesterday, month_begin, check, export, handle, proxy_handle, proxy_id):
        match = {'gat': '港台', 'slsc': '品牌'}
        emailAdd = {'gat': 'giikinliujun@163.com', 'slsc': 'sunyaru@giikin.com'}
        today = datetime.date.today().strftime('%Y.%m.%d')
        # print('正在清 测试使用的 订单…………"GT230517100501MWLEK8","GT230517100520A6XW26","GT230517100534R0BU66","GT230517100551Q32PX8"')
        # sql = '''DELETE FROM gat_order_list WHERE gat_order_list.`订单编号` IN ("GT230517100501MWLEK8","GT230517100520A6XW26","GT230517100534R0BU66","GT230517100551Q32PX8");'''
        # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=20000)
        print('查询开始时间：', datetime.datetime.now())
        if check == '是':
            print('正在第一次检查父级分类为空的信息---')
            sql = '''SELECT 订单编号,商品id,dp.`product_id`, dp.`name` product_name, dp.third_cate_id, dc.`ppname` cate, dc.`pname` second_cate, dc.`name` third_cate
                    FROM (SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
                            FROM (SELECT id,日期,`订单编号`,`商品id`,`产品id`,`父级分类`,`系统订单状态`
								  FROM {0}_order_list
								  WHERE `日期` >= '{1}'
							) sl
                            WHERE (sl.`父级分类` IS NULL or sl.`父级分类`= '') AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'))
                         ) s
                    LEFT JOIN dim_product_gat dp ON  dp.product_id = s.`产品id`
                    LEFT JOIN dim_cate dc ON  dc.id = dp.third_cate_id;'''.format(team, month_begin)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            if df.empty:
                print('  第一次检查没有为空的………… ')
            else:
                print('正在更新父级分类的详情…………')
                df.to_sql('tem_product_id', con=self.engine1, index=False, if_exists='replace')
                sql = '''update {0}_order_list a, tem_product_id b
                            set a.`父级分类`= IF(b.`cate` = '', a.`父级分类`, b.`cate`),
                                a.`二级分类`= IF(b.`second_cate` = '', a.`二级分类`, b.`second_cate`),
                                a.`三级分类`= IF(b.`third_cate` = '', a.`三级分类`, b.`third_cate`)
                        where a.`订单编号`= b.`订单编号`;'''.format(team)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                print('更新完成+++')

                print('正在第二次检查父级分类为空的信息---')
                sql = '''SELECT 订单编号,商品id,dp.`product_id`, dp.`name` product_name, dp.third_cate_id, dc.`ppname` cate, dc.`pname` second_cate, dc.`name` third_cate
                        FROM (SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
                                FROM (SELECT id,日期,`订单编号`,`商品id`,`产品id`,`父级分类`,`系统订单状态`
									  FROM {0}_order_list
								      WHERE `日期` >= '{1}'
								) sl
                                WHERE (sl.`父级分类` IS NULL or sl.`父级分类`= '') AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'))
                             ) s
                        LEFT JOIN dim_product_gat dp ON  dp.product_id = s.`产品id`
                        LEFT JOIN (SELECT * FROM dim_cate GROUP BY pid ) dc ON  dc.pid = dp.second_cate_id;'''.format(team, month_begin)
                df = pd.read_sql_query(sql=sql, con=self.engine1)
                if df.empty:
                    print('  第二次检查没有为空的………… ')
                else:
                    print('正在更新父级分类的详情…………')
                    sql = '''update {0}_order_list a, tem_product_id b
                                set a.`父级分类`= IF(b.`cate` = '', a.`父级分类`, b.`cate`),
                                    a.`二级分类`= IF(b.`second_cate` = '', a.`二级分类`, b.`second_cate`),
                                    a.`三级分类`= IF(b.`third_cate` = '', a.`三级分类`, b.`third_cate`)
                            where a.`订单编号`= b.`订单编号`;'''.format(team)
                    pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                    print('更新完成+++')

            print('正在第一次检查产品id为空的信息---')
            sql = '''SELECT 订单编号,商品id,dp.product_id, dp.`name` product_name, dp.third_cate_id
                    FROM (SELECT id,日期,`订单编号`,`商品id`,sl.`产品id`
                        FROM (SELECT id,日期,`订单编号`,`商品id`,`产品id`,`产品名称`,`父级分类`,`系统订单状态`
								FROM {0}_order_list
								WHERE `日期` >= '{1}'
						) sl
                        WHERE (sl.`产品名称` IS NULL or sl.`产品名称`= '') AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'))
                    ) s
                    LEFT JOIN dim_product_gat dp ON dp.product_id = s.`产品id`;'''.format(team, month_begin)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            if df.empty:
                print('  第一次检查没有为空的………… ')
            else:
                print('正在更新产品详情…………')
                df.to_sql('tem_product_id', con=self.engine1, index=False, if_exists='replace')
                sql = '''update {0}_order_list a, tem_product_id b
                            set a.`产品id`= IF(b.`product_id` = '',a.`产品id`, b.`product_id`),
                                a.`产品名称`= IF(b.`product_name` = '',a.`产品名称`, b.`product_name`)
                    where a.`订单编号`= b.`订单编号`;'''.format(team)
                pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                print('更新完成+++')

            print('正在综合检查 父级分类、产品id 为空的信息---')
            sql = '''SELECT id,日期,`订单编号`,`商品id`,`产品id`,`产品名称`,`父级分类`,`二级分类`,`三级分类`
                    FROM (SELECT id,日期,`订单编号`,`商品id`,`产品id`,`产品名称`,`父级分类`,`二级分类`,`三级分类`,`系统订单状态`
							FROM {0}_order_list 
							WHERE `日期` >= '{1}'
					) sl
                    WHERE (sl.`父级分类` IS NULL or sl.`父级分类`= '' OR sl.`产品名称` IS NULL or sl.`产品名称`= '')
                        AND ( NOT sl.`系统订单状态` IN ('已删除', '问题订单', '支付失败', '未支付'));'''.format(team, month_begin)
            data = pd.read_sql_query(sql=sql, con=self.engine1)
            data.to_sql('tem_product_cp', con=self.engine1, index=False, if_exists='replace')

            sql = '''SELECT DISTINCT 产品id FROM tem_product_cp;'''.format(team, month_begin)
            ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
            if ordersDict.empty:
                print(' ****** 没有要补充的信息; ****** ')
            else:
                print('！！！ 请再次补充缺少的数据中！！！')
                login_TmpCode = 'login_TmpCode'
                if handle == '手动':
                    print('请输入口令Token:  回车确认')
                    login_TmpCode = str(input())
                lw = QueryTwoT('+86-17596568562', 'xhy123456', login_TmpCode, handle, proxy_handle, proxy_id)
                lw.productInfo('gat_order_list', ordersDict, proxy_handle, proxy_id)
        print('查询结束时间：', datetime.datetime.now())
        if team in ('gat'):
            del_time = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y%m')
            sql = '''DELETE FROM gat_zqsb gt
                    WHERE gt.年月 >= {0}
                      and gt.`订单编号` IN (SELECT 订单编号 
                                        FROM gat_order_list gs
                                        WHERE gs.年月 >= {0} and gs.`系统订单状态` NOT IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'));'''.format(del_time)
            print('正在清除港澳台-总表的可能删除了的订单…………')
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=20000)
            print('正在获取---' + match[team] + '---更新数据内容…………')
            sql = '''SELECT 年月, 旬, 日期, 团队, 所属团队, 币种, null 区域, 订单来源, a.订单编号, 电话号码, a.运单编号,
                            -- IF(ISNULL(a.仓储扫描时间), IF(出货时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), null, 出货时间), a.仓储扫描时间) 出货时间,
                            IF(出货时间='1990-01-01 00:00:00' or 出货时间='1899-12-29 00:00:00' or 出货时间='1899-12-30 00:00:00' or 出货时间='0000-00-00 00:00:00', a.仓储扫描时间, 出货时间) 出货时间,
                            IF(ISNULL(c.标准物流状态), b.物流状态, c.标准物流状态) 物流状态, c.`物流状态代码` 物流状态代码,
                            IF(状态时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), '', 状态时间) 状态时间,
                            IF(ISNULL(a.上线时间), IF(b.上线时间 in ('1990-01-01 00:00:00','1899-12-29 00:00:00','1899-12-30 00:00:00','0000-00-00 00:00:00'), null,b.上线时间), a.上线时间) 上线时间, 系统订单状态,
                            IF(ISNULL(d.订单编号), 系统物流状态, '已退货') 系统物流状态,
                            IF(ISNULL(d.订单编号), NULL, '已退货') 退货登记,
                            IF(ISNULL(d.订单编号), IF(ISNULL(系统物流状态), IF(ISNULL(c.标准物流状态) OR c.标准物流状态 = '未上线', IF(系统订单状态 IN ('已转采购', '待发货'), '未发货', '未上线') , IF(物流方式 like '%天马%' and c.签收表物流状态 = '在途','未上线', c.标准物流状态)), 系统物流状态), '已退货') 最终状态,
                            IF(是否改派='二次改派', '改派', 是否改派) 是否改派,
                            物流方式,物流渠道,物流名称,null 运输方式,null 货物类型,是否低价,付款方式,产品id,产品名称,父级分类, 二级分类,三级分类, 下单时间,审核时间,仓储扫描时间,完结状态时间,价格,价格RMB, null 价格区间, 包裹重量, null 包裹体积,null 邮编, 
                            IF(ISNULL(b.运单编号), '否', '是') 签收表是否存在, null 签收表订单编号, null 签收表运单编号, null 原运单号, b.物流状态 签收表物流状态, null 添加时间, null 成本价, null 物流花费, null 打包花费, null 其它花费, 添加物流单号时间,
                            省洲,市区,数量, a.下架时间, a.物流提货时间, a.完结状态, a.回款时间, a.支付类型, a.是否盲盒, a.克隆类型, a.主订单, a.发货仓库
                        FROM (SELECT * 
                            FROM {0}_order_list g
                            WHERE g.日期 >= '{2}' AND g.日期 <= '{3}' AND g.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)') AND g.物流方式 <> '盲盒专用物流渠道'
                        ) a
                        LEFT JOIN gat_wl_data b ON a.`查件单号` = b.`运单编号`
                        LEFT JOIN {0}_logisitis_match c ON b.物流状态 = c.签收表物流状态
                        LEFT JOIN (SELECT 订单编号 FROM {0}_return r WHERE r.`订单编号` IS NOT NULL AND r.`订单编号` <> "") d ON a.订单编号 = d.订单编号
                        ORDER BY a.`下单时间`;'''.format(team, month_begin, month_last, month_yesterday)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print('正在写入---' + match[team] + ' ---临时缓存…………')  # 备用临时缓存表
            df.to_sql('d1_{0}'.format(team), con=self.engine1, index=False, if_exists='replace', chunksize=20000)
            if export == '导表':
                print('正在写入excel…………')
                df = df[['日期', '团队', '所属团队', '币种', '订单编号', '电话号码', '运单编号', '出货时间', '物流状态', '物流状态代码', '状态时间', '上线时间','系统订单状态', '系统物流状态', '最终状态',
                         '是否改派', '物流方式', '物流渠道', '物流名称', '签收表物流状态', '付款方式', '产品id', '产品名称', '父级分类', '二级分类', '下单时间','审核时间', '仓储扫描时间', '完结状态时间', '发货仓库']]
                old_path = 'F:\\输出文件\\{} {} 更新-签收表.xlsx'.format(today, match[team])
                df.to_excel(old_path, sheet_name=match[team], index=False)
                new_path = "F:\\神龙签收率\\" + (datetime.datetime.now()).strftime('%m.%d') + '\\{} {} 更新-签收表.xlsx'.format(today, match[team])
                shutil.copyfile(old_path, new_path)  # copy到指定位置
                print('----已写入excel; 并复制到指定文件夹中')
            else:
                print('不 写入excel…………')
        print('查询开始时间：', datetime.datetime.now())
        sql = '''DELETE FROM d1_gat gt WHERE gt.`订单编号` IN (SELECT 订单编号 FROM gat_易速配退运);'''
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=20000)
        print('已清除不参与计算的易速配退运订单（总表中）…………')

        print('查询开始时间：', datetime.datetime.now())
        print('正在写入' + match[team] + ' 全部签收表中…………')
        sql = 'REPLACE INTO {0}_zqsb SELECT *, NOW() 更新时间 FROM d1_{0};'.format(team)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=20000)
        try:
            print('正在 回滚更新订单总表......')
            sql = '''update {0}_order_list a, d1_gat b
                            set a.`系统物流状态`= IF(b.`最终状态` = '', NULL, b.`最终状态`),
                                a.`上线时间`= IF(a.`上线时间` = '0000-00-00 00:00:00' OR a.`上线时间` IS NULL, IF(b.`上线时间` = '0000-00-00 00:00:00', NULL, b.`上线时间`), a.`上线时间`)
                    where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=20000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))

        print('查询开始时间：', datetime.datetime.now())
        sql = '''DELETE FROM gat_zqsb gz 
                 WHERE gz.`系统订单状态` = '已转采购' and gz.`是否改派` = '改派'
                   and gz.`审核时间` >= '{0} 00:00:00' AND gz.`日期` >= '{1}';'''.format(month_yesterday, month_last)
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        print('已清除不参与计算的今日改派订单…………')

        print('查询开始时间：', datetime.datetime.now())
        print('正在检查总表 订单重量异常 信息 ......')
        sql = '''SELECT 订单编号, 币种, 下单时间, 电话号码, 产品id, 产品名称, 包裹重量
                FROM {0} s 
                WHERE s.包裹重量 > 5000 AND s.`订单编号` not in (SELECT 订单编号 FROM {1});'''.format('d1_gat', '订单重量异常')
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        if df is not None and len(df) > 0:
            df.to_sql('d1_cpy', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO {0}(订单编号, 币种, 下单时间, 电话号码, 产品id, 产品名称, 记录时间) 
                               SELECT 订单编号, 币种, 下单时间, 电话号码, 产品id, 产品名称, NOW() 添加时间 FROM d1_cpy; '''.format('订单重量异常')
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

            orderId = list(df['订单编号'])
            orderId = ','.join(orderId)
            url = "https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e"  # url为机器人的webhook  个人 小海
            # # url = "https://oapi.dingtalk.com/robot/send?access_token=fa74c55267674d9281f705b6fde624818c9977287cb590891ef2691714a9ceda"  # url为机器人的webhook  审单问题群
            content = r"订单重量异常, 请核实；" + orderId  # 钉钉消息内容，注意test是自定义的关键字，需要在钉钉机器人设置中添加，这样才能接收到消息
            mobile_list = ['18538110674']  # 要@的人的手机号，可以是多个，注意：钉钉机器人设置中需要添加这些人，否则不会接收到消息
            isAtAll = '单个'  # 是、 否、 单个、 @所有人
            self.dk.send_dingtalk_message(url, content, mobile_list, isAtAll)
            print('订单重量异常 信息 已发送 请注意查看......')
        else:
            print('无 订单重量异常 信息！！！')
        print('查询开始时间：', datetime.datetime.now())

    # 导出总的签收表---各家族-港澳台(三)
    def EportOrderBook(self, team, month_last, month_yesterday):
        today = datetime.date.today().strftime('%Y.%m.%d')
        match = {'gat': '港台','slsc': '品牌'}
        day = datetime.datetime.now().strftime('%d')
        if int(day) < 10:
            print('****** 当前日期小于10, 为本月上旬 --- 起止时间：' + month_last + ' - ' + month_yesterday + ' ******')
            month_last = (datetime.datetime.now() - relativedelta(months=2)).strftime('%Y-%m') + '-01'
        else:
            print('****** 当前日期大于10, 为本月中下旬 --- 起止时间：' + month_last + ' - ' + month_yesterday + ' ******')
            month_last = month_last
        print('****** 转存数据      起止时间：' + month_last + ' - ' + month_yesterday + ' ******')
        # print('正在修改-港澳台-物流渠道…………')
        # self.update_logistics_name()

        print('正在获取---' + match[team] + ' ---全部数据内容…………')
        sql = '''SELECT * FROM {0}_zqsb a WHERE a.日期 >= '{1}' AND a.日期 <= '{2}' ORDER BY a.`下单时间`;'''.format(team, month_last, month_yesterday)     # 港台查询函数导出
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        print('正在写入---' + match[team] + ' ---临时缓存…………')             # 备用临时缓存表
        df.to_sql('d1_{0}'.format(team), con=self.engine1, index=False, if_exists='replace', chunksize=10000)

        sql = '''SELECT DISTINCT 所属团队 FROM d1_gat;'''.format(team)     # 港台- 获取导出 团队名称
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        tem_name = list(df['所属团队'])
        for tem in tem_name:
            sql = '''SELECT * FROM d1_{0} sl WHERE sl.`所属团队`in ("{1}");'''.format(team, tem)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            old_path = 'F:\\输出文件\\{} {}签收表.xlsx'.format(today, tem)
            df.to_excel(old_path, sheet_name=tem, index=False)
            new_path = "F:\\神龙签收率\\" + (datetime.datetime.now()).strftime('%m.%d') + '\\{} {}签收表.xlsx'.format(today, tem)
            shutil.copyfile(old_path, new_path)     # copy到指定位置
            print(tem + '----已写入excel; 并复制到指定文件夹中')
        try:
            print('正在转存中' + month_last + ' - ' + month_yesterday + '的订单数据......')
            sql = '''SELECT 年月, 旬, 日期, 团队,所属团队,币种, 订单来源, 订单编号, 出货时间, IF(`状态时间` = '',NULL,状态时间) as 状态时间, 上线时间, 最终状态,是否改派,物流方式,物流渠道,
                            产品id,父级分类,二级分类,三级分类,下单时间, 审核时间,仓储扫描时间,下架时间, 物流提货时间, 完结状态, 完结状态时间,回款时间, 价格RMB, 运单编号, curdate() 记录时间
                    FROM d1_{0} a WHERE a.`运单编号` is not null ;'''.format(team)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print('正在添加缓存中......')
            df.to_sql('gat_update_cp', con=self.engine1, index=False, if_exists='replace')
            print('正在转存数据中......')
            sql = '''REPLACE INTO qsb_{0} SELECT * FROM gat_update_cp; '''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('转存失败：', str(Exception) + str(e))
        print('转存成功…………')

        print('正在获取预估签收率的数据......')
        week: datetime = datetime.datetime.now()
        if week.isoweekday() == 0 or week.isoweekday() == 0:
            time_path: datetime = datetime.datetime.now()
            mkpath = "F:\\神龙签收率\\A预估签收率\\" + time_path.strftime('%m.%d')
            isExists = os.path.exists(mkpath)
            if not isExists:
                os.makedirs(mkpath)
            else:
                print(mkpath + ' 目录已存在')
            file_path = mkpath + '\\{} 预测_产品签收率_使用版.xlsx'.format(time_path.strftime('%m.%d'))
            sql = '''SELECT 所属团队 AS 家族, 币种, 产品id, 产品名称, 
			                concat(ROUND(产品金额团队占比* 100,2),'%') AS 产品金额团队占比, 最近3天单量,
			                concat(ROUND(IF(历史平均 = 0 OR 历史平均 IS NULL,预测, 历史平均)* 100,2),'%') AS '预测签收',
			                concat(ROUND(IF(历史平均 = 0 OR 历史平均 IS NULL,产品金额团队占比 * 预测影响, 产品金额团队占比 * 历史平均影响)* 100,2),'%') AS '预测签收影响',
			                IF(历史平均 = 0 OR 历史平均 IS NULL,'预测取值', '历史取值') AS 取值,
			                concat(ROUND(目标签收率* 100,2),'%') AS 目标签收率
                    FROM(SELECT s.所属团队, s.币种, s.产品id, s.产品名称, 
			                    总金额 / 团队金额 AS 产品金额团队占比,
			                    最近单量 AS 最近3天单量,
			                    avg_sign_rate AS 历史平均,
			                    IF(s.团队 = '神龙港台' AND s.币种 = '台湾',(avg_sign_rate-0.825),
			                    IF(s.团队 = '神龙港台' AND s.币种 = '香港',(avg_sign_rate-0.89) ,
			                    IF(s.团队 = '火凤凰港台' AND s.币种 = '台湾',(avg_sign_rate-0.87),
			                    IF(s.团队 = '火凤凰港台' AND s.币种 = '香港',(avg_sign_rate-0.89), 
			                    IF(s.团队 = '神龙港台' AND s.币种 = '台湾',(avg_sign_rate-0.86), 
			                    IF(s.团队 = '神龙港台' AND s.币种 = '香港',(avg_sign_rate-0.88), NULL)))))) AS 历史平均影响, 
			                    sign_rate AS 预测, 
			                    IF(s.团队 = '神龙港台' AND s.币种 = '台湾',(sign_rate-0.825),
			                    IF(s.团队 = '神龙港台' AND s.币种 = '香港',(sign_rate-0.89) ,
			                    IF(s.团队 = '火凤凰港台' AND s.币种 = '台湾',(sign_rate-0.87),
			                    IF(s.团队 = '火凤凰港台' AND s.币种 = '香港',(sign_rate-0.89), 
			                    IF(s.团队 = '神龙港台' AND s.币种 = '台湾',(sign_rate-0.86), 
			                    IF(s.团队 = '神龙港台' AND s.币种 = '香港',(sign_rate-0.88), NULL)))))) AS 预测影响,
			                    IF(s.团队 = '神龙港台' AND s.币种 = '台湾',0.825,
			                    IF(s.团队 = '神龙港台' AND s.币种 = '香港',0.89,
			                    IF(s.团队 = '火凤凰港台' AND s.币种 = '台湾',0.87,
			                    IF(s.团队 = '火凤凰港台' AND s.币种 = '香港',0.89, 
			                    IF(s.团队 = '神龙港台' AND s.币种 = '台湾',0.86, 
			                    IF(s.团队 = '神龙港台' AND s.币种 = '香港',0.88, NULL)))))) AS 目标签收率
                        FROM (SELECT cc.`所属团队`, cc.`币种`, cc.`产品id`, cc.`产品名称`, cc.`三级分类`,
						            COUNT(订单编号) AS 本月单量,
						            SUM(`价格RMB`) AS 总金额,
						            SUM(IF(最终状态 = "已签收",价格RMB,0)) as 签收金额,
						            SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
						            SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),价格RMB,0)) as 完成金额						
			                FROM gat_zqsb cc
			                where cc.年月 = DATE_FORMAT(curdate(),'%Y%m')
			                GROUP BY cc.`所属团队`, cc.`币种`, cc.`产品id`
                        ) s
                        LEFT JOIN 
                        (   SELECT cc.`所属团队`, cc.`币种`, SUM(`价格RMB`) AS 团队金额
	                        FROM gat_zqsb cc
                            where cc.年月 = DATE_FORMAT(curdate(),'%Y%m')
	                        GROUP BY cc.`所属团队`, cc.`币种`
                        ) s1  ON s.所属团队 = s1.所属团队 AND s.币种 = s1.币种
                        LEFT JOIN
                        (   SELECT cc.`所属团队`, cc.`币种`, cc.`产品id`, COUNT(订单编号) AS 最近单量
	                        FROM gat_zqsb cc
	                        where cc.日期  BETWEEN DATE_ADD(CURRENT_DATE(), INTERVAL -3 DAY) AND CURDATE()
	                        GROUP BY cc.`所属团队`, cc.`币种`, cc.`产品id`
                        ) s2 ON s.所属团队 = s2.所属团队 AND s.币种 = s2.币种 AND s.产品id = s2.产品id
                        LEFT JOIN
                        (   SELECT * FROM gk_stat_sign_rate) s3 ON s.所属团队 = s3.area_id AND s.币种 = s3.currency_id AND s.产品id = s3.goods_id
                        LEFT JOIN
                        (   SELECT * FROM gk_bi_estimate_goods) s4 ON s.所属团队 = s4.area_id AND s.币种 = s4.currency_id AND s.产品id = s4.goods_id
                    ) z;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            df.to_excel(file_path, sheet_name='使用', index=False)
            print('输出成功…………')
            try:
                print('正在运行 预估签收率 表宏…………')
                app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
                app.display_alerts = False
                wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
                wbsht1 = app.books.open(file_path)
                wbsht.macro('预估签收率修饰_使用')()
                wbsht1.save()
                wbsht1.close()
                wbsht.close()
                app.quit()
            except Exception as e:
                print('运行失败：', str(Exception) + str(e))
            print('运行成功…………')
        else:
            print('今日  无需获取预估签收率的数据！！！')

        print('正在获取同产品各团队对比的数据......')
        week: datetime = datetime.datetime.now()
        if week.isoweekday() == 2:
            time_path: datetime = datetime.datetime.now()
            mkpath = "F:\\神龙签收率\\A同产品各团队对比\\" + time_path.strftime('%m.%d')
            isExists = os.path.exists(mkpath)
            if not isExists:
                os.makedirs(mkpath)
            else:
                print(mkpath + ' 目录已存在')
            file_path = mkpath + '\\{} 同产品各团队对比_神龙.xlsx'.format(time_path.strftime('%m.%d'))
            sql = '''SELECT *
					FROM(SELECT	IFNULL(月份, '总计') 月份, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id, IFNULL(产品名称, '总计') 产品名称,
							SUM(神龙单量) 神龙单量, 
                                concat(ROUND(SUM(神龙签收) / SUM(神龙总量) * 100,2),'%') as 神龙总计签收,
                                concat(ROUND(SUM(神龙完成) / SUM(神龙总量) * 100,2),'%') as 神龙完成占比,					
							SUM(火凤凰单量) 火凤凰单量, 
                                concat(ROUND(SUM(火凤凰签收) / SUM(火凤凰总量) * 100,2),'%') as 火凤凰总计签收,
                                concat(ROUND(SUM(火凤凰完成) / SUM(火凤凰总量) * 100,2),'%') as 火凤凰完成占比,					
							SUM(雪豹港台单量) 雪豹港台单量, 
                                concat(ROUND(SUM(雪豹港台签收) / SUM(雪豹港台总量) * 100,2),'%') as 雪豹港台总计签收,
                                concat(ROUND(SUM(雪豹港台完成) / SUM(雪豹港台总量) * 100,2),'%') as 雪豹港台完成占比					
                        FROM(SELECT 年月 月份,币种 地区, 产品id, 产品名称,
                                    SUM(IF(所属团队 = '神龙港台',1,0)) as 神龙单量,
									SUM(IF(所属团队 = '神龙港台',价格,0)) as 神龙总量,
                                    SUM(IF(所属团队 = '神龙港台' AND 最终状态 = "已签收",价格,0)) as 神龙签收,
                                    SUM(IF(所属团队 = '神龙港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 神龙完成,
                                    SUM(IF(所属团队 = '火凤凰港台',1,0)) as 火凤凰单量,
									SUM(IF(所属团队 = '火凤凰港台',价格,0)) as 火凤凰总量,
                                    SUM(IF(所属团队 = '火凤凰港台' AND 最终状态 = "已签收",价格,0)) as 火凤凰签收,
                                    SUM(IF(所属团队 = '火凤凰港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 火凤凰完成,
                                    SUM(IF(所属团队 = '雪豹港台',1,0)) as 雪豹港台单量,
									SUM(IF(所属团队 = '雪豹港台',价格,0)) as 雪豹港台总量,
                                    SUM(IF(所属团队 = '雪豹港台' AND 最终状态 = "已签收",价格,0)) as 雪豹港台签收,
                                    SUM(IF(所属团队 = '雪豹港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 雪豹港台完成,
                                    SUM(IF(所属团队 = '金鹏港台',1,0)) as 小虎队单量,
									SUM(IF(所属团队 = '金鹏港台',价格,0)) as 小虎队总量,
                                    SUM(IF(所属团队 = '金鹏港台' AND 最终状态 = "已签收",价格,0)) as 小虎队签收,
                                    SUM(IF(所属团队 = '金鹏港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 小虎队完成
                            FROM gat_zqsb cc
							WHERE cc.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m')
							GROUP BY cc.年月,cc.币种,cc.产品id
						) s
					GROUP BY 月份,地区,产品id		
--                   WITH ROLLUP 
					) ss
                   ORDER BY FIELD(月份,DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'),'总计'),
                            FIELD(地区,'台湾','香港','总计'),
                            神龙单量 DESC;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            df.to_excel(file_path, sheet_name='使用', index=False)
            print('输出成功…………')
            try:
                print('正在运行 同产品各团队对比 表宏…………')
                app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
                app.display_alerts = False
                wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
                wbsht1 = app.books.open(file_path)
                wbsht.macro('同产品各团队对比_使用')()
                wbsht1.save()
                wbsht1.close()
                wbsht.close()
                app.quit()
            except Exception as e:
                print('运行失败：', str(Exception) + str(e))
            print('运行成功…………')

            file_path = mkpath + '\\{} 同产品各团队对比_火凤凰.xlsx'.format(time_path.strftime('%m.%d'))
            sql = '''SELECT *
        					FROM(SELECT	IFNULL(月份, '总计') 月份, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id, IFNULL(产品名称, '总计') 产品名称,
        							SUM(火凤凰单量) 火凤凰单量, 
                                        concat(ROUND(SUM(火凤凰签收) / SUM(火凤凰总量) * 100,2),'%') as 火凤凰总计签收,
                                        concat(ROUND(SUM(火凤凰完成) / SUM(火凤凰总量) * 100,2),'%') as 火凤凰完成占比,	
        							SUM(神龙单量) 神龙单量, 
                                        concat(ROUND(SUM(神龙签收) / SUM(神龙总量) * 100,2),'%') as 神龙总计签收,
                                        concat(ROUND(SUM(神龙完成) / SUM(神龙总量) * 100,2),'%') as 神龙完成占比,									
        							SUM(雪豹港台单量) 雪豹港台单量, 
                                        concat(ROUND(SUM(雪豹港台签收) / SUM(雪豹港台总量) * 100,2),'%') as 雪豹港台总计签收,
                                        concat(ROUND(SUM(雪豹港台完成) / SUM(雪豹港台总量) * 100,2),'%') as 雪豹港台完成占比					
                        FROM(SELECT 年月 月份,币种 地区, 产品id, 产品名称,
                                    SUM(IF(所属团队 = '神龙港台',1,0)) as 神龙单量,
									SUM(IF(所属团队 = '神龙港台',价格,0)) as 神龙总量,
                                    SUM(IF(所属团队 = '神龙港台' AND 最终状态 = "已签收",价格,0)) as 神龙签收,
                                    SUM(IF(所属团队 = '神龙港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 神龙完成,
                                    SUM(IF(所属团队 = '火凤凰港台',1,0)) as 火凤凰单量,
									SUM(IF(所属团队 = '火凤凰港台',价格,0)) as 火凤凰总量,
                                    SUM(IF(所属团队 = '火凤凰港台' AND 最终状态 = "已签收",价格,0)) as 火凤凰签收,
                                    SUM(IF(所属团队 = '火凤凰港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 火凤凰完成,
                                    SUM(IF(所属团队 = '雪豹港台',1,0)) as 雪豹港台单量,
									SUM(IF(所属团队 = '雪豹港台',价格,0)) as 雪豹港台总量,
                                    SUM(IF(所属团队 = '雪豹港台' AND 最终状态 = "已签收",价格,0)) as 雪豹港台签收,
                                    SUM(IF(所属团队 = '雪豹港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 雪豹港台完成,
                                    SUM(IF(所属团队 = '金鹏港台',1,0)) as 小虎队单量,
									SUM(IF(所属团队 = '金鹏港台',价格,0)) as 小虎队总量,
                                    SUM(IF(所属团队 = '金鹏港台' AND 最终状态 = "已签收",价格,0)) as 小虎队签收,
                                    SUM(IF(所属团队 = '金鹏港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格,0)) as 小虎队完成
                            FROM gat_zqsb cc
        							WHERE cc.年月 >= DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m')
        							GROUP BY cc.年月,cc.币种,cc.产品id
        						) s
        					GROUP BY 月份,地区,产品id		
        --                   WITH ROLLUP 
        					) ss
                           ORDER BY FIELD(月份,DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'),'总计'),
                                    FIELD(地区,'台湾','香港','总计'),
                                    火凤凰单量 DESC;'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            df.to_excel(file_path, sheet_name='使用', index=False)
            print('输出成功…………')
            try:
                print('正在运行 同产品各团队对比 表宏…………')
                app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
                app.display_alerts = False
                wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
                wbsht1 = app.books.open(file_path)
                wbsht.macro('同产品各团队对比_使用')()
                wbsht1.save()
                wbsht1.close()
                wbsht.close()
                app.quit()
            except Exception as e:
                print('运行失败：', str(Exception) + str(e))
            print('运行成功…………')

        else:
            print('今日  无需获取 同产品各团队对比 的数据！！！')

    # 新版签收率-报表(自己看的) - 单量计算
    def gat_new(self, team, month_last, month_yesterday, currency_id):  # 报表各团队近两个月的物流数据
        month_now = datetime.datetime.now().strftime('%Y-%m-%d')
        match = {'gat': '港台'}
        emailAdd = {'台湾': 'giikinliujun@163.com',
                    '香港': 'giikinliujun@163.com',
                    '品牌': 'sunyaru@giikin.com'}
        print(month_last + '---' + month_yesterday)
        currency = None
        if currency_id == '全部付款':
            currency = '"货到付款","货到付款（含税）","Pacypay信用卡支付【波兰】","钱海支付","gleepay","AsiaBill信用卡支付","Asiabill信用卡直接支付","Asiabill信用卡2.5方支付","Asiabill2.5方支付","paypal快捷支付","Cropay信用卡支付","空中云汇直连信用卡"'
        elif currency_id == '货到付款':
            currency = '"货到付款","货到付款（含税）"'
        elif currency_id == '在线付款':
            currency = '"Pacypay信用卡支付【波兰】","钱海支付","gleepay","AsiaBill信用卡支付","Asiabill信用卡直接支付","Asiabill信用卡2.5方支付","Asiabill2.5方支付","paypal快捷支付","Cropay信用卡支付","空中云汇直连信用卡"'
        # print('正在修改-港澳台-物流渠道…………')
        # self.update_logistics_name()

        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 物流分类
        print('正在获取---物流分类…………')
        sql0 = '''SELECT s2.家族, s2.币种, s2.年月, s2.是否改派, s2.物流方式,
                        IF(s2.签收=0,NULL,s2.签收) as 签收,
                        IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
                        IF(s2.在途=0,NULL,s2.在途) as 在途,				
                        IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
                        IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
                        IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
                        IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
                        IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
                        IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
                        IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
                        IF(s2.总订单=0,NULL,s2.总订单) as 全部,					
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                    concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                        concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                        concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                FROM (  SELECT  IFNULL(cx.所属团队,'合计') as 家族, IFNULL(cx.币种,'合计') as 币种, IFNULL(cx.年月,'合计') as 年月, IFNULL(cx.是否改派,'合计') as 是否改派, IFNULL(cx.物流渠道,'合计') as 物流方式,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *
                                    FROM {0}_zqsb cc 
                                    WHERE cc.日期 BETWEEN '{1}' AND '{2}'
                            ) cx
                            WHERE cx.`运单编号` is not null AND cx.付款方式 in ({3})
                            GROUP BY cx.`所属团队`, cx.`币种`, cx.`年月`, cx.`是否改派`, cx.`物流渠道`
                            with rollup
                ) s2
                GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`物流方式` 
                HAVING s2.年月 <> '合计'
                ORDER BY FIELD(s2.`家族`, {5},'合计'),
                        FIELD(s2.`币种`,'台湾','香港','合计'),
                        s2.`年月`,
                        FIELD(s2.`是否改派`,'改派','直发','合计'),
                        FIELD(s2.`物流方式`, {6},'合计'),
                        s2.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2, self.logistics_name)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)
        # 物流分旬
        print('正在获取---物流分旬…………')
        sql11 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.物流方式,s2.旬,
                        IF(s2.签收=0,NULL,s2.签收) as 签收,
                        IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
                        IF(s2.在途=0,NULL,s2.在途) as 在途,				
                        IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
                        IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
                        IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
                        IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
                        IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
                        IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
                        IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
                        IF(s2.总订单=0,NULL,s2.总订单) as 全部,					
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
                    concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                        concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                        concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                FROM ( SELECT  IFNULL(cx.家族,'合计') as 家族, IFNULL(cx.币种,'合计') as 币种, IFNULL(cx.年月,'合计') as 年月, IFNULL(cx.是否改派,'合计') as 是否改派, IFNULL(cx.物流渠道,'合计') as 物流方式,
							    IFNULL(IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))),'合计') as 旬,
							    SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单, 总订单量, 已发货单量, 已完成单量,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *, 所属团队 as 家族
                                    FROM {0}_zqsb cc 
                                    WHERE cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发' AND cc.付款方式 in ({3})
                            ) cx
							LEFT JOIN 
							(  SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
							    FROM (SELECT *, 所属团队 as 家族
							            FROM {0}_zqsb cc 
							            WHERE cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'  AND cc.`是否改派` = '直发' AND cc.付款方式 in ({3})
							    ) dg  
							    GROUP BY dg.币种,dg.家族,dg.年月
							) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
							GROUP BY cx.`家族`, cx.`币种`, cx.`年月`, cx.`是否改派`, cx.`物流渠道`, cx.`旬`
							with rollup
                ) s2
                GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`物流方式` , s2.`旬` 
                HAVING s2.是否改派 <> '合计'
                ORDER BY FIELD(s2.`家族`, {5},'合计'),
                        FIELD(s2.`币种`,'台湾','香港','合计'),
                        s2.`年月`,
                        FIELD(s2.`是否改派`,'改派','直发','合计'),
                        FIELD(s2.`物流方式`, {6},'合计'),
                        FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                        s2.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2, self.logistics_name)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)
        listT.append(df11)

        # 父级分旬
        print('正在获取---父级分旬…………')
        sql12 = '''SELECT s2.家族,s2.币种,s2.年月,s2.父级分类,s2.旬,
                                IF(s2.签收=0,NULL,s2.签收) as 签收,
                                IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
                                IF(s2.在途=0,NULL,s2.在途) as 在途,				
                                IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
                                IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
                                IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
                                IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
                                IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
                                IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
                                IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
                                IF(s2.总订单=0,NULL,s2.总订单) as 全部,					
                            concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                                concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                                concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                                concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                                concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                                concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                                concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                                concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
                            concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                                concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                                concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                                concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                                concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                        FROM (  SELECT  IFNULL(cx.家族,'合计') as 家族, IFNULL(cx.币种,'合计') as 币种, IFNULL(cx.年月,'合计') as 年月,
        								IFNULL(cx.父级分类,'合计') as 父级分类,
        								IFNULL(IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))),'合计') as 旬,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                        SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                        SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                        SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                        SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                        SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                        SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                        count(订单编号) as 总订单,总订单量,已发货单量,已完成单量,
                                        count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                        SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                        SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                        SUM(`价格RMB`) as 总计金额,
                                        SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                                FROM (SELECT *, 所属团队 as 家族
                                        FROM {0}_zqsb cc 
                                        WHERE cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'  AND cc.`是否改派` = '直发' AND cc.付款方式 in ({3})
                                ) cx
                                LEFT JOIN 
                                (  SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                                    FROM (SELECT *, 所属团队 as 家族
                                            FROM {0}_zqsb cc 
                                            WHERE cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'  AND cc.`是否改派` = '直发' AND cc.付款方式 in ({3})
                                    ) dg  
                                    GROUP BY dg.币种,dg.家族,dg.年月
                                ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`家族`, cx.`币种`, cx.`年月`, cx.`父级分类`, cx.`旬`
                                with rollup
                        ) s2
                        GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`父级分类` , s2.`旬` 
                        HAVING s2.年月 <> '合计'
                        ORDER BY FIELD(s2.`家族`, {5},'合计'),
                                FIELD(s2.`币种`,'台湾','香港','合计'),
                                s2.`年月`,
                                FIELD(s2.父级分类, '居家百货', '电子电器', '服饰', '医药保健',  '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','包材类', '合计' ),
                                FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                                s2.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)

        df12 = pd.read_sql_query(sql=sql12, con=self.engine1)
        listT.append(df12)
        # 二级分旬
        print('正在获取---二级分旬…………')
        sql13 = '''SELECT s2.家族,s2.币种,s2.年月,s2.父级分类,s2.二级分类, s2.旬,
                        IF(s2.签收=0,NULL,s2.签收) as 签收,
                        IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
                        IF(s2.在途=0,NULL,s2.在途) as 在途,				
                        IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
                        IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
                        IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
                        IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
                        IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
                        IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
                        IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
                        IF(s2.总订单=0,NULL,s2.总订单) as 全部,					
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
                    concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                        concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                        concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                FROM (  SELECT  IFNULL(cx.家族,'合计') as 家族, IFNULL(cx.币种,'合计') as 币种, IFNULL(cx.年月,'合计') as 年月, IFNULL(cx.父级分类,'合计') as 父级分类, IFNULL(cx.二级分类,'合计') as 二级分类,
                                IFNULL(IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))),'合计') as 旬,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,总订单量,已发货单量,已完成单量,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                        FROM (SELECT *, 所属团队 as 家族
                                FROM {0}_zqsb cc 
                                WHERE cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发' AND cc.付款方式 in ({3})
                        ) cx
                        LEFT JOIN 
                        (  SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                            FROM (SELECT *, 所属团队 as 家族
                                    FROM {0}_zqsb cc 
                                    WHERE cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发' AND cc.付款方式 in ({3})
                            ) dg  
                            GROUP BY dg.币种,dg.家族,dg.年月
                        ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                        GROUP BY cx.`家族`, cx.`币种`, cx.`年月`, cx.`父级分类`, cx.`二级分类`, cx.`旬`
                        with rollup
                ) s2
                GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`父级分类` ,  s2.`二级分类` ,  s2.`旬` 
                HAVING s2.年月 <> '合计'
                ORDER BY FIELD(s2.`家族`, {5},'合计'),
                        FIELD(s2.`币种`,'台湾','香港','合计'),
                        s2.`年月`,
                        FIELD(s2.父级分类, "居家百货", "电子电器", "服饰", "医药保健", "鞋类", "美容个护", "包类","钟表珠宝","母婴玩具","包材类","合计"),
                        FIELD(s2.二级分类, "上衣","下装","内衣","套装","裙子","配饰","母婴服饰","凉/拖鞋","皮鞋","休闲运动鞋","靴子",
                                           "单肩包","双肩包","钱包","行李箱包","厨房用品","日用百货","布艺家纺","宠物用品","户外运动","汽车用品","家装建材","办公/文化",
                                           "手表手环","影音娱乐","电脑外设","手机外设","家用电器","个护电器","智能设备","彩妆","护肤","个人洗护",
                                           "保健食品","护理护具","保健器械","药品","成人保健","手表","钟表","饰品","玩具","母婴用品","仓库包材","仓库耗材","合计"),
                        FIELD(s2.`旬`,"上旬","中旬","下旬","合计"),
                        s2.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)

        df13 = pd.read_sql_query(sql=sql13, con=self.engine1)
        listT.append(df13)

        # 产品整月 台湾
        print('正在获取---产品整月 台湾…………')
        sql14 = '''SELECT *
                FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份, IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                            SUM(s1.已签收) as 已签收,  SUM(s1.拒收) as 拒收, SUM(s1.已退货) as 已退货, SUM(s1.已完成) as 已完成, SUM(s1.总订单) as 总订单,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						    concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
						SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
						    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收', SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
						    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成', SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
						    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
						    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
						    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
						SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
						    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收', SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
						    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成', SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
						    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
						    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
						    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						    SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收', SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						    SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成', SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						    concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						    concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						    concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
						SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
						    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收', SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
						    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成', SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
						    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
						    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
						    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
						SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
						    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收', SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
						    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成', SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
						    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
						    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
						    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
						SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
						    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收', SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
						    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成', SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
						    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
						    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
						    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
						SUM(s1.速派711超商已签收) as '台湾-速派-711超商已签收',
						    SUM(s1.速派711超商拒收) as '台湾-速派-711超商拒收',  SUM(s1.速派711超商已退货) as '台湾-速派-711超商已退货',
						    SUM(s1.速派711超商已完成) as '台湾-速派-711超商已完成', SUM(s1.速派711超商总订单) as '台湾-速派-711超商总订单',
						    concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
						    concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
						    concat(ROUND(SUM(s1.速派711超商已完成) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
						    concat(ROUND(SUM(s1.速派711超商已退货) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
						    concat(ROUND(SUM(s1.速派711超商拒收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
						SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
						    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收', SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
						    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成', SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
						    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
				    		concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
					    	concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
					    	concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
					    	concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
						SUM(s1.速派黑猫已签收) as '台湾-速派-黑猫已签收',
						    SUM(s1.速派黑猫拒收) as '台湾-速派-黑猫拒收',  SUM(s1.速派黑猫已退货) as '台湾-速派-黑猫已退货',
						    SUM(s1.速派黑猫已完成) as '台湾-速派-黑猫已完成', SUM(s1.速派黑猫总订单) as '台湾-速派-黑猫总订单',
						    concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫完成签收',
				    		concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫总计签收',
					    	concat(ROUND(SUM(s1.速派黑猫已完成) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫完成占比',
					    	concat(ROUND(SUM(s1.速派黑猫已退货) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫退货率',
					    	concat(ROUND(SUM(s1.速派黑猫拒收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫拒收率',
						SUM(s1.速派宅配通已签收) as '台湾-速派宅配通已签收',
						    SUM(s1.速派宅配通拒收) as '台湾-速派宅配通拒收',  SUM(s1.速派宅配通已退货) as '台湾-速派宅配通已退货',
						    SUM(s1.速派宅配通已完成) as '台湾-速派宅配通已完成', SUM(s1.速派宅配通总订单) as '台湾-速派宅配通总订单',
						    concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通完成签收',
				    		concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通总计签收',
					    	concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通完成占比',
					    	concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通退货率',
					    	concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通拒收率',
						SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
						    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收', SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
					    	SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成', SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
				    		concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
				    		concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
					    	concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
						SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
					    	SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
					    	SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成', SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
					    	concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
					    	concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
					    	concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
						SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
					    	SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收', SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
					    	SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成', SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
					    	concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
					    	concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
					    	concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
						SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
					    	SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收', SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
					    	SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成', SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
					    	concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
					    	concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
					    	concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
						SUM(s1.铱熙无敌新竹已签收) as '铱熙无敌-新竹已签收',
					    	SUM(s1.铱熙无敌新竹拒收) as '铱熙无敌-新竹拒收', SUM(s1.铱熙无敌新竹已退货) as '铱熙无敌-新竹已退货',
					    	SUM(s1.铱熙无敌新竹已完成) as '铱熙无敌-新竹已完成', SUM(s1.铱熙无敌新竹总订单) as '铱熙无敌-新竹总订单',
					    	concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹完成签收',
					    	concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹总计签收',
					    	concat(ROUND(SUM(s1.铱熙无敌新竹已完成) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹完成占比',
					    	concat(ROUND(SUM(s1.铱熙无敌新竹已退货) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹退货率',
					    	concat(ROUND(SUM(s1.铱熙无敌新竹拒收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹拒收率',
						SUM(s1.铱熙无敌711超商已签收) as '铱熙无敌-711超商已签收',
					    	SUM(s1.铱熙无敌711超商拒收) as '铱熙无敌-711超商拒收', SUM(s1.铱熙无敌711超商已退货) as '铱熙无敌-711超商已退货',
					    	SUM(s1.铱熙无敌711超商已完成) as '铱熙无敌-711超商已完成', SUM(s1.铱熙无敌711超商总订单) as '铱熙无敌-711超商总订单',
					    	concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商完成签收',
					    	concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商总计签收',
					    	concat(ROUND(SUM(s1.铱熙无敌711超商已完成) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商完成占比',
					    	concat(ROUND(SUM(s1.铱熙无敌711超商已退货) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商退货率',
					    	concat(ROUND(SUM(s1.铱熙无敌711超商拒收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商拒收率',
						SUM(s1.铱熙无敌黑猫已签收) as '铱熙无敌-黑猫已签收',
					    	SUM(s1.铱熙无敌黑猫拒收) as '铱熙无敌-黑猫拒收', SUM(s1.铱熙无敌黑猫已退货) as '铱熙无敌-黑猫已退货',
					    	SUM(s1.铱熙无敌黑猫已完成) as '铱熙无敌-黑猫已完成', SUM(s1.铱熙无敌黑猫总订单) as '铱熙无敌-黑猫总订单',
					    	concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫完成签收',
					    	concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫总计签收',
					    	concat(ROUND(SUM(s1.铱熙无敌黑猫已完成) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫完成占比',
					    	concat(ROUND(SUM(s1.铱熙无敌黑猫已退货) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫退货率',
					    	concat(ROUND(SUM(s1.铱熙无敌黑猫拒收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫拒收率',
						SUM(s1.铱熙无敌宅配通已签收) as '铱熙无敌-宅配通已签收',
					    	SUM(s1.铱熙无敌宅配通拒收) as '铱熙无敌-宅配通拒收', SUM(s1.铱熙无敌宅配通已退货) as '铱熙无敌-宅配通已退货',
					    	SUM(s1.铱熙无敌宅配通已完成) as '铱熙无敌-宅配通已完成', SUM(s1.铱熙无敌宅配通总订单) as '铱熙无敌-宅配通总订单',
					    	concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通完成签收',
					    	concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通总计签收',
					    	concat(ROUND(SUM(s1.铱熙无敌宅配通已完成) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通完成占比',
					    	concat(ROUND(SUM(s1.铱熙无敌宅配通已退货) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通退货率',
					    	concat(ROUND(SUM(s1.铱熙无敌宅配通拒收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通拒收率',
						SUM(s1.铱熙无敌尾已签收) as '易速配头程-铱熙无敌尾已签收',
					    	SUM(s1.铱熙无敌尾拒收) as '易速配头程-铱熙无敌尾拒收', SUM(s1.铱熙无敌尾已退货) as '易速配头程-铱熙无敌尾已退货',
					    	SUM(s1.铱熙无敌尾已完成) as '易速配头程-铱熙无敌尾已完成',SUM(s1.铱熙无敌尾总订单) as '易速配头程-铱熙无敌尾总订单',
					    	concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾完成签收',
					    	concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾总计签收',
					    	concat(ROUND(SUM(s1.铱熙无敌尾已完成) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾完成占比',
					    	concat(ROUND(SUM(s1.铱熙无敌尾已退货) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾退货率',
					    	concat(ROUND(SUM(s1.铱熙无敌尾拒收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾拒收率',
						SUM(s1.龟山改派已签收) as '龟山改派已签收',
					    	SUM(s1.龟山改派拒收) as '龟山改派拒收',	SUM(s1.龟山改派已退货) as '龟山改派已退货',	
					    	SUM(s1.龟山改派已完成) as '龟山改派已完成',SUM(s1.龟山改派总订单) as '龟山改派总订单',
					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
					    	concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
					    	concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
					    	concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
				    	SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
					    	SUM(s1.森鸿改派拒收) as '森鸿改派拒收',	SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
					    	SUM(s1.森鸿改派已完成) as '森鸿改派已完成',	SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
					    	concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
					    	concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
					    	concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
						SUM(s1.速派新竹改派已签收) as '速派新竹改派已签收',
					    	SUM(s1.速派新竹改派拒收) as '速派新竹改派拒收',	SUM(s1.速派新竹改派已退货) as '速派新竹改派已退货',
					    	SUM(s1.速派新竹改派已完成) as '速派新竹改派已完成',	SUM(s1.速派新竹改派总订单) as '速派新竹改派总订单',
					    	concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派完成签收',
					    	concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派总计签收',
					    	concat(ROUND(SUM(s1.速派新竹改派已完成) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派完成占比',
					    	concat(ROUND(SUM(s1.速派新竹改派已退货) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派退货率',
					    	concat(ROUND(SUM(s1.速派新竹改派拒收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派拒收率',
						SUM(s1.速派宅配通改派已签收) as '速派宅配通改派已签收',
					    	SUM(s1.速派宅配通改派拒收) as '速派宅配通改派拒收',	SUM(s1.速派宅配通改派已退货) as '速派宅配通改派已退货',
					    	SUM(s1.速派宅配通改派已完成) as '速派宅配通改派已完成',	SUM(s1.速派宅配通改派总订单) as '速派宅配通改派总订单',
					    	concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派完成签收',
					    	concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派总计签收',
					    	concat(ROUND(SUM(s1.速派宅配通改派已完成) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派完成占比',
					    	concat(ROUND(SUM(s1.速派宅配通改派已退货) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派退货率',
					    	concat(ROUND(SUM(s1.速派宅配通改派拒收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派拒收率',
						SUM(s1.速派黑猫改派已签收) as '速派黑猫改派已签收',
					    	SUM(s1.速派黑猫改派拒收) as '速派黑猫改派拒收',	SUM(s1.速派黑猫改派已退货) as '速派黑猫改派已退货',
					    	SUM(s1.速派黑猫改派已完成) as '速派黑猫改派已完成',	SUM(s1.速派黑猫改派总订单) as '速派黑猫改派总订单',
					    	concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派完成签收',
					    	concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派总计签收',
					    	concat(ROUND(SUM(s1.速派黑猫改派已完成) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派完成占比',
					    	concat(ROUND(SUM(s1.速派黑猫改派已退货) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派退货率',
					    	concat(ROUND(SUM(s1.速派黑猫改派拒收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派拒收率',
						SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
					    	SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',	SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
					    	SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',	SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
					    	concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
					    	concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
					    	concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
						SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
					    	SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',	SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
					    	SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',	SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
					    	concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
					    	concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
					    	concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
						SUM(s1.天马黑猫改派已签收) as '天马黑猫改派已签收',
					    	SUM(s1.天马黑猫改派拒收) as '天马黑猫改派拒收',	SUM(s1.天马黑猫改派已退货) as '天马黑猫改派已退货',
					    	SUM(s1.天马黑猫改派已完成) as '天马黑猫改派已完成',	SUM(s1.天马黑猫改派总订单) as '天马黑猫改派总订单',
					    	concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派完成签收',
					    	concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派总计签收',
					    	concat(ROUND(SUM(s1.天马黑猫改派已完成) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派完成占比',
					    	concat(ROUND(SUM(s1.天马黑猫改派已退货) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派退货率',
					    	concat(ROUND(SUM(s1.天马黑猫改派拒收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派拒收率',
						SUM(s1.铱熙无敌黑猫改派已签收) as '铱熙无敌黑猫改派已签收',
					    	SUM(s1.铱熙无敌黑猫改派拒收) as '铱熙无敌黑猫改派拒收',	SUM(s1.铱熙无敌黑猫改派已退货) as '铱熙无敌黑猫改派已退货',
					    	SUM(s1.铱熙无敌黑猫改派已完成) as '铱熙无敌黑猫改派已完成',	SUM(s1.铱熙无敌黑猫改派总订单) as '铱熙无敌黑猫改派总订单',
					    	concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
					    	concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
					    	concat(ROUND(SUM(s1.铱熙无敌黑猫改派已完成) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
					    	concat(ROUND(SUM(s1.铱熙无敌黑猫改派已退货) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
					    	concat(ROUND(SUM(s1.铱熙无敌黑猫改派拒收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率',
						SUM(s1.铱熙无敌新竹改派已签收) as '铱熙无敌新竹改派已签收',
					    	SUM(s1.铱熙无敌新竹改派拒收) as '铱熙无敌新竹改派拒收',	SUM(s1.铱熙无敌新竹改派已退货) as '铱熙无敌新竹改派已退货',
					    	SUM(s1.铱熙无敌新竹改派已完成) as '铱熙无敌新竹改派已完成',	SUM(s1.铱熙无敌新竹改派总订单) as '铱熙无敌新竹改派总订单',
					    	concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派完成签收',
					    	concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派总计签收',
					    	concat(ROUND(SUM(s1.铱熙无敌新竹改派已完成) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派完成占比',
					    	concat(ROUND(SUM(s1.铱熙无敌新竹改派已退货) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派退货率',
					    	concat(ROUND(SUM(s1.铱熙无敌新竹改派拒收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派拒收率'
                    FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,	IFNULL(cx.币种, '合计') 地区,	IFNULL(cx.`年月`, '合计') 月份,	IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称, IFNULL(cx.父级分类, '合计') 父级分类, IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,	SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,	SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,	SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,	SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
							SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
							SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
							SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
							SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
								SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
								SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
								SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
								SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
							SUM(IF(cx.物流渠道 = "台湾-速派-711超商" ,1,0)) AS 速派711超商总订单,
								SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派711超商已签收,
								SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派711超商拒收,
								SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派711超商已退货,
								SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派711超商已完成,
							SUM(IF(cx.物流渠道 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
								SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
								SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
								SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
								SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
							SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" ,1,0)) AS 速派黑猫总订单,
								SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫已签收,
								SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫拒收,
								SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫已退货,
								SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫已完成,
							SUM(IF(cx.物流渠道 = "台湾-速派宅配通" ,1,0)) AS 速派宅配通总订单,
								SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
								SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
								SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
								SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
							SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
								SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
								SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
								SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
								SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
							SUM(IF(cx.物流渠道 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
								SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
								SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
								SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
								SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
							SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
								SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
								SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
								SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
								SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
							SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
								SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
								SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
								SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
								SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌新竹总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" ,1,0)) AS 铱熙无敌711超商总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已签收",1,0)) as 铱熙无敌711超商已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "拒收",1,0)) as 铱熙无敌711超商拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已退货",1,0)) as 铱熙无敌711超商已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌711超商已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" ,1,0)) AS 铱熙无敌黑猫总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" ,1,0)) AS 铱熙无敌宅配通总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已签收",1,0)) as 铱熙无敌宅配通已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "拒收",1,0)) as 铱熙无敌宅配通拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已退货",1,0)) as 铱熙无敌宅配通已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌宅配通已完成,
							SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" ,1,0)) AS 铱熙无敌尾总订单,
								SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已签收",1,0)) as 铱熙无敌尾已签收,
								SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "拒收",1,0)) as 铱熙无敌尾拒收,
								SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已退货",1,0)) as 铱熙无敌尾已退货,
								SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌尾已完成,
							SUM(IF(cx.物流渠道 = "龟山" ,1,0)) AS 龟山改派总订单,
								SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
								SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
								SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
								SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
							SUM(IF(cx.物流渠道 = "森鸿" ,1,0)) AS 森鸿改派总订单,
								SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
								SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
								SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
								SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
							SUM(IF(cx.物流渠道 = "速派新竹" ,1,0)) AS 速派新竹改派总订单,
								SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹改派已签收,
								SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹改派拒收,
								SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹改派已退货,
								SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹改派已完成,
							SUM(IF(cx.物流渠道 = "速派宅配通" ,1,0)) AS 速派宅配通改派总订单,
								SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通改派已签收,
								SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通改派拒收,
								SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通改派已退货,
								SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通改派已完成,
							SUM(IF(cx.物流渠道 = "速派黑猫" ,1,0)) AS 速派黑猫改派总订单,
								SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫改派已签收,
								SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫改派拒收,
								SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫改派已退货,
								SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫改派已完成,
							SUM(IF(cx.物流渠道 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
								SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
								SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
								SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
								SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
							SUM(IF(cx.物流渠道 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
								SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
								SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
								SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
								SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
							SUM(IF(cx.物流渠道 = "天马黑猫" ,1,0)) AS 天马黑猫改派总订单,
								SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫改派已签收,
								SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫改派拒收,
								SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫改派已退货,
								SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫改派已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" ,1,0)) AS 铱熙无敌黑猫改派总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫改派已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫改派拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫改派已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫改派已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌新竹改派总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹改派已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹改派拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹改派已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹改派已完成
				            FROM (SELECT *, 所属团队 as 家族
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                            ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                        ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                        WITH ROLLUP 
                ) s HAVING s.月份 != '合计'
        ORDER BY FIELD(s.`家族`, {5},'合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df14 = pd.read_sql_query(sql=sql14, con=self.engine1)
        listT.append(df14)
        # 产品分旬 台湾
        print('正在获取---产品分旬 台湾…………')
        sql15 = '''SELECT *
                 FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
						SUM(s1.已签收) as 已签收,	SUM(s1.拒收) as 拒收, SUM(s1.已退货) as 已退货,	SUM(s1.已完成) as 已完成,	SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                    SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
                        SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',  SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
                        SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',  SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
                        concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
                        concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
                        concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
                        concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
                        concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
                    SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
                        SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
                        SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
                        SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
                        SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
                        concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
                        concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
                        concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
                        concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
                        concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
                    SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
                        SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',  SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
                        SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',  SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
                        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
                        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
                        concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
                        concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
                        concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
                    SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
                        SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',   SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
                        SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',  SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
                        concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
                        concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
                        concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
                        concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
                        concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
                    SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
                        SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收', SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
                        SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',  SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
                        concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
                        concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
                        concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
                        concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
                        concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
                    SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
                        SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收', SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
                        SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',  SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
                        concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
                        concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
                        concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
                        concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
                        concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
                    SUM(s1.速派711超商已签收) as '台湾-速派-711超商已签收',
                        SUM(s1.速派711超商拒收) as '台湾-速派-711超商拒收',
                        SUM(s1.速派711超商已退货) as '台湾-速派-711超商已退货',
                        SUM(s1.速派711超商已完成) as '台湾-速派-711超商已完成',
                        SUM(s1.速派711超商总订单) as '台湾-速派-711超商总订单',
                        concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
                        concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
                        concat(ROUND(SUM(s1.速派711超商已完成) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
                        concat(ROUND(SUM(s1.速派711超商已退货) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
                        concat(ROUND(SUM(s1.速派711超商拒收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
                    SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
                        SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',  SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
                        SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',   SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
                        concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
                        concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
                        concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
                        concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
                        concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
                    SUM(s1.速派黑猫已签收) as '台湾-速派-黑猫已签收',
                        SUM(s1.速派黑猫拒收) as '台湾-速派-黑猫拒收',
                        SUM(s1.速派黑猫已退货) as '台湾-速派-黑猫已退货',
                        SUM(s1.速派黑猫已完成) as '台湾-速派-黑猫已完成',
                        SUM(s1.速派黑猫总订单) as '台湾-速派-黑猫总订单',
                        concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫完成签收',
                        concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫总计签收',
                        concat(ROUND(SUM(s1.速派黑猫已完成) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫完成占比',
                        concat(ROUND(SUM(s1.速派黑猫已退货) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫退货率',
                        concat(ROUND(SUM(s1.速派黑猫拒收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫拒收率',
                    SUM(s1.速派宅配通已签收) as '台湾-速派宅配通已签收',
                        SUM(s1.速派宅配通拒收) as '台湾-速派宅配通拒收', SUM(s1.速派宅配通已退货) as '台湾-速派宅配通已退货',
                        SUM(s1.速派宅配通已完成) as '台湾-速派宅配通已完成',  SUM(s1.速派宅配通总订单) as '台湾-速派宅配通总订单',
                        concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通完成签收',
                        concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通总计签收',
                        concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通完成占比',
                        concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通退货率',
                        concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通拒收率',
                    SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
                        SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
                        SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
                        SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
                        SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
                        concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
                        concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
                        concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
                        concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
                        concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
                    SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
                        SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收', SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
                        SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',  SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
                        concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
                        concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
                        concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
                        concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
                        concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
                    SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
                        SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',  SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
                        SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',  SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
                        concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
                        concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
                        concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
                        concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
                        concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
                    SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
                        SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
                        SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
                        SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
                        SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
                        concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
                        concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
                        concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
                        concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
                        concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
                    SUM(s1.铱熙无敌新竹已签收) as '铱熙无敌-新竹已签收',
                        SUM(s1.铱熙无敌新竹拒收) as '铱熙无敌-新竹拒收',  SUM(s1.铱熙无敌新竹已退货) as '铱熙无敌-新竹已退货',
                        SUM(s1.铱熙无敌新竹已完成) as '铱熙无敌-新竹已完成',  SUM(s1.铱熙无敌新竹总订单) as '铱熙无敌-新竹总订单',
                        concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹完成签收',
                        concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹总计签收',
                        concat(ROUND(SUM(s1.铱熙无敌新竹已完成) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹完成占比',
                        concat(ROUND(SUM(s1.铱熙无敌新竹已退货) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹退货率',
                        concat(ROUND(SUM(s1.铱熙无敌新竹拒收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹拒收率',
                    SUM(s1.铱熙无敌711超商已签收) as '铱熙无敌-711超商已签收',
                        SUM(s1.铱熙无敌711超商拒收) as '铱熙无敌-711超商拒收', SUM(s1.铱熙无敌711超商已退货) as '铱熙无敌-711超商已退货',
                        SUM(s1.铱熙无敌711超商已完成) as '铱熙无敌-711超商已完成',  SUM(s1.铱熙无敌711超商总订单) as '铱熙无敌-711超商总订单',
                        concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商完成签收',
                        concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商总计签收',
                        concat(ROUND(SUM(s1.铱熙无敌711超商已完成) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商完成占比',
                        concat(ROUND(SUM(s1.铱熙无敌711超商已退货) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商退货率',
                        concat(ROUND(SUM(s1.铱熙无敌711超商拒收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商拒收率',
                    SUM(s1.铱熙无敌黑猫已签收) as '铱熙无敌-黑猫已签收',
                        SUM(s1.铱熙无敌黑猫拒收) as '铱熙无敌-黑猫拒收', SUM(s1.铱熙无敌黑猫已退货) as '铱熙无敌-黑猫已退货',
                        SUM(s1.铱熙无敌黑猫已完成) as '铱熙无敌-黑猫已完成', SUM(s1.铱熙无敌黑猫总订单) as '铱熙无敌-黑猫总订单',
                        concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫完成签收',
                        concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫总计签收',
                        concat(ROUND(SUM(s1.铱熙无敌黑猫已完成) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫完成占比',
                        concat(ROUND(SUM(s1.铱熙无敌黑猫已退货) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫退货率',
                        concat(ROUND(SUM(s1.铱熙无敌黑猫拒收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫拒收率',
                    SUM(s1.铱熙无敌宅配通已签收) as '铱熙无敌-宅配通已签收',
                        SUM(s1.铱熙无敌宅配通拒收) as '铱熙无敌-宅配通拒收', SUM(s1.铱熙无敌宅配通已退货) as '铱熙无敌-宅配通已退货',
                        SUM(s1.铱熙无敌宅配通已完成) as '铱熙无敌-宅配通已完成', SUM(s1.铱熙无敌宅配通总订单) as '铱熙无敌-宅配通总订单',
                        concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通完成签收',
                        concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通总计签收',
                        concat(ROUND(SUM(s1.铱熙无敌宅配通已完成) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通完成占比',
                        concat(ROUND(SUM(s1.铱熙无敌宅配通已退货) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通退货率',
                        concat(ROUND(SUM(s1.铱熙无敌宅配通拒收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通拒收率',
                    SUM(s1.铱熙无敌尾已签收) as '易速配头程-铱熙无敌尾已签收',
                        SUM(s1.铱熙无敌尾拒收) as '易速配头程-铱熙无敌尾拒收',  SUM(s1.铱熙无敌尾已退货) as '易速配头程-铱熙无敌尾已退货',
                        SUM(s1.铱熙无敌尾已完成) as '易速配头程-铱熙无敌尾已完成',  SUM(s1.铱熙无敌尾总订单) as '易速配头程-铱熙无敌尾总订单',
                        concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾完成签收',
                        concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾总计签收',
                        concat(ROUND(SUM(s1.铱熙无敌尾已完成) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾完成占比',
                        concat(ROUND(SUM(s1.铱熙无敌尾已退货) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾退货率',
                        concat(ROUND(SUM(s1.铱熙无敌尾拒收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾拒收率',
                    SUM(s1.龟山改派已签收) as '龟山改派已签收',
                        SUM(s1.龟山改派拒收) as '龟山改派拒收',  SUM(s1.龟山改派已退货) as '龟山改派已退货',
                        SUM(s1.龟山改派已完成) as '龟山改派已完成', SUM(s1.龟山改派总订单) as '龟山改派总订单',
                        concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
                        concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
                        concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
                        concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
                        concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
                    SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
                        SUM(s1.森鸿改派拒收) as '森鸿改派拒收',  SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
                        SUM(s1.森鸿改派已完成) as '森鸿改派已完成',  SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
                        concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
                        concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
                        concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
                        concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
                        concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
                    SUM(s1.速派新竹改派已签收) as '速派新竹改派已签收',
                        SUM(s1.速派新竹改派拒收) as '速派新竹改派拒收',  SUM(s1.速派新竹改派已退货) as '速派新竹改派已退货',
                        SUM(s1.速派新竹改派已完成) as '速派新竹改派已完成',  SUM(s1.速派新竹改派总订单) as '速派新竹改派总订单',
                        concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派完成签收',
                        concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派总计签收',
                        concat(ROUND(SUM(s1.速派新竹改派已完成) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派完成占比',
                        concat(ROUND(SUM(s1.速派新竹改派已退货) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派退货率',
                        concat(ROUND(SUM(s1.速派新竹改派拒收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派拒收率',
                    SUM(s1.速派宅配通改派已签收) as '速派宅配通改派已签收',
                        SUM(s1.速派宅配通改派拒收) as '速派宅配通改派拒收',
                        SUM(s1.速派宅配通改派已退货) as '速派宅配通改派已退货',
                        SUM(s1.速派宅配通改派已完成) as '速派宅配通改派已完成',
                        SUM(s1.速派宅配通改派总订单) as '速派宅配通改派总订单',
                        concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派完成签收',
                        concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派总计签收',
                        concat(ROUND(SUM(s1.速派宅配通改派已完成) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派完成占比',
                        concat(ROUND(SUM(s1.速派宅配通改派已退货) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派退货率',
                        concat(ROUND(SUM(s1.速派宅配通改派拒收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派拒收率',
                    SUM(s1.速派黑猫改派已签收) as '速派黑猫改派已签收',
                        SUM(s1.速派黑猫改派拒收) as '速派黑猫改派拒收',  SUM(s1.速派黑猫改派已退货) as '速派黑猫改派已退货',
                        SUM(s1.速派黑猫改派已完成) as '速派黑猫改派已完成',  SUM(s1.速派黑猫改派总订单) as '速派黑猫改派总订单',
                        concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派完成签收',
                        concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派总计签收',
                        concat(ROUND(SUM(s1.速派黑猫改派已完成) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派完成占比',
                        concat(ROUND(SUM(s1.速派黑猫改派已退货) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派退货率',
                        concat(ROUND(SUM(s1.速派黑猫改派拒收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派拒收率',
                    SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
                        SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',  SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
                        SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',  SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
                        concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
                        concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
                        concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
                        concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
                        concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
                    SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
                        SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',    SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
                        SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',   SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
                        concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
                        concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
                        concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
                        concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
                        concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
                    SUM(s1.天马黑猫改派已签收) as '天马黑猫改派已签收',
                        SUM(s1.天马黑猫改派拒收) as '天马黑猫改派拒收',   SUM(s1.天马黑猫改派已退货) as '天马黑猫改派已退货',
                        SUM(s1.天马黑猫改派已完成) as '天马黑猫改派已完成',   SUM(s1.天马黑猫改派总订单) as '天马黑猫改派总订单',
                        concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派完成签收',
                        concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派总计签收',
                        concat(ROUND(SUM(s1.天马黑猫改派已完成) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派完成占比',
                        concat(ROUND(SUM(s1.天马黑猫改派已退货) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派退货率',
                        concat(ROUND(SUM(s1.天马黑猫改派拒收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派拒收率',
                    SUM(s1.铱熙无敌黑猫改派已签收) as '铱熙无敌黑猫改派已签收',
                        SUM(s1.铱熙无敌黑猫改派拒收) as '铱熙无敌黑猫改派拒收',    SUM(s1.铱熙无敌黑猫改派已退货) as '铱熙无敌黑猫改派已退货',
                        SUM(s1.铱熙无敌黑猫改派已完成) as '铱熙无敌黑猫改派已完成',   SUM(s1.铱熙无敌黑猫改派总订单) as '铱熙无敌黑猫改派总订单',
                        concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
                        concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
                        concat(ROUND(SUM(s1.铱熙无敌黑猫改派已完成) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
                        concat(ROUND(SUM(s1.铱熙无敌黑猫改派已退货) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
                        concat(ROUND(SUM(s1.铱熙无敌黑猫改派拒收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率',
                    SUM(s1.铱熙无敌新竹改派已签收) as '铱熙无敌新竹改派已签收',
                        SUM(s1.铱熙无敌新竹改派拒收) as '铱熙无敌新竹改派拒收',   SUM(s1.铱熙无敌新竹改派已退货) as '铱熙无敌新竹改派已退货',
                        SUM(s1.铱熙无敌新竹改派已完成) as '铱熙无敌新竹改派已完成',   SUM(s1.铱熙无敌新竹改派总订单) as '铱熙无敌新竹改派总订单',
                        concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派完成签收',
                        concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派总计签收',
                        concat(ROUND(SUM(s1.铱熙无敌新竹改派已完成) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派完成占比',
                        concat(ROUND(SUM(s1.铱熙无敌新竹改派已退货) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派退货率',
                        concat(ROUND(SUM(s1.铱熙无敌新竹改派拒收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派拒收率'
                FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,IFNULL(cx.币种, '合计') 地区,IFNULL(cx.`年月`, '合计') 月份,IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
								IFNULL(cx.产品id, '合计') 产品id,IFNULL(cx.产品名称, '合计') 产品名称,IFNULL(cx.父级分类, '合计') 父级分类,IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,	SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,	SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,	SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,	SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
							SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
								SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
							SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
							SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
								SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
							SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
								SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
								SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
								SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
								SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
							SUM(IF(cx.物流渠道 = "台湾-速派-711超商" ,1,0)) AS 速派711超商总订单,
								SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派711超商已签收,
								SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派711超商拒收,
								SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派711超商已退货,
								SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派711超商已完成,
							SUM(IF(cx.物流渠道 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
								SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
								SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
								SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
								SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
							SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" ,1,0)) AS 速派黑猫总订单,
								SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫已签收,
								SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫拒收,
								SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫已退货,
								SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫已完成,
							SUM(IF(cx.物流渠道 = "台湾-速派宅配通" ,1,0)) AS 速派宅配通总订单,
								SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
								SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
								SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
								SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
							SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
								SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
								SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
								SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
								SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
							SUM(IF(cx.物流渠道 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
								SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
								SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
								SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
								SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
							SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
								SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
								SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
								SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
								SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
							SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
								SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
								SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
								SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
								SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌新竹总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" ,1,0)) AS 铱熙无敌711超商总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已签收",1,0)) as 铱熙无敌711超商已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "拒收",1,0)) as 铱熙无敌711超商拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已退货",1,0)) as 铱熙无敌711超商已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌711超商已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" ,1,0)) AS 铱熙无敌黑猫总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" ,1,0)) AS 铱熙无敌宅配通总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已签收",1,0)) as 铱熙无敌宅配通已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "拒收",1,0)) as 铱熙无敌宅配通拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已退货",1,0)) as 铱熙无敌宅配通已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌宅配通已完成,
							SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" ,1,0)) AS 铱熙无敌尾总订单,
								SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已签收",1,0)) as 铱熙无敌尾已签收,
								SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "拒收",1,0)) as 铱熙无敌尾拒收,
								SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已退货",1,0)) as 铱熙无敌尾已退货,
								SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌尾已完成,
							SUM(IF(cx.物流渠道 = "龟山" ,1,0)) AS 龟山改派总订单,
								SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
								SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
								SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
								SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
							SUM(IF(cx.物流渠道 = "森鸿" ,1,0)) AS 森鸿改派总订单,
								SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
								SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
								SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
								SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
							SUM(IF(cx.物流渠道 = "速派新竹" ,1,0)) AS 速派新竹改派总订单,
								SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹改派已签收,
								SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹改派拒收,
								SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹改派已退货,
								SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹改派已完成,
							SUM(IF(cx.物流渠道 = "速派宅配通" ,1,0)) AS 速派宅配通改派总订单,
								SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通改派已签收,
								SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通改派拒收,
								SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通改派已退货,
								SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通改派已完成,
							SUM(IF(cx.物流渠道 = "速派黑猫" ,1,0)) AS 速派黑猫改派总订单,
								SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫改派已签收,
								SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫改派拒收,
								SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫改派已退货,
								SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫改派已完成,
							SUM(IF(cx.物流渠道 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
								SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
								SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
								SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
								SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
							SUM(IF(cx.物流渠道 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
								SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
								SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
								SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
								SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
							SUM(IF(cx.物流渠道 = "天马黑猫" ,1,0)) AS 天马黑猫改派总订单,
								SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫改派已签收,
								SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫改派拒收,
								SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫改派已退货,
								SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫改派已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" ,1,0)) AS 铱熙无敌黑猫改派总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫改派已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫改派拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫改派已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫改派已完成,
							SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌新竹改派总订单,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹改派已签收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹改派拒收,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹改派已退货,
								SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹改派已完成
				        FROM (SELECT *, 所属团队 as 家族
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                        ) cx WHERE cx.`币种` = '台湾'
                    GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                    ) s1
                GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                WITH ROLLUP 
            ) s HAVING s.旬 != '合计'
        ORDER BY FIELD(s.`家族`, {5},'合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df15 = pd.read_sql_query(sql=sql15, con=self.engine1)
        listT.append(df15)

        # 产品整月 香港
        print('正在获取---产品整月 香港…………')
        sql16 = '''SELECT *
                FROM(SELECT IFNULL(s1.家族, '合计') 家族, IFNULL(s1.地区, '合计') 地区, IFNULL(s1.月份, '合计') 月份,  IFNULL(s1.产品id, '合计') 产品id,
                            IFNULL(s1.产品名称, '合计') 产品名称, IFNULL(s1.父级分类, '合计') 父级分类, IFNULL(s1.二级分类, '合计') 二级分类,
                            SUM(s1.已签收) as 已签收, SUM(s1.拒收) as 拒收, SUM(s1.已退货) as 已退货,SUM(s1.已完成) as 已完成, SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.香港圆通已签收) as '香港-圆通已签收',
						SUM(s1.香港圆通拒收) as '香港-圆通拒收',
						SUM(s1.香港圆通已退货) as '香港-圆通拒收已退货',
						SUM(s1.香港圆通已完成) as '香港-圆通拒收已完成',
						SUM(s1.香港圆通总订单) as '香港-圆通拒收总订单',
						concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收完成签收',
						concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收总计签收',
						concat(ROUND(SUM(s1.香港圆通已完成) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收完成占比',
						concat(ROUND(SUM(s1.香港圆通已退货) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收退货率',
						concat(ROUND(SUM(s1.香港圆通拒收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收拒收率',
					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
					SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
						SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
						SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
						SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
						SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
						concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
						concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
						concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
						concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
						concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流渠道 = "香港-圆通" ,1,0)) AS 香港圆通总订单,
								SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已签收",1,0)) as 香港圆通已签收,
								SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "拒收",1,0)) as 香港圆通拒收,
								SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已退货",1,0)) as 香港圆通已退货,
								SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 香港圆通已完成,
							SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
							SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
							SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
							SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
							SUM(IF(cx.物流渠道 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
							SUM(IF(cx.物流渠道 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
								SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
								SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
								SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
								SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
							SUM(IF(cx.物流渠道 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
				            FROM (SELECT *, 所属团队 as 家族
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                            ) cx WHERE cx.`币种` = '香港'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                        ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.月份 != '合计'
        ORDER BY FIELD(s.`家族`, {5},'合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df16 = pd.read_sql_query(sql=sql16, con=self.engine1)
        listT.append(df16)
        # 产品分旬 香港
        print('正在获取---产品分旬 香港…………')
        sql17 = '''SELECT *
                FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
					SUM(s1.已签收) as 已签收,SUM(s1.拒收) as 拒收,SUM(s1.已退货) as 已退货,SUM(s1.已完成) as 已完成,SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.香港圆通已签收) as '香港-圆通已签收',
						SUM(s1.香港圆通拒收) as '香港-圆通拒收',
						SUM(s1.香港圆通已退货) as '香港-圆通拒收已退货',
						SUM(s1.香港圆通已完成) as '香港-圆通拒收已完成',
						SUM(s1.香港圆通总订单) as '香港-圆通拒收总订单',
						concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收完成签收',
						concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收总计签收',
						concat(ROUND(SUM(s1.香港圆通已完成) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收完成占比',
						concat(ROUND(SUM(s1.香港圆通已退货) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收退货率',
						concat(ROUND(SUM(s1.香港圆通拒收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收拒收率',
					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
					SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
						SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
						SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
						SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
						SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
						concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
						concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
						concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
						concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
						concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流渠道 = "香港-圆通" ,1,0)) AS 香港圆通总订单,
								SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已签收",1,0)) as 香港圆通已签收,
								SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "拒收",1,0)) as 香港圆通拒收,
								SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已退货",1,0)) as 香港圆通已退货,
								SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 香港圆通已完成,
							SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
							SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
							SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
							SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
							SUM(IF(cx.物流渠道 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
							SUM(IF(cx.物流渠道 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
								SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
								SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
								SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
								SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
							SUM(IF(cx.物流渠道 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
				        FROM (SELECT *, 所属团队 as 家族
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                        ) cx WHERE cx.`币种` = '香港'
                        GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                    ) s1
                    GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                    WITH ROLLUP 
            ) s HAVING s.旬 <> '合计'
        ORDER BY FIELD(s.`家族`, {5}),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df17 = pd.read_sql_query(sql=sql17, con=self.engine1)
        listT.append(df17)

        # 产品整月_直发 台湾
        print('正在获取---产品整月_直发 台湾…………')
        sql18 = '''SELECT *
                        FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                                    IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                                    SUM(s1.已签收) as 已签收, SUM(s1.拒收) as 拒收, SUM(s1.已退货) as 已退货, SUM(s1.已完成) as 已完成, SUM(s1.总订单) as 总订单,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						            concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                                SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
                                    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
                                    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
                                    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
                                    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
                                    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
                                    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
                                    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
                                    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
                                    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
                                SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
                                    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
                                    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
                                    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
                                    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
                                    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
                                    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
                                    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
                                    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
                                    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
                                SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
                                    SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
                                    SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
                                    SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
                                    SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
                                    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
                                    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
                                    concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
                                    concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
                                    concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
                                SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
                                    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
                                    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
                                    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
                                    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
                                    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
                                    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
                                    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
                                    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
                                    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
                                SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
                                    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
                                    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
                                    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
                                    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
                                    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
                                    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
                                    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
                                    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
                                    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
                                SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
                                    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
                                    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
                                    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
                                    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
                                    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
                                    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
                                    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
                                    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
                                    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
                                SUM(s1.速派711超商已签收) as '台湾-速派-711超商已签收',
                                    SUM(s1.速派711超商拒收) as '台湾-速派-711超商拒收',
                                    SUM(s1.速派711超商已退货) as '台湾-速派-711超商已退货',
                                    SUM(s1.速派711超商已完成) as '台湾-速派-711超商已完成',
                                    SUM(s1.速派711超商总订单) as '台湾-速派-711超商总订单',
                                    concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
                                    concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
                                    concat(ROUND(SUM(s1.速派711超商已完成) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
                                    concat(ROUND(SUM(s1.速派711超商已退货) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
                                    concat(ROUND(SUM(s1.速派711超商拒收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
                                SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
                                    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
                                    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
                                    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
                                    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
                                    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
                                    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
                                    concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
                                    concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
                                    concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
                                SUM(s1.速派黑猫已签收) as '台湾-速派-黑猫已签收',
                                    SUM(s1.速派黑猫拒收) as '台湾-速派-黑猫拒收',
                                    SUM(s1.速派黑猫已退货) as '台湾-速派-黑猫已退货',
                                    SUM(s1.速派黑猫已完成) as '台湾-速派-黑猫已完成',
                                    SUM(s1.速派黑猫总订单) as '台湾-速派-黑猫总订单',
                                    concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫完成签收',
                                    concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫总计签收',
                                    concat(ROUND(SUM(s1.速派黑猫已完成) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫完成占比',
                                    concat(ROUND(SUM(s1.速派黑猫已退货) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫退货率',
                                    concat(ROUND(SUM(s1.速派黑猫拒收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫拒收率',
                                SUM(s1.速派宅配通已签收) as '台湾-速派宅配通已签收',
                                    SUM(s1.速派宅配通拒收) as '台湾-速派宅配通拒收',
                                    SUM(s1.速派宅配通已退货) as '台湾-速派宅配通已退货',
                                    SUM(s1.速派宅配通已完成) as '台湾-速派宅配通已完成',
                                    SUM(s1.速派宅配通总订单) as '台湾-速派宅配通总订单',
                                    concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通完成签收',
                                    concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通总计签收',
                                    concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通完成占比',
                                    concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通退货率',
                                    concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通拒收率',
                                SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
                                    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
                                    SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
                                    SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
                                    SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
                                    concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
                                    concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
                                    concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
                                    concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
                                    concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
                                SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
                                    SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
                                    SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
                                    SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
                                    SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
                                    concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
                                    concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
                                    concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
                                    concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
                                    concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
                                SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
                                    SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
                                    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
                                    SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
                                    SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
                                    concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
                                    concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
                                    concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
                                    concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
                                    concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
                                SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
                                    SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
                                    SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
                                    SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
                                    SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
                                    concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
                                    concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
                                    concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
                                    concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
                                    concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
                                SUM(s1.铱熙无敌新竹已签收) as '铱熙无敌-新竹已签收',
                                    SUM(s1.铱熙无敌新竹拒收) as '铱熙无敌-新竹拒收',
                                    SUM(s1.铱熙无敌新竹已退货) as '铱熙无敌-新竹已退货',
                                    SUM(s1.铱熙无敌新竹已完成) as '铱熙无敌-新竹已完成',
                                    SUM(s1.铱熙无敌新竹总订单) as '铱熙无敌-新竹总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹已完成) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹已退货) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹拒收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹拒收率',
                                SUM(s1.铱熙无敌711超商已签收) as '铱熙无敌-711超商已签收',
                                    SUM(s1.铱熙无敌711超商拒收) as '铱熙无敌-711超商拒收',
                                    SUM(s1.铱熙无敌711超商已退货) as '铱熙无敌-711超商已退货',
                                    SUM(s1.铱熙无敌711超商已完成) as '铱熙无敌-711超商已完成',
                                    SUM(s1.铱熙无敌711超商总订单) as '铱熙无敌-711超商总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌711超商已完成) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌711超商已退货) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌711超商拒收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商拒收率',
                                SUM(s1.铱熙无敌黑猫已签收) as '铱熙无敌-黑猫已签收',
                                    SUM(s1.铱熙无敌黑猫拒收) as '铱熙无敌-黑猫拒收',
                                    SUM(s1.铱熙无敌黑猫已退货) as '铱熙无敌-黑猫已退货',
                                    SUM(s1.铱熙无敌黑猫已完成) as '铱熙无敌-黑猫已完成',
                                    SUM(s1.铱熙无敌黑猫总订单) as '铱熙无敌-黑猫总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫已完成) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫已退货) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫拒收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫拒收率',
                                SUM(s1.铱熙无敌宅配通已签收) as '铱熙无敌-宅配通已签收',
                                    SUM(s1.铱熙无敌宅配通拒收) as '铱熙无敌-宅配通拒收', SUM(s1.铱熙无敌宅配通已退货) as '铱熙无敌-宅配通已退货',
                                    SUM(s1.铱熙无敌宅配通已完成) as '铱熙无敌-宅配通已完成', SUM(s1.铱熙无敌宅配通总订单) as '铱熙无敌-宅配通总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌宅配通已完成) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌宅配通已退货) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌宅配通拒收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通拒收率',
                                SUM(s1.铱熙无敌尾已签收) as '易速配头程-铱熙无敌尾已签收',
                                    SUM(s1.铱熙无敌尾拒收) as '易速配头程-铱熙无敌尾拒收',
                                    SUM(s1.铱熙无敌尾已退货) as '易速配头程-铱熙无敌尾已退货',
                                    SUM(s1.铱熙无敌尾已完成) as '易速配头程-铱熙无敌尾已完成',
                                    SUM(s1.铱熙无敌尾总订单) as '易速配头程-铱熙无敌尾总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌尾已完成) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌尾已退货) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌尾拒收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾拒收率',
                                SUM(s1.龟山改派已签收) as '龟山改派已签收',
                                    SUM(s1.龟山改派拒收) as '龟山改派拒收',
                                    SUM(s1.龟山改派已退货) as '龟山改派已退货',
                                    SUM(s1.龟山改派已完成) as '龟山改派已完成',
                                    SUM(s1.龟山改派总订单) as '龟山改派总订单',
                                    concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
                                    concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
                                    concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
                                    concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
                                    concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
                                SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
                                    SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
                                    SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
                                    SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
                                    SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
                                    concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
                                    concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
                                    concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
                                    concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
                                    concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
                                SUM(s1.速派新竹改派已签收) as '速派新竹改派已签收',
                                    SUM(s1.速派新竹改派拒收) as '速派新竹改派拒收',
                                    SUM(s1.速派新竹改派已退货) as '速派新竹改派已退货',
                                    SUM(s1.速派新竹改派已完成) as '速派新竹改派已完成',
                                    SUM(s1.速派新竹改派总订单) as '速派新竹改派总订单',
                                    concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派完成签收',
                                    concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派总计签收',
                                    concat(ROUND(SUM(s1.速派新竹改派已完成) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派完成占比',
                                    concat(ROUND(SUM(s1.速派新竹改派已退货) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派退货率',
                                    concat(ROUND(SUM(s1.速派新竹改派拒收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派拒收率',
                                SUM(s1.速派宅配通改派已签收) as '速派宅配通改派已签收',
                                    SUM(s1.速派宅配通改派拒收) as '速派宅配通改派拒收',
                                    SUM(s1.速派宅配通改派已退货) as '速派宅配通改派已退货',
                                    SUM(s1.速派宅配通改派已完成) as '速派宅配通改派已完成',
                                    SUM(s1.速派宅配通改派总订单) as '速派宅配通改派总订单',
                                    concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派完成签收',
                                    concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派总计签收',
                                    concat(ROUND(SUM(s1.速派宅配通改派已完成) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派完成占比',
                                    concat(ROUND(SUM(s1.速派宅配通改派已退货) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派退货率',
                                    concat(ROUND(SUM(s1.速派宅配通改派拒收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派拒收率',
                                SUM(s1.速派黑猫改派已签收) as '速派黑猫改派已签收',
                                    SUM(s1.速派黑猫改派拒收) as '速派黑猫改派拒收',
                                    SUM(s1.速派黑猫改派已退货) as '速派黑猫改派已退货',
                                    SUM(s1.速派黑猫改派已完成) as '速派黑猫改派已完成',
                                    SUM(s1.速派黑猫改派总订单) as '速派黑猫改派总订单',
                                    concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派完成签收',
                                    concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派总计签收',
                                    concat(ROUND(SUM(s1.速派黑猫改派已完成) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派完成占比',
                                    concat(ROUND(SUM(s1.速派黑猫改派已退货) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派退货率',
                                    concat(ROUND(SUM(s1.速派黑猫改派拒收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派拒收率',
                                SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
                                    SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
                                    SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
                                    SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
                                    SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
                                    concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
                                    concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
                                    concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
                                    concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
                                    concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
                                SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
                                    SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
                                    SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
                                    SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
                                    SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
                                    concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
                                    concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
                                    concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
                                    concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
                                    concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
                                SUM(s1.天马黑猫改派已签收) as '天马黑猫改派已签收',
                                    SUM(s1.天马黑猫改派拒收) as '天马黑猫改派拒收',
                                    SUM(s1.天马黑猫改派已退货) as '天马黑猫改派已退货',
                                    SUM(s1.天马黑猫改派已完成) as '天马黑猫改派已完成',
                                    SUM(s1.天马黑猫改派总订单) as '天马黑猫改派总订单',
                                    concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派完成签收',
                                    concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派总计签收',
                                    concat(ROUND(SUM(s1.天马黑猫改派已完成) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派完成占比',
                                    concat(ROUND(SUM(s1.天马黑猫改派已退货) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派退货率',
                                    concat(ROUND(SUM(s1.天马黑猫改派拒收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派拒收率',
                                SUM(s1.铱熙无敌黑猫改派已签收) as '铱熙无敌黑猫改派已签收',
                                    SUM(s1.铱熙无敌黑猫改派拒收) as '铱熙无敌黑猫改派拒收',
                                    SUM(s1.铱熙无敌黑猫改派已退货) as '铱熙无敌黑猫改派已退货',
                                    SUM(s1.铱熙无敌黑猫改派已完成) as '铱熙无敌黑猫改派已完成',
                                    SUM(s1.铱熙无敌黑猫改派总订单) as '铱熙无敌黑猫改派总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫改派已完成) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫改派已退货) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫改派拒收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率',
                                SUM(s1.铱熙无敌新竹改派已签收) as '铱熙无敌新竹改派已签收',
                                    SUM(s1.铱熙无敌新竹改派拒收) as '铱熙无敌新竹改派拒收',
                                    SUM(s1.铱熙无敌新竹改派已退货) as '铱熙无敌新竹改派已退货',
                                    SUM(s1.铱熙无敌新竹改派已完成) as '铱熙无敌新竹改派已完成',
                                    SUM(s1.铱熙无敌新竹改派总订单) as '铱熙无敌新竹改派总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹改派已完成) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹改派已退货) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹改派拒收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派拒收率'
                            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,IFNULL(cx.币种, '合计') 地区,IFNULL(cx.`年月`, '合计') 月份,IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,IFNULL(cx.父级分类, '合计') 父级分类,IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
                                    SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-711超商" ,1,0)) AS 速派711超商总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派711超商已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派711超商拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派711超商已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派711超商已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" ,1,0)) AS 速派黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派宅配通" ,1,0)) AS 速派宅配通总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" ,1,0)) AS 铱熙无敌711超商总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已签收",1,0)) as 铱熙无敌711超商已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "拒收",1,0)) as 铱熙无敌711超商拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已退货",1,0)) as 铱熙无敌711超商已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌711超商已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" ,1,0)) AS 铱熙无敌黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" ,1,0)) AS 铱熙无敌宅配通总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已签收",1,0)) as 铱熙无敌宅配通已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "拒收",1,0)) as 铱熙无敌宅配通拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已退货",1,0)) as 铱熙无敌宅配通已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌宅配通已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" ,1,0)) AS 铱熙无敌尾总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已签收",1,0)) as 铱熙无敌尾已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "拒收",1,0)) as 铱熙无敌尾拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已退货",1,0)) as 铱熙无敌尾已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌尾已完成,
                                    SUM(IF(cx.物流渠道 = "龟山" ,1,0)) AS 龟山改派总订单,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
                                    SUM(IF(cx.物流渠道 = "森鸿" ,1,0)) AS 森鸿改派总订单,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派新竹" ,1,0)) AS 速派新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派宅配通" ,1,0)) AS 速派宅配通改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派黑猫" ,1,0)) AS 速派黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马黑猫" ,1,0)) AS 天马黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" ,1,0)) AS 铱熙无敌黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹改派已完成
        				            FROM (SELECT *, 所属团队 as 家族
                                        FROM {0}_zqsb cc where cc.`是否改派` = '直发' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                                    ) cx WHERE cx.`币种` = '台湾'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                        ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`, {5},'合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df18 = pd.read_sql_query(sql=sql18, con=self.engine1)
        listT.append(df18)
        # 产品分旬_直发 台湾
        print('正在获取---产品分旬_直发 台湾…………')
        sql19 = '''SELECT *
                    FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,SUM(s1.拒收) as 拒收,SUM(s1.已退货) as 已退货,SUM(s1.已完成) as 已完成,SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						        concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                            SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
                                SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
                                SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
                                SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
                                SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
                                concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
                                concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
                                concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
                                concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
                                concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
                            SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
                                SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
                                SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
                                SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
                                SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
                                concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
                                concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
                                concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
                                concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
                                concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
                            SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
                                SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
                                SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
                                SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
                                SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
                                concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
                                concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
                                concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
                                concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
                                concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
                            SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
                                SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
                                SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
                                SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
                                SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
                                concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
                                concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
                                concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
                                concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
                                concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
                            SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
                                SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
                                SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
                                SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
                                SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
                                concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
                                concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
                                concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
                                concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
                                concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
                            SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
                                SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
                                SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
                                SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
                                SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
                                concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
                                concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
                                concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
                                concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
                                concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
                            SUM(s1.速派711超商已签收) as '台湾-速派-711超商已签收',
                                SUM(s1.速派711超商拒收) as '台湾-速派-711超商拒收',
                                SUM(s1.速派711超商已退货) as '台湾-速派-711超商已退货',
                                SUM(s1.速派711超商已完成) as '台湾-速派-711超商已完成',
                                SUM(s1.速派711超商总订单) as '台湾-速派-711超商总订单',
                                concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
                                concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
                                concat(ROUND(SUM(s1.速派711超商已完成) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
                                concat(ROUND(SUM(s1.速派711超商已退货) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
                                concat(ROUND(SUM(s1.速派711超商拒收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
                            SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
                                SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
                                SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
                                SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
                                SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
                                concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
                                concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
                                concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
                                concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
                                concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
                            SUM(s1.速派黑猫已签收) as '台湾-速派-黑猫已签收',
                                SUM(s1.速派黑猫拒收) as '台湾-速派-黑猫拒收',
                                SUM(s1.速派黑猫已退货) as '台湾-速派-黑猫已退货',
                                SUM(s1.速派黑猫已完成) as '台湾-速派-黑猫已完成',
                                SUM(s1.速派黑猫总订单) as '台湾-速派-黑猫总订单',
                                concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫完成签收',
                                concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫总计签收',
                                concat(ROUND(SUM(s1.速派黑猫已完成) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫完成占比',
                                concat(ROUND(SUM(s1.速派黑猫已退货) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫退货率',
                                concat(ROUND(SUM(s1.速派黑猫拒收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫拒收率',
                            SUM(s1.速派宅配通已签收) as '台湾-速派宅配通已签收',
                                SUM(s1.速派宅配通拒收) as '台湾-速派宅配通拒收',
                                SUM(s1.速派宅配通已退货) as '台湾-速派宅配通已退货',
                                SUM(s1.速派宅配通已完成) as '台湾-速派宅配通已完成',
                                SUM(s1.速派宅配通总订单) as '台湾-速派宅配通总订单',
                                concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通完成签收',
                                concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通总计签收',
                                concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通完成占比',
                                concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通退货率',
                                concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通拒收率',
                            SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
                                SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
                                SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
                                SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
                                SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
                                concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
                                concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
                                concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
                                concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
                                concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
                            SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
                                SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
                                SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
                                SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
                                SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
                                concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
                                concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
                                concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
                                concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
                                concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
                            SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
                                SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
                                SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
                                SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
                                SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
                                concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
                                concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
                                concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
                                concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
                                concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
                            SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
                                SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
                                SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
                                SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
                                SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
                                concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
                                concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
                                concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
                                concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
                                concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
                            SUM(s1.铱熙无敌新竹已签收) as '铱熙无敌-新竹已签收',
                                SUM(s1.铱熙无敌新竹拒收) as '铱熙无敌-新竹拒收',
                                SUM(s1.铱熙无敌新竹已退货) as '铱熙无敌-新竹已退货',
                                SUM(s1.铱熙无敌新竹已完成) as '铱熙无敌-新竹已完成',
                                SUM(s1.铱熙无敌新竹总订单) as '铱熙无敌-新竹总订单',
                                concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌新竹已完成) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌新竹已退货) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹退货率',
                                concat(ROUND(SUM(s1.铱熙无敌新竹拒收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹拒收率',
                            SUM(s1.铱熙无敌711超商已签收) as '铱熙无敌-711超商已签收',
                                SUM(s1.铱熙无敌711超商拒收) as '铱熙无敌-711超商拒收',
                                SUM(s1.铱熙无敌711超商已退货) as '铱熙无敌-711超商已退货',
                                SUM(s1.铱熙无敌711超商已完成) as '铱熙无敌-711超商已完成',
                                SUM(s1.铱熙无敌711超商总订单) as '铱熙无敌-711超商总订单',
                                concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌711超商已完成) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌711超商已退货) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商退货率',
                                concat(ROUND(SUM(s1.铱熙无敌711超商拒收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商拒收率',
                            SUM(s1.铱熙无敌黑猫已签收) as '铱熙无敌-黑猫已签收',
                                SUM(s1.铱熙无敌黑猫拒收) as '铱熙无敌-黑猫拒收',
                                SUM(s1.铱熙无敌黑猫已退货) as '铱熙无敌-黑猫已退货',
                                SUM(s1.铱熙无敌黑猫已完成) as '铱熙无敌-黑猫已完成',
                                SUM(s1.铱熙无敌黑猫总订单) as '铱熙无敌-黑猫总订单',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫已完成) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫已退货) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫退货率',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫拒收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫拒收率',
                            SUM(s1.铱熙无敌宅配通已签收) as '铱熙无敌-宅配通已签收',
                                SUM(s1.铱熙无敌宅配通拒收) as '铱熙无敌-宅配通拒收', 
                                SUM(s1.铱熙无敌宅配通已退货) as '铱熙无敌-宅配通已退货',
                                SUM(s1.铱熙无敌宅配通已完成) as '铱熙无敌-宅配通已完成', 
                                SUM(s1.铱熙无敌宅配通总订单) as '铱熙无敌-宅配通总订单',
                                concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌宅配通已完成) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌宅配通已退货) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通退货率',
                                concat(ROUND(SUM(s1.铱熙无敌宅配通拒收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通拒收率',
                            SUM(s1.铱熙无敌尾已签收) as '易速配头程-铱熙无敌尾已签收',
                                SUM(s1.铱熙无敌尾拒收) as '易速配头程-铱熙无敌尾拒收',
                                SUM(s1.铱熙无敌尾已退货) as '易速配头程-铱熙无敌尾已退货',
                                SUM(s1.铱熙无敌尾已完成) as '易速配头程-铱熙无敌尾已完成',
                                SUM(s1.铱熙无敌尾总订单) as '易速配头程-铱熙无敌尾总订单',
                                concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌尾已完成) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌尾已退货) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾退货率',
                                concat(ROUND(SUM(s1.铱熙无敌尾拒收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾拒收率',
                            SUM(s1.龟山改派已签收) as '龟山改派已签收',
                                SUM(s1.龟山改派拒收) as '龟山改派拒收',
                                SUM(s1.龟山改派已退货) as '龟山改派已退货',
                                SUM(s1.龟山改派已完成) as '龟山改派已完成',
                                SUM(s1.龟山改派总订单) as '龟山改派总订单',
                                concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
                                concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
                                concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
                                concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
                                concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
                            SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
                                SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
                                SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
                                SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
                                SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
                                concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
                                concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
                                concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
                                concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
                                concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
                            SUM(s1.速派新竹改派已签收) as '速派新竹改派已签收',
                                SUM(s1.速派新竹改派拒收) as '速派新竹改派拒收',
                                SUM(s1.速派新竹改派已退货) as '速派新竹改派已退货',
                                SUM(s1.速派新竹改派已完成) as '速派新竹改派已完成',
                                SUM(s1.速派新竹改派总订单) as '速派新竹改派总订单',
                                concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派完成签收',
                                concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派总计签收',
                                concat(ROUND(SUM(s1.速派新竹改派已完成) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派完成占比',
                                concat(ROUND(SUM(s1.速派新竹改派已退货) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派退货率',
                                concat(ROUND(SUM(s1.速派新竹改派拒收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派拒收率',
                            SUM(s1.速派宅配通改派已签收) as '速派宅配通改派已签收',
                                SUM(s1.速派宅配通改派拒收) as '速派宅配通改派拒收',
                                SUM(s1.速派宅配通改派已退货) as '速派宅配通改派已退货',
                                SUM(s1.速派宅配通改派已完成) as '速派宅配通改派已完成',
                                SUM(s1.速派宅配通改派总订单) as '速派宅配通改派总订单',
                                concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派完成签收',
                                concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派总计签收',
                                concat(ROUND(SUM(s1.速派宅配通改派已完成) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派完成占比',
                                concat(ROUND(SUM(s1.速派宅配通改派已退货) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派退货率',
                                concat(ROUND(SUM(s1.速派宅配通改派拒收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派拒收率',
                            SUM(s1.速派黑猫改派已签收) as '速派黑猫改派已签收',
                                SUM(s1.速派黑猫改派拒收) as '速派黑猫改派拒收',
                                SUM(s1.速派黑猫改派已退货) as '速派黑猫改派已退货',
                                SUM(s1.速派黑猫改派已完成) as '速派黑猫改派已完成',
                                SUM(s1.速派黑猫改派总订单) as '速派黑猫改派总订单',
                                concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派完成签收',
                                concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派总计签收',
                                concat(ROUND(SUM(s1.速派黑猫改派已完成) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派完成占比',
                                concat(ROUND(SUM(s1.速派黑猫改派已退货) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派退货率',
                                concat(ROUND(SUM(s1.速派黑猫改派拒收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派拒收率',
                            SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
                                SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
                                SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
                                SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
                                SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
                                concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
                                concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
                                concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
                                concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
                                concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
                            SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
                                SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
                                SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
                                SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
                                SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
                                concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
                                concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
                                concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
                                concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
                                concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
                            SUM(s1.天马黑猫改派已签收) as '天马黑猫改派已签收',
                                SUM(s1.天马黑猫改派拒收) as '天马黑猫改派拒收',
                                SUM(s1.天马黑猫改派已退货) as '天马黑猫改派已退货',
                                SUM(s1.天马黑猫改派已完成) as '天马黑猫改派已完成',
                                SUM(s1.天马黑猫改派总订单) as '天马黑猫改派总订单',
                                concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派完成签收',
                                concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派总计签收',
                                concat(ROUND(SUM(s1.天马黑猫改派已完成) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派完成占比',
                                concat(ROUND(SUM(s1.天马黑猫改派已退货) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派退货率',
                                concat(ROUND(SUM(s1.天马黑猫改派拒收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派拒收率',
                            SUM(s1.铱熙无敌黑猫改派已签收) as '铱熙无敌黑猫改派已签收',
                                SUM(s1.铱熙无敌黑猫改派拒收) as '铱熙无敌黑猫改派拒收',
                                SUM(s1.铱熙无敌黑猫改派已退货) as '铱熙无敌黑猫改派已退货',
                                SUM(s1.铱熙无敌黑猫改派已完成) as '铱熙无敌黑猫改派已完成',
                                SUM(s1.铱熙无敌黑猫改派总订单) as '铱熙无敌黑猫改派总订单',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫改派已完成) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫改派已退货) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫改派拒收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率',
                            SUM(s1.铱熙无敌新竹改派已签收) as '铱熙无敌新竹改派已签收',
                                SUM(s1.铱熙无敌新竹改派拒收) as '铱熙无敌新竹改派拒收',
                                SUM(s1.铱熙无敌新竹改派已退货) as '铱熙无敌新竹改派已退货',
                                SUM(s1.铱熙无敌新竹改派已完成) as '铱熙无敌新竹改派已完成',
                                SUM(s1.铱熙无敌新竹改派总订单) as '铱熙无敌新竹改派总订单',
                                concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌新竹改派已完成) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌新竹改派已退货) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派退货率',
                                concat(ROUND(SUM(s1.铱熙无敌新竹改派拒收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派拒收率'
                        FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
                                    SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-711超商" ,1,0)) AS 速派711超商总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派711超商已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派711超商拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派711超商已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派711超商已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" ,1,0)) AS 速派黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派宅配通" ,1,0)) AS 速派宅配通总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" ,1,0)) AS 铱熙无敌711超商总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已签收",1,0)) as 铱熙无敌711超商已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "拒收",1,0)) as 铱熙无敌711超商拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已退货",1,0)) as 铱熙无敌711超商已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌711超商已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" ,1,0)) AS 铱熙无敌黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" ,1,0)) AS 铱熙无敌宅配通总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已签收",1,0)) as 铱熙无敌宅配通已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "拒收",1,0)) as 铱熙无敌宅配通拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已退货",1,0)) as 铱熙无敌宅配通已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌宅配通已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" ,1,0)) AS 铱熙无敌尾总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已签收",1,0)) as 铱熙无敌尾已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "拒收",1,0)) as 铱熙无敌尾拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已退货",1,0)) as 铱熙无敌尾已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌尾已完成,
                                    SUM(IF(cx.物流渠道 = "龟山" ,1,0)) AS 龟山改派总订单,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
                                    SUM(IF(cx.物流渠道 = "森鸿" ,1,0)) AS 森鸿改派总订单,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派新竹" ,1,0)) AS 速派新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派宅配通" ,1,0)) AS 速派宅配通改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派黑猫" ,1,0)) AS 速派黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马黑猫" ,1,0)) AS 天马黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" ,1,0)) AS 铱熙无敌黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹改派已完成
        				        FROM (SELECT *, 所属团队 as 家族
                                    FROM {0}_zqsb cc where  cc.`是否改派` = '直发' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                                ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.旬 != '合计'
                ORDER BY FIELD(s.`家族`,{5},'合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df19 = pd.read_sql_query(sql=sql19, con=self.engine1)
        listT.append(df19)

        # 产品整月_改派 台湾
        print('正在获取---产品整月_直发 台湾…………')
        sql20 = '''SELECT *
                        FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                                    IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                                    SUM(s1.已签收) as 已签收, SUM(s1.拒收) as 拒收, SUM(s1.已退货) as 已退货, SUM(s1.已完成) as 已完成, SUM(s1.总订单) as 总订单,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						            concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                                SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
                                    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
                                    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
                                    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
                                    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
                                    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
                                    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
                                    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
                                    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
                                    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
                                SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
                                    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
                                    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
                                    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
                                    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
                                    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
                                    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
                                    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
                                    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
                                    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
                                SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
                                    SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
                                    SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
                                    SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
                                    SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
                                    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
                                    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
                                    concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
                                    concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
                                    concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
                                SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
                                    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
                                    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
                                    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
                                    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
                                    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
                                    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
                                    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
                                    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
                                    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
                                SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
                                    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
                                    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
                                    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
                                    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
                                    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
                                    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
                                    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
                                    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
                                    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
                                SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
                                    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
                                    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
                                    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
                                    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
                                    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
                                    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
                                    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
                                    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
                                    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
                                SUM(s1.速派711超商已签收) as '台湾-速派-711超商已签收',
                                    SUM(s1.速派711超商拒收) as '台湾-速派-711超商拒收',
                                    SUM(s1.速派711超商已退货) as '台湾-速派-711超商已退货',
                                    SUM(s1.速派711超商已完成) as '台湾-速派-711超商已完成',
                                    SUM(s1.速派711超商总订单) as '台湾-速派-711超商总订单',
                                    concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
                                    concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
                                    concat(ROUND(SUM(s1.速派711超商已完成) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
                                    concat(ROUND(SUM(s1.速派711超商已退货) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
                                    concat(ROUND(SUM(s1.速派711超商拒收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
                                SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
                                    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
                                    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
                                    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
                                    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
                                    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
                                    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
                                    concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
                                    concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
                                    concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
                                SUM(s1.速派黑猫已签收) as '台湾-速派-黑猫已签收',
                                    SUM(s1.速派黑猫拒收) as '台湾-速派-黑猫拒收',
                                    SUM(s1.速派黑猫已退货) as '台湾-速派-黑猫已退货',
                                    SUM(s1.速派黑猫已完成) as '台湾-速派-黑猫已完成',
                                    SUM(s1.速派黑猫总订单) as '台湾-速派-黑猫总订单',
                                    concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫完成签收',
                                    concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫总计签收',
                                    concat(ROUND(SUM(s1.速派黑猫已完成) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫完成占比',
                                    concat(ROUND(SUM(s1.速派黑猫已退货) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫退货率',
                                    concat(ROUND(SUM(s1.速派黑猫拒收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫拒收率',
                                SUM(s1.速派宅配通已签收) as '台湾-速派宅配通已签收',
                                    SUM(s1.速派宅配通拒收) as '台湾-速派宅配通拒收',
                                    SUM(s1.速派宅配通已退货) as '台湾-速派宅配通已退货',
                                    SUM(s1.速派宅配通已完成) as '台湾-速派宅配通已完成',
                                    SUM(s1.速派宅配通总订单) as '台湾-速派宅配通总订单',
                                    concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通完成签收',
                                    concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通总计签收',
                                    concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通完成占比',
                                    concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通退货率',
                                    concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通拒收率',
                                SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
                                    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
                                    SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
                                    SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
                                    SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
                                    concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
                                    concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
                                    concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
                                    concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
                                    concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
                                SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
                                    SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
                                    SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
                                    SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
                                    SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
                                    concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
                                    concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
                                    concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
                                    concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
                                    concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
                                SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
                                    SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
                                    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
                                    SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
                                    SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
                                    concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
                                    concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
                                    concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
                                    concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
                                    concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
                                SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
                                    SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
                                    SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
                                    SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
                                    SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
                                    concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
                                    concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
                                    concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
                                    concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
                                    concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
                                SUM(s1.铱熙无敌新竹已签收) as '铱熙无敌-新竹已签收',
                                    SUM(s1.铱熙无敌新竹拒收) as '铱熙无敌-新竹拒收',
                                    SUM(s1.铱熙无敌新竹已退货) as '铱熙无敌-新竹已退货',
                                    SUM(s1.铱熙无敌新竹已完成) as '铱熙无敌-新竹已完成',
                                    SUM(s1.铱熙无敌新竹总订单) as '铱熙无敌-新竹总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹已完成) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹已退货) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹拒收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹拒收率',
                                SUM(s1.铱熙无敌711超商已签收) as '铱熙无敌-711超商已签收',
                                    SUM(s1.铱熙无敌711超商拒收) as '铱熙无敌-711超商拒收',
                                    SUM(s1.铱熙无敌711超商已退货) as '铱熙无敌-711超商已退货',
                                    SUM(s1.铱熙无敌711超商已完成) as '铱熙无敌-711超商已完成',
                                    SUM(s1.铱熙无敌711超商总订单) as '铱熙无敌-711超商总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌711超商已完成) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌711超商已退货) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌711超商拒收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商拒收率',
                                SUM(s1.铱熙无敌黑猫已签收) as '铱熙无敌-黑猫已签收',
                                    SUM(s1.铱熙无敌黑猫拒收) as '铱熙无敌-黑猫拒收',
                                    SUM(s1.铱熙无敌黑猫已退货) as '铱熙无敌-黑猫已退货',
                                    SUM(s1.铱熙无敌黑猫已完成) as '铱熙无敌-黑猫已完成',
                                    SUM(s1.铱熙无敌黑猫总订单) as '铱熙无敌-黑猫总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫已完成) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫已退货) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫拒收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫拒收率',
                                SUM(s1.铱熙无敌宅配通已签收) as '铱熙无敌-宅配通已签收',
                                    SUM(s1.铱熙无敌宅配通拒收) as '铱熙无敌-宅配通拒收', SUM(s1.铱熙无敌宅配通已退货) as '铱熙无敌-宅配通已退货',
                                    SUM(s1.铱熙无敌宅配通已完成) as '铱熙无敌-宅配通已完成', SUM(s1.铱熙无敌宅配通总订单) as '铱熙无敌-宅配通总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌宅配通已完成) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌宅配通已退货) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌宅配通拒收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通拒收率',
                                SUM(s1.铱熙无敌尾已签收) as '易速配头程-铱熙无敌尾已签收',
                                    SUM(s1.铱熙无敌尾拒收) as '易速配头程-铱熙无敌尾拒收',
                                    SUM(s1.铱熙无敌尾已退货) as '易速配头程-铱熙无敌尾已退货',
                                    SUM(s1.铱熙无敌尾已完成) as '易速配头程-铱熙无敌尾已完成',
                                    SUM(s1.铱熙无敌尾总订单) as '易速配头程-铱熙无敌尾总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌尾已完成) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌尾已退货) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌尾拒收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾拒收率',
                                SUM(s1.龟山改派已签收) as '龟山改派已签收',
                                    SUM(s1.龟山改派拒收) as '龟山改派拒收',
                                    SUM(s1.龟山改派已退货) as '龟山改派已退货',
                                    SUM(s1.龟山改派已完成) as '龟山改派已完成',
                                    SUM(s1.龟山改派总订单) as '龟山改派总订单',
                                    concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
                                    concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
                                    concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
                                    concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
                                    concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
                                SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
                                    SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
                                    SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
                                    SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
                                    SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
                                    concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
                                    concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
                                    concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
                                    concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
                                    concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
                                SUM(s1.速派新竹改派已签收) as '速派新竹改派已签收',
                                    SUM(s1.速派新竹改派拒收) as '速派新竹改派拒收',
                                    SUM(s1.速派新竹改派已退货) as '速派新竹改派已退货',
                                    SUM(s1.速派新竹改派已完成) as '速派新竹改派已完成',
                                    SUM(s1.速派新竹改派总订单) as '速派新竹改派总订单',
                                    concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派完成签收',
                                    concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派总计签收',
                                    concat(ROUND(SUM(s1.速派新竹改派已完成) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派完成占比',
                                    concat(ROUND(SUM(s1.速派新竹改派已退货) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派退货率',
                                    concat(ROUND(SUM(s1.速派新竹改派拒收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派拒收率',
                                SUM(s1.速派宅配通改派已签收) as '速派宅配通改派已签收',
                                    SUM(s1.速派宅配通改派拒收) as '速派宅配通改派拒收',
                                    SUM(s1.速派宅配通改派已退货) as '速派宅配通改派已退货',
                                    SUM(s1.速派宅配通改派已完成) as '速派宅配通改派已完成',
                                    SUM(s1.速派宅配通改派总订单) as '速派宅配通改派总订单',
                                    concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派完成签收',
                                    concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派总计签收',
                                    concat(ROUND(SUM(s1.速派宅配通改派已完成) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派完成占比',
                                    concat(ROUND(SUM(s1.速派宅配通改派已退货) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派退货率',
                                    concat(ROUND(SUM(s1.速派宅配通改派拒收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派拒收率',
                                SUM(s1.速派黑猫改派已签收) as '速派黑猫改派已签收',
                                    SUM(s1.速派黑猫改派拒收) as '速派黑猫改派拒收',
                                    SUM(s1.速派黑猫改派已退货) as '速派黑猫改派已退货',
                                    SUM(s1.速派黑猫改派已完成) as '速派黑猫改派已完成',
                                    SUM(s1.速派黑猫改派总订单) as '速派黑猫改派总订单',
                                    concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派完成签收',
                                    concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派总计签收',
                                    concat(ROUND(SUM(s1.速派黑猫改派已完成) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派完成占比',
                                    concat(ROUND(SUM(s1.速派黑猫改派已退货) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派退货率',
                                    concat(ROUND(SUM(s1.速派黑猫改派拒收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派拒收率',
                                SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
                                    SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
                                    SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
                                    SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
                                    SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
                                    concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
                                    concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
                                    concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
                                    concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
                                    concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
                                SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
                                    SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
                                    SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
                                    SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
                                    SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
                                    concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
                                    concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
                                    concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
                                    concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
                                    concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
                                SUM(s1.天马黑猫改派已签收) as '天马黑猫改派已签收',
                                    SUM(s1.天马黑猫改派拒收) as '天马黑猫改派拒收',
                                    SUM(s1.天马黑猫改派已退货) as '天马黑猫改派已退货',
                                    SUM(s1.天马黑猫改派已完成) as '天马黑猫改派已完成',
                                    SUM(s1.天马黑猫改派总订单) as '天马黑猫改派总订单',
                                    concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派完成签收',
                                    concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派总计签收',
                                    concat(ROUND(SUM(s1.天马黑猫改派已完成) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派完成占比',
                                    concat(ROUND(SUM(s1.天马黑猫改派已退货) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派退货率',
                                    concat(ROUND(SUM(s1.天马黑猫改派拒收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派拒收率',
                                SUM(s1.铱熙无敌黑猫改派已签收) as '铱熙无敌黑猫改派已签收',
                                    SUM(s1.铱熙无敌黑猫改派拒收) as '铱熙无敌黑猫改派拒收',
                                    SUM(s1.铱熙无敌黑猫改派已退货) as '铱熙无敌黑猫改派已退货',
                                    SUM(s1.铱熙无敌黑猫改派已完成) as '铱熙无敌黑猫改派已完成',
                                    SUM(s1.铱熙无敌黑猫改派总订单) as '铱熙无敌黑猫改派总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫改派已完成) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫改派已退货) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌黑猫改派拒收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率',
                                SUM(s1.铱熙无敌新竹改派已签收) as '铱熙无敌新竹改派已签收',
                                    SUM(s1.铱熙无敌新竹改派拒收) as '铱熙无敌新竹改派拒收',
                                    SUM(s1.铱熙无敌新竹改派已退货) as '铱熙无敌新竹改派已退货',
                                    SUM(s1.铱熙无敌新竹改派已完成) as '铱熙无敌新竹改派已完成',
                                    SUM(s1.铱熙无敌新竹改派总订单) as '铱熙无敌新竹改派总订单',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派完成签收',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派总计签收',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹改派已完成) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派完成占比',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹改派已退货) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派退货率',
                                    concat(ROUND(SUM(s1.铱熙无敌新竹改派拒收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派拒收率'
                            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,IFNULL(cx.币种, '合计') 地区,IFNULL(cx.`年月`, '合计') 月份,IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,IFNULL(cx.父级分类, '合计') 父级分类,IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
                                    SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-711超商" ,1,0)) AS 速派711超商总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派711超商已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派711超商拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派711超商已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派711超商已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" ,1,0)) AS 速派黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派宅配通" ,1,0)) AS 速派宅配通总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" ,1,0)) AS 铱熙无敌711超商总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已签收",1,0)) as 铱熙无敌711超商已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "拒收",1,0)) as 铱熙无敌711超商拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已退货",1,0)) as 铱熙无敌711超商已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌711超商已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" ,1,0)) AS 铱熙无敌黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" ,1,0)) AS 铱熙无敌宅配通总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已签收",1,0)) as 铱熙无敌宅配通已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "拒收",1,0)) as 铱熙无敌宅配通拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已退货",1,0)) as 铱熙无敌宅配通已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌宅配通已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" ,1,0)) AS 铱熙无敌尾总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已签收",1,0)) as 铱熙无敌尾已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "拒收",1,0)) as 铱熙无敌尾拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已退货",1,0)) as 铱熙无敌尾已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌尾已完成,
                                    SUM(IF(cx.物流渠道 = "龟山" ,1,0)) AS 龟山改派总订单,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
                                    SUM(IF(cx.物流渠道 = "森鸿" ,1,0)) AS 森鸿改派总订单,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派新竹" ,1,0)) AS 速派新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派宅配通" ,1,0)) AS 速派宅配通改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派黑猫" ,1,0)) AS 速派黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马黑猫" ,1,0)) AS 天马黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" ,1,0)) AS 铱熙无敌黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹改派已完成
        				            FROM (SELECT *, 所属团队 as 家族
                                        FROM {0}_zqsb cc where cc.`是否改派` = '改派' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                                    ) cx WHERE cx.`币种` = '台湾'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                        ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`,{5},'合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df20 = pd.read_sql_query(sql=sql20, con=self.engine1)
        listT.append(df20)
        # 产品分旬_改派 台湾
        print('正在获取---产品分旬_直发 台湾…………')
        sql21 = '''SELECT *
                    FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,SUM(s1.拒收) as 拒收,SUM(s1.已退货) as 已退货,SUM(s1.已完成) as 已完成,SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						        concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                            SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
                                SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
                                SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
                                SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
                                SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
                                concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
                                concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
                                concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
                                concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
                                concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
                            SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
                                SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
                                SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
                                SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
                                SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
                                concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
                                concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
                                concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
                                concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
                                concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
                            SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
                                SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
                                SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
                                SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
                                SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
                                concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
                                concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
                                concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
                                concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
                                concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
                            SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
                                SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
                                SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
                                SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
                                SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
                                concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
                                concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
                                concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
                                concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
                                concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
                            SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
                                SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
                                SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
                                SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
                                SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
                                concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
                                concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
                                concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
                                concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
                                concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
                            SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
                                SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
                                SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
                                SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
                                SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
                                concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
                                concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
                                concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
                                concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
                                concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
                            SUM(s1.速派711超商已签收) as '台湾-速派-711超商已签收',
                                SUM(s1.速派711超商拒收) as '台湾-速派-711超商拒收',
                                SUM(s1.速派711超商已退货) as '台湾-速派-711超商已退货',
                                SUM(s1.速派711超商已完成) as '台湾-速派-711超商已完成',
                                SUM(s1.速派711超商总订单) as '台湾-速派-711超商总订单',
                                concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
                                concat(ROUND(SUM(s1.速派711超商已签收) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
                                concat(ROUND(SUM(s1.速派711超商已完成) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
                                concat(ROUND(SUM(s1.速派711超商已退货) / SUM(s1.速派711超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
                                concat(ROUND(SUM(s1.速派711超商拒收) / SUM(s1.速派711超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
                            SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
                                SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
                                SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
                                SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
                                SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
                                concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
                                concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
                                concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
                                concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
                                concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
                            SUM(s1.速派黑猫已签收) as '台湾-速派-黑猫已签收',
                                SUM(s1.速派黑猫拒收) as '台湾-速派-黑猫拒收',
                                SUM(s1.速派黑猫已退货) as '台湾-速派-黑猫已退货',
                                SUM(s1.速派黑猫已完成) as '台湾-速派-黑猫已完成',
                                SUM(s1.速派黑猫总订单) as '台湾-速派-黑猫总订单',
                                concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫完成签收',
                                concat(ROUND(SUM(s1.速派黑猫已签收) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫总计签收',
                                concat(ROUND(SUM(s1.速派黑猫已完成) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫完成占比',
                                concat(ROUND(SUM(s1.速派黑猫已退货) / SUM(s1.速派黑猫总订单) * 100,2),'%') as '台湾-速派-黑猫退货率',
                                concat(ROUND(SUM(s1.速派黑猫拒收) / SUM(s1.速派黑猫已完成) * 100,2),'%') as '台湾-速派-黑猫拒收率',
                            SUM(s1.速派宅配通已签收) as '台湾-速派宅配通已签收',
                                SUM(s1.速派宅配通拒收) as '台湾-速派宅配通拒收',
                                SUM(s1.速派宅配通已退货) as '台湾-速派宅配通已退货',
                                SUM(s1.速派宅配通已完成) as '台湾-速派宅配通已完成',
                                SUM(s1.速派宅配通总订单) as '台湾-速派宅配通总订单',
                                concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通完成签收',
                                concat(ROUND(SUM(s1.速派宅配通已签收) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通总计签收',
                                concat(ROUND(SUM(s1.速派宅配通已完成) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通完成占比',
                                concat(ROUND(SUM(s1.速派宅配通已退货) / SUM(s1.速派宅配通总订单) * 100,2),'%') as '台湾-速派宅配通退货率',
                                concat(ROUND(SUM(s1.速派宅配通拒收) / SUM(s1.速派宅配通已完成) * 100,2),'%') as '台湾-速派宅配通拒收率',
                            SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
                                SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
                                SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
                                SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
                                SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
                                concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
                                concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
                                concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
                                concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
                                concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
                            SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
                                SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
                                SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
                                SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
                                SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
                                concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
                                concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
                                concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
                                concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
                                concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
                            SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
                                SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
                                SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
                                SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
                                SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
                                concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
                                concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
                                concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
                                concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
                                concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
                            SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
                                SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
                                SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
                                SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
                                SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
                                concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
                                concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
                                concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
                                concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
                                concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
                            SUM(s1.铱熙无敌新竹已签收) as '铱熙无敌-新竹已签收',
                                SUM(s1.铱熙无敌新竹拒收) as '铱熙无敌-新竹拒收',
                                SUM(s1.铱熙无敌新竹已退货) as '铱熙无敌-新竹已退货',
                                SUM(s1.铱熙无敌新竹已完成) as '铱熙无敌-新竹已完成',
                                SUM(s1.铱熙无敌新竹总订单) as '铱熙无敌-新竹总订单',
                                concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌新竹已签收) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌新竹已完成) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌新竹已退货) / SUM(s1.铱熙无敌新竹总订单) * 100,2),'%') as '铱熙无敌-新竹退货率',
                                concat(ROUND(SUM(s1.铱熙无敌新竹拒收) / SUM(s1.铱熙无敌新竹已完成) * 100,2),'%') as '铱熙无敌-新竹拒收率',
                            SUM(s1.铱熙无敌711超商已签收) as '铱熙无敌-711超商已签收',
                                SUM(s1.铱熙无敌711超商拒收) as '铱熙无敌-711超商拒收',
                                SUM(s1.铱熙无敌711超商已退货) as '铱熙无敌-711超商已退货',
                                SUM(s1.铱熙无敌711超商已完成) as '铱熙无敌-711超商已完成',
                                SUM(s1.铱熙无敌711超商总订单) as '铱熙无敌-711超商总订单',
                                concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌711超商已签收) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌711超商已完成) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌711超商已退货) / SUM(s1.铱熙无敌711超商总订单) * 100,2),'%') as '铱熙无敌-711超商退货率',
                                concat(ROUND(SUM(s1.铱熙无敌711超商拒收) / SUM(s1.铱熙无敌711超商已完成) * 100,2),'%') as '铱熙无敌-711超商拒收率',
                            SUM(s1.铱熙无敌黑猫已签收) as '铱熙无敌-黑猫已签收',
                                SUM(s1.铱熙无敌黑猫拒收) as '铱熙无敌-黑猫拒收',
                                SUM(s1.铱熙无敌黑猫已退货) as '铱熙无敌-黑猫已退货',
                                SUM(s1.铱熙无敌黑猫已完成) as '铱熙无敌-黑猫已完成',
                                SUM(s1.铱熙无敌黑猫总订单) as '铱熙无敌-黑猫总订单',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫已签收) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫已完成) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫已退货) / SUM(s1.铱熙无敌黑猫总订单) * 100,2),'%') as '铱熙无敌-黑猫退货率',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫拒收) / SUM(s1.铱熙无敌黑猫已完成) * 100,2),'%') as '铱熙无敌-黑猫拒收率',
                            SUM(s1.铱熙无敌宅配通已签收) as '铱熙无敌-宅配通已签收',
                                SUM(s1.铱熙无敌宅配通拒收) as '铱熙无敌-宅配通拒收', SUM(s1.铱熙无敌宅配通已退货) as '铱熙无敌-宅配通已退货',
                                SUM(s1.铱熙无敌宅配通已完成) as '铱熙无敌-宅配通已完成', SUM(s1.铱熙无敌宅配通总订单) as '铱熙无敌-宅配通总订单',
                                concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌宅配通已签收) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌宅配通已完成) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌宅配通已退货) / SUM(s1.铱熙无敌宅配通总订单) * 100,2),'%') as '铱熙无敌-宅配通退货率',
                                concat(ROUND(SUM(s1.铱熙无敌宅配通拒收) / SUM(s1.铱熙无敌宅配通已完成) * 100,2),'%') as '铱熙无敌-宅配通拒收率',
                            SUM(s1.铱熙无敌尾已签收) as '易速配头程-铱熙无敌尾已签收',
                                SUM(s1.铱熙无敌尾拒收) as '易速配头程-铱熙无敌尾拒收',
                                SUM(s1.铱熙无敌尾已退货) as '易速配头程-铱熙无敌尾已退货',
                                SUM(s1.铱熙无敌尾已完成) as '易速配头程-铱熙无敌尾已完成',
                                SUM(s1.铱熙无敌尾总订单) as '易速配头程-铱熙无敌尾总订单',
                                concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌尾已签收) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌尾已完成) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌尾已退货) / SUM(s1.铱熙无敌尾总订单) * 100,2),'%') as '易速配头程-铱熙无敌尾退货率',
                                concat(ROUND(SUM(s1.铱熙无敌尾拒收) / SUM(s1.铱熙无敌尾已完成) * 100,2),'%') as '易速配头程-铱熙无敌尾拒收率',
                            SUM(s1.龟山改派已签收) as '龟山改派已签收',
                                SUM(s1.龟山改派拒收) as '龟山改派拒收',
                                SUM(s1.龟山改派已退货) as '龟山改派已退货',
                                SUM(s1.龟山改派已完成) as '龟山改派已完成',
                                SUM(s1.龟山改派总订单) as '龟山改派总订单',
                                concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
                                concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
                                concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
                                concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
                                concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
                            SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
                                SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
                                SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
                                SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
                                SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
                                concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
                                concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
                                concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
                                concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
                                concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
                            SUM(s1.速派新竹改派已签收) as '速派新竹改派已签收',
                                SUM(s1.速派新竹改派拒收) as '速派新竹改派拒收',
                                SUM(s1.速派新竹改派已退货) as '速派新竹改派已退货',
                                SUM(s1.速派新竹改派已完成) as '速派新竹改派已完成',
                                SUM(s1.速派新竹改派总订单) as '速派新竹改派总订单',
                                concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派完成签收',
                                concat(ROUND(SUM(s1.速派新竹改派已签收) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派总计签收',
                                concat(ROUND(SUM(s1.速派新竹改派已完成) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派完成占比',
                                concat(ROUND(SUM(s1.速派新竹改派已退货) / SUM(s1.速派新竹改派总订单) * 100,2),'%') as '速派新竹改派退货率',
                                concat(ROUND(SUM(s1.速派新竹改派拒收) / SUM(s1.速派新竹改派已完成) * 100,2),'%') as '速派新竹改派拒收率',
                            SUM(s1.速派宅配通改派已签收) as '速派宅配通改派已签收',
                                SUM(s1.速派宅配通改派拒收) as '速派宅配通改派拒收',
                                SUM(s1.速派宅配通改派已退货) as '速派宅配通改派已退货',
                                SUM(s1.速派宅配通改派已完成) as '速派宅配通改派已完成',
                                SUM(s1.速派宅配通改派总订单) as '速派宅配通改派总订单',
                                concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派完成签收',
                                concat(ROUND(SUM(s1.速派宅配通改派已签收) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派总计签收',
                                concat(ROUND(SUM(s1.速派宅配通改派已完成) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派完成占比',
                                concat(ROUND(SUM(s1.速派宅配通改派已退货) / SUM(s1.速派宅配通改派总订单) * 100,2),'%') as '速派宅配通改派退货率',
                                concat(ROUND(SUM(s1.速派宅配通改派拒收) / SUM(s1.速派宅配通改派已完成) * 100,2),'%') as '速派宅配通改派拒收率',
                            SUM(s1.速派黑猫改派已签收) as '速派黑猫改派已签收',
                                SUM(s1.速派黑猫改派拒收) as '速派黑猫改派拒收',
                                SUM(s1.速派黑猫改派已退货) as '速派黑猫改派已退货',
                                SUM(s1.速派黑猫改派已完成) as '速派黑猫改派已完成',
                                SUM(s1.速派黑猫改派总订单) as '速派黑猫改派总订单',
                                concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派完成签收',
                                concat(ROUND(SUM(s1.速派黑猫改派已签收) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派总计签收',
                                concat(ROUND(SUM(s1.速派黑猫改派已完成) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派完成占比',
                                concat(ROUND(SUM(s1.速派黑猫改派已退货) / SUM(s1.速派黑猫改派总订单) * 100,2),'%') as '速派黑猫改派退货率',
                                concat(ROUND(SUM(s1.速派黑猫改派拒收) / SUM(s1.速派黑猫改派已完成) * 100,2),'%') as '速派黑猫改派拒收率',
                            SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
                                SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
                                SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
                                SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
                                SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
                                concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
                                concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
                                concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
                                concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
                                concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
                            SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
                                SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
                                SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
                                SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
                                SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
                                concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
                                concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
                                concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
                                concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
                                concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率',
                            SUM(s1.天马黑猫改派已签收) as '天马黑猫改派已签收',
                                SUM(s1.天马黑猫改派拒收) as '天马黑猫改派拒收',
                                SUM(s1.天马黑猫改派已退货) as '天马黑猫改派已退货',
                                SUM(s1.天马黑猫改派已完成) as '天马黑猫改派已完成',
                                SUM(s1.天马黑猫改派总订单) as '天马黑猫改派总订单',
                                concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派完成签收',
                                concat(ROUND(SUM(s1.天马黑猫改派已签收) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派总计签收',
                                concat(ROUND(SUM(s1.天马黑猫改派已完成) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派完成占比',
                                concat(ROUND(SUM(s1.天马黑猫改派已退货) / SUM(s1.天马黑猫改派总订单) * 100,2),'%') as '天马黑猫改派退货率',
                                concat(ROUND(SUM(s1.天马黑猫改派拒收) / SUM(s1.天马黑猫改派已完成) * 100,2),'%') as '天马黑猫改派拒收率',
                            SUM(s1.铱熙无敌黑猫改派已签收) as '铱熙无敌黑猫改派已签收',
                                SUM(s1.铱熙无敌黑猫改派拒收) as '铱熙无敌黑猫改派拒收',
                                SUM(s1.铱熙无敌黑猫改派已退货) as '铱熙无敌黑猫改派已退货',
                                SUM(s1.铱熙无敌黑猫改派已完成) as '铱熙无敌黑猫改派已完成',
                                SUM(s1.铱熙无敌黑猫改派总订单) as '铱熙无敌黑猫改派总订单',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫改派已签收) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫改派已完成) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫改派已退货) / SUM(s1.铱熙无敌黑猫改派总订单) * 100,2),'%') as '铱熙无敌改派退货率',
                                concat(ROUND(SUM(s1.铱熙无敌黑猫改派拒收) / SUM(s1.铱熙无敌黑猫改派已完成) * 100,2),'%') as '铱熙无敌改派拒收率',
                            SUM(s1.铱熙无敌新竹改派已签收) as '铱熙无敌新竹改派已签收',
                                SUM(s1.铱熙无敌新竹改派拒收) as '铱熙无敌新竹改派拒收',
                                SUM(s1.铱熙无敌新竹改派已退货) as '铱熙无敌新竹改派已退货',
                                SUM(s1.铱熙无敌新竹改派已完成) as '铱熙无敌新竹改派已完成',
                                SUM(s1.铱熙无敌新竹改派总订单) as '铱熙无敌新竹改派总订单',
                                concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派完成签收',
                                concat(ROUND(SUM(s1.铱熙无敌新竹改派已签收) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派总计签收',
                                concat(ROUND(SUM(s1.铱熙无敌新竹改派已完成) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派完成占比',
                                concat(ROUND(SUM(s1.铱熙无敌新竹改派已退货) / SUM(s1.铱熙无敌新竹改派总订单) * 100,2),'%') as '铱熙无敌新竹改派退货率',
                                concat(ROUND(SUM(s1.铱熙无敌新竹改派拒收) / SUM(s1.铱熙无敌新竹改派已完成) * 100,2),'%') as '铱熙无敌新竹改派拒收率'
                        FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,IFNULL(cx.币种, '合计') 地区,IFNULL(cx.`年月`, '合计') 月份,IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        							IFNULL(cx.产品id, '合计') 产品id,IFNULL(cx.产品名称, '合计') 产品名称,IFNULL(cx.父级分类, '合计') 父级分类,IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        							SUM(IF(最终状态 = "已签收",1,0)) as 已签收,SUM(IF(最终状态 = "拒收",1,0)) as 拒收,SUM(IF(最终状态 = "已退货",1,0)) as 已退货,SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
                                    SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-711超商" ,1,0)) AS 速派711超商总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派711超商已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派711超商拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派711超商已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派711超商已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" ,1,0)) AS 速派黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-速派宅配通" ,1,0)) AS 速派宅配通总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" ,1,0)) AS 铱熙无敌新竹总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" ,1,0)) AS 铱熙无敌711超商总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已签收",1,0)) as 铱熙无敌711超商已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "拒收",1,0)) as 铱熙无敌711超商拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 = "已退货",1,0)) as 铱熙无敌711超商已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌711超商已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" ,1,0)) AS 铱熙无敌黑猫总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" ,1,0)) AS 铱熙无敌宅配通总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已签收",1,0)) as 铱熙无敌宅配通已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "拒收",1,0)) as 铱熙无敌宅配通拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 = "已退货",1,0)) as 铱熙无敌宅配通已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌宅配通已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" ,1,0)) AS 铱熙无敌尾总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已签收",1,0)) as 铱熙无敌尾已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "拒收",1,0)) as 铱熙无敌尾拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 = "已退货",1,0)) as 铱熙无敌尾已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-易速配头程-铱熙无敌尾" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌尾已完成,
                                    SUM(IF(cx.物流渠道 = "龟山" ,1,0)) AS 龟山改派总订单,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
                                        SUM(IF(cx.物流渠道 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
                                    SUM(IF(cx.物流渠道 = "森鸿" ,1,0)) AS 森鸿改派总订单,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
                                        SUM(IF(cx.物流渠道 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派新竹" ,1,0)) AS 速派新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派宅配通" ,1,0)) AS 速派宅配通改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已签收",1,0)) as 速派宅配通改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "拒收",1,0)) as 速派宅配通改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 = "已退货",1,0)) as 速派宅配通改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派宅配通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派宅配通改派已完成,
                                    SUM(IF(cx.物流渠道 = "速派黑猫" ,1,0)) AS 速派黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已签收",1,0)) as 速派黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "拒收",1,0)) as 速派黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 = "已退货",1,0)) as 速派黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "速派黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成,
                                    SUM(IF(cx.物流渠道 = "天马黑猫" ,1,0)) AS 天马黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "天马黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" ,1,0)) AS 铱熙无敌黑猫改派总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌黑猫改派已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌黑猫改派拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌黑猫改派已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-黑猫改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌黑猫改派已完成,
                                    SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" ,1,0)) AS 铱熙无敌新竹改派总订单,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已签收",1,0)) as 铱熙无敌新竹改派已签收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "拒收",1,0)) as 铱熙无敌新竹改派拒收,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 = "已退货",1,0)) as 铱熙无敌新竹改派已退货,
                                        SUM(IF(cx.物流渠道 = "台湾-铱熙无敌-新竹改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 铱熙无敌新竹改派已完成
        				        FROM (SELECT *,所属团队 as 家族
                                    FROM {0}_zqsb cc where  cc.`是否改派` = '改派' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                                ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.旬 != '合计'
                ORDER BY FIELD(s.`家族`,{5},'合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df21 = pd.read_sql_query(sql=sql21, con=self.engine1)
        listT.append(df21)

        # 产品整月-直发 香港
        print('正在获取---产品整月_直发 香港…………')
        sql31 = '''SELECT *
                     FROM(SELECT IFNULL(s1.家族, '合计') 家族, IFNULL(s1.地区, '合计') 地区, IFNULL(s1.月份, '合计') 月份,
                                IFNULL(s1.产品id, '合计') 产品id, IFNULL(s1.产品名称, '合计') 产品名称, IFNULL(s1.父级分类, '合计') 父级分类, IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,SUM(s1.拒收) as 拒收,SUM(s1.已退货) as 已退货,SUM(s1.已完成) as 已完成,SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
        					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
        						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                            SUM(s1.香港圆通已签收) as '香港-圆通已签收',
                                SUM(s1.香港圆通拒收) as '香港-圆通拒收',
                                SUM(s1.香港圆通已退货) as '香港-圆通拒收已退货',
                                SUM(s1.香港圆通已完成) as '香港-圆通拒收已完成',
                                SUM(s1.香港圆通总订单) as '香港-圆通拒收总订单',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收完成签收',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收总计签收',
                                concat(ROUND(SUM(s1.香港圆通已完成) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收完成占比',
                                concat(ROUND(SUM(s1.香港圆通已退货) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收退货率',
                                concat(ROUND(SUM(s1.香港圆通拒收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收拒收率',
        					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
        						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
        						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
        						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
        						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
        						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
        						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
        						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
        					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
        						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
        						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
        						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
        						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
        						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
        						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
        						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
        					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
        						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
        						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
        						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
        						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
        						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
        					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
        						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
        						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
        						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
        						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
        						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
        					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
        						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
        						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
        						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
        						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
        						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
        						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
        						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
                            SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
                                SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
                                SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
                                SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
                                SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
                                concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
                                concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
                                concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
        					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
        						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
        						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
        						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
        						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
        						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
        						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
        						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
        		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
        							SUM(cx.`价格RMB`) as 总订单金额,
        								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
        								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
        								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
                                    SUM(IF(cx.物流渠道 = "香港-圆通" ,1,0)) AS 香港圆通总订单,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已签收",1,0)) as 香港圆通已签收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "拒收",1,0)) as 香港圆通拒收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已退货",1,0)) as 香港圆通已退货,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 香港圆通已完成,
        							SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
        							SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
                                    SUM(IF(cx.物流渠道 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
        							SUM(IF(cx.物流渠道 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
        				            FROM (SELECT *, 所属团队 as 家族
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.`是否改派` = '直发' AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                                    ) cx WHERE cx.`币种` = '香港'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                            ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`,{5},'合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df31 = pd.read_sql_query(sql=sql31, con=self.engine1)
        listT.append(df31)
        # 产品分旬-直发 香港
        print('正在获取---产品分旬_直发 香港…………')
        sql32 = '''SELECT *
                    FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
        					SUM(s1.已签收) as 已签收,SUM(s1.拒收) as 拒收,SUM(s1.已退货) as 已退货,SUM(s1.已完成) as 已完成,SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
        					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
        						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                            SUM(s1.香港圆通已签收) as '香港-圆通已签收',
                                SUM(s1.香港圆通拒收) as '香港-圆通拒收',
                                SUM(s1.香港圆通已退货) as '香港-圆通拒收已退货',
                                SUM(s1.香港圆通已完成) as '香港-圆通拒收已完成',
                                SUM(s1.香港圆通总订单) as '香港-圆通拒收总订单',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收完成签收',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收总计签收',
                                concat(ROUND(SUM(s1.香港圆通已完成) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收完成占比',
                                concat(ROUND(SUM(s1.香港圆通已退货) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收退货率',
                                concat(ROUND(SUM(s1.香港圆通拒收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收拒收率',
        					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
        						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
        						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
        						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
        						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
        						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
        						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
        						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
        					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
        						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
        						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
        						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
        						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
        						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
        						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
        						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
        					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
        						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
        						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
        						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
        						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
        						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
        					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
        						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
        						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
        						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
        						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
        						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
        					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
        						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
        						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
        						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
        						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
        						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
        						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
        						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
                            SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
                                SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
                                SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
                                SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
                                SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
                                concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
                                concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
                                concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
        					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
        						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
        						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
        						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
        						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
        						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
        						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
        						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
        		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
        							SUM(cx.`价格RMB`) as 总订单金额,
        								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
        								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
        								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
                                    SUM(IF(cx.物流渠道 = "香港-圆通" ,1,0)) AS 香港圆通总订单,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已签收",1,0)) as 香港圆通已签收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "拒收",1,0)) as 香港圆通拒收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已退货",1,0)) as 香港圆通已退货,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 香港圆通已完成,
        							SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
        							SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
                                    SUM(IF(cx.物流渠道 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
        							SUM(IF(cx.物流渠道 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
        				        FROM (SELECT *, 所属团队 as 家族
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.`是否改派` = '直发' AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                                ) cx WHERE cx.`币种` = '香港'
                                GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                            GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                            WITH ROLLUP 
                    ) s HAVING s.旬 <> '合计'
                ORDER BY FIELD(s.`家族`,{5},'合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df32 = pd.read_sql_query(sql=sql32, con=self.engine1)
        listT.append(df32)

        # 产品整月-改派 香港
        print('正在获取---产品整月_改派 香港…………')
        sql41 = '''SELECT *
                 FROM(SELECT IFNULL(s1.家族, '合计') 家族, IFNULL(s1.地区, '合计') 地区, IFNULL(s1.月份, '合计') 月份,
                            IFNULL(s1.产品id, '合计') 产品id, IFNULL(s1.产品名称, '合计') 产品名称,  IFNULL(s1.父级分类, '合计') 父级分类, IFNULL(s1.二级分类, '合计') 二级分类,
        					SUM(s1.已签收) as 已签收,SUM(s1.拒收) as 拒收,SUM(s1.已退货) as 已退货,SUM(s1.已完成) as 已完成,SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
        					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
        						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                            SUM(s1.香港圆通已签收) as '香港-圆通已签收',
                                SUM(s1.香港圆通拒收) as '香港-圆通拒收',
                                SUM(s1.香港圆通已退货) as '香港-圆通拒收已退货',
                                SUM(s1.香港圆通已完成) as '香港-圆通拒收已完成',
                                SUM(s1.香港圆通总订单) as '香港-圆通拒收总订单',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收完成签收',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收总计签收',
                                concat(ROUND(SUM(s1.香港圆通已完成) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收完成占比',
                                concat(ROUND(SUM(s1.香港圆通已退货) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收退货率',
                                concat(ROUND(SUM(s1.香港圆通拒收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收拒收率',
        					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
        						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
        						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
        						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
        						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
        						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
        						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
        						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
        					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
        						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
        						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
        						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
        						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
        						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
        						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
        						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
        					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
        						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
        						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
        						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
        						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
        						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
        					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
        						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
        						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
        						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
        						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
        						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
        					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
        						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
        						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
        						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
        						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
        						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
        						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
        						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
                            SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
                                SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
                                SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
                                SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
                                SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
                                concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
                                concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
                                concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
        					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
        						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
        						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
        						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
        						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
        						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
        						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
        						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
        		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,IFNULL(cx.币种, '合计') 地区,IFNULL(cx.`年月`, '合计') 月份,
        								IFNULL(cx.产品id, '合计') 产品id,IFNULL(cx.产品名称, '合计') 产品名称,IFNULL(cx.父级分类, '合计') 父级分类,IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
        							SUM(cx.`价格RMB`) as 总订单金额,
        								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
        								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
        								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
                                    SUM(IF(cx.物流渠道 = "香港-圆通" ,1,0)) AS 香港圆通总订单,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已签收",1,0)) as 香港圆通已签收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "拒收",1,0)) as 香港圆通拒收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已退货",1,0)) as 香港圆通已退货,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 香港圆通已完成,
        							SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
        							SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
                                    SUM(IF(cx.物流渠道 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
        							SUM(IF(cx.物流渠道 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
        				            FROM (SELECT *, 所属团队 as 家族
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.`是否改派` = '改派' AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                                    ) cx WHERE cx.`币种` = '香港'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                            ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`, {5},'合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df41 = pd.read_sql_query(sql=sql41, con=self.engine1)
        listT.append(df41)
        # 产品分旬-改派 香港
        print('正在获取---产品分旬_改派 香港…………')
        sql42 = '''SELECT *
                 FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
        					IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
        					SUM(s1.已签收) as 已签收,SUM(s1.拒收) as 拒收,SUM(s1.已退货) as 已退货,SUM(s1.已完成) as 已完成,SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
        					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
        						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
        						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
                            SUM(s1.香港圆通已签收) as '香港-圆通已签收',
                                SUM(s1.香港圆通拒收) as '香港-圆通拒收',
                                SUM(s1.香港圆通已退货) as '香港-圆通拒收已退货',
                                SUM(s1.香港圆通已完成) as '香港-圆通拒收已完成',
                                SUM(s1.香港圆通总订单) as '香港-圆通拒收总订单',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收完成签收',
                                concat(ROUND(SUM(s1.香港圆通已签收) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收总计签收',
                                concat(ROUND(SUM(s1.香港圆通已完成) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收完成占比',
                                concat(ROUND(SUM(s1.香港圆通已退货) / SUM(s1.香港圆通总订单) * 100,2),'%') as '香港-圆通拒收退货率',
                                concat(ROUND(SUM(s1.香港圆通拒收) / SUM(s1.香港圆通已完成) * 100,2),'%') as '香港-圆通拒收拒收率',
        					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
        						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
        						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
        						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
        						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
        						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
        						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
        						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
        						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
        					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
        						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
        						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
        						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
        						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
        						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
        						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
        						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
        						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
        					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
        						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
        						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
        						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
        						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
        						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
        					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
        						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
        						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
        						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
        						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
        						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
        						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
        						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
        					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
        						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
        						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
        						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
        						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
        						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
        						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
        						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
        						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
                            SUM(s1.圆通改派已签收) as '香港-圆通-改派已签收',
                                SUM(s1.圆通改派拒收) as '香港-圆通-改派拒收',
                                SUM(s1.圆通改派已退货) as '香港-圆通-改派已退货',
                                SUM(s1.圆通改派已完成) as '香港-圆通-改派已完成',
                                SUM(s1.圆通改派总订单) as '香港-圆通-改派总订单',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派完成签收',
                                concat(ROUND(SUM(s1.圆通改派已签收) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派总计签收',
                                concat(ROUND(SUM(s1.圆通改派已完成) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派完成占比',
                                concat(ROUND(SUM(s1.圆通改派已退货) / SUM(s1.圆通改派总订单) * 100,2),'%') as '香港-圆通-改派退货率',
                                concat(ROUND(SUM(s1.圆通改派拒收) / SUM(s1.圆通改派已完成) * 100,2),'%') as '香港-圆通-改派拒收率',
        					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
        						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
        						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
        						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
        						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
        						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
        						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
        						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
        						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
        		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
        							SUM(cx.`价格RMB`) as 总订单金额,
        								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
        								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
        								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
                                    SUM(IF(cx.物流渠道 = "香港-圆通" ,1,0)) AS 香港圆通总订单,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已签收",1,0)) as 香港圆通已签收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "拒收",1,0)) as 香港圆通拒收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 = "已退货",1,0)) as 香港圆通已退货,
                                        SUM(IF(cx.物流渠道 = "香港-圆通" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 香港圆通已完成,
        							SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
        							SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
        								SUM(IF(cx.物流渠道 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
        							SUM(IF(cx.物流渠道 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
        								SUM(IF(cx.物流渠道 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
                                    SUM(IF(cx.物流渠道 = "香港-圆通-改派" ,1,0)) AS 圆通改派总订单,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已签收",1,0)) as 圆通改派已签收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "拒收",1,0)) as 圆通改派拒收,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 = "已退货",1,0)) as 圆通改派已退货,
                                        SUM(IF(cx.物流渠道 = "香港-圆通-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 圆通改派已完成,
        							SUM(IF(cx.物流渠道 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
        								SUM(IF(cx.物流渠道 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
        				        FROM (SELECT *,所属团队 as 家族
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.`是否改派` = '改派' AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.付款方式 in ({3})
                                ) cx WHERE cx.`币种` = '香港'
                                GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                            GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                            WITH ROLLUP 
                    ) s HAVING s.旬 <> '合计'
                ORDER BY FIELD(s.`家族`,{5},'合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday, currency, 'team_name', self.team_name2)
        df42 = pd.read_sql_query(sql=sql42, con=self.engine1)
        listT.append(df42)

        today = datetime.date.today().strftime('%Y.%m.%d')
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾', '产品分旬台湾', '产品整月香港', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾']
        print('正在将物流品类写入excel…………')
        file_path = 'F:\\输出文件\\港台-签收率.xlsx'
        if currency_id == '全部付款':
            file_path = 'F:\\输出文件\\{} {} 物流品类-签收率.xlsx'.format(today, match[team])
        elif currency_id == '货到付款':
            file_path = 'F:\\输出文件\\{} {} 物流品类-签收率-COD.xlsx'.format(today, match[team])
        elif currency_id == '在线付款':
            file_path = 'F:\\输出文件\\{} {} 物流品类-签收率-在线.xlsx'.format(today, match[team])
        # df0 = pd.DataFrame([])                                    # 创建空的dataframe数据框
        # df0.to_excel(file_path, index=False)                         # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')        # 初始化写入对象
        # book = load_workbook(file_path)                             # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book                                          # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # listT[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
        # listT[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
        # listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
        # listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        # if 'Sheet1' in book.sheetnames:                                 # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            listT[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
            listT[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
            listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
            listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        print('正在运行' + match[team] + '表宏…………（xlwings方法一）')
        try:
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.screen_updating = False
            # app.display_alerts = False
            wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('zl_gat_report_new2.gat_总_品类_物流_两月签收率')()
            wbsht1.save()
            wbsht.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        new_path = 'F:\\神龙签收率\\' + (datetime.datetime.now()).strftime('%m.%d') + '\\签收率\\{} {} 物流品类-签收率.xlsx'.format(today, match[team])
        shutil.copyfile(file_path, new_path)        # copy到指定位置
        print('----已写入excel; 并复制到指定文件夹中')

        print('正在将品类分旬写入excel…………')
        if currency_id == '全部付款':
            file_path = 'F:\\输出文件\\{} {} 品类分旬-签收率.xlsx'.format(today, match[team])
        elif currency_id == '货到付款':
            file_path = 'F:\\输出文件\\{} {} 品类分旬-签收率-COD.xlsx'.format(today, match[team])
        elif currency_id == '在线付款':
            file_path = 'F:\\输出文件\\{} {} 品类分旬-签收率-在线.xlsx'.format(today, match[team])
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾', '产品整月香港', '产品分旬台湾', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾']
        # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
        # listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
            listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        print('正在运行' + match[team] + '表宏…………')
        try:
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.screen_updating = False
            # app.display_alerts = False
            wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('zl_gat_report_new2.gat_品类直发分旬签收率')()
            wbsht1.save()
            wbsht1.close()
            wbsht.save()
            wbsht.close()
            app.quit()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        new_path = 'F:\\神龙签收率\\' + (datetime.datetime.now()).strftime('%m.%d') + '\\签收率\\{} {} 品类分旬-签收率.xlsx'.format(today, match[team])
        shutil.copyfile(file_path, new_path)        # copy到指定位置
        print('----已写入excel; 并复制到指定文件夹中')

        print('正在将产品写入excel…………')
        if currency_id == '全部付款':
            file_path = 'F:\\输出文件\\{} {} 产品明细-签收率.xlsx'.format(today, match[team])
        elif currency_id == '货到付款':
            file_path = 'F:\\输出文件\\{} {} 产品明细-签收率-COD.xlsx'.format(today, match[team])
        elif currency_id == '在线付款':
            file_path = 'F:\\输出文件\\{} {} 产品明细-签收率-在线.xlsx'.format(today, match[team])
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾','产品分旬台湾',  '产品整月香港', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾', '产品月_直发香港', '产品旬_直发香港', '产品月_改派香港', '产品旬_改派香港']
        # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # listT[4].to_excel(excel_writer=writer, sheet_name=sheet_name[4], index=False)       # 产品整月台湾
        # listT[5].to_excel(excel_writer=writer, sheet_name=sheet_name[5], index=False)       # 产品分旬台湾
        # listT[6].to_excel(excel_writer=writer, sheet_name=sheet_name[6], index=False)       # 产品整月香港
        # listT[7].to_excel(excel_writer=writer, sheet_name=sheet_name[7], index=False)       # 产品分旬香港
        # listT[8].to_excel(excel_writer=writer, sheet_name=sheet_name[8], index=False)       # 产品月_直发台湾
        # listT[9].to_excel(excel_writer=writer, sheet_name=sheet_name[9], index=False)       # 产品旬_直发台湾
        # listT[10].to_excel(excel_writer=writer, sheet_name=sheet_name[10], index=False)     # 产品月_改派台湾
        # listT[11].to_excel(excel_writer=writer, sheet_name=sheet_name[11], index=False)     # 产品旬_改派台湾
        # listT[12].to_excel(excel_writer=writer, sheet_name=sheet_name[12], index=False)     # 产品月_直发香港
        # listT[13].to_excel(excel_writer=writer, sheet_name=sheet_name[13], index=False)     # 产品旬_直发香港
        # listT[14].to_excel(excel_writer=writer, sheet_name=sheet_name[14], index=False)     # 产品月_改派香港
        # listT[15].to_excel(excel_writer=writer, sheet_name=sheet_name[15], index=False)     # 产品旬_改派香港
        # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            listT[4].to_excel(excel_writer=writer, sheet_name=sheet_name[4], index=False)  # 产品整月台湾
            listT[5].to_excel(excel_writer=writer, sheet_name=sheet_name[5], index=False)  # 产品分旬台湾
            listT[6].to_excel(excel_writer=writer, sheet_name=sheet_name[6], index=False)  # 产品整月香港
            listT[7].to_excel(excel_writer=writer, sheet_name=sheet_name[7], index=False)  # 产品分旬香港
            listT[8].to_excel(excel_writer=writer, sheet_name=sheet_name[8], index=False)  # 产品月_直发台湾
            listT[9].to_excel(excel_writer=writer, sheet_name=sheet_name[9], index=False)  # 产品旬_直发台湾
            listT[10].to_excel(excel_writer=writer, sheet_name=sheet_name[10], index=False)  # 产品月_改派台湾
            listT[11].to_excel(excel_writer=writer, sheet_name=sheet_name[11], index=False)  # 产品旬_改派台湾
            listT[12].to_excel(excel_writer=writer, sheet_name=sheet_name[12], index=False)  # 产品月_直发香港
            listT[13].to_excel(excel_writer=writer, sheet_name=sheet_name[13], index=False)  # 产品旬_直发香港
            listT[14].to_excel(excel_writer=writer, sheet_name=sheet_name[14], index=False)  # 产品月_改派香港
            listT[15].to_excel(excel_writer=writer, sheet_name=sheet_name[15], index=False)  # 产品旬_改派香港
        print('正在运行' + match[team] + '表宏…………')
        try:
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.screen_updating = False
            # app.display_alerts = False
            wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('zl_gat_report_new2.gat_产品签收率_总')()
            wbsht1.save()
            wbsht1.close()
            wbsht.save()
            wbsht.close()
            app.quit()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        new_path = 'F:\\神龙签收率\\' + (datetime.datetime.now()).strftime('%m.%d') + '\\签收率\\{} {} 产品明细-签收率.xlsx'.format(today, match[team])
        shutil.copyfile(file_path, new_path)        # copy到指定位置
        print('----已写入excel; 并复制到指定文件夹中')

        print("强制关闭Execl后台进程中......")
        system('taskkill /F /IM EXCEL.EXE')

    # 新版签收率-报表(刘姐看的)- 单量计算
    def qsb_new(self, team, month_last):  # 报表各团队近两个月的物流数据
        month_now = datetime.datetime.now().strftime('%Y-%m-%d')
        gat_time = (datetime.datetime.now() - relativedelta(months=12)).strftime('%Y%m')
        match = {'gat': '港台-每日'}
        not_team = '"客服中心港台","奥创队","神龙主页运营","APP运营","Line运营","红杉港台","郑州北美","研发部港台","金鹏港台","金狮港台","翼虎港台"'
        del_time = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y%m')

        sql = '''DELETE FROM gat_zqsb gt
                WHERE gt.年月 >= {0}
                  and gt.`订单编号` IN 
                        ( SELECT 订单编号 
                            FROM gat_order_list gs
                            WHERE gs.年月 >= {0}
                              and gs.`系统订单状态` NOT IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)')
                        );'''.format(del_time)
        # print('正在清除港澳台-总表的可能删除了的订单…………')
        # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        sql = '''DELETE FROM gat_zqsb gz 
                WHERE gz.`系统订单状态` = '已转采购' and gz.`是否改派` = '改派' and gz.`审核时间` >= '{0} 00:00:00' AND gz.`日期` >= '{1}';'''.format(month_now, month_last)
        # print('正在清除不参与计算的今日改派订单…………')
        # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        # print('正在修改-港澳台-物流渠道…………')
        # self.update_logistics_name()

        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 0、每日-各团队
        print('正在获取---0、每日各团队…………')
        # sql0 = '''SELECT 月份,地区, 家族,
        #                     SUM(s.昨日订单量) as 有运单号,
        #                     SUM(s.直发签收) as 直发签收,
        #                     SUM(s.直发拒收) as 直发拒收,
        #                     SUM(s.直发完成) as 直发完成,
        #                     SUM(s.直发总订单) as 直发总订单,
        #                     concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发完成), 0) * 100,2),'%') as 直发完成签收,
        #                     concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发总订单), 0) * 100,2),'%') as 直发总计签收,
        #                     concat(ROUND(IFNULL(SUM(s.直发完成) / SUM(s.直发总订单), 0) * 100,2),'%')as 直发完成占比,
        #                     SUM(s.改派签收) as 改派签收,
        #                     SUM(s.改派拒收) as 改派拒收,
        #                     SUM(s.改派完成) as 改派完成,
        #                     SUM(s.改派总订单) as 改派总订单,
        #                     concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派完成), 0) * 100,2),'%') as 改派完成签收,
        #                     concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派总计签收,
        #                     concat(ROUND(IFNULL(SUM(s.改派完成) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派完成占比
        #             FROM( SELECT IFNULL(cx.`年月`, '总计') 月份,
        #                         IFNULL(cx.币种, '总计') 地区,
        #                         IFNULL(cx.家族, '总计') 家族,
        #                         SUM(IF(cx.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY),1,0)) as 昨日订单量,
        #                         SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) as 直发签收,
        #                         SUM(IF(`是否改派` = '直发' AND 最终状态 = "拒收",1,0)) as 直发拒收,
        #                         SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 直发完成,
        #                         SUM(IF(`是否改派` = '直发',1,0)) as 直发总订单,
        #                         SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) as 改派签收,
        #                         SUM(IF(`是否改派` = '改派' AND 最终状态 = "拒收",1,0)) as 改派拒收,
        #                         SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 改派完成,
        #                         SUM(IF(`是否改派` = '改派',1,0)) as 改派总订单
        #                     FROM (SELECT *,
        #                                 IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","金鹏",cc.团队)))))) as 家族
        #                             FROM gat_zqsb cc
        #                             where cc.日期 >= '{0}' and cc.`运单编号` is not null
        #                           ) cx
        #                     GROUP BY cx.年月,cx.币种,cx.家族
        #                     WITH ROLLUP
        #                 ) s
        #                 GROUP BY 月份,地区,家族
        #                 ORDER BY 月份 DESC,
        #                         FIELD( 地区, '台湾', '香港', '总计' ),
        #                         FIELD( 家族, '神龙', '火凤凰', '金狮', '金鹏','神龙香港', '红杉', '总计');'''.format(month_last, team)
        sql0 = '''SELECT *
                FROM (SELECT IFNULL(s.`年月`, '总计') 月份,  IFNULL(s.币种, '总计') 地区, IFNULL(s.家族, '总计') 家族,  
                            SUM(昨日单量) 昨日单量,
                            SUM(s.直发签收) as 直发签收,
                            SUM(s.直发拒收) as 直发拒收,
                            SUM(s.直发完成) as 直发完成,
                            SUM(s.直发总订单) as 直发总订单,
                            concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发完成), 0) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(IFNULL(SUM(s.直发签收) / SUM(s.直发总订单), 0) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(IFNULL(SUM(s.直发完成) / SUM(s.直发总订单), 0) * 100,2),'%')as 直发完成占比,
                            SUM(s.改派签收) as 改派签收,
                            SUM(s.改派拒收) as 改派拒收,
                            SUM(s.改派完成) as 改派完成,
                            SUM(s.改派总订单) as 改派总订单,
                            concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派完成), 0) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(IFNULL(SUM(s.改派签收) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(IFNULL(SUM(s.改派完成) / SUM(s.改派总订单), 0) * 100,2),'%') as 改派完成占比
                    FROM( SELECT cx.`年月`, cx.`币种`, cx.`家族`,   总订单量 昨日单量,
                                SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) as 直发签收,
                                SUM(IF(`是否改派` = '直发' AND 最终状态 = "拒收",1,0)) as 直发拒收,
                                SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 直发完成,
                                SUM(IF(`是否改派` = '直发',1,0)) as 直发总订单,
                                SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) as 改派签收,
                                SUM(IF(`是否改派` = '改派' AND 最终状态 = "拒收",1,0)) as 改派拒收,
                                SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 改派完成,
                                SUM(IF(`是否改派` = '改派',1,0)) as 改派总订单
                            FROM (  SELECT *, 所属团队 as 家族 
                                    FROM {1}_zqsb cc
                                    where cc.日期 >= '{0}' and cc.`运单编号` is not null AND cc.所属团队 NOT IN ({2})
                            ) cx	
                            LEFT JOIN 
                            (  SELECT 年月,币种, 所属团队 AS 家族,count(订单编号) as 总订单量
                                FROM {1}_order_list cc 
                                WHERE cc.日期 = DATE_SUB(CURDATE(), INTERVAL 1 DAY) AND cc.所属团队 NOT IN ({2})
                                GROUP BY cc.年月, cc.币种, cc.所属团队
                            ) cx2 
                            ON  cx.年月 = cx2.年月 AND cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族   
                            GROUP BY cx.年月,cx.币种,cx.家族
                    ) s						
                    GROUP BY s.年月,s.币种,s.家族
                    WITH ROLLUP 
                    HAVING `地区` <> '总计'
                ) ss					
                ORDER BY 月份 DESC,
                        FIELD( 地区, '台湾', '香港', '总计' ),
                        FIELD( 家族, {4}, '总计'),
                        直发总订单 DESC;'''.format(month_last, team, not_team, 'team_name', self.team_name2)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)

        # 01、各团队-审核率 删单率
        print('正在获取---01、各团队-审核率_删单率…………')
        sql01 = '''SELECT gs.币种,SUBSTRING(删除原因,2) as 删除原因, 
                            COUNT(订单编号) as 单量, 
                            SUM(IF(gs.`审单类型` = '是',1,0)) as 自动审单量,
                            concat(ROUND(SUM(IF(gs.`审单类型` = '是',1,0)) / 总订单量 * 100,2),'%') as 自动审单量率,
                            SUM(IF(gs.`审单类型` = '否' or gs.`审单类型` IS NULL,1,0)) as 人工审单量,
                            concat(ROUND(SUM(IF(gs.`审单类型` = '否' or gs.`审单类型` IS NULL,1,0)) / 总订单量 * 100,2),'%') as 人工审单量率,
                            SUM(IF(gs.`问题原因` IS NOT NULL,1,0)) as 问题订单量,
                            SUM(IF(gs.`问题原因` IS NOT NULL AND gs.`系统订单状态` IN ("已删除","支付失败","未支付"),1,0)) as 问题订单删单量,
                            concat(ROUND(SUM(IF(gs.`问题原因` IS NOT NULL AND gs.`系统订单状态` NOT IN ("已删除","支付失败","未支付"),1,0)) / SUM(IF(gs.`问题原因` IS NOT NULL,1,0))  * 100,2),'%') as 问题订单转化率,
                            SUM(IF(gs.`系统订单状态` = '已删除',1,0)) as 删单量,
                            concat(ROUND(SUM(IF(gs.`系统订单状态` = '已删除',1,0)) / 总订单量 * 100,2),'%') as 删单率
                    FROM (  SELECT *
                            FROM gat_order_list cc 
                            WHERE cc.日期 = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                    ) gs
                    LEFT JOIN 
                    (   SELECT 币种, COUNT(订单编号)  as 总订单量
                        FROM  gat_order_list gss
                        WHERE gss.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY) AND gss.所属团队 NOT IN ({0})
                        GROUP BY gss.`币种`
                    ) gs2 ON gs.`币种` = gs2.`币种`
                    GROUP BY gs.`币种`,gs.`删除原因`
                    WITH ROLLUP
                    HAVING gs.`币种` IS NOT null
                    ORDER BY gs.币种, 单量 DESC;'''.format(not_team)
        df01 = pd.read_sql_query(sql=sql01, con=self.engine1)
        listT.append(df01)

        # 1、各月-各团队
        print('正在获取---1、各月各团队…………')
        sql10 = '''SELECT *
                FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(家族, '总计') 家族,		
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,		
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,										
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                    FROM( SELECT 年月 月份,币种 地区, 所属团队 AS 家族,
                                    COUNT(订单编号) as 总单量,
									SUM(IF(最终状态 = "已签收",1,0)) 签收,
									SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
									SUM(IF(最终状态 = "已退货",1,0)) 退货,
									SUM(价格RMB) AS 金额,
									SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                                    SUM(IF(是否改派 = '直发',1,0))  as 直发单量,
									SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
									SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                                    SUM(IF(是否改派 = '改派',1,0))  as 改派单量,
									SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
									SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM {0}_zqsb cx   
                        WHERE cx.年月 >= '{3}' and cx.`运单编号` is not null AND cx.所属团队 NOT IN ({1})             
                        GROUP BY cx.年月,cx.币种, cx.所属团队
                    ) s
					GROUP BY 月份,地区,家族
                    WITH ROLLUP		
                ) ss			
                ORDER BY 月份 DESC,
                    FIELD( 地区, '台湾', '香港', '总计' ),
                    FIELD( 家族, {2}, '总计'),
                    总单量 DESC;'''.format(team, not_team, self.team_name2, gat_time)
        df10 = pd.read_sql_query(sql=sql10, con=self.engine1)
        listT.append(df10)
        # 2、各月各团队---分旬
        print('正在获取---2、各月各团队---分旬…………')
        sql11 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(旬, '总计') 旬,IFNULL(地区, '总计') 地区,IFNULL(家族, '总计') 家族,		
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,		
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,										
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                        FROM(SELECT 年月 月份, 旬,币种 地区, 所属团队 AS 家族,
                                    COUNT(订单编号) as 总单量,
									SUM(IF(最终状态 = "已签收",1,0)) 签收,
									SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
									SUM(IF(最终状态 = "已退货",1,0)) 退货,
									SUM(价格RMB) AS 金额,
									SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                                    SUM(IF(是否改派 = '直发',1,0))  as 直发单量,
									SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
									SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                                    SUM(IF(是否改派 = '改派',1,0))  as 改派单量,
									SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
									SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                            FROM {0}_zqsb cx       
                            WHERE cx.年月 >= '{3}' and cx.`运单编号` is not null AND cx.所属团队 NOT IN ({1})       
                            GROUP BY cx.年月,cx.旬,cx.币种, cx.所属团队
                        ) s
						GROUP BY 月份,旬,地区,家族
                        WITH ROLLUP	
                ) ss				
                ORDER BY 月份 DESC,旬,
                        FIELD( 地区, '台湾', '香港', '总计' ),
                        FIELD( 家族, {2}),
                        总单量 DESC;'''.format(team, not_team, self.team_name2, gat_time)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)
        listT.append(df11)

        # 3、各团队-各品类
        print('正在获取---3、各团队-各品类…………')
        sql20 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(s.家族, '总计') 家族,IFNULL(父级分类, '总计') 父级分类, SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							concat(ROUND(SUM(总单量) / 总订单量 * 100,2),'%') as 品类占比,						
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,									
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
							concat(ROUND(SUM(直发单量) / 直发总单量 * 100,2),'%') as 直发品类占比,
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比,
							concat(ROUND(SUM(改派单量) / 改派总单量 * 100,2),'%') as 改派品类占比
                    FROM(SELECT cx.年月 月份, cx.币种 地区, cx.所属团队 as 家族, cx.父级分类, 总订单量, 
                                COUNT(cx.订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,		
							直发总单量,							
                                SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,
							改派总单量,
                                SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM (SELECT 年月,币种,所属团队,父级分类,订单编号,最终状态,价格RMB,是否改派
							    FROM gat_zqsb cc 
								WHERE  cc.年月 >= '{4}' and  cc.`运单编号` is not null AND cc.所属团队 NOT IN ({2})
						) cx 
                        LEFT JOIN 
						( SELECT 币种,所属团队,年月,count(订单编号) as 总订单量,SUM(IF(`是否改派`= '直发',1,0)) as 直发总单量,SUM(IF(`是否改派` = '改派',1,0)) as 改派总单量
                            FROM gat_zqsb cc 
                            WHERE  cc.年月 >= '{4}' and cc.`运单编号` is not null AND cc.所属团队 NOT IN ({2})
							GROUP BY cc.币种, cc.所属团队, cc.年月
						) cx2 ON cx.币种 = cx2.币种 AND  cx.所属团队 = cx2.所属团队 AND  cx.年月 = cx2.年月   
						GROUP BY cx.年月,cx.币种,cx.所属团队,cx.父级分类
                    ) s				
					GROUP BY 月份,地区,s.家族,父级分类
                    WITH ROLLUP
            ) ss
			ORDER BY 月份 DESC,
                    FIELD( 地区, '台湾', '香港', '总计' ),
                    FIELD( 家族, {3}, '总计'),
                    FIELD( 父级分类, '居家百货', '电子电器', '服饰', '医药保健',  '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','包材类','总计' ),
                    总单量 DESC;'''.format(month_last, team, not_team, self.team_name2, gat_time)
        df20 = pd.read_sql_query(sql=sql20, con=self.engine1)
        listT.append(df20)
        # 4、各团队-各物流
        print('正在获取---4、各团队-各物流…………')
        sql21 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(是否改派, '总计') 是否改派,IFNULL(家族, '总计') 家族,IFNULL(物流方式, '总计') 物流方式,
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,											
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                FROM(SELECT 年月 月份, 币种 地区, 是否改派, 所属团队 as 家族, 物流渠道 as 物流方式,
                                COUNT(订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                            SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                            SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM gat_zqsb cx
						WHERE cx.年月 >= '{5}' and cx.`运单编号` is not null  AND cx.所属团队 NOT IN ({2})       
						GROUP BY cx.年月,cx.币种,cx.是否改派,cx.所属团队,cx.物流渠道
                ) s
				GROUP BY 月份,地区,是否改派,家族,物流方式
                WITH ROLLUP
            ) ss
            ORDER BY FIELD(月份, DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'), 
                                DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 4 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 5 MONTH),'%Y%m'), 
			                    DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 6 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 7 MONTH),'%Y%m'), '总计' ),
                            FIELD(地区, '台湾', '香港', '总计' ),
                            FIELD(是否改派, '直发', '改派', '总计' ),
                            FIELD(家族, {4}, '总计'),
                            FIELD(物流方式, {3}, '总计'),
                            总单量 DESC;'''.format(month_last, team, not_team, self.logistics_name, self.team_name2, gat_time)
        df21 = pd.read_sql_query(sql=sql21, con=self.engine1)
        listT.append(df21)

        # 5、各团队-各平台
        print('正在获取---5、各团队-各平台…………')
        sql30 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(家族, '总计') 家族,IFNULL(平台, '总计') 平台,
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,											
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                    FROM(SELECT 年月 月份, 币种 地区, 所属团队 as 家族,订单来源 平台,
                                COUNT(订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                            SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                            SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM gat_zqsb cx
						WHERE cx.年月 >= '{4}' and  cx.`运单编号` is not null AND cx.所属团队 NOT IN ({2})                 
                        GROUP BY cx.年月,cx.币种,cx.所属团队,cx.订单来源
                        ) s
					GROUP BY 月份,地区,家族,平台
                    WITH ROLLUP
            ) ss
            ORDER BY 月份 DESC,
                    FIELD(地区, '台湾', '香港', '总计' ),
                    FIELD(家族, {3}, '总计'),
                    FIELD(平台, "google","facebook","line","tiktok","clone","Criteo","interpark","kol","tiktokpage","mercado","Qoo10","ozon","w_service","s_service","e_service",
                                "allegro","hepsi","tikshop","gmarket","11st","shopline","aws","pre_sale_clone","coupang","detain_goods","aliexpress","bigo","refuse_clone","line_natural",
                                "youtube_natural","facebook_natural","instagram_natural","tiktok_natural","postsaleclone","outplay","outbrain","email","shangwutong","lazada","headline",
                                "shopee","recommend","propellerads","snapchat","tenmax","shopify","Dragon","taboola","naver","mf","速卖通发货","topbuzz",
                                "vivishop","sms","edm","facebookpage","recomm","native","twitter","yahoo","bing","总计"),
                    总单量 DESC;'''.format(month_last, team, not_team, self.team_name2, gat_time)
        df30 = pd.read_sql_query(sql=sql30, con=self.engine1)
        listT.append(df30)
        # 6、各平台-各团队
        print('正在获取---6、各平台-各团队…………')
        sql31 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(平台, '总计') 平台,IFNULL(家族, '总计') 家族,									
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,											
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                    FROM(SELECT 年月 月份, 币种 地区, 订单来源 平台,所属团队 as 家族,
                                COUNT(订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                            SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                            SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM gat_zqsb cx
						WHERE cx.年月 >= '{4}' and cx.`运单编号` is not null AND cx.所属团队 NOT IN ({2})
                        GROUP BY cx.年月,cx.币种,cx.订单来源,cx.所属团队
                        ) s
					GROUP BY 月份,地区,平台,家族
                    WITH ROLLUP
            ) ss
            ORDER BY 月份 DESC,
                    FIELD(地区, '台湾', '香港', '总计' ),
                    FIELD(平台, "google","facebook","line","tiktok","clone","Criteo","interpark","kol","tiktokpage","mercado","Qoo10","ozon","w_service","s_service","e_service",
                                "allegro","hepsi","tikshop","gmarket","11st","shopline","aws","pre_sale_clone","coupang","detain_goods","aliexpress","bigo","refuse_clone","line_natural",
                                "youtube_natural","facebook_natural","instagram_natural","tiktok_natural","postsaleclone","outplay","outbrain","email","shangwutong","lazada","headline",
                                "shopee","recommend","propellerads","snapchat","tenmax","shopify","Dragon","taboola","naver","mf","速卖通发货","topbuzz",
                                "vivishop","sms","edm","facebookpage","recomm","native","twitter","yahoo","bing","总计"),
                    FIELD(家族, {3}, '总计'),
                    总单量 DESC;'''.format(month_last, team, not_team, self.team_name2, gat_time)
        df31 = pd.read_sql_query(sql=sql31, con=self.engine1)
        listT.append(df31)

        # 7、各品类-各团队
        print('正在获取---7、各品类-各团队…………')
        sql40 = '''SELECT *
                FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(父级分类, '总计') 父级分类,IFNULL(家族, '总计') 家族,									
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,											
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                    FROM(SELECT 年月 月份, 币种 地区, 父级分类,所属团队 as 家族,
                                COUNT(订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                            SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                            SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM gat_zqsb cx
						WHERE cx.年月 >= '{4}' and cx.`运单编号` is not null  AND cx.所属团队 NOT IN ({2})               
                        GROUP BY cx.年月,cx.币种,cx.父级分类,cx.所属团队
                    ) s
					GROUP BY 月份,地区,父级分类,家族
                    WITH ROLLUP
                ) ss
                ORDER BY 月份 DESC,
                        FIELD(地区, '台湾', '香港', '总计' ),
                        FIELD(父级分类, '居家百货', '电子电器', '服饰', '医药保健',  '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','包材类','总计' ),
                        FIELD(家族, {3}, '总计'),
                        总单量 DESC;'''.format(month_last, team,not_team, self.team_name2, gat_time)
        df40 = pd.read_sql_query(sql=sql40, con=self.engine1)
        listT.append(df40)
        # 8、各物流-各团队
        print('正在获取---8、各物流-各团队…………')
        sql41 = '''SELECT *
            FROM(SELECT IFNULL(月份, '总计') 月份,IFNULL(地区, '总计') 地区,IFNULL(是否改派, '总计') 是否改派,IFNULL(物流方式, '总计') 物流方式,IFNULL(家族, '总计') 家族,									
						SUM(总单量) as 总单量,
                            concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                            concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                            concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,
                            concat(ROUND(SUM(退货) / SUM(总单量) * 100,2),'%') as 退款率,
                            concat(ROUND(SUM(签收金额) / SUM(金额) * 100,2),'%') as '总计签收(金额)',
							ROUND(SUM(金额) / SUM(总单量),2) as 平均客单价,											
						SUM(直发单量) as 直发单量,
                            concat(ROUND(SUM(直发签收) / SUM(直发完成) * 100,2),'%') as 直发完成签收,
                            concat(ROUND(SUM(直发签收) / SUM(直发单量) * 100,2),'%') as 直发总计签收,
                            concat(ROUND(SUM(直发完成) / SUM(直发单量) * 100,2),'%') as 直发完成占比,									
                        concat(ROUND(SUM(改派单量) / SUM(总单量) * 100,2),'%') as 改派占比,
                            concat(ROUND(SUM(改派签收) / SUM(改派完成) * 100,2),'%') as 改派完成签收,
                            concat(ROUND(SUM(改派签收) / SUM(改派单量) * 100,2),'%') as 改派总计签收,
                            concat(ROUND(SUM(改派完成) / SUM(改派单量) * 100,2),'%') as 改派完成占比
                    FROM(SELECT 年月 月份, 币种 地区, 是否改派, 物流渠道 as 物流方式, 所属团队 as 家族,
                                COUNT(订单编号) as 总单量,
								SUM(IF(最终状态 = "已签收",1,0)) 签收,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 完成,
								SUM(IF(最终状态 = "已退货",1,0)) 退货,
								SUM(价格RMB) AS 金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) 签收金额,									
                            SUM(IF(是否改派 = '直发',1,0)) as 直发单量,
								SUM(IF(是否改派 = '直发' AND 最终状态 = "已签收",1,0)) 直发签收,
								SUM(IF(是否改派 = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 直发完成,								
                            SUM(IF(是否改派 = '改派',1,0)) as 改派单量,
								SUM(IF(是否改派 = '改派' AND 最终状态 = "已签收",1,0)) 改派签收,
								SUM(IF(是否改派 = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) 改派完成
                        FROM gat_zqsb cx
						WHERE cx.年月 >= '{4}' and cx.`运单编号` is not null  AND cx.所属团队 NOT IN ({2})                         
                        GROUP BY cx.年月,cx.币种,cx.是否改派,cx.物流渠道,cx.所属团队
                    ) s
					GROUP BY 月份, 地区, 是否改派, 物流方式, 家族
                    WITH ROLLUP
            ) ss
            ORDER BY FIELD(月份, DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'), 
                                DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 4 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 5 MONTH),'%Y%m'), 
			                    DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 6 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 7 MONTH),'%Y%m'), '总计' ),
                    FIELD(地区, '台湾', '香港', '总计' ),
                    FIELD(是否改派, '直发', '改派', '总计' ),
                    FIELD(物流方式, {3}, '总计'),
                    FIELD(家族, {4}, '总计'),
                    总单量 DESC;'''.format(month_last, team, not_team, self.logistics_name, self.team_name2, gat_time)
        df41 = pd.read_sql_query(sql=sql41, con=self.engine1)
        listT.append(df41)

        # 9、同产品各团队的对比
        print('正在获取---9、同产品各团队的对比…………')
        sql50 = '''SELECT *, IF(神龙完成签收 = '0.00%' OR 神龙完成签收 IS NULL, 神龙完成签收, concat(ROUND(神龙完成签收-完成签收,2),'%')) as 神龙港台对比,
    			            IF(火凤凰完成签收 = '0.00%' OR 火凤凰完成签收 IS NULL, 火凤凰完成签收, concat(ROUND(火凤凰完成签收-完成签收,2),'%')) as 火凤凰港台对比,
    			            IF(雪豹完成签收 = '0.00%' OR 雪豹完成签收 IS NULL, 雪豹完成签收, concat(ROUND(雪豹完成签收-完成签收,2),'%')) as 雪豹港台对比,
    			            IF(金蝉项目完成签收 = '0.00%' OR 金蝉项目完成签收 IS NULL, 金蝉项目完成签收, concat(ROUND(金蝉项目完成签收-完成签收,2),'%')) as 金蝉项目组对比,
    			            IF(金蝉优化完成签收 = '0.00%' OR 金蝉优化完成签收 IS NULL, 金蝉优化完成签收, concat(ROUND(金蝉优化完成签收-完成签收,2),'%')) as 金蝉家族优化组对比,
    			        IF(金蝉公共完成签收 = '0.00%' OR 金蝉公共完成签收 IS NULL, 金蝉公共完成签收, concat(ROUND(金蝉公共完成签收-完成签收,2),'%')) as 金蝉家族公共团队对比,
    			            IF(客服中心完成签收 = '0.00%' OR 客服中心完成签收 IS NULL, 客服中心完成签收, concat(ROUND(客服中心完成签收-完成签收,2),'%')) as 客服中心港台对比,
    			            IF(奥创队完成签收 = '0.00%' OR 奥创队完成签收 IS NULL, 奥创队完成签收, concat(ROUND(奥创队完成签收-完成签收,2),'%')) as 奥创队对比,
    			            IF(神龙主页运营完成签收 = '0.00%' OR 神龙主页运营完成签收 IS NULL, 神龙主页运营完成签收, concat(ROUND(神龙主页运营完成签收-完成签收,2),'%')) as 神龙主页运营对比,
    			        IF(APP完成签收 = '0.00%' OR APP完成签收 IS NULL, APP完成签收, concat(ROUND(APP完成签收-完成签收,2),'%')) as APP运营对比,
    			            IF(Line完成签收 = '0.00%' OR Line完成签收 IS NULL, Line完成签收, concat(ROUND(Line完成签收-完成签收,2),'%')) as Line运营对比,
    			            IF(红杉完成签收 = '0.00%' OR 红杉完成签收 IS NULL, 红杉完成签收, concat(ROUND(红杉完成签收-完成签收,2),'%')) as 红杉港台对比,
    			            IF(郑州北美完成签收 = '0.00%' OR 郑州北美完成签收 IS NULL, 郑州北美完成签收, concat(ROUND(郑州北美完成签收-完成签收,2),'%')) as 郑州北美对比,
    			        IF(研发部完成签收 = '0.00%' OR 研发部完成签收 IS NULL, 研发部完成签收, concat(ROUND(研发部完成签收-完成签收,2),'%')) as 研发部港台对比,
    			            IF(金鹏完成签收 = '0.00%' OR 金鹏完成签收 IS NULL, 金鹏完成签收, concat(ROUND(金鹏完成签收-完成签收,2),'%')) as 金鹏港台对比,
    			            IF(金狮完成签收 = '0.00%' OR 金狮完成签收 IS NULL, 金狮完成签收, concat(ROUND(金狮完成签收-完成签收,2),'%')) as 金狮港台对比,
    			            IF(翼虎完成签收 = '0.00%' OR 翼虎完成签收 IS NULL, 翼虎完成签收, concat(ROUND(翼虎完成签收-完成签收,2),'%')) as 翼虎港台对比
					FROM(SELECT	IFNULL(月份, '总计') 月份, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id, IFNULL(产品名称, '总计') 产品名称, IFNULL(父级分类, '总计') 父级分类,
							SUM(总单量) 总单量, SUM(签收) 签收, SUM(拒收) 拒收,
                                concat(ROUND(SUM(改派) / SUM(总单量) * 100,2),'%') as 改派占比,
                                concat(ROUND(SUM(签收) / SUM(总单量) * 100,2),'%') as 总计签收,
                                concat(ROUND(SUM(签收) / SUM(完成) * 100,2),'%') as 完成签收,
                                concat(ROUND(SUM(完成) / SUM(总单量) * 100,2),'%') as 完成占比,		
							SUM(神龙单量) 神龙单量, SUM(神龙签收) 神龙签收, SUM(神龙拒收) 神龙拒收,
                                concat(ROUND(SUM(神龙改派) / SUM(神龙单量) * 100,2),'%') as 神龙改派占比,
                                concat(ROUND(SUM(神龙签收) / SUM(神龙单量) * 100,2),'%') as 神龙总计签收,
                                concat(ROUND(SUM(神龙签收) / SUM(神龙完成) * 100,2),'%') as 神龙完成签收,
                                concat(ROUND(SUM(神龙完成) / SUM(神龙单量) * 100,2),'%') as 神龙完成占比,
							SUM(火凤凰单量) 火凤凰单量, SUM(火凤凰签收) 火凤凰签收, SUM(火凤凰拒收) 火凤凰拒收,
                                concat(ROUND(SUM(火凤凰改派) / SUM(火凤凰单量) * 100,2),'%') as 火凤凰改派占比,
                                concat(ROUND(SUM(火凤凰签收) / SUM(火凤凰单量) * 100,2),'%') as 火凤凰总计签收,
                                concat(ROUND(SUM(火凤凰签收) / SUM(火凤凰完成) * 100,2),'%') as 火凤凰完成签收,
                                concat(ROUND(SUM(火凤凰完成) / SUM(火凤凰单量) * 100,2),'%') as 火凤凰完成占比,	
							SUM(雪豹单量) 雪豹单量, SUM(雪豹签收) 雪豹签收, SUM(雪豹拒收) 雪豹拒收,
                                concat(ROUND(SUM(雪豹改派) / SUM(雪豹单量) * 100,2),'%') as 雪豹改派占比,
                                concat(ROUND(SUM(雪豹签收) / SUM(雪豹单量) * 100,2),'%') as 雪豹总计签收,
                                concat(ROUND(SUM(雪豹签收) / SUM(雪豹完成) * 100,2),'%') as 雪豹完成签收,
                                concat(ROUND(SUM(雪豹完成) / SUM(雪豹单量) * 100,2),'%') as 雪豹完成占比,	
							SUM(金蝉项目单量) 金蝉项目单量, SUM(金蝉项目签收) 金蝉项目签收, SUM(金蝉项目拒收) 金蝉项目拒收,
                                concat(ROUND(SUM(金蝉项目改派) / SUM(金蝉项目单量) * 100,2),'%') as 金蝉项目改派占比,
                                concat(ROUND(SUM(金蝉项目签收) / SUM(金蝉项目单量) * 100,2),'%') as 金蝉项目总计签收,
                                concat(ROUND(SUM(金蝉项目签收) / SUM(金蝉项目完成) * 100,2),'%') as 金蝉项目完成签收,
                                concat(ROUND(SUM(金蝉项目完成) / SUM(金蝉项目单量) * 100,2),'%') as 金蝉项目完成占比,	
							SUM(金蝉优化单量) 金蝉优化单量, SUM(金蝉优化签收) 金蝉优化签收, SUM(金蝉优化拒收) 金蝉优化拒收,
                                concat(ROUND(SUM(金蝉优化改派) / SUM(金蝉优化单量) * 100,2),'%') as 金蝉优化改派占比,
                                concat(ROUND(SUM(金蝉优化签收) / SUM(金蝉优化单量) * 100,2),'%') as 金蝉优化总计签收,
                                concat(ROUND(SUM(金蝉优化签收) / SUM(金蝉优化完成) * 100,2),'%') as 金蝉优化完成签收,
                                concat(ROUND(SUM(金蝉优化完成) / SUM(金蝉优化单量) * 100,2),'%') as 金蝉优化完成占比,
							SUM(金蝉公共单量) 金蝉公共单量, SUM(金蝉公共签收) 金蝉公共签收, SUM(金蝉公共拒收) 金蝉公共拒收,
                                concat(ROUND(SUM(金蝉公共改派) / SUM(金蝉公共单量) * 100,2),'%') as 金蝉公共改派占比,
                                concat(ROUND(SUM(金蝉公共签收) / SUM(金蝉公共单量) * 100,2),'%') as 金蝉公共总计签收,
                                concat(ROUND(SUM(金蝉公共签收) / SUM(金蝉公共完成) * 100,2),'%') as 金蝉公共完成签收,
                                concat(ROUND(SUM(金蝉公共完成) / SUM(金蝉公共单量) * 100,2),'%') as 金蝉公共完成占比,
								
							SUM(客服中心单量) 客服中心单量, SUM(客服中心签收) 客服中心签收, SUM(客服中心拒收) 客服中心拒收,
                                concat(ROUND(SUM(客服中心改派) / SUM(客服中心单量) * 100,2),'%') as 客服中心改派占比,
                                concat(ROUND(SUM(客服中心签收) / SUM(客服中心单量) * 100,2),'%') as 客服中心总计签收,
                                concat(ROUND(SUM(客服中心签收) / SUM(客服中心完成) * 100,2),'%') as 客服中心完成签收,
                                concat(ROUND(SUM(客服中心完成) / SUM(客服中心单量) * 100,2),'%') as 客服中心完成占比,
							SUM(奥创队单量) 奥创队单量, SUM(奥创队签收) 奥创队签收, SUM(奥创队拒收) 奥创队拒收,
                                concat(ROUND(SUM(奥创队改派) / SUM(奥创队单量) * 100,2),'%') as 奥创队改派占比,
                                concat(ROUND(SUM(奥创队签收) / SUM(奥创队单量) * 100,2),'%') as 奥创队总计签收,
                                concat(ROUND(SUM(奥创队签收) / SUM(奥创队完成) * 100,2),'%') as 奥创队完成签收,
                                concat(ROUND(SUM(奥创队完成) / SUM(奥创队单量) * 100,2),'%') as 奥创队完成占比,
							SUM(神龙主页运营单量) 神龙主页运营单量, SUM(神龙主页运营签收) 神龙主页运营签收, SUM(神龙主页运营拒收) 神龙主页运营拒收,
                                concat(ROUND(SUM(神龙主页运营改派) / SUM(神龙主页运营单量) * 100,2),'%') as 神龙主页运营改派占比,
                                concat(ROUND(SUM(神龙主页运营签收) / SUM(神龙主页运营单量) * 100,2),'%') as 神龙主页运营总计签收,
                                concat(ROUND(SUM(神龙主页运营签收) / SUM(神龙主页运营完成) * 100,2),'%') as 神龙主页运营完成签收,
                                concat(ROUND(SUM(神龙主页运营完成) / SUM(神龙主页运营单量) * 100,2),'%') as 神龙主页运营完成占比,
							SUM(APP单量) APP单量, SUM(APP签收) APP签收, SUM(APP拒收) APP拒收,
                                concat(ROUND(SUM(APP改派) / SUM(APP单量) * 100,2),'%') as APP改派占比,
                                concat(ROUND(SUM(APP签收) / SUM(APP单量) * 100,2),'%') as APP总计签收,
                                concat(ROUND(SUM(APP签收) / SUM(APP完成) * 100,2),'%') as APP完成签收,
                                concat(ROUND(SUM(APP完成) / SUM(APP单量) * 100,2),'%') as APP完成占比,
							SUM(Line单量) Line单量, SUM(Line签收) Line签收, SUM(Line拒收) Line拒收,
                                concat(ROUND(SUM(Line改派) / SUM(Line单量) * 100,2),'%') as Line改派占比,
                                concat(ROUND(SUM(Line签收) / SUM(Line单量) * 100,2),'%') as Line总计签收,
                                concat(ROUND(SUM(Line签收) / SUM(Line完成) * 100,2),'%') as Line完成签收,
                                concat(ROUND(SUM(Line完成) / SUM(Line单量) * 100,2),'%') as Line完成占比,
							SUM(红杉单量) 红杉单量, SUM(红杉签收) 红杉签收, SUM(红杉拒收) 红杉拒收,
                                concat(ROUND(SUM(红杉改派) / SUM(红杉单量) * 100,2),'%') as 红杉改派占比,
                                concat(ROUND(SUM(红杉签收) / SUM(红杉单量) * 100,2),'%') as 红杉总计签收,
                                concat(ROUND(SUM(红杉签收) / SUM(红杉完成) * 100,2),'%') as 红杉完成签收,
                                concat(ROUND(SUM(红杉完成) / SUM(红杉单量) * 100,2),'%') as 红杉完成占比,
							SUM(郑州北美单量) 郑州北美单量, SUM(郑州北美签收) 郑州北美签收, SUM(郑州北美拒收) 郑州北美拒收,
                                concat(ROUND(SUM(郑州北美改派) / SUM(郑州北美单量) * 100,2),'%') as 郑州北美改派占比,
                                concat(ROUND(SUM(郑州北美签收) / SUM(郑州北美单量) * 100,2),'%') as 郑州北美总计签收,
                                concat(ROUND(SUM(郑州北美签收) / SUM(郑州北美完成) * 100,2),'%') as 郑州北美完成签收,
                                concat(ROUND(SUM(郑州北美完成) / SUM(郑州北美单量) * 100,2),'%') as 郑州北美完成占比,
							SUM(研发部单量) 研发部单量, SUM(研发部签收) 研发部签收, SUM(研发部拒收) 研发部拒收,
                                concat(ROUND(SUM(研发部改派) / SUM(研发部单量) * 100,2),'%') as 研发部改派占比,
                                concat(ROUND(SUM(研发部签收) / SUM(研发部单量) * 100,2),'%') as 研发部总计签收,
                                concat(ROUND(SUM(研发部签收) / SUM(研发部完成) * 100,2),'%') as 研发部完成签收,
                                concat(ROUND(SUM(研发部完成) / SUM(研发部单量) * 100,2),'%') as 研发部完成占比,
							SUM(金鹏单量) 金鹏单量, SUM(金鹏签收) 金鹏签收, SUM(金鹏拒收) 金鹏拒收,
                                concat(ROUND(SUM(金鹏改派) / SUM(金鹏单量) * 100,2),'%') as 金鹏改派占比,
                                concat(ROUND(SUM(金鹏签收) / SUM(金鹏单量) * 100,2),'%') as 金鹏总计签收,
                                concat(ROUND(SUM(金鹏签收) / SUM(金鹏完成) * 100,2),'%') as 金鹏完成签收,
                                concat(ROUND(SUM(金鹏完成) / SUM(金鹏单量) * 100,2),'%') as 金鹏完成占比,	
							SUM(金狮单量) 金狮单量, SUM(金狮签收) 金狮签收, SUM(金狮拒收) 金狮拒收,
                                concat(ROUND(SUM(金狮改派) / SUM(金狮单量) * 100,2),'%') as 金狮改派占比,
                                concat(ROUND(SUM(金狮签收) / SUM(金狮单量) * 100,2),'%') as 金狮总计签收,
                                concat(ROUND(SUM(金狮签收) / SUM(金狮完成) * 100,2),'%') as 金狮完成签收,
                                concat(ROUND(SUM(金狮完成) / SUM(金狮单量) * 100,2),'%') as 金狮完成占比,	
							SUM(翼虎单量) 翼虎单量, SUM(翼虎签收) 翼虎签收, SUM(翼虎拒收) 翼虎拒收,
                                concat(ROUND(SUM(翼虎改派) / SUM(翼虎单量) * 100,2),'%') as 翼虎改派占比,
                                concat(ROUND(SUM(翼虎签收) / SUM(翼虎单量) * 100,2),'%') as 翼虎总计签收,
                                concat(ROUND(SUM(翼虎签收) / SUM(翼虎完成) * 100,2),'%') as 翼虎完成签收,
                                concat(ROUND(SUM(翼虎完成) / SUM(翼虎单量) * 100,2),'%') as 翼虎完成占比	
                        FROM(SELECT 年月 月份,币种 地区, 产品id, 产品名称, 父级分类,
                                COUNT(订单编号) as 总单量,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 完成,
                                SUM(IF(是否改派 = '改派',1,0)) as 改派,
                            SUM(IF(所属团队 = '神龙港台',1,0)) as 神龙单量,
                                SUM(IF(所属团队 = '神龙港台' AND 最终状态 = "已签收",1,0)) as 神龙签收,
                                SUM(IF(所属团队 = '神龙港台' AND 最终状态 = "拒收",1,0)) as 神龙拒收,
                                SUM(IF(所属团队 = '神龙港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 神龙完成,
                                SUM(IF(所属团队 = '神龙港台' AND 是否改派 = '改派',1,0)) as 神龙改派,
                            SUM(IF(所属团队 = '火凤凰港台',1,0)) as 火凤凰单量,
                                SUM(IF(所属团队 = '火凤凰港台' AND 最终状态 = "已签收",1,0)) as 火凤凰签收,
                                SUM(IF(所属团队 = '火凤凰港台' AND 最终状态 = "拒收",1,0)) as 火凤凰拒收,
                                SUM(IF(所属团队 = '火凤凰港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 火凤凰完成,
                                SUM(IF(所属团队 = '火凤凰港台' AND 是否改派 = '改派',1,0)) as 火凤凰改派,
                            SUM(IF(所属团队 = '雪豹港台',1,0)) as 雪豹单量,
                                SUM(IF(所属团队 = '雪豹港台' AND 最终状态 = "已签收",1,0)) as 雪豹签收,
                                SUM(IF(所属团队 = '雪豹港台' AND 最终状态 = "拒收",1,0)) as 雪豹拒收,
                                SUM(IF(所属团队 = '雪豹港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 雪豹完成,
                                SUM(IF(所属团队 = '雪豹港台' AND 是否改派 = '改派',1,0)) as 雪豹改派,
                            SUM(IF(所属团队 = '金蝉项目组',1,0)) as 金蝉项目单量,
                                SUM(IF(所属团队 = '金蝉项目组' AND 最终状态 = "已签收",1,0)) as 金蝉项目签收,
                                SUM(IF(所属团队 = '金蝉项目组' AND 最终状态 = "拒收",1,0)) as 金蝉项目拒收,
                                SUM(IF(所属团队 = '金蝉项目组' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 金蝉项目完成,
                                SUM(IF(所属团队 = '金蝉项目组' AND 是否改派 = '改派',1,0)) as 金蝉项目改派,
                            SUM(IF(所属团队 = '金蝉家族优化组',1,0)) as 金蝉优化单量,
                                SUM(IF(所属团队 = '金蝉家族优化组' AND 最终状态 = "已签收",1,0)) as 金蝉优化签收,
                                SUM(IF(所属团队 = '金蝉家族优化组' AND 最终状态 = "拒收",1,0)) as 金蝉优化拒收,
                                SUM(IF(所属团队 = '金蝉家族优化组' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 金蝉优化完成,
                                SUM(IF(所属团队 = '金蝉家族优化组' AND 是否改派 = '改派',1,0)) as 金蝉优化改派,
                            SUM(IF(所属团队 = '金蝉家族公共团队',1,0)) as 金蝉公共单量,
                                SUM(IF(所属团队 = '金蝉家族公共团队' AND 最终状态 = "已签收",1,0)) as 金蝉公共签收,
                                SUM(IF(所属团队 = '金蝉家族公共团队' AND 最终状态 = "拒收",1,0)) as 金蝉公共拒收,
                                SUM(IF(所属团队 = '金蝉家族公共团队' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 金蝉公共完成,
                                SUM(IF(所属团队 = '金蝉家族公共团队' AND 是否改派 = '改派',1,0)) as 金蝉公共改派,
                            SUM(IF(所属团队 = '客服中心港台',1,0)) as 客服中心单量,
                                SUM(IF(所属团队 = '客服中心港台' AND 最终状态 = "已签收",1,0)) as 客服中心签收,
                                SUM(IF(所属团队 = '客服中心港台' AND 最终状态 = "拒收",1,0)) as 客服中心拒收,
                                SUM(IF(所属团队 = '客服中心港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 客服中心完成,
                                SUM(IF(所属团队 = '客服中心港台' AND 是否改派 = '改派',1,0)) as 客服中心改派,
                            SUM(IF(所属团队 = '奥创队',1,0)) as 奥创队单量,
                                SUM(IF(所属团队 = '奥创队' AND 最终状态 = "已签收",1,0)) as 奥创队签收,
                                SUM(IF(所属团队 = '奥创队' AND 最终状态 = "拒收",1,0)) as 奥创队拒收,
                                SUM(IF(所属团队 = '奥创队' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 奥创队完成,
                                SUM(IF(所属团队 = '奥创队' AND 是否改派 = '改派',1,0)) as 奥创队改派,
                            SUM(IF(所属团队 = '神龙主页运营',1,0)) as 神龙主页运营单量,
                                SUM(IF(所属团队 = '神龙主页运营' AND 最终状态 = "已签收",1,0)) as 神龙主页运营签收,
                                SUM(IF(所属团队 = '神龙主页运营' AND 最终状态 = "拒收",1,0)) as 神龙主页运营拒收,
                                SUM(IF(所属团队 = '神龙主页运营' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 神龙主页运营完成,
                                SUM(IF(所属团队 = '神龙主页运营' AND 是否改派 = '改派',1,0)) as 神龙主页运营改派,
                            SUM(IF(所属团队 = 'APP运营',1,0)) as APP单量,
                                SUM(IF(所属团队 = 'APP运营' AND 最终状态 = "已签收",1,0)) as APP签收,
                                SUM(IF(所属团队 = 'APP运营' AND 最终状态 = "拒收",1,0)) as APP拒收,
                                SUM(IF(所属团队 = 'APP运营' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as APP完成,
                                SUM(IF(所属团队 = 'APP运营' AND 是否改派 = '改派',1,0)) as APP改派,
                            SUM(IF(所属团队 = 'Line运营',1,0)) as Line单量,
                                SUM(IF(所属团队 = 'Line运营' AND 最终状态 = "已签收",1,0)) as Line签收,
                                SUM(IF(所属团队 = 'Line运营' AND 最终状态 = "拒收",1,0)) as Line拒收,
                                SUM(IF(所属团队 = 'Line运营' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as Line完成,
                                SUM(IF(所属团队 = 'Line运营' AND 是否改派 = '改派',1,0)) as Line改派,
                            SUM(IF(所属团队 = '红杉港台',1,0)) as 红杉单量,
                                SUM(IF(所属团队 = '红杉港台' AND 最终状态 = "已签收",1,0)) as 红杉签收,
                                SUM(IF(所属团队 = '红杉港台' AND 最终状态 = "拒收",1,0)) as 红杉拒收,
                                SUM(IF(所属团队 = '红杉港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 红杉完成,
                                SUM(IF(所属团队 = '红杉港台' AND 是否改派 = '改派',1,0)) as 红杉改派,
                            SUM(IF(所属团队 = '郑州北美',1,0)) as 郑州北美单量,
                                SUM(IF(所属团队 = '郑州北美' AND 最终状态 = "已签收",1,0)) as 郑州北美签收,
                                SUM(IF(所属团队 = '郑州北美' AND 最终状态 = "拒收",1,0)) as 郑州北美拒收,
                                SUM(IF(所属团队 = '郑州北美' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 郑州北美完成,
                                SUM(IF(所属团队 = '郑州北美' AND 是否改派 = '改派',1,0)) as 郑州北美改派,
                            SUM(IF(所属团队 = '研发部港台',1,0)) as 研发部单量,
                                SUM(IF(所属团队 = '研发部港台' AND 最终状态 = "已签收",1,0)) as 研发部签收,
                                SUM(IF(所属团队 = '研发部港台' AND 最终状态 = "拒收",1,0)) as 研发部拒收,
                                SUM(IF(所属团队 = '研发部港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 研发部完成,
                                SUM(IF(所属团队 = '研发部港台' AND 是否改派 = '改派',1,0)) as 研发部改派,
                            SUM(IF(所属团队 = '金鹏港台',1,0)) as 金鹏单量,
                                SUM(IF(所属团队 = '金鹏港台' AND 最终状态 = "已签收",1,0)) as 金鹏签收,
                                SUM(IF(所属团队 = '金鹏港台' AND 最终状态 = "拒收",1,0)) as 金鹏拒收,
                                SUM(IF(所属团队 = '金鹏港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 金鹏完成,
                                SUM(IF(所属团队 = '金鹏港台' AND 是否改派 = '改派',1,0)) as 金鹏改派,
                            SUM(IF(所属团队 = '金狮港台',1,0)) as 金狮单量,
                                SUM(IF(所属团队 = '金狮港台' AND 最终状态 = "已签收",1,0)) as 金狮签收,
                                SUM(IF(所属团队 = '金狮港台' AND 最终状态 = "拒收",1,0)) as 金狮拒收,
                                SUM(IF(所属团队 = '金狮港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 金狮完成,
                                SUM(IF(所属团队 = '金狮港台' AND 是否改派 = '改派',1,0)) as 金狮改派,
                            SUM(IF(所属团队 = '翼虎港台',1,0)) as 翼虎单量,
                                SUM(IF(所属团队 = '翼虎港台' AND 最终状态 = "已签收",1,0)) as 翼虎签收,
                                SUM(IF(所属团队 = '翼虎港台' AND 最终状态 = "拒收",1,0)) as 翼虎拒收,
                                SUM(IF(所属团队 = '翼虎港台' AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 翼虎完成,
                                SUM(IF(所属团队 = '翼虎港台' AND 是否改派 = '改派',1,0)) as 翼虎改派
                            FROM {1}_zqsb cc
						    WHERE cc.日期 >= '{0}' and cc.`运单编号` is not null AND cc.所属团队 NOT IN ({2})
                            GROUP BY cc.年月,cc.币种,cc.产品id
					    ) s
						GROUP BY 月份,地区,产品id		
                        WITH ROLLUP 
					) ss
                   ORDER BY FIELD(月份,DATE_FORMAT(CURDATE(),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(CURDATE(),INTERVAL 2 MONTH),'%Y%m'),'总计'),
                            FIELD(地区,'台湾','香港','总计'),
                            总单量 DESC;'''.format(month_last, team, not_team)
        df50 = pd.read_sql_query(sql=sql50, con=self.engine1)
        listT.append(df50)

        # 10、同产品各月的对比
        print('正在获取---10、同产品各月的对比…………')
        t1 = (datetime.datetime.now() - relativedelta(months=12)).strftime('%Y%m')
        t2 = (datetime.datetime.now() - relativedelta(months=11)).strftime('%Y%m')
        t3 = (datetime.datetime.now() - relativedelta(months=10)).strftime('%Y%m')
        t4 = (datetime.datetime.now() - relativedelta(months=9)).strftime('%Y%m')
        t5 = (datetime.datetime.now() - relativedelta(months=8)).strftime('%Y%m')
        t6 = (datetime.datetime.now() - relativedelta(months=7)).strftime('%Y%m')
        t7 = (datetime.datetime.now() - relativedelta(months=6)).strftime('%Y%m')
        t8 = (datetime.datetime.now() - relativedelta(months=5)).strftime('%Y%m')
        t9 = (datetime.datetime.now() - relativedelta(months=4)).strftime('%Y%m')
        t10 = (datetime.datetime.now() - relativedelta(months=3)).strftime('%Y%m')
        t11 = (datetime.datetime.now() - relativedelta(months=2)).strftime('%Y%m')
        t12 = (datetime.datetime.now() - relativedelta(months=1)).strftime('%Y%m')
        t13 = datetime.datetime.now().strftime('%Y%m')
        sql51 = '''SELECT *
                FROM (SELECT IFNULL(家族, '总计') 家族, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id,  IFNULL(产品名称, '总计') 产品名称, IFNULL(父级分类, '总计') 父级分类, SUM(总单量) 总单量,
						SUM(04总量) 11总单量,
							concat(ROUND(SUM(04签收量) / SUM(04总量) * 100,2),'%') as 11总计签收,
							concat(ROUND(SUM(04签收量) / SUM(04完成量) * 100,2),'%') as 11完成签收,
							concat(ROUND(SUM(04完成量) / SUM(04总量) * 100,2),'%') as 11完成占比,
						SUM(05总量) 22总单量,
							concat(ROUND(SUM(05签收量) / SUM(05总量) * 100,2),'%') as 22总计签收,
							concat(ROUND(SUM(05签收量) / SUM(05完成量) * 100,2),'%') as 22完成签收,
							concat(ROUND(SUM(05完成量) / SUM(05总量) * 100,2),'%') as 22完成占比,
						SUM(06总量) 33总单量,
							concat(ROUND(SUM(06签收量) / SUM(06总量) * 100,2),'%') as 33总计签收,
							concat(ROUND(SUM(06签收量) / SUM(06完成量) * 100,2),'%') as 33完成签收,
							concat(ROUND(SUM(06完成量) / SUM(06总量) * 100,2),'%') as 33完成占比,
						SUM(07总量) 44总单量,
							concat(ROUND(SUM(07签收量) / SUM(07总量) * 100,2),'%') as 44总计签收,
							concat(ROUND(SUM(07签收量) / SUM(07完成量) * 100,2),'%') as 44完成签收,
							concat(ROUND(SUM(07完成量) / SUM(07总量) * 100,2),'%') as 44完成占比,
						SUM(08总量) 55总单量,
							concat(ROUND(SUM(08签收量) / SUM(08总量) * 100,2),'%') as 55总计签收,
							concat(ROUND(SUM(08签收量) / SUM(08完成量) * 100,2),'%') as 55完成签收,
							concat(ROUND(SUM(08完成量) / SUM(08总量) * 100,2),'%') as 55完成占比,
						SUM(09总量) 66总单量,
							concat(ROUND(SUM(09签收量) / SUM(09总量) * 100,2),'%') as 66总计签收,
							concat(ROUND(SUM(09签收量) / SUM(09完成量) * 100,2),'%') as 66完成签收,
							concat(ROUND(SUM(09完成量) / SUM(09总量) * 100,2),'%') as 66完成占比,
						SUM(10总量) 77总单量,
							concat(ROUND(SUM(10签收量) / SUM(10总量) * 100,2),'%') as 77总计签收,
							concat(ROUND(SUM(10签收量) / SUM(10完成量) * 100,2),'%') as 77完成签收,
							concat(ROUND(SUM(10完成量) / SUM(10总量) * 100,2),'%') as 77完成占比,
						SUM(11总量) 88总单量,
							concat(ROUND(SUM(11签收量) / SUM(11总量) * 100,2),'%') as 88总计签收,
							concat(ROUND(SUM(11签收量) / SUM(11完成量) * 100,2),'%') as 88完成签收,
							concat(ROUND(SUM(11完成量) / SUM(11总量) * 100,2),'%') as 88完成占比,
						SUM(12总量) 99总单量,
							concat(ROUND(SUM(12签收量) / SUM(12总量) * 100,2),'%') as 99总计签收,
							concat(ROUND(SUM(12签收量) / SUM(12完成量) * 100,2),'%') as 99完成签收,
							concat(ROUND(SUM(12完成量) / SUM(12总量) * 100,2),'%') as 99完成占比,
						SUM(13总量) 100总单量,
							concat(ROUND(SUM(12签收量) / SUM(12总量) * 100,2),'%') as 100总计签收,
							concat(ROUND(SUM(12签收量) / SUM(12完成量) * 100,2),'%') as 100完成签收,
							concat(ROUND(SUM(12完成量) / SUM(12总量) * 100,2),'%') as 100完成占比,
						SUM(14总量) 110总单量,
							concat(ROUND(SUM(14签收量) / SUM(14总量) * 100,2),'%') as 110总计签收,
							concat(ROUND(SUM(14签收量) / SUM(14完成量) * 100,2),'%') as 110完成签收,
							concat(ROUND(SUM(14完成量) / SUM(14总量) * 100,2),'%') as 110完成占比,
						SUM(15总量) 120总单量,
							concat(ROUND(SUM(15签收量) / SUM(15总量) * 100,2),'%') as 120总计签收,
							concat(ROUND(SUM(15签收量) / SUM(15完成量) * 100,2),'%') as 120完成签收,
							concat(ROUND(SUM(15完成量) / SUM(15总量) * 100,2),'%') as 120完成占比,
						SUM(16总量) 130总单量,
							concat(ROUND(SUM(16签收量) / SUM(16总量) * 100,2),'%') as 130总计签收,
							concat(ROUND(SUM(16签收量) / SUM(16完成量) * 100,2),'%') as 130完成签收,
							concat(ROUND(SUM(16完成量) / SUM(16总量) * 100,2),'%') as 130完成占比
                    FROM(SELECT 所属团队 as 家族,币种 地区, 产品id, 产品名称, 父级分类,
                                COUNT(cx.`订单编号`) as 总单量,
                            SUM(IF(年月 = {1},1,0)) as “04总量“,
                                SUM(IF(年月 = {1} AND 最终状态 = "已签收",1,0)) as “04签收量“,
                                SUM(IF(年月 = {1} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “04完成量“,
                            SUM(IF(年月 = {2},1,0)) as “05总量“,
                                SUM(IF(年月 = {2} AND 最终状态 = "已签收",1,0)) as “05签收量“,
                                SUM(IF(年月 = {2} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “05完成量“,
                            SUM(IF(年月 = {3},1,0)) as “06总量“,
                                SUM(IF(年月 = {3} AND 最终状态 = "已签收",1,0)) as “06签收量“,
                                SUM(IF(年月 = {3} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “06完成量“,
                            SUM(IF(年月 = {4},1,0)) as “07总量“,
                                SUM(IF(年月 = {4} AND 最终状态 = "已签收",1,0)) as “07签收量“,
                                SUM(IF(年月 = {4} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “07完成量“,
                            SUM(IF(年月 = {5},1,0)) as “08总量“,
                                SUM(IF(年月 = {5} AND 最终状态 = "已签收",1,0)) as “08签收量“,
                                SUM(IF(年月 = {5} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “08完成量“,
                            SUM(IF(年月 = {6},1,0)) as “09总量“,
                                SUM(IF(年月 = {6} AND 最终状态 = "已签收",1,0)) as “09签收量“,
                                SUM(IF(年月 = {6} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “09完成量“,
                            SUM(IF(年月 = {7},1,0)) as “10总量“,
                                SUM(IF(年月 = {7} AND 最终状态 = "已签收",1,0)) as “10签收量“,
                                SUM(IF(年月 = {7} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “10完成量“,
                            SUM(IF(年月 = {8},1,0)) as “11总量“,
                                SUM(IF(年月 = {8} AND 最终状态 = "已签收",1,0)) as “11签收量“,
                                SUM(IF(年月 = {8} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “11完成量“,
                            SUM(IF(年月 = {9},1,0)) as “12总量“,
                                SUM(IF(年月 = {9} AND 最终状态 = "已签收",1,0)) as “12签收量“,
                                SUM(IF(年月 = {9} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “12完成量“,
                            SUM(IF(年月 = {10},1,0)) as “13总量“,
                                SUM(IF(年月 = {10} AND 最终状态 = "已签收",1,0)) as “13签收量“,
                                SUM(IF(年月 = {10} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “13完成量“,
                            SUM(IF(年月 = {11},1,0)) as “14总量“,
                                SUM(IF(年月 = {11} AND 最终状态 = "已签收",1,0)) as “14签收量“,
                                SUM(IF(年月 = {11} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “14完成量“,
                            SUM(IF(年月 = {12},1,0)) as "“15总量“",
                                SUM(IF(年月 = {12} AND 最终状态 = "已签收",1,0)) as “15签收量“,
                                SUM(IF(年月 = {12} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “15完成量“,
                            SUM(IF(年月 = {13},1,0)) as “16总量“,
                                SUM(IF(年月 = {13} AND 最终状态 = "已签收",1,0)) as “16签收量“,
                                SUM(IF(年月 = {13} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as “16完成量“
                        FROM gat_zqsb cx
                        where cx.年月 >= '{1}' AND cx.`运单编号` is not null AND cx.所属团队 NOT IN ({0})
                        GROUP BY cx.所属团队,cx.币种,cx.产品id
                    ) s
					GROUP BY s.家族,s.地区,s.产品id
					WITH ROLLUP
				) ss
                ORDER BY FIELD(ss.`家族`,{14}, '总计'),
                        FIELD(ss.地区, '台湾', '香港', '总计' ),
                        ss.总单量 DESC;'''.format(not_team, t1,t2,t3,t4,t5,t6,t7,t8,t9,t10,t11,t12,t13, self.team_name2)
        sql51 = '''SELECT *
                FROM (SELECT IFNULL(家族, '总计') 家族, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id,  IFNULL(产品名称, '总计') 产品名称, IFNULL(父级分类, '总计') 父级分类, SUM(总单量) 总单量,
						SUM("04总量") "11总单量",
                            concat(ROUND(SUM("04签收量") / SUM("04总量") * 100,2),'%') as "11总计签收",
							concat(ROUND(SUM("04签收量") / SUM("04完成量") * 100,2),'%') as "11完成签收",
							concat(ROUND(SUM("04完成量") / SUM("04总量") * 100,2),'%') as "11完成占比",
						SUM("05总量") "22总单量",
							concat(ROUND(SUM("05签收量") / SUM("05总量") * 100,2),'%') as "22总计签收",
							concat(ROUND(SUM("05签收量") / SUM("05完成量") * 100,2),'%') as "22完成签收",
							concat(ROUND(SUM("05完成量") / SUM("05总量") * 100,2),'%') as "22完成占比",
						SUM("06总量") "33总单量",
							concat(ROUND(SUM("06签收量") / SUM("06总量") * 100,2),'%') as "33总计签收",
							concat(ROUND(SUM("06签收量") / SUM("06完成量") * 100,2),'%') as "33完成签收",
							concat(ROUND(SUM("06完成量") / SUM("06总量") * 100,2),'%') as "33完成占比",
						SUM("07总量") "44总单量",
							concat(ROUND(SUM("07签收量") / SUM("07总量") * 100,2),'%') as "44总计签收",
							concat(ROUND(SUM("07签收量") / SUM("07完成量") * 100,2),'%') as "44完成签收",
							concat(ROUND(SUM("07完成量") / SUM("07总量") * 100,2),'%') as "44完成占比",
						SUM("08总量") "55总单量",
							concat(ROUND(SUM("08签收量") / SUM("08总量") * 100,2),'%') as "55总计签收",
							concat(ROUND(SUM("08签收量") / SUM("08完成量") * 100,2),'%') as "55完成签收",
							concat(ROUND(SUM("08完成量") / SUM("08总量") * 100,2),'%') as "55完成占比",
						SUM("09总量") "66总单量",
							concat(ROUND(SUM("09签收量") / SUM("09总量") * 100,2),'%') as "66总计签收",
							concat(ROUND(SUM("09签收量") / SUM("09完成量") * 100,2),'%') as "66完成签收",
							concat(ROUND(SUM("09完成量") / SUM("09总量") * 100,2),'%') as "66完成占比",
						SUM("10总量") "77总单量",
							concat(ROUND(SUM("10签收量") / SUM("10总量") * 100,2),'%') as "77总计签收",
							concat(ROUND(SUM("10签收量") / SUM("10完成量") * 100,2),'%') as "77完成签收",
							concat(ROUND(SUM("10完成量") / SUM("10总量") * 100,2),'%') as "77完成占比",
						SUM("11总量") "88总单量",
							concat(ROUND(SUM("11签收量") / SUM("11总量") * 100,2),'%') as "88总计签收",
							concat(ROUND(SUM("11签收量") / SUM("11完成量") * 100,2),'%') as "88完成签收",
							concat(ROUND(SUM("11完成量") / SUM("11总量") * 100,2),'%') as "88完成占比",
						SUM("12总量") "99总单量",
							concat(ROUND(SUM("12签收量") / SUM("12总量") * 100,2),'%') as "99总计签收",
							concat(ROUND(SUM("12签收量") / SUM("12完成量") * 100,2),'%') as "99完成签收",
							concat(ROUND(SUM("12完成量") / SUM("12总量") * 100,2),'%') as "99完成占比",
						SUM("13总量") "100总单量",
							concat(ROUND(SUM("12签收量") / SUM("12总量") * 100,2),'%') as "100总计签收",
							concat(ROUND(SUM("12签收量") / SUM("12完成量") * 100,2),'%') as "100完成签收",
							concat(ROUND(SUM("12完成量") / SUM("12总量") * 100,2),'%') as "100完成占比",
						SUM("14总量") "110总单量",
							concat(ROUND(SUM("14签收量") / SUM("14总量") * 100,2),'%') as "110总计签收",
							concat(ROUND(SUM("14签收量") / SUM("14完成量") * 100,2),'%') as "110完成签收",
							concat(ROUND(SUM("14完成量") / SUM("14总量") * 100,2),'%') as "110完成占比",
						SUM("15总量") "120总单量",
							concat(ROUND(SUM("15签收量") / SUM("15总量") * 100,2),'%') as "120总计签收",
							concat(ROUND(SUM("15签收量") / SUM("15完成量") * 100,2),'%') as "120完成签收",
							concat(ROUND(SUM("15完成量") / SUM("15总量") * 100,2),'%') as "120完成占比",
						SUM("16总量") "130总单量",
							concat(ROUND(SUM("16签收量") / SUM("16总量") * 100,2),'%') as "130总计签收",
							concat(ROUND(SUM("16签收量") / SUM("16完成量") * 100,2),'%') as "130完成签收",
							concat(ROUND(SUM("16完成量") / SUM("16总量") * 100,2),'%') as "130完成占比"
                    FROM(SELECT 所属团队 as 家族,币种 地区, 产品id, 产品名称, 父级分类,
                                COUNT(cx.`订单编号`) as 总单量,
                            SUM(IF(年月 = {1},1,0)) as "04总量",
                                SUM(IF(年月 = {1} AND 最终状态 = "已签收",1,0)) as "04签收量",
                                SUM(IF(年月 = {1} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "04完成量",
                            SUM(IF(年月 = {2},1,0)) as "05总量",
                                SUM(IF(年月 = {2} AND 最终状态 = "已签收",1,0)) as "05签收量",
                                SUM(IF(年月 = {2} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "05完成量",
                            SUM(IF(年月 = {3},1,0)) as "06总量",
                                SUM(IF(年月 = {3} AND 最终状态 = "已签收",1,0)) as "06签收量",
                                SUM(IF(年月 = {3} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "06完成量",
                            SUM(IF(年月 = {4},1,0)) as "07总量",
                                SUM(IF(年月 = {4} AND 最终状态 = "已签收",1,0)) as "07签收量",
                                SUM(IF(年月 = {4} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "07完成量",
                            SUM(IF(年月 = {5},1,0)) as "08总量",
                                SUM(IF(年月 = {5} AND 最终状态 = "已签收",1,0)) as "08签收量",
                                SUM(IF(年月 = {5} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "08完成量",
                            SUM(IF(年月 = {6},1,0)) as "09总量",
                                SUM(IF(年月 = {6} AND 最终状态 = "已签收",1,0)) as "09签收量",
                                SUM(IF(年月 = {6} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "09完成量",
                            SUM(IF(年月 = {7},1,0)) as "10总量",
                                SUM(IF(年月 = {7} AND 最终状态 = "已签收",1,0)) as "10签收量",
                                SUM(IF(年月 = {7} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "10完成量",
                            SUM(IF(年月 = {8},1,0)) as "11总量",
                                SUM(IF(年月 = {8} AND 最终状态 = "已签收",1,0)) as "11签收量",
                                SUM(IF(年月 = {8} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "11完成量",
                            SUM(IF(年月 = {9},1,0)) as "12总量",
                                SUM(IF(年月 = {9} AND 最终状态 = "已签收",1,0)) as "12签收量",
                                SUM(IF(年月 = {9} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "12完成量",
                            SUM(IF(年月 = {10},1,0)) as "13总量",
                                SUM(IF(年月 = {10} AND 最终状态 = "已签收",1,0)) as "13签收量",
                                SUM(IF(年月 = {10} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "13完成量",
                            SUM(IF(年月 = {11},1,0)) as "14总量",
                                SUM(IF(年月 = {11} AND 最终状态 = "已签收",1,0)) as "14签收量",
                                SUM(IF(年月 = {11} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "14完成量",
                            SUM(IF(年月 = {12},1,0)) as "15总量",
                                SUM(IF(年月 = {12} AND 最终状态 = "已签收",1,0)) as "15签收量",
                                SUM(IF(年月 = {12} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "15完成量",
                            SUM(IF(年月 = {13},1,0)) as "16总量",
                                SUM(IF(年月 = {13} AND 最终状态 = "已签收",1,0)) as "16签收量",
                                SUM(IF(年月 = {13} AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as "16完成量"
                        FROM gat_zqsb cx
                        where cx.年月 >= '{1}' AND cx.`运单编号` is not null AND cx.所属团队 NOT IN ({0})
                        GROUP BY cx.所属团队,cx.币种,cx.产品id
                    ) s
					GROUP BY s.家族,s.地区,s.产品id
					WITH ROLLUP
				) ss
                ORDER BY FIELD(ss.`家族`,{14}, '总计'),
                        FIELD(ss.地区, '台湾', '香港', '总计' ),
                        ss.总单量 DESC;'''.format(not_team, t1,t2,t3,t4,t5,t6,t7,t8,t9,t10,t11,t12,t13, self.team_name2)
        sql51 = '''SELECT *
                FROM (SELECT IFNULL(家族, '总计') 家族, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id,  IFNULL(产品名称, '总计') 产品名称, IFNULL(父级分类, '总计') 父级分类, SUM(总单量) 总单量,
                    -- 	SUM(04总量) AS 040总单量, concat(ROUND(SUM(04签收量) / SUM(04总量) * 100,2),'%') as 040总计签收,
                    -- 		concat(ROUND(SUM(04签收量) / SUM(04完成量) * 100,2),'%') as 040完成签收,concat(ROUND(SUM(04完成量) / SUM(04总量) * 100,2),'%') as 040完成占比,
                    -- 	SUM(05总量) AS 050总单量,concat(ROUND(SUM(05签收量) / SUM(05总量) * 100,2),'%') as 050总计签收,
                    -- 		concat(ROUND(SUM(05签收量) / SUM(05完成量) * 100,2),'%') as 050完成签收,concat(ROUND(SUM(05完成量) / SUM(05总量) * 100,2),'%') as 050完成占比,
                    -- 	SUM(06总量) AS 060总单量,concat(ROUND(SUM(06签收量) / SUM(06总量) * 100,2),'%') as 060总计签收,
                    -- 		concat(ROUND(SUM(06签收量) / SUM(06完成量) * 100,2),'%') as 060完成签收,concat(ROUND(SUM(06完成量) / SUM(06总量) * 100,2),'%') as 060完成占比,
                    -- 	SUM(07总量) AS 070总单量,concat(ROUND(SUM(07签收量) / SUM(07总量) * 100,2),'%') as 070总计签收,
                    -- 		concat(ROUND(SUM(07签收量) / SUM(07完成量) * 100,2),'%') as 070完成签收,concat(ROUND(SUM(07完成量) / SUM(07总量) * 100,2),'%') as 070完成占比,
                    -- 	SUM(08总量) AS 080总单量,concat(ROUND(SUM(08签收量) / SUM(08总量) * 100,2),'%') as 080总计签收,
                    -- 		concat(ROUND(SUM(08签收量) / SUM(08完成量) * 100,2),'%') as 080完成签收,concat(ROUND(SUM(08完成量) / SUM(08总量) * 100,2),'%') as 080完成占比,
                    -- 	SUM(09总量) AS 090总单量,concat(ROUND(SUM(09签收量) / SUM(09总量) * 100,2),'%') as 090总计签收,
                    -- 		concat(ROUND(SUM(09签收量) / SUM(09完成量) * 100,2),'%') as 090完成签收,concat(ROUND(SUM(09完成量) / SUM(09总量) * 100,2),'%') as 090完成占比,
						SUM(10总量) AS 100总单量,
							concat(ROUND(SUM(10签收量) / SUM(10总量) * 100,2),'%') as 100总计签收,
							concat(ROUND(SUM(10签收量) / SUM(10完成量) * 100,2),'%') as 100完成签收,
							concat(ROUND(SUM(10完成量) / SUM(10总量) * 100,2),'%') as 100完成占比,
						SUM(11总量) AS 110总单量,
							concat(ROUND(SUM(11签收量) / SUM(11总量) * 100,2),'%') as 110总计签收,
							concat(ROUND(SUM(11签收量) / SUM(11完成量) * 100,2),'%') as 110完成签收,
							concat(ROUND(SUM(11完成量) / SUM(11总量) * 100,2),'%') as 110完成占比,
						SUM(12总量) AS 120总单量,
							concat(ROUND(SUM(12签收量) / SUM(12总量) * 100,2),'%') as 120总计签收,
							concat(ROUND(SUM(12签收量) / SUM(12完成量) * 100,2),'%') as 120完成签收,
							concat(ROUND(SUM(12完成量) / SUM(12总量) * 100,2),'%') as 120完成占比,
						SUM(13总量) AS 130总单量,
							concat(ROUND(SUM(13签收量) / SUM(13总量) * 100,2),'%') as 130总计签收,
							concat(ROUND(SUM(13签收量) / SUM(13完成量) * 100,2),'%') as 130完成签收,
							concat(ROUND(SUM(13完成量) / SUM(13总量) * 100,2),'%') as 130完成占比,
						SUM(14总量) AS 140总单量,
							concat(ROUND(SUM(14签收量) / SUM(14总量) * 100,2),'%') as 140总计签收,
							concat(ROUND(SUM(14签收量) / SUM(14完成量) * 100,2),'%') as 140完成签收,
							concat(ROUND(SUM(14完成量) / SUM(14总量) * 100,2),'%') as 140完成占比,
						SUM(15总量) AS 150总单量,
							concat(ROUND(SUM(15签收量) / SUM(15总量) * 100,2),'%') as 150总计签收,
							concat(ROUND(SUM(15签收量) / SUM(15完成量) * 100,2),'%') as 150完成签收,
							concat(ROUND(SUM(15完成量) / SUM(15总量) * 100,2),'%') as 150完成占比,
						SUM(16总量) AS 160总单量,
							concat(ROUND(SUM(16签收量) / SUM(16总量) * 100,2),'%') as 160总计签收,
							concat(ROUND(SUM(16签收量) / SUM(16完成量) * 100,2),'%') as 160完成签收,
							concat(ROUND(SUM(16完成量) / SUM(16总量) * 100,2),'%') as 160完成占比
                    FROM(SELECT 所属团队 as 家族,币种 地区, 产品id, 产品名称, 父级分类,
                                COUNT(cx.`订单编号`) as 总单量,
                    --        SUM(IF(年月 = '{1}',1,0)) as 04总量,
                    --            SUM(IF(年月 = '{1}' AND 最终状态 = 已签收,1,0)) as 04签收量,
                    --            SUM(IF(年月 = '{1}' AND 最终状态 IN (已签收,拒收,已退货,理赔, 自发头程丢件),1,0)) as 04完成量,
                    --        SUM(IF(年月 = '{2}',1,0)) as 05总量,
                    --            SUM(IF(年月 = '{2}' AND 最终状态 = 已签收,1,0)) as 05签收量,
                    --            SUM(IF(年月 = '{2}' AND 最终状态 IN (已签收,拒收,已退货,理赔, 自发头程丢件),1,0)) as 05完成量,
                    --        SUM(IF(年月 = '{3}',1,0)) as 06总量,
                    --            SUM(IF(年月 = '{3}' AND 最终状态 = 已签收,1,0)) as 06签收量,
                    --            SUM(IF(年月 = '{3}' AND 最终状态 IN (已签收,拒收,已退货,理赔, 自发头程丢件),1,0)) as 06完成量,
                    --        SUM(IF(年月 = '{4}',1,0)) as 07总量,
                    --            SUM(IF(年月 = '{4}' AND 最终状态 = 已签收,1,0)) as 07签收量,
                    --            SUM(IF(年月 = '{4}' AND 最终状态 IN (已签收,拒收,已退货,理赔, 自发头程丢件),1,0)) as 07完成量,
                    --        SUM(IF(年月 = '{5}',1,0)) as 08总量,
                    --            SUM(IF(年月 = '{5}' AND 最终状态 = 已签收,1,0)) as 08签收量,
                    --            SUM(IF(年月 = '{5}' AND 最终状态 IN (已签收,拒收,已退货,理赔, 自发头程丢件),1,0)) as 08完成量,
                    --        SUM(IF(年月 = '{6}',1,0)) as 09总量,
                    --            SUM(IF(年月 = '{6}' AND 最终状态 = 已签收,1,0)) as 09签收量,
                    --            SUM(IF(年月 = '{6}' AND 最终状态 IN (已签收,拒收,已退货,理赔, 自发头程丢件),1,0)) as 09完成量,
                            SUM(IF(年月 = '{7}',1,0)) as 10总量,
                                SUM(IF(年月 = '{7}' AND 最终状态 = '已签收',1,0)) as 10签收量,
                                SUM(IF(年月 = '{7}' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 10完成量,
                            SUM(IF(年月 = '{8}',1,0)) as 11总量,
                                SUM(IF(年月 = '{8}' AND 最终状态 = '已签收',1,0)) as 11签收量,
                                SUM(IF(年月 = '{8}' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 11完成量,
                            SUM(IF(年月 = '{9}',1,0)) as 12总量,
                                SUM(IF(年月 = '{9}' AND 最终状态 = '已签收',1,0)) as 12签收量,
                                SUM(IF(年月 = '{9}' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 12完成量,
                            SUM(IF(年月 = '{10}',1,0)) as 13总量,
                                SUM(IF(年月 = '{10}' AND 最终状态 = '已签收',1,0)) as 13签收量,
                                SUM(IF(年月 = '{10}' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 13完成量,
                            SUM(IF(年月 = '{11}',1,0)) as 14总量,
                                SUM(IF(年月 = '{11}' AND 最终状态 = '已签收',1,0)) as 14签收量,
                                SUM(IF(年月 = '{11}' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 14完成量,
                            SUM(IF(年月 = '{12}',1,0)) as 15总量,
                                SUM(IF(年月 = '{12}' AND 最终状态 = '已签收',1,0)) as 15签收量,
                                SUM(IF(年月 = '{12}' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 15完成量,
                            SUM(IF(年月 = '{13}',1,0)) as 16总量,
                                SUM(IF(年月 = '{13}' AND 最终状态 = '已签收',1,0)) as 16签收量,
                                SUM(IF(年月 = '{13}' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 16完成量
                        FROM gat_zqsb cx
                        where cx.年月 >= '{1}' AND cx.`运单编号` is not null AND cx.所属团队 NOT IN ({0})
                        GROUP BY cx.所属团队,cx.币种,cx.产品id
                    ) s
					GROUP BY s.家族,s.地区,s.产品id
					WITH ROLLUP
				) ss
                ORDER BY FIELD(ss.`家族`,{14}, '总计'),
                        FIELD(ss.地区, '台湾', '香港', '总计' ),
                        ss.总单量 DESC;'''.format(not_team, t1,t2,t3,t4,t5,t6,t7,t8,t9,t10,t11,t12,t13, self.team_name2)
        # sql51 = '''SELECT *
        #         FROM (SELECT IFNULL(家族, '总计') 家族, IFNULL(地区, '总计') 地区, IFNULL(产品id, '总计') 产品id,  IFNULL(产品名称, '总计') 产品名称, IFNULL(父级分类, '总计') 父级分类,
		# 				    SUM(总单量) 总单量,
		# 				SUM(04总量) 202201总单量,
		# 					concat(ROUND(SUM(04签收量) / SUM(04总量) * 100,2),'%') as 202201总计签收,
		# 					concat(ROUND(SUM(04签收量) / SUM(04完成量) * 100,2),'%') as 202201完成签收,
		# 					concat(ROUND(SUM(04完成量) / SUM(04总量) * 100,2),'%') as 202201完成占比,
		# 				SUM(05总量) 202202总单量,
		# 					concat(ROUND(SUM(05签收量) / SUM(05总量) * 100,2),'%') as 202202总计签收,
		# 					concat(ROUND(SUM(05签收量) / SUM(05完成量) * 100,2),'%') as 202202完成签收,
		# 					concat(ROUND(SUM(05完成量) / SUM(05总量) * 100,2),'%') as 202202完成占比,
		# 				SUM(06总量) 202203总单量,
		# 					concat(ROUND(SUM(06签收量) / SUM(06总量) * 100,2),'%') as 202203总计签收,
		# 					concat(ROUND(SUM(06签收量) / SUM(06完成量) * 100,2),'%') as 202203完成签收,
		# 					concat(ROUND(SUM(06完成量) / SUM(06总量) * 100,2),'%') as 202203完成占比,
		# 				SUM(07总量) 202204总单量,
		# 					concat(ROUND(SUM(07签收量) / SUM(07总量) * 100,2),'%') as 202204总计签收,
		# 					concat(ROUND(SUM(07签收量) / SUM(07完成量) * 100,2),'%') as 202204完成签收,
		# 					concat(ROUND(SUM(07完成量) / SUM(07总量) * 100,2),'%') as 202204完成占比,
		# 				SUM(08总量) 202205总单量,
		# 					concat(ROUND(SUM(08签收量) / SUM(08总量) * 100,2),'%') as 202205总计签收,
		# 					concat(ROUND(SUM(08签收量) / SUM(08完成量) * 100,2),'%') as 202205完成签收,
		# 					concat(ROUND(SUM(08完成量) / SUM(08总量) * 100,2),'%') as 202205完成占比,
		# 				SUM(09总量) 202206总单量,
		# 					concat(ROUND(SUM(09签收量) / SUM(09总量) * 100,2),'%') as 202206总计签收,
		# 					concat(ROUND(SUM(09签收量) / SUM(09完成量) * 100,2),'%') as 202206完成签收,
		# 					concat(ROUND(SUM(09完成量) / SUM(09总量) * 100,2),'%') as 202206完成占比,
		# 				SUM(10总量) 202207总单量,
		# 					concat(ROUND(SUM(10签收量) / SUM(10总量) * 100,2),'%') as 202207总计签收,
		# 					concat(ROUND(SUM(10签收量) / SUM(10完成量) * 100,2),'%') as 202207完成签收,
		# 					concat(ROUND(SUM(10完成量) / SUM(10总量) * 100,2),'%') as 202207完成占比,
		# 				SUM(11总量) 202208总单量,
		# 					concat(ROUND(SUM(11签收量) / SUM(11总量) * 100,2),'%') as 202208总计签收,
		# 					concat(ROUND(SUM(11签收量) / SUM(11完成量) * 100,2),'%') as 202208完成签收,
		# 					concat(ROUND(SUM(11完成量) / SUM(11总量) * 100,2),'%') as 202208完成占比,
		# 				SUM(12总量) 202209总单量,
		# 					concat(ROUND(SUM(12签收量) / SUM(12总量) * 100,2),'%') as 202209总计签收,
		# 					concat(ROUND(SUM(12签收量) / SUM(12完成量) * 100,2),'%') as 202209完成签收,
		# 					concat(ROUND(SUM(12完成量) / SUM(12总量) * 100,2),'%') as 202209完成占比,
		# 				SUM(13总量) 202210总单量,
		# 					concat(ROUND(SUM(12签收量) / SUM(12总量) * 100,2),'%') as 202210总计签收,
		# 					concat(ROUND(SUM(12签收量) / SUM(12完成量) * 100,2),'%') as 202210完成签收,
		# 					concat(ROUND(SUM(12完成量) / SUM(12总量) * 100,2),'%') as 202210完成占比,
		# 				SUM(14总量) 202211总单量,
		# 					concat(ROUND(SUM(14签收量) / SUM(14总量) * 100,2),'%') as 202211总计签收,
		# 					concat(ROUND(SUM(14签收量) / SUM(14完成量) * 100,2),'%') as 202211完成签收,
		# 					concat(ROUND(SUM(14完成量) / SUM(14总量) * 100,2),'%') as 202211完成占比,
		# 				SUM(15总量) 202212总单量,
		# 					concat(ROUND(SUM(15签收量) / SUM(15总量) * 100,2),'%') as 202212总计签收,
		# 					concat(ROUND(SUM(15签收量) / SUM(15完成量) * 100,2),'%') as 202212完成签收,
		# 					concat(ROUND(SUM(15完成量) / SUM(15总量) * 100,2),'%') as 202212完成占比,
		# 				SUM(16总量) 202301总单量,
		# 					concat(ROUND(SUM(16签收量) / SUM(16总量) * 100,2),'%') as 202301总计签收,
		# 					concat(ROUND(SUM(16签收量) / SUM(16完成量) * 100,2),'%') as 202301完成签收,
		# 					concat(ROUND(SUM(16完成量) / SUM(16总量) * 100,2),'%') as 202301完成占比
        #             FROM(SELECT 家族,币种 地区, 产品id, 产品名称, 父级分类,
        #                         COUNT(cx.`订单编号`) as 总单量,
        #                     SUM(IF(年月 = 202201,1,0)) as 04总量,
        #                         SUM(IF(年月 = 202201 AND 最终状态 = "已签收",1,0)) as 04签收量,
        #                         SUM(IF(年月 = 202201 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 04完成量,
        #                     SUM(IF(年月 = 202202,1,0)) as 05总量,
        #                         SUM(IF(年月 = 202202 AND 最终状态 = "已签收",1,0)) as 05签收量,
        #                         SUM(IF(年月 = 202202 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 05完成量,
        #                     SUM(IF(年月 = 202203,1,0)) as 06总量,
        #                         SUM(IF(年月 = 202203 AND 最终状态 = "已签收",1,0)) as 06签收量,
        #                         SUM(IF(年月 = 202203 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 06完成量,
        #                     SUM(IF(年月 = 202204,1,0)) as 07总量,
        #                         SUM(IF(年月 = 202204 AND 最终状态 = "已签收",1,0)) as 07签收量,
        #                         SUM(IF(年月 = 202204 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 07完成量,
        #                     SUM(IF(年月 = 202205,1,0)) as 08总量,
        #                         SUM(IF(年月 = 202205 AND 最终状态 = "已签收",1,0)) as 08签收量,
        #                         SUM(IF(年月 = 202205 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 08完成量,
        #                     SUM(IF(年月 = 202206,1,0)) as 09总量,
        #                         SUM(IF(年月 = 202206 AND 最终状态 = "已签收",1,0)) as 09签收量,
        #                         SUM(IF(年月 = 202206 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 09完成量,
        #                     SUM(IF(年月 = 202207,1,0)) as 10总量,
        #                         SUM(IF(年月 = 202207 AND 最终状态 = "已签收",1,0)) as 10签收量,
        #                         SUM(IF(年月 = 202207 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 10完成量,
        #                     SUM(IF(年月 = 202208,1,0)) as 11总量,
        #                         SUM(IF(年月 = 202208 AND 最终状态 = "已签收",1,0)) as 11签收量,
        #                         SUM(IF(年月 = 202208 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 11完成量,
        #                     SUM(IF(年月 = 202209,1,0)) as 12总量,
        #                         SUM(IF(年月 = 202209 AND 最终状态 = "已签收",1,0)) as 12签收量,
        #                         SUM(IF(年月 = 202209 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 12完成量,
        #                     SUM(IF(年月 = 202210,1,0)) as 13总量,
        #                         SUM(IF(年月 = 202210 AND 最终状态 = "已签收",1,0)) as 13签收量,
        #                         SUM(IF(年月 = 202210 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 13完成量,
        #                     SUM(IF(年月 = 202211,1,0)) as 14总量,
        #                         SUM(IF(年月 = 202211 AND 最终状态 = "已签收",1,0)) as 14签收量,
        #                         SUM(IF(年月 = 202211 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 14完成量,
        #                     SUM(IF(年月 = 202212,1,0)) as 15总量,
        #                         SUM(IF(年月 = 202212 AND 最终状态 = "已签收",1,0)) as 15签收量,
        #                         SUM(IF(年月 = 202212 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 15完成量,
        #                     SUM(IF(年月 = 202301,1,0)) as 16总量,
        #                         SUM(IF(年月 = 202301 AND 最终状态 = "已签收",1,0)) as 16签收量,
        #                         SUM(IF(年月 = 202301 AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 16完成量
        #                 FROM gat_zqsb_cache cx
        #                 where cx.`运单编号` is not null AND cx.团队 NOT IN ({0})
        #                 GROUP BY cx.家族,cx.币种,cx.产品id
        #             ) s
		# 			GROUP BY s.家族,s.地区,s.产品id
		# 			WITH ROLLUP
		# 		) ss
        #         ORDER BY FIELD(ss.`家族`, {1}),
        #                 FIELD(ss.地区, '台湾', '香港', '总计' ),
        #                 ss.总单量 DESC;'''.format(not_team, team_name2)
        df51 = pd.read_sql_query(sql=sql51, con=self.engine1)
        print(df51.columns)
        # df51.rename(columns={'11总单量': t1 + '总单量', '11总计签收': t1 + '总计签收', '11完成签收': t1 + '完成签收', '11完成占比': t1 + '完成占比',
        #                      '22总单量': t2 + '总单量', '22总计签收': t2 + '总计签收', '22完成签收': t2 + '完成签收', '22完成占比': t2 + '完成占比',
        #                      '33总单量': t3 + '总单量', '33总计签收': t3 + '总计签收', '33完成签收': t3 + '完成签收', '33完成占比': t3 + '完成占比',
        #                      '44总单量': t4 + '总单量', '44总计签收': t4 + '总计签收', '44完成签收': t4 + '完成签收', '44完成占比': t4 + '完成占比',
        #                      '55总单量': t5 + '总单量', '55总计签收': t5 + '总计签收', '55完成签收': t5 + '完成签收', '55完成占比': t5 + '完成占比',
        #                      '66总单量': t6 + '总单量', '66总计签收': t6 + '总计签收', '66完成签收': t6 + '完成签收', '66完成占比': t6 + '完成占比',
        #                      '77总单量': t7 + '总单量', '77总计签收': t7 + '总计签收', '77完成签收': t7 + '完成签收', '77完成占比': t7 + '完成占比',
        #                      '88总单量': t8 + '总单量', '88总计签收': t8 + '总计签收', '88完成签收': t8 + '完成签收', '88完成占比': t8 + '完成占比',
        #                      '99总单量': t9 + '总单量', '99总计签收': t9 + '总计签收', '99完成签收': t9 + '完成签收', '99完成占比': t9 + '完成占比',
        #                      '100总单量': t10 + '总单量', '100总计签收': t10 + '总计签收', '100完成签收': t10 + '完成签收', '100完成占比': t10 + '完成占比',
        #                      '110总单量': t11 + '总单量', '110总计签收': t11 + '总计签收', '110完成签收': t11 + '完成签收', '110完成占比': t11 + '完成占比',
        #                      '120总单量': t12 + '总单量', '120总计签收': t12 + '总计签收', '120完成签收': t12 + '完成签收', '120完成占比': t12 + '完成占比',
        #                      '130总单量': t13 + '总单量', '130总计签收': t13 + '总计签收', '130完成签收': t13 + '完成签收', '130完成占比': t13 + '完成占比'
        #                      }, inplace=True)
        df51.rename(columns={'100总单量': t7 + '总单量', '100总计签收': t7 + '总计签收', '100完成签收': t7 + '完成签收', '100完成占比': t7 + '完成占比',
                             '110总单量': t8 + '总单量', '110总计签收': t8 + '总计签收', '110完成签收': t8 + '完成签收', '110完成占比': t8 + '完成占比',
                             '120总单量': t9 + '总单量', '120总计签收': t9 + '总计签收', '120完成签收': t9 + '完成签收', '120完成占比': t9 + '完成占比',
                             '130总单量': t10 + '总单量', '130总计签收': t10 + '总计签收', '130完成签收': t10 + '完成签收', '130完成占比': t10 + '完成占比',
                             '140总单量': t11 + '总单量', '140总计签收': t11 + '总计签收', '140完成签收': t11 + '完成签收', '140完成占比': t11 + '完成占比',
                             '150总单量': t12 + '总单量', '150总计签收': t12 + '总计签收', '150完成签收': t12 + '完成签收', '150完成占比': t12 + '完成占比',
                             '160总单量': t13 + '总单量', '160总计签收': t13 + '总计签收', '160完成签收': t13 + '完成签收', '160完成占比': t13 + '完成占比' }, inplace=True)
        print(df51.columns)
        listT.append(df51)

        # 13、各团队-问题率
        # print('正在获取---3、各团队-问题率…………')
        sql02 = '''SELECT *
                FROM (
                    (SELECT 币种,'核实地址' AS 问题原因,核实地址 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实地址%",1,0)) as 核实地址
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实姓名' AS 问题原因,核实姓名 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实姓名%",1,0)) as 核实姓名
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实规格' AS 问题原因,核实规格 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实规格%",1,0)) as 核实规格
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实数量' AS 问题原因,核实数量 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实数量%",1,0)) as 核实数量
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实电话' AS 问题原因,核实电话 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实电话%",1,0)) as 核实电话
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'重复下单' AS 问题原因,重复下单 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%重复下单%",1,0)) as 重复下单
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实拉黑率' AS 问题原因,核实拉黑率 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实拉黑率%",1,0)) as 核实拉黑率
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实IP' AS 问题原因,核实IP AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实IP%",1,0)) as 核实IP
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'回复留言' AS 问题原因,回复留言 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%回复留言%",1,0)) as 回复留言
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实金额' AS 问题原因,核实金额 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实金额%",1,0)) as 核实金额
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'删运单号' AS 问题原因,删运单号 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%删运单号%",1,0)) as 删运单号
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'可疑订单' AS 问题原因,可疑订单 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%可疑订单%",1,0)) as 可疑订单
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实邮箱' AS 问题原因,核实邮箱 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实邮箱%",1,0)) as 核实邮箱
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'无法派送地区' AS 问题原因,无法派送地区 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%无法派送地区%",1,0)) as 无法派送地区
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = '2021-10-02'
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'核实邮编' AS 问题原因,核实邮编 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%核实邮编%",1,0)) as 核实邮编
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'拼团未完成' AS 问题原因,拼团未完成 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%拼团未完成%",1,0)) as 拼团未完成
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'支付失败' AS 问题原因,支付失败 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%支付失败%",1,0)) as 支付失败
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )
                UNION
                    (SELECT 币种,'未支付' AS 问题原因,未支付 AS 数量
                        FROM (SELECT 币种,SUM(IF(问题原因 LIKE "%未支付%",1,0)) as 未支付
                                FROM  gat_order_list gs
                                WHERE gs.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY)
                                GROUP BY gs.`币种`
                        ) ss
                    )					
                ) scs
                ORDER BY 币种 , 数量 DESC;'''
        # df02 = pd.read_sql_query(sql=sql02, con=self.engine1)
        # listT.append(df02)

        # 11、各团队-各二级品类
        # print('正在获取---3、各团队-各二级品类…………')
        sql20 = '''SELECT *
                FROM(SELECT IFNULL(cx.`年月`, '总计') 月份,
                                IFNULL(cx.`币种`, '总计') 地区,
                                IFNULL(cx.`所属团队`, '总计') 家族,
                                IFNULL(cx.`父级分类`, '总计') 父级分类,
                                IFNULL(cx.`二级分类`, '总计') 二级分类,
                                COUNT(cx.`订单编号`) as 总单量,
                                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF( 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
                                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 总计签收,
                                concat(ROUND(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 完成占比,
                                concat(ROUND(SUM(IF(最终状态 = "已退货",1,0)) / COUNT(cx.`订单编号`) * 100,2),'%') as 退款率,
                                concat(ROUND(SUM(IF(最终状态 = "已签收",价格RMB,0)) / SUM(价格RMB) * 100,2),'%') as '总计签收(金额)',
                                concat(ROUND(COUNT(cx.`订单编号`) / 总订单量 * 100,2),'%') as 品类占比,
                                ROUND(SUM(价格RMB) / COUNT(cx.`订单编号`),2) as 平均客单价,
                            SUM(IF(`是否改派` = '直发',1,0))  as 直发单量,
                                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 直发完成签收,
                                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发总计签收,
                                concat(ROUND(SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '直发',1,0)) * 100,2),'%') as 直发完成占比,
                                concat(ROUND(SUM(IF(`是否改派` = '直发',1,0)) / 直发总单量 * 100,2),'%') as 直发品类占比,
                            concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / COUNT(cx.`订单编号`) * 100,2),'%')as 改派占比,
                                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 改派完成签收,
                                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派总计签收,
                                concat(ROUND(SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / SUM(IF(`是否改派` = '改派',1,0)) * 100,2),'%') as 改派完成占比,
                                concat(ROUND(SUM(IF(`是否改派` = '改派',1,0)) / 改派总单量 * 100,2),'%') as 改派品类占比
                        FROM (SELECT *, 所属团队 AS 家族
                                FROM {0}_zqsb cc
                                where cc.年月 >= '{1}' AND cc.`运单编号` is not null AND cc.团队 NOT IN ({2})
                        ) cx 
                        LEFT JOIN 
        				( SELECT 币种, 所属团队,年月,count(订单编号) as 总订单量,SUM(IF(`是否改派`= '直发',1,0)) as 直发总单量,SUM(IF(`是否改派` = '改派',1,0)) as 改派总单量
        					FROM {0}_zqsb cc
        					where cc.年月 >= '{1}' AND cc.`运单编号` is not null AND cc.团队 NOT IN ({2})
        					GROUP BY cc.币种,cc.所属团队,cc.年月
        				) cx2 ON cx.币种 = cx2.币种 AND  cx.所属团队 = cx2.所属团队 AND  cx.年月 = cx2.年月                       
                        GROUP BY cx.年月,cx.币种,cx.家族,cx.父级分类,cx.二级分类
                        WITH ROLLUP 
                ) s
                ORDER BY 月份 DESC,
                        FIELD( 地区, '台湾', '香港', '总计' ),
                        FIELD( s.家族, {3}, '总计'),
                        FIELD(s.父级分类, "居家百货", "电子电器", "服饰", "医药保健", "鞋类", "美容个护", "包类","钟表珠宝","母婴玩具","包材类","合计"),
                        FIELD(s.二级分类, "上衣","下装","内衣","套装","裙子","配饰","母婴服饰","凉/拖鞋","皮鞋","休闲运动鞋","靴子",
                                           "单肩包","双肩包","钱包","行李箱包","厨房用品","日用百货","布艺家纺","宠物用品","户外运动","汽车用品","家装建材","办公/文化",
                                           "手表手环","影音娱乐","电脑外设","手机外设","家用电器","个护电器","智能设备","彩妆","护肤","个人洗护",
                                           "保健食品","护理护具","保健器械","药品","成人保健","手表","钟表","饰品","玩具","母婴用品","仓库包材","仓库耗材","合计"),
                        s.总单量 DESC;'''.format(team, gat_time, not_team, self.team_name2)
        # df20 = pd.read_sql_query(sql=sql20, con=self.engine1)
        # listT.append(df20)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        file_path = 'F:\\输出文件\\{} {}-签收率.xlsx'.format(today, match[team])
        sheet_name = ['每日各团队', '审核率_删单率', '各月各团队', '各月各团队分旬', '各团队各品类', '各团队各物流', '各团队各平台', '各平台各团队', '各品类各团队', '各物流各团队', '同产品各团队','同产品各月', '各团队二级品类']
        # df0 = pd.DataFrame([])                                          # 创建空的dataframe数据框
        # df0.to_excel(file_path, index=False)                            # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')           # 初始化写入对象
        # book = load_workbook(file_path)                                 # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book                                              # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # for i in range(len(listT)):
        #     listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        # if 'Sheet1' in book.sheetnames:                                 # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for i in range(len(listT)):
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        print('正在运行' + match[team] + '表宏…………（xlwings方法一）')
        try:
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            # app.display_alerts = False
            app.screen_updating = False
            wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('zl_report_now.zl_report_day')()
            wbsht1.save()
            wbsht1.close()
            wbsht.save()
            wbsht.close()
            app.quit()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))

        # print('正在运行' + match[team] + '表宏…………（win32com方法二）')
        # xls = win32com.client.Dispatch("Excel.Application")
        # wb = xls.workbooks.open('E:/桌面文件/新版-格式转换(python表).xlsm')  ##存储vba代码的文件
        # wb.Application.DisplayAlerts = False
        # try:
        #     wb1 = xls.workbooks.open(file_path)
        #     wb.Application.Run('zl_gat_report_new.gat_总_品类_物流_两月签收率')  ##开始调用vba宏
        #     wb1.Application.Save()
        #     wb1.Application.close()
        #     wb.Application.Save()
        #     wb.Application.close()
        # except Exception as e:
        #     print(e)
        # xls.Application.Quit()

        new_path = 'F:\\神龙签收率\\' + (datetime.datetime.now()).strftime('%m.%d') + '\\{} {}-签收率.xlsx'.format(today, match[team])
        shutil.copyfile(file_path, new_path)        # copy到指定位置
        print('----已写入excel; 并复制到指定文件夹中')

        print("强制关闭Execl后台进程中......")
        system('taskkill /F /IM EXCEL.EXE')

    # 更新-地区签收率(自己看的)
    def address_repot(self, team, month_last, month_yesterday):    # 更新-地区签收率
        today = datetime.date.today().strftime('%Y.%m.%d')
        match = {'gat': '港台'}
        print(month_last + '---' + month_yesterday)
        try:
            print('正在更新单表中......')
            sql = '''update {0}_order_list a, gat_update b
                            set a.`省洲`= IF(b.`省洲` = '', NULL, b.`省洲`),
                                a.`市区`= IF(b.`市区` = '', NULL, b.`市区`)
        		            where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('正在更新总表中......')
            sql = '''update {0}_zqsb a, gat_update b
                            set a.`省洲`= IF(b.`省洲` = '', NULL, b.`省洲`),
                                a.`市区`= IF(b.`市区` = '', NULL, b.`市区`)
                    		where a.`订单编号`= b.`订单编号`;'''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')

        # 地区分类
        sheet_name = ['地区分类', '地区总']
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---地区签收率…………')
        sql0 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.省洲,
    						IF(s2.签收=0,NULL,s2.签收) as 签收,
    						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
    						IF(s2.在途=0,NULL,s2.在途) as 在途,
    						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
    						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
    						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,
    						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
    						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
    						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
    						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
    						IF(s2.总订单=0,NULL,s2.总订单) as 全部,
                        concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                            concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                            concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                            concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                            concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                            concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                            concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                            concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                            concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                    FROM ( SELECT IFNULL(s1.币种,'合计') as 币种,
                                IFNULL(s1.家族,'合计') as 家族,
                                IFNULL(s1.年月,'合计') as 年月,
                                IFNULL(s1.是否改派,'合计') as 是否改派,
                                IFNULL(s1.省洲,'合计') as 省洲,
                                SUM(s1.签收) as 签收,
                                SUM(s1.拒收) as 拒收,
                                SUM(s1.在途) as 在途,
                                SUM(s1.未发货) as 未发货,
                                SUM(s1.未上线) as 未上线,
                                SUM(s1.已退货) as 已退货,
                                SUM(s1.理赔) as 理赔,
                                SUM(s1.自发头程丢件) as 自发头程丢件,
                                SUM(s1.已发货) as 已发货,
                                SUM(s1.已完成) as 已完成,
                                SUM(s1.总订单) as 总订单,
                                SUM(s1.签收金额) as 签收金额,
                                SUM(s1.退货金额) as 退货金额,
                                SUM(s1.完成金额) as 完成金额,
                                SUM(s1.发货金额) as 发货金额,
                                SUM(s1.总计金额) as 总计金额
                        FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.省洲 as 省洲,
                                    SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                    SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                    SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                    SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                    SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                    SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                    SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                    SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                    count(订单编号) as 总订单,
                                    count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                    SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                    SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                    SUM(`价格RMB`) as 总计金额,
                                    SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                                FROM (SELECT *, 所属团队 as 家族
                                        FROM {0}_zqsb cc 
                                        where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                ) cx
                                GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派`, cx.`省洲`
                                ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`是否改派`, s1.`省洲`
                        with rollup
                    ) s2
                    GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`省洲`
                    HAVING s2.年月 <> '合计'
        ORDER BY FIELD(s2.`家族`,{3}, '总计'),
                FIELD(s2.`币种`,'台湾','香港','合计'),
                s2.`年月`,
                FIELD(s2.`是否改派`,'改派','直发','合计'),
                FIELD(s2.`省洲`,'屏东县','高雄市','新竹市','宜兰县','新北市','花莲县','台东县','基隆市','台北市','新竹县',
                                '桃园市','苗栗县','台中市','彰化县','南投县','嘉义市','嘉义县','云林县','台南市','合计'),
                s2.总订单 DESC;'''.format(team, month_last, month_yesterday, self.team_name2)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)

        sql1 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.省洲,
            						IF(s2.签收=0,NULL,s2.签收) as 签收,
            						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
            						IF(s2.在途=0,NULL,s2.在途) as 在途,
            						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
            						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
            						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,
            						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
            						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
            						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
            						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
            						IF(s2.总订单=0,NULL,s2.总订单) as 全部,
                                concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                                    concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                                    concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                                    concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                                    concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                                concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                                    concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                                    concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                                    concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                                    concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                            FROM ( SELECT IFNULL(s1.币种,'合计') as 币种,
                                        IFNULL(s1.家族,'合计') as 家族,
                                        IFNULL(s1.年月,'合计') as 年月,
                                        IFNULL(s1.是否改派,'合计') as 是否改派,
                                        IFNULL(s1.省洲,'合计') as 省洲,
                                        SUM(s1.签收) as 签收,
                                        SUM(s1.拒收) as 拒收,
                                        SUM(s1.在途) as 在途,
                                        SUM(s1.未发货) as 未发货,
                                        SUM(s1.未上线) as 未上线,
                                        SUM(s1.已退货) as 已退货,
                                        SUM(s1.理赔) as 理赔,
                                        SUM(s1.自发头程丢件) as 自发头程丢件,
                                        SUM(s1.已发货) as 已发货,
                                        SUM(s1.已完成) as 已完成,
                                        SUM(s1.总订单) as 总订单,
                                        SUM(s1.签收金额) as 签收金额,
                                        SUM(s1.退货金额) as 退货金额,
                                        SUM(s1.完成金额) as 完成金额,
                                        SUM(s1.发货金额) as 发货金额,
                                        SUM(s1.总计金额) as 总计金额
                                FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.省洲 as 省洲,
                                            SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                            SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                            SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                            SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                            SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                            SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                            SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                            SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                            SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                            count(订单编号) as 总订单,
                                            count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                            SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                            SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                            SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                            SUM(`价格RMB`) as 总计金额,
                                            SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                                        FROM (SELECT *,所属团队 as 家族
                                                FROM {0}_zqsb cc 
                                                where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                        ) cx
                                        GROUP BY cx.`币种`,cx.`年月`, cx.`是否改派`, cx.`省洲`
                                        ORDER BY cx.`币种`,cx.`年月`, cx.`是否改派` DESC,总订单 DESC
                                ) s1
                                GROUP BY s1.`币种`, s1.`年月`, s1.`省洲`
                                with rollup
                            ) s2
                            GROUP BY s2.`币种`, s2.`年月`, s2.`省洲`
                            HAVING s2.年月 <> '合计'
                ORDER BY FIELD(s2.`币种`,'台湾','香港','合计'),
                        s2.`年月`,
                        FIELD(s2.`是否改派`,'改派','直发','合计'),
                        FIELD(s2.`省洲`,'屏東縣','高雄市','新竹市','宜蘭縣','新北市','花蓮縣','臺東縣','基隆市','臺北市','新竹縣',
                                        '桃園市','苗栗縣','臺中市','彰化縣','南投縣','嘉義市','嘉義縣','雲林縣','臺南市','合计'),
                        s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df1 = pd.read_sql_query(sql=sql1, con=self.engine1)
        listT.append(df1)

        print('正在将 地区签收率 写入excel…………')
        file_path = 'F:\\输出文件\\{} {} 地区-签收率.xlsx'.format(today, match[team])
        # df0 = pd.DataFrame([])                                  # 创建空的dataframe数据框
        # df0.to_excel(file_path, index=False)                    # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')   # 初始化写入对象
        # book = load_workbook(file_path)                         # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book                                      # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # for i in range(len(listT)):
        #     listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        # if 'Sheet1' in book.sheetnames:                         # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for i in range(len(listT)):
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_总_地区_两月签收率')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')
    # 更新上期-总表 （备用）
    def replaceHostbefore(self, team, last_time):
        try:
            print('正在获取往昔数据中......')
            sql = '''SELECT 年月, 旬, 日期, 团队,币种, 订单来源, 订单编号, 出货时间, 状态时间, 上线时间, 最终状态,是否改派,物流方式, 产品id, 
                            父级分类,二级分类,三级分类,下单时间, 审核时间,仓储扫描时间,完结状态时间,IF(价格RMB = '',null,价格RMB) as 价格RMB, '{0}' as 记录时间
                    FROM gat_update;'''.format(last_time)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            print('正在添加缓存中......')
            df.to_sql('gat_update_cp', con=self.engine1, index=False, if_exists='replace')
            print('正在数据添加中......')
            sql = '''REPLACE INTO qsb_{0} SELECT * FROM gat_update_cp; '''.format(team)
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        except Exception as e:
            print('更新失败：', str(Exception) + str(e))
        print('更新成功…………')
    # report报表（备用）
    def qsb_report(self, team, day_yesterday, day_last):  # 获取各团队近两个月的物流数据
        match = {'gat': '港台'}
        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 每日
        sql0 = '''SELECT 月份,地区, 家族,
                        SUM(s.昨日订单量) as 昨日订单量,
                        SUM(s.直发签收) as 直发签收,
                        SUM(s.直发完成) as 直发完成,
                        SUM(s.直发总订单) as 直发总订单,
                        IFNULL(SUM(s.直发签收) / SUM(s.直发完成), 0) as 直发完成签收,
                        IFNULL(SUM(s.直发签收) / SUM(s.直发总订单), 0) as 直发总计签收,
                        IFNULL(SUM(s.直发完成) / SUM(s.直发总订单), 0) as 直发完成占比,
                        SUM(s.改派签收) as 改派签收,
                        SUM(s.改派完成) as 改派完成,
                        SUM(s.改派总订单) as 改派总订单,
                        IFNULL(SUM(s.改派签收) / SUM(s.改派完成), 0) as 改派完成签收,
                        IFNULL(SUM(s.改派签收) / SUM(s.改派总订单), 0) as 改派总计签收,
                        IFNULL(SUM(s.改派完成) / SUM(s.改派总订单), 0) as 改派完成占比
                FROM( SELECT IFNULL(cx.`年月`, '总计') 月份,
                            IFNULL(cx.币种, '总计') 地区,
                            IFNULL(cx.团队, '总计') 家族,
                            SUM(IF(cx.`日期` = DATE_SUB(CURDATE(), INTERVAL 1 DAY),1,0)) as 昨日订单量,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 = "已签收",1,0)) as 直发签收,
                            SUM(IF(`是否改派` = '直发' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 直发完成,
                            SUM(IF(`是否改派` = '直发',1,0)) as 直发总订单,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 = "已签收",1,0)) as 改派签收,
                            SUM(IF(`是否改派` = '改派' AND 最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 改派完成,
                            SUM(IF(`是否改派` = '改派',1,0)) as 改派总订单
                        FROM  qsb_gat cx
                        WHERE cx.`记录时间` = '{1}'
                        GROUP BY cx.年月,cx.币种,cx.团队
                        WITH ROLLUP 
                    ) s
                    GROUP BY 月份,地区,家族
                    ORDER BY 月份 DESC,
                            FIELD( 地区, '台湾', '香港', '总计' ),
                            FIELD( 家族, '神龙', '火凤凰', '红杉', '金狮', '总计' );'''.format(team, day_yesterday)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)

        # 总表
        sql = '''SELECT cx.币种 线路,
			                cx.团队 家族,
			                cx.年月 月份,
			                count(订单编号) as 总订单,
			                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
			                concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) * 100,2),'%') as 总计签收,
			                concat(ROUND(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) * 100,2),'%') as 完成占比,
			                null 序号
                    FROM qsb_gat cx
                    WHERE cx.`记录时间` = '{1}'
                    GROUP BY cx.币种,cx.团队,cx.年月
                    ORDER BY cx.币种,cx.团队,cx.年月;'''.format(team, day_yesterday)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df)
        # 总表-上月
        sql2 = '''SELECT 线路,家族,月份,总订单,完成签收,总计签收,完成占比,@rownum:=@rownum+1 AS 序号
	            FROM (SELECT cx.币种 线路,
        			        cx.团队 家族,
        			        cx.年月 月份,
        			        count(订单编号) as 总订单,
        			        concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) * 100,2),'%') as 完成签收,
        			        concat(ROUND(SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) * 100,2),'%') as 总计签收,
        			        concat(ROUND(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) * 100,2),'%') as 完成占比,
        			        @rownum:=0 
                        FROM qsb_gat cx
                        WHERE cx.`记录时间` = '{1}'
                        GROUP BY cx.币种,cx.团队,cx.年月
                    ) s
                ORDER BY s.线路,s.家族,s.月份;'''.format(team, day_last)
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
        listT.append(df2)

        # 物流
        sql3 = '''SELECT s2.币种,s2.团队 家族,s2.年月,s2.是否改派,s2.物流方式,
						s2.总订单,
						concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
						concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
						concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
						concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
						concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
						concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			            null 序号
				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.物流方式,'总计') as 物流方式,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.物流方式 as 物流方式,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                          ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
						    ) s1
						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
					   	    with rollup
					    ) s2
                ORDER BY    FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-台湾','火凤凰-台湾','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`物流方式`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)
        # 物流-上月
        sql4 = '''SELECT 币种,团队 家族,年月,是否改派,物流方式,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as '总计签收(金额)',累计占比, @rownum:=@rownum+1 AS 序号
		        FROM ( SELECT s2.币种,
        							s2.团队,
        							s2.年月,
        							s2.是否改派,
        							s2.物流方式,
        							s2.总订单,
        							concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        							concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        							concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        							concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        							concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
        							concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.物流方式,'总计') as 物流方式,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.物流方式 as 物流方式,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        						    ) s1
        						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
        					   	    with rollup
        					    ) s2
                        ) s
                        ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        					    FIELD(s.`团队`,'神龙家族-台湾','火凤凰-台湾','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        					    FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        					    FIELD(s.`是否改派`,'直发','改派','总计'),
        					    FIELD(s.`物流方式`,'总计'),
        					    s.总订单 DESC;'''.format(team, day_last)
        df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        listT.append(df4)

        # 品类
        sql5 = '''SELECT s2.币种,
								s2.团队 家族,
								s2.年月,
								s2.是否改派,
								s2.父级分类,
								s2.总订单,
								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
                                concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			                    null 序号
				 FROM (
                        SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.父级分类,'总计') as 父级分类,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	 SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.父级分类 as 父级分类,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                            ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
							) s1
						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
					   	with rollup
				 ) s2
				 ORDER BY	FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-台湾','火凤凰-台湾','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`父级分类`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)
        # 品类-上月
        sql5 = '''SELECT 币种,团队 家族,年月,是否改派,父级分类,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as `总计签收(金额)`,累计占比, @rownum:=@rownum+1 AS 序号
		        FROM (SELECT s2.币种,
        								s2.团队,
        								s2.年月,
        								s2.是否改派,
        								s2.父级分类,
        								s2.总订单,
        								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
                                        concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM (
                                SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.父级分类,'总计') as 父级分类,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	 SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.父级分类 as 父级分类,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                    ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        							) s1
        						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
        					   	with rollup
        				) s2 
        		) s
                ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        				FIELD(s.`团队`,'神龙家族-台湾','火凤凰-台湾','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        				FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        				FIELD(s.`是否改派`,'直发','改派','总计'),
        				FIELD(s.`父级分类`,'总计'),
        				s.总订单 DESC;'''.format(team, day_last)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)

        # 产品
        sql6 = '''SELECT * 
				    FROM ( SELECT   IFNULL( cx.`币种`,'总计') as 币种,
                                    IFNULL( cx.`团队`,'总计') as 家族,
                                    IFNULL( cx.`年月`,'总计') as 年月,
                                    IFNULL( cx.`产品id`,'总计') as 产品id,
                                    cx.`产品名称`,
							        cx.`父级分类`,
                                    count(订单编号) as 总订单,
                                    SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成签收,
                                    SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) as 总计签收,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) as 完成占比,
                                    count(订单编号) /总订单2 单量占比,
                                    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)),0) as 直发完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发完成占比,
                                    IFNULL(SUM(IF(是否改派 = '直发',1,0))  / 直发总订单2,0) as 直发单量占比,
                                    SUM(IF(是否改派 = '改派',1,0)) as 改派总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)),0) as 改派完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派完成占比,
                                    IFNULL(SUM(IF(是否改派 = '改派',1,0)) / 改派总订单2,0) 改派单量占比
                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                            LEFT JOIN  (SELECT 币种,团队,年月,count(订单编号) as 总订单2 , 
											    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单2 , 
												SUM(IF(是否改派 = '改派',1,0)) as 改派总订单2 
										FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') da GROUP BY da.币种,da.团队,da.年月
									) cx2  ON cx.币种 = cx2.币种 AND cx.团队 = cx2.团队 AND cx.年月 = cx2.年月
                            GROUP BY cx.币种,cx.团队,cx.年月,`产品id`
	                        with rollup
					) s1
	                ORDER BY	FIELD(s1.`币种`,'台湾','香港','总计'),
								FIELD(s1.`团队`,'神龙家族-台湾','火凤凰-台湾','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
								FIELD(s1.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
								总订单 DESC;'''.format(team, day_yesterday)
        df6 = pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)

        # 产品明细-台湾
        sql7 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
			            concat(ROUND(速派签收量 / 速派完成量 * 100,2),'%')  AS 速派完成签收,
			            concat(ROUND(速派签收量 / 速派单量 * 100,2),'%')  AS 速派总计签收,
			            concat(ROUND(速派完成量 / 速派单量 * 100,2),'%')  AS 速派完成占比,
			            concat(ROUND(速派单量 / 订单量 * 100,2),'%')  AS 速派单量占比,
			            concat(ROUND(711签收量 / 711完成量 * 100,2),'%')  AS 711完成签收,
			            concat(ROUND(711签收量 / 711单量 * 100,2),'%')  AS 711总计签收,
			            concat(ROUND(711完成量 / 711单量 * 100,2),'%')  AS 711完成占比,
			            concat(ROUND(711单量 / 订单量 * 100,2),'%')  AS 711单量占比,
			            concat(ROUND(天马签收量 / 天马完成量 * 100,2),'%')  AS 天马完成签收,
			            concat(ROUND(天马签收量 / 天马单量 * 100,2),'%')  AS 天马总计签收,
			            concat(ROUND(天马完成量 / 天马单量 * 100,2),'%')  AS 天马完成占比,
			            concat(ROUND(天马单量 / 订单量 * 100,2),'%')  AS 天马单量占比,
			            concat(ROUND(易速配签收量 / 易速配完成量 * 100,2),'%')  AS 易速配完成签收,
			            concat(ROUND(易速配签收量 / 易速配单量 * 100,2),'%')  AS 易速配总计签收,
			            concat(ROUND(易速配完成量 / 易速配单量 * 100,2),'%')  AS 易速配完成占比,
			            concat(ROUND(易速配单量 / 订单量 * 100,2),'%')  AS 易速配单量占比,
			            concat(ROUND(森鸿签收量 / 森鸿完成量 * 100,2),'%')  AS 森鸿完成签收,
			            concat(ROUND(森鸿签收量 / 森鸿单量 * 100,2),'%')  AS 森鸿总计签收,
			            concat(ROUND(森鸿完成量 / 森鸿单量 * 100,2),'%')  AS 森鸿完成占比,
			            concat(ROUND(森鸿单量 / 订单量 * 100,2),'%')  AS 森鸿单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
				            SUM(速派单量) 速派单量,  SUM(速派签收量) 速派签收量,  SUM(速派完成量) 速派完成量,
				            SUM(711单量) 711单量,  SUM(711签收量) 711签收量,  SUM(711完成量) 711完成量,
				            SUM(天马单量) 天马单量,  SUM(天马签收量) 天马签收量,  SUM(天马完成量) 天马完成量,
				            SUM(易速配单量) 易速配单量,  SUM(易速配签收量) 易速配签收量,  SUM(易速配完成量) 易速配完成量,
				            SUM(森鸿单量) 森鸿单量,  SUM(森鸿签收量) 森鸿签收量,  SUM(森鸿完成量) 森鸿完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派签收量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  速派完成量,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS '711单量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as  '711签收量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  '711完成量',
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马单量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as  天马签收量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  天马完成量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配单量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as  易速配签收量,
							    SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配完成量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿单量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as  森鸿签收量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '台湾'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-台湾', '火凤凰-台湾', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df7 = pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)
        # 产品明细-香港
        sql8 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
						concat(ROUND(立邦签收量 / 立邦完成量 * 100,2),'%')  AS 立邦完成签收,
						concat(ROUND(立邦签收量 / 立邦单量 * 100,2),'%')  AS 立邦总计签收,
						concat(ROUND(立邦完成量 / 立邦单量 * 100,2),'%')  AS 立邦完成占比,
						concat(ROUND(立邦单量 / 订单量 * 100,2),'%')  AS 立邦单量占比,
						concat(ROUND(森鸿SF签收量 / 森鸿SF完成量 * 100,2),'%')  AS 森鸿SF完成签收,
						concat(ROUND(森鸿SF签收量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF总计签收,
						concat(ROUND(森鸿SF完成量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF完成占比,
						concat(ROUND(森鸿SF单量 / 订单量 * 100,2),'%')  AS 森鸿SF单量占比,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH完成量 * 100,2),'%')  AS 森鸿SH完成签收,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH总计签收,
					    concat(ROUND(森鸿SH完成量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH完成占比,
					    concat(ROUND(森鸿SH单量 / 订单量 * 100,2),'%')  AS 森鸿SH单量占比,
					    concat(ROUND(易速配SF签收量 / 易速配SF完成量 * 100,2),'%')  AS 易速配SF完成签收,
					    concat(ROUND(易速配SF签收量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF总计签收,
					    concat(ROUND(易速配SF完成量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF完成占比,
					    concat(ROUND(易速配SF单量 / 订单量 * 100,2),'%')  AS 易速配SF单量占比,
					    concat(ROUND(易速配YC签收量 / 易速配YC完成量 * 100,2),'%')  AS 易速配YC完成签收,
					    concat(ROUND(易速配YC签收量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC总计签收,
					    concat(ROUND(易速配YC完成量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC完成占比,
					    concat(ROUND(易速配YC单量 / 订单量 * 100,2),'%')  AS 易速配YC单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
							SUM(立邦单量) 立邦单量,  SUM(立邦签收量) 立邦签收量,  SUM(立邦完成量) 立邦完成量,
				            SUM(森鸿SF单量) 森鸿SF单量,  SUM(森鸿SF签收量) 森鸿SF签收量,  SUM(森鸿SF完成量) 森鸿SF完成量,
				            SUM(森鸿SH单量) 森鸿SH单量,  SUM(森鸿SH签收量) 森鸿SH签收量,  SUM(森鸿SH完成量) 森鸿SH完成量,					
				            SUM(易速配SF单量) 易速配SF单量,  SUM(易速配SF签收量) 易速配SF签收量,  SUM(易速配SF完成量) 易速配SF完成量,				
				            SUM(易速配YC单量) 易速配YC单量,  SUM(易速配YC签收量) 易速配YC签收量,  SUM(易速配YC完成量) 易速配YC完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦签收量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  立邦完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿SF单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SF签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SF完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SH签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SH完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配SF单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as  易速配SF签收量,
							    SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配SF完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" ,1,0)) AS 易速配YC单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 = "已签收",1,0)) as  易速配YC签收量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配YC完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '香港'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-台湾', '火凤凰-台湾', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df8 = pd.read_sql_query(sql=sql8, con=self.engine1)
        listT.append(df8)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        for wbbook in ['神龙', '火凤凰', '红杉', '金狮']:
            file_path = 'F:\\输出文件\\{} {}-签收率.xlsx'.format(today, wbbook)
            sheet_name = ['每日', '总表', '总表上月', '物流', '物流上月', '品类', '品类上月', '产品', '产品明细台湾', '产品明细香港']
            # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            # for i in range(len(listT)):
            #     listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            #     del book['Sheet1']
            # writer.save()
            # writer.close()
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for i in range(len(listT)):
                    listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)

            # print('正在运行' + wbbook + '表宏…………')
            # app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            # app.display_alerts = False
            # wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
            # wbsht1 = app.books.open(file_path)
            # wbsht.macro('py_sl_总运行')()
            # wbsht1.save()
            # wbsht1.close()
            # wbsht.close()
            # app.quit()
        print('----已写入excel ')
    # 获取各团队近两个月的物流数据（备用）
    def qsb_report_T(self, team, day_yesterday, day_last):
        match = {'gat': '港台'}
        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 物流
        sql3 = '''SELECT s2.币种,s2.团队 家族,s2.年月,s2.是否改派,s2.物流方式,
						s2.总订单,
						concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
						concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
						concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
						concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
						concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
						concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			            null 序号
				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.物流方式,'总计') as 物流方式,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.物流方式 as 物流方式,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                          ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
						    ) s1
						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
					   	    with rollup
					    ) s2
                ORDER BY    FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-台湾','火凤凰-台湾','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`物流方式`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)
        # 物流-上月
        sql4 = '''SELECT 币种,团队 家族,年月,是否改派,物流方式,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as '总计签收(金额)',累计占比, @rownum:=@rownum+1 AS 序号
		        FROM ( SELECT s2.币种,
        							s2.团队,
        							s2.年月,
        							s2.是否改派,
        							s2.物流方式,
        							s2.总订单,
        							concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        							concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        							concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        							concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        							concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
        							concat(ROUND(IF(s2.物流方式 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM ( SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.物流方式,'总计') as 物流方式,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.物流方式 as 物流方式,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN (SELECT 币种,团队,年月,物流方式,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                ) cx2 ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        						    ) s1
        						    GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
        					   	    with rollup
        					    ) s2
                        ) s
                        ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        					    FIELD(s.`团队`,'神龙家族-台湾','火凤凰-台湾','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        					    FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        					    FIELD(s.`是否改派`,'直发','改派','总计'),
        					    FIELD(s.`物流方式`,'总计'),
        					    s.总订单 DESC;'''.format(team, day_last)
        df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        listT.append(df4)

        # 品类
        sql5 = '''SELECT s2.币种,
								s2.团队 家族,
								s2.年月,
								s2.是否改派,
								s2.父级分类,
								s2.总订单,
								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') '总计签收(金额)',
                                concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比,
			                    null 序号
				 FROM (
                        SELECT  IFNULL(s1.币种,'总计') as 币种,
                                IFNULL(s1.团队,'总计') as 团队,
                                IFNULL(s1.年月,'总计') as 年月,
                                IFNULL(s1.是否改派,'总计') as 是否改派,
                                IFNULL(s1.父级分类,'总计') as 父级分类,
								SUM(s1.签收) 签收,
								SUM(s1.完成) 完成,
								SUM(s1.总订单) 总订单,
								SUM(s1.总订单量) 总订单量,
								SUM(s1.签收金额) 签收金额,
								SUM(s1.总计金额) 总计金额
                        FROM (	 SELECT cx.币种 as 币种,
								        cx.团队 as 团队,
								        cx.年月 as 年月,
								        cx.是否改派 as 是否改派,
								        cx.父级分类 as 父级分类,
                                        SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                        count(订单编号) as 总订单,
                                        总订单量,
                                        @i:=0,
										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
										SUM(`价格RMB`) as 总计金额
                                FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                            GROUP BY dg.币种,dg.团队,dg.年月
                                            ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
							) s1
						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
					   	with rollup
				 ) s2
				 ORDER BY	FIELD(s2.`币种`,'台湾','香港','总计'),
							FIELD(s2.`团队`,'神龙家族-台湾','火凤凰-台湾','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
							FIELD(s2.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
							FIELD(s2.`是否改派`,'直发','改派','总计'),
							FIELD(s2.`父级分类`,'总计'),
							s2.总订单 DESC;'''.format(team, day_yesterday)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)
        # 品类-上月
        sql5 = '''SELECT 币种,团队 家族,年月,是否改派,父级分类,总订单,完成签收,总计签收,完成占比,单量占比,总计签收金额 as `总计签收(金额)`,累计占比, @rownum:=@rownum+1 AS 序号
		        FROM (SELECT s2.币种,
        								s2.团队,
        								s2.年月,
        								s2.是否改派,
        								s2.父级分类,
        								s2.总订单,
        								concat(ROUND(s2.签收 / s2.完成 * 100,2),'%') 完成签收,
        								concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') 总计签收,
        								concat(ROUND(s2.完成 / s2.总订单 * 100,2),'%') 完成占比,
        								concat(ROUND(s2.总订单 / s2.总订单量 * 100,2),'%') 单量占比,
        								concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') 总计签收金额,
                                        concat(ROUND(IF(s2.父级分类 like "总计", IF(@i > 1 ,@i - SUBSTRING_INDEX(@i,'.',1), @i), IF((@i:=@i + 总订单 / 总订单量) >1, @i - SUBSTRING_INDEX(@i,'.',1),  @i)) * 100,2),'%')  as 累计占比
        				 FROM (
                                SELECT  IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.团队,'总计') as 团队,
                                        IFNULL(s1.年月,'总计') as 年月,
                                        IFNULL(s1.是否改派,'总计') as 是否改派,
                                        IFNULL(s1.父级分类,'总计') as 父级分类,
        								SUM(s1.签收) 签收,
        								SUM(s1.完成) 完成,
        								SUM(s1.总订单) 总订单,
        								SUM(s1.总订单量) 总订单量,
        								SUM(s1.签收金额) 签收金额,
        								SUM(s1.总计金额) 总计金额
                                FROM (	 SELECT cx.币种 as 币种,
        								        cx.团队 as 团队,
        								        cx.年月 as 年月,
        								        cx.是否改派 as 是否改派,
        								        cx.父级分类 as 父级分类,
                                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成,
                                                count(订单编号) as 总订单,
                                                总订单量,
                                                @i:=0,
        										SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
        										SUM(`价格RMB`) as 总计金额
                                        FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                                        LEFT JOIN  (SELECT 币种,团队,年月,父级分类,count(订单编号) as 总订单量
                                                    FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') dg
                                                    GROUP BY dg.币种,dg.团队,dg.年月
                                                    ) cx2  ON cx.币种 = cx2.币种 AND  cx.团队 = cx2.团队 AND  cx.年月 = cx2.年月
                                        GROUP BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派`, cx.`父级分类`
                                        ORDER BY cx.`币种`,cx.`团队`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
        							) s1
        						GROUP BY s1.`币种`,s1.`团队`, s1.`年月`, s1.`是否改派`, s1.`父级分类`
        					   	with rollup
        				) s2 
        		) s
                ORDER BY FIELD(s.`币种`,'台湾','香港','总计'),
        				FIELD(s.`团队`,'神龙家族-台湾','火凤凰-台湾','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
        				FIELD(s.`年月`, DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'总计'),
        				FIELD(s.`是否改派`,'直发','改派','总计'),
        				FIELD(s.`父级分类`,'总计'),
        				s.总订单 DESC;'''.format(team, day_last)
        df5 = pd.read_sql_query(sql=sql5, con=self.engine1)
        listT.append(df5)

        # 产品
        sql6 = '''SELECT * 
				    FROM ( SELECT   IFNULL( cx.`币种`,'总计') as 币种,
                                    IFNULL( cx.`团队`,'总计') as 家族,
                                    IFNULL( cx.`年月`,'总计') as 年月,
                                    IFNULL( cx.`产品id`,'总计') as 产品id,
                                    cx.`产品名称`,
							        cx.`父级分类`,
                                    count(订单编号) as 总订单,
                                    SUM(IF(最终状态 = "已签收",1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) as 完成签收,
                                    SUM(IF(最终状态 = "已签收",1,0)) /  count(订单编号) as 总计签收,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件"),1,0)) / count(订单编号) as 完成占比,
                                    count(订单编号) /总订单2 单量占比,
                                    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)),0) as 直发完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '直发',1,0)) / SUM(IF(是否改派 = '直发',1,0)),0) as 直发完成占比,
                                    IFNULL(SUM(IF(是否改派 = '直发',1,0))  / 直发总订单2,0) as 直发单量占比,
                                    SUM(IF(是否改派 = '改派',1,0)) as 改派总订单,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)),0) as 改派完成签收,
                                    IFNULL(SUM(IF(最终状态 = "已签收" AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派总计签收,
                                    IFNULL(SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔", "自发头程丢件") AND 是否改派 = '改派',1,0)) / SUM(IF(是否改派 = '改派',1,0)),0) as 改派完成占比,
                                    IFNULL(SUM(IF(是否改派 = '改派',1,0)) / 改派总订单2,0) 改派单量占比
                            FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx
                            LEFT JOIN  (SELECT 币种,团队,年月,count(订单编号) as 总订单2 , 
											    SUM(IF(是否改派 = '直发',1,0)) as 直发总订单2 , 
												SUM(IF(是否改派 = '改派',1,0)) as 改派总订单2 
										FROM (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') da GROUP BY da.币种,da.团队,da.年月
									) cx2  ON cx.币种 = cx2.币种 AND cx.团队 = cx2.团队 AND cx.年月 = cx2.年月
                            GROUP BY cx.币种,cx.团队,cx.年月,`产品id`
	                        with rollup
					) s1
	                ORDER BY	FIELD(s1.`币种`,'台湾','香港','总计'),
								FIELD(s1.`团队`,'神龙家族-台湾','火凤凰-台湾','红杉家族-港澳台','红杉家族-港澳台2','金狮-港澳台','总计'),
								FIELD(s1.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
								总订单 DESC;'''.format(team, day_yesterday)
        df6 = pd.read_sql_query(sql=sql6, con=self.engine1)
        listT.append(df6)

        # 产品明细-台湾
        sql7 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
			            concat(ROUND(速派签收量 / 速派完成量 * 100,2),'%')  AS 速派完成签收,
			            concat(ROUND(速派签收量 / 速派单量 * 100,2),'%')  AS 速派总计签收,
			            concat(ROUND(速派完成量 / 速派单量 * 100,2),'%')  AS 速派完成占比,
			            concat(ROUND(速派单量 / 订单量 * 100,2),'%')  AS 速派单量占比,
			            concat(ROUND(711签收量 / 711完成量 * 100,2),'%')  AS 711完成签收,
			            concat(ROUND(711签收量 / 711单量 * 100,2),'%')  AS 711总计签收,
			            concat(ROUND(711完成量 / 711单量 * 100,2),'%')  AS 711完成占比,
			            concat(ROUND(711单量 / 订单量 * 100,2),'%')  AS 711单量占比,
			            concat(ROUND(天马签收量 / 天马完成量 * 100,2),'%')  AS 天马完成签收,
			            concat(ROUND(天马签收量 / 天马单量 * 100,2),'%')  AS 天马总计签收,
			            concat(ROUND(天马完成量 / 天马单量 * 100,2),'%')  AS 天马完成占比,
			            concat(ROUND(天马单量 / 订单量 * 100,2),'%')  AS 天马单量占比,
			            concat(ROUND(易速配签收量 / 易速配完成量 * 100,2),'%')  AS 易速配完成签收,
			            concat(ROUND(易速配签收量 / 易速配单量 * 100,2),'%')  AS 易速配总计签收,
			            concat(ROUND(易速配完成量 / 易速配单量 * 100,2),'%')  AS 易速配完成占比,
			            concat(ROUND(易速配单量 / 订单量 * 100,2),'%')  AS 易速配单量占比,
			            concat(ROUND(森鸿签收量 / 森鸿完成量 * 100,2),'%')  AS 森鸿完成签收,
			            concat(ROUND(森鸿签收量 / 森鸿单量 * 100,2),'%')  AS 森鸿总计签收,
			            concat(ROUND(森鸿完成量 / 森鸿单量 * 100,2),'%')  AS 森鸿完成占比,
			            concat(ROUND(森鸿单量 / 订单量 * 100,2),'%')  AS 森鸿单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
				            SUM(速派单量) 速派单量,  SUM(速派签收量) 速派签收量,  SUM(速派完成量) 速派完成量,
				            SUM(711单量) 711单量,  SUM(711签收量) 711签收量,  SUM(711完成量) 711完成量,
				            SUM(天马单量) 天马单量,  SUM(天马签收量) 天马签收量,  SUM(天马完成量) 天马完成量,
				            SUM(易速配单量) 易速配单量,  SUM(易速配签收量) 易速配签收量,  SUM(易速配完成量) 易速配完成量,
				            SUM(森鸿单量) 森鸿单量,  SUM(森鸿签收量) 森鸿签收量,  SUM(森鸿完成量) 森鸿完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派单量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派签收量,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  速派完成量,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS '711单量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as  '711签收量',
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  '711完成量',
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马单量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as  天马签收量,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  天马完成量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配单量,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as  易速配签收量,
							    SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配完成量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿单量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as  森鸿签收量,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '台湾'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-台湾', '火凤凰-台湾', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df7 = pd.read_sql_query(sql=sql7, con=self.engine1)
        listT.append(df7)
        # 产品明细-香港
        sql8 = '''SELECT 币种,团队 家族,年月,产品id,产品名称,父级分类,订单量,
			            concat(ROUND(IF(SUBSTRING_INDEX(订单量 / 品类总订单,'.',1) > 1 ,1,订单量 / 品类总订单 ) * 100,2),'%')  AS 订单品类占比,
						concat(ROUND(立邦签收量 / 立邦完成量 * 100,2),'%')  AS 立邦完成签收,
						concat(ROUND(立邦签收量 / 立邦单量 * 100,2),'%')  AS 立邦总计签收,
						concat(ROUND(立邦完成量 / 立邦单量 * 100,2),'%')  AS 立邦完成占比,
						concat(ROUND(立邦单量 / 订单量 * 100,2),'%')  AS 立邦单量占比,
						concat(ROUND(森鸿SF签收量 / 森鸿SF完成量 * 100,2),'%')  AS 森鸿SF完成签收,
						concat(ROUND(森鸿SF签收量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF总计签收,
						concat(ROUND(森鸿SF完成量 / 森鸿SF单量 * 100,2),'%')  AS 森鸿SF完成占比,
						concat(ROUND(森鸿SF单量 / 订单量 * 100,2),'%')  AS 森鸿SF单量占比,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH完成量 * 100,2),'%')  AS 森鸿SH完成签收,
					    concat(ROUND(森鸿SH签收量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH总计签收,
					    concat(ROUND(森鸿SH完成量 / 森鸿SH单量 * 100,2),'%')  AS 森鸿SH完成占比,
					    concat(ROUND(森鸿SH单量 / 订单量 * 100,2),'%')  AS 森鸿SH单量占比,
					    concat(ROUND(易速配SF签收量 / 易速配SF完成量 * 100,2),'%')  AS 易速配SF完成签收,
					    concat(ROUND(易速配SF签收量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF总计签收,
					    concat(ROUND(易速配SF完成量 / 易速配SF单量 * 100,2),'%')  AS 易速配SF完成占比,
					    concat(ROUND(易速配SF单量 / 订单量 * 100,2),'%')  AS 易速配SF单量占比,
					    concat(ROUND(易速配YC签收量 / 易速配YC完成量 * 100,2),'%')  AS 易速配YC完成签收,
					    concat(ROUND(易速配YC签收量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC总计签收,
					    concat(ROUND(易速配YC完成量 / 易速配YC单量 * 100,2),'%')  AS 易速配YC完成占比,
					    concat(ROUND(易速配YC单量 / 订单量 * 100,2),'%')  AS 易速配YC单量占比
                FROM (SELECT IFNULL(s1.`币种`, '总计') AS 币种,
							IFNULL(s1.`团队`, '总计') AS 团队,
				            IFNULL(s1.`年月`, '总计') AS 年月,
				            IFNULL(s1.`产品id`, '总计') AS 产品id, 
							IFNULL(s1.`产品名称`, '总计') AS 产品名称,
							IFNULL(s1.`父级分类`, '总计') AS 父级分类,
				            SUM(订单量) 订单量, 
							品类总订单,
							SUM(立邦单量) 立邦单量,  SUM(立邦签收量) 立邦签收量,  SUM(立邦完成量) 立邦完成量,
				            SUM(森鸿SF单量) 森鸿SF单量,  SUM(森鸿SF签收量) 森鸿SF签收量,  SUM(森鸿SF完成量) 森鸿SF完成量,
				            SUM(森鸿SH单量) 森鸿SH单量,  SUM(森鸿SH签收量) 森鸿SH签收量,  SUM(森鸿SH完成量) 森鸿SH完成量,					
				            SUM(易速配SF单量) 易速配SF单量,  SUM(易速配SF签收量) 易速配SF签收量,  SUM(易速配SF完成量) 易速配SF完成量,				
				            SUM(易速配YC单量) 易速配YC单量,  SUM(易速配YC签收量) 易速配YC签收量,  SUM(易速配YC完成量) 易速配YC完成量
                    FROM ( SELECT cx.`币种`, cx.`团队`, cx.`年月`, cx.`产品id`,  cx.`产品名称`,	cx.`父级分类`, 
								count(订单编号) AS 订单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦单量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦签收量,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  立邦完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿SF单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SF签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SF完成量,

								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH单量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as  森鸿SH签收量,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  森鸿SH完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配SF单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as  易速配SF签收量,
							    SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配SF完成量,

								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" ,1,0)) AS 易速配YC单量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 = "已签收",1,0)) as  易速配YC签收量,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰YC" AND 最终状态 IN ( "已签收", "拒收", "已退货", "理赔", "自发头程丢件") ,1,0)) as  易速配YC完成量
	                        FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                        WHERE cx.`是否改派` = '直发'
	                        GROUP BY cx.币种,cx.团队,cx.年月,cx.`产品id` 
	                    ) s1
                    LEFT JOIN 
						(SELECT cx.`币种`,  cx.`团队`, cx.`年月`,  cx.`父级分类`,  count(订单编号) AS 品类总订单 
	                    FROM  (SELECT * FROM qsb_gat WHERE qsb_gat.`记录时间` = '{1}') cx 
	                    WHERE  cx.`是否改派` = '直发'
	                    GROUP BY	cx.币种,	cx.团队,	cx.年月,cx.`父级分类` 	
	                    ) s2 ON s1.`币种` = s2.`币种` AND s1.`团队` = s2.`团队` AND s1.`年月` = s2.`年月` AND s1.`父级分类` = s2.`父级分类` 	
                    GROUP BY	s1.币种,	s1.团队,	s1.年月,	s1.产品id
                    WITH ROLLUP 	
                ) s WHERE s.`币种` = '香港'
                ORDER BY FIELD( s.`币种`, '台湾', '香港', '总计' ),
	                    FIELD( s.`团队`, '神龙家族-台湾', '火凤凰-台湾', '红杉家族-港澳台', '红杉家族-港澳台2','金狮-港澳台','总计'),
	                    FIELD(s.`年月`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'),'总计'),
	                    s.`订单量` DESC;'''.format(team, day_yesterday)
        df8 = pd.read_sql_query(sql=sql8, con=self.engine1)
        listT.append(df8)

        print('正在写入excel…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        for wbbook in ['神龙', '火凤凰', '红杉', '金狮']:
            file_path = 'F:\\输出文件\\{} {}-签收率.xlsx'.format(today, wbbook)
            sheet_name = ['每日', '总表', '总表上月', '物流', '物流上月', '品类', '品类上月', '产品', '产品明细台湾', '产品明细香港']
            # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            # for i in range(len(listT)):
            #     listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            #     del book['Sheet1']
            # writer.save()
            # writer.close()
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for i in range(len(listT)):
                    listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            # print('正在运行' + wbbook + '表宏…………')
            # app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            # app.display_alerts = False
            # wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
            # wbsht1 = app.books.open(file_path)
            # wbsht.macro('py_sl_总运行')()
            # wbsht1.save()
            # wbsht1.close()
            # wbsht.close()
            # app.quit()
        print('----已写入excel ')


    # 拒收核实-查询需要的产品id
    def jushou(self):
        print('正在查询需核实订单…………')
        listT = []  # 查询sql的结果 存放池
        sql = '''SELECT *
                FROM (SELECT g.*,c.`家族`,c.`月份`,c.`拒收`,c.`总订单`,c.`退货率`,c.`拒收率`
			            FROM  需核实拒收_每日新增订单 g
			            LEFT JOIN (SELECT *
								 FROM(SELECT IFNULL(s1.家族, '合计') 家族, IFNULL(s1.地区, '合计') 地区, IFNULL(s1.月份, '合计') 月份,
											IFNULL(s1.产品id, '合计') 产品id,
											IFNULL(s1.产品名称, '合计') 产品名称,
											IFNULL(s1.父级分类, '合计') 父级分类,
											IFNULL(s1.二级分类, '合计') 二级分类,
											SUM(s1.已签收) as 已签收,
											SUM(s1.拒收) as 拒收,
											SUM(s1.已退货) as 已退货,
											SUM(s1.已完成) as 已完成,
						                    SUM(s1.总订单) as 总订单,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						                    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						                    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						                    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率
                                    FROM(SELECT IFNULL(cx.`家族`, '合计') 家族, IFNULL(cx.币种, '合计') 地区, IFNULL(cx.`年月`, '合计') 月份,
						                        IFNULL(cx.产品id, '合计') 产品id,
						                        IFNULL(cx.产品名称, '合计') 产品名称,
						                        IFNULL(cx.父级分类, '合计') 父级分类,
						                        IFNULL(cx.二级分类, '合计') 二级分类,
						                        COUNT(cx.`订单编号`) as 总订单,
						                        SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
						                        SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
						                        SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
						                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成		
		                                FROM (SELECT *,IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族
                                            FROM gat_zqsb cc 
					                        WHERE cc.年月 >=  DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m') AND cc.`币种` = '台湾' AND cc.`运单编号` is not null
		                                ) cx
                                        GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                    ) s1
                                    GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                    WITH ROLLUP 
                                ) s 
                                HAVING s.月份 != '合计' AND s.产品id != '合计' AND s.`拒收` >= '1'
                                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙香港','Line运营','金蝉家族','金蝉项目组','APP运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                                FIELD(s.`地区`,'台湾','香港','合计'),
                                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'合计'),
                                FIELD(s.`产品id`,'合计'),
                                s.拒收 DESC
			            ) c ON g.`团队` = c.`家族` AND EXTRACT(YEAR_MONTH FROM g.`下单时间`) = c.`月份` AND g.`产品Id` =c.`产品Id`
                ) s WHERE s.`家族` is not null;'''
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df)
        print('正在查询两月拒收订单…………')
        sql2 = '''SELECT * FROM 需核实拒收_获取最近两个月订单;'''
        df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
        listT.append(df2)
        print('正在查询两月拒收产品id…………')
        sql3 = '''SELECT *
								 FROM(SELECT IFNULL(s1.家族, '合计') 家族, IFNULL(s1.地区, '合计') 地区, IFNULL(s1.月份, '合计') 月份,
											IFNULL(s1.产品id, '合计') 产品id,
											IFNULL(s1.产品名称, '合计') 产品名称,
											IFNULL(s1.父级分类, '合计') 父级分类,
											IFNULL(s1.二级分类, '合计') 二级分类,
											SUM(s1.已签收) as 已签收,
											SUM(s1.拒收) as 拒收,
											SUM(s1.已退货) as 已退货,
											SUM(s1.已完成) as 已完成,
						                    SUM(s1.总订单) as 总订单,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						                    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						                    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						                    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						                    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率
                                    FROM(SELECT IFNULL(cx.`家族`, '合计') 家族, IFNULL(cx.币种, '合计') 地区, IFNULL(cx.`年月`, '合计') 月份,
						                        IFNULL(cx.产品id, '合计') 产品id,
						                        IFNULL(cx.产品名称, '合计') 产品名称,
						                        IFNULL(cx.父级分类, '合计') 父级分类,
						                        IFNULL(cx.二级分类, '合计') 二级分类,
						                        COUNT(cx.`订单编号`) as 总订单,
						                        SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
						                        SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
						                        SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
						                        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成		
		                                FROM (SELECT *,IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族
                                            FROM gat_zqsb cc 
					                        WHERE cc.年月 >=  DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m') AND cc.`币种` = '台湾' AND cc.`运单编号` is not null
		                                ) cx
                                        GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                    ) s1
                                    GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                    WITH ROLLUP 
                                ) s 
                                HAVING s.月份 != '合计' AND s.产品id != '合计' AND s.`拒收` >= '1'
                                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙香港','Line运营','金蝉家族','金蝉项目组','APP运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                                FIELD(s.`地区`,'台湾','香港','合计'),
                                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'),'合计'),
                                FIELD(s.`产品id`,'合计'),
                                s.拒收 DESC;'''
        df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
        listT.append(df3)
        print('正在查询需核实拒收_每日新增订单…………')
        sql4 = '''SELECT * FROM 需核实拒收_每日新增订单;'''
        df4 = pd.read_sql_query(sql=sql4, con=self.engine1)
        listT.append(df4)
        print('正在写入excel…………')
        today = datetime.date.today().strftime('%m.%d')
        file_path = 'F:\\输出文件\\{} 需核实拒收-每日数据源.xlsx'.format(today)
        if os.path.exists(file_path):  # 判断是否有需要的表格
            print("正在清除重复文件......")
            os.remove(file_path)
        sheet_name = ['查询', '两月拒收', '两月拒收产品id', '每日新增订单']
        # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # for i in range(len(listT)):
        #     listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for i in range(len(listT)):
                listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
        print('----已写入excel ')

    # 获取电话核实日报表 周报表
    def phone_report(self, handle, month_last, month_yesterday):
        today = datetime.date.today().strftime('%Y.%m.%d')
        match = {'gat': '港台'}
        week: datetime = datetime.datetime.now()
        if week.isoweekday() == 1 or handle == '手动':
            week_time1 = (datetime.datetime.now() - datetime.timedelta(days=7)).strftime('%m.%d')
            week_time2 = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%m.%d')
            listT = []  # 查询sql的结果 存放池
            print('正在获取 日报表 数据内容…………')
            sql = '''SELECT 日期31天, ss.*, ss1.*, ss2.*, ss3.*, ss4.* , ss5.*
                    FROM date
                    LEFT JOIN
                    (SELECT 日期 AS 系统问题, COUNT(订单编号) AS 问题订单,
                            SUM(IF(g.`系统订单状态` NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货'),1,0)) AS 正常出货, SUM(IF(g.`系统订单状态` = '已删除',1,0)) AS 删除订单, SUM(IF(g.`系统物流状态` = '已签收',1,0)) AS 实际签收
                        FROM gat_order_list g
                        WHERE (g.日期 BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY)) AND  (g.`问题时间` BETWEEN TIMESTAMP(DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY)) AND TIMESTAMP(CURDATE()))  AND g.`问题原因` IS NOT NULL AND g.币种 = '台湾'
--                            WHERE (g.日期 BETWEEN '2023-03-01' AND '2023-04-01') AND (g.`问题时间` BETWEEN TIMESTAMP('2023-03-01') AND TIMESTAMP('2023-04-01')) AND g.`问题原因` IS NOT NULL AND g.币种 = '台湾'
                        GROUP BY DATE(日期) 
                        ORDER BY DATE(日期)
                    ) ss ON  date.`日期31天` = EXTRACT(day FROM ss.`系统问题`)
                    LEFT JOIN 
                    (	SELECT ww.* ,物流问题总量, 约派送, 核实拒收, 再派签收, 未接听, 无效号码
                        FROM (SELECT 处理时间 AS 物流问题, COUNT(订单编号) AS 物流问题联系量
                                    FROM 物流问题件 cg
                                    WHERE cg.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) AND cg.币种 = '台币'  AND cg.问题类型 <> '订单压单（giikin内部专用）'
--                                     WHERE cg.`处理时间` BETWEEN '2023-03-01' AND '2023-03-31' AND cg.币种 = '台币' AND cg.问题类型 <> '订单压单（giikin内部专用）'
                                    GROUP BY 处理时间
                                    ORDER BY 处理时间
                        ) ww
                        LEFT JOIN 
                        (SELECT 处理时间 AS 物流问题, COUNT(订单编号) AS 物流问题总量, SUM(IF(ks.`处理结果` LIKE '%送货%' or ks.`处理结果` LIKE '%配送%' or ks.`处理结果` LIKE '%自取%',1,0)) AS 约派送, 
                                SUM(IF(ks.`处理结果` LIKE '%拒收%' ,1,0)) AS 核实拒收, SUM(IF((ks.`处理结果` LIKE '%送货%' or ks.`处理结果` LIKE '%配送%') AND ks.`系统物流状态` LIKE '已签收%',1,0)) AS 再派签收, 
                                SUM(IF(ks.`处理结果` LIKE '%无人接听%',1,0)) AS 未接听, SUM(IF(ks.`处理结果` LIKE '%无效号码%',1,0)) AS 无效号码
                            FROM (SELECT wt.*, g.`系统订单状态`, g.`系统物流状态`, g.`完结状态`
                                    FROM (SELECT * 
                                            FROM 物流问题件 
                                            WHERE id IN ( SELECT MAX(id) 
                                                            FROM 物流问题件 w 
                                                            WHERE w.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY)  AND w.问题类型 <> '订单压单（giikin内部专用）'
-- 																WHERE w.`处理时间` BETWEEN '2023-03-01' AND '2023-03-31'  AND w.问题类型 <> '订单压单（giikin内部专用）'
                                                            GROUP BY 订单编号) 
                                             ORDER BY id
                                    ) wt 
                                    LEFT JOIN gat_order_list g ON  wt.`订单编号` = g.`订单编号`
                                    WHERE wt.币种 = '台币'
                            ) ks
                            GROUP BY ks.处理时间
                            ORDER BY 处理时间
                        ) ww2  ON ww.`物流问题` = ww2.`物流问题`
                    ) ss1 ON  date.`日期31天` = EXTRACT(day FROM ss1.`物流问题`)
                    LEFT JOIN
                    ( SELECT cc.* ,客诉总量, 挽回单数, 未确认, 退款单数, 实际退款单数, 实际挽回单数
                        FROM (SELECT 处理时间 AS 物流客诉, COUNT(订单编号) AS 物流客诉联系量
                                    FROM 物流客诉件 cg
                                    WHERE cg.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) and cg.币种 = '台币'
--                                     WHERE cg.`处理时间` BETWEEN '2023-03-01' AND '2023-03-31' and cg.币种 = '台币'
                                    GROUP BY 处理时间
                                    ORDER BY 处理时间
                        ) cc 
                    LEFT JOIN
                        (SELECT 处理时间 AS 物流客诉, COUNT(订单编号) AS 客诉总量, SUM(IF(ks.`处理方案` LIKE '%不退款%' or ks.`处理方案` LIKE '%赠品%' or ks.`处理方案` LIKE '%补发%' or ks.`处理方案` LIKE '%换货%',1,0)) AS 挽回单数,
                                SUM(IF(ks.`处理结果` LIKE '%转语音%' or ks.`处理结果` LIKE '%空号%' or ks.`处理结果` LIKE '%挂断电话%' or ks.`处理结果` LIKE '%无人接听%',1,0)) AS 未确认,
                                 SUM(IF(ks.`处理方案` LIKE '%退款%' AND ks.`处理方案` NOT LIKE '%不%',1,0)) AS 退款单数, SUM(IF(ks.`完结状态` = '退款',1,0)) AS 实际退款单数, SUM(IF(ks.`完结状态` = '收款',1,0)) AS 实际挽回单数
                            FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`, g.`完结状态`
                                    FROM (SELECT * 
                                            FROM 物流客诉件 
                                            WHERE id IN (SELECT MAX(id) 
                                                            FROM 物流客诉件 w 
                                                            WHERE w.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) 
-- 																	WHERE w.`处理时间` BETWEEN '2023-03-01' AND '2023-03-31'
                                                            GROUP BY 订单编号) 
                                            ORDER BY id
                                        ) cg
                                    LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                    WHERE cg.币种 = '台币'
                            ) ks
                            GROUP BY ks.处理时间
                            ORDER BY 处理时间
                        ) cc2  ON cc.`物流客诉` = cc2.`物流客诉`
                    ) ss2 ON  date.`日期31天` = EXTRACT(day FROM ss2.`物流客诉`)
                    LEFT JOIN
                    (SELECT gg.* ,异常单量, 正常发货, 取消订单
                        FROM (SELECT cg.处理时间 AS 采购异常, COUNT(cg.订单编号) AS 采购异常联系量
                                FROM 采购异常 cg
                                LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                WHERE cg.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) AND g.是否改派 = '直发'
--                                     WHERE cg.`处理时间` BETWEEN '2023-03-01' AND '2023-03-31' AND g.是否改派 = '直发'
                                GROUP BY cg.处理时间
                                ORDER BY cg.处理时间
                        ) gg 
                        LEFT JOIN
                        (SELECT DATE(s.处理时间) AS 采购异常, COUNT(订单编号) AS 异常单量,  SUM(IF(s.`系统订单状态` NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货') AND s.处理结果 <> '跟进',1,0)) AS 正常发货,
                                SUM(IF(s.`系统订单状态` = '已删除',1,0)) AS 取消订单, SUM(IF(s.处理结果 = '跟进',1,0)) AS 跟进, SUM(IF(s.`反馈内容` NOT like '%取消%',1,0)) AS 正常发货22, SUM(IF(s.`反馈内容` like '%取消%',1,0)) AS 取消订单22
                            FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`, g.`币种`, g.`是否改派`
                                    FROM (SELECT * 
                                            FROM 采购异常 
                                            WHERE id IN (SELECT MAX(id) 
                                                            FROM 采购异常 w 
                                                            WHERE w.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) 
-- 																	WHERE w.`处理时间` BETWEEN '2023-03-01' AND '2023-03-31'
                                                            GROUP BY 订单编号) 
                                            ORDER BY id
                                        ) cg
                                    LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                    WHERE  g.是否改派 = '直发'
                            ) s
                            GROUP BY DATE(s.处理时间) 
                            ORDER BY DATE(s.处理时间) 
                        ) gg2 ON gg.`采购异常` = gg2.`采购异常`
                    ) ss3 ON  date.`日期31天` = EXTRACT(day FROM ss3.`采购异常`)
                    LEFT JOIN
                    (SELECT DATE(s.处理时间) AS 拒收问题件, COUNT(订单编号) AS '联系量（有结果）', SUM(IF(s.`再次克隆下单` IS NOT NULL,1,0)) AS 挽单量
                        FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`
                                FROM (SELECT * 
                                        FROM 拒收问题件 
                                        WHERE 联系方式 = '电话' AND 币种 = '台币' 
                                             AND id IN (SELECT MAX(id) 
                                                            FROM 拒收问题件 w 
                                                            WHERE w.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY)  AND w.处理人 IN ('蔡利英','杨嘉仪','张陈平','李晓青')
-- 																WHERE w.`处理时间` BETWEEN '2023-03-01' AND '2023-03-31' AND w.处理人 IN ('蔡利英','杨嘉仪','张陈平','李晓青')
                                                            GROUP BY 订单编号)	
                                        ORDER BY id
                                    ) cg
                                LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                WHERE g.币种 = '台湾' 
                        ) s
                        WHERE  s.核实原因 <> '未联系上客户'
                        GROUP BY DATE(s.处理时间) 
                        ORDER BY DATE(s.处理时间) 
                    ) ss4 ON  date.`日期31天` = EXTRACT(day FROM ss4.`拒收问题件`)
					LEFT JOIN 
					(	SELECT ww.* , 压单核实总量, 有效订单, 签收量, 删单量, 联系取消订单, 无人接听
                        FROM (SELECT 处理时间 AS 压单核实, COUNT(订单编号) AS 压单核实联系量
                                FROM 物流问题件 cg
                                WHERE cg.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY) AND cg.币种 = '台币'  AND cg.问题类型 = '订单压单（giikin内部专用）'
--                                     WHERE cg.`处理时间` BETWEEN '2023-03-01' AND '2023-03-31' AND cg.币种 = '台币' AND cg.问题类型 = '订单压单（giikin内部专用）'
                                GROUP BY 处理时间
                                ORDER BY 处理时间
                        ) ww
                        LEFT JOIN 
                        (SELECT 处理时间 AS 压单核实,  COUNT(订单编号) AS 压单核实总量, SUM(IF(ks.系统订单状态 IN ('已审核', '已转采购', '已发货', '已收货', '已完成', '已退货(销售)', '已退货(物流)', '已退货(不拆包物流)'),1,0)) AS 有效订单,
								SUM(IF(ks.系统物流状态 = '已签收',1,0)) AS 签收量, SUM(IF(ks.系统订单状态 IN ('未支付', '支付失败', '已删除'),1,0)) AS 删单量,
                                SUM(IF(ks.`处理结果` LIKE '%取消%' OR ks.`处理结果` LIKE '%无效号码%',1,0)) AS 联系取消订单, SUM(IF(ks.`处理结果` LIKE '%无人接听%',1,0)) AS 无人接听
                            FROM (SELECT wt.*, g.`系统订单状态`, g.`系统物流状态`, g.`完结状态`
                                    FROM (SELECT * 
                                                FROM 物流问题件 
                                                WHERE id IN ( SELECT MAX(id) 
                                                                FROM 物流问题件 w 
                                                                WHERE w.`处理时间` BETWEEN DATE_SUB(CURDATE(), INTERVAL DAY(CURDATE())-1 DAY) AND DATE_SUB(CURDATE(), INTERVAL 1 DAY)  AND w.问题类型 <> '订单压单（giikin内部专用）'
-- 																	WHERE w.`处理时间` BETWEEN '2023-03-01' AND '2023-03-31'  AND w.问题类型 = '订单压单（giikin内部专用）'
                                                                GROUP BY 订单编号) 
                                                ORDER BY id
                                    ) wt 
                                    LEFT JOIN gat_order_list g ON  wt.`订单编号` = g.`订单编号`
                                    WHERE wt.币种 = '台币'
                            ) ks
                            GROUP BY ks.处理时间
                            ORDER BY 处理时间
                        ) ww2  ON ww.`压单核实` = ww2.`压单核实`
                    ) ss5 ON date.`日期31天` = EXTRACT(day FROM ss5.`压单核实`)
                    GROUP BY 日期31天
                    ORDER BY 日期31天;'''.format()     # 港台查询函数导出
            df0 = pd.read_sql_query(sql=sql, con=self.engine1)
            listT.append(df0)
            print('正在获取 周报表 数据内容…………')
            sql = '''SELECT 日期31天,ss.问题订单,ss.正常出货,ss.删除订单,concat(ROUND(IFNULL(ss.删除订单/ss.问题订单,0) * 100,2),'%') as 取消占比,ss.实际签收, ss1.约派送,ss1.核实拒收 as 核实拒收原因,ss1.再派签收,ss2.挽回单数,ss2.未确认,ss2.退款单数,ss2.实际挽回单数,ss3.正常发货,ss3.取消订单,
                            ss5.客言需要, ss5.客言取消, ss5.未接听, ss4.`联系量（有结果）`,ss4.挽单量,  ss4.张联系量 AS '张陈平-联系量(有结果)',ss4.张挽单量 AS '张陈平-挽单量', ss4.蔡联系量 AS '蔡利英-联系量(有结果)',ss4.蔡挽单量 AS '蔡利英-挽单量', 
                            ss4.杨联系量 AS '杨嘉仪-联系量(有结果)',ss4.杨挽单量 AS '杨嘉仪-挽单量',   ss4.李联系量 AS '李晓青-联系量(有结果)',ss4.李挽单量 AS '李晓青-挽单量',
                            NULL 联系量,NULL 客户接听量,  NULL '张陈平-联系量',NULL '张陈平-客户接听量', NULL '蔡利英-联系量',NULL '蔡利英-客户接听量', NULL '杨嘉仪-联系量',NULL '杨嘉仪-客户接听量', NULL '李晓青-联系量',NULL '李晓青-客户接听量'
                    FROM date
                    LEFT JOIN
                    (SELECT 日期 AS 系统问题,COUNT(订单编号) AS 问题订单,
                            SUM(IF(g.`系统订单状态` NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货'),1,0)) AS 正常出货,
                            SUM(IF(g.`系统订单状态` = '已删除',1,0)) AS 删除订单, SUM(IF(g.`系统物流状态` = '已签收',1,0)) AS 实际签收
                        FROM gat_order_list g
                        WHERE (g.日期  BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0)) AND 
                            (g.`问题时间` BETWEEN TIMESTAMP(subdate(curdate(),date_format(curdate(),'%w')+6)) AND TIMESTAMP(subdate(curdate(),date_format(curdate(),'%w')-1))) 
                            AND g.`问题原因` IS NOT NULL AND g.币种 = '台湾'
                        GROUP BY DATE(日期) 
                    ) ss ON  date.`日期31天` = EXTRACT(day FROM ss.`系统问题`)
                    LEFT JOIN 
                    (SELECT ww.* ,物流问题总量, 约派送, 核实拒收, 再派签收, 未接听, 无效号码
                        FROM (SELECT 处理时间 AS 物流问题, COUNT(订单编号) AS 物流问题联系量
                                FROM 物流问题件 cg
                                WHERE cg.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) AND cg.币种 = '台币' AND cg.`问题类型` NOT IN ('订单压单（giikin内部专用）','订单追踪（giikin内部专用）')
                                GROUP BY 处理时间
                        ) ww
                        LEFT JOIN 
                        (SELECT 处理时间 AS 物流问题, COUNT(订单编号) AS 物流问题总量, SUM(IF(ks.`处理结果` LIKE '%送货%' or ks.`处理结果` LIKE '%配送%' or ks.`处理结果` LIKE '%自取%',1,0)) AS 约派送,
                                SUM(IF(ks.`处理结果` LIKE '%拒收%' OR ks.`处理结果` LIKE '%无人接听%' OR ks.`处理结果` LIKE '%无效号码%',1,0)) AS 核实拒收, SUM(IF((ks.`处理结果` LIKE '%送货%' or ks.`处理结果` LIKE '%配送%') AND ks.`系统物流状态` LIKE '已签收%',1,0)) AS 再派签收,
                                SUM(IF(ks.`处理结果` LIKE '%无人接听%',1,0)) AS 未接听,  SUM(IF(ks.`处理结果` LIKE '%无效号码%',1,0)) AS 无效号码
                            FROM (SELECT wt.*, g.`系统订单状态`, g.`系统物流状态`, g.`完结状态`
                                    FROM (SELECT * 
                                            FROM 物流问题件 
                                            WHERE id IN (SELECT MAX(id) FROM 物流问题件 w WHERE w.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) 
                                                                                            AND w.`问题类型` NOT IN ('订单压单（giikin内部专用）','订单追踪（giikin内部专用）')GROUP BY 订单编号) 
                                            ORDER BY id
                                    ) wt 
                                LEFT JOIN gat_order_list g ON  wt.`订单编号` = g.`订单编号`
                                WHERE wt.币种 = '台币'
                            ) ks
                            GROUP BY ks.处理时间
                        ) ww2  ON ww.`物流问题` = ww2.`物流问题`
                    ) ss1 ON  date.`日期31天` = EXTRACT(day FROM ss1.`物流问题`)
                    LEFT JOIN
                    ( SELECT cc.* ,客诉总量, 挽回单数, 未确认, 退款单数, 实际退款单数, 实际挽回单数
                        FROM (SELECT 处理时间 AS 物流客诉, COUNT(订单编号) AS 物流客诉联系量
                                FROM 物流客诉件 cg
                                WHERE cg.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) and cg.币种 = '台币'
                                GROUP BY 处理时间
                        ) cc 
                        LEFT JOIN
                        (SELECT 处理时间 AS 物流客诉, COUNT(订单编号) AS 客诉总量, SUM(IF(ks.`处理方案` LIKE '%不退款%' or ks.`处理方案` LIKE '%赠品%' or ks.`处理方案` LIKE '%补发%' or ks.`处理方案` LIKE '%换货%',1,0)) AS 挽回单数,
                                SUM(IF(ks.`处理结果` LIKE '%转语音%' or ks.`处理结果` LIKE '%空号%' or ks.`处理结果` LIKE '%挂断电话%' or ks.`处理结果` LIKE '%无人接听%',1,0)) AS 未确认,
                                SUM(IF(ks.`处理方案` LIKE '%退款%' AND ks.`处理方案` NOT LIKE '%不%',1,0)) AS 退款单数, SUM(IF(ks.`完结状态` = '退款',1,0)) AS 实际退款单数, SUM(IF(ks.`完结状态` = '收款',1,0)) AS 实际挽回单数
                            FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`, g.`完结状态`
                                    FROM (SELECT * 
                                            FROM 物流客诉件 
                                            WHERE id IN (SELECT MAX(id) FROM 物流客诉件 w WHERE w.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) GROUP BY 订单编号) 
                                            ORDER BY id
                                    ) cg
                                    LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                    WHERE cg.币种 = '台币'
                            ) ks
                            GROUP BY ks.处理时间
                        ) cc2  ON cc.`物流客诉` = cc2.`物流客诉`
                    ) ss2 ON  date.`日期31天` = EXTRACT(day FROM ss2.`物流客诉`)
                    LEFT JOIN
                    (SELECT gg.* ,异常单量, 正常发货, 取消订单
                        FROM (SELECT 处理时间 AS 采购异常, COUNT(cg.订单编号) AS 采购异常联系量
                                FROM 采购异常 cg
                                LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                WHERE cg.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) AND g.是否改派 = '直发'
                                GROUP BY 处理时间
                        ) gg 
                        LEFT JOIN
                        (SELECT DATE(s.处理时间) AS 采购异常, COUNT(订单编号) AS 异常单量,
                                SUM(IF(s.`系统订单状态` NOT IN ('未支付','待审核','已取消','截单','支付失败','已删除','问题订单','问题订单审核','待发货') AND s.处理结果 <> '跟进',1,0)) AS 正常发货,
                                SUM(IF(s.`系统订单状态` = '已删除',1,0)) AS 取消订单,
                                SUM(IF(s.处理结果 = '跟进',1,0)) AS 跟进,
                                SUM(IF(s.`反馈内容` NOT like '%取消%',1,0)) AS 正常发货22,
                                SUM(IF(s.`反馈内容` like '%取消%',1,0)) AS 取消订单22
                            FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`, g.`币种`
                                    FROM (SELECT * 
                                            FROM 采购异常 
                                            WHERE id IN (SELECT MAX(id) FROM 采购异常 w WHERE w.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) GROUP BY 订单编号) 
                                            ORDER BY id
                                    ) cg
                                     LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                    WHERE g.是否改派 = '直发'
                            ) s
                            GROUP BY DATE(s.处理时间) 
                        ) gg2 ON gg.`采购异常` = gg2.`采购异常`
                    ) ss3 ON  date.`日期31天` = EXTRACT(day FROM ss3.`采购异常`)
                    LEFT JOIN
                    (SELECT DATE(s.处理时间) AS 拒收问题件, COUNT(订单编号) AS '联系量（有结果）', SUM(IF(s.`再次克隆下单` IS NOT NULL,1,0)) AS 挽单量, SUM(IF(处理人='张陈平',1,0)) AS 张联系量, SUM(IF(处理人='张陈平' AND s.`再次克隆下单` IS NOT NULL,1,0)) AS 张挽单量,
                            SUM(IF(处理人='蔡利英',1,0)) AS 蔡联系量, SUM(IF(处理人='蔡利英' AND s.`再次克隆下单` IS NOT NULL,1,0)) AS 蔡挽单量,  SUM(IF(处理人='杨嘉仪',1,0)) AS 杨联系量, SUM(IF(处理人='杨嘉仪' AND s.`再次克隆下单` IS NOT NULL,1,0)) AS 杨挽单量,
                            SUM(IF(处理人='李晓青',1,0)) AS 李联系量, SUM(IF(处理人='李晓青' AND s.`再次克隆下单` IS NOT NULL,1,0)) AS 李挽单量
                        FROM (SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`
                                FROM (SELECT * 
                                        FROM 拒收问题件 
                                        WHERE 联系方式 = '电话' AND 币种 = '台币' AND id IN (SELECT MAX(id) FROM 拒收问题件 w WHERE w.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) GROUP BY 订单编号) 
                                        ORDER BY id
                                    ) cg
                                LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                                WHERE g.币种 = '台湾' 
                        ) s
                        WHERE  s.核实原因 <> '未联系上客户'
                        GROUP BY DATE(s.处理时间) 
                    ) ss4 ON  date.`日期31天` = EXTRACT(day FROM ss4.`拒收问题件`)
					LEFT JOIN
					(SELECT ww.* ,压单核实总量, 客言需要, 客言取消, 未接听
                        FROM (SELECT 处理时间 AS 压单核实, COUNT(订单编号) AS 压单核实联系量
                                FROM 物流问题件 cg
                                WHERE cg.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) 
																	AND cg.币种 = '台币' AND cg.`问题类型` IN ('订单压单（giikin内部专用）','订单追踪（giikin内部专用）')
                                GROUP BY 处理时间
                        ) ww
                        LEFT JOIN 
                        (SELECT 处理时间 AS 压单核实, COUNT(订单编号) AS 压单核实总量, SUM(IF(ks.`处理结果` LIKE '%需要%' AND ks.`处理结果` NOT LIKE '%不%' or ks.`处理结果` LIKE '%等货%' or ks.`处理结果` LIKE '%修改%',1,0)) AS 客言需要,
                                SUM(IF(ks.`处理结果` LIKE '%取消%' or ks.`处理结果` LIKE '%无订购%',1,0)) AS 客言取消, SUM(IF(ks.`处理结果` LIKE '%无人接听%' or ks.`处理结果` LIKE '%无人%接听%' or ks.`处理结果` LIKE '%无效%' or ks.`处理结果` LIKE '%挂断%' or ks.`处理结果` LIKE '%未说话%' or ks.`处理结果` LIKE '%无效%',1,0)) AS 未接听
                        FROM (SELECT wt.*, g.`系统订单状态`, g.`系统物流状态`, g.`完结状态`
                                FROM (SELECT * 
                                        FROM 物流问题件 
                                        WHERE id IN (SELECT MAX(id) FROM 物流问题件 w WHERE w.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) 
																						AND w.`问题类型` IN ('订单压单（giikin内部专用）','订单追踪（giikin内部专用）')GROUP BY 订单编号) 
                                        ORDER BY id
                                ) wt 
                            LEFT JOIN gat_order_list g ON  wt.`订单编号` = g.`订单编号`
                            WHERE wt.币种 = '台币'
                        ) ks
                        GROUP BY ks.处理时间
                        ) ww2  ON ww.`压单核实` = ww2.`压单核实`
                    ) ss5 on date.`日期31天` = EXTRACT(day FROM ss5.`压单核实`)
                    WHERE ss.系统问题 IS NOT NULL
                    GROUP BY 日期31天
                    ORDER BY 系统问题;'''.format()  # 港台查询函数导出
            df1 = pd.read_sql_query(sql=sql, con=self.engine1)
            listT.append(df1)

            print('正在获取 周报表 拒收明细数据内容…………')
            sql = '''SELECT cg.*, g.`系统订单状态`, g.`系统物流状态`
                    FROM (SELECT * 
                            FROM 拒收问题件 
                            WHERE 联系方式 = '电话' AND 币种 = '台币' AND id IN (SELECT MAX(id) FROM 拒收问题件 w WHERE w.`处理时间` BETWEEN subdate(curdate(),date_format(curdate(),'%w')+6) AND subdate(curdate(),date_format(curdate(),'%w')-0) GROUP BY 订单编号) 
                            ORDER BY id
                        ) cg
                    LEFT JOIN gat_order_list g ON  cg.`订单编号` = g.`订单编号`
                    WHERE g.币种 = '台湾' ;'''
            df12 = pd.read_sql_query(sql=sql, con=self.engine1)
            listT.append(df12)

            print('正在写入excel…………')
            today = datetime.date.today().strftime('%Y.%m.%d')
            file_path = r'''F:\\输出文件\\台湾电话核实({0}-{1})周报表{2}.xlsx'''.format(week_time1, week_time2, today)
            sheet_name = ['日报表', '周报表', '拒收明细']
            # df0 = pd.DataFrame([])                                          # 创建空的dataframe数据框
            # df0.to_excel(file_path, index=False)                            # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            # writer = pd.ExcelWriter(file_path, engine='openpyxl')           # 初始化写入对象
            # book = load_workbook(file_path)                                 # 可以向不同的sheet写入数据（对现有工作表的追加）
            # writer.book = book                                              # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            # for i in range(len(listT)):
            #     listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            # if 'Sheet1' in book.sheetnames:                                 # 删除新建文档时的第一个工作表
            #     del book['Sheet1']
            # writer.save()
            # writer.close()
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for i in range(len(listT)):
                    listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            try:
                print('正在运行 日报表、周报表 宏…………')
                app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
                app.display_alerts = False
                wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
                wbsht1 = app.books.open(file_path)
                wbsht.macro('zl_gat_report_new2.电话核实日报表_周报表')()
                wbsht1.save()
                wbsht1.close()
                wbsht.close()
                app.quit()
            except Exception as e:
                print('运行失败：', str(Exception) + str(e))
            print('----已写入excel ')

        if week.isoweekday() == 2 or week.isoweekday() == 4 or handle == '手动':
            month = datetime.datetime.now().strftime('%Y%m')
            time_bengin = ((datetime.datetime.now() - relativedelta(months=1)) - datetime.timedelta(days=10)).strftime('%Y-%m-%d')
            time_end = (datetime.datetime.now() - datetime.timedelta(days=10)).strftime('%Y-%m-%d')
            listT = []  # 查询sql的结果 存放池
            print("正在获取 物流签收率（产品前50单）" + time_bengin + "-" + time_end + " 数据内容…………")
            sql = '''SELECT IFNULL(s.家族,'合计') as 家族, 
                            IFNULL(s.币种,'合计') as 币种, 
                            '{0}' as 月份, 
                            IFNULL(s.产品ID,'合计') as 产品ID, 
                            IFNULL(s.产品名称,'合计') as 产品名称,
                            IFNULL(s.物流方式,'合计') as 物流方式,
                            s.总单量,
                            SUM(s.签收) as 签收, 
                            SUM(s.拒收) as 拒收, 
                            SUM(s.已退货) as 已退货,  
                            SUM(s.已完成) as 已完成, 
                            SUM(s.总订单) as 总订单,
                            concat(ROUND(IFNULL(SUM(s.签收) / SUM(s.已完成),NULL) * 100,2),'%') as 完成签收,
                            concat(ROUND(IFNULL(SUM(s.签收) / SUM(s.总订单),NULL) * 100,2),'%') as 总计签收,
                            concat(ROUND(IFNULL(SUM(s.已完成) / SUM(s.总订单),NULL) * 100,2),'%') as 完成占比,
                            concat(ROUND(IFNULL(SUM(s.已退货) / SUM(s.总订单),NULL) * 100,2),'%') as 退货率,
                            concat(ROUND(IFNULL(SUM(s.总订单) / s.总单量,NULL) * 100,2),'%') as 订单占比
					FROM ( SELECT ss1.*,ss2.物流方式, ss2.总订单, ss2.签收, ss2.拒收, ss2.已退货, ss2.已完成
						    FROM ( SELECT s1.币种,s1.家族,s1.年月,s1.产品ID,s1.产品名称, SUM(s1.总订单) as 总单量
                                    FROM ( SELECT cx.币种, cx.所属团队 as 家族, cx.年月, cx.产品ID, cx.产品名称, count(订单编号) as 总订单, IF(count(订单编号) >=100 ,"头部产品",IF(count(订单编号) < 50 ,"尾部产品","中间产品")) 产品类型
                                            FROM gat_zqsb cx
                                            WHERE cx.是否改派 = '直发' AND cx.物流方式 <> '台湾-速派-711超商' AND cx.`运单编号` is not null AND cx.日期 >= '{1}' AND cx.日期 <= '{2}'
                                            GROUP BY cx.`币种`,cx.`所属团队`, cx.`产品ID`
                                    ) s1
									WHERE s1.`产品类型` IN ("头部产品","中间产品")
                                    GROUP BY s1.`家族`,s1.`币种`,  s1.`产品ID`
                            ) ss1
                            LEFT JOIN 
                            ( SELECT cx.币种,cx.所属团队 as 家族,cx.年月,cx.产品ID, cx.物流方式, 
                                    count(订单编号) as 总订单, 
                                    SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                    SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                    SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成
                                FROM gat_zqsb cx
                                WHERE cx.是否改派 = '直发' AND cx.物流方式 <> '台湾-速派-711超商' AND cx.`运单编号` is not null AND cx.日期 >= '{1}' AND cx.日期 <= '{2}'
                                GROUP BY cx.`币种`,cx.`所属团队`,  cx.`产品ID`, cx.`物流方式`
                            ) ss2 ON ss1.币种 = ss2.币种 AND ss1.家族 = ss2.家族 AND ss1.产品ID = ss2.产品ID
				    ) s
				    GROUP BY s.家族, s.币种, s.产品ID, s.物流方式
				    WITH ROLLUP
				    HAVING s.币种 <> '合计';'''.format(month, time_bengin, time_end, 'team_name', self.team_name2)  # 港台查询函数导出
            df5 = pd.read_sql_query(sql=sql, con=self.engine1)
            listT.append(df5)

            print('正在写入excel…………')
            today = datetime.date.today().strftime('%Y.%m.%d')
            file_path = 'F:\\输出文件\\{} 物流签收率-头部产品.xlsx'.format(today)
            sheet_name = ['查询']
            # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            # for i in range(len(listT)):
            #     listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            #     del book['Sheet1']
            # writer.save()
            # writer.close()
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for i in range(len(listT)):
                    listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            new_path = "F:\\神龙签收率\\" + (datetime.datetime.now()).strftime('%m.%d') + '\\物流签收率\\{} {} 物流签收率-头部产品.xlsx'.format(today,match[team])
            shutil.copyfile(file_path, new_path)     # copy到指定位置
            try:
                print('正在运行 物流头部产品签收率 宏…………')
                app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
                app.display_alerts = False
                wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
                wbsht1 = app.books.open(file_path)
                wbsht.macro('zl_gat_report_new2.物流头程产品签收率_月')()
                wbsht1.save()
                wbsht1.close()
                wbsht.close()
                app.quit()
            except Exception as e:
                print('运行失败：', str(Exception) + str(e))
            print('----已写入excel ')

        # if week.isoweekday() == 3 or handle == '手动':
        if week.isoweekday() != 0 or handle == '手动':
            listT = []  # 查询sql的结果 存放池
            print("正在获取 在线签收率" + month_last + "-" + month_yesterday + " 数据内容…………")
            sql = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.物流方式,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,					
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / 总量,0) * 100,2),'%') as 在线占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                    concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                        concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                        concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)',总量
                FROM ( SELECT IFNULL(s1.币种,'合计') as 币种,
                            IFNULL(s1.家族,'合计') as 家族,
                            IFNULL(s1.年月,'合计') as 年月,
                            IFNULL(s1.是否改派,'合计') as 是否改派,
                            IFNULL(s1.物流方式,'合计') as 物流方式,
                            SUM(s1.签收) as 签收,
                            SUM(s1.拒收) as 拒收,
                            SUM(s1.在途) as 在途,
                            SUM(s1.未发货) as 未发货,
                            SUM(s1.未上线) as 未上线,
                            SUM(s1.已退货) as 已退货,
                            SUM(s1.理赔) as 理赔,
                            SUM(s1.自发头程丢件) as 自发头程丢件,
                            SUM(s1.已发货) as 已发货,
                            SUM(s1.已完成) as 已完成,
                            SUM(s1.总订单) as 总订单,
                            SUM(s1.签收金额) as 签收金额,
                            SUM(s1.退货金额) as 退货金额,
                            SUM(s1.完成金额) as 完成金额,
                            SUM(s1.发货金额) as 发货金额,
                            SUM(s1.总计金额) as 总计金额
                    FROM (SELECT cx.币种 as 币种,cx.所属团队 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.物流渠道 as 物流方式,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM gat_zqsb cx
							where cx.支付类型 = '在线付款' and cx.`运单编号` is not null AND cx.日期 >= '{0}' AND cx.日期 <= '{1}'
                            GROUP BY cx.`币种`,cx.`所属团队`, cx.`年月`, cx.`是否改派`, cx.`物流渠道`
                            ORDER BY cx.`币种`,cx.`所属团队`, cx.`年月`, cx.`是否改派` DESC, 总订单 DESC
                    ) s1
                    GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`是否改派`,  s1.`物流方式`
                    with rollup
                ) s2
                LEFT JOIN 
                (  SELECT IFNULL(币种,'合计') as 币种, IFNULL(所属团队,'合计') as 家族, IFNULL(年月,'合计') as 年月, IFNULL(是否改派,'合计') as 是否改派, IFNULL(物流方式,'合计') as 物流方式, COUNT(订单编号) AS 总量
                    FROM gat_zqsb z
                    where z.`运单编号` is not null AND z.日期 >= '{0}' AND z.日期 <= '{1}'
                    GROUP BY 所属团队, 币种, 年月
                    with rollup
                ) s3 ON  s2.家族 = s3.家族 AND s2.币种 = s3.币种 AND s2.年月 = s3.年月
                GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`,  s2.`物流方式`
                HAVING s2.年月 <> '合计'
                ORDER BY FIELD(s2.`家族`,{3},'合计'),
                        FIELD(s2.`币种`,'台湾','香港','合计'),
                        s2.`年月`,
                        FIELD(s2.`是否改派`,'改派','直发','合计'),
                        FIELD(s2.`物流方式`, {4},'合计'),
                        s2.总订单 DESC;'''.format(month_last, month_yesterday, 'team_name',self.team_name2, self.logistics_name)  # 港台查询函数导出
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            listT.append(df)

            print('正在写入excel…………')
            today = datetime.date.today().strftime('%Y.%m.%d')
            file_path = 'F:\\输出文件\\{} 在线签收率_查询.xlsx'.format(today)
            sheet_name = ['物流分类']
            # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            # for i in range(len(listT)):
            #     listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            #     del book['Sheet1']
            # writer.save()
            # writer.close()
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for i in range(len(listT)):
                    listT[i].to_excel(excel_writer=writer, sheet_name=sheet_name[i], index=False)
            new_path = "F:\\神龙签收率\\" + (datetime.datetime.now()).strftime('%m.%d') + '\\签收率\\{} {} 在线签收率_查询.xlsx'.format(today, match[team])
            shutil.copyfile(file_path, new_path)  # copy到指定位置
            try:
                print('正在运行 在线签收率 宏…………')
                app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
                app.display_alerts = False
                wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
                wbsht1 = app.books.open(new_path)
                wbsht.macro('zl_gat_report_new2.gat_总_品类_物流_两月签收率')()
                wbsht1.save()
                wbsht1.close()
                wbsht.close()
                app.quit()
            except Exception as e:
                print('运行失败：', str(Exception) + str(e))
            print('----已写入excel ')

    def slrb_new(self, team, month_last, month_yesterday):  # 报表各团队近两个月的物流数据
        month_now = datetime.datetime.now().strftime('%Y-%m-%d')
        match = {'gat': '港台'}
        emailAdd = {'台湾': 'giikinliujun@163.com',
                    '香港': 'giikinliujun@163.com',
                    '品牌': 'sunyaru@giikin.com'}
        # if team == 'gat9':
        #     month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        #     month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
        # else:
        #     month_last = '2021-08-01'
        #     month_yesterday = '2021-09-30'
        print(month_last)
        print(month_yesterday)
        sql = '''UPDATE gat_zqsb d
                        SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-易速配-顺丰%','香港-易速配-顺丰', IF(d.`物流方式` LIKE '台湾-天马-711%','台湾-天马-新竹', d.`物流方式`) )
                        WHERE d.`是否改派` ='直发';'''
        print('正在修改-直发的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''UPDATE gat_zqsb d
                        SET d.`物流方式`= IF(d.`物流方式` LIKE '香港-森鸿%','香港-森鸿-改派',
                                        IF(d.`物流方式` LIKE '香港-立邦%','香港-立邦-改派',
            							IF(d.`物流方式` LIKE '香港-易速配%','香港-易速配-改派',
            							IF(d.`物流方式` LIKE '台湾-立邦普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-大黄蜂普货头程-森鸿尾程%' OR d.`物流方式` LIKE '台湾-森鸿-新竹%','森鸿',
            							IF(d.`物流方式` LIKE '台湾-天马-顺丰%','天马顺丰',
            							IF(d.`物流方式` LIKE '台湾-天马-新竹%' OR d.`物流方式` LIKE '台湾-天马-711%','天马新竹',
            							IF(d.`物流方式` LIKE '台湾-天马-黑猫%','天马黑猫',
            							IF(d.`物流方式` LIKE '台湾-易速配-龟山%' OR d.`物流方式` LIKE '台湾-易速配-新竹%' OR d.`物流方式` = '易速配','龟山',
            							IF(d.`物流方式` LIKE '台湾-速派-新竹%' OR d.`物流方式` LIKE '台湾-速派-711超商%','速派',
            							IF(d.`物流方式` LIKE '台湾-大黄蜂普货头程-易速配尾程%' OR d.`物流方式` LIKE '台湾-立邦普货头程-易速配尾程%','龟山', d.`物流方式`)))  )  )  )  )  )  )  )
                        WHERE d.`是否改派` ='改派';'''
        print('正在修改-改派的物流渠道…………')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        filePath = []
        listT = []  # 查询sql的结果 存放池
        print('正在获取---' + match[team] + '---签收率…………')
        # 物流分类
        print('正在获取---物流分类…………')
        sql0 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.物流方式,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,					
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                    concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                        concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                        concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
                FROM ( SELECT IFNULL(s1.币种,'合计') as 币种,
                            IFNULL(s1.家族,'合计') as 家族,
                            IFNULL(s1.年月,'合计') as 年月,
                            IFNULL(s1.是否改派,'合计') as 是否改派,
                            IFNULL(s1.物流方式,'合计') as 物流方式,
                            SUM(s1.签收) as 签收,
                            SUM(s1.拒收) as 拒收,
                            SUM(s1.在途) as 在途,
                            SUM(s1.未发货) as 未发货,
                            SUM(s1.未上线) as 未上线,
                            SUM(s1.已退货) as 已退货,
                            SUM(s1.理赔) as 理赔,
                            SUM(s1.自发头程丢件) as 自发头程丢件,
                            SUM(s1.已发货) as 已发货,
                            SUM(s1.已完成) as 已完成,
                            SUM(s1.总订单) as 总订单,
                            SUM(s1.签收金额) as 签收金额,
                            SUM(s1.退货金额) as 退货金额,
                            SUM(s1.完成金额) as 完成金额,
                            SUM(s1.发货金额) as 发货金额,
                            SUM(s1.总计金额) as 总计金额
                    FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.物流方式 as 物流方式,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                            ) cx
                            GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派`, cx.`物流方式`
                            ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
                    ) s1
                    GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`是否改派`, s1.`物流方式`
                    with rollup
                ) s2
                GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`物流方式` 
                HAVING s2.年月 <> '合计'
    ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','小虎队','神龙香港','Line运营','神龙主页运营','红杉','金狮','合计'),
            FIELD(s2.`币种`,'台湾','香港','合计'),
            s2.`年月`,
            FIELD(s2.`是否改派`,'改派','直发','合计'),
            FIELD(s2.`物流方式`, '台湾-大黄蜂普货头程-森鸿尾程','台湾-大黄蜂普货头程-易速配尾程', '台湾-立邦普货头程-森鸿尾程','台湾-易速配-TW海快','台湾-立邦普货头程-易速配尾程', 
                                '台湾-森鸿-新竹-自发头程', '台湾-速派-711超商', '台湾-速派-新竹','台湾-天马-新竹','台湾-天马-顺丰','台湾-天马-黑猫','台湾-易速配-新竹',
                                '香港-立邦-顺丰','香港-易速配-顺丰','香港-易速配-顺丰YC', '香港-森鸿-SH渠道','香港-森鸿-顺丰渠道',
                                '龟山','森鸿','速派','天马顺丰','天马新竹','香港-立邦-改派','香港-森鸿-改派','香港-易速配-改派','合计' ),
            s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df0 = pd.read_sql_query(sql=sql0, con=self.engine1)
        listT.append(df0)
        # 物流分旬
        print('正在获取---物流分旬…………')
        sql11 = '''SELECT s2.家族,s2.币种,s2.年月,s2.是否改派,s2.物流方式,s2.旬,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,	
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
                    concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
                        concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
                        concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
                        concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'    
                FROM (SELECT IFNULL(s1.币种,'合计') as 币种,
                            IFNULL(s1.家族,'合计') as 家族,
                            IFNULL(s1.年月,'合计') as 年月,
                            IFNULL(s1.是否改派,'合计') as 是否改派,
                            IFNULL(s1.物流方式,'合计') as 物流方式,
                            IFNULL(s1.旬,'合计') as 旬,
                            SUM(s1.签收) as 签收,
                            SUM(s1.拒收) as 拒收,
                            SUM(s1.在途) as 在途,
                            SUM(s1.未发货) as 未发货,
                            SUM(s1.未上线) as 未上线,
                            SUM(s1.已退货) as 已退货,
                            SUM(s1.理赔) as 理赔,
                            SUM(s1.自发头程丢件) as 自发头程丢件,
                            SUM(s1.已发货) as 已发货,
                            SUM(s1.已完成) as 已完成,
                            SUM(s1.总订单) as 总订单,
                            s1.总订单量,
							s1.已发货单量,
							s1.已完成单量,
                            SUM(s1.签收金额) as 签收金额,
                            SUM(s1.退货金额) as 退货金额,
                            SUM(s1.完成金额) as 完成金额,
                            SUM(s1.发货金额) as 发货金额,
                            SUM(s1.总计金额) as 总计金额
                    FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.是否改派 as 是否改派,cx.物流方式 as 物流方式,
                                IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
								总订单量,
								已发货单量,
								已完成单量,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                    FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                            ) cx
                            LEFT JOIN 
							    (SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                                FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                    ) dg  
								    GROUP BY dg.币种,dg.家族,dg.年月
                            ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                            GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派`, cx.`物流方式`, cx.`旬`
                            ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`是否改派` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`是否改派`, s1.`物流方式`, s1.`旬`
                        with rollup
                    ) s2 
                    GROUP BY s2.`家族`,s2.`币种`, s2.`年月`, s2.`是否改派`, s2.`物流方式`, s2.`旬`
                    HAVING s2.是否改派 <> '合计'
        ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','小虎队','神龙香港','Line运营','神龙主页运营','红杉','金狮','合计'),
                FIELD(s2.`币种`,'台湾','香港','合计'),
                s2.`年月`,
                FIELD(s2.`是否改派`,'改派','直发','合计'),
                FIELD(s2.`物流方式`,'台湾-大黄蜂普货头程-森鸿尾程','台湾-大黄蜂普货头程-易速配尾程','台湾-立邦普货头程-森鸿尾程','台湾-易速配-TW海快','台湾-立邦普货头程-易速配尾程',
                        '台湾-森鸿-新竹-自发头程','台湾-速派-711超商','台湾-速派-新竹','台湾-天马-新竹','台湾-天马-顺丰','台湾-天马-黑猫','台湾-易速配-新竹',
                        '香港-立邦-顺丰','香港-易速配-顺丰','香港-易速配-顺丰YC','香港-森鸿-SH渠道','香港-森鸿-顺丰渠道','合计'),   
                FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df11 = pd.read_sql_query(sql=sql11, con=self.engine1)
        listT.append(df11)

        # 父级分旬
        print('正在获取---父级分旬…………')
        sql12 = '''SELECT s2.家族,s2.币种,s2.年月,s2.父级分类,s2.旬,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,	
                    concat(ROUND(s2.签收 / s2.已完成 * 100,2),'%') as 完成签收,
                        concat(ROUND(s2.签收 / s2.总订单 * 100,2),'%') as 总计签收,
                        concat(ROUND(s2.已完成 / s2.总订单 * 100,2),'%') as 完成占比,
                        concat(ROUND(s2.已完成 / s2.已发货 * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(s2.已退货 / s2.总订单 * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
					concat(ROUND(s2.签收金额 / s2.完成金额 * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(s2.签收金额 / s2.总计金额 * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(s2.完成金额 / s2.总计金额 * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(s2.完成金额 / s2.发货金额 * 100,2),'%') as '已完成/已发货(金额)',
						concat(ROUND(s2.退货金额 / s2.总计金额 * 100,2),'%') as '退货率(金额)'
				 FROM ( SELECT  IFNULL(s1.币种,'合计') as 币种,IFNULL(s1.家族,'合计') as 家族,IFNULL(s1.年月,'合计') as 年月,IFNULL(s1.父级分类,'合计') as 父级分类,IFNULL(s1.旬,'合计') as 旬,
								SUM(s1.签收) as 签收,
								SUM(s1.拒收) as 拒收,
								SUM(s1.在途) as 在途,
								SUM(s1.未发货) as 未发货,
								SUM(s1.未上线) as 未上线,
								SUM(s1.已退货) as 已退货,
								SUM(s1.理赔) as 理赔,
								SUM(s1.自发头程丢件) as 自发头程丢件,
								SUM(s1.已发货) as 已发货,
								SUM(s1.已完成) as 已完成,
								SUM(s1.总订单) as 总订单,
                                s1.总订单量,s1.已发货单量,s1.已完成单量,
								SUM(s1.签收金额) as 签收金额,
								SUM(s1.退货金额) as 退货金额,
								SUM(s1.完成金额) as 完成金额,
								SUM(s1.发货金额) as 发货金额,
								SUM(s1.总计金额) as 总计金额
                        FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.父级分类 as 父级分类,IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
                                    SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                    SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                    SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                    SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                    SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                    SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                    SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                    SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                    count(订单编号) as 总订单,
                                    总订单量,已发货单量,已完成单量,
                                    count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                    SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                    SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                    SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                    SUM(`价格RMB`) as 总计金额,
                                    SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                                FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                ) cx
                                LEFT JOIN 
							        (SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                                    FROM (SELECT *,
                                                IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                            FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                        ) dg  
								        GROUP BY dg.币种,dg.家族,dg.年月
                                ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                                GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类`, cx.`旬`
                                ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`父级分类`, s1.`旬`
                        with rollup
                ) s2 HAVING s2.年月 <> '合计'
            ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','小虎队','神龙香港','Line运营','神龙主页运营','红杉','金狮','合计'),
                    FIELD(s2.`币种`,'台湾','香港','合计'),
                    s2.`年月`,
                    FIELD(s2.父级分类, '居家百货', '电子电器', '服饰', '医药保健',  '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','包材类','合计' ),
                    FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                    s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df12 = pd.read_sql_query(sql=sql12, con=self.engine1)
        listT.append(df12)
        # 二级分旬
        print('正在获取---二级分旬…………')
        sql13 = '''SELECT s2.家族,s2.币种,s2.年月,s2.父级分类,s2.二级分类,s2.旬,
						IF(s2.签收=0,NULL,s2.签收) as 签收,
						IF(s2.拒收=0,NULL,s2.拒收) as 拒收,
						IF(s2.在途=0,NULL,s2.在途) as 在途,				
						IF(s2.未发货=0,NULL,s2.未发货) as 未发货,
						IF(s2.未上线=0,NULL,s2.未上线) as 未上线,
						IF(s2.已退货=0,NULL,s2.已退货) as 已退货,					
						IF(s2.理赔=0,NULL,s2.理赔) as 理赔,
						IF(s2.自发头程丢件=0,NULL,s2.自发头程丢件) as 自发头程丢件,
						IF(s2.已发货=0,NULL,s2.已发货) as 已发货,
						IF(s2.已完成=0,NULL,s2.已完成) as 已完成,
						IF(s2.总订单=0,NULL,s2.总订单) as 全部,	
                    concat(ROUND(IFNULL(s2.签收 / s2.已完成,0) * 100,2),'%') as 完成签收,
                        concat(ROUND(IFNULL(s2.签收 / s2.总订单,0) * 100,2),'%') as 总计签收,
                        concat(ROUND(IFNULL(s2.已完成 / s2.总订单,0) * 100,2),'%') as 完成占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已发货,0) * 100,2),'%') as '已完成/已发货',
                        concat(ROUND(IFNULL(s2.已退货 / s2.总订单,0) * 100,2),'%') as 退货率,
                        concat(ROUND(IFNULL(s2.已发货 / s2.已发货单量,0) * 100,2),'%') as 已发货占比,
                        concat(ROUND(IFNULL(s2.已完成 / s2.已完成单量,0) * 100,2),'%') as 已完成占比,
                        concat(ROUND(IFNULL(s2.总订单 / s2.总订单量,0) * 100,2),'%') as 全部占比,
						concat(ROUND(IFNULL(s2.签收金额 / s2.完成金额,0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(s2.签收金额 / s2.总计金额,0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(s2.完成金额 / s2.总计金额,0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(s2.完成金额 / s2.发货金额,0) * 100,2),'%') as '已完成/已发货(金额)',
						concat(ROUND(IFNULL(s2.退货金额 / s2.总计金额,0) * 100,2),'%') as '退货率(金额)'
				 FROM ( SELECT  IFNULL(s1.币种,'合计') as 币种,IFNULL(s1.家族,'合计') as 家族,IFNULL(s1.年月,'合计') as 年月,IFNULL(s1.父级分类,'合计') as 父级分类,IFNULL(s1.二级分类,'合计') as 二级分类,IFNULL(s1.旬,'合计') as 旬,
								SUM(s1.签收) as 签收,
								SUM(s1.拒收) as 拒收,
								SUM(s1.在途) as 在途,
								SUM(s1.未发货) as 未发货,
								SUM(s1.未上线) as 未上线,
								SUM(s1.已退货) as 已退货,
								SUM(s1.理赔) as 理赔,
								SUM(s1.自发头程丢件) as 自发头程丢件,
								SUM(s1.已发货) as 已发货,
								SUM(s1.已完成) as 已完成,
								SUM(s1.总订单) as 总订单,
                                s1.总订单量,s1.已发货单量,s1.已完成单量,
								SUM(s1.签收金额) as 签收金额,
								SUM(s1.退货金额) as 退货金额,
								SUM(s1.完成金额) as 完成金额,
								SUM(s1.发货金额) as 发货金额,
								SUM(s1.总计金额) as 总计金额
                        FROM (SELECT cx.币种 as 币种,cx.家族 as 家族,cx.年月 as 年月,cx.父级分类 as 父级分类,cx.二级分类 as 二级分类,IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
                                SUM(IF(最终状态 = "已签收",1,0)) as 签收,
                                SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(最终状态 = "在途",1,0)) as 在途,
                                SUM(IF(最终状态 = "未发货",1,0)) as 未发货,
                                SUM(IF(最终状态 = "未上线",1,0)) as 未上线,
                                SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(最终状态 = "理赔",1,0)) as 理赔,
                                SUM(IF(最终状态 = "自发头程丢件",1,0)) as 自发头程丢件,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                count(订单编号) as 总订单,
								总订单量,已发货单量,已完成单量,
                                count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货,
                                SUM(IF(最终状态 = "已签收",`价格RMB`,0)) as 签收金额,
                                SUM(IF(最终状态 = "已退货",`价格RMB`,0)) as 退货金额,
                                SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),`价格RMB`,0)) as 完成金额,
                                SUM(`价格RMB`) as 总计金额,
                                SUM(`价格RMB`) - SUM(IF(最终状态 = "未发货",`价格RMB`,0)) as 发货金额
                            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                            ) cx
                            LEFT JOIN 
							    (SELECT 币种,家族,年月,物流方式,count(订单编号) as 总订单量, count(订单编号)-SUM(IF(最终状态 = "未发货",1,0)) as 已发货单量, SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成单量
                                FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}' AND cc.`是否改派` = '直发'
                                    ) dg  
								    GROUP BY dg.币种,dg.家族,dg.年月
                            ) cx2 ON cx.币种 = cx2.币种 AND  cx.家族 = cx2.家族 AND  cx.年月 = cx2.年月
                            GROUP BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类`, cx.`二级分类`, cx.`旬`
                            ORDER BY cx.`币种`,cx.`家族`, cx.`年月`, cx.`父级分类`, cx.`二级分类` DESC,总订单 DESC
                        ) s1
                        GROUP BY s1.`家族`,s1.`币种`, s1.`年月`, s1.`父级分类`, s1.`二级分类`, s1.`旬`
                        with rollup
                ) s2 HAVING s2.年月 <> '合计'
        ORDER BY FIELD(s2.`家族`,'神龙','火凤凰','小虎队','神龙香港','Line运营','神龙主页运营','红杉','金狮','合计'),
                FIELD(s2.`币种`,'台湾','香港','合计'),
                s2.`年月`,
                FIELD(s2.父级分类, '居家百货', '电子电器', '服饰', '医药保健', '鞋类', '美容个护', '包类','钟表珠宝','母婴玩具','包材类','合计' ),
                FIELD(s2.二级分类,'个人洗护','皮鞋','日用百货','影音娱乐','家用电器','药品','上衣','下装'
                            ,'饰品','保健器械','保健食品','彩妆','钱包','休闲运动鞋','内衣','护理护具','凉/拖鞋'
                            ,'裙子','个护电器','配饰','护肤','布艺家纺','母婴用品','厨房用品','汽车用品','双肩包'
                            ,'单肩包','手机外设','电脑外设','成人保健','套装','靴子','手表手环','行李箱包','户外运动'
                            ,'玩具','手表','宠物用品','智能设备','家装建材','母婴服饰','办公/文化','仓库包材','合计'),
                FIELD(s2.`旬`,'上旬','中旬','下旬','合计'),
                s2.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df13 = pd.read_sql_query(sql=sql13, con=self.engine1)
        listT.append(df13)

        # 产品整月 台湾
        print('正在获取---产品整月 台湾…………')
        sql14 = '''SELECT *
                FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                            IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                            SUM(s1.已签收) as 已签收,
						    SUM(s1.拒收) as 拒收,
						    SUM(s1.已退货) as 已退货,
						    SUM(s1.已完成) as 已完成,
						    SUM(s1.总订单) as 总订单,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						    concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						    concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
						SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
						    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
						    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
						    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
						    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
						    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
						    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
						    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
						SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
						    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
						    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
						    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
						    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
						    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
						    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
						    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						    SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						    SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						    SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						    SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						    concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						    concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						    concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						    concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
						SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
						    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
						    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
						    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
						    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
						    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
						    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
						    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
						SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
						    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
						    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
						    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
						    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
						    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
						    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
						    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
						SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
						    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
						    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
						    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
						    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
						    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
						    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
						    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
						SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
						    SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
						    SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
						    SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
						    SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
						    concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
						    concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
						    concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
						SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
						    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
						    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
						    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
						    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
						    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
				    		concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
					    	concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
					    	concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
					    	concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
						SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
						    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
					    	SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
					    	SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
					    	SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
				    		concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
				    		concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
					    	concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
						SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
					    	SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
					    	SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
					    	SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
					    	SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
					    	concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
					    	concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
					    	concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
						SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
					    	SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
						    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
					    	SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
					        SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
					    	concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
					    	concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
					    	concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
						SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
					    	SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
					    	SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
					    	SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
					    	SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
					    	concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
					    	concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
					    	concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
						SUM(s1.龟山改派已签收) as '龟山改派已签收',
					    	SUM(s1.龟山改派拒收) as '龟山改派拒收',
					    	SUM(s1.龟山改派已退货) as '龟山改派已退货',
					    	SUM(s1.龟山改派已完成) as '龟山改派已完成',
					    	SUM(s1.龟山改派总订单) as '龟山改派总订单',
					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
					    	concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
					    	concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
					    	concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
				    	SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
					    	SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
					    	SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
					    	SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
					    	SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
					    	concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
					    	concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
					    	concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
						SUM(s1.速派改派已签收) as '速派改派已签收',
					    	SUM(s1.速派改派拒收) as '速派改派拒收',
					    	SUM(s1.速派改派已退货) as '速派改派已退货',
					    	SUM(s1.速派改派已完成) as '速派改派已完成',
					    	SUM(s1.速派改派总订单) as '速派改派总订单',
					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
					    	concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
					    	concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
					    	concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
						SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
					    	SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
					    	SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
					    	SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
					    	SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
					    	concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
					    	concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
					    	concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
						SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
					    	SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
					    	SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
					    	SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
					    	SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
					    	concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
					    	concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
					    	concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                    FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
				            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                            ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                        ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                        WITH ROLLUP 
                ) s HAVING s.月份 != '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙香港','Line运营','金蝉家族','金蝉项目组','APP运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df14 = pd.read_sql_query(sql=sql14, con=self.engine1)
        listT.append(df14)
        # 产品分旬 台湾
        print('正在获取---产品分旬 台湾…………')
        sql15 = '''SELECT *
                    FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
						SUM(s1.已签收) as 已签收,
						SUM(s1.拒收) as 拒收,
						SUM(s1.已退货) as 已退货,
						SUM(s1.已完成) as 已完成,
						SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
						SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
						SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
						SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
						SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
						concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
						concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
						concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
					SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
						SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
						SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
						SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
						SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
						concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
						concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
						concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
					SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
					SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
						SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
						SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
						SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
						SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
						concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
						concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
						concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
					SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
						SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
						SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
						SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
						SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
						concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
						concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
						concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
					SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
						SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
						SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
						SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
						SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
						concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
						concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
						concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
					SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
						SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
						SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
						SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
						SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
						concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
						concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
						concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
					SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
						SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
						SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
						SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
						SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
						concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
						concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
						concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
					SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
						SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
						SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
						SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
						SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
						concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
						concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
						concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
					SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
						SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
						SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
						SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
						SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
						concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
						concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
						concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
					SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
						SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
						SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
						SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
						SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
						concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
						concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
						concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
					SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
						SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
						SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
						SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
						SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
						concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
						concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
						concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
					SUM(s1.龟山改派已签收) as '龟山改派已签收',
						SUM(s1.龟山改派拒收) as '龟山改派拒收',
						SUM(s1.龟山改派已退货) as '龟山改派已退货',
						SUM(s1.龟山改派已完成) as '龟山改派已完成',
						SUM(s1.龟山改派总订单) as '龟山改派总订单',
						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
						concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
						concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
						concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
					SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
						SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
						SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
						SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
						SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
						concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
						concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
						concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
					SUM(s1.速派改派已签收) as '速派改派已签收',
						SUM(s1.速派改派拒收) as '速派改派拒收',
						SUM(s1.速派改派已退货) as '速派改派已退货',
						SUM(s1.速派改派已完成) as '速派改派已完成',
						SUM(s1.速派改派总订单) as '速派改派总订单',
						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
						concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
						concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
						concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
					SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
						SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
						SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
						SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
						SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
						concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
						concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
						concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
					SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
						SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
						SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
						SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
						SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
						concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
						concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
						concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
				        FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                        ) cx WHERE cx.`币种` = '台湾'
                    GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                    ) s1
                GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                WITH ROLLUP 
            ) s HAVING s.旬 != '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙香港','Line运营','金蝉家族','金蝉项目组','APP运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df15 = pd.read_sql_query(sql=sql15, con=self.engine1)
        listT.append(df15)

        # 产品整月 香港
        print('正在获取---产品整月 香港…………')
        sql16 = '''SELECT *
                    FROM(SELECT IFNULL(s1.家族, '合计') 家族,
                                IFNULL(s1.地区, '合计') 地区,
                                IFNULL(s1.月份, '合计') 月份,
                        IFNULL(s1.产品id, '合计') 产品id,
                        IFNULL(s1.产品名称, '合计') 产品名称,
                        IFNULL(s1.父级分类, '合计') 父级分类,
                        IFNULL(s1.二级分类, '合计') 二级分类,
						SUM(s1.已签收) as 已签收,
						SUM(s1.拒收) as 拒收,
						SUM(s1.已退货) as 已退货,
						SUM(s1.已完成) as 已完成,
						SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
				            FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                            ) cx WHERE cx.`币种` = '香港'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                        ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.月份 != '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙香港','Line运营','金蝉家族','金蝉项目组','APP运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df16 = pd.read_sql_query(sql=sql16, con=self.engine1)
        listT.append(df16)
        # 产品分旬 香港
        print('正在获取---产品分旬 香港…………')
        sql17 = '''SELECT *
                    FROM(SELECT 
						IFNULL(s1.家族, '合计') 家族,
						IFNULL(s1.地区, '合计') 地区,
						IFNULL(s1.月份, '合计') 月份,
						IFNULL(s1.旬, '合计') 旬,
						IFNULL(s1.产品id, '合计') 产品id,
						IFNULL(s1.产品名称, '合计') 产品名称,
						IFNULL(s1.父级分类, '合计') 父级分类,
						IFNULL(s1.二级分类, '合计') 二级分类,
					SUM(s1.已签收) as 已签收,
						SUM(s1.拒收) as 拒收,
						SUM(s1.已退货) as 已退货,
						SUM(s1.已完成) as 已完成,
						SUM(s1.总订单) as 总订单,
					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
					concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
					SUM(s1.立邦顺丰已签收) as '香港-立邦-顺丰已签收',
						SUM(s1.立邦顺丰拒收) as '香港-立邦-顺丰拒收',
						SUM(s1.立邦顺丰已退货) as '香港-立邦-顺丰已退货',
						SUM(s1.立邦顺丰已完成) as '香港-立邦-顺丰已完成',
						SUM(s1.立邦顺丰总订单) as '香港-立邦-顺丰总订单',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰完成签收',
						concat(ROUND(SUM(s1.立邦顺丰已签收) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰总计签收',
						concat(ROUND(SUM(s1.立邦顺丰已完成) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰完成占比',
						concat(ROUND(SUM(s1.立邦顺丰已退货) / SUM(s1.立邦顺丰总订单) * 100,2),'%') as '香港-立邦-顺丰退货率',
						concat(ROUND(SUM(s1.立邦顺丰拒收) / SUM(s1.立邦顺丰已完成) * 100,2),'%') as '香港-立邦-顺丰拒收率',
					SUM(s1.易速配顺丰已签收) as '香港-易速配-顺丰已签收',
						SUM(s1.易速配顺丰拒收) as '香港-易速配-顺丰拒收',
						SUM(s1.易速配顺丰已退货) as '香港-易速配-顺丰已退货',
						SUM(s1.易速配顺丰已完成) as '香港-易速配-顺丰已完成',
						SUM(s1.易速配顺丰总订单) as '香港-易速配-顺丰总订单',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰完成签收',
						concat(ROUND(SUM(s1.易速配顺丰已签收) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰总计签收',
						concat(ROUND(SUM(s1.易速配顺丰已完成) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰完成占比',
						concat(ROUND(SUM(s1.易速配顺丰已退货) / SUM(s1.易速配顺丰总订单) * 100,2),'%') as '香港-易速配-顺丰退货率',
						concat(ROUND(SUM(s1.易速配顺丰拒收) / SUM(s1.易速配顺丰已完成) * 100,2),'%') as '香港-易速配-顺丰拒收率',
					SUM(s1.森鸿SH已签收) as '香港-森鸿-SH渠道已签收',
						SUM(s1.森鸿SH拒收) as '香港-森鸿-SH渠道拒收',
						SUM(s1.森鸿SH已退货) as '香港-森鸿-SH渠道已退货',
						SUM(s1.森鸿SH已完成) as '香港-森鸿-SH渠道已完成',
						SUM(s1.森鸿SH总订单) as '香港-森鸿-SH渠道总订单',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道完成签收',
						concat(ROUND(SUM(s1.森鸿SH已签收) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道总计签收',
						concat(ROUND(SUM(s1.森鸿SH已完成) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道完成占比',
						concat(ROUND(SUM(s1.森鸿SH已退货) / SUM(s1.森鸿SH总订单) * 100,2),'%') as '香港-森鸿-SH渠道退货率',
						concat(ROUND(SUM(s1.森鸿SH拒收) / SUM(s1.森鸿SH已完成) * 100,2),'%') as '香港-森鸿-SH渠道拒收率',
					SUM(s1.森鸿顺丰已签收) as '香港-森鸿-顺丰渠道已签收',
						SUM(s1.森鸿顺丰拒收) as '香港-森鸿-顺丰渠道拒收',
						SUM(s1.森鸿顺丰已退货) as '香港-森鸿-顺丰渠道已退货',
						SUM(s1.森鸿顺丰已完成) as '香港-森鸿-顺丰渠道已完成',
						SUM(s1.森鸿顺丰总订单) as '香港-森鸿-顺丰渠道总订单',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道完成签收',
						concat(ROUND(SUM(s1.森鸿顺丰已签收) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道总计签收',
						concat(ROUND(SUM(s1.森鸿顺丰已完成) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道完成占比',
						concat(ROUND(SUM(s1.森鸿顺丰已退货) / SUM(s1.森鸿顺丰总订单) * 100,2),'%') as '香港-森鸿-顺丰渠道退货率',
						concat(ROUND(SUM(s1.森鸿顺丰拒收) / SUM(s1.森鸿顺丰已完成) * 100,2),'%') as '香港-森鸿-顺丰渠道拒收率',
					SUM(s1.立邦改派已签收) as '香港-立邦-改派已签收',
						SUM(s1.立邦改派拒收) as '香港-立邦-改派拒收',
						SUM(s1.立邦改派已退货) as '香港-立邦-改派已退货',
						SUM(s1.立邦改派已完成) as '香港-立邦-改派已完成',
						SUM(s1.立邦改派总订单) as '香港-立邦-改派总订单',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派完成签收',
						concat(ROUND(SUM(s1.立邦改派已签收) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派总计签收',
						concat(ROUND(SUM(s1.立邦改派已完成) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派完成占比',
						concat(ROUND(SUM(s1.立邦改派已退货) / SUM(s1.立邦改派总订单) * 100,2),'%') as '香港-立邦-改派退货率',
						concat(ROUND(SUM(s1.立邦改派拒收) / SUM(s1.立邦改派已完成) * 100,2),'%') as '香港-立邦-改派拒收率',
					SUM(s1.易速配改派已签收) as '香港-易速配-改派已签收',
						SUM(s1.易速配改派拒收) as '香港-易速配-改派拒收',
						SUM(s1.易速配改派已退货) as '香港-易速配-改派已退货',
						SUM(s1.易速配改派已完成) as '香港-易速配-改派已完成',
						SUM(s1.易速配改派总订单) as '香港-易速配-改派总订单',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派完成签收',
						concat(ROUND(SUM(s1.易速配改派已签收) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派总计签收',
						concat(ROUND(SUM(s1.易速配改派已完成) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派完成占比',
						concat(ROUND(SUM(s1.易速配改派已退货) / SUM(s1.易速配改派总订单) * 100,2),'%') as '香港-易速配-改派退货率',
						concat(ROUND(SUM(s1.易速配改派拒收) / SUM(s1.易速配改派已完成) * 100,2),'%') as '香港-易速配-改派拒收率'
		            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
								IFNULL(cx.币种, '合计') 地区,
								IFNULL(cx.`年月`, '合计') 月份,
								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
								IFNULL(cx.产品id, '合计') 产品id,
								IFNULL(cx.产品名称, '合计') 产品名称,
								IFNULL(cx.父级分类, '合计') 父级分类,
								IFNULL(cx.二级分类, '合计') 二级分类,
							COUNT(cx.`订单编号`) as 总订单,
								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
							SUM(cx.`价格RMB`) as 总订单金额,
								SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
							SUM(IF(cx.物流方式 = "香港-立邦-顺丰" ,1,0)) AS 立邦顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已签收",1,0)) as 立邦顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "拒收",1,0)) as 立邦顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 = "已退货",1,0)) as 立邦顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-顺丰" ,1,0)) AS 易速配顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已签收",1,0)) as 易速配顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "拒收",1,0)) as 易速配顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 = "已退货",1,0)) as 易速配顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" ,1,0)) AS 森鸿SH总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已签收",1,0)) as 森鸿SH已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "拒收",1,0)) as 森鸿SH拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 = "已退货",1,0)) as 森鸿SH已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-SH渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿SH已完成,
							SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" ,1,0)) AS 森鸿顺丰总订单,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已签收",1,0)) as 森鸿顺丰已签收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "拒收",1,0)) as 森鸿顺丰拒收,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 = "已退货",1,0)) as 森鸿顺丰已退货,
								SUM(IF(cx.物流方式 = "香港-森鸿-顺丰渠道" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿顺丰已完成,
							SUM(IF(cx.物流方式 = "香港-立邦-改派" ,1,0)) AS 立邦改派总订单,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已签收",1,0)) as 立邦改派已签收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "拒收",1,0)) as 立邦改派拒收,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 = "已退货",1,0)) as 立邦改派已退货,
								SUM(IF(cx.物流方式 = "香港-立邦-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦改派已完成,
							SUM(IF(cx.物流方式 = "香港-易速配-改派" ,1,0)) AS 易速配改派总订单,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已签收",1,0)) as 易速配改派已签收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "拒收",1,0)) as 易速配改派拒收,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 = "已退货",1,0)) as 易速配改派已退货,
								SUM(IF(cx.物流方式 = "香港-易速配-改派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配改派已完成
				        FROM (SELECT *,
                                    IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                FROM {0}_zqsb cc where cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                        ) cx WHERE cx.`币种` = '香港'
                        GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                    ) s1
                    GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                    WITH ROLLUP 
            ) s HAVING s.旬 <> '合计'
        ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙香港','Line运营','金蝉家族','金蝉项目组','APP运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                FIELD(s.`地区`,'台湾','香港','合计'),
                FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                FIELD(s.`产品id`,'合计'),
                s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df17 = pd.read_sql_query(sql=sql17, con=self.engine1)
        listT.append(df17)

        # 产品整月_直发 台湾
        print('正在获取---产品整月_直发 台湾…………')
        sql18 = '''SELECT *
                        FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                                    IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                                    SUM(s1.已签收) as 已签收,
        						    SUM(s1.拒收) as 拒收,
        						    SUM(s1.已退货) as 已退货,
        						    SUM(s1.已完成) as 已完成,
        						    SUM(s1.总订单) as 总订单,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						            concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        						SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        						SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						        SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						            SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						            SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						            SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						            SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						            concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						            concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						            concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        						SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        						SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        						SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        						SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						    SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						    SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						    SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						    SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						    concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						    concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						    concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        						SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        				    		concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        					    	concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        					    	concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        					    	concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
        						SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        					    	SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        					    	SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        					    	SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        				    		concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        				    		concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        						SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        					    	SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        					    	SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        					    	SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        					    	SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        					    	concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        						SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        					    	SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        					    	SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        					        SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        					    	concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        					    	concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        						SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        					    	SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        					    	SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        					    	SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        					    	SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        					    	concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        					    	concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        						SUM(s1.龟山改派已签收) as '龟山改派已签收',
        					    	SUM(s1.龟山改派拒收) as '龟山改派拒收',
        					    	SUM(s1.龟山改派已退货) as '龟山改派已退货',
        					    	SUM(s1.龟山改派已完成) as '龟山改派已完成',
        					    	SUM(s1.龟山改派总订单) as '龟山改派总订单',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        					    	concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        					    	concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        					    	concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        				    	SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        					    	SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        					    	SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        					    	SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        					    	SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        					    	concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        					    	concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        						SUM(s1.速派改派已签收) as '速派改派已签收',
        					    	SUM(s1.速派改派拒收) as '速派改派拒收',
        					    	SUM(s1.速派改派已退货) as '速派改派已退货',
        					    	SUM(s1.速派改派已完成) as '速派改派已完成',
        					    	SUM(s1.速派改派总订单) as '速派改派总订单',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        					    	concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        					    	concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        					    	concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
        						SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        					    	SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        					    	SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        					    	SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        					    	SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        					    	concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        						SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        					    	SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        					    	SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        					    	SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        					    	SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
        				            FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`是否改派` = '直发' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                    ) cx WHERE cx.`币种` = '台湾'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                        ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙香港','Line运营','金蝉家族','金蝉项目组','APP运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df18 = pd.read_sql_query(sql=sql18, con=self.engine1)
        listT.append(df18)
        # 产品分旬_直发 台湾
        print('正在获取---产品分旬_直发 台湾…………')
        sql19 = '''SELECT *
                            FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,
        						SUM(s1.拒收) as 拒收,
        						SUM(s1.已退货) as 已退货,
        						SUM(s1.已完成) as 已完成,
        						SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						        concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        					SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        					SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						    SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						        SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						        SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						        SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						        SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						        concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						        concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						        concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        					SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        					SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        					SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        					SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        					SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        						concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        						concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        						concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
        					SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        						SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        						SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        						SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        						concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        						concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        						concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        					SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        						SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        						SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        						SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        						SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        						concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        						concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        						concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        					SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        						SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        						SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        						SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        						concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        						concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        						concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        					SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        						SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        						SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        						SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        						SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        						concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        						concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        						concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        					SUM(s1.龟山改派已签收) as '龟山改派已签收',
        						SUM(s1.龟山改派拒收) as '龟山改派拒收',
        						SUM(s1.龟山改派已退货) as '龟山改派已退货',
        						SUM(s1.龟山改派已完成) as '龟山改派已完成',
        						SUM(s1.龟山改派总订单) as '龟山改派总订单',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        						concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        						concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        						concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        					SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        						SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        						SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        						SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        						SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        						concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        						concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        						concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        					SUM(s1.速派改派已签收) as '速派改派已签收',
        						SUM(s1.速派改派拒收) as '速派改派拒收',
        						SUM(s1.速派改派已退货) as '速派改派已退货',
        						SUM(s1.速派改派已完成) as '速派改派已完成',
        						SUM(s1.速派改派总订单) as '速派改派总订单',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        						concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        						concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        						concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
        					SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        						SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        						SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        						SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        						SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        						concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        						concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        						concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        					SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        						SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        						SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        						SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        						SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        						concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        						concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                        FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
        				        FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                    FROM {0}_zqsb cc where  cc.`是否改派` = '直发' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.旬 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙香港','Line运营','金蝉家族','金蝉项目组','APP运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df19 = pd.read_sql_query(sql=sql19, con=self.engine1)
        listT.append(df19)

        # 产品整月_改派 台湾
        print('正在获取---产品整月_直发 台湾…………')
        sql20 = '''SELECT *
                        FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,
                                    IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
                                    SUM(s1.已签收) as 已签收,
        						    SUM(s1.拒收) as 拒收,
        						    SUM(s1.已退货) as 已退货,
        						    SUM(s1.已完成) as 已完成,
        						    SUM(s1.总订单) as 总订单,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						    concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						    concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						    concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						    concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						            concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						            concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        						SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						    SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						    SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						    SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						    SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        						SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						    SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						    SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						    SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						    SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						        SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						            SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						            SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						            SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						            SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						            concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						            concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						            concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						            concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        						SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						    SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						    SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						    SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						    SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        						SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						    SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						    SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						    SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						    SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						    concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						    concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						    concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        						SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						    SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						    SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						    SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						    SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						    concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						    concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						    concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        						SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						    SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						    SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						    SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						    SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						    concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						    concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						    concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						    concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        						SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						    SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						    SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						    SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						    SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						    concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        				    		concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        					    	concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        					    	concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        					    	concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
        						SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						    SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        					    	SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        					    	SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        					    	SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        				    		concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        				    		concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        						SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        					    	SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        					    	SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        					    	SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        					    	SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        					    	concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        						SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        					    	SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						    SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        					    	SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        					        SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        					    	concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        					    	concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        					    	concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        						SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        					    	SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        					    	SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        					    	SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        					    	SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        					    	concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        					    	concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        					    	concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        						SUM(s1.龟山改派已签收) as '龟山改派已签收',
        					    	SUM(s1.龟山改派拒收) as '龟山改派拒收',
        					    	SUM(s1.龟山改派已退货) as '龟山改派已退货',
        					    	SUM(s1.龟山改派已完成) as '龟山改派已完成',
        					    	SUM(s1.龟山改派总订单) as '龟山改派总订单',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        					    	concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        					    	concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        					    	concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        					    	concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        				    	SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        					    	SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        					    	SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        					    	SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        					    	SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        					    	concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        					    	concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        					    	concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        						SUM(s1.速派改派已签收) as '速派改派已签收',
        					    	SUM(s1.速派改派拒收) as '速派改派拒收',
        					    	SUM(s1.速派改派已退货) as '速派改派已退货',
        					    	SUM(s1.速派改派已完成) as '速派改派已完成',
        					    	SUM(s1.速派改派总订单) as '速派改派总订单',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        					    	concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        					    	concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        					    	concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        					    	concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
        						SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        					    	SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        					    	SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        					    	SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        					    	SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        					    	concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        					    	concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        					    	concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        						SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        					    	SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        					    	SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        					    	SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        					    	SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        					    	concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        					    	concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                            FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
        				            FROM (SELECT *,
                                            IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                        FROM {0}_zqsb cc where cc.`是否改派` = '改派' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                    ) cx WHERE cx.`币种` = '台湾'
                                    GROUP BY cx.家族,cx.币种,cx.年月,cx.产品id
                                ) s1
                                GROUP BY s1.家族,s1.地区,s1.月份,s1.产品id
                                WITH ROLLUP 
                        ) s HAVING s.月份 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙香港','Line运营','金蝉家族','金蝉项目组','APP运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df20 = pd.read_sql_query(sql=sql20, con=self.engine1)
        listT.append(df20)
        # 产品分旬_改派 台湾
        print('正在获取---产品分旬_直发 台湾…………')
        sql21 = '''SELECT *
                            FROM(SELECT IFNULL(s1.家族, '合计') 家族,IFNULL(s1.地区, '合计') 地区,IFNULL(s1.月份, '合计') 月份,IFNULL(s1.旬, '合计') 旬,
        						IFNULL(s1.产品id, '合计') 产品id,IFNULL(s1.产品名称, '合计') 产品名称,IFNULL(s1.父级分类, '合计') 父级分类,IFNULL(s1.二级分类, '合计') 二级分类,
        						SUM(s1.已签收) as 已签收,
        						SUM(s1.拒收) as 拒收,
        						SUM(s1.已退货) as 已退货,
        						SUM(s1.已完成) as 已完成,
        						SUM(s1.总订单) as 总订单,
        					concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.已完成),0) * 100,2),'%') as 完成签收,
        						concat(ROUND(IFNULL(SUM(s1.已签收) / SUM(s1.总订单),0) * 100,2),'%') as 总计签收,
        						concat(ROUND(IFNULL(SUM(s1.已完成) / SUM(s1.总订单),0) * 100,2),'%') as 完成占比,
        						concat(ROUND(IFNULL(SUM(s1.已退货) / SUM(s1.总订单),0) * 100,2),'%') as 退货率,
        						concat(ROUND(IFNULL(SUM(s1.拒收) / SUM(s1.已完成),0) * 100,2),'%') as 拒收率,
						    concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '完成签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已签收金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '总计签收(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已完成金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '完成占比(金额)',
						        concat(ROUND(IFNULL(SUM(s1.已退货金额) / SUM(s1.总订单金额),0) * 100,2),'%') as '退货率(金额)',
						        concat(ROUND(IFNULL(SUM(s1.拒收金额) / SUM(s1.已完成金额),0) * 100,2),'%') as '拒收率(金额)',
        					SUM(s1.大黄蜂已签收) as '台湾-大黄蜂普货头程-森鸿尾程已签收',
        						SUM(s1.大黄蜂拒收) as '台湾-大黄蜂普货头程-森鸿尾程拒收',
        						SUM(s1.大黄蜂已退货) as '台湾-大黄蜂普货头程-森鸿尾程已退货',
        						SUM(s1.大黄蜂已完成) as '台湾-大黄蜂普货头程-森鸿尾程已完成',
        						SUM(s1.大黄蜂总订单) as '台湾-大黄蜂普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂已签收) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂已完成) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂已退货) / SUM(s1.大黄蜂总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂拒收) / SUM(s1.大黄蜂已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-森鸿尾程拒收率',
        					SUM(s1.大黄蜂易速配已签收) as '台湾-大黄蜂普货头程-易速配尾程已签收',
        						SUM(s1.大黄蜂易速配拒收) as '台湾-大黄蜂普货头程-易速配尾程拒收',
        						SUM(s1.大黄蜂易速配已退货) as '台湾-大黄蜂普货头程-易速配尾程已退货',
        						SUM(s1.大黄蜂易速配已完成) as '台湾-大黄蜂普货头程-易速配尾程已完成',
        						SUM(s1.大黄蜂易速配总订单) as '台湾-大黄蜂普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已签收) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.大黄蜂易速配已完成) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.大黄蜂易速配已退货) / SUM(s1.大黄蜂易速配总订单) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.大黄蜂易速配拒收) / SUM(s1.大黄蜂易速配已完成) * 100,2),'%') as '台湾-大黄蜂普货头程-易速配尾程拒收率',
						    SUM(s1.TW海快易速配已签收) as '台湾-易速配-TW海快已签收',
						        SUM(s1.TW海快易速配拒收) as '台湾-易速配-TW海快拒收',
						        SUM(s1.TW海快易速配已退货) as '台湾-易速配-TW海快已退货',
						        SUM(s1.TW海快易速配已完成) as '台湾-易速配-TW海快已完成',
						        SUM(s1.TW海快易速配总订单) as '台湾-易速配-TW海快总订单',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快完成签收',
						        concat(ROUND(SUM(s1.TW海快易速配已签收) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快总计签收',
						        concat(ROUND(SUM(s1.TW海快易速配已完成) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快完成占比',
						        concat(ROUND(SUM(s1.TW海快易速配已退货) / SUM(s1.TW海快易速配总订单) * 100,2),'%') as '台湾-易速配-TW海快退货率',
						        concat(ROUND(SUM(s1.TW海快易速配拒收) / SUM(s1.TW海快易速配已完成) * 100,2),'%') as '台湾-易速配-TW海快拒收率',
        					SUM(s1.立邦普货已签收) as '台湾-立邦普货头程-森鸿尾程已签收',
        						SUM(s1.立邦普货拒收) as '台湾-立邦普货头程-森鸿尾程拒收',
        						SUM(s1.立邦普货已退货) as '台湾-立邦普货头程-森鸿尾程已退货',
        						SUM(s1.立邦普货已完成) as '台湾-立邦普货头程-森鸿尾程已完成',
        						SUM(s1.立邦普货总订单) as '台湾-立邦普货头程-森鸿尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货已签收) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货已完成) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货已退货) / SUM(s1.立邦普货总订单) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货拒收) / SUM(s1.立邦普货已完成) * 100,2),'%') as '台湾-立邦普货头程-森鸿尾程拒收率',
        					SUM(s1.立邦普货易速配已签收) as '台湾-立邦普货头程-易速配尾程已签收',
        						SUM(s1.立邦普货易速配拒收) as '台湾-立邦普货头程-易速配尾程拒收',
        						SUM(s1.立邦普货易速配已退货) as '台湾-立邦普货头程-易速配尾程已退货',
        						SUM(s1.立邦普货易速配已完成) as '台湾-立邦普货头程-易速配尾程已完成',
        						SUM(s1.立邦普货易速配总订单) as '台湾-立邦普货头程-易速配尾程总订单',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已签收) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程总计签收',
        						concat(ROUND(SUM(s1.立邦普货易速配已完成) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程完成占比',
        						concat(ROUND(SUM(s1.立邦普货易速配已退货) / SUM(s1.立邦普货易速配总订单) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程退货率',
        						concat(ROUND(SUM(s1.立邦普货易速配拒收) / SUM(s1.立邦普货易速配已完成) * 100,2),'%') as '台湾-立邦普货头程-易速配尾程拒收率',
        					SUM(s1.森鸿新竹已签收) as '台湾-森鸿-新竹-自发头程已签收',
        						SUM(s1.森鸿新竹拒收) as '台湾-森鸿-新竹-自发头程拒收',
        						SUM(s1.森鸿新竹已退货) as '台湾-森鸿-新竹-自发头程已退货',
        						SUM(s1.森鸿新竹已完成) as '台湾-森鸿-新竹-自发头程已完成',
        						SUM(s1.森鸿新竹总订单) as '台湾-森鸿-新竹-自发头程总订单',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成签收',
        						concat(ROUND(SUM(s1.森鸿新竹已签收) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程总计签收',
        						concat(ROUND(SUM(s1.森鸿新竹已完成) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程完成占比',
        						concat(ROUND(SUM(s1.森鸿新竹已退货) / SUM(s1.森鸿新竹总订单) * 100,2),'%') as '台湾-森鸿-新竹-自发头程退货率',
        						concat(ROUND(SUM(s1.森鸿新竹拒收) / SUM(s1.森鸿新竹已完成) * 100,2),'%') as '台湾-森鸿-新竹-自发头程拒收率',
        					SUM(s1.速派超商已签收) as '台湾-速派-711超商已签收',
        						SUM(s1.速派超商拒收) as '台湾-速派-711超商拒收',
        						SUM(s1.速派超商已退货) as '台湾-速派-711超商已退货',
        						SUM(s1.速派超商已完成) as '台湾-速派-711超商已完成',
        						SUM(s1.速派超商总订单) as '台湾-速派-711超商总订单',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商完成签收',
        						concat(ROUND(SUM(s1.速派超商已签收) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商总计签收',
        						concat(ROUND(SUM(s1.速派超商已完成) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商完成占比',
        						concat(ROUND(SUM(s1.速派超商已退货) / SUM(s1.速派超商总订单) * 100,2),'%') as '台湾-速派-711超商退货率',
        						concat(ROUND(SUM(s1.速派超商拒收) / SUM(s1.速派超商已完成) * 100,2),'%') as '台湾-速派-711超商拒收率',
        					SUM(s1.速派新竹已签收) as '台湾-速派-新竹已签收',
        						SUM(s1.速派新竹拒收) as '台湾-速派-新竹拒收',
        						SUM(s1.速派新竹已退货) as '台湾-速派-新竹已退货',
        						SUM(s1.速派新竹已完成) as '台湾-速派-新竹已完成',
        						SUM(s1.速派新竹总订单) as '台湾-速派-新竹总订单',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹完成签收',
        						concat(ROUND(SUM(s1.速派新竹已签收) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹总计签收',
        						concat(ROUND(SUM(s1.速派新竹已完成) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹完成占比',
        						concat(ROUND(SUM(s1.速派新竹已退货) / SUM(s1.速派新竹总订单) * 100,2),'%') as '台湾-速派-新竹退货率',
        						concat(ROUND(SUM(s1.速派新竹拒收) / SUM(s1.速派新竹已完成) * 100,2),'%') as '台湾-速派-新竹拒收率',
        					SUM(s1.天马顺丰已签收) as '台湾-天马-顺丰已签收',
        						SUM(s1.天马顺丰拒收) as '台湾-天马-顺丰拒收',
        						SUM(s1.天马顺丰已退货) as '台湾-天马-顺丰已退货',
        						SUM(s1.天马顺丰已完成) as '台湾-天马-顺丰已完成',
        						SUM(s1.天马顺丰总订单) as '台湾-天马-顺丰总订单',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰完成签收',
        						concat(ROUND(SUM(s1.天马顺丰已签收) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰总计签收',
        						concat(ROUND(SUM(s1.天马顺丰已完成) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰完成占比',
        						concat(ROUND(SUM(s1.天马顺丰已退货) / SUM(s1.天马顺丰总订单) * 100,2),'%') as '台湾-天马-顺丰退货率',
        						concat(ROUND(SUM(s1.天马顺丰拒收) / SUM(s1.天马顺丰已完成) * 100,2),'%') as '台湾-天马-顺丰拒收率',
        					SUM(s1.天马新竹已签收) as '台湾-天马-新竹已签收',
        						SUM(s1.天马新竹拒收) as '台湾-天马-新竹拒收',
        						SUM(s1.天马新竹已退货) as '台湾-天马-新竹已退货',
        						SUM(s1.天马新竹已完成) as '台湾-天马-新竹已完成',
        						SUM(s1.天马新竹总订单) as '台湾-天马-新竹总订单',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹完成签收',
        						concat(ROUND(SUM(s1.天马新竹已签收) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹总计签收',
        						concat(ROUND(SUM(s1.天马新竹已完成) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹完成占比',
        						concat(ROUND(SUM(s1.天马新竹已退货) / SUM(s1.天马新竹总订单) * 100,2),'%') as '台湾-天马-新竹退货率',
        						concat(ROUND(SUM(s1.天马新竹拒收) / SUM(s1.天马新竹已完成) * 100,2),'%') as '台湾-天马-新竹拒收率',
        					SUM(s1.天马黑猫已签收) as '台湾-天马-黑猫已签收',
        						SUM(s1.天马黑猫拒收) as '台湾-天马-黑猫拒收',
        						SUM(s1.天马黑猫已退货) as '台湾-天马-黑猫已退货',
        						SUM(s1.天马黑猫已完成) as '台湾-天马-黑猫已完成',
        						SUM(s1.天马黑猫总订单) as '台湾-天马-黑猫总订单',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫完成签收',
        						concat(ROUND(SUM(s1.天马黑猫已签收) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫总计签收',
        						concat(ROUND(SUM(s1.天马黑猫已完成) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫完成占比',
        						concat(ROUND(SUM(s1.天马黑猫已退货) / SUM(s1.天马黑猫总订单) * 100,2),'%') as '台湾-天马-黑猫退货率',
        						concat(ROUND(SUM(s1.天马黑猫拒收) / SUM(s1.天马黑猫已完成) * 100,2),'%') as '台湾-天马-黑猫拒收率',
        					SUM(s1.易速配新竹已签收) as '台湾-易速配-新竹已签收',
        						SUM(s1.易速配新竹拒收) as '台湾-易速配-新竹拒收',
        						SUM(s1.易速配新竹已退货) as '台湾-易速配-新竹已退货',
        						SUM(s1.易速配新竹已完成) as '台湾-易速配-新竹已完成',
        						SUM(s1.易速配新竹总订单) as '台湾-易速配-新竹总订单',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹完成签收',
        						concat(ROUND(SUM(s1.易速配新竹已签收) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹总计签收',
        						concat(ROUND(SUM(s1.易速配新竹已完成) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹完成占比',
        						concat(ROUND(SUM(s1.易速配新竹已退货) / SUM(s1.易速配新竹总订单) * 100,2),'%') as '台湾-易速配-新竹退货率',
        						concat(ROUND(SUM(s1.易速配新竹拒收) / SUM(s1.易速配新竹已完成) * 100,2),'%') as '台湾-易速配-新竹拒收率',
        					SUM(s1.龟山改派已签收) as '龟山改派已签收',
        						SUM(s1.龟山改派拒收) as '龟山改派拒收',
        						SUM(s1.龟山改派已退货) as '龟山改派已退货',
        						SUM(s1.龟山改派已完成) as '龟山改派已完成',
        						SUM(s1.龟山改派总订单) as '龟山改派总订单',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派完成签收',
        						concat(ROUND(SUM(s1.龟山改派已签收) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派总计签收',
        						concat(ROUND(SUM(s1.龟山改派已完成) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派完成占比',
        						concat(ROUND(SUM(s1.龟山改派已退货) / SUM(s1.龟山改派总订单) * 100,2),'%') as '龟山改派退货率',
        						concat(ROUND(SUM(s1.龟山改派拒收) / SUM(s1.龟山改派已完成) * 100,2),'%') as '龟山改派拒收率',
        					SUM(s1.森鸿改派已签收) as '森鸿改派已签收',
        						SUM(s1.森鸿改派拒收) as '森鸿改派拒收',
        						SUM(s1.森鸿改派已退货) as '森鸿改派已退货',
        						SUM(s1.森鸿改派已完成) as '森鸿改派已完成',
        						SUM(s1.森鸿改派总订单) as '森鸿改派总订单',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派完成签收',
        						concat(ROUND(SUM(s1.森鸿改派已签收) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派总计签收',
        						concat(ROUND(SUM(s1.森鸿改派已完成) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派完成占比',
        						concat(ROUND(SUM(s1.森鸿改派已退货) / SUM(s1.森鸿改派总订单) * 100,2),'%') as '森鸿改派退货率',
        						concat(ROUND(SUM(s1.森鸿改派拒收) / SUM(s1.森鸿改派已完成) * 100,2),'%') as '森鸿改派拒收率',
        					SUM(s1.速派改派已签收) as '速派改派已签收',
        						SUM(s1.速派改派拒收) as '速派改派拒收',
        						SUM(s1.速派改派已退货) as '速派改派已退货',
        						SUM(s1.速派改派已完成) as '速派改派已完成',
        						SUM(s1.速派改派总订单) as '速派改派总订单',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派完成签收',
        						concat(ROUND(SUM(s1.速派改派已签收) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派总计签收',
        						concat(ROUND(SUM(s1.速派改派已完成) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派完成占比',
        						concat(ROUND(SUM(s1.速派改派已退货) / SUM(s1.速派改派总订单) * 100,2),'%') as '速派改派退货率',
        						concat(ROUND(SUM(s1.速派改派拒收) / SUM(s1.速派改派已完成) * 100,2),'%') as '速派改派拒收率',
        					SUM(s1.天马新竹改派已签收) as '天马新竹改派已签收',
        						SUM(s1.天马新竹改派拒收) as '天马新竹改派拒收',
        						SUM(s1.天马新竹改派已退货) as '天马新竹改派已退货',
        						SUM(s1.天马新竹改派已完成) as '天马新竹改派已完成',
        						SUM(s1.天马新竹改派总订单) as '天马新竹改派总订单',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派完成签收',
        						concat(ROUND(SUM(s1.天马新竹改派已签收) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派总计签收',
        						concat(ROUND(SUM(s1.天马新竹改派已完成) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派完成占比',
        						concat(ROUND(SUM(s1.天马新竹改派已退货) / SUM(s1.天马新竹改派总订单) * 100,2),'%') as '天马新竹改派退货率',
        						concat(ROUND(SUM(s1.天马新竹改派拒收) / SUM(s1.天马新竹改派已完成) * 100,2),'%') as '天马新竹改派拒收率',
        					SUM(s1.天马顺丰改派已签收) as '天马顺丰改派已签收',
        						SUM(s1.天马顺丰改派拒收) as '天马顺丰改派拒收',
        						SUM(s1.天马顺丰改派已退货) as '天马顺丰改派已退货',
        						SUM(s1.天马顺丰改派已完成) as '天马顺丰改派已完成',
        						SUM(s1.天马顺丰改派总订单) as '天马顺丰改派总订单',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派完成签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已签收) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派总计签收',
        						concat(ROUND(SUM(s1.天马顺丰改派已完成) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派完成占比',
        						concat(ROUND(SUM(s1.天马顺丰改派已退货) / SUM(s1.天马顺丰改派总订单) * 100,2),'%') as '天马顺丰改派退货率',
        						concat(ROUND(SUM(s1.天马顺丰改派拒收) / SUM(s1.天马顺丰改派已完成) * 100,2),'%') as '天马顺丰改派拒收率'
                        FROM(SELECT IFNULL(cx.`家族`, '合计') 家族,
        								IFNULL(cx.币种, '合计') 地区,
        								IFNULL(cx.`年月`, '合计') 月份,
        								IF(cx.旬 =1,'上旬',IF(cx.旬 =2,'中旬',IF(cx.旬 =3,'下旬',cx.旬))) as 旬,
        								IFNULL(cx.产品id, '合计') 产品id,
        								IFNULL(cx.产品名称, '合计') 产品名称,
        								IFNULL(cx.父级分类, '合计') 父级分类,
        								IFNULL(cx.二级分类, '合计') 二级分类,
        							COUNT(cx.`订单编号`) as 总订单,
        								SUM(IF(最终状态 = "已签收",1,0)) as 已签收,
        								SUM(IF(最终状态 = "拒收",1,0)) as 拒收,
        								SUM(IF(最终状态 = "已退货",1,0)) as 已退货,
        								SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
								    SUM(cx.`价格RMB`) as 总订单金额,
								        SUM(IF(最终状态 = "已签收",价格RMB,0)) as 已签收金额,
								        SUM(IF(最终状态 = "拒收",价格RMB,0)) as 拒收金额,
								        SUM(IF(最终状态 = "已退货",价格RMB,0)) as 已退货金额,
								        SUM(IF(最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),价格RMB,0)) as 已完成金额,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" ,1,0)) AS 大黄蜂总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂已完成,
        							SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" ,1,0)) AS 大黄蜂易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 大黄蜂易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 大黄蜂易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 大黄蜂易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-大黄蜂普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 大黄蜂易速配已完成,
							        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" ,1,0)) AS TW海快易速配总订单,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已签收",1,0)) as TW海快易速配已签收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "拒收",1,0)) as TW海快易速配拒收,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 = "已退货",1,0)) as TW海快易速配已退货,
								        SUM(IF(cx.物流方式 = "台湾-易速配-TW海快" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as TW海快易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" ,1,0)) AS 立邦普货总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-森鸿尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货已完成,
        							SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" ,1,0)) AS 立邦普货易速配总订单,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已签收",1,0)) as 立邦普货易速配已签收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "拒收",1,0)) as 立邦普货易速配拒收,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 = "已退货",1,0)) as 立邦普货易速配已退货,
        								SUM(IF(cx.物流方式 = "台湾-立邦普货头程-易速配尾程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 立邦普货易速配已完成,
        							SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" ,1,0)) AS 森鸿新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已签收",1,0)) as 森鸿新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "拒收",1,0)) as 森鸿新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 = "已退货",1,0)) as 森鸿新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-森鸿-新竹-自发头程" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-711超商" ,1,0)) AS 速派超商总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已签收",1,0)) as 速派超商已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "拒收",1,0)) as 速派超商拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 = "已退货",1,0)) as 速派超商已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-711超商" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派超商已完成,
        							SUM(IF(cx.物流方式 = "台湾-速派-新竹" ,1,0)) AS 速派新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已签收",1,0)) as 速派新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "拒收",1,0)) as 速派新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 = "已退货",1,0)) as 速派新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-速派-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-顺丰" ,1,0)) AS 天马顺丰总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-新竹" ,1,0)) AS 天马新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹已完成,
        							SUM(IF(cx.物流方式 = "台湾-天马-黑猫" ,1,0)) AS 天马黑猫总订单,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已签收",1,0)) as 天马黑猫已签收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "拒收",1,0)) as 天马黑猫拒收,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 = "已退货",1,0)) as 天马黑猫已退货,
        								SUM(IF(cx.物流方式 = "台湾-天马-黑猫" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马黑猫已完成,
        							SUM(IF(cx.物流方式 = "台湾-易速配-新竹" ,1,0)) AS 易速配新竹总订单,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已签收",1,0)) as 易速配新竹已签收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "拒收",1,0)) as 易速配新竹拒收,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 = "已退货",1,0)) as 易速配新竹已退货,
        								SUM(IF(cx.物流方式 = "台湾-易速配-新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 易速配新竹已完成,
        							SUM(IF(cx.物流方式 = "龟山" ,1,0)) AS 龟山改派总订单,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已签收",1,0)) as 龟山改派已签收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "拒收",1,0)) as 龟山改派拒收,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 = "已退货",1,0)) as 龟山改派已退货,
        								SUM(IF(cx.物流方式 = "龟山" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 龟山改派已完成,
        							SUM(IF(cx.物流方式 = "森鸿" ,1,0)) AS 森鸿改派总订单,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已签收",1,0)) as 森鸿改派已签收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "拒收",1,0)) as 森鸿改派拒收,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 = "已退货",1,0)) as 森鸿改派已退货,
        								SUM(IF(cx.物流方式 = "森鸿" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 森鸿改派已完成,
        							SUM(IF(cx.物流方式 = "速派" ,1,0)) AS 速派改派总订单,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已签收",1,0)) as 速派改派已签收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "拒收",1,0)) as 速派改派拒收,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 = "已退货",1,0)) as 速派改派已退货,
        								SUM(IF(cx.物流方式 = "速派" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 速派改派已完成,
        							SUM(IF(cx.物流方式 = "天马新竹" ,1,0)) AS 天马新竹改派总订单,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已签收",1,0)) as 天马新竹改派已签收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "拒收",1,0)) as 天马新竹改派拒收,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 = "已退货",1,0)) as 天马新竹改派已退货,
        								SUM(IF(cx.物流方式 = "天马新竹" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马新竹改派已完成,
        							SUM(IF(cx.物流方式 = "天马顺丰" ,1,0)) AS 天马顺丰改派总订单,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已签收",1,0)) as 天马顺丰改派已签收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "拒收",1,0)) as 天马顺丰改派拒收,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 = "已退货",1,0)) as 天马顺丰改派已退货,
        								SUM(IF(cx.物流方式 = "天马顺丰" AND 最终状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 天马顺丰改派已完成
        				        FROM (SELECT *,
                                        IF(cc.团队 LIKE "%红杉%","红杉",IF(cc.团队 LIKE "火凤凰%","火凤凰",IF(cc.团队 LIKE "神龙家族%","神龙",IF(cc.团队 LIKE "金狮%","金狮",IF(cc.团队 LIKE "神龙-香港%","神龙香港",IF(cc.团队 LIKE "金鹏%","小虎队",IF(cc.团队 LIKE "神龙-主页运营%","神龙主页运营",IF(cc.团队 LIKE "金蝉家族%","金蝉家族",cc.团队)))))))) as 家族 
                                    FROM {0}_zqsb cc where  cc.`是否改派` = '改派' AND cc.`运单编号` is not null AND cc.日期 >= '{1}' AND cc.日期 <= '{2}'
                                ) cx WHERE cx.`币种` = '台湾'
                            GROUP BY cx.家族,cx.币种,cx.年月,cx.旬,cx.产品id
                            ) s1
                        GROUP BY s1.家族,s1.地区,s1.月份,s1.旬,s1.产品id
                        WITH ROLLUP 
                    ) s HAVING s.旬 != '合计'
                ORDER BY FIELD(s.`家族`,'神龙','火凤凰','神龙香港','Line运营','金蝉家族','金蝉项目组','APP运营','神龙主页运营','小虎队','红杉','金狮','合计'),
                        FIELD(s.`地区`,'台湾','香港','合计'),
                        FIELD(s.`月份`, DATE_FORMAT(curdate(),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 1 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 2 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 3 MONTH),'%Y%m'), DATE_FORMAT(DATE_SUB(curdate(), INTERVAL 4 MONTH),'%Y%m'),'合计'),
                        FIELD(s.`旬`,'上旬','中旬','下旬','合计'),
                        FIELD(s.`产品id`,'合计'),
                        s.总订单 DESC;'''.format(team, month_last, month_yesterday)
        df21 = pd.read_sql_query(sql=sql21, con=self.engine1)
        listT.append(df21)

        today = datetime.date.today().strftime('%Y.%m.%d')
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾', '产品整月香港', '产品分旬台湾', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾']
        print('正在将物流品类写入excel…………')
        file_path = 'F:\\输出文件\\{} {} 物流品类-签收率.xlsx'.format(today, match[team])
        # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # listT[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
        # listT[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
        # listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
        # listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            listT[0].to_excel(excel_writer=writer, sheet_name=sheet_name[0], index=False)
            listT[1].to_excel(excel_writer=writer, sheet_name=sheet_name[1], index=False)
            listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
            listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_总_品类_物流_两月签收率')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')

        print('正在将品类分旬写入excel…………')
        file_path = 'F:\\输出文件\\{} {} 品类分旬-签收率.xlsx'.format(today, match[team])
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾', '产品整月香港', '产品分旬台湾', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾']
        # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
        # listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            listT[2].to_excel(excel_writer=writer, sheet_name=sheet_name[2], index=False)
            listT[3].to_excel(excel_writer=writer, sheet_name=sheet_name[3], index=False)
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_品类直发分旬签收率')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')

        print('正在将产品写入excel…………')
        file_path = 'F:\\输出文件\\{} {} 产品明细-签收率.xlsx'.format(today, match[team])
        sheet_name = ['物流分类', '物流分旬', '一级分旬', '二级分旬', '产品整月台湾', '产品分旬台湾', '产品整月香港', '产品分旬香港', '产品月_直发台湾', '产品旬_直发台湾', '产品月_改派台湾', '产品旬_改派台湾']
        # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
        # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
        # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
        # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
        # listT[4].to_excel(excel_writer=writer, sheet_name=sheet_name[4], index=False)
        # listT[5].to_excel(excel_writer=writer, sheet_name=sheet_name[5], index=False)
        # listT[6].to_excel(excel_writer=writer, sheet_name=sheet_name[6], index=False)
        # listT[7].to_excel(excel_writer=writer, sheet_name=sheet_name[7], index=False)
        # listT[8].to_excel(excel_writer=writer, sheet_name=sheet_name[8], index=False)
        # listT[9].to_excel(excel_writer=writer, sheet_name=sheet_name[9], index=False)
        # listT[10].to_excel(excel_writer=writer, sheet_name=sheet_name[10], index=False)
        # listT[11].to_excel(excel_writer=writer, sheet_name=sheet_name[11], index=False)
        # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
        #     del book['Sheet1']
        # writer.save()
        # writer.close()
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            listT[4].to_excel(excel_writer=writer, sheet_name=sheet_name[4], index=False)
            listT[5].to_excel(excel_writer=writer, sheet_name=sheet_name[5], index=False)
            listT[6].to_excel(excel_writer=writer, sheet_name=sheet_name[6], index=False)
            listT[7].to_excel(excel_writer=writer, sheet_name=sheet_name[7], index=False)
            listT[8].to_excel(excel_writer=writer, sheet_name=sheet_name[8], index=False)
            listT[9].to_excel(excel_writer=writer, sheet_name=sheet_name[9], index=False)
            listT[10].to_excel(excel_writer=writer, sheet_name=sheet_name[10], index=False)
            listT[11].to_excel(excel_writer=writer, sheet_name=sheet_name[11], index=False)
        try:
            print('正在运行' + match[team] + '表宏…………')
            app = xlwings.App(visible=False, add_book=False)  # 运行宏调整
            app.display_alerts = False
            wbsht = app.books.open('E:/桌面文件/新版-格式转换(python表).xlsm')
            wbsht1 = app.books.open(file_path)
            wbsht.macro('gat_产品签收率_总')()
            wbsht1.save()
            wbsht1.close()
            wbsht.close()
            app.quit()
        except Exception as e:
            print('运行失败：', str(Exception) + str(e))
        print('----已写入excel ')

if __name__ == '__main__':
    m = QueryUpdate()
    start: datetime = datetime.datetime.now()
    match1 = {'gat': '港台'}
    team = 'gat'
    print('开始时间：', datetime.datetime.now())
    '''  
        -----------------------------------------------手动导入状态运行（一）-----------------------------------------
        初始化配置>>> 
        1、dim_product： 切换：总产品- 不包含直发改派；分产品- 包含直发改派 ；
        2、write：       切换：本期- 本期最近两个月的数据 ； 本期并转存-本期最近两个月的数据的转存； 上期 -上期最近两个月的数据的转存
        3、last_time：   切换：更新上传时间；
    '''
    select = 99
    handle_time = '自动0'
    if int(select) == 99:
        if handle_time == '自动':
            month_last = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            month_old = (datetime.datetime.now() - relativedelta(months=2)).strftime('%Y-%m') + '-01'     # 获取-每日-报表 各产品各团队 最近三个月的 开始的时间
            month_yesterday = datetime.datetime.now().strftime('%Y-%m-%d')
        else:
            month_last = '2023-05-01'
            month_old = '2023-05-01'                                # 获取-每日-报表 各产品各团队 最近三个月的 开始的时间
            month_yesterday = '2023-07-17'
        print('****** 签收率      起止时间：' + month_last + ' - ' + month_yesterday + ' ******')

        last_time = '2021-01-01'

        up_time = '2022-09-02'                      # 手动更新数据库 --历史总表的记录日期
        write = '本期'
        m.readFormHost(team, write, last_time, up_time)  # 更新签收表---港澳台（一）

        currency_id = '全部付款'
        m.gat_new(team, month_last, month_yesterday, currency_id)   # 获取-货到付款& 在线付款 签收率-报表
        m.qsb_new(team, month_old)                                  # 获取-每日-报表
        m.EportOrderBook(team, month_last, month_yesterday)         # 导出-总的-签收
        m.phone_report('handle', month_last, month_yesterday)       # 获取电话核实日报表 周报表 handle=手动 自定义时间（以及 物流签收率-产品前50单对比、 以及每周三 在线签收率、同产品各家族对比）

        # currency_id = '在线付款'
        # m.gat_new(team, month_last, month_yesterday, currency_id)  # 获取-在线付款 签收率-报表


    elif int(select) == 1:
        last_time = '2021-01-01'
        up_time = '2022-10-20'                      # 手动更新数据库 --历史总表的记录日期
        write = '手动更新数据库'
        m.readFormHost(team, write, last_time, up_time)  # 更新签收表---港澳台（一）

    elif int(select) == 2:
        last_time = '2021-01-01'
        up_time = '2022-10-20'
        write = '在线支付'
        m.readFormHost(team, write, last_time, up_time)  # 在线支付 读表---港澳台（一）
        # m.online_paly()                                  # 在线支付 获取

    elif int(select) == 3:
        last_time = '2021-01-01'
        up_time = '2022-10-20'
        write = '在线支付'
        m.readFormHost(team, write, last_time, up_time)  # 在线支付 读表---港澳台（一）
        write = '线付重复订单'
        m.readFormHost(team, write, last_time, up_time)  # 线付重复订单 读表---港澳台（一）



    elif int(select) == 88:
        # m.jushou()                                            #  拒收核实-查询需要的产品id
        month_last = '2023-03-01'
        month_yesterday = '2023-03-31'
        last_time = '2021-01-01'
        up_time = '2022-10-20'
        write = '地区签收率'
        m.readFormHost(team, write, last_time, up_time)  # 更新签收表---港澳台（一）
        m.address_repot(team, month_last, month_yesterday)                       #  获取-地区签收率-报表

        # 停用备用使用
        # m.EportOrder(team)       #  导出需要更新的签收表
        m.qsb_report(team, '2021-06-26', '2021-05-26')
        pass
    print('结束时间：', datetime.datetime.now())
    print('耗时：', datetime.datetime.now() - start)
    # win32api.MessageBox(0, "注意:>>>    程序运行结束， 请查看表  ！！！", "提 醒",win32con.MB_OK)