import pandas as pd
import os
import datetime
import time
import xlwings
import xlsxwriter
import math
import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel
import win32api,win32con
import win32com.client as win32
import win32com.client
from urllib.parse import urlencode      # body中的data参数是用urlencoded形式传过去的，用urlencode处理一下

from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色
from bs4 import BeautifulSoup # 抓标签里面元素的方法

# -*- coding:utf-8 -*-
class QueryTwoLower(Settings, Settings_sso):
    def __init__(self, userMobile, password, login_TmpCode,handle, proxy_id, proxy_handle):
        Settings.__init__(self)
        Settings_sso.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        # self.sso_online_cang()
        # self.bulid_file()

        if proxy_handle == '代理服务器':
            if handle == '手动':
                self.sso_online_cang_handle(login_TmpCode)
            else:
                self.sso_online_cang_auto()
        else:
            if handle == '手动':
                self.sso_online_cang_handle(login_TmpCode)
            else:
                self.sso_online_cang_auto()

        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
    def readFile(self,select):
        start: datetime = datetime.datetime.now()
        path = ''
        if select == 1:
            path = r'F:\神龙签收率\(未发货) 直发-仓库-压单\每日压单核实汇总'
        elif select == 2:
            path = r'F:\需要用到的文件\需要用到的文件\B导入头程提货单号'
        elif select == 4:
            path = r'F:\需要用到的文件\需要用到的文件\A查询导表'
        elif select == 9:
            path = r'F:\需要用到的文件\无敌'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                rq = ''
                if select == 1:
                    if '压单上传表' in filePath:
                        # excel = win32.gencache.EnsureDispatch('Excel.Application')
                        # wb = excel.Workbooks.Open(filePath)
                        # file_path = os.path.join(path, "~$ " + dir)
                        # wb.SaveAs(file_path, FileFormat=51)              # FileFormat = 51 is for .xlsx extension
                        # wb.Close()                                      # FileFormat = 56 is for .xls extension
                        # excel.Application.Quit()
                        # os.remove(filePath)
                        excel = win32.DispatchEx('Excel.Application')
                        wb = excel.Workbooks.Open(filePath)
                        file_path = os.path.join(path, "~$ " + dir)
                        wb.SaveAs(file_path, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
                        wb.Close()  # FileFormat = 56 is for .xls extension
                        excel.Application.Quit()
                        os.remove(filePath)
                    else:
                        rq = (dir.split('压单')[0]).strip()
                        rq = datetime.datetime.strptime(rq, '%Y.%m.%d')
                        rq = rq.strftime('%Y.%m.%d')
                        self._readFile(filePath, rq)                            # 工作表的   压单   信息
                        excel = win32.DispatchEx('Excel.Application')
                        wb = excel.Workbooks.Open(filePath)
                        file_path = os.path.join(path, "~$ " + dir)
                        wb.SaveAs(file_path, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
                        wb.Close()  # FileFormat = 56 is for .xls extension
                        excel.Application.Quit()
                        os.remove(filePath)
                elif select == 2:
                    if '海运' in filePath:
                        tem = '超峰国际'
                    else:
                        tem = '立邦国际'
                    self._readFile_select(filePath, rq, tem)                # 工作表的   头程物流  信息
                elif select == 4:
                    self._readFile_Yixiwudi(filePath)  # 工作表的   头程物流  信息
                    excel = win32.DispatchEx('Excel.Application')
                    wb = excel.Workbooks.Open(filePath)
                    file_path = os.path.join(path, "~$ " + dir)
                    wb.SaveAs(file_path, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
                    wb.Close()  # FileFormat = 56 is for .xls extension
                    excel.Application.Quit()
                    os.remove(filePath)
                elif select == 9:
                    rq = (dir.split('2023-')[1]).strip()
                    rq = '2023-' + rq[:5].strip()
                    # rq = datetime.datetime.strptime(rq, '%Y.%m.%d')
                    # rq = rq.strftime('%Y.%m.%d')
                    if '海运' in filePath:
                        tem = '超峰国际'
                    else:
                        tem = '立邦国际'
                    self._readFile_chaofeng(filePath, rq, tem)                # 工作表的   头程物流  信息
                    os.remove(filePath)
        print('处理耗时：', datetime.datetime.now() - start)

    # 工作表的   压单   信息
    def _readFile(self, filePath, rq):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    # print(db.columns)
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在导入查询：' + sht.name + '表； 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    if '处理时间' not in db.columns:
                        db.insert(0, '处理时间', rq)
                    db = db[['订单编号', '处理时间', '备注（压单核实是否需要）','处理人']]
                    db.rename(columns={'备注（压单核实是否需要）': '处理结果'}, inplace=True)
                    db.dropna(axis=0, subset=['处理结果'], how='any', inplace=True)
                    db.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
                    # sql = '''update 压单表 a, customer b set a.`处理时间` = b.`处理时间`, a.`处理结果` = b.`处理结果` where a.`订单编号`= b.`订单编号`;'''
                    sql = '''REPLACE INTO 压单表_已核实(订单编号,处理时间,处理结果,处理人, 记录时间) 
                            SELECT 订单编号,处理时间,处理结果,IF(处理人 = '' OR 处理人 IS NULL,'-',处理人) 处理人, NOW() 记录时间 
                            FROM customer;'''
                    # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                    print('++++成功导入：' + sht.name + '--->>>到压单表')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()

    # 超峰国际   出货   信息
    def _readFile_chaofeng(self, filePath, rq, tem):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                if sht.name == 'Worksheet':
                    try:
                        db = None
                        db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                        # print(db.columns)
                    except Exception as e:
                        print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                    if db is not None and len(db) > 0:
                        print('++++正在导入查询：' + sht.name + '表； 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        if '提货物流' not in db.columns:
                            db.insert(0, '提货物流', tem)
                        if '提货时间' not in db.columns:
                            db.insert(0, '提货时间', rq)
                        db = db[['提货时间', '提货物流', '运单号','箱号', '装箱条码']]
                        db.dropna(axis=0, subset=['运单号'], how='any', inplace=True)
                        db.to_sql('cache', con=self.engine1, index=False, if_exists='replace')
                        sql = '''REPLACE INTO gat_take_delivery_info(id, 提货单号,提货时间,提货物流,运单号,箱号, 装箱条码, 记录时间) 
                                SELECT null id,null 提货单号, 提货时间,提货物流,运单号,箱号,装箱条码, NOW() 记录时间 
                                FROM cache;'''
                        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                        print('++++成功导入：' + sht.name + ':' + rq + '--->>>到头程发货明细表')
                    else:
                        print('----------数据为空导入失败：' + sht.name)
                else:
                    print('----------不需要导入：' + sht.name)
            wb.close()
        app.quit()

    # 工作表的   头程物流  信息
    def _readFile_select(self, filePath, rq , tem):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                if sht.api.Visible == -1:
                    try:
                        db = None
                        # db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                        if tem == '立邦国际':
                            db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                        elif tem == '超峰国际':
                            db = sht.used_range.options(pd.DataFrame, header=2, numbers=int, index=False).value
                            db.columns = db.columns.droplevel(0)  # 直接将指定的层级索引drop掉
                            # db = pd.read_excel(filePath, sheet_name=sht.name)
                            db.dropna(subset=["提单号"], axis=0, inplace=True)           # 滤除指定列中含有缺失的行
                        print(db.columns)
                        print(db)
                    except Exception as e:
                        print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                    if db is not None and len(db) > 0:
                        print('++++正在导入更新：' + sht.name + '表； 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        if tem == '立邦国际':
                            if '提货物流' not in db.columns:
                                db.insert(0, '提货物流', tem)
                            db = db[['提货物流', '出貨日期', '件數', '主號', '航班號', '航班情况', '清關情況', '全清時間', '出貨日期']]
                            db.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
                            sql = '''update gat_take_delivery a, customer b 
                                    set a.`主號` = IF(b.`主號` = '' or  b.`主號` is NULL, a.`主號`, b.`主號`),
                                        a.`航班號` = IF(b.`航班號` = '' or  b.`航班號` is NULL, a.`航班號`, b.`航班號`)
                                    where a.`提货日期`= b.`出貨日期` and a.`提货物流`= b.`提货物流`;'''
                            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                        elif tem == '超峰国际':
                            if '提货物流' not in db.columns:
                                db.insert(0, '提货物流', tem)
                            if '提單號' in db.columns:
                                db.rename(columns={'提單號': '提单号', '出貨時間': '出货时间', '開船時間': '开船时间', '到達時間': '到达时间'}, inplace=True)
                            db = db[['提货物流', '出货时间', '提单号', '开船时间', '到达时间']]
                            db['开船时间'] = db['开船时间'].apply(lambda x: self._fun_time(x))       # 时间函数
                            db['到达时间'] = db['到达时间'].apply(lambda x: self._fun_time(x))
                            #   db['开船时间'] = db['开船时间'].apply(lambda x: datetime.datetime.now().strftime("%Y/") + x.split("晚上")[0] + " 23:59:59" if "晚上" in x else 0)
                            #   db['开船时间'] = db['开船时间'].apply(lambda x: (datetime.datetime.strptime(x, '%Y/%m/%d %H:%M:%S')).strftime("%Y-%m-%d %H:%M:%S"))
                            #   db['到达时间'] = db['到达时间'].apply(lambda x: datetime.datetime.now().strftime("%Y/") + x.split("早上")[0] + " 03:00:00" if "早上" in x else 0)
                            #   db['到达时间'] = db['到达时间'].apply(lambda x: (datetime.datetime.strptime(x, '%Y/%m/%d %H:%M:%S')).strftime("%Y-%m-%d %H:%M:%S"))
                            # print(db)
                            db.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
                            sql = '''update gat_take_delivery a, customer b 
                                    set a.`主號` = IF(b.`提单号` = '' or  b.`提单号` is NULL, a.`主號`, b.`提单号`),
                                        a.`出货时间` = IF(b.`开船时间` is NULL, a.`出货时间`, b.`开船时间`),
                                        a.`交货时间` = IF(b.`到达时间` is NULL, a.`交货时间`, b.`到达时间`)
                                    where a.`提货日期`= b.`出货时间` and a.`提货物流`= b.`提货物流`;'''
                            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                        print('++++成功更新：' + sht.name + '--->>>到头程物流表')
                    else:
                        print('----------数据为空导入失败：' + sht.name)
                else:
                    print('----不用导入：' + sht.name)
            wb.close()
        app.quit()
    def _fun_time(self, val):    # 时间函数
        val_time = ''
        if "晚上" in val and "12点" in val or "晚上" in val and "12點" in val:
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("晚上")[0] + " 23:59:59"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")
        elif "晚上" in val:
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("晚上")[0] + " 23:59:59"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")

        elif "早上" in val and "3点" in val:
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("早上")[0] + " 03:00:00"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")

        elif "早上" in val and ("8点" in val or "8點" in val):
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("早上")[0] + " 08:00:00"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")
        elif "早上" in val and ("11点半" in val or "11點半" in val):
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("早上")[0] + " 11:30:00"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")
        elif "早上" in val and ("12点" in val or "12點" in val):
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("早上")[0] + " 12:00:00"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")


        elif "上午" in val and ("3点" in val or "3點" in val):
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("上午")[0] + " 03:00:00"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")

        elif "上午" in val and ("8点" in val or "8點" in val):
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("上午")[0] + " 08:00:00"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")
        elif "上午" in val and ("11点" in val or "11點" in val):
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("上午")[0] + " 11:00:00"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")
        elif "上午" in val and ("11点半" in val or "11點半" in val):
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("上午")[0] + " 11:30:00"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")
        elif "上午" in val and ("12点" in val or "12點" in val):
            val_time = datetime.datetime.now().strftime("%Y/") + val.split("上午")[0] + " 12:00:00"
            # 将字符串转化为datetime
            val_time = datetime.datetime.strptime(val_time, '%Y/%m/%d %H:%M:%S')
            # 将datetime转化为字符串
            val_time = val_time.strftime("%Y-%m-%d %H:%M:%S")
        return val_time

    # 工作表的   协来运截单换渠道
    def _readFile_Yixiwudi(self, filePath):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    # print(db.columns)
                    if '运单编号' in db.columns:
                        db.rename(columns={'运单编号': '运单号'}, inplace=True)
                    elif '查件单号' in db.columns:
                        db.rename(columns={'查件单号': '运单号'}, inplace=True)
                    elif '运单号' in db.columns:
                        db.rename(columns={'运单号': '运单号'}, inplace=True)
                    elif '物流单号' in db.columns:
                        db.rename(columns={'物流单号': '运单号'}, inplace=True)
                    db = db[['运单号']]
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在查询：' + sht.name + '表； 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    orderId = list(db['运单号'])
                    max_count = len(orderId)  # 使用len()获取列表的长度，上节学的
                    if max_count > 0:
                        df = pd.DataFrame([])  # 创建空的dataframe数据框
                        dlist = []
                        for ord in orderId:
                            print(ord)
                            data = self._SearchGoods(ord)
                            if data is not None and len(data) > 0:
                                dlist.append(data)
                        dp = df.append(dlist, ignore_index=True)
                    else:
                        dp = None
                    print(dp)
                    dp.to_excel('F:\\输出文件\\新竹快递-查询{}.xlsx'.format(rq), sheet_name='查询', index=False,engine='xlsxwriter')  # Xlsx是python用来构造xlsx文件的模块，可以向excel2007+中写text，numbers，formulas 公式以及hyperlinks超链接。
                    print('查询已导出+++')
                    print('*' * 50)
                    print('++++成功导入：' + sht.name + '--->>>到压单表')
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()
    def _order_spec_Yixiwudi(self, ord):  # 进入压单检索界面
        timeStart = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
        timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
        print('+++正在查询订单信息中')
        url = r'http://gwms-v3.giikin.cn/order/order/search'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/search'}
        data = {'page': 1,
                'limit': 20,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1 and oc.billno=' + ord}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        if max_count != [] or max_count != 0:
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersdict)
            # print(data)
        else:
            data = None
            print('****** 没有信息！！！')
        return data



    # 进入 压单反馈 界面 （仓储的获取）
    def order_spec(self):  # 进入   压单反馈  界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        timeStart = ((datetime.datetime.now() + datetime.timedelta(days=1)) - relativedelta(months=3)).strftime('%Y-%m-%d')
        timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
        print('正在查询 港台 压单订单信息中......')
        url = r'http://gwms-v3.giikin.cn/order/pressure/index'
        url = r'http://gwms-v3.giikin.cn/order/pressure/index2'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/pressure/index2'}
        data = {'page': '1',
                'limit': 500,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1 and a.reason in (1,2,3,4)'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        print('++++++本次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        if max_count != [] or max_count != 0:
            # 首次查询
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            df = pd.json_normalize(ordersdict)
            # 剩余查询
            if max_count > 500:
                in_count = math.ceil(max_count/500)
                dlist = []
                n = 1
                while n < in_count:  # 这里用到了一个while循环，穿越过来的
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                    data = self._order_spec(timeStart, timeEnd, n)                     # 分页获取详情
                    dlist.append(data)
                dp = df.append(dlist, ignore_index=True)
            else:
                dp = df
            print('正在写入......')
            dp = dp[['order_number', 'goods_id', 'goods_name', 'currency_id', 'area_id', 'intime', 'purid', 'other_reason', 'buyer', 'intime', 'addtime', 'cate', 'sku', 'spec', 'yd_day', 'billno']]
            dp.columns = ['订单编号', '产品ID', '产品名称', '币种', '团队', '反馈时间', '压单原因', '其他原因', '采购员', '入库时间', '下单时间', '品类', 'SKU', '产品规格', '压单天数', '运单号']
            dp = dp[(dp['币种'].str.contains('港币|台币', na=False))]
            # print(dp)
            dp.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 压单表(订单编号,产品ID,产品名称,币种,团队, 反馈时间, 压单原因, 其他原因, 采购员, 入库时间, 下单时间, 是否下架, 下架时间, 品类, SKU, 产品规格, 压单天数, 运单号,记录时间) 
                    SELECT 订单编号,产品ID,产品名称,币种,团队, 反馈时间, 压单原因, 其他原因, 采购员, 入库时间, 下单时间, NULL as 是否下架, NULL as 下架时间, 品类, SKU, 产品规格, 压单天数, 运单号,NOW() 记录时间
                    FROM customer;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('共有 ' + str(len(dp)) + '条 成功写入数据库+++++++')


            print('正在获取 压单反馈 信息中......')
            time_path: datetime = datetime.datetime.now()
            mkpath = r"F:\神龙签收率\(未发货) 直发-仓库-压单\\" + time_path.strftime('%m.%d')
            mkpath = r"F:\神龙签收率\(未发货) 直发-仓库-压单\\" + time_path.strftime('%m') + '月'
            sql = '''SELECT s.订单编号,s.产品ID,s.产品名称,NULL SKU,NULL 产品规格,NULL 运单号,s.币种,家族 AS 团队,'港台' AS 团队2, NULL 状态,s.反馈时间,s.压单原因,s.其他原因,	s.采购员, 品类, s.入库时间,s.下单时间,s.是否下架,
                            s.下架时间,s.记录时间, s1.处理结果,s1.处理时间,NULL 备注,  DATEDIFF(curdate(),入库时间) 压单天数, DATE_FORMAT(入库时间,'%Y-%m-%d') 入库,采购异常, 处理进度, 详细处理时间,反馈内容, 压单短信发送时间
                     FROM ( SELECT * , {0}
                            FROM 压单表 g
                            WHERE g.`记录时间` >= CURDATE() 
                            # and g.是否下架 <> '已下架'
                    ) s
                    LEFT JOIN (SELECT *
                                FROM 压单表_已核实
                                WHERE id IN (SELECT MAX(id) FROM 压单表_已核实 y WHERE y.处理时间 >= DATE_SUB(CURDATE(), INTERVAL 2 month) GROUP BY 订单编号) 
                    ) s1 ON s.订单编号 = s1.订单编号
                    LEFT JOIN ( SELECT 订单编号 AS 采购异常, 处理结果 AS 处理进度, 详细处理时间, 反馈内容 FROM 采购异常 WHERE id IN (SELECT MAX(id) FROM 采购异常 GROUP BY 订单编号 ) ORDER BY id) s2 ON s.订单编号 = s2.采购异常
                    LEFT JOIN ( SELECT 订单编号 AS 订单号, 发送时间 AS 压单短信发送时间
                                FROM 短信日志_发送时间 m1
                                WHERE id IN (SELECT MAX(id) FROM 短信日志_发送时间 m WHERE m.`短信模板` = '压单短信' GROUP BY 订单编号)
                    ) s3 ON s.订单编号 = s3.订单号
                    ORDER BY 压单天数;'''.format(self.team_name)
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            sql2 = '''SELECT *  FROM 压单表 g WHERE g.`记录时间` >= CURDATE();'''
            df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
            sql3 = '''SELECT  币种, 团队,压单天数, COUNT(订单编号) AS 单量
                    FROM (
                           SELECT s.订单编号,s.产品ID,s.产品名称,NULL SKU,NULL 产品规格,NULL 运单号,s.币种,家族 AS 团队,'港台' AS 团队2, NULL 状态,s.反馈时间,s.压单原因,s.其他原因,	s.采购员, 品类, s.入库时间,s.下单时间,s.是否下架,
                                s.下架时间,s.记录时间,NULL 备注,  DATEDIFF(curdate(),入库时间) 压单天数, DATE_FORMAT(入库时间,'%Y-%m-%d') 入库
                        FROM ( SELECT * , IF(团队 LIKE '火凤凰-港%','火凤凰-港澳台',团队) AS 家族
                                FROM 压单表 g
                                WHERE g.`记录时间` >= CURDATE() and g.是否下架 <> '已下架'
                        ) s
                        ORDER BY 压单天数
                    ) ss
                    GROUP BY 币种, 团队, 压单天数
                    ORDER BY 币种, 团队, 压单天数;'''

            isExists = os.path.exists(mkpath)
            if not isExists:
                os.makedirs(mkpath)
            else:
                print(mkpath + ' 目录已存在')
            file_path = mkpath + '\\压单反馈 {0}.xlsx'.format(rq)
            # df0 = pd.DataFrame([])                                  # 创建空的dataframe数据框
            # df0.to_excel(file_path, index=False)                    # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            # writer = pd.ExcelWriter(file_path, engine='openpyxl')   # 初始化写入对象
            # book = load_workbook(file_path)                         # 可以向不同的sheet写入数据（对现有工作表的追加）
            # writer.book = book                                      # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            # df.to_excel(excel_writer=writer, sheet_name='查询', index=False)
            # df2.to_excel(excel_writer=writer, sheet_name='明细', index=False)
            # if 'Sheet1' in book.sheetnames:                         # 删除新建文档时的第一个工作表
            #     del book['Sheet1']
            # writer.save()
            # writer.close()
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(excel_writer=writer, sheet_name='查询', index=False)
                df2.to_excel(excel_writer=writer, sheet_name='明细', index=False)
            print('输出成功......')
            print('*' * 50)
        else:
            print('****** 没有新增的改派订单！！！')
            return None
        print('*' * 50)
    def _order_spec(self, timeStart, timeEnd, n):  # 进入压单检索界面
        print('+++正在查询订单信息中')
        url = r'http://gwms-v3.giikin.cn/order/pressure/index2'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/pressure/index2'}
        data = {'page': n,
                'limit': 500,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1 and a.reason in (1,2,3,4)'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        if max_count != [] or max_count != 0:
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
                sso = Settings_sso()
                sso.send_dingtalk_message(
                    "https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e", "压单反馈-更新失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
            data = pd.json_normalize(ordersdict)
            # print(data)
        else:
            data = None
            print('****** 没有信息！！！')
        return data


    # 进入 二次改派订单 界面 （仓储的获取）
    def second_sendorder(self, handle, timeStart, timeEnd):  # 进入   压单反馈  界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        if handle == '手动':
            timeStart = timeStart
            timeEnd = timeEnd
        else:
            timeStart = ((datetime.datetime.now() + datetime.timedelta(days=1)) - relativedelta(months=3)).strftime('%Y-%m-%d')
            timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
        print('正在查询  改派原运单号： ' + timeStart + ' - ' + timeEnd + '信息中......')
        url = r'http://gwms-v3.giikin.cn/order/order/secondsendorder'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'Host': 'gwms-v3.giikin.cn',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/secondsendorder'}
        data = {'page': 1,
                'limit': 20,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1'}
        print(data)
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=urlencode(data))
        print(req.headers)
        print(req.text)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        print(req)
        max_count = req['count']
        print('++++++本次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        if max_count != [] and max_count != 0:
            in_count = math.ceil(max_count/500)
            dlist = []
            df = pd.DataFrame([])
            n = 1
            while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                print('剩余查询次数' + str(in_count - n))
                data = self._second_sendorder(timeStart, timeEnd, n)                     # 分页获取详情
                dlist.append(data)
                n = n + 1
            dp = df.append(dlist, ignore_index=True)
            print('正在写入......')
            # dp = dp[['order_number', 'goods_id', 'goods_name', 'currency_id', 'area_id', 'ydtime', 'purid', 'other_reason', 'buyer', 'intime', 'addtime', 'is_lower', 'below_time', 'cate']]
            # dp.columns = ['订单编号', '产品ID', '产品名称', '币种', '团队', '反馈时间', '压单原因', '其他原因', '采购员', '入库时间', '下单时间', '是否下架', '下架时间', '品类']
            # dp = dp[(dp['币种'].str.contains('港币|台币', na=False))]
            # print(dp)
            dp.to_sql('cache', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 压单表(订单编号,产品ID,产品名称,币种,团队, 反馈时间, 压单原因, 其他原因, 采购员, 入库时间, 下单时间, 是否下架, 下架时间, 品类, 记录时间) 
                    SELECT 订单编号,产品ID,产品名称,币种,团队, 反馈时间, 压单原因, 其他原因, 采购员, 入库时间, 下单时间, 是否下架, IF(下架时间 = '',NULL,下架时间) 下架时间, 品类, NOW() 记录时间
                    FROM customer'''
            # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('共有 ' + str(len(dp)) + '条 成功写入数据库+++++++')

            print('正在更新表总表中......')
            sql = '''update {0} a, cache b
                        set a.`改派原运单号`= b.`改派原运单号`
                     where a.`订单编号`=b.`订单编号`;'''.format('gat_order_list')
            # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('更新成功......')
            print('*' * 50)
        else:
            print('****** 没有新增的改派订单！！！')
            return None
        print('*' * 50)
    def _second_sendorder(self, timeStart, timeEnd, n):  # 进入压单检索界面
        print('+++正在查询订单信息中')
        url = r'http://gwms-v3.giikin.cn/order/order/secondsendorder'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/secondsendorder'}
        data = {'page': n,
                'limit': 500,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        if max_count != [] or max_count != 0:
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersdict)
        else:
            data = None
            print('****** 没有信息！！！')
        return data


    # 进入 已下架 界面  （仓储的获取）（一）
    def order_lower(self, timeStart, timeEnd, auto_time):  # 进入已下架界面
        start: datetime = datetime.datetime.now()
        team_whid = ['速派八股仓', '天马新竹仓', '立邦香港顺丰', '香港易速配', '天马顺丰仓', '协来运', '香港圆通仓', '神龙备货-铱熙无敌', '火凤凰备货-铱熙无敌']
        # team_whid = ['协来运']
        team_stock_type = [1, 2]
        # team_stock_type = [2]
        match = {1: 'SKU库存',
                 2: '组合库存',
                 3: '混合库存'}
        match2 = {'速派八股仓': 95,
                  '天马新竹仓': 102,
                  '立邦香港顺丰': 117,
                  '香港易速配': 134,
                  '天马顺丰仓': 204,
                  '协来运': 241,
                  '香港圆通仓': 274,
                  '神龙备货-铱熙无敌': 307,
                  '火凤凰备货-铱熙无敌': 308
                  }
        if auto_time == '自动':
            # sql = '''SELECT DISTINCT 统计时间 FROM 已下架表 d GROUP BY 统计时间 ORDER BY 统计时间 DESC'''
            # rq = pd.read_sql_query(sql=sql, con=self.engine1)
            # rq = pd.to_datetime(rq['统计时间'][0])
            #
            # begin = (rq + datetime.timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
            # begin = datetime.datetime.strptime(begin, '%Y-%m-%d %H:%M:%S')
            # end = (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M:%S')
            # end = datetime.datetime.strptime(end, '%Y-%m-%d %H:%M:%S')
            # print('****** 总起止时间：' + begin.strftime('%Y-%m-%d') + ' - ' + end.strftime('%Y-%m-%d') + ' ******')
            #
            # for i in range((end - begin).days + 1):  # 按天循环获取订单状态
            #     day = begin + datetime.timedelta(days=i)
            #     timeStart = (day - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            #     timeEnd = day.strftime('%Y-%m-%d')
            timeStart = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m') + '-01'
            timeStart = (datetime.datetime.now() - relativedelta(months=2)).strftime('%Y-%m') + '-01'
            timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
            print('正在查询日期---起止时间：' + timeStart + ' - ' + timeEnd)
            for tem in team_whid:
                if tem in ('神龙备货-铱熙无敌', '火凤凰备货-铱熙无敌'):
                    for tem_type in team_stock_type:
                        print('+++正在查询仓库： ' + tem + '；库存类型:' + match[tem_type] + ' 信息')
                        self._order_lower_info(match2[tem], tem_type, timeStart, timeEnd, tem, match[tem_type])
                else:
                    print('+++正在查询仓库： ' + tem + '；库存类型:组合库存 信息')
                    self._order_lower_info(match2[tem], 2, timeStart, timeEnd, tem, '组合库存')
        else:
            print('正在查询日期---起止时间：' + timeStart + ' - ' + timeEnd)
            for tem in team_whid:
                if tem in ('神龙备货-铱熙无敌', '火凤凰备货-铱熙无敌'):
                    for tem_type in team_stock_type:
                        print('+++正在查询仓库： ' + tem + '；库存类型:' + match[tem_type] + ' 信息')
                        self._order_lower_info(match2[tem], tem_type, timeStart, timeEnd, tem, match[tem_type])
                else:
                    print('+++正在查询仓库： ' + tem + '；库存类型:组合库存 信息')
                    self._order_lower_info(match2[tem], 2, timeStart, timeEnd, tem, '组合库存')
        print('查询耗时：', datetime.datetime.now() - start)
    # 进入 已下架 界面
    def _order_lower_info(self, tem, tem_type, timeStart, timeEnd, tem_name, type_name):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        # print('+++正在查询信息中')
        url = r'http://gwms-v3.giikin.cn/order/order/shelves'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': 1, 'limit': 500, 'startDate': timeStart + ' 09:00:00', 'endDate':  timeEnd + ' 23:59:59', 'selectStr': '1=1 and ob.whid = ' + str(tem) + ' and ob.stock_type = ' + str(tem_type)}
        proxy = '47.75.114.218:10020'  # 使用代理服务器
        # proxies = {'http': 'socks5://' + proxy, 'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print(data)
        # print(req.text)
        # print(json.loads(f'"{req.text}"'))
        # req = req.text.encode('utf-8').decode("unicode_escape")
        # print('+++已成功发送请求......')              # 转码使用
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        max_count = req['data']
        max_count2 = req['count']
        if max_count != []:
            if max_count2 > 500:
                in_count = math.ceil(max_count2/500)
                dlist = []
                df = pd.DataFrame([])
                n = 1
                while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                    data = self._order_lower_infoT(n, tem, tem_type, timeStart, timeEnd, tem_name, type_name)
                    dlist.append(data)
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                dp = df.append(dlist, ignore_index=True)
            else:
                dp = self._order_lower_infoT(1, tem, tem_type, timeStart, timeEnd, tem_name, type_name)
            # dp = pd.json_normalize(ordersDict)
            # data = dp[['order_number', 'addtime', 'billno', 'old_billno', 'goods_id', 'product_name', 'intime', 'whid', 'waill_name', 'currency_id', 'area_id', 'product_spec', 'quantity', 'ship_name', 'ship_address', 'ship_phone', 'amount', 'userId', 'in_sqs', 'count_time']]
            # data.columns = ['订单编号', '下单时间', '新运单号', '原运单号', '产品id', '商品名称', '下架时间', '仓库', '物流渠道', '币种', '团队', '商品规格', '购买数量', '收货人', '收货地址', '联系电话', '订单金额', '下架人', '获取单号结果', '统计时间']
            print(dp)
            print('>>>' + tem_name + '-' + type_name + ' <<< 查询完结！！！')
            dp.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 已下架表(订单编号,下单时间,新运单号,查件单号,原运单号, 退货单号,产品id, 商品名称, 下架时间, 仓库, 物流渠道,币种, 团队,商品规格, 购买数量, 收货人, 收货地址, 联系电话,订单金额,下架人,获取单号结果,统计时间,记录时间)
                    SELECT 订单编号,下单时间,新运单号, IF(仓库 LIKE "%天马%" AND LENGTH(新运单号) = 20, CONCAT(861, RIGHT(新运单号, 8)), IF((仓库 LIKE "%速派%" or 仓库 LIKE "%易速配%") AND (新运单号 LIKE "A%" OR 新运单号 LIKE "B%"), RIGHT(新运单号, LENGTH(新运单号) - 1), UPPER(新运单号))) 查件单号,
                           原运单号,NULL 退货单号, 产品id, 商品名称, 下架时间, 仓库, 物流渠道,币种, 团队, 商品规格, 购买数量, 收货人, 收货地址, 联系电话, 订单金额,下架人,获取单号结果,统计时间,NOW() 记录时间
                    FROM customer'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
            dp.to_excel('F:\\输出文件\\已下架 {0} {1}-{2}.xlsx'.format(tem_name, type_name, rq), sheet_name='查询', index=False, engine='xlsxwriter')
            # self.stockcompose_upload()
            # print('补充退货单号成功......')
        else:
            print('****** 没有新增的改派订单！！！')
            return None

        print('获取每日新增 龟山备货 表......')
        rq = datetime.datetime.now().strftime('%m.%d')
        sql = '''SELECT CURDATE() '序號(無用途)',NULL '訂單號長度限制: 20碼請勿使用中文）', 收货人 AS '收件人姓名(必填)長度限制: 20碼', 收货地址 AS '收件人地址(必填)中文限制: 50字', 
                                联系电话 AS '收件人電話長度限制: 15碼',商品名称 AS '託運備註中文限制: 50字', NULL '(商品別編號)勿填', 购买数量 AS '商品數量(必填)(限數字)', NULL '才積重量限數字', 
                                订单金额 AS '代收貨款限數字',NULL '指定配送日期YYYYMMDD範例: 20140220    ->2月20號', NULL '指定配送時間範例:   1   (上午 -> 09~13) 2   (下午 -> 13~17)3   (晚上 -> 17~20)',
                                订单编号 , 商品规格, 产品id AS '产品ID', NULL '原运单号', 团队,下架时间,统计时间
                        FROM 已下架表 yx
                        WHERE yx.记录时间 >= TIMESTAMP(CURDATE()) AND yx.物流渠道 = '龟山备货' AND yx.`新运单号` IS NULL;'''
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        if df is not None and len(df) > 0:
            df.to_excel('F:\\输出文件\\{} 龟山备货.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
            print('获取成功......')
        else:
            print('****** 今日无新增龟山备货数据！！！')

        print('*' * 50)
    def _order_lower_infoT(self, n, tem, tem_type, timeStart, timeEnd, tem_name, type_name):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        # print('+++正在查询信息中')
        url = r'http://gwms-v3.giikin.cn/order/order/shelves'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': n, 'limit': 500, 'startDate': timeStart + ' 09:00:00', 'endDate':  timeEnd + ' 23:59:59', 'selectStr': '1=1 and ob.whid = ' + str(tem) + ' and ob.stock_type = ' + str(tem_type)}
        proxy = '47.75.114.218:10020'  # 使用代理服务器
        # proxies = {'http': 'socks5://' + proxy, 'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print(req.text)
        # print(json.loads(f'"{req.text}"'))
        # req = req.text.encode('utf-8').decode("unicode_escape")
        # print('+++已成功发送请求......')              # 转码使用
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        max_count = req['data']
        if max_count != []:
            ordersDict = []
            try:
                for result in req['data']:              # 添加新的字典键-值对，为下面的重新赋值
                    if '协来运' in tem_name or '神龙备货-铱熙无敌' in tem_name or '火凤凰备货-铱熙无敌' in tem_name:
                        result['count_time'] = (datetime.datetime.strptime(result['intime'], '%Y-%m-%d %H:%M:%S') + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
                    else:
                        if result['intime'] > (result['intime']).split()[0] + ' 08:30:00':      # 判断修改统计时间
                            result['count_time'] = (datetime.datetime.strptime(result['intime'], '%Y-%m-%d %H:%M:%S') + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
                        else:
                            result['count_time'] = (result['intime']).split()[0]

                    # result['count_time'] = timeEnd
                    if type_name == 'SKU库存':
                        if '神龙备货-铱熙无敌' in result['whid']:
                            result['waill_name'] = '铱熙无敌备货'
                        elif '火凤凰备货-铱熙无敌' in result['whid']:
                            result['waill_name'] = '铱熙无敌备货'
                        else:
                            result['waill_name'] = '备货'
                    else:
                        if '龟山易速配' in result['whid']:
                            result['waill_name'] = '龟山'
                        if '易速配-桃园仓' in result['whid']:
                            result['waill_name'] = '易速配桃园'
                        elif '速派八股仓' in result['whid']:
                            result['waill_name'] = '速派'
                        elif '天马新竹仓' in result['whid']:
                            result['waill_name'] = '天马新竹'
                        elif '立邦香港顺丰' in result['whid']:
                            result['waill_name'] = '立邦'
                        elif '香港圆通仓' in result['whid']:
                            result['waill_name'] = '圆通'
                        elif '香港易速配' in result['whid']:
                            result['waill_name'] = '易速配'
                        elif '龟山-神龙备货' in result['whid']:
                            result['waill_name'] = '龟山备货'
                        elif '龟山-火凤凰备货' in result['whid']:
                            result['waill_name'] = '龟山备货'
                        elif '天马顺丰仓' in result['whid']:
                            result['waill_name'] = '天马顺丰'
                        elif '协来运' in result['whid']:
                            result['waill_name'] = '协来运'
                        elif '神龙备货-铱熙无敌' in result['whid']:
                            result['waill_name'] = '铱熙无敌备货'
                        elif '火凤凰备货-铱熙无敌' in result['whid']:
                            result['waill_name'] = '铱熙无敌备货'
                    # print(result)
                    ordersDict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
                sso = Settings_sso()
                sso.send_dingtalk_message(
                    "https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e", "已下架-更新失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
            data = pd.json_normalize(ordersDict)
            data = data[['order_number', 'addtime', 'billno', 'old_billno', 'goods_id', 'product_name', 'intime', 'whid', 'waill_name', 'currency_id', 'area_id', 'product_spec', 'quantity', 'ship_name', 'ship_address', 'ship_phone', 'amount', 'userId', 'in_sqs', 'count_time']]
            data.columns = ['订单编号', '下单时间', '新运单号', '原运单号', '产品id', '商品名称', '下架时间', '仓库', '物流渠道', '币种', '团队', '商品规格', '购买数量', '收货人', '收货地址', '联系电话', '订单金额', '下架人', '获取单号结果', '统计时间']
            # print(data)
        else:
            print('****** 没有新增的改派订单！！！')
            data = None
        print('*' * 50)
        return data


    # 进入 组合库存界面  补充已下架的退货单号  （仓储的获取）（二）'qyz1404039293@163.com'
    def stockcompose_upload(self):
        emailAdd = {'gat': 'jikenyin666@163.com'}
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        today = datetime.date.today().strftime('%Y.%m.%d')
        print('++++正在获取 退货单号： ......')
        sql = '''SELECT * FROM 已下架表 x WHERE x.记录时间 >= TIMESTAMP (CURDATE()) AND x.`原运单号` IS NOT NULL AND x.`仓库` = '易速配-桃园仓';'''
        df0 = pd.read_sql_query(sql=sql, con=self.engine1)
        orderId = list(df0['原运单号'])
        max_count = len(orderId)    # 使用len()获取列表的长度，上节学的
        if max_count > 0:
            df = pd.DataFrame([])
            dlist = []
            for ord in orderId:
                print(ord)
                if ";" in ord:
                    ordersdict = []
                    res = {}
                    billno = ''
                    refund_number = ''
                    for od in ord.split(';'):
                        status = 1
                        bill, refund = self._stockcompose_upload(od, status)
                        if bill == "":
                            status = 2
                            bill, refund = self._stockcompose_upload(od, status)
                        billno = billno + ';' + bill
                        refund_number = refund_number + ';' + refund
                    res['原运单号'] = billno[1:]
                    res['退货单号'] = refund_number[1:]
                    ordersdict.append(res)
                    data = pd.json_normalize(ordersdict)
                else:
                    status = 1          # 已锁定
                    data = self._stockcompose_upload_two(ord, status)
                    if data is None or len(data) == 0:
                        status = 2                          # 已使用
                        data = self._stockcompose_upload_two(ord, status)
                dlist.append(data)
            dp = df.append(dlist, ignore_index=True)
            print(dp)
            dp.to_excel('F:\\输出文件\\组合库存-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
            dp.to_sql('customer_cp', con=self.engine1, index=False, if_exists='replace')
            sql = '''update `已下架表` a, `customer_cp` b
                        set a.`退货单号`= IF(b.`退货单号` = '' OR b.`退货单号` IS NULL,NULL, b.`退货单号`)
                    WHERE a.记录时间 >= TIMESTAMP (CURDATE()) AND a.`原运单号` = b.`原运单号`;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('++++正在获取 收货人信息： ......')
        self.sso__online_auto()
        orderId = list(df0['订单编号'])
        max_count = len(orderId)  # 使用len()获取列表的长度，上节学的
        if 0 < max_count <= 500:
            ord = ','.join(orderId)
            df = pd.DataFrame([])
            dp = self._stockcompose_upload_three(ord)
            print(dp)
            print(11)
            dp.to_sql('customer_cp', con=self.engine1, index=False, if_exists='replace')
            sql = '''update `已下架表` a, `customer_cp` b
                        set a.`收货人`= IF(b.`收货人` = '' OR b.`收货人` IS NULL,NULL, b.`收货人`),
                            a.`收货地址`= IF(b.`收货地址` = '' OR b.`收货地址` IS NULL,NULL, b.`收货地址`)
                    WHERE a.记录时间 >= TIMESTAMP (CURDATE()) AND a.`订单编号` = b.`订单编号`;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('获取每日新增 桃园仓重出 表......')
            rq = datetime.datetime.now().strftime('%m.%d')
            sql = '''SELECT NULL AS 客代,原运单号 AS 原單號, 退货单号 AS 退單號, 订单编号, 收货人 AS 收件人, 联系电话 AS 收件人電話, 收货地址 AS 收件人地址, 商品名称 AS 品名, 购买数量 AS 件數, 订单金额 AS 台幣代收款, 团队
                     FROM 已下架表 yx
                     WHERE yx.记录时间 >= TIMESTAMP(CURDATE()) AND yx.仓库 = '易速配-桃园仓' AND (yx.`新运单号` IS NULL OR yx.`新运单号` = '');'''
            df = pd.read_sql_query(sql=sql, con=self.engine1)
            df.to_excel('F:\\输出文件\\组合库存-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
            if df is not None and len(df) > 0:
                file_path = r'F:\\输出文件\\{0} 桃园仓重出 {1}单.xlsx'.format(rq, str(len(df)))
                df.to_excel(file_path, sheet_name='查询', index=False, engine='xlsxwriter')
                print('正在运行宏…………')
                # 通过Win32的方式并不限制xls和xlsx（因为操作是wps在做）  https://wenku.baidu.com/view/3d298b06de36a32d7375a417866fb84ae45cc3ef.html
                # excel =win32com.client.Dispatch('Excel.Application')  # word、excel、powerpoint对应的是微软的文字、表格和演示
                excel = win32com.client.Dispatch('Ket.Application')  # wps、et、wpp对应的是金山文件、表格和演示
                excel.Visible = False  # 可视化选项
                Path = r"D:/Users/Administrator/Desktop/slgat_签收计算(ver5.24).xlsm"
                workbook = excel.Workbooks.Open(Path)
                workbook1 = excel.Workbooks.Open(file_path)
                workbook.Application.Run("'D:/Users/Administrator/Desktop/slgat_签收计算(ver5.24).xlsm'!总审核解析.A解析")
                workbook1.Save()
                excel.Quit()

                print('获取成功发送中......')
                filepath = [file_path]
                # self.e.send('{0} 桃园仓重出 {1}单.xlsx'.format(today, str(len(df))), filepath, emailAdd['gat'])
            else:
                print('****** 今日无新增 桃园仓重出 数据！！！')
    def _stockcompose_upload(self, waybill, status):                            # 进入压单检索界面
        print('+++正在查询订单信息中')
        url = r'http://gwms-v3.giikin.cn/stock/stockcompose/index'
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
            'origin': 'http://gwms-v3.giikin.cn',
            'Referer': 'http://gwms-v3.giikin.cn/order/refund/sale'}
        data = {'page': 1,
                'limit': 20,
                'selectStr': "1=1 and scb.billno= '" + waybill + "'and scb.status= '" + str(status) + "'",
                'relateNumber': None}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print(req)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        print(req['comment'])
        max_count = req['count']
        billno = ''
        refund_number = ''
        if req['data'] != [] and max_count != 0:
            for result in req['data']:
                billno = result['billno']
                refund_number = result['refund_number']
        return billno, refund_number            # 进入订单检索界面  获取收货人信息
    def _stockcompose_upload_two(self, waybill, status):  # 进入压单检索界面
        print('+++正在查询订单信息中')
        url = r'http://gwms-v3.giikin.cn/stock/stockcompose/index'
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
            'origin': 'http://gwms-v3.giikin.cn',
            'Referer': 'http://gwms-v3.giikin.cn/order/refund/sale'}
        data = {'page': 1,
                'limit': 20,
                'selectStr': "1=1 and scb.billno= '" + waybill + "'and scb.status= '" + str(status) + "'",
                'relateNumber': None}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print(req)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        print(req['comment'])
        # print(req['data'])
        max_count = req['count']
        data = None
        if req['data'] != [] and max_count != 0:
            ordersdict = []
            for result in req['data']:
                ordersdict.append(result)
            data = pd.json_normalize(ordersdict)
            data = data[['billno', 'refund_number']]
            data.columns = ['原运单号', '退货单号']
        return data
    def _stockcompose_upload_three(self, ord):          # 进入订单检索界面  获取收货人信息
        # print('......正在查询信息中......')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/deletedOrder'}
        data = {'page': 1, 'pageSize': 500, 'order_number': ord, 'shippingNumber': None, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None,'currencyId': None,
                'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '', 'warehouse': None, 'isEmptyWayBillNumber': None,  'logisticsStatus': None,
                'orderStatus': None, 'tuan': None,  'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None,  'chooser_id': None, 'service_id': None,
                'autoVerifyStatus': None, 'shipZip': None, 'remark': None, 'shipState': None,'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None,
                'order': None, 'sortField': None, 'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None,'timeStart': None, 'timeEnd': None}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print('......已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        try:
            for result in req['data']['list']:
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersdict)
        df = data[['orderNumber', 'shipInfo.shipName', 'shipInfo.shipAddress']]
        df.columns = ['订单编号', '收货人', '收货地址']
        print('******本批次查询成功')
        return df

    # 进入组合库存查询界面 ？？？？
    def gp_order_stockcompose(self):
        print('正在获取 改派未发货…………')
        today = datetime.date.today().strftime('%Y.%m.%d')
        sql = '''SELECT *
                 FROM 已下架表 x
                 WHERE x.记录时间 >= TIMESTAMP ( CURDATE( ) );'''

        df = pd.read_sql_query(sql=sql, con=self.engine1)
        df = df.loc[df["币种"] == "台币"]
        df.to_sql('cache', con=self.engine1, index=False, if_exists='replace')

        print('正在查询 改派未发货…………')
        sql = '''SELECT 订单编号 FROM {0};'''.format('cache')
        ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
        if ordersDict.empty:
            print('无需要更新订单信息！！！')
            # sys.exit()
            return
        orderId = list(ordersDict['订单编号'])
        max_count = len(orderId)    # 使用len()获取列表的长度，上节学的
        n = 0
        df = pd.DataFrame([])  # 创建空的dataframe数据框
        dlist = []
        while n < max_count:        # 这里用到了一个while循环，穿越过来的
            ord = ','.join(orderId[n:n + 500])
            n = n + 500
            data =self._gp_order_stockcompose(ord)
            if data is not None and len(data) > 0:
                dlist.append(data)
        dp = df.append(dlist, ignore_index=True)
        dp.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')

        print('正在更新 改派未发货......')
        sql = '''update `cache` a, `cache_cp` b
                        set a.`系统订单状态`= b.`orderStatus`,
                            a.`系统物流状态`= b.`logisticsStatus`
                where a.`订单编号`=b.`orderNumber`;'''.format('cache')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        print('正在导出 改派未发货…………')
        sql = '''SELECT * FROM cache;'''.format('team')
        dt = pd.read_sql_query(sql=sql, con=self.engine1)
        file_path = 'F:\\神龙签收率\\(未发货) 改派-物流\\{} 改派未发货.xlsx'.format(today)
        dt.to_excel(file_path, sheet_name='台湾', index=False, engine='xlsxwriter')
        print('----已写入excel ')
    # 进入组合库存查询界面（新后台的获取） ？？？？
    def _gp_order_stockcompose(self, ord):  # 进入订单检索界面
        print('+++正在查询 信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500, 'orderPrefix': ord, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'shippingNumber': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None,'type': None, 'collId': None, 'isClone': None, 'currencyId': None,
                'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '','warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None,
                'orderStatus': None, 'tuan': None,  'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None,
                'service_id': None, 'autoVerifyStatus': None, 'shipZip': None, 'remark': None, 'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None,
                'estimateWeightEnd': None, 'order': None, 'sortField': None, 'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        try:
            for result in req['data']['list']:
                # print(result)
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersdict)
        df = data[['orderNumber', 'orderStatus', 'wayBillNumber', 'logisticsName', 'logisticsStatus', 'warehouse', 'update_time']]
        # print('++++++本批次查询成功+++++++')
        print('*' * 50)
        # print(df)
        return df


    # 查询改派无运单好（仓储的获取）  停用
    def get_billno_res(self):  # 进入仓储界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        timeStart = ((datetime.datetime.now() + datetime.timedelta(days=1)) - relativedelta(months=2)).strftime('%Y-%m-%d')
        timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
        print('正在查询 港台 改派无运单号......')
        url = r'http://gwms-v3.giikin.cn/order/order/secondsendbillnonone'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': '1',
                'limit': 500,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        print('++++++本次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        if max_count != [] or max_count != 0:
            # 首次查询
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            df = pd.json_normalize(ordersdict)
            # 剩余查询
            if max_count > 500:
                in_count = math.ceil(max_count/500)
                dlist = []
                n = 1
                while n < in_count:  # 这里用到了一个while循环，穿越过来的
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                    data = self._order_spec(timeStart, timeEnd, n)                     # 分页获取详情
                    dlist.append(data)
                dp = df.append(dlist, ignore_index=True)
            else:
                dp = df
            print('正在写入......')
            dp.to_excel('F:\\输出文件\\改派无运单号 {}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
            # dp = dp[['order_number', 'goods_id', 'goods_name', 'currency_id', 'area_id', 'ydtime', 'purid', 'other_reason', 'buyer', 'intime', 'addtime', 'is_lower', 'below_time', 'cate']]
            # dp.columns = ['订单编号', '产品ID', '产品名称', '币种', '团队', '反馈时间', '压单原因', '其他原因', '采购员', '入库时间', '下单时间', '是否下架', '下架时间', '品类']
            # dp = dp[(dp['币种'].str.contains('港币|台币', na=False))]
            # print(dp)
            # dp.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO 压单表(订单编号,产品ID,产品名称,币种,团队, 反馈时间, 压单原因, 其他原因, 采购员, 入库时间, 下单时间, 是否下架, 下架时间,记录时间) 
                    SELECT 订单编号,产品ID,产品名称,币种,团队, 反馈时间, 压单原因, 其他原因, 采购员, 入库时间, 下单时间, 是否下架, IF(下架时间 = '',NULL,下架时间) 下架时间, NOW() 记录时间
                    FROM customer'''
            # pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('共有 ' + str(len(dp)) + '条 成功写入数据库+++++++')

            # print('正在获取 压单反馈 信息中......')
            # time_path: datetime = datetime.datetime.now()
            # mkpath = r"F:\神龙签收率\(未发货) 直发-仓库-压单\\" + time_path.strftime('%m.%d')
            # sql = '''SELECT s.*,s1.处理结果,s1.处理时间
            #         FROM ( SELECT * FROM 压单表 g WHERE g.`记录时间` >= CURDATE() and g.是否下架 <> '已下架'
            #         ) s
            #         LEFT JOIN 压单表_已核实 s1 ON s.订单编号 = s1.订单编号;'''
            # df = pd.read_sql_query(sql=sql, con=self.engine1)
            # isExists = os.path.exists(mkpath)
            # if not isExists:
            #     os.makedirs(mkpath)
            # else:
            #     print(mkpath + ' 目录已存在')
            # file_path = mkpath + '\\压单反馈 {0}.xlsx'.format(rq)
            # df.to_excel(file_path, sheet_name='查询', index=False, engine='xlsxwriter')
            # print('输出成功......')
            # print('*' * 50)
        else:
            print('****** 没有新增的改派订单！！！')
            return None
        print('*' * 50)
    # 进入仓储检索界面                停用
    def _get_billno_res(self, timeStart, timeEnd, n):
        print('+++正在查询订单信息中')
        url = r'http://gwms-v3.giikin.cn/order/order/secondsendbillnonone'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': '1',
                'limit': 500,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        if max_count != [] or max_count != 0:
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersdict)
            # print(data)
        else:
            data = None
            print('****** 没有信息！！！')
        return data


    # 创建每日文件
    def bulid_file(self):
        print('正在生成每日新文件夹......')
        time_path: datetime = datetime.datetime.now()
        mkpath = "F:\\神龙签收率\\" + time_path.strftime('%m.%d')
        isExists = os.path.exists(mkpath)
        if not isExists:
            os.makedirs(mkpath)
            os.makedirs(mkpath + "\\产品签收率")
            os.makedirs(mkpath + "\\产品签收率\\直发&改派")
            os.makedirs(mkpath + "\\导运单号&提货时间")
            os.makedirs(mkpath + "\\导状态")
            os.makedirs(mkpath + "\\签收率")
            os.makedirs(mkpath + "\\物流签收率")
            os.makedirs(mkpath + "\\物流表")
            print('创建成功')
            file_path = mkpath + '\\导运单号&提货时间\\{} 龟山 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path1 = mkpath + '\\导运单号&提货时间\\{} 圆通 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path2 = mkpath + '\\导运单号&提货时间\\{} 立邦 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path3 = mkpath + '\\导运单号&提货时间\\{} 天马 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path4 = mkpath + '\\导运单号&提货时间\\{} 速派 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path5 = mkpath + '\\导运单号&提货时间\\{} 协来运普货 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path50 = mkpath + '\\导运单号&提货时间\\{} 协来运特货 无运单号.xlsx'.format(time_path.strftime('%m.%d'))
            df = pd.DataFrame([['', '']], columns=['订单编号', '物流单号'])
            df.to_excel(file_path, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path1, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path2, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path3, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path4, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path5, sheet_name='查询', index=False, engine='xlsxwriter')
            df.to_excel(file_path50, sheet_name='查询', index=False, engine='xlsxwriter')

            file_path31 = mkpath + '\\导运单号&提货时间\\{} 天马 换新运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path32 = mkpath + '\\导运单号&提货时间\\{} 协来运 换新运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path33 = mkpath + '\\导运单号&提货时间\\{} 立邦 换新运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path34 = mkpath + '\\导运单号&提货时间\\{} 速派 换新运单号.xlsx'.format(time_path.strftime('%m.%d'))
            file_path35 = mkpath + '\\导运单号&提货时间\\{} 龟山 换新运单号.xlsx'.format(time_path.strftime('%m.%d'))
            df2 = pd.DataFrame([['', '', '']], columns=['订单编号', '旧运单号', '新运单号'])
            df2.to_excel(file_path31, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path32, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path33, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path34, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path35, sheet_name='查询', index=False, engine='xlsxwriter')

            file_path91 = mkpath + '\\导运单号&提货时间\\{} 导入提货时间 龟山.xlsx'.format(time_path.strftime('%m.%d'))
            file_path92 = mkpath + '\\导运单号&提货时间\\{} 导入提货时间 天马.xlsx'.format(time_path.strftime('%m.%d'))
            file_path93 = mkpath + '\\导运单号&提货时间\\{} 导入提货时间 速派.xlsx'.format(time_path.strftime('%m.%d'))
            file_path94 = mkpath + '\\导运单号&提货时间\\{} 导入提货时间 协来运.xlsx'.format(time_path.strftime('%m.%d'))
            file_path95 = mkpath + '\\导运单号&提货时间\\{} 导入提货时间 立邦.xlsx'.format(time_path.strftime('%m.%d'))
            df2 = pd.DataFrame([['', '', '']], columns=['订单号', '物流单号', '提货时间'])
            df2.to_excel(file_path91, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path92, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path93, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path94, sheet_name='查询', index=False, engine='xlsxwriter')
            df2.to_excel(file_path95, sheet_name='查询', index=False, engine='xlsxwriter')
            print('创建文件')
        else:
            print(mkpath + ' 目录已存在')
        print('*' * 50)




    # 进入 头程物流跟踪 界面 （仓储的获取- 明细）
    def get_take_delivery_list(self):
        print('+++正在查询头程物流信息中')
        sql = '''SELECT id,`订单编号`  FROM {0} sl WHERE sl.`日期` = '{1}';'''.format('gat_order_list', last_month)
        ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
        if ordersDict.empty:
            print('无需要更新订单信息！！！')
            return
        print(ordersDict['订单编号'][0])
        orderId = list(ordersDict['订单编号'])
        max_count = len(orderId)  # 使用len()获取列表的长度，上节学的
        n = 0
        while n < max_count:  # 这里用到了一个while循环，穿越过来的
            ord = ', '.join(orderId[n:n + 500])
            n = n + 500
            self.orderInfoQuery(ord)
        print('单日查询耗时：', datetime.datetime.now() - start)


        timeStart = (datetime.datetime.now() - datetime.timedelta(days=10)).strftime('%Y-%m-%d')
        timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
        url = r'http://gwms-v3.giikin.cn/order/delivery/firstLegTrace'
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
            'origin': 'http://gwms-v3.giikin.cn',
            'Referer': 'http://gwms-v3.giikin.cn/order/delivery/takeDeliveryRegister?id=8755'}
        data = {'page': '1',
                'limit': 100,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1 and a.country= "TW"'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print(req)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        max_count = req['count']
        if max_count != [] or max_count != 0:
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersdict)
            # print(data)
            df = data[['id', 'take_delivery_no', 'take_delivery_date', 'take_delivery_company', 'take_delivery_company_id', 'transport_mode', 'product_type', 'transport_type',
                       'batch', 'barcode', 'country', 'boxCount', 'analy', 'deliverytime', 'send_first_logistics_comment', 'uptime']]
            df.columns = ['id', '提货单号', '提货时间', '提货物流', '提货物流id', '运输方式', '货物类型', '运输公司',
                          '运输班次', '箱号', '线路', '箱数', '统计', '交货时间', '报关资料发送结果', '更新时间']

            print('共有 ' + str(len(df)) + '条 正在写入......')
            df.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            # df.to_excel('F:\\输出文件\\{0}-查询{1}.xlsx'.format(match[team], rq), sheet_name='查询', index=False,engine='xlsxwriter')
            sql = '''REPLACE INTO gat_take_delivery(id,提货单号,提货时间,提货日期,提货物流,提货物流id,运输方式,货物类型,运输公司,运输班次,箱号,线路,箱数,统计,出货时间, 交货时间,报关资料发送结果,更新时间, 主號,航班號,记录时间)
                     SELECT id,提货单号,提货时间,DATE_FORMAT(提货时间,'%Y-%m-%d') 提货日期,提货物流,提货物流id,运输方式,货物类型,IF(运输公司 = '',NULL,运输公司) 运输公司,IF(运输班次 = '',NULL,运输班次) 运输班次,箱号,线路,箱数,统计,
                            NULL 出货时间, IF(交货时间 = '',NULL,交货时间) 交货时间,报关资料发送结果,更新时间,NULL 主號,NULL 航班號,NOW() 记录时间
                    FROM customer;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
        else:
            df = None
            print('****** 没有信息！！！')
        return df

    # 进入 头程物流跟踪 界面 （仓储的获取）
    def get_take_delivery_no(self):
        print('+++正在查询头程物流信息中')
        timeStart = (datetime.datetime.now() - datetime.timedelta(days=10)).strftime('%Y-%m-%d')
        timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
        url = r'http://gwms-v3.giikin.cn/order/delivery/firstLegTrace'
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
            'origin': 'http://gwms-v3.giikin.cn',
            'Referer': 'http://gwms-v3.giikin.cn/order/delivery/takeDeliveryRegister?id=8755'}
        data = {'page': '1',
                'limit': 100,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': '1=1 and a.country= "TW"'}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print(req)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        max_count = req['count']
        if max_count != [] or max_count != 0:
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersdict)
            # print(data)
            df = data[['id', 'take_delivery_no', 'take_delivery_date', 'take_delivery_company', 'take_delivery_company_id', 'transport_mode', 'product_type', 'transport_type',
                       'batch', 'barcode', 'country', 'boxCount', 'analy', 'deliverytime', 'send_first_logistics_comment', 'uptime']]
            df.columns = ['id', '提货单号', '提货时间', '提货物流', '提货物流id', '运输方式', '货物类型', '运输公司',
                          '运输班次', '箱号', '线路', '箱数', '统计', '交货时间', '报关资料发送结果', '更新时间']

            print('共有 ' + str(len(df)) + '条 正在写入......')
            df.to_sql('customer', con=self.engine1, index=False, if_exists='replace')
            # df.to_excel('F:\\输出文件\\{0}-查询{1}.xlsx'.format(match[team], rq), sheet_name='查询', index=False,engine='xlsxwriter')
            sql = '''REPLACE INTO gat_take_delivery(id,提货单号,提货时间,提货日期,提货物流,提货物流id,运输方式,货物类型,运输公司,运输班次,箱号,线路,箱数,统计,出货时间, 交货时间,报关资料发送结果,更新时间, 主號,航班號,记录时间)
                     SELECT id,提货单号,提货时间,DATE_FORMAT(提货时间,'%Y-%m-%d') 提货日期,提货物流,提货物流id,运输方式,货物类型,IF(运输公司 = '',NULL,运输公司) 运输公司,IF(运输班次 = '',NULL,运输班次) 运输班次,箱号,线路,箱数,统计,
                            NULL 出货时间, IF(交货时间 = '',NULL,交货时间) 交货时间,报关资料发送结果,更新时间,NULL 主號,NULL 航班號,NOW() 记录时间
                    FROM customer;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('写入成功......')
        else:
            df = None
            print('****** 没有信息！！！')
        return df
    # 进入 头程检索界面
    def _get_take_delivery_no(self):
        timeStart = (datetime.datetime.now() - datetime.timedelta(days=10)).strftime('%Y-%m-%d')
        start = datetime.datetime.now()
        print('正在更新 头程提货单号 (立邦国际)信息…………')
        sql = '''SELECT id, 提货单号,主號, 航班號, 提货日期  
                FROM {0} g 
                WHERE g.运输公司 IS NULL AND g.`航班號` IS NOT NULL AND g.`提货日期` >= '{1}' AND g.提货物流 = '立邦国际';'''.format('gat_take_delivery', timeStart)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        if df.empty:
            print('无需要更新订单信息！！！')
        else:
            for row in df.itertuples():
                tem = '立邦国际'
                ord_id = getattr(row, '提货单号')
                id = getattr(row, 'id')
                take_delivery_no = getattr(row, '主號')
                batch = getattr(row, '航班號')
                departed_time = ''
                arrived_time = ''
                self._upload_take_delivery_no(ord_id, id, take_delivery_no, batch, tem, departed_time, arrived_time)

        print('正在更新 头程提货单号 (超峰国际)信息…………')
        sql = '''SELECT id, 提货单号,主號, 出货时间, 交货时间
                FROM {0} g 
                WHERE g.运输公司 IS NULL AND g.`主號` IS NOT NULL AND g.`提货日期` >= '{1}' AND g.提货物流 = '超峰国际';'''.format('gat_take_delivery', timeStart)
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        if df.empty:
            print('无需要更新订单信息！！！')
        else:
            for row in df.itertuples():
                tem = '超峰国际'
                ord_id = getattr(row, '提货单号')
                id = getattr(row, 'id')
                take_delivery_no = getattr(row, '主號')
                batch = ''
                departed_time = getattr(row, '出货时间')
                arrived_time = getattr(row, '交货时间')
                self._upload_take_delivery_no(ord_id, id, take_delivery_no, batch, tem, departed_time, arrived_time)
        print('单次更新耗时：', datetime.datetime.now() - start)
    # 更新 头程提货单号
    def _upload_take_delivery_no(self, ord_id, id, take_delivery_no, batch, tem, departed_time, arrived_time):
        print('正在更新中')
        url = r'http://gwms-v3.giikin.cn/order/delivery/takedeliveryregister'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/delivery/takeDeliveryRegister?id=8755'}
        data = ''
        if tem == '立邦国际':
            transport_type = batch[:2]
            print('提货单号：' + ord_id, 'id：' + str(id), ';主號：' + take_delivery_no, '；航班號：'+transport_type, '；航班信息：'+batch)
            data = {'id': id,
                    'take_delivery_no': take_delivery_no,
                    'transport_type': transport_type,
                    'batch': batch,
                    'departed_time': '',
                    'departed_place': '',
                    'arrived_time': '',
                    'arrived_place': '',
                    'product_type': ''
                    }
        elif tem == '超峰国际':
            print('提货单号：' + ord_id, 'id：' + str(id), ';主號：' + take_delivery_no, '；开船时间：' + str(departed_time), '；到达时间：' + str(arrived_time))
            data = {'id': id,
                    'take_delivery_no': take_delivery_no,
                    'transport_type': '',
                    'batch': '',
                    'departed_time': departed_time,
                    'departed_place': '',
                    'arrived_time': arrived_time,
                    'arrived_place': '',
                    'product_type': ''
                    }
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print(req)
        # print(req.text)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        # print(req['code'])
        # print(req['comment'])
        val = req['comment']
        if val == 'success':
            print('头程物流 更新成功！！！')
        else:
            print('头程物流 更新失败！！！')

    # 驳回发货-改派使用
    def delivery_rejectDelivery(self):
        print(datetime.datetime.now())
        rq = datetime.datetime.now().strftime('%Y-%m-%d')
        file_data = input('驳回发货:  请输入 运单编号,多个以逗号分割:  ')
        # file_data ='7464302731,7464320043,7464314104'
        if ',' in file_data:  # 定义一行数据
            dlist = file_data.split(',')
        else:
            dlist = [file_data]
        ordersdict = []
        if len(dlist) > 40:
            n = 0
            while n < len(dlist) - 40:  # 这里用到了一个while循环，穿越过来的
                n = n + 40
                dl = dlist[n:n + 40]    # 分批查询
                print(n)
                for d in dl:
                    result = self._delivery_rejectDelivery(d)
                    ordersdict.append(result)
                print('暂停30秒')
                time.sleep(30)             # 暂停30秒
        else:
            for d in dlist:
                result = self._delivery_rejectDelivery(d)
                ordersdict.append(result)
        df = pd.json_normalize(ordersdict)
        df.to_excel('F:\\输出文件\\{0} 驳回发货-查询.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
        print('写入完成++++++')
    def _delivery_rejectDelivery(self, waybill):
        print('+++正在驳回中')
        url = r'http://gwms-v3.giikin.cn/order/delivery/rejectDelivery'
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
            'Host': 'gwms-v3.giikin.cn',
            'origin': 'http://gwms-v3.giikin.cn',
            'Referer': 'http://gwms-v3.giikin.cn/order/delivery/rejectdelivery'}
        data = {'orderno': waybill}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print(req)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型 或者 str字符串  数据转换为dict字典
        print(req)
        print(req['comment'])
        ordersdict = {}
        ordersdict[str(waybill)] = req['comment']
        return ordersdict


    # 查询出库更新 以订单编号（仓储的获取）
    def gwms_chuku(self,timeStart, timeEnd):  # 进入压单检索界面
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        print('正在获取需 更新订单信息…………')
        sql = '''SELECT 订单编号 FROM 查询 g;'''.format('', '')
        df = pd.read_sql_query(sql=sql, con=self.engine1)
        orderId = list(df['订单编号'])
        max_count = len(orderId)
        print('++++++本批次查询成功;  总计： ' + str(max_count) + ' 条信息+++++++')  # 获取总单量
        if max_count > 500:
            ord = "', '".join(orderId[0:500])
            df = self._gwms_chuku(ord, timeStart, timeEnd)
            dlist = []
            n = 0
            while n < max_count - 500:  # 这里用到了一个while循环，穿越过来的
                n = n + 500
                ord = "', '".join(orderId[n:n + 500])
                data = self._gwms_chuku(ord, timeStart, timeEnd)
                print(data)
                if data is not None and len(data) > 0:
                    dlist.append(data)
            dp = df.append(dlist, ignore_index=True)
        else:
            print(22)
            ord = "','".join(orderId[0:max_count])
            dp = self._gwms_chuku(ord, timeStart, timeEnd)
        if dp is None or len(dp) == 0:
            print('查询为空，不需更新+++')
        else:
            print('正在写入临时缓存表......')
            print(dp)
            print('查询已导出+++')
            dp.to_excel('F:\\输出文件\\出库-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
        # 进入运单扫描导出 界面
    def _gwms_chuku(self, ord ,timeStart, timeEnd):
        print('+++正在查询订单信息中')
        # timeStart = ((datetime.datetime.now() + datetime.timedelta(days=1)) - relativedelta(months=2)).strftime('%Y-%m-%d')
        # timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')
        url = r'http://gwms-v3.giikin.cn/order/delivery/deliverylog'
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
            'origin': 'http://gwms-v3.giikin.cn',
            'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': 1,
                'limit': 500,
                'startDate': timeStart + ' 00:00:00',
                'endDate': timeEnd + ' 23:59:59',
                'selectStr': "1=1 and bs.order_number in ('" + ord + "')"
                }
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型 或者 str字符串  数据转换为dict字典
        max_count = req['count']
        print(max_count)
        if max_count != [] and max_count != 0:
            ordersdict = []
            try:
                for result in req['data']:
                    ordersdict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersdict)
        else:
            data = None
            print('****** 没有信息！！！')
        return data

    # 仓储库存查询 （仓储的获取）
    def inquire_info(self, n, tem, tem_type, timeStart, timeEnd, tem_name, type_name):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        # print('+++正在查询信息中')
        url = r'http://gwms-v3.giikin.cn/order/order/shelves'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'http://gwms-v3.giikin.cn',
                    'Referer': 'http://gwms-v3.giikin.cn/order/order/shelves'}
        data = {'page': n, 'limit': 500, 'startDate': timeStart + ' 09:00:00', 'endDate':  timeEnd + ' 23:59:59', 'selectStr': '1=1 and ob.whid = ' + str(tem) + ' and ob.stock_type = ' + str(tem_type)}
        proxy = '47.75.114.218:10020'  # 使用代理服务器
        # proxies = {'http': 'socks5://' + proxy, 'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print(req.text)
        # print(json.loads(f'"{req.text}"'))
        # req = req.text.encode('utf-8').decode("unicode_escape")
        # print('+++已成功发送请求......')              # 转码使用
        req = json.loads(req.text)                           # json类型 或者 str字符串  数据转换为dict字典
        # print(req)
        max_count = req['data']
        if max_count != []:
            ordersDict = []
            try:
                for result in req['data']:              # 添加新的字典键-值对，为下面的重新赋值
                    if '协来运' in tem_name or '神龙备货-铱熙无敌' in tem_name or '火凤凰备货-铱熙无敌' in tem_name:
                        result['count_time'] = (datetime.datetime.strptime(result['intime'], '%Y-%m-%d %H:%M:%S') + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
                    else:
                        if result['intime'] > (result['intime']).split()[0] + ' 08:30:00':      # 判断修改统计时间
                            result['count_time'] = (datetime.datetime.strptime(result['intime'], '%Y-%m-%d %H:%M:%S') + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
                        else:
                            result['count_time'] = (result['intime']).split()[0]

                    # result['count_time'] = timeEnd
                    if type_name == 'SKU库存':
                        if '神龙备货-铱熙无敌' in result['whid']:
                            result['waill_name'] = '铱熙无敌备货'
                        elif '火凤凰备货-铱熙无敌' in result['whid']:
                            result['waill_name'] = '铱熙无敌备货'
                        else:
                            result['waill_name'] = '备货'
                    else:
                        if '龟山易速配' in result['whid']:
                            result['waill_name'] = '龟山'
                        if '易速配-桃园仓' in result['whid']:
                            result['waill_name'] = '易速配桃园'
                        elif '速派八股仓' in result['whid']:
                            result['waill_name'] = '速派'
                        elif '天马新竹仓' in result['whid']:
                            result['waill_name'] = '天马新竹'
                        elif '立邦香港顺丰' in result['whid']:
                            result['waill_name'] = '立邦'
                        elif '香港圆通仓' in result['whid']:
                            result['waill_name'] = '圆通'
                        elif '香港易速配' in result['whid']:
                            result['waill_name'] = '易速配'
                        elif '龟山-神龙备货' in result['whid']:
                            result['waill_name'] = '龟山备货'
                        elif '龟山-火凤凰备货' in result['whid']:
                            result['waill_name'] = '龟山备货'
                        elif '天马顺丰仓' in result['whid']:
                            result['waill_name'] = '天马顺丰'
                        elif '协来运' in result['whid']:
                            result['waill_name'] = '协来运'
                        elif '神龙备货-铱熙无敌' in result['whid']:
                            result['waill_name'] = '铱熙无敌备货'
                        elif '火凤凰备货-铱熙无敌' in result['whid']:
                            result['waill_name'] = '铱熙无敌备货'
                    # print(result)
                    ordersDict.append(result)
            except Exception as e:
                print('转化失败： 重新获取中', str(Exception) + str(e))
            data = pd.json_normalize(ordersDict)
            data = data[['order_number', 'addtime', 'billno', 'old_billno', 'goods_id', 'product_name', 'intime', 'whid', 'waill_name', 'currency_id', 'area_id', 'product_spec', 'quantity', 'ship_name', 'ship_address', 'ship_phone', 'amount', 'userId', 'in_sqs', 'count_time']]
            data.columns = ['订单编号', '下单时间', '新运单号', '原运单号', '产品id', '商品名称', '下架时间', '仓库', '物流渠道', '币种', '团队', '商品规格', '购买数量', '收货人', '收货地址', '联系电话', '订单金额', '下架人', '获取单号结果', '统计时间']
            # print(data)
        else:
            print('****** 没有新增的改派订单！！！')
            data = None
        print('*' * 50)
        return data


if __name__ == '__main__':
    proxy_handle = '代理服务器0'
    proxy_id = '192.168.13.89:37466'  # 输入代理服务器节点和端口
    handle = '手0动'
    login_TmpCode = '0bd57ce215513982b1a984d363469e30'  # 输入登录口令Tkoen
    m = QueryTwoLower('+86-18538110674', 'qyz04163510.','1dd6113668f83ab5b79797b87b8bcc41','手0动', proxy_id, proxy_handle)
    # m.bulid_file()
    start: datetime = datetime.datetime.now()
    match1 = {'gat': '港台', 'gat_order_list': '港台', 'slsc': '品牌'}

    # -----------------------------------------------手动设置时间；若无法查询，切换代理和直连的网络-----------------------------------------

    # m.order_lower('2022-02-17', '2022-02-18', '自动')   # 已下架
    select = 9
    if select == 1:
        # m.readFile(select)            # 上传每日压单核实结果
        m.order_spec()                # 压单反馈  （备注（压单核实是否需要））

    elif select == 2:
        m.readFile(select)             # 读取头程时效表
        m._get_take_delivery_no()      # 头程导入提货单号

    elif select == 3:
        m.stockcompose_upload()

    elif select == 4:
        m.delivery_rejectDelivery()     # 驳回发货-改派使用

    elif select == 5:
        # m.readFile(select)  # 上传每日压单核实结果

        m.gwms_chuku('2022-10-01', '2022-12-23')
        pass
        # m.get_take_delivery_no()
        # m.readFile(select)
        # m._get_take_delivery_no()

        # m. _upload_take_delivery_no(8637, '297-82680091', 'CI', 'CI6844')

    elif select == 9:
        m.readFile(select)             # 读取头程时效表

    # m.get_billno_res()      # 改派无运单号

    #h  ttp://gwms-v3.giikin.cn/order/delivery/batchImportLegDeliveryBoxData

    print('查询耗时：', datetime.datetime.now() - start)