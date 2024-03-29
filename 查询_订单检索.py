# coding=utf-8
import pandas as pd
import os
import re
import datetime
import time
import xlwings
import xlsxwriter
import math
import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
import pandas.io.formats.excel
import win32api,win32con
import math
from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色
from 轨迹查询_单点 import QueryTwo

# -*- coding:utf-8 -*-
class QueryOrder(Settings, Settings_sso):
    def __init__(self, userMobile, password, login_TmpCode, handle, proxy_id, proxy_handle):
        Settings.__init__(self)
        Settings_sso.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue(maxsize=10)  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        # self.sso_online_Two()
        # self._online_Two()

        # self.sso__online_auto()

        if proxy_handle == '代理服务器':
            if handle == '手动':
                self.sso__online_handle_proxy(login_TmpCode, proxy_id)
            else:
                self.sso__online_auto_proxy(proxy_id)
        else:
            if handle == '手动':
                self.sso__online_handle(login_TmpCode)
            else:
                self.sso__online_auto()

        # self.sso__online_handle(login_TmpCode)
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
        self.dk = Settings_sso()    # 钉钉发送
    def reSetEngine(self):
        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
    def readFormHost(self, team, searchType,pople_Query, timeStart, timeEnd, to_sql, proxy_id, proxy_handle):
        start = datetime.datetime.now()
        path = r'F:\需要用到的文件\A查询导表'
        dirs = os.listdir(path=path)
        # ---读取execl文件---
        for dir in dirs:
            filePath = os.path.join(path, dir)
            if dir[:2] != '~$':
                print(filePath)
                if pople_Query == '客服查询':
                    self.wbsheetHost_pople(filePath, team, searchType, proxy_id, proxy_handle)
                elif pople_Query == '电话检索':
                    self.wbsheetHost_iphone(filePath, team, searchType, timeStart, timeEnd, proxy_id, proxy_handle)
                else:
                    self.wbsheetHost(filePath, team, searchType, to_sql, proxy_id, proxy_handle)
                # self.cs_wbsheetHost(filePath, team, searchType)
        print('处理耗时：', datetime.datetime.now() - start)

    # 工作表的订单信息
    def wbsheetHost(self, filePath, team, searchType, to_sql, proxy_id, proxy_handle):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                if sht.api.Visible == -1:
                    try:
                        tem = None
                        db = None
                        db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                        print(db.columns)
                        # db = db[['订单编号']]
                        columns_value = list(db.columns)                             # 获取数据的标题名，转为列表
                        if searchType == '订单号':
                            tem = '订单编号'
                            if '订单号' in columns_value:
                                db.rename(columns={'订单号': '订单编号'}, inplace=True)
                            for column_val in columns_value:
                                if '订单编号' != column_val:
                                    db.drop(labels=[column_val], axis=1, inplace=True)  # 去掉多余的旬列表
                            db.dropna(axis=0, how='any', inplace=True)                  # 空值（缺失值），将空值所在的行/列删除后
                        elif searchType == '运单号':
                            tem = '运单编号'
                            if '运单号' in columns_value:
                                db.rename(columns={'运单号': '运单编号'}, inplace=True)
                            elif '查件单号' in columns_value:
                                db.rename(columns={'查件单号': '运单编号'}, inplace=True)
                    except Exception as e:
                        print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                    if db is not None and len(db) > 0:
                        # print(db)
                        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
                        print('++++正在获取：' + sht.name + ' 表；共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        orderId = list(db[tem])
                        max_count = len(orderId)                                    # 使用len()获取列表的长度，上节学的
                        # print(orderId)
                        # print(max_count)
                        if max_count > 500:
                            ord = ','.join(orderId[0:500])
                            # print(ord)
                            df = self.orderInfoQuery(ord, searchType, proxy_id, proxy_handle)
                            # print(df)
                            dlist = []
                            n = 0
                            while n < max_count-500:                                # 这里用到了一个while循环，穿越过来的
                                n = n + 500
                                ord = ','.join(orderId[n:n + 500])
                                data = self.orderInfoQuery(ord, searchType, proxy_id, proxy_handle)
                                dlist.append(data)
                                # print(dlist)
                                dp = df.append(dlist, ignore_index=True)
                                if to_sql == '写入':
                                    print('正在写入......')
                                    dp = dp[['orderNumber', 'area', 'shipInfo.shipEmail', 'addTime','logisticsStatus','orderStatus']]
                                    dp.columns = ['订单编号', '运营团队', '邮箱', '下单时间','物流状态','订单状态']
                                    dp.to_sql('cache', con=self.engine1, index=False, if_exists='replace')
                                    sql = '''REPLACE INTO 订单检索(订单编号, 运营团队, 邮箱, 下单时间, 物流状态, 订单状态) SELECT 订单编号, 运营团队, 邮箱, 下单时间 , 物流状态, 订单状态
                                            FROM cache;'''
                                    pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
                        else:
                            ord = ','.join(orderId[0:max_count])
                            dp = self.orderInfoQuery(ord, searchType, proxy_id, proxy_handle)
                        if to_sql != '写入':
                            dp = dp[['id','orderNumber', 'currency', 'area', 'productId', 'saleProduct', 'saleName', 'spec', 'shipInfo.shipName', 'shipInfo.shipPhone', 'percent', 'phoneLength',
                                     'shipInfo.shipAddress','amount', 'quantity', 'orderStatus', 'wayBillNumber', 'payType', 'addTime', 'username', 'verifyTime', 'logisticsName', 'dpeStyle',
                                     'hasLowPrice', 'collId', 'saleId', 'reassignmentTypeName', 'logisticsStatus', 'weight', 'delReason', 'questionReason', 'service', 'transferTime','deliveryTime', 'onlineTime',
                                     'finishTime', 'refundTime', 'remark', 'ip', 'volume', 'shipInfo.shipState', 'shipInfo.shipCity', 'chooser', 'optimizer','autoVerify', 'cloneUser', 'isClone', 'warehouse', 'smsStatus',
                                     'logisticsControl', 'logisticsRefuse', 'logisticsUpdateTime', 'stateTime', 'collDomain', 'typeName', 'update_time', 'autoVerifyTip', 'auto_VerifyTip', 'auto_VerifyTip_zl', 'auto_VerifyTip_qs', 'auto_VerifyTip_js','notes']]
                            dp.columns = ['id','订单编号', '币种', '运营团队', '产品id', '产品名称', '出货单名称', '规格(中文)', '收货人', '联系电话', '拉黑率', '电话长度',
                                          '配送地址', '应付金额', '数量', '订单状态', '运单号', '支付方式', '下单时间', '审核人', '审核时间', '物流渠道', '货物类型',
                                          '是否低价', '站点ID', '商品ID', '订单类型', '物流状态', '重量', '删除原因', '问题原因', '下单人', '转采购时间', '发货时间', '上线时间',
                                          '完成时间', '销售退货时间', '备注', 'IP', '体积', '省洲', '市/区', '选品人', '优化师', '审单类型', '克隆人', '克隆ID', '发货仓库', '是否发送短信',
                                          '物流渠道预设方式', '拒收原因', '物流更新时间', '状态时间', '来源域名', '订单来源类型', '更新时间', '异常提示', '异常拉黑率',
                                          '拉黑率总量','拉黑率签收','拉黑率拒收','留言']
                            dp.to_excel('F:\\输出文件\\订单检索-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')   # Xlsx是python用来构造xlsx文件的模块，可以向excel2007+中写text，numbers，formulas 公式以及hyperlinks超链接。
                        print('查询已导出+++')
                    else:
                        print('----------数据为空,查询失败：' + sht.name)
                else:
                    print('----不需查询：' + sht.name)
            wb.close()
        app.quit()

    def wbsheetHost_pople(self, filePath, team, searchType, proxy_id, proxy_handle):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                if sht.api.Visible == -1:
                    db = None
                    tem = None
                    try:
                        db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                        print(db.columns)
                        # db = db[['订单编号']]
                        if searchType == '订单号':
                            tem = '订单编号'
                            columns_value = list(db.columns)                             # 获取数据的标题名，转为列表
                            if '订单号' in columns_value:
                                db.rename(columns={'订单号': '订单编号'}, inplace=True)
                            for column_val in columns_value:
                                if '订单编号' != column_val:
                                    db.drop(labels=[column_val], axis=1, inplace=True)  # 去掉多余的旬列表
                        elif searchType == '运单号':
                            tem = '运单编号'
                            columns_value = list(db.columns)                            # 获取数据的标题名，转为列表
                            if '运单号' in columns_value:
                                db.rename(columns={'运单号': '运单编号'}, inplace=True)
                            for column_val in columns_value:
                                if '运单编号' != column_val:
                                    db.drop(labels=[column_val], axis=1, inplace=True)  # 去掉多余的旬列表

                        db.dropna(axis=0, how='any', inplace=True)                      # 空值（缺失值），将空值所在的行/列删除后
                    except Exception as e:
                        print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                    if db is not None and len(db) > 0:
                        # print(db)
                        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
                        print('++++正在获取：' + sht.name + ' 表；共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        orderId = list(db[tem])
                        max_count = len(orderId)                                    # 使用len()获取列表的长度，上节学的
                        if max_count > 90:
                            ord = ','.join(orderId[0:90])
                            df = self.orderInfo_pople(ord, searchType, proxy_id, proxy_handle)
                            # print(df)
                            dlist = []
                            n = 0
                            while n < max_count-90:                                # 这里用到了一个while循环，穿越过来的
                                n = n + 90
                                ord = ','.join(orderId[n:n + 90])
                                data = self.orderInfo_pople(ord, searchType, proxy_id, proxy_handle)
                                dlist.append(data)
                            print('正在写入......')
                            # print(dlist)
                            dp = df.append(dlist, ignore_index=True)
                        else:
                            ord = ','.join(orderId[0:max_count])
                            dp = self.orderInfo_pople(ord, searchType, proxy_id, proxy_handle)
                        # dp.columns = ['订单编号', '币种', '运营团队', '产品id', '产品名称', '出货单名称', '规格(中文)', '收货人', '联系电话', '拉黑率', '电话长度',
                        #               '配送地址', '应付金额', '数量', '订单状态', '运单号', '支付方式', '下单时间', '审核人', '审核时间', '物流渠道', '货物类型',
                        #               '是否低价', '站点ID', '商品ID', '订单类型', '物流状态', '重量', '删除原因', '问题原因', '下单人', '转采购时间', '发货时间', '上线时间',
                        #               '完成时间', '销售退货时间', '备注', 'IP', '体积', '省洲', '市/区', '选品人', '优化师', '审单类型', '克隆人', '克隆ID', '发货仓库', '是否发送短信',
                        #               '物流渠道预设方式', '拒收原因', '物流更新时间', '状态时间', '来源域名', '订单来源类型', '更新时间', '异常提示', '异常拉黑率',
                        #               '拉黑率总量','拉黑率签收','拉黑率拒收','留言']
                        dp.to_excel('F:\\输出文件\\订单检索-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')   # Xlsx是python用来构造xlsx文件的模块，可以向excel2007+中写text，numbers，formulas 公式以及hyperlinks超链接。
                        print('查询已导出+++')
                    else:
                        print('----------数据为空,查询失败：' + sht.name)
                else:
                    print('----不需查询：' + sht.name)
            wb.close()
        app.quit()

    def wbsheetHost_iphone(self, filePath, team, searchType, timeStart, timeEnd, proxy_id, proxy_handle):
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                if sht.api.Visible == -1:
                    try:
                        db = None
                        db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                        print(db.columns)
                    except Exception as e:
                        print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                    if db is not None and len(db) > 0:
                        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
                        print('++++正在获取：' + sht.name + ' 表；共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                        orderId = list(db['电话'])
                        # print(orderId)
                        max_count = len(orderId)                                    # 使用len()获取列表的长度，上节学的
                        if max_count > 10:
                            # ord = ','.join(orderId[0:100])
                            ord = ','.join('%s' % d for d in orderId[0:10])
                            # print(ord)
                            df = self.orderInfo_iphone(ord, searchType, timeStart, timeEnd, proxy_id, proxy_handle)
                            dlist = []
                            n = 0
                            while n < max_count-10:                                # 这里用到了一个while循环，穿越过来的
                                n = n + 10
                                # ord = ','.join(orderId[n:n + 100])
                                ord = ','.join('%s' % d for d in orderId[n:n + 10])
                                # print(ord)
                                data = self.orderInfo_iphone(ord, searchType, timeStart, timeEnd, proxy_id, proxy_handle)
                                dlist.append(data)
                            print('正在写入......')
                            # print(dlist)
                            dp = df.append(dlist, ignore_index=True)
                        else:
                            # print(orderId[0:max_count])
                            # ord = ','.join(orderId[0:max_count])
                            ord = ','.join('%s' %d for d in orderId[0:max_count])
                            # print(ord)
                            dp = self.orderInfo_iphone(ord, searchType, timeStart, timeEnd, proxy_id, proxy_handle)
                        # dp.columns = ['订单编号', '币种', '运营团队', '产品id', '产品名称', '出货单名称', '规格(中文)', '收货人', '联系电话', '拉黑率', '电话长度',
                        #               '配送地址', '应付金额', '数量', '订单状态', '运单号', '支付方式', '下单时间', '审核人', '审核时间', '物流渠道', '货物类型',
                        #               '是否低价', '站点ID', '商品ID', '订单类型', '物流状态', '重量', '删除原因', '问题原因', '下单人', '转采购时间', '发货时间', '上线时间',
                        #               '完成时间', '销售退货时间', '备注', 'IP', '体积', '省洲', '市/区', '选品人', '优化师', '审单类型', '克隆人', '克隆ID', '发货仓库', '是否发送短信',
                        #               '物流渠道预设方式', '拒收原因', '物流更新时间', '状态时间', '来源域名', '订单来源类型', '更新时间', '异常提示', '异常拉黑率',
                        #               '拉黑率总量','拉黑率签收','拉黑率拒收','留言']
                        dp.to_excel('F:\\输出文件\\订单检索-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')   # Xlsx是python用来构造xlsx文件的模块，可以向excel2007+中写text，numbers，formulas 公式以及hyperlinks超链接。
                        print('查询已导出+++')
                    else:
                        print('----------数据为空,查询失败：' + sht.name)
                else:
                    print('----不需查询：' + sht.name)
            wb.close()
        app.quit()

    # 测试 self.InfoQuery(db, searchType)
    def cs_wbsheetHost(self, filePath, team, searchType):
        match2 = {'gat': '港台'}
        print('---正在获取 ' + match2[team] + ' 签收表的详情++++++')
        fileType = os.path.splitext(filePath)[1]
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        if 'xls' in fileType:
            wb = app.books.open(filePath, update_links=False, read_only=True)
            for sht in wb.sheets:
                try:
                    db = None
                    db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
                    columns_value = list(db.columns)  # 获取数据的标题名，转为列表
                    if '订单号' in columns_value:
                        db.rename(columns={'订单号': '订单编号'}, inplace=True)
                    for column_val in columns_value:
                        if '订单编号' != column_val:
                            db.drop(labels=[column_val], axis=1, inplace=True)  # 去掉多余的旬列表
                except Exception as e:
                    print('xxxx查看失败：' + sht.name, str(Exception) + str(e))
                if db is not None and len(db) > 0:
                    print('++++正在查询：' + sht.name + ' 共：' + str(len(db)) + '行', 'sheet共：' + str(sht.used_range.last_cell.row) + '行')
                    self.cs_InfoQuery(db, searchType)
                else:
                    print('----------数据为空导入失败：' + sht.name)
            wb.close()
        app.quit()
    def cs_InfoQuery(self, db, searchType):  # 调用多线程
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        orderId = list(db['订单编号'])
        max_count = len(orderId)  # 使用len()获取列表的长度，上节学的
        if max_count > 500:
            ord = ', '.join(orderId[0:500])
            # df = self.cs_orderInfoQuery(ord, searchType)
            self.cs_orderInfoQuery(ord, searchType)
            df = self.q.get()
            print(df)
            print(99)
            # self.q.join()
            print('主线程开始执行……………………')
            threads = []  # 多线程用线程池--
            n = 0
            count = 1
            while n < max_count - 500:  # 这里用到了一个while循环，穿越过来的
                n = n + 500
                ord = ','.join(orderId[n:n + 500])
                print(str(count) + '次')
                # threads.append(Thread(target=self.cs_orderInfoQuery, args=(ord, searchType)))  # -----也即是子线程
                t = Thread(target=self.cs_orderInfoQuery, args=(ord, searchType))  # -----也即是子线程
                threads.append(t)
                t.start()
                count = count + 1
            for th in threads:
                th.join()
            print('主线程执行结束---------')
            dlist = []
            for j in range(self.q.qsize()):
                dlist.append(self.q.get())
            print(dlist)
            print('正在写入......')
            dp = df.append(dlist, ignore_index=True)
            print(dp)
        else:
            ord = ','.join(orderId[0:max_count])
            df = self.cs_orderInfoQuery(ord, searchType)
        dp.to_excel('F:\\输出文件\\订单检索-查询{}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
    def cs_orderInfoQuery(self, ord, searchType):  # 进入订单检索界面
        print('+++正在查询订单信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500,
                'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None,
                'type': None, 'collId': None, 'isClone': None,
                'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '',
                'warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None,
                'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None,
                'service_id': None, 'autoVerifyStatus': None, 'shipZip': None, 'remark': None, 'shipState': None, 'weightStart': None,
                'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None}
        if searchType == '订单号':
            data.update({'orderPrefix': ord,
                         'shippingNumber': None})
        elif searchType == '运单号':
            data.update({'order_number': None,
                         'shippingNumber': ord})
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        ordersdict = []
        # print('正在处理json数据转化为dataframe…………')
        try:
            for result in req['data']['list']:
                result['saleId'] = 0        # 添加新的字典键-值对，为下面的重新赋值用
                result['saleName'] = 0
                result['productId'] = 0
                result['saleProduct'] = 0
                result['spec'] = 0
                result['chooser'] = 0
                result['saleId'] = result['specs'][0]['saleId']
                result['saleName'] = result['specs'][0]['saleName']
                result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
                result['saleProduct'] = (result['specs'][0]['saleProduct']).split('#')[2]
                result['spec'] = result['specs'][0]['spec']
                result['chooser'] = result['specs'][0]['chooser']
                quest = ''
                for re in result['questionReason']:
                    quest = quest + ';' + re
                result['questionReason'] = quest
                delr = ''
                for re in result['delReason']:
                    delr = delr + ';' + re
                result['delReason'] = delr
                auto = ''
                for re in result['autoVerify']:
                    auto = auto + ';' + re
                result['autoVerify'] = auto
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersdict)
        df = None
        try:
            df = data[['orderNumber', 'currency', 'area', 'productId', 'saleProduct', 'saleName', 'spec',
                    'shipInfo.shipName', 'shipInfo.shipPhone', 'percent', 'phoneLength', 'shipInfo.shipAddress',
                    'amount', 'quantity', 'orderStatus', 'wayBillNumber', 'payType', 'addTime', 'username', 'verifyTime',
                    'logisticsName', 'dpeStyle', 'hasLowPrice', 'collId', 'saleId', 'reassignmentTypeName',
                    'logisticsStatus', 'weight', 'delReason', 'questionReason', 'service', 'transferTime', 'deliveryTime', 'onlineTime',
                    'finishTime', 'remark', 'ip', 'volume', 'shipInfo.shipState', 'shipInfo.shipCity', 'chooser', 'optimizer',
                    'autoVerify', 'cloneUser', 'isClone', 'warehouse', 'smsStatus', 'logisticsControl',
                    'logisticsRefuse', 'logisticsUpdateTime', 'stateTime', 'collDomain', 'typeName', 'update_time']]
            # print(df)
            self.q.put(df)
        except Exception as e:
            print('------查询为空')
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        # return df

    # 一、订单——查询更新（单点获取）
    def orderInfoQuery(self, ord, searchType, proxy_id, proxy_handle):  # 进入订单检索界面
        print('+++正在查询订单信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500,
                'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None,
                'type': None, 'collId': None, 'isClone': None,
                'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '',
                'warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None,
                'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None,
                'service_id': None, 'autoVerifyStatus': None, 'shipZip': None, 'remark': None, 'shipState': None, 'weightStart': None,
                'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None}
        if searchType == '订单号':
            data.update({'orderPrefix': ord, 'shippingNumber': None})
        elif searchType == '运单号':
            data.update({'order_number': None, 'shippingNumber': ord})
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        # print('正在处理json数据转化为dataframe…………')
        try:
            for result in req['data']['list']:
                # print(result['orderNumber'])
                if result['specs'] != []:
                    result['saleId'] = 0        # 添加新的字典键-值对，为下面的重新赋值用
                    result['saleName'] = 0
                    result['productId'] = 0
                    result['saleProduct'] = 0
                    result['spec'] = 0
                    result['chooser'] = 0
                    result['saleId'] = result['specs'][0]['saleId']
                    result['saleName'] = result['specs'][0]['saleName']
                    result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
                    result['saleProduct'] = (result['specs'][0]['saleProduct']).split('#')[2]
                    result['spec'] = result['specs'][0]['spec']
                    result['chooser'] = result['specs'][0]['chooser']
                else:
                    result['saleId'] = ''        # 添加新的字典键-值对，为下面的重新赋值用
                    result['saleName'] = ''
                    result['productId'] = ''
                    result['saleProduct'] = ''
                    result['spec'] = ''
                    result['chooser'] = ''
                quest = ''
                for re in result['questionReason']:
                    quest = quest + ';' + re
                result['questionReason'] = quest
                delr = ''
                for re in result['delReason']:
                    delr = delr + ';' + re
                result['delReason'] = delr
                auto = ''
                for re in result['autoVerify']:
                    auto = auto + ';' + re
                result['autoVerify'] = auto

                result['auto_VerifyTip'] = ''
                result['auto_VerifyTip_zl'] = ''
                result['auto_VerifyTip_qs'] = ''
                result['auto_VerifyTip_js'] = ''
                if result['autoVerifyTip'] == "":
                    result['auto_VerifyTip'] = '0.00%'
                else:
                    if '未读到拉黑表记录' in result['autoVerifyTip']:
                        result['auto_VerifyTip'] = '0.00%'
                    else:
                        t3 = result['autoVerifyTip']
                        result['auto_VerifyTip_zl'] = (t3.split('订单配送总量：')[1]).split(',')[0]
                        result['auto_VerifyTip_qs'] = (t3.split('送达订单量：')[1]).split(',')[0]
                        result['auto_VerifyTip_js'] = (t3.split('拒收订单量：')[1]).split(',')[0]
                        if '拉黑率问题' not in result['autoVerifyTip']:
                            t2 = result['autoVerifyTip'].split(',拉黑率')[1]
                            result['auto_VerifyTip'] = t2.split('%;')[0] + '%'
                        else:
                            t2 = result['autoVerifyTip'].split('拒收订单量：')[1]
                            t2 = t2.split('%;')[0]
                            result['auto_VerifyTip'] = t2.split('拉黑率')[1] + '%'
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        df = pd.json_normalize(ordersdict)
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return df
    # 一、订单——客服查询（单点获取）
    def orderInfo_pople(self, ord, searchType, proxy_id, proxy_handle):  # 进入订单检索界面
        print('+++正在查询订单信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getQueryOrder'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 90, 'orderPrefix': None, 'shippingNumber': None, 'phone': None, 'email': None, 'ip': None }
        if searchType == '订单号':
            data.update({'orderPrefix': ord, 'shippingNumber': None})
        elif searchType == '运单号':
            data.update({'order_number': None, 'shippingNumber': ord})
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        # print('正在处理json数据转化为dataframe…………')
        try:
            for result in req['data']['list']:
                result['saleId'] = 0        # 添加新的字典键-值对，为下面的重新赋值用
                result['saleName'] = 0
                result['productId'] = 0
                result['saleProduct'] = 0
                result['spec'] = 0
                result['saleId'] = result['specs'][0]['saleId']
                result['saleName'] = result['specs'][0]['saleName']
                result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
                result['saleProduct'] = (result['specs'][0]['saleProduct']).split('#')[2]
                result['spec'] = result['specs'][0]['spec']
                quest = ''
                for re in result['questionReason']:
                    quest = quest + ';' + re
                result['questionReason'] = quest
                delr = ''
                for re in result['delReason']:
                    delr = delr + ';' + re
                result['delReason'] = delr
                auto = ''
                for re in result['autoVerify']:
                    auto = auto + ';' + re
                result['autoVerify'] = auto
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            # time.sleep(10)
            # print(team)
            # print(searchType)
            # self.readFormHost(team, searchType)
            # self.orderInfoQuery(ord, searchType)
        #     self.q.put(result)
        # for i in range(len(req['data']['list'])):
        #     ordersdict.append(self.q.get())
        data = pd.json_normalize(ordersdict)
        # df = None
        # print(df)
        # try:
        #     df = data[['orderNumber', 'currency', 'area', 'productId', 'saleProduct', 'saleName', 'spec',
        #             'shipInfo.shipName', 'shipInfo.shipPhone', 'percent', 'phoneLength', 'shipInfo.shipAddress',
        #             'amount', 'quantity', 'orderStatus', 'wayBillNumber', 'payType', 'addTime', 'username', 'verifyTime',
        #             'logisticsName', 'dpeStyle', 'hasLowPrice', 'collId', 'saleId', 'reassignmentTypeName',
        #             'logisticsStatus', 'weight', 'delReason', 'questionReason', 'service', 'transferTime', 'deliveryTime', 'onlineTime',
        #             'finishTime', 'refundTime', 'remark', 'ip', 'volume', 'shipInfo.shipState', 'shipInfo.shipCity', 'optimizer',
        #             'autoVerify', 'cloneUser', 'isClone', 'warehouse', 'smsStatus', 'logisticsControl','logisticsRefuse', 'logisticsUpdateTime',
        #             'stateTime', 'collDomain', 'typeName', 'update_time','notes']]
        # except Exception as e:
        #     print('------查询为空')
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return data
    # 一、订单——电话查询（单点获取）
    def orderInfo_iphone(self, ord, searchType, timeStart, timeEnd, proxy_id, proxy_handle):  # 进入订单检索界面
        print('+++正在查询订单信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500, 'orderPrefix': None,'shippingNumber': None, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': ord, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None, 'emailStatus': None,
                'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '','warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None,
                'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None,'service_id': None, 'autoVerifyStatus': None, 'shipZip': None,
                'remark': None, 'shipState': None, 'weightStart': None,'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'isChangeMark': None, 'percentStart': None, 'percentEnd': None, 'userid': None, 'questionId': None,
                'delUserId': None, 'transferNumber': None, 'smsStatus': None, 'designer_id': None, 'logistics_remarks': None, 'clone_type': None, 'categoryId': None, 'addressType': None,
                'timeStart': timeStart,  'timeEnd': timeEnd }
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,  'https': 'socks5://' + proxy}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        # print('正在处理json数据转化为dataframe…………')
        try:
            for result in req['data']['list']:
                result['saleId'] = 0        # 添加新的字典键-值对，为下面的重新赋值用
                result['saleName'] = 0
                result['productId'] = 0
                result['saleProduct'] = 0
                result['spec'] = 0
                result['chooser'] = 0
                result['saleId'] = result['specs'][0]['saleId']
                result['saleName'] = result['specs'][0]['saleName']
                result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
                result['saleProduct'] = (result['specs'][0]['saleProduct']).split('#')[2]
                result['spec'] = result['specs'][0]['spec']
                result['chooser'] = result['specs'][0]['chooser']
                quest = ''
                for re in result['questionReason']:
                    quest = quest + ';' + re
                result['questionReason'] = quest
                delr = ''
                for re in result['delReason']:
                    delr = delr + ';' + re
                result['delReason'] = delr
                auto = ''
                for re in result['autoVerify']:
                    auto = auto + ';' + re
                result['autoVerify'] = auto

                result['auto_VerifyTip'] = ''
                result['auto_VerifyTip_zl'] = ''
                result['auto_VerifyTip_qs'] = ''
                result['auto_VerifyTip_js'] = ''
                if result['autoVerifyTip'] == "":
                    result['auto_VerifyTip'] = '0.00%'
                else:
                    if '未读到拉黑表记录' in result['autoVerifyTip']:
                        result['auto_VerifyTip'] = '0.00%'
                    else:
                        t3 = result['autoVerifyTip']
                        result['auto_VerifyTip_zl'] = (t3.split('订单配送总量：')[1]).split(',')[0]
                        result['auto_VerifyTip_qs'] = (t3.split('送达订单量：')[1]).split(',')[0]
                        result['auto_VerifyTip_js'] = (t3.split('拒收订单量：')[1]).split(',')[0]
                        if '拉黑率问题' not in result['autoVerifyTip']:
                            t2 = result['autoVerifyTip'].split(',拉黑率')[1]
                            result['auto_VerifyTip'] = t2.split('%;')[0] + '%'
                        else:
                            t2 = result['autoVerifyTip'].split('拒收订单量：')[1]
                            t2 = t2.split('%;')[0]
                            result['auto_VerifyTip'] = t2.split('拉黑率')[1] + '%'
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersdict)
        df = None
        try:
            df = data[['orderNumber', 'currency', 'area', 'productId', 'saleProduct', 'saleName', 'spec',
                    'shipInfo.shipName', 'shipInfo.shipPhone', 'percent', 'phoneLength', 'shipInfo.shipAddress',
                    'amount', 'quantity', 'orderStatus', 'wayBillNumber', 'payType', 'addTime', 'username', 'verifyTime',
                    'logisticsName', 'dpeStyle', 'hasLowPrice', 'collId', 'saleId', 'reassignmentTypeName',
                    'logisticsStatus', 'weight', 'delReason', 'questionReason', 'service', 'transferTime', 'deliveryTime', 'onlineTime',
                    'finishTime', 'refundTime', 'remark', 'ip', 'volume', 'shipInfo.shipState', 'shipInfo.shipCity', 'chooser', 'optimizer',
                    'autoVerify', 'cloneUser', 'isClone', 'warehouse', 'smsStatus', 'logisticsControl','logisticsRefuse', 'logisticsUpdateTime',
                    'stateTime', 'collDomain', 'typeName', 'update_time', 'autoVerifyTip','auto_VerifyTip','auto_VerifyTip_zl','auto_VerifyTip_qs','auto_VerifyTip_js','notes']]
        except Exception as e:
            print('------查询为空')
        print('++++++本批次查询成功+++++++')
        print('*' * 50)
        return df


    # 单点获取 line运营  时间-查询
    def order_TimeQuery(self, timeStart, timeEnd, areaId, query, proxy_id, proxy_handle, logisticsId, currencyId):  # 进入订单检索界面
        print('+++正在查询订单信息中起止： ' + timeStart + ':' + timeEnd)
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500, 'order_number': None, 'shippingNumber': None, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None,
                'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '', 'warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None,
                'tuan': None,'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': None, 'autoVerifyStatus': None,
                'shipZip': None, 'remark': None, 'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'timeStart': None, 'timeEnd': None}
        if query == '下单时间':
            data.update({'timeStart': timeStart + '00:00:00', 'timeEnd': timeEnd + '23:59:59', 'finishTimeStart': None, 'finishTimeEnd': None,
                         'areaId': areaId, 'currencyId': currencyId, 'logisticsId': logisticsId})
        elif query == '完成时间':
            data.update({'timeStart': None, 'timeEnd': None, 'finishTimeStart': timeStart + '00:00:00', 'finishTimeEnd': timeEnd + '23:59:59',
                         'areaId': areaId, 'currencyId': currencyId, 'logisticsId': logisticsId})
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        max_count = req['data']['count']
        print(max_count)
        if max_count > 0:
            in_count = math.ceil(max_count/500)
            df = pd.DataFrame([])
            dlist = []
            n = 1
            while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                print('剩余查询次数' + str(in_count - n))
                data = self._timeQuery_format(timeStart, timeEnd, n, areaId, query, proxy_id, proxy_handle, logisticsId, currencyId)
                n = n + 1
                dlist.append(data)
            print('正在写入......')
            dp = df.append(dlist, ignore_index=True)
            dp = dp[['orderNumber', 'area', 'logisticsStatus', 'orderStatus', 'abbreviation', 'addTime', 'saleProduct']]
            dp.columns = ['订单编号', '运营团队', '物流状态', '订单状态', '商品简称', '下单时间', '产品名称']
            # columns = list(dp.columns)
            # columns = ','.join(columns)
            dp.to_sql('cache', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO {0}(订单编号,运营团队,物流状态,订单状态,商品简称,下单时间,产品名称) 
                               SELECT 订单编号,运营团队,物流状态,订单状态,商品简称,下单时间,产品名称 FROM cache;'''.format('订单检索')
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('查询已导出+++')
        else:
            print('查询无资料+++')

    # 单点获取 全部昨日订单 分析 头程渠道 & 天马711 渠道
    def order_Query_Yiwudi(self, timeStart, timeEnd, areaId, query, proxy_id, proxy_handle):  # 进入订单检索界面
        print('+++正在获取 ' + timeStart + ' 到 ' + timeEnd + ' 号订单明细......')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500, 'order_number': None, 'shippingNumber': None, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None, 'emailStatus': None,
                'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '', 'warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None,
                'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': None, 'autoVerifyStatus': None, 'shipZip': None,
                'remark': None, 'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'timeStart':None, 'timeEnd': None}
        if query == '下单时间':
            data.update({'timeStart': timeStart + '00:00:00', 'timeEnd': timeEnd + '23:59:59', 'finishTimeStart': None, 'finishTimeEnd': None})
        elif query == '完成时间':
            data.update({'timeStart': None,'timeEnd': None,  'finishTimeStart': timeStart + '00:00:00', 'finishTimeEnd': timeEnd + '23:59:59'})
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        max_count = req['data']['count']
        print(max_count)
        if max_count > 0:
            in_count = math.ceil(max_count/500)
            df = pd.DataFrame([])
            dlist = []
            n = 1
            while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                print('查询第 ' + str(n) + ' 页中，剩余次数' + str(in_count - n))
                data = self._timeQuery(timeStart, timeEnd, n, areaId, query, proxy_id, proxy_handle)
                dlist.append(data)
                n = n + 1
            print('正在写入......')
            dp = df.append(dlist, ignore_index=True)

            print('检查渠道：天马711')
            db9 = dp[(dp['订单类型'].str.contains('直发下架|未下架未改派'))]
            db91 = db9[(db9['订单状态'].str.contains('已转采购|截单中(面单已打印,等待仓库审核)|待审核|待发货|问题订单|问题订单审核|截单'))]
            # db92 = db91[(db91['物流渠道'].str.contains('台湾-速派-711超商'))]
            db92 = db91[(db91['物流渠道'].str.contains('台湾-天马-711'))]
            if len(db92) > 0:
                print('天马-711超商 订单出现，请拦截！！！')
                orderId = list(db92['订单编号'])
                orderId = ','.join(orderId)
                # url = "https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e"  # url为机器人的webhook  个人 小海
                url = "https://oapi.dingtalk.com/robot/send?access_token=fa74c55267674d9281f705b6fde624818c9977287cb590891ef2691714a9ceda"  # url为机器人的webhook  审单问题群
                content = r"天马711超商订单： 解绑超商，重新审核；" + orderId  # 钉钉消息内容，注意test是自定义的关键字，需要在钉钉机器人设置中添加，这样才能接收到消息
                mobile_list = ['18538110674']  # 要@的人的手机号，可以是多个，注意：钉钉机器人设置中需要添加这些人，否则不会接收到消息
                isAtAll = '单个'  # 是、 否、 单个、 @所有人
                self.dk.send_dingtalk_message(url, content, mobile_list, isAtAll)
            else:
                print('---无 天马-711超商 订单')

            print('检查渠道：头程直发渠道')
            db1 = dp[(dp['币种'].str.contains('台币'))]
            db12 = db1[(db1['订单类型'].str.contains('直发下架|未下架未改派'))]
            db13 = db12[(db12['订单状态'].str.contains('已转采购|截单中(面单已打印,等待仓库审核)|待审核|待发货|问题订单|问题订单审核|截单'))]
            db14 = db13[(db13['物流渠道'].str.contains('台湾-铱熙无敌-新竹特货|台湾-铱熙无敌-新竹普货|台湾-立邦普货头程-易速配尾程|台湾-铱熙无敌-黑猫改派|台湾-铱熙无敌-黑猫特货|台湾-铱熙无敌-黑猫普货|台湾-铱熙无敌-711敏感货|台湾-易速配头程-铱熙无敌尾'))]
            if len(db14) > 0:
                print('铱熙无敌 订单出现，请拦截！！！')
            else:
                print('---无 铱熙无敌 订单')

            # ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]') # ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
            file_path = 'F:\\输出文件\\{0} 订单检索-明细.xlsx'.format(timeEnd)
            # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            # db1.to_excel(excel_writer=writer, sheet_name='明细', index=False)
            # db92.to_excel(excel_writer=writer, sheet_name='天马711', index=False)
            # db14.to_excel(excel_writer=writer, sheet_name='协来运直发', index=False)
            # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            #     del book['Sheet1']
            # writer.save()
            # writer.close()
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                db1.to_excel(excel_writer=writer, sheet_name='明细', index=False)
                db92.to_excel(excel_writer=writer, sheet_name='天马711', index=False)
                db14.to_excel(excel_writer=writer, sheet_name='协来运直发', index=False)
            print('昨日明细 查询已导出+++')
        else:
            print('无信息+++')

    #   --  -- --  --  -- -- 时间查询的公用函数  --  订单检索 --  -- --  --  -- --  --  -- --  --  -- --
    def _timeQuery(self, timeStart, timeEnd, n, areaId, query, proxy_id, proxy_handle):  # 进入订单检索界面
        # print('......正在查询信息中......')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': n, 'pageSize': 500, 'order_number': None, 'shippingNumber': None, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None,
                'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '', 'warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None,
                'tuan': None,'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': None, 'autoVerifyStatus': None,
                'shipZip': None, 'remark': None, 'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'timeStart': None, 'timeEnd': None}
        if query == '下单时间':
            data.update({'timeStart': timeStart + '00:00:00', 'timeEnd': timeEnd + '23:59:59', 'finishTimeStart': None, 'finishTimeEnd': None,
                         'areaId': areaId, 'currencyId': None, 'logisticsId': None})
        elif query == '完成时间':
            data.update({'timeStart': None, 'timeEnd': None, 'finishTimeStart': timeStart + '00:00:00', 'finishTimeEnd': timeEnd + '23:59:59',
                         'areaId': areaId, 'currencyId': None, 'logisticsId': None})

        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('......已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        try:
            for result in req['data']['list']:
                e_val =result['orderNumber']
                # print(result['orderNumber'])
                if result['specs'] != '' and result['specs'] != []:
                    result['saleId'] = 0        # 添加新的字典键-值对，为下面的重新赋值用
                    result['saleName'] = 0
                    result['productId'] = 0
                    result['saleProduct'] = 0
                    result['spec'] = 0
                    result['chooser'] = 0
                    result['saleId'] = result['specs'][0]['saleId']
                    result['saleName'] = result['specs'][0]['saleName']
                    result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
                    result['saleProduct'] = (result['specs'][0]['saleProduct']).split('#')[2]
                    result['spec'] = result['specs'][0]['spec']
                    result['chooser'] = result['specs'][0]['chooser']
                else:
                    result['saleId'] = ''
                    result['saleProduct'] = ''
                    result['productId'] = ''
                    result['spec'] = ''
                quest = ''
                for re in result['questionReason']:
                    quest = quest + ';' + re
                result['questionReason'] = quest
                delr = ''
                for re in result['delReason']:
                    delr = delr + ';' + re
                result['delReason'] = delr
                auto = ''
                for re in result['autoVerify']:
                    auto = auto + ';' + re
                result['autoVerify'] = auto
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e) + str(e_val))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e", "订单检索-获取数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        data = pd.json_normalize(ordersdict)
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        # data.to_excel('F:\\输出文件\\明细{0}.xlsx'.format(rq), sheet_name='查询', index=False, engine='xlsxwriter')
        df = None
        try:
            df = data[['orderNumber', 'befrom', 'currency', 'area', 'productId', 'saleProduct', 'saleName', 'spec', 'shipInfo.shipName', 'shipInfo.shipPhone', 'tel_phone','percent', 'phoneLength',
                       'shipInfo.shipAddress', 'amount', 'quantity', 'orderStatus', 'wayBillNumber', 'payType', 'addTime', 'username', 'verifyTime','logisticsName', 'dpeStyle',
                       'hasLowPrice', 'collId', 'saleId', 'reassignmentTypeName', 'logisticsStatus', 'weight', 'delReason', 'questionReason', 'service', 'transferTime', 'deliveryTime', 'onlineTime',
                       'finishTime', 'refundTime', 'remark', 'ip', 'volume', 'shipInfo.shipState', 'shipInfo.shipCity', 'chooser', 'optimizer', 'autoVerify', 'autoVerifyTip', 'cloneUser', 'isClone', 'warehouse', 'smsStatus',
                       'logisticsControl', 'logisticsRefuse', 'logisticsUpdateTime', 'stateTime', 'collDomain', 'typeName','update_time']]
            df.columns = ['订单编号', '平台', '币种', '运营团队', '产品id', '产品名称', '出货单名称', '规格(中文)', '收货人', '联系电话', '标准电话','拉黑率', '电话长度',
                          '配送地址', '应付金额', '数量', '订单状态', '运单号', '支付方式', '下单时间', '审核人', '审核时间', '物流渠道', '货物类型',
                          '是否低价', '站点ID', '商品ID', '订单类型', '物流状态', '重量', '删除原因', '问题原因', '下单人', '转采购时间', '发货时间', '上线时间',
                          '完成时间', '销售退货时间', '备注', 'IP', '体积', '省洲', '市/区', '选品人', '优化师', '审单类型', '异常提示', '克隆人', '克隆ID', '发货仓库', '是否发送短信',
                          '物流渠道预设方式', '拒收原因', '物流更新时间', '状态时间', '来源域名', '订单来源类型',  '更新时间']
        except Exception as e:
            print('！！！！！！ 转化失败，请检查过程 ！！！！！！')
        print('******本批次查询成功')
        return df

    def _timeQuery_format(self, timeStart, timeEnd, n, areaId, query, proxy_id, proxy_handle, logisticsId, currencyId):  # 进入订单检索界面
        # print('......正在查询信息中......')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': n, 'pageSize': 500, 'order_number': None, 'shippingNumber': None, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None,
                'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '', 'warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None,
                'tuan': None,'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': None, 'autoVerifyStatus': None,
                'shipZip': None, 'remark': None, 'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'timeStart': None, 'timeEnd': None}
        if query == '下单时间':
            data.update({'timeStart': timeStart + '00:00:00', 'timeEnd': timeEnd + '23:59:59', 'finishTimeStart': None, 'finishTimeEnd': None,
                         'areaId': areaId, 'currencyId': currencyId, 'logisticsId': logisticsId})
        elif query == '完成时间':
            data.update({'timeStart': None, 'timeEnd': None, 'finishTimeStart': timeStart + '00:00:00', 'finishTimeEnd': timeEnd + '23:59:59',
                         'areaId': areaId, 'currencyId': currencyId, 'logisticsId': logisticsId})
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('......已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        try:
            for result in req['data']['list']:
                # print(result['orderNumber'])
                if result['specs'] != '' and result['specs'] != []:
                    result['saleId'] = 0        # 添加新的字典键-值对，为下面的重新赋值用
                    result['saleName'] = 0
                    result['productId'] = 0
                    result['saleProduct'] = 0
                    result['spec'] = 0
                    result['chooser'] = 0
                    result['saleId'] = result['specs'][0]['saleId']
                    result['saleName'] = result['specs'][0]['saleName']
                    result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
                    result['saleProduct'] = (result['specs'][0]['saleProduct']).split('#')[2]
                    result['spec'] = result['specs'][0]['spec']
                    result['chooser'] = result['specs'][0]['chooser']
                else:
                    result['saleId'] = ''
                    result['saleName'] = ''
                    result['productId'] = ''
                    result['saleProduct'] = ''
                    result['spec'] = ''
                    result['chooser'] = ''
                quest = ''
                for re in result['questionReason']:
                    quest = quest + ';' + re
                result['questionReason'] = quest
                delr = ''
                for re in result['delReason']:
                    delr = delr + ';' + re
                result['delReason'] = delr
                auto = ''
                for re in result['autoVerify']:
                    auto = auto + ';' + re
                result['autoVerify'] = auto
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersdict)
        print('******本批次查询成功')
        return data

    #  1.0 单点获取 最近三天订单 删除原因分析
    def order_Query_Delete(self, timeStart, timeEnd, areaId, query, proxy_id, proxy_handle, time_handle):  # 进入订单检索界面
        if time_handle == '自动':
            sql = '''SELECT DISTINCT 下单日期 FROM day_delete_cache_copy1 d GROUP BY 下单日期 ORDER BY 下单日期 DESC;'''.format(
                'day_delete_cache_copy1')
            rq = pd.read_sql_query(sql=sql, con=self.engine1)
            rq = pd.to_datetime(rq['下单日期'][0])
            timeStart = (rq + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            timeEnd = (datetime.datetime.now()).strftime('%Y-%m-%d')

        timeStart = datetime.datetime.strptime(timeStart, '%Y-%m-%d').date()  # 按天循环获取
        timeEnd = datetime.datetime.strptime(timeEnd, '%Y-%m-%d').date()
        for i in range((timeEnd - timeStart).days):  # 按天循环获取订单状态
            day = timeStart + datetime.timedelta(days=i)
            day_time = str(day)
            self.order_Query_Delete_Update(day_time, day_time, areaId, query, proxy_id, proxy_handle)

    #  1.1 单点获取 最近三天订单 删除原因分析
    def order_Query_Delete_Update(self, timeStart, timeEnd, areaId, query, proxy_id, proxy_handle):  # 进入订单检索界面
        print('+++正在检查 ' + timeStart + ' 到 ' + timeEnd + ' 号 订单删除 信息中')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500, 'order_number': None, 'shippingNumber': None, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None, 'emailStatus': None,
                'befrom': None, 'areaId': areaId, 'reassignmentType': None, 'lowerstatus': '', 'warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None,
                'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': None, 'autoVerifyStatus': None, 'shipZip': None,
                'remark': None, 'shipState': None, 'weightStart': None, 'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'timeStart': timeStart +' 00:00:00', 'timeEnd': timeEnd + ' 23:59:59'}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        max_count = req['data']['count']
        print(max_count)
        if max_count > 0:
            in_count = math.ceil(max_count/500)
            df = pd.DataFrame([])
            dlist = []
            n = 1
            while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                print('查询第 ' + str(n) + ' 页中，剩余次数' + str(in_count - n))
                data = self._timeQuery(timeStart, timeEnd, n, areaId, query, proxy_id, proxy_handle)
                dlist.append(data)
                n = n + 1
            print('正在写入......')
            dp = df.append(dlist, ignore_index=True)
            db0 = dp[(dp['运营团队'].str.contains('神龙家族-港澳台|神龙家族-台湾|神龙-香港|火凤凰-港澳台|火凤凰-台湾|火凤凰-香港'))]
            db0 = db0[(db0['币种'].str.contains('台币'))]
            print('正在导入临时表中......')
            db0 = db0[['订单编号', '平台', '币种', '运营团队', '产品id', '产品名称', '收货人', '联系电话', '标准电话', '拉黑率', '配送地址', '应付金额', '数量', '订单状态', '运单号', '支付方式',
                       '下单时间', '审核人', '审核时间', '物流渠道', '货物类型', '订单类型', '物流状态', '重量', '删除原因', '问题原因', '下单人', '备注', 'IP', '体积',
                       '审单类型', '异常提示', '克隆人', '克隆ID', '发货仓库', '拒收原因', '物流更新时间', '状态时间', '更新时间']]
            db0.insert(0, '删除人', '')
            db0.to_sql('cache', con=self.engine1, index=False, if_exists='replace')
            db0.to_excel('F:\\输出文件\\{0} 神龙-火凤凰 删单明细.xlsx'.format(timeEnd), sheet_name='查询', index=False, engine='xlsxwriter')

            print('正在分析订单删除的原因......')
            sql = '''DELETE FROM `cache` gt WHERE gt.`订单编号` IN (SELECT * FROM gat_地址邮编错误);'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            self.del_people(proxy_id, proxy_handle)
            self.del_order_day_new(timeStart)

            print('查询已导出+++')
        else:
            print('无信息+++')

    # 1.2 分析订单删除的原因
    def del_people(self, proxy_id, proxy_handle):
        print('正在更新 订单删除 信息……………………………………………………………………………………………………………………………………………………………………………………')
        start = datetime.datetime.now()
        sql = '''SELECT 订单编号 FROM {0} s WHERE s.`订单状态` = '已删除';'''.format('cache')
        db = pd.read_sql_query(sql=sql, con=self.engine1)
        if db.empty:
            print('无需要更新订单信息！！！')
            return
        print(db['订单编号'][0])
        orderId = list(db['订单编号'])
        max_count = len(orderId)  # 使用len()获取列表的长度，上节学的
        if max_count > 0:
            df = pd.DataFrame([])
            dlist = []
            n = 1
            in_count = math.ceil(max_count / 100)
            while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                print('查询第 ' + str(n) + ' 页中，剩余次数' + str(in_count - n))
                data = self._del_people(n, None, proxy_id, proxy_handle)
                dlist.append(data)
                n = n + 1
            print('正在写入......')
            dp = df.append(dlist, ignore_index=True)
            # n = 0
            # while n <= max_count:  # 这里用到了一个while循环，穿越过来的
            #     n = n + 500
            #     ord = ','.join(orderId[n:n + 500])
            #     data = self._del_people(n, ord, None, proxy_id, proxy_handle)
            #     dlist.append(data)
            # print('正在写入......')
            # dp = df.append(dlist, ignore_index=True)
            dp.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
            print('正在更新删除原因中......')
            sql = '''update cache a, cache_cp b set a.`删除人`= IF(b.`删除人` = '', NULL, b.`删除人`) where a.`订单编号`=b.`订单编号`;'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            sql = '''update `cache` a
                    SET a.`删除人` = IF( a.`删除人` = '', NULL, a.`删除人` ),
                        a.`订单状态` = IF( a.`订单状态` = '', NULL, a.`订单状态` ),
                        a.`审核人` = IF( a.`审核人` = '', NULL, a.`审核人` ),
                        a.`删除原因` = IF( a.`删除原因` = '', NULL, a.`删除原因` ),
                        a.`问题原因` = IF( a.`问题原因` = '', NULL, a.`问题原因` ),
                        a.`下单人` = IF( a.`下单人` = '', NULL, a.`下单人` ),
                        a.`IP` = IF( a.`IP` = '', NULL, a.`IP` ),
                        a.`删除原因` = IF(a.`删除原因` LIKE ';%', RIGHT(a.`删除原因`,CHAR_LENGTH(a.`删除原因`)-1), a.`删除原因`);'''
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        else:
            print('查询为空，不需更新+++')
        print('查询耗时：', datetime.datetime.now() - start)
    def _del_people(self, n, areaId, proxy_id, proxy_handle):  # 进入订单检索界面
        # print('......正在查询信息中......')
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getRemoveOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/deletedOrder'}
        data = {'page': n, 'pageSize': 100, 'orderPrefix': None, 'shipUsername': None, 'shippingNumber': None, 'email': None, 'saleIds': None, 'ip': None,
                'productIds': None, 'phone': None, 'optimizer': None, 'payment': None, 'type': None, 'collId': None, 'isClone': None, 'currencyId': None,
                'emailStatus': None, 'befrom': None, 'areaId': areaId, 'orderStatus': None, 'timeStart': None, 'timeEnd': None, 'payType': None,
                'questionId': None, 'reassignmentType': None, 'delUserId': None, 'delReasonIds': None,'delTimeStart':None, 'delTimeEnd': None}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        # print('......已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        try:
            for result in req['data']['list']:
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersdict)
        df = None
        try:
            df = data[['orderNumber', 'username']]
            df.columns = ['订单编号', '删除人']
        except Exception as e:
            print('------查询为空')
        print('******本批次查询成功')
        return df # 更新删除订单的原因 -函数调用
    # 1.3 分析订单删除的  分析导出
    def del_order_day_new(self, timeStart):
        print('正在分析 ' + timeStart + ' 删单原因中')
        sql ='''SELECT 币种, 运营团队, 删单原因, 订单量, 总订单量, 总删单量, 系统删单量,
                        concat(ROUND(SUM(IF(删单原因 IS NULL OR 删单原因 = '',总订单量-订单量,订单量)) / SUM(总订单量) * 100,2),'%') as '删单率'
                FROM (SELECT s1.*,总订单量,总删单量, 系统删单量
                      FROM (SELECT 币种,运营团队,删单原因,COUNT(订单编号) AS 订单量
                            FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                  FROM `cache` c
                            ) w
                            GROUP BY 币种,运营团队,删单原因
                      ) s1
                      LEFT JOIN
                      ( SELECT 币种,运营团队,COUNT(订单编号) AS 总订单量, SUM(IF(订单状态 = '已删除',1,0)) AS 总删单量, SUM(IF(订单状态 = '已删除' AND 删除人 IS NULL,1,0)) AS 系统删单量
                        FROM `cache` w
                        GROUP BY 币种,运营团队
                      ) s2 
                      ON s1.`币种`=s2.`币种` AND s1.`运营团队`=s2.`运营团队`
                ) s
                WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港')
                GROUP BY 币种,运营团队,删单原因
                ORDER BY FIELD(币种,'台币','港币','合计'),
                         FIELD(运营团队,'神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港','神龙-运营1组','Line运营','金鹏家族-小虎队','合计'),
                         订单量 DESC;'''
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        df1.to_sql('cache_online', con=self.engine1, index=False, if_exists='replace')

        sql = '''SELECT 币种, 运营团队, 删单原因,  订单量, 总订单量, 总删单量, 系统删单量, 删单率
                FROM (  SELECT dt.*, 
                                @_rn := if(@_prev_a = 币种 and @_prev_b = 运营团队, @_rn + 1, 1) as _rn,
                                @_prev_a := 币种 as _prev_a,  @_prev_b := 运营团队 as _prev_b
                        FROM cache_online dt 
                        JOIN (
                                SELECT @_rn:=0, @_prev_a:=null, @_prev_b:=null, @_prev_c:=null
                        ) tmp  
                ) s1 where _rn <= 5;'''
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)

        # print(df1)
        # 初始化设置
        sl_tem, hfh_tem = '', ''
        sl_tem_lh, hfh_tem_lh = '\n            其中占比较多的是：', '\n            其中占比较多的是：'
        for date, row in df1.iterrows():
            tem = row['运营团队']
            delreson = row['删单原因']
            count = row['订单量']
            count_z = row['总订单量']
            count_zd = row['总删单量']
            count_zdl = row['删单率']
            count_xtdl = row['系统删单量']
            if tem == '神龙家族-台湾':
                if delreson == None:
                    sl_tem = '*神  龙:   有效单量：' + str(int(count_z)) + ';  删单量：' + str(int(count_zd)) + ';  删单率：' + str(count_zdl) + ';  系统删单量：' + str(int(count_xtdl)) + '单; '
                else:
                    sl_tem_lh = sl_tem_lh + delreson + '：' + str(int(count)) + '单; '
            elif tem == '火凤凰-台湾':
                if delreson == None:
                    hfh_tem = '*火凤凰:   有效单量：' + str(int(count_z)) + ';  删单量：' + str(int(count_zd)) + ';  删单率：' + str(count_zdl) + ';  系统删单量：' + str(int(count_xtdl)) + '单; '
                else:
                    hfh_tem_lh = hfh_tem_lh + delreson + '：' + str(int(count)) + '单; '
        print('*' * 50)
        print(sl_tem + sl_tem_lh)
        print(hfh_tem + hfh_tem_lh)

        print('正在获取 删单明细 拉黑率信息 二…………')
        sql ='''SELECT 币种, 删单原因, IF(标准电话 = '总计',NULL,标准电话) AS 拉黑率订单, 订单量, 拉黑率70以上, 拉黑率70以下, 下单日期
                FROM(
                        (	SELECT s1.*
                            FROM (SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(标准电话,'总计') 标准电话, COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下, 下单日期
                                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因, DATE_FORMAT(下单时间, '%Y-%m-%d' ) AS 下单日期
                                            FROM `cache` c
                                            WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除原因 LIKE '拉黑率%'
                                    ) w
                                    GROUP BY 币种,删单原因, 标准电话
                                    WITH ROLLUP
                            )  s1
                            WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                            GROUP BY 币种,删单原因, 标准电话
                            ORDER BY 订单量 desc
                            LIMIT 5
                        ) 
                        UNION ALL
                        (   SELECT s1.*
                            FROM (  SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(ip,'总计') ip, COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下, 下单日期
                                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因, DATE_FORMAT(下单时间, '%Y-%m-%d' ) AS 下单日期
                                            FROM `cache` c
                                            WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除原因 LIKE '拉黑率%'
                                    ) w
                                    GROUP BY 币种,删单原因, ip
                                    WITH ROLLUP
                            )  s1
                            WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                            GROUP BY 币种,删单原因, ip
                            ORDER BY 订单量 desc
                            LIMIT 5
                        ) 
                ) s;'''
        df2 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df2)
        df2.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        print('正在记录 拉黑率信息......')
        sql = '''REPLACE INTO {0}(币种, 删单类型, 删单明细, 单量, 下单日期, 记录日期, 更新时间) SELECT 币种, 删单原因, 拉黑率订单, 订单量, 下单日期, CURDATE() 记录日期, NOW() 更新时间 FROM cache_cp s WHERE s.拉黑率订单 IS NOT NULL;'''.format('day_delete_cache_copy1')
        pd.read_sql_query(sql=sql, con=self.engine1,chunksize=10000)

        sl_Black ,sl_Black_iphone , sl_Black_ip = '','',''
        k = 0
        k2 = 0
        for row in df2.itertuples():
            tem_Black = getattr(row, '拉黑率订单')
            count = getattr(row, '订单量')
            if tem_Black == None:
                sl_Black = '*拉黑率删除:  ' + str(int(getattr(row, '订单量'))) + '单；拉黑率70以上的：' + str(int(getattr(row, '拉黑率70以上'))) +'单；'
            else:
                if '.' not in tem_Black:
                    if k == 0:
                        sl_Black_iphone = sl_Black + '\n           同一电话有：(' + str(int(tem_Black)) + ':' + str(int(count)) +'单), '
                        k = k + 1
                    elif k > 0:
                        sl_Black_iphone =sl_Black_iphone + '(' + str(int(tem_Black)) + ':' + str(int(count)) + '单), '
                        k = k + 1
                elif '.' in tem_Black:
                    if k2 == 0:
                        sl_Black_ip = sl_Black_iphone + '\n           同一ip有：   (' + str(tem_Black) + ':' + str(int(count)) +'单), '
                        k2 = k2 + 1
                    elif k > 0:
                        sl_Black_ip = sl_Black_ip + '（' + str(tem_Black) + ':' + str(int(count)) + '单);'
                        k2 = k2 + 1
        print('*' * 50)
        print(sl_Black_ip)

        print('正在获取 删单明细 恶意删除信息 三…………')
        sql ='''SELECT 币种, 删单原因, IF(标准电话 = '总计',NULL,标准电话) AS 恶意删除, 订单量, 拉黑率70以上, 拉黑率70以下, 下单日期
                FROM(
                    (SELECT s1.*
                        FROM (  SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(标准电话,'总计') 标准电话,COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下, 下单日期
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因, DATE_FORMAT(下单时间, '%Y-%m-%d' ) AS 下单日期
                                        FROM `cache` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除原因 LIKE '恶意%'
                                ) w
                                GROUP BY 币种,删单原因, 标准电话
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                        GROUP BY 币种,删单原因, 标准电话
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                    UNION ALL
                    (SELECT s1.*
                        FROM (  SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(ip,'总计') ip,COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下, 下单日期
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因, DATE_FORMAT(下单时间, '%Y-%m-%d' ) AS 下单日期
                                        FROM `cache` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除原因 LIKE '恶意%'
                                ) w
                                GROUP BY 币种,删单原因, ip
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                        GROUP BY 币种,删单原因, ip
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                ) s;'''
        df3 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df3)
        df3.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        print('正在记录 恶意删除信息......')
        sql = '''REPLACE INTO {0}(币种, 删单类型, 删单明细, 单量, 下单日期, 记录日期, 更新时间) SELECT 币种, 删单原因, 恶意删除, 订单量, 下单日期, CURDATE() 记录日期, NOW() 更新时间 FROM cache_cp s WHERE s.恶意删除 IS NOT NULL;'''.format('day_delete_cache_copy1')
        pd.read_sql_query(sql=sql, con=self.engine1,chunksize=10000)

        st_ey, st_ey_iphone, st_ey_ip = '','',''
        k = 0
        k2 = 0
        for row in df3.itertuples():
            tem_Black = getattr(row, '恶意删除')
            count = getattr(row, '订单量')
            if tem_Black == None:
                st_ey = '*恶意删除： ' + str(int(count)) + '单；拉黑率70以上的：' + str(int(getattr(row, '拉黑率70以上'))) + '单；低于70的：' + str(int(getattr(row, '拉黑率70以下'))) + '单；'
            else:
                if '.' not in tem_Black:
                    if k == 0:
                        st_ey_iphone = st_ey + ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) +'单), '
                        k = k + 1
                    elif k > 0:
                        st_ey_iphone = st_ey_iphone + '(' + str(tem_Black) + ': ' + str(int(count)) + '单), '
                        k = k + 1
                if '.' in tem_Black:
                    if k2 == 0:
                        st_ey_ip = st_ey_iphone + ';\n           同一ip有：    (' + str(tem_Black) + ': ' + str(int(count)) +'单), '
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_ey_ip = st_ey_ip + '(' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
        print('*' * 50)
        print(st_ey_ip)

        print('正在获取 删单明细 重复删除信息 四…………')
        sql = '''SELECT IFNULL(币种,'币种') 币种,IFNULL(删单原因,'总计') 重复删除,COUNT(订单编号) AS 订单量
                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                            FROM `cache` c
                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除原因 LIKE '重复订单%'
                    ) w
                GROUP BY 币种,删单原因;'''
        df4 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df4)
        cf_del = ''
        for row in df4.itertuples():
            tem_Black = getattr(row, '重复删除')
            count = getattr(row, '订单量')
            # cf_del = '*重复删除：' + str(count) + '单, 查询后都是客户上笔未收到或者是连续订多笔订单重复删除；'
            cf_del = '*重复删除：' + str(count) + '单；'
        print('*' * 50)
        print(cf_del)

        print('正在获取 删单明细 系统删除信息 五…………')
        sql = '''SELECT s1.*
                        FROM ( 
        					    (SELECT *
                                    FROM (SELECT IFNULL(币种,NULL) 币种,IFNULL(删单原因,NULL) 系统删除,COUNT(订单编号) AS 订单量, 下单日期
                                            FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因, DATE_FORMAT(下单时间, '%Y-%m-%d' ) AS 下单日期
                                                    FROM `cache` c
                                                    WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除人 IS NULL
                                            ) w
                                            GROUP BY 币种,删单原因
                                            WITH ROLLUP
                                    ) w1
                                    WHERE 币种 <> '币种'
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                                UNION ALL
                                 ( SELECT 币种,标准电话 AS 系统删除,COUNT(订单编号) AS 订单量, 下单日期
                                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因, DATE_FORMAT(下单时间, '%Y-%m-%d' ) AS 下单日期
                                            FROM `cache` c
                                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除人 IS NULL
                                    ) w
                                    GROUP BY 币种,标准电话
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                                UNION ALL
                                (SELECT 币种,ip AS 系统删除,COUNT(订单编号) AS 订单量, 下单日期
                                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因, DATE_FORMAT(下单时间, '%Y-%m-%d' ) AS 下单日期
                                            FROM `cache` c
                                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除人 IS NULL
                                    ) w
                                    GROUP BY 币种,ip
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                        ) s1;'''
        df5 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df5)
        df5.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        print('正在记录 系统删除信息......')
        sql = '''REPLACE INTO {0}(币种, 删单类型, 删单明细, 单量, 下单日期, 记录日期, 更新时间) 
                SELECT 币种,'系统删除'  删单类型, 系统删除, 订单量, 下单日期, CURDATE() 记录日期, NOW() 更新时间
				FROM cache_cp s 
				WHERE s.系统删除 LIKE '%.%' OR s.系统删除 LIKE '%9%' OR s.系统删除 LIKE '%8%' OR s.系统删除 LIKE '%7%' OR s.系统删除 LIKE '%6%' OR s.系统删除 LIKE '%5%' 
				   OR s.系统删除 LIKE '%4%' OR s.系统删除 LIKE '%3%' OR s.系统删除 LIKE '%2%' OR s.系统删除 LIKE '%1%' OR s.系统删除 LIKE '%0%';'''.format('day_delete_cache_copy1')
        pd.read_sql_query(sql=sql, con=self.engine1,chunksize=10000)

        st_del, st_del_iphone, st_del_ip = '', '', ''
        k = 0
        k2 = 0
        for row in df5.itertuples():
            tem_Black = getattr(row, '系统删除')
            count = getattr(row, '订单量')
            if tem_Black == None:
                st_del = '*系统删除： ' + str(int(count)) + '单；其中比较多的是：'
            elif '订单' in tem_Black:
                st_del = st_del + str(tem_Black) + ':' + str(int(count)) + '单, '
            else:
                if '.' not in tem_Black:
                    if k == 0:
                        st_del_iphone = st_del + ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单), '
                        k = k + 1
                    elif k > 0:
                        st_del_iphone = st_del_iphone + '(' + str(tem_Black) + ': ' + str(int(count)) + '单), '
                        k = k + 1
                if '.' in tem_Black:
                    if k2 == 0:
                        st_del_ip = st_del_iphone + ';\n           同一ip有：   (' + str(tem_Black) + ': ' + str(int(count)) + '单), '
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_del_ip = st_del_ip + '(' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
        print('*' * 50)
        print(st_del_ip)

        print('正在获取 连续3天以上的黑名单 电话、IP信息 六…………')
        sql = '''SELECT *
                FROM (
                        SELECT 币种,删单类型,删单明细, COUNT(下单日期) AS 次数
                        FROM day_delete_cache_copy1 d
                        WHERE d.`下单日期` >= IF(DATE_FORMAT('2022-06-29','%w') = 2 ,DATE_SUB(CURDATE(), INTERVAL 6 DAY),IF(DATE_FORMAT('2022-06-29','%w') = 1 ,DATE_SUB(CURDATE(), INTERVAL 7 DAY),DATE_SUB(CURDATE(), INTERVAL 3 DAY)))
                        GROUP BY 币种,删单类型,删单明细
                        ORDER BY 币种,删单类型,删单明细, 次数 DESC
                ) s
                WHERE s.次数 >= 3 AND s.删单类型 <> '系统删除'
                ORDER BY 币种,删单类型,次数 DESC;'''
        df6 = pd.read_sql_query(sql=sql, con=self.engine1)
        db61 = df6[~(df6['删单类型'].str.contains('系统删除'))]
        # day_del = db61.to_markdown()
        day_del = '''注意：连续3天同电话\IP的信息>>>'''
        day_del2 = '恶意订单:'
        day_del3 = '拉黑率订单:'
        for row in db61.itertuples():
            tem = getattr(row, '删单类型')
            info = getattr(row, '删单明细')
            count = getattr(row, '次数')
            if '恶意订单' in tem:
                day_del2 = day_del2 + '\n' + info + ':  ' + str(int(count)) + '单,'
            elif '拉黑率订单' in tem:
                day_del3 = day_del3 + '\n' + info + ':  ' + str(int(count)) + '单,'
        day_del = day_del + '\n' + day_del2 + '\n' + day_del3
        print('*' * 50)
        print(day_del)


        # url = "https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e"  # url为机器人的webhook  小嗨
        url = "https://oapi.dingtalk.com/robot/send?access_token=9a92f00296846dcd3ec8b52d7bacce114a9e34cb2d5dbfad9ce3371ab8d037f9"  # url为机器人的webhook  港台客服
        content = r'r"H:\桌面\需要用到的文件\文件夹\out2.jpeg"'  # 钉钉消息内容，注意test是自定义的关键字，需要在钉钉机器人设置中添加，这样才能接收到消息
        mobile_list = ['18538110674']  # 要@的人的手机号，可以是多个，注意：钉钉机器人设置中需要添加这些人，否则不会接收到消息
        isAtAll = '是'  # 是否@所有人
        headers = {'Content-Type': 'application/json', "Charset": "UTF-8"}
        data = {"msgtype": "text",
                # "markdown": {# 要发送的内容【支持markdown】【！注意：content内容要包含机器人自定义关键字，不然消息不会发送出去，这个案例中是test字段】
                #         "title": 'TEST',
                #         "text": "#### 昨日删单率分析" + "\n" +
                #         "* " + sl_tem +
                #         "   + " + sl_tem_lh + sl_tem_ey + sl_tem_cf + "\n" +
                #         "* " + hfh_tem +
                #         "   + " + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf
                # },
                "text": {        # 要发送的内容【支持markdown】【！注意：content内容要包含机器人自定义关键字，不然消息不会发送出去，这个案例中是test字段】
                    "content": timeStart + ' >>> 神龙 - 火凤凰 <<< 台湾订单 删单分析' + '\n' + '\n' +
                               sl_tem + sl_tem_lh + '\n' +
                               hfh_tem + hfh_tem_lh + '\n' +
                               sl_Black_ip + '\n' +
                               st_ey_ip + '\n' +
                               cf_del + '\n' +
                               st_del_ip + '\n' + '\n' +
                               day_del
                    # "content": 'TEST'
                },
                "at": {# 要@的人
                        # "atMobiles": mobile_list,
                        # 是否@所有人
                        "isAtAll": False  # @全体成员（在此可设置@特定某人）
                }
        }
        # 4、对请求的数据进行json封装
        sendData = json.dumps(data)  # 将字典类型数据转化为json格式
        sendData = sendData.encode("utf-8")  # python3的Request要求data为byte类型
        r = requests.post(url, headers=headers, data=json.dumps(data))
        req = json.loads(r.text)  # json类型数据转换为dict字典
        print(req['errmsg'])


    def del_order_day(self):
        print('正在分析 昨日 删除原因')
        sql = '''SELECT *, concat(ROUND(SUM(IF(删单原因 IS NULL OR 删单原因 = '',总订单量-订单量,订单量)) / SUM(总订单量) * 100,2),'%') as '删单率'
                FROM (SELECT s1.*, 总订单量, 总删单量, 系统删单量
                      FROM (SELECT 币种, 运营团队, 删单原因, COUNT(订单编号) AS 订单量
                            FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                  FROM `cache` c
                            ) w
                            GROUP BY 币种,运营团队,删单原因
                      ) s1
                      LEFT JOIN
                      ( SELECT 币种,运营团队,COUNT(订单编号) AS 总订单量, SUM(IF(订单状态 = '已删除',1,0)) AS 总删单量, SUM(IF(订单状态 = '已删除' AND 删除人 IS NULL,1,0)) AS 系统删单量
                        FROM `cache` w
                        GROUP BY 币种,运营团队
                      ) s2 
                      ON s1.`币种`=s2.`币种` AND s1.`运营团队`=s2.`运营团队`
                ) s
                WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港')
                GROUP BY 币种,运营团队,删单原因
                ORDER BY FIELD(币种,'台币','港币','合计'),
                         FIELD(运营团队,'神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港','神龙-运营1组','Line运营','金鹏家族-小虎队','合计'),
                         订单量 DESC;'''
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df1)
        # 初始化设置
        sl_tem, sl_tem_lh, sl_tem_ey, sl_tem_cf = '', '', '', ''
        hfh_tem, hfh_tem_lh, hfh_tem_ey, hfh_tem_cf = '', '', '', ''
        for row in df1.itertuples():
            tem = getattr(row, '运营团队')
            delreson = getattr(row, '删单原因')
            count = getattr(row, '订单量')
            if tem == '神龙家族-台湾' and delreson == None:
                sl_tem = '*神  龙:   昨日单量：' + str(int(getattr(row, '总订单量'))) + '；删单量：' + str(
                    int(getattr(row, '总删单量'))) + '；删单率：' + str(getattr(row, '删单率')) + '；系统删单量：' + str(
                    int(getattr(row, '系统删单量'))) + '单;'
                # print(sl_tem)
            elif tem == '神龙家族-台湾' and '拉黑率订单' == delreson:
                sl_tem_lh = '，\n            其中占比较多的是：拉黑率订单：' + str(int(count)) + '单, '
                # print(sl_tem_lh)
            elif tem == '神龙家族-台湾' and '恶意订单' == delreson:
                sl_tem_ey = '恶意订单：' + str(int(count)) + '单, '
                # print(sl_tem_ey)
            elif tem == '神龙家族-台湾' and '重复订单' == delreson:
                sl_tem_cf = '重复订单：' + str(int(count)) + '单;'
                # print(sl_tem_cf)

            elif tem == '火凤凰-台湾' and delreson == None:
                hfh_tem = '*火凤凰:  昨日单量：' + str(int(getattr(row, '总订单量'))) + '；删单量：' + str(
                    int(getattr(row, '总删单量'))) + '；删单率：' + str(getattr(row, '删单率')) + '；系统删单量：' + str(
                    int(getattr(row, '系统删单量'))) + '单;'
                # print(hfh_tem)
            elif tem == '火凤凰-台湾' and '拉黑率订单' == delreson:
                hfh_tem_lh = '，\n            其中占比较多的是：拉黑率订单：' + str(int(count)) + '单, '
                # print(hfh_tem_lh)
            elif tem == '火凤凰-台湾' and '恶意订单' == delreson:
                hfh_tem_ey = '恶意订单：' + str(int(count)) + '单, '
                # print(hfh_tem_ey)
            elif tem == '火凤凰-台湾' and '重复订单' == delreson:
                hfh_tem_cf = '重复订单：' + str(int(count)) + '单;'
                # print(hfh_tem_cf)
        print('*' * 50)
        print(sl_tem + sl_tem_lh + sl_tem_ey + sl_tem_cf)
        print(hfh_tem + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf)

        print('正在获取 删单明细 拉黑率信息 二…………')
        sql = '''SELECT 币种, 删除原因, IF(联系电话 = '总计',NULL,联系电话) AS 拉黑率订单, 订单量, 拉黑率70以上, 拉黑率70以下
                FROM(
                    (SELECT s1.*
                        FROM (SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(联系电话,'总计') 联系电话, COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                        FROM `cache` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除原因 LIKE '拉黑率%'
                                ) w
                                GROUP BY 币种,删单原因, 联系电话
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                        GROUP BY 币种,删单原因, 联系电话
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                    UNION ALL
                    (SELECT s1.*
                        FROM (  SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(ip,'总计') ip, COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                        FROM `cache` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除原因 LIKE '拉黑率%'
                                ) w
                                GROUP BY 币种,删单原因, ip
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                        GROUP BY 币种,删单原因, ip
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                ) s;'''
        df2 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df2)
        df2.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        print('正在记录 拉黑率信息......')
        sql = '''REPLACE INTO {0}(币种, 删单类型, 删单明细, 单量, 记录日期, 更新时间) SELECT 币种, 删单原因, 拉黑率订单, 订单量, CURDATE() 记录日期, NOW() 更新时间 FROM cache_cp s WHERE s.拉黑率订单 IS NOT NULL;'''.format(
            'day_delete_cache')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        sl_Black, sl_Black_iphone, sl_Black_ip = '', '', ''
        k = 0
        k2 = 0
        for row in df2.itertuples():
            tem_Black = getattr(row, '拉黑率订单')
            count = getattr(row, '订单量')
            if tem_Black == None:
                sl_Black = '*拉黑率删除:  ' + str(int(getattr(row, '订单量'))) + '单；拉黑率70以上的：' + str(
                    int(getattr(row, '拉黑率70以上'))) + '单；'
                # print(sl_Black)
            elif tem_Black != None and '.' not in tem_Black:
                if count >= 10:
                    if k == 0:
                        sl_Black_iphone = '\n           同一电话有：(0' + str(int(tem_Black)) + ':' + str(
                            int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        sl_Black_iphone = sl_Black_iphone + '(0' + str(int(tem_Black)) + ':' + str(
                            int(count)) + '单);'
                        k = k + 1
                else:
                    if k == 0:
                        sl_Black_iphone = '\n           同一电话有：(0' + str(int(tem_Black)) + ':' + str(
                            int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        sl_Black_iphone = sl_Black_iphone + '(0' + str(int(tem_Black)) + ':' + str(
                            int(count)) + '单);'
                        k = k + 1
                # print(sl_Black_iphone)
            elif tem_Black != None and '.' in tem_Black:
                if count >= 10:
                    if k2 == 0:
                        sl_Black_ip = '\n           同一ip有：   (' + str(getattr(row, '拉黑率订单')) + ':' + str(
                            int(getattr(row, '订单量'))) + '单),'
                        k2 = k2 + 1
                    elif k > 0:
                        sl_Black_ip = sl_Black_ip + '(' + str(tem_Black) + ':' + str(int(count)) + '单);'
                        k2 = k2 + 1
                else:
                    if k2 == 0:
                        sl_Black_ip = '\n           同一ip有：   (' + str(tem_Black) + ':' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k > 0:
                        sl_Black_ip = sl_Black_ip + '（' + str(tem_Black) + ':' + str(int(count)) + '单);'
                        k2 = k2 + 1
                # print(sl_Black_ip)
        print('*' * 50)
        print(sl_Black + sl_Black_iphone + sl_Black_ip)

        print('正在获取 删单明细 恶意删除信息 三…………')
        sql = '''SELECT 币种, 删除原因, IF(联系电话 = '总计',NULL,联系电话) AS 恶意删除, 订单量, 拉黑率70以上, 拉黑率70以下
                FROM(
                    (SELECT s1.*
                        FROM (  SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(联系电话,'总计') 联系电话,COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                        FROM `cache` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除原因 LIKE '恶意%'
                                ) w
                                GROUP BY 币种,删单原因, 联系电话
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                        GROUP BY 币种,删单原因, 联系电话
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                    UNION ALL
                    (SELECT s1.*
                        FROM (  SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(ip,'总计') ip,COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                        FROM `cache` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除原因 LIKE '恶意%'
                                ) w
                                GROUP BY 币种,删单原因, ip
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                        GROUP BY 币种,删单原因, ip
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                ) s;'''
        df3 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df3)
        df3.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        print('正在记录 恶意删除信息......')
        sql = '''REPLACE INTO {0}(币种, 删单类型, 删单明细, 单量, 记录日期, 更新时间) SELECT 币种, 删单原因, 恶意删除, 订单量, CURDATE() 记录日期, NOW() 更新时间 FROM cache_cp s WHERE s.恶意删除 IS NOT NULL;'''.format(
            'day_delete_cache')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        st_ey, st_ey_iphone, st_ey_ip = '', '', ''
        k = 0
        k2 = 0
        for row in df3.itertuples():
            tem_Black = getattr(row, '恶意删除')
            count = getattr(row, '订单量')
            if tem_Black == None:
                st_ey = '*恶意删除： ' + str(int(count)) + '单；拉黑率70以上的：' + str(
                    int(getattr(row, '拉黑率70以上'))) + '单；低于70的：' + str(int(getattr(row, '拉黑率70以下'))) + '单；'
            elif tem_Black != None and '.' not in tem_Black:
                if count >= 10:
                    if k == 0:
                        st_ey_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        st_ey_iphone = st_ey_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1
                else:
                    if k == 0:
                        st_ey_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        st_ey_iphone = st_ey_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1

            elif tem_Black != None and '.' in tem_Black:
                if count >= 10:
                    if k2 == 0:
                        st_ey_ip = ';\n           同一ip有：    (' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_ey_ip = st_ey_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
                else:
                    if k2 == 0:
                        st_ey_ip = ';\n           同一ip有：    (' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_ey_ip = st_ey_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
        print('*' * 50)
        print(st_ey + st_ey_iphone + st_ey_ip)

        print('正在获取 删单明细 重复删除信息 四…………')
        sql = '''SELECT IFNULL(币种,'币种') 币种,IFNULL(删单原因,'总计') 重复删除,COUNT(订单编号) AS 订单量
                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                            FROM `cache` c
                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除原因 LIKE '重复订单%'
                    ) w
                GROUP BY 币种,删除原因;'''
        df4 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df4)
        cf_del = ''
        for row in df4.itertuples():
            tem_Black = getattr(row, '重复删除')
            count = getattr(row, '订单量')
            # cf_del = '*重复删除：' + str(count) + '单, 查询后都是客户上笔未收到或者是连续订多笔订单重复删除；'
            cf_del = '*重复删除：' + str(count) + '单；'
        print('*' * 50)
        print(cf_del)

        print('正在获取 删单明细 系统删除信息 五…………')
        sql = '''SELECT s1.*
                        FROM ( 
                                (SELECT *
                                    FROM (SELECT IFNULL(币种,'币种') 币种,IFNULL(删单原因,'总计') 系统删除,COUNT(订单编号) AS 订单量
                                            FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                                    FROM `cache` c
                                                    WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除人 IS NULL
                                            ) w
                                            GROUP BY 币种,删单原因
                                            WITH ROLLUP
                                    ) w1
                                    WHERE 币种 <> '币种'
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                                UNION ALL
                                 ( SELECT 币种,联系电话 AS 系统删除,COUNT(订单编号) AS 订单量
                                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                            FROM `cache` c
                                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除人 IS NULL
                                    ) w
                                    GROUP BY 币种,联系电话
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                                UNION ALL
                                (SELECT 币种,ip AS 系统删除,COUNT(订单编号) AS 订单量
                                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                            FROM `cache` c
                                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','神龙-香港','火凤凰-台湾','火凤凰-香港') AND 删除人 IS NULL
                                    ) w
                                    GROUP BY 币种,ip
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                        ) s1;'''
        df5 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df5)
        df5.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        print('正在记录 系统删除信息......')
        sql = '''REPLACE INTO {0}(币种, 删单类型, 删单明细, 单量, 记录日期, 更新时间) 
                SELECT 币种,'系统删除'  删单类型, 系统删除, 订单量, CURDATE() 记录日期, NOW() 更新时间
                FROM cache_cp s 
                WHERE s.系统删除 LIKE '%.%' OR s.系统删除 LIKE '%9%' OR s.系统删除 LIKE '%8%' OR s.系统删除 LIKE '%7%' OR s.系统删除 LIKE '%6%' OR s.系统删除 LIKE '%5%' 
                   OR s.系统删除 LIKE '%4%' OR s.系统删除 LIKE '%3%' OR s.系统删除 LIKE '%2%' OR s.系统删除 LIKE '%1%' OR s.系统删除 LIKE '%0%';'''.format(
            'day_delete_cache')
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)

        st_del, st_del_iphone, st_del_ip = '', '', ''
        k = 0
        k2 = 0
        for row in df5.itertuples():
            tem_Black = getattr(row, '系统删除')
            count = getattr(row, '订单量')
            if '总计' in tem_Black:
                st_del = '*系统删除： ' + str(int(count)) + '单；其中比较多的是：'
            elif '订单' in tem_Black:
                st_del = st_del + str(tem_Black) + ':' + str(int(count)) + '单,'

            elif tem_Black != None and '.' not in tem_Black:
                if count >= 10:
                    if k == 0:
                        st_del_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        st_del_iphone = st_del_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1
                else:
                    if k == 0:
                        st_del_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        st_del_iphone = st_del_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1

            elif tem_Black != None and '.' in tem_Black:
                if count >= 10:
                    if k2 == 0:
                        st_del_ip = ';\n           同一ip有：   (' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_del_ip = st_del_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
                else:
                    if k2 == 0:
                        st_del_ip = ';\n           同一ip有：   (' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_del_ip = st_del_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
        print('*' * 50)
        print(st_del + st_del_iphone + st_del_ip)

        print('正在获取 连续3天以上的黑名单 电话、IP信息 六…………')
        sql = '''UPDATE day_delete_cache d 
                    SET d.`删单明细` = IF(d.`删单明细` LIKE '0%', RIGHT(d.`删单明细`,LENGTH(d.`删单明细`)-1),d.`删单明细`);'''
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''SELECT *
                FROM (
                        SELECT 币种,删单类型,删单明细, COUNT(记录日期) AS 次数
                        FROM day_delete_cache d
                        WHERE d.`记录日期` >= IF(DATE_FORMAT('2022-06-29','%w') = 2 ,DATE_SUB(CURDATE(), INTERVAL 5 DAY),IF(DATE_FORMAT('2022-06-29','%w') = 1 ,DATE_SUB(CURDATE(), INTERVAL 6 DAY),DATE_SUB(CURDATE(), INTERVAL 2 DAY)))
                        GROUP BY 币种,删单类型,删单明细
                ) s
                WHERE s.次数 >= 3
                ORDER BY 币种,删单类型,删单明细, 次数 DESC;'''
        df6 = pd.read_sql_query(sql=sql, con=self.engine1)
        db61 = df6[~(df6['删单类型'].str.contains('系统删除'))]
        # day_del = db61.to_markdown()
        day_del = '''注意：连续3天同电话\IP的信息>>>'''
        day_del2 = '恶意订单:'
        day_del3 = '拉黑率订单:'
        for row in db61.itertuples():
            tem = getattr(row, '删单类型')
            info = getattr(row, '删单明细')
            count = getattr(row, '次数')
            # day_del = day_del + '\n' + tem + ':' + info + ':  ' + str(int(count)) + '单,'
            if '恶意订单' in tem:
                day_del2 = day_del2 + '\n' + info + ':  ' + str(int(count)) + '单,'
            elif '拉黑率订单' in tem:
                day_del3 = day_del3 + '\n' + info + ':  ' + str(int(count)) + '单,'

        print('*' * 50)
        print(day_del + '\n' + day_del2 + '\n' + day_del3)

        # url = "https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e"  # url为机器人的webhook
        url = "https://oapi.dingtalk.com/robot/send?access_token=9a92f00296846dcd3ec8b52d7bacce114a9e34cb2d5dbfad9ce3371ab8d037f9"  # url为机器人的webhook  港台客服
        content = r'r"H:\桌面\需要用到的文件\文件夹\out2.jpeg"'  # 钉钉消息内容，注意test是自定义的关键字，需要在钉钉机器人设置中添加，这样才能接收到消息
        mobile_list = ['18538110674']  # 要@的人的手机号，可以是多个，注意：钉钉机器人设置中需要添加这些人，否则不会接收到消息
        isAtAll = '是'  # 是否@所有人
        headers = {'Content-Type': 'application/json', "Charset": "UTF-8"}
        data = {"msgtype": "text",
                # "markdown": {# 要发送的内容【支持markdown】【！注意：content内容要包含机器人自定义关键字，不然消息不会发送出去，这个案例中是test字段】
                #         "title": 'TEST',
                #         "text": "#### 昨日删单率分析" + "\n" +
                #         "* " + sl_tem +
                #         "   + " + sl_tem_lh + sl_tem_ey + sl_tem_cf + "\n" +
                #         "* " + hfh_tem +
                #         "   + " + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf
                # },
                "text": {  # 要发送的内容【支持markdown】【！注意：content内容要包含机器人自定义关键字，不然消息不会发送出去，这个案例中是test字段】
                    "content": '神龙 - 火凤凰 昨日台湾订单 删除分析' + '\n' +
                               sl_tem + sl_tem_lh + sl_tem_ey + sl_tem_cf + '\n' +
                               hfh_tem + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf + '\n' +
                               sl_Black + sl_Black_iphone + sl_Black_ip + '\n' +
                               st_ey + st_ey_iphone + st_ey_ip + '\n' +
                               cf_del + '\n' +
                               st_del + st_del_iphone + st_del_ip + '\n' + '\n' +
                               day_del + '\n' + day_del2 + '\n' + day_del3
                    # "content": 'TEST'
                },
                "at": {  # 要@的人
                    # "atMobiles": mobile_list,
                    # 是否@所有人
                    "isAtAll": False  # @全体成员（在此可设置@特定某人）
                }
                }
        # 4、对请求的数据进行json封装
        sendData = json.dumps(data)  # 将字典类型数据转化为json格式
        sendData = sendData.encode("utf-8")  # python3的Request要求data为byte类型
        r = requests.post(url, headers=headers, data=json.dumps(data))
        req = json.loads(r.text)  # json类型数据转换为dict字典
        print(req['errmsg'])

    def del_order_day_test(self):
        print('正在分析 昨日 删单原因中')
        sql ='''SELECT *,concat(ROUND(SUM(IF(删单原因 IS NULL OR 删单原因 = '',总订单量-订单量,订单量)) / SUM(总订单量) * 100,2),'%') as '删单率'
                FROM (SELECT s1.*,总订单量,总删单量, 系统删单量
                      FROM (SELECT 币种,运营团队,删单原因,COUNT(订单编号) AS 订单量
                            FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                  FROM `cache` c
                            ) w
                            GROUP BY 币种,运营团队,删单原因
                      ) s1
                      LEFT JOIN
                      ( SELECT 币种,运营团队,COUNT(订单编号) AS 总订单量, SUM(IF(订单状态 = '已删除',1,0)) AS 总删单量, SUM(IF(订单状态 = '已删除' AND 删除人 IS NULL,1,0)) AS 系统删单量
                        FROM `cache` w
                        GROUP BY 币种,运营团队
                      ) s2 
                      ON s1.`币种`=s2.`币种` AND s1.`运营团队`=s2.`运营团队`
                ) s
                WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾')
                GROUP BY 币种,运营团队,删单原因
                ORDER BY FIELD(币种,'台币','港币','合计'),
                         FIELD(运营团队,'神龙家族-台湾','火凤凰-台湾','神龙-运营1组','Line运营','金鹏家族-小虎队','合计'),
                         订单量 DESC;'''
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df1)
        # 初始化设置
        sl_tem, sl_tem_lh, sl_tem_ey, sl_tem_cf = '', '', '', ''
        hfh_tem, hfh_tem_lh, hfh_tem_ey, hfh_tem_cf = '', '', '', ''
        for row in df1.itertuples():
            tem = getattr(row, '运营团队')
            delreson = getattr(row, '删单原因')
            count = getattr(row, '订单量')
            if tem == '神龙家族-台湾' and delreson == None:
                sl_tem = '*神  龙:   昨日单量：' + str(int(getattr(row, '总订单量'))) + '；删单量：' + str(int(getattr(row, '总删单量'))) + '；删单率：' + str(getattr(row, '删单率')) + '；系统删单量：' + str(int(getattr(row, '系统删单量'))) + '单;'
                # print(sl_tem)
            elif tem == '神龙家族-台湾' and '拉黑率订单' == delreson:
                sl_tem_lh = '，\n            其中占比较多的是：拉黑率订单：' + str(int(count)) + '单, '
                # print(sl_tem_lh)
            elif tem == '神龙家族-台湾' and '恶意订单' == delreson:
                sl_tem_ey = '恶意订单：' + str(int(count)) + '单, '
                # print(sl_tem_ey)
            elif tem == '神龙家族-台湾' and '重复订单' == delreson:
                sl_tem_cf = '重复订单：' + str(int(count)) + '单;'
                # print(sl_tem_cf)

            elif tem == '火凤凰-台湾' and delreson == None:
                hfh_tem = '*火凤凰:  昨日单量：' + str(int(getattr(row, '总订单量'))) + '；删单量：' + str(int(getattr(row, '总删单量'))) + '；删单率：' + str(getattr(row, '删单率')) + '；系统删单量：' + str(int(getattr(row, '系统删单量'))) + '单;'
                # print(hfh_tem)
            elif tem == '火凤凰-台湾' and '拉黑率订单' == delreson:
                hfh_tem_lh = '，\n            其中占比较多的是：拉黑率订单：' + str(int(count)) + '单, '
                # print(hfh_tem_lh)
            elif tem == '火凤凰-台湾' and '恶意订单' == delreson:
                hfh_tem_ey = '恶意订单：' + str(int(count)) + '单, '
                # print(hfh_tem_ey)
            elif tem == '火凤凰-台湾' and '重复订单' == delreson:
                hfh_tem_cf = '重复订单：' + str(int(count)) + '单;'
                # print(hfh_tem_cf)
        print('*' * 50)
        print(sl_tem + sl_tem_lh + sl_tem_ey + sl_tem_cf)
        print(hfh_tem + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf)

        print('正在获取 删单明细 拉黑率信息 二…………')
        sql ='''SELECT 币种, 删单原因, IF(联系电话 = '总计',NULL,联系电话) AS 拉黑率订单, 订单量, 拉黑率70以上, 拉黑率70以下
                FROM(
                    (SELECT s1.*
					    FROM (SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(联系电话,'总计') 联系电话, COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                        FROM `cache` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '拉黑率%'
                                ) w
                                GROUP BY 币种,删单原因, 联系电话
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                        GROUP BY 币种,删单原因, 联系电话
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                    UNION ALL
                    (SELECT s1.*
                        FROM (  SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(ip,'总计') ip, COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                        FROM `cache` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '拉黑率%'
                                ) w
                                GROUP BY 币种,删单原因, ip
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                        GROUP BY 币种,删单原因, ip
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                ) s;'''
        df2 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df2)
        df2.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        print('正在记录 拉黑率信息......')
        sql = '''REPLACE INTO {0}(币种, 删单类型, 删单明细, 单量, 记录日期, 更新时间) SELECT 币种, 删单原因, 拉黑率订单, 订单量, CURDATE() 记录日期, NOW() 更新时间 FROM cache_cp s WHERE s.拉黑率订单 IS NOT NULL;'''.format('day_delete_cache')
        pd.read_sql_query(sql=sql, con=self.engine1,chunksize=10000)

        sl_Black ,sl_Black_iphone , sl_Black_ip = '','',''
        k = 0
        k2 = 0
        for row in df2.itertuples():
            tem_Black = getattr(row, '拉黑率订单')
            count = getattr(row, '订单量')
            if tem_Black == None:
                sl_Black = '*拉黑率删除:  ' + str(int(getattr(row, '订单量'))) + '单；拉黑率70以上的：' + str(int(getattr(row, '拉黑率70以上'))) +'单；'
                # print(sl_Black)
            elif tem_Black != None and '.' not in tem_Black:
                if count >= 10:
                    if k == 0:
                        sl_Black_iphone = '\n           同一电话有：(0' + str(int(tem_Black)) + ':' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        sl_Black_iphone =sl_Black_iphone + '(0' + str(int(tem_Black)) + ':' + str(int(count)) + '单);'
                        k = k + 1
                else:
                    if k == 0:
                        sl_Black_iphone = '\n           同一电话有：(0' + str(int(tem_Black)) + ':' + str(int(count)) +'单),'
                        k = k + 1
                    elif k > 0:
                        sl_Black_iphone =sl_Black_iphone + '(0' + str(int(tem_Black)) + ':' + str(int(count)) + '单);'
                        k = k + 1
                # print(sl_Black_iphone)
            elif tem_Black != None and '.' in tem_Black:
                if count >= 10:
                    if k2 == 0:
                        sl_Black_ip = '\n           同一ip有：   (' + str(getattr(row, '拉黑率订单')) + ':' + str(int(getattr(row, '订单量'))) + '单),'
                        k2 = k2 + 1
                    elif k > 0:
                        sl_Black_ip = sl_Black_ip + '(' + str(tem_Black) + ':' + str(int(count)) + '单);'
                        k2 = k2 + 1
                else:
                    if k2 == 0:
                        sl_Black_ip = '\n           同一ip有：   (' + str(tem_Black) + ':' + str(int(count)) +'单),'
                        k2 = k2 + 1
                    elif k > 0:
                        sl_Black_ip = sl_Black_ip + '（' + str(tem_Black) + ':' + str(int(count)) + '单);'
                        k2 = k2 + 1
                # print(sl_Black_ip)
        print('*' * 50)
        print(sl_Black + sl_Black_iphone + sl_Black_ip)

        print('正在获取 删单明细 恶意删除信息 三…………')
        sql ='''SELECT 币种, 删单原因, IF(联系电话 = '总计',NULL,联系电话) AS 恶意删除, 订单量, 拉黑率70以上, 拉黑率70以下
                FROM(
                    (SELECT s1.*
                        FROM (  SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(联系电话,'总计') 联系电话,COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                        FROM `cache` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '恶意%'
                                ) w
                                GROUP BY 币种,删单原因, 联系电话
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                        GROUP BY 币种,删单原因, 联系电话
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                    UNION ALL
                    (SELECT s1.*
                        FROM (  SELECT IFNULL(币种,'总计') 币种,IFNULL(删单原因,'总计') 删单原因,IFNULL(ip,'总计') ip,COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                        FROM `cache` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '恶意%'
                                ) w
                                GROUP BY 币种,删单原因, ip
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 <> "总计" AND 删单原因 <> "总计"
                        GROUP BY 币种,删单原因, ip
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                ) s;'''
        df3 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df3)
        df3.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        print('正在记录 恶意删除信息......')
        sql = '''REPLACE INTO {0}(币种, 删单类型, 删单明细, 单量, 记录日期, 更新时间) SELECT 币种, 删单原因, 恶意删除, 订单量, CURDATE() 记录日期, NOW() 更新时间 FROM cache_cp s WHERE s.恶意删除 IS NOT NULL;'''.format('day_delete_cache')
        pd.read_sql_query(sql=sql, con=self.engine1,chunksize=10000)

        st_ey, st_ey_iphone, st_ey_ip = '','',''
        k = 0
        k2 = 0
        for row in df3.itertuples():
            tem_Black = getattr(row, '恶意删除')
            count = getattr(row, '订单量')
            if tem_Black == None:
                st_ey = '*恶意删除： ' + str(int(count)) + '单；拉黑率70以上的：' + str(int(getattr(row, '拉黑率70以上'))) + '单；低于70的：' + str(int(getattr(row, '拉黑率70以下'))) + '单；'
            elif tem_Black != None and '.' not in tem_Black:
                if count >= 10:
                    if k == 0:
                        st_ey_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        st_ey_iphone = st_ey_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1
                else:
                    if k == 0:
                        st_ey_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) +'单),'
                        k = k + 1
                    elif k > 0:
                        st_ey_iphone = st_ey_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1

            elif tem_Black != None and '.' in tem_Black:
                if count >= 10:
                    if k2 == 0:
                        st_ey_ip = ';\n           同一ip有：    (' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_ey_ip = st_ey_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
                else:
                    if k2 == 0:
                        st_ey_ip = ';\n           同一ip有：    (' + str(tem_Black) + ': ' + str(int(count)) +'单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_ey_ip = st_ey_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
        print('*' * 50)
        print(st_ey + st_ey_iphone + st_ey_ip)

        print('正在获取 删单明细 重复删除信息 四…………')
        sql = '''SELECT IFNULL(币种,'币种') 币种,IFNULL(删单原因,'总计') 重复删除,COUNT(订单编号) AS 订单量
                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                            FROM `cache` c
                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '重复订单%'
                    ) w
                GROUP BY 币种,删单原因;'''
        df4 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df4)
        cf_del = ''
        for row in df4.itertuples():
            tem_Black = getattr(row, '重复删除')
            count = getattr(row, '订单量')
            cf_del = '*重复删除：' + str(count) + '单, 查询后都是客户上笔未收到或者是连续订多笔订单重复删除；'
        print('*' * 50)
        print(cf_del)

        print('正在获取 删单明细 系统删除信息 五…………')
        sql = '''SELECT s1.*
                        FROM ( 
        					    (SELECT *
                                    FROM (SELECT IFNULL(币种,'币种') 币种,IFNULL(删单原因,'总计') 系统删除,COUNT(订单编号) AS 订单量
                                            FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                                    FROM `cache` c
                                                    WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除人 IS NULL
                                            ) w
                                            GROUP BY 币种,删单原因
                                            WITH ROLLUP
                                    ) w1
                                    WHERE 币种 <> '币种'
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                                UNION ALL
                                 ( SELECT 币种,联系电话 AS 系统删除,COUNT(订单编号) AS 订单量
                                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                            FROM `cache` c
                                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除人 IS NULL
                                    ) w
                                    GROUP BY 币种,联系电话
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                                UNION ALL
                                (SELECT 币种,ip AS 系统删除,COUNT(订单编号) AS 订单量
                                    FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '重复订单%','重复订单',删除原因))) 删单原因
                                            FROM `cache` c
                                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除人 IS NULL
                                    ) w
                                    GROUP BY 币种,ip
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                        ) s1;'''
        df5 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df5)
        df5.to_sql('cache_cp', con=self.engine1, index=False, if_exists='replace')
        print('正在记录 系统删除信息......')
        sql = '''REPLACE INTO {0}(币种, 删单类型, 删单明细, 单量, 记录日期, 更新时间) 
                SELECT 币种,'系统删除'  删单类型, 系统删除, 订单量, CURDATE() 记录日期, NOW() 更新时间
				FROM cache_cp s 
				WHERE s.系统删除 LIKE '%.%' OR s.系统删除 LIKE '%9%' OR s.系统删除 LIKE '%8%' OR s.系统删除 LIKE '%7%' OR s.系统删除 LIKE '%6%' OR s.系统删除 LIKE '%5%' 
				   OR s.系统删除 LIKE '%4%' OR s.系统删除 LIKE '%3%' OR s.系统删除 LIKE '%2%' OR s.系统删除 LIKE '%1%' OR s.系统删除 LIKE '%0%';'''.format('day_delete_cache')
        pd.read_sql_query(sql=sql, con=self.engine1,chunksize=10000)

        st_del, st_del_iphone, st_del_ip = '', '', ''
        k = 0
        k2 = 0
        for row in df5.itertuples():
            tem_Black = getattr(row, '系统删除')
            count = getattr(row, '订单量')
            if '总计' in tem_Black:
                st_del = '*系统删除： ' + str(int(count)) + '单；其中比较多的是：'
            elif '订单' in tem_Black:
                st_del = st_del + str(tem_Black) + ':' + str(int(count)) + '单,'

            elif tem_Black != None and '.' not in tem_Black:
                if count >= 10:
                    if k == 0:
                        st_del_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        st_del_iphone = st_del_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1
                else:
                    if k == 0:
                        st_del_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        st_del_iphone = st_del_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1

            elif tem_Black != None and '.' in tem_Black:
                if count >= 10:
                    if k2 == 0:
                        st_del_ip = ';\n           同一ip有：   (' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_del_ip = st_del_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
                else:
                    if k2 == 0:
                        st_del_ip = ';\n           同一ip有：   (' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_del_ip = st_del_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
        print('*' * 50)
        print(st_del + st_del_iphone + st_del_ip)

        print('正在获取 连续3天以上的黑名单 电话、IP信息 六…………')
        sql ='''UPDATE day_delete_cache d 
                    SET d.`删单明细` = IF(d.`删单明细` LIKE '0%', RIGHT(d.`删单明细`,LENGTH(d.`删单明细`)-1),d.`删单明细`);'''
        pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
        sql = '''SELECT *
                FROM (
                        SELECT 币种,删单类型,删单明细, COUNT(记录日期) AS 次数
                        FROM day_delete_cache d
                        WHERE d.`记录日期` >= IF(DATE_FORMAT('2022-06-29','%w') = 2 ,DATE_SUB(CURDATE(), INTERVAL 5 DAY),IF(DATE_FORMAT('2022-06-29','%w') = 1 ,DATE_SUB(CURDATE(), INTERVAL 6 DAY),DATE_SUB(CURDATE(), INTERVAL 2 DAY)))
                        GROUP BY 币种,删单类型,删单明细
                ) s
                WHERE s.次数 >= 3
                ORDER BY 币种,删单类型,删单明细, 次数 DESC;'''
        df6 = pd.read_sql_query(sql=sql, con=self.engine1)
        db61 = df6[~(df6['删单类型'].str.contains('系统删除'))]
        # day_del = db61.to_markdown()
        day_del = '''注意：连续3天同电话\IP的信息>>>'''
        day_del2 = '恶意订单:'
        day_del3 = '拉黑率订单:'
        for row in db61.itertuples():
            tem = getattr(row, '删单类型')
            info = getattr(row, '删单明细')
            count = getattr(row, '次数')
            # day_del = day_del + '\n' + tem + ':' + info + ':  ' + str(int(count)) + '单,'
            if '恶意订单' in tem:
                day_del2 = day_del2 + '\n' + info + ':  ' + str(int(count)) + '单,'
            elif '拉黑率订单' in tem:
                day_del3 = day_del3 + '\n' + info + ':  ' + str(int(count)) + '单,'

        print('*' * 50)
        print(day_del + '\n' + day_del2 + '\n' + day_del3)


        url = "https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e"  # url为机器人的webhook    个人小嗨
        # url = "https://oapi.dingtalk.com/robot/send?access_token=9a92f00296846dcd3ec8b52d7bacce114a9e34cb2d5dbfad9ce3371ab8d037f9"  # url为机器人的webhook  港台客服
        content = r'r"H:\桌面\需要用到的文件\文件夹\out2.jpeg"'  # 钉钉消息内容，注意test是自定义的关键字，需要在钉钉机器人设置中添加，这样才能接收到消息
        mobile_list = ['18538110674']  # 要@的人的手机号，可以是多个，注意：钉钉机器人设置中需要添加这些人，否则不会接收到消息
        isAtAll = '是'  # 是否@所有人
        headers = {'Content-Type': 'application/json', "Charset": "UTF-8"}
        data = {"msgtype": "text",
                # "markdown": {# 要发送的内容【支持markdown】【！注意：content内容要包含机器人自定义关键字，不然消息不会发送出去，这个案例中是test字段】
                #         "title": 'TEST',
                #         "text": "#### 昨日删单率分析" + "\n" +
                #         "* " + sl_tem +
                #         "   + " + sl_tem_lh + sl_tem_ey + sl_tem_cf + "\n" +
                #         "* " + hfh_tem +
                #         "   + " + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf
                # },
                "text": {        # 要发送的内容【支持markdown】【！注意：content内容要包含机器人自定义关键字，不然消息不会发送出去，这个案例中是test字段】
                    "content": '神龙 - 火凤凰 昨日台湾订单 删除分析' + '\n' +
                               sl_tem + sl_tem_lh + sl_tem_ey + sl_tem_cf + '\n' +
                               hfh_tem + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf + '\n' +
                               sl_Black + sl_Black_iphone + sl_Black_ip + '\n' +
                               st_ey + st_ey_iphone + st_ey_ip + '\n' +
                               cf_del + '\n' +
                               st_del + st_del_iphone + st_del_ip + '\n' + '\n' +
                               day_del + '\n' + day_del2 + '\n' + day_del3
                    # "content": 'TEST'
                },
                "at": {# 要@的人
                        # "atMobiles": mobile_list,
                        # 是否@所有人
                        "isAtAll": False  # @全体成员（在此可设置@特定某人）
                }
        }
        # 4、对请求的数据进行json封装
        sendData = json.dumps(data)  # 将字典类型数据转化为json格式
        sendData = sendData.encode("utf-8")  # python3的Request要求data为byte类型
        r = requests.post(url, headers=headers, data=json.dumps(data))
        req = json.loads(r.text)  # json类型数据转换为dict字典
        print(req['errmsg'])


    # 删除订单的  分析导出 测试
    def del_order_day_two(self):
        print('正在分析 昨日 删单原因中')
        sql ='''SELECT *,concat(ROUND(SUM(IF(删单原因 IS NULL OR 删单原因 = '',总订单量-订单量,订单量)) / SUM(总订单量) * 100,2),'%') as '删单率'
                FROM (SELECT s1.*,总订单量,总删单量, 系统删单量
                      FROM (SELECT 币种,运营团队,删单原因,COUNT(订单编号) AS 订单量
                            FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',IF(删除原因 LIKE '拉黑率%','拉黑率订单',删除原因)) 删单原因
                                  FROM `worksheet` c
                            ) w
                            GROUP BY 币种,运营团队,删单原因
                      ) s1
                      LEFT JOIN
                      ( SELECT 币种,运营团队,COUNT(订单编号) AS 总订单量, SUM(IF(订单状态 = '已删除',1,0)) AS 总删单量, SUM(IF(订单状态 = '已删除' AND 删除人 IS NULL,1,0)) AS 系统删单量
                        FROM `worksheet` w
                        GROUP BY 币种,运营团队
                      ) s2 
                      ON s1.`币种`=s2.`币种` AND s1.`运营团队`=s2.`运营团队`
                ) s
                WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾')
                GROUP BY 币种,运营团队,删单原因
                ORDER BY FIELD(币种,'台币','港币','合计'),
                         FIELD(运营团队,'神龙家族-台湾','火凤凰-台湾','神龙-运营1组','Line运营','金鹏家族-小虎队','合计'),
                         订单量 DESC;'''
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df1)
        # 初始化设置
        sl_tem, sl_tem_lh, sl_tem_ey, sl_tem_cf = '', '', '', ''
        hfh_tem, hfh_tem_lh, hfh_tem_ey, hfh_tem_cf = '', '', '', ''
        for row in df1.itertuples():
            tem = getattr(row, '运营团队')
            delreson = getattr(row, '删单原因')
            count = getattr(row, '订单量')
            if tem == '神龙家族-台湾' and delreson == None:
                sl_tem = '*神  龙:   昨日单量：' + str(int(getattr(row, '总订单量'))) + '；删单量：' + str(int(getattr(row, '总删单量'))) + '；删单率：' + str(getattr(row, '删单率')) + '；系统删单量：' + str(int(getattr(row, '系统删单量'))) + '单;'
                # print(sl_tem)
            elif tem == '神龙家族-台湾' and '拉黑率订单' in delreson:
                sl_tem_lh = '，\n            其中占比较多的是：拉黑率订单：' + str(int(count)) + '单, '
                # print(sl_tem_lh)
            elif tem == '神龙家族-台湾' and '恶意订单' in delreson:
                sl_tem_ey = '恶意订单：' + str(int(count)) + '单, '
                # print(sl_tem_ey)
            elif tem == '神龙家族-台湾' and '重复订单' in delreson:
                sl_tem_cf = '重复订单：' + str(int(count)) + '单;'
                # print(sl_tem_cf)

            elif tem == '火凤凰-台湾' and delreson == None:
                hfh_tem = '*火凤凰:  昨日单量：' + str(int(getattr(row, '总订单量'))) + '；删单量：' + str(int(getattr(row, '总删单量'))) + '；删单率：' + str(getattr(row, '删单率')) + '；系统删单量：' + str(int(getattr(row, '系统删单量'))) + '单;'
                # print(hfh_tem)
            elif tem == '火凤凰-台湾' and '拉黑率订单' in delreson:
                hfh_tem_lh = '，\n            其中占比较多的是：拉黑率订单：' + str(int(count)) + '单, '
                # print(hfh_tem_lh)
            elif tem == '火凤凰-台湾' and '恶意订单' in delreson:
                hfh_tem_ey = '恶意订单：' + str(int(count)) + '单, '
                # print(hfh_tem_ey)
            elif tem == '火凤凰-台湾' and '重复订单' in delreson:
                hfh_tem_cf = '重复订单：' + str(int(count)) + '单;'
                # print(hfh_tem_cf)
        print('*' * 50)
        print(sl_tem + sl_tem_lh + sl_tem_ey + sl_tem_cf)
        print(hfh_tem + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf)

        print('正在获取 删单明细 拉黑率信息 二…………')
        sql ='''SELECT 币种, 删单原因, 联系电话 AS 拉黑率订单, 订单量, 拉黑率70以上, 拉黑率70以下
                FROM(
                    (SELECT s1.*
                        FROM (  SELECT 币种,删单原因, 联系电话,COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '拉黑率%','拉黑率订单',删除原因) 删单原因
                                        FROM `worksheet` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '拉黑率%'
                                ) w
                                GROUP BY 币种,删单原因, 联系电话
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 IS NOT NULL AND 删单原因 IS NOT NULL AND 删单原因 <> ""
                        GROUP BY 币种,删单原因, 联系电话
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                    UNION ALL
                    (SELECT s1.*
                        FROM (  SELECT 币种,删单原因, ip,COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '拉黑率%','拉黑率订单',删除原因) 删单原因
                                        FROM `worksheet` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '拉黑率%'
                                ) w
                                GROUP BY 币种,删单原因, ip
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 IS NOT NULL AND 删单原因 IS NOT NULL AND 删单原因 <> ""
                        GROUP BY 币种,删单原因, ip
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                ) s;'''
        df2 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df2)
        sl_Black ,sl_Black_iphone , sl_Black_ip = '','',''
        k = 0
        k2 = 0
        for row in df2.itertuples():
            tem_Black = getattr(row, '拉黑率订单')
            count = getattr(row, '订单量')
            if tem_Black == None:
                sl_Black = '*拉黑率删除:  ' + str(int(getattr(row, '订单量'))) + '单；拉黑率70以上的：' + str(int(getattr(row, '拉黑率70以上'))) +'单；'
                # print(sl_Black)
            elif tem_Black != None and '.' not in tem_Black:
                if count >= 10:
                    if k == 0:
                        sl_Black_iphone = '\n           同一电话有：(0' + str(int(tem_Black)) + ':' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        sl_Black_iphone =sl_Black_iphone + '(0' + str(int(tem_Black)) + ':' + str(int(count)) + '单);'
                        k = k + 1
                else:
                    if k == 0:
                        sl_Black_iphone = '\n           同一电话有：(0' + str(int(tem_Black)) + ':' + str(int(count)) +'单),'
                        k = k + 1
                    elif k > 0:
                        sl_Black_iphone =sl_Black_iphone + '(0' + str(int(tem_Black)) + ':' + str(int(count)) + '单);'
                        k = k + 1
                    # print(sl_Black_iphone)

            elif tem_Black != None and '.' in tem_Black:
                if count >= 10:
                    if k2 == 0:
                        sl_Black_ip = '\n           同一ip有：   (' + str(getattr(row, '拉黑率订单')) + ':' + str(int(getattr(row, '订单量'))) + '单),'
                        k2 = k2 + 1
                    elif k > 0:
                        sl_Black_ip = sl_Black_ip + '(' + str(tem_Black) + ':' + str(int(count)) + '单);'
                        k2 = k2 + 1
                else:
                    if k2 == 0:
                        sl_Black_ip = '\n           同一ip有：   (' + str(tem_Black) + ':' + str(int(count)) +'单),'
                        k2 = k2 + 1
                    elif k > 0:
                        sl_Black_ip = sl_Black_ip + '（' + str(tem_Black) + ':' + str(int(count)) + '单);'
                        k2 = k2 + 1
                    # print(sl_Black_ip)
        print('*' * 50)
        print(sl_Black + sl_Black_iphone + sl_Black_ip)


        print('正在获取 删单明细 恶意删除信息 三…………')
        sql ='''SELECT 币种, 删单原因, 联系电话 AS 恶意删除, 订单量, 拉黑率70以上, 拉黑率70以下
                FROM(
                    (SELECT s1.*
                        FROM (  SELECT 币种,删单原因, 联系电话,COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',删除原因) 删单原因
                                        FROM `worksheet` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '恶意%'
                                ) w
                                GROUP BY 币种,删单原因, 联系电话
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 IS NOT NULL AND 删单原因 IS NOT NULL AND 删单原因 <> ""
                        GROUP BY 币种,删单原因, 联系电话
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                    UNION ALL
                    (SELECT s1.*
                        FROM (  SELECT 币种,删单原因, ip,COUNT(订单编号) AS 订单量, SUM(IF(拉黑率 > 70 ,1,0)) AS 拉黑率70以上,SUM(IF(拉黑率 < 70 ,1,0)) AS 拉黑率70以下
                                FROM (SELECT *,IF(删除原因 LIKE '恶意%','恶意订单',删除原因) 删单原因
                                        FROM `worksheet` c
                                        WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '恶意%'
                                ) w
                                GROUP BY 币种,删单原因, ip
                                WITH ROLLUP
                        )  s1
                        WHERE 币种 IS NOT NULL AND 删单原因 IS NOT NULL AND 删单原因 <> ""
                        GROUP BY 币种,删单原因, ip
                        ORDER BY 订单量 desc
                        LIMIT 5
                    ) 
                ) s;'''
        df3 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df3)
        st_ey, st_ey_iphone, st_ey_ip = '','',''
        k = 0
        k2 = 0
        for row in df3.itertuples():
            tem_Black = getattr(row, '恶意删除')
            count = getattr(row, '订单量')
            if tem_Black == None:
                st_ey = '*恶意删除： ' + str(int(count)) + '单；拉黑率70以上的：' + str(int(getattr(row, '拉黑率70以上'))) + '单；低于70的：' + str(int(getattr(row, '拉黑率70以下'))) + '单；'
            elif tem_Black != None and '.' not in tem_Black:
                if count >= 10:
                    if k == 0:
                        st_ey_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        st_ey_iphone = st_ey_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1
                else:
                    if k == 0:
                        st_ey_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) +'单),'
                        k = k + 1
                    elif k > 0:
                        st_ey_iphone = st_ey_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1

            elif tem_Black != None and '.' in tem_Black:
                if count >= 10:
                    if k2 == 0:
                        st_ey_ip = ';\n           同一ip有：    (' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_ey_ip = st_ey_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
                else:
                    if k2 == 0:
                        st_ey_ip = ';\n           同一ip有：    (' + str(tem_Black) + ': ' + str(int(count)) +'单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_ey_ip = st_ey_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
        print('*' * 50)
        print(st_ey + st_ey_iphone + st_ey_ip)

        print('正在获取 删单明细 重复删除信息 四…………')
        sql = '''SELECT IFNULL(币种,'币种') 币种,IFNULL(删单原因,'总计') 重复删除,COUNT(订单编号) AS 订单量
                    FROM (SELECT *,IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '恶意%','恶意订单',删除原因)) 删单原因
                            FROM `worksheet` c
                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '重复订单%'
                    ) w
                GROUP BY 币种,删单原因;'''
        df4 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df4)
        cf_del = ''
        for row in df4.itertuples():
            tem_Black = getattr(row, '重复删除')
            count = getattr(row, '订单量')
            cf_del = '*重复删除：' + str(count) + '单, 查询后都是客户上笔未收到或者是连续订多笔订单重复删除；'
        print('*' * 50)
        print(cf_del)


        print('正在获取 删单明细 系统删除信息 物五…………')
        sql = '''SELECT s1.*
                        FROM ( 
        					    (SELECT *
                                    FROM (SELECT IFNULL(币种,'币种') 币种,IFNULL(删单原因,'总计') 系统删除,COUNT(订单编号) AS 订单量
                                            FROM (SELECT *,IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '恶意%','恶意订单',删除原因)) 删单原因
                                                    FROM `worksheet` c
                                                    WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除人 IS NULL
                                            ) w
                                            GROUP BY 币种,删单原因
                                            WITH ROLLUP
                                    ) w1
                                    WHERE 币种 <> '币种'
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                                UNION ALL
                                 ( SELECT 币种,联系电话 AS 系统删除,COUNT(订单编号) AS 订单量
                                    FROM (SELECT *,IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '恶意%','恶意订单',删除原因)) 删单原因
                                            FROM `worksheet` c
                                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除人 IS NULL
                                    ) w
                                    GROUP BY 币种,联系电话
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                                UNION ALL
                                (SELECT 币种,ip AS 系统删除,COUNT(订单编号) AS 订单量
                                    FROM (SELECT *,IF(删除原因 LIKE '拉黑率%','拉黑率订单',IF(删除原因 LIKE '恶意%','恶意订单',删除原因)) 删单原因
                                            FROM `worksheet` c
                                            WHERE 币种 = '台币' AND 订单状态 = '已删除' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除人 IS NULL
                                    ) w
                                    GROUP BY 币种,ip
                                    ORDER BY 订单量 DESC
                                    LIMIT 4
                                )
                        ) s1;'''
        df5 = pd.read_sql_query(sql=sql, con=self.engine1)
        # print(df5)
        st_del, st_del_iphone, st_del_ip = '', '', ''
        k = 0
        k2 = 0
        for row in df5.itertuples():
            tem_Black = getattr(row, '系统删除')
            count = getattr(row, '订单量')
            if '总计' in tem_Black:
                st_del = '*系统删除： ' + str(int(count)) + '单；其中比较多的是：'
            elif '订单' in tem_Black:
                st_del = st_del + str(tem_Black) + ':' + str(int(count)) + '单,'

            elif tem_Black != None and '.' not in tem_Black:
                if count >= 10:
                    if k == 0:
                        st_del_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        st_del_iphone = st_del_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1
                else:
                    if k == 0:
                        st_del_iphone = ';\n           同一电话有：(' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k = k + 1
                    elif k > 0:
                        st_del_iphone = st_del_iphone + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k = k + 1

            elif tem_Black != None and '.' in tem_Black:
                if count >= 10:
                    if k2 == 0:
                        st_del_ip = ';\n           同一ip有：   (' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_del_ip = st_del_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
                else:
                    if k2 == 0:
                        st_del_ip = ';\n           同一ip有：   (' + str(tem_Black) + ': ' + str(int(count)) + '单),'
                        k2 = k2 + 1
                    elif k2 > 0:
                        st_del_ip = st_del_ip + '(0' + str(tem_Black) + ': ' + str(int(count)) + '单);'
                        k2 = k2 + 1
        print('*' * 50)
        print(st_del + st_del_iphone + st_del_ip)


        url = "https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e"  # url为机器人的webhook
        content = r'r"H:\桌面\需要用到的文件\文件夹\out2.jpeg"'  # 钉钉消息内容，注意test是自定义的关键字，需要在钉钉机器人设置中添加，这样才能接收到消息
        mobile_list = ['18538110674']  # 要@的人的手机号，可以是多个，注意：钉钉机器人设置中需要添加这些人，否则不会接收到消息
        isAtAll = '是'  # 是否@所有人
        headers = {'Content-Type': 'application/json', "Charset": "UTF-8"}
        data = {"msgtype": "markdown",
                # "markdown": {# 要发送的内容【支持markdown】【！注意：content内容要包含机器人自定义关键字，不然消息不会发送出去，这个案例中是test字段】
                #         "title": 'TEST',
                #         "text": "#### 昨日删单率分析" + "\n" +
                #         "* " + sl_tem +
                #         "   + " + sl_tem_lh + sl_tem_ey + sl_tem_cf + "\n" +
                #         "* " + hfh_tem +
                #         "   + " + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf
                # },
                "markdown": {  # 要发送的内容【支持markdown】【！注意：content内容要包含机器人自定义关键字，不然消息不会发送出去，这个案例中是test字段】
                    "title": 'TEST',
                    "text": "#### 昨日删单率分析" + "\n" +
                            "| 一个普通标题 | 一个普通标题 | 一个普通标题 |一个普通标题 | 一个普通标题 | 一个普通标题 |一个普通标题 | 一个普通标题 |"
                            # "* " + sl_tem + sl_tem_lh + sl_tem_ey + sl_tem_cf + '\n' +
                            # "   + " + hfh_tem + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf + '\n' +
                            # "* " + sl_Black + sl_Black_iphone + sl_Black_ip + '\n' +
                            # "   + " + st_ey + st_ey_iphone + st_ey_ip + '\n' +
                            # "* " + cf_del + '\n' +
                            # "   + " + st_del + st_del_iphone + st_del_ip
                },
                # "text": {
                #     # 要发送的内容【支持markdown】【！注意：content内容要包含机器人自定义关键字，不然消息不会发送出去，这个案例中是test字段】
                #     "content": 'TEST' + '\n' +
                #                sl_tem + sl_tem_lh + sl_tem_ey + sl_tem_cf + '\n' +
                #                hfh_tem + hfh_tem_lh + hfh_tem_ey + hfh_tem_cf + '\n' +
                #                sl_Black + sl_Black_iphone + sl_Black_ip + '\n' +
                #                st_ey + st_ey_iphone + st_ey_ip + '\n' +
                #                cf_del + '\n' +
                #                st_del + st_del_iphone + st_del_ip
                # },
                "at": {# 要@的人
                        # "atMobiles": mobile_list,
                        # 是否@所有人
                        "isAtAll": False  # @全体成员（在此可设置@特定某人）
                }
        }

        # 4、对请求的数据进行json封装
        sendData = json.dumps(data)  # 将字典类型数据转化为json格式
        sendData = sendData.encode("utf-8")  # python3的Request要求data为byte类型

        r = requests.post(url, headers=headers, data=json.dumps(data))
        req = json.loads(r.text)  # json类型数据转换为dict字典
        print(req['errmsg'])

        # return req['errmsg']
        # ax = df1.plot()
        # fig = ax.get_figure()
        # fig.savefig(r"H:\桌面\需要用到的文件\文件夹\out2.jpeg")

        # plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']  # 显示中文字体
        # fig = plt.figure(figsize=(3, 4), dpi=1400)  # dpi表示清晰度
        # ax = fig.add_subplot(111, frame_on=False)
        # ax.xaxis.set_visible(False)  # hide the x axis
        # ax.yaxis.set_visible(False)  # hide the y axis
        # table(ax, df1, loc='center')  # 将df换成需要保存的dataframe即可
        # plt.savefig(r"H:\桌面\需要用到的文件\文件夹\out.jpeg")

        # im = Image.fromarray(df1)
        # im.save(r"H:\桌面\需要用到的文件\文件夹\out.jpeg")
        # listT.append(df1)

        # print('正在获取 删单原因汇总 信息…………')
        # sql ='''SELECT s1.*
        #       FROM (
        #             SELECT 币种,删除原因,COUNT(订单编号) AS 订单量,
        #             		SUM(IF(拉黑率 > 80 ,1,0)) AS 拉黑率80以上,
    def del_order(self):
        print('+++正在分析 昨日 删单原因中')
        listT = []  # 查询sql的结果 存放池
        print('正在获取 删单明细…………')
        # sql ='''SELECT * FROM `cache` c  WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾');'''
        # df0 = pd.read_sql_query(sql=sql, con=self.engine1)
        # listT.append(df0)

        print('正在获取 删单明细 信息…………')
        sql ='''SELECT *,concat(ROUND(SUM(IF(删单原因 IS NULL OR 删单原因 = '',总订单量-订单量,订单量)) / SUM(总订单量) * 100,2),'%') as '删单率'
                FROM (
                      SELECT s1.*,总订单量,总删单量
                      FROM (
                            SELECT 币种,运营团队,删单原因,COUNT(订单编号) AS 订单量
                            FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                                  FROM `worksheet` c
                            ) w
                            GROUP BY 币种,运营团队,删单原因
                      ) s1
                      LEFT JOIN
                      (
                            SELECT 币种,运营团队,COUNT(订单编号) AS 总订单量,
                                  SUM(IF(订单状态 = '已删除',1,0)) AS 总删单量
                            FROM `worksheet` w
                            GROUP BY 币种,运营团队
                      ) s2 ON s1.`币种`=s2.`币种` AND s1.`运营团队`=s2.`运营团队`
                ) s
                WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾')
                GROUP BY 币种,运营团队,删单原因
                ORDER BY FIELD(币种,'台币','港币','合计'),
                         FIELD(运营团队,'神龙家族-台湾','火凤凰-台湾','神龙-运营1组','Line运营','金鹏家族-小虎队','合计'),
                         订单量 DESC;'''
        sql ='''SELECT 币种,运营团队,删单原因2 AS '删单原因(单)',订单量2 AS '删单量(单)',删单率
FROM (
SELECT 币种,运营团队,
      删单原因,
--       IF(删单原因 IS NULL ,CONCAT('总订单量：',总订单量,'单; 总删单量：',总删单量,'单;'),删单原因) AS 删单原因2,
      IF(删单原因 IS NULL ,CONCAT('总订单量：',总订单量),删单原因) AS 删单原因2,
      IF(删单原因 IS NULL ,CONCAT('总删单量：',总删单量),订单量) AS 订单量2,
      订单量,总订单量,总删单量,
      concat(ROUND(SUM(IF(删单原因 IS NULL OR 删单原因 = '',总订单量-订单量,订单量)) / SUM(总订单量) * 100,2),'%') as '删单率'
                FROM (
                      SELECT s1.*,总订单量,总删单量
                      FROM (
                            SELECT 币种,运营团队,删单原因,COUNT(订单编号) AS 订单量
                            FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                                  FROM `worksheet` c
                            ) w
                            GROUP BY 币种,运营团队,删单原因
                      ) s1
                      LEFT JOIN
                      (
                            SELECT 币种,运营团队,COUNT(订单编号) AS 总订单量,
                                  SUM(IF(订单状态 = '已删除',1,0)) AS 总删单量
                            FROM `worksheet` w
                            GROUP BY 币种,运营团队
                      ) s2 ON s1.`币种`=s2.`币种` AND s1.`运营团队`=s2.`运营团队`
                ) s
                WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾')
                GROUP BY 币种,运营团队,删单原因
--                 WITH rollup
                ORDER BY FIELD(币种,'台币','港币','合计'),
                         FIELD(运营团队,'神龙家族-台湾','火凤凰-台湾','神龙-运营1组','Line运营','金鹏家族-小虎队','合计'),
                         订单量 DESC
) ss;'''
        df1 = pd.read_sql_query(sql=sql, con=self.engine1)
        print(df1)
        # da = df1['总订单量'].values[0]
        da = df1.总订单量.values[0]
        print(da)

        for row in df1.itertuples():
            tem = getattr(row, '运营团队')
            delreson = getattr(row, '删单原因')
            if tem == '神龙家族-台湾' and delreson == None:
                tem_count = getattr(row, '总订单量')
                tem_count2 = getattr(row, '总删单量')
                tem_count3 = getattr(row, '删单率')
                print(tem_count)
                print(tem_count2)
                print(tem_count3)


        ax = df1.plot()
        fig = ax.get_figure()
        fig.savefig(r"H:\桌面\需要用到的文件\文件夹\out2.jpeg")

        plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']  # 显示中文字体
        fig = plt.figure(figsize=(3, 4), dpi=1400)  # dpi表示清晰度
        ax = fig.add_subplot(111, frame_on=False)
        ax.xaxis.set_visible(False)  # hide the x axis
        ax.yaxis.set_visible(False)  # hide the y axis
        table(ax, df1, loc='center')  # 将df换成需要保存的dataframe即可
        plt.savefig(r"H:\桌面\需要用到的文件\文件夹\out.jpeg")

        im = Image.fromarray(df1)
        im.save(r"H:\桌面\需要用到的文件\文件夹\out.jpeg")

        listT.append(df1)

        print('正在获取 删单原因汇总 信息…………')
        sql ='''SELECT s1.*
              FROM (
                    SELECT 币种,删除原因,COUNT(订单编号) AS 订单量,
                    		SUM(IF(拉黑率 > 80 ,1,0)) AS 拉黑率80以上,
							SUM(IF(拉黑率 < 80 ,1,0)) AS 拉黑率80以下
                    FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                                            FROM `cache` c
                                            WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾')
                                ) w
                    GROUP BY 币种,删单原因
              )  s1
                    WHERE 删除原因 IS NOT NULL AND 删除原因 <> ""
                    GROUP BY 币种,删除原因
              ORDER BY 订单量 desc
              LIMIT 5;'''
        df2 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df2)

        print('正在获取 删单原因明细（恶意订单-电话） 信息…………')
        sql ='''SELECT 币种,删单原因,联系电话,COUNT(订单编号) AS 订单量
                FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                      FROM `cache` c
                      WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '%恶意%'
                    ) w
                GROUP BY 币种,`联系电话`
				ORDER BY 订单量 desc;'''
        df30 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df30)
        print('正在获取 删单原因明细（恶意订单-ip） 信息…………')
        sql = '''SELECT 币种,删单原因,IP,COUNT(订单编号) AS 订单量
                        FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                              FROM `cache` c
                              WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '%恶意%'
                            ) w
                        GROUP BY 币种,`IP`
        				ORDER BY 订单量 desc;'''
        df31 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df31)

        print('正在获取 删单原因明细（拉黑率-电话） 信息…………')
        sql = '''SELECT 币种,删单原因,联系电话,COUNT(订单编号) AS 订单量
                        FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                              FROM `cache` c
                              WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '%拉黑率%'
                            ) w
                        GROUP BY 币种,`联系电话`
        				ORDER BY 订单量 desc;'''
        df40 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df40)
        print('正在获取 删单原因明细（拉黑率-ip） 信息…………')
        sql = '''SELECT 币种,删单原因,IP,COUNT(订单编号) AS 订单量
                                FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                                      FROM `cache` c
                                      WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '%拉黑率%'
                                    ) w
                                GROUP BY 币种,`IP`
                				ORDER BY 订单量 desc;'''
        df41 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df41)

        print('正在获取 删单原因明细（系统删除-删单原因） 信息…………')
        sql = '''SELECT 币种,删单原因,COUNT(订单编号) AS 订单量
                FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                     FROM `cache` c
                     WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND `删除人` IS NULL
                ) w
                GROUP BY 币种,`删单原因`
				ORDER BY 订单量 desc;'''
        df50 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df50)
        print('正在获取 删单原因明细（系统删除-ip） 信息…………')
        sql = '''SELECT 币种,删单原因,IP,COUNT(订单编号) AS 订单量
                FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                     FROM `cache` c
                     WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND `删除人` IS NULL
                ) w
                GROUP BY 币种,`IP`
                ORDER BY 订单量 desc;'''
        df51 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df51)
        print('正在获取 删单原因明细（系统删除-电话） 信息…………')
        sql = '''SELECT 币种,删单原因,联系电话,COUNT(订单编号) AS 订单量
                FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                     FROM `cache` c
                     WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND `删除人` IS NULL
                ) w
                GROUP BY 币种,`联系电话`
                ORDER BY 订单量 desc;'''
        df52 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df52)

        print('正在获取 删单原因明细（重复订单-电话） 信息…………')
        sql = '''SELECT 币种,删单原因,联系电话,COUNT(订单编号) AS 订单量
                        FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                              FROM `cache` c
                              WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '%恶意%'
                            ) w
                        GROUP BY 币种,`联系电话`
        				ORDER BY 订单量 desc;'''
        df60 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df60)
        print('正在获取 删单原因明细（重复订单-ip） 信息…………')
        sql = '''SELECT 币种,删单原因,IP,COUNT(订单编号) AS 订单量
                                FROM (SELECT *,IF(删除原因 LIKE '%恶意%',';恶意订单',IF(删除原因 LIKE '%拉黑率%',';拉黑率订单',删除原因)) 删单原因
                                      FROM `cache` c
                                      WHERE 币种 = '台币' AND 运营团队 IN ('神龙家族-台湾','火凤凰-台湾') AND 删除原因 LIKE '%恶意%'
                                    ) w
                                GROUP BY 币种,`IP`
                				ORDER BY 订单量 desc;'''
        df61 = pd.read_sql_query(sql=sql, con=self.engine1)
        listT.append(df61)

        print('正在写入excel…………')
        today = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        file_path = 'H:\\桌面\\需要用到的文件\\输出文件\\工作数量 {}.xlsx'.format(today)

        writer2 = pd.ExcelWriter(file_path, engine='openpyxl')
        df1.to_excel(writer2, index=False)                  # 删单
        df2.to_excel(writer2, index=False, startrow=20)     # 删单原因汇总

        df30.to_excel(writer2, index=False, startcol=9)     # 恶意订单-电话
        df31.to_excel(writer2, index=False, startcol=14)    # 恶意订单-ip

        df40.to_excel(writer2, index=False, startcol=19)     # 拉黑率-电话
        df41.to_excel(writer2, index=False, startcol=24)     # 拉黑率-ip

        df50.to_excel(writer2, index=False, startcol=29)     # 系统删除-删单原因
        df51.to_excel(writer2, index=False, startcol=34)     # 系统删除-ip
        df52.to_excel(writer2, index=False, startcol=39)     # 系统删除-电话

        df60.to_excel(writer2, index=False, startcol=44)     # 重复订单-电话）
        df61.to_excel(writer2, index=False, startcol=49)     # 重复订单-ip

        writer2.save()
        writer2.close()
        print()

    # 进入订单检索界面     促单查询
    def order_track_Query(self, time_handle, timeStart, timeEnd, proxy_id, proxy_handle):
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        team = '促单_分析'
        if time_handle == '自动':
            # sql = '''SELECT DISTINCT 下单时间 FROM {0} d GROUP BY 下单时间 ORDER BY 下单时间 DESC'''.format(team)
            # rq = pd.read_sql_query(sql=sql, con=self.engine1)
            # rq = pd.to_datetime(rq['下单时间'][0])
            # timeStart = (rq + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            # timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            timeStart = (datetime.datetime.now() - relativedelta(months=2)).strftime('%Y-%m') + '-01'
            timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            timeStart = timeStart
            timeEnd = timeEnd
            # if (datetime.datetime.now()).strftime('%d') == 1:
            #     timeStart = (datetime.datetime.now() - relativedelta(months=1)).strftime('%Y-%m-%d')
            #     timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            # else:
            #     timeStart = (datetime.datetime.now().replace(day=1)).strftime('%Y-%m-%d')
            #     timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        print('+++正在查询 促单订单 信息中：' + str(timeStart) + " *** " + str(timeEnd))

        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500, 'orderPrefix': None, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None,'isClone': None,
                'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None,  'reassignmentType': None, 'lowerstatus': None, 'warehouse': None,
                'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None, 'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None,
                'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': -1, 'autoVerifyStatus': None, 'shipZip': None, 'remark': None,
                'shipState': None, 'weightStart': None,'weightEnd': None,  'estimateWeightStart': None,  'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'isChangeMark': None,
                'timeStart': timeStart + ' 00:00:00', 'timeEnd': timeEnd + ' 23:59:59'}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        req = json.loads(req.text)  # json类型数据转换为dict字典
        max_count = req['data']['count']
        print('共...' + str(max_count) + '...单量')
        if max_count != 0:
            df = pd.DataFrame([])
            n = 1
            if max_count > 500:
                in_count = math.ceil(max_count / 500)
                print(in_count)
                dlist = []
                while n <= in_count:  # 这里用到了一个while循环，穿越过来的
                    data = self._order_track_Query(timeStart, timeEnd, n, proxy_handle, proxy_id)
                    dlist.append(data)
                    print('剩余查询次数' + str(in_count - n))
                    n = n + 1
                dp = df.append(dlist, ignore_index=True)
            else:
                dp = self._order_track_Query(timeStart, timeEnd, n, proxy_handle, proxy_id)
            dp = dp[['id','orderNumber', 'currency', 'wayBillNumber','addTime', 'orderStatus', 'logisticsStatus', 'service', 'cloneUser','befrom', 'cloneTypeName']]
            dp.columns = ['id','订单编号', '币种', '运单编号', '下单时间', '订单状态', '物流状态', '代下单客服', '克隆人','来源渠道', '克隆类型']
            dp.to_sql('cache', con=self.engine1, index=False, if_exists='replace')
            sql = '''REPLACE INTO {0}(id,订单编号,币种,下单时间,订单状态, 物流状态, 代下单客服,克隆人,来源渠道, 克隆类型,记录时间) 
                               SELECT id,订单编号,币种,下单时间,订单状态, 物流状态, 代下单客服,克隆人,来源渠道, 克隆类型, NOW() 记录时间 
                    FROM cache;'''.format('促单_分析')
            pd.read_sql_query(sql=sql, con=self.engine1, chunksize=10000)
            print('正在查询 - 促单分析；挽单列表；拒收问题件 数据')
            listT = []
            listT.append(dp)    # 0 明细单量
            print('不分币种 分月份 的整体 - 促单分析')
            sql1 = '''SELECT IFNULL(月份,'总计') as 月份,
                            COUNT(订单编号) as 总代下单量,
                            SUM(IF(订单状态 NOT IN ("已删除","问题订单审核","问题订单","待审核","未支付","待发货","支付失败","已取消","截单","截单中（面单已打印，等待仓库审核）","待发货转审核"),1,0)) AS 有效代下单量,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 完成签收,
                            SUM(IF(物流状态 = "已签收",1,0)) as 签收,
                            SUM(IF(物流状态 = "拒收",1,0)) as 拒收,
                            SUM(IF(物流状态 = "已退货",1,0)) as 已退货,
                            SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / COUNT(订单编号),0) * 100,2),'%') as 总计签收,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) / COUNT(订单编号),0) * 100,2),'%') as 完成占比,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已退货",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 退货率
                    FROM ( SELECT *,  DATE_FORMAT(下单时间, '%Y%m') AS 月份
                            FROM `促单_分析` s
                            WHERE (s.克隆人 IS NULL OR s.克隆人 = "") or (s.克隆类型 = "扣货克隆")
                    ) ss1
                    GROUP BY 月份
                    WITH ROLLUP 
                    ORDER BY 月份 DESC;'''
            df1 = pd.read_sql_query(sql=sql1, con=self.engine1)
            listT.append(df1)  # 1 不分币种

            print('不分币种 分月份 的个人 - 促单分析')
            sql2 = '''SELECT IFNULL(月份,'总计') as 月份,
                            IFNULL(代下单客服,'总计') 代下单客服,
                            COUNT(订单编号) as 总代下单量,
                            SUM(IF(订单状态 NOT IN ("已删除","问题订单审核","问题订单","待审核","未支付","待发货","支付失败","已取消","截单","截单中（面单已打印，等待仓库审核）","待发货转审核"),1,0)) AS 有效代下单量,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 完成签收,
                            SUM(IF(物流状态 = "已签收",1,0)) as 签收,
                            SUM(IF(物流状态 = "拒收",1,0)) as 拒收,
                            SUM(IF(物流状态 = "已退货",1,0)) as 已退货,
                            SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / COUNT(订单编号),0) * 100,2),'%') as 总计签收,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) / COUNT(订单编号),0) * 100,2),'%') as 完成占比,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已退货",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 退货率
                    FROM ( SELECT *,  DATE_FORMAT(下单时间, '%Y%m') AS 月份
                            FROM `促单_分析` s
                            WHERE (s.克隆人 IS NULL OR s.克隆人 = "") or (s.克隆类型 = "扣货克隆")
                    ) ss1
                    GROUP BY 月份, 代下单客服
                    WITH ROLLUP 
                    ORDER BY 月份 DESC, FIELD(代下单客服,'马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','刘文君','曲开拓','侯振峰','蔡利英','杨嘉仪','张陈平','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df2 = pd.read_sql_query(sql=sql2, con=self.engine1)
            listT.append(df2)   # 2 不分币种

            print('不分币种 分月份 分日期 的个人 - 促单分析')
            sql3 = '''SELECT IFNULL(月份,'总计') as 月份,
                                IFNULL(日期,'总计') as 日期,
                                IFNULL(代下单客服,'总计') 代下单客服,
                                COUNT(订单编号) as 总代下单量,
                                SUM(IF(订单状态 NOT IN ("已删除","问题订单审核","问题订单","待审核","未支付","待发货","支付失败","已取消","截单","截单中（面单已打印，等待仓库审核）","待发货转审核"),1,0)) AS 有效代下单量,
                                concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 完成签收,
                                SUM(IF(物流状态 = "已签收",1,0)) as 签收,
                                SUM(IF(物流状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(物流状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / COUNT(订单编号),0) * 100,2),'%') as 总计签收,
                                concat(ROUND(IFNULL(SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) / COUNT(订单编号),0) * 100,2),'%') as 完成占比,
                                concat(ROUND(IFNULL(SUM(IF(物流状态 = "已退货",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 退货率
                        FROM ( SELECT *,  DATE_FORMAT(下单时间, '%Y%m') AS 月份, DATE_FORMAT(下单时间, '%Y-%m-%d' ) AS 日期
                                FROM `促单_分析` s
                                WHERE (s.克隆人 IS NULL OR s.克隆人 = "") or (s.克隆类型 = "扣货克隆")
                        ) ss1
                        GROUP BY 月份, 日期, 代下单客服
                        WITH ROLLUP 
            	        ORDER BY 月份 DESC, 日期 DESC, FIELD(代下单客服,'马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','刘文君','曲开拓','侯振峰','蔡利英','杨嘉仪','张陈平','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df3 = pd.read_sql_query(sql=sql3, con=self.engine1)
            listT.append(df3)  # 3 不分币种


            print('分币种 分月份 的整体 - 促单分析')
            sql11 = '''SELECT IFNULL(币种,'总计') as 币种,
                            IFNULL(月份,'总计') as 月份,
                            COUNT(订单编号) as 总代下单量,
                            SUM(IF(订单状态 NOT IN ("已删除","问题订单审核","问题订单","待审核","未支付","待发货","支付失败","已取消","截单","截单中（面单已打印，等待仓库审核）","待发货转审核"),1,0)) AS 有效代下单量,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 完成签收,
                            SUM(IF(物流状态 = "已签收",1,0)) as 签收,
                            SUM(IF(物流状态 = "拒收",1,0)) as 拒收,
                            SUM(IF(物流状态 = "已退货",1,0)) as 已退货,
                            SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / COUNT(订单编号),0) * 100,2),'%') as 总计签收,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) / COUNT(订单编号),0) * 100,2),'%') as 完成占比,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已退货",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 退货率
                    FROM ( SELECT *,  DATE_FORMAT(下单时间, '%Y%m') AS 月份
                            FROM `促单_分析` s
                            WHERE (s.克隆人 IS NULL OR s.克隆人 = "") or (s.克隆类型 = "扣货克隆")
                    ) ss1
                    GROUP BY 币种, 月份
                    WITH ROLLUP 
                    ORDER BY 币种, 月份 DESC;'''
            df11 = pd.read_sql_query(sql=sql11, con=self.engine1)
            listT.append(df11)  # 1 分币种

            print('分币种 分月份 的个人 - 促单分析')
            sql22 = '''SELECT IFNULL(币种,'总计') as 币种,
                            IFNULL(月份,'总计') as 月份,
                            IFNULL(代下单客服,'总计') 代下单客服,
                            COUNT(订单编号) as 总代下单量,
                            SUM(IF(订单状态 NOT IN ("已删除","问题订单审核","问题订单","待审核","未支付","待发货","支付失败","已取消","截单","截单中（面单已打印，等待仓库审核）","待发货转审核"),1,0)) AS 有效代下单量,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 完成签收,
                            SUM(IF(物流状态 = "已签收",1,0)) as 签收,
                            SUM(IF(物流状态 = "拒收",1,0)) as 拒收,
                            SUM(IF(物流状态 = "已退货",1,0)) as 已退货,
                            SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / COUNT(订单编号),0) * 100,2),'%') as 总计签收,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) / COUNT(订单编号),0) * 100,2),'%') as 完成占比,
                            concat(ROUND(IFNULL(SUM(IF(物流状态 = "已退货",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 退货率
                    FROM ( SELECT *,  DATE_FORMAT(下单时间, '%Y%m') AS 月份
                            FROM `促单_分析` s
                            WHERE (s.克隆人 IS NULL OR s.克隆人 = "") or (s.克隆类型 = "扣货克隆")
                    ) ss1
                    GROUP BY 币种, 月份, 代下单客服
                    WITH ROLLUP 
                    ORDER BY 币种, 月份 DESC, FIELD(代下单客服,'马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','刘文君','曲开拓','侯振峰','蔡利英','杨嘉仪','张陈平','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df22 = pd.read_sql_query(sql=sql22, con=self.engine1)
            listT.append(df22)   # 2 分币种

            print('分币种 分月份 分日期 的个人 - 促单分析')
            sql33 = '''SELECT IFNULL(币种,'总计') as 币种,
                                IFNULL(月份,'总计') as 月份,
                                IFNULL(日期,'总计') as 日期,
                                IFNULL(代下单客服,'总计') 代下单客服,
                                COUNT(订单编号) as 总代下单量,
                                SUM(IF(订单状态 NOT IN ("已删除","问题订单审核","问题订单","待审核","未支付","待发货","支付失败","已取消","截单","截单中（面单已打印，等待仓库审核）","待发货转审核"),1,0)) AS 有效代下单量,
                                concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 完成签收,
                                SUM(IF(物流状态 = "已签收",1,0)) as 签收,
                                SUM(IF(物流状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(物流状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                concat(ROUND(IFNULL(SUM(IF(物流状态 = "已签收",1,0)) / COUNT(订单编号),0) * 100,2),'%') as 总计签收,
                                concat(ROUND(IFNULL(SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) / COUNT(订单编号),0) * 100,2),'%') as 完成占比,
                                concat(ROUND(IFNULL(SUM(IF(物流状态 = "已退货",1,0)) / SUM(IF(物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 退货率
                        FROM ( SELECT *,  DATE_FORMAT(下单时间, '%Y%m') AS 月份, DATE_FORMAT(下单时间, '%Y-%m-%d' ) AS 日期
                                FROM `促单_分析` s
                                WHERE (s.克隆人 IS NULL OR s.克隆人 = "") or (s.克隆类型 = "扣货克隆")
                        ) ss1
                        GROUP BY 币种, 月份, 日期, 代下单客服
                        WITH ROLLUP 
            	        ORDER BY 币种, 月份 DESC, 日期 DESC, FIELD(代下单客服,'马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','刘文君','曲开拓','侯振峰','蔡利英','杨嘉仪','张陈平','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df33 = pd.read_sql_query(sql=sql33, con=self.engine1)
            listT.append(df33)  # 3 分币种


            print('不分币种 分月份 不分类型 的整体 - 挽单列表分析')
            sql111 = '''SELECT IFNULL(月份,'总计') as 月份,
                                IFNULL(创建人,'总计') 创建人,
                                COUNT(订单编号) as 总代下单量,
                                SUM(IF(当前订单状态 NOT IN ("已删除","问题订单审核","问题订单","待审核","未支付","待发货","支付失败","已取消","截单","截单中（面单已打印，等待仓库审核）","待发货转审核"),1,0)) AS 有效代下单量,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已签收",1,0)) / SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 完成签收,
                                SUM(IF(当前物流状态 = "已签收",1,0)) as 签收,
                                SUM(IF(当前物流状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(当前物流状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已签收",1,0)) / COUNT(订单编号),0) * 100,2),'%') as 总计签收,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) / COUNT(订单编号),0) * 100,2),'%') as 完成占比,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已退货",1,0)) / SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 退货率
                        FROM ( SELECT *,  DATE_FORMAT(创建时间, '%Y%m') AS 月份, DATE_FORMAT(创建时间, '%Y-%m-%d' ) AS 日期
                                FROM `挽单列表_分析` s
                                WHERE s.删除人 IS NULL OR s.删除人 = ""
                        ) ss1
                        GROUP BY 月份,  创建人
                        WITH ROLLUP 
            	        ORDER BY 月份 DESC,  FIELD(创建人,'马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','刘文君','曲开拓','侯振峰','蔡利英','杨嘉仪','张陈平','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df111 = pd.read_sql_query(sql=sql111, con=self.engine1)
            listT.append(df111)  # 111 不分币种

            print('不分币种 分月份 分类型 的整体 - 挽单列表分析')
            sql222 = '''SELECT  IFNULL(月份,'总计') as 月份,
                                IFNULL(挽单类型,'总计') as 挽单类型,
                                IFNULL(创建人,'总计') 创建人,
                                COUNT(订单编号) as 总代下单量,
                                SUM(IF(当前订单状态 NOT IN ("已删除","问题订单审核","问题订单","待审核","未支付","待发货","支付失败","已取消","截单","截单中（面单已打印，等待仓库审核）","待发货转审核"),1,0)) AS 有效代下单量,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已签收",1,0)) / SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 完成签收,
                                SUM(IF(当前物流状态 = "已签收",1,0)) as 签收,
                                SUM(IF(当前物流状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(当前物流状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已签收",1,0)) / COUNT(订单编号),0) * 100,2),'%') as 总计签收,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) / COUNT(订单编号),0) * 100,2),'%') as 完成占比,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已退货",1,0)) / SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 退货率
                        FROM ( SELECT *,  DATE_FORMAT(创建时间, '%Y%m') AS 月份, DATE_FORMAT(创建时间, '%Y-%m-%d' ) AS 日期
                                FROM `挽单列表_分析` s
                                WHERE s.删除人 IS NULL OR s.删除人 = ""
                        ) ss1
                        GROUP BY 月份, 挽单类型, 创建人
                        WITH ROLLUP 
            	        ORDER BY 月份 DESC, 挽单类型, FIELD(创建人,'马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','刘文君','曲开拓','侯振峰','蔡利英','杨嘉仪','张陈平','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df222 = pd.read_sql_query(sql=sql222, con=self.engine1)
            listT.append(df222)  # 222 不分币种

            print('不分币种 分月份 分日期 不分类型 的整体 - 挽单列表分析')
            sql333 = '''SELECT  IFNULL(月份,'总计') as 月份,
                                IFNULL(日期,'总计') as 日期,
                                IFNULL(创建人,'总计') 创建人,
                                COUNT(订单编号) as 总代下单量,
                                SUM(IF(当前订单状态 NOT IN ("已删除","问题订单审核","问题订单","待审核","未支付","待发货","支付失败","已取消","截单","截单中（面单已打印，等待仓库审核）","待发货转审核"),1,0)) AS 有效代下单量,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已签收",1,0)) / SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 完成签收,
                                SUM(IF(当前物流状态 = "已签收",1,0)) as 签收,
                                SUM(IF(当前物流状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(当前物流状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已签收",1,0)) / COUNT(订单编号),0) * 100,2),'%') as 总计签收,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) / COUNT(订单编号),0) * 100,2),'%') as 完成占比,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已退货",1,0)) / SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 退货率
                        FROM ( SELECT *,  DATE_FORMAT(创建时间, '%Y%m') AS 月份, DATE_FORMAT(创建时间, '%Y-%m-%d' ) AS 日期
                                FROM `挽单列表_分析` s
                                WHERE s.删除人 IS NULL OR s.删除人 = ""
                        ) ss1
                        GROUP BY 月份, 日期, 创建人
                        WITH ROLLUP 
            	        ORDER BY 月份 DESC, 日期 DESC, FIELD(创建人,'马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','刘文君','曲开拓','侯振峰','蔡利英','杨嘉仪','张陈平','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df333 = pd.read_sql_query(sql=sql333, con=self.engine1)
            listT.append(df333)  # 333 不分币种

            print('不分币种 分月份 分日期 分类型 的整体 - 挽单列表分析')
            sql444 = '''SELECT IFNULL(月份,'总计') as 月份,
                                IFNULL(日期,'总计') as 日期,
                                IFNULL(挽单类型,'总计') as 挽单类型,
                                IFNULL(创建人,'总计') 创建人,
                                COUNT(订单编号) as 总代下单量,
                                SUM(IF(当前订单状态 NOT IN ("已删除","问题订单审核","问题订单","待审核","未支付","待发货","支付失败","已取消","截单","截单中（面单已打印，等待仓库审核）","待发货转审核"),1,0)) AS 有效代下单量,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已签收",1,0)) / SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 完成签收,
                                SUM(IF(当前物流状态 = "已签收",1,0)) as 签收,
                                SUM(IF(当前物流状态 = "拒收",1,0)) as 拒收,
                                SUM(IF(当前物流状态 = "已退货",1,0)) as 已退货,
                                SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) as 已完成,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已签收",1,0)) / COUNT(订单编号),0) * 100,2),'%') as 总计签收,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)) / COUNT(订单编号),0) * 100,2),'%') as 完成占比,
                                concat(ROUND(IFNULL(SUM(IF(当前物流状态 = "已退货",1,0)) / SUM(IF(当前物流状态 IN ("已签收","拒收","已退货","理赔","自发头程丢件"),1,0)),0) * 100,2),'%') as 退货率
                        FROM ( SELECT *,  DATE_FORMAT(创建时间, '%Y%m') AS 月份, DATE_FORMAT(创建时间, '%Y-%m-%d' ) AS 日期
                                FROM `挽单列表_分析` s
                                WHERE s.删除人 IS NULL OR s.删除人 = ""
                        ) ss1
                        GROUP BY 月份, 日期, 挽单类型, 创建人
                        WITH ROLLUP 
            	        ORDER BY 月份 DESC, 日期 DESC, 挽单类型, FIELD(创建人,'马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','刘文君','曲开拓','侯振峰','蔡利英','杨嘉仪','张陈平','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df444 = pd.read_sql_query(sql=sql444, con=self.engine1)
            listT.append(df444)  # 444 不分币种


            print('不分币种 分月份 分登记处理人 的整体 - 拒收挽单分析')
            sql1111 = '''SELECT 月份, 登记处理人,联系单量, 有效联系量, 挽单量,
                                concat(ROUND(有效联系量 / 联系单量 * 100,2),'%') AS 有效联系率,	
                                concat(ROUND(挽单量 / 有效联系量 * 100,2),'%') AS 挽单率, 
                                concat(ROUND(挽单签收量 / 挽单完成量 * 100,2),'%') AS 挽单完成签收,
                                concat(ROUND(挽单签收量 / 挽单量 * 100,2),'%') AS 挽单总计签收, 
                                concat(ROUND(挽单完成量 / 挽单量 * 100,2),'%') AS 挽单完成占比,
                                挽单签收量, 挽单完成量
                        FROM  (	SELECT IFNULL(s1.月份,'总计') as 月份,
                                        IFNULL(s1.登记处理人,'总计') as 登记处理人,
                                        COUNT(s1.`订单编号`) AS 联系单量, 
                                        SUM(IF(s1.`具体原因` NOT IN ('无人接听','无效号码','未联系上客户','联系不上客户') ,1,0)) AS 有效联系量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '',1,0)) AS 挽单量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '' AND gt.`系统物流状态` = '已签收',1,0)) AS 挽单签收量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '' AND gt.`系统物流状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) AS 挽单完成量
                                FROM ( SELECT *
                                        FROM (SELECT *,  DATE_FORMAT(处理时间, '%Y%m') AS 月份, DATE_FORMAT(处理时间, '%Y-%m-%d' ) AS 日期, IF(处理人 = "" OR 处理人 IS NULL,联系方式,处理人) AS 登记处理人
                                                    FROM 拒收问题件 j
                                                    WHERE j.id IN (SELECT MAX(id) FROM 拒收问题件 w  GROUP BY 订单编号)
                                        ) lp
                                        WHERE lp.`月份` >= DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 12 month),'%Y%m')
                                ) s1
                                LEFT JOIN gat_order_list gt ON s1.`再次克隆下单` = gt.`订单编号` 
                                GROUP BY 月份,  登记处理人
                                WITH ROLLUP
                        ) ss1
                        WHERE ss1.月份 <> '总计'
                        GROUP BY 月份, 登记处理人
                        ORDER BY 月份 DESC,  FIELD(登记处理人,'蔡利英','杨嘉仪','张陈平','邮件','电话','Line','短信','whatsapp','客户问题','刘文君','曲开拓','侯振峰','马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df1111 = pd.read_sql_query(sql=sql1111, con=self.engine1)
            listT.append(df1111)  # 1111 不分币种

            print('不分币种 分月份 分新单克隆人 的整体 - 拒收挽单分析')
            sql2222 = '''SELECT 月份, 新单克隆人,联系单量, 有效联系量, 挽单量,
                                concat(ROUND(有效联系量 / 联系单量 * 100,2),'%') AS 有效联系率,	
                                concat(ROUND(挽单量 / 有效联系量 * 100,2),'%') AS 挽单率, 
                                concat(ROUND(挽单签收量 / 挽单完成量 * 100,2),'%') AS 挽单完成签收,
                                concat(ROUND(挽单签收量 / 挽单量 * 100,2),'%') AS 挽单总计签收, 
                                concat(ROUND(挽单完成量 / 挽单量 * 100,2),'%') AS 挽单完成占比,
                                挽单签收量, 挽单完成量
                        FROM  (	SELECT IFNULL(s1.月份,'总计') as 月份,
                                        IFNULL(s1.新单克隆人,'总计') as 新单克隆人,
                                        COUNT(s1.`订单编号`) AS 联系单量, 
                                        SUM(IF(s1.`具体原因` NOT IN ('无人接听','无效号码','未联系上客户','联系不上客户') ,1,0)) AS 有效联系量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '',1,0)) AS 挽单量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '' AND gt.`系统物流状态` = '已签收',1,0)) AS 挽单签收量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '' AND gt.`系统物流状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) AS 挽单完成量
                                FROM ( SELECT *
                                        FROM (SELECT *,  DATE_FORMAT(处理时间, '%Y%m') AS 月份, DATE_FORMAT(处理时间, '%Y-%m-%d' ) AS 日期, IF(处理人 = "" OR 处理人 IS NULL,联系方式,处理人) AS 登记处理人
                                                    FROM 拒收问题件 j
                                                    WHERE j.id IN (SELECT MAX(id) FROM 拒收问题件 w  GROUP BY 订单编号)
                                        ) lp
                                        WHERE lp.`月份` >= DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 12 month),'%Y%m')
                                ) s1
                                LEFT JOIN gat_order_list gt ON s1.`再次克隆下单` = gt.`订单编号` 
                                GROUP BY 月份,  新单克隆人
                                WITH ROLLUP
                        ) ss1
                        WHERE ss1.月份 <> '总计'
                        GROUP BY 月份, 新单克隆人
                        ORDER BY 月份 DESC,  FIELD(新单克隆人,'蔡利英','杨嘉仪','张陈平','邮件','电话','Line','短信','whatsapp','客户问题','刘文君','曲开拓','侯振峰','马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df2222 = pd.read_sql_query(sql=sql2222, con=self.engine1)
            listT.append(df2222)  # 2222 不分币种

            print('分币种 分月份 分登记处理人 的整体 - 拒收挽单分析')
            sql3333 = '''SELECT 币种, 月份, 登记处理人,联系单量, 有效联系量, 挽单量,
                                concat(ROUND(有效联系量 / 联系单量 * 100,2),'%') AS 有效联系率,	
                                concat(ROUND(挽单量 / 有效联系量 * 100,2),'%') AS 挽单率, 
                                concat(ROUND(挽单签收量 / 挽单完成量 * 100,2),'%') AS 挽单完成签收,
                                concat(ROUND(挽单签收量 / 挽单量 * 100,2),'%') AS 挽单总计签收, 
                                concat(ROUND(挽单完成量 / 挽单量 * 100,2),'%') AS 挽单完成占比,
                                挽单签收量, 挽单完成量
                        FROM (	SELECT IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.月份,'总计') as 月份,
                                        IFNULL(s1.登记处理人,'总计') as 登记处理人,
                                        COUNT(s1.`订单编号`) AS 联系单量, 
                                        SUM(IF(s1.`具体原因` NOT IN ('无人接听','无效号码','未联系上客户','联系不上客户') ,1,0)) AS 有效联系量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '',1,0)) AS 挽单量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '' AND gt.`系统物流状态` = '已签收',1,0)) AS 挽单签收量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '' AND gt.`系统物流状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) AS 挽单完成量
                                FROM (  SELECT *
                                        FROM (SELECT *,  DATE_FORMAT(处理时间, '%Y%m') AS 月份, DATE_FORMAT(处理时间, '%Y-%m-%d' ) AS 日期, IF(处理人 = "" OR 处理人 IS NULL,联系方式,处理人) AS 登记处理人
                                                    FROM 拒收问题件 j
                                                    WHERE j.id IN (SELECT MAX(id) FROM 拒收问题件 w  GROUP BY 订单编号)
                                        ) lp
                                        WHERE lp.`月份` >= DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 12 month),'%Y%m')
                                ) s1
                                LEFT JOIN gat_order_list gt ON s1.`再次克隆下单` = gt.`订单编号` 
                                GROUP BY 币种, 月份,  登记处理人
                                WITH ROLLUP
                        ) ss1
                        WHERE ss1.币种 IS NOT NULL
                        GROUP BY 币种, 月份, 登记处理人
                        ORDER BY 币种, 月份 DESC,  FIELD(登记处理人,'蔡利英','杨嘉仪','张陈平','邮件','电话','Line','短信','whatsapp','客户问题','刘文君','曲开拓','侯振峰','马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df3333 = pd.read_sql_query(sql=sql3333, con=self.engine1)
            listT.append(df3333)  # 3333 不分币种

            print('分币种 分月份 分新单克隆人 的整体 - 拒收挽单分析')
            sql4444 = '''SELECT 币种, 月份, 新单克隆人,联系单量, 有效联系量, 挽单量,
                                concat(ROUND(有效联系量 / 联系单量 * 100,2),'%') AS 有效联系率,	
                                concat(ROUND(挽单量 / 有效联系量 * 100,2),'%') AS 挽单率, 
                                concat(ROUND(挽单签收量 / 挽单完成量 * 100,2),'%') AS 挽单完成签收,
                                concat(ROUND(挽单签收量 / 挽单量 * 100,2),'%') AS 挽单总计签收, 
                                concat(ROUND(挽单完成量 / 挽单量 * 100,2),'%') AS 挽单完成占比,
                                挽单签收量, 挽单完成量
                        FROM (	SELECT IFNULL(s1.币种,'总计') as 币种,
                                        IFNULL(s1.月份,'总计') as 月份,
                                        IFNULL(s1.新单克隆人,'总计') as 新单克隆人,
                                        COUNT(s1.`订单编号`) AS 联系单量, 
                                        SUM(IF(s1.`具体原因` NOT IN ('无人接听','无效号码','未联系上客户','联系不上客户') ,1,0)) AS 有效联系量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '',1,0)) AS 挽单量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '' AND gt.`系统物流状态` = '已签收',1,0)) AS 挽单签收量,
                                        SUM(IF(s1.`再次克隆下单` IS NOT NULL AND s1.`再次克隆下单` <> '' AND gt.`系统物流状态` IN ('拒收', '理赔', '已签收', '已退货', '自发头程丢件'),1,0)) AS 挽单完成量
                                FROM (  SELECT *
                                        FROM (SELECT *,  DATE_FORMAT(处理时间, '%Y%m') AS 月份, DATE_FORMAT(处理时间, '%Y-%m-%d' ) AS 日期, IF(处理人 = "" OR 处理人 IS NULL,联系方式,处理人) AS 登记处理人
                                                    FROM 拒收问题件 j
                                                    WHERE j.id IN (SELECT MAX(id) FROM 拒收问题件 w  GROUP BY 订单编号)
                                        ) lp
                                        WHERE lp.`月份` >= DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 12 month),'%Y%m')
                                ) s1
                                LEFT JOIN gat_order_list gt ON s1.`再次克隆下单` = gt.`订单编号` 
                                GROUP BY 币种, 月份,  新单克隆人
                                WITH ROLLUP
                        ) ss1
                        WHERE ss1.币种 IS NOT NULL
                        GROUP BY 币种, 月份, 新单克隆人
                        ORDER BY 币种, 月份 DESC,  FIELD(新单克隆人,'蔡利英','杨嘉仪','张陈平','邮件','电话','Line','短信','whatsapp','客户问题','刘文君','曲开拓','侯振峰','马育慧','闫凯歌','杨昊','于海洋','周浩迪','曹可可','曹玉婉','刘君','齐元章','李若兰','袁焕欣','张雨诺','史永巧','康晓雅','蔡贵敏','关梦楠','王苏楠','孙亚茹','夏绍琛','周思文','张静','张登锋','苑亚平','谢玲玲','王正正','王芬','汤楠英','宋晓利','秦小燕','孟芮羽','吕龙飞','李晓青','李青','李丹妮','惠珣','何金蓉','郝淑蓉','丁娜','代下单客服','总计');'''
            df4444 = pd.read_sql_query(sql=sql4444, con=self.engine1)
            listT.append(df4444)  # 2222 不分币种

            file_path = 'F:\\输出文件\\促单&挽单列表查询-分析 {}.xlsx'.format(rq)
            # df0 = pd.DataFrame([])  # 创建空的dataframe数据框
            # df0.to_excel(file_path, index=False)  # 备用：可以向不同的sheet写入数据（创建新的工作表并进行写入）
            # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # 初始化写入对象
            # book = load_workbook(file_path)  # 可以向不同的sheet写入数据（对现有工作表的追加）
            # writer.book = book  # 将数据写入excel中的sheet2表,sheet_name改变后即是新增一个sheet
            # listT[4].to_excel(excel_writer=writer, sheet_name='汇总', index=False)  # 个人
            # listT[6].to_excel(excel_writer=writer, sheet_name='汇总', index=False, startrow=22)      # 不分币种 分月份 的整体
            # listT[3].to_excel(excel_writer=writer, sheet_name='汇总', index=False, startcol=13)      # 不分币种 分月份 的个人
            # listT[1].to_excel(excel_writer=writer, sheet_name='汇总', index=False, startcol=19)      # 有效单量
            # listT[2].to_excel(excel_writer=writer, sheet_name='汇总', index=False, startcol=22)      # 总下单量
            # listT[5].to_excel(excel_writer=writer, sheet_name='汇总', index=False, startcol=25)      # 分币种单量
            # listT[0].to_excel(excel_writer=writer, sheet_name='明细', index=False)     # 明细单量
            # if 'Sheet1' in book.sheetnames:  # 删除新建文档时的第一个工作表
            #     del book['Sheet1']
            # writer.save()
            # writer.close()
            # df.to_excel('F:\\输出文件\\促单查询 {}.xlsx'.format(rq), sheet_name='有效单量', index=False, engine='xlsxwriter')

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df1.to_excel(excel_writer=writer, sheet_name='促单分析', index=False)               # 不分币种 分月份 的整体
                df2.to_excel(excel_writer=writer, sheet_name='促单分析', index=False, startcol=16)  # 不分币种 分月份 的个人
                df3.to_excel(excel_writer=writer, sheet_name='促单分析', index=False, startcol=31)  # 不分币种 分月份 分日期 的个人
                df11.to_excel(excel_writer=writer, sheet_name='促单分析', index=False, startcol=46)  # 分币种 分月份 的整体
                df22.to_excel(excel_writer=writer, sheet_name='促单分析', index=False, startcol=61)  # 分币种 分月份 的个人
                df33.to_excel(excel_writer=writer, sheet_name='促单分析', index=False, startcol=76)  # 分币种 分月份 分日期 的个人
                df111.to_excel(excel_writer=writer, sheet_name='挽单分析', index=False)              # 挽单明细
                df222.to_excel(excel_writer=writer, sheet_name='挽单分析', index=False, startcol=17)  # 分币种 分月份 的整体
                df333.to_excel(excel_writer=writer, sheet_name='挽单分析', index=False, startcol=32)  # 分币种 分月份 的个人
                df444.to_excel(excel_writer=writer, sheet_name='挽单分析', index=False, startcol=47)  # 分币种 分月份 分日期 的个人
                df1111.to_excel(excel_writer=writer, sheet_name='拒收挽单分析', index=False)               # 拒收挽单明细
                df2222.to_excel(excel_writer=writer, sheet_name='拒收挽单分析', index=False, startcol=16)  # 分币种 分月份 的整体
                df3333.to_excel(excel_writer=writer, sheet_name='拒收挽单分析', index=False, startcol=31)  # 分币种 分月份 的整体
                df4444.to_excel(excel_writer=writer, sheet_name='拒收挽单分析', index=False, startcol=46)  # 分币种 分月份 的整体
            # sql = '''SELECT 运单编号
            #         FROM (
            #                 SELECT *
            #                 FROM `cache` s
            #                 WHERE (s.克隆人 IS NULL OR s.克隆人 = "") and s.代下单客服 = "刘文君"
            #         ) s1
            #         WHERE s1.物流状态 <> '已签收' and (s1.运单编号 IS not NULL and s1.运单编号 <> "");'''
            # ordersDict = pd.read_sql_query(sql=sql, con=self.engine1)
            # if ordersDict.empty:
            #     print(' ****** 没有要查询物流轨迹的信息; ****** ')
            # else:
            #     print('！！！ 查询物流轨迹数据中！！！')
            #     handle = '自动'
            #     login_TmpCode = 'login_TmpCode'
            #     if handle == '手动':
            #         print('请输入口令Token:  回车确认')
            #         login_TmpCode = str(input())
            #     login_TmpCode = '0b04de569eb6395e88a34a2e9cde8e92'  # 输入登录口令Tkoen
            #     proxy_handle = '代理服务器0'
            #     proxy_id = '192.168.13.89:37466'  # 输入代理服务器节点和端口
            #     lw = QueryTwo('+86-18538110674', 'qyz04163510.', login_TmpCode, handle, proxy_handle, proxy_id)
            #     lw.Search_online(ordersDict, 1, '运单编号', proxy_handle, proxy_id, 0)
            #     print('物流轨迹已输出！！！')


        print('++++++本批次查询成功+++++++')
        print('*' * 50)
    def _order_track_Query(self, timeStart, timeEnd, n, proxy_id, proxy_handle):
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': n, 'pageSize': 500, 'orderPrefix': None, 'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None, 'type': None, 'collId': None,'isClone': None,
                'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None,  'reassignmentType': None, 'lowerstatus': None, 'warehouse': None,
                'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None, 'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None,
                'volumeEnd': None, 'volumeStart': None, 'chooser_id': None, 'service_id': -1, 'autoVerifyStatus': None, 'shipZip': None, 'remark': None,
                'shipState': None, 'weightStart': None,'weightEnd': None,  'estimateWeightStart': None,  'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None, 'isChangeMark': None,
                'timeStart': timeStart + ' 00:00:00', 'timeEnd': timeEnd + ' 23:59:59'}
        if proxy_handle == '代理服务器':
            proxies = {'http': 'socks5://' + proxy_id, 'https': 'socks5://' + proxy_id}
            req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        else:
            req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        # print('正在处理json数据转化为dataframe…………')
        try:
            for result in req['data']['list']:
                # print(result['orderNumber'])
                if result['specs'] != []:
                    result['saleId'] = 0        # 添加新的字典键-值对，为下面的重新赋值用
                    result['saleName'] = 0
                    result['productId'] = 0
                    result['saleProduct'] = 0
                    result['spec'] = 0
                    result['chooser'] = 0
                    result['saleId'] = result['specs'][0]['saleId']
                    result['saleName'] = result['specs'][0]['saleName']
                    result['productId'] = (result['specs'][0]['saleProduct']).split('#')[1]
                    result['saleProduct'] = (result['specs'][0]['saleProduct']).split('#')[2]
                    result['spec'] = result['specs'][0]['spec']
                    result['chooser'] = result['specs'][0]['chooser']
                else:
                    result['saleId'] = ''        # 添加新的字典键-值对，为下面的重新赋值用
                    result['saleName'] = ''
                    result['productId'] = ''
                    result['saleProduct'] = ''
                    result['spec'] = ''
                    result['chooser'] = ''
                quest = ''
                for re in result['questionReason']:
                    quest = quest + ';' + re
                result['questionReason'] = quest
                delr = ''
                for re in result['delReason']:
                    delr = delr + ';' + re
                result['delReason'] = delr
                auto = ''
                for re in result['autoVerify']:
                    auto = auto + ';' + re
                result['autoVerify'] = auto

                result['auto_VerifyTip'] = ''
                result['auto_VerifyTip_zl'] = ''
                result['auto_VerifyTip_qs'] = ''
                result['auto_VerifyTip_js'] = ''
                if result['autoVerifyTip'] == "":
                    result['auto_VerifyTip'] = '0.00%'
                else:
                    if '未读到拉黑表记录' in result['autoVerifyTip']:
                        result['auto_VerifyTip'] = '0.00%'
                    else:
                        t3 = result['autoVerifyTip']
                        result['auto_VerifyTip_zl'] = (t3.split('订单配送总量：')[1]).split(',')[0]
                        result['auto_VerifyTip_qs'] = (t3.split('送达订单量：')[1]).split(',')[0]
                        result['auto_VerifyTip_js'] = (t3.split('拒收订单量：')[1]).split(',')[0]
                        if '拉黑率问题' not in result['autoVerifyTip']:
                            t2 = result['autoVerifyTip'].split(',拉黑率')[1]
                            result['auto_VerifyTip'] = t2.split('%;')[0] + '%'
                        else:
                            t2 = result['autoVerifyTip'].split('拒收订单量：')[1]
                            t2 = t2.split('%;')[0]
                            result['auto_VerifyTip'] = t2.split('拉黑率')[1] + '%'
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
            sso = Settings_sso()
            sso.send_dingtalk_message("https://oapi.dingtalk.com/robot/send?access_token=68eeb5baf4625d0748b15431800b185fec8056a3dbac2755457f3905b0c8ea1e", "订单检索-获取数据 失败，请检查原因》》》本地数据库：：", ['18538110674'], "是")
        df = pd.json_normalize(ordersdict)
        print('++++++查询输出成功+++++++')
        print('*' * 50)
        return df



if __name__ == '__main__':
    # TODO------------------------------------单点更新配置------------------------------------
    ''' 198.台湾-天马-711
        199.台湾-天马-新竹
        229.台湾-天马-顺丰
        356.台湾-天马-新竹改派
        376.台湾-天马-顺丰改派
        380.台湾-天马-黑猫
        
        85.台湾-速派-新竹
        191.台湾-速派-711超商
        348.台湾-速派-新竹改派
        703.台湾-速派宅配通
        711.台湾-速派宅配通-改派
        722.台湾-速派-黑猫
        
        555.台湾-铱熙无敌-新竹普货
        556.台湾-衣熙无敌-新竹特货
        557.台湾-铱熙无敌-新竹改派
        724.台湾-铱熙无敌-711敏感货
        768.台湾-铱熙无敌-黑猫普货
        769.台湾-铱熙无敌-黑猫特货
        770台湾-铱熙无敌-黑猫改派
        802台湾-铱熙无敌-黑猫改派备货
        
        230.香港-立邦-顺丰
        277香港-立邦-改派
        
        665.香港-圆通
        693.香港-圆通-改派
        
        374台湾-立邦普货头程-森鸿尾程
        383.台湾-立邦普货头程-易速配尾程 
    '''
    proxy_handle = '代理服务器0'
    proxy_id = '192.168.13.89:37466'  # 输入代理服务器节点和端口
    handle = '手0动'
    login_TmpCode = '0bd57ce215513982b1a984d363469e30'  # 输入登录口令Tkoen

    m = QueryOrder('+86-17596568562', 'xhy123456.', login_TmpCode, handle, proxy_id, proxy_handle)
    # m = QueryOrder('+86-15565053520', 'sunan1022wang.@&')
    start: datetime = datetime.datetime.now()
    match1 = {'gat': '港台', 'gat_order_list': '港台', 'slsc': '品牌'}
    # -----------------------------------------------查询状态运行（一）-----------------------------------------
    # 1、按订单查询状态
    # team = 'gat'
    # searchType = '订单号'运单号
    # m.readFormHost(team, searchType)        # 导入；，更新--->>数据更新切换
    # 2、按时间查询状态
    # m.order_TimeQuery('2021-11-01', '2021-11-09')auto_VerifyTip
    # m.del_reson()

    select = 1                                 # 1、 正在按订单查询；2、正在按时间查询；--->>数据更新切换
    if int(select) == 1:
        print("1-->>> 正在按订单查询+++")
        team = 'gat'
        searchType = '订单号'
        pople_Query = '订单检索'                # 客服查询；订单检索
        to_sql = '写0入'                       # 写入：汇总到数据库表； 不写入：直接导出表格
        m.readFormHost(team, searchType, pople_Query, 'timeStart', 'timeEnd', to_sql, proxy_id, proxy_handle)        # 导入；，更新--->>数据更新切换

    elif int(select) == 2:
        print("1-->>> 正在按运单号查询+++")
        team = 'gat'
        searchType = '订单号'
        pople_Query = '客服查询'  # 客服查询；订单检索 运单号
        to_sql = '不写入'  # 写入：汇总到数据库表； 不写入：直接导出表格
        m.readFormHost(team, searchType, pople_Query, 'timeStart', 'timeEnd', to_sql, proxy_id, proxy_handle)  # 导入；，更新--->>数据更新切换

    elif int(select) == 3:
        print("1-->>> 正在按下单时间查询+++")
        timeStart = datetime.date(2023, 4, 1)  # 单点更新
        timeEnd = datetime.date(2023, 4, 14)
        areaId = None
        query = '下单时间'
        # logisticsId = "85,348,199,356"      # 物流名称
        # currencyId = "13"                     # 币种名称：13台湾，6香港
        logisticsId = None      # 物流名称
        currencyId = None                    # 币种名称：13台湾，6香港
        for i in range((timeEnd - timeStart).days):  # 按天循环获取订单状态
            day = timeStart + datetime.timedelta(days=i)
            day_time = str(day)
            m.order_TimeQuery(day_time, day_time, areaId, query, proxy_id, proxy_handle, logisticsId, currencyId)

    elif int(select) == 33:
        print("2-->>> 正在按完成时间查询+++")
        timeStart = datetime.date(2023, 3, 15)  # 单点更新
        timeEnd = datetime.date(2023, 3, 28)
        areaId = ''
        query = '完成时间'
        logisticsId = "85,348,199,356"      # 物流名称
        currencyId = "13"                     # 币种名称
        for i in range((timeEnd - timeStart).days):  # 按天循环获取订单状态
            day = timeStart + datetime.timedelta(days=i)
            day_time = str(day)
            m.order_TimeQuery(day_time, day_time, areaId, query, proxy_id, proxy_handle, logisticsId, currencyId)

    elif int(select) == 4:
        print("1-->>> 正在按电话查询+++")
        team = 'gat'
        searchType = '电话'
        pople_Query = '电话检索'                # 电话查询；订单检索
        timeStart = '2022-10-01 00:00:00'
        timeEnd = '2022-11-30 23:59:59'
        to_sql = '不写入'  # 写入：汇总到数据库表； 不写入：直接导出表格
        m.readFormHost(team, searchType, pople_Query, timeStart, timeEnd, to_sql, proxy_id, proxy_handle)

    # 促单查询；订单检索
    elif int(select) == 5:
        hanlde = '自0动'
        timeStart = '2023-02-01'
        timeEnd = '2023-02-28'

        # timeStart = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        # timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        print(timeStart + "---" + timeEnd)
        m.order_track_Query(hanlde, timeStart, timeEnd, proxy_id, proxy_handle)

    # 订单检索  根据 gat_order_list 源表查询
    elif int(select) == 6:
        hanlde = '自0动'
        timeStart = '2023-01-01'
        timeEnd = '2023-04-11'

        # timeStart = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        # timeEnd = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        print(timeStart + "---" + timeEnd)
        m.order_track_Query(hanlde, timeStart, timeEnd, proxy_id, proxy_handle)

    elif int(select) == 9:
        m.del_order_day_new()


    print('查询耗时：', datetime.datetime.now() - start)