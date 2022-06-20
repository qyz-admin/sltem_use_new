import pandas as pd
import os
import datetime
import time
import xlwings
import xlsxwriter
import math
import requests
import json
import sys
from queue import Queue
from dateutil.relativedelta import relativedelta
from threading import Thread #  使用 threading 模块创建线程
import pandas.io.formats.excel
import win32api,win32con
import win32com.client as win32

from sqlalchemy import create_engine
from settings import Settings
from settings_sso import Settings_sso
from emailControl import EmailControl
from openpyxl import load_workbook  # 可以向不同的sheet写入数据
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment  # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色
from bs4 import BeautifulSoup # 抓标签里面元素的方法

# -*- coding:utf-8 -*-
class import_Data(Settings, Settings_sso):
    def __init__(self, userMobile, password, login_TmpCode,handle):
        Settings.__init__(self)
        Settings_sso.__init__(self)
        self.session = requests.session()  # 实例化session，维持会话,可以让我们在跨请求时保存某些参数
        self.q = Queue()  # 多线程调用的函数不能用return返回值，用来保存返回值
        self.userMobile = userMobile
        self.password = password
        # self.sso_online_cang()
        # self.bulid_file()
        #
        # if handle == '手动':
        #     self.sso_online_cang_handle(login_TmpCode)
        # else:
        #     self.sso_online_cang_auto()

        self.engine1 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql1['user'],
                                                                                    self.mysql1['password'],
                                                                                    self.mysql1['host'],
                                                                                    self.mysql1['port'],
                                                                                    self.mysql1['datebase']))
        self.engine2 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql2['user'],
                                                                                    self.mysql2['password'],
                                                                                    self.mysql2['host'],
                                                                                    self.mysql2['port'],
                                                                                    self.mysql2['datebase']))
        self.engine20 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql20['user'],
                                                                                    self.mysql20['password'],
                                                                                    self.mysql20['host'],
                                                                                    self.mysql20['port'],
                                                                                    self.mysql20['datebase']))
        self.engine3 = create_engine('mysql+mysqlconnector://{}:{}@{}:{}/{}'.format(self.mysql3['user'],
                                                                                    self.mysql3['password'],
                                                                                    self.mysql3['host'],
                                                                                    self.mysql3['port'],
                                                                                    self.mysql3['datebase']))
        self.e = EmailControl()
    # 创建上传 的临时文件
    def _build_file(self):
        print(datetime.datetime.now())
        file_data = input('重新下单:  请输入 运单编号,多个以逗号分割:  ')
        mkpath = r"D:\Users\Administrator\Desktop\需要用到的文件\Temp"
        isExists = os.path.exists(mkpath)
        if not isExists:
            os.makedirs(mkpath)
            print('已成功创建......')
        print('清除临时文件......')
        ls = os.listdir(mkpath)
        for i in ls:
            c_path = os.path.join(mkpath, i)
            os.remove(c_path)
        print('创建临时文件......')
        if file_data[:1].isdigit():         #用isalpha判断是否字母
            file_data = self._waybill_Info(file_data)
            data = [list(file_data['orderNumber'])]
        else:
            if ',' in file_data:                                            # 定义一行数据
                data = [file_data.split(',')]
            else:
                data = [[file_data]]
        print('订单编号: ' + str(data))
        file_path = mkpath + '\\临时写入文件.xlsx'
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()                            # add_worksheet() 添加工作表
        titles = ['订单编号']                                             # 定义 excel 表头信息
        worksheet.write_row(0, 0, titles)                               # 将表头信息写入 excel 的第 0 行、第 0列
        worksheet.write_column(1, 0, data[0])                           # 将数据写入第 1 行、第 0 列
        workbook.close()
        print('写入完成++++++')
        return file_path
    # 获取写入的数据
    def _waybill_Info(self, ord):  # 进入订单检索界面
        self.sso__online_auto()
        print('+++正在查询订单信息中')
        url = r'https://gimp.giikin.com/service?service=gorder.customer&action=getOrderList'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'page': 1, 'pageSize': 500, 'order_number': None, 'shippingNumber': ord,
                'orderNumberFuzzy': None, 'shipUsername': None, 'phone': None, 'email': None, 'ip': None, 'productIds': None,
                'saleIds': None, 'payType': None, 'logisticsId': None, 'logisticsStyle': None, 'logisticsMode': None,
                'type': None, 'collId': None, 'isClone': None,
                'currencyId': None, 'emailStatus': None, 'befrom': None, 'areaId': None, 'reassignmentType': None, 'lowerstatus': '',
                'warehouse': None, 'isEmptyWayBillNumber': None, 'logisticsStatus': None, 'orderStatus': None, 'tuan': None,
                'tuanStatus': None, 'hasChangeSale': None, 'optimizer': None, 'volumeEnd': None, 'volumeStart': None, 'chooser_id': None,
                'service_id': None, 'autoVerifyStatus': None, 'shipZip': None, 'remark': None, 'shipState': None, 'weightStart': None,
                'weightEnd': None, 'estimateWeightStart': None, 'estimateWeightEnd': None, 'order': None, 'sortField': None,
                'orderMark': None, 'remarkCheck': None, 'preSecondWaybill': None, 'whid': None}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        # print('正在处理json数据转化为dataframe…………')
        try:
            for result in req['data']['list']:
                ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersdict)
        df = data[['orderNumber']]
        print('*' * 50)
        return df

    # 导入发货使用 仓储
    def import_delivery_no(self):
        self.sso_online_cang_auto()
        print('正在导入发货中')
        url = r'http://gwms-v3.giikin.cn/order/delivery/importDelivery'
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
            'origin': 'http://gwms-v3.giikin.cn',
            'Referer': 'http://gwms-v3.giikin.cn/order/delivery/importdelivery'}
        file_path = r"H:\桌面\需要用到的文件\文件夹\新建 XLSX 工作表.xlsx"
        up_load_data = {"file": open(file_path, "rb")}
        data = {'0': 1158951430,
                '1': '6/17/2022 10:11'
                }
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, files=up_load_data)
        print(req)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型 或者 str字符串  数据转换为dict字典
        print(req)
        comment = req['success']
        if comment == 1:
            print('上传成功')
        else:
            comment = req['errorMessage']
        return comment

    # 重新下单使用 仓储-上传文件
    def order_change_logistics(self):
        print('正在初始化配置.......')
        self.sso_online_cang_auto()             # 登录系统
        file_path = self._build_file()          # 获取上传文件
        print(file_path)
        print('正在重新下单......')
        url = r'http://gwms-v3.giikin.cn/order/delivery/importDelivery'
        r_header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
            'origin': 'http://gwms-v3.giikin.cn',
            'Referer': 'http://gwms-v3.giikin.cn/order/delivery/importdelivery'}
        up_load_data = {"file": open(file_path, "rb")}
        data = {'logistics_id': 199,
                'dpe_style': 'P'
                }
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data = data, files=up_load_data)
        print(req)
        print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型 或者 str字符串  数据转换为dict字典
        print(req)
        comment = req['success']
        if comment == 1:
            print('上传成功')
        else:
            comment = req['errorMessage']
        return comment

    # 重新下单使用 仓储-获取新单号
    def wb_sheet(self, filePath):
        self.sso__online_auto()
        print('获取新单号.......')
        rq = datetime.datetime.now().strftime('%Y%m%d.%H%M%S')
        app = xlwings.App(visible=False, add_book=False)
        app.display_alerts = False
        wb = app.books.open(filePath, update_links=False, read_only=True)
        sht = wb.sheets[0]
        db = sht.used_range.options(pd.DataFrame, header=1, numbers=int, index=False).value
        orderId = list(db['订单编号'])
        df = pd.DataFrame([])
        dlist = []
        for ord in orderId:
            data = self._wb_sheet(ord)
            dlist.append(data)
        dp = df.append(dlist, ignore_index=True)
        print(dp)
        print('重新下单: 新单号已输出,请复制使用')
        print(dp['重新下单'].to_string(index=False))
        dp.to_excel('G:\输出文件\\重新下单-新单号{}.xlsx'.format(rq), sheet_name='查询', index=False,  engine='xlsxwriter')  # Xlsx是python用来构造xlsx文件的模块，可以向excel2007+中写text，numbers，formulas 公式以及hyperlinks超链接。
        print('查询已导出+++')
        wb.close()
        app.quit()
    def _wb_sheet(self, ord):  # 进入订单检索界面
        print('正在查询......')
        url = r'https://gimp.giikin.com/service?service=gorder.order&action=getOrderLog'
        r_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                    'origin': 'https: // gimp.giikin.com',
                    'Referer': 'https://gimp.giikin.com/front/orderToolsOrderSearch'}
        data = {'orderKey': ord}
        proxy = '39.105.167.0:40005'  # 使用代理服务器
        proxies = {'http': 'socks5://' + proxy,
                   'https': 'socks5://' + proxy}
        # req = self.session.post(url=url, headers=r_header, data=data, proxies=proxies)
        req = self.session.post(url=url, headers=r_header, data=data)
        # print('+++已成功发送请求......')
        req = json.loads(req.text)  # json类型数据转换为dict字典
        # print(req)
        ordersdict = []
        # print('正在处理json数据转化为dataframe…………')
        try:
            result = req['data'][0]
            result['订单编号'] = result['orderNumber']
            val = result['remark'].split('waybill_number,')[1]
            result['旧运单号'] = val.split('->')[0]
            result['新单号'] = val.split('->')[1]
            result['重新下单'] = val.replace('->', '新单号')
            ordersdict.append(result)
        except Exception as e:
            print('转化失败： 重新获取中', str(Exception) + str(e))
        data = pd.json_normalize(ordersdict)
        data = data[['订单编号','旧运单号','新单号','重新下单']]
        print('*' * 50)
        return data


if __name__ == '__main__':
    m = import_Data('+86-18538110674', 'qyz35100416','84c3a0212a7b3de386b2a20d4a46b0ea','手0动')

    # m._build_file()
    # m.order_change_logistics()

    # 重新下单 - 上传文件
    m.order_change_logistics()
    # 重新下单 - 获取新单号
    m.wb_sheet(r'D:\Users\Administrator\Desktop\需要用到的文件\Temp\临时写入文件.xlsx')
